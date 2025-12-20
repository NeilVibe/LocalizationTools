"""
LDM (LanguageData Manager) API Endpoints

REST API for managing localization projects, folders, files, and rows.
Supports real-time collaboration via WebSocket (see websocket.py).
"""

import asyncio
import hashlib
from collections import defaultdict
from functools import partial
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from sqlalchemy.orm import selectinload
from typing import Optional, List
from datetime import datetime
from pydantic import BaseModel
from loguru import logger
from io import BytesIO

from server.utils.dependencies import get_async_db, get_current_active_user_async, get_db
from server.database.db_utils import normalize_text_for_hash
from server.database.models import (
    User, LDMProject, LDMFolder, LDMFile, LDMRow, LDMEditHistory, LDMActiveSession,
    LDMTranslationMemory, LDMTMEntry
)
from server.tools.ldm.websocket import broadcast_cell_update


# ============================================================================
# Pydantic Schemas
# ============================================================================

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None

class ProjectResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    owner_id: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class FolderCreate(BaseModel):
    project_id: int
    parent_id: Optional[int] = None
    name: str

class FolderResponse(BaseModel):
    id: int
    project_id: int
    parent_id: Optional[int]
    name: str
    created_at: datetime

    class Config:
        from_attributes = True

class FileResponse(BaseModel):
    id: int
    project_id: int
    folder_id: Optional[int]
    name: str
    original_filename: str
    format: str
    row_count: int
    source_language: str
    target_language: Optional[str]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class RowResponse(BaseModel):
    id: int
    file_id: int
    row_num: int
    string_id: Optional[str]
    source: Optional[str]
    target: Optional[str]
    status: str
    updated_at: datetime

    class Config:
        from_attributes = True

class RowUpdate(BaseModel):
    target: Optional[str] = None
    status: Optional[str] = None

class PaginatedRows(BaseModel):
    rows: List[RowResponse]
    total: int
    page: int
    limit: int
    total_pages: int


class TMResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    source_lang: str
    target_lang: str
    entry_count: int
    status: str
    created_at: Optional[datetime]

    class Config:
        from_attributes = True


class TMUploadResponse(BaseModel):
    tm_id: int
    name: str
    entry_count: int
    status: str
    time_seconds: float
    rate_per_second: int


class TMSearchResult(BaseModel):
    source_text: str
    target_text: str
    similarity: float
    tier: int
    strategy: str


class DeleteResponse(BaseModel):
    """Standard response for delete operations."""
    message: str


class TMSuggestion(BaseModel):
    """Single TM suggestion."""
    id: int
    source: str
    target: str
    file_id: int
    file_name: str
    similarity: float


class TMSuggestResponse(BaseModel):
    """Response from TM suggest endpoint."""
    suggestions: List[TMSuggestion]
    count: int


class PretranslateRequest(BaseModel):
    """Request for pretranslation."""
    file_id: int
    engine: str  # "standard" | "xls_transfer" | "kr_similar"
    dictionary_id: int
    threshold: float = 0.92
    skip_existing: bool = True


class PretranslateResponse(BaseModel):
    """Response from pretranslation."""
    file_id: int
    engine: str
    matched: int
    skipped: int
    total: int
    threshold: float
    time_seconds: float


# ============================================================================
# Router
# ============================================================================

router = APIRouter(prefix="/api/ldm", tags=["LDM"])


# ============================================================================
# Health Check
# ============================================================================

@router.get("/health")
async def health():
    """Health check endpoint for LDM module."""
    logger.info("LDM health check requested")
    return {
        "status": "ok",
        "module": "LDM (LanguageData Manager)",
        "version": "1.0.0",
        "features": {
            "projects": True,
            "folders": True,
            "files": True,
            "rows": True,
            "websocket": True,  # Phase 3: Real-time sync enabled
            "row_locking": True,
            "presence": True
        }
    }


# ============================================================================
# Projects
# ============================================================================

@router.get("/projects", response_model=List[ProjectResponse])
async def list_projects(
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """List all projects for current user."""
    user_id = current_user["user_id"]
    logger.info(f"Listing projects for user {user_id}")

    result = await db.execute(
        select(LDMProject).where(LDMProject.owner_id == current_user["user_id"])
    )
    projects = result.scalars().all()

    return projects


@router.post("/projects", response_model=ProjectResponse)
async def create_project(
    project: ProjectCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Create a new project."""
    user_id = current_user["user_id"]
    logger.info(f"Creating project '{project.name}' for user {user_id}")

    new_project = LDMProject(
        name=project.name,
        description=project.description,
        owner_id=current_user["user_id"]
    )

    db.add(new_project)
    await db.commit()
    await db.refresh(new_project)

    logger.success(f"Project created: id={new_project.id}, name='{new_project.name}'")
    return new_project


@router.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(
    project_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Get a project by ID."""
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return project


@router.delete("/projects/{project_id}", response_model=DeleteResponse)
async def delete_project(
    project_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Delete a project and all its contents."""
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    await db.delete(project)
    await db.commit()

    logger.info(f"Project deleted: id={project_id}")
    return {"message": "Project deleted"}


# ============================================================================
# Folders
# ============================================================================

@router.get("/projects/{project_id}/folders", response_model=List[FolderResponse])
async def list_folders(
    project_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """List all folders in a project."""
    # Verify project ownership
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Project not found")

    result = await db.execute(
        select(LDMFolder).where(LDMFolder.project_id == project_id)
    )
    folders = result.scalars().all()

    return folders


@router.post("/folders", response_model=FolderResponse)
async def create_folder(
    folder: FolderCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Create a new folder in a project."""
    # Verify project ownership
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == folder.project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Project not found")

    new_folder = LDMFolder(
        project_id=folder.project_id,
        parent_id=folder.parent_id,
        name=folder.name
    )

    db.add(new_folder)
    await db.commit()
    await db.refresh(new_folder)

    logger.info(f"Folder created: id={new_folder.id}, name='{new_folder.name}'")
    return new_folder


@router.delete("/folders/{folder_id}", response_model=DeleteResponse)
async def delete_folder(
    folder_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Delete a folder and all its contents."""
    result = await db.execute(
        select(LDMFolder).options(selectinload(LDMFolder.project)).where(
            LDMFolder.id == folder_id
        )
    )
    folder = result.scalar_one_or_none()

    if not folder:
        raise HTTPException(status_code=404, detail="Folder not found")

    if folder.project.owner_id != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    await db.delete(folder)
    await db.commit()

    logger.info(f"Folder deleted: id={folder_id}")
    return {"message": "Folder deleted"}


# ============================================================================
# Files
# ============================================================================

@router.get("/projects/{project_id}/files", response_model=List[FileResponse])
async def list_files(
    project_id: int,
    folder_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """List files in a project, optionally filtered by folder."""
    # Verify project ownership
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Project not found")

    query = select(LDMFile).where(LDMFile.project_id == project_id)
    if folder_id is not None:
        query = query.where(LDMFile.folder_id == folder_id)

    result = await db.execute(query)
    files = result.scalars().all()

    return files


@router.get("/files/{file_id}", response_model=FileResponse)
async def get_file(
    file_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Get file metadata by ID."""
    result = await db.execute(
        select(LDMFile).options(selectinload(LDMFile.project)).where(
            LDMFile.id == file_id
        )
    )
    file = result.scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    if file.project.owner_id != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    return file


@router.post("/files/upload", response_model=FileResponse)
async def upload_file(
    project_id: int = Form(...),
    folder_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    # Excel column mapping (optional, only used for Excel files)
    source_col: Optional[int] = Form(None),      # Column index for source (0=A, 1=B, etc.)
    target_col: Optional[int] = Form(None),      # Column index for target
    stringid_col: Optional[int] = Form(None),    # Column index for StringID (None = 2-column mode)
    has_header: Optional[bool] = Form(True),     # Whether first row is header
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Upload a localization file (TXT/XML/Excel), parse it, and store rows in database.

    Supported formats:
    - TXT/TSV: Tab-delimited, columns 0-4=StringID, 5=Source(KR), 6=Target, 7+=extra
    - XML: LocStr elements with StringId, StrOrigin(source), Str(target), other attrs=extra
    - Excel: User-defined columns via source_col, target_col, stringid_col parameters
      - 2-column mode: Source + Target (default: A, B)
      - 3-column mode: Source + Target + StringID (e.g., A, B, C)

    Full Structure Preservation:
    - File-level metadata stored in LDMFile.extra_data (encoding, root element, etc.)
    - Row-level extra data stored in LDMRow.extra_data (extra columns/attributes)
    - Enables FULL reconstruction of original file format
    """
    # Verify project ownership
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Project not found")

    # Verify folder exists (if provided)
    if folder_id:
        result = await db.execute(
            select(LDMFolder).where(
                LDMFolder.id == folder_id,
                LDMFolder.project_id == project_id
            )
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="Folder not found")

    # Determine file type and parse
    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    file_metadata = None
    if ext in ('txt', 'tsv'):
        from server.tools.ldm.file_handlers.txt_handler import parse_txt_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_txt_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext == 'xml':
        from server.tools.ldm.file_handlers.xml_handler import parse_xml_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_xml_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext in ('xlsx', 'xls'):
        from server.tools.ldm.file_handlers.excel_handler import parse_excel_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        # Use provided column mappings or defaults (A=0, B=1)
        rows_data = parse_excel_file(
            file_content,
            filename,
            source_col=source_col if source_col is not None else 0,
            target_col=target_col if target_col is not None else 1,
            stringid_col=stringid_col,  # None means 2-column mode
            has_header=has_header if has_header is not None else True
        )
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {ext}. Use TXT, TSV, XML, or Excel (XLSX/XLS)."
        )

    if not rows_data:
        raise HTTPException(
            status_code=400,
            detail="No valid rows found in file"
        )

    # TASK-001: Add TrackedOperation for progress tracking on large files
    # Run row insertion in thread with progress tracking
    user_id = current_user["user_id"]
    username = current_user.get("username", "unknown")
    total_rows = len(rows_data)

    def _insert_file_and_rows():
        from server.utils.progress_tracker import TrackedOperation

        sync_db = next(get_db())
        try:
            with TrackedOperation(
                operation_name=f"Upload: {filename}",
                user_id=user_id,
                username=username,
                tool_name="LDM",
                function_name="upload_file",
                parameters={"filename": filename, "row_count": total_rows}
            ) as tracker:
                # Create file record
                new_file = LDMFile(
                    project_id=project_id,
                    folder_id=folder_id,
                    name=filename,
                    original_filename=filename,
                    format=file_format,
                    row_count=total_rows,
                    source_language=source_lang,
                    target_language=None,
                    extra_data=file_metadata
                )
                sync_db.add(new_file)
                sync_db.flush()  # Get the file ID

                tracker.update(10, "File record created", 1, total_rows + 1)

                # Create row records with progress
                for i, row_data in enumerate(rows_data):
                    row = LDMRow(
                        file_id=new_file.id,
                        row_num=row_data["row_num"],
                        string_id=row_data["string_id"],
                        source=row_data["source"],
                        target=row_data["target"],
                        status=row_data["status"],
                        extra_data=row_data.get("extra_data")
                    )
                    sync_db.add(row)

                    # Update progress every 100 rows
                    if (i + 1) % 100 == 0 or i == total_rows - 1:
                        progress_pct = 10 + ((i + 1) / total_rows) * 90
                        tracker.update(progress_pct, f"Row {i + 1}/{total_rows}", i + 1, total_rows)

                sync_db.commit()
                sync_db.refresh(new_file)

                return {
                    "id": new_file.id,
                    "project_id": new_file.project_id,
                    "folder_id": new_file.folder_id,
                    "name": new_file.name,
                    "original_filename": new_file.original_filename,
                    "format": new_file.format,
                    "row_count": new_file.row_count,
                    "source_language": new_file.source_language,
                    "target_language": new_file.target_language,
                    "created_at": new_file.created_at,
                    "updated_at": new_file.updated_at
                }
        finally:
            sync_db.close()

    result = await asyncio.to_thread(_insert_file_and_rows)

    logger.success(f"File uploaded: id={result['id']}, name='{filename}', rows={total_rows}, has_metadata={file_metadata is not None}")

    # Return FileResponse-compatible dict
    return result


@router.post("/files/excel-preview")
async def excel_preview(
    file: UploadFile = File(...),
    max_rows: int = Query(5, ge=1, le=20),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get a preview of an Excel file for column mapping UI.

    Returns:
        - sheet_name: Name of the active sheet
        - headers: Column headers (from first row)
        - sample_rows: First N rows of data
        - total_rows: Total row count

    Use this before uploading to let users select which columns
    are Source, Target, and StringID.
    """
    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    if ext not in ('xlsx', 'xls'):
        raise HTTPException(
            status_code=400,
            detail=f"Excel preview only supports XLSX/XLS files, got: {ext}"
        )

    try:
        from server.tools.ldm.file_handlers.excel_handler import get_excel_preview
        file_content = await file.read()
        preview = get_excel_preview(file_content, max_rows=max_rows)
        return preview
    except Exception as e:
        logger.error(f"Excel preview failed: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================================
# Rows (Pagination)
# ============================================================================

@router.get("/files/{file_id}/rows", response_model=PaginatedRows)
async def list_rows(
    file_id: int,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Get paginated rows for a file."""
    # Verify file access
    result = await db.execute(
        select(LDMFile).options(selectinload(LDMFile.project)).where(
            LDMFile.id == file_id
        )
    )
    file = result.scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    if file.project.owner_id != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Build query
    query = select(LDMRow).where(LDMRow.file_id == file_id)
    count_query = select(func.count(LDMRow.id)).where(LDMRow.file_id == file_id)

    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            (LDMRow.source.ilike(search_pattern)) |
            (LDMRow.target.ilike(search_pattern)) |
            (LDMRow.string_id.ilike(search_pattern))
        )
        count_query = count_query.where(
            (LDMRow.source.ilike(search_pattern)) |
            (LDMRow.target.ilike(search_pattern)) |
            (LDMRow.string_id.ilike(search_pattern))
        )

    if status:
        query = query.where(LDMRow.status == status)
        count_query = count_query.where(LDMRow.status == status)

    # Get total count
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Get paginated rows
    offset = (page - 1) * limit
    query = query.order_by(LDMRow.row_num).offset(offset).limit(limit)

    result = await db.execute(query)
    rows = result.scalars().all()

    total_pages = (total + limit - 1) // limit

    return PaginatedRows(
        rows=rows,
        total=total,
        page=page,
        limit=limit,
        total_pages=total_pages
    )


@router.put("/rows/{row_id}", response_model=RowResponse)
async def update_row(
    row_id: int,
    update: RowUpdate,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Update a row's target text or status (source is READ-ONLY)."""
    # Get row with file and project
    result = await db.execute(
        select(LDMRow).options(
            selectinload(LDMRow.file).selectinload(LDMFile.project)
        ).where(LDMRow.id == row_id)
    )
    row = result.scalar_one_or_none()

    if not row:
        raise HTTPException(status_code=404, detail="Row not found")

    if row.file.project.owner_id != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Save history before update
    old_target = row.target
    old_status = row.status

    # Update row
    if update.target is not None:
        row.target = update.target
        # Auto-set status to translated if target is set and was pending
        if row.status == "pending" and update.target:
            row.status = "translated"

    if update.status is not None:
        row.status = update.status

    row.updated_by = current_user["user_id"]
    row.updated_at = datetime.utcnow()

    # Create edit history
    history = LDMEditHistory(
        row_id=row_id,
        user_id=current_user["user_id"],
        old_target=old_target,
        new_target=row.target,
        old_status=old_status,
        new_status=row.status
    )
    db.add(history)

    await db.commit()
    await db.refresh(row)

    # Broadcast cell update to all viewers (real-time sync)
    try:
        await broadcast_cell_update(
            file_id=row.file_id,
            row_id=row.id,
            row_num=row.row_num,
            target=row.target,
            status=row.status,
            updated_by=current_user["user_id"],
            updated_by_username=current_user["username"]
        )
    except Exception as e:
        # WebSocket broadcast failure shouldn't fail the API call
        logger.warning(f"WebSocket broadcast failed for row {row_id}: {e}")

    user_id = current_user["user_id"]
    logger.info(f"Row updated: id={row_id}, user={user_id}")
    return row


# ============================================================================
# Project Tree (Full structure for File Explorer)
# ============================================================================

@router.get("/projects/{project_id}/tree")
async def get_project_tree(
    project_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Get full project tree structure (folders + files) for File Explorer."""
    # Verify project ownership
    result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get all folders
    result = await db.execute(
        select(LDMFolder).where(LDMFolder.project_id == project_id)
    )
    folders = result.scalars().all()

    # Get all files
    result = await db.execute(
        select(LDMFile).where(LDMFile.project_id == project_id)
    )
    files = result.scalars().all()

    # Build lookup dicts for O(n) tree building instead of O(n*m)
    folders_by_parent = defaultdict(list)
    files_by_folder = defaultdict(list)

    for folder in folders:
        folders_by_parent[folder.parent_id].append(folder)
    for file in files:
        files_by_folder[file.folder_id].append(file)

    def build_tree(parent_id=None):
        tree = []

        # Add folders at this level (O(1) lookup)
        for folder in folders_by_parent.get(parent_id, []):
            tree.append({
                "type": "folder",
                "id": folder.id,
                "name": folder.name,
                "children": build_tree(folder.id)
            })

        # Add files at this level (O(1) lookup)
        for file in files_by_folder.get(parent_id, []):
            tree.append({
                "type": "file",
                "id": file.id,
                "name": file.name,
                "format": file.format,
                "row_count": file.row_count
            })

        return tree

    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "description": project.description
        },
        "tree": build_tree(None)
    }


# ============================================================================
# Translation Memory (TM) API
# ============================================================================

@router.get("/tm/suggest", response_model=TMSuggestResponse)
async def get_tm_suggestions(
    source: str,
    file_id: Optional[int] = None,
    project_id: Optional[int] = None,
    exclude_row_id: Optional[int] = None,
    threshold: float = Query(0.70, ge=0.0, le=1.0, description="Similarity threshold (0.0-1.0)"),
    max_results: int = Query(5, ge=1, le=50, description="Maximum suggestions to return"),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get Translation Memory suggestions for a source text.

    Finds similar source texts in the database and returns their translations.

    Args:
        source: Korean source text to find matches for
        file_id: Optional - limit search to same file
        project_id: Optional - limit search to same project
        exclude_row_id: Optional - exclude this row from results
        threshold: Minimum similarity (0.0-1.0, default 0.70)
        max_results: Maximum suggestions (default 5)

    Returns:
        List of TM suggestions with source, target, similarity, etc.
    """
    from server.tools.ldm.tm import TranslationMemory

    logger.info(f"TM suggest: source={source[:30]}..., file={file_id}, project={project_id}")

    try:
        # Use PostgreSQL pg_trgm similarity() for efficient fuzzy matching
        # This is O(log n) with GIN index vs O(n) Python loop
        from sqlalchemy import text, literal_column

        # Build SQL with pg_trgm similarity
        # Note: Requires pg_trgm extension and GIN index on source column
        sql_params = {
            'search_text': source.strip(),
            'threshold': threshold,
            'max_results': max_results
        }

        # Base conditions - use parameterized queries (NO f-strings for SQL!)
        conditions = ["r.target IS NOT NULL", "r.target != ''"]
        if file_id:
            conditions.append("r.file_id = :file_id")
            sql_params['file_id'] = file_id
        elif project_id:
            conditions.append("f.project_id = :project_id")
            sql_params['project_id'] = project_id
        if exclude_row_id:
            conditions.append("r.id != :exclude_row_id")
            sql_params['exclude_row_id'] = exclude_row_id

        where_clause = " AND ".join(conditions)

        # Use pg_trgm similarity() for fuzzy matching
        # Falls back to exact match if pg_trgm not available
        sql = text(f"""
            SELECT
                r.id,
                r.source,
                r.target,
                r.file_id,
                f.name as file_name,
                similarity(lower(r.source), lower(:search_text)) as sim
            FROM ldm_rows r
            JOIN ldm_files f ON r.file_id = f.id
            WHERE {where_clause}
              AND similarity(lower(r.source), lower(:search_text)) >= :threshold
            ORDER BY sim DESC
            LIMIT :max_results
        """)

        result = await db.execute(sql, sql_params)
        rows = result.fetchall()

        suggestions = [
            {
                'source': row.source,
                'target': row.target,
                'similarity': round(float(row.sim), 3),
                'row_id': row.id,
                'file_name': row.file_name
            }
            for row in rows
        ]

        logger.info(f"TM found {len(suggestions)} suggestions (pg_trgm)")
        return {"suggestions": suggestions, "count": len(suggestions)}

    except Exception as e:
        logger.error(f"TM suggest failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="TM search failed. Check server logs.")


# ============================================================================
# Translation Memory CRUD (TMManager-based)
# ============================================================================

@router.post("/tm/upload", response_model=TMUploadResponse)
async def upload_tm(
    name: str = Form(...),
    source_lang: str = Form("ko"),
    target_lang: str = Form("en"),
    description: Optional[str] = Form(None),
    file: UploadFile = File(...),
    # TM mode: "standard" (duplicates merged) or "stringid" (all variations kept)
    mode: str = Form("standard"),
    # Excel column mapping (only used for Excel files)
    source_col: Optional[int] = Form(None),      # Column index for source (0=A)
    target_col: Optional[int] = Form(None),      # Column index for target (1=B)
    stringid_col: Optional[int] = Form(None),    # Column index for StringID (2=C)
    has_header: Optional[bool] = Form(True),     # Whether first row is header
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Upload a Translation Memory file.

    Supported formats:
    - TXT/TSV: Column 5 = Source, Column 6 = Target
    - XML: StrOrigin = Source, Str = Target
    - XLSX: User-defined columns via source_col, target_col, stringid_col

    Modes:
    - standard: Duplicates merged (most frequent target wins)
    - stringid: All variations kept (same source, different StringIDs)

    Performance: ~20,000+ entries/second bulk insert
    """
    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    if ext not in ('txt', 'tsv', 'xml', 'xlsx', 'xls'):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported TM format: {ext}. Use TXT, TSV, XML, or XLSX."
        )

    logger.info(f"TM upload started: name={name}, file={filename}, user={current_user['user_id']}")

    # TASK-001: Add TrackedOperation for progress tracking
    user_id = current_user["user_id"]
    username = current_user.get("username", "unknown")

    try:
        file_content = await file.read()

        # Run sync TMManager in threadpool to avoid blocking event loop
        def _upload_tm():
            from server.utils.progress_tracker import TrackedOperation

            sync_db = next(get_db())
            try:
                with TrackedOperation(
                    operation_name=f"Upload TM: {name}",
                    user_id=user_id,
                    username=username,
                    tool_name="LDM",
                    function_name="upload_tm",
                    parameters={"name": name, "filename": filename, "mode": mode}
                ) as tracker:
                    tracker.update(10, "Parsing file...")

                    from server.tools.ldm.tm_manager import TMManager
                    tm_manager = TMManager(sync_db)

                    tracker.update(30, "Inserting entries...")

                    result = tm_manager.upload_tm(
                        file_content=file_content,
                        filename=filename,
                        name=name,
                        owner_id=user_id,
                        source_lang=source_lang,
                        target_lang=target_lang,
                        description=description,
                        mode=mode,
                        source_col=source_col if source_col is not None else 0,
                        target_col=target_col if target_col is not None else 1,
                        stringid_col=stringid_col,
                        has_header=has_header if has_header is not None else True
                    )

                    tracker.update(100, f"Complete: {result['entry_count']} entries")
                    return result
            finally:
                sync_db.close()

        result = await asyncio.to_thread(_upload_tm)
        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid TM data provided")
    except Exception as e:
        logger.error(f"TM upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="TM upload failed. Check server logs.")


@router.get("/tm", response_model=List[TMResponse])
async def list_tms(
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """List all Translation Memories for current user."""
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.owner_id == current_user["user_id"]
        ).order_by(LDMTranslationMemory.created_at.desc())
    )
    tms = result.scalars().all()

    # Debug: Log actual status values from DB
    for tm in tms:
        logger.debug(f"TM {tm.id} '{tm.name}': status='{tm.status}', entry_count={tm.entry_count}")

    logger.info(f"Listed {len(tms)} TMs for user {current_user['user_id']}")
    return tms


@router.get("/tm/{tm_id}", response_model=TMResponse)
async def get_tm(
    tm_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Get a Translation Memory by ID."""
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    return tm


@router.delete("/tm/{tm_id}", response_model=DeleteResponse)
async def delete_tm(
    tm_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """Delete a Translation Memory and all its entries."""
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    entry_count = tm.entry_count
    await db.delete(tm)
    await db.commit()

    logger.info(f"Deleted TM: id={tm_id}, entries={entry_count}")
    return {"message": "Translation Memory deleted", "entries_deleted": entry_count}


# =============================================================================
# TM Viewer Endpoints (FEAT-003)
# =============================================================================

@router.get("/tm/{tm_id}/entries")
async def get_tm_entries(
    tm_id: int,
    page: int = 1,
    limit: int = 100,
    sort_by: str = "id",
    sort_order: str = "asc",
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get paginated TM entries for TM Viewer.

    Query params:
    - page: Page number (1-indexed)
    - limit: Items per page (max 500)
    - sort_by: Field to sort by (id, source_text, target_text, string_id, created_at)
    - sort_order: asc or desc
    - search: Search term (searches source, target, and string_id)
    """
    from sqlalchemy import or_

    # Verify TM ownership
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = tm_result.scalar_one_or_none()
    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Validate and cap limit
    limit = min(limit, 500)
    page = max(page, 1)
    offset = (page - 1) * limit

    # Build query
    query = select(LDMTMEntry).where(LDMTMEntry.tm_id == tm_id)

    # Apply search filter
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            or_(
                LDMTMEntry.source_text.ilike(search_pattern),
                LDMTMEntry.target_text.ilike(search_pattern),
                LDMTMEntry.string_id.ilike(search_pattern)
            )
        )

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    count_result = await db.execute(count_query)
    total = count_result.scalar()

    # Apply sorting
    sort_column = {
        "id": LDMTMEntry.id,
        "source_text": LDMTMEntry.source_text,
        "target_text": LDMTMEntry.target_text,
        "string_id": LDMTMEntry.string_id,
        "created_at": LDMTMEntry.created_at
    }.get(sort_by, LDMTMEntry.id)

    if sort_order.lower() == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    # Apply pagination
    query = query.offset(offset).limit(limit)

    # Execute
    result = await db.execute(query)
    entries = result.scalars().all()

    # Format entries
    formatted_entries = [
        {
            "id": e.id,
            "source_text": e.source_text,
            "target_text": e.target_text,
            "string_id": e.string_id,
            "created_at": e.created_at.isoformat() if e.created_at else None
        }
        for e in entries
    ]

    total_pages = (total + limit - 1) // limit if total > 0 else 1

    return {
        "entries": formatted_entries,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "sort_by": sort_by,
        "sort_order": sort_order,
        "search": search,
        "tm_name": tm.name
    }


@router.put("/tm/{tm_id}/entries/{entry_id}")
async def update_tm_entry(
    tm_id: int,
    entry_id: int,
    source_text: Optional[str] = None,
    target_text: Optional[str] = None,
    string_id: Optional[str] = None,
    background_tasks: BackgroundTasks = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Update a single TM entry (for inline editing in TM Viewer).
    Q-001: Auto-syncs TM indexes after update.
    """
    # Verify TM ownership
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = tm_result.scalar_one_or_none()
    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Find entry
    entry_result = await db.execute(
        select(LDMTMEntry).where(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        )
    )
    entry = entry_result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    # Update fields if provided
    if source_text is not None:
        entry.source_text = source_text
        # Recalculate hash
        normalized = normalize_text_for_hash(source_text)
        entry.source_hash = hashlib.sha256(normalized.encode('utf-8')).hexdigest()

    if target_text is not None:
        entry.target_text = target_text

    if string_id is not None:
        entry.string_id = string_id

    # BUG-020: Track who made the update
    entry.updated_at = datetime.utcnow()
    entry.updated_by = current_user.get("username", "unknown")

    # Mark TM as updated (triggers index rebuild on next pretranslate)
    tm.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(entry)

    logger.info(f"Updated TM entry: tm_id={tm_id}, entry_id={entry_id}, by={current_user.get('username')}")

    # Q-001: Auto-sync indexes in background
    if background_tasks:
        background_tasks.add_task(_auto_sync_tm_indexes, tm_id, current_user["user_id"])

    return {
        "id": entry.id,
        "source_text": entry.source_text,
        "target_text": entry.target_text,
        "string_id": entry.string_id,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "updated_by": entry.updated_by,
        "is_confirmed": entry.is_confirmed,
        "confirmed_at": entry.confirmed_at.isoformat() if entry.confirmed_at else None,
        "confirmed_by": entry.confirmed_by
    }


@router.delete("/tm/{tm_id}/entries/{entry_id}")
async def delete_tm_entry(
    tm_id: int,
    entry_id: int,
    background_tasks: BackgroundTasks = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Delete a single TM entry.
    Q-001: Auto-syncs TM indexes after delete.
    """
    # Verify TM ownership
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = tm_result.scalar_one_or_none()
    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Find entry
    entry_result = await db.execute(
        select(LDMTMEntry).where(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        )
    )
    entry = entry_result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    await db.delete(entry)

    # Update TM entry count and updated_at
    tm.entry_count = max(0, (tm.entry_count or 0) - 1)
    tm.updated_at = datetime.utcnow()

    await db.commit()

    logger.info(f"Deleted TM entry: tm_id={tm_id}, entry_id={entry_id}")

    # Q-001: Auto-sync indexes in background
    if background_tasks:
        background_tasks.add_task(_auto_sync_tm_indexes, tm_id, current_user["user_id"])

    return {"message": "Entry deleted", "entry_id": entry_id}


# BUG-020: Confirm/Unconfirm endpoints for memoQ-style workflow
@router.post("/tm/{tm_id}/entries/{entry_id}/confirm")
async def confirm_tm_entry(
    tm_id: int,
    entry_id: int,
    confirm: bool = True,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    BUG-020: Confirm or unconfirm a TM entry (memoQ-style workflow).

    When user approves a translation, it gets marked as confirmed with metadata.
    """
    # Verify TM ownership
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Find entry
    entry_result = await db.execute(
        select(LDMTMEntry).where(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        )
    )
    entry = entry_result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    username = current_user.get("username", "unknown")

    if confirm:
        entry.is_confirmed = True
        entry.confirmed_at = datetime.utcnow()
        entry.confirmed_by = username
    else:
        entry.is_confirmed = False
        entry.confirmed_at = None
        entry.confirmed_by = None

    await db.commit()
    await db.refresh(entry)

    logger.info(f"{'Confirmed' if confirm else 'Unconfirmed'} TM entry: tm_id={tm_id}, entry_id={entry_id}, by={username}")

    return {
        "id": entry.id,
        "source_text": entry.source_text,
        "target_text": entry.target_text,
        "string_id": entry.string_id,
        "is_confirmed": entry.is_confirmed,
        "confirmed_at": entry.confirmed_at.isoformat() if entry.confirmed_at else None,
        "confirmed_by": entry.confirmed_by
    }


@router.post("/tm/{tm_id}/entries/bulk-confirm")
async def bulk_confirm_tm_entries(
    tm_id: int,
    entry_ids: List[int],
    confirm: bool = True,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    BUG-020: Bulk confirm/unconfirm multiple TM entries.
    """
    from sqlalchemy import update

    # Verify TM ownership
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    username = current_user.get("username", "unknown")
    now = datetime.utcnow()

    if confirm:
        stmt = update(LDMTMEntry).where(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.id.in_(entry_ids)
        ).values(
            is_confirmed=True,
            confirmed_at=now,
            confirmed_by=username
        )
    else:
        stmt = update(LDMTMEntry).where(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.id.in_(entry_ids)
        ).values(
            is_confirmed=False,
            confirmed_at=None,
            confirmed_by=None
        )

    result = await db.execute(stmt)
    await db.commit()

    updated_count = result.rowcount

    logger.info(f"Bulk {'confirmed' if confirm else 'unconfirmed'} {updated_count} TM entries: tm_id={tm_id}, by={username}")

    return {
        "updated_count": updated_count,
        "action": "confirmed" if confirm else "unconfirmed"
    }


@router.get("/tm/{tm_id}/search/exact")
async def search_tm_exact(
    tm_id: int,
    source: str,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Search for exact match in a Translation Memory.

    Uses hash-based O(1) lookup for maximum speed.
    """
    logger.info(f"TM exact search: tm_id={tm_id}, source={source[:30]}...")

    # Verify TM ownership (async)
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Generate hash for O(1) lookup (async query)
    normalized = normalize_text_for_hash(source)
    source_hash = hashlib.sha256(normalized.encode('utf-8')).hexdigest()

    entry_result = await db.execute(
        select(LDMTMEntry).where(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.source_hash == source_hash
        )
    )
    entry = entry_result.scalar_one_or_none()

    if entry:
        return {
            "match": {
                "source_text": entry.source_text,
                "target_text": entry.target_text,
                "similarity": 1.0,
                "tier": 1,
                "strategy": "perfect_whole_match"
            },
            "found": True
        }
    return {"match": None, "found": False}


@router.get("/tm/{tm_id}/search")
async def search_tm(
    tm_id: int,
    pattern: str,
    limit: int = Query(10, ge=1, le=100),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Search a Translation Memory using LIKE pattern.

    For fuzzy/similar text searching (not exact match).
    """
    logger.info(f"TM search: tm_id={tm_id}, pattern={pattern[:30]}...")

    # Verify TM ownership (async)
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Async LIKE search
    entries_result = await db.execute(
        select(LDMTMEntry).where(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.source_text.ilike(f"%{pattern}%")
        ).limit(limit)
    )
    entries = entries_result.scalars().all()

    results = [
        {
            "source_text": e.source_text,
            "target_text": e.target_text,
            "similarity": 0.0,  # LIKE doesn't provide similarity
            "tier": 0,
            "strategy": "like_search"
        }
        for e in entries
    ]
    return {"results": results, "count": len(results)}


@router.post("/tm/{tm_id}/entries")
async def add_tm_entry(
    tm_id: int,
    source_text: str = Form(...),
    target_text: str = Form(...),
    background_tasks: BackgroundTasks = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Add a single entry to a Translation Memory (Adaptive TM).

    Used to add translations as they're created during editing.
    Q-001: Auto-syncs TM indexes after add.
    """
    logger.info(f"Adding TM entry: tm_id={tm_id}, source={source_text[:30]}...")

    # Verify TM ownership (async)
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Run sync bulk_copy in threadpool to avoid blocking event loop
    def _add_entry():
        sync_db = next(get_db())
        try:
            from server.tools.ldm.tm_manager import TMManager
            tm_manager = TMManager(sync_db)
            return tm_manager.add_entry(tm_id, source_text, target_text)
        finally:
            sync_db.close()

    result = await asyncio.to_thread(_add_entry)

    if not result:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    logger.success(f"TM entry added: tm_id={tm_id}, total entries={result['entry_count']}")

    # Q-001: Auto-sync indexes in background
    if background_tasks:
        background_tasks.add_task(_auto_sync_tm_indexes, tm_id, current_user["user_id"])

    return result


@router.get("/tm/{tm_id}/export")
async def export_tm(
    tm_id: int,
    format: str = Query("text", regex="^(text|excel|tmx)$"),
    columns: Optional[str] = Query(None, description="Comma-separated list of columns: source_text,target_text,string_id,created_at"),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Export a Translation Memory in specified format.

    Formats:
    - text: Tab-separated values (TSV)
    - excel: Excel spreadsheet (.xlsx)
    - tmx: Translation Memory eXchange format

    Columns (optional, comma-separated):
    - source_text (required, always included)
    - target_text (required, always included)
    - string_id
    - created_at

    Returns file download.
    """
    from fastapi.responses import Response

    logger.info(f"Exporting TM: tm_id={tm_id}, format={format}, columns={columns}")

    # Verify TM ownership (async)
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    if not tm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Parse columns if provided
    column_list = None
    if columns:
        column_list = [c.strip() for c in columns.split(",") if c.strip()]

    # Run export in threadpool (uses sync ORM)
    def _export_tm():
        sync_db = next(get_db())
        try:
            from server.tools.ldm.tm_manager import TMManager
            tm_manager = TMManager(sync_db)
            return tm_manager.export_tm(tm_id, format=format, columns=column_list)
        finally:
            sync_db.close()

    result = await asyncio.to_thread(_export_tm)

    logger.success(f"TM exported: tm_id={tm_id}, entries={result['entry_count']}, format={format}")

    return Response(
        content=result["content"],
        media_type=result["mime_type"],
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"'
        }
    )


@router.post("/tm/{tm_id}/build-indexes")
async def build_tm_indexes(
    tm_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Build FAISS indexes for a Translation Memory.

    Creates:
    - whole_text_lookup (hash index for Tier 1 exact match)
    - line_lookup (hash index for Tier 3 line match)
    - whole.index (FAISS HNSW for Tier 2 semantic search)
    - line.index (FAISS HNSW for Tier 4 line-by-line search)

    This is required before using the 5-Tier Cascade TM search.
    Building indexes can take several minutes for large TMs (50k+ entries).

    Returns operation_id for progress tracking via WebSocket/TaskManager.
    """
    from server.tools.ldm.tm_indexer import TMIndexer
    from server.utils.progress_tracker import TrackedOperation

    logger.info(f"Building TM indexes: tm_id={tm_id}, user={current_user['user_id']}")

    # Verify TM ownership (async)
    tm_result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id
        )
    )
    tm = tm_result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    if tm.owner_id != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    tm_name = tm.name  # Capture for use in executor

    # Run indexing in threadpool to avoid blocking event loop
    def _build_indexes():
        sync_db = next(get_db())
        try:
            # Create operation for progress tracking
            with TrackedOperation(
                operation_name=f"Build TM Indexes: {tm_name}",
                user_id=current_user["user_id"],
                username=current_user.get("username", "unknown"),
                tool_name="LDM",
                function_name="build_tm_indexes",
                total_steps=4,
                parameters={"tm_id": tm_id}
            ) as tracker:
                # Build indexes with progress callback
                def progress_callback(stage: str, current: int, total: int):
                    progress_pct = (current / total) * 100 if total > 0 else 0
                    tracker.update(progress_pct, stage, current, total)

                indexer = TMIndexer(sync_db)
                result = indexer.build_indexes(tm_id, progress_callback=progress_callback)

                logger.success(f"TM indexes built: tm_id={tm_id}, entries={result['entry_count']}")
                return result
        finally:
            sync_db.close()

    try:
        result = await asyncio.to_thread(_build_indexes)
        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid TM configuration")
    except Exception as e:
        logger.error(f"TM index build failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Index build failed. Check server logs.")


@router.get("/tm/{tm_id}/indexes")
async def get_tm_index_status(
    tm_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get index status for a Translation Memory.

    Returns list of indexes and their status.
    """
    from server.database.models import LDMTMIndex

    # Verify TM ownership
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Get indexes
    result = await db.execute(
        select(LDMTMIndex).where(LDMTMIndex.tm_id == tm_id)
    )
    indexes = result.scalars().all()

    return {
        "tm_id": tm_id,
        "tm_status": tm.status,
        "indexes": [
            {
                "type": idx.index_type,
                "status": idx.status,
                "file_size": idx.file_size,
                "built_at": idx.built_at.isoformat() if idx.built_at else None
            }
            for idx in indexes
        ]
    }


# ============================================================================
# FEAT-004: TM Sync Protocol (2025-12-18)
# ============================================================================

@router.get("/tm/{tm_id}/sync-status")
async def get_tm_sync_status(
    tm_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Check if TM indexes are stale (DB has newer changes than local indexes).

    Returns:
        - is_stale: True if DB was updated after last sync
        - pending_changes: Estimated number of changes (0 if up-to-date)
        - last_synced: When indexes were last synced
        - tm_updated_at: When TM was last modified in DB
    """
    import json
    from pathlib import Path

    # Verify TM ownership
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    # Check local metadata
    data_dir = Path(__file__).parent.parent.parent / "data" / "ldm_tm" / str(tm_id)
    metadata_path = data_dir / "metadata.json"

    last_synced = None
    synced_entry_count = 0

    if metadata_path.exists():
        try:
            with open(metadata_path, 'r') as f:
                metadata = json.load(f)
            last_synced = metadata.get("synced_at")
            synced_entry_count = metadata.get("entry_count", 0)
        except Exception as e:
            logger.warning(f"Failed to read TM metadata: {e}")

    # Get current DB entry count
    from sqlalchemy import func
    result = await db.execute(
        select(func.count()).select_from(LDMTMEntry).where(LDMTMEntry.tm_id == tm_id)
    )
    db_entry_count = result.scalar() or 0

    # Determine if stale
    is_stale = False
    pending_changes = 0

    if last_synced is None:
        # Never synced
        is_stale = True
        pending_changes = db_entry_count
    elif tm.updated_at:
        # Compare timestamps
        from datetime import datetime
        try:
            synced_dt = datetime.fromisoformat(last_synced.replace('Z', '+00:00'))
            if tm.updated_at.replace(tzinfo=None) > synced_dt.replace(tzinfo=None):
                is_stale = True
                # Estimate pending changes (rough: diff in counts + assume some updates)
                pending_changes = max(1, abs(db_entry_count - synced_entry_count))
        except Exception:
            is_stale = True
            pending_changes = db_entry_count

    return {
        "tm_id": tm_id,
        "is_stale": is_stale,
        "pending_changes": pending_changes,
        "last_synced": last_synced,
        "tm_updated_at": tm.updated_at.isoformat() if tm.updated_at else None,
        "db_entry_count": db_entry_count,
        "synced_entry_count": synced_entry_count
    }


# Q-001: Auto-sync helper for background task
def _auto_sync_tm_indexes(tm_id: int, user_id: int):
    """
    Background task to auto-sync TM indexes after entry modifications.
    Model2Vec is fast (~29k sentences/sec), so this runs quickly.

    BUG-032/BUG-034: Now also updates TM status to 'ready' after successful sync.
    """
    from server.tools.ldm.tm_indexer import TMSyncManager
    from datetime import datetime

    sync_db = next(get_db())
    try:
        # Verify TM still belongs to user
        from server.database.models import LDMTranslationMemory
        tm = sync_db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == user_id
        ).first()

        if not tm:
            logger.warning(f"Auto-sync skipped: TM {tm_id} not found or not owned by user {user_id}")
            return

        sync_manager = TMSyncManager(sync_db, tm_id)
        result = sync_manager.sync()

        # BUG-032/BUG-034: Update TM status to 'ready' after successful sync
        tm.status = "ready"
        tm.updated_at = datetime.utcnow()
        sync_db.commit()

        logger.info(
            f"Auto-sync TM {tm_id}: INSERT={result['stats']['insert']}, "
            f"UPDATE={result['stats']['update']}, time={result['time_seconds']:.2f}s, status=ready"
        )
    except Exception as e:
        logger.error(f"Auto-sync failed for TM {tm_id}: {e}")
    finally:
        sync_db.close()


@router.post("/tm/{tm_id}/sync")
async def sync_tm_indexes(
    tm_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Smart sync TM indexes with DB.

    Uses TMSyncManager for efficient sync:
    - Only re-embeds INSERT/UPDATE entries
    - Copies existing embeddings for UNCHANGED entries
    - Rebuilds FAISS/hash indexes at the end

    Returns:
        Sync results including stats (insert, update, delete, unchanged)
    """
    from server.tools.ldm.tm_indexer import TMSyncManager

    # Verify TM ownership
    result = await db.execute(
        select(LDMTranslationMemory).where(
            LDMTranslationMemory.id == tm_id,
            LDMTranslationMemory.owner_id == current_user["user_id"]
        )
    )
    tm = result.scalar_one_or_none()

    if not tm:
        raise HTTPException(status_code=404, detail="Translation Memory not found")

    logger.info(f"Starting TM sync for TM {tm_id} (user: {current_user['username']})")

    # Run sync in threadpool to avoid blocking
    # BUG-033: Also update TM status after successful sync
    def _sync_tm():
        sync_db = next(get_db())
        try:
            sync_manager = TMSyncManager(sync_db, tm_id)
            result = sync_manager.sync()

            # BUG-033/BUG-034: Update TM status to 'ready' after successful sync
            from server.database.models import LDMTranslationMemory
            tm_record = sync_db.query(LDMTranslationMemory).filter(
                LDMTranslationMemory.id == tm_id
            ).first()
            if tm_record:
                tm_record.status = "ready"
                tm_record.updated_at = datetime.utcnow()
                sync_db.commit()

            return result
        finally:
            sync_db.close()

    try:
        result = await asyncio.to_thread(_sync_tm)

        logger.success(
            f"TM {tm_id} sync complete: "
            f"INSERT={result['stats']['insert']}, UPDATE={result['stats']['update']}, "
            f"UNCHANGED={result['stats']['unchanged']}, time={result['time_seconds']}s, status=ready"
        )

        return result

    except Exception as e:
        logger.error(f"TM sync failed for TM {tm_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"TM sync failed: {str(e)}")


# ============================================================================
# Pretranslation
# ============================================================================

@router.post("/pretranslate", response_model=PretranslateResponse)
async def pretranslate_file(
    request: PretranslateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Pretranslate a file using selected engine.

    Engines:
    - standard: TM 5-tier cascade (hash + FAISS + ngram)
    - xls_transfer: XLS Transfer logic with code preservation
    - kr_similar: KR Similar logic with structure adaptation

    All engines use LOCAL processing (Qwen embeddings + FAISS).
    No external translation API required.

    Args:
        file_id: LDM file ID to pretranslate
        engine: "standard" | "xls_transfer" | "kr_similar"
        dictionary_id: TM ID (for standard) or dictionary ID (for xls/kr)
        threshold: Similarity threshold (default 0.92)
        skip_existing: Skip rows that already have translations (default True)

    Returns:
        PretranslateResponse with stats: matched, skipped, total, time_seconds
    """
    from server.tools.ldm.pretranslate import pretranslate_file as do_pretranslate

    logger.info(f"Pretranslate request: file_id={request.file_id}, engine={request.engine}, "
                f"dict_id={request.dictionary_id}, threshold={request.threshold}")

    # Validate engine
    valid_engines = ["standard", "xls_transfer", "kr_similar"]
    if request.engine not in valid_engines:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid engine. Must be one of: {valid_engines}"
        )

    # Verify file ownership
    file_result = await db.execute(
        select(LDMFile).where(LDMFile.id == request.file_id)
    )
    file = file_result.scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Verify project ownership
    project_result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == file.project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    project = project_result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=403, detail="Access denied to file's project")

    # Run pretranslation in threadpool to avoid blocking
    # TASK-001: Add TrackedOperation for progress tracking
    file_name = file.name  # Capture for use in thread
    user_id = current_user["user_id"]
    username = current_user.get("username", "unknown")

    def _do_pretranslate():
        from server.utils.progress_tracker import TrackedOperation

        sync_db = next(get_db())
        try:
            with TrackedOperation(
                operation_name=f"Pretranslate: {file_name}",
                user_id=user_id,
                username=username,
                tool_name="LDM",
                function_name="pretranslate",
                parameters={
                    "file_id": request.file_id,
                    "engine": request.engine,
                    "dictionary_id": request.dictionary_id,
                    "threshold": request.threshold
                }
            ) as tracker:
                def progress_callback(current: int, total: int):
                    progress_pct = (current / total) * 100 if total > 0 else 0
                    tracker.update(progress_pct, f"Row {current}/{total}", current, total)

                result = do_pretranslate(
                    db=sync_db,
                    file_id=request.file_id,
                    engine=request.engine,
                    dictionary_id=request.dictionary_id,
                    threshold=request.threshold,
                    skip_existing=request.skip_existing,
                    progress_callback=progress_callback
                )
                return result
        finally:
            sync_db.close()

    try:
        result = await asyncio.to_thread(_do_pretranslate)

        return PretranslateResponse(
            file_id=result["file_id"],
            engine=result["engine"],
            matched=result["matched"],
            skipped=result["skipped"],
            total=result["total"],
            threshold=result["threshold"],
            time_seconds=result["time_seconds"]
        )

    except ValueError as e:
        logger.error(f"Pretranslation validation error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Pretranslation failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Pretranslation failed. Check server logs.")


# ============================================================================
# File to TM Conversion Endpoint
# ============================================================================

class FileToTMRequest(BaseModel):
    name: str
    project_id: Optional[int] = None
    language: str = "en"
    description: Optional[str] = None


@router.post("/files/{file_id}/register-as-tm")
async def register_file_as_tm(
    file_id: int,
    request: FileToTMRequest,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Convert an LDM file into a Translation Memory.

    Takes all source/target pairs from the file and creates a new TM.
    """
    # Get file info
    result = await db.execute(
        select(LDMFile).where(LDMFile.id == file_id)
    )
    file = result.scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    logger.info(f"Converting file to TM: file_id={file_id}, name={request.name}")

    try:
        # Get all rows from the file
        result = await db.execute(
            select(LDMRow).where(
                LDMRow.file_id == file_id,
                LDMRow.source.isnot(None),
                LDMRow.source != "",
                LDMRow.target.isnot(None),
                LDMRow.target != ""
            ).order_by(LDMRow.row_num)
        )
        rows = result.scalars().all()

        if not rows:
            raise HTTPException(status_code=400, detail="File has no translatable rows")

        # Create TM entries data
        entries_data = [
            {"source": row.source, "target": row.target}
            for row in rows
        ]

        # Run sync TMManager in threadpool to avoid blocking event loop
        def _create_tm():
            sync_db = next(get_db())
            try:
                from server.tools.ldm.tm_manager import TMManager
                tm_manager = TMManager(sync_db)

                # Create TM
                tm = tm_manager.create_tm(
                    name=request.name,
                    owner_id=current_user["user_id"],
                    source_lang="ko",  # Default, can be extended
                    target_lang=request.language,
                    description=request.description or f"Created from file: {file.name}"
                )

                # Add entries in bulk
                import time
                start_time = time.time()
                entry_count = tm_manager.add_entries_bulk(tm.id, entries_data)
                elapsed = time.time() - start_time

                return {
                    "tm_id": tm.id,
                    "name": tm.name,
                    "entry_count": entry_count,
                    "status": tm.status,
                    "time_seconds": round(elapsed, 2),
                    "rate_per_second": int(entry_count / elapsed) if elapsed > 0 else 0,
                    "source_file": file.name
                }
            finally:
                sync_db.close()

        result = await asyncio.to_thread(_create_tm)

        logger.info(f"TM created from file: tm_id={result['tm_id']}, entries={result['entry_count']}")
        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid file format or data")
    except Exception as e:
        logger.error(f"File to TM conversion failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Conversion failed. Check server logs.")


# ============================================================================
# File Download/Export Endpoints
# ============================================================================

@router.get("/files/{file_id}/download")
async def download_file(
    file_id: int,
    status_filter: Optional[str] = Query(None, description="Filter by status: reviewed, translated, all"),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Download a file with current translations.

    Rebuilds the file from database rows in the original format.
    Uses extra_data for FULL structure preservation (extra columns, attributes, etc.)

    Query params:
    - status_filter: "reviewed" (confirmed only), "translated" (translated+reviewed), "all" (everything)
    """
    # Get file info
    result = await db.execute(
        select(LDMFile).where(LDMFile.id == file_id)
    )
    file = result.scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Get file-level metadata for reconstruction
    file_metadata = file.extra_data or {}

    # Get all rows for this file
    status_conditions = []
    if status_filter == "reviewed":
        # Only confirmed/reviewed strings
        status_conditions = ["reviewed", "approved"]
    elif status_filter == "translated":
        # Translated and reviewed
        status_conditions = ["translated", "reviewed", "approved"]
    # else: all rows

    query = select(LDMRow).where(LDMRow.file_id == file_id).order_by(LDMRow.row_num)

    if status_conditions:
        query = query.where(LDMRow.status.in_(status_conditions))

    result = await db.execute(query)
    rows = result.scalars().all()

    if not rows:
        raise HTTPException(status_code=404, detail="No rows found for this file")

    # Build file content based on format (case-insensitive)
    # Pass file_metadata for full structure preservation
    format_lower = file.format.lower() if file.format else ""
    if format_lower == "txt":
        content = _build_txt_file(rows, file_metadata)
        media_type = "text/plain"
        extension = ".txt"
    elif format_lower == "xml":
        content = _build_xml_file(rows, file_metadata)
        media_type = "application/xml"
        extension = ".xml"
    elif format_lower in ["xlsx", "excel"]:
        content = _build_excel_file(rows, file_metadata)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = ".xlsx"
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {file.format}")

    # Create filename
    base_name = file.original_filename.rsplit('.', 1)[0] if '.' in file.original_filename else file.original_filename
    download_name = f"{base_name}_export{extension}"

    logger.info(f"LDM: Downloading file {file_id} ({len(rows)} rows) as {download_name}")

    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={download_name}"}
    )


def _build_txt_file(rows: List[LDMRow], file_metadata: dict = None) -> bytes:
    """
    Rebuild TXT file from rows with FULL structure preservation.

    TXT format: tab-separated columns
    Original format: idx0\tidx1\tidx2\tidx3\tidx4\tsource\ttarget\t[extra columns...]

    Uses:
    - file_metadata.encoding: Original file encoding
    - file_metadata.total_columns: Total column count
    - row.extra_data: Extra columns beyond 0-6
    """
    file_metadata = file_metadata or {}
    lines = []

    for row in rows:
        # Reconstruct string_id parts - stored with spaces, output with tabs
        # Upload stores as "0 100 0 0 1" (space-joined), we need "0\t100\t0\t0\t1"
        string_id_parts = row.string_id.split(' ') if row.string_id else [""] * 5

        # Ensure we have 5 parts for the index
        while len(string_id_parts) < 5:
            string_id_parts.append("")

        # Build base columns: idx0\tidx1\tidx2\tidx3\tidx4\tsource\ttarget
        source = row.source or ""
        target = row.target or ""

        # Replace newlines back to actual newlines in the file
        source = source.replace("↵", "\n")
        target = target.replace("↵", "\n")

        # Start with standard 7 columns
        parts = string_id_parts[:5] + [source, target]

        # Add extra columns from extra_data (col7, col8, etc.)
        if row.extra_data:
            # Get the total columns from file metadata, or calculate from extra_data
            total_cols = file_metadata.get("total_columns", 7)
            for i in range(7, total_cols):
                col_key = f"col{i}"
                parts.append(row.extra_data.get(col_key, ""))

        line = "\t".join(parts)
        lines.append(line)

    content = "\n".join(lines)

    # Use original encoding if available
    encoding = file_metadata.get("encoding", "utf-8")
    return content.encode(encoding)


def _build_xml_file(rows: List[LDMRow], file_metadata: dict = None) -> bytes:
    """
    Rebuild XML file from rows with FULL structure preservation.

    XML format:
    <?xml version="1.0" encoding="UTF-8"?>
    <RootElement rootAttr="...">
        <ElementTag stringid="ID" strorigin="source" str="target" otherAttr="..."/>
    </RootElement>

    Uses:
    - file_metadata.root_element: Original root element tag name
    - file_metadata.root_attributes: Original root element attributes
    - file_metadata.element_tag: Original localization element tag name
    - row.extra_data: Extra attributes beyond stringid/strorigin/str
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    file_metadata = file_metadata or {}

    # Use original root element or default to LangData
    root_tag = file_metadata.get("root_element", "LangData")
    root = ET.Element(root_tag)

    # Add original root attributes if preserved
    root_attribs = file_metadata.get("root_attributes")
    if root_attribs:
        for key, val in root_attribs.items():
            root.set(key, val)

    # Use original element tag or default to String
    element_tag = file_metadata.get("element_tag", "String")

    for row in rows:
        string_elem = ET.SubElement(root, element_tag)

        # Set core attributes
        string_elem.set("stringid", row.string_id or "")
        string_elem.set("strorigin", row.source or "")
        string_elem.set("str", row.target or "")

        # Add extra attributes from extra_data
        if row.extra_data:
            for attr_name, attr_val in row.extra_data.items():
                string_elem.set(attr_name, attr_val or "")

    # Pretty print with original encoding
    encoding = file_metadata.get("encoding", "UTF-8")
    xml_str = ET.tostring(root, encoding='unicode')
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ", encoding=encoding)

    return pretty_xml


def _build_excel_file(rows: List[LDMRow], file_metadata: dict = None) -> bytes:
    """
    Rebuild Excel file from rows with FULL structure preservation.

    Excel format: Source in column A, Target in column B, extra columns in C+

    Uses:
    - file_metadata.headers: Original column headers
    - file_metadata.sheet_name: Original sheet name
    - row.extra_data: Extra columns beyond A-B
    """
    import openpyxl

    file_metadata = file_metadata or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = file_metadata.get("sheet_name", "Translations")

    # Get headers from metadata or use defaults
    headers = file_metadata.get("headers", ["Source", "Target"])
    if len(headers) < 2:
        headers = ["Source", "Target"]

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row.source or "")
        ws.cell(row=row_idx, column=2, value=row.target or "")

        # Add extra columns from extra_data (C, D, E, etc.)
        if row.extra_data:
            for col_letter, val in row.extra_data.items():
                # Convert column letter to index (C=3, D=4, etc.)
                if col_letter.isalpha() and len(col_letter) == 1:
                    col_num = ord(col_letter.upper()) - ord('A') + 1
                    if col_num > 2:  # Skip A and B (source/target)
                        ws.cell(row=row_idx, column=col_num, value=val or "")

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


# ============================================================================
# Sync to Central Server (Offline → Online)
# ============================================================================

class SyncFileToCentralRequest(BaseModel):
    """Request body for syncing a local file to central PostgreSQL."""
    file_id: int  # File ID in local SQLite
    destination_project_id: int  # Project ID in central PostgreSQL


class SyncFileToCentralResponse(BaseModel):
    """Response for sync operation."""
    success: bool
    new_file_id: Optional[int] = None
    rows_synced: int = 0
    message: str


class SyncTMToCentralRequest(BaseModel):
    """Request body for syncing a local TM to central PostgreSQL."""
    tm_id: int  # TM ID in local SQLite


class SyncTMToCentralResponse(BaseModel):
    """Response for TM sync operation."""
    success: bool
    new_tm_id: Optional[int] = None
    entries_synced: int = 0
    message: str


@router.post("/sync-to-central", response_model=SyncFileToCentralResponse)
async def sync_file_to_central(
    request: SyncFileToCentralRequest,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Sync a file from local SQLite to central PostgreSQL.

    This endpoint:
    1. Reads file metadata + all rows from local SQLite
    2. Creates new file record in PostgreSQL (destination project)
    3. Bulk inserts all rows to PostgreSQL
    4. Returns the new file_id in central DB

    Use this when:
    - User worked offline (SQLite mode)
    - User reconnected (went online)
    - User wants to upload local work to central server

    The file data is passed as JSON (not re-read from disk).
    """
    from server.config import ACTIVE_DATABASE_TYPE
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session
    import os

    logger.info(f"Sync to central: file_id={request.file_id}, dest_project={request.destination_project_id}")

    # Verify we're online (connected to PostgreSQL)
    if ACTIVE_DATABASE_TYPE != "postgresql":
        raise HTTPException(
            status_code=400,
            detail="Cannot sync to central server while in offline mode. Connect to server first."
        )

    # Verify destination project exists and user has access
    project_result = await db.execute(
        select(LDMProject).where(
            LDMProject.id == request.destination_project_id,
            LDMProject.owner_id == current_user["user_id"]
        )
    )
    if not project_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Destination project not found or access denied")

    # Read from local SQLite database
    # The SQLite file is at server/data/locanext.db
    sqlite_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "locanext.db")

    if not os.path.exists(sqlite_path):
        raise HTTPException(
            status_code=400,
            detail="No local database found. You may not have worked offline."
        )

    try:
        # Create SQLite engine and session
        sqlite_engine = create_engine(f"sqlite:///{sqlite_path}")
        sqlite_session = Session(sqlite_engine)

        try:
            # Read the file from SQLite
            local_file = sqlite_session.query(LDMFile).filter(LDMFile.id == request.file_id).first()

            if not local_file:
                raise HTTPException(status_code=404, detail="File not found in local database")

            # Read all rows for this file from SQLite
            local_rows = sqlite_session.query(LDMRow).filter(
                LDMRow.file_id == request.file_id
            ).order_by(LDMRow.row_num).all()

            logger.info(f"Read {len(local_rows)} rows from local SQLite for file {request.file_id}")

            # Create new file in PostgreSQL
            new_file = LDMFile(
                project_id=request.destination_project_id,
                folder_id=None,  # Goes to project root
                name=local_file.name,
                original_filename=local_file.original_filename,
                format=local_file.format,
                row_count=len(local_rows),
                source_language=local_file.source_language,
                target_language=local_file.target_language,
                extra_data=local_file.extra_data,
                created_by=current_user["user_id"]
            )
            db.add(new_file)
            await db.flush()

            # Create rows in PostgreSQL
            for local_row in local_rows:
                new_row = LDMRow(
                    file_id=new_file.id,
                    row_num=local_row.row_num,
                    string_id=local_row.string_id,
                    source=local_row.source,
                    target=local_row.target,
                    status=local_row.status,
                    extra_data=local_row.extra_data
                )
                db.add(new_row)

            await db.commit()

            logger.success(f"Synced file to central: local_id={request.file_id} → central_id={new_file.id}, rows={len(local_rows)}")

            return SyncFileToCentralResponse(
                success=True,
                new_file_id=new_file.id,
                rows_synced=len(local_rows),
                message=f"Successfully synced {len(local_rows)} rows to central server"
            )

        finally:
            sqlite_session.close()
            sqlite_engine.dispose()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Sync to central failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Sync failed. Check server logs.")


@router.post("/tm/sync-to-central", response_model=SyncTMToCentralResponse)
async def sync_tm_to_central(
    request: SyncTMToCentralRequest,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Sync a Translation Memory from local SQLite to central PostgreSQL.

    This endpoint:
    1. Reads TM metadata + all entries from local SQLite
    2. Creates new TM record in PostgreSQL
    3. Bulk inserts all entries to PostgreSQL
    4. Returns the new tm_id in central DB
    """
    from server.config import ACTIVE_DATABASE_TYPE
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session
    import os

    logger.info(f"Sync TM to central: tm_id={request.tm_id}")

    # Verify we're online (connected to PostgreSQL)
    if ACTIVE_DATABASE_TYPE != "postgresql":
        raise HTTPException(
            status_code=400,
            detail="Cannot sync to central server while in offline mode. Connect to server first."
        )

    # Read from local SQLite database
    sqlite_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "locanext.db")

    if not os.path.exists(sqlite_path):
        raise HTTPException(
            status_code=400,
            detail="No local database found. You may not have worked offline."
        )

    try:
        # Create SQLite engine and session
        sqlite_engine = create_engine(f"sqlite:///{sqlite_path}")
        sqlite_session = Session(sqlite_engine)

        try:
            # Read the TM from SQLite
            local_tm = sqlite_session.query(LDMTranslationMemory).filter(
                LDMTranslationMemory.id == request.tm_id
            ).first()

            if not local_tm:
                raise HTTPException(status_code=404, detail="Translation Memory not found in local database")

            # Read all entries for this TM from SQLite
            local_entries = sqlite_session.query(LDMTMEntry).filter(
                LDMTMEntry.tm_id == request.tm_id
            ).all()

            logger.info(f"Read {len(local_entries)} entries from local SQLite for TM {request.tm_id}")

            # Create new TM in PostgreSQL
            new_tm = LDMTranslationMemory(
                name=local_tm.name,
                description=local_tm.description,
                owner_id=current_user["user_id"],
                source_lang=local_tm.source_lang,
                target_lang=local_tm.target_lang,
                entry_count=len(local_entries),
                status="pending"  # Will need re-indexing on server
            )
            db.add(new_tm)
            await db.flush()

            # Bulk insert entries to PostgreSQL
            for local_entry in local_entries:
                new_entry = LDMTMEntry(
                    tm_id=new_tm.id,
                    source_text=local_entry.source_text,
                    target_text=local_entry.target_text,
                    source_hash=local_entry.source_hash,
                    created_by=local_entry.created_by,
                    change_date=local_entry.change_date
                )
                db.add(new_entry)

            await db.commit()

            logger.success(f"Synced TM to central: local_id={request.tm_id} → central_id={new_tm.id}, entries={len(local_entries)}")

            return SyncTMToCentralResponse(
                success=True,
                new_tm_id=new_tm.id,
                entries_synced=len(local_entries),
                message=f"Successfully synced {len(local_entries)} TM entries to central server. Run 'Build Indexes' on the server to enable semantic search."
            )

        finally:
            sqlite_session.close()
            sqlite_engine.dispose()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"TM sync to central failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="TM sync failed. Check server logs.")


# ============================================================================
# Embedding Engine Settings (FEAT-005)
# ============================================================================

class EmbeddingEngineInfo(BaseModel):
    """Info about an available embedding engine."""
    id: str
    name: str
    description: str
    dimension: int
    memory_mb: int
    default: bool


class EmbeddingEngineResponse(BaseModel):
    """Response for current engine."""
    current_engine: str
    engine_name: str


class SetEngineRequest(BaseModel):
    """Request to change embedding engine."""
    engine: str


@router.get("/settings/embedding-engines", response_model=List[EmbeddingEngineInfo])
async def list_embedding_engines(
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    List available embedding engines.

    Returns info about each engine for UI display.
    """
    from server.tools.shared import get_available_engines
    return get_available_engines()


@router.get("/settings/embedding-engine", response_model=EmbeddingEngineResponse)
async def get_current_embedding_engine(
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get the currently selected embedding engine.
    """
    from server.tools.shared import get_current_engine_name, get_embedding_engine

    engine_name = get_current_engine_name()
    engine = get_embedding_engine(engine_name)

    return EmbeddingEngineResponse(
        current_engine=engine_name,
        engine_name=engine.name
    )


@router.post("/settings/embedding-engine", response_model=EmbeddingEngineResponse)
async def set_embedding_engine(
    request: SetEngineRequest,
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Change the embedding engine.

    Options:
    - "model2vec": Fast (79x), lightweight, good for real-time (DEFAULT)
    - "qwen": Deep semantic, heavy, good for batch/quality work

    Note: Changing engine requires re-indexing TMs for best results.
    Existing indexes will still work but may have dimension mismatches.
    """
    from server.tools.shared import set_current_engine, get_embedding_engine, get_current_engine_name

    try:
        set_current_engine(request.engine)
        engine = get_embedding_engine(request.engine)

        logger.info(f"User {current_user['username']} switched embedding engine to: {request.engine}")

        return EmbeddingEngineResponse(
            current_engine=get_current_engine_name(),
            engine_name=engine.name
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
