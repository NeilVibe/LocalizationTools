"""
LDM Translation Memory Manager

Handles TM upload, parsing, and management.
REUSES existing file handlers for TXT and XML parsing.

Supported formats:
- TXT/TSV: Column 5 = Source, Column 6 = Target (via txt_handler.py)
- XML: StrOrigin = Source, Str = Target (via xml_handler.py)
- Excel: Column A (0) = Source, Column B (1) = Target (simple parser)

Uses bulk_insert_tm_entries() for high-performance import (20k+ entries/sec).
"""

import os
from pathlib import Path
from typing import List, Dict, Optional, Callable, BinaryIO
from datetime import datetime

from loguru import logger
from sqlalchemy.orm import Session

# Reuse existing handlers
from server.tools.ldm.file_handlers.txt_handler import parse_txt_file
from server.tools.ldm.file_handlers.xml_handler import parse_xml_file

# Database utilities - COPY TEXT for 3-5x faster uploads (P21)
from server.database.db_utils import bulk_copy_tm_entries, bulk_insert_tm_entries
from server.database.models import LDMTranslationMemory, LDMTMEntry


class TMManager:
    """
    Translation Memory Manager for LDM.

    Handles:
    - TM file upload and parsing (TXT, XML, Excel)
    - Bulk import with progress tracking
    - TM CRUD operations
    """

    def __init__(self, db: Session):
        self.db = db

    # =========================================================================
    # TM Upload and Parsing
    # =========================================================================

    def upload_tm(
        self,
        file_content: bytes,
        filename: str,
        name: str,
        owner_id: int,
        source_lang: str = "ko",
        target_lang: str = "en",
        description: Optional[str] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
        mode: str = "standard",
        source_col: int = 0,
        target_col: int = 1,
        stringid_col: Optional[int] = None,
        has_header: bool = True
    ) -> Dict:
        """
        Upload and parse a TM file.

        Args:
            file_content: Raw file bytes
            filename: Original filename (determines parser)
            name: Display name for the TM
            owner_id: User ID of TM owner
            source_lang: Source language code (default: ko)
            target_lang: Target language code (default: en)
            description: Optional TM description
            progress_callback: Optional callback for progress updates
            mode: "standard" (duplicates merged) or "stringid" (all variations)
            source_col: Source column index for Excel (default: 0 = A)
            target_col: Target column index for Excel (default: 1 = B)
            stringid_col: StringID column index for Excel (None = 2-column mode)
            has_header: Whether Excel has header row

        Returns:
            Dict with tm_id, entry_count, status, time_seconds
        """
        start_time = datetime.now()

        # Determine file type and parse
        file_ext = Path(filename).suffix.lower()

        logger.info(f"Uploading TM: {name} ({filename}) - {len(file_content):,} bytes, mode={mode}")

        try:
            # Parse file based on type
            if file_ext in ['.txt', '.tsv']:
                entries = self._parse_txt_for_tm(file_content, filename)
            elif file_ext == '.xml':
                entries = self._parse_xml_for_tm(file_content, filename)
            elif file_ext in ['.xlsx', '.xls']:
                entries = self._parse_excel_for_tm(
                    file_content, filename,
                    source_col=source_col,
                    target_col=target_col,
                    stringid_col=stringid_col,
                    has_header=has_header,
                    mode=mode
                )
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")

            if not entries:
                raise ValueError("No valid entries found in file")

            logger.info(f"Parsed {len(entries):,} TM entries from {filename}")

            # Create TM record
            tm = LDMTranslationMemory(
                name=name,
                description=description,
                owner_id=owner_id,
                source_lang=source_lang,
                target_lang=target_lang,
                entry_count=0,  # Will be updated after insert
                status="indexing",
                mode=mode
            )
            self.db.add(tm)
            self.db.commit()

            logger.info(f"Created TM record: id={tm.id}")

            # Bulk insert entries using COPY TEXT (3-5x faster for PostgreSQL)
            inserted = bulk_copy_tm_entries(
                self.db,
                tm.id,
                entries,
                progress_callback=progress_callback
            )

            # Update TM record
            tm.entry_count = inserted
            tm.status = "ready"
            self.db.commit()

            elapsed = (datetime.now() - start_time).total_seconds()
            rate = inserted / elapsed if elapsed > 0 else 0

            logger.info(f"TM import complete: {inserted:,} entries in {elapsed:.2f}s ({rate:,.0f}/sec)")

            return {
                "tm_id": tm.id,
                "name": tm.name,
                "entry_count": inserted,
                "status": tm.status,
                "time_seconds": round(elapsed, 2),
                "rate_per_second": round(rate)
            }

        except Exception as e:
            logger.error(f"TM upload failed: {e}")
            raise

    # =========================================================================
    # Parsers - Convert to TM Entry Format
    # =========================================================================

    def _parse_txt_for_tm(self, file_content: bytes, filename: str) -> List[Dict]:
        """
        Parse TXT file for TM entries.
        REUSES txt_handler.py (Column 5 = Source, Column 6 = Target).
        """
        rows = parse_txt_file(file_content, filename)

        # Convert to TM entry format
        entries = []
        for row in rows:
            source = row.get("source")
            target = row.get("target")

            # TM entries need both source and target
            if source and target:
                entries.append({
                    "source_text": source,
                    "target_text": target
                })

        return entries

    def _parse_xml_for_tm(self, file_content: bytes, filename: str) -> List[Dict]:
        """
        Parse XML file for TM entries.
        REUSES xml_handler.py (StrOrigin = Source, Str = Target).
        """
        rows = parse_xml_file(file_content, filename)

        # Convert to TM entry format
        entries = []
        for row in rows:
            source = row.get("source")
            target = row.get("target")

            # TM entries need both source and target
            if source and target:
                entries.append({
                    "source_text": source,
                    "target_text": target
                })

        return entries

    def _parse_excel_for_tm(
        self,
        file_content: bytes,
        filename: str,
        source_col: int = 0,
        target_col: int = 1,
        stringid_col: Optional[int] = None,
        has_header: bool = True,
        mode: str = "standard"
    ) -> List[Dict]:
        """
        Parse Excel file for TM entries.

        Supports two modes:
        - standard: Source + Target (duplicates merged to most frequent)
        - stringid: Source + Target + StringID (all variations kept)

        Args:
            file_content: Raw Excel bytes
            filename: For logging
            source_col: Source column index (default: 0 = Column A)
            target_col: Target column index (default: 1 = Column B)
            stringid_col: StringID column index (None = 2-column mode)
            has_header: Whether first row is header (skip it)
            mode: "standard" or "stringid"
        """
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active

            entries = []
            required_cols = max(source_col, target_col)
            if stringid_col is not None:
                required_cols = max(required_cols, stringid_col)

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                # Skip header row if specified
                if row_idx == 1 and has_header:
                    continue

                if not row or len(row) <= required_cols:
                    continue

                source = str(row[source_col] or "").strip()
                target = str(row[target_col] or "").strip()

                # TM entries need both source and target
                if not source or not target:
                    continue

                entry = {
                    "source_text": source,
                    "target_text": target
                }

                # Add StringID if in stringid mode
                if mode == "stringid" and stringid_col is not None:
                    string_id = str(row[stringid_col] or "").strip() if len(row) > stringid_col else None
                    entry["string_id"] = string_id if string_id else None

                entries.append(entry)

            wb.close()
            logger.info(f"Parsed Excel {filename}: {len(entries)} TM entries, mode={mode}")
            return entries

        except ImportError:
            logger.error("openpyxl not installed - cannot parse Excel files")
            raise ValueError("Excel support requires openpyxl package")
        except Exception as e:
            logger.error(f"Failed to parse Excel file {filename}: {e}")
            raise

    # =========================================================================
    # TM CRUD Operations
    # =========================================================================

    def list_tms(self, owner_id: Optional[int] = None) -> List[Dict]:
        """
        List all Translation Memories.

        Args:
            owner_id: Filter by owner (optional)

        Returns:
            List of TM info dicts
        """
        query = self.db.query(LDMTranslationMemory)

        if owner_id:
            query = query.filter(LDMTranslationMemory.owner_id == owner_id)

        tms = query.order_by(LDMTranslationMemory.created_at.desc()).all()

        return [
            {
                "id": tm.id,
                "name": tm.name,
                "description": tm.description,
                "source_lang": tm.source_lang,
                "target_lang": tm.target_lang,
                "entry_count": tm.entry_count,
                "status": tm.status,
                "created_at": tm.created_at.isoformat() if tm.created_at else None
            }
            for tm in tms
        ]

    def get_tm(self, tm_id: int) -> Optional[Dict]:
        """Get a single TM by ID."""
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()

        if not tm:
            return None

        return {
            "id": tm.id,
            "name": tm.name,
            "description": tm.description,
            "source_lang": tm.source_lang,
            "target_lang": tm.target_lang,
            "entry_count": tm.entry_count,
            "status": tm.status,
            "created_at": tm.created_at.isoformat() if tm.created_at else None
        }

    def delete_tm(self, tm_id: int) -> bool:
        """
        Delete a Translation Memory and all its entries.

        Args:
            tm_id: TM ID to delete

        Returns:
            True if deleted, False if not found
        """
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()

        if not tm:
            return False

        # Entries are deleted via CASCADE
        self.db.delete(tm)
        self.db.commit()

        logger.info(f"Deleted TM: id={tm_id} name={tm.name}")
        return True

    def add_entry(
        self,
        tm_id: int,
        source_text: str,
        target_text: str
    ) -> Optional[Dict]:
        """
        Add a single entry to a TM (for Adaptive TM feature).

        Args:
            tm_id: TM ID
            source_text: Source text
            target_text: Target text

        Returns:
            Entry info dict or None if TM not found
        """
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()

        if not tm:
            return None

        # Use COPY TEXT for consistency (handles hash generation)
        entries = [{"source_text": source_text, "target_text": target_text}]
        inserted = bulk_copy_tm_entries(self.db, tm_id, entries)

        if inserted > 0:
            # Update entry count
            tm.entry_count = (tm.entry_count or 0) + 1
            self.db.commit()

            return {
                "tm_id": tm_id,
                "source_text": source_text,
                "target_text": target_text,
                "entry_count": tm.entry_count
            }

        return None

    def create_tm(
        self,
        name: str,
        owner_id: int,
        source_lang: str = "ko",
        target_lang: str = "en",
        description: Optional[str] = None,
        mode: str = "standard"
    ) -> LDMTranslationMemory:
        """
        Create a new empty TM record.

        Args:
            name: TM display name
            owner_id: User ID of TM owner
            source_lang: Source language code
            target_lang: Target language code
            description: Optional description
            mode: "standard" (duplicates merged) or "stringid" (all variations kept)

        Returns:
            Created TM model object
        """
        tm = LDMTranslationMemory(
            name=name,
            description=description,
            owner_id=owner_id,
            source_lang=source_lang,
            target_lang=target_lang,
            entry_count=0,
            status="pending",
            mode=mode
        )
        self.db.add(tm)
        self.db.commit()

        logger.info(f"Created TM record: id={tm.id} name={name}")
        return tm

    def add_entries_bulk(
        self,
        tm_id: int,
        entries: List[Dict],
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> int:
        """
        Add multiple entries to a TM in bulk.

        Args:
            tm_id: TM ID to add entries to
            entries: List of dicts with 'source' and 'target' keys
            progress_callback: Optional callback for progress updates

        Returns:
            Number of entries inserted
        """
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()

        if not tm:
            raise ValueError(f"TM not found: id={tm_id}")

        # Convert to expected format (source_text, target_text, string_id)
        formatted_entries = []
        for e in entries:
            source = e.get("source") or e.get("source_text")
            target = e.get("target") or e.get("target_text")
            if source and target:
                entry = {
                    "source_text": source,
                    "target_text": target
                }
                # Include string_id if present (for stringid mode)
                if "string_id" in e:
                    entry["string_id"] = e.get("string_id")
                formatted_entries.append(entry)

        if not formatted_entries:
            logger.warning(f"No valid entries to add to TM {tm_id}")
            return 0

        # Use COPY TEXT for bulk insert
        inserted = bulk_copy_tm_entries(
            self.db,
            tm_id,
            formatted_entries,
            progress_callback=progress_callback
        )

        # Update TM record
        tm.entry_count = (tm.entry_count or 0) + inserted
        tm.status = "ready"
        self.db.commit()

        logger.info(f"Bulk added {inserted} entries to TM {tm_id}")
        return inserted

    # =========================================================================
    # TM Search (Basic - Full 5-tier cascade in tm_search.py)
    # =========================================================================

    def search_exact(self, tm_id: int, source_text: str) -> Optional[Dict]:
        """
        Search for exact match in TM using hash.
        O(1) lookup using source_hash index.

        Args:
            tm_id: TM ID to search
            source_text: Source text to find

        Returns:
            Match dict or None
        """
        import hashlib
        from server.database.db_utils import normalize_text_for_hash

        # Generate hash for lookup
        normalized = normalize_text_for_hash(source_text)
        source_hash = hashlib.sha256(normalized.encode('utf-8')).hexdigest()

        entry = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.source_hash == source_hash
        ).first()

        if entry:
            return {
                "source_text": entry.source_text,
                "target_text": entry.target_text,
                "similarity": 1.0,
                "tier": 1,
                "strategy": "perfect_whole_match"
            }

        return None

    def search_like(
        self,
        tm_id: int,
        pattern: str,
        limit: int = 10
    ) -> List[Dict]:
        """
        Search TM using LIKE pattern (for quick suggestions).

        Args:
            tm_id: TM ID to search
            pattern: Search pattern (will be wrapped with %)
            limit: Max results

        Returns:
            List of matches
        """
        entries = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == tm_id,
            LDMTMEntry.source_text.ilike(f"%{pattern}%")
        ).limit(limit).all()

        return [
            {
                "source_text": e.source_text,
                "target_text": e.target_text,
                "similarity": 0.0,  # LIKE doesn't provide similarity
                "tier": 0,
                "strategy": "like_search"
            }
            for e in entries
        ]

    # =========================================================================
    # TM Viewer - Paginated Entries (FEAT-003)
    # =========================================================================

    def get_entries_paginated(
        self,
        tm_id: int,
        page: int = 1,
        limit: int = 100,
        sort_by: str = "id",
        sort_order: str = "asc",
        search: Optional[str] = None,
        metadata_field: str = "string_id"
    ) -> Dict:
        """
        Get paginated TM entries with sorting and search.

        Args:
            tm_id: TM ID
            page: Page number (1-indexed)
            limit: Items per page (max 500)
            sort_by: Field to sort by (id, source_text, target_text, string_id, created_at)
            sort_order: Sort order (asc, desc)
            search: Optional search term (searches source and target)
            metadata_field: Which metadata field to include (string_id, created_at, etc.)

        Returns:
            Dict with entries, total, page, limit, total_pages
        """
        from sqlalchemy import func, or_

        # Validate and cap limit
        limit = min(limit, 500)
        page = max(page, 1)
        offset = (page - 1) * limit

        # Base query
        query = self.db.query(LDMTMEntry).filter(LDMTMEntry.tm_id == tm_id)

        # Apply search filter
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    LDMTMEntry.source_text.ilike(search_pattern),
                    LDMTMEntry.target_text.ilike(search_pattern),
                    LDMTMEntry.string_id.ilike(search_pattern)
                )
            )

        # Get total count (before pagination)
        total = query.count()

        # Apply sorting (BUG-020: added is_confirmed, updated_at, confirmed_at)
        sort_column = {
            "id": LDMTMEntry.id,
            "source_text": LDMTMEntry.source_text,
            "target_text": LDMTMEntry.target_text,
            "string_id": LDMTMEntry.string_id,
            "created_at": LDMTMEntry.created_at,
            "is_confirmed": LDMTMEntry.is_confirmed,
            "updated_at": LDMTMEntry.updated_at,
            "confirmed_at": LDMTMEntry.confirmed_at
        }.get(sort_by, LDMTMEntry.id)

        if sort_order.lower() == "desc":
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())

        # Apply pagination
        entries = query.offset(offset).limit(limit).all()

        # Format entries (BUG-020: include confirmation metadata)
        formatted_entries = []
        for e in entries:
            entry_data = {
                "id": e.id,
                "source_text": e.source_text,
                "target_text": e.target_text,
                "string_id": e.string_id,
                "created_at": e.created_at.isoformat() if e.created_at else None,
                "created_by": e.created_by,
                "updated_at": e.updated_at.isoformat() if e.updated_at else None,
                "updated_by": e.updated_by,
                "is_confirmed": e.is_confirmed,
                "confirmed_at": e.confirmed_at.isoformat() if e.confirmed_at else None,
                "confirmed_by": e.confirmed_by
            }
            formatted_entries.append(entry_data)

        total_pages = (total + limit - 1) // limit  # Ceiling division

        return {
            "entries": formatted_entries,
            "total": total,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
            "sort_by": sort_by,
            "sort_order": sort_order,
            "search": search
        }

    def update_entry(
        self,
        tm_id: int,
        entry_id: int,
        source_text: Optional[str] = None,
        target_text: Optional[str] = None,
        string_id: Optional[str] = None,
        updated_by: Optional[str] = None
    ) -> Optional[Dict]:
        """
        Update a single TM entry (for inline editing).

        Args:
            tm_id: TM ID (for validation)
            entry_id: Entry ID to update
            source_text: New source text (optional)
            target_text: New target text (optional)
            string_id: New string ID (optional)
            updated_by: Username of user making the update (optional)

        Returns:
            Updated entry dict or None if not found
        """
        entry = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        ).first()

        if not entry:
            return None

        # Update fields if provided
        if source_text is not None:
            entry.source_text = source_text
            # Recalculate hash
            import hashlib
            from server.database.db_utils import normalize_text_for_hash
            normalized = normalize_text_for_hash(source_text)
            entry.source_hash = hashlib.sha256(normalized.encode('utf-8')).hexdigest()

        if target_text is not None:
            entry.target_text = target_text

        if string_id is not None:
            entry.string_id = string_id

        # BUG-020: Track who made the update
        entry.updated_at = datetime.utcnow()
        if updated_by:
            entry.updated_by = updated_by

        # Mark TM as updated (triggers index rebuild on next pretranslate)
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()
        if tm:
            tm.updated_at = datetime.utcnow()

        self.db.commit()

        logger.info(f"Updated TM entry: tm_id={tm_id}, entry_id={entry_id}, by={updated_by}")

        return {
            "id": entry.id,
            "source_text": entry.source_text,
            "target_text": entry.target_text,
            "string_id": entry.string_id,
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
            "updated_by": entry.updated_by,
            "is_confirmed": entry.is_confirmed,
            "confirmed_at": entry.confirmed_at.isoformat() if entry.confirmed_at else None,
            "confirmed_by": entry.confirmed_by
        }

    def delete_entry(self, tm_id: int, entry_id: int) -> bool:
        """
        Delete a single TM entry.

        Args:
            tm_id: TM ID (for validation)
            entry_id: Entry ID to delete

        Returns:
            True if deleted, False if not found
        """
        entry = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        ).first()

        if not entry:
            return False

        self.db.delete(entry)

        # Update TM entry count and updated_at
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()
        if tm:
            tm.entry_count = max(0, (tm.entry_count or 0) - 1)
            tm.updated_at = datetime.utcnow()

        self.db.commit()

        logger.info(f"Deleted TM entry: tm_id={tm_id}, entry_id={entry_id}")
        return True

    def confirm_entry(
        self,
        tm_id: int,
        entry_id: int,
        confirmed_by: str,
        confirm: bool = True
    ) -> Optional[Dict]:
        """
        BUG-020: Confirm or unconfirm a TM entry (memoQ-style workflow).

        When user approves a translation, it gets marked as confirmed with metadata.

        Args:
            tm_id: TM ID (for validation)
            entry_id: Entry ID to confirm
            confirmed_by: Username of user confirming
            confirm: True to confirm, False to unconfirm

        Returns:
            Updated entry dict or None if not found
        """
        entry = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.id == entry_id,
            LDMTMEntry.tm_id == tm_id
        ).first()

        if not entry:
            return None

        if confirm:
            entry.is_confirmed = True
            entry.confirmed_at = datetime.utcnow()
            entry.confirmed_by = confirmed_by
        else:
            entry.is_confirmed = False
            entry.confirmed_at = None
            entry.confirmed_by = None

        self.db.commit()

        logger.info(f"{'Confirmed' if confirm else 'Unconfirmed'} TM entry: tm_id={tm_id}, entry_id={entry_id}, by={confirmed_by}")

        return {
            "id": entry.id,
            "source_text": entry.source_text,
            "target_text": entry.target_text,
            "string_id": entry.string_id,
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
            "updated_by": entry.updated_by,
            "is_confirmed": entry.is_confirmed,
            "confirmed_at": entry.confirmed_at.isoformat() if entry.confirmed_at else None,
            "confirmed_by": entry.confirmed_by
        }

    def bulk_confirm_entries(
        self,
        tm_id: int,
        entry_ids: List[int],
        confirmed_by: str,
        confirm: bool = True
    ) -> Dict:
        """
        BUG-020: Bulk confirm/unconfirm multiple TM entries.

        Args:
            tm_id: TM ID (for validation)
            entry_ids: List of entry IDs to confirm
            confirmed_by: Username of user confirming
            confirm: True to confirm, False to unconfirm

        Returns:
            Dict with count of updated entries
        """
        now = datetime.utcnow()

        if confirm:
            updated = self.db.query(LDMTMEntry).filter(
                LDMTMEntry.tm_id == tm_id,
                LDMTMEntry.id.in_(entry_ids)
            ).update({
                LDMTMEntry.is_confirmed: True,
                LDMTMEntry.confirmed_at: now,
                LDMTMEntry.confirmed_by: confirmed_by
            }, synchronize_session=False)
        else:
            updated = self.db.query(LDMTMEntry).filter(
                LDMTMEntry.tm_id == tm_id,
                LDMTMEntry.id.in_(entry_ids)
            ).update({
                LDMTMEntry.is_confirmed: False,
                LDMTMEntry.confirmed_at: None,
                LDMTMEntry.confirmed_by: None
            }, synchronize_session=False)

        self.db.commit()

        logger.info(f"Bulk {'confirmed' if confirm else 'unconfirmed'} {updated} TM entries: tm_id={tm_id}, by={confirmed_by}")

        return {
            "updated_count": updated,
            "action": "confirmed" if confirm else "unconfirmed"
        }

    # =========================================================================
    # TM Export (FEAT-002)
    # =========================================================================

    def export_tm(
        self,
        tm_id: int,
        format: str = "text",
        columns: Optional[List[str]] = None
    ) -> Dict:
        """
        Export TM entries in specified format.

        Args:
            tm_id: TM ID to export
            format: Export format - "text" (TSV), "excel" (.xlsx), or "tmx"
            columns: List of columns to include (default: source_text, target_text)
                     Available: source_text, target_text, string_id, created_at, created_by

        Returns:
            Dict with content (bytes), filename, and mime_type
        """
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == tm_id
        ).first()

        if not tm:
            raise ValueError(f"TM not found: id={tm_id}")

        # Default columns if not specified
        if not columns:
            columns = ["source_text", "target_text"]

        # Always include source and target
        if "source_text" not in columns:
            columns.insert(0, "source_text")
        if "target_text" not in columns:
            columns.insert(1, "target_text")

        # Get all entries
        entries = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == tm_id
        ).order_by(LDMTMEntry.id).all()

        logger.info(f"Exporting TM {tm_id} ({tm.name}): {len(entries)} entries, format={format}")

        if format == "text":
            return self._export_text(tm, entries, columns)
        elif format == "excel":
            return self._export_excel(tm, entries, columns)
        elif format == "tmx":
            return self._export_tmx(tm, entries)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    def _export_text(
        self,
        tm: LDMTranslationMemory,
        entries: List[LDMTMEntry],
        columns: List[str]
    ) -> Dict:
        """Export TM as TSV (tab-separated values)."""
        lines = []

        # Header row
        header_map = {
            "source_text": "Source",
            "target_text": "Target",
            "string_id": "StringID",
            "created_at": "Created At",
            "created_by": "Created By"
        }
        headers = [header_map.get(col, col) for col in columns]
        lines.append("\t".join(headers))

        # Data rows
        for entry in entries:
            row = []
            for col in columns:
                value = getattr(entry, col, "")
                if col == "created_at" and value:
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                row.append(str(value or ""))
            lines.append("\t".join(row))

        content = "\n".join(lines).encode("utf-8")

        # Clean filename
        safe_name = "".join(c for c in tm.name if c.isalnum() or c in "._- ")
        filename = f"{safe_name}_export.txt"

        return {
            "content": content,
            "filename": filename,
            "mime_type": "text/tab-separated-values",
            "entry_count": len(entries)
        }

    def _export_excel(
        self,
        tm: LDMTranslationMemory,
        entries: List[LDMTMEntry],
        columns: List[str]
    ) -> Dict:
        """Export TM as Excel (.xlsx)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TM Export"

            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, color="FFFFFF")

            # Column headers
            header_map = {
                "source_text": "Source",
                "target_text": "Target",
                "string_id": "StringID",
                "created_at": "Created At",
                "created_by": "Created By"
            }

            for col_idx, col in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header_map.get(col, col)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, entry in enumerate(entries, start=2):
                for col_idx, col in enumerate(columns, start=1):
                    value = getattr(entry, col, "")
                    if col == "created_at" and value:
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_idx, column=col_idx).value = str(value or "")

            # Auto-width columns
            for col_idx, col in enumerate(columns, start=1):
                max_length = len(header_map.get(col, col))
                for row in range(2, min(102, len(entries) + 2)):  # Sample first 100 rows
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)[:50]))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            content = output.getvalue()
            output.close()

            # Clean filename
            safe_name = "".join(c for c in tm.name if c.isalnum() or c in "._- ")
            filename = f"{safe_name}_export.xlsx"

            return {
                "content": content,
                "filename": filename,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "entry_count": len(entries)
            }

        except ImportError:
            logger.error("openpyxl not installed - cannot export Excel")
            raise ValueError("Excel export requires openpyxl package")

    def _export_tmx(
        self,
        tm: LDMTranslationMemory,
        entries: List[LDMTMEntry]
    ) -> Dict:
        """Export TM as TMX (Translation Memory eXchange) format."""
        import xml.etree.ElementTree as ET
        from datetime import datetime

        # Create root element
        root = ET.Element("tmx", version="1.4")

        # Header
        header = ET.SubElement(root, "header")
        header.set("creationtool", "LocaNext")
        header.set("creationtoolversion", "1.0")
        header.set("datatype", "plaintext")
        header.set("segtype", "sentence")
        header.set("adminlang", "en")
        header.set("srclang", tm.source_lang or "ko")
        header.set("creationdate", datetime.utcnow().strftime("%Y%m%dT%H%M%SZ"))

        # Body
        body = ET.SubElement(root, "body")

        for entry in entries:
            tu = ET.SubElement(body, "tu")

            # Add StringID as property if available
            if entry.string_id:
                prop = ET.SubElement(tu, "prop", type="x-string-id")
                prop.text = entry.string_id

            # Source segment
            tuv_src = ET.SubElement(tu, "tuv")
            tuv_src.set("{http://www.w3.org/XML/1998/namespace}lang", tm.source_lang or "ko")
            seg_src = ET.SubElement(tuv_src, "seg")
            seg_src.text = entry.source_text

            # Target segment
            tuv_tgt = ET.SubElement(tu, "tuv")
            tuv_tgt.set("{http://www.w3.org/XML/1998/namespace}lang", tm.target_lang or "en")
            seg_tgt = ET.SubElement(tuv_tgt, "seg")
            seg_tgt.text = entry.target_text

        # Convert to string with XML declaration
        tree = ET.ElementTree(root)
        from io import BytesIO
        output = BytesIO()
        tree.write(output, encoding="utf-8", xml_declaration=True)
        content = output.getvalue()
        output.close()

        # Clean filename
        safe_name = "".join(c for c in tm.name if c.isalnum() or c in "._- ")
        filename = f"{safe_name}_export.tmx"

        return {
            "content": content,
            "filename": filename,
            "mime_type": "application/x-tmx+xml",
            "entry_count": len(entries)
        }
