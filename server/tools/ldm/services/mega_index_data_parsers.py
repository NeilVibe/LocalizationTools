"""
MegaIndex Data Parsers Mixin - Phase 1, 3, and 5 parsing methods.

Handles: DDS texture scanning, WEM audio scanning, knowledge parsing,
localization (strorigin, translations), export events/loc, and devmemo scanning.

Extracted from mega_index.py during ARCH-02 decomposition.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Set

from loguru import logger

from server.tools.ldm.services.mega_index_helpers import (
    _get_export_key,
    _get_stringid,
    _normalize_strorigin,
    _safe_parse_xml,
)
from server.tools.ldm.services.mega_index_schemas import (
    KnowledgeEntry,
    KnowledgeGroupNode,
)


def _ci_attrs(elem) -> dict:
    """Case-insensitive XML attribute extraction — lowercase BOTH keys and values used as IDs."""
    return {k.lower(): v for k, v in elem.attrib.items()}


class DataParsersMixin:
    """Mixin providing Phase 1 (foundation), Phase 3 (localization), and Phase 5 (broad scan) parsers."""

    # Type hints for self.* dicts (populated by MegaIndex.__init__)
    knowledge_by_strkey: Dict[str, KnowledgeEntry]
    dds_by_stem: Dict[str, Path]
    wem_by_event: Dict[str, Path]
    wem_by_event_en: Dict[str, Path]
    wem_by_event_kr: Dict[str, Path]
    wem_by_event_zh: Dict[str, Path]
    knowledge_group_hierarchy: Dict[str, KnowledgeGroupNode]
    event_to_stringid: Dict[str, str]
    stringid_to_strorigin: Dict[str, str]
    stringid_to_translations: Dict[str, Dict[str, str]]
    export_file_stringids: Dict[str, Set[str]]
    ordered_export_index: Dict[str, Dict[str, List[str]]]
    event_to_export_path: Dict[str, str]
    event_to_xml_order: Dict[str, int]
    strkey_to_devmemo: Dict[str, str]

    # =========================================================================
    # Phase 1: Foundation Parsers
    # =========================================================================

    def _scan_dds_textures(self, texture_folder: Path) -> None:
        """D9: Scan texture folder for DDS files.

        GRAFTED from MDG DDSIndex.scan_folder() — dual-key strategy:
        Key 1 = stem.lower() (e.g. "character_varon")
        Key 2 = full filename.lower() (e.g. "character_varon.dds")
        First-seen-wins dedup per key slot.
        """
        try:
            if not texture_folder.is_dir():
                logger.warning(f"[MEGAINDEX] Texture folder not found: {texture_folder}")
                return
            count = 0
            for dds_path in sorted(texture_folder.rglob("*.dds")):
                # MDG dual-key: index by BOTH stem and full filename
                stem_lower = dds_path.stem.lower()
                name_lower = dds_path.name.lower()
                if stem_lower not in self.dds_by_stem:
                    self.dds_by_stem[stem_lower] = dds_path
                if name_lower not in self.dds_by_stem:
                    self.dds_by_stem[name_lower] = dds_path
                count += 1
            logger.debug(f"[MEGAINDEX] DDS scan: {count} files, {len(self.dds_by_stem)} index keys from {texture_folder}")
        except Exception as e:
            logger.warning(f"[MEGAINDEX] DDS scan failed: {e}")

    def _scan_wem_files(self, audio_folder: Path) -> None:
        """D10: Legacy single-folder scan. Use _scan_wem_files_all_languages for multi-lang."""
        self._scan_wem_into(audio_folder, self.wem_by_event, "EN")

    def _scan_wem_files_all_languages(self, audio_en: Path, audio_kr: Path, audio_zh: Path) -> None:
        """D10a/b/c: Scan 3 audio folders for WEM files, one per language."""
        self._scan_wem_into(audio_en, self.wem_by_event_en, "EN")
        self._scan_wem_into(audio_kr, self.wem_by_event_kr, "KR")
        self._scan_wem_into(audio_zh, self.wem_by_event_zh, "ZH")

    def _scan_wem_into(self, audio_folder: Path, target: Dict[str, Path], label: str) -> None:
        """Scan a single audio folder into the target dict. Key=stem.lower(), Value=Path."""
        try:
            if not audio_folder.is_dir():
                logger.debug(f"[MEGAINDEX] Audio folder ({label}) not found: {audio_folder}")
                return
            count = 0
            dupes = 0
            for wem_path in audio_folder.rglob("*.wem"):
                stem_lower = wem_path.stem.lower()
                if stem_lower in target:
                    dupes += 1
                    continue  # MDG pattern: first-seen-wins dedup
                target[stem_lower] = wem_path
                count += 1
            logger.debug(f"[MEGAINDEX] WEM scan ({label}): {count} files indexed ({dupes} dupes) from {audio_folder}")
        except Exception as e:
            logger.exception(f"[MEGAINDEX] WEM scan ({label}) failed: {e}")

    def _parse_knowledge_info(self, knowledge_folder: Path) -> None:
        """D1+D15: Parse KnowledgeInfo and KnowledgeGroupInfo in single pass."""
        try:
            if not knowledge_folder.is_dir():
                logger.warning(
                    f"[MEGAINDEX] Knowledge folder not found: {knowledge_folder}"
                )
                return
            xml_files = list(knowledge_folder.rglob("*.xml"))
            logger.debug(f"[MEGAINDEX] Knowledge XMLs found: {len(xml_files)}")
            parse_ok = 0
            parse_fail = 0
            for xml_path in xml_files:
                root = _safe_parse_xml(xml_path)
                if root is None:
                    parse_fail += 1
                    logger.debug(f"[MEGAINDEX] Knowledge XML parse failed: {xml_path.name}")
                    continue
                parse_ok += 1
                source_file = xml_path.name

                # D1: KnowledgeInfo entries
                # GRAFTED from MDG: "best value wins" dedup — if existing entry
                # has UITextureName and new one doesn't, KEEP the existing entry.
                for elem in root.iter("KnowledgeInfo"):
                    a = _ci_attrs(elem)
                    strkey = (a.get("strkey") or "").lower()
                    if not strkey:
                        continue
                    ui_texture = a.get("uitexturename") or ""
                    existing = self.knowledge_by_strkey.get(strkey)
                    if existing and existing.ui_texture_name and not ui_texture:
                        logger.debug(f"[MEGAINDEX] Keeping existing UITexture for {strkey}: '{existing.ui_texture_name}' (skipping empty)")
                        continue
                    self.knowledge_by_strkey[strkey] = KnowledgeEntry(
                        strkey=strkey,
                        name=a.get("name") or "",
                        desc=a.get("desc") or "",
                        ui_texture_name=ui_texture,
                        group_key=(a.get("knowledgegroupkey")
                        or a.get("groupkey")
                        or "").lower(),
                        source_file=source_file,
                    )

                # D15: KnowledgeGroupInfo entries
                for elem in root.iter("KnowledgeGroupInfo"):
                    a = _ci_attrs(elem)
                    strkey = (a.get("strkey") or "").lower()
                    if not strkey:
                        continue
                    child_strkeys: List[str] = []
                    for child in elem.iter("KnowledgeInfo"):
                        ca = _ci_attrs(child)
                        csk = (ca.get("strkey") or "").lower()
                        if csk:
                            child_strkeys.append(csk)
                    self.knowledge_group_hierarchy[strkey] = KnowledgeGroupNode(
                        strkey=strkey,
                        group_name=a.get("groupname") or a.get("name") or "",
                        child_strkeys=tuple(child_strkeys),
                    )
            logger.debug(f"[MEGAINDEX] Knowledge parse: {parse_ok} OK, {parse_fail} failed out of {len(xml_files)} XMLs")
        except Exception as e:
            logger.exception(f"[MEGAINDEX] Knowledge parse failed: {e}")

    # =========================================================================
    # Phase 3: Localization Parsers
    # =========================================================================

    def _parse_loc_strorigin(self, loc_folder: Path) -> None:
        """D12: Parse languagedata_KOR.xml for StringId -> StrOrigin mapping."""
        try:
            if not loc_folder.is_dir():
                logger.warning(f"[MEGAINDEX] Loc folder not found: {loc_folder}")
                return
            # Look for Korean language data file
            kor_files = list(loc_folder.glob("**/languagedata_[Kk][Oo][Rr]*.xml"))
            if not kor_files:
                # Try broader pattern
                kor_files = [
                    f for f in loc_folder.rglob("*.xml")
                    if "kor" in f.stem.lower() and "languagedata" in f.stem.lower()
                ]
            logger.debug(f"[MEGAINDEX] Korean loc files found: {len(kor_files)} — {[f.name for f in kor_files]}")
            for kor_path in kor_files:
                root = _safe_parse_xml(kor_path)
                if root is None:
                    continue
                for elem in root.iter("LocStr"):
                    sid = _get_stringid(elem)
                    if not sid:
                        continue
                    a = _ci_attrs(elem)
                    strorigin = a.get("strorigin") or ""
                    if strorigin:
                        self.stringid_to_strorigin[sid] = strorigin
        except Exception as e:
            logger.exception(f"[MEGAINDEX] StrOrigin parse failed: {e}")

    def _parse_loc_translations(
        self, loc_folder: Path, langs: List[str]
    ) -> None:
        """D13: Parse languagedata_{code}.xml for each preloaded language."""
        try:
            if not loc_folder.is_dir():
                logger.warning(f"[MEGAINDEX] Loc folder not found for translations: {loc_folder}")
                return
            langs_found: Dict[str, str] = {}  # lang -> filename
            langs_skipped: List[str] = []
            for xml_path in loc_folder.rglob("*.xml"):
                stem = xml_path.stem.lower()
                if not stem.startswith("languagedata_"):
                    continue
                # Extract language code
                lang = stem.split("_", 1)[1]
                if lang == "kor":
                    continue  # Korean is strorigin, not translation
                if lang not in langs:
                    if lang not in langs_skipped:
                        langs_skipped.append(lang)
                    continue
                langs_found[lang] = xml_path.name
                root = _safe_parse_xml(xml_path)
                if root is None:
                    continue
                for elem in root.iter("LocStr"):
                    sid = _get_stringid(elem)
                    if not sid:
                        continue
                    a = _ci_attrs(elem)
                    text = a.get("str") or ""
                    if text:
                        if sid not in self.stringid_to_translations:
                            self.stringid_to_translations[sid] = {}
                        self.stringid_to_translations[sid][lang] = text
            logger.debug(f"[MEGAINDEX] Translation parse: loaded {list(langs_found.keys())}, skipped {langs_skipped}")
        except Exception as e:
            logger.exception(f"[MEGAINDEX] Translation parse failed: {e}")

    def _parse_export_events(self, export_folder: Path) -> None:
        """D11+D20+D21: Parse export XMLs for event->stringid mapping."""
        try:
            if not export_folder.is_dir():
                logger.warning(
                    f"[MEGAINDEX] Export folder not found: {export_folder}"
                )
                return
            global_order = 0
            export_xmls = sorted(export_folder.rglob("*.xml"))
            export_data_count = sum(1 for f in export_xmls if not f.name.endswith(".loc.xml"))
            export_loc_count = sum(1 for f in export_xmls if f.name.endswith(".loc.xml"))
            logger.debug(f"[MEGAINDEX] Export folder: {len(export_xmls)} XMLs total ({export_data_count} data, {export_loc_count} .loc.xml)")
            for xml_path in export_xmls:
                # Skip .loc.xml files (handled by _parse_export_loc)
                if xml_path.name.endswith(".loc.xml"):
                    continue
                root = _safe_parse_xml(xml_path)
                if root is None:
                    continue

                # Relative path from export root
                try:
                    rel_dir = str(xml_path.relative_to(export_folder).parent)
                except ValueError:
                    rel_dir = ""

                for elem in root.iter():
                    a = _ci_attrs(elem)
                    event_name = (
                        a.get("soundeventname")
                        or a.get("eventname")
                        or ""
                    ).strip().lower()
                    sid = (a.get("stringid") or "").strip().lower()
                    if event_name and sid:
                        self.event_to_stringid[event_name] = sid  # D11
                        self.event_to_export_path[event_name] = rel_dir  # D20
                        self.event_to_xml_order[event_name] = global_order  # D21
                        global_order += 1
        except Exception as e:
            logger.exception(f"[MEGAINDEX] Export events parse failed: {e}")

    def _parse_export_loc(self, export_folder: Path) -> None:
        """D17+D18+D11: Parse export .loc.xml files for StringId sets, ordered index, and audio links.

        CRITICAL FIX: Uses RELATIVE PATH (not just filename) as D17 key so that
        categorize_by_export_path() can detect Sequencer/Dialog from the path prefix.
        Also extracts SoundEventName → D11 event_to_stringid for audio linking.
        """
        try:
            if not export_folder.is_dir():
                logger.debug("[MEGAINDEX] Export folder not found for .loc.xml: %s", export_folder)
                return
            loc_files = sorted(export_folder.rglob("*.loc.xml"))
            logger.info(f"[MEGAINDEX] Export .loc.xml files found: {len(loc_files)}")
            parsed_ok = 0
            audio_links = 0
            for xml_path in loc_files:
                root = _safe_parse_xml(xml_path)
                if root is None:
                    continue

                # Use RELATIVE PATH as key (preserves Sequencer/Dialog/ prefix for categorization)
                try:
                    rel_path = str(xml_path.relative_to(export_folder))
                    # Normalize: forward slashes, remove .loc.xml extension
                    filename_key = rel_path.replace("\\", "/").lower()
                    if filename_key.endswith(".loc.xml"):
                        filename_key = filename_key[:-8]  # strip .loc.xml
                    elif filename_key.endswith(".xml"):
                        filename_key = filename_key[:-4]
                except ValueError:
                    filename_key = _get_export_key(xml_path.name)
                    logger.warning(
                        "[MEGAINDEX] Export .loc.xml not relative to export folder, "
                        "category detection may be wrong: %s", xml_path,
                    )

                # Also compute the relative directory for D20 (export_path category)
                try:
                    rel_dir = str(xml_path.relative_to(export_folder).parent)
                    if rel_dir == ".":
                        rel_dir = ""
                except ValueError:
                    rel_dir = ""

                sids: Set[str] = set()
                kor_map: Dict[str, List[str]] = {}

                for elem in root.iter("LocStr"):
                    sid = _get_stringid(elem)
                    a = _ci_attrs(elem)
                    origin = a.get("strorigin") or ""
                    if sid:
                        sids.add(sid)
                    if origin and sid:
                        norm = _normalize_strorigin(origin)
                        if norm:
                            kor_map.setdefault(norm, []).append(sid)

                    # Extract SoundEventName → D11 audio linking (CRITICAL FIX)
                    if sid:
                        sound_event = (a.get("soundeventname") or "").strip().lower()
                        if sound_event and sound_event not in self.event_to_stringid:
                            self.event_to_stringid[sound_event] = sid  # D11
                            self.event_to_export_path[sound_event] = rel_dir  # D20
                            audio_links += 1

                if sids:
                    if filename_key in self.export_file_stringids:
                        self.export_file_stringids[filename_key].update(sids)
                    else:
                        self.export_file_stringids[filename_key] = sids

                if kor_map:
                    if filename_key in self.ordered_export_index:
                        existing = self.ordered_export_index[filename_key]
                        for norm_text, sid_list in kor_map.items():
                            existing.setdefault(norm_text, []).extend(sid_list)
                    else:
                        self.ordered_export_index[filename_key] = kor_map
                parsed_ok += 1
            logger.info(
                f"[MEGAINDEX] Export .loc.xml: {parsed_ok}/{len(loc_files)} parsed, "
                f"{len(self.export_file_stringids)} file sets, "
                f"{len(self.ordered_export_index)} ordered indexes, "
                f"{audio_links} SoundEventName→StringID audio links (D11)"
            )
        except Exception as e:
            logger.exception(f"[MEGAINDEX] Export loc parse failed: {e}")

    # =========================================================================
    # Phase 5: Broad Scan
    # =========================================================================

    def _scan_devmemo(self, staticinfo_folder: Path) -> None:
        """D19: Broad scan of StaticInfo for StrKey -> DevMemo/DevComment."""
        try:
            if not staticinfo_folder.is_dir():
                logger.warning(
                    f"[MEGAINDEX] StaticInfo folder not found: {staticinfo_folder}"
                )
                return
            xml_files = list(staticinfo_folder.rglob("*.xml"))
            logger.debug(f"[MEGAINDEX] DevMemo broad scan: {len(xml_files)} XMLs in {staticinfo_folder}")
            for xml_path in xml_files:
                root = _safe_parse_xml(xml_path)
                if root is None:
                    continue
                for elem in root.iter():
                    a = _ci_attrs(elem)
                    strkey = (a.get("strkey") or "").lower()
                    if not strkey:
                        continue
                    korean = a.get("devmemo") or a.get("devcomment") or ""
                    if korean and strkey not in self.strkey_to_devmemo:
                        self.strkey_to_devmemo[strkey] = korean
        except Exception as e:
            logger.warning(f"[MEGAINDEX] DevMemo scan failed: {e}")

    # =========================================================================
    # Phase 4: VRS (VoiceRecordingSheet) chronological reorder
    # =========================================================================

    def _apply_vrs_order(self, vrs_folder: Path) -> None:
        """Reorder D21 (event_to_xml_order) using VoiceRecordingSheet Excel.

        Exact MDG logic: events found in VRS get their row position as xml_order.
        Events NOT in VRS keep their original xml_order offset after VRS entries.
        """
        if not vrs_folder or not vrs_folder.exists():
            logger.info(f"[MEGAINDEX] VRS folder not found: {vrs_folder} — skipping")
            return

        # Find most recent Excel file
        xlsx_files = sorted(vrs_folder.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
        if not xlsx_files:
            logger.info(f"[MEGAINDEX] No Excel files in VRS folder: {vrs_folder}")
            return

        vrs_file = xlsx_files[0]
        logger.info(f"[MEGAINDEX] Loading VRS order from: {vrs_file.name}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("[MEGAINDEX] openpyxl not installed — VRS ordering disabled")
            return

        try:
            wb = load_workbook(vrs_file, data_only=True, read_only=True)
            try:
                ws = wb.active

                # Find EventName column by header (row 1)
                event_col_idx = None
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if header_row:
                    for idx, cell_value in enumerate(header_row):
                        if cell_value and str(cell_value).strip().lower() == "eventname":
                            event_col_idx = idx
                            break

                if event_col_idx is None:
                    logger.warning(f"[MEGAINDEX] 'EventName' column not found in VRS header: {header_row}")
                    return

                # Build event_name -> VRS row position
                vrs_order: Dict[str, int] = {}
                position = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > event_col_idx:
                        event_name = row[event_col_idx]
                        if event_name:
                            event_lower = str(event_name).strip().lower()
                            if event_lower and event_lower not in vrs_order:
                                vrs_order[event_lower] = position
                                position += 1

                logger.info(f"[MEGAINDEX] Loaded {len(vrs_order)} EventNames from VRS")
            finally:
                wb.close()

            if not vrs_order:
                return

            # Apply VRS order: matched events get VRS position,
            # unmatched events get offset to sort after all VRS entries
            vrs_max = len(vrs_order)
            reordered = 0
            for event_name, current_order in list(self.event_to_xml_order.items()):
                vrs_pos = vrs_order.get(event_name)
                if vrs_pos is not None:
                    self.event_to_xml_order[event_name] = vrs_pos
                    reordered += 1
                else:
                    self.event_to_xml_order[event_name] = vrs_max + (current_order or 0)

            logger.info(
                f"[MEGAINDEX] VRS reordered {reordered} / {len(self.event_to_xml_order)} events "
                f"({len(self.event_to_xml_order) - reordered} unmatched appended after)"
            )

        except Exception as e:
            logger.exception(f"[MEGAINDEX] VRS processing failed: {e}")
