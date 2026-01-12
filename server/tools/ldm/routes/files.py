"""
File endpoints - Upload, download, list, preview files.

Migrated from api.py lines 392-658, 2448-2779
P10: DB Abstraction Layer - Uses repositories for FULL PARITY.
"""

import asyncio
from io import BytesIO
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from loguru import logger

from server.utils.dependencies import get_async_db, get_current_active_user_async, get_db
from server.database.models import LDMFile, LDMRow
from server.tools.ldm.schemas import FileResponse, FileToTMRequest
from server.tools.ldm.permissions import can_access_project, can_access_file
from server.repositories import (
    FileRepository, get_file_repository,
    ProjectRepository, get_project_repository,
    FolderRepository, get_folder_repository,
    TMRepository, get_tm_repository
)

router = APIRouter(tags=["LDM"])


# =============================================================================
# File CRUD Endpoints
# =============================================================================

@router.get("/projects/{project_id}/files", response_model=List[FileResponse])
async def list_files(
    project_id: int,
    folder_id: Optional[int] = None,
    repo: FileRepository = Depends(get_file_repository),
    project_repo: ProjectRepository = Depends(get_project_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    List files in a project, optionally filtered by folder.

    P10: Uses Repository Pattern for database operations.
    """
    logger.debug(f"[FILES] list_files: project_id={project_id}, folder_id={folder_id}")

    # P10: Verify project exists using repository (online mode)
    if project_id > 0:
        project = await project_repo.get(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        if not await can_access_project(db, project_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    # Get files via repository
    files = await repo.get_all(project_id=project_id, folder_id=folder_id)

    logger.debug(f"[FILES] list_files complete: project_id={project_id}, files={len(files)}")
    return files


@router.get("/files", response_model=List[FileResponse])
async def list_all_files(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    UI-074 FIX: List all files across all accessible projects.
    Used by ReferenceSettingsModal to show available reference files.

    DB Abstraction: Uses FileRepository for database operations.
    Permission filtering is done at route level (requires project access check).
    """
    from server.tools.ldm.permissions import get_accessible_projects

    # Get all accessible projects for user
    accessible_projects = await get_accessible_projects(db, current_user)
    project_ids = [p.id for p in accessible_projects]

    if not project_ids:
        return []

    # Get files from all accessible projects via repository
    # Note: We need to get files for multiple projects, so we iterate
    all_files = []
    for project_id in project_ids:
        files = await repo.get_all(project_id=project_id, limit=limit, offset=0)
        all_files.extend(files)

    # Sort by updated_at descending and apply pagination
    all_files.sort(key=lambda f: f.get("updated_at") or "", reverse=True)
    return all_files[offset:offset + limit]


@router.get("/files/{file_id}", response_model=FileResponse)
async def get_file(
    file_id: int,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get file metadata by ID.

    P10-REPO: Uses Repository Pattern - automatically selects PostgreSQL or SQLite
    based on user's online/offline mode.
    """
    # Repository handles both PostgreSQL and SQLite
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Permission check (online mode only - offline is single user)
    # Local files have negative IDs, skip permission check
    if file_id > 0:
        if not await can_access_file(db, file_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    return file


@router.post("/files/upload", response_model=FileResponse)
async def upload_file(
    project_id: Optional[int] = Form(None),  # Optional when storage=local
    folder_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    # P9: Storage destination - 'server' (PostgreSQL) or 'local' (SQLite Offline Storage)
    storage: Optional[str] = Form("server"),
    # Excel column mapping (optional, only used for Excel files)
    source_col: Optional[int] = Form(None),      # Column index for source (0=A, 1=B, etc.)
    target_col: Optional[int] = Form(None),      # Column index for target
    stringid_col: Optional[int] = Form(None),    # Column index for StringID (None = 2-column mode)
    has_header: Optional[bool] = Form(True),     # Whether first row is header
    db: AsyncSession = Depends(get_async_db),
    project_repo: ProjectRepository = Depends(get_project_repository),
    folder_repo: FolderRepository = Depends(get_folder_repository),
    file_repo: FileRepository = Depends(get_file_repository),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Upload a localization file (TXT/XML/Excel), parse it, and store rows in database.

    P9: Unified endpoint for both online (PostgreSQL) and offline (SQLite) storage.
    - storage='server' (default): Saves to PostgreSQL, requires project_id
    - storage='local': Saves to SQLite Offline Storage, project_id optional

    P10: Uses repository pattern for verification steps.

    Supported formats:
    - TXT/TSV: Tab-delimited, columns 0-4=StringID, 5=Source(KR), 6=Target, 7+=extra
    - XML: LocStr elements with StringId, StrOrigin(source), Str(target), other attrs=extra
    - Excel: User-defined columns via source_col, target_col, stringid_col parameters
      - 2-column mode: Source + Target (default: A, B)
      - 3-column mode: Source + Target + StringID (e.g., A, B, C)

    Full Structure Preservation:
    - File-level metadata stored in LDMFile.extra_data (encoding, root element, etc.)
    - Row-level extra data stored in LDMRow.extra_data (extra columns/attributes)
    - Enables FULL reconstruction of original file format
    """
    logger.debug(f"[FILES] upload_file: storage={storage}, project_id={project_id}, folder_id={folder_id}")

    # P9: Route to local storage if requested
    storage_type = storage or "server"
    if storage_type == "local":
        return await _upload_to_local_storage(
            file=file,
            source_col=source_col,
            target_col=target_col,
            stringid_col=stringid_col,
            has_header=has_header,
            folder_id=folder_id  # P9-FIX: Pass folder_id for creating files inside local folders
        )

    # Server storage requires project_id
    if not project_id:
        raise HTTPException(status_code=400, detail="project_id is required for server storage")

    # P10: Verify project exists using repository
    project = await project_repo.get(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Check access permission
    if not await can_access_project(db, project_id, current_user):
        raise HTTPException(status_code=404, detail="Resource not found")

    # P10: Verify folder exists using repository (if provided)
    if folder_id:
        folder = await folder_repo.get(folder_id)
        if not folder or folder.get("project_id") != project_id:
            raise HTTPException(status_code=404, detail="Folder not found")

    # P10: DB-002 unique names using repository
    filename = file.filename or "unknown"
    filename = await file_repo.generate_unique_name(
        base_name=filename,
        project_id=project_id,
        folder_id=folder_id
    )

    # Determine file type and parse
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    file_metadata = None
    if ext in ('txt', 'tsv'):
        from server.tools.ldm.file_handlers.txt_handler import parse_txt_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_txt_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext == 'xml':
        from server.tools.ldm.file_handlers.xml_handler import parse_xml_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_xml_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext in ('xlsx', 'xls'):
        from server.tools.ldm.file_handlers.excel_handler import parse_excel_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        # Use provided column mappings or defaults (A=0, B=1)
        rows_data = parse_excel_file(
            file_content,
            filename,
            source_col=source_col if source_col is not None else 0,
            target_col=target_col if target_col is not None else 1,
            stringid_col=stringid_col,  # None means 2-column mode
            has_header=has_header if has_header is not None else True
        )
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {ext}. Use TXT, TSV, XML, or Excel (XLSX/XLS)."
        )

    if not rows_data:
        raise HTTPException(
            status_code=400,
            detail="No valid rows found in file"
        )

    # TASK-001: Add TrackedOperation for progress tracking on large files
    # Run row insertion in thread with progress tracking
    user_id = current_user["user_id"]
    username = current_user.get("username", "unknown")
    total_rows = len(rows_data)

    def _insert_file_and_rows():
        from server.utils.progress_tracker import TrackedOperation

        sync_db = next(get_db())
        try:
            with TrackedOperation(
                operation_name=f"Upload: {filename}",
                user_id=user_id,
                username=username,
                tool_name="LDM",
                function_name="upload_file",
                parameters={"filename": filename, "row_count": total_rows}
            ) as tracker:
                # Create file record
                new_file = LDMFile(
                    project_id=project_id,
                    folder_id=folder_id,
                    name=filename,
                    original_filename=filename,
                    format=file_format,
                    row_count=total_rows,
                    source_language=source_lang,
                    target_language=None,
                    extra_data=file_metadata
                )
                sync_db.add(new_file)
                sync_db.flush()  # Get the file ID

                tracker.update(10, "File record created", 1, total_rows + 1)

                # Create row records with progress
                for i, row_data in enumerate(rows_data):
                    row = LDMRow(
                        file_id=new_file.id,
                        row_num=row_data["row_num"],
                        string_id=row_data["string_id"],
                        source=row_data["source"],
                        target=row_data["target"],
                        status=row_data["status"],
                        extra_data=row_data.get("extra_data")
                    )
                    sync_db.add(row)

                    # Update progress every 100 rows
                    if (i + 1) % 100 == 0 or i == total_rows - 1:
                        progress_pct = 10 + ((i + 1) / total_rows) * 90
                        tracker.update(progress_pct, f"Row {i + 1}/{total_rows}", i + 1, total_rows)

                sync_db.commit()
                sync_db.refresh(new_file)

                return {
                    "id": new_file.id,
                    "project_id": new_file.project_id,
                    "folder_id": new_file.folder_id,
                    "name": new_file.name,
                    "original_filename": new_file.original_filename,
                    "format": new_file.format,
                    "row_count": new_file.row_count,
                    "source_language": new_file.source_language,
                    "target_language": new_file.target_language,
                    "created_at": new_file.created_at,
                    "updated_at": new_file.updated_at
                }
        finally:
            sync_db.close()

    result = await asyncio.to_thread(_insert_file_and_rows)

    logger.success(f"[FILES] File uploaded: id={result['id']}, name='{filename}', rows={total_rows}, has_metadata={file_metadata is not None}")

    # Return FileResponse-compatible dict
    return result


@router.patch("/files/{file_id}/move")
async def move_file_to_folder(
    file_id: int,
    folder_id: Optional[int] = None,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Move a file to a different folder (or root of project if folder_id is None).

    Used for drag-and-drop file organization in FileExplorer.
    P10-REPO: Uses Repository Pattern for database operations.
    """
    # Get file via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Permission check (online mode only - local files have negative IDs)
    if file_id > 0:
        if not await can_access_project(db, file["project_id"], current_user):
            raise HTTPException(status_code=403, detail="Not authorized to modify this file")

    # Move via repository (handles folder validation)
    try:
        updated_file = await repo.move(file_id, folder_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    logger.success(f"[FILES] File moved: id={file_id}, new_folder={folder_id}")

    return {"success": True, "file_id": file_id, "folder_id": folder_id}


@router.patch("/files/{file_id}/move-cross-project")
async def move_file_cross_project(
    file_id: int,
    target_project_id: int,
    target_folder_id: Optional[int] = None,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    EXPLORER-005: Move a file to a different project.
    EXPLORER-009: Requires 'cross_project_move' capability.

    P10-REPO: Uses Repository Pattern for database operations.
    Note: Cross-project move not supported for local files.
    """
    # Get file via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Local files can't be moved cross-project
    if file_id < 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot move local files between projects. Upload to server first."
        )

    # Check access to source project
    if not await can_access_project(db, file["project_id"], current_user):
        raise HTTPException(status_code=404, detail="File not found")

    # Check access to destination project
    if not await can_access_project(db, target_project_id, current_user):
        raise HTTPException(status_code=404, detail="Destination project not found")

    # EXPLORER-009: Check capability for cross-project move
    from ..permissions import require_capability
    await require_capability(db, current_user, "cross_project_move")

    source_project_id = file["project_id"]

    # Move via repository (handles folder validation and auto-rename)
    try:
        updated_file = await repo.move_cross_project(file_id, target_project_id, target_folder_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    logger.success(f"[FILES] File moved cross-project: id={file_id}, from project {source_project_id} to {target_project_id}")
    return {
        "success": True,
        "file_id": file_id,
        "new_name": updated_file["name"],
        "target_project_id": target_project_id,
        "target_folder_id": target_folder_id
    }


@router.post("/files/{file_id}/copy")
async def copy_file(
    file_id: int,
    target_project_id: Optional[int] = None,
    target_folder_id: Optional[int] = None,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Copy a file to a different location.
    EXPLORER-001: Ctrl+C/V file operations.

    If target_project_id is None, copies to same project.
    If target_folder_id is None, copies to project root.
    Auto-renames if duplicate name exists.

    P10-REPO: Uses Repository Pattern for database operations.
    """
    # Get source file via repository
    source_file = await repo.get(file_id)

    if not source_file:
        raise HTTPException(status_code=404, detail="File not found")

    # Permission checks (online mode only)
    if file_id > 0:
        # Check access permission for source
        if not await can_access_project(db, source_file["project_id"], current_user):
            raise HTTPException(status_code=404, detail="File not found")

        # Check access permission for destination
        dest_project_id = target_project_id or source_file["project_id"]
        if target_project_id and target_project_id != source_file["project_id"]:
            if not await can_access_project(db, target_project_id, current_user):
                raise HTTPException(status_code=404, detail="Destination project not found")

    # Copy via repository (handles unique naming and row copying)
    try:
        new_file = await repo.copy(file_id, target_project_id, target_folder_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    logger.success(f"[FILES] File copied: {source_file['name']} -> {new_file['name']}, id={new_file['id']}")
    return {
        "success": True,
        "new_file_id": new_file["id"],
        "name": new_file["name"],
        "row_count": new_file.get("row_count", 0)
    }


@router.patch("/files/{file_id}/rename")
async def rename_file(
    file_id: int,
    name: str,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Rename a file.

    P10-REPO: Uses Repository Pattern for database operations.
    """
    # Get file via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    old_name = file["name"]

    # Permission check (online mode only - local files have negative IDs)
    if file_id > 0:
        if not await can_access_project(db, file["project_id"], current_user):
            raise HTTPException(status_code=403, detail="Not authorized to modify this file")

    # Rename via repository (handles uniqueness check)
    try:
        updated_file = await repo.rename(file_id, name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    logger.success(f"[FILES] File renamed: id={file_id}, '{old_name}' -> '{name}'")
    return {"success": True, "file_id": file_id, "name": updated_file["name"]}


@router.delete("/files/{file_id}")
async def delete_file(
    file_id: int,
    permanent: bool = Query(False, description="If true, permanently delete instead of moving to trash"),
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Delete a file and all its rows.
    EXPLORER-008: By default, moves to trash (soft delete). Use permanent=true for hard delete.

    P10-REPO: Uses Repository Pattern for database operations.
    """
    # Get file via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Permission check (online mode only - local files have negative IDs)
    if file_id > 0:
        if not await can_access_project(db, file["project_id"], current_user):
            raise HTTPException(status_code=403, detail="Not authorized to delete this file")

    file_name = file["name"]

    # Local files (negative ID) - repository handles trash
    if file_id < 0:
        success = await repo.delete(file_id, permanent=permanent)
        if not success:
            raise HTTPException(status_code=500, detail="Failed to delete file")
        action = "permanently deleted" if permanent else "moved to trash"
        logger.info(f"Local file {action}: id={file_id}, name='{file_name}'")
        return {"message": f"File {action}", "file_id": file_id}

    # Server files - use trash serialization (requires SQLAlchemy model)
    if not permanent:
        # EXPLORER-008: Soft delete - move to trash
        from .trash import move_to_trash, serialize_file_for_trash

        # Get SQLAlchemy model for trash serialization
        result = await db.execute(select(LDMFile).where(LDMFile.id == file_id))
        file_model = result.scalar_one_or_none()

        if file_model:
            # Serialize file data for restore
            file_data = await serialize_file_for_trash(db, file_model)

            # Move to trash
            await move_to_trash(
                db,
                item_type="file",
                item_id=file_model.id,
                item_name=file_model.name,
                item_data=file_data,
                parent_project_id=file_model.project_id,
                parent_folder_id=file_model.folder_id,
                deleted_by=current_user["user_id"]
            )

    # Clean up sync subscription if it exists
    try:
        from server.database.offline import get_offline_db
        offline_db = get_offline_db()
        offline_db.remove_subscription("file", file_id)
        logger.debug(f"Cleaned up sync subscription for file {file_id}")
    except Exception as e:
        logger.debug(f"No subscription to clean up for file {file_id}: {e}")

    # Delete via repository
    await repo.delete(file_id, permanent=True)  # Already serialized to trash

    action = "permanently deleted" if permanent else "moved to trash"
    logger.info(f"File {action}: id={file_id}, name='{file_name}'")
    return {"message": f"File {action}", "file_id": file_id}


@router.post("/files/excel-preview")
async def excel_preview(
    file: UploadFile = File(...),
    max_rows: int = Query(5, ge=1, le=20),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Get a preview of an Excel file for column mapping UI.

    Returns:
        - sheet_name: Name of the active sheet
        - headers: Column headers (from first row)
        - sample_rows: First N rows of data
        - total_rows: Total row count

    Use this before uploading to let users select which columns
    are Source, Target, and StringID.
    """
    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    if ext not in ('xlsx', 'xls'):
        raise HTTPException(
            status_code=400,
            detail=f"Excel preview only supports XLSX/XLS files, got: {ext}"
        )

    try:
        from server.tools.ldm.file_handlers.excel_handler import get_excel_preview
        file_content = await file.read()
        preview = get_excel_preview(file_content, max_rows=max_rows)
        return preview
    except Exception as e:
        logger.error(f"[FILES] Excel preview failed: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# =============================================================================
# File to TM Conversion
# =============================================================================

@router.post("/files/{file_id}/register-as-tm")
async def register_file_as_tm(
    file_id: int,
    request: FileToTMRequest,
    repo: FileRepository = Depends(get_file_repository),
    tm_repo: TMRepository = Depends(get_tm_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Convert an LDM file into a Translation Memory.

    DB Abstraction: Uses FileRepository + TMRepository for FULL PARITY.
    - Offline files (negative IDs) → SQLite TM
    - Online files (positive IDs) → PostgreSQL TM

    Takes all source/target pairs from the file and creates a new TM.
    """
    # Get file info via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check access permission (online mode only - local files have negative IDs)
    if file_id > 0:
        if not await can_access_file(db, file_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    file_name = file.get("name") or "unknown"
    is_offline = file_id < 0  # Local files have negative IDs
    logger.info(f"Converting file to TM: file_id={file_id}, name={request.name}, offline={is_offline}")

    try:
        # Get all rows from the file via repository
        rows = await repo.get_rows_for_export(file_id)

        # Filter to only rows with both source and target
        rows = [r for r in rows if r.get("source") and r.get("target")]

        if not rows:
            raise HTTPException(status_code=400, detail="File has no translatable rows")

        # Create TM entries data
        entries_data = [
            {"source": row.get("source"), "target": row.get("target")}
            for row in rows
        ]

        import time
        start_time = time.time()

        if is_offline:
            # P10-FIX: Use SQLite TM repository for offline files
            tm = await tm_repo.create(
                name=request.name,
                source_lang="ko",
                target_lang=request.language,
                owner_id=current_user.get("user_id")
            )

            # Bulk add entries (executemany = instant, like PostgreSQL COPY)
            entry_count = await tm_repo.add_entries_bulk(tm["id"], entries_data)

            # P11-UX: Assign TM to same scope as source file for intuitive experience
            # User registers file from a folder → TM appears in that folder
            from server.repositories.interfaces.tm_repository import AssignmentTarget
            file_folder_id = file.get("folder_id")
            file_project_id = file.get("project_id")
            if file_folder_id:
                await tm_repo.assign(tm["id"], AssignmentTarget(folder_id=file_folder_id))
                logger.info(f"Assigned TM {tm['id']} to folder {file_folder_id}")
            elif file_project_id:
                await tm_repo.assign(tm["id"], AssignmentTarget(project_id=file_project_id))
                logger.info(f"Assigned TM {tm['id']} to project {file_project_id}")

            elapsed = time.time() - start_time
            result = {
                "tm_id": tm["id"],
                "name": tm["name"],
                "entry_count": entry_count,
                "status": tm.get("status", "active"),
                "time_seconds": round(elapsed, 2),
                "rate_per_second": int(entry_count / elapsed) if elapsed > 0 else 0,
                "source_file": file_name
            }
        else:
            # Online mode: Use PostgreSQL TMManager
            file_folder_id = file.get("folder_id")  # Get before thread
            file_project_id = file.get("project_id")

            def _create_tm():
                sync_db = next(get_db())
                try:
                    from server.tools.ldm.tm_manager import TMManager
                    tm_manager = TMManager(sync_db)

                    # Create TM
                    tm = tm_manager.create_tm(
                        name=request.name,
                        owner_id=current_user["user_id"],
                        source_lang="ko",
                        target_lang=request.language,
                        description=request.description or f"Created from file: {file_name}"
                    )

                    # Add entries in bulk
                    entry_count = tm_manager.add_entries_bulk(tm.id, entries_data)
                    elapsed = time.time() - start_time

                    return {
                        "tm_id": tm.id,
                        "name": tm.name,
                        "entry_count": entry_count,
                        "status": tm.status,
                        "time_seconds": round(elapsed, 2),
                        "rate_per_second": int(entry_count / elapsed) if elapsed > 0 else 0,
                        "source_file": file_name
                    }
                finally:
                    sync_db.close()

            result = await asyncio.to_thread(_create_tm)

            # P11-FIX: Assign TM to same scope as source file (PARITY with offline mode)
            # User registers file from a folder → TM appears in that folder
            from server.repositories.interfaces.tm_repository import AssignmentTarget
            if file_folder_id:
                await tm_repo.assign(result["tm_id"], AssignmentTarget(folder_id=file_folder_id))
                logger.info(f"[ONLINE] Assigned TM {result['tm_id']} to folder {file_folder_id}")
            elif file_project_id:
                await tm_repo.assign(result["tm_id"], AssignmentTarget(project_id=file_project_id))
                logger.info(f"[ONLINE] Assigned TM {result['tm_id']} to project {file_project_id}")

        logger.info(f"TM created from file: tm_id={result['tm_id']}, entries={result['entry_count']}, offline={is_offline}")
        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid file format or data")
    except Exception as e:
        logger.error(f"[FILES] File to TM conversion failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Conversion failed. Check server logs.")


# =============================================================================
# File Download/Export
# =============================================================================

@router.get("/files/{file_id}/download")
async def download_file(
    file_id: int,
    status_filter: Optional[str] = Query(None, description="Filter by status: reviewed, translated, all"),
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Download a file with current translations.

    Rebuilds the file from database rows in the original format.
    Uses extra_data for FULL structure preservation (extra columns, attributes, etc.)

    DB Abstraction: Uses FileRepository - works for both PostgreSQL and SQLite.

    Query params:
    - status_filter: "reviewed" (confirmed only), "translated" (translated+reviewed), "all" (everything)
    """
    # Get file info via repository (handles both PostgreSQL and SQLite)
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check access permission (online mode only - local files have negative IDs)
    if file_id > 0:
        if not await can_access_file(db, file_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    # Get file-level metadata for reconstruction
    file_metadata = file.get("extra_data") or {}

    # Get all rows for export via repository
    rows = await repo.get_rows_for_export(file_id, status_filter=status_filter)

    if not rows:
        raise HTTPException(status_code=404, detail="No rows found for this file")

    # Build file content based on format (case-insensitive)
    # Use _from_dicts builders since repository returns dicts
    format_lower = (file.get("format") or "").lower()
    if format_lower == "txt":
        content = _build_txt_file_from_dicts(rows, file_metadata)
        media_type = "text/plain"
        extension = ".txt"
    elif format_lower == "xml":
        content = _build_xml_file_from_dicts(rows, file_metadata)
        media_type = "application/xml"
        extension = ".xml"
    elif format_lower in ["xlsx", "excel"]:
        content = _build_excel_file_from_dicts(rows, file_metadata)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = ".xlsx"
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {file.get('format')}")

    # Create filename
    original_filename = file.get("original_filename") or file.get("name") or "file"
    base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
    download_name = f"{base_name}_export{extension}"

    logger.info(f"LDM: Downloading file {file_id} ({len(rows)} rows) as {download_name}")

    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={download_name}"}
    )


# =============================================================================
# File Merge (P3: MERGE System)
# =============================================================================

@router.post("/files/{file_id}/merge")
async def merge_file(
    file_id: int,
    original_file: UploadFile = File(..., description="Original file to merge into"),
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Merge reviewed translations from LDM back into original file.

    DB Abstraction: Uses FileRepository - works for both PostgreSQL and SQLite.

    Flow:
    1. User uploads original file (from LanguageData)
    2. System gets reviewed rows from LDM database
    3. Matches by StringID + Source
    4. EDIT: Updates target for matches
    5. ADD: Appends new rows not in original
    6. Returns merged file for download

    Supported formats: TXT, XML (Excel not supported - no StringID)
    """
    from server.tools.ldm.file_handlers.txt_handler import parse_txt_file
    from server.tools.ldm.file_handlers.xml_handler import parse_xml_file

    # Get LDM file info via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check access permission (online mode only - local files have negative IDs)
    if file_id > 0:
        project_id = file.get("project_id")
        if project_id and not await can_access_project(db, project_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    # Check format compatibility
    format_lower = (file.get("format") or "").lower()
    if format_lower not in ["txt", "xml"]:
        raise HTTPException(
            status_code=400,
            detail=f"Merge not supported for format: {file.get('format')}. Only TXT and XML supported."
        )

    # Read original file content
    original_content = await original_file.read()
    original_filename = original_file.filename or "merged_file"

    # Determine format from original file extension
    original_ext = original_filename.rsplit('.', 1)[-1].lower() if '.' in original_filename else ""

    # Validate format match
    if original_ext == "txt" and format_lower != "txt":
        raise HTTPException(status_code=400, detail="Format mismatch: LDM file is XML, original is TXT")
    if original_ext == "xml" and format_lower != "xml":
        raise HTTPException(status_code=400, detail="Format mismatch: LDM file is TXT, original is XML")

    # Parse original file
    if format_lower == "txt":
        original_rows = parse_txt_file(original_content, original_filename)
    else:  # xml
        original_rows = parse_xml_file(original_content, original_filename)

    if not original_rows:
        raise HTTPException(status_code=400, detail="Could not parse original file or file is empty")

    # Get reviewed rows from LDM database via repository
    db_rows = await repo.get_rows_for_export(file_id, status_filter="reviewed")

    if not db_rows:
        raise HTTPException(status_code=400, detail="No reviewed rows to merge")

    # Build lookup from DB rows: (string_id, source) -> row
    db_lookup = {}
    for row in db_rows:
        key = (row.get("string_id") or "", row.get("source") or "")
        db_lookup[key] = row

    # Track merge statistics
    edited_count = 0
    added_count = 0
    original_keys = set()

    # Merge: Update original rows with reviewed translations
    merged_rows = []
    for orig in original_rows:
        key = (orig.get("string_id") or "", orig.get("source") or "")
        original_keys.add(key)

        if key in db_lookup:
            # EDIT: Replace target with reviewed translation
            db_row = db_lookup[key]
            orig["target"] = db_row.get("target")
            orig["extra_data"] = db_row.get("extra_data") or orig.get("extra_data")
            edited_count += 1

        merged_rows.append(orig)

    # ADD: Append new rows from DB that don't exist in original
    for key, db_row in db_lookup.items():
        if key not in original_keys:
            merged_rows.append({
                "row_num": len(merged_rows) + 1,
                "string_id": db_row.get("string_id"),
                "source": db_row.get("source"),
                "target": db_row.get("target"),
                "extra_data": db_row.get("extra_data")
            })
            added_count += 1

    # Build merged file content
    file_extra_data = file.get("extra_data") or {}
    if format_lower == "txt":
        content = _build_txt_file_from_dicts(merged_rows, file_extra_data)
        media_type = "text/plain"
        extension = ".txt"
    else:  # xml
        content = _build_xml_file_from_dicts(merged_rows, file_extra_data)
        media_type = "application/xml"
        extension = ".xml"

    # Create filename
    base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
    download_name = f"{base_name}_merged{extension}"

    logger.info(f"LDM MERGE: file_id={file_id}, edited={edited_count}, added={added_count}, total={len(merged_rows)}")

    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={download_name}",
            "X-Merge-Edited": str(edited_count),
            "X-Merge-Added": str(added_count),
            "X-Merge-Total": str(len(merged_rows))
        }
    )


# =============================================================================
# File Format Conversion (P4)
# =============================================================================

@router.get("/files/{file_id}/convert")
async def convert_file(
    file_id: int,
    format: str = Query(..., regex="^(xlsx|xml|txt|tmx)$", description="Target format"),
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Convert a file to a different format.
    DB Abstraction: Uses FileRepository - works for both PostgreSQL and SQLite.

    Supported conversions:
    - TXT → Excel, XML, TMX
    - XML → Excel, TMX
    - Excel → XML, TMX

    NOT supported (StringID loss):
    - XML → TXT
    - Excel → TXT
    """
    # Get file info via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check access permission (online mode only - local files have negative IDs)
    if file_id > 0:
        project_id = file.get("project_id")
        if project_id and not await can_access_project(db, project_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    source_format = (file.get("format") or "").lower()
    target_format = format.lower()

    # Validate conversion is allowed
    if target_format == "txt" and source_format in ["xml", "xlsx", "excel"]:
        raise HTTPException(
            status_code=400,
            detail="Cannot convert to TXT: StringID information would be lost"
        )

    if source_format == target_format:
        raise HTTPException(
            status_code=400,
            detail=f"File is already in {target_format.upper()} format. Use Download instead."
        )

    # Get all rows for this file via repository
    rows = await repo.get_rows_for_export(file_id)

    if not rows:
        raise HTTPException(status_code=404, detail="No rows found for this file")

    # Get file metadata
    file_metadata = file.get("extra_data") or {}

    # Build file in target format (using _from_dicts builders)
    if target_format == "xlsx":
        content = _build_excel_file_from_dicts(rows, file_metadata)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = ".xlsx"
    elif target_format == "xml":
        content = _build_xml_file_from_dicts(rows, file_metadata)
        media_type = "application/xml"
        extension = ".xml"
    elif target_format == "txt":
        content = _build_txt_file_from_dicts(rows, file_metadata)
        media_type = "text/plain"
        extension = ".txt"
    elif target_format == "tmx":
        # _build_tmx_file already handles both obj and dict rows
        # Create a simple object-like wrapper for language access
        class FileInfo:
            def __init__(self, file_dict):
                self.source_language = file_dict.get("source_language")
                self.target_language = file_dict.get("target_language")
        content = _build_tmx_file(rows, file_metadata, FileInfo(file))
        media_type = "application/x-tmx+xml"
        extension = ".tmx"
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")

    # Create filename
    original_filename = file.get("original_filename") or file.get("name") or "file"
    base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
    download_name = f"{base_name}_converted{extension}"

    logger.info(f"LDM CONVERT: file_id={file_id}, {source_format} → {target_format}, {len(rows)} rows")

    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={download_name}",
            "X-Source-Format": source_format,
            "X-Target-Format": target_format,
            "X-Row-Count": str(len(rows))
        }
    )


@router.get("/files/{file_id}/extract-glossary")
async def extract_glossary(
    file_id: int,
    repo: FileRepository = Depends(get_file_repository),
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(get_current_active_user_async)
):
    """
    Extract glossary terms from a file and download as Excel.

    DB Abstraction: Uses FileRepository - works for both PostgreSQL and SQLite.

    Filtering rules:
    - Length ≤ 21 characters
    - Minimum 2 occurrences
    - No punctuation endings (.?!)

    Returns Excel with Source/Target columns.
    """
    from collections import Counter

    # Get file info via repository
    file = await repo.get(file_id)

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check access permission (online mode only - local files have negative IDs)
    if file_id > 0:
        project_id = file.get("project_id")
        if project_id and not await can_access_project(db, project_id, current_user):
            raise HTTPException(status_code=404, detail="Resource not found")

    # Get all rows for this file via repository
    rows = await repo.get_rows_for_export(file_id)

    # Filter to only rows with source (repository returns all rows)
    rows = [r for r in rows if r.get("source")]

    if not rows:
        raise HTTPException(status_code=404, detail="No rows found in file")

    # Count occurrences of each source term
    source_counts = Counter(row.get("source", "").strip() for row in rows if row.get("source"))

    # Build source → target mapping (first occurrence wins)
    source_to_target = {}
    for row in rows:
        source = (row.get("source") or "").strip()
        if source and source not in source_to_target:
            source_to_target[source] = (row.get("target") or "").strip()

    # Filter glossary terms
    glossary = []
    for source, count in source_counts.items():
        # Rule 1: Length ≤ 21 characters
        if len(source) > 21:
            continue
        # Rule 2: Minimum 2 occurrences
        if count < 2:
            continue
        # Rule 3: No punctuation endings
        if source.endswith(('.', '?', '!')):
            continue

        target = source_to_target.get(source, "")
        glossary.append((source, target, count))

    if not glossary:
        raise HTTPException(status_code=404, detail="No glossary terms found (check filtering rules)")

    # Sort by frequency (most common first)
    glossary.sort(key=lambda x: -x[2])

    # Build Excel file
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glossary"

    # Header row
    ws.cell(row=1, column=1, value="Source")
    ws.cell(row=1, column=2, value="Target")

    # Data rows
    for idx, (source, target, _count) in enumerate(glossary, start=2):
        ws.cell(row=idx, column=1, value=source)
        ws.cell(row=idx, column=2, value=target)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create filename
    original_filename = file.get("original_filename") or file.get("name") or "file"
    base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
    download_name = f"{base_name}_glossary.xlsx"

    logger.info(f"LDM: Extracted glossary from file {file_id}: {len(glossary)} terms")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={download_name}"}
    )


# =============================================================================
# Dict-based File Builders (for Repository Pattern)
# =============================================================================

def _build_txt_file_from_dicts(rows: List[dict], file_metadata: dict = None) -> bytes:
    """
    Build TXT file from dict rows.

    P10: Works with dict rows from Repository Pattern.
    """
    file_metadata = file_metadata or {}
    lines = []

    for row in rows:
        # Reconstruct string_id parts
        string_id = row.get("string_id") or ""
        string_id_parts = string_id.split(' ') if string_id else [""] * 5

        # Ensure we have 5 parts
        while len(string_id_parts) < 5:
            string_id_parts.append("")

        source = row.get("source") or ""
        target = row.get("target") or ""

        # Replace newline markers
        source = source.replace("↵", "\n")
        target = target.replace("↵", "\n")

        # Start with standard 7 columns
        parts = string_id_parts[:5] + [source, target]

        # Add extra columns
        extra_data = row.get("extra_data")
        if extra_data:
            total_cols = file_metadata.get("total_columns", 7)
            for i in range(7, total_cols):
                col_key = f"col{i}"
                parts.append(extra_data.get(col_key, ""))

        line = "\t".join(parts)
        lines.append(line)

    content = "\n".join(lines)
    encoding = file_metadata.get("encoding", "utf-8")
    return content.encode(encoding)


def _build_xml_file_from_dicts(rows: List[dict], file_metadata: dict = None) -> bytes:
    """
    Build XML file from dict rows (used by merge).

    Same as _build_xml_file but works with dicts instead of LDMRow objects.
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    file_metadata = file_metadata or {}

    root_tag = file_metadata.get("root_element", "LangData")
    root = ET.Element(root_tag)

    root_attribs = file_metadata.get("root_attributes")
    if root_attribs:
        for key, val in root_attribs.items():
            root.set(key, val)

    element_tag = file_metadata.get("element_tag", "String")

    for row in rows:
        string_elem = ET.SubElement(root, element_tag)

        string_elem.set("stringid", row.get("string_id") or "")
        string_elem.set("strorigin", row.get("source") or "")
        string_elem.set("str", row.get("target") or "")

        extra_data = row.get("extra_data")
        if extra_data:
            for attr_name, attr_val in extra_data.items():
                string_elem.set(attr_name, attr_val or "")

    encoding = file_metadata.get("encoding", "UTF-8")
    xml_str = ET.tostring(root, encoding='unicode')
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ", encoding=encoding)

    return pretty_xml


def _build_excel_file_from_dicts(rows: List[dict], file_metadata: dict = None) -> bytes:
    """
    Build Excel file from dict rows (used by Repository Pattern).

    Same as _build_excel_file but works with dicts instead of LDMRow objects.
    """
    import openpyxl

    file_metadata = file_metadata or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = file_metadata.get("sheet_name", "Translations")

    # Get headers from metadata or use defaults
    headers = file_metadata.get("headers", ["Source", "Target"])
    if len(headers) < 2:
        headers = ["Source", "Target"]

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row.get("source") or "")
        ws.cell(row=row_idx, column=2, value=row.get("target") or "")

        # Add extra columns from extra_data (C, D, E, etc.)
        extra_data = row.get("extra_data")
        if extra_data:
            for col_letter, val in extra_data.items():
                # Convert column letter to index (C=3, D=4, etc.)
                if col_letter.isalpha() and len(col_letter) == 1:
                    col_num = ord(col_letter.upper()) - ord('A') + 1
                    if col_num > 2:  # Skip A and B (source/target)
                        ws.cell(row=row_idx, column=col_num, value=val or "")

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


# =============================================================================
# TMX Builder (P4: File Conversions)
# =============================================================================

def _build_tmx_file(rows: List[LDMRow], file_metadata: dict = None, file: LDMFile = None) -> bytes:
    """
    Build TMX (Translation Memory eXchange) file from rows.

    TMX format:
    <?xml version="1.0" encoding="UTF-8"?>
    <tmx version="1.4">
      <header ... />
      <body>
        <tu>
          <prop type="x-string-id">StringID</prop>
          <tuv xml:lang="ko"><seg>Source</seg></tuv>
          <tuv xml:lang="en"><seg>Target</seg></tuv>
        </tu>
      </body>
    </tmx>
    """
    import xml.etree.ElementTree as ET
    from datetime import datetime

    file_metadata = file_metadata or {}

    # Determine languages
    source_lang = file.source_language if file else "ko"
    target_lang = file.target_language if file else "en"

    # Create root element
    root = ET.Element("tmx", version="1.4")

    # Header
    header = ET.SubElement(root, "header")
    header.set("creationtool", "LocaNext")
    header.set("creationtoolversion", "1.0")
    header.set("datatype", "plaintext")
    header.set("segtype", "sentence")
    header.set("adminlang", "en")
    header.set("srclang", source_lang or "ko")
    header.set("creationdate", datetime.utcnow().strftime("%Y%m%dT%H%M%SZ"))

    # Body
    body = ET.SubElement(root, "body")

    for row in rows:
        tu = ET.SubElement(body, "tu")

        # Add StringID as property if available
        string_id = row.string_id if hasattr(row, 'string_id') else row.get("string_id")
        if string_id:
            prop = ET.SubElement(tu, "prop", type="x-string-id")
            prop.text = string_id

        # Source segment
        tuv_src = ET.SubElement(tu, "tuv")
        tuv_src.set("{http://www.w3.org/XML/1998/namespace}lang", source_lang or "ko")
        seg_src = ET.SubElement(tuv_src, "seg")
        source = row.source if hasattr(row, 'source') else row.get("source")
        seg_src.text = source or ""

        # Target segment
        tuv_tgt = ET.SubElement(tu, "tuv")
        tuv_tgt.set("{http://www.w3.org/XML/1998/namespace}lang", target_lang or "en")
        seg_tgt = ET.SubElement(tuv_tgt, "seg")
        target = row.target if hasattr(row, 'target') else row.get("target")
        seg_tgt.text = target or ""

    # Convert to bytes with XML declaration
    tree = ET.ElementTree(root)
    output = BytesIO()
    tree.write(output, encoding="utf-8", xml_declaration=True)
    content = output.getvalue()
    output.close()

    return content


# =============================================================================
# P9: Local Storage Upload (Offline Storage)
# =============================================================================

async def _upload_to_local_storage(
    file: UploadFile,
    source_col: Optional[int] = None,
    target_col: Optional[int] = None,
    stringid_col: Optional[int] = None,
    has_header: Optional[bool] = True,
    folder_id: Optional[int] = None  # P9-FIX: Support creating files inside local folders
) -> dict:
    """
    P9: Upload a file to SQLite Offline Storage with proper parsing.

    This uses the SAME file parsers as PostgreSQL upload, ensuring
    consistent parsing for both online and offline modes.

    Args:
        folder_id: Optional local folder ID to place file in. If None, file is at root.

    Returns a FileResponse-compatible dict.
    """
    from server.database.offline import get_offline_db
    from datetime import datetime

    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ''

    # Parse file using existing handlers (same as PostgreSQL upload)
    file_metadata = None
    if ext in ('txt', 'tsv'):
        from server.tools.ldm.file_handlers.txt_handler import parse_txt_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_txt_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext == 'xml':
        from server.tools.ldm.file_handlers.xml_handler import parse_xml_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_xml_file(file_content, filename)
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    elif ext in ('xlsx', 'xls'):
        from server.tools.ldm.file_handlers.excel_handler import parse_excel_file, get_file_format, get_source_language, get_file_metadata
        file_content = await file.read()
        rows_data = parse_excel_file(
            file_content,
            filename,
            source_col=source_col if source_col is not None else 0,
            target_col=target_col if target_col is not None else 1,
            stringid_col=stringid_col,
            has_header=has_header if has_header is not None else True
        )
        file_format = get_file_format()
        source_lang = get_source_language()
        file_metadata = get_file_metadata()
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: {ext}. Use TXT, TSV, XML, or Excel (XLSX/XLS)."
        )

    if not rows_data:
        raise HTTPException(
            status_code=400,
            detail="No valid rows found in file"
        )

    try:
        offline_db = get_offline_db()

        # Create local file in Offline Storage (may auto-rename if duplicate)
        # P9-FIX: Support placing files inside local folders
        result = offline_db.create_local_file(
            name=filename,
            original_filename=filename,
            file_format=file_format,
            source_language=source_lang,
            target_language=None,
            folder_id=folder_id  # P9-FIX: Place in specified folder
        )
        file_id = result["id"]
        final_name = result["name"]  # May be different from filename if renamed

        # Add parsed rows to the local file
        offline_db.add_rows_to_local_file(file_id, rows_data)

        logger.success(f"[FILES] P9: File uploaded to Offline Storage: id={file_id}, name='{final_name}', rows={len(rows_data)}")

        # Return FileResponse-compatible dict
        now = datetime.utcnow()
        return {
            "id": file_id,
            "project_id": None,  # Local files have no project
            "folder_id": folder_id,  # P9-FIX: Include folder_id
            "name": final_name,  # P9-FIX: Use actual name (may be auto-renamed)
            "original_filename": filename,
            "format": file_format,
            "row_count": len(rows_data),
            "source_language": source_lang,
            "target_language": None,
            "created_at": now,
            "updated_at": now
        }

    except Exception as e:
        logger.error(f"[FILES] P9: Failed to upload to Offline Storage: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")
