"""
LDM Excel File Handler

Parses Excel localization files with flexible column mapping:
- 2-column mode: Source (A) + Target (B)
- 3-column mode: Source (A) + Target (B) + StringID (C)
- Extra columns preserved in extra_data for full reconstruction

User selects columns via UI - not hardcoded.

Returns list of row dicts ready for database insertion.
Also returns file-level metadata for full file reconstruction.
"""

from typing import List, Dict, Optional
from loguru import logger

# Factor Power: Use centralized text utils
from server.utils.text_utils import normalize_text


# Module-level state for file metadata
_file_metadata: Dict = {}


def parse_excel_file(
    file_content: bytes,
    filename: str,
    source_col: int = 0,      # Column A by default
    target_col: int = 1,      # Column B by default
    stringid_col: Optional[int] = None,  # Column C if 3-column mode
    has_header: bool = True   # Skip first row if header
) -> List[Dict]:
    """
    Parse Excel file content and extract rows.

    Preserves ALL data for full file reconstruction:
    - Mapped columns (source, target, stringid) parsed to dedicated fields
    - Extra columns stored in extra_data for reconstruction

    Structures supported:
    - 2-column: Source (A) + Target (B)
    - 3-column: Source (A) + Target (B) + StringID (C)

    Args:
        file_content: Raw Excel file bytes
        filename: Original filename (for logging)
        source_col: Column index for source text (0 = A, 1 = B, etc.)
        target_col: Column index for target text
        stringid_col: Column index for StringID (None if 2-column mode)
        has_header: Whether first row is header (skip it)

    Returns:
        List of row dicts with keys: row_num, string_id, source, target, extra_data
    """
    global _file_metadata
    rows = []

    try:
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        sheet_name = ws.title

        # Determine which columns are "core" (mapped) vs "extra"
        core_cols = {source_col, target_col}
        if stringid_col is not None:
            core_cols.add(stringid_col)

        row_num = 0
        max_columns = 0
        headers = []

        for excel_row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Track max columns
            if row:
                max_columns = max(max_columns, len(row))

            # Handle header row
            if excel_row_idx == 1 and has_header:
                headers = [str(cell) if cell else f"Column_{i}" for i, cell in enumerate(row)]
                continue

            # Skip empty rows
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            row_num += 1

            # Extract source text
            source = ""
            if source_col < len(row) and row[source_col] is not None:
                source = normalize_text(str(row[source_col]))

            # Extract target text
            target = ""
            if target_col < len(row) and row[target_col] is not None:
                target = normalize_text(str(row[target_col]))

            # Extract StringID (if 3-column mode)
            string_id = None
            if stringid_col is not None and stringid_col < len(row) and row[stringid_col] is not None:
                string_id = str(row[stringid_col]).strip()
                if not string_id:
                    string_id = None

            # Skip if both source and target are empty
            if not source and not target:
                continue

            # Capture extra columns for full reconstruction
            extra_data = None
            extra_cols = {}
            for i, val in enumerate(row):
                if i not in core_cols and val is not None:
                    col_name = headers[i] if i < len(headers) else f"col{i}"
                    extra_cols[col_name] = str(val)
            if extra_cols:
                extra_data = extra_cols

            rows.append({
                "row_num": row_num,
                "string_id": string_id,
                "source": source if source else None,
                "target": target if target else None,
                "status": "translated" if target else "pending",
                "extra_data": extra_data
            })

        wb.close()

        # Store file-level metadata for reconstruction
        _file_metadata = {
            "encoding": "utf-8",  # openpyxl handles encoding
            "sheet_name": sheet_name,
            "headers": headers,
            "total_columns": max_columns,
            "source_col": source_col,
            "target_col": target_col,
            "stringid_col": stringid_col,
            "has_header": has_header,
            "format_version": "1.0"
        }

        logger.info(f"Parsed Excel {filename}: {len(rows)} rows, sheet='{sheet_name}', cols={max_columns}")

    except ImportError:
        logger.error("openpyxl not installed - cannot parse Excel files")
        raise ValueError("Excel support requires openpyxl package")
    except Exception as e:
        logger.error(f"Failed to parse Excel file {filename}: {e}")
        raise

    return rows


def get_file_metadata() -> Dict:
    """
    Get file-level metadata for full reconstruction.

    Returns dict with:
    - sheet_name: Original sheet name
    - headers: Column headers from first row
    - total_columns: Maximum column count
    - source_col, target_col, stringid_col: Column mappings used
    """
    global _file_metadata
    return _file_metadata.copy()


def get_source_language() -> str:
    """Return the source language for Excel files."""
    return "KR"  # Korean is source


def get_file_format() -> str:
    """Return the file format identifier."""
    return "XLSX"


def get_excel_preview(
    file_content: bytes,
    max_rows: int = 5
) -> Dict:
    """
    Get a preview of the Excel file for column mapping UI.

    Returns:
        Dict with sheet_name, headers, and sample rows
    """
    try:
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        preview = {
            "sheet_name": ws.title,
            "headers": [],
            "sample_rows": [],
            "total_rows": 0
        }

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # First row as headers
                preview["headers"] = [str(cell) if cell else f"Column {i+1}" for i, cell in enumerate(row)]
            elif row_idx <= max_rows + 1:
                # Sample rows
                preview["sample_rows"].append([str(cell) if cell else "" for cell in row])

            preview["total_rows"] = row_idx

        wb.close()
        return preview

    except Exception as e:
        logger.error(f"Failed to get Excel preview: {e}")
        raise
