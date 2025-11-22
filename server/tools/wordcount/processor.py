"""
WordCount Processor

Core logic for processing XML translation files and generating word count reports.
Simplified version adapted from wordcount_diff_master.py for web API use.
"""

import os
import re
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Set, Optional

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from loguru import logger


class WordCountProcessor:
    """
    Process XML translation files and generate word count comparison reports.

    Features:
    - Parse XML language files
    - Calculate word counts and coverage
    - Compare with historical data
    - Generate Excel reports with smart categorization
    - Maintain history in JSON file
    """

    def __init__(self, history_file: str, output_dir: str):
        """
        Initialize processor.

        Args:
            history_file: Path to history JSON file
            output_dir: Directory to save generated reports
        """
        self.history_file = Path(history_file)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)

        # XML parsing helpers
        self._bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')
        self._korean_re = re.compile(r'[\uac00-\ud7a3]')

        logger.info(f"WordCountProcessor initialized", {
            "history_file": str(history_file),
            "output_dir": str(output_dir)
        })

    # ========================================================================
    # XML PARSING
    # ========================================================================

    def _fix_bad_entities(self, xml_text: str) -> str:
        """Fix malformed XML entities."""
        return self._bad_entity_re.sub("&amp;", xml_text)

    def _parse_xml_file(self, path: Path) -> ET._Element:
        """Parse XML file with error recovery."""
        raw = path.read_text(encoding="utf-8", errors="ignore")
        wrapped = f"<ROOT>\n{self._fix_bad_entities(raw)}\n</ROOT>"
        parser = ET.XMLParser(recover=True, huge_tree=True)
        return ET.fromstring(wrapped.encode("utf-8"), parser=parser)

    def _count_words(self, text: str) -> int:
        """Count words in text."""
        return len([w for w in re.split(r'\s+', text.strip()) if w])

    def _is_korean(self, text: str) -> bool:
        """Check if text contains Korean characters."""
        return bool(self._korean_re.search(text))

    def _analyze_file(self, path: Path) -> Tuple[int, int]:
        """
        Analyze a single language file.

        Returns:
            (total_words, completed_words)
        """
        root = self._parse_xml_file(path)
        total_words = completed_words = 0

        for loc in root.iter("LocStr"):
            origin = (loc.get("StrOrigin") or "").strip()
            if not origin:
                continue
            origin_wc = self._count_words(origin)
            total_words += origin_wc

            trans = (loc.get("Str") or "").strip()
            if trans and not self._is_korean(trans):
                completed_words += origin_wc

        return total_words, completed_words

    # ========================================================================
    # FILE PROCESSING
    # ========================================================================

    def _detect_language(self, filename: str) -> Optional[str]:
        """
        Detect language code from filename.

        Expected format: LanguageData_XXX.xml where XXX is language code.

        Args:
            filename: File name to analyze

        Returns:
            Language code (e.g., "ENG", "FRA") or None if not detected
        """
        # Pattern: LanguageData_XXX.xml
        match = re.search(r'LanguageData_([A-Z]{2,3})\.xml', filename, re.IGNORECASE)
        if match:
            return match.group(1).upper()
        return None

    def _process_language_files(self, file_paths: List[str]) -> Dict[str, Dict[str, int]]:
        """
        Process all language files and calculate word counts.

        Args:
            file_paths: List of XML file paths

        Returns:
            Dict mapping language code to stats:
            {
                "ENG": {"total_words": 125000, "completed_words": 98000},
                "FRA": {"total_words": 125000, "completed_words": 95000},
                ...
            }
        """
        logger.info(f"Processing {len(file_paths)} XML files")

        language_stats = {}

        for file_path in file_paths:
            path = Path(file_path)
            if not path.exists():
                logger.warning(f"File not found: {file_path}")
                continue

            # Detect language from filename
            lang_code = self._detect_language(path.name)
            if not lang_code:
                logger.warning(f"Could not detect language from filename: {path.name}")
                continue

            # Skip Korean files
            if lang_code == "KOR":
                logger.debug(f"Skipping Korean file: {path.name}")
                continue

            logger.debug(f"Processing {lang_code}: {path.name}")

            try:
                total_words, completed_words = self._analyze_file(path)

                # Aggregate stats for this language
                if lang_code not in language_stats:
                    language_stats[lang_code] = {
                        "total_words": 0,
                        "completed_words": 0
                    }

                language_stats[lang_code]["total_words"] += total_words
                language_stats[lang_code]["completed_words"] += completed_words

                logger.debug(f"{lang_code}: {total_words} total, {completed_words} completed")

            except Exception as e:
                logger.error(f"Error processing {path.name}: {e}")
                continue

        logger.success(f"Processed {len(language_stats)} languages")
        return language_stats

    # ========================================================================
    # HISTORY MANAGEMENT
    # ========================================================================

    def _load_history(self) -> List[Dict]:
        """
        Load history from JSON file.

        Returns:
            List of historical runs
        """
        if not self.history_file.exists():
            logger.info("No history file found, starting fresh")
            return []

        try:
            with open(self.history_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                history = data.get('history', [])
                logger.info(f"Loaded {len(history)} historical runs")
                return history
        except Exception as e:
            logger.error(f"Error loading history: {e}")
            return []

    def _save_history(self, history: List[Dict]):
        """Save history to JSON file."""
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump({"history": history}, f, indent=2, ensure_ascii=False)
            logger.success(f"History saved: {len(history)} runs")
        except Exception as e:
            logger.error(f"Error saving history: {e}")

    def _find_past_data(self, history: List[Dict], past_date: str) -> Optional[Dict]:
        """
        Find historical data for a specific date.

        Args:
            history: List of historical runs
            past_date: Date to find (YYYY-MM-DD)

        Returns:
            Historical data dict or None if not found
        """
        for entry in history:
            if entry.get('date') == past_date:
                logger.info(f"Found historical data for {past_date}")
                return entry

        logger.warning(f"No historical data found for {past_date}")
        return None

    # ========================================================================
    # PERIOD CATEGORIZATION
    # ========================================================================

    def _determine_category(self, days_diff: int) -> str:
        """
        Determine if period should be categorized as Weekly or Monthly.

        Logic: Compare distance to 7 days vs 30 days
        - If closer to 7: Weekly
        - If closer to 30: Monthly

        Args:
            days_diff: Number of days between dates

        Returns:
            "weekly" or "monthly"
        """
        dist_weekly = abs(days_diff - 7)
        dist_monthly = abs(days_diff - 30)

        if dist_weekly < dist_monthly:
            category = "weekly"
        else:
            category = "monthly"

        logger.info(f"Period categorized as {category.upper()}", {
            "days_diff": days_diff,
            "dist_weekly": dist_weekly,
            "dist_monthly": dist_monthly
        })

        return category

    # ========================================================================
    # EXCEL REPORT GENERATION
    # ========================================================================

    def _generate_excel_report(
        self,
        current_data: Dict[str, Dict[str, int]],
        past_data: Optional[Dict],
        past_date: str,
        category: str,
        output_filename: str
    ) -> Path:
        """
        Generate Excel report with 4 sheets.

        Args:
            current_data: Current language stats
            past_data: Historical data (or None)
            past_date: Past date string
            category: "weekly" or "monthly"
            output_filename: Output file name

        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating Excel report: {output_filename}")

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Period info
        today = datetime.now().strftime("%Y-%m-%d")
        past_date_obj = datetime.strptime(past_date, "%Y-%m-%d")
        today_obj = datetime.strptime(today, "%Y-%m-%d")
        days_diff = (today_obj - past_date_obj).days
        period_title = f"Period: {today} to {past_date} ({days_diff} days)"

        # Create 4 sheets
        sheets = [
            ("Weekly Diff - Full", "weekly", "full"),
            ("Monthly Diff - Full", "monthly", "full"),
            ("Weekly Diff - Detailed", "weekly", "detailed"),
            ("Monthly Diff - Detailed", "monthly", "detailed")
        ]

        for sheet_name, sheet_category, sheet_type in sheets:
            ws = wb.create_sheet(sheet_name)

            if sheet_category == category:
                # Active sheet - show data
                self._populate_active_sheet(
                    ws=ws,
                    current_data=current_data,
                    past_data=past_data,
                    period_title=period_title,
                    sheet_type=sheet_type
                )
            else:
                # Inactive sheet - show N/A message
                self._populate_na_sheet(ws, category, sheet_category)

        # Save file
        output_path = self.output_dir / output_filename
        wb.save(str(output_path))

        logger.success(f"Excel report generated: {output_path}")
        return output_path

    def _populate_active_sheet(
        self,
        ws,
        current_data: Dict,
        past_data: Optional[Dict],
        period_title: str,
        sheet_type: str
    ):
        """Populate active sheet with data."""
        # Title row
        ws["A1"] = period_title
        ws["A1"].font = Font(bold=True, color="0000FF", size=12)

        # Headers (row 3)
        headers = ["Language", "Total Words", "Completed", "Coverage %"]
        if past_data:
            headers.extend(["Diff Words", "Diff %"])

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Data rows
        row = 4
        for lang, stats in sorted(current_data.items()):
            total = stats["total_words"]
            completed = stats["completed_words"]
            coverage = (completed / total * 100) if total > 0 else 0

            ws.cell(row=row, column=1, value=lang)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=completed)
            ws.cell(row=row, column=4, value=f"{coverage:.1f}%")

            # Calculate diffs if past data exists
            if past_data:
                past_lang_data = past_data.get('languages', {}).get(lang, {})
                past_completed = past_lang_data.get('completed_words', 0)

                diff_words = completed - past_completed
                diff_pct = (diff_words / past_completed * 100) if past_completed > 0 else 0

                ws.cell(row=row, column=5, value=diff_words)
                ws.cell(row=row, column=6, value=f"{diff_pct:.1f}%")

                # Color code diffs
                if diff_words > 0:
                    ws.cell(row=row, column=5).font = Font(color="008000")  # Green
                elif diff_words < 0:
                    ws.cell(row=row, column=5).font = Font(color="FF0000")  # Red

            row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _populate_na_sheet(self, ws, actual_category: str, sheet_category: str):
        """Populate N/A sheet with message."""
        ws["A1"] = "N/A - Select appropriate comparison period"
        ws["A1"].font = Font(bold=True, color="FF0000", size=12)

        ws["A3"] = f"This sheet is for comparisons around {'7' if sheet_category == 'weekly' else '30'} days apart."
        ws["A4"] = f"Your selected period was categorized as {actual_category.upper()}."

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 70

    # ========================================================================
    # MAIN PROCESSING
    # ========================================================================

    def process_files(
        self,
        file_paths: List[str],
        past_date: str
    ) -> Tuple[str, Dict]:
        """
        Process files and generate word count report.

        Args:
            file_paths: List of XML file paths
            past_date: Past date to compare against (YYYY-MM-DD)

        Returns:
            Tuple of (report_path, stats_dict)
        """
        logger.info(f"Starting WordCount processing", {
            "files_count": len(file_paths),
            "past_date": past_date
        })

        # Step 1: Process current files
        current_data = self._process_language_files(file_paths)

        if not current_data:
            raise ValueError("No valid language files found")

        # Step 2: Load history
        history = self._load_history()

        # Step 3: Find past data
        past_data = self._find_past_data(history, past_date)

        # Step 4: Determine category
        today = datetime.now()
        past_date_obj = datetime.strptime(past_date, "%Y-%m-%d")
        days_diff = (today - past_date_obj).days
        category = self._determine_category(days_diff)

        # Step 5: Generate Excel report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"WordCountAnalysis_{timestamp}.xlsx"
        report_path = self._generate_excel_report(
            current_data=current_data,
            past_data=past_data,
            past_date=past_date,
            category=category,
            output_filename=output_filename
        )

        # Step 6: Update history
        today_str = today.strftime("%Y-%m-%d")
        new_entry = {
            "date": today_str,
            "languages": {
                lang: {
                    "total_words": stats["total_words"],
                    "completed_words": stats["completed_words"],
                    "coverage_percent": (stats["completed_words"] / stats["total_words"] * 100)
                    if stats["total_words"] > 0 else 0
                }
                for lang, stats in current_data.items()
            }
        }

        # Replace or append
        found = False
        for idx, entry in enumerate(history):
            if entry["date"] == today_str:
                history[idx] = new_entry
                found = True
                break

        if not found:
            history.append(new_entry)

        self._save_history(history)

        # Calculate overall stats
        total_words = sum(stats["total_words"] for stats in current_data.values())
        completed_words = sum(stats["completed_words"] for stats in current_data.values())
        coverage_percent = (completed_words / total_words * 100) if total_words > 0 else 0

        stats = {
            "languages_processed": len(current_data),
            "total_words": total_words,
            "completed_words": completed_words,
            "coverage_percent": round(coverage_percent, 2),
            "category": category.title(),
            "days_diff": days_diff
        }

        logger.success(f"WordCount processing complete", stats)

        return str(report_path), stats
