#!/usr/bin/env python3
"""
Process XLSTransfer operations
Replicates original XLSTransfer0225.py functionality exactly

Usage: python process_operation.py <operation_type> <selections_json> <threshold>
"""

import sys
import json
import os
import pickle
import numpy as np
import pandas as pd
import openpyxl
import shutil
from pathlib import Path
import faiss
from copy import copy
from typing import Optional, TYPE_CHECKING

# Lazy import for SentenceTransformer (takes ~30s to load PyTorch)
# Import only when model is actually needed, not at module load time
if TYPE_CHECKING:
    from sentence_transformers import SentenceTransformer

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent))
# Add project root for centralized imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

from core import clean_text, excel_column_to_index
import config
# Factor Power: Use centralized progress tracker
from server.utils.progress_tracker import ProgressTracker


def safe_most_frequent(x):
    """Get most frequent value from Series"""
    if x.empty or x.isna().all():
        return pd.NA
    return x.value_counts().index[0]


def create_dictionary(selections, operation_id: Optional[int] = None):
    """
    Create dictionary from Excel files
    Exact replica of process_excel_and_generate_embeddings from original (lines 197-308)

    Args:
        selections: Dict mapping file paths to sheet/column selections
        operation_id: Optional ActiveOperation ID for progress tracking
    """
    # Initialize progress tracker
    tracker = ProgressTracker(operation_id)
    tracker.log_milestone("Starting dictionary creation operation")

    all_kr_texts_split = []
    all_fr_texts_split = []
    all_kr_texts_whole = []
    all_fr_texts_whole = []

    # Load model (lazy import to avoid slow startup)
    tracker.update(5.0, "Loading Korean BERT model...")
    print("Loading Korean BERT model...", file=sys.stderr)
    from sentence_transformers import SentenceTransformer
    model = SentenceTransformer('snunlp/KR-SBERT-V40K-klueNLI-augSTS')
    tracker.log_milestone("Korean BERT model loaded successfully")

    tracker.update(10.0, "Starting concatenation of tabs...")
    print("Starting concatenation of tabs...", file=sys.stderr)
    total_rows = 0
    total_files = len(selections)
    current_file_idx = 0

    for excel_file_path, file_selections in selections.items():
        current_file_idx += 1
        filename = os.path.basename(excel_file_path)

        # Progress from 10% to 40% across all files
        file_progress = 10.0 + (30.0 * current_file_idx / total_files)
        tracker.update(file_progress, f"Processing file {current_file_idx}/{total_files}: {filename}")

        for sheet_name, columns in file_selections.items():
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None)
            col_kr_index = excel_column_to_index(columns['kr_column'])
            col_fr_index = excel_column_to_index(columns['trans_column'])

            # Drop rows where either KR or FR is NaN
            df = df.iloc[:, [col_kr_index, col_fr_index]].dropna()

            kr_texts = df.iloc[:, 0].apply(clean_text).tolist()
            fr_texts = df.iloc[:, 1].apply(clean_text).tolist()
            total_rows += len(kr_texts)

            for kr, fr in zip(kr_texts, fr_texts):
                kr_lines = kr.split('\n')
                fr_lines = fr.split('\n')
                if len(kr_lines) == len(fr_lines):
                    all_kr_texts_split.extend(kr_lines)
                    all_fr_texts_split.extend(fr_lines)
                else:
                    all_kr_texts_whole.append(kr)
                    all_fr_texts_whole.append(fr)

            tracker.log_milestone(f"Sheet '{sheet_name}' in '{filename}' processed: {len(kr_texts)} rows")
            print(f"Tab '{sheet_name}' in file '{os.path.basename(excel_file_path)}' processed: {len(kr_texts)} rows", file=sys.stderr)

    tracker.log_milestone(f"Processed {total_files} file(s), {total_rows} total rows")
    print(f"Total number of files processed: {len(selections)}", file=sys.stderr)
    print(f"Total number of rows concatenated: {total_rows}", file=sys.stderr)

    # Process split texts
    tracker.update(40.0, "Processing split texts (grouping by Korean text)...")
    split_translation_pairs = pd.DataFrame({'KR': all_kr_texts_split, 'FR': all_fr_texts_split})
    split_most_freq_trans = split_translation_pairs.groupby('KR')['FR'].agg(safe_most_frequent).reset_index()
    split_most_freq_trans = split_most_freq_trans.dropna()
    unique_kr_texts_split = split_most_freq_trans['KR'].tolist()
    tracker.log_milestone(f"Found {len(unique_kr_texts_split)} unique split texts")

    # Process whole texts
    tracker.update(45.0, "Processing whole texts (grouping by Korean text)...")
    whole_translation_pairs = pd.DataFrame({'KR': all_kr_texts_whole, 'FR': all_fr_texts_whole})
    whole_most_freq_trans = whole_translation_pairs.groupby('KR')['FR'].agg(safe_most_frequent).reset_index()
    whole_most_freq_trans = whole_most_freq_trans.dropna()
    unique_kr_texts_whole = whole_most_freq_trans['KR'].tolist()
    tracker.log_milestone(f"Found {len(unique_kr_texts_whole)} unique whole texts")

    print(f"Total unique split texts: {len(unique_kr_texts_split)}", file=sys.stderr)
    print(f"Total unique whole texts: {len(unique_kr_texts_whole)}", file=sys.stderr)

    # Generate embeddings for split
    tracker.update(50.0, "Generating embeddings for split texts...")
    split_embeddings = []
    for index, text in enumerate(unique_kr_texts_split):
        batch_embeddings = model.encode([text], convert_to_tensor=False)
        split_embeddings.extend(batch_embeddings)
        if (index + 1) % 100 == 0 or (index + 1) == len(unique_kr_texts_split):
            # Progress from 50% to 75% for split embeddings
            embedding_progress = 50.0 + (25.0 * (index + 1) / len(unique_kr_texts_split))
            tracker.update(
                embedding_progress,
                f"Generating split embeddings: {index + 1}/{len(unique_kr_texts_split)} texts"
            )
            print(f"Split embedding progress: {index + 1}/{len(unique_kr_texts_split)} lines", file=sys.stderr)

    # Generate embeddings for whole
    tracker.update(75.0, "Generating embeddings for whole texts...")
    whole_embeddings = []
    for index, text in enumerate(unique_kr_texts_whole):
        batch_embeddings = model.encode([text], convert_to_tensor=False)
        whole_embeddings.extend(batch_embeddings)
        if (index + 1) % 10 == 0 or (index + 1) == len(unique_kr_texts_whole):
            # Progress from 75% to 90% for whole embeddings
            embedding_progress = 75.0 + (15.0 * (index + 1) / len(unique_kr_texts_whole)) if len(unique_kr_texts_whole) > 0 else 90.0
            tracker.update(
                embedding_progress,
                f"Generating whole embeddings: {index + 1}/{len(unique_kr_texts_whole)} blocks"
            )
            print(f"Whole embedding progress: {index + 1}/{len(unique_kr_texts_whole)} blocks", file=sys.stderr)

    # Save embeddings and dictionaries
    tracker.update(90.0, "Saving split embeddings and dictionary...")
    split_embeddings = np.array(split_embeddings)
    np.save('SplitExcelEmbeddings.npy', split_embeddings)
    split_translation_dict = dict(zip(split_most_freq_trans['KR'], split_most_freq_trans['FR']))
    with open('SplitExcelDictionary.pkl', 'wb') as f:
        pickle.dump(split_translation_dict, f)
    tracker.log_milestone(f"Saved SplitExcelEmbeddings.npy ({len(unique_kr_texts_split)} entries)")

    if unique_kr_texts_whole:
        tracker.update(93.0, "Saving whole embeddings and dictionary...")
        whole_embeddings = np.array(whole_embeddings)
        np.save('WholeExcelEmbeddings.npy', whole_embeddings)
        whole_translation_dict = dict(zip(whole_most_freq_trans['KR'], whole_most_freq_trans['FR']))
        with open('WholeExcelDictionary.pkl', 'wb') as f:
            pickle.dump(whole_translation_dict, f)
        tracker.log_milestone(f"Saved WholeExcelEmbeddings.npy ({len(unique_kr_texts_whole)} entries)")
    else:
        print("No whole embeddings found. Skipping whole mode.", file=sys.stderr)

    tracker.update(95.0, "Dictionary creation complete! Finalizing...")
    tracker.log_milestone(f"Dictionary created: {len(unique_kr_texts_split)} split + {len(unique_kr_texts_whole)} whole entries")

    return {
        "success": True,
        "split_entries": len(unique_kr_texts_split),
        "whole_entries": len(unique_kr_texts_whole),
        "message": f"Dictionary created with {len(unique_kr_texts_split)} split + {len(unique_kr_texts_whole)} whole entries"
    }


def translate_excel(selections, threshold, operation_id: Optional[int] = None):
    """
    Translate Excel to Excel
    Exact replica of translate_excel_to_excel from original (lines 648-778)

    Args:
        selections: Dict mapping file paths to sheet/column selections
        threshold: FAISS similarity threshold (0.0-1.0)
        operation_id: Optional ActiveOperation ID for progress tracking
    """
    # Initialize progress tracker
    tracker = ProgressTracker(operation_id)
    tracker.log_milestone("Starting Excel translation operation")

    # Load dictionaries and create index
    if not os.path.exists('SplitExcelDictionary.pkl') or not os.path.exists('SplitExcelEmbeddings.npy'):
        tracker.log_error("Dictionary files not found", "FileNotFoundError")
        return {"success": False, "error": "Dictionary not loaded. Please load dictionary first."}

    tracker.update(5.0, "Loading Korean BERT model...")
    print("Loading model and dictionary...", file=sys.stderr)
    from sentence_transformers import SentenceTransformer
    model = SentenceTransformer('snunlp/KR-SBERT-V40K-klueNLI-augSTS')

    tracker.update(10.0, "Loading translation dictionaries...")
    # Load split mode
    with open('SplitExcelDictionary.pkl', 'rb') as file:
        translation_dict_split = pickle.load(file)
    ref_kr_embeddings_split = np.load('SplitExcelEmbeddings.npy')
    ref_kr_sentences_split = pd.Series(list(translation_dict_split.keys()), dtype=str)
    faiss.normalize_L2(ref_kr_embeddings_split)
    index_split = faiss.IndexFlatIP(ref_kr_embeddings_split.shape[1])
    index_split.add(ref_kr_embeddings_split)

    # Load whole mode if available
    whole_mode_available = False
    if os.path.exists('WholeExcelDictionary.pkl') and os.path.exists('WholeExcelEmbeddings.npy'):
        with open('WholeExcelDictionary.pkl', 'rb') as file:
            translation_dict_whole = pickle.load(file)
        ref_kr_embeddings_whole = np.load('WholeExcelEmbeddings.npy')
        ref_kr_sentences_whole = pd.Series(list(translation_dict_whole.keys()), dtype=str)
        faiss.normalize_L2(ref_kr_embeddings_whole)
        index_whole = faiss.IndexFlatIP(ref_kr_embeddings_whole.shape[1])
        index_whole.add(ref_kr_embeddings_whole)
        whole_mode_available = True

    faiss_threshold = float(threshold)

    tracker.update(15.0, "Starting translation process...")
    tracker.log_milestone(f"Processing {len(selections)} file(s) with threshold {faiss_threshold}")

    # Track progress across files
    total_files = len(selections)
    current_file_idx = 0

    for excel_file_path, file_selections in selections.items():
        current_file_idx += 1
        filename = os.path.basename(excel_file_path)

        # Update progress for this file (15% to 85% range)
        file_progress = 15.0 + (70.0 * (current_file_idx - 1) / total_files)
        tracker.update(file_progress, f"Processing file {current_file_idx}/{total_files}: {filename}")
        temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
        shutil.copy2(excel_file_path, temp_file_path)

        temp_wb = openpyxl.load_workbook(temp_file_path, data_only=False)

        for sheet_name, columns in file_selections.items():
            sheet = temp_wb[sheet_name]
            col_to_translate = openpyxl.utils.column_index_from_string(columns['kr_column'])
            write_col_index = openpyxl.utils.column_index_from_string(columns['trans_column'])
            print(f"Translating column {columns['kr_column']} in tab '{sheet_name}' and writing to column {columns['trans_column']}", file=sys.stderr)

            total_rows = sheet.max_row
            tracker.log_milestone(f"Translating sheet '{sheet_name}' with {total_rows} rows")

            for idx in range(1, total_rows + 1):
                # Update progress every 50 rows for smooth UI without spam
                if idx % 50 == 0 or idx == 1 or idx == total_rows:
                    row_progress = idx / total_rows
                    current_progress = file_progress + (70.0 / total_files) * row_progress
                    tracker.update(
                        current_progress,
                        f"{filename} | Sheet '{sheet_name}' | Row {idx}/{total_rows}"
                    )
                cell = sheet.cell(row=idx, column=col_to_translate)
                if not isinstance(cell.value, str):
                    continue

                cleaned_input_text = clean_text(cell.value)

                # First attempt whole match
                if whole_mode_available:
                    sentence_embeddings = model.encode([cleaned_input_text])
                    faiss.normalize_L2(sentence_embeddings)
                    distances, indices = index_whole.search(sentence_embeddings, 1)
                    best_match_index = indices[0][0]
                    distance_score = distances[0][0]
                    if distance_score >= faiss_threshold and best_match_index < len(ref_kr_sentences_whole):
                        best_match_korean = ref_kr_sentences_whole.iloc[best_match_index]
                        translated_text = translation_dict_whole.get(best_match_korean, "")
                    else:
                        translated_text = ""
                else:
                    translated_text = ""

                if not translated_text:
                    # Fallback to split match
                    lines = cleaned_input_text.split('\n')
                    translated_lines = []
                    for line in lines:
                        if not line.strip():
                            continue
                        sentence_embeddings = model.encode([line])
                        faiss.normalize_L2(sentence_embeddings)
                        distances, indices = index_split.search(sentence_embeddings, 1)
                        best_match_index = indices[0][0]
                        distance_score = distances[0][0]
                        if distance_score >= faiss_threshold and best_match_index < len(ref_kr_sentences_split):
                            best_match_korean = ref_kr_sentences_split.iloc[best_match_index]
                            translated_line = translation_dict_split.get(best_match_korean, "")
                            if translated_line.strip():
                                translated_lines.append(translated_line)
                    translated_text = '\n'.join(translated_lines) if translated_lines else ''

                # Clean and write
                cleaned_translated_text = clean_text(translated_text)
                if cleaned_translated_text:
                    write_cell = sheet.cell(row=idx, column=write_col_index)
                    write_cell.value = cleaned_translated_text

                print(f"Processed {idx}/{total_rows} rows in tab '{sheet_name}'", file=sys.stderr)

        # Save to organized output directory
        tracker.update(85.0 + (5.0 * current_file_idx / total_files), f"Saving {filename}...")

        output_dir = config.get_output_directory(app_name="xlstransfer")
        original_filename = os.path.basename(excel_file_path)
        output_filename = f"{os.path.splitext(original_filename)[0]}_translated.xlsx"
        output_file_path = str(output_dir / output_filename)

        temp_wb.save(output_file_path)
        temp_wb.close()
        os.remove(temp_file_path)
        print(f"Translation completed for {excel_file_path}!", file=sys.stderr)
        print(f"Output saved to: {output_file_path}", file=sys.stderr)

        tracker.log_milestone(f"Saved: {output_file_path}")

    tracker.update(95.0, "Translation complete! Finalizing...")
    tracker.log_milestone(f"Successfully processed {total_files} file(s)")

    return {
        "success": True,
        "message": "Translation completed for all files!",
        "output_dir": str(output_dir)
    }


def check_newlines(selections):
    """
    Check newline consistency
    Exact replica of check_newlines from original (lines 782-865)
    """
    flagged_rows = {}

    for excel_file_path, file_selections in selections.items():
        temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
        shutil.copy2(excel_file_path, temp_file_path)

        wb = openpyxl.load_workbook(temp_file_path)
        file_name = os.path.basename(excel_file_path)
        flagged_rows[file_name] = {}

        for sheet_name, columns in file_selections.items():
            sheet = wb[sheet_name]
            col1 = openpyxl.utils.column_index_from_string(columns['kr_column'])
            col2 = openpyxl.utils.column_index_from_string(columns['trans_column'])

            flagged_rows[file_name][sheet_name] = []

            for idx in range(1, sheet.max_row + 1):
                cell1 = sheet.cell(row=idx, column=col1).value
                cell2 = sheet.cell(row=idx, column=col2).value

                if cell1 is None or cell2 is None:
                    continue

                newlines1 = str(cell1).count('\n')
                newlines2 = str(cell2).count('\n')
                if newlines1 != newlines2:
                    flagged_rows[file_name][sheet_name].append(idx)

        wb.close()
        os.remove(temp_file_path)

    # Generate report
    if any(sheet for file in flagged_rows.values() for sheet in file.values()):
        report_path = "newline_mismatch_report.txt"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("Newline Mismatch Report:\n\n")
            for file_name, sheets in flagged_rows.items():
                if any(sheets.values()):
                    f.write(f"File: {file_name}\n")
                    for sheet_name, rows in sheets.items():
                        if rows:
                            f.write(f"  Tab: {sheet_name}\n")
                            for row in rows:
                                f.write(f"    Row: {row}\n")
                    f.write("\n")

        return {
            "success": True,
            "report_file": report_path,
            "message": f"Newline check complete. Report saved to {report_path}"
        }
    else:
        return {
            "success": True,
            "message": "No rows with different number of newlines found."
        }


def combine_excel(selections):
    """
    Combine Excel files
    Exact replica of combine_excel_files from original (lines 869-941)
    """
    combined_wb = openpyxl.Workbook()
    combined_sheet = combined_wb.active
    combined_sheet.title = "Combined Data"
    row_index = 1

    for excel_file_path, file_selections in selections.items():
        temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
        shutil.copy2(excel_file_path, temp_file_path)

        source_wb = openpyxl.load_workbook(temp_file_path, data_only=False)

        for sheet_name, columns in file_selections.items():
            sheet = source_wb[sheet_name]
            selected_columns = [openpyxl.utils.column_index_from_string(col) for col in columns.values()]

            for row_num in range(1, sheet.max_row + 1):
                if all(sheet.cell(row=row_num, column=col).value is None for col in selected_columns):
                    continue  # Skip empty rows

                for dest_col, src_col in enumerate(selected_columns, start=1):
                    src_cell = sheet.cell(row=row_num, column=src_col)
                    dest_cell = combined_sheet.cell(row=row_index, column=dest_col)

                    # Clean the cell value before copying
                    cleaned_value = clean_text(src_cell.value)
                    dest_cell.value = cleaned_value

                    # Copy cell format
                    if src_cell.has_style:
                        dest_cell.font = copy(src_cell.font)
                        dest_cell.border = copy(src_cell.border)
                        dest_cell.fill = copy(src_cell.fill)
                        dest_cell.number_format = copy(src_cell.number_format)
                        dest_cell.protection = copy(src_cell.protection)
                        dest_cell.alignment = copy(src_cell.alignment)

                row_index += 1

        source_wb.close()
        os.remove(temp_file_path)

    # Save combined file
    base_name = os.path.splitext(os.path.basename(list(selections.keys())[0]))[0]
    save_path = f"{os.path.dirname(list(selections.keys())[0])}/{base_name}_combined.xlsx"
    combined_wb.save(save_path)

    return {
        "success": True,
        "output_file": save_path,
        "message": f"Combined Excel file saved as {save_path}"
    }


def newline_auto_adapt(selections):
    """
    Newline auto adaptation
    Exact replica of newline_auto_adapt from original (lines 946-1098)
    """
    def preprocess_text(text):
        if not isinstance(text, str):
            return text
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        return '\n'.join(lines)

    def add_newlines(text, target_newlines):
        words = text.split()
        if not words:
            return text

        target_lines = target_newlines + 1
        avg_words_per_line = len(words) / target_lines

        punctuations = '.?!,'
        new_lines = []
        current_line = []

        for i, word in enumerate(words):
            current_line.append(word)

            if len(current_line) >= avg_words_per_line or i == len(words) - 1:
                break_index = -1
                for j, w in enumerate(reversed(current_line)):
                    if any(p in w for p in punctuations):
                        break_index = len(current_line) - j - 1
                        break

                if break_index != -1:
                    new_lines.append(' '.join(current_line[:break_index+1]))
                    current_line = current_line[break_index+1:]
                else:
                    new_lines.append(' '.join(current_line))
                    current_line = []

        if current_line:
            new_lines[-1] += ' ' + ' '.join(current_line)

        while len(new_lines) > target_lines:
            shortest_line_index = min(range(len(new_lines)), key=lambda i: len(new_lines[i]))
            if shortest_line_index > 0:
                new_lines[shortest_line_index-1] += ' ' + new_lines.pop(shortest_line_index)
            else:
                new_lines[0] += ' ' + new_lines.pop(1)

        return '\n'.join(new_lines)

    def remove_newlines(text, target_newlines):
        text = text.replace('\n', ' ')
        return add_newlines(text, target_newlines)

    def process_text(kr_text, fr_text):
        if not isinstance(kr_text, str) or not isinstance(fr_text, str):
            return fr_text

        kr_text = clean_text(kr_text)
        fr_text = clean_text(fr_text)

        kr_lines = kr_text.split('\n')
        fr_lines = fr_text.split('\n')

        kr_newlines = len(kr_lines) - 1
        fr_newlines = len(fr_lines) - 1

        if kr_newlines == fr_newlines:
            return fr_text

        if kr_newlines == 0:
            return fr_text.replace('\n', ' ').replace('  ', ' ')

        if kr_newlines > fr_newlines:
            return add_newlines(fr_text, kr_newlines)

        if kr_newlines < fr_newlines:
            return remove_newlines(fr_text, kr_newlines)

        return fr_text

    for excel_file_path, file_selections in selections.items():
        temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
        shutil.copy2(excel_file_path, temp_file_path)

        source_wb = openpyxl.load_workbook(temp_file_path, data_only=False)

        for sheet_name, columns in file_selections.items():
            sheet = source_wb[sheet_name]
            kr_col = openpyxl.utils.column_index_from_string(columns['kr_column'])
            fr_col = openpyxl.utils.column_index_from_string(columns['trans_column'])

            for idx in range(1, sheet.max_row + 1):
                kr_cell = sheet.cell(row=idx, column=kr_col)
                fr_cell = sheet.cell(row=idx, column=fr_col)

                if kr_cell.value and fr_cell.value:
                    kr_value = preprocess_text(clean_text(str(kr_cell.value)))
                    fr_value = preprocess_text(clean_text(str(fr_cell.value)))

                    kr_cell.value = kr_value
                    fr_cell.value = fr_value

                    new_fr_value = process_text(kr_value, fr_value)

                    if new_fr_value != fr_value:
                        fr_cell.value = new_fr_value

        output_file_path = f"{os.path.splitext(excel_file_path)[0]}_adapted.xlsx"
        source_wb.save(output_file_path)
        source_wb.close()
        os.remove(temp_file_path)

    return {
        "success": True,
        "message": "Newline adaptation completed for all files!"
    }


def main():
    try:
        if len(sys.argv) < 4:
            result = {
                "success": False,
                "error": "Usage: process_operation.py <operation_type> <selections_json> <threshold>"
            }
            print(json.dumps(result))
            sys.exit(1)

        operation_type = sys.argv[1]
        selections_json = sys.argv[2]
        threshold = sys.argv[3] if len(sys.argv) > 3 else "0.99"

        # Parse selections
        selections = json.loads(selections_json)

        # Execute operation
        if operation_type == "create_dictionary":
            result = create_dictionary(selections)
        elif operation_type == "translate_excel":
            result = translate_excel(selections, threshold)
        elif operation_type == "check_newlines":
            result = check_newlines(selections)
        elif operation_type == "combine_excel":
            result = combine_excel(selections)
        elif operation_type == "newline_auto_adapt":
            result = newline_auto_adapt(selections)
        else:
            result = {
                "success": False,
                "error": f"Unknown operation type: {operation_type}"
            }

        print(json.dumps(result))

        if not result.get("success"):
            sys.exit(1)

    except Exception as e:
        import traceback
        result = {
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }
        print(json.dumps(result))
        sys.exit(1)


if __name__ == "__main__":
    main()
