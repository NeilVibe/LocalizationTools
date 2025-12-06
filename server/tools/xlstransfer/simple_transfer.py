#!/usr/bin/env python3
"""
Simple Excel Transfer
Exact replica of simple_excel_transfer from original XLSTransfer0225.py (lines 1110-1372)

This module transfers data between Excel files by mapping values from
source columns to destination columns based on matching file identifiers.
"""

import os
import tempfile
import shutil
import random
from typing import List, Dict, Any, Tuple
from pathlib import Path
from loguru import logger

try:
    import openpyxl
    import pandas as pd
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - simple transfer disabled")


def clean_text(text: Any) -> str:
    """Clean text value for transfer (matches monolith clean_text)."""
    if text is None:
        return ""
    if isinstance(text, (int, float)):
        return str(text)
    return str(text).strip()


def convert_cell_value(value: Any) -> Any:
    """Convert cell value for writing (matches monolith convert_cell_value)."""
    if value is None:
        return None
    if isinstance(value, str):
        # Try to convert numeric strings
        try:
            if '.' in value:
                return float(value)
            return int(value)
        except ValueError:
            return value
    return value


def create_temp_excel(original_file: str, suffix: str) -> str:
    """
    Create temp file in system's temp directory with unique suffix.
    Matches monolith lines 1102-1108.
    """
    temp_dir = tempfile.gettempdir()
    temp_name = f"xlstransfer_temp_{suffix}_{random.randint(1000, 9999)}_{os.path.basename(original_file)}"
    temp_path = os.path.join(temp_dir, temp_name)
    shutil.copy2(original_file, temp_path)
    return temp_path


def analyze_files(source_path: str, dest_path: str) -> Dict[str, Any]:
    """
    Analyze source and destination Excel files.
    Returns sheet names for both files.

    Args:
        source_path: Path to source Excel file
        dest_path: Path to destination Excel file

    Returns:
        Dictionary with source_sheets, dest_sheets, and temp file paths
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not available")

    temp_files = []

    try:
        # Create temp copies
        temp_source = create_temp_excel(source_path, "source")
        temp_files.append(temp_source)

        temp_dest = create_temp_excel(dest_path, "dest")
        temp_files.append(temp_dest)

        # Get sheet names
        source_xls = pd.ExcelFile(temp_source)
        dest_xls = pd.ExcelFile(temp_dest)

        source_sheets = source_xls.sheet_names
        dest_sheets = dest_xls.sheet_names

        source_xls.close()
        dest_xls.close()

        logger.info(f"Analyzed files: source={len(source_sheets)} sheets, dest={len(dest_sheets)} sheets")

        return {
            "success": True,
            "source_file": source_path,
            "dest_file": dest_path,
            "temp_source": temp_source,
            "temp_dest": temp_dest,
            "source_sheets": source_sheets,
            "dest_sheets": dest_sheets
        }

    except Exception as e:
        # Clean up temp files on error
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
        raise e


def validate_transfer_settings(settings_list: List[Dict[str, str]]) -> Tuple[bool, str]:
    """
    Validate transfer settings.
    Matches monolith lines 1267-1289.

    Args:
        settings_list: List of transfer setting dictionaries

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not settings_list:
        return False, "At least one transfer must be configured."

    for settings in settings_list:
        # Check all fields filled
        required_fields = ['source_tab', 'source_file_col', 'source_note_col',
                          'dest_tab', 'dest_file_col', 'dest_note_col']
        if not all(settings.get(f) for f in required_fields):
            return False, "All fields must be filled for each transfer."

        # Validate column letters
        col_fields = ['source_file_col', 'source_note_col', 'dest_file_col', 'dest_note_col']
        for field in col_fields:
            col = settings.get(field, '')
            if not col.isalpha():
                return False, f"Invalid column letter: {col}"

    # Check for conflicts (same dest sheet + dest column)
    seen_combinations = set()
    for settings in settings_list:
        dest_combo = (settings['dest_tab'], settings['dest_note_col'].upper())
        if dest_combo in seen_combinations:
            return False, "Conflict detected: Cannot write to the same destination column in the same sheet."
        seen_combinations.add(dest_combo)

    return True, ""


def execute_transfer(
    temp_source_file: str,
    temp_dest_file: str,
    dest_file: str,
    settings_list: List[Dict[str, str]],
    cleanup_temps: bool = True
) -> Dict[str, Any]:
    """
    Execute the transfer operations.
    Matches monolith lines 1309-1361.

    Args:
        temp_source_file: Path to temp source Excel file
        temp_dest_file: Path to temp destination Excel file
        dest_file: Original destination file path (for output naming)
        settings_list: List of transfer settings
        cleanup_temps: Whether to clean up temp files after

    Returns:
        Dictionary with success status and output file path
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not available")

    temp_files = [temp_source_file, temp_dest_file]

    try:
        source_wb = openpyxl.load_workbook(temp_source_file, data_only=False)
        dest_wb = openpyxl.load_workbook(temp_dest_file, data_only=False)

        total_transferred = 0

        for settings in settings_list:
            source_ws = source_wb[settings['source_tab']]
            dest_ws = dest_wb[settings['dest_tab']]

            source_file_col = settings['source_file_col'].upper()
            source_note_col = settings['source_note_col'].upper()
            dest_file_col = settings['dest_file_col'].upper()
            dest_note_col = settings['dest_note_col'].upper()

            # Create mapping from source (monolith lines 1316-1325)
            file_map = {}
            for row in range(1, source_ws.max_row + 1):
                file_name = source_ws.cell(
                    row=row,
                    column=openpyxl.utils.column_index_from_string(source_file_col)
                ).value
                note = source_ws.cell(
                    row=row,
                    column=openpyxl.utils.column_index_from_string(source_note_col)
                ).value

                if file_name:
                    cleaned_note = clean_text(note) if note is not None else None
                    converted_note = convert_cell_value(cleaned_note)
                    file_map[file_name] = converted_note

            # Transfer data (monolith lines 1326-1332)
            transfer_count = 0
            for row in range(1, dest_ws.max_row + 1):
                dest_file_name = dest_ws.cell(
                    row=row,
                    column=openpyxl.utils.column_index_from_string(dest_file_col)
                ).value

                if dest_file_name in file_map:
                    value = file_map[dest_file_name]
                    converted_value = convert_cell_value(value)
                    dest_ws.cell(
                        row=row,
                        column=openpyxl.utils.column_index_from_string(dest_note_col)
                    ).value = converted_value
                    transfer_count += 1

            logger.info(f"Transfer {settings['source_tab']}->{settings['dest_tab']}: {transfer_count} rows")
            total_transferred += transfer_count

        # Save output (monolith line 1333)
        output_file_path = f"{os.path.splitext(dest_file)[0]}_transferred.xlsx"
        dest_wb.save(output_file_path)

        source_wb.close()
        dest_wb.close()

        logger.success(f"Transfer complete: {total_transferred} total rows, saved to {output_file_path}")

        # Clean up temp files
        if cleanup_temps:
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass

        return {
            "success": True,
            "output_file": output_file_path,
            "transfers_count": len(settings_list),
            "total_rows_transferred": total_transferred
        }

    except Exception as e:
        # Clean up on error
        try:
            source_wb.close()
            dest_wb.close()
        except:
            pass

        if cleanup_temps:
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass

        logger.error(f"Transfer failed: {e}")
        raise e


def simple_excel_transfer_api(
    source_path: str,
    dest_path: str,
    settings_list: List[Dict[str, str]]
) -> Dict[str, Any]:
    """
    API entry point for simple Excel transfer.
    Combines analyze and execute in a single call.

    Args:
        source_path: Path to source Excel file
        dest_path: Path to destination Excel file
        settings_list: List of transfer settings, each with:
            - source_tab: Source sheet name
            - source_file_col: Column with file identifiers (e.g., 'A')
            - source_note_col: Column with values to transfer (e.g., 'B')
            - dest_tab: Destination sheet name
            - dest_file_col: Column with file identifiers (e.g., 'A')
            - dest_note_col: Column to write values to (e.g., 'B')

    Returns:
        Dictionary with success status and output file
    """
    # Validate settings
    is_valid, error = validate_transfer_settings(settings_list)
    if not is_valid:
        return {"success": False, "error": error}

    # Analyze and create temp files
    analysis = analyze_files(source_path, dest_path)

    # Execute transfer
    result = execute_transfer(
        temp_source_file=analysis['temp_source'],
        temp_dest_file=analysis['temp_dest'],
        dest_file=dest_path,
        settings_list=settings_list,
        cleanup_temps=True
    )

    return result


if __name__ == "__main__":
    # CLI mode for testing
    import sys
    import json

    try:
        if len(sys.argv) < 4:
            print(json.dumps({
                "success": False,
                "error": "Usage: python simple_transfer.py <source.xlsx> <dest.xlsx> '<settings_json>'"
            }))
            sys.exit(1)

        source = sys.argv[1]
        dest = sys.argv[2]
        settings = json.loads(sys.argv[3])

        result = simple_excel_transfer_api(source, dest, settings)
        print(json.dumps(result))
        sys.exit(0 if result.get("success") else 1)

    except Exception as e:
        import traceback
        result = {
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }
        print(json.dumps(result))
        sys.exit(1)
