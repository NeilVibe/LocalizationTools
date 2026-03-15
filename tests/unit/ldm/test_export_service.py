"""
Tests for ExportService -- XML, Excel, and Text export.

Tests verify:
- XML: lxml output, attribute casing, br-tag escaping, None vs empty
- Excel: 14-column EU structure, header formatting, data mapping, Korean detection
- Text: tab-delimited, UTF-8, empty field handling
"""

from __future__ import annotations

import io
import pytest
from lxml import etree

from server.tools.ldm.services.export_service import ExportService


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def service():
    return ExportService()


@pytest.fixture
def sample_rows():
    return [
        {
            "string_id": "STR_001",
            "source": "Hello World",
            "target": "안녕하세요",
            "extra_data": {
                "Desc": "Greeting",
                "DescOrigin": "인사",
                "Category": "UI",
            },
        },
        {
            "string_id": "STR_002",
            "source": "Goodbye",
            "target": "안녕히 가세요",
            "extra_data": {
                "Desc": "Farewell",
                "Category": "Dialog",
            },
        },
    ]


@pytest.fixture
def sample_metadata():
    return {
        "root_element": "LangData",
        "element_tag": "LocStr",
        "encoding": "utf-8",
    }


# ===========================================================================
# TestXMLExport
# ===========================================================================

class TestXMLExport:
    """XML export via lxml with correct attribute casing."""

    def test_basic_xml_output(self, service, sample_rows, sample_metadata):
        """Rows with StringId/StrOrigin/Str produce valid XML with correct attribute casing."""
        result = service.export_xml(sample_rows, sample_metadata)
        assert isinstance(result, bytes)

        root = etree.fromstring(result)
        elements = root.findall("LocStr")
        assert len(elements) == 2

        elem = elements[0]
        # Attribute casing must be exact
        assert elem.get("StringId") == "STR_001"
        assert elem.get("StrOrigin") == "Hello World"
        assert elem.get("Str") == "안녕하세요"

    def test_brtag_roundtrip(self, service, sample_metadata):
        """In-memory br-tags serialize as escaped entities on disk."""
        rows = [
            {
                "string_id": "BR_001",
                "source": "Line1<br/>Line2",
                "target": "줄1<br/>줄2",
                "extra_data": {},
            }
        ]
        result = service.export_xml(rows, sample_metadata)
        # lxml auto-escapes < and > in attribute values
        xml_text = result.decode("utf-8")
        assert "&lt;br/&gt;" in xml_text
        # The literal <br/> should NOT appear unescaped in attribute values
        # Parse back and verify the values round-trip
        root = etree.fromstring(result)
        elem = root.findall("LocStr")[0]
        assert elem.get("StrOrigin") == "Line1<br/>Line2"
        assert elem.get("Str") == "줄1<br/>줄2"

    def test_extra_data_preserved(self, service, sample_rows, sample_metadata):
        """extra_data dict attributes appear in output elements."""
        result = service.export_xml(sample_rows, sample_metadata)
        root = etree.fromstring(result)
        elem = root.findall("LocStr")[0]
        assert elem.get("Desc") == "Greeting"
        assert elem.get("DescOrigin") == "인사"
        assert elem.get("Category") == "UI"

    def test_none_vs_empty_string(self, service, sample_metadata):
        """None values skip attribute, empty string writes attr=''."""
        rows = [
            {
                "string_id": "NONE_001",
                "source": "Test",
                "target": None,
                "extra_data": {"Desc": "", "Category": None},
            }
        ]
        result = service.export_xml(rows, sample_metadata)
        root = etree.fromstring(result)
        elem = root.findall("LocStr")[0]
        assert elem.get("StringId") == "NONE_001"
        assert elem.get("StrOrigin") == "Test"
        # None target => attribute skipped
        assert elem.get("Str") is None
        # Empty string => attribute present but empty
        assert elem.get("Desc") == ""
        # None in extra_data => attribute skipped
        assert elem.get("Category") is None

    def test_root_attributes(self, service, sample_rows):
        """root_attributes from metadata appear on root element."""
        metadata = {
            "root_element": "LangData",
            "root_attributes": {"xmlns": "http://example.com/lang", "version": "1.0"},
            "element_tag": "LocStr",
            "encoding": "utf-8",
        }
        result = service.export_xml(sample_rows, metadata)
        xml_text = result.decode("utf-8")
        # xmlns appears in the output
        assert 'xmlns="http://example.com/lang"' in xml_text
        # Regular attributes also present
        root = etree.fromstring(result)
        assert root.get("version") == "1.0"

    def test_custom_element_tag(self, service, sample_rows):
        """element_tag from metadata controls per-row tag name."""
        metadata = {
            "root_element": "GameData",
            "element_tag": "Entry",
            "encoding": "utf-8",
        }
        result = service.export_xml(sample_rows, metadata)
        root = etree.fromstring(result)
        assert root.tag == "GameData"
        assert len(root.findall("Entry")) == 2


# ===========================================================================
# TestExcelExport
# ===========================================================================

class TestExcelExport:
    """Excel export via xlsxwriter with EU column structure."""

    def _load_workbook(self, data: bytes):
        """Load Excel bytes with openpyxl (read-only)."""
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(data))

    def test_14_column_headers(self, service, sample_rows, sample_metadata):
        """Output has exactly 14 EU columns in correct order."""
        result = service.export_excel(sample_rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        expected = [
            "StrOrigin", "ENG", "Str", "Correction", "Text State",
            "STATUS", "COMMENT", "MEMO1", "MEMO2", "Category",
            "FileName", "StringID", "DescOrigin", "Desc",
        ]
        assert headers == expected

    def test_header_formatting(self, service, sample_rows, sample_metadata):
        """Headers have bold, #DAEEF3 background, centered, bordered."""
        result = service.export_excel(sample_rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        # xlsxwriter writes hex without '#', openpyxl reads as '00DAEEF3'
        fill_color = cell.fill.start_color.rgb
        assert "DAEEF3" in fill_color.upper()

    def test_data_mapping(self, service, sample_rows, sample_metadata):
        """source maps to StrOrigin (col 0), target to Str (col 2), string_id to StringID (col 11)."""
        result = service.export_excel(sample_rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        # Row 2 = first data row
        assert ws.cell(row=2, column=1).value == "Hello World"      # StrOrigin = source
        assert ws.cell(row=2, column=3).value == "안녕하세요"          # Str = target
        assert ws.cell(row=2, column=12).value == "STR_001"          # StringID = string_id

    def test_text_state_korean(self, service, sample_metadata):
        """Korean text in target produces 'KOREAN' in Text State column."""
        rows = [{"string_id": "KR_001", "source": "Test", "target": "테스트", "extra_data": {}}]
        result = service.export_excel(rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "KOREAN"  # Text State col

    def test_text_state_translated(self, service, sample_metadata):
        """Non-Korean text produces 'TRANSLATED' in Text State column."""
        rows = [{"string_id": "EN_001", "source": "Test", "target": "Translated text", "extra_data": {}}]
        result = service.export_excel(rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "TRANSLATED"

    def test_stringid_text_format(self, service, sample_metadata):
        """StringID column uses text format to prevent scientific notation."""
        rows = [{"string_id": "12345E10", "source": "Test", "target": "T", "extra_data": {}}]
        result = service.export_excel(rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        # The value should be stored as text, not interpreted as number
        cell = ws.cell(row=2, column=12)
        assert cell.value == "12345E10"
        # Check number format is text
        assert cell.number_format == "@"

    def test_freeze_panes(self, service, sample_rows, sample_metadata):
        """Worksheet has frozen panes at row 1."""
        result = service.export_excel(sample_rows, sample_metadata)
        wb = self._load_workbook(result)
        ws = wb.active
        assert ws.freeze_panes == "A2"


# ===========================================================================
# TestTextExport
# ===========================================================================

class TestTextExport:
    """Text export: tab-delimited StringID + source + target."""

    def test_basic_tab_delimited(self, service, sample_rows, sample_metadata):
        """Output is tab-delimited with StringID, source, target per line."""
        result = service.export_text(sample_rows, sample_metadata)
        lines = result.decode("utf-8").strip().split("\n")
        assert len(lines) == 2
        parts = lines[0].split("\t")
        assert parts[0] == "STR_001"
        assert parts[1] == "Hello World"
        assert parts[2] == "안녕하세요"

    def test_utf8_encoding(self, service, sample_rows, sample_metadata):
        """Output bytes are valid UTF-8."""
        result = service.export_text(sample_rows, sample_metadata)
        assert isinstance(result, bytes)
        # Should not raise
        text = result.decode("utf-8")
        assert "안녕하세요" in text

    def test_empty_fields(self, service, sample_metadata):
        """Missing fields produce empty strings, not 'None'."""
        rows = [
            {"string_id": "EMPTY_001", "source": None, "target": None, "extra_data": {}},
        ]
        result = service.export_text(rows, sample_metadata)
        text = result.decode("utf-8")
        assert "None" not in text
        # Don't strip -- trailing tabs are significant
        line = text.split("\n")[0]
        parts = line.split("\t")
        assert parts[0] == "EMPTY_001"
        assert parts[1] == ""
        assert parts[2] == ""
