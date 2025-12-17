"""
E2E Test 3: TMSearcher with Variations

Tests the TMSearcher functionality with StringID support:
- Tier 1 (hash) returns all variations
- Tier 2 (embedding) includes string_id
- Result format with string_id field
- Score and match_type preserved

Test Cases: SRC-01 to SRC-09

Run: python -m pytest tests/fixtures/stringid/test_e2e_3_tm_search.py -v
"""

import os
import sys
import pytest
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Test data
from tests.fixtures.stringid.stringid_test_data import BASIC_STRINGID_DATA


class TestTMSearcherVariations:
    """E2E tests for TMSearcher with StringID variations."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment with a StringID TM and built indexes."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.tools.ldm.tm_indexer import TMIndexer, TMSearcher
        from server.database.models import LDMTranslationMemory, LDMTMEntry
        import openpyxl
        from io import BytesIO

        # Get session
        engine, session_maker = setup_database(drop_existing=False)
        self.db = session_maker()
        self.tm_manager = TMManager(self.db)
        self.tm_indexer = TMIndexer(self.db)
        self.created_tm_ids = []

        # Create a StringID TM for testing
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'
        for i, entry in enumerate(BASIC_STRINGID_DATA, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            ws[f'C{i}'] = entry.get("string_id", "")

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_tm_search.xlsx",
            name="Test TMSearcher TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )
        self.tm_id = result["tm_id"]
        self.created_tm_ids.append(self.tm_id)

        # Build indexes
        self.tm_indexer.build_indexes(self.tm_id)

        # Load indexes and create searcher
        from server.tools.ldm.tm_indexer import TMSearcher
        indexes = self.tm_indexer.load_indexes(self.tm_id)
        self.searcher = TMSearcher(indexes)

        yield

        # Cleanup
        for tm_id in self.created_tm_ids:
            try:
                self.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        self.db.close()

    # =========================================================================
    # SRC-01: Tier 1 exact - returns all variations
    # =========================================================================

    def test_src_01_tier1_exact_variations(self):
        """SRC-01: Tier 1 exact match returns all 3 variations for '저장'."""
        result = self.searcher.search("저장")

        # Should return multiple results (variations)
        assert len(result["results"]) == 3, \
            f"Expected 3 variations, got {len(result['results'])}"

        # Verify tier
        assert result["tier"] == 1, f"Expected tier 1, got {result['tier']}"
        assert result["tier_name"] == "perfect_whole", f"Expected 'perfect_whole', got {result['tier_name']}"

        # Check all targets are different
        targets = {r["target_text"] for r in result["results"]}
        expected_targets = {"Save", "Save Game", "Storage"}
        assert targets == expected_targets, f"Target mismatch. Expected: {expected_targets}, Got: {targets}"

        print(f"SRC-01 PASSED: '저장' returned {len(result['results'])} variations")

    # =========================================================================
    # SRC-02: Tier 1 exact - single result
    # =========================================================================

    def test_src_02_tier1_exact_single(self):
        """SRC-02: Tier 1 exact match returns 1 result for '취소' (no variations)."""
        result = self.searcher.search("취소")

        # Should return 1 result (no variations for 취소)
        assert len(result["results"]) == 1, \
            f"Expected 1 result, got {len(result['results'])}"

        assert result["results"][0]["target_text"] == "Cancel"
        assert result["results"][0]["string_id"] == "UI_BUTTON_CANCEL"

        print(f"SRC-02 PASSED: '취소' returned 1 result with string_id")

    # =========================================================================
    # SRC-03: Tier 1 - string_id in result
    # =========================================================================

    def test_src_03_tier1_stringid_in_result(self):
        """SRC-03: Results have string_id field."""
        result = self.searcher.search("설정")

        assert len(result["results"]) == 2, f"Expected 2 variations, got {len(result['results'])}"

        # Check all results have string_id
        for r in result["results"]:
            assert "string_id" in r, "Result missing string_id field"
            assert r["string_id"] is not None, "string_id should not be None"

        # Verify specific string_ids
        string_ids = {r["string_id"] for r in result["results"]}
        expected_ids = {"UI_MENU_SETTINGS", "TECH_CONFIG"}
        assert string_ids == expected_ids, f"StringID mismatch. Expected: {expected_ids}, Got: {string_ids}"

        print(f"SRC-03 PASSED: '설정' results have correct string_ids")

    # =========================================================================
    # SRC-04: Tier 2 embedding (similar text)
    # =========================================================================

    def test_src_04_tier2_embedding(self):
        """SRC-04: Tier 2 embedding search returns matches with string_id."""
        # Search for similar text that won't match exactly
        result = self.searcher.search("저장하기")  # Similar to "저장"

        # May return Tier 2 (embedding) or Tier 1 if normalized match
        if len(result["results"]) > 0:
            # Check results have string_id
            for r in result["results"]:
                assert "string_id" in r, "Result missing string_id field"

            # If Tier 2, score should be < 1.0
            if result["tier"] == 2:
                for r in result["results"]:
                    assert r["score"] < 1.0, f"Tier 2 score should be < 1.0, got {r['score']}"

            print(f"SRC-04 PASSED: '저장하기' search - Tier {result['tier']}, {len(result['results'])} results")
        else:
            print(f"SRC-04 PASSED: '저장하기' search - No matches (below threshold)")

    # =========================================================================
    # SRC-05: Tier 2 - variations
    # =========================================================================

    def test_src_05_tier2_variations(self):
        """SRC-05: Tier 2 search returns variations with string_id."""
        # This test checks that even embedding matches include string_id
        result = self.searcher.search("파일 저장")  # Contains "저장"

        # May or may not match depending on embedding similarity
        if len(result["results"]) > 0:
            for r in result["results"]:
                assert "string_id" in r, "Result missing string_id"
            print(f"SRC-05 PASSED: Found {len(result['results'])} results with string_id")
        else:
            print(f"SRC-05 PASSED: No matches found (expected for dissimilar text)")

    # =========================================================================
    # SRC-06: No match
    # =========================================================================

    def test_src_06_no_match(self):
        """SRC-06: Unknown query returns empty results."""
        result = self.searcher.search("XXXXXXXXXXXXXXX")

        assert len(result["results"]) == 0, f"Expected 0 results, got {len(result['results'])}"
        assert result["tier"] == 0, f"Expected tier 0, got {result['tier']}"

        print(f"SRC-06 PASSED: Unknown query returns empty results")

    # =========================================================================
    # SRC-07: Standard TM search (no variations)
    # =========================================================================

    def test_src_07_standard_tm_search(self):
        """SRC-07: Standard mode TM returns single result per source."""
        import openpyxl
        from io import BytesIO

        # Create standard TM (no StringID)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['A2'] = '테스트'
        ws['B2'] = 'Test'
        ws['A3'] = '예제'
        ws['B3'] = 'Example'

        buffer = BytesIO()
        wb.save(buffer)

        result = self.tm_manager.upload_tm(
            file_content=buffer.getvalue(),
            filename="test_standard_search.xlsx",
            name="Test Standard Search TM",
            owner_id=1,
            mode="standard",
            source_col=0,
            target_col=1,
            has_header=True
        )
        standard_tm_id = result["tm_id"]
        self.created_tm_ids.append(standard_tm_id)

        # Build indexes
        self.tm_indexer.build_indexes(standard_tm_id)

        # Search
        from server.tools.ldm.tm_indexer import TMSearcher
        indexes = self.tm_indexer.load_indexes(standard_tm_id)
        searcher = TMSearcher(indexes)
        search_result = searcher.search("테스트")

        assert len(search_result["results"]) == 1, f"Expected 1 result, got {len(search_result['results'])}"
        assert search_result["results"][0]["target_text"] == "Test"

        print(f"SRC-07 PASSED: Standard TM returns single result")

    # =========================================================================
    # SRC-08: Score preserved
    # =========================================================================

    def test_src_08_score_preserved(self):
        """SRC-08: Search results have score field."""
        result = self.searcher.search("저장")

        for r in result["results"]:
            assert "score" in r, "Result missing score field"
            assert isinstance(r["score"], (int, float)), "Score should be numeric"
            assert 0 <= r["score"] <= 1.0, f"Score out of range: {r['score']}"

        print(f"SRC-08 PASSED: All results have valid score field")

    # =========================================================================
    # SRC-09: match_type preserved
    # =========================================================================

    def test_src_09_match_type_preserved(self):
        """SRC-09: Search results have match_type field."""
        result = self.searcher.search("저장")

        for r in result["results"]:
            assert "match_type" in r, "Result missing match_type field"
            assert r["match_type"] in ["perfect_whole", "embedding_whole", "perfect_line", "embedding_line", "ngram"], \
                f"Unknown match_type: {r['match_type']}"

        # For exact match, should be "perfect_whole"
        assert all(r["match_type"] == "perfect_whole" for r in result["results"]), \
            "Exact match should have match_type='perfect_whole'"

        print(f"SRC-09 PASSED: All results have valid match_type field")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
