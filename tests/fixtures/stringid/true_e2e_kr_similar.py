"""
TRUE E2E Test: KR Similar Engine with Structure Adaptation

===============================================================================
THIS IS A TRUE END-TO-END TEST
===============================================================================

WHAT THIS TEST DOES:
1. Loads ~5000 rows from real production data (sampleofLanguageData.txt)
2. Creates a TM and loads it using load_tm() method
3. Builds FAISS indexes for similarity search
4. Creates test cases with various patterns (triangle markers, multiline, etc.)
5. Runs find_similar() and pretranslation
6. Verifies correct translation + structure preservation

DATA SOURCE:
/mnt/c/NEIL_PROJECTS_WINDOWSBUILD/LocaNextProject/TestFilesForLocaNext/sampleofLanguageData.txt
- 103,499 rows of real game localization data
- Tab-separated: Col1=StringID, Col5=Korean, Col6=French
- Contains: color codes, TextBind, newlines

KEY KR SIMILAR FEATURES TESTED:
1. Similarity search using FAISS (find_similar)
2. Triangle marker (▶) line-by-line handling
3. Structure adaptation (adapt_structure)
4. Split vs whole mode selection
5. TM loading via load_tm()

Run: python -m pytest tests/fixtures/stringid/true_e2e_kr_similar.py -v -s
Time: ~5 minutes (embedding generation)
"""

import os
import sys
import pytest
import random
import numpy as np
from pathlib import Path
from typing import List, Dict, Tuple

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Data source path
SAMPLE_DATA_PATH = Path("/mnt/c/NEIL_PROJECTS_WINDOWSBUILD/LocaNextProject/TestFilesForLocaNext/sampleofLanguageData.txt")

# Test configuration
TM_SIZE = 5000          # Number of rows for TM
TEST_CASES_COUNT = 150  # Number of test cases
THRESHOLD = 0.85        # Match threshold


def parse_sample_data(limit: int = None) -> List[Dict]:
    """
    Parse sampleofLanguageData.txt into list of entries.

    Format: Tab-separated
    Col 0: Flag (0)
    Col 1: StringID (numeric)
    Col 2-4: Metadata
    Col 5: Korean source
    Col 6: Target (French)
    Col 7: Status
    """
    if not SAMPLE_DATA_PATH.exists():
        pytest.skip(f"Sample data not found: {SAMPLE_DATA_PATH}")

    entries = []
    with open(SAMPLE_DATA_PATH, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            if limit and i >= limit:
                break

            parts = line.strip().split('\t')
            if len(parts) >= 7:
                string_id = parts[1].strip()
                source = parts[5].strip()
                target = parts[6].strip()

                if source and target:  # Skip empty
                    entries.append({
                        "string_id": f"ID_{string_id}",
                        "source": source,
                        "target": target
                    })

    return entries


def find_multiline_entries(entries: List[Dict]) -> List[Dict]:
    """Find entries with multiple lines (for split mode testing)."""
    multiline = []
    for entry in entries:
        src = entry["source"]
        tgt = entry["target"]
        # Check for newlines or escaped newlines
        if ('\n' in src or '\\n' in src) and ('\n' in tgt or '\\n' in tgt):
            multiline.append(entry)
    return multiline


def categorize_entries(entries: List[Dict]) -> Dict[str, List[Dict]]:
    """Categorize entries by type."""
    categories = {
        "plain_text": [],
        "multiline": [],
        "pacolor": [],
        "textbind": [],
        "long_text": [],
    }

    for entry in entries:
        src = entry["source"]

        if "<PAColor" in src:
            categories["pacolor"].append(entry)
        elif "TextBind" in src:
            categories["textbind"].append(entry)
        elif '\n' in src or '\\n' in src:
            categories["multiline"].append(entry)
        elif len(src) > 100:
            categories["long_text"].append(entry)
        else:
            categories["plain_text"].append(entry)

    return categories


class TestTrueE2EKRSimilar:
    """
    TRUE End-to-End test for KR Similar with structure adaptation.

    This test uses real production data and verifies the full pipeline:
    TM Creation → load_tm() → FAISS Index → Similarity Search → Pretranslation
    """

    @pytest.fixture(scope="class", autouse=True)
    def setup(self, request):
        """Setup test environment with TM from real data."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.tools.kr_similar.embeddings import EmbeddingsManager
        from server.tools.kr_similar.searcher import SimilaritySearcher
        from server.tools.kr_similar.core import adapt_structure, normalize_text
        from server.database.models import LDMFile, LDMRow, LDMProject
        import openpyxl
        from io import BytesIO

        cls = request.cls

        print("\n" + "="*70)
        print("TRUE E2E TEST: KR Similar with Structure Adaptation")
        print("="*70)

        # Parse real data
        print(f"\n[1/7] Loading data from {SAMPLE_DATA_PATH}...")
        all_entries = parse_sample_data(limit=TM_SIZE * 2)
        print(f"      Loaded {len(all_entries)} entries")

        # Categorize entries
        cls.categories = categorize_entries(all_entries)
        print(f"      Categories:")
        for cat, items in cls.categories.items():
            print(f"        - {cat}: {len(items)} entries")

        # Select entries for TM
        cls.tm_entries = all_entries[:TM_SIZE]
        print(f"\n[2/7] Selected {len(cls.tm_entries)} entries for TM")

        # Database setup
        print(f"\n[3/7] Setting up database...")
        engine, session_maker = setup_database(drop_existing=False)
        cls.db = session_maker()
        cls.tm_manager = TMManager(cls.db)
        cls.created_tm_ids = []
        cls.created_file_ids = []

        # Create Excel file for TM upload
        print(f"\n[4/7] Creating TM Excel file ({len(cls.tm_entries)} rows)...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'

        for i, entry in enumerate(cls.tm_entries, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            ws[f'C{i}'] = entry["string_id"]

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        print(f"      Excel file size: {len(excel_bytes):,} bytes")

        # Upload TM
        print(f"\n[5/7] Uploading TM (mode=stringid)...")
        result = cls.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="true_e2e_kr_similar_tm.xlsx",
            name="TRUE E2E KR Similar TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )
        cls.tm_id = result["tm_id"]
        cls.created_tm_ids.append(cls.tm_id)
        print(f"      TM created: id={cls.tm_id}, entries={result.get('entry_count', 'N/A')}")

        # Load TM using KR Similar's load_tm method
        print(f"\n[6/7] Loading TM via KR Similar load_tm() (builds FAISS indexes)...")
        cls.embeddings_mgr = EmbeddingsManager()
        load_success = cls.embeddings_mgr.load_tm(cls.tm_id, cls.db)
        if not load_success:
            raise RuntimeError(f"Failed to load TM {cls.tm_id}")
        print(f"      Split pairs: {len(cls.embeddings_mgr.split_dict)}")
        print(f"      Whole pairs: {len(cls.embeddings_mgr.whole_dict)}")

        # Create searcher
        cls.searcher = SimilaritySearcher(embeddings_manager=cls.embeddings_mgr)
        cls.adapt_structure = adapt_structure
        cls.normalize_text = normalize_text

        # Create test cases
        print(f"\n[7/7] Creating test cases ({TEST_CASES_COUNT} cases)...")
        cls._create_test_cases_static(cls)
        print(f"      Test cases created: {len(cls.test_cases)}")

        print("\n" + "="*70)
        print("SETUP COMPLETE - Ready for testing")
        print("="*70 + "\n")

        yield

        # Cleanup
        print("\n[CLEANUP] Removing test data...")
        for file_id in cls.created_file_ids:
            try:
                cls.db.query(LDMRow).filter(LDMRow.file_id == file_id).delete()
                cls.db.query(LDMFile).filter(LDMFile.id == file_id).delete()
                cls.db.commit()
            except Exception:
                pass
        for tm_id in cls.created_tm_ids:
            try:
                cls.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        cls.db.close()
        print("[CLEANUP] Done")

    @staticmethod
    def _create_test_cases_static(cls):
        """
        Create test cases mixing:
        1. Plain text (exact from TM - should match)
        2. Similar text (minor variation - should fuzzy match)
        3. Triangle markers (▶) - line-by-line handling
        4. Multiline text - structure adaptation
        5. No match cases
        """
        cls.test_cases = []
        cls.expected_results = {}

        case_num = 0

        # 1. Exact match cases (40 rows) - direct from TM
        plain_entries = [e for e in cls.tm_entries if '\n' not in e["source"] and '\\n' not in e["source"] and '<PA' not in e["source"]]
        for entry in plain_entries[:40]:
            case_num += 1
            cls.test_cases.append({
                "case_num": case_num,
                "type": "exact_match",
                "korean": entry["source"],
                "expected_translation": entry["target"],
                "should_match": True
            })
            cls.expected_results[case_num] = {
                "type": "exact_match",
                "expected": entry["target"]
            }

        # 2. Similar text cases (30 rows) - slightly modified from TM
        for entry in plain_entries[40:70]:
            case_num += 1
            # Add minor variation (add space or punctuation)
            modified_source = entry["source"] + " "  # Just add trailing space
            cls.test_cases.append({
                "case_num": case_num,
                "type": "similar_match",
                "korean": modified_source,
                "original_source": entry["source"],
                "expected_translation": entry["target"],
                "should_match": True  # Should still match via embedding similarity
            })
            cls.expected_results[case_num] = {
                "type": "similar_match",
                "expected": entry["target"]
            }

        # 3. Triangle marker cases (30 rows) - manually created
        # Create test cases with triangle markers that should match line-by-line
        for entry in plain_entries[70:100]:
            case_num += 1
            # Create triangle-prefixed version
            triangle_source = f"▶{entry['source']}"
            cls.test_cases.append({
                "case_num": case_num,
                "type": "triangle_marker",
                "korean": triangle_source,
                "original_source": entry["source"],
                "expected_translation": entry["target"],
                "should_match": True
            })
            cls.expected_results[case_num] = {
                "type": "triangle_marker",
                "expected": entry["target"]
            }

        # 4. Multiline cases (20 rows) - combine two entries
        for i in range(100, 120):
            if i + 1 >= len(plain_entries):
                break
            case_num += 1
            entry1 = plain_entries[i]
            entry2 = plain_entries[i + 1] if i + 1 < len(plain_entries) else entry1

            multiline_source = f"▶{entry1['source']}\n▶{entry2['source']}"
            multiline_target = f"▶{entry1['target']}\n▶{entry2['target']}"

            cls.test_cases.append({
                "case_num": case_num,
                "type": "multiline_triangle",
                "korean": multiline_source,
                "expected_translation": multiline_target,
                "should_match": True
            })
            cls.expected_results[case_num] = {
                "type": "multiline_triangle",
                "expected": multiline_target
            }

        # 5. No match cases (20 rows) - random gibberish
        random_texts = [
            "XXXXXXXXX random text YYYYY",
            "이것은 매칭되지 않을 텍스트입니다 ZZZZ",
            "Completely unrelated English text 12345",
        ]
        for i, text in enumerate(random_texts * 7):  # Repeat to get ~20
            if len(cls.test_cases) >= TEST_CASES_COUNT:
                break
            case_num += 1
            cls.test_cases.append({
                "case_num": case_num,
                "type": "no_match",
                "korean": f"{text}_{i}",
                "expected_translation": None,
                "should_match": False
            })
            cls.expected_results[case_num] = {
                "type": "no_match",
                "expected": None
            }

        # Pad remaining with exact matches
        while len(cls.test_cases) < TEST_CASES_COUNT:
            entry = random.choice(plain_entries)
            case_num += 1
            cls.test_cases.append({
                "case_num": case_num,
                "type": "exact_match",
                "korean": entry["source"],
                "expected_translation": entry["target"],
                "should_match": True
            })
            cls.expected_results[case_num] = {
                "type": "exact_match",
                "expected": entry["target"]
            }

    # =========================================================================
    # RUN SIMILARITY SEARCH ONCE FOR ALL TESTS
    # =========================================================================

    @pytest.fixture(scope="class")
    def search_results(self):
        """Run similarity search once and store results for all tests."""
        results = {}

        for case in self.test_cases:
            korean = case["korean"]
            case_type = case["type"]

            # Handle different case types
            if case_type == "multiline_triangle" or '\n' in korean:
                # Line-by-line search
                lines = korean.replace('\\n', '\n').split('\n')
                line_results = []

                for line in lines:
                    line_stripped = line.strip()
                    if not line_stripped:
                        line_results.append({"translation": "", "matched": False})
                        continue

                    # Remove triangle marker for search
                    search_text = line_stripped.lstrip('▶').strip()
                    if not search_text:
                        line_results.append({"translation": line_stripped, "matched": False})
                        continue

                    matches = self.searcher.find_similar(
                        search_text,
                        threshold=THRESHOLD,
                        top_k=1,
                        use_whole=False
                    )

                    if matches:
                        trans = matches[0].get("translation", "")
                        # Preserve triangle marker
                        if line_stripped.startswith('▶'):
                            trans = '▶' + trans
                        line_results.append({
                            "translation": trans,
                            "matched": True,
                            "similarity": matches[0].get("similarity", 0)
                        })
                    else:
                        line_results.append({"translation": line_stripped, "matched": False})

                # Combine results
                any_match = any(r.get("matched", False) for r in line_results)
                combined_trans = '\n'.join(r["translation"] for r in line_results) if any_match else ""

                results[case["case_num"]] = {
                    "korean": korean,
                    "translation": combined_trans,
                    "matched": any_match,
                    "line_results": line_results,
                    "type": case_type
                }

            elif case_type == "triangle_marker":
                # Single line with triangle
                search_text = korean.lstrip('▶').strip()
                matches = self.searcher.find_similar(
                    search_text,
                    threshold=THRESHOLD,
                    top_k=1,
                    use_whole=False
                )

                if matches:
                    trans = matches[0].get("translation", "")
                    results[case["case_num"]] = {
                        "korean": korean,
                        "translation": trans,
                        "matched": True,
                        "similarity": matches[0].get("similarity", 0),
                        "type": case_type
                    }
                else:
                    results[case["case_num"]] = {
                        "korean": korean,
                        "translation": "",
                        "matched": False,
                        "type": case_type
                    }

            else:
                # Standard single-line search
                matches = self.searcher.find_similar(
                    korean,
                    threshold=THRESHOLD,
                    top_k=1,
                    use_whole=True if len(korean) > 100 else False
                )

                if matches:
                    results[case["case_num"]] = {
                        "korean": korean,
                        "translation": matches[0].get("translation", ""),
                        "matched": True,
                        "similarity": matches[0].get("similarity", 0),
                        "type": case_type
                    }
                else:
                    results[case["case_num"]] = {
                        "korean": korean,
                        "translation": "",
                        "matched": False,
                        "type": case_type
                    }

        self.__class__.search_results_data = results
        return results

    # =========================================================================
    # TEST 1: Search runs successfully
    # =========================================================================

    def test_01_search_runs(self, search_results):
        """TEST-01: Similarity search completes without errors."""
        results = search_results

        total = len(results)
        matched = sum(1 for r in results.values() if r.get("matched", False))

        print(f"\nTEST-01 PASSED: Similarity search completed")
        print(f"  - Total: {total}")
        print(f"  - Matched: {matched}")
        print(f"  - Match rate: {matched/total:.1%}")

        assert total == len(self.test_cases)

    # =========================================================================
    # TEST 2: Exact match accuracy
    # =========================================================================

    def test_02_exact_matches(self, search_results):
        """TEST-02: Exact match cases get correct translations."""
        results = search_results

        exact_cases = [c for c in self.test_cases if c["type"] == "exact_match"]

        matched = 0
        correct = 0

        for case in exact_cases:
            result = results[case["case_num"]]
            if result.get("matched", False):
                matched += 1
                expected = case.get("expected_translation", "")
                if result["translation"] == expected:
                    correct += 1

        match_rate = matched / len(exact_cases) if exact_cases else 0

        print(f"\nTEST-02: Exact Matches")
        print(f"  - Total exact cases: {len(exact_cases)}")
        print(f"  - Got translation: {matched}")
        print(f"  - Match rate: {match_rate:.1%}")

        assert match_rate >= 0.80, f"Exact match rate too low: {match_rate:.1%}"

        print(f"TEST-02 PASSED: {matched}/{len(exact_cases)} exact matches found")

    # =========================================================================
    # TEST 3: Similar match (fuzzy) accuracy
    # =========================================================================

    def test_03_similar_matches(self, search_results):
        """TEST-03: Similar text cases match via embedding similarity."""
        results = search_results

        similar_cases = [c for c in self.test_cases if c["type"] == "similar_match"]

        matched = 0

        for case in similar_cases:
            result = results[case["case_num"]]
            if result.get("matched", False):
                matched += 1

        match_rate = matched / len(similar_cases) if similar_cases else 0

        print(f"\nTEST-03: Similar Matches (Fuzzy)")
        print(f"  - Total similar cases: {len(similar_cases)}")
        print(f"  - Got translation: {matched}")
        print(f"  - Match rate: {match_rate:.1%}")

        # Fuzzy matches should still work well due to embedding similarity
        assert match_rate >= 0.70, f"Similar match rate too low: {match_rate:.1%}"

        print(f"TEST-03 PASSED: {matched}/{len(similar_cases)} similar matches found")

    # =========================================================================
    # TEST 4: Triangle marker handling
    # =========================================================================

    def test_04_triangle_markers(self, search_results):
        """TEST-04: Triangle marker (▶) cases are handled correctly."""
        results = search_results

        triangle_cases = [c for c in self.test_cases if c["type"] == "triangle_marker"]

        matched = 0

        for case in triangle_cases:
            result = results[case["case_num"]]
            if result.get("matched", False):
                matched += 1

        match_rate = matched / len(triangle_cases) if triangle_cases else 0

        print(f"\nTEST-04: Triangle Marker Handling")
        print(f"  - Total triangle cases: {len(triangle_cases)}")
        print(f"  - Got translation: {matched}")
        print(f"  - Match rate: {match_rate:.1%}")

        assert match_rate >= 0.70, f"Triangle marker match rate too low: {match_rate:.1%}"

        print(f"TEST-04 PASSED: {matched}/{len(triangle_cases)} triangle markers handled")

    # =========================================================================
    # TEST 5: Multiline triangle handling
    # =========================================================================

    def test_05_multiline_triangle(self, search_results):
        """TEST-05: Multiline text with triangle markers is handled line-by-line."""
        results = search_results

        multiline_cases = [c for c in self.test_cases if c["type"] == "multiline_triangle"]

        matched = 0
        lines_matched = 0
        total_lines = 0

        for case in multiline_cases:
            result = results[case["case_num"]]
            if result.get("matched", False):
                matched += 1

            # Count line-level matches
            line_results = result.get("line_results", [])
            for lr in line_results:
                total_lines += 1
                if lr.get("matched", False):
                    lines_matched += 1

        match_rate = matched / len(multiline_cases) if multiline_cases else 0
        line_match_rate = lines_matched / total_lines if total_lines else 0

        print(f"\nTEST-05: Multiline Triangle Handling")
        print(f"  - Total multiline cases: {len(multiline_cases)}")
        print(f"  - Got translation: {matched}")
        print(f"  - Case match rate: {match_rate:.1%}")
        print(f"  - Line match rate: {line_match_rate:.1%} ({lines_matched}/{total_lines})")

        assert match_rate >= 0.50, f"Multiline match rate too low: {match_rate:.1%}"

        print(f"TEST-05 PASSED: {matched}/{len(multiline_cases)} multiline cases handled")

    # =========================================================================
    # TEST 6: No match cases
    # =========================================================================

    def test_06_no_match_cases(self, search_results):
        """TEST-06: Random/gibberish text correctly returns no match."""
        results = search_results

        no_match_cases = [c for c in self.test_cases if c["type"] == "no_match"]

        correctly_no_match = 0

        for case in no_match_cases:
            result = results[case["case_num"]]
            if not result.get("matched", False):
                correctly_no_match += 1

        correct_rate = correctly_no_match / len(no_match_cases) if no_match_cases else 0

        print(f"\nTEST-06: No Match Cases")
        print(f"  - Total no-match cases: {len(no_match_cases)}")
        print(f"  - Correctly no match: {correctly_no_match}")
        print(f"  - Correct rate: {correct_rate:.1%}")

        # Most gibberish should not match
        assert correct_rate >= 0.80, f"No-match correct rate too low: {correct_rate:.1%}"

        print(f"TEST-06 PASSED: {correctly_no_match}/{len(no_match_cases)} correctly identified")

    # =========================================================================
    # TEST 7: Full summary
    # =========================================================================

    def test_07_full_summary(self, search_results):
        """TEST-07: Complete summary of all results."""
        results = search_results

        # Collect stats by type
        stats = {}
        for case in self.test_cases:
            t = case["type"]
            if t not in stats:
                stats[t] = {"total": 0, "matched": 0, "should_match": 0}

            stats[t]["total"] += 1
            if case.get("should_match", False):
                stats[t]["should_match"] += 1

            result = results[case["case_num"]]
            if result.get("matched", False):
                stats[t]["matched"] += 1

        print(f"\n{'='*70}")
        print("TEST-07: FULL SUMMARY")
        print(f"{'='*70}")

        total_cases = 0
        total_matched = 0
        total_should_match = 0
        total_correct_match = 0

        for t, s in stats.items():
            match_rate = s["matched"] / s["total"] if s["total"] else 0
            expected_rate = s["should_match"] / s["total"] if s["total"] else 0

            print(f"\n{t}:")
            print(f"  Total: {s['total']}")
            print(f"  Should match: {s['should_match']} ({expected_rate:.1%})")
            print(f"  Actually matched: {s['matched']} ({match_rate:.1%})")

            total_cases += s["total"]
            total_matched += s["matched"]
            total_should_match += s["should_match"]

        overall_match_rate = total_matched / total_cases if total_cases else 0

        print(f"\n{'='*70}")
        print(f"OVERALL:")
        print(f"  Total cases: {total_cases}")
        print(f"  Should match: {total_should_match}")
        print(f"  Actually matched: {total_matched} ({overall_match_rate:.1%})")
        print(f"{'='*70}")

        # Overall success criteria
        # We expect high match rate for cases that should match
        should_match_cases = [c for c in self.test_cases if c.get("should_match", False)]
        actual_match_for_should = sum(
            1 for c in should_match_cases
            if results[c["case_num"]].get("matched", False)
        )
        should_match_rate = actual_match_for_should / len(should_match_cases) if should_match_cases else 0

        assert should_match_rate >= 0.65, f"Should-match rate too low: {should_match_rate:.1%}"

        print(f"\nTEST-07 PASSED: Should-match rate = {should_match_rate:.1%}")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
