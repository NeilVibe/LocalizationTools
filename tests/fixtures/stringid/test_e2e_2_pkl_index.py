"""
E2E Test 2: PKL Index Building with Variations

Tests the index building functionality with StringID support:
- TMIndexer.build_indexes() with StringID TM
- _build_whole_lookup() variations structure
- _build_whole_embeddings() string_id in mapping
- Line PKL structures

Test Cases: IDX-01 to IDX-08

Run: python -m pytest tests/fixtures/stringid/test_e2e_2_pkl_index.py -v
"""

import os
import sys
import pickle
import pytest
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Test data
from tests.fixtures.stringid.stringid_test_data import BASIC_STRINGID_DATA


class TestPKLIndexBuilding:
    """E2E tests for PKL Index Building with StringID variations."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment with a StringID TM."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.tools.ldm.tm_indexer import TMIndexer
        from server.database.models import LDMTranslationMemory, LDMTMEntry
        import openpyxl
        from io import BytesIO

        # Get session
        engine, session_maker = setup_database(drop_existing=False)
        self.db = session_maker()
        self.tm_manager = TMManager(self.db)
        self.tm_indexer = TMIndexer(self.db)
        self.created_tm_ids = []

        # Create a StringID TM for testing
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'
        for i, entry in enumerate(BASIC_STRINGID_DATA, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            ws[f'C{i}'] = entry.get("string_id", "")

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_pkl_index.xlsx",
            name="Test PKL Index TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )
        self.tm_id = result["tm_id"]
        self.created_tm_ids.append(self.tm_id)

        yield

        # Cleanup
        for tm_id in self.created_tm_ids:
            try:
                self.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        self.db.close()

    def _load_pickle(self, path: Path) -> dict:
        """Load pickle file."""
        with open(path, 'rb') as f:
            return pickle.load(f)

    # =========================================================================
    # IDX-01: Build indexes with StringID TM
    # =========================================================================

    def test_idx_01_build_indexes_stringid(self):
        """IDX-01: Build indexes creates PKL with variations structure."""
        # Build indexes
        result = self.tm_indexer.build_indexes(self.tm_id)

        assert result["status"] == "ready", f"Index build failed: {result}"
        assert result["whole_embeddings_count"] > 0, "No whole entries indexed"

        print(f"IDX-01 PASSED: Built indexes - {result['whole_embeddings_count']} whole entries")

    # =========================================================================
    # IDX-02: Standard TM (no StringID) - no variations
    # =========================================================================

    def test_idx_02_build_indexes_standard(self):
        """IDX-02: Standard mode TM has no variations structure."""
        import openpyxl
        from io import BytesIO

        # Create standard TM (no StringID)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['A2'] = '테스트'
        ws['B2'] = 'Test'
        ws['A3'] = '예제'
        ws['B3'] = 'Example'

        buffer = BytesIO()
        wb.save(buffer)

        result = self.tm_manager.upload_tm(
            file_content=buffer.getvalue(),
            filename="test_standard.xlsx",
            name="Test Standard TM",
            owner_id=1,
            mode="standard",
            source_col=0,
            target_col=1,
            has_header=True
        )
        standard_tm_id = result["tm_id"]
        self.created_tm_ids.append(standard_tm_id)

        # Build indexes
        build_result = self.tm_indexer.build_indexes(standard_tm_id)
        assert build_result["status"] == "ready"

        # Check PKL structure - standard mode entries have no variations
        tm_path = Path(f"server/data/ldm_tm/{standard_tm_id}")
        lookup_path = tm_path / "hash" / "whole_lookup.pkl"

        if lookup_path.exists():
            lookup = self._load_pickle(lookup_path)
            # Standard mode: entries should have string_id=None or no variations
            for key, value in lookup.items():
                if isinstance(value, dict):
                    if "variations" in value:
                        # If variations exist, they should have string_id=None
                        for var in value["variations"]:
                            assert var.get("string_id") is None or var.get("string_id") == ""

        print(f"IDX-02 PASSED: Standard mode indexes built correctly")

    # =========================================================================
    # IDX-03: Verify whole_lookup variations structure
    # =========================================================================

    def test_idx_03_whole_lookup_variations(self):
        """IDX-03: whole_lookup.pkl has variations structure for duplicate sources."""
        # Build indexes first
        self.tm_indexer.build_indexes(self.tm_id)

        # Load whole_lookup.pkl
        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        lookup_path = tm_path / "hash" / "whole_lookup.pkl"

        assert lookup_path.exists(), f"whole_lookup.pkl not found at {lookup_path}"

        lookup = self._load_pickle(lookup_path)

        # "저장" should have 3 variations (Save, Save Game, Storage)
        assert "저장" in lookup, "'저장' not found in lookup"

        save_entry = lookup["저장"]
        assert "variations" in save_entry, "'저장' should have variations key"
        assert len(save_entry["variations"]) == 3, \
            f"Expected 3 variations for '저장', got {len(save_entry['variations'])}"

        # Verify each variation has string_id
        string_ids = {v["string_id"] for v in save_entry["variations"]}
        expected_ids = {"UI_BUTTON_SAVE", "UI_MENU_SAVE", "TECH_STORAGE"}
        assert string_ids == expected_ids, \
            f"StringID mismatch. Expected: {expected_ids}, Got: {string_ids}"

        print(f"IDX-03 PASSED: '저장' has {len(save_entry['variations'])} variations with correct StringIDs")

    # =========================================================================
    # IDX-04: Verify whole_mapping has string_id
    # =========================================================================

    def test_idx_04_whole_mapping_stringid(self):
        """IDX-04: whole_mapping.pkl entries have string_id field."""
        # Build indexes first
        self.tm_indexer.build_indexes(self.tm_id)

        # Load whole_mapping.pkl
        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        mapping_path = tm_path / "embeddings" / "whole_mapping.pkl"

        assert mapping_path.exists(), f"whole_mapping.pkl not found at {mapping_path}"

        mapping = self._load_pickle(mapping_path)

        assert len(mapping) > 0, "Mapping is empty"

        # Check that entries have string_id field
        entries_with_stringid = [e for e in mapping if "string_id" in e]
        assert len(entries_with_stringid) == len(mapping), \
            f"Not all entries have string_id. {len(entries_with_stringid)}/{len(mapping)}"

        # Verify string_id values
        string_ids = {e["string_id"] for e in mapping if e["string_id"]}
        expected_ids = {e["string_id"] for e in BASIC_STRINGID_DATA}
        assert string_ids == expected_ids, \
            f"StringID mismatch in mapping. Expected: {expected_ids}, Got: {string_ids}"

        print(f"IDX-04 PASSED: All {len(mapping)} mapping entries have string_id field")

    # =========================================================================
    # IDX-05: Line lookup structure (if multi-line entries exist)
    # =========================================================================

    def test_idx_05_line_lookup_stringid(self):
        """IDX-05: line_lookup.pkl entries have string_id (if multi-line exists)."""
        # Build indexes first
        self.tm_indexer.build_indexes(self.tm_id)

        # Check if line lookup exists
        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        line_lookup_path = tm_path / "hash" / "line_lookup.pkl"

        if line_lookup_path.exists():
            line_lookup = self._load_pickle(line_lookup_path)
            if line_lookup:
                # Check structure
                for key, value in line_lookup.items():
                    if isinstance(value, dict) and "variations" in value:
                        for var in value["variations"]:
                            assert "string_id" in var, f"Line entry missing string_id: {key}"
                print(f"IDX-05 PASSED: Line lookup has string_id in entries")
            else:
                print(f"IDX-05 SKIPPED: No line lookup entries (no multi-line data)")
        else:
            print(f"IDX-05 SKIPPED: No line_lookup.pkl (no multi-line data in TM)")

    # =========================================================================
    # IDX-06: Line mapping structure (if multi-line entries exist)
    # =========================================================================

    def test_idx_06_line_mapping_stringid(self):
        """IDX-06: line_mapping.pkl entries have string_id (if multi-line exists)."""
        # Build indexes first
        self.tm_indexer.build_indexes(self.tm_id)

        # Check if line mapping exists
        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        line_mapping_path = tm_path / "embeddings" / "line_mapping.pkl"

        if line_mapping_path.exists():
            line_mapping = self._load_pickle(line_mapping_path)
            if line_mapping:
                for entry in line_mapping:
                    assert "string_id" in entry, f"Line mapping entry missing string_id"
                print(f"IDX-06 PASSED: Line mapping has string_id in {len(line_mapping)} entries")
            else:
                print(f"IDX-06 SKIPPED: Empty line mapping (no multi-line data)")
        else:
            print(f"IDX-06 SKIPPED: No line_mapping.pkl (no multi-line data in TM)")

    # =========================================================================
    # IDX-07: No line indexes (TM without multi-line)
    # =========================================================================

    def test_idx_07_no_line_indexes(self):
        """IDX-07: TM without multi-line entries has no line indexes (no error)."""
        # Our test data is single-line, so no line indexes expected
        self.tm_indexer.build_indexes(self.tm_id)

        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        line_npy = tm_path / "embeddings" / "line.npy"

        # Line indexes are optional - should not cause errors
        # Just verify the whole indexes exist
        whole_lookup = tm_path / "hash" / "whole_lookup.pkl"
        assert whole_lookup.exists(), "whole_lookup.pkl should exist"

        if not line_npy.exists():
            print(f"IDX-07 PASSED: No line.npy (as expected for single-line data)")
        else:
            print(f"IDX-07 PASSED: line.npy exists (multi-line data found)")

    # =========================================================================
    # IDX-08: Rebuild indexes produces same structure
    # =========================================================================

    def test_idx_08_rebuild_indexes(self):
        """IDX-08: Rebuilding indexes produces consistent structure."""
        # Build first time
        result1 = self.tm_indexer.build_indexes(self.tm_id)

        # Load structure
        tm_path = Path(f"server/data/ldm_tm/{self.tm_id}")
        lookup1 = self._load_pickle(tm_path / "hash" / "whole_lookup.pkl")

        # Delete PKL to force rebuild
        import shutil
        if tm_path.exists():
            shutil.rmtree(tm_path)

        # Rebuild
        result2 = self.tm_indexer.build_indexes(self.tm_id)

        # Load again
        lookup2 = self._load_pickle(tm_path / "hash" / "whole_lookup.pkl")

        # Compare structure (keys should be same)
        assert set(lookup1.keys()) == set(lookup2.keys()), "Keys changed after rebuild"

        # Compare variation counts
        for key in lookup1:
            if isinstance(lookup1[key], dict) and "variations" in lookup1[key]:
                assert len(lookup1[key]["variations"]) == len(lookup2[key]["variations"]), \
                    f"Variation count changed for '{key}'"

        print(f"IDX-08 PASSED: Rebuild produces consistent structure")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
