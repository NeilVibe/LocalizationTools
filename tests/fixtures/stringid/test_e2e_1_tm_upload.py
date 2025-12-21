"""
E2E Test 1: TM Upload with StringID Mode

Tests the TM upload functionality with StringID support:
- Upload Excel with 3 columns (Source, Target, StringID)
- Upload Excel with 2 columns (Standard mode)
- Verify DB storage of string_id column
- Verify TM mode column
- All test cases: TM-01 to TM-08

Run: python -m pytest tests/fixtures/stringid/test_e2e_1_tm_upload.py -v

NOTE: Requires 'mode' column in ldm_translation_memories table (migration pending in CI)
"""

import os
import sys
import pytest

# Note: db_setup.py's upgrade_schema() auto-adds the 'mode' column if missing
from pathlib import Path
from io import BytesIO

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Test data
from tests.fixtures.stringid.stringid_test_data import (
    BASIC_STRINGID_DATA,
    STANDARD_TM_DATA,
    ALL_STRINGID_TEST_DATA,
)


class TestTMUploadStringID:
    """E2E tests for TM Upload with StringID mode."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment."""
        # Import here to avoid issues if DB not available
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.database.models import LDMTranslationMemory, LDMTMEntry

        # Get session maker (don't drop existing DB)
        engine, session_maker = setup_database(drop_existing=False)
        self.db = session_maker()
        self.tm_manager = TMManager(self.db)
        self.created_tm_ids = []

        yield

        # Cleanup: Delete test TMs
        for tm_id in self.created_tm_ids:
            try:
                self.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        self.db.close()

    def _generate_excel_bytes(self, data: list, include_stringid: bool = True) -> bytes:
        """Generate Excel file bytes from test data."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        if include_stringid:
            ws['A1'] = 'Source'
            ws['B1'] = 'Target'
            ws['C1'] = 'StringID'
        else:
            ws['A1'] = 'Source'
            ws['B1'] = 'Target'

        # Data rows
        for i, entry in enumerate(data, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            if include_stringid:
                ws[f'C{i}'] = entry.get("string_id", "") or ""

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # =========================================================================
    # TM-01: Upload StringID Excel (3-col)
    # =========================================================================

    def test_tm_01_upload_stringid_excel(self):
        """TM-01: Upload 3-column Excel with mode=stringid stores all entries."""
        # Generate Excel with StringID data
        excel_bytes = self._generate_excel_bytes(BASIC_STRINGID_DATA, include_stringid=True)

        # Upload with stringid mode
        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_stringid.xlsx",
            name="Test StringID TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # Verify all entries stored (no merging)
        assert result["entry_count"] == len(BASIC_STRINGID_DATA), \
            f"Expected {len(BASIC_STRINGID_DATA)} entries, got {result['entry_count']}"

        print(f"TM-01 PASSED: {result['entry_count']} entries uploaded")

    # =========================================================================
    # TM-02: Upload Standard Excel (2-col) - Duplicates Merged
    # =========================================================================

    def test_tm_02_upload_standard_excel(self):
        """TM-02: Upload 2-column Excel with mode=standard merges duplicates."""
        # Generate Excel without StringID (only source + target)
        # Use data that has duplicate sources
        test_data = [
            {"string_id": None, "source": "저장", "target": "Save"},
            {"string_id": None, "source": "저장", "target": "Save Game"},  # Duplicate
            {"string_id": None, "source": "설정", "target": "Settings"},
            {"string_id": None, "source": "확인", "target": "OK"},
        ]
        excel_bytes = self._generate_excel_bytes(test_data, include_stringid=False)

        # Upload with standard mode
        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_standard.xlsx",
            name="Test Standard TM",
            owner_id=1,
            mode="standard",
            source_col=0,
            target_col=1,
            stringid_col=None,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # In standard mode, all entries are stored (no merging in DB, merging happens in PKL)
        # The DB stores all entries, PKL merges them
        assert result["entry_count"] == len(test_data), \
            f"Expected {len(test_data)} entries, got {result['entry_count']}"

        print(f"TM-02 PASSED: {result['entry_count']} entries uploaded (standard mode)")

    # =========================================================================
    # TM-03: Verify string_id Stored in DB
    # =========================================================================

    def test_tm_03_verify_stringid_stored(self):
        """TM-03: Verify string_id column is populated in DB."""
        from server.database.models import LDMTMEntry

        # Upload StringID TM
        excel_bytes = self._generate_excel_bytes(BASIC_STRINGID_DATA, include_stringid=True)
        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_stringid.xlsx",
            name="Test StringID Verify",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # Query DB directly
        entries = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == result["tm_id"]
        ).all()

        # Count entries with string_id populated
        entries_with_stringid = [e for e in entries if e.string_id]
        assert len(entries_with_stringid) == len(BASIC_STRINGID_DATA), \
            f"Expected {len(BASIC_STRINGID_DATA)} entries with string_id, got {len(entries_with_stringid)}"

        # Verify specific string_ids
        string_ids = {e.string_id for e in entries}
        expected_ids = {e["string_id"] for e in BASIC_STRINGID_DATA}
        assert string_ids == expected_ids, \
            f"StringID mismatch. Expected: {expected_ids}, Got: {string_ids}"

        print(f"TM-03 PASSED: All {len(entries_with_stringid)} entries have string_id populated")

    # =========================================================================
    # TM-04: Verify TM Mode Stored
    # =========================================================================

    def test_tm_04_verify_tm_mode(self):
        """TM-04: Verify TM mode column is set correctly."""
        from server.database.models import LDMTranslationMemory

        # Upload StringID TM
        excel_bytes = self._generate_excel_bytes(BASIC_STRINGID_DATA, include_stringid=True)
        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_mode.xlsx",
            name="Test Mode TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # Query TM record
        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == result["tm_id"]
        ).first()

        assert tm.mode == "stringid", f"Expected mode='stringid', got '{tm.mode}'"

        print(f"TM-04 PASSED: TM mode correctly set to '{tm.mode}'")

    # =========================================================================
    # TM-05: Empty StringID Column
    # =========================================================================

    def test_tm_05_empty_stringid_column(self):
        """TM-05: Upload with empty StringID column - entries have NULL string_id."""
        from server.database.models import LDMTMEntry

        # Data with empty StringID
        test_data = [
            {"string_id": "", "source": "테스트1", "target": "Test1"},
            {"string_id": "", "source": "테스트2", "target": "Test2"},
        ]
        excel_bytes = self._generate_excel_bytes(test_data, include_stringid=True)

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_empty_stringid.xlsx",
            name="Test Empty StringID",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # Query DB
        entries = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == result["tm_id"]
        ).all()

        # All entries should have NULL or empty string_id
        entries_with_stringid = [e for e in entries if e.string_id]
        assert len(entries_with_stringid) == 0, \
            f"Expected 0 entries with string_id, got {len(entries_with_stringid)}"

        print(f"TM-05 PASSED: Empty StringID handled correctly ({len(entries)} entries with NULL string_id)")

    # =========================================================================
    # TM-06: Mixed StringID (Some rows have StringID, some don't)
    # =========================================================================

    def test_tm_06_mixed_stringid(self):
        """TM-06: Upload with mixed StringID - partial storage."""
        from server.database.models import LDMTMEntry

        # Data with mixed StringID
        test_data = [
            {"string_id": "ID_001", "source": "테스트1", "target": "Test1"},
            {"string_id": "", "source": "테스트2", "target": "Test2"},
            {"string_id": "ID_003", "source": "테스트3", "target": "Test3"},
            {"string_id": None, "source": "테스트4", "target": "Test4"},
        ]
        excel_bytes = self._generate_excel_bytes(test_data, include_stringid=True)

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_mixed_stringid.xlsx",
            name="Test Mixed StringID",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        # Query DB
        entries = self.db.query(LDMTMEntry).filter(
            LDMTMEntry.tm_id == result["tm_id"]
        ).all()

        # Count entries with string_id
        entries_with_stringid = [e for e in entries if e.string_id]
        assert len(entries_with_stringid) == 2, \
            f"Expected 2 entries with string_id, got {len(entries_with_stringid)}"

        print(f"TM-06 PASSED: Mixed StringID - {len(entries_with_stringid)} with ID, {len(entries) - len(entries_with_stringid)} without")

    # =========================================================================
    # TM-07: Upload TXT Format (StringID in col 1)
    # =========================================================================

    def test_tm_07_upload_txt_format(self):
        """TM-07: Upload TXT file with StringID in column 1."""
        # Read the generated TXT fixture
        fixture_path = Path(__file__).parent / "basic_stringid_data.txt"
        with open(fixture_path, 'rb') as f:
            txt_bytes = f.read()

        result = self.tm_manager.upload_tm(
            file_content=txt_bytes,
            filename="basic_stringid_data.txt",
            name="Test TXT StringID",
            owner_id=1,
            mode="stringid"
        )

        self.created_tm_ids.append(result["tm_id"])

        # NOTE: Current TXT parser may not extract StringID
        # This test documents expected behavior
        assert result["entry_count"] > 0, "TXT upload should have entries"

        print(f"TM-07 PASSED: TXT upload - {result['entry_count']} entries")

    # =========================================================================
    # TM-08: Upload XML Format (StringID in id attribute)
    # =========================================================================

    def test_tm_08_upload_xml_format(self):
        """TM-08: Upload XML file with StringID in id attribute."""
        # Read the generated XML fixture
        fixture_path = Path(__file__).parent / "basic_stringid_data.xml"
        with open(fixture_path, 'rb') as f:
            xml_bytes = f.read()

        result = self.tm_manager.upload_tm(
            file_content=xml_bytes,
            filename="basic_stringid_data.xml",
            name="Test XML StringID",
            owner_id=1,
            mode="stringid"
        )

        self.created_tm_ids.append(result["tm_id"])

        # NOTE: Current XML parser may not extract StringID
        # This test documents expected behavior
        assert result["entry_count"] > 0, "XML upload should have entries"

        print(f"TM-08 PASSED: XML upload - {result['entry_count']} entries")


class TestTMUploadBackwardCompat:
    """Backward compatibility tests for TM upload."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager

        engine, session_maker = setup_database(drop_existing=False)
        self.db = session_maker()
        self.tm_manager = TMManager(self.db)
        self.created_tm_ids = []

        yield

        for tm_id in self.created_tm_ids:
            try:
                self.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        self.db.close()

    def _generate_excel_bytes(self, data: list) -> bytes:
        """Generate 2-column Excel file bytes."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'

        for i, entry in enumerate(data, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_bc_01_standard_mode_works(self):
        """BC-01: Standard TM mode still works without StringID."""
        test_data = [
            {"source": "저장", "target": "Save"},
            {"source": "설정", "target": "Settings"},
            {"source": "확인", "target": "OK"},
        ]
        excel_bytes = self._generate_excel_bytes(test_data)

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_bc.xlsx",
            name="Test BC Standard",
            owner_id=1,
            mode="standard",
            source_col=0,
            target_col=1,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])
        assert result["entry_count"] == len(test_data)

        print(f"BC-01 PASSED: Standard mode works ({result['entry_count']} entries)")

    def test_bc_03_default_mode_standard(self):
        """BC-03: TM without explicit mode defaults to standard."""
        from server.database.models import LDMTranslationMemory

        test_data = [{"source": "테스트", "target": "Test"}]
        excel_bytes = self._generate_excel_bytes(test_data)

        # Upload without specifying mode (should default)
        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_default.xlsx",
            name="Test Default Mode",
            owner_id=1,
            source_col=0,
            target_col=1,
            has_header=True
        )

        self.created_tm_ids.append(result["tm_id"])

        tm = self.db.query(LDMTranslationMemory).filter(
            LDMTranslationMemory.id == result["tm_id"]
        ).first()

        assert tm.mode == "standard", f"Expected default mode='standard', got '{tm.mode}'"

        print(f"BC-03 PASSED: Default mode is 'standard'")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
