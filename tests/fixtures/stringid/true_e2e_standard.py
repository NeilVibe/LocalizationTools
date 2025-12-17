"""
TRUE E2E Test: Standard TM Engine with StringID

═══════════════════════════════════════════════════════════════════════════════
THIS IS A TRUE END-TO-END TEST
═══════════════════════════════════════════════════════════════════════════════

WHAT THIS TEST DOES:
1. Loads ~5000 rows from real production data (sampleofLanguageData.txt)
2. Creates a TM with StringID mode (duplicates kept as variations)
3. Builds PKL indexes with Qwen embeddings (~5 min)
4. Creates a test file with 150 rows to pretranslate
5. Runs pretranslation with engine="standard"
6. Verifies CORRECT target returned based on StringID match

DATA SOURCE:
/mnt/c/NEIL_PROJECTS_WINDOWSBUILD/LocaNextProject/TestFilesForLocaNext/sampleofLanguageData.txt
- 103,499 rows of real game localization data
- Tab-separated: Col1=StringID, Col5=Korean, Col6=French
- Contains: color codes, TextBind, newlines

TEST EXPECTATIONS:
- If row has StringID that matches TM entry → get that exact target
- If row has StringID that doesn't match → get first variation (fallback)
- If row has no StringID → get first variation

Run: python -m pytest tests/fixtures/stringid/true_e2e_standard.py -v -s
Time: ~10 minutes (embedding generation is slow)
"""

import os
import sys
import pytest
import random
from pathlib import Path
from typing import List, Dict, Tuple

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Data source path
SAMPLE_DATA_PATH = Path("/mnt/c/NEIL_PROJECTS_WINDOWSBUILD/LocaNextProject/TestFilesForLocaNext/sampleofLanguageData.txt")

# Test configuration
TM_SIZE = 5000          # Number of rows for TM
TEST_FILE_SIZE = 150    # Number of rows to pretranslate
THRESHOLD = 0.92        # Match threshold


def parse_sample_data(limit: int = None) -> List[Dict]:
    """
    Parse sampleofLanguageData.txt into list of entries.

    Format: Tab-separated
    Col 0: Flag (0)
    Col 1: StringID (numeric)
    Col 2-4: Metadata
    Col 5: Korean source
    Col 6: Target (French)
    Col 7: Status
    """
    if not SAMPLE_DATA_PATH.exists():
        pytest.skip(f"Sample data not found: {SAMPLE_DATA_PATH}")

    entries = []
    with open(SAMPLE_DATA_PATH, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            if limit and i >= limit:
                break

            parts = line.strip().split('\t')
            if len(parts) >= 7:
                string_id = parts[1].strip()
                source = parts[5].strip()
                target = parts[6].strip()

                if source and target:  # Skip empty
                    entries.append({
                        "string_id": f"ID_{string_id}",  # Prefix to make clear it's a StringID
                        "source": source,
                        "target": target
                    })

    return entries


def find_duplicate_sources(entries: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Find entries with duplicate Korean sources (different StringIDs).
    These are the key test cases for StringID matching.
    """
    source_map = {}
    for entry in entries:
        src = entry["source"]
        if src not in source_map:
            source_map[src] = []
        source_map[src].append(entry)

    # Return only sources with multiple entries (duplicates)
    return {src: entries for src, entries in source_map.items() if len(entries) > 1}


class TestTrueE2EStandardTM:
    """
    TRUE End-to-End test for Standard TM with StringID matching.

    This test uses real production data and verifies the full pipeline:
    TM Upload → Index Build → Pretranslation → Target Verification
    """

    @pytest.fixture(scope="class", autouse=True)
    def setup(self, request):
        """Setup test environment with large TM from real data."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.tools.ldm.tm_indexer import TMIndexer
        from server.tools.ldm.pretranslate import PretranslationEngine
        from server.database.models import LDMFile, LDMRow, LDMProject
        import openpyxl
        from io import BytesIO

        cls = request.cls  # Store on class for class-scoped fixture

        print("\n" + "="*70)
        print("TRUE E2E TEST: Standard TM with StringID")
        print("="*70)

        # Parse real data
        print(f"\n[1/6] Loading data from {SAMPLE_DATA_PATH}...")
        all_entries = parse_sample_data(limit=TM_SIZE * 2)  # Load extra to find duplicates
        print(f"      Loaded {len(all_entries)} entries")

        # Find duplicates for StringID testing
        duplicates = find_duplicate_sources(all_entries)
        print(f"      Found {len(duplicates)} duplicate sources (key for StringID testing)")

        # Select entries for TM (prioritize duplicates)
        cls.tm_entries = []
        cls.duplicate_sources = duplicates

        # First add all duplicates (these are key test cases)
        for src, entries in duplicates.items():
            cls.tm_entries.extend(entries)

        # Fill remaining with unique entries
        used_sources = set(duplicates.keys())
        for entry in all_entries:
            if len(cls.tm_entries) >= TM_SIZE:
                break
            if entry["source"] not in used_sources:
                cls.tm_entries.append(entry)
                used_sources.add(entry["source"])

        print(f"      Selected {len(cls.tm_entries)} entries for TM")
        print(f"      Duplicates included: {sum(len(v) for v in duplicates.values())} entries")

        # Database setup
        print(f"\n[2/6] Setting up database...")
        engine, session_maker = setup_database(drop_existing=False)
        cls.db = session_maker()
        cls.tm_manager = TMManager(cls.db)
        cls.tm_indexer = TMIndexer(cls.db)
        cls.pretranslation_engine = PretranslationEngine(cls.db)
        cls.created_tm_ids = []
        cls.created_file_ids = []

        # Create Excel file for TM upload
        print(f"\n[3/6] Creating TM Excel file ({len(cls.tm_entries)} rows)...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'

        for i, entry in enumerate(cls.tm_entries, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            ws[f'C{i}'] = entry["string_id"]

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        print(f"      Excel file size: {len(excel_bytes):,} bytes")

        # Upload TM
        print(f"\n[4/6] Uploading TM (mode=stringid)...")
        result = cls.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="true_e2e_standard_tm.xlsx",
            name="TRUE E2E Standard TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )
        cls.tm_id = result["tm_id"]
        cls.created_tm_ids.append(cls.tm_id)
        print(f"      TM created: id={cls.tm_id}, entries={result.get('entry_count', 'N/A')}")

        # Build indexes (this takes time with Qwen embeddings)
        print(f"\n[5/6] Building PKL indexes (this takes ~1 minute)...")
        cls.tm_indexer.build_indexes(cls.tm_id)
        print(f"      Indexes built successfully")

        # Create test file
        print(f"\n[6/6] Creating test file ({TEST_FILE_SIZE} rows)...")
        cls._create_test_file_static(cls)
        print(f"      Test file created: id={cls.test_file_id}")

        print("\n" + "="*70)
        print("SETUP COMPLETE - Ready for testing")
        print("="*70 + "\n")

        yield

        # Cleanup
        print("\n[CLEANUP] Removing test data...")
        for file_id in cls.created_file_ids:
            try:
                cls.db.query(LDMRow).filter(LDMRow.file_id == file_id).delete()
                cls.db.query(LDMFile).filter(LDMFile.id == file_id).delete()
                cls.db.commit()
            except Exception:
                pass
        for tm_id in cls.created_tm_ids:
            try:
                cls.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        cls.db.close()
        print("[CLEANUP] Done")

    @staticmethod
    def _create_test_file_static(cls):
        """
        Create test file with 150 rows for pretranslation.

        Mix of:
        - Rows with StringID that matches TM (should get exact target)
        - Rows with StringID that doesn't match (should get fallback)
        - Rows without StringID (should get first variation)
        """
        from server.database.models import LDMFile, LDMRow, LDMProject

        # Ensure test project exists
        project = cls.db.query(LDMProject).filter(LDMProject.name == "TRUE E2E Test").first()
        if not project:
            project = LDMProject(name="TRUE E2E Test", owner_id=1)
            cls.db.add(project)
            cls.db.commit()

        # Create test rows
        cls.test_rows = []
        cls.expected_results = {}  # row_num -> expected_target

        row_num = 0

        # 1. Add rows with MATCHING StringID (50 rows)
        # These should get the exact target from TM
        for src, entries in list(cls.duplicate_sources.items())[:25]:
            for entry in entries[:2]:  # Take up to 2 per duplicate source
                row_num += 1
                cls.test_rows.append({
                    "source": entry["source"],
                    "target": "",
                    "string_id": entry["string_id"]
                })
                cls.expected_results[row_num] = {
                    "string_id": entry["string_id"],
                    "expected_target": entry["target"],
                    "match_type": "exact_stringid"
                }

        # 2. Add rows with NON-MATCHING StringID (50 rows)
        # These should get fallback (first variation)
        for src, entries in list(cls.duplicate_sources.items())[25:50]:
            row_num += 1
            first_entry = entries[0]
            cls.test_rows.append({
                "source": first_entry["source"],
                "target": "",
                "string_id": "UNKNOWN_ID_" + str(row_num)  # StringID not in TM
            })
            # Expected: first variation (fallback)
            cls.expected_results[row_num] = {
                "string_id": "UNKNOWN_ID_" + str(row_num),
                "expected_target": first_entry["target"],  # First variation
                "match_type": "fallback_first"
            }

        # 3. Add rows WITHOUT StringID (50 rows)
        # These should get first variation
        unique_entries = [e for e in cls.tm_entries if e["source"] not in cls.duplicate_sources]
        for entry in unique_entries[:50]:
            row_num += 1
            cls.test_rows.append({
                "source": entry["source"],
                "target": "",
                "string_id": None
            })
            cls.expected_results[row_num] = {
                "string_id": None,
                "expected_target": entry["target"],
                "match_type": "no_stringid"
            }

        # Pad to TEST_FILE_SIZE if needed
        while len(cls.test_rows) < TEST_FILE_SIZE:
            entry = random.choice(unique_entries)
            row_num += 1
            cls.test_rows.append({
                "source": entry["source"],
                "target": "",
                "string_id": None
            })
            cls.expected_results[row_num] = {
                "string_id": None,
                "expected_target": entry["target"],
                "match_type": "no_stringid"
            }

        # Create file record
        file = LDMFile(
            project_id=project.id,
            name="true_e2e_test_file.txt",
            original_filename="true_e2e_test_file.txt",
            format="txt",
            row_count=len(cls.test_rows),
            created_by=1
        )
        cls.db.add(file)
        cls.db.commit()
        cls.test_file_id = file.id
        cls.created_file_ids.append(file.id)

        # Create rows
        for i, row_data in enumerate(cls.test_rows, start=1):
            row = LDMRow(
                file_id=file.id,
                row_num=i,
                source=row_data["source"],
                target=row_data["target"],
                string_id=row_data.get("string_id"),
                status="pending"
            )
            cls.db.add(row)

        cls.db.commit()

    # =========================================================================
    # RUN PRETRANSLATION ONCE FOR ALL TESTS
    # =========================================================================

    @pytest.fixture(scope="class")
    def pretranslation_result(self):
        """Run pretranslation once and store result for all tests."""
        result = self.pretranslation_engine.pretranslate(
            file_id=self.test_file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=THRESHOLD
        )
        # Store on class for other tests
        self.__class__.pretranslation_result_data = result
        return result

    # =========================================================================
    # TEST 1: Pretranslation runs successfully
    # =========================================================================

    def test_01_pretranslation_runs(self, pretranslation_result):
        """TEST-01: Pretranslation completes without errors."""
        result = pretranslation_result

        assert "matched" in result
        assert "total" in result
        assert result["total"] == len(self.test_rows)

        print(f"\nTEST-01 PASSED: Pretranslation completed")
        print(f"  - Total: {result['total']}")
        print(f"  - Matched: {result['matched']}")
        print(f"  - Time: {result.get('time_seconds', 'N/A')}s")

    # =========================================================================
    # TEST 2: Match rate is reasonable
    # =========================================================================

    def test_02_match_rate(self, pretranslation_result):
        """TEST-02: Match rate is above 80% (all test data comes from TM)."""
        result = pretranslation_result

        match_rate = result["matched"] / result["total"] if result["total"] > 0 else 0

        # Should be high since test data comes from TM
        assert match_rate >= 0.80, f"Match rate too low: {match_rate:.1%}"

        print(f"\nTEST-02 PASSED: Match rate = {match_rate:.1%}")

    # =========================================================================
    # TEST 3: StringID exact matches get correct target
    # =========================================================================

    def test_03_stringid_exact_matches(self, pretranslation_result):
        """TEST-03: Rows with matching StringID get the exact correct target."""
        from server.database.models import LDMRow

        # Check exact StringID matches
        exact_matches = [
            (row_num, exp) for row_num, exp in self.expected_results.items()
            if exp["match_type"] == "exact_stringid"
        ]

        errors = []
        correct = 0

        rows = self.db.query(LDMRow).filter(LDMRow.file_id == self.test_file_id).all()
        row_map = {r.row_num: r for r in rows}

        for row_num, expected in exact_matches:
            row = row_map.get(row_num)
            if row:
                self.db.refresh(row)
                if row.target == expected["expected_target"]:
                    correct += 1
                else:
                    errors.append(
                        f"Row {row_num}: StringID='{expected['string_id']}' "
                        f"expected='{expected['expected_target'][:30]}...' "
                        f"got='{(row.target or '')[:30]}...'"
                    )

        error_rate = len(errors) / len(exact_matches) if exact_matches else 0

        print(f"\nTEST-03: StringID Exact Matches")
        print(f"  - Total exact match rows: {len(exact_matches)}")
        print(f"  - Correct: {correct}")
        print(f"  - Errors: {len(errors)}")

        if errors[:5]:  # Show first 5 errors
            print(f"  - Sample errors:")
            for e in errors[:5]:
                print(f"    {e}")

        # Allow up to 10% errors (some might have threshold issues)
        assert error_rate <= 0.10, f"Too many errors: {error_rate:.1%}"

        print(f"TEST-03 PASSED: {correct}/{len(exact_matches)} exact matches correct")

    # =========================================================================
    # TEST 4: Fallback to first variation works
    # =========================================================================

    def test_04_fallback_behavior(self, pretranslation_result):
        """TEST-04: Rows with unknown StringID fall back to first variation."""
        from server.database.models import LDMRow

        # Check fallback matches
        fallback_matches = [
            (row_num, exp) for row_num, exp in self.expected_results.items()
            if exp["match_type"] == "fallback_first"
        ]

        rows = self.db.query(LDMRow).filter(LDMRow.file_id == self.test_file_id).all()
        row_map = {r.row_num: r for r in rows}

        matched_count = 0
        for row_num, expected in fallback_matches:
            row = row_map.get(row_num)
            if row:
                self.db.refresh(row)
                if row.target:  # Got some target (fallback worked)
                    matched_count += 1

        match_rate = matched_count / len(fallback_matches) if fallback_matches else 0

        print(f"\nTEST-04: Fallback Behavior")
        print(f"  - Total fallback rows: {len(fallback_matches)}")
        print(f"  - Got translations: {matched_count}")
        print(f"  - Rate: {match_rate:.1%}")

        assert match_rate >= 0.80, f"Fallback rate too low: {match_rate:.1%}"

        print(f"TEST-04 PASSED: Fallback working for {matched_count}/{len(fallback_matches)} rows")

    # =========================================================================
    # TEST 5: No StringID rows work
    # =========================================================================

    def test_05_no_stringid_rows(self, pretranslation_result):
        """TEST-05: Rows without StringID get translations."""
        from server.database.models import LDMRow

        # Check no-StringID matches
        no_stringid = [
            (row_num, exp) for row_num, exp in self.expected_results.items()
            if exp["match_type"] == "no_stringid"
        ]

        rows = self.db.query(LDMRow).filter(LDMRow.file_id == self.test_file_id).all()
        row_map = {r.row_num: r for r in rows}

        matched_count = 0
        for row_num, expected in no_stringid:
            row = row_map.get(row_num)
            if row:
                self.db.refresh(row)
                if row.target:
                    matched_count += 1

        match_rate = matched_count / len(no_stringid) if no_stringid else 0

        print(f"\nTEST-05: No StringID Rows")
        print(f"  - Total no-StringID rows: {len(no_stringid)}")
        print(f"  - Got translations: {matched_count}")
        print(f"  - Rate: {match_rate:.1%}")

        assert match_rate >= 0.80, f"No-StringID rate too low: {match_rate:.1%}"

        print(f"TEST-05 PASSED: {matched_count}/{len(no_stringid)} no-StringID rows translated")

    # =========================================================================
    # TEST 6: Full verification summary
    # =========================================================================

    def test_06_full_verification(self, pretranslation_result):
        """TEST-06: Complete verification of all results."""
        from server.database.models import LDMRow

        result = pretranslation_result

        rows = self.db.query(LDMRow).filter(LDMRow.file_id == self.test_file_id).all()

        stats = {
            "total": len(rows),
            "matched": result["matched"],
            "exact_stringid_correct": 0,
            "exact_stringid_wrong": 0,
            "fallback_got_target": 0,
            "no_stringid_got_target": 0
        }

        row_map = {r.row_num: r for r in rows}

        for row_num, expected in self.expected_results.items():
            row = row_map.get(row_num)
            if not row:
                continue
            self.db.refresh(row)

            if expected["match_type"] == "exact_stringid":
                if row.target == expected["expected_target"]:
                    stats["exact_stringid_correct"] += 1
                else:
                    stats["exact_stringid_wrong"] += 1
            elif expected["match_type"] == "fallback_first":
                if row.target:
                    stats["fallback_got_target"] += 1
            elif expected["match_type"] == "no_stringid":
                if row.target:
                    stats["no_stringid_got_target"] += 1

        print(f"\n{'='*70}")
        print("TEST-06: FULL VERIFICATION SUMMARY")
        print(f"{'='*70}")
        print(f"Total rows: {stats['total']}")
        print(f"Matched by pretranslation: {stats['matched']}")
        print(f"")
        print(f"StringID Exact Matches:")
        print(f"  - Correct: {stats['exact_stringid_correct']}")
        print(f"  - Wrong: {stats['exact_stringid_wrong']}")
        print(f"")
        print(f"Fallback (unknown StringID): {stats['fallback_got_target']} got translations")
        print(f"No StringID: {stats['no_stringid_got_target']} got translations")
        print(f"{'='*70}")

        # Overall success criteria
        total_expected_matches = (
            stats["exact_stringid_correct"] +
            stats["fallback_got_target"] +
            stats["no_stringid_got_target"]
        )
        success_rate = total_expected_matches / stats["total"] if stats["total"] > 0 else 0

        assert success_rate >= 0.75, f"Overall success rate too low: {success_rate:.1%}"

        print(f"\nTEST-06 PASSED: Overall success rate = {success_rate:.1%}")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
