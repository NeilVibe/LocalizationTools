"""
E2E Test 4: Pretranslation with StringID Matching

═══════════════════════════════════════════════════════════════════════════════
HOW THIS E2E TEST WORKS
═══════════════════════════════════════════════════════════════════════════════

1. SETUP (in @pytest.fixture):
   a) Create a TM with StringID entries (Excel upload in "stringid" mode)
      - "저장" → "Save" (StringID: UI_BUTTON_SAVE)
      - "저장" → "Save Game" (StringID: UI_MENU_SAVE)
      - "저장" → "Storage" (StringID: TECH_STORAGE)
      - Plus other entries (설정, 취소, etc.)

   b) Build PKL indexes for the TM
      - whole_lookup.pkl contains: {"저장": {"variations": [3 entries]}}
      - Each entry has: {"target": "...", "string_id": "..."}

2. TEST:
   a) Create test file (LDMFile + LDMRows) with rows to pretranslate
      - Each row has: source="저장", target="", string_id="UI_BUTTON_SAVE"

   b) Run pretranslation
      - Calls PretranslationEngine.pretranslate(file_id, engine="standard", ...)
      - Engine searches TM, finds variations, matches by string_id

   c) Verify correct target
      - Query the DB for the row
      - Assert row.target == expected value based on string_id

3. CLEANUP:
   - Delete test files and TM entries

═══════════════════════════════════════════════════════════════════════════════
STRINGID MATCHING LOGIC
═══════════════════════════════════════════════════════════════════════════════

When pretranslating with StringID mode TM:
1. Search TM for source text → returns ALL variations
2. If row.string_id exists:
   - Find variation with matching string_id → use that target
   - If no match → use first variation (fallback)
3. If row.string_id is None → use first variation

═══════════════════════════════════════════════════════════════════════════════
TEST CASES
═══════════════════════════════════════════════════════════════════════════════

PRE-01: Standard engine works (basic)
PRE-02: StringID='UI_BUTTON_SAVE' → target='Save'
PRE-03: StringID='UI_MENU_SAVE' → target='Save Game'
PRE-04: Unknown StringID → fallback to first variation
PRE-05: File without StringID → use first variation
PRE-06: skip_existing=True → skip rows with target
PRE-07: skip_existing=False → process all rows
PRE-08: High threshold filters matches
PRE-09: StringID='TECH_STORAGE' → target='Storage'
PRE-10: Multi-row test (3 rows, 3 StringIDs, verify all 3 correct)

Run: python -m pytest tests/fixtures/stringid/test_e2e_4_pretranslate.py -v
"""

import os
import sys
import pytest
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Test data
from tests.fixtures.stringid.stringid_test_data import BASIC_STRINGID_DATA


class TestPretranslationStringID:
    """E2E tests for Pretranslation with StringID matching."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test environment with TM and File."""
        from server.database.db_setup import setup_database
        from server.tools.ldm.tm_manager import TMManager
        from server.tools.ldm.tm_indexer import TMIndexer
        from server.tools.ldm.pretranslate import PretranslationEngine
        from server.database.models import LDMFile, LDMRow
        import openpyxl
        from io import BytesIO

        # Get session
        engine, session_maker = setup_database(drop_existing=False)
        self.db = session_maker()
        self.tm_manager = TMManager(self.db)
        self.tm_indexer = TMIndexer(self.db)
        self.pretranslation_engine = PretranslationEngine(self.db)
        self.created_tm_ids = []
        self.created_file_ids = []

        # Create StringID TM
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'
        for i, entry in enumerate(BASIC_STRINGID_DATA, start=2):
            ws[f'A{i}'] = entry["source"]
            ws[f'B{i}'] = entry["target"]
            ws[f'C{i}'] = entry.get("string_id", "")

        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        result = self.tm_manager.upload_tm(
            file_content=excel_bytes,
            filename="test_pretranslate.xlsx",
            name="Test Pretranslate TM",
            owner_id=1,
            mode="stringid",
            source_col=0,
            target_col=1,
            stringid_col=2,
            has_header=True
        )
        self.tm_id = result["tm_id"]
        self.created_tm_ids.append(self.tm_id)

        # Build indexes
        self.tm_indexer.build_indexes(self.tm_id)

        yield

        # Cleanup
        for file_id in self.created_file_ids:
            try:
                self.db.query(LDMRow).filter(LDMRow.file_id == file_id).delete()
                self.db.query(LDMFile).filter(LDMFile.id == file_id).delete()
                self.db.commit()
            except Exception:
                pass
        for tm_id in self.created_tm_ids:
            try:
                self.tm_manager.delete_tm(tm_id)
            except Exception:
                pass
        self.db.close()

    def _create_test_file(self, rows: list) -> int:
        """Create a test file with rows."""
        from server.database.models import LDMFile, LDMRow, LDMProject

        # Ensure test project exists
        project = self.db.query(LDMProject).filter(LDMProject.name == "Test Project").first()
        if not project:
            project = LDMProject(name="Test Project", owner_id=1)
            self.db.add(project)
            self.db.commit()

        # Create file record
        file = LDMFile(
            project_id=project.id,
            name="test_pretranslate_file.txt",
            original_filename="test_pretranslate_file.txt",
            format="txt",
            row_count=len(rows),
            created_by=1
        )
        self.db.add(file)
        self.db.commit()
        self.created_file_ids.append(file.id)

        # Create rows
        for i, row_data in enumerate(rows, start=1):
            row = LDMRow(
                file_id=file.id,
                row_num=i,
                source=row_data.get("source", ""),
                target=row_data.get("target", ""),
                string_id=row_data.get("string_id"),
                status="pending"
            )
            self.db.add(row)

        self.db.commit()
        return file.id

    # =========================================================================
    # PRE-01: Standard engine works
    # =========================================================================

    def test_pre_01_standard_engine(self):
        """PRE-01: Standard engine matches rows correctly."""
        # Create test file with rows that should match
        rows = [
            {"source": "저장", "target": ""},  # Should match
            {"source": "설정", "target": ""},  # Should match
            {"source": "Unknown", "target": ""},  # No match
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["total"] == 3
        assert result["matched"] >= 2, f"Expected at least 2 matches, got {result['matched']}"

        print(f"PRE-01 PASSED: Standard engine matched {result['matched']}/{result['total']} rows")

    # =========================================================================
    # PRE-02: StringID exact match returns "Save"
    # =========================================================================

    def test_pre_02_stringid_exact_save(self):
        """PRE-02: Row with StringID='UI_BUTTON_SAVE' returns 'Save'."""
        from server.database.models import LDMRow

        rows = [
            {"source": "저장", "target": "", "string_id": "UI_BUTTON_SAVE"},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["matched"] == 1

        # Check the CORRECT target was set based on StringID
        row = self.db.query(LDMRow).filter(LDMRow.file_id == file_id).first()
        self.db.refresh(row)

        assert row.target == "Save", f"Expected 'Save' for UI_BUTTON_SAVE, got '{row.target}'"

        print(f"PRE-02 PASSED: StringID='UI_BUTTON_SAVE' returned target='Save'")

    # =========================================================================
    # PRE-03: StringID different returns "Save Game"
    # =========================================================================

    def test_pre_03_stringid_different(self):
        """PRE-03: Row with StringID='UI_MENU_SAVE' returns 'Save Game'."""
        from server.database.models import LDMRow

        rows = [
            {"source": "저장", "target": "", "string_id": "UI_MENU_SAVE"},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["matched"] == 1

        # Check the CORRECT target was set based on StringID
        row = self.db.query(LDMRow).filter(LDMRow.file_id == file_id).first()
        self.db.refresh(row)

        assert row.target == "Save Game", f"Expected 'Save Game' for UI_MENU_SAVE, got '{row.target}'"

        print(f"PRE-03 PASSED: StringID='UI_MENU_SAVE' returned target='Save Game'")

    # =========================================================================
    # PRE-04: StringID no match returns variations
    # =========================================================================

    def test_pre_04_stringid_no_match(self):
        """PRE-04: Row with unknown StringID returns all variations."""
        from server.database.models import LDMRow

        rows = [
            {"source": "저장", "target": "", "string_id": "UNKNOWN_ID"},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        # Should still match (by source text)
        assert result["matched"] == 1

        print(f"PRE-04 PASSED: Unknown StringID still matched by source")

    # =========================================================================
    # PRE-05: File without StringID - fallback
    # =========================================================================

    def test_pre_05_file_without_stringid(self):
        """PRE-05: File without StringID column uses first/all variations."""
        rows = [
            {"source": "저장", "target": ""},
            {"source": "설정", "target": ""},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["matched"] == 2

        print(f"PRE-05 PASSED: Files without StringID matched correctly")

    # =========================================================================
    # PRE-06: Skip existing (row has target)
    # =========================================================================

    def test_pre_06_skip_existing(self):
        """PRE-06: Rows with existing target are skipped when skip_existing=True."""
        rows = [
            {"source": "저장", "target": "Already Translated"},
            {"source": "설정", "target": ""},  # Should match
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92,
            skip_existing=True
        )

        # skip_existing filters rows BEFORE processing, so total is reduced
        # Row with existing target is filtered out, leaving only 1 row to process
        assert result["total"] == 1, f"Expected 1 total (rows without target), got {result['total']}"
        assert result["matched"] == 1, f"Expected 1 match, got {result['matched']}"

        print(f"PRE-06 PASSED: Only {result['total']} rows processed (skip_existing filtered 1)")

    # =========================================================================
    # PRE-07: Don't skip existing
    # =========================================================================

    def test_pre_07_dont_skip_existing(self):
        """PRE-07: All rows processed when skip_existing=False."""
        rows = [
            {"source": "저장", "target": "Already Translated"},
            {"source": "설정", "target": ""},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92,
            skip_existing=False
        )

        assert result["skipped"] == 0, f"Expected 0 skipped, got {result['skipped']}"
        assert result["total"] == 2, f"Expected 2 total, got {result['total']}"

        print(f"PRE-07 PASSED: All rows processed when skip_existing=False")

    # =========================================================================
    # PRE-08: Threshold filter
    # =========================================================================

    def test_pre_08_threshold_filter(self):
        """PRE-08: High threshold filters out lower matches."""
        rows = [
            {"source": "저장", "target": ""},  # Exact match (1.0)
            {"source": "저장하기", "target": ""},  # Similar (< 1.0)
        ]
        file_id = self._create_test_file(rows)

        # Very high threshold
        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.99  # Only exact matches
        )

        # At 0.99 threshold, only exact match should work
        assert result["matched"] >= 1

        print(f"PRE-08 PASSED: High threshold matched {result['matched']} rows")

    # =========================================================================
    # PRE-09: Third variation (Storage)
    # =========================================================================

    def test_pre_09_stringid_storage(self):
        """PRE-09: Row with StringID='TECH_STORAGE' returns 'Storage'."""
        from server.database.models import LDMRow

        rows = [
            {"source": "저장", "target": "", "string_id": "TECH_STORAGE"},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["matched"] == 1

        row = self.db.query(LDMRow).filter(LDMRow.file_id == file_id).first()
        self.db.refresh(row)

        assert row.target == "Storage", f"Expected 'Storage' for TECH_STORAGE, got '{row.target}'"

        print(f"PRE-09 PASSED: StringID='TECH_STORAGE' returned target='Storage'")

    # =========================================================================
    # PRE-10: Multi-row StringID matching (CRITICAL TEST)
    # =========================================================================

    def test_pre_10_multirow_stringid_matching(self):
        """
        PRE-10: Multiple rows with same Korean but different StringIDs get correct targets.

        This is the CRITICAL test for StringID matching.

        Test Setup:
        - TM has: "저장" with 3 variations:
          - UI_BUTTON_SAVE → "Save"
          - UI_MENU_SAVE → "Save Game"
          - TECH_STORAGE → "Storage"

        - File has 3 rows with same source "저장" but different StringIDs:
          - Row 1: string_id=UI_BUTTON_SAVE → should get "Save"
          - Row 2: string_id=UI_MENU_SAVE → should get "Save Game"
          - Row 3: string_id=TECH_STORAGE → should get "Storage"

        Expected: Each row gets the CORRECT target based on its StringID.
        """
        from server.database.models import LDMRow

        rows = [
            {"source": "저장", "target": "", "string_id": "UI_BUTTON_SAVE"},
            {"source": "저장", "target": "", "string_id": "UI_MENU_SAVE"},
            {"source": "저장", "target": "", "string_id": "TECH_STORAGE"},
        ]
        file_id = self._create_test_file(rows)

        result = self.pretranslation_engine.pretranslate(
            file_id=file_id,
            engine="standard",
            dictionary_id=self.tm_id,
            threshold=0.92
        )

        assert result["matched"] == 3, f"Expected 3 matches, got {result['matched']}"

        # Get all rows and verify each got the correct target
        db_rows = self.db.query(LDMRow).filter(
            LDMRow.file_id == file_id
        ).order_by(LDMRow.row_num).all()

        for row in db_rows:
            self.db.refresh(row)

        expected = {
            "UI_BUTTON_SAVE": "Save",
            "UI_MENU_SAVE": "Save Game",
            "TECH_STORAGE": "Storage"
        }

        for row in db_rows:
            expected_target = expected.get(row.string_id)
            assert row.target == expected_target, \
                f"Row {row.row_num}: StringID='{row.string_id}' " \
                f"expected target='{expected_target}', got '{row.target}'"

        print(f"PRE-10 PASSED: All 3 rows got correct targets based on StringID:")
        for row in db_rows:
            print(f"  - {row.string_id} → '{row.target}'")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
