"""
Fixture Generator for StringID Tests

Generates test fixture files in multiple formats:
- TXT (tab-separated, matches sampleofLanguageData.txt format)
- XML (LanguageData format)
- Excel (XLSX with 3 columns)

Usage:
    python generate_fixtures.py
"""

import os
import sys
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from tests.fixtures.stringid.stringid_test_data import (
    ALL_STRINGID_TEST_DATA,
    BASIC_STRINGID_DATA,
    STANDARD_TM_DATA,
)


def generate_txt_file(data: list, output_path: Path, include_stringid: bool = True):
    """
    Generate TXT file in sampleofLanguageData.txt format.

    Format (tab-separated):
    Col 0: Flag (0)
    Col 1: StringID
    Col 2-4: Metadata (0, 0, 1)
    Col 5: Korean source text
    Col 6: Target translation
    Col 7: Status ("Current")
    """
    lines = []
    for entry in data:
        string_id = entry.get("string_id", "") or ""
        source = entry["source"]
        target = entry["target"]

        if include_stringid:
            # Full format with StringID
            line = f"0\t{string_id}\t0\t0\t1\t{source}\t{target}\tCurrent"
        else:
            # Standard format without StringID (col 1 = numeric ID)
            line = f"0\t0\t0\t0\t1\t{source}\t{target}\tCurrent"

        lines.append(line)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"Generated TXT: {output_path} ({len(lines)} entries)")


def generate_xml_file(data: list, output_path: Path):
    """
    Generate XML file in LocStr format (what xml_handler.py expects).

    Format:
    <LanguageData>
        <LocStr StringId="xxx" StrOrigin="Korean text" Str="Translation" />
    </LanguageData>
    """
    lines = ['<?xml version="1.0" encoding="utf-8"?>', '<LanguageData>']

    for entry in data:
        string_id = escape_xml(entry.get("string_id", "NO_ID") or "NO_ID")
        source = escape_xml(entry["source"])
        target = escape_xml(entry["target"])

        # Use LocStr element with attributes (matching xml_handler.py expected format)
        lines.append(f'    <LocStr StringId="{string_id}" StrOrigin="{source}" Str="{target}" />')

    lines.append('</LanguageData>')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"Generated XML: {output_path} ({len(data)} entries)")


def escape_xml(text: str) -> str:
    """Escape special XML characters."""
    return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&apos;'))


def generate_excel_file(data: list, output_path: Path, include_stringid: bool = True):
    """
    Generate Excel file with 2 or 3 columns.

    3-column format (StringID mode):
    Col A: Source
    Col B: Target
    Col C: StringID

    2-column format (Standard mode):
    Col A: Source
    Col B: Target
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Excel generation")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Headers
    if include_stringid:
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'
        ws['C1'] = 'StringID'
    else:
        ws['A1'] = 'Source'
        ws['B1'] = 'Target'

    # Data rows
    for i, entry in enumerate(data, start=2):
        ws[f'A{i}'] = entry["source"]
        ws[f'B{i}'] = entry["target"]
        if include_stringid:
            ws[f'C{i}'] = entry.get("string_id", "") or ""

    wb.save(output_path)
    print(f"Generated Excel: {output_path} ({len(data)} entries, {'3-col' if include_stringid else '2-col'})")


def main():
    """Generate all fixture files."""
    fixture_dir = Path(__file__).parent

    print("=" * 60)
    print("Generating StringID Test Fixtures")
    print("=" * 60)

    # 1. StringID TXT (full format)
    generate_txt_file(
        ALL_STRINGID_TEST_DATA,
        fixture_dir / "stringid_test_data.txt",
        include_stringid=True
    )

    # 2. Standard TXT (no StringID)
    generate_txt_file(
        STANDARD_TM_DATA,
        fixture_dir / "standard_test_data.txt",
        include_stringid=False
    )

    # 3. StringID XML
    generate_xml_file(
        ALL_STRINGID_TEST_DATA,
        fixture_dir / "stringid_test_data.xml"
    )

    # 4. StringID Excel (3-column)
    generate_excel_file(
        ALL_STRINGID_TEST_DATA,
        fixture_dir / "stringid_test_data.xlsx",
        include_stringid=True
    )

    # 5. Standard Excel (2-column)
    generate_excel_file(
        STANDARD_TM_DATA,
        fixture_dir / "standard_test_data.xlsx",
        include_stringid=False
    )

    # 6. Basic StringID data only (smaller test set)
    generate_txt_file(
        BASIC_STRINGID_DATA,
        fixture_dir / "basic_stringid_data.txt",
        include_stringid=True
    )

    generate_xml_file(
        BASIC_STRINGID_DATA,
        fixture_dir / "basic_stringid_data.xml"
    )

    generate_excel_file(
        BASIC_STRINGID_DATA,
        fixture_dir / "basic_stringid_data.xlsx",
        include_stringid=True
    )

    print("=" * 60)
    print("Done! Generated fixture files:")
    for f in fixture_dir.glob("*.txt"):
        print(f"  - {f.name}")
    for f in fixture_dir.glob("*.xml"):
        print(f"  - {f.name}")
    for f in fixture_dir.glob("*.xlsx"):
        print(f"  - {f.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
