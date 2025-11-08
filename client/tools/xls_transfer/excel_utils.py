"""
XLSTransfer Excel Utilities

Excel file reading, writing, and manipulation operations.
CLEAN, modular functions for all Excel file operations.
"""

from typing import List, Tuple, Dict, Optional
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string
from copy import copy
from loguru import logger

from client.tools.xls_transfer import config
from client.tools.xls_transfer.core import clean_text, excel_column_to_index, convert_cell_value, count_newlines
from client.utils.file_handler import create_temp_copy, ensure_output_path


# ============================================
# Excel Reading
# ============================================

def read_excel_columns(
    file_path: str,
    sheet_name: str,
    kr_column: str,
    trans_column: Optional[str] = None,
    header: Optional[int] = None
) -> Tuple[List[str], List[str]]:
    """
    Read Korean and translation columns from Excel file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read
        kr_column: Korean column letter (e.g., 'A')
        trans_column: Translation column letter (e.g., 'B'), optional
        header: Row to use as header (None = no header)

    Returns:
        Tuple of (kr_texts, trans_texts) as lists

    Example:
        >>> kr_texts, trans_texts = read_excel_columns("data.xlsx", "Sheet1", "A", "B")
        >>> len(kr_texts)
        100
    """
    logger.info(f"Reading Excel: {Path(file_path).name} - Sheet: {sheet_name}")

    # Read Excel
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)

    # Get column indices
    kr_idx = excel_column_to_index(kr_column)

    # Extract Korean texts
    kr_texts = df.iloc[:, kr_idx].apply(clean_text).dropna().tolist()

    # Extract translation texts if column specified
    trans_texts = []
    if trans_column:
        trans_idx = excel_column_to_index(trans_column)
        trans_texts = df.iloc[:, trans_idx].apply(clean_text).dropna().tolist()

    logger.info(f"Read {len(kr_texts)} Korean texts, {len(trans_texts)} translations")

    return kr_texts, trans_texts


def get_sheet_names(file_path: str) -> List[str]:
    """
    Get list of sheet names from Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names

    Example:
        >>> sheets = get_sheet_names("data.xlsx")
        >>> print(sheets)
        ['Sheet1', 'Sheet2', 'Sheet3']
    """
    xl_file = pd.ExcelFile(file_path)
    return xl_file.sheet_names


# ============================================
# Excel Writing
# ============================================

def write_translations_to_excel(
    input_file: str,
    sheet_name: str,
    kr_column: str,
    trans_column: str,
    translations: Dict[int, str],
    output_file: Optional[str] = None
) -> str:
    """
    Write translations to Excel file.

    Args:
        input_file: Input Excel file path
        sheet_name: Sheet name to write to
        kr_column: Korean column letter
        trans_column: Translation column letter to write
        translations: Dictionary mapping row index to translation
        output_file: Output file path (auto-generated if None)

    Returns:
        Path to output file

    Example:
        >>> translations = {1: "Hello", 2: "Thanks"}
        >>> output = write_translations_to_excel("input.xlsx", "Sheet1", "A", "B", translations)
    """
    logger.info(f"Writing translations to {Path(input_file).name}")

    # Generate output path if not specified
    if output_file is None:
        output_file = ensure_output_path(input_file, config.OUTPUT_SUFFIXES['translate_excel'])

    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    sheet = wb[sheet_name]

    # Get column indices
    kr_col_idx = column_index_from_string(kr_column)
    trans_col_idx = column_index_from_string(trans_column)

    # Write translations
    for row_idx, translation in translations.items():
        if translation:  # Only write non-empty translations
            cell = sheet.cell(row=row_idx, column=trans_col_idx)
            cell.value = clean_text(translation)

    # Save
    wb.save(output_file)
    wb.close()

    logger.info(f"Saved {len(translations)} translations to {output_file}")

    return output_file


# ============================================
# Newline Checking
# ============================================

def check_newline_mismatches(
    file_path: str,
    sheet_name: str,
    kr_column: str,
    trans_column: str
) -> List[Dict[str, any]]:
    """
    Check for newline count mismatches between Korean and translation.

    Args:
        file_path: Excel file path
        sheet_name: Sheet name
        kr_column: Korean column letter
        trans_column: Translation column letter

    Returns:
        List of dicts with mismatch info:
            - row: Row number
            - kr_text: Korean text
            - trans_text: Translation text
            - kr_newlines: Newline count in Korean
            - trans_newlines: Newline count in translation

    Example:
        >>> mismatches = check_newline_mismatches("data.xlsx", "Sheet1", "A", "B")
        >>> len(mismatches)
        5
    """
    logger.info(f"Checking newline mismatches in {Path(file_path).name}")

    mismatches = []

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Get column indices
    kr_col_idx = column_index_from_string(kr_column)
    trans_col_idx = column_index_from_string(trans_column)

    # Check each row
    for row_idx in range(1, sheet.max_row + 1):
        kr_cell = sheet.cell(row=row_idx, column=kr_col_idx)
        trans_cell = sheet.cell(row=row_idx, column=trans_col_idx)

        kr_text = kr_cell.value
        trans_text = trans_cell.value

        # Skip if either is None
        if not kr_text or not trans_text:
            continue

        kr_newlines = count_newlines(str(kr_text))
        trans_newlines = count_newlines(str(trans_text))

        # Check for mismatch
        if kr_newlines != trans_newlines:
            mismatches.append({
                'row': row_idx,
                'kr_text': str(kr_text)[:100],  # Truncate for display
                'trans_text': str(trans_text)[:100],
                'kr_newlines': kr_newlines,
                'trans_newlines': trans_newlines
            })

    wb.close()

    logger.info(f"Found {len(mismatches)} newline mismatches")

    return mismatches


# ============================================
# Excel Combining
# ============================================

def combine_excel_files(
    file_paths: List[str],
    output_file: Optional[str] = None
) -> str:
    """
    Combine multiple Excel files into one.

    Args:
        file_paths: List of Excel file paths to combine
        output_file: Output file path (auto-generated if None)

    Returns:
        Path to combined Excel file

    Example:
        >>> files = ["file1.xlsx", "file2.xlsx", "file3.xlsx"]
        >>> combined = combine_excel_files(files)
    """
    logger.info(f"Combining {len(file_paths)} Excel files")

    if not file_paths:
        raise ValueError("No files provided to combine")

    # Generate output path if not specified
    if output_file is None:
        output_file = ensure_output_path(file_paths[0], config.OUTPUT_SUFFIXES['combine_excel'])

    # Read all DataFrames
    all_dfs = []
    for file_path in file_paths:
        logger.info(f"Reading {Path(file_path).name}")
        xl_file = pd.ExcelFile(file_path)

        # Read all sheets
        for sheet_name in xl_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            all_dfs.append(df)

    # Concatenate all DataFrames
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # Save to Excel
    combined_df.to_excel(output_file, index=False, header=False)

    logger.info(f"Combined {len(file_paths)} files into {output_file}")
    logger.info(f"Total rows: {len(combined_df)}")

    return output_file


# ============================================
# Newline Auto-Adaptation
# ============================================

def auto_adapt_newlines(
    file_path: str,
    sheet_name: str,
    kr_column: str,
    trans_column: str,
    output_file: Optional[str] = None
) -> Tuple[str, int]:
    """
    Automatically adapt newlines in translation to match Korean text.

    Args:
        file_path: Excel file path
        sheet_name: Sheet name
        kr_column: Korean column letter
        trans_column: Translation column letter
        output_file: Output file path (auto-generated if None)

    Returns:
        Tuple of (output_file_path, num_adapted)

    Example:
        >>> output, count = auto_adapt_newlines("data.xlsx", "Sheet1", "A", "B")
        >>> print(f"Adapted {count} rows")
    """
    logger.info(f"Auto-adapting newlines in {Path(file_path).name}")

    # Generate output path if not specified
    if output_file is None:
        output_file = ensure_output_path(file_path, config.OUTPUT_SUFFIXES['newline_adapt'])

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Get column indices
    kr_col_idx = column_index_from_string(kr_column)
    trans_col_idx = column_index_from_string(trans_column)

    num_adapted = 0

    # Process each row
    for row_idx in range(1, sheet.max_row + 1):
        kr_cell = sheet.cell(row=row_idx, column=kr_col_idx)
        trans_cell = sheet.cell(row=row_idx, column=trans_col_idx)

        kr_text = kr_cell.value
        trans_text = trans_cell.value

        # Skip if either is None
        if not kr_text or not trans_text:
            continue

        kr_text_str = str(kr_text)
        trans_text_str = str(trans_text)

        kr_newlines = count_newlines(kr_text_str)
        trans_newlines = count_newlines(trans_text_str)

        # Adapt if mismatch
        if kr_newlines != trans_newlines:
            # Simple approach: split translation and rejoin with same newline count
            trans_words = trans_text_str.replace('\n', ' ').split()

            if kr_newlines > 0 and trans_words:
                # Distribute words across lines
                words_per_line = max(1, len(trans_words) // (kr_newlines + 1))
                adapted_lines = []

                for i in range(0, len(trans_words), words_per_line):
                    line_words = trans_words[i:i + words_per_line]
                    adapted_lines.append(' '.join(line_words))

                adapted_text = '\n'.join(adapted_lines)
                trans_cell.value = adapted_text
                num_adapted += 1

    # Save
    wb.save(output_file)
    wb.close()

    logger.info(f"Adapted {num_adapted} rows, saved to {output_file}")

    return output_file, num_adapted


# ============================================
# Simple Transfer
# ============================================

def simple_excel_transfer(
    source_file: str,
    dest_file: str,
    transfers: List[Dict[str, str]],
    output_file: Optional[str] = None
) -> str:
    """
    Transfer data from source Excel to destination Excel.

    Args:
        source_file: Source Excel file path
        dest_file: Destination Excel file path
        transfers: List of transfer configs with keys:
            - source_sheet: Source sheet name
            - source_column: Source column letter
            - dest_sheet: Destination sheet name
            - dest_column: Destination column letter
        output_file: Output file path (auto-generated if None)

    Returns:
        Path to output file

    Example:
        >>> transfers = [{
        ...     'source_sheet': 'Sheet1',
        ...     'source_column': 'B',
        ...     'dest_sheet': 'Sheet2',
        ...     'dest_column': 'C'
        ... }]
        >>> output = simple_excel_transfer("source.xlsx", "dest.xlsx", transfers)
    """
    logger.info(f"Transferring data from {Path(source_file).name} to {Path(dest_file).name}")

    # Generate output path if not specified
    if output_file is None:
        output_file = ensure_output_path(dest_file, config.OUTPUT_SUFFIXES['simple_transfer'])

    # Load destination workbook
    dest_wb = openpyxl.load_workbook(dest_file)

    # Process each transfer
    for transfer_config in transfers:
        source_sheet_name = transfer_config['source_sheet']
        source_col = transfer_config['source_column']
        dest_sheet_name = transfer_config['dest_sheet']
        dest_col = transfer_config['dest_column']

        logger.info(f"Transfer: {source_sheet_name}:{source_col} -> {dest_sheet_name}:{dest_col}")

        # Read source data
        source_df = pd.read_excel(source_file, sheet_name=source_sheet_name, header=None)
        source_col_idx = excel_column_to_index(source_col)
        source_data = source_df.iloc[:, source_col_idx].apply(convert_cell_value).tolist()

        # Write to destination
        dest_sheet = dest_wb[dest_sheet_name]
        dest_col_idx = column_index_from_string(dest_col)

        for row_idx, value in enumerate(source_data, start=1):
            cell = dest_sheet.cell(row=row_idx, column=dest_col_idx)
            cell.value = value

    # Save
    dest_wb.save(output_file)
    dest_wb.close()

    logger.info(f"Transfer complete: {len(transfers)} transfers saved to {output_file}")

    return output_file


# ============================================
# Excel Validation
# ============================================

def validate_excel_file(file_path: str) -> Tuple[bool, str]:
    """
    Validate that file is a valid Excel file.

    Args:
        file_path: Path to file

    Returns:
        Tuple of (is_valid, error_message)

    Example:
        >>> is_valid, error = validate_excel_file("data.xlsx")
        >>> print(is_valid)
        True
    """
    file_path = Path(file_path)

    if not file_path.exists():
        return False, f"File not found: {file_path}"

    if file_path.suffix.lower() not in config.SUPPORTED_EXCEL_EXTENSIONS:
        return False, f"Invalid file type. Expected: {', '.join(config.SUPPORTED_EXCEL_EXTENSIONS)}"

    try:
        # Try to read the file
        pd.ExcelFile(str(file_path))
        return True, ""
    except Exception as e:
        return False, f"Error reading Excel file: {str(e)}"


# Example usage
if __name__ == "__main__":
    print("Excel Utils Module")
    print("=" * 50)
    print("Functions available:")
    print("- read_excel_columns(): Read KR and translation columns")
    print("- get_sheet_names(): Get all sheet names from file")
    print("- write_translations_to_excel(): Write translations to file")
    print("- check_newline_mismatches(): Find newline mismatches")
    print("- combine_excel_files(): Combine multiple Excel files")
    print("- auto_adapt_newlines(): Auto-fix newline mismatches")
    print("- simple_excel_transfer(): Transfer data between files")
    print("- validate_excel_file(): Validate Excel file")
