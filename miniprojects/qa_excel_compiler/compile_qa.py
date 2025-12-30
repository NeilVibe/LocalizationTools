#!/usr/bin/env python3
"""
QA Excel Compiler
=================
Compiles QA tester Excel files into master sheets.

Usage:
    python3 compile_qa.py

Input:  QAfolder/{Username}_{Category}.xlsx
Output: Masterfolder/Master_{Category}.xlsx

Categories: Quest, Knowledge, Item, Node, System
"""

import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter


# === CONFIGURATION ===
SCRIPT_DIR = Path(__file__).parent
QA_FOLDER = SCRIPT_DIR / "QAfolder"
MASTER_FOLDER = SCRIPT_DIR / "Masterfolder"
CATEGORIES = ["Quest", "Knowledge", "Item", "Node", "System"]

# Column indices (1-based for openpyxl)
COL_STATUS = 5      # E
COL_COMMENT = 6     # F
COL_SCREENSHOT = 7  # G

# Valid STATUS values (only these count as "filled")
VALID_STATUS = ["ISSUE", "NO ISSUE", "BLOCKED"]


def discover_qa_files():
    """
    Find all QA Excel files and parse their metadata.

    Returns: List of dicts with {filepath, username, category}
    """
    files = []

    if not QA_FOLDER.exists():
        print(f"ERROR: QAfolder not found: {QA_FOLDER}")
        return files

    for f in QA_FOLDER.glob("*.xlsx"):
        # Skip temp files
        if f.name.startswith("~$"):
            continue

        # Parse: Username_Category.xlsx
        parts = f.stem.split("_")
        if len(parts) >= 2:
            username = parts[0]
            category = parts[1]

            if category in CATEGORIES:
                files.append({
                    "filepath": f,
                    "username": username,
                    "category": category
                })
            else:
                print(f"WARN: Unknown category '{category}' in {f.name}")
        else:
            print(f"WARN: Invalid filename format: {f.name} (expected: Username_Category.xlsx)")

    return files


def ensure_master_folder():
    """Create Masterfolder if it doesn't exist."""
    MASTER_FOLDER.mkdir(exist_ok=True)


def get_or_create_master(category, template_file=None):
    """
    Load existing master file or create from template.

    Args:
        category: Category name (Quest, Knowledge, etc.)
        template_file: Path to first QA file to use as template

    Returns: openpyxl Workbook
    """
    master_path = MASTER_FOLDER / f"Master_{category}.xlsx"

    if master_path.exists():
        print(f"  Loading existing: {master_path.name}")
        return openpyxl.load_workbook(master_path), master_path
    elif template_file:
        print(f"  Creating new master from: {template_file.name}")
        wb = openpyxl.load_workbook(template_file)

        # Clear STATUS, COMMENT, SCREENSHOT columns (keep structure)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=COL_STATUS).value = None
                ws.cell(row=row, column=COL_COMMENT).value = None
                ws.cell(row=row, column=COL_SCREENSHOT).value = None

        return wb, master_path
    else:
        print(f"  ERROR: No template file for {category}")
        return None, master_path


def get_user_column(ws, username, col_type="COMMENT"):
    """
    Find or create {col_type}_{username} column.

    Args:
        ws: Worksheet
        username: User identifier
        col_type: "COMMENT" or "STATUS"

    Returns: Column index (1-based)
    """
    col_name = f"{col_type}_{username}"
    base_col = COL_STATUS if col_type == "STATUS" else COL_COMMENT

    # Check if column already exists
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == col_name:
            return col

    # Find last column of this type or use position after base column
    last_col = base_col
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).startswith(f"{col_type}_"):
            last_col = col

    # Insert new column after last column of this type
    new_col = last_col + 1
    ws.insert_cols(new_col)
    ws.cell(row=1, column=new_col).value = col_name

    print(f"    Created column: {col_name} at {get_column_letter(new_col)}")
    return new_col


def format_comment(new_comment, existing_comment=None):
    """
    Format comment with datetime, append to existing if present.

    Format: "comment text" (date: YYMMDD HHMM)

    Duplicate check: If the exact comment text already exists (ignoring date),
    we skip to avoid duplicating on re-runs.
    """
    if not new_comment or str(new_comment).strip() == "":
        return existing_comment

    new_text = str(new_comment).strip()

    # Check if this exact comment already exists (avoid duplicates on re-run)
    if existing_comment:
        existing_str = str(existing_comment)
        # Extract comment texts (between quotes) from existing
        if f'"{new_text}"' in existing_str:
            # Comment already exists, skip
            return existing_comment

    timestamp = datetime.now().strftime("%y%m%d %H%M")
    formatted = f'"{new_text}" (date: {timestamp})'

    if existing_comment and str(existing_comment).strip():
        # New on top, old below
        return f"{formatted}\n\n{existing_comment}"
    else:
        return formatted


def process_sheet(master_ws, qa_ws, username):
    """
    Process a single sheet: copy STATUS and COMMENT from QA to master.

    Returns: Dict with counts {status: n, comments: n}
    """
    # Get or create user's columns (STATUS first, then COMMENT)
    status_col = get_user_column(master_ws, username, "STATUS")
    comment_col = get_user_column(master_ws, username, "COMMENT")

    counts = {"status": 0, "comments": 0}
    max_row = min(master_ws.max_row, qa_ws.max_row)

    for row in range(2, max_row + 1):  # Skip header
        # Get QA STATUS
        qa_status = qa_ws.cell(row=row, column=COL_STATUS).value
        if qa_status and str(qa_status).strip().upper() in VALID_STATUS:
            # Overwrite status (latest wins)
            master_ws.cell(row=row, column=status_col).value = str(qa_status).strip().upper()
            counts["status"] += 1

        # Get QA COMMENT
        qa_comment = qa_ws.cell(row=row, column=COL_COMMENT).value
        if qa_comment and str(qa_comment).strip():
            # Get existing comment in master
            existing = master_ws.cell(row=row, column=comment_col).value

            # Format and update (appends if different)
            new_value = format_comment(qa_comment, existing)
            master_ws.cell(row=row, column=comment_col).value = new_value
            counts["comments"] += 1

    return counts


def calculate_completion(ws, status_col):
    """
    Calculate completion % based on STATUS column.

    Completion = rows with valid STATUS (ISSUE/NO ISSUE/BLOCKED) / total rows

    Returns: Percentage (0-100)
    """
    total = 0
    filled = 0

    for row in range(2, ws.max_row + 1):
        total += 1
        value = ws.cell(row=row, column=status_col).value
        if value and str(value).strip().upper() in VALID_STATUS:
            filled += 1

    if total == 0:
        return 100.0

    return round(filled / total * 100, 1)


def update_status_sheet(wb, users, sheet_stats):
    """
    Create/update STATUS sheet with completion tracking.

    Args:
        wb: Master workbook
        users: Set of usernames
        sheet_stats: {sheet_name: {username: percentage}}
    """
    # Create or clear STATUS sheet
    if "STATUS" in wb.sheetnames:
        del wb["STATUS"]

    ws = wb.create_sheet("STATUS")

    # Get sheet names (excluding STATUS)
    sheets = [s for s in sheet_stats.keys()]

    # Header row
    headers = ["User"] + sheets + ["Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header

    # Data rows
    for row, user in enumerate(sorted(users), 2):
        ws.cell(row=row, column=1).value = user

        percentages = []
        for col, sheet in enumerate(sheets, 2):
            pct = sheet_stats.get(sheet, {}).get(user, 0.0)
            ws.cell(row=row, column=col).value = f"{pct}%"
            percentages.append(pct)

        # Total average
        avg = round(sum(percentages) / len(percentages), 1) if percentages else 0.0
        ws.cell(row=row, column=len(headers)).value = f"{avg}%"

    print(f"  Updated STATUS sheet with {len(users)} users")


def process_category(category, qa_files):
    """
    Process all QA files for one category.
    """
    print(f"\n{'='*50}")
    print(f"Processing: {category} ({len(qa_files)} files)")
    print(f"{'='*50}")

    # Get or create master
    first_file = qa_files[0]["filepath"]
    master_wb, master_path = get_or_create_master(category, first_file)

    if master_wb is None:
        return

    # Track users and stats
    all_users = set()
    sheet_stats = defaultdict(dict)  # {sheet: {user: pct}}

    # Process each QA file
    for qf in qa_files:
        username = qf["username"]
        filepath = qf["filepath"]
        all_users.add(username)

        print(f"\n  File: {filepath.name}")

        qa_wb = openpyxl.load_workbook(filepath)

        for sheet_name in qa_wb.sheetnames:
            # Skip STATUS sheets
            if sheet_name == "STATUS":
                continue

            # Check if sheet exists in master
            if sheet_name not in master_wb.sheetnames:
                print(f"    WARN: Sheet '{sheet_name}' not in master, skipping")
                continue

            # Process sheet
            counts = process_sheet(master_wb[sheet_name], qa_wb[sheet_name], username)
            print(f"    {sheet_name}: {counts['status']} status, {counts['comments']} comments")

            # Calculate completion (based on STATUS column)
            status_col = get_user_column(master_wb[sheet_name], username, "STATUS")
            pct = calculate_completion(master_wb[sheet_name], status_col)
            sheet_stats[sheet_name][username] = pct

        qa_wb.close()

    # Update STATUS sheet
    update_status_sheet(master_wb, all_users, sheet_stats)

    # Save master
    master_wb.save(master_path)
    print(f"\n  Saved: {master_path}")


def main():
    """Main entry point."""
    print("="*60)
    print("QA Excel Compiler")
    print("="*60)

    # Ensure folders exist
    ensure_master_folder()

    # Discover QA files
    qa_files = discover_qa_files()

    if not qa_files:
        print("\nNo valid QA files found in QAfolder/")
        print("Expected format: Username_Category.xlsx")
        print(f"Valid categories: {', '.join(CATEGORIES)}")
        return

    print(f"\nFound {len(qa_files)} QA file(s)")

    # Group by category
    by_category = defaultdict(list)
    for qf in qa_files:
        by_category[qf["category"]].append(qf)

    # Process each category
    for category in CATEGORIES:
        if category in by_category:
            process_category(category, by_category[category])
        else:
            print(f"\nSKIP: No files for category '{category}'")

    print("\n" + "="*60)
    print("Compilation complete!")
    print("="*60)


if __name__ == "__main__":
    main()
