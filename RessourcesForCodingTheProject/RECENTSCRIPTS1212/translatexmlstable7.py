#!/usr/bin/env python3
import os
import re
import time
import html
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
import threading
import openpyxl

# ---- GLOBALS ----
xml_folder_path = None
xml_file_path   = None
tmx_file_path   = None

btn_xml_folder     = None
btn_tmx_file       = None
btn_simple         = None
btn_two            = None
btn_excel_simple   = None
btn_excel_2step    = None
btn_check          = None
btn_insert_excel   = None
btn_id             = None
status_label       = None
info_panel         = None
mode_var           = None

# ---- CORE NORMALIZATION FUNCTION ----
def normalize_text(txt):
    """
    Ensures consistent text normalization:
    1. Unescape HTML entities (&lt; → <, &amp; → &, etc.)
    2. Strip leading/trailing whitespace
    3. Collapse all internal whitespace (spaces, tabs, newlines) to single space
    
    This guarantees TMX and XML text are IDENTICAL for matching.
    """
    if not txt:
        return ""
    # Unescape HTML entities
    txt = html.unescape(str(txt))
    # Collapse all whitespace to single space and strip
    txt = re.sub(r'\s+', ' ', txt.strip())
    return txt

# ---- CRITICAL: TMX PREPROCESSING (FROM OLD CODE) ----
def preprocess_tmx_content(raw_content):
    """
    Pre-process TMX content BEFORE any XML parsing.
    Replace newlines inside <seg> tags with &lt;br/&gt;
    This ensures lxml treats <br/> as ESCAPED TEXT, not XML elements.
    """
    def replace_in_seg(match):
        seg_content = match.group(1)
        # Replace newlines with &lt;br/&gt;
        cleaned = seg_content.replace('\n', '&lt;br/&gt;')
        cleaned = cleaned.replace('\r\n', '&lt;br/&gt;')
        return f'<seg>{cleaned}</seg>'
    
    # Apply regex to find and clean all <seg> contents
    cleaned_content = re.sub(
        r'<seg>(.*?)</seg>', 
        replace_in_seg, 
        raw_content, 
        flags=re.DOTALL
    )
    
    return cleaned_content

# ---- HELPERS ----
def get_all_xml_files(folder):
    files = []
    for dp, _, fns in os.walk(folder):
        for fn in fns:
            if fn.lower().endswith(".xml"):
                files.append(os.path.join(dp, fn))
    return files

def get_xml_files():
    if mode_var.get() == "folder":
        return get_all_xml_files(xml_folder_path) if xml_folder_path else []
    else:
        return [xml_file_path] if xml_file_path else []

def parse_tmx(tmx_path):
    """
    Parse TMX and return two lists of dicts:
      - translation_units: normal segments (context, kr, en)
      - desc_units       : description segments (context, kr, en)

    This version PRESERVES the raw '<seg>' text (including '&desc;' and '&lt;br/&gt;')
    so that matching against XML @DescOrigin will succeed.
    """
    print(f"[TMX] Parsing TMX file: {tmx_path}")
    # 1) read raw file
    raw = open(tmx_path, "r", encoding="utf-8", errors="ignore").read()
    # 2) escape any real <br/> in <seg> so they stay text
    pre = preprocess_tmx_content(raw)
    # 3) parse with recovery
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                             no_network=True, recover=True)
    root = etree.fromstring(pre.encode("utf-8"), parser=parser)
    body = root.find("body")
    if body is None:
        raise RuntimeError("TMX missing <body>")
    # 4) walk TUs
    translation_units = []
    desc_units        = []
    for tu in body.findall("tu"):
        ctx = None; kr = en = desc_ko = None
        # context = StringId
        for prop in tu.findall("prop"):
            if prop.get("type") == "x-context" and prop.text:
                ctx = prop.text.strip()
                break
        if not ctx:
            continue
        for tuv in tu.findall("tuv"):
            # pick up lang
            lang = (
                tuv.get("{http://www.w3.org/XML/1998/namespace}lang")
                or tuv.get("xml:lang")
                or tuv.get("lang")
                or ""
            ).lower()
            seg = tuv.find("seg")
            if seg is None:
                continue
            # take raw text (includes &desc; or &lt;br/&gt;)
            txt = seg.text.strip() if seg.text else ""
            if lang.startswith("ko"):
                # description marker?
                if txt.lower().startswith("&desc;") or txt.lower().startswith("&amp;desc;"):
                    desc_ko = txt
                else:
                    kr = txt
            elif lang.startswith("en"):
                en = seg.text.strip() if seg.text else ""
        # classify
        if desc_ko:
            desc_units.append({"context": ctx, "kr": desc_ko, "en": en})
        elif kr is not None or en is not None:
            translation_units.append({"context": ctx, "kr": kr, "en": en})
    print(f"[TMX]   normal units: {len(translation_units)}   description units: {len(desc_units)}")
    return translation_units, desc_units

def robust_parse_xml(path):
    """Parse XML file with recovery mode."""
    try:
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        tree = etree.parse(path, parser)
        return tree, tree.getroot()
    except Exception as e:
        print(f"[WARN] Cannot parse {path}: {e}")
        return None, None

# ---- SIMPLE KR+ID TRANSLATE ----
def simple_translate_krid():
    """
    Simple KR+ID translate:
      match (StringId, normalized StrOrigin) → En
      match (StringId, normalized DescOrigin) → En (descriptions)
    """
    global xml_folder_path, xml_file_path, tmx_file_path, status_label, mode_var

    if not tmx_file_path or not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select XML and TMX file/folder.")
        return

    status_label.config(text="Simple translate: parsing TMX...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX:\n{e}")
        return

    # build lookup dicts
    perfect_str = {
        (u["context"], normalize_text(u["kr"])): u["en"]
        for u in tmx_trans
        if u.get("kr") and u.get("en")
    }
    perfect_desc = {
        (u["context"], normalize_text(u["kr"])): u["en"]
        for u in tmx_desc
        if u.get("kr") and u.get("en")
    }

    # collect XML files
    xml_files = (
        get_all_xml_files(xml_folder_path)
        if mode_var.get() == "folder"
        else [xml_file_path]
    )
    updated_s = updated_d = 0

    for path in xml_files:
        tree, root = robust_parse_xml(path)
        if root is None:
            continue
        changed = False
        for loc in root.iter("LocStr"):
            sid    = loc.get("StringId", "").strip()
            orig   = normalize_text(loc.get("StrOrigin", ""))
            dorig  = normalize_text(loc.get("DescOrigin", ""))
            old_s  = loc.get("Str", "")
            old_d  = loc.get("Desc", "")

            # Str
            new_s = perfect_str.get((sid, orig))
            if new_s and new_s != old_s:
                loc.set("Str", new_s)
                updated_s += 1
                changed = True

            # Desc
            new_d = perfect_desc.get((sid, dorig))
            if new_d and new_d != old_d:
                loc.set("Desc", new_d)
                updated_d += 1
                changed = True

        if changed:
            tree.write(path, encoding="utf-8",
                       xml_declaration=False, pretty_print=True)

    status_label.config(
        text=f"Simple translate done: Str={updated_s}, Desc={updated_d}"
    )
    messagebox.showinfo(
        "Simple Translate Complete",
        f"Updated Str: {updated_s}\nUpdated Desc: {updated_d}"
    )

# ---- 2-STEP TRANSLATE ----
def two_step_match_krid_then_kr():
    """
    Two-step translate:
      1) perfect ID+KR match
      2) KR-only fallback
      also transfers descriptions by ID+DescOrigin
    """
    global xml_folder_path, xml_file_path, tmx_file_path, status_label, mode_var

    if not tmx_file_path or not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select XML and TMX file/folder.")
        return

    status_label.config(text="2-step translate: parsing TMX...")
    # run in background so UI stays responsive
    def worker():
        try:
            tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse TMX:\n{e}")
            return

        # perfect ID+KR
        perfect_str = {
            (u["context"], normalize_text(u["kr"])): u["en"]
            for u in tmx_trans
            if u.get("kr") and u.get("en")
        }
        perfect_desc = {
            (u["context"], normalize_text(u["kr"])): u["en"]
            for u in tmx_desc
            if u.get("kr") and u.get("en")
        }

        # KR-only fallback
        kr_only       = {normalize_text(u["kr"]): u["en"] for u in tmx_trans if u.get("kr") and u.get("en")}
        kr_only_desc  = {normalize_text(u["kr"]): u["en"] for u in tmx_desc if u.get("kr") and u.get("en")}

        xml_files = (
            get_all_xml_files(xml_folder_path)
            if mode_var.get() == "folder" else [xml_file_path]
        )
        upd1 = upd2 = updd = 0

        for path in xml_files:
            tree, root = robust_parse_xml(path)
            if root is None:
                continue
            changed = False
            for loc in root.iter("LocStr"):
                sid   = loc.get("StringId", "").strip()
                orig  = normalize_text(loc.get("StrOrigin", ""))
                dorig = normalize_text(loc.get("DescOrigin", ""))

                old_s = loc.get("Str", "")
                old_d = loc.get("Desc", "")

                # STEP 1: perfect ID+KR
                key = (sid, orig)
                if key in perfect_str:
                    new = perfect_str[key]
                    if new and new != old_s:
                        loc.set("Str", new)
                        upd1 += 1
                        changed = True
                else:
                    # STEP 2: KR-only fallback
                    if orig in kr_only:
                        new = kr_only[orig]
                        if new and new != old_s:
                            loc.set("Str", new)
                            upd2 += 1
                            changed = True

                # DESCRIPTIONS by perfect ID+DescOrigin
                keyd = (sid, dorig)
                if keyd in perfect_desc:
                    new = perfect_desc[keyd]
                    if new and new != old_d:
                        loc.set("Desc", new)
                        updd += 1
                        changed = True
                else:
                    # desc KR-only fallback
                    if dorig in kr_only_desc:
                        new = kr_only_desc[dorig]
                        if new and new != old_d:
                            loc.set("Desc", new)
                            updd += 1
                            changed = True

            if changed:
                tree.write(path, encoding="utf-8",
                           xml_declaration=False, pretty_print=True)

        status_label.config(
            text=f"2-Step done: Perfect={upd1}, KR-only={upd2}, Desc={updd}"
        )
        messagebox.showinfo(
            "2-Step Translate Complete",
            f"Perfect matches: {upd1}\nKR-only matches: {upd2}\nDescription matches: {updd}"
        )

    threading.Thread(target=worker, daemon=True).start()





# ---- STRINGID ONLY TRANSLATE ----
def id_translate():
    """
    XML translation by StringId only.
    Transfers Str and Desc from TMX for any matching <LocStr StringId="...">.
    """
    global xml_folder_path, xml_file_path, tmx_file_path, status_label

    if not tmx_file_path or not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select XML and TMX file/folder.")
        return

    status_label.config(text="Parsing TMX file (ID-only)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    # build lookup by context (StringId) → EN
    trans_by_id = {u['context']: u['en'] for u in tmx_trans if u.get('context') and u.get('en')}
    desc_by_id  = {u['context']: u['en'] for u in tmx_desc  if u.get('context') and u.get('en')}

    xml_files = get_xml_files()
    updated_str = 0
    updated_desc = 0

    for path in xml_files:
        tree, root = robust_parse_xml(path)
        if root is None:
            continue

        changed = False
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()
            if not sid:
                continue

            # Str
            old_str = loc.get("Str", "")
            new_str = trans_by_id.get(sid)
            if new_str is not None and new_str != old_str:
                loc.set("Str", new_str)
                updated_str += 1
                changed = True

            # Desc
            old_desc = loc.get("Desc", "")
            new_desc = desc_by_id.get(sid)
            if new_desc is not None and new_desc != old_desc:
                loc.set("Desc", new_desc)
                updated_desc += 1
                changed = True

        if changed:
            tree.write(path, encoding="utf-8", xml_declaration=False, pretty_print=True)

    status_label.config(text=f"ID-only done! Str: {updated_str}   Desc: {updated_desc}")
    messagebox.showinfo(
        "ID-Only Translate",
        f"Updated Str: {updated_str}\n"
        f"Updated Desc: {updated_desc}\n"
        f"in {len(xml_files)} files."
    )



# ---- EXCEL TRANSLATE – SIMPLE ----
def excel_translate_simple_krid():
    """
    Excel translation (Simple KR+ID):
      Col1 = KR text (StrOrigin)
      Col2 = EN translation (Str)
      Col3 = StringId
      Col4 = KR text (DescOrigin)
      Col5 = EN translation (Desc)
    Now uses normalize_text() + a nospace‐fallback dict for both Str and Desc.
    """
    global tmx_file_path, status_label
    if not tmx_file_path:
        messagebox.showerror("Error", "Please select a TMX file first.")
        return

    status_label.config(text="Parsing TMX for Excel (Simple KR+ID)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    # Build perfect‐match dicts + nospace fallbacks for Str
    perfect_str = {}
    perfect_str_nospace = {}
    for u in tmx_trans:
        ctx, kr, en = u.get('context'), u.get('kr'), u.get('en')
        if ctx and en:
            kr_norm = normalize_text(kr)
            perfect_str[(ctx, kr_norm)] = en
            kr_ns = re.sub(r'\s+', '', kr_norm)
            perfect_str_nospace[(ctx, kr_ns)] = en

    # Build perfect‐match dicts + nospace fallbacks for Desc
    perfect_desc = {}
    perfect_desc_nospace = {}
    for u in tmx_desc:
        ctx, kr, en = u.get('context'), u.get('kr'), u.get('en')
        if ctx and en:
            kr_norm = normalize_text(kr)
            # strip leading &desc; marker if present
            if kr_norm.lower().startswith("&desc;"):
                kr_norm = kr_norm[6:].lstrip()
            perfect_desc[(ctx, kr_norm)] = en
            kr_ns = re.sub(r'\s+', '', kr_norm)
            perfect_desc_nospace[(ctx, kr_ns)] = en

    if not (perfect_str or perfect_desc):
        messagebox.showerror("Error", "No usable translation units found in TMX.")
        return

    excel_path = filedialog.askopenfilename(
        title="Select Excel File", filetypes=[("Excel Files","*.xlsx")])
    if not excel_path:
        status_label.config(text="Excel translation cancelled.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        updated_str = updated_desc = total_rows = 0

        for row_idx in range(1, ws.max_row + 1):
            str_kr_val  = ws.cell(row_idx, 1).value
            strid_val   = ws.cell(row_idx, 3).value
            desc_kr_val = ws.cell(row_idx, 4).value
            if not strid_val:
                continue
            total_rows += 1
            sid = strid_val.strip()

            # Str translation (col1→col2)
            if str_kr_val is not None:
                kr_norm = normalize_text(str_kr_val)
                kr_ns   = re.sub(r'\s+', '', kr_norm)
                old_en  = (ws.cell(row_idx, 2).value or "").strip()

                new_en = None
                key    = (sid, kr_norm)
                if key in perfect_str:
                    new_en = perfect_str[key]
                else:
                    key_ns = (sid, kr_ns)
                    new_en = perfect_str_nospace.get(key_ns)

                if new_en and new_en != old_en:
                    ws.cell(row_idx, 2).value = new_en
                    updated_str += 1

            # Desc translation (col4→col5)
            if desc_kr_val is not None:
                kr_norm = normalize_text(desc_kr_val)
                # strip leading &desc; if any
                if kr_norm.lower().startswith("&desc;"):
                    kr_norm = kr_norm[6:].lstrip()
                kr_ns  = re.sub(r'\s+', '', kr_norm)
                old_en = (ws.cell(row_idx, 5).value or "").strip()

                new_en = None
                key    = (sid, kr_norm)
                if key in perfect_desc:
                    new_en = perfect_desc[key]
                else:
                    key_ns = (sid, kr_ns)
                    new_en = perfect_desc_nospace.get(key_ns)

                if new_en and new_en != old_en:
                    ws.cell(row_idx, 5).value = new_en
                    updated_desc += 1

        wb.save(excel_path)
        status_label.config(
            text=f"Excel Simple done: Str {updated_str}, Desc {updated_desc} updated."
        )
        messagebox.showinfo(
            "Excel Translate – Simple",
            f"Rows processed: {total_rows}\n"
            f"Str updated: {updated_str}\n"
            f"Desc updated: {updated_desc}\n"
            f"File saved: {excel_path}"
        )
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel file: {e}")
        status_label.config(text="Excel translation failed.")


# ---- EXCEL TRANSLATE – 2-STEP ----
def excel_translate_2step_krid_then_kr():
    """
    Excel translation (2-STEP):
      1) Perfect match (ID+KR, with nospace fallback)
      2) KR‐only fallback (with nospace)
      Col1 = KR text (StrOrigin)
      Col2 = EN translation (Str)
      Col3 = StringId
      Col4 = KR text (DescOrigin)
      Col5 = EN translation (Desc)
    Now applies normalize_text() + both perfect‐nospace and kr‐only‐nospace lookups.
    """
    global tmx_file_path, status_label
    if not tmx_file_path:
        messagebox.showerror("Error", "Please select a TMX file first.")
        return

    status_label.config(text="Parsing TMX for Excel (2-STEP)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    # Build Str dicts
    perfect_str = {}
    perfect_str_nospace = {}
    kr_only = {}
    kr_only_nospace = {}
    for u in tmx_trans:
        ctx, kr, en = u.get('context'), u.get('kr'), u.get('en')
        if ctx and en:
            kn = normalize_text(kr)
            ns = re.sub(r'\s+', '', kn)
            perfect_str[(ctx, kn)] = en
            perfect_str_nospace[(ctx, ns)] = en
            kr_only[kn] = en
            kr_only_nospace[ns] = en

    # Build Desc dicts
    perfect_desc = {}
    perfect_desc_nospace = {}
    kr_only_desc = {}
    kr_only_desc_nospace = {}
    for u in tmx_desc:
        ctx, kr, en = u.get('context'), u.get('kr'), u.get('en')
        if ctx and en:
            kn = normalize_text(kr)
            if kn.lower().startswith("&desc;"):
                kn = kn[6:].lstrip()
            ns = re.sub(r'\s+', '', kn)
            perfect_desc[(ctx, kn)] = en
            perfect_desc_nospace[(ctx, ns)] = en
            kr_only_desc[kn] = en
            kr_only_desc_nospace[ns] = en

    if not (perfect_str or perfect_desc):
        messagebox.showerror("Error", "No usable translation units found in TMX.")
        return

    excel_path = filedialog.askopenfilename(
        title="Select Excel File", filetypes=[("Excel Files","*.xlsx")])
    if not excel_path:
        status_label.config(text="Excel translation cancelled.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        upd_perf_str = upd_kronly_str = upd_perf_desc = upd_kronly_desc = total_rows = 0

        for row_idx in range(1, ws.max_row + 1):
            str_kr_val  = ws.cell(row_idx, 1).value
            strid_val   = ws.cell(row_idx, 3).value
            desc_kr_val = ws.cell(row_idx, 4).value
            if not strid_val:
                continue
            total_rows += 1
            sid = strid_val.strip()

            # STR column
            if str_kr_val is not None:
                kn     = normalize_text(str_kr_val)
                ns     = re.sub(r'\s+', '', kn)
                old_en = normalize_text(ws.cell(row_idx, 2).value)

                # Step 1: perfect ID+KR
                key      = (sid, kn)
                key_ns   = (sid, ns)
                new_en   = None
                used_perf = False

                if key in perfect_str:
                    new_en = perfect_str[key]
                    used_perf = True
                elif key_ns in perfect_str_nospace:
                    new_en = perfect_str_nospace[key_ns]
                    used_perf = True

                if used_perf and new_en and new_en != old_en:
                    ws.cell(row_idx, 2).value = new_en
                    upd_perf_str += 1

                # Step 2: KR‐only fallback if no perfect match
                if not used_perf:
                    if kn in kr_only:
                        new_en = kr_only[kn]
                        if new_en and new_en != old_en:
                            ws.cell(row_idx, 2).value = new_en
                            upd_kronly_str += 1
                    else:
                        if ns in kr_only_nospace:
                            new_en = kr_only_nospace[ns]
                            if new_en and new_en != old_en:
                                ws.cell(row_idx, 2).value = new_en
                                upd_kronly_str += 1

            # DESC column
            if desc_kr_val is not None:
                kn     = normalize_text(desc_kr_val)
                if kn.lower().startswith("&desc;"):
                    kn = kn[6:].lstrip()
                ns      = re.sub(r'\s+', '', kn)
                old_en  = normalize_text(ws.cell(row_idx, 5).value)

                # perfect ID+KR for Desc
                key     = (sid, kn)
                key_ns  = (sid, ns)
                new_en  = None
                used_pd = False

                if key in perfect_desc:
                    new_en = perfect_desc[key]
                    used_pd = True
                elif key_ns in perfect_desc_nospace:
                    new_en = perfect_desc_nospace[key_ns]
                    used_pd = True

                if used_pd and new_en and new_en != old_en:
                    ws.cell(row_idx, 5).value = new_en
                    upd_perf_desc += 1

                # KR‐only Desc fallback
                if not used_pd:
                    if kn in kr_only_desc:
                        new_en = kr_only_desc[kn]
                        if new_en and new_en != old_en:
                            ws.cell(row_idx, 5).value = new_en
                            upd_kronly_desc += 1
                    else:
                        if ns in kr_only_desc_nospace:
                            new_en = kr_only_desc_nospace[ns]
                            if new_en and new_en != old_en:
                                ws.cell(row_idx, 5).value = new_en
                                upd_kronly_desc += 1

        wb.save(excel_path)
        status_label.config(
            text=f"Excel 2-STEP done: perfectStr={upd_perf_str}, krStr={upd_kronly_str}, "
                 f"perfectDesc={upd_perf_desc}, krDesc={upd_kronly_desc}"
        )
        messagebox.showinfo(
            "Excel Translate – 2-STEP",
            f"Rows: {total_rows}\n"
            f"Perfect Str: {upd_perf_str}\nKR-only Str: {upd_kronly_str}\n"
            f"Perfect Desc: {upd_perf_desc}\nKR-only Desc: {upd_kronly_desc}\n"
            f"File saved: {excel_path}"
        )
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel file: {e}")
        status_label.config(text="Excel 2-STEP translation failed.")



# ---- INSERT FROM EXCEL ----
def insert_translations_from_excel():
    try:
        excel_path = filedialog.askopenfilename(title="Select Excel File",
                                                filetypes=[("Excel Files", "*.xlsx")])
        if not excel_path:
            return

        xml_path = filedialog.askopenfilename(title="Select XML File",
                                              filetypes=[("XML Files", "*.xml")])
        if not xml_path:
            return

        folder_path = filedialog.askdirectory(title="Select XML Folder for VoiceId Mapping")
        if not folder_path:
            return

        try:
            os.chmod(xml_path, 0o666)
        except:
            pass

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        excel_rows = []
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                excel_rows.append((str(row[0]).strip(), str(row[1]).strip()))
        print(f"[INFO] Loaded {len(excel_rows)} rows from Excel.")

        tree, root = robust_parse_xml(xml_path)
        if tree is None:
            messagebox.showerror("Error", "Failed to parse XML file.")
            return

        stringid_map = {}
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()
            if sid:
                stringid_map[sid.lower()] = loc
        print(f"[INFO] Found {len(stringid_map)} StringIds in target XML.")

        voiceid_to_stringid = {}
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        for dirpath, _, files in os.walk(folder_path):
            for fn in files:
                if fn.lower().endswith(".xml"):
                    try:
                        t = etree.parse(os.path.join(dirpath, fn), parser)
                        for loc in t.getroot().iter("LocStr"):
                            sid = loc.get("StringId", "").strip()
                            vid = loc.get("VoiceId", "").strip()
                            snd = loc.get("SoundEventName", "").strip()
                            if sid:
                                if vid:
                                    voiceid_to_stringid[vid.lower()] = sid
                                if snd:
                                    voiceid_to_stringid[snd.lower()] = sid
                    except:
                        pass
        print(f"[INFO] Built {len(voiceid_to_stringid)} VoiceId/SoundEventName mappings.")

        updated = 0
        for translation, id_or_voice in excel_rows:
            key = id_or_voice.lower()

            if key in stringid_map:
                loc = stringid_map[key]
                loc.set("Str", translation)
                updated += 1
            elif key in voiceid_to_stringid:
                sid = voiceid_to_stringid[key]
                if sid.lower() in stringid_map:
                    loc = stringid_map[sid.lower()]
                    loc.set("Str", translation)
                    updated += 1

        tree.write(xml_path, encoding="utf-8", xml_declaration=False, pretty_print=True)
        messagebox.showinfo("Done",
                            f"Inserted translations for {updated} entries.\n"
                            f"File overwritten:\n{xml_path}")
        print(f"[INFO] Updated {updated} entries in {xml_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed: {e}")

# ---- COVERAGE CHECK ----
def check_translation_coverage():
    global xml_folder_path, xml_file_path, status_label
    if not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select an XML file or folder first.")
        return
    status_label.config(text="Calculating translation coverage...")

    xml_files = get_xml_files()
    total = translated = untranslated = 0
    korean_re = re.compile(r'[\uac00-\ud7a3]')
    locstr_re = re.compile(r'<LocStr\b([^>]*)>', re.IGNORECASE | re.DOTALL)
    str_attr_re = re.compile(r'Str="([^"]*)"')

    for path in xml_files:
        try:
            with open(path, encoding="utf-8", errors="ignore") as f:
                content = f.read()
            for m in locstr_re.finditer(content):
                total += 1
                attrs = m.group(1)
                str_m = str_attr_re.search(attrs)
                val = (str_m.group(1).strip() if str_m else "")
                if not val or korean_re.search(val):
                    untranslated += 1
                else:
                    translated += 1
        except:
            pass

    coverage = (translated / total * 100) if total else 0.0
    msg = (f"Total LocStrs: {total}\nTranslated: {translated}\n"
           f"Untranslated: {untranslated}\nCoverage: {coverage:.1f}%"
           if coverage < 99.95 else
           f"Total LocStrs: {total}\nTranslated: {translated}\n"
           f"Untranslated: {untranslated}\nCoverage: {coverage:.3f}%")
    status_label.config(text=msg)
    messagebox.showinfo("Translation Coverage", msg)

# ---- GUI CALLBACKS ----
def upload_xml():
    global xml_folder_path, xml_file_path
    if mode_var.get() == "folder":
        folder = filedialog.askdirectory(title="Select XML Folder")
        if folder:
            xml_folder_path, xml_file_path = folder, None
            for p in get_all_xml_files(folder):
                try:
                    os.chmod(p, 0o666)
                except:
                    pass
            btn_xml_folder.config(bg="green")
            status_label.config(text="XML folder selected.")
    else:
        file = filedialog.askopenfilename(title="Select XML File",
                                          filetypes=[("XML Files", "*.xml")])
        if file:
            xml_file_path, xml_folder_path = file, None
            try:
                os.chmod(file, 0o666)
            except:
                pass
            btn_xml_folder.config(bg="green")
            status_label.config(text="XML file selected.")
    update_buttons()
    update_info_panel()

def upload_tmx_file():
    global tmx_file_path
    file = filedialog.askopenfilename(title="Select TMX File",
                                      filetypes=[("TMX Files", "*.tmx")])
    if file:
        tmx_file_path = file
        btn_tmx_file.config(bg="green")
        status_label.config(text="TMX file selected.")
    update_buttons()
    update_info_panel()

def update_buttons():
    if xml_folder_path or xml_file_path:
        btn_check.config(state="normal", bg="purple")
    else:
        btn_check.config(state="disabled", bg="grey")

    if tmx_file_path:
        btn_excel_simple.config(state="normal", bg="#32cd32")
        btn_excel_2step.config(state="normal", bg="#ffb347")
    else:
        btn_excel_simple.config(state="disabled", bg="grey")
        btn_excel_2step.config(state="disabled", bg="grey")

    if (xml_folder_path or xml_file_path) and tmx_file_path:
        btn_simple.config(state="normal", bg="orange")
        btn_two.config(state="normal", bg="#ffb347")
        btn_id.config(state="normal", bg="#ff4d4d")
    else:
        btn_simple.config(state="disabled", bg="grey")
        btn_two.config(state="disabled", bg="grey")
        btn_id.config(state="disabled", bg="grey")

def update_info_panel():
    info_panel.config(state="normal")
    info_panel.delete(1.0, tk.END)
    info_panel.insert(tk.END, "Current Selections:\n")
    info_panel.insert(tk.END, f"Mode       : {'Folder' if mode_var.get() == 'folder' else 'File'}\n")
    info_panel.insert(tk.END, f"XML Folder : {xml_folder_path if xml_folder_path else '[Not selected]'}\n")
    info_panel.insert(tk.END, f"XML File   : {xml_file_path if xml_file_path else '[Not selected]'}\n")
    info_panel.insert(tk.END, f"TMX File   : {tmx_file_path if tmx_file_path else '[Not selected]'}\n")
    info_panel.config(state="disabled")

def switch_mode():
    global xml_folder_path, xml_file_path
    xml_folder_path = None
    xml_file_path = None
    btn_xml_folder.config(bg="SystemButtonFace")
    update_buttons()
    update_info_panel()
    status_label.config(text=f"Switched to {'Folder' if mode_var.get() == 'folder' else 'File'} mode. Please select XML.")

# ---- MAIN GUI ----
def main():
    global btn_xml_folder, btn_tmx_file, btn_simple, btn_two
    global btn_excel_simple, btn_excel_2step, btn_check, btn_insert_excel, btn_id
    global status_label, info_panel, mode_var

    root = tk.Tk()
    root.title("Batch XML Updater + Excel Helpers")
    root.geometry("700x780")

    mode_var = tk.StringVar(value="folder")
    mode_frame = tk.Frame(root)
    mode_frame.pack(pady=8)
    tk.Label(mode_frame, text="Mode:").pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="Folder Mode", variable=mode_var,
                   value="folder", command=switch_mode).pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="File Mode", variable=mode_var,
                   value="file", command=switch_mode).pack(side="left", padx=4)

    btn_xml_folder = tk.Button(root, text="Upload XML Folder/File",
                               command=upload_xml, width=30, height=2)
    btn_xml_folder.pack(pady=8)

    btn_tmx_file = tk.Button(root, text="Upload TMX File",
                             command=upload_tmx_file, width=30, height=2)
    btn_tmx_file.pack(pady=8)

    btn_simple = tk.Button(root, text="KR+ID Simple Translate",
                           command=simple_translate_krid,
                           state="disabled", bg="grey", width=30, height=2)
    btn_simple.pack(pady=8)

    btn_two = tk.Button(root, text="2 STEP MATCH – KR+ID then KR",
                        command=two_step_match_krid_then_kr,
                        state="disabled", bg="grey", width=30, height=2)
    btn_two.pack(pady=8)

    # NEW ID-ONLY TRANSLATE BUTTON
    btn_id = tk.Button(root, text="ID-Only Translate",
                       command=id_translate,
                       state="disabled", bg="grey", width=30, height=2)
    btn_id.pack(pady=8)

    btn_excel_simple = tk.Button(root,
                                 text="Excel Translate SIMPLE (KR+ID)",
                                 command=excel_translate_simple_krid,
                                 state="disabled", bg="grey", width=30, height=2)
    btn_excel_simple.pack(pady=8)

    btn_excel_2step = tk.Button(root,
                                text="Excel Translate 2-STEP (ID+KR ➜ KR)",
                                command=excel_translate_2step_krid_then_kr,
                                state="disabled", bg="grey", width=30, height=2)
    btn_excel_2step.pack(pady=8)

    btn_check = tk.Button(root, text="Check Translation Coverage",
                          command=check_translation_coverage,
                          state="disabled", bg="grey", width=30, height=2)
    btn_check.pack(pady=8)

    btn_insert_excel = tk.Button(root, text="Excel ➜ XML (StringID only)",
                                 command=insert_translations_from_excel,
                                 width=30, height=2, bg="#87ceeb")
    btn_insert_excel.pack(pady=8)

    status_label = tk.Label(root, text="Please upload XML and/or TMX.", fg="blue")
    status_label.pack(pady=12)

    info_panel = tk.Text(root, height=6, width=80, wrap="none",
                         state="disabled", bg="#f0f0f0")
    info_panel.pack(pady=8, padx=8, fill="x")

    update_info_panel()
    update_buttons()
    root.mainloop()

if __name__ == "__main__":
    main()