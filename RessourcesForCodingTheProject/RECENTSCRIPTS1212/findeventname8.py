
#!/usr/bin/env python3
# coding: utf-8
"""
GUI tool with multiple processes:
1. AudioFileNameToEventName (UPDATED LOGIC, robust XML parsing)
2. SoundEventName ↔ StringID mapping
3. Filename Replace Tool
4. Extract .wem / .wav filenames to Excel
5. EventName → ExcelAudio LQA (UPDATED LOGIC)
6. QuestDialog / AIDialog Prefix  🔥 NEW
"""

import sys
import re
import shutil
import datetime as _dt
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
import tkinter as tk
from tkinter import messagebox, filedialog
from typing import Optional  # for Python 3.8 compatibility

# --- CONFIG ------------------------------------------------------------------
WEM_FOLDER              = Path(r"F:\perforce\cd\cd_kappa\resource\Sound\Windows\English(US)")
MAINLINE_EXPORT_FOLDER  = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
KAPPA_EXPORT_FOLDER     = Path(r"F:\perforce\cd\cd_kappa\resource\GameData\stringtable\export__")
MAINLINE_LOC_FILE       = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc\languagedata_ZHO-CN.xml")
KAPPA_LOC_FILE          = Path(r"F:\perforce\cd\cd_kappa\resource\GameData\stringtable\loc\languagedata_ZHO-CN.xml")
OUTPUT_FILE_AUDIO       = Path(__file__).parent / "SoundEventName_KAPPA_Matches.xlsx"
TOP_N                   = 1  # top‐N similarity candidates
# -----------------------------------------------------------------------------


# ========== SMALL HELPERS ====================================================
def debug(msg: str):
    """Print debug line to stdout."""
    print(f"[DEBUG] {msg}")


def _norm(txt: Optional[str]) -> str:
    """Trim + lowercase helper (safe for None)."""
    return txt.strip().lower() if txt else ""


# ========== GENERIC XML HELPERS =============================================
def robust_parse_xml(path: Path):
    """
    Parse XML ignoring BOM / DTD / encoding issues.
    Returns (etree.ElementTree, root_element) or (None, None) on failure.
    """
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as fh:
            content = fh.read()

        # strip <?xml ... ?> or <!DOCTYPE ... > if present
        content = re.sub(r'^<\?xml[^>]*\?>\s*', '', content, flags=re.MULTILINE)
        content = re.sub(r'<!DOCTYPE[^>]*>\s*', '', content, flags=re.MULTILINE)

        parser = etree.XMLParser(resolve_entities=False,
                                 load_dtd=False,
                                 no_network=True,
                                 recover=True)
        root = etree.fromstring(content.encode("utf-8"), parser=parser)
        return etree.ElementTree(root), root
    except Exception as exc:
        debug(f"XML parse failed for {path}: {exc}")
        return None, None


# ========== MAINLINE / KAPPA PARSING ========================================
def parse_export_folder(folder: Path):
    """
    Parse every *.xml in `folder` recursively and collect:
        • event2origin:  {event_name_lower → StrOrigin}
        • id2events:     {StringId → [event_name_lower, ...]}
        • all_events:    {event_name_lower, ...}
    """
    event2origin = {}
    id2events    = defaultdict(list)
    all_events   = set()

    for xml_file in folder.rglob("*.xml"):
        _, root = robust_parse_xml(xml_file)
        if root is None:
            continue

        for node in root.iter():
            evt = node.get("SoundEventName")
            if not evt:
                continue
            evt_lc         = _norm(evt)
            all_events.add(evt_lc)

            so             = node.get("StrOrigin") or "EMPTY"
            event2origin[evt_lc] = so

            sid = node.get("StringId")
            if sid:
                id2events[sid].append(evt_lc)

    return event2origin, id2events, all_events


def parse_loc_file(xml_path: Path):
    """
    Parse a localisation XML and return:
        • origin2str: {StrOrigin → ChineseString}
        • str2ids:    {ChineseString → [StringId, ...]}
    """
    origin2str = {}
    str2ids    = defaultdict(list)

    _, root = robust_parse_xml(xml_path)
    if root is None:
        return origin2str, str2ids

    for node in root.iter():
        string = node.get("Str")
        origin = node.get("StrOrigin") or "EMPTY"
        sid    = node.get("StringId")

        if string and origin:
            origin2str[origin] = string
        if string and sid:
            str2ids[string].append(sid)

    return origin2str, str2ids


# ========== TOOL 1: AudioFileName → EventName ===============================
def _collect_audio_event_names(folder: Path):
    """Return sorted list of distinct *.wem stems inside `folder`."""
    return sorted({_norm(p.stem) for p in folder.rglob("*.wem")})


def process_audiofilename_to_eventname():
    """
    UPDATED AudioFileNameToEventName.
    """
    try:
        debug("Starting UPDATED AudioFileNameToEventName ...")

        original_events                               = _collect_audio_event_names(WEM_FOLDER)
        ml_event2origin, _, _                         = parse_export_folder(MAINLINE_EXPORT_FOLDER)
        ml_origin2str, _                              = parse_loc_file(MAINLINE_LOC_FILE)
        th_event2origin, th_sid2events, th_event_set  = parse_export_folder(KAPPA_EXPORT_FOLDER)
        _, th_str2ids                                 = parse_loc_file(KAPPA_LOC_FILE)

        perfect_matches, str_matches, no_match = [], [], []

        for evt in original_events:
            evt_lc = _norm(evt)

            # 1) perfect match in KAPPA
            if evt_lc in th_event_set:
                ml_so = ml_event2origin.get(evt_lc)
                th_so = th_event2origin.get(evt_lc)
                if ml_so is None:
                    cmp = "NOT FOUND IN MAINLINE"
                elif th_so is None:
                    cmp = "NOT FOUND IN KAPPA"
                elif ml_so == th_so:
                    cmp = "IDENTICAL StrOrigin"
                else:
                    cmp = "DIFFERENT StrOrigin"
                perfect_matches.append((evt, evt, cmp))
                continue

            # 2) try via StrOrigin → Chinese string → KAPPA → Event
            str_origin = ml_event2origin.get(evt_lc)
            if not str_origin:
                no_match.append(evt)
                continue

            cn_string = ml_origin2str.get(str_origin)
            if not cn_string:
                no_match.append(evt)
                continue

            kappa_ids = th_str2ids.get(cn_string)
            if not kappa_ids:
                no_match.append(evt)
                continue

            # 3) similarity rank
            candidates = []
            for tid in kappa_ids:
                for cand_evt in th_sid2events.get(tid, []):
                    score = SequenceMatcher(None, evt_lc, cand_evt).ratio()
                    candidates.append((score, cand_evt))
            if not candidates:
                no_match.append(evt)
                continue

            candidates.sort(key=lambda t: t[0], reverse=True)
            for score, best_evt in candidates[:TOP_N]:
                str_matches.append((evt, best_evt, round(score * 100, 2)))

        # ---------- SAVE ----------------------------------------------------
        perfect_matches.sort(key=lambda t: t[0])
        str_matches.sort(key=lambda t: t[2], reverse=True)
        no_match.sort()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "100% match"
        ws1.append(["EventName", "MatchedEvent", "StrOriginCompare"])
        for row in perfect_matches:
            ws1.append(row)

        ws2 = wb.create_sheet("Matched by Str")
        ws2.append(["OriginalEvent", "MatchedEvent", "Similarity(%)"])
        for row in str_matches:
            ws2.append(row)

        ws3 = wb.create_sheet("No match")
        ws3.append(["EventName"])
        for evt in no_match:
            ws3.append([evt])

        wb.save(OUTPUT_FILE_AUDIO)
        debug(f"Output saved to {OUTPUT_FILE_AUDIO}")

        messagebox.showinfo(
            "Done",
            f"AudioFileNameToEventName completed.\n"
            f"Perfect matches: {len(perfect_matches)}\n"
            f"Str-based matches: {len(str_matches)}\n"
            f"No match: {len(no_match)}"
        )
    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 2: SE ↔ SID mappings =======================================
def build_full_index_for_folder(folder: Path):
    """
    Parse all XML files in `folder` to build mapping dictionaries.
    Returns (se_to_sid, sid_to_se)
    """
    se_to_sid, sid_to_se = {}, {}

    files = list(folder.rglob("*.xml"))
    debug(f"Parsing {len(files)} XMLs in {folder}")

    for idx, xml_file in enumerate(files, 1):
        _, root = robust_parse_xml(xml_file)
        if root is None:
            continue

        for node in root.iter():
            se  = (node.get("SoundEventName") or "").strip()
            sid = (node.get("StringId") or "").strip()
            if se and sid:
                se_to_sid[se.lower()]    = sid
                sid_to_se[sid.lower()]   = se
        debug(f"[{idx}/{len(files)}] Parsed {xml_file.name}")

    debug(f"Folder {folder} → SE→SID={len(se_to_sid)}  SID→SE={len(sid_to_se)}")
    return se_to_sid, sid_to_se


def process_soundeventname_to_stringid(direction: str):
    """
    Convert between SoundEventName and StringId using MAINLINE or KAPPA data.
    direction must be either "SE_TO_SID" or "SID_TO_SE".
    """
    try:
        debug(f"Starting mapping: {direction}")

        excel_in = filedialog.askopenfilename(
            title="Select Input Excel", filetypes=[("Excel files", "*.xlsx")]
        )
        if not excel_in:
            return
        excel_in = Path(excel_in)

        wb_in = load_workbook(excel_in)
        ws_in = wb_in.active
        rows  = list(ws_in.iter_rows(values_only=True))
        debug(f"Loaded {len(rows)} rows")

        # Build indexes
        ml_se2sid, ml_sid2se = build_full_index_for_folder(MAINLINE_EXPORT_FOLDER)
        th_se2sid, th_sid2se = build_full_index_for_folder(KAPPA_EXPORT_FOLDER)

        # popup for per-column dataset selection
        root_sel = tk.Toplevel()
        root_sel.title("Select Dataset for Each Column")
        tk.Label(root_sel, text="Choose dataset (MAINLINE / KAPPA) for every column").pack(pady=4)

        col_vars = []
        for ci in range(len(rows[0])):
            fr = tk.Frame(root_sel); fr.pack(anchor="w", padx=5, pady=2)
            tk.Label(fr, text=f"Column {ci+1}:").pack(side=tk.LEFT)
            var = tk.StringVar(value="MAINLINE")
            tk.OptionMenu(fr, var, "MAINLINE", "KAPPA").pack(side=tk.LEFT)
            col_vars.append(var)

        tk.Button(root_sel, text="OK", command=root_sel.destroy).pack(pady=8)
        root_sel.grab_set()
        root_sel.wait_window()

        # --- mapping -------------------------------------------------------
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Mapped"
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for row in rows:
            out_row = []
            for ci, cell in enumerate(row):
                value = (cell or "").strip()
                low   = value.lower()
                chosen_ds = col_vars[ci].get()
                mapped = None
                if direction == "SE_TO_SID":
                    mapped = (ml_se2sid if chosen_ds == "MAINLINE" else th_se2sid).get(low)
                else:
                    mapped = (ml_sid2se if chosen_ds == "MAINLINE" else th_sid2se).get(low)

                out_row.append(mapped if mapped else value)
            ws_out.append(out_row)

        # highlight misses
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                value = (cell or "").strip().lower()
                chosen_ds = col_vars[c_idx - 1].get()
                if direction == "SE_TO_SID":
                    miss = value and value not in (ml_se2sid if chosen_ds == "MAINLINE" else th_se2sid)
                else:
                    miss = value and value not in (ml_sid2se if chosen_ds == "MAINLINE" else th_sid2se)

                if miss:
                    ws_out.cell(row=r_idx, column=c_idx).fill = red_fill

        out_path = excel_in.with_name(f"{excel_in.stem}_OUTPUT.xlsx")
        wb_out.save(out_path)
        debug(f"Saved: {out_path}")
        messagebox.showinfo("Done", f"Mapping finished.\nOutput: {out_path}")

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 3: Filename Replace ========================================
def process_replace_filenames():
    """
    Excel input must have two columns:
        current_name | new_name   (names WITHOUT extension)
    """
    try:
        debug("Starting Filename Replace Tool ...")

        excel_in = filedialog.askopenfilename(
            title="Select Excel mapping file", filetypes=[("Excel files", "*.xlsx")]
        )
        if not excel_in:
            return
        excel_in = Path(excel_in)

        folder = filedialog.askdirectory(title="Select folder to rename files in")
        if not folder:
            return
        folder = Path(folder)

        wb = load_workbook(excel_in)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        mapping = {
            (r[0] or "").strip().lower(): (r[1] or "").strip().lower()
            for r in rows
            if r and r[0] and r[1]
        }

        renamed = set()
        for fp in folder.rglob("*.*"):
            stem_lc = fp.stem.lower()
            if stem_lc in mapping:
                new_fp = fp.with_name(mapping[stem_lc] + fp.suffix.lower())
                debug(f"Renaming {fp.name} → {new_fp.name}")
                fp.rename(new_fp)
                renamed.add(stem_lc)

        # mark renamed in Excel
        orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        for r_idx, row in enumerate(rows, start=1):
            if row and (row[0] or "").strip().lower() in renamed:
                ws.cell(row=r_idx, column=1).fill = orange

        out_path = excel_in.with_name(f"{excel_in.stem}_REPLACED.xlsx")
        wb.save(out_path)
        debug(f"Saved: {out_path}")
        messagebox.showinfo("Done", f"Renaming completed.\nOutput: {out_path}")

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 4: Extract audio filenames =================================
def process_extract_audio_filenames():
    """
    Extract unique stems of .wem or .wav in a chosen folder → Excel.
    """
    try:
        debug("Starting Extract Audio Filenames ...")

        # ask which extension
        win = tk.Toplevel()
        win.title("Select audio type")
        tk.Label(win, text="Pick audio file extension to extract:").pack(pady=6)
        ext_var = tk.StringVar(value=".wem")
        for ext in (".wem", ".wav"):
            tk.Radiobutton(win, text=ext, variable=ext_var, value=ext).pack(anchor="w", padx=20)
        tk.Button(win, text="OK", command=win.destroy).pack(pady=8)
        win.grab_set()
        win.wait_window()

        ext = ext_var.get()
        if not ext:
            return

        folder = filedialog.askdirectory(title=f"Select folder to scan for *{ext}")
        if not folder:
            return
        folder = Path(folder)

        names = sorted({p.stem for p in folder.rglob(f"*{ext}")})
        if not names:
            messagebox.showinfo("No Files", f"No {ext} files found.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = f"{ext[1:].upper()}_Filenames"
        ws.append(["FilenameWithoutExtension"])
        for name in names:
            ws.append([name])

        out_path = Path(__file__).parent / f"{ext[1:]}_Filenames.xlsx"
        wb.save(out_path)
        debug(f"Extracted {len(names)} filenames to {out_path}")
        messagebox.showinfo("Done", f"Extracted {len(names)} filenames.\nSaved: {out_path}")

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 5: EventName → ExcelAudio LQA (REPAIRED) ====================
def process_eventname_to_excelaudio_lqa():
    """
    Build an LQA Excel and copy matching *.wav files.

    Steps:
        1. Ask for SOURCE Excel (contains EventNames column).
        2. Ask user which column contains the EventNames.
        3. Ask for folder that actually holds the *.wav files.
        4. Ask for TARGET Excel (provides Group / SequenceName / StrOrigin / Text / EventName).
        5. Intersect datasets → copy wavs & generate output Excel with hyperlinks.
    """
    try:
        debug("Starting EventName → ExcelAudio LQA ...")

        # 1) SOURCE EXCEL --------------------------------------------------
        src_path = filedialog.askopenfilename(
            title="Select SOURCE Excel (contains EventNames)",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not src_path:
            return
        src_path = Path(src_path)
        wb_src = load_workbook(src_path)
        ws_src = wb_src.active

        # read header row safely
        src_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_src[1]]
        if not any(src_headers):
            messagebox.showerror("Error", "First row of SOURCE Excel is empty.")
            return

        # Column picker ----------------------------------------------------
        col_var = tk.StringVar()
        picker  = tk.Toplevel()
        picker.title("Select EventName Column")
        tk.Label(picker, text="Select the column that contains EventNames:").pack(pady=6)

        for idx, header in enumerate(src_headers, 1):
            letter = get_column_letter(idx)
            label  = f"{letter}: {header}" if header else letter
            tk.Radiobutton(picker, text=label, variable=col_var, value=str(idx)).pack(anchor="w", padx=18)

        # If only one column, pre-select it
        if ws_src.max_column == 1:
            col_var.set("1")

        tk.Button(picker, text="OK", command=picker.destroy).pack(pady=10)
        picker.grab_set()
        picker.wait_window()

        if not col_var.get():
            return
        src_col_idx = int(col_var.get())
        debug(f"Chosen SOURCE column: {get_column_letter(src_col_idx)}")

        # Build lookup set --------------------------------------------------
        lookup_set = set()
        for r in range(2, ws_src.max_row + 1):
            val = ws_src.cell(row=r, column=src_col_idx).value
            if not val:
                continue
            name = _norm(str(val))
            if name.lower().endswith(".wav"):
                name = _norm(Path(name).stem)
            lookup_set.add(name)

        if not lookup_set:
            messagebox.showerror("Error", "No EventNames found in chosen SOURCE column.")
            return
        debug(f"Collected {len(lookup_set)} EventNames from SOURCE")

        # 2) WAV DIRECTORY --------------------------------------------------
        wav_dir = filedialog.askdirectory(title="Select folder containing *.wav files")
        if not wav_dir:
            return
        wav_dir = Path(wav_dir)
        wav_paths = list(wav_dir.rglob("*.wav"))
        if not wav_paths:
            messagebox.showerror("Error", "No *.wav files found in the selected directory.")
            return
        stem2wav = {_norm(p.stem): p for p in wav_paths}
        debug(f"Indexed {len(stem2wav)} wav files.")

        # 3) TARGET EXCEL ---------------------------------------------------
        tgt_path = filedialog.askopenfilename(
            title="Select TARGET Excel (with Group / SequenceName / StrOrigin / Text / EventName)",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not tgt_path:
            return
        tgt_path = Path(tgt_path)
        wb_tgt   = load_workbook(tgt_path)
        ws_tgt   = wb_tgt.active

        # find required columns (case-insensitive)
        tgt_headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws_tgt[1]]

        required = ["group", "sequencename", "strorigin", "text", "eventname"]
        col_map  = {}
        for req in required:
            try:
                col_map[req] = tgt_headers.index(req) + 1
            except ValueError:
                messagebox.showerror("Error", f"TARGET Excel missing header: '{req}'")
                return
        debug(f"TGT columns: {col_map}")

        # 4) Find matches ---------------------------------------------------
        matches = []  # (seq_sort, group, seq, so, text, wav_path)
        for r in range(2, ws_tgt.max_row + 1):
            evt_raw = ws_tgt.cell(row=r, column=col_map["eventname"]).value
            if not evt_raw:
                continue
            evt_norm = _norm(str(evt_raw))
            if evt_norm.endswith(".wav"):
                evt_norm = _norm(Path(evt_norm).stem)

            if evt_norm in lookup_set and evt_norm in stem2wav:
                grp = ws_tgt.cell(row=r, column=col_map["group"]).value
                seq = ws_tgt.cell(row=r, column=col_map["sequencename"]).value
                so  = ws_tgt.cell(row=r, column=col_map["strorigin"]).value
                txt = ws_tgt.cell(row=r, column=col_map["text"]).value
                matches.append((evt_norm, grp, seq, so, txt, stem2wav[evt_norm]))

        debug(f"Total matches: {len(matches)}")
        if not matches:
            messagebox.showinfo("No Match", "No matches found between data sets.")
            return

        matches.sort(key=lambda t: t[0])  # sort by normalized SequenceName

        # 5) Prepare output directory --------------------------------------
        out_dir = Path(__file__).parent / f"LQA_OUTPUT_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        out_dir.mkdir(parents=True, exist_ok=True)
        debug(f"Output directory: {out_dir}")

        # Copy wavs and collect rows
        final_rows = []  # (Group, SequenceName, StrOrigin, Text, EventName)
        for _, grp, seq, so, txt, wav_path in matches:
            dst = out_dir / wav_path.name
            if not dst.exists():
                shutil.copy2(wav_path, dst)
            final_rows.append((grp, seq, so, txt, wav_path.name))

        # 6) Build output Excel --------------------------------------------
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "LQA"
        ws_out.append(["Group", "SequenceName", "StrOrigin", "Text", "EventName"])

        for row in final_rows:
            ws_out.append(row)
            # add hyperlink to the wav in EventName cell
            c = ws_out.cell(row=ws_out.max_row, column=5)
            c.hyperlink = (out_dir / row[4]).resolve().as_uri()
            c.style = "Hyperlink"

        out_excel = out_dir / "LQA_Output.xlsx"
        wb_out.save(out_excel)

        debug(f"Excel saved: {out_excel}")
        messagebox.showinfo(
            "Done",
            f"LQA package completed.\n"
            f"Copied wav files: {len(final_rows)}\n"
            f"Output folder:\n{out_dir}"
        )

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 6: QuestDialog / AIDialog Prefix ============================
def process_dialog_prefix():
    """
    For each row, inspect the first column string.
    If it contains 'aidialog' or 'questdialog' (case-insensitive),
    write the substring starting from that keyword into column 2.
    Save as *_PREFIX_OUTPUT.xlsx next to the source.
    """
    try:
        debug("Starting QuestDialog / AIDialog Prefix ...")

        in_path = filedialog.askopenfilename(
            title="Select Excel (first column will be processed)",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not in_path:
            return
        in_path = Path(in_path)

        wb = load_workbook(in_path)
        ws = wb.active
        processed = 0

        for r in range(1, ws.max_row + 1):
            raw = ws.cell(row=r, column=1).value
            if raw is None:
                continue
            txt   = str(raw)
            lower = txt.lower()

            idx = lower.find("aidialog")
            if idx == -1:
                idx = lower.find("questdialog")

            extracted = txt[idx:].lower() if idx != -1 else ""
            ws.cell(row=r, column=2, value=extracted)
            processed += 1

        out_path = in_path.with_name(f"{in_path.stem}_PREFIX_OUTPUT.xlsx")
        wb.save(out_path)
        debug(f"Processed {processed} rows → {out_path}")
        messagebox.showinfo("Done", f"Prefix extraction finished.\nSaved: {out_path}")

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== TOOL 7: Extract Audio File From Excel (case-insensitive + highlights) =========
def process_extract_audio_from_excel():
    """
    Read EventNames from an Excel file, find matching .wem files in a folder (case-insensitive),
    copy them into a new auto-generated output folder, and produce a new Excel
    with missing EventNames highlighted in red.
    """
    try:
        debug("Starting Extract Audio File From Excel ...")

        # 1) Ask for Excel file
        excel_path = filedialog.askopenfilename(
            title="Select Excel containing EventNames",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not excel_path:
            return
        excel_path = Path(excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active

        # Collect event names from all cells (case-insensitive)
        eventnames = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    name = _norm(str(cell))
                    if name.endswith(".wem"):
                        name = _norm(Path(name).stem)
                    eventnames.add(name.lower())

        if not eventnames:
            messagebox.showerror("Error", "No EventNames found in Excel.")
            return
        debug(f"Collected {len(eventnames)} EventNames from Excel.")

        # 2) Ask for folder containing .wem files
        wem_folder = filedialog.askdirectory(title="Select folder containing .wem files")
        if not wem_folder:
            return
        wem_folder = Path(wem_folder)

        wem_paths = list(wem_folder.rglob("*.wem"))
        if not wem_paths:
            messagebox.showerror("Error", "No .wem files found in the selected folder.")
            return

        # Map lowercase stem → path for case-insensitive matching
        stem2wem = {_norm(p.stem): p for p in wem_paths}
        debug(f"Indexed {len(stem2wem)} .wem files.")

        # 3) Create output folder
        out_dir = Path(__file__).parent / f"WEM_FROM_EXCEL_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        out_dir.mkdir(parents=True, exist_ok=True)

        # 4) Prepare highlight style
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 5) Copy files & mark missing
        copied_count = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                if cell.value:
                    name = _norm(str(cell.value))
                    if name.endswith(".wem"):
                        name = _norm(Path(name).stem)
                    name_lc = name.lower()
                    if name_lc in stem2wem:
                        # copy file
                        src = stem2wem[name_lc]
                        dst = out_dir / src.name
                        if not dst.exists():
                            shutil.copy2(src, dst)
                            copied_count += 1
                    else:
                        # highlight missing
                        ws.cell(row=r_idx, column=c_idx).fill = red_fill

        # 6) Save highlighted Excel in output folder
        out_excel = out_dir / f"{excel_path.stem}_HIGHLIGHTED.xlsx"
        wb.save(out_excel)

        debug(f"Copied {copied_count} .wem files to {out_dir}")
        debug(f"Highlighted Excel saved: {out_excel}")
        messagebox.showinfo(
            "Done",
            f"Copied {copied_count} .wem files.\n"
            f"Output folder:\n{out_dir}\n"
            f"Highlighted Excel:\n{out_excel}"
        )

    except Exception as exc:
        messagebox.showerror("Error", str(exc))


# ========== GUI launcher =====================================================
def run_gui():
    root = tk.Tk()
    root.title("SoundEventName Tools")

    tk.Button(root, text="AudioFileNameToEventName (UPDATED)", width=40,
              command=process_audiofilename_to_eventname).pack(pady=6)

    tk.Button(root, text="SoundEventName → StringID", width=40,
              command=lambda: process_soundeventname_to_stringid("SE_TO_SID")).pack(pady=6)

    tk.Button(root, text="StringID → SoundEventName", width=40,
              command=lambda: process_soundeventname_to_stringid("SID_TO_SE")).pack(pady=6)

    tk.Button(root, text="Replace Filenames (Excel Map)", width=40,
              command=process_replace_filenames).pack(pady=6)

    tk.Button(root, text="Extract Audio Filenames to Excel", width=40,
              command=process_extract_audio_filenames).pack(pady=6)

    tk.Button(root, text="Extract Audio File From Excel", width=40,
              command=process_extract_audio_from_excel).pack(pady=6)

    tk.Button(root, text="EventName → ExcelAudio LQA (UPDATED)", width=40,
              command=process_eventname_to_excelaudio_lqa).pack(pady=6)

    tk.Button(root, text="QuestDialog / AIDialog Prefix", width=40,
              command=process_dialog_prefix).pack(pady=6)

    root.mainloop()


# -----------------------------------------------------------------------------
if __name__ == "__main__":
    run_gui()
