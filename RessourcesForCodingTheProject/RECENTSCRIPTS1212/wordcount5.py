
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ============================================================
#  Translation Utilities – Word-Count  &  StrOrigin Comparison
# ============================================================
#
#  1.  Word-Count Process
#      • counts origin / translation words for every LanguageData_*.xml
#      • writes BOTH a compact and a detailed Excel summary
#      • writes one global Missing_From_Export.xml  +  per-file DIFF xml
#
#  2.  StrOrigin Comparison
#      • Full or per-sub-folder (“Granular”) comparison between BEFORE/AFTER trees
#      • counts files added / deleted / modified, StrOrigin changes & additions
#      • builds nicely formatted Excel report with TOP-50 stats
#
#  ------------------------------------------------------------
#  Author:  (your name)
#  Requires:  Python 3,  lxml,  openpyxl
# ============================================================

import os
import re
import sys
import copy
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from pathlib import Path
from datetime import datetime
import xml.etree.ElementTree as ET

from lxml import etree as LET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ----------------------------------------------------------------------
#  HELPERS TO HANDLE “FROZEN” EXECUTABLES  (PyInstaller, cx_Freeze, etc.)
# ----------------------------------------------------------------------
def get_base_dir() -> Path:
    """
    Return the directory that should be considered the “program folder”.

    • When running from source  :  path of the *.py file (Path(__file__).parent)
    • When running as EXE (frozen):  folder that contains the executable
      (Path(sys.executable).parent).  Using this instead of the temporary
      extraction dir (sys._MEIPASS) makes it possible to place the BEFORE /
      AFTER folders next to the exe.
    """
    if getattr(sys, "frozen", False):          # frozen by PyInstaller / cx_Freeze
        return Path(sys.executable).parent
    return Path(__file__).parent

# ----------------------------------------------------------------------
# CONSTANTS – adjust to your project layout
# ----------------------------------------------------------------------
# ---- Word-count -------------------------------------------------------
LANGUAGE_FOLDER     = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
EXPORT_FOLDER       = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
DIFF_OUTPUT_FOLDER  = Path.cwd() / "diff_entries_output"
MISSING_XML_FILE    = Path.cwd() / "Missing_From_Export.xml"   # GLOBAL list of missing nodes

# ---- StrOrigin comparison --------------------------------------------
BASE_DIR      = get_base_dir()
BEFORE_FOLDER = BASE_DIR / "BEFORE"
AFTER_FOLDER  = BASE_DIR / "AFTER"

# ----------------------------------------------------------------------
# REGEX / SMALL HELPERS
# ----------------------------------------------------------------------
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')
korean_re      = re.compile(r'[\uac00-\ud7a3]')

def fix_bad_entities(x: str) -> str:
    return _bad_entity_re.sub("&amp;", x)

def count_words(txt: str) -> int:
    return len([w for w in re.split(r"\s+", txt.strip()) if w])

def parse_xml_file(p: Path) -> LET._Element:
    raw     = p.read_text(encoding="utf-8", errors="ignore")
    wrapped = f"<ROOT>\n{fix_bad_entities(raw)}\n</ROOT>"
    parser  = LET.XMLParser(recover=True, huge_tree=True)
    return LET.fromstring(wrapped.encode("utf-8"), parser=parser)

def is_korean(txt: str) -> bool:
    return bool(korean_re.search(txt))

def iter_language_files(folder: Path):
    for r, _, fs in os.walk(folder):
        for f in fs:
            lf = f.lower()
            if lf.startswith("languagedata_") and lf.endswith(".xml"):
                yield Path(r) / f

def build_export_index(export_folder: Path):
    """
    Map StringId → category   (Dialog/xxx, Sequencer/xxx, ...)
    """
    idx = {}
    for xml in export_folder.rglob("*.xml"):
        try:
            root = parse_xml_file(xml)
        except Exception:
            continue
        parts = xml.relative_to(export_folder).parts
        if parts and parts[0].lower() == "dialog" and len(parts) > 1:
            cat = f"Dialog/{parts[1]}"
        elif parts and parts[0].lower() == "sequencer" and len(parts) > 1:
            cat = f"Sequencer/{parts[1]}"
        else:
            cat = parts[0] if parts else "Unknown"
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                idx[sid] = cat
    return idx

def collect_strorigin(folder: Path):
    """
    Walk folder tree and build:  {rel_file : {StringId: StrOrigin}}
    """
    mapping = {}
    for r, _, fs in os.walk(folder):
        for f in fs:
            if not f.lower().endswith(".xml"):
                continue
            fp  = Path(r) / f
            rel = fp.relative_to(folder).as_posix()
            try:
                tree = ET.parse(str(fp))
                rt   = tree.getroot()
                mp   = {}
                for loc in rt.findall(".//LocStr"):
                    sid = loc.get("StringId")
                    so  = loc.get("StrOrigin","")
                    if sid:
                        mp[sid] = so
                if mp:
                    mapping[rel] = mp
            except Exception:
                pass
    return mapping

def style_header(ws, row_idx, headers):
    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(bold=True, color="FFFFFF")
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    for col, txt in enumerate(headers, start=1):
        c = ws.cell(row_idx, col, txt)
        c.fill = fill
        c.font = font
        c.border = thick_border
        ws.column_dimensions[c.column_letter].width = max(len(str(txt)) + 4, 18)

def style_separator(ws, row_idx, num_cols: int):
    fill = PatternFill("solid", fgColor="FFFF00")
    for col in range(1, num_cols+1):
        c = ws.cell(row_idx, col, "")
        c.fill = fill

# ----------------------------------------------------------------------
#  GUI APP
# ----------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Translation Utilities – WordCount & StrOrigin")
        self.root.geometry("950x720")

        # -------------------------------------------------- Buttons
        fbtn = tk.Frame(root); fbtn.pack(pady=10)
        tk.Button(fbtn, text="Word-Count Process",
                  width=25, command=self.start_word_count).pack(side=tk.LEFT, padx=10)
        tk.Button(fbtn, text="StrOrigin Comparison",
                  width=25, command=self.start_strorigin_comparison).pack(side=tk.LEFT, padx=10)

        # -------------------------------------------------- Progress + Output
        self.progress = ttk.Progressbar(root, mode="determinate")
        self.progress.pack(fill=tk.X, padx=15, pady=5)

        tk.Label(root, text="Log / Output:").pack(anchor=tk.W, padx=15)
        self.out = scrolledtext.ScrolledText(root, height=30)
        self.out.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

    # ------------ small logger
    def log(self, *msg):
        self.out.insert(tk.END, " ".join(str(x) for x in msg) + "\n")
        self.out.see(tk.END)
        self.root.update_idletasks()

    # ==============================================================
    #  WORD-COUNT PROCESS
    # ==============================================================
    def start_word_count(self):
        threading.Thread(target=self.run_word_count, daemon=True).start()

    def run_word_count(self):
        self.out.delete(1.0, tk.END)
        ts_start = datetime.now()
        self.log(f"[{ts_start:%Y-%m-%d %H:%M:%S}] Word-count process started")
        self.log(f"LANGUAGE_FOLDER = {LANGUAGE_FOLDER}")
        self.log(f"EXPORT_FOLDER   = {EXPORT_FOLDER}\n")
        self.progress['value'] = 0

        # ---------- Build export StringId index
        self.log("Building export index ...")
        export_index = build_export_index(EXPORT_FOLDER)
        self.log(f"Indexed {len(export_index):,} StringIds\n")

        files = list(iter_language_files(LANGUAGE_FOLDER))
        if not files:
            messagebox.showerror("Word-count", "No LanguageData_*.xml found.")
            return

        total_files = len(files)
        DIFF_OUTPUT_FOLDER.mkdir(exist_ok=True)

        # accumulators
        per_file_rows = []    # (lang, file, origin_words, completed_origin_words, translated_words)
        category_stats = {}   # lang → cat → counters
        global_missing_nodes = []  # all nodes missing from export (for single XML)

        for idx, xml_path in enumerate(files, 1):
            self.progress['value'] = idx / total_files * 100
            self.log(f"({idx}/{total_files}) {xml_path.name}")

            stem = xml_path.stem
            parts = stem.split("_", 1)
            if len(parts) != 2:
                continue
            lang = parts[1].upper()
            if lang == "KOR":          # skip Korean source
                continue

            total_words = completed_origin_words = translated_words = 0
            diff_nodes = []

            try:
                tree = parse_xml_file(xml_path)
            except Exception:
                continue

            for loc in tree.iter("LocStr"):
                origin = (loc.get("StrOrigin") or "").strip()
                if not origin:
                    continue
                wc_origin = count_words(origin)
                total_words += wc_origin

                sid = loc.get("StringId")
                in_export = sid in export_index
                if not in_export:
                    diff_nodes.append(loc)
                    global_missing_nodes.append(copy.deepcopy(loc))

                trans = (loc.get("Str") or "").strip()
                cat   = export_index.get(sid, "Unknown")

                stats = category_stats\
                        .setdefault(lang, {})\
                        .setdefault(cat, {
                            "total_words":0, "total_nodes":0,
                            "completed_nodes":0, "completed_origin_words":0,
                            "translated_words":0
                        })
                stats["total_words"] += wc_origin
                stats["total_nodes"] += 1

                if trans and not is_korean(trans):
                    completed_origin_words += wc_origin
                    wc_trans = count_words(trans)
                    translated_words += wc_trans
                    stats["completed_nodes"] += 1
                    stats["completed_origin_words"] += wc_origin
                    stats["translated_words"] += wc_trans

            per_file_rows.append((
                lang, xml_path.name, total_words,
                completed_origin_words, translated_words
            ))

            # -------------- write DIFF xml (nodes missing from export) – per file
            if diff_nodes:
                root_diff = LET.Element("ROOT")
                for loc in diff_nodes:
                    root_diff.append(loc)
                out_path = DIFF_OUTPUT_FOLDER / f"{xml_path.stem}_DIFF.xml"
                LET.ElementTree(root_diff).write(
                    str(out_path), encoding="utf-8", xml_declaration=True)
                self.log(f"   -> diff written : {out_path}")

        # ------------------------------------------------ GLOBAL missing xml
        if global_missing_nodes:
            root_all = LET.Element("ROOT")
            for loc in global_missing_nodes:
                root_all.append(loc)
            LET.ElementTree(root_all).write(str(MISSING_XML_FILE),
                                            encoding="utf-8",
                                            xml_declaration=True)
            self.log(f"\nGlobal missing-from-export XML saved → {MISSING_XML_FILE}")

        # ------------------------------------------------ Excel summary
        self.progress['value'] = 95
        self.log("\nGenerating Excel summary ...")
        wb = Workbook()

        # ----------- Sheet 1 – compact per-file summary
        ws = wb.active
        ws.title = "Per-File"

        hdr = ["Language", "File",
               "Total Origin Words",
               "Completed Origin Words",
               "Translated Words",
               "Coverage %"]
        style_header(ws, 1, hdr)

        r = 2
        for lang, fn, tw, cow, tws in sorted(per_file_rows):
            cov = round(100 * cow / tw, 2) if tw else 0
            ws.cell(r, 1, lang)
            ws.cell(r, 2, fn)
            ws.cell(r, 3, tw)
            ws.cell(r, 4, cow)
            ws.cell(r, 5, tws)
            ws.cell(r, 6, cov)
            r += 1

        # autosize
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                15, max(len(str(c.value or "")) for c in col) + 2)

        # ----------- Sheet 2 – detailed per-language / per-category stats
        ws_det = wb.create_sheet("Detailed Summary")
        hdr_det = ["Category",
                   "Total Origin Words",
                   "Completed Origin Words",
                   "Translated Words",
                   "Coverage %"]

        row = 1
        for lang in sorted(category_stats):
            ws_det.cell(row, 1, f"Language: {lang}").font = Font(bold=True)
            row += 1
            style_header(ws_det, row, hdr_det)
            row += 1
            for cat, st in sorted(category_stats[lang].items()):
                cov = round(100 * st["completed_origin_words"] / st["total_words"], 2) \
                      if st["total_words"] else 0
                ws_det.cell(row, 1, cat)
                ws_det.cell(row, 2, st["total_words"])
                ws_det.cell(row, 3, st["completed_origin_words"])
                ws_det.cell(row, 4, st["translated_words"])
                ws_det.cell(row, 5, cov)
                row += 1
            style_separator(ws_det, row, len(hdr_det))
            row += 2

        for col in ws_det.columns:
            ws_det.column_dimensions[col[0].column_letter].width = max(
                15, max(len(str(c.value or "")) for c in col) + 2)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xlsx = Path.cwd() / f"Translation_Coverage_SUMMARY_{ts}.xlsx"
        wb.save(out_xlsx)
        self.log(f"\nExcel saved → {out_xlsx}")
        self.progress['value'] = 100
        messagebox.showinfo("Word-count", f"Done!\nReport: {out_xlsx}")

    # ==============================================================
    #  STRORIGIN COMPARISON
    # ==============================================================
    def start_strorigin_comparison(self):
        before = BEFORE_FOLDER
        after  = AFTER_FOLDER

        if not before.is_dir() or not after.is_dir():
            messagebox.showerror(
                "StrOrigin",
                "Required folders not found:\n\n"
                f"BEFORE : {before}\n"
                f"AFTER  : {after}\n\n"
                "Both folders must sit next to the executable."
            )
            return

        common = sorted({d.name for d in before.iterdir() if d.is_dir()} &
                        {d.name for d in after.iterdir()  if d.is_dir()})

        if not common:
            messagebox.showerror(
                "StrOrigin",
                "No matching top-level sub-folders found under "
                "BEFORE / AFTER."
            )
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("StrOrigin Comparison – options")
        dlg.grab_set()

        tk.Label(
            dlg,
            text=("Select the sub-folders you want to compare and tick "
                  "“Granular” if each of their sub-folders should be compared "
                  "individually:")
        ).pack(padx=10, pady=8)

        rows = []
        for name in common:
            fr   = tk.Frame(dlg); fr.pack(fill=tk.X, padx=15, pady=2)
            sel  = tk.BooleanVar(value=True)
            gran = tk.BooleanVar(value=False)

            tk.Checkbutton(fr, text=name, variable=sel).pack(side=tk.LEFT)
            tk.Label(fr, text="Granular").pack(side=tk.LEFT, padx=(25, 0))
            tk.Checkbutton(fr, variable=gran).pack(side=tk.LEFT)

            rows.append((name, sel, gran))

        selection = {"value": None}

        def on_ok():
            picked = [(n, g.get()) for n, s, g in rows if s.get()]
            if not picked:
                messagebox.showerror("StrOrigin", "Select at least one folder.")
                return
            selection["value"] = picked
            dlg.destroy()

        tk.Frame(dlg).pack(pady=5)
        tk.Button(dlg, text="OK",     width=10, command=on_ok)      \
            .pack(side=tk.LEFT,  padx=15, pady=10)
        tk.Button(dlg, text="Cancel", width=10, command=dlg.destroy) \
            .pack(side=tk.RIGHT, padx=15, pady=10)

        self.root.wait_window(dlg)
        if not selection["value"]:
            return

        todo = [(before / name, after / name, granular)
                for name, granular in selection["value"]]

        threading.Thread(
            target=self.run_strorigin_report,
            args=(todo,),
            daemon=True
        ).start()

    # --------------------------------------------------------------
    def run_strorigin_report(self, selections):
        self.out.delete(1.0, tk.END)
        self.progress['value'] = 0
        self.log(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] StrOrigin comparison started")
        total_sel = len(selections)

        global_change = {}
        global_add = {}
        summary_rows = []

        files_added_list = []
        files_modified_list = []

        for idx, (bdir, adir, gran) in enumerate(selections, 1):
            self.progress['value'] = idx / total_sel * 100
            label = f"{bdir.name}"
            self.log(f"\n[{idx}/{total_sel}] {label}")

            if not gran:
                stats, chg, add, added_files, modified_files = self.compare_pair_collect_files(
                    bdir, adir, label, global_change, global_add)
                summary_rows.append(stats)
                files_added_list.extend(added_files)
                files_modified_list.extend(modified_files)
            else:
                subs_b = {d.name for d in bdir.iterdir() if d.is_dir()}
                subs_a = {d.name for d in adir.iterdir() if d.is_dir()}
                com_sub = sorted(subs_b & subs_a)
                if not com_sub:
                    self.log("  (no common sub-folders)")
                for sub in com_sub:
                    self.log(f"  · Subfolder {sub}")
                    stats, chg, add, added_files, modified_files = self.compare_pair_collect_files(
                        bdir / sub, adir / sub,
                        f"{bdir.name}/{sub}",
                        global_change, global_add)
                    summary_rows.append(stats)
                    files_added_list.extend(added_files)
                    files_modified_list.extend(modified_files)

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        hdr = ["Folder", "Files Compared", "Files Modified", "Files Added",
               "Files Deleted", "StrOrigin Changes (Rows)", "Changed Words",
               "StrOrigin Additions (Rows)", "Added Words", "Grand Total Words"]

        header_fill = PatternFill("solid", fgColor="A7C7E7")
        medium_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        for c, val in enumerate(hdr, 1):
            cell = ws.cell(1, c, val)
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = medium_border

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        palette = ["F8CBAD", "C6E0B4", "FFD966",
                   "D9D2E9", "F4B183", "A9D08E", "FFE699"]
        parent_colors = {}
        color_index = 0

        r = 2
        for row_data in summary_rows:
            folder_name = row_data[0]
            parent = folder_name.split("/", 1)[0]
            if parent not in parent_colors:
                parent_colors[parent] = palette[color_index % len(palette)]
                color_index += 1
            fill = PatternFill("solid", fgColor=parent_colors[parent])

            for c, val in enumerate(row_data, 1):
                cell = ws.cell(r, c, val)
                cell.fill = fill
                cell.border = thin_border
                cell.font = Font(size=12, bold=(c == 1))
            r += 1

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        def make_top(name, data):
            w = wb.create_sheet(name)
            header_fill_top = PatternFill("solid", fgColor="A7C7E7")
            for c, val in enumerate(["File", "Count"], 1):
                cell = w.cell(1, c, val)
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill_top
            for i, (f, cnt) in enumerate(sorted(data.items(),
                                                key=lambda x: x[1],
                                                reverse=True)[:50], 2):
                w.cell(i, 1, f).font = Font(size=12)
                w.cell(i, 2, cnt).font = Font(size=12)
            for col in w.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                w.column_dimensions[col[0].column_letter].width = max_len + 4

        make_top("Top-50 Changes", global_change)
        make_top("Top-50 Additions", global_add)

        ws_added = wb.create_sheet("Files Added")
        ws_added.cell(1, 1, "File Path").font = Font(size=12, bold=True)
        ws_added.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_added.cell(1, 1).fill = header_fill
        for i, f in enumerate(sorted(files_added_list), 2):
            ws_added.cell(i, 1, f).font = Font(size=12)
        if files_added_list:
            ws_added.auto_filter.ref = f"A1:A{len(files_added_list)+1}"
        ws_added.column_dimensions['A'].width = max((len(f) for f in files_added_list), default=15) + 4

        ws_mod = wb.create_sheet("Files Modified")
        ws_mod.cell(1, 1, "File Path").font = Font(size=12, bold=True)
        ws_mod.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_mod.cell(1, 1).fill = header_fill
        for i, f in enumerate(sorted(files_modified_list), 2):
            ws_mod.cell(i, 1, f).font = Font(size=12)
        if files_modified_list:
            ws_mod.auto_filter.ref = f"A1:A{len(files_modified_list)+1}"
        ws_mod.column_dimensions['A'].width = max((len(f) for f in files_modified_list), default=15) + 4

        out = Path.cwd() / f"StrOrigin_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(out)
        self.log(f"\nExcel saved → {out}")
        self.progress['value'] = 100
        messagebox.showinfo("StrOrigin", "Done!\nReport:\n" + str(out))

    def compare_pair_collect_files(self, before_dir: Path, after_dir: Path,
                                   label: str,
                                   global_ch: dict,
                                   global_add: dict):
        bmap = collect_strorigin(before_dir)
        amap = collect_strorigin(after_dir)

        bf = set(bmap)
        af = set(amap)
        common = bf & af

        files_added = list(af - bf)
        files_deleted = len(bf - af)
        files_modified = 0

        total_changes_rows = total_changes_words = 0
        total_add_rows = total_add_words = 0

        change_counts = {}
        add_counts = {}

        modified_files_list = []
        added_files_list = []

        for rel in common:
            s = bmap[rel]
            t = amap[rel]
            changes = [sid for sid in s if sid in t and s[sid] != t[sid]]
            adds = [sid for sid in t if sid not in s]

            if changes:
                files_modified += 1
                change_counts[rel] = len(changes)
                total_changes_rows += len(changes)
                total_changes_words += sum(count_words(t[sid]) for sid in changes)
                modified_files_list.append(f"{label}/{rel}")

            if adds:
                add_counts[rel] = len(adds)
                total_add_rows += len(adds)
                total_add_words += sum(count_words(t[sid]) for sid in adds)

        for rel in af - bf:
            d = amap[rel]
            rows = len(d)
            if rows:
                add_counts[rel] = rows
                total_add_rows += rows
                total_add_words += sum(count_words(v) for v in d.values())
                added_files_list.append(f"{label}/{rel}")

        for f, c in change_counts.items():
            global_ch[f"{label}/{f}"] = global_ch.get(f"{label}/{f}", 0) + c
        for f, c in add_counts.items():
            global_add[f"{label}/{f}"] = global_add.get(f"{label}/{f}", 0) + c

        self.log(f"    Files compared : {len(common)}")
        self.log(f"    Files modified : {files_modified}")
        self.log(f"    Files added    : {len(files_added)}")
        self.log(f"    Files deleted  : {files_deleted}")
        self.log(f"    StrOrigin changes   : {total_changes_rows} rows / {total_changes_words} words")
        self.log(f"    StrOrigin additions : {total_add_rows} rows / {total_add_words} words")

        stats_row = [label, len(common), files_modified, len(files_added), files_deleted,
                     total_changes_rows, total_changes_words,
                     total_add_rows, total_add_words,
                     total_changes_words + total_add_words]
        return stats_row, change_counts, add_counts, added_files_list, modified_files_list

# ----------------------------------------------------------------------
def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
