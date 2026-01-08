#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Gimmick LQA Extractor – v2.0 (Multi-language output)

Extracts Gimmick data from StaticInfo XMLs with proper hierarchy:
  GimmickAttributeGroup (GimmickName) → GimmickInfo (StrKey, GimmickName) → DropItem (Key)

FILTER: Only includes gimmicks that have ALL of:
  - GimmickInfo with StrKey
  - GimmickInfo with GimmickName (not empty)
  - DropItem with Key

Output per-language Excel files with:
  Sheet 1: GimmickDropItem – Hierarchical view with Group→Gimmick→Item
  Sheet 2: GimmickFullHierarchy – Complete gimmick tree view
"""

from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable, Set
from dataclasses import dataclass, field

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION – FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
STATICINFO_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"
)
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\cd_lambda\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "Gimmick_LQA_Output"
LOG_FILE = base_path / "gimmick_lqa.log"


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING – FILE + CONSOLE
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("GimmickLQA")
log.setLevel(logging.DEBUG)

# _file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
# _file_handler.setFormatter(
    # logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
# )
# _file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

# log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ─────────────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class ItemData:
    """Item info for DropItem lookups."""
    strkey: str
    item_name: str       # KOR
    item_desc: str       # KOR


@dataclass
class GimmickEntry:
    """
    A valid gimmick entry with all required fields.
    Organized by: AttributeGroup (group_name) → GimmickInfo → DropItems
    """
    # From GimmickAttributeGroup
    group_name_kor: str          # GimmickName on AttributeGroup (group level)
    
    # From GimmickInfo
    gimmick_strkey: str          # StrKey
    gimmick_name_kor: str        # GimmickName on GimmickInfo
    
    # From DropItem(s)
    drop_item_keys: List[str] = field(default_factory=list)
    
    # For ordering (file + element order)
    source_file: str = ""
    order_index: int = 0


# ─────────────────────────────────────────────────────────────────────────────
# TEXT UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)


def normalize_placeholders(text: str) -> str:
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text


def contains_korean(text: str) -> bool:
    return bool(_korean_re.search(text))


def is_good_translation(text: str) -> bool:
    return bool(text) and not contains_korean(text)


# ─────────────────────────────────────────────────────────────────────────────
# XML UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)


def _preprocess_newlines(raw_content: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1)
        inner = inner.replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw_content, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(
        r'="([^"]*<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*&[^ltgapoqu][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )
    tag_stack: List[str] = []
    tag_open = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    fixed_lines: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            tag_stack.append(mo.group(1))
            fixed_lines.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if tag_stack and tag_stack[-1] == mc.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            else:
                fixed_lines.append(line)
            continue
        fixed_lines.append(line)
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.debug("Cannot read %s", path)
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    parser_strict = ET.XMLParser(huge_tree=True)
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=parser_strict)
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retrying with recovery", path.name)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True),
            )
        except ET.XMLSyntaxError:
            log.debug("Even recovery parse failed: %s", path)
            return None


def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),  # YELLOW - group
    1: PatternFill("solid", fgColor="D9E1F2"),  # BLUE - gimmick
    2: PatternFill("solid", fgColor="E2EFDA"),  # GREEN - item
}
_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ─────────────────────────────────────────────────────────────────────────────
def load_single_language(folder: Path, lang_code: str) -> Dict[str, Tuple[str, str]]:
    tbl: Dict[str, Tuple[str, str]] = {}
    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith(f"languagedata_{lang_code}"):
            continue
        log.info("Loading language file [%s] – %s", lang_code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            continue
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            tr = el.get("Str") or ""
            sid = el.get("StringId") or ""
            if not origin:
                continue
            normalized_origin = normalize_placeholders(origin)
            if normalized_origin in tbl:
                existing_tr, _ = tbl[normalized_origin]
                if is_good_translation(tr) and not is_good_translation(existing_tr):
                    tbl[normalized_origin] = (tr, sid)
            else:
                tbl[normalized_origin] = (tr, sid)
    log.info("Language %s loaded – %d entries", lang_code.upper(), len(tbl))
    return tbl


def parse_language_folder(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    log.info("Scanning language folder: %s", folder)
    langs: Dict[str, Dict[str, Tuple[str, str]]] = {}
    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_") or stem.startswith("languagedata_kor"):
            continue
        code = stem.split("_", 1)[1].lower()
        if code in langs:
            continue
        log.info("Loading language file [%s] – %s", code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            continue
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            tr = el.get("Str") or ""
            sid = el.get("StringId") or ""
            if not origin:
                continue
            normalized_origin = normalize_placeholders(origin)
            if normalized_origin in tbl:
                existing_tr, _ = tbl[normalized_origin]
                if is_good_translation(tr) and not is_good_translation(existing_tr):
                    tbl[normalized_origin] = (tr, sid)
            else:
                tbl[normalized_origin] = (tr, sid)
        langs[code] = tbl
        log.info("Language %s loaded – %d entries", code.upper(), len(tbl))
    log.info("Language tables loaded: %d", len(langs))
    return langs


# ─────────────────────────────────────────────────────────────────────────────
# STRING-KEY TABLE
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        log.error("Cannot parse StringKeyTable")
        return {}
    tbl: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        num = el.get("Key") or ""
        sk = el.get("StrKey") or ""
        if num and sk:
            tbl[sk.lower()] = num
    log.info("StringKeyTable entries: %d", len(tbl))
    return tbl


# ─────────────────────────────────────────────────────────────────────────────
# INDEX ITEMINFO
# ─────────────────────────────────────────────────────────────────────────────
def index_iteminfo(static_folder: Path) -> Dict[str, ItemData]:
    log.info("Indexing ItemInfo from StaticInfo...")
    items: Dict[str, ItemData] = {}
    file_count = 0
    for xml in iter_xml_files(static_folder):
        root = parse_xml_file(xml)
        if root is None:
            continue
        file_count += 1
        for item in root.iter("ItemInfo"):
            key = item.get("StrKey")
            name = item.get("ItemName") or ""
            desc = item.get("ItemDesc") or ""
            if key and key not in items:
                items[key] = ItemData(strkey=key, item_name=name, item_desc=desc)
    log.info("Indexed %d ItemInfo entries from %d files", len(items), file_count)
    return items


# ─────────────────────────────────────────────────────────────────────────────
# INDEX GIMMICKS – NEW STRUCTURE
# Parses: GimmickAttributeGroup → GimmickInfo → DropItem
# Only valid if: GimmickInfo.StrKey + GimmickInfo.GimmickName + DropItem.Key
# ─────────────────────────────────────────────────────────────────────────────
def find_parent_attribute_group(el: ET._Element) -> Tuple[str, ET._Element | None]:
    """
    Walk up from GimmickInfo to find nearest GimmickAttributeGroup with GimmickName.
    Returns (group_name, group_element) or ("", None).
    """
    parent = el.getparent()
    while parent is not None:
        if parent.tag == "GimmickAttributeGroup":
            gname = parent.get("GimmickName") or ""
            if gname:
                return gname, parent
        parent = parent.getparent()
    return "", None


def index_gimmicks_v2(static_folder: Path) -> List[GimmickEntry]:
    """
    Index gimmicks with the required structure.
    
    FILTER: Only include when ALL are present:
      - GimmickInfo with StrKey
      - GimmickInfo with GimmickName (not empty)
      - DropItem with Key
    
    Returns list of GimmickEntry in document order.
    """
    log.info("Indexing Gimmicks (v2) from StaticInfo...")
    entries: List[GimmickEntry] = []
    
    file_count = 0
    gimmick_count = 0
    skipped_no_name = 0
    skipped_no_drop = 0
    
    for xml in sorted(iter_xml_files(static_folder)):
        root = parse_xml_file(xml)
        if root is None:
            continue
        file_count += 1
        
        # Find all GimmickInfo elements
        for idx, gim_el in enumerate(root.iter("GimmickInfo")):
            strkey = gim_el.get("StrKey") or ""
            gim_name = gim_el.get("GimmickName") or ""
            
            if not strkey:
                continue
            
            # FILTER: Must have GimmickName
            if not gim_name:
                skipped_no_name += 1
                continue
            
            # Collect DropItem keys
            drop_keys: List[str] = []
            for drop_el in gim_el.findall(".//DropItem"):
                dk = drop_el.get("Key")
                if dk:
                    drop_keys.append(dk)
            
            # FILTER: Must have at least one DropItem
            if not drop_keys:
                skipped_no_drop += 1
                continue
            
            # Find parent GimmickAttributeGroup with GimmickName
            group_name, _ = find_parent_attribute_group(gim_el)
            
            entry = GimmickEntry(
                group_name_kor=group_name,
                gimmick_strkey=strkey,
                gimmick_name_kor=gim_name,
                drop_item_keys=drop_keys,
                source_file=xml.name,
                order_index=idx,
            )
            entries.append(entry)
            gimmick_count += 1
    
    log.info(
        "Indexed %d valid gimmicks from %d files (skipped: %d no name, %d no dropitem)",
        gimmick_count, file_count, skipped_no_name, skipped_no_drop
    )
    return entries


# ─────────────────────────────────────────────────────────────────────────────
# TRANSLATION HELPER
# ─────────────────────────────────────────────────────────────────────────────
def translate(
    lang_tbl: Dict[str, Tuple[str, str]],
    kor_text: str,
    fallback_to_kor: bool = True,
) -> str:
    if not kor_text:
        return ""
    norm = normalize_placeholders(kor_text)
    result = lang_tbl.get(norm, ("", ""))[0]
    if result and is_good_translation(result):
        return result
    return kor_text if fallback_to_kor else ""


def get_string_id(lang_tbl: Dict[str, Tuple[str, str]], kor_text: str) -> str:
    if not kor_text:
        return ""
    norm = normalize_placeholders(kor_text)
    return lang_tbl.get(norm, ("", ""))[1]


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: GIMMICK DROPITEM – HIERARCHICAL VIEW
# ─────────────────────────────────────────────────────────────────────────────
def write_dropitem_sheet(
    wb: Workbook,
    lang_code: str,
    entries: List[GimmickEntry],
    items: Dict[str, ItemData],
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
) -> None:
    """
    Sheet 1: GimmickDropItem with STATUS column (ISSUE/NO ISSUE/BLOCKED only)
    """
    code = lang_code.upper()
    ws = wb.create_sheet(title="GimmickDropItem")

    # Helper to add STATUS drop-down
    def _add_status_validation(sh, status_col_idx: int, max_row: int) -> None:
        col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED"',
            allow_blank=True,
            showErrorMessage=True,
        )
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)
        sh.add_data_validation(dv)

    # Build headers
    headers = [
        "Depth",
        "Type",
        "GroupName(KOR)",
        f"GroupName({code})" if lang_code != "eng" else "GroupName(ENG)",
        "GimmickKey",
        "GimmickName(KOR)",
        f"GimmickName({code})" if lang_code != "eng" else "GimmickName(ENG)",
        "ItemKey",
        "ItemName(KOR)",
        f"ItemName({code})" if lang_code != "eng" else "ItemName(ENG)",
        "ItemDesc(KOR)",
        f"ItemDesc({code})" if lang_code != "eng" else "ItemDesc(ENG)",
        "Command",
        "StringID",
        "STATUS",  # NEW COLUMN
    ]

    # Write headers
    for col, txt in enumerate(headers, start=1):
        cell = ws.cell(1, col, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border

    # Hidden columns
    hidden_cols = {"Depth", "GroupName(KOR)", "GimmickName(KOR)", "ItemName(KOR)", "ItemDesc(KOR)"}
    for idx, h in enumerate(headers, 1):
        if h in hidden_cols:
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    # Build hierarchical rows
    rows_data: List[Tuple[List, int, str]] = []
    last_group = None
    last_gimmick = None

    for entry in entries:
        if entry.group_name_kor and entry.group_name_kor != last_group:
            group_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.group_name_kor)
            row = [0, "Group", entry.group_name_kor, group_loc, "", "", "", "", "", "", "", "", "", "", ""]
            rows_data.append((row, 0, "Group"))
            last_group = entry.group_name_kor
            last_gimmick = None

        if entry.gimmick_strkey != last_gimmick:
            gim_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.gimmick_name_kor)
            row = [1, "Gimmick", entry.group_name_kor, translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.group_name_kor),
                   entry.gimmick_strkey, entry.gimmick_name_kor, gim_loc, "", "", "", "", "", "", "", ""]
            rows_data.append((row, 1, "Gimmick"))
            last_gimmick = entry.gimmick_strkey

        for item_key in entry.drop_item_keys:
            itm = items.get(item_key)
            item_kor = itm.item_name if itm else ""
            item_desc_kor = itm.item_desc if itm else ""
            item_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, item_kor)
            desc_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, item_desc_kor)
            cmd = f"/create item {item_key}"
            sid = get_string_id(lang_tbl, item_kor) or get_string_id(eng_tbl, item_kor)
            row = [2, "Item", entry.group_name_kor, translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.group_name_kor),
                   entry.gimmick_strkey, entry.gimmick_name_kor, translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.gimmick_name_kor),
                   item_key, item_kor, item_loc, item_desc_kor, desc_loc, cmd, sid, ""]
            rows_data.append((row, 2, "Item"))

    # Write rows
    for r_idx, (row, depth, row_type) in enumerate(rows_data, start=2):
        fill = _depth_fill.get(depth, _depth_fill[2])
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.fill = fill
            cell.border = _border
            if row_type in ("Group", "Gimmick"):
                cell.font = _bold_font
            cell.alignment = Alignment(
                horizontal="left" if c_idx > 1 else "center",
                vertical="top",
                wrap_text=True,
                indent=depth if c_idx == 2 else 0
            )

    # Add STATUS validation
    status_col_idx = headers.index("STATUS") + 1
    _add_status_validation(ws, status_col_idx, ws.max_row)

    # Finalize
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows_data)+1}"
    ws.freeze_panes = "A2"

    # Column widths
    width_map = {
        "Depth": 8, "Type": 10, "GroupName(KOR)": 20, f"GroupName({code})": 25, "GroupName(ENG)": 25,
        "GimmickKey": 40, "GimmickName(KOR)": 20, f"GimmickName({code})": 25, "GimmickName(ENG)": 25,
        "ItemKey": 30, "ItemName(KOR)": 20, f"ItemName({code})": 25, "ItemName(ENG)": 25,
        "ItemDesc(KOR)": 25, f"ItemDesc({code})": 35, "ItemDesc(ENG)": 35,
        "Command": 35, "StringID": 15, "STATUS": 15
    }
    for idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width_map.get(h, 20)

    log.info("  Sheet GimmickDropItem: %d rows", len(rows_data))


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: GIMMICK FULL HIERARCHY (simpler flat view)
# ─────────────────────────────────────────────────────────────────────────────
def write_flat_sheet(
    wb: Workbook,
    lang_code: str,
    entries: List[GimmickEntry],
    items: Dict[str, ItemData],
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
) -> None:
    """
    Sheet 2: GimmickFlat
    
    One row per Gimmick→Item pair, flat view for easy filtering.
    
    Columns:
    - GroupName(KOR) [hidden]
    - GroupName(LOC)
    - GimmickKey
    - GimmickName(KOR) [hidden]
    - GimmickName(LOC)
    - ItemKey
    - ItemName(KOR) [hidden]
    - ItemName(LOC)
    - ItemDesc(KOR) [hidden]
    - ItemDesc(LOC)
    - Command
    - StringID
    """
    code = lang_code.upper()
    ws = wb.create_sheet(title="GimmickFlat")

    # Build headers
    headers = [
        "GroupName(KOR)",
        f"GroupName({code})" if lang_code != "eng" else "GroupName(ENG)",
        "GimmickKey",
        "GimmickName(KOR)",
        f"GimmickName({code})" if lang_code != "eng" else "GimmickName(ENG)",
        "ItemKey",
        "ItemName(KOR)",
        f"ItemName({code})" if lang_code != "eng" else "ItemName(ENG)",
        "ItemDesc(KOR)",
        f"ItemDesc({code})" if lang_code != "eng" else "ItemDesc(ENG)",
        "Command",
        "StringID",
    ]

    # Write headers
    for col, txt in enumerate(headers, start=1):
        cell = ws.cell(1, col, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border

    # Hidden columns
    hidden_cols = {"GroupName(KOR)", "GimmickName(KOR)", "ItemName(KOR)", "ItemDesc(KOR)"}
    for idx, h in enumerate(headers, 1):
        if h in hidden_cols:
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    # Build flat rows
    rows_data: List[List] = []
    
    for entry in entries:
        group_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.group_name_kor)
        gim_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, entry.gimmick_name_kor)
        
        for item_key in entry.drop_item_keys:
            itm = items.get(item_key)
            item_kor = itm.item_name if itm else ""
            item_desc_kor = itm.item_desc if itm else ""
            
            item_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, item_kor)
            desc_loc = translate(lang_tbl if lang_code != "eng" else eng_tbl, item_desc_kor)
            cmd = f"/create item {item_key}"
            sid = get_string_id(lang_tbl, item_kor) or get_string_id(eng_tbl, item_kor)
            
            row = [
                entry.group_name_kor, group_loc,
                entry.gimmick_strkey,
                entry.gimmick_name_kor, gim_loc,
                item_key,
                item_kor, item_loc,
                item_desc_kor, desc_loc,
                cmd, sid
            ]
            rows_data.append(row)

    # Write rows with alternating colors
    fill_a = PatternFill("solid", fgColor="E2EFDA")
    fill_b = PatternFill("solid", fgColor="FCE4D6")
    current_fill = fill_a
    last_gimmick = None

    for r_idx, row in enumerate(rows_data, start=2):
        gim_key = row[2]  # GimmickKey
        if last_gimmick is not None and gim_key != last_gimmick:
            current_fill = fill_b if current_fill == fill_a else fill_a
        last_gimmick = gim_key

        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.fill = current_fill
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Finalize
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows_data)+1}"
    ws.freeze_panes = "A2"

    # Column widths
    width_map = {
        "GroupName(KOR)": 20,
        f"GroupName({code})": 25,
        "GroupName(ENG)": 25,
        "GimmickKey": 40,
        "GimmickName(KOR)": 20,
        f"GimmickName({code})": 25,
        "GimmickName(ENG)": 25,
        "ItemKey": 30,
        "ItemName(KOR)": 20,
        f"ItemName({code})": 25,
        "ItemName(ENG)": 25,
        "ItemDesc(KOR)": 25,
        f"ItemDesc({code})": 35,
        "ItemDesc(ENG)": 35,
        "Command": 35,
        "StringID": 15,
    }
    for idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width_map.get(h, 20)

    log.info("  Sheet GimmickFlat: %d rows", len(rows_data))


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 70)
    log.info("Gimmick LQA Extractor – v2.0 (Multi-language output)")
    log.info("=" * 70)

    # Validate paths
    if not STATICINFO_FOLDER.exists():
        log.error("STATICINFO_FOLDER not found: %s", STATICINFO_FOLDER)
        sys.exit(1)
    if not LANGUAGE_FOLDER.exists():
        log.error("LANGUAGE_FOLDER not found: %s", LANGUAGE_FOLDER)
        sys.exit(1)

    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # Load English table first
    eng_tbl = load_single_language(LANGUAGE_FOLDER, "eng")
    if not eng_tbl:
        log.warning("No English table found – translations may be incomplete")

    # Load all language tables
    lang_tables = parse_language_folder(LANGUAGE_FOLDER)
    if not lang_tables:
        log.error("No language files found!")
        sys.exit(1)

    # Load string key table
    id_tbl = load_string_key_table(STRINGKEYTABLE_FILE)

    # Index items
    items = index_iteminfo(STATICINFO_FOLDER)

    # Index gimmicks (v2 - proper filtering)
    entries = index_gimmicks_v2(STATICINFO_FOLDER)

    if not entries:
        log.error("No valid gimmick entries found!")
        sys.exit(1)

    # Process each language
    log.info("Processing languages...")
    total = len(lang_tables)
    for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
        log.info("(%d/%d) Language %s", idx, total, code.upper())

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Hierarchical DropItem view
        write_dropitem_sheet(wb, code, entries, items, tbl, eng_tbl, id_tbl)

        # Sheet 2: Flat view
        write_flat_sheet(wb, code, entries, items, tbl, eng_tbl, id_tbl)

        # Save
        out_path = OUTPUT_FOLDER / f"Gimmick_LQA_{code.upper()}.xlsx"
        wb.save(out_path)
        log.info("  Saved: %s", out_path.name)

    log.info("=" * 70)
    log.info("Done! Output: %s", OUTPUT_FOLDER)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
