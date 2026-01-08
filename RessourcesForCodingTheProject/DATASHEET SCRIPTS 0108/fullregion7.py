#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Region-data extractor v3.1
-----------------------------------------------
COMPLETE RESTRUCTURE - Now follows natural XML hierarchy:

FactionGroup → Faction → FactionNode (nested)

Structure:
• ONE Excel sheet per FactionGroup (using GroupName as tab name)
• Each sheet contains Factions as major sections (INCLUDING 세력 없음 if inside FactionGroup)
• FactionNodes nested under their parent Factions
• Proper indentation showing hierarchy depth

Key changes from v3.0:
• 세력 없음 inside FactionGroup → stays in that sheet as a Faction section
• 세력 없음 standalone (not in any FactionGroup) → separate sheet

Key changes from v2.x:
• REMOVED: Territory/RegionInfo-based logic
• REMOVED: Territory code extraction matching
• NEW: Direct FactionGroup → Faction → FactionNode parsing
• NEW: FactionNode.Name comes from KnowledgeInfo lookup (KnowledgeKey → StrKey match)
• NEW: FactionNode.Desc used directly as description
• KEPT: Shop parsing (separate sheet)

Hierarchy in Excel:
  [FactionGroup.GroupName as Sheet Tab]
    Faction.Name (depth 0, bold header) - including "세력 없음" if present
      FactionNode.Name (depth 1) - from KnowledgeInfo lookup
        Description (depth 2, italic) - from FactionNode.Desc
        Child FactionNode (depth 2)
          Description (depth 3)
          ...
  
  [세력 없음 Sheet - only for standalone Empty Factions]
    FactionNode entries from Factions not inside any FactionGroup
"""

from __future__ import annotations

import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple
from collections import OrderedDict

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"
)
SHOP_FILE = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\staticinfo_quest\funcnpc\shop_world.staticinfo.xml"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "Region_LQA_v3"
LOG_FILE = base_path / "region_scan_v3.log"

# Empty Faction identifier (Korean)
EMPTY_FACTION_NAME = "세력 없음"

# ──────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────
log = logging.getLogger("RegionLQA")
log.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

log.addHandler(_console_handler)
log.propagate = False

# ──────────────────────────────────────────────────────────────────────
# NORMALIZATION (FOR TRANSLATION MATCHING)
# ──────────────────────────────────────────────────────────────────────
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)


def normalize_placeholders(text: str) -> str:
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text


# ──────────────────────────────────────────────────────────────────────
# KOREAN DETECTION
# ──────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')


def contains_korean(text: str) -> bool:
    return bool(_korean_re.search(text))


def is_good_translation(text: str) -> bool:
    return bool(text) and not contains_korean(text)


# ──────────────────────────────────────────────────────────────────────
# XML SANITIZATION
# ──────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r"&(?!lt;|gt;|amp;|apos;|quot;)")


def _fix_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)


def _escape_newlines_in_seg(txt: str) -> str:
    def repl(m: re.Match) -> str:
        seg = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r", "")
        return "<seg>{}</seg>".format(seg)
    return re.sub(r"<seg>(.*?)</seg>", repl, txt, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = _fix_entities(raw)
    raw = _escape_newlines_in_seg(raw)

    raw = re.sub(
        r'="([^"]*?<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*?&[^ltgapoqu"][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    tag_stack: List[str] = []
    o_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    c_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed: List[str] = []
    for line in raw.splitlines():
        s = line.strip()
        m_open = o_re.match(s)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed.append(line)
            continue
        m_close = c_re.match(s)
        if m_close:
            while tag_stack and tag_stack[-1] != m_close.group(1):
                fixed.append("</{}>".format(tag_stack.pop()))
            if tag_stack:
                tag_stack.pop()
            fixed.append(line)
            continue
        if s.startswith("</>") and tag_stack:
            fixed.append(line.replace("</>", "</{}>".format(tag_stack.pop())))
            continue
        fixed.append(line)
    while tag_stack:
        fixed.append("</{}>".format(tag_stack.pop()))
    return "\n".join(fixed)


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None

    wrapped = "<ROOT>\n{}\n</ROOT>".format(sanitize_xml(raw))
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retry with recover", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True, recover=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse failed: %s", path)
            return None


def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn


# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES (v3.0 - Simplified hierarchy)
# ──────────────────────────────────────────────────────────────────────
@dataclass
class FactionNodeData:
    """A single FactionNode with nested children"""
    strkey: str
    name: str              # Display name (from KnowledgeInfo.Name lookup)
    original_name: str     # Original Name attribute (fallback)
    description: str       # From FactionNode.Desc
    knowledge_key: str
    node_type: str         # Main, Sub, etc.
    children: List["FactionNodeData"] = field(default_factory=list)


@dataclass
class FactionData:
    """A Faction containing multiple FactionNodes"""
    strkey: str
    name: str              # Faction Name attribute
    knowledge_key: str
    nodes: List[FactionNodeData] = field(default_factory=list)


@dataclass
class FactionGroupData:
    """A FactionGroup (top-level, becomes sheet tab)"""
    strkey: str
    group_name: str        # GroupName attribute - used as sheet tab name
    knowledge_key: str
    factions: List[FactionData] = field(default_factory=list)


@dataclass
class ShopStage:
    """A single shop stage"""
    strkey: str
    name: str
    description: str
    category: str = ""
    npc_key: str = ""
    dev_comment: str = ""


@dataclass
class ShopGroup:
    """A shop group"""
    tag_name: str
    name: str
    group: str
    stages: List[ShopStage] = field(default_factory=list)


# ──────────────────────────────────────────────────────────────────────
# KNOWLEDGE NAME LOOKUP (KnowledgeKey → Name)
# ──────────────────────────────────────────────────────────────────────
def build_knowledge_name_lookup(folder: Path) -> Dict[str, str]:
    """
    Build lookup: KnowledgeInfo.StrKey → KnowledgeInfo.Name
    
    Used to get correct display names for FactionNodes via their KnowledgeKey.
    """
    log.info("Building Knowledge Name lookup (StrKey → Name)…")
    
    name_lookup: Dict[str, str] = {}
    duplicates = 0
    
    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue
        
        for ki in root_el.iter("KnowledgeInfo"):
            strkey = ki.get("StrKey") or ""
            name = ki.get("Name") or ""
            
            if not strkey or not name:
                continue
            
            if strkey in name_lookup:
                duplicates += 1
                continue
            
            name_lookup[strkey] = name
    
    log.info("  → %d entries (%d duplicates ignored)", len(name_lookup), duplicates)
    return name_lookup


# ──────────────────────────────────────────────────────────────────────
# PARSE FACTION HIERARCHY (v3.0 - Main logic)
# ──────────────────────────────────────────────────────────────────────
def parse_faction_node_recursive(
    elem: ET._Element,
    knowledge_lookup: Dict[str, str],
    global_seen: Set[str],
    depth: int = 0
) -> Optional[FactionNodeData]:
    """
    Parse a FactionNode element and all its nested FactionNode children.
    
    Name resolution:
    1. Get KnowledgeKey from FactionNode
    2. Look up KnowledgeInfo where StrKey == KnowledgeKey
    3. Use KnowledgeInfo.Name as display name
    4. Fallback to FactionNode.Name if no match
    
    Description: Use FactionNode.Desc directly
    """
    strkey = elem.get("StrKey") or ""
    original_name = elem.get("Name") or ""
    knowledge_key = elem.get("KnowledgeKey") or ""
    description = elem.get("Desc") or ""
    node_type = elem.get("Type") or ""
    
    # Resolve display name via KnowledgeInfo lookup
    if knowledge_key and knowledge_key in knowledge_lookup:
        display_name = knowledge_lookup[knowledge_key]
    else:
        display_name = original_name
    
    if not display_name:
        return None
    
    # Duplicate check
    if strkey:
        if strkey in global_seen:
            log.debug("  Duplicate FactionNode skipped: %s", strkey)
            return None
        global_seen.add(strkey)
    
    node = FactionNodeData(
        strkey=strkey,
        name=display_name,
        original_name=original_name,
        description=description,
        knowledge_key=knowledge_key,
        node_type=node_type,
    )
    
    # Parse nested FactionNodes
    for child_elem in elem:
        if child_elem.tag == "FactionNode":
            child_node = parse_faction_node_recursive(
                child_elem, knowledge_lookup, global_seen, depth + 1
            )
            if child_node:
                node.children.append(child_node)
    
    return node


def parse_faction_element(
    elem: ET._Element,
    knowledge_lookup: Dict[str, str],
    global_seen: Set[str]
) -> Optional[FactionData]:
    """
    Parse a Faction element and all its FactionNode children.
    """
    strkey = elem.get("StrKey") or ""
    name = elem.get("Name") or ""
    knowledge_key = elem.get("KnowledgeKey") or ""
    
    if not name:
        return None
    
    faction = FactionData(
        strkey=strkey,
        name=name,
        knowledge_key=knowledge_key,
    )
    
    # Parse direct FactionNode children
    for child_elem in elem:
        if child_elem.tag == "FactionNode":
            node = parse_faction_node_recursive(child_elem, knowledge_lookup, global_seen)
            if node:
                faction.nodes.append(node)
    
    return faction


def parse_faction_group_element(
    elem: ET._Element,
    knowledge_lookup: Dict[str, str],
    global_seen: Set[str]
) -> Optional[FactionGroupData]:
    """
    Parse a FactionGroup element and all its Faction children.
    """
    strkey = elem.get("StrKey") or ""
    group_name = elem.get("GroupName") or ""
    knowledge_key = elem.get("KnowledgeKey") or ""
    
    if not group_name:
        return None
    
    faction_group = FactionGroupData(
        strkey=strkey,
        group_name=group_name,
        knowledge_key=knowledge_key,
    )
    
    # Parse Faction children
    for child_elem in elem:
        if child_elem.tag == "Faction":
            faction = parse_faction_element(child_elem, knowledge_lookup, global_seen)
            if faction:
                faction_group.factions.append(faction)
    
    return faction_group


def parse_all_faction_groups(
    folder: Path,
    knowledge_lookup: Dict[str, str],
    global_seen: Set[str]
) -> List[FactionGroupData]:
    """
    Parse all FactionGroup elements from all XML files.
    
    Returns list of FactionGroupData, each becoming a sheet tab.
    """
    log.info("Parsing FactionGroup → Faction → FactionNode hierarchy…")
    
    faction_groups: List[FactionGroupData] = []
    seen_group_strkeys: Set[str] = set()
    
    total_factions = 0
    total_nodes = 0
    empty_factions = 0
    
    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue
        
        for fg_elem in root_el.iter("FactionGroup"):
            strkey = fg_elem.get("StrKey") or ""
            
            # Skip duplicate FactionGroups
            if strkey and strkey in seen_group_strkeys:
                continue
            if strkey:
                seen_group_strkeys.add(strkey)
            
            fg = parse_faction_group_element(fg_elem, knowledge_lookup, global_seen)
            if fg and fg.factions:
                faction_groups.append(fg)
                
                for faction in fg.factions:
                    total_factions += 1
                    if faction.name == EMPTY_FACTION_NAME:
                        empty_factions += 1
                    total_nodes += count_nodes_recursive(faction.nodes)
    
    log.info("  → %d FactionGroups, %d Factions (%d empty), %d FactionNodes",
             len(faction_groups), total_factions, empty_factions, total_nodes)
    
    return faction_groups


def count_nodes_recursive(nodes: List[FactionNodeData]) -> int:
    """Count total nodes including nested children."""
    count = len(nodes)
    for node in nodes:
        count += count_nodes_recursive(node.children)
    return count


# ──────────────────────────────────────────────────────────────────────
# PARSE STANDALONE FACTIONS (not inside any FactionGroup)
# ──────────────────────────────────────────────────────────────────────
def parse_standalone_factions(
    folder: Path,
    knowledge_lookup: Dict[str, str],
    global_seen: Set[str],
    faction_group_strkeys: Set[str]
) -> List[FactionData]:
    """
    Parse Faction elements that are NOT inside any FactionGroup.
    These are standalone factions (like Faction_Empty_Akapen, Faction_Empty_Crimson_Desert).
    
    Args:
        folder: Path to XML files
        knowledge_lookup: KnowledgeKey → Name lookup
        global_seen: Global duplicate tracking set
        faction_group_strkeys: Set of FactionGroup StrKeys already processed
    
    Returns:
        List of standalone FactionData (typically 세력 없음 factions)
    """
    log.info("Parsing standalone Factions (not in any FactionGroup)…")
    
    standalone_factions: List[FactionData] = []
    seen_faction_strkeys: Set[str] = set()
    
    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue
        
        # First, collect all Faction StrKeys that ARE inside FactionGroups
        factions_in_groups: Set[str] = set()
        for fg_elem in root_el.iter("FactionGroup"):
            for faction_elem in fg_elem.iter("Faction"):
                sk = faction_elem.get("StrKey") or ""
                if sk:
                    factions_in_groups.add(sk)
        
        # Now find Factions that are direct children of ROOT (not in FactionGroup)
        for child in root_el:
            if child.tag == "Faction":
                strkey = child.get("StrKey") or ""
                
                # Skip if already processed
                if strkey and strkey in seen_faction_strkeys:
                    continue
                
                # Skip if this faction is inside a FactionGroup (we found it elsewhere)
                if strkey and strkey in factions_in_groups:
                    continue
                
                if strkey:
                    seen_faction_strkeys.add(strkey)
                
                faction = parse_faction_element(child, knowledge_lookup, global_seen)
                if faction and faction.nodes:
                    standalone_factions.append(faction)
                    log.debug("  Standalone Faction: %s (%d nodes)", 
                             faction.name, len(faction.nodes))
    
    # Count nodes
    total_nodes = sum(count_nodes_recursive(f.nodes) for f in standalone_factions)
    empty_factions = [f for f in standalone_factions if f.name == EMPTY_FACTION_NAME]
    
    log.info("  → %d standalone Factions found (%d are 세력 없음), %d total nodes",
             len(standalone_factions), len(empty_factions), total_nodes)
    
    return standalone_factions


# ──────────────────────────────────────────────────────────────────────
# SHOP PARSING (kept from v2.x)
# ──────────────────────────────────────────────────────────────────────
def parse_shop_file(shop_path: Path, global_seen: Set[str]) -> List[ShopGroup]:
    if not shop_path.exists():
        log.warning("Shop file not found: %s", shop_path)
        return []

    log.info("Parsing shop file: %s", shop_path)

    root_el = parse_xml_file(shop_path)
    if root_el is None:
        log.error("Failed to parse shop file")
        return []

    merged_groups: OrderedDict[str, ShopGroup] = OrderedDict()
    duplicate_count = 0

    for child in root_el:
        if not isinstance(child.tag, str):
            continue

        name = child.get("Name") or ""
        group = child.get("Group") or ""

        if not name:
            continue

        new_stages: List[ShopStage] = []
        for stage_el in child.iter("Stage"):
            strkey = stage_el.get("StrKey") or ""
            stage_name = stage_el.get("Name") or ""
            desc = stage_el.get("Desc") or ""
            category = stage_el.get("Category") or ""
            npc_key = stage_el.get("NPCShopCharacterKey") or ""
            dev_comment = stage_el.get("DevComment") or ""

            if not stage_name:
                continue

            if strkey:
                if strkey in global_seen:
                    duplicate_count += 1
                    continue
                global_seen.add(strkey)

            stage = ShopStage(
                strkey=strkey,
                name=stage_name,
                description=desc,
                category=category,
                npc_key=npc_key,
                dev_comment=dev_comment,
            )
            new_stages.append(stage)

        if not new_stages:
            continue

        if name in merged_groups:
            merged_groups[name].stages.extend(new_stages)
        else:
            merged_groups[name] = ShopGroup(
                tag_name=child.tag,
                name=name,
                group=group,
                stages=new_stages,
            )

    shop_groups = list(merged_groups.values())
    log.info("  → %d groups, %d stages (%d duplicates)",
             len(shop_groups), sum(len(g.stages) for g in shop_groups), duplicate_count)

    return shop_groups


# ──────────────────────────────────────────────────────────────────────
# ROW GENERATION (v3.0)
# ──────────────────────────────────────────────────────────────────────
# Row format: (depth, text, style_type, is_description)
RowItem = Tuple[int, str, str, bool]


def emit_faction_node_rows(node: FactionNodeData, depth: int) -> List[RowItem]:
    """Generate rows for a FactionNode and its children."""
    rows: List[RowItem] = []
    
    # Node name
    style = f"FactionNode_{node.node_type}" if node.node_type else "FactionNode"
    rows.append((depth, node.name, style, False))
    
    # Description (if exists)
    if node.description:
        rows.append((depth + 1, node.description, "Description", True))
    
    # Child nodes
    for child in node.children:
        rows.extend(emit_faction_node_rows(child, depth + 1))
    
    return rows


def emit_faction_rows(faction: FactionData, depth: int = 0) -> List[RowItem]:
    """Generate rows for a Faction and all its FactionNodes."""
    rows: List[RowItem] = []
    
    # Faction header
    rows.append((depth, faction.name, "Faction", False))
    
    # FactionNodes
    for node in faction.nodes:
        rows.extend(emit_faction_node_rows(node, depth + 1))
    
    return rows


def emit_faction_group_rows(fg: FactionGroupData) -> List[RowItem]:
    """Generate all rows for a FactionGroup (one sheet's content)."""
    rows: List[RowItem] = []
    
    for faction in fg.factions:
        rows.extend(emit_faction_rows(faction, 0))
    
    return rows


def emit_standalone_faction_rows(standalone_factions: List[FactionData]) -> List[RowItem]:
    """Generate rows for standalone Factions (not in any FactionGroup)."""
    rows: List[RowItem] = []
    
    for faction in standalone_factions:
        # Add Faction name as header
        rows.append((0, faction.name, "Faction", False))
        
        for node in faction.nodes:
            rows.extend(emit_faction_node_rows(node, 1))
    
    return rows


def emit_shop_rows(shop_groups: List[ShopGroup]) -> List[RowItem]:
    """Generate rows for Shop data."""
    rows: List[RowItem] = []
    
    for group in shop_groups:
        rows.append((0, group.name, "ShopGroup", False))
        for stage in group.stages:
            rows.append((1, stage.name, "ShopStage", False))
            if stage.description:
                rows.append((2, stage.description, "Description", True))
    
    return rows


# ──────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ──────────────────────────────────────────────────────────────────────
def load_language_tables(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    tables: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            normalized_origin = normalize_placeholders(origin)

            if normalized_origin in tbl:
                existing_tr, _ = tbl[normalized_origin]
                if is_good_translation(tr) and not is_good_translation(existing_tr):
                    tbl[normalized_origin] = (tr, sid)
            else:
                tbl[normalized_origin] = (tr, sid)

        tables[lang] = tbl
        log.info("  Language %s: %d entries", lang.upper(), len(tbl))

    return tables


# ──────────────────────────────────────────────────────────────────────
# EXCEL STYLING (v3.0 - Simplified)
# ──────────────────────────────────────────────────────────────────────
# Style definitions
STYLES = {
    "Faction": {
        "fill": PatternFill("solid", fgColor="4472C4"),  # Dark blue
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "row_height": 35,
    },
    "FactionNode_Main": {
        "fill": PatternFill("solid", fgColor="5B9BD5"),  # Medium blue
        "font": Font(bold=True, size=11, color="FFFFFF"),
        "row_height": 30,
    },
    "FactionNode_Sub": {
        "fill": PatternFill("solid", fgColor="B4C6E7"),  # Light blue
        "font": Font(bold=True),
        "row_height": None,
    },
    "FactionNode": {
        "fill": PatternFill("solid", fgColor="D6DCE4"),  # Very light blue
        "font": Font(),
        "row_height": None,
    },
    "Description": {
        "fill": PatternFill("solid", fgColor="F5F5F5"),  # Light gray
        "font": Font(italic=False),
        "row_height": None,
    },
    "ShopGroup": {
        "fill": PatternFill("solid", fgColor="00B0B0"),  # Teal
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "row_height": 35,
    },
    "ShopStage": {
        "fill": PatternFill("solid", fgColor="E0F7FA"),  # Light cyan
        "font": Font(bold=True),
        "row_height": None,
    },
}

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_default_font = Font()
_default_fill = PatternFill("solid", fgColor="FFFFFF")

_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_style(style_type: str) -> Tuple[PatternFill, Font, Optional[float]]:
    """Get style for a row type."""
    style = STYLES.get(style_type, {})
    return (
        style.get("fill", _default_fill),
        style.get("font", _default_font),
        style.get("row_height"),
    )


def apply_cell_style(cell, fill: PatternFill, indent: int, font: Font) -> None:
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
    cell.border = _border


# ──────────────────────────────────────────────────────────────────────
# HELPER: Get translated tab name
# ──────────────────────────────────────────────────────────────────────
def get_translated_tab_name(
    korean_name: str,
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    fallback: str = "Sheet"
) -> str:
    """Get translated sheet tab name."""
    normalized = normalize_placeholders(korean_name)
    is_eng = lang_code.lower() == "eng"
    
    if is_eng:
        trans, _ = eng_tbl.get(normalized, ("", ""))
        title = trans.strip() if trans else korean_name.strip()
    else:
        if lang_tbl:
            trans, _ = lang_tbl.get(normalized, ("", ""))
            if trans and is_good_translation(trans):
                title = trans.strip()
            else:
                trans_eng, _ = eng_tbl.get(normalized, ("", ""))
                title = trans_eng.strip() if trans_eng else korean_name.strip()
        else:
            trans_eng, _ = eng_tbl.get(normalized, ("", ""))
            title = trans_eng.strip() if trans_eng else korean_name.strip()
    
    if not title:
        title = fallback
    
    # Sanitize for Excel
    title = title[:31]
    title = re.sub(r"[\\/*?:\[\]]", "_", title)
    
    return title


# ──────────────────────────────────────────────────────────────────────
# EXCEL WRITER (v3.0)
# ──────────────────────────────────────────────────────────────────────
def write_sheet_content(
    sheet,
    rows: List[RowItem],
    is_eng: bool,
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str
) -> None:
    """Write rows to a sheet with proper formatting."""
    
    # Headers
    headers = []
    headers.append(sheet.cell(1, 1, "Original (KR)"))
    headers.append(sheet.cell(1, 2, "English (ENG)"))
    if not is_eng:
        headers.append(sheet.cell(1, 3, f"Translation ({lang_code.upper()})"))
    
    extra_headers = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    start_col = len(headers) + 1
    for idx, name in enumerate(extra_headers, start=start_col):
        headers.append(sheet.cell(1, idx, name))
    
    for c in headers:
        c.font = _header_font
        c.fill = _header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border
    sheet.row_dimensions[1].height = 25
    
    # Data rows
    for r_idx, (depth, text, style_type, is_desc) in enumerate(rows, start=2):
        fill, font, row_height = get_style(style_type)
        normalized = normalize_placeholders(text)
        
        trans_eng, sid_eng = eng_tbl.get(normalized, ("", ""))
        trans_other = sid_other = ""
        if not is_eng and lang_tbl:
            trans_other, sid_other = lang_tbl.get(normalized, ("", ""))
        
        # Original
        c_orig = sheet.cell(r_idx, 1, text)
        # English
        c_eng = sheet.cell(r_idx, 2, trans_eng)
        # Target language (if not English)
        c_other = None
        if not is_eng:
            c_other = sheet.cell(r_idx, 3, trans_other)
        
        # Extra columns
        col_offset = 3 if not is_eng else 2
        c_status = sheet.cell(r_idx, col_offset + 1, "")
        c_comment = sheet.cell(r_idx, col_offset + 2, "")
        c_stringid = sheet.cell(r_idx, col_offset + 3, sid_other if not is_eng else sid_eng)
        c_screenshot = sheet.cell(r_idx, col_offset + 4, "")
        
        # Apply styles
        apply_cell_style(c_orig, fill, indent=depth, font=font)
        apply_cell_style(c_eng, fill, indent=depth, font=font)
        if c_other:
            apply_cell_style(c_other, fill, indent=depth, font=font)
        
        apply_cell_style(c_status, fill, indent=0, font=_default_font)
        c_status.alignment = Alignment(horizontal="center", vertical="center")
        apply_cell_style(c_comment, fill, indent=0, font=_default_font)
        apply_cell_style(c_stringid, fill, indent=0, font=Font(bold=True))
        apply_cell_style(c_screenshot, fill, indent=0, font=_default_font)
        
        if row_height:
            sheet.row_dimensions[r_idx].height = row_height
    
    # Column widths
    widths = [40, 80] + ([] if is_eng else [80]) + [15, 70, 25, 25]
    for idx, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w
    
    # Hide English column for non-English workbooks
    if not is_eng:
        sheet.column_dimensions["B"].hidden = True
    
    # Status dropdown
    status_col = 3 if is_eng else 4
    dv = DataValidation(
        type="list",
        formula1='"ISSUE,NO ISSUE,BLOCKED"',
        allow_blank=True,
    )
    col_letter = get_column_letter(status_col)
    dv.add(f"{col_letter}2:{col_letter}{sheet.max_row}")
    sheet.add_data_validation(dv)


def write_workbook(
    faction_groups: List[FactionGroupData],
    standalone_factions: List[FactionData],
    shop_groups: List[ShopGroup],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """Generate Excel workbook for a language."""
    wb = Workbook()
    wb.remove(wb.active)
    
    is_eng = lang_code.lower() == "eng"
    used_titles: Set[str] = set()
    
    def get_unique_title(base: str) -> str:
        title = base
        counter = 1
        while title in used_titles:
            title = f"{base[:28]}_{counter}"
            counter += 1
        used_titles.add(title)
        return title
    
    # ─────────────────────────────────────────────────────────────────
    # FactionGroup sheets
    # ─────────────────────────────────────────────────────────────────
    for fg in faction_groups:
        rows = emit_faction_group_rows(fg)
        if not rows:
            continue
        
        title = get_translated_tab_name(fg.group_name, eng_tbl, lang_tbl, lang_code, "Group")
        title = get_unique_title(title)
        
        sheet = wb.create_sheet(title=title)
        write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code)
        
        log.info("    Sheet '%s': %d rows", title, len(rows))
    
    # ─────────────────────────────────────────────────────────────────
    # Standalone Faction sheet (factions not in any FactionGroup)
    # ─────────────────────────────────────────────────────────────────
    if standalone_factions:
        rows = emit_standalone_faction_rows(standalone_factions)
        if rows:
            title = get_translated_tab_name(EMPTY_FACTION_NAME, eng_tbl, lang_tbl, lang_code, "No Faction")
            title = get_unique_title(title)
            
            sheet = wb.create_sheet(title=title)
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code)
            
            log.info("    Sheet '%s': %d rows", title, len(rows))
    
    # ─────────────────────────────────────────────────────────────────
    # Shop sheet
    # ─────────────────────────────────────────────────────────────────
    if shop_groups:
        rows = emit_shop_rows(shop_groups)
        if rows:
            sheet = wb.create_sheet(title="Shop")
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code)
            log.info("    Sheet 'Shop': %d rows", len(rows))
    
    # ─────────────────────────────────────────────────────────────────
    # Save
    # ─────────────────────────────────────────────────────────────────
    if wb.worksheets:
        wb.save(out_path)
        log.info("  → Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))
    else:
        log.warning("  → Skipped: %s (no data)", out_path.name)


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 70)
    log.info("Region LQA Extractor v3.1")
    log.info("FactionGroup → Faction → FactionNode hierarchy")
    log.info("세력 없음 in FactionGroup stays in sheet; standalone → separate sheet")
    log.info("=" * 70)
    
    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)
    
    # Global duplicate tracking
    global_seen: Set[str] = set()
    
    # ─────────────────────────────────────────────────────────────────
    # Step 1: Knowledge lookup
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 1: Building Knowledge lookup…")
    knowledge_lookup = build_knowledge_name_lookup(RESOURCE_FOLDER)
    
    # ─────────────────────────────────────────────────────────────────
    # Step 2: Parse FactionGroup hierarchy
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 2: Parsing FactionGroup hierarchy from: %s", RESOURCE_FOLDER)
    faction_groups = parse_all_faction_groups(RESOURCE_FOLDER, knowledge_lookup, global_seen)
    
    # Collect FactionGroup StrKeys for standalone faction detection
    fg_strkeys = {fg.strkey for fg in faction_groups}
    
    # ─────────────────────────────────────────────────────────────────
    # Step 3: Parse Standalone Factions (not inside any FactionGroup)
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 3: Parsing Standalone Factions…")
    standalone_factions = parse_standalone_factions(RESOURCE_FOLDER, knowledge_lookup, global_seen, fg_strkeys)
    
    # ─────────────────────────────────────────────────────────────────
    # Step 4: Parse Shop data
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 4: Parsing Shop data…")
    shop_groups = parse_shop_file(SHOP_FILE, global_seen)
    
    # ─────────────────────────────────────────────────────────────────
    # Step 5: Load language tables
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 5: Loading language tables from: %s", LANGUAGE_FOLDER)
    lang_tables = load_language_tables(LANGUAGE_FOLDER)
    
    if not lang_tables:
        sys.exit("ERROR: No language tables found!")
    
    eng_tbl = lang_tables.get("eng", {})
    
    # ─────────────────────────────────────────────────────────────────
    # Step 6: Generate workbooks
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("Step 6: Generating Excel workbooks…")
    
    # Summary before generation
    log.info("")
    log.info("Data summary:")
    for fg in faction_groups:
        node_count = sum(count_nodes_recursive(f.nodes) for f in fg.factions)
        empty_in_fg = sum(1 for f in fg.factions if f.name == EMPTY_FACTION_NAME)
        log.info("  • %s: %d factions (%d 세력 없음), %d nodes", 
                 fg.group_name, len(fg.factions), empty_in_fg, node_count)
    
    standalone_node_count = sum(count_nodes_recursive(f.nodes) for f in standalone_factions)
    log.info("  • Standalone Factions: %d factions, %d nodes", len(standalone_factions), standalone_node_count)
    log.info("  • Shop: %d groups", len(shop_groups))
    log.info("")
    
    total = len(lang_tables)
    for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
        log.info("(%d/%d) Generating: %s", idx, total, code.upper())
        out_path = OUTPUT_FOLDER / f"Region_LQA_{code.upper()}.xlsx"
        
        if code.lower() == "eng":
            write_workbook(faction_groups, standalone_factions, shop_groups, eng_tbl, None, code, out_path)
        else:
            write_workbook(faction_groups, standalone_factions, shop_groups, eng_tbl, tbl, code, out_path)
    
    # ─────────────────────────────────────────────────────────────────
    # Summary
    # ─────────────────────────────────────────────────────────────────
    log.info("")
    log.info("=" * 70)
    log.info("DONE – %d workbook(s) generated", total)
    log.info("=" * 70)
    log.info("")
    log.info("Final statistics:")
    log.info("  • Unique StrKeys processed: %d", len(global_seen))
    log.info("  • FactionGroups: %d", len(faction_groups))
    log.info("  • Standalone Factions: %d", len(standalone_factions))
    log.info("  • Shop groups: %d", len(shop_groups))
    log.info("=" * 70)


if __name__ == "__main__":
    main()
