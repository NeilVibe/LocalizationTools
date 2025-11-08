print("- XSL Transfer ver. 0225 by Neil -")
print("- XSL Transfer ver. 0225 by Neil -")
print("- XSL Transfer ver. 0225 by Neil -")
print("- XSL Transfer ver. 0225 by Neil -")
print("Launching...")
import re
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog, scrolledtext, Tk, ttk
import threading
import pandas as pd
import numpy as np
print("Halfway there... Getting other packages ready...")
from sentence_transformers import SentenceTransformer
import faiss
import pickle
print("Almost done...!")
import torch
import traceback
from traceback import print_exc
import sys
import os
import random
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
from openpyxl.writer.excel import save_workbook
import shutil
import tempfile

try:
    import pyi_splash
    splash_active = True
except ImportError:
    splash_active = False

def on_closing():
    root.destroy()
    sys.exit()

stop_processing = False
model = SentenceTransformer('snunlp/KR-SBERT-V40K-klueNLI-augSTS')

def clean_text(text):
    if text is None:
        return None
    # Convert to string if the input is not already a string
    if not isinstance(text, str):
        text = str(text)
    return text.replace('_x000D_', '')



def excel_column_to_index(column_letter):
    return ord(column_letter.upper()) - ord('A')

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f'{width}x{height}+{x}+{y}')

def create_upload_settings_gui(excel_file_paths, operation_type):
    if not excel_file_paths:
        messagebox.showerror("Error", "No files selected.")
        return

    settings_window = tk.Toplevel()
    settings_window.title("Upload Settings")
    center_window(settings_window, 800, 600)

    main_frame = ttk.Frame(settings_window)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    canvas = tk.Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    main_frame.pack(fill=tk.BOTH, expand=True)
    canvas.pack(side="left", fill=tk.BOTH, expand=True)
    scrollbar.pack(side="right", fill="y")

    all_file_selections = {}

    for file_path in excel_file_paths:
        file_frame = ttk.LabelFrame(scrollable_frame, text=os.path.basename(file_path))
        file_frame.pack(fill=tk.X, padx=5, pady=5, anchor="w")

        # Create a temporary file
        temp_file_path = f"{os.path.splitext(file_path)[0]}_temp.xlsx"
        shutil.copy2(file_path, temp_file_path)

        # Use openpyxl to get sheet names without fully loading the file
        wb = openpyxl.load_workbook(temp_file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # Remove the temporary file
        os.remove(temp_file_path)

        tab_vars = []
        kr_entries = []
        trans_entries = []

        for i, sheet in enumerate(sheet_names):
            frame = ttk.Frame(file_frame)
            frame.pack(fill=tk.X, padx=5, pady=2, anchor="w")

            var = tk.BooleanVar()
            tab_vars.append(var)
            
            def toggle_entries(index, var, kr_entry, trans_entry):
                kr_entry.config(state='normal' if var.get() else 'disabled')
                trans_entry.config(state='normal' if var.get() else 'disabled')
            
            kr_entry = ttk.Entry(frame, width=5, state='disabled')
            trans_entry = ttk.Entry(frame, width=5, state='disabled')
            
            cb = ttk.Checkbutton(frame, text=sheet, variable=var, 
                                 command=lambda i=i, v=var, k=kr_entry, t=trans_entry: toggle_entries(i, v, k, t))
            cb.pack(side=tk.LEFT)

            ttk.Label(frame, text="KR Column:").pack(side=tk.LEFT, padx=(10, 0))
            kr_entry.pack(side=tk.LEFT, padx=(0, 10))
            kr_entries.append(kr_entry)

            ttk.Label(frame, text="Translation Column:").pack(side=tk.LEFT)
            trans_entry.pack(side=tk.LEFT)
            trans_entries.append(trans_entry)

        all_file_selections[file_path] = {
            'tab_vars': tab_vars,
            'kr_entries': kr_entries,
            'trans_entries': trans_entries,
            'sheet_names': sheet_names
        }

    def validate_inputs():
        for file_path, file_data in all_file_selections.items():
            selected_tabs = [i for i, var in enumerate(file_data['tab_vars']) if var.get()]
            if not selected_tabs:
                messagebox.showerror("Error", f"Please select at least one tab in {os.path.basename(file_path)}.")
                return False

            for i in selected_tabs:
                if not file_data['kr_entries'][i].get() or not file_data['trans_entries'][i].get():
                    messagebox.showerror("Error", f"Please enter both column letters for {file_data['sheet_names'][i]} in {os.path.basename(file_path)}.")
                    return False
                if not file_data['kr_entries'][i].get().isalpha() or not file_data['trans_entries'][i].get().isalpha():
                    messagebox.showerror("Error", f"Invalid column letter for {file_data['sheet_names'][i]} in {os.path.basename(file_path)}.")
                    return False

        return True

    def on_ok():
        if validate_inputs():
            selections = {}
            for file_path, file_data in all_file_selections.items():
                file_selections = {
                    file_data['sheet_names'][i]: {
                        'kr_column': file_data['kr_entries'][i].get().upper(),
                        'trans_column': file_data['trans_entries'][i].get().upper()
                    }
                    for i in range(len(file_data['sheet_names']))
                    if file_data['tab_vars'][i].get()
                }
                selections[file_path] = file_selections
            
            settings_window.destroy()
            if operation_type == "create_dictionary":
                process_excel_and_generate_embeddings(selections)
            elif operation_type == "translate_excel":
                translate_excel_to_excel(selections)
            elif operation_type == "check_newlines":
                check_newlines(selections)        
            elif operation_type == "combine_excel":
                combine_excel_files(selections)               
            elif operation_type == "newline_auto_adapt":
                newline_auto_adapt(selections)

    ok_button = ttk.Button(settings_window, text="OK", command=on_ok)
    ok_button.pack(side=tk.RIGHT, padx=10, pady=10)

    cancel_button = ttk.Button(settings_window, text="Cancel", command=settings_window.destroy)
    cancel_button.pack(side=tk.RIGHT, padx=10, pady=10)

    settings_window.mainloop()

def process_excel_and_generate_embeddings(selections):
    global stop_processing
    stop_processing = False

    def safe_most_frequent(x):
        if x.empty or x.isna().all():
            return pd.NA
        return x.value_counts().index[0]

    def run_embedding_process(selections):
        all_kr_texts_split = []
        all_fr_texts_split = []
        all_kr_texts_whole = []
        all_fr_texts_whole = []

        try:
            print("Starting concatenation of tabs...")
            total_rows = 0
            for excel_file_path, file_selections in selections.items():
                for sheet_name, columns in file_selections.items():
                    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None)
                    col_kr_letter = columns['kr_column']
                    col_fr_letter = columns['trans_column']
                    col_kr_index = excel_column_to_index(col_kr_letter)
                    col_fr_index = excel_column_to_index(col_fr_letter)

                    # Drop rows where either KR or FR is NaN
                    df = df.iloc[:, [col_kr_index, col_fr_index]].dropna()

                    kr_texts = df.iloc[:, 0].apply(clean_text).tolist()
                    fr_texts = df.iloc[:, 1].apply(clean_text).tolist()
                    total_rows += len(kr_texts)

                    for kr, fr in zip(kr_texts, fr_texts):
                        kr_lines = kr.split('\n')
                        fr_lines = fr.split('\n')
                        if len(kr_lines) == len(fr_lines):
                            all_kr_texts_split.extend(kr_lines)
                            all_fr_texts_split.extend(fr_lines)
                        else:
                            all_kr_texts_whole.append(kr)
                            all_fr_texts_whole.append(fr)

                    print(f"Tab '{sheet_name}' in file '{os.path.basename(excel_file_path)}' processed: {len(kr_texts)} rows")

            print(f"Total number of files processed: {len(selections)}")
            print(f"Total number of rows concatenated: {total_rows}")

            # Process split texts
            split_translation_pairs = pd.DataFrame({'KR': all_kr_texts_split, 'FR': all_fr_texts_split})
            split_most_freq_trans = split_translation_pairs.groupby('KR')['FR'].agg(safe_most_frequent).reset_index()
            split_most_freq_trans = split_most_freq_trans.dropna()
            unique_kr_texts_split = split_most_freq_trans['KR'].tolist()

            # Process whole texts
            whole_translation_pairs = pd.DataFrame({'KR': all_kr_texts_whole, 'FR': all_fr_texts_whole})
            whole_most_freq_trans = whole_translation_pairs.groupby('KR')['FR'].agg(safe_most_frequent).reset_index()
            whole_most_freq_trans = whole_most_freq_trans.dropna()
            unique_kr_texts_whole = whole_most_freq_trans['KR'].tolist()

            print(f"Total unique split texts: {len(unique_kr_texts_split)}")
            print(f"Total unique whole texts: {len(unique_kr_texts_whole)}")

            # Generate embeddings
            split_embeddings = []
            for index, text in enumerate(unique_kr_texts_split):
                if stop_processing:
                    print("Embedding process stopped.")
                    return
                batch_embeddings = model.encode([text], convert_to_tensor=False)
                split_embeddings.extend(batch_embeddings)
                if (index + 1) % 100 == 0 or (index + 1) == len(unique_kr_texts_split):
                    print(f"Split embedding progress: {index + 1}/{len(unique_kr_texts_split)} lines")

            whole_embeddings = []
            for index, text in enumerate(unique_kr_texts_whole):
                if stop_processing:
                    print("Embedding process stopped.")
                    return
                batch_embeddings = model.encode([text], convert_to_tensor=False)
                whole_embeddings.extend(batch_embeddings)
                if (index + 1) % 10 == 0 or (index + 1) == len(unique_kr_texts_whole):
                    print(f"Whole embedding progress: {index + 1}/{len(unique_kr_texts_whole)} blocks")

            # Save embeddings and dictionaries
            split_embeddings = np.array(split_embeddings)
            np.save('SplitExcelEmbeddings.npy', split_embeddings)
            split_translation_dict = dict(zip(split_most_freq_trans['KR'], split_most_freq_trans['FR']))
            with open('SplitExcelDictionary.pkl', 'wb') as f:
                pickle.dump(split_translation_dict, f)

            if unique_kr_texts_whole:
                whole_embeddings = np.array(whole_embeddings)
                np.save('WholeExcelEmbeddings.npy', whole_embeddings)
                whole_translation_dict = dict(zip(whole_most_freq_trans['KR'], whole_most_freq_trans['FR']))
                with open('WholeExcelDictionary.pkl', 'wb') as f:
                    pickle.dump(whole_translation_dict, f)
            else:
                print("No whole embeddings found. Skipping whole mode.")

            messagebox.showinfo("완료", "Embeddings and translation dictionaries have been generated and saved.")
            print("Embeddings and translation dictionaries have been generated and saved!")
            print(f"Total unique split texts processed: {len(unique_kr_texts_split)}")
            print(f"Total unique whole texts processed: {len(unique_kr_texts_whole)}")

        except Exception as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    # Start the embedding process in a separate thread
    embedding_thread = threading.Thread(target=run_embedding_process, args=(selections,))
    embedding_thread.start()

def load_embeddings_and_create_index():
    global ref_kr_embeddings_split, translation_dict_split, index_split, ref_kr_sentences_split
    global ref_kr_embeddings_whole, translation_dict_whole, index_whole, ref_kr_sentences_whole
    global split_mode_available, whole_mode_available

    split_mode_available = False
    whole_mode_available = False

    try:
        # Try to load split mode
        if os.path.exists('SplitExcelDictionary.pkl') and os.path.exists('SplitExcelEmbeddings.npy'):
            with open('SplitExcelDictionary.pkl', 'rb') as file:
                translation_dict_split = pickle.load(file)
            ref_kr_embeddings_split = np.load('SplitExcelEmbeddings.npy')
            ref_kr_sentences_split = pd.Series(list(translation_dict_split.keys()), dtype=str)
            faiss.normalize_L2(ref_kr_embeddings_split)
            index_split = faiss.IndexFlatIP(ref_kr_embeddings_split.shape[1])
            index_split.add(ref_kr_embeddings_split)
            split_mode_available = True
            print("Split mode loaded successfully.")

        # Try to load whole mode
        if os.path.exists('WholeExcelDictionary.pkl') and os.path.exists('WholeExcelEmbeddings.npy'):
            with open('WholeExcelDictionary.pkl', 'rb') as file:
                translation_dict_whole = pickle.load(file)
            ref_kr_embeddings_whole = np.load('WholeExcelEmbeddings.npy')
            ref_kr_sentences_whole = pd.Series(list(translation_dict_whole.keys()), dtype=str)
            faiss.normalize_L2(ref_kr_embeddings_whole)
            index_whole = faiss.IndexFlatIP(ref_kr_embeddings_whole.shape[1])
            index_whole.add(ref_kr_embeddings_whole)
            whole_mode_available = True
            print("Whole mode loaded successfully.")

        if split_mode_available or whole_mode_available:
            translate_file_button.config(state=tk.NORMAL)
            excel_to_excel_button.config(state=tk.NORMAL)
            load_embeddings_button.config(bg="green")
            messagebox.showinfo("Success", f"Loaded modes: {'Split' if split_mode_available else ''} {'Whole' if whole_mode_available else ''}")
        else:
            messagebox.showwarning("Warning", "No embedding modes were loaded. Please create a dictionary first.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while loading the embeddings or dictionary: {str(e)}")


def stop_processing_function():
    global stop_processing
    stop_processing = True
    messagebox.showinfo("Process Stopped", "Translation process has been stopped.")



def translate_file():
    global whole_mode_available, split_mode_available

    def convert_newlines(text):
        return text.replace('\n', '\\n')

    def analyze_code_patterns(text):
        # Dictionary to store all code pattern relationships
        patterns = {
            'start_codes': set(),
            'next_levels': {}  # Stores code relationships for any level depth
        }
        
        current_pos = 0
        
        while current_pos < len(text):
            if text[current_pos] == '{' or text[current_pos:].startswith("<PAColor"):
                # Process consecutive code blocks
                code_positions = []  # Store positions of all code blocks
                
                while current_pos < len(text):
                    if text[current_pos] == '{' or text[current_pos:].startswith("<PAColor"):
                        code_start = current_pos
                        if text[current_pos:].startswith("<PAColor"):
                            code_end = text.find(">", current_pos) + 1
                        else:
                            code_end = text.find("}", current_pos) + 1
                        
                        if code_end > code_start:
                            code = text[code_start:code_end].split('(')[0]
                            code_positions.append((code, code_start, code_end))
                            current_pos = code_end
                        else:
                            current_pos += 1
                    else:
                        # Just move past any punctuation or other characters
                        current_pos += 1
                        break
                
                # Add codes to patterns
                if code_positions:
                    first_code, _, _ = code_positions[0]
                    patterns['start_codes'].add(first_code)
                    
                    # Process relationships between codes
                    for i in range(len(code_positions) - 1):
                        current_code, _, _ = code_positions[i]
                        next_code, _, _ = code_positions[i+1]
                        
                        key = tuple(cp[0] for cp in code_positions[:i+1])
                        
                        if key not in patterns['next_levels']:
                            patterns['next_levels'][key] = set()
                        patterns['next_levels'][key].add(next_code)
            else:
                current_pos += 1
        
        return patterns

    def simple_number_replace(original, translated):
        if not isinstance(original, str):
            return translated

        # Handle text + code(s) case
        first_code_start = original.find("{")
        if first_code_start > 0 or original.startswith("<PAColor"):
            codes = []
            current_pos = first_code_start if first_code_start > 0 else 0
            
            while current_pos < len(original):
                if original[current_pos:].startswith("<PAColor"):
                    end_pos = original.find(">", current_pos)
                    codes.append(original[current_pos:end_pos+1])
                    current_pos = end_pos + 1
                elif original[current_pos] == '{':
                    end_pos = original.find("}", current_pos)
                    codes.append(original[current_pos:end_pos+1])
                    current_pos = end_pos + 1
                else:
                    break
                            
            if codes:
                return ''.join(codes) + translated

        # Extract only code blocks at the beginning (without punctuation)
        prefix = ""
        pos = 0
        
        while pos < len(original):
            if original[pos:].startswith("<PAColor") or original[pos] == '{':
                if original[pos:].startswith("<PAColor"):
                    end_pos = original.find(">", pos) + 1
                else:
                    end_pos = original.find("}", pos) + 1
                if end_pos > pos:
                    prefix += original[pos:end_pos]
                    pos = end_pos
                else:
                    break
            else:
                # Stop as soon as we encounter anything that's not a code block
                break
        
        if prefix:
            result = prefix + translated
            if original.endswith("<PAOldColor>"):
                result += "<PAOldColor>"
            return result

        # Handle PAOldColor ending
        if original.endswith("<PAOldColor>"):
            translated += "<PAOldColor>"

        return translated

    def clean_audiovoice_tags(text):
        text = re.sub(r'\{[^}]*\}', '', text)  # Remove any code that starts with {
        text = re.sub(r'<PAColor[^>]*>', '', text)  # Remove color code tags
        text = re.sub(r'<PAOldColor>', '', text)  # Remove <PAOldColor> tag
        return text

    def process_batch(batch, mode='whole'):
        main_kr_batch = batch.tolist()
        results = []

        if mode == 'whole' and not whole_mode_available:
            mode = 'split'
        
        if mode == 'split' and not split_mode_available:
            return [(sentence, "") for sentence in main_kr_batch]

        if mode == 'whole':
            index = index_whole
            ref_kr_sentences = ref_kr_sentences_whole
            translation_dict = translation_dict_whole
        else:
            index = index_split
            ref_kr_sentences = ref_kr_sentences_split
            translation_dict = translation_dict_split

        for sentence in main_kr_batch:
            if not isinstance(sentence, str):
                results.append((sentence, ""))
                continue
            clean_sentence = clean_audiovoice_tags(sentence)
            try:
                sentence_embeddings = model.encode([clean_sentence])
                faiss.normalize_L2(sentence_embeddings)
                distances, indices = index.search(sentence_embeddings, 1)
                best_match_index = indices[0][0]
                distance_score = distances[0][0]
                if best_match_index >= len(ref_kr_sentences):
                    results.append((sentence, ""))
                    continue
                best_match_korean = ref_kr_sentences.iloc[best_match_index]
                best_match_translation = translation_dict.get(best_match_korean, "")
                faiss_threshold = float(faiss_threshold_entry.get())
                if distance_score >= faiss_threshold:
                    results.append((sentence, best_match_translation))
                else:
                    results.append((sentence, ""))
            except Exception as e:
                results.append((sentence, ""))

        return results

    def translation_task():
        global stop_processing
        stop_processing = False

        def trim_columns(path_in):
            columns_to_keep = list(range(9))
            trimmed_lines = []
            with open(path_in, 'r', encoding='utf-8') as f_in:
                for line in f_in:
                    columns = line.rstrip('\n').split('\t')
                    if len(columns) < 9:
                        columns += [''] * (9 - len(columns))
                    new_line = '\t'.join([columns[i] for i in columns_to_keep])
                    trimmed_lines.append(new_line)
            return trimmed_lines

        try:
            if not whole_mode_available and not split_mode_available:
                messagebox.showerror("Error", "No translation index available. Please load the dictionary first.")
                return

            trimmed_lines = trim_columns(file_path)
            
            total_sentences = len(trimmed_lines)
            print(f"Total sentences: {total_sentences}")

            updated_data_rows = []

            for idx, line in enumerate(trimmed_lines):
                if stop_processing:
                    print("File processing stopped.")
                    return

                row = line.split('\t')
                korean_text = row[5]
                
                if not korean_text.strip():
                    continue

                lines = korean_text.split("\\n")
                final_reconstruction = []
                translation_inserted = False

                if whole_mode_available:
                    results = process_batch(pd.Series([korean_text]), mode='whole')
                    whole_translated_text = results[0][1]
                else:
                    whole_translated_text = ""

                if whole_translated_text:
                    whole_translated_text = convert_newlines(whole_translated_text)
                    final_reconstruction.append(whole_translated_text)
                    translation_inserted = True
                else:
                    for line in lines:
                        if stop_processing:
                            print("File processing stopped.")
                            return

                        if line.strip() == "":
                            final_reconstruction.append("")
                            continue

                        results = process_batch(pd.Series([line]), mode='split')

                        for original_sentence, translated_sentence in results:
                            if translated_sentence:
                                final_translation = simple_number_replace(original_sentence, translated_sentence)
                                final_reconstruction.append(final_translation)
                                translation_inserted = True
                            else:
                                final_reconstruction.append(line)

                if translation_inserted:
                    row[6] = "\\n".join(final_reconstruction)
                    updated_data_rows.append(row)

                print(f"\rProcessed {idx + 1}/{total_sentences} rows", end="")

            print("\nWriting updated rows to the output file...")
            if updated_data_rows:
                with open(output_file_path, 'w', newline='', encoding='utf-8') as f:
                    for row in updated_data_rows:
                        f.write('\t'.join(row) + '\n')
                messagebox.showinfo("번역 전환 작업 완료했습니다!", f"Translation completed with insertion! {len(updated_data_rows)} rows modified.", icon=messagebox.INFO)
                print(f"\nTranslation completed with insertion! {len(updated_data_rows)} rows modified.")
            else:
                print("No translations inserted. Output file not created.")

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    file_path = filedialog.askopenfilename()
    if not file_path:
        messagebox.showerror("Error", "No file selected.")
        return
    output_file_path = f"{os.path.splitext(file_path)[0]}_translated.txt"
    if not output_file_path:
        messagebox.showerror("Error", "No output file selected.")
        return

    translation_thread = threading.Thread(target=translation_task)
    translation_thread.start()


def convert_cell_value(value):
    if isinstance(value, str):
        # Try to convert string numbers to actual numbers
        try:
            # Check if it's a pure number string
            float(value)
            return float(value)
        except ValueError:
            # If conversion fails, it's not a pure number, keep as string
            return value
    return value



def translate_excel_to_excel(selections):
    global stop_processing
    
    def process_batch(batch, mode='whole'):
        main_kr_batch = batch.tolist()
        results = []

        if mode == 'whole' and not whole_mode_available:
            mode = 'split'
        
        if mode == 'split' and not split_mode_available:
            return [(sentence, "") for sentence in main_kr_batch]

        if mode == 'whole':
            index = index_whole
            ref_kr_sentences = ref_kr_sentences_whole
            translation_dict = translation_dict_whole
        else:
            index = index_split
            ref_kr_sentences = ref_kr_sentences_split
            translation_dict = translation_dict_split

        for sentence in main_kr_batch:
            cleaned_sentence = clean_text(sentence)
            if not isinstance(cleaned_sentence, str):
                results.append((cleaned_sentence, ""))
                continue
            try:
                sentence_embeddings = model.encode([cleaned_sentence])
                faiss.normalize_L2(sentence_embeddings)
                distances, indices = index.search(sentence_embeddings, 1)
                best_match_index = indices[0][0]
                distance_score = distances[0][0]
                if best_match_index >= len(ref_kr_sentences):
                    results.append((cleaned_sentence, ""))
                    continue
                best_match_korean = ref_kr_sentences.iloc[best_match_index]
                best_match_translation = translation_dict.get(best_match_korean, "")
                faiss_threshold = float(faiss_threshold_entry.get())
                if distance_score >= faiss_threshold:
                    results.append((best_match_korean, best_match_translation))
                else:
                    results.append((cleaned_sentence, ""))
            except Exception as e:
                results.append((cleaned_sentence, ""))

        return results

    def translation_task():
        global stop_processing
        stop_processing = False

        try:
            for excel_file_path, file_selections in selections.items():
                temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
                shutil.copy2(excel_file_path, temp_file_path)
                
                temp_wb = openpyxl.load_workbook(temp_file_path, data_only=False)
                
                for sheet_name, columns in file_selections.items():
                    if stop_processing:
                        messagebox.showinfo("Process Stopped", "Translation process has been stopped.")
                        os.remove(temp_file_path)
                        return

                    sheet = temp_wb[sheet_name]
                    col_to_translate = openpyxl.utils.column_index_from_string(columns['kr_column'])
                    write_col_index = openpyxl.utils.column_index_from_string(columns['trans_column'])
                    print(f"Translating column {columns['kr_column']} in tab '{sheet_name}' and writing to column {columns['trans_column']}")

                    total_rows = sheet.max_row

                    for idx in range(1, total_rows + 1):
                        if stop_processing:
                            messagebox.showinfo("Process Stopped", "Translation process has been stopped.")
                            os.remove(temp_file_path)
                            return

                        cell = sheet.cell(row=idx, column=col_to_translate)
                        if not isinstance(cell.value, str):
                            continue

                        cleaned_input_text = clean_text(cell.value)

                        # First attempt whole match
                        whole_results = process_batch(pd.Series([cleaned_input_text]), mode='whole')
                        whole_translated_text = whole_results[0][1]

                        if whole_translated_text != "":
                            translated_text = whole_translated_text
                        else:
                            # Fallback to split match
                            lines = cleaned_input_text.split('\n')
                            translated_lines = []
                            for line in lines:
                                if not line.strip():
                                    continue
                                results = process_batch(pd.Series([line]), mode='split')
                                translated_line = results[0][1]
                                if translated_line.strip():
                                    translated_lines.append(translated_line)
                            translated_text = '\n'.join(translated_lines) if translated_lines else ''

                        # Clean the output text before writing
                        cleaned_translated_text = clean_text(translated_text)

                        # Only write to the cell if we have a translation
                        if cleaned_translated_text:
                            write_cell = sheet.cell(row=idx, column=write_col_index)
                            write_cell.value = cleaned_translated_text

                        print(f"Processed {idx}/{total_rows} rows in tab '{sheet_name}'")

                output_file_path = f"{os.path.splitext(excel_file_path)[0]}_translated.xlsx"
                temp_wb.save(output_file_path)
                temp_wb.close()
                os.remove(temp_file_path)
                print(f"Translation completed for {excel_file_path}! Output saved to {output_file_path}")

            if not stop_processing:
                messagebox.showinfo("완료", "Translation completed for all files!")
                print("Translation completed for all files!")

        except Exception as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)

    translation_thread = threading.Thread(target=translation_task)
    translation_thread.start()

    

def check_newlines(selections):
    flagged_rows = {}
    total_files = len(selections)
    processed_files = 0

    try:
        for excel_file_path, file_selections in selections.items():
            processed_files += 1
            print(f"Processing file {processed_files} of {total_files}: {os.path.basename(excel_file_path)}")
            
            # Create a temporary file
            temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
            shutil.copy2(excel_file_path, temp_file_path)
            
            # Load the temporary workbook
            wb = openpyxl.load_workbook(temp_file_path)
            file_name = os.path.basename(excel_file_path)
            flagged_rows[file_name] = {}

            total_sheets = len(file_selections)
            processed_sheets = 0

            for sheet_name, columns in file_selections.items():
                processed_sheets += 1
                print(f"  Processing sheet {processed_sheets} of {total_sheets}: {sheet_name}")
                
                sheet = wb[sheet_name]
                col1 = openpyxl.utils.column_index_from_string(columns['kr_column'])
                col2 = openpyxl.utils.column_index_from_string(columns['trans_column'])

                flagged_rows[file_name][sheet_name] = []

                total_rows = sheet.max_row
                for idx in range(1, total_rows + 1):
                    cell1 = sheet.cell(row=idx, column=col1).value
                    cell2 = sheet.cell(row=idx, column=col2).value

                    if cell1 is None or cell2 is None:
                        continue
                    
                    newlines1 = str(cell1).count('\n')
                    newlines2 = str(cell2).count('\n')
                    if newlines1 != newlines2:
                        flagged_rows[file_name][sheet_name].append(idx)

                    if idx % 100 == 0 or idx == total_rows:
                        print(f"    Processed {idx}/{total_rows} rows")

                print(f"    Rows flagged in {sheet_name}: {len(flagged_rows[file_name][sheet_name])}")

            # Close the workbook and remove the temporary file
            wb.close()
            os.remove(temp_file_path)

        if any(sheet for file in flagged_rows.values() for sheet in file.values()):
            print("Generating report...")
            report_path = filedialog.asksaveasfilename(defaultextension=".txt", title="Save Newline Comparison Report")
            if report_path:
                with open(report_path, 'w', encoding='utf-8') as f:
                    f.write("Newline Mismatch Report:\n\n")
                    for file_name, sheets in flagged_rows.items():
                        if any(sheets.values()):
                            f.write(f"File: {file_name}\n")
                            for sheet_name, rows in sheets.items():
                                if rows:
                                    f.write(f"  Tab: {sheet_name}\n")
                                    for row in rows:
                                        f.write(f"    Row: {row}\n")
                            f.write("\n")
                print(f"Report saved to: {report_path}")
                messagebox.showinfo("Report Generated", f"Newline comparison report saved to {report_path}")
            else:
                print("Report generation cancelled by user.")
        else:
            print("No rows with different number of newlines found.")
            messagebox.showinfo("Report", "No rows with different number of newlines found.")

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
        if 'temp_file_path' in locals():
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)  # Remove temp file if it exists




def combine_excel_files(selections):
    def combination_task():
        combined_wb = openpyxl.Workbook()
        combined_sheet = combined_wb.active
        combined_sheet.title = "Combined Data"
        row_index = 1

        try:
            total_files = len(selections)
            for file_index, (excel_file_path, file_selections) in enumerate(selections.items(), 1):
                print(f"\nProcessing file {file_index}/{total_files}: {os.path.basename(excel_file_path)}")
                
                # Create a temporary file
                temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
                shutil.copy2(excel_file_path, temp_file_path)
                
                # Load the temporary workbook
                source_wb = openpyxl.load_workbook(temp_file_path, data_only=False)
                
                total_sheets = len(file_selections)
                for sheet_index, (sheet_name, columns) in enumerate(file_selections.items(), 1):
                    print(f"  Processing sheet {sheet_index}/{total_sheets}: {sheet_name}")
                    sheet = source_wb[sheet_name]
                    selected_columns = [openpyxl.utils.column_index_from_string(col) for col in columns.values()]
                    
                    total_rows = sheet.max_row
                    for row_num in range(1, total_rows + 1):
                        if all(sheet.cell(row=row_num, column=col).value is None for col in selected_columns):
                            continue  # Skip empty rows
                        
                        for dest_col, src_col in enumerate(selected_columns, start=1):
                            src_cell = sheet.cell(row=row_num, column=src_col)
                            dest_cell = combined_sheet.cell(row=row_index, column=dest_col)
                            
                            # Clean the cell value before copying
                            cleaned_value = clean_text(src_cell.value)
                            dest_cell.value = cleaned_value
                            
                            # Copy cell format
                            if src_cell.has_style:
                                dest_cell.font = copy(src_cell.font)
                                dest_cell.border = copy(src_cell.border)
                                dest_cell.fill = copy(src_cell.fill)
                                dest_cell.number_format = copy(src_cell.number_format)
                                dest_cell.protection = copy(src_cell.protection)
                                dest_cell.alignment = copy(src_cell.alignment)
                        
                        row_index += 1
                        
                        if row_num % 100 == 0 or row_num == total_rows:
                            print(f"    Processed {row_num}/{total_rows} rows")
                
                # Close the workbook and remove the temporary file
                source_wb.close()
                os.remove(temp_file_path)

            # Automatically save the combined file
            base_name = os.path.splitext(os.path.basename(list(selections.keys())[0]))[0]
            save_path = f"{os.path.dirname(list(selections.keys())[0])}/{base_name}_combined.xlsx"
            combined_wb.save(save_path)
            print(f"\nCombined Excel file saved as {save_path}")
            messagebox.showinfo("Success", f"Combined Excel file saved as {save_path}")

        except Exception as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")
            if 'temp_file_path' in locals():
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)  # Remove temp file if it exists

    combination_thread = threading.Thread(target=combination_task)
    combination_thread.start()





def newline_auto_adapt(selections):
    def preprocess_text(text):
        if not isinstance(text, str):
            return text
        # Remove empty lines and trim whitespace
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        return '\n'.join(lines)

    def process_text(kr_text, fr_text):
        if not isinstance(kr_text, str) or not isinstance(fr_text, str):
            return fr_text
        
        # Clean texts before processing
        kr_text = clean_text(kr_text)
        fr_text = clean_text(fr_text)
        
        kr_lines = kr_text.split('\n')
        fr_lines = fr_text.split('\n')
        
        kr_newlines = len(kr_lines) - 1
        fr_newlines = len(fr_lines) - 1
        
        if kr_newlines == fr_newlines:
            return fr_text
        
        if kr_newlines == 0:
            # Remove all newlines from FR
            return fr_text.replace('\n', ' ').replace('  ', ' ')
        
        if kr_newlines > fr_newlines:
            # Need to add newlines to FR
            return add_newlines(fr_text, kr_newlines)
        
        if kr_newlines < fr_newlines:
            # Need to remove newlines from FR
            return remove_newlines(fr_text, kr_newlines)
        
        return fr_text

    def add_newlines(text, target_newlines):
        words = text.split()
        if not words:
            return text
        
        target_lines = target_newlines + 1
        avg_words_per_line = len(words) / target_lines
        
        punctuations = '.?!,'
        new_lines = []
        current_line = []
        
        for i, word in enumerate(words):
            current_line.append(word)
            
            if len(current_line) >= avg_words_per_line or i == len(words) - 1:
                # Check if we can break at a punctuation
                break_index = -1
                for j, w in enumerate(reversed(current_line)):
                    if any(p in w for p in punctuations):
                        break_index = len(current_line) - j - 1
                        break
                
                if break_index != -1:
                    new_lines.append(' '.join(current_line[:break_index+1]))
                    current_line = current_line[break_index+1:]
                else:
                    new_lines.append(' '.join(current_line))
                    current_line = []
        
        # Add any remaining words
        if current_line:
            new_lines[-1] += ' ' + ' '.join(current_line)
        
        # Adjust the number of lines if necessary
        while len(new_lines) > target_lines:
            shortest_line_index = min(range(len(new_lines)), key=lambda i: len(new_lines[i]))
            if shortest_line_index > 0:
                new_lines[shortest_line_index-1] += ' ' + new_lines.pop(shortest_line_index)
            else:
                new_lines[0] += ' ' + new_lines.pop(1)
        
        return '\n'.join(new_lines)

    def remove_newlines(text, target_newlines):
        # First, remove all newlines
        text = text.replace('\n', ' ')
        # Then, add newlines as needed
        return add_newlines(text, target_newlines)

    def adaptation_task():
        try:
            total_files = len(selections)
            for file_index, (excel_file_path, file_selections) in enumerate(selections.items(), 1):
                print(f"\nProcessing file {file_index}/{total_files}: {os.path.basename(excel_file_path)}")
                
                # Create a temporary file
                temp_file_path = f"{os.path.splitext(excel_file_path)[0]}_temp.xlsx"
                shutil.copy2(excel_file_path, temp_file_path)
                
                # Load the temporary workbook
                source_wb = openpyxl.load_workbook(temp_file_path, data_only=False)
                
                total_sheets = len(file_selections)
                for sheet_index, (sheet_name, columns) in enumerate(file_selections.items(), 1):
                    print(f"  Processing sheet {sheet_index}/{total_sheets}: {sheet_name}")
                    sheet = source_wb[sheet_name]
                    kr_col = openpyxl.utils.column_index_from_string(columns['kr_column'])
                    fr_col = openpyxl.utils.column_index_from_string(columns['trans_column'])
                    
                    total_rows = sheet.max_row
                    for idx in range(1, total_rows + 1):
                        kr_cell = sheet.cell(row=idx, column=kr_col)
                        fr_cell = sheet.cell(row=idx, column=fr_col)
                        
                        if kr_cell.value and fr_cell.value:
                            # Preprocess and clean both KR and FR text
                            kr_value = preprocess_text(clean_text(str(kr_cell.value)))
                            fr_value = preprocess_text(clean_text(str(fr_cell.value)))
                            
                            # Update the cells with preprocessed text
                            kr_cell.value = kr_value
                            fr_cell.value = fr_value
                            
                            # Apply newline adaptation only to FR text
                            new_fr_value = process_text(kr_value, fr_value)
                            
                            if new_fr_value != fr_value:
                                fr_cell.value = new_fr_value
                        
                        if idx % 100 == 0 or idx == total_rows:
                            print(f"    Processed {idx}/{total_rows} rows")
                
                # Save the output workbook
                output_file_path = f"{os.path.splitext(excel_file_path)[0]}_adapted.xlsx"
                source_wb.save(output_file_path)
                
                # Close the workbook and remove the temporary file
                source_wb.close()
                os.remove(temp_file_path)
                
                print(f"Newline adaptation completed for {excel_file_path}! Output saved to {output_file_path}")
            
            messagebox.showinfo("완료", "Newline adaptation completed for all files!")
            print("\nNewline adaptation completed for all files!")
        except Exception as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")
            if 'temp_file_path' in locals():
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)  # Remove temp file if it exists

    adaptation_thread = threading.Thread(target=adaptation_task)
    adaptation_thread.start()



def create_temp_excel(original_file, suffix):
    # Create temp file in system's temp directory with unique suffix
    temp_dir = tempfile.gettempdir()
    temp_name = f"xlstransfer_temp_{suffix}_{random.randint(1000,9999)}_{os.path.basename(original_file)}"
    temp_path = os.path.join(temp_dir, temp_name)
    shutil.copy2(original_file, temp_path)
    return temp_path

def simple_excel_transfer():
    temp_files = []  # Keep track of all temp files
    
    try:
        # Select source file and create temp immediately
        source_file = filedialog.askopenfilename(title="Select Source Excel File", filetypes=[("Excel files", "*.xlsx")])
        if not source_file:
            messagebox.showerror("Error", "No source file selected.")
            return
            
        # Create source temp file in temp directory
        temp_source_file = create_temp_excel(source_file, "source")
        temp_files.append(temp_source_file)
        
        # Try to open the temp file to verify access
        source_xls = pd.ExcelFile(temp_source_file)

        # Select destination file
        dest_file = filedialog.askopenfilename(title="Select Destination Excel File", filetypes=[("Excel files", "*.xlsx")])
        if not dest_file:
            for temp_file in temp_files:  # Clean up temp files
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            messagebox.showerror("Error", "No destination file selected.")
            return
            
        # Create destination temp file in temp directory
        temp_dest_file = create_temp_excel(dest_file, "dest")
        temp_files.append(temp_dest_file)
            
        # Try to open the temp file to verify access
        dest_xls = pd.ExcelFile(temp_dest_file)

        class TransferSettings:
            def __init__(self, parent_frame, source_sheet_names, dest_sheet_names, active_transfers_list):
                self.active_transfers_list = active_transfers_list
                self.active_transfers_list.append(self)
                
                self.frame = ttk.LabelFrame(parent_frame, text=f"Transfer {len(self.active_transfers_list)}")
                self.frame.pack(fill=tk.X, padx=5, pady=5, anchor="w")

                # Source settings
                source_frame = ttk.Frame(self.frame)
                source_frame.pack(fill=tk.X, padx=5, pady=2)

                self.source_tab_var = tk.StringVar(value=source_sheet_names[0])
                self.source_file_col_var = tk.StringVar()
                self.source_note_col_var = tk.StringVar()

                ttk.Label(source_frame, text="Source Sheet:").pack(side=tk.LEFT, padx=(5, 5))
                ttk.Combobox(source_frame, textvariable=self.source_tab_var, values=source_sheet_names, state='readonly', width=15).pack(side=tk.LEFT, padx=(5, 15))

                ttk.Label(source_frame, text="Source Column:").pack(side=tk.LEFT)
                ttk.Entry(source_frame, textvariable=self.source_file_col_var, width=5).pack(side=tk.LEFT, padx=(5, 15))

                ttk.Label(source_frame, text="Target Column:").pack(side=tk.LEFT)
                ttk.Entry(source_frame, textvariable=self.source_note_col_var, width=5).pack(side=tk.LEFT, padx=(5, 15))

                # Destination settings
                dest_frame = ttk.Frame(self.frame)
                dest_frame.pack(fill=tk.X, padx=5, pady=2)

                self.dest_tab_var = tk.StringVar(value=dest_sheet_names[0])
                self.dest_file_col_var = tk.StringVar()
                self.dest_note_col_var = tk.StringVar()

                ttk.Label(dest_frame, text="Dest Sheet:").pack(side=tk.LEFT, padx=(5, 5))
                ttk.Combobox(dest_frame, textvariable=self.dest_tab_var, values=dest_sheet_names, state='readonly', width=15).pack(side=tk.LEFT, padx=(5, 15))

                ttk.Label(dest_frame, text="Source Column:").pack(side=tk.LEFT)
                ttk.Entry(dest_frame, textvariable=self.dest_file_col_var, width=5).pack(side=tk.LEFT, padx=(5, 15))

                ttk.Label(dest_frame, text="Target Column:").pack(side=tk.LEFT)
                ttk.Entry(dest_frame, textvariable=self.dest_note_col_var, width=5).pack(side=tk.LEFT, padx=(5, 15))

                # Delete button
                self.delete_button = ttk.Button(self.frame, text="Delete", command=self.remove_frame)
                self.delete_button.pack(side=tk.RIGHT, padx=5, pady=2)

            def remove_frame(self):
                self.active_transfers_list.remove(self)
                self.frame.destroy()
                # Update the numbering of remaining transfers
                for i, transfer in enumerate(self.active_transfers_list, 1):
                    transfer.frame.configure(text=f"Transfer {i}")

            def get_settings(self):
                return {
                    'source_tab': self.source_tab_var.get(),
                    'source_file_col': self.source_file_col_var.get().upper(),
                    'source_note_col': self.source_note_col_var.get().upper(),
                    'dest_tab': self.dest_tab_var.get(),
                    'dest_file_col': self.dest_file_col_var.get().upper(),
                    'dest_note_col': self.dest_note_col_var.get().upper()
                }

        def create_transfer_settings_gui():
            settings_window = tk.Toplevel()
            settings_window.title("Multiple Transfer Settings")
            center_window(settings_window, 900, 700)

            def on_window_close():
                try:
                    source_xls.close()
                    dest_xls.close()
                except:
                    pass
                    
                for temp_file in temp_files:
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                    except:
                        pass
                settings_window.destroy()

            settings_window.protocol("WM_DELETE_WINDOW", on_window_close)

            main_frame = ttk.Frame(settings_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            main_frame.pack(fill=tk.BOTH, expand=True)
            canvas.pack(side="left", fill=tk.BOTH, expand=True)
            scrollbar.pack(side="right", fill="y")

            source_sheet_names = source_xls.sheet_names
            dest_sheet_names = dest_xls.sheet_names

            active_transfers = []

            def add_transfer():
                TransferSettings(scrollable_frame, source_sheet_names, dest_sheet_names, active_transfers)

            def check_conflicts(settings_list):
                seen_combinations = set()
                for settings in settings_list:
                    dest_combo = (settings['dest_tab'], settings['dest_note_col'])
                    if dest_combo in seen_combinations:
                        return True
                    seen_combinations.add(dest_combo)
                return False

            def validate_and_start_transfer():
                if not active_transfers:
                    messagebox.showerror("Error", "At least one transfer must be configured.")
                    return

                settings_list = [ts.get_settings() for ts in active_transfers]
                
                # Validate all fields are filled
                for settings in settings_list:
                    if not all(settings.values()):
                        messagebox.showerror("Error", "All fields must be filled for each transfer.")
                        return
                    
                    # Validate column letters
                    if not all(col.isalpha() for col in [settings['source_file_col'], settings['source_note_col'], 
                                                        settings['dest_file_col'], settings['dest_note_col']]):
                        messagebox.showerror("Error", "Invalid column letter(s) found.")
                        return

                # Check for conflicts
                if check_conflicts(settings_list):
                    messagebox.showerror("Error", "Conflict detected: Cannot write to the same destination column in the same sheet.")
                    return

                settings_window.destroy()
                start_transfer(settings_list)

            buttons_frame = ttk.Frame(settings_window)
            buttons_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

            add_button = ttk.Button(buttons_frame, text="Add Transfer", command=add_transfer)
            add_button.pack(side=tk.LEFT, padx=5)

            ok_button = ttk.Button(buttons_frame, text="Start Transfer", command=validate_and_start_transfer)
            ok_button.pack(side=tk.RIGHT, padx=5)

            cancel_button = ttk.Button(buttons_frame, text="Cancel", command=on_window_close)
            cancel_button.pack(side=tk.RIGHT, padx=5)

            add_transfer()
            settings_window.mainloop()

        def start_transfer(settings_list):
            try:
                source_wb = openpyxl.load_workbook(temp_source_file, data_only=False)
                dest_wb = openpyxl.load_workbook(temp_dest_file, data_only=False)
                for settings in settings_list:
                    source_ws = source_wb[settings['source_tab']]
                    dest_ws = dest_wb[settings['dest_tab']]
                    # Create mapping from source
                    file_map = {}
                    for row in range(1, source_ws.max_row + 1):
                        file_name = source_ws.cell(row=row, column=openpyxl.utils.column_index_from_string(settings['source_file_col'])).value
                        note = source_ws.cell(row=row, column=openpyxl.utils.column_index_from_string(settings['source_note_col'])).value
                        
                        if file_name:
                            cleaned_note = clean_text(note) if note is not None else None
                            converted_note = convert_cell_value(cleaned_note)
                            file_map[file_name] = converted_note
                    # Transfer data
                    for row in range(1, dest_ws.max_row + 1):
                        dest_file_name = dest_ws.cell(row=row, column=openpyxl.utils.column_index_from_string(settings['dest_file_col'])).value
                        if dest_file_name in file_map:
                            value = file_map[dest_file_name]
                            converted_value = convert_cell_value(value)
                            dest_ws.cell(row=row, column=openpyxl.utils.column_index_from_string(settings['dest_note_col'])).value = converted_value
                output_file_path = f"{os.path.splitext(dest_file)[0]}_transferred.xlsx"
                dest_wb.save(output_file_path)
                
                source_wb.close()
                dest_wb.close()
                
                for temp_file in temp_files:
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                    except:
                        pass
                
                messagebox.showinfo("완료", f"Multiple transfers completed successfully!\nSaved to: {output_file_path}")
            except Exception as e:
                try:
                    source_wb.close()
                    dest_wb.close()
                except:
                    pass
                    
                for temp_file in temp_files:
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                    except:
                        pass
                        
                messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

        create_transfer_settings_gui()
    except Exception as e:
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
        messagebox.showerror("Error", f"Failed to access Excel files: {str(e)}")
        return


# Create GUI
root = tk.Tk()
root.title("XLS Transfer - by Neil (ver. 0225)")
root.geometry("600x500")

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width // 2) - (600 // 2)
y = (screen_height // 2) - (500 // 2)
root.geometry(f"600x500+{x}+{y}")

frame = tk.Frame(root)
frame.pack(expand=True, fill=tk.BOTH)

# Excel embedding button
excel_embedding_button = tk.Button(frame, text="Create dictionary", command=lambda: create_upload_settings_gui(filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")]), "create_dictionary"))
excel_embedding_button.pack(padx=10, pady=10)

# Load embeddings and create index button
load_embeddings_button = tk.Button(frame, text="Load dictionary", command=load_embeddings_and_create_index)
load_embeddings_button.pack(padx=10, pady=10)

# File translation button (initially disabled)
translate_file_button = tk.Button(frame, text="Transfer to Close", command=translate_file, state=tk.DISABLED)
translate_file_button.pack(padx=10, pady=10)

# Create a new frame for the threshold entry field and label
threshold_frame = tk.Frame(root)
threshold_frame.pack(padx=10, pady=10)

faiss_threshold_entry = tk.Entry(threshold_frame)
faiss_threshold_entry.grid(row=0, column=1)
tk.Label(threshold_frame, text="최소 일치율").grid(row=0, column=0)
faiss_threshold_entry.insert(0, "0.99")

stop_button = tk.Button(frame, text="STOP", command=stop_processing_function)
stop_button.pack(padx=10, pady=10)

# Add the new button for Excel to Excel translation
excel_to_excel_button = tk.Button(frame, text="Transfer to Excel", command=lambda: create_upload_settings_gui(filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")]), "translate_excel"), state=tk.DISABLED)
excel_to_excel_button.pack(padx=10, pady=10)

check_newlines_button = tk.Button(frame, text="Check Newlines", command=lambda: create_upload_settings_gui(filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")]), "check_newlines"))
check_newlines_button.pack(padx=10, pady=10)

combine_excel_button = tk.Button(frame, text="Combine Excel Files", command=lambda: create_upload_settings_gui(filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")]), "combine_excel"))
combine_excel_button.pack(padx=10, pady=10)

newline_adapt_button = tk.Button(frame, text="Newline Auto Adapt", command=lambda: create_upload_settings_gui(filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")]), "newline_auto_adapt"))
newline_adapt_button.pack(padx=10, pady=10)

# Add the new button for Simple Excel Transfer
simple_excel_transfer_button = tk.Button(frame, text="Simple Excel Transfer", command=simple_excel_transfer)
simple_excel_transfer_button.pack(padx=10, pady=10)


print("실행 완료 !")
root.protocol("WM_DELETE_WINDOW", on_closing)
if splash_active:
    pyi_splash.close()  # Close the splash screen

root.mainloop()