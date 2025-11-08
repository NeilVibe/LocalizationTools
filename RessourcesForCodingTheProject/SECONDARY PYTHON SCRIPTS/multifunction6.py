import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
import openpyxl

# ----------------- FILE HELPERS -----------------
def get_all_files(folder, ext):
    files = []
    for dp, _, fns in os.walk(folder):
        for fn in fns:
            if fn.lower().endswith(ext.lower()):
                files.append(os.path.join(dp, fn))
    return files

def robust_parse_xml(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        content = re.sub(r'^<\?xml[^>]*\?>\s*', '', content, flags=re.MULTILINE)
        content = re.sub(r'<!DOCTYPE[^>]*>\s*', '', content, flags=re.MULTILINE)
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False, no_network=True, recover=True)
        root = etree.fromstring(content.encode("utf-8"), parser=parser)
        return etree.ElementTree(root), root
    except Exception as e:
        print(f"[WARN] Cannot parse {path}: {e}")
        return None, None

# ----------------- STRING HELPERS -----------------
def is_korean(text):
    return bool(text and re.search(r'[\uac00-\ud7af\u1100-\u11ff\u3130-\u318f]', text))

def postprocess_tmx_string(xml_str):
    def replace_outside_ph(pattern, repl, text):
        parts, last = [], 0
        for m in re.finditer(r'<ph>.*?</ph>', text, flags=re.DOTALL):
            chunk = re.sub(pattern, repl, text[last:m.start()])
            parts.append(chunk)
            parts.append(m.group(0))
            last = m.end()
        parts.append(re.sub(pattern, repl, text[last:]))
        return ''.join(parts)

    xml_str = replace_outside_ph(r'%(\d)#',
        lambda m: f'<ph>&lt;mq:rxt-req displaytext="Param{m.group(1)}" val="%{m.group(1)}#" /&gt;</ph>', xml_str)
    xml_str = replace_outside_ph(r'\{([^}]+)\}',
        lambda m: f'<ph>&lt;mq:rxt-req displaytext="{m.group(1)}" val="{{{m.group(1)}}}" /&gt;</ph>', xml_str)
    xml_str = replace_outside_ph(r'\\(\w)',
        lambda m: f'<ph>&lt;mq:rxt displaytext="\\{m.group(1)}" val="\\{m.group(1)}" /&gt;</ph>', xml_str)
    xml_str = replace_outside_ph(r'&lt;br/&gt;',
        lambda m: '<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>', xml_str)
    xml_str = replace_outside_ph(r'&amp;desc;',
        lambda m: '<ph>&lt;mq:rxt-req displaytext="desc" val="&amp;desc;" /&gt;</ph>', xml_str)

    return re.sub(
        r'(<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>)(\s{2,})(\S)',
        lambda m: m.group(1) + f"<ph type='fmt'>{m.group(2)}</ph>" + m.group(3),
        xml_str
    )

# ----------------- XML to Excel -----------------
def xml_to_excel():
    mode = mode_var.get()
    if mode == "folder":
        folder = filedialog.askdirectory(title="Select XML Folder")
        if not folder: return
        xml_files = get_all_files(folder, ".xml")
    else:
        file = filedialog.askopenfilename(title="Select XML File", filetypes=[("XML Files", "*.xml")])
        if not file: return
        xml_files = [file]

    if not xml_files:
        messagebox.showerror("Error", "No XML files found.")
        return

    rows = []
    for path in xml_files:
        _, root = robust_parse_xml(path)
        if root is None: continue
        for locstr in root.iter("LocStr"):
            rows.append((
                locstr.get("StrOrigin", "").strip(),
                locstr.get("Str", "").strip(),
                locstr.get("StringId", "").strip(),
                locstr.get("Desc", "").strip()
            ))

    if not rows:
        messagebox.showerror("Error", "No <LocStr> entries found.")
        return

    out_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if not out_path: return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["StrOrigin", "Str", "StringId", "Desc"])
    for row in rows:
        ws.append(row)
    wb.save(out_path)
    messagebox.showinfo("Success", f"Extracted {len(rows)} rows to:\n{out_path}")

# ----------------- Excel to TMX -----------------
def excel_to_tmx(is_memoq=False):
    mode = mode_var.get()
    if mode == "folder":
        folder = filedialog.askdirectory(title="Select Excel Folder")
        if not folder: return
        excel_files = get_all_files(folder, ".xlsx")
    else:
        file = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
        if not file: return
        excel_files = [file]

    if not excel_files:
        messagebox.showerror("Error", "No Excel files found.")
        return

    XML_NS = "http://www.w3.org/XML/1998/namespace"
    tmx_root = etree.Element("tmx", version="1.4")
    etree.SubElement(tmx_root, "header", creationtool="Excel2TMX", creationtoolversion="1.0",
                     segtype="sentence", o_tmf="UTF-8", adminlang="en", srclang="ko", datatype="PlainText")
    body_elem = etree.SubElement(tmx_root, "body")

    for excel_file in excel_files:
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file {excel_file}:\n{e}")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            src         = str(row[0]).strip() if row[0] else ""
            tgt         = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            ctx         = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            desc_origin = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            desc_text   = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            # MAIN TU
            if src and tgt and not is_korean(tgt):
                tu = etree.SubElement(body_elem, "tu", creationid="Excel2TMX", changeid="Excel2TMX")
                tuv_ko = etree.SubElement(tu, "tuv", {f"{{{XML_NS}}}lang": "ko"})
                etree.SubElement(tuv_ko, "seg").text = src
                tuv_tgt = etree.SubElement(tu, "tuv", {f"{{{XML_NS}}}lang": "en-us"})
                etree.SubElement(tuv_tgt, "seg").text = tgt
                etree.SubElement(tu, "prop", type="x-document").text = os.path.basename(excel_file)
                etree.SubElement(tu, "prop", type="x-context").text = ctx or "NoStringId"

            # DESCRIPTION TU
            if ctx and desc_origin and desc_text and not is_korean(desc_text):
                if desc_origin in ("&desc;", "&amp;desc;") and desc_text in ("&desc;", "&amp;desc;"):
                    continue

                if not desc_origin.startswith("&desc;") and not desc_origin.startswith("&amp;desc;"):
                    desc_origin = "&desc;" + desc_origin

                desc_tu = etree.SubElement(body_elem, "tu", creationid="PAAT", changeid="PAAT")
                tuv_ko_d = etree.SubElement(desc_tu, "tuv", {f"{{{XML_NS}}}lang": "ko"})
                tuv_tgt_d = etree.SubElement(desc_tu, "tuv", {f"{{{XML_NS}}}lang": "en-us"})
                seg_ko = etree.SubElement(tuv_ko_d, "seg")
                seg_tgt = etree.SubElement(tuv_tgt_d, "seg")

                if is_memoq:
                    ph_ko = etree.Element("ph")
                    ph_ko.text = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
                    seg_ko.append(ph_ko)
                    ph_ko.tail = desc_origin.replace("&amp;desc;", "").replace("&desc;", "")

                    ph_tgt = etree.Element("ph")
                    ph_tgt.text = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
                    seg_tgt.append(ph_tgt)
                    ph_tgt.tail = desc_text.replace("&amp;desc;", "").replace("&desc;", "")
                else:
                    seg_ko.text = desc_origin  # raw, so only single escape
                    seg_tgt.text = desc_text

                etree.SubElement(desc_tu, "prop", type="x-document").text = os.path.basename(excel_file)
                etree.SubElement(desc_tu, "prop", type="x-context").text = ctx

    xml_bytes = etree.tostring(tmx_root, pretty_print=True, encoding='UTF-8', xml_declaration=True)
    xml_str = xml_bytes.decode('utf-8')
    if is_memoq:
        xml_str = postprocess_tmx_string(xml_str)

    if mode == "folder":
        out_path = filedialog.asksaveasfilename(defaultextension=".tmx", filetypes=[("TMX Files", "*.tmx")])
        if not out_path: return
    else:
        out_path = os.path.splitext(excel_files[0])[0] + ("_memoq.tmx" if is_memoq else ".tmx")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(xml_str)

    messagebox.showinfo("Success", f"Excel to TMX conversion completed.\nSaved to:\n{out_path}")

# ----------------- TMX to Excel -----------------
def tmx_to_excel():
    mode = mode_var.get()
    if mode == "folder":
        folder = filedialog.askdirectory(title="Select TMX Folder")
        if not folder: return
        tmx_files = get_all_files(folder, ".tmx")
    else:
        file = filedialog.askopenfilename(title="Select TMX File", filetypes=[("TMX Files", "*.tmx")])
        if not file: return
        tmx_files = [file]

    if not tmx_files:
        messagebox.showerror("Error", "No TMX files found.")
        return

    XML_NS = "{http://www.w3.org/XML/1998/namespace}"
    all_rows = []

    for tmx_file in tmx_files:
        _, root = robust_parse_xml(tmx_file)
        if root is None: continue
        for tu in root.findall(".//tu"):
            src = tgt = ctx = desc = ""
            for tuv in tu.findall("tuv"):
                lang = tuv.attrib.get(f"{XML_NS}lang") or tuv.attrib.get("xml:lang") or ""
                seg = tuv.find("seg")
                if lang.lower().startswith("ko") and seg is not None and seg.text:
                    src = seg.text.strip()
                elif lang.lower().startswith("en") and seg is not None and seg.text:
                    tgt = seg.text.strip()
            for prop in tu.findall("prop"):
                if prop.attrib.get("type") == "x-context" and prop.text:
                    ctx = prop.text.strip()
                elif prop.attrib.get("type") == "x-description" and prop.text:
                    desc = prop.text.strip()
            all_rows.append((src, tgt, ctx, desc))

    if not all_rows:
        messagebox.showerror("Error", "No translation units found.")
        return

    if mode == "folder":
        out_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not out_path: return
    else:
        out_path = os.path.splitext(tmx_files[0])[0] + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Source", "Target", "Context", "Description"])
    for row in all_rows:
        ws.append(row)
    wb.save(out_path)

    messagebox.showinfo("Success", f"TMX to Excel conversion completed.\nSaved to:\n{out_path}")

# ----------------- MAIN GUI -----------------
def main():
    global mode_var
    root = tk.Tk()
    root.title("Excel <-> TMX & XML to Excel Converter")
    root.geometry("420x420")

    tk.Label(root, text="Excel <-> TMX & XML to Excel Converter", font=("Arial", 14, "bold")).pack(pady=10)

    mode_var = tk.StringVar(value="file")
    mode_frame = tk.Frame(root)
    mode_frame.pack(pady=5)
    tk.Label(mode_frame, text="Mode:").pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="File Mode", variable=mode_var, value="file").pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="Folder Mode", variable=mode_var, value="folder").pack(side="left", padx=4)

    tk.Button(root, text="Excel to TMX", width=30, height=2,
              command=lambda: excel_to_tmx(is_memoq=False), bg="#d0ffd0").pack(pady=5)
    tk.Button(root, text="Excel to MemoQ-TMX", width=30, height=2,
              command=lambda: excel_to_tmx(is_memoq=True), bg="#ffe080").pack(pady=5)
    tk.Button(root, text="TMX to Excel", width=30, height=2,
              command=tmx_to_excel, bg="#d0e0ff").pack(pady=5)
    tk.Button(root, text="XML to Excel", width=30, height=2,
              command=xml_to_excel, bg="#ffcccc").pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    main()