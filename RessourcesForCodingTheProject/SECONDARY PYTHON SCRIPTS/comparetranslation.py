import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def pick_excel_file(title: str) -> Path:
    """Open a native file dialog and return the selected .xlsx path."""
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel workbooks", "*.xlsx")],
    )
    if not file_path:
        return None
    return Path(file_path)


def normalize(text) -> str:
    """Convert to string, strip, collapse whitespace."""
    if text is None:
        return ""
    return " ".join(str(text).strip().split())


def get_headers(file_path: Path):
    """Return list of headers from first row of the active sheet."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")
    return headers


def process_excel(source_path: Path, target_path: Path,
                  source_col: int, target_col: int, mode: str) -> None:
    """
    Compare source_col in SOURCE file vs target_col in TARGET file.
    mode = 'matches' -> color matches green
    mode = 'differences' -> color differences red
    Colors applied directly to TARGET file.
    """
    wb_source = load_workbook(source_path, data_only=True)
    ws_source = wb_source.active

    wb_target = load_workbook(target_path)
    ws_target = wb_target.active

    green_fill = PatternFill(start_color="FF00FF00",
                             end_color="FF00FF00",
                             fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000",
                           end_color="FFFF0000",
                           fill_type="solid")

    rows_changed = 0

    # Build mapping from source values for quick lookup
    source_values = {}
    for row in ws_source.iter_rows(min_row=2):
        val = normalize(row[source_col - 1].value)
        if val:
            source_values[val.lower()] = True

    # Compare target values against source values
    for row in ws_target.iter_rows(min_row=2):
        tgt_cell = row[target_col - 1]
        tgt_val = normalize(tgt_cell.value)
        if not tgt_val:
            continue

        match_found = tgt_val.lower() in source_values

        if mode == "matches" and match_found:
            tgt_cell.fill = green_fill
            rows_changed += 1
        elif mode == "differences" and not match_found:
            tgt_cell.fill = red_fill
            rows_changed += 1

    wb_target.save(target_path)
    messagebox.showinfo(
        "Done",
        f"Processed TARGET: {target_path.name}\n"
        f"Cells colored: {rows_changed}"
    )


def run_comparison(mode: str, src_entry, tgt_entry, src_col_idx, tgt_col_idx):
    source_file = Path(src_entry.get())
    target_file = Path(tgt_entry.get())

    if not source_file.exists() or not target_file.exists():
        messagebox.showerror("Missing Files", "Please select valid source and target files.")
        return

    process_excel(source_file, target_file, src_col_idx + 1, tgt_col_idx + 1, mode)


def browse_file(entry_widget, title):
    path = pick_excel_file(title)
    if path:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, str(path))


def choose_columns_popup(src_entry, tgt_entry, mode):
    """Popup window to choose columns from headers."""
    source_file = Path(src_entry.get())
    target_file = Path(tgt_entry.get())

    if not source_file.exists() or not target_file.exists():
        messagebox.showerror("Missing Files", "Please select valid source and target files first.")
        return

    src_headers = get_headers(source_file)
    tgt_headers = get_headers(target_file)

    popup = tk.Toplevel()
    popup.title("Select Columns")
    popup.geometry("400x200")
    popup.grab_set()

    tk.Label(popup, text="Source Column:").pack(pady=5)
    src_combo = ttk.Combobox(popup, values=src_headers, state="readonly")
    src_combo.pack(pady=5)

    tk.Label(popup, text="Target Column:").pack(pady=5)
    tgt_combo = ttk.Combobox(popup, values=tgt_headers, state="readonly")
    tgt_combo.pack(pady=5)

    def confirm_selection():
        src_idx = src_combo.current()
        tgt_idx = tgt_combo.current()
        if src_idx < 0 or tgt_idx < 0:
            messagebox.showerror("Selection Error", "Please select both columns.")
            return
        popup.destroy()
        run_comparison(mode, src_entry, tgt_entry, src_idx, tgt_idx)

    tk.Button(popup, text="Run", command=confirm_selection).pack(pady=15)


def main():
    root = tk.Tk()
    root.title("Excel Source vs Target Comparator")

    tk.Label(root, text="Source File:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    src_entry = tk.Entry(root, width=50)
    src_entry.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(src_entry, "Select SOURCE Excel file")).grid(row=0, column=2, padx=5, pady=5)

    tk.Label(root, text="Target File:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    tgt_entry = tk.Entry(root, width=50)
    tgt_entry.grid(row=1, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(tgt_entry, "Select TARGET Excel file")).grid(row=1, column=2, padx=5, pady=5)

    tk.Button(root, text="Highlight Matches (Green)",
              width=30,
              command=lambda: choose_columns_popup(src_entry, tgt_entry, "matches")).grid(row=2, column=0, columnspan=3, pady=10)

    tk.Button(root, text="Highlight Differences (Red)",
              width=30,
              command=lambda: choose_columns_popup(src_entry, tgt_entry, "differences")).grid(row=3, column=0, columnspan=3, pady=5)

    tk.Button(root, text="Exit", width=30, command=root.destroy).grid(row=4, column=0, columnspan=3, pady=20)

    root.mainloop()


if __name__ == "__main__":
    main()