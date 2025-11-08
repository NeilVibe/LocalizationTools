import os
import difflib
import openpyxl
from lxml import etree
from tkinter import filedialog, Tk, messagebox, Button, Label, Frame

# -------------------- COMMON HELPERS --------------------

def read_excel_data(path, max_cols=4):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, max_col=max_cols, values_only=True):
        cleaned = [str(c).strip() if c else "" for c in row]
        if cleaned[0]:
            rows.append(tuple(cleaned))
    return rows

def read_xml_locstr(path):
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False, no_network=True, recover=True)
    try:
        tree = etree.parse(path, parser)
    except Exception as e:
        print(f"[ERROR] Failed to parse XML: {e}")
        return []
    root = tree.getroot()
    entries = []
    for loc in root.iter("LocStr"):
        entries.append({
            "StrOrigin": loc.get("StrOrigin", "").strip(),
            "StringId": loc.get("StringId", "").strip(),
            "Str": loc.get("Str", "").strip(),
            "Desc": loc.get("Desc", "").strip()
        })
    return entries

def find_soundeventnames(folder_path, xml_entries):
    """Map StringId+StrOrigin to SoundEventName from folder XMLs."""
    sound_map = {}
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False, no_network=True, recover=True)
    # Build lookup set for faster match
    lookup_set = {(e["StringId"], e["StrOrigin"]) for e in xml_entries}
    for root_dir, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(".xml"):
                file_path = os.path.join(root_dir, file)
                try:
                    tree = etree.parse(file_path, parser)
                except Exception:
                    continue
                for loc in tree.getroot().iter("LocStr"):
                    sid = loc.get("StringId", "").strip()
                    sorigin = loc.get("StrOrigin", "").strip()
                    if (sid, sorigin) in lookup_set:
                        sound_event = loc.get("SoundEventName", "").strip()
                        if sound_event:
                            sound_map.setdefault((sid, sorigin), set()).add(sound_event)
    return sound_map

def save_matches_to_excel(matches, output_path, headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in matches:
        ws.append(row)
    wb.save(output_path)

# -------------------- PROCESS 1 (Main Process) --------------------

def match_excel_to_xml(excel_rows, xml_entries, sound_map):
    matches = []
    used_ids = set()
    for (col1, col2, col3, col4) in excel_rows:
        search_terms = [t.strip().lower() for t in (col2, col3, col4) if t]
        potential_matches = [
            e for e in xml_entries
            if any(term in (e["Str"].strip().lower(), e["StrOrigin"].strip().lower(), e["Desc"].strip().lower())
                   for term in search_terms)
        ]
        potential_matches = [e for e in potential_matches if e["StringId"] not in used_ids]
        if not potential_matches:
            matches.append((col1, "NO_MATCH", ""))
            continue
        if len(potential_matches) == 1:
            e = potential_matches[0]
            used_ids.add(e["StringId"])
            sound_events = sound_map.get((e["StringId"], e["StrOrigin"]), [""])
            sound_event = list(sound_events)[0] if sound_events else ""
            matches.append((col1, e["StringId"], sound_event))
        else:
            best_entry = None
            best_score = -1
            for e in potential_matches:
                sound_events = sound_map.get((e["StringId"], e["StrOrigin"]), [""])
                for se in sound_events:
                    score = difflib.SequenceMatcher(None, col1.lower(), se.lower()).ratio()
                    if score > best_score:
                        best_score = score
                        best_entry = (e["StringId"], se)
            if best_entry:
                used_ids.add(best_entry[0])
                matches.append((col1, best_entry[0], best_entry[1]))
            else:
                matches.append((col1, "NO_MATCH", ""))
    return matches

def run_main_process():
    excel_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path: return
    xml_path = filedialog.askopenfilename(title="Select XML File", filetypes=[("XML Files", "*.xml")])
    if not xml_path: return
    folder_path = filedialog.askdirectory(title="Select Folder Containing XML Files for SoundEventName Mapping")
    if not folder_path: return

    excel_rows = read_excel_data(excel_path, max_cols=4)
    xml_entries = read_xml_locstr(xml_path)
    sound_map = find_soundeventnames(folder_path, xml_entries)
    matches = match_excel_to_xml(excel_rows, xml_entries, sound_map)

    output_path = os.path.join(os.path.dirname(excel_path), "matched_with_soundevent.xlsx")
    save_matches_to_excel(matches, output_path, ["Excel_Col1", "Matched_StringId", "Mapped_SoundEventName"])
    messagebox.showinfo("Done", f"Main Process complete.\nSaved to:\n{output_path}")

# -------------------- PROCESS 2 (Custom Process with Detailed Tracking) --------------------

def custom_match_process(excel_rows, xml_entries, sound_map):
    matches = []
    print("\n========== CUSTOM MATCH PROCESS START ==========\n")
    for idx, (col1, col2, col3) in enumerate(excel_rows, start=1):
        print(f"[ROW {idx}] Processing:")
        print(f"  Col1_StrOrigin: '{col1}'")
        print(f"  Col2_Str:       '{col2}'")
        print(f"  Col3_BaseID:    '{col3}'")

        potential_matches = [e for e in xml_entries if e["StrOrigin"].strip().lower() == col1.strip().lower()]
        print(f"  → Found {len(potential_matches)} potential matches in XML.")

        if not potential_matches:
            print("  ❌ No matches found. Assigning NO_MATCH.\n")
            matches.append((col1, col2, col3, "NO_MATCH"))
            continue

        if len(potential_matches) == 1:
            e = potential_matches[0]
            sound_events = sound_map.get((e["StringId"], e["StrOrigin"]), [""])
            sound_event = list(sound_events)[0] if sound_events else ""
            print(f"  ✅ Single match found: StringId='{e['StringId']}', SoundEventName='{sound_event}'\n")
            matches.append((col1, col2, col3, e["StringId"]))
        else:
            print("  ⚠ Multiple matches found. Comparing SoundEventNames for best match...")
            best_entry = None
            best_score = -1
            for e in potential_matches:
                sound_events = sound_map.get((e["StringId"], e["StrOrigin"]), [""])
                for se in sound_events:
                    score = difflib.SequenceMatcher(None, col3.lower(), se.lower()).ratio()
                    print(f"    Comparing BaseID='{col3}' with SoundEventName='{se}' → Score={score:.4f}")
                    if score > best_score:
                        best_score = score
                        best_entry = e["StringId"]
            if best_entry:
                print(f"  ✅ Best match: StringId='{best_entry}' with score {best_score:.4f}\n")
                matches.append((col1, col2, col3, best_entry))
            else:
                print("  ❌ No suitable SoundEventName match found. Assigning NO_MATCH.\n")
                matches.append((col1, col2, col3, "NO_MATCH"))
    print("\n========== CUSTOM MATCH PROCESS END ==========\n")
    return matches

def run_custom_process():
    excel_path = filedialog.askopenfilename(title="Select Excel File (3 columns)", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path: return
    xml_path = filedialog.askopenfilename(title="Select XML File", filetypes=[("XML Files", "*.xml")])
    if not xml_path: return
    folder_path = filedialog.askdirectory(title="Select Folder Containing XML Files for SoundEventName Mapping")
    if not folder_path: return

    excel_rows = read_excel_data(excel_path, max_cols=3)
    xml_entries = read_xml_locstr(xml_path)
    sound_map = find_soundeventnames(folder_path, xml_entries)

    matches = custom_match_process(excel_rows, xml_entries, sound_map)

    output_path = os.path.join(os.path.dirname(excel_path), "custom_matched.xlsx")
    save_matches_to_excel(matches, output_path, ["Col1_StrOrigin", "Col2_Str", "Col3_BaseID", "Matched_StringId"])
    messagebox.showinfo("Done", f"Custom Process complete.\nSaved to:\n{output_path}")

# -------------------- GUI --------------------

def main_gui():
    root = Tk()
    root.title("XML/Excel Matcher")
    root.geometry("400x200")

    Label(root, text="Select a process to run:", font=("Arial", 14)).pack(pady=10)

    frame = Frame(root)
    frame.pack(pady=20)

    Button(frame, text="Main Process", width=20, command=run_main_process).grid(row=0, column=0, padx=10, pady=5)
    Button(frame, text="Custom Process", width=20, command=run_custom_process).grid(row=1, column=0, padx=10, pady=5)

    root.mainloop()

if __name__ == "__main__":
    main_gui()