#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Knowledge-data extractor – ENG only, with proper nesting, depth anomaly detection,
deep debug always ON, logs written to file (source XML only).
Aggressive preprocessing to fix bad XML issues.
Splits Excel output into multiple sheets based on depth=2 rows.
Ignores banned strings.
Highlights KnowledgeGroupInfo nodes that have KnowledgeGroupIcon but are not already depth-colored — in BLUE.
Bold any row that is colored.
Adds very light yellow fill to cells with no color, auto word wrap, and borders for all cells.
Also bolds any cell corresponding to Name="" attribute values (not Desc).
"""

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\knowledgeinfo")
LANGUAGE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
OUTPUT_EXCEL    = Path.cwd() / "Knowledge_LQA_ENG.xlsx"
LOG_FILE        = Path.cwd() / "knowledge_scan.log"

IGNORE_LIST = [
    "해제/면역 연금 포션",
]

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")]
)
log = logging.getLogger("KnowledgeLQA")

# ──────────────────────────────────────────────────────────────────────────────
# SANITIZATION HELPERS
# ──────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)

def preprocess_newlines_in_tags(raw_content: str) -> str:
    def replace_in_seg(match):
        seg_content = match.group(1)
        cleaned = seg_content.replace("\n", "&lt;br/&gt;").replace("\\n", "&lt;br/&gt;")
        return f"<seg>{cleaned}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", replace_in_seg, raw_content, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = preprocess_newlines_in_tags(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"', lambda m: '="' + m.group(1).replace('<', '&lt;') + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"', lambda m: '="' + m.group(1).replace('&', '&amp;') + '"', raw)

    tag_stack = []
    tag_open_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed_lines = []
    for line in raw.splitlines():
        stripped = line.strip()
        m_open = tag_open_re.match(stripped)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed_lines.append(line)
            continue
        m_close = tag_close_re.match(stripped)
        if m_close:
            if tag_stack and tag_stack[-1] == m_close.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(line.replace("</>", f"</{tag_stack.pop()}>"))
                continue
        fixed_lines.append(line)
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)

# ──────────────────────────────────────────────────────────────────────────────
# XML PARSER
# ──────────────────────────────────────────────────────────────────────────────
def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Failed to read %s", path)
        return None
    raw_fixed = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{raw_fixed}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(recover=False, huge_tree=True))
    except ET.XMLSyntaxError:
        log.exception("Strict parse failed for %s", path)
        try:
            return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(recover=True, huge_tree=True))
        except ET.XMLSyntaxError:
            log.exception("Recover parse failed for %s", path)
            return None

# ──────────────────────────────────────────────────────────────────────────────
# LOAD ENG LANGUAGE TABLE
# ──────────────────────────────────────────────────────────────────────────────
def parse_eng_language(folder: Path) -> Dict[str, Tuple[str, str]]:
    for xml_path in sorted(folder.rglob("LanguageData_eng.xml")):
        root = parse_xml_file(xml_path)
        if root is None:
            return {}
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            trans  = el.get("Str") or ""
            sid    = el.get("StringId") or ""
            if origin:
                tbl[origin] = (trans, sid)
        return tbl
    return {}

# ──────────────────────────────────────────────────────────────────────────────
# SCAN HIERARCHY
# ──────────────────────────────────────────────────────────────────────────────
RowItem = Tuple[int, str, bool, bool, bool]  # depth, text, needs_trans, is_icon, is_name_attr

def should_ignore(text: str) -> bool:
    return any(banned in text for banned in IGNORE_LIST)

def scan_knowledge_hierarchy(folder: Path) -> List[RowItem]:
    xml_files = sorted(folder.rglob("*.xml"))
    if not xml_files:
        sys.exit(f"No XML files found in {folder}")
    rows: List[RowItem] = []
    last_depth = None
    def check_depth(depth: int, text: str):
        nonlocal last_depth
        if last_depth is not None and abs(depth - last_depth) > 2:
            log.warning("Depth anomaly: %d -> %d for '%s'", last_depth, depth, text)
        last_depth = depth
    def recurse(node: ET._Element, depth: int) -> None:
        if node.tag == "br":
            return
        if node.tag == "KnowledgeGroupInfo":
            is_icon = bool(node.get("KnowledgeGroupIcon"))
            name = node.get("GroupName") or ""
            desc = node.get("Desc") or ""
            if name and not should_ignore(name):
                check_depth(depth, name)
                rows.append((depth, name, True, is_icon, True))  # Name attr bold
            if desc and not should_ignore(desc):
                check_depth(depth + 1, desc)
                rows.append((depth + 1, desc, True, is_icon, False))
            for child in node:
                recurse(child, depth + 1)
        elif node.tag == "KnowledgeInfo":
            kn_name = node.get("Name") or ""
            kn_desc = node.get("Desc") or ""
            if kn_name and not should_ignore(kn_name):
                check_depth(depth, kn_name)
                rows.append((depth, kn_name, True, False, True))  # Name attr bold
            if kn_desc and not should_ignore(kn_desc):
                check_depth(depth + 1, kn_desc)
                rows.append((depth + 1, kn_desc, True, False, False))
            for ld in node.findall("LevelData"):
                lvl = ld.get("Level") or ""
                if not should_ignore(f"Level {lvl}"):
                    check_depth(depth + 1, f"Level {lvl}")
                    rows.append((depth + 1, f"Level {lvl}", False, False, False))
                text = ld.get("Desc") or ""
                if text and not should_ignore(text):
                    check_depth(depth + 2, text)
                    rows.append((depth + 2, text, True, False, False))
        else:
            for child in node:
                recurse(child, depth + 1)
    for xml_path in xml_files:
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        recurse(root, 0)
    return rows

# ──────────────────────────────────────────────────────────────────────────────
# WRITE EXCEL
# ──────────────────────────────────────────────────────────────────────────────
def write_multi_excel(rows: List[RowItem], eng_table: Dict[str, Tuple[str, str]], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fills = {
        0: PatternFill("solid", fgColor="FFD966"),
        1: PatternFill("solid", fgColor="D9E1F2"),
        2: PatternFill("solid", fgColor="E2EFDA"),
        3: PatternFill("solid", fgColor="FCE4D6"),
    }
    icon_fill = PatternFill("solid", fgColor="9BC2E6")  # blue only if no depth color
    light_yellow_fill = PatternFill("solid", fgColor="FFFDEB")  # very light yellow for no color
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    current_sheet = None
    r = 2
    for idx, (depth, text, needs_trans, is_icon, is_name_attr) in enumerate(rows):
        fg = fills.get(depth)  # None if no depth color
        if fg is None and is_icon:
            fg = icon_fill  # only blue if no depth color
        translation = eng_table.get(text, ("", ""))[0] if needs_trans else ""
        if depth == 2:
            if current_sheet is not None:
                for col_cells in current_sheet.iter_cols(min_col=1, max_col=2):
                    maxw = max((len(str(cell.value)) if cell.value else 0 for cell in col_cells), default=0)
                    current_sheet.column_dimensions[col_cells[0].column_letter].width = min(maxw * 1.1 + 2, 80)
            sheet_name = translation.strip() or text.strip() or f"Sheet{idx}"
            sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
            current_sheet = wb.create_sheet(title=sheet_name)
            current_sheet.cell(1, 1, "Original").font = header_font
            current_sheet.cell(1, 1).fill = header_fill
            current_sheet.cell(1, 1).border = border_style
            current_sheet.cell(1, 2, "Translation").font = header_font
            current_sheet.cell(1, 2).fill = header_fill
            current_sheet.cell(1, 2).border = border_style
            r = 2
        if current_sheet is None:
            continue
        co = current_sheet.cell(r, 1, text)
        co.fill = fg if fg else light_yellow_fill
        if fg or is_name_attr:
            co.font = Font(bold=True)
        co.alignment = Alignment(indent=depth, wrap_text=True)  # default left alignment
        co.border = border_style
        ct = current_sheet.cell(r, 2, translation)
        ct.fill = fg if fg else light_yellow_fill
        if fg or is_name_attr:
            ct.font = Font(bold=True)
        ct.alignment = Alignment(indent=depth, wrap_text=True)  # default left alignment
        ct.border = border_style
        r += 1
    wb.save(out_path)
    log.info("Saved Excel → %s", out_path)

# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main() -> None:
    rows = scan_knowledge_hierarchy(RESOURCE_FOLDER)
    eng_table = parse_eng_language(LANGUAGE_FOLDER)
    write_multi_excel(rows, eng_table, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()