#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generates an Excel report with:
    • "Normal Summary" sheet: per-file stats for all languages (excluding KOR)
    • "Detailed Summary" sheet: all languages' detailed category coverage stacked vertically
      with a yellow separator row between each language block (excluding KOR).
"""

import os
import re
from pathlib import Path
from typing import Dict, List, Tuple, Set

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
LANGUAGE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
EXPORT_FOLDER   = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
OUTPUT_FILE     = Path.cwd() / "Translation_Coverage_SUMMARY.xlsx"

# ─────────────────────────────────────────────────────────────
# XML PARSE HELPERS
# ─────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)

def parse_xml_file(path: Path) -> ET._Element:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    wrapped = f"<ROOT>\n{fix_bad_entities(raw)}\n</ROOT>"
    parser = ET.XMLParser(recover=True, huge_tree=True)
    return ET.fromstring(wrapped.encode("utf-8"), parser=parser)

# ─────────────────────────────────────────────────────────────
# FILE ITERATION
# ─────────────────────────────────────────────────────────────
def iter_language_files(folder: Path):
    for dirpath, _, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().startswith("languagedata_") and fn.lower().endswith(".xml"):
                yield Path(dirpath) / fn

# ─────────────────────────────────────────────────────────────
# WORD-COUNT & TRANSLATION DETECTION
# ─────────────────────────────────────────────────────────────
korean_re = re.compile(r'[\uac00-\ud7a3]')

def count_words(text: str) -> int:
    return len([w for w in re.split(r'\s+', text.strip()) if w])

def is_korean(text: str) -> bool:
    return bool(korean_re.search(text))

def analyse_file(path: Path) -> Tuple[int, int, int, int]:
    root = parse_xml_file(path)
    total_words = total_nodes = completed_nodes = completed_words = 0
    for loc in root.iter("LocStr"):
        origin = (loc.get("StrOrigin") or "").strip()
        if not origin:
            continue
        origin_wc = count_words(origin)
        total_nodes  += 1
        total_words  += origin_wc
        trans = (loc.get("Str") or "").strip()
        if trans and not is_korean(trans):
            completed_nodes += 1
            completed_words += origin_wc
    return total_words, total_nodes, completed_nodes, completed_words

def collect_completed_ids_for_paths(xml_paths: List[Path]) -> Set[str]:
    completed: Set[str] = set()
    for xml_path in xml_paths:
        root = parse_xml_file(xml_path)
        for loc in root.iter("LocStr"):
            origin = (loc.get("StrOrigin") or "").strip()
            if not origin:
                continue
            trans = (loc.get("Str") or "").strip()
            if trans and not is_korean(trans):
                sid = loc.get("StringId")
                if sid:
                    completed.add(sid)
    return completed

def analyse_export_file(path: Path, completed_ids: Set[str]) -> Tuple[int, int, int, int]:
    root = parse_xml_file(path)
    tw = tn = cn = cw = 0
    for loc in root.iter("LocStr"):
        origin = (loc.get("StrOrigin") or "").strip()
        if not origin:
            continue
        origin_wc = count_words(origin)
        tn += 1
        tw += origin_wc
        sid = loc.get("StringId")
        if sid in completed_ids:
            cn += 1
            cw += origin_wc
    return tw, tn, cn, cw

# ─────────────────────────────────────────────────────────────
# EXCEL UTILITIES
# ─────────────────────────────────────────────────────────────
def style_header(ws, row: int, headers: List[str]) -> None:
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 2)

def style_separator(ws, row: int, num_cols: int) -> None:
    sep_fill = PatternFill("solid", fgColor="FFFF00")  # Yellow
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col, "")
        cell.fill = sep_fill

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main() -> None:
    print(f"[INFO] Scanning language folder: {LANGUAGE_FOLDER}")

    per_file: List[Tuple[str, str, int, int, int, int]] = []
    xmls_by_lang: Dict[str, List[Path]] = {}

    for xml_path in iter_language_files(LANGUAGE_FOLDER):
        stem = xml_path.stem
        parts = stem.split("_", 1)
        if len(parts) != 2:
            continue
        lang_code = parts[1].upper()

        # ─── FILTER OUT KOR ───
        if lang_code == "KOR":
            continue

        tw, tn, cn, cw = analyse_file(xml_path)
        per_file.append((lang_code, xml_path.name, tw, tn, cn, cw))
        xmls_by_lang.setdefault(lang_code, []).append(xml_path)
        print(f"    [{lang_code}] {xml_path.name}: {cn}/{tn} nodes  •  {cw}/{tw} words translated")

    languages = sorted(xmls_by_lang.keys())

    # ───── Build Excel workbook ─────
    wb = Workbook()

    # Normal Summary sheet
    ws_normal = wb.active
    ws_normal.title = "Normal Summary"
    headers_file = [
        "Language", "File", "Total Words", "Total Nodes",
        "Completed Nodes", "Completed Words", "Node Coverage %", "Word Coverage %"
    ]
    style_header(ws_normal, 1, headers_file)
    for row, (lang, fname, tw, tn, cn, cw) in enumerate(sorted(per_file), start=2):
        node_cov = 100 * cn / tn if tn else 0.0
        word_cov = 100 * cw / tw if tw else 0.0
        ws_normal.cell(row, 1, lang)
        ws_normal.cell(row, 2, fname)
        ws_normal.cell(row, 3, tw)
        ws_normal.cell(row, 4, tn)
        ws_normal.cell(row, 5, cn)
        ws_normal.cell(row, 6, cw)
        ws_normal.cell(row, 7, round(node_cov, 2))
        ws_normal.cell(row, 8, round(word_cov, 2))

    # ───── Prepare categories ─────
    print(f"[INFO] Scanning export folder: {EXPORT_FOLDER}")
    categories: Dict[str, List[Path]] = {}
    for child in EXPORT_FOLDER.iterdir():
        if not child.is_dir():
            continue
        if child.name == "Sequencer":
            for name in ("Faction", "Main"):
                sub = child / name
                if sub.is_dir():
                    categories[name] = [sub]
            group_paths = []
            for name in ("Sequencer", "Other"):
                sub = child / name
                if sub.is_dir():
                    group_paths.append(sub)
            if group_paths:
                categories["Sequencer + Other"] = group_paths
        else:
            categories[child.name] = [child]

    headers_det = [
        "Category", "Total Words", "Total Nodes",
        "Completed Nodes", "Completed Words", "Node Coverage %", "Word Coverage %"
    ]

    # Detailed Summary sheet
    ws_det = wb.create_sheet("Detailed Summary")
    current_row = 1
    for lang in languages:
        print(f"[INFO] Adding detailed block for: {lang}")
        completed_ids = collect_completed_ids_for_paths(xmls_by_lang[lang])
        # Language title row
        ws_det.cell(current_row, 1, f"Language: {lang}").font = Font(bold=True)
        current_row += 1
        # Header row
        style_header(ws_det, current_row, headers_det)
        current_row += 1
        # Data rows
        for cat, paths in sorted(categories.items()):
            total_words = total_nodes = comp_nodes = comp_words = 0
            for base in paths:
                for xml_path in base.rglob("*.xml"):
                    tw, tn, cn, cw = analyse_export_file(xml_path, completed_ids)
                    total_words += tw
                    total_nodes  += tn
                    comp_nodes   += cn
                    comp_words   += cw
            node_cov = 100 * comp_nodes / total_nodes if total_nodes else 0.0
            word_cov = 100 * comp_words / total_words if total_words else 0.0
            ws_det.cell(current_row, 1, cat)
            ws_det.cell(current_row, 2, total_words)
            ws_det.cell(current_row, 3, total_nodes)
            ws_det.cell(current_row, 4, comp_nodes)
            ws_det.cell(current_row, 5, comp_words)
            ws_det.cell(current_row, 6, round(node_cov, 2))
            ws_det.cell(current_row, 7, round(word_cov, 2))
            current_row += 1
        # Separator row
        style_separator(ws_det, current_row, len(headers_det))
        current_row += 1

    wb.save(OUTPUT_FILE)
    print(f"[INFO] Report written to: {OUTPUT_FILE.resolve()}")

if __name__ == "__main__":
    main()