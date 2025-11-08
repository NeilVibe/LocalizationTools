
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Item-data extractor / LQA helper – multi-language edition.

Key features
============
1.  Robust XML loader identical to the Quest-LQA script
    (illegal-entity and tag-sanitising, strict→recover fallback).
2.  Uses pathlib everywhere + dual logging (file + console).
3.  Proper depth calculation:
        depth 0  = root ItemGroup
        depth 1  = nested ItemGroup
        depth 2  = ItemInfo (under whichever group it belongs)
4.  Excel output per language             (.../Item_LQA_<LANG>.xlsx)
    • Header: white on blue, borders, freeze-pane, auto-filter
    • Depth-based row fills, bold group rows, indented cells
    • Columns:
        A) depth   (hidden)
        B) GroupKey
        C) GroupName-KOR
        D) GroupName-LOC
        E) ItemKey
        F) Item#   (from StringKeyTable, “<MISSING>” if absent)
        G) ItemName-KOR
        H) ItemName-LOC
        I) StringID
"""

from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION – FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\iteminfo"
)
ITEMGROUPINFO_FILE = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\iteminfo\itemgroupinfo.staticinfo.xml"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
)
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\mainline\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)
OUTPUT_FOLDER = Path.cwd() / "ItemData_Map_All"
LOG_FILE = Path.cwd() / "item_scan.log"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING  – FILE + CONSOLE
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("ItemLQA")
log.setLevel(logging.DEBUG)

_file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
_file_handler.setFormatter(
    logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
)
_file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ─────────────────────────────────────────────────────────────────────────────
# XML-SANITISATION UTILITIES  (identical to Quest script)
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)


def _preprocess_newlines(raw_content: str) -> str:
    """
    Replace embedded newlines inside <seg>...</seg> with &lt;br/&gt; so that the
    parser does not treat them as tag starts.
    """
    def repl(m: re.Match) -> str:
        inner = m.group(1)
        inner = inner.replace("\n", "&lt;br/&gt;").replace("\\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"

    return re.sub(r"<seg>(.*?)</seg>", repl, raw_content, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)

    # Escape stray "<" or "&" inside attribute values
    raw = re.sub(
        r'="([^"]*<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*&[^ltgapoqu][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    # Auto-close broken tags
    tag_stack: List[str] = []
    tag_open = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed_lines: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            tag_stack.append(mo.group(1))
            fixed_lines.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if tag_stack and tag_stack[-1] == mc.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            else:
                fixed_lines.append(line)
            continue
        fixed_lines.append(line)
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    parser_strict = ET.XMLParser(huge_tree=True)
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=parser_strict)
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retrying with recovery", path.name)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Even recovery parse failed: %s", path)
            return None


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS  –  FILE ITERATION, INDENT, AUTOFIT
# ─────────────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn


def autofit(ws) -> None:
    """Simple auto-fit based on max string length per column (capped)."""
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len * 1.15 + 2, 80)


# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE  TABLES
# ─────────────────────────────────────────────────────────────────────────────
def parse_language_folder(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    log.info("Scanning language folder: %s", folder)
    langs: Dict[str, Dict[str, Tuple[str, str]]] = {}
    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_") or stem.startswith("languagedata_kor"):
            continue
        code = stem.split("_", 1)[1].lower()
        log.info("Loading language file [%s] – %s", code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            log.error("Skipping language file due to parse error: %s", path.name)
            continue
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            sid = el.get("StringId") or ""
            txt = el.get("Str") or ""
            if origin:
                tbl[origin] = (txt, sid)
        langs[code] = tbl
    log.info("Language tables loaded: %d", len(langs))
    return langs


# ─────────────────────────────────────────────────────────────────────────────
# STRING-KEY TABLE
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Aborting – cannot parse StringKeyTable")
    tbl: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        num = el.get("Key") or ""
        sk = el.get("StrKey") or ""
        if num and sk:
            tbl[sk.lower()] = num
    log.info("StringKeyTable entries: %d", len(tbl))
    return tbl


# ─────────────────────────────────────────────────────────────────────────────
# MASTER  ITEM-GROUP  INFO
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_groups(path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    log.info("Parsing ItemGroupInfo master: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse ItemGroupInfo master file")
    group_name: Dict[str, str] = {}       # StrKey -> GroupName (KOR)
    parent_of: Dict[str, str] = {}        # child StrKey -> parent StrKey
    for el in root.iter("ItemGroupInfo"):
        sk = el.get("StrKey") or ""
        name = el.get("GroupName") or ""
        if sk:
            group_name[sk] = name
            parent_el = el.getparent()
            if parent_el is not None and parent_el.tag == "ItemGroupInfo":
                parent_of[sk] = parent_el.get("StrKey") or ""
    log.info("Groups parsed: %d  |  parent links: %d", len(group_name), len(parent_of))
    return group_name, parent_of


# ─────────────────────────────────────────────────────────────────────────────
# RESOURCE  SCAN
# ─────────────────────────────────────────────────────────────────────────────
#   Returns: dict[GroupStrKey, list[tuple[item_strkey, item_kor_name]]]
def scan_resource_folder(folder: Path) -> Dict[str, List[Tuple[str, str]]]:
    log.info("Scanning resource folder for items: %s", folder)
    group_items: Dict[str, List[Tuple[str, str]]] = {}
    for idx, path in enumerate(iter_xml_files(folder), 1):
        root = parse_xml_file(path)
        if root is None:
            continue
        for g_el in root.iter("ItemGroupInfo"):
            g_key = g_el.get("StrKey") or ""
            if not g_key:
                continue
            bucket = group_items.setdefault(g_key, [])
            for item in g_el.iter("ItemInfo"):
                ik = item.get("StrKey") or ""
                name = item.get("ItemName") or ""
                if ik:
                    bucket.append((ik, name))
        if idx % 200 == 0:
            log.debug("... scanned %d XML files", idx)
    log.info("Group-item mapping built: Groups=%d  Total items=%d",
             len(group_items),
             sum(len(v) for v in group_items.values()))
    return group_items


# ─────────────────────────────────────────────────────────────────────────────
# TREE  HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def build_children_map(parent_of: Dict[str, str]) -> Dict[str, List[str]]:
    children: Dict[str, List[str]] = {}
    for child, parent in parent_of.items():
        children.setdefault(parent, []).append(child)
    return children


def calc_depth(node: str, parent_of: Dict[str, str]) -> int:
    d = 0
    while node in parent_of and parent_of[node]:
        node = parent_of[node]
        d += 1
    return d


# Row tuple: depth, group_key, grp_kor, grp_loc, item_key, item_num,
#            item_kor, item_loc, string_id, is_group_row
Row = Tuple[int, str, str, str, str, str, str, str, str, bool]


def build_rows_for_language(
    code: str,
    group_name_kor: Dict[str, str],
    parent_of: Dict[str, str],
    group_items: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str],
) -> List[Row]:
    log.info("Building rows for language: %s", code.upper())

    def t(text: str) -> str:
        return lang_tbl.get(text, ("", ""))[0]

    rows: List[Row] = []
    children_of = build_children_map(parent_of)

    # Identify root groups (no parent) and sort for deterministic order
    root_groups = sorted([g for g in group_name_kor if g not in parent_of])

    def recurse(group_key: str) -> None:
        depth = calc_depth(group_key, parent_of)
        grp_kor = group_name_kor.get(group_key, "")
        grp_loc = t(grp_kor)
        rows.append(
            (
                depth,
                group_key,
                grp_kor,
                grp_loc,
                "",
                "",
                "",
                "",
                "",
                True,
            )
        )
        # Items
        for ik, in_kor in sorted(group_items.get(group_key, []), key=lambda x: x[0]):
            in_loc = t(in_kor)
            num = id_table.get(ik.lower(), "<MISSING>")
            sid = lang_tbl.get(in_kor, ("", ""))[1]
            rows.append(
                (
                    depth + 1,
                    group_key,
                    grp_kor,
                    grp_loc,
                    ik,
                    num,
                    in_kor,
                    in_loc,
                    sid,
                    False,
                )
            )
        # Recurse into child groups
        for child in sorted(children_of.get(group_key, [])):
            recurse(child)

    for g in root_groups:
        recurse(g)

    log.info("Rows built for %s: %d", code.upper(), len(rows))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL  RENDERING
# ─────────────────────────────────────────────────────────────────────────────
_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),  # root group
    1: PatternFill("solid", fgColor="D9E1F2"),  # subgroup
    2: PatternFill("solid", fgColor="E2EFDA"),  # item rows (will be overriden)
}
_item_fill = PatternFill("solid", fgColor="FCE4D6")
_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_sheet(
    wb: Workbook,
    title: str,
    rows: List[Row],
    hide_depth_col: bool = True,
) -> None:
    ws = wb.create_sheet(title=title[:31])

    headers = [
        "Depth",          # hidden if requested
        "GroupKey",
        "GroupName(KOR)",
        f"GroupName({title})",
        "ItemKey",
        "ItemKey#",
        "ItemName(KOR)",
        f"ItemName({title})",
        "StringID",
    ]
    for col_idx, txt in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border

    r = 2
    for depth, gkey, gk_kor, gk_loc, ikey, inum, in_kor, in_loc, sid, is_group in rows:
        cells = (
            depth,
            gkey,
            gk_kor,
            gk_loc,
            ikey,
            inum,
            in_kor,
            in_loc,
            sid,
        )
        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(r, col_idx, val)
            c.alignment = Alignment(indent=depth if col_idx not in (1, 5, 6, 9) else 0,
                                    wrap_text=True,
                                    vertical="top")
            c.border = _border
            if is_group and col_idx in (3, 4):  # group names only
                c.font = _bold_font
            if is_group:
                c.fill = _depth_fill.get(depth, _depth_fill[1])
            else:
                c.fill = _item_fill
        r += 1

    # Cosmetic
    ws.auto_filter.ref = f"B1:I{r-1}"
    ws.freeze_panes = "A2"
    autofit(ws)
    if hide_depth_col:
        ws.column_dimensions["A"].hidden = True


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("Item LQA extraction started")
    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # 1. Load languages --------------------------------------------------------
    lang_tables = parse_language_folder(LANGUAGE_FOLDER)
    if not lang_tables:
        sys.exit("No language files found!")

    # 2. Load StringKey table --------------------------------------------------
    id_tbl = load_string_key_table(STRINGKEYTABLE_FILE)

    # 3. Master item groups ----------------------------------------------------
    group_names, parent_of = parse_master_groups(ITEMGROUPINFO_FILE)

    # 4. Scan resource items ---------------------------------------------------
    group_items = scan_resource_folder(RESOURCE_FOLDER)

    # 5. Generate Excel per language ------------------------------------------
    for idx, (code, lang_tbl) in enumerate(lang_tables.items(), 1):
        log.info("(%d/%d) Processing language: %s", idx, len(lang_tables), code.upper())

        rows = build_rows_for_language(
            code,
            group_names,
            parent_of,
            group_items,
            lang_tbl,
            id_tbl,
        )

        wb = Workbook()
        wb.remove(wb.active)
        write_sheet(wb, code.upper(), rows)

        out_path = OUTPUT_FOLDER / f"Item_LQA_{code.upper()}.xlsx"
        wb.save(out_path)
        log.info("Excel saved: %s", out_path)

    log.info("All done – processed %d language(s).", len(lang_tables))


if __name__ == "__main__":
    main()
