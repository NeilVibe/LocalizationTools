import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from lxml import etree

# --- XML Utilities ---

def fix_bad_entities(xml_text):
    # Replace & with &amp; unless it's a known entity
    return re.sub(r'&(?!lt;|gt;|amp;|apos;|quot;)', '&amp;', xml_text)

def get_all_xml_files(input_folder):
    xml_files = []
    for dirpath, _, filenames in os.walk(input_folder):
        for file in filenames:
            if file.lower().endswith(".xml"):
                xml_files.append(os.path.join(dirpath, file))
    return xml_files

def parse_xml_file(file_path):
    try:
        txt = open(file_path, encoding='utf-8').read()
    except Exception as e:
        logging.error(f"Error reading {file_path!r}: {e}")
        return None
    txt = fix_bad_entities(txt)
    wrapped = "<ROOT>\n" + txt + "\n</ROOT>"
    rec_parser = etree.XMLParser(recover=True)
    try:
        recovered = etree.fromstring(wrapped.encode('utf-8'), parser=rec_parser)
    except etree.XMLSyntaxError as e:
        logging.error(f"Fatal parse error (recover mode) on {file_path!r}: {e}")
        return None
    strict_parser = etree.XMLParser(recover=False)
    blob = etree.tostring(recovered, encoding='utf-8')
    try:
        return etree.fromstring(blob, parser=strict_parser)
    except etree.XMLSyntaxError:
        return recovered

def build_key_and_translation_maps(xml_folder):
    """
    Recursively parse all XML files in xml_folder.
    - key_map: strkey.lower() -> key (from <StringKeyMap Key=... StrKey=...>)
    - translation_map: Korean string (StrOrigin) -> English translation (Str)
    """
    key_map = {}
    translation_map = {}
    xml_files = get_all_xml_files(xml_folder)
    for xml_file in xml_files:
        root = parse_xml_file(xml_file)
        if root is None:
            continue
        # StringKeyMap for key mapping
        for elem in root.iter():
            if elem.tag == "StringKeyMap":
                key = elem.attrib.get("Key")
                strkey = elem.attrib.get("StrKey")
                if key and strkey:
                    key_map[strkey.strip().lower()] = key
        # LocStr for translation mapping
        for elem in root.iter():
            if elem.tag == "LocStr":
                kr = elem.attrib.get("StrOrigin", "").strip()
                en = elem.attrib.get("Str", "").strip()
                if kr and en:
                    translation_map[kr] = en
    return key_map, translation_map

# --- Structure Parsing (NEW FORMAT, REVERSED CHAPTER LOGIC) ---

def parse_text_structure(filepath):
    """
    Parses the structure with reversed chapter logic:
    - Everything between two <...Group="..."> lines is a chapter, and the group/chapter name is taken from the NEXT such line.
    Returns a list of chapters, each with group, quests, missions, and submissions.
    """
    # Patterns
    chapter_tag_re = re.compile(r'^<([^ >]+)(?: Group="([^"]+)")?>')
    sep_re = re.compile(r'^-+$')
    quest_re = re.compile(r'^([A-Za-z0-9_]+)$')
    quest_kr_re = re.compile(r'^\s*-\s*(.+)$')
    mission_block_re = re.compile(r'^\s*Mission\s*$')
    mission_kr_re = re.compile(r'^\s*-\s*(.+)$')
    mission_strkey_re = re.compile(r'^\s*-\s*([A-Za-z0-9_]+)\s*$')
    submission_block_re = re.compile(r'^\s*SubMission\s*$')
    submission_kr_re = re.compile(r'^\s*-\s*(.+)$')

    # Read all lines
    with open(filepath, encoding='utf-8') as f:
        lines = [line.rstrip('\n') for line in f]

    # Find all chapter tag lines and their indices
    chapter_tags = []
    for idx, line in enumerate(lines):
        m = chapter_tag_re.match(line)
        if m:
            chapter_tags.append((idx, m.group(1), m.group(2) or ""))

    # Add a dummy "end" at the end for easier parsing
    chapter_tags.append((len(lines), None, None))

    structure = []

    # For each chapter block (between two <...Group=...> lines)
    for i in range(len(chapter_tags) - 1):
        start_idx = chapter_tags[i][0] + 1  # Start after the <...Group=...> line
        end_idx = chapter_tags[i+1][0]      # Up to (but not including) the next <...Group=...> line

        # The group/chapter name is from the NEXT tag
        chapter_name = chapter_tags[i+1][1]
        group = chapter_tags[i+1][2]

        # If no chapter_name (e.g. first block before any <...Group=...>), skip
        if not chapter_name:
            continue

        # Parse the block
        block_lines = lines[start_idx:end_idx]
        block_structure = parse_chapter_block(block_lines)
        if block_structure:
            block_structure["chapter"] = chapter_name
            block_structure["group"] = group
            structure.append(block_structure)

    return structure

def parse_chapter_block(lines):
    """
    Parses a block of lines representing a chapter (without the <...Group=...> line).
    Returns a dict: {"chapter":..., "group":..., "quests":[...]}
    """
    quest_re = re.compile(r'^([A-Za-z0-9_]+)$')
    quest_kr_re = re.compile(r'^\s*-\s*(.+)$')
    mission_block_re = re.compile(r'^\s*Mission\s*$')
    mission_kr_re = re.compile(r'^\s*-\s*(.+)$')
    mission_strkey_re = re.compile(r'^\s*-\s*([A-Za-z0-9_]+)\s*$')
    submission_block_re = re.compile(r'^\s*SubMission\s*$')
    submission_kr_re = re.compile(r'^\s*-\s*(.+)$')
    sep_re = re.compile(r'^-+$')

    quests = []
    current_quest = None
    current_mission = None
    in_mission_block = False
    in_submission_block = False

    idx = 0
    while idx < len(lines):
        line = lines[idx]
        if not line.strip() or sep_re.match(line):
            idx += 1
            continue

        # Quest line (not indented, not a separator)
        if not line.startswith(' ') and quest_re.match(line):
            quest_strkey = line.strip()
            # Next line is likely the Korean name
            quest_kr = ""
            if idx + 1 < len(lines):
                m = quest_kr_re.match(lines[idx + 1])
                if m:
                    quest_kr = m.group(1).strip()
                    idx += 1
            current_quest = {
                "quest_strkey": quest_strkey,
                "quest_kr": quest_kr,
                "missions": []
            }
            quests.append(current_quest)
            idx += 1
            continue

        # Mission block (expect two lines after "Mission": Korean name, then stringkey)
        if mission_block_re.match(line):
            in_mission_block = True
            in_submission_block = False
            idx += 1
            while idx + 1 < len(lines):
                line_kr = lines[idx]
                line_key = lines[idx + 1]
                m_kr = mission_kr_re.match(line_kr)
                m_key = mission_strkey_re.match(line_key)
                if m_kr and m_key:
                    mission_kr = m_kr.group(1).strip()
                    mission_strkey = m_key.group(1).strip()
                    current_mission = {
                        "mission_strkey": mission_strkey,
                        "mission_kr": mission_kr,
                        "submissions": []
                    }
                    if current_quest is not None:
                        current_quest["missions"].append(current_mission)
                    idx += 2
                else:
                    break
            continue

        # Submission block
        if submission_block_re.match(line):
            in_submission_block = True
            in_mission_block = False
            idx += 1
            continue

        # Submission Korean name (in submission block)
        if in_submission_block and submission_kr_re.match(line):
            submission_kr = submission_kr_re.match(line).group(1).strip()
            if current_mission is not None:
                current_mission["submissions"].append({
                    "submission_kr": submission_kr
                })
            idx += 1
            continue

        idx += 1

    if not quests:
        return None
    return {"quests": quests}

# --- Text Output with Keys and Translations ---

def process_and_write_text(input_path, key_map, translation_map, output_path):
    """
    Reads the input text file, and writes a new file with keys and translations next to mission/submission names.
    """
    structure = parse_text_structure(input_path)

    # Build mission key map: mission_strkey.lower() -> key
    mission_key_map = {}
    for chapter in structure:
        for quest in chapter["quests"]:
            for mission in quest["missions"]:
                mission_strkey_lc = mission["mission_strkey"].strip().lower()
                mission_key = key_map.get(mission_strkey_lc)
                mission_key_map[mission_strkey_lc] = mission_key

    # For submissions, assign key as "missionkey idx"
    submission_key_map = {}
    for chapter in structure:
        for quest in chapter["quests"]:
            for mission in quest["missions"]:
                mission_strkey_lc = mission["mission_strkey"].strip().lower()
                mission_key = mission_key_map.get(mission_strkey_lc)
                for idx, sub in enumerate(mission["submissions"]):
                    submission_key = f"{mission_key} {idx}" if mission_key else "?"
                    submission_key_map[(mission_strkey_lc, sub["submission_kr"])] = submission_key

    # Now, output in the new format, with keys and translations
    with open(output_path, 'w', encoding='utf-8') as outfile:
        for chapter in structure:
            outfile.write(f"<{chapter['chapter']}" + (f' Group="{chapter["group"]}"' if chapter["group"] else "") + ">\n")
            outfile.write("-" * 80 + "\n\n")
            for quest in chapter["quests"]:
                outfile.write(f"{quest['quest_strkey']}\n")
                quest_kr = quest.get("quest_kr", "")
                quest_tr = translation_map.get(quest_kr, "")
                if quest_kr:
                    outfile.write(f"  - {quest_kr}" + (f" ({quest_tr})" if quest_tr else "") + "\n")
                for mission in quest["missions"]:
                    outfile.write("  Mission\n")
                    mission_strkey = mission["mission_strkey"]
                    mission_kr = mission["mission_kr"]
                    mission_tr = translation_map.get(mission_kr, "")
                    mission_key = mission_key_map.get(mission_strkey.strip().lower(), "?")
                    # Write mission name, translation, stringkey, and key
                    outfile.write(f"    - {mission_kr}" + (f" ({mission_tr})" if mission_tr else "") + f" [{mission_strkey}] (Key = {mission_key})\n")
                    if mission["submissions"]:
                        outfile.write("    SubMission\n")
                        for idx, sub in enumerate(mission["submissions"]):
                            sub_kr = sub["submission_kr"]
                            sub_tr = translation_map.get(sub_kr, "")
                            submission_key = submission_key_map.get((mission_strkey.strip().lower(), sub_kr), "?")
                            outfile.write(f"      - {sub_kr}" + (f" ({sub_tr})" if sub_tr else "") + f" (Key = {submission_key})\n")
                outfile.write("\n")
            outfile.write("\n")

# --- Excel Output ---

def flatten_structure_for_excel(structure, key_map, translation_map):
    """
    Flattens the structure for Excel output.
    Returns a list of rows:
    [chapter, group, quest_strkey, quest_kr, quest_tr, mission_strkey, mission_kr, mission_tr, mission_key, submission_kr, submission_tr, submission_key]
    """
    rows = []
    for chapter in structure:
        chapter_name = chapter["chapter"]
        group = chapter.get("group", "")
        for quest in chapter["quests"]:
            quest_strkey = quest["quest_strkey"]
            quest_kr = quest.get("quest_kr", "")
            quest_tr = translation_map.get(quest_kr, "")
            for mission in quest["missions"]:
                mission_strkey = mission["mission_strkey"]
                mission_kr = mission["mission_kr"]
                mission_tr = translation_map.get(mission_kr, "")
                mission_key = key_map.get(mission_strkey.strip().lower(), "")
                if not mission["submissions"]:
                    rows.append([
                        chapter_name, group, quest_strkey, quest_kr, quest_tr,
                        mission_strkey, mission_kr, mission_tr, mission_key,
                        "", "", ""
                    ])
                for idx, sub in enumerate(mission["submissions"]):
                    sub_kr = sub["submission_kr"]
                    sub_tr = translation_map.get(sub_kr, "")
                    submission_key = f"{mission_key} {idx}" if mission_key else ""
                    rows.append([
                        chapter_name, group, quest_strkey, quest_kr, quest_tr,
                        mission_strkey, mission_kr, mission_tr, mission_key,
                        sub_kr, sub_tr, submission_key
                    ])
    return rows

def save_structure_to_excel(structure, key_map, translation_map, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Missions"

    # Header (added QUEST TRANSLATED NAME)
    headers = [
        "CHAPTER", "GROUP", "QUEST STRKEY", "QUEST NAME (KR)", "QUEST NAME (EN)",
        "MISSION STRKEY", "MISSION NAME (KR)", "MISSION NAME (EN)",
        "MISSION KEY", "SUB MISSION NAME (KR)", "SUB MISSION NAME (EN)", "SUB MISSION KEY"
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    rows = flatten_structure_for_excel(structure, key_map, translation_map)
    for row in rows:
        ws.append(row)

    # Adjust column widths
    widths = [20, 15, 35, 35, 35, 30, 35, 35, 15, 35, 35, 15]
    for col, width in zip(range(1, len(headers)+1), widths):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)

# --- GUI LOGIC ---

class MissionKeyMapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mission Key Mapper (Renewed)")

        frm = tk.Frame(root, padx=10, pady=10)
        frm.pack()

        # Text file (mission structure)
        tk.Label(frm, text="Mission Structure File:").grid(row=0, column=0, sticky="e")
        self.entry_text_file = tk.Entry(frm, width=50)
        self.entry_text_file.grid(row=0, column=1)
        tk.Button(frm, text="Browse...", command=self.browse_text_file).grid(row=0, column=2)

        # XML folder (for keys and translations)
        tk.Label(frm, text="XML Data Folder:").grid(row=1, column=0, sticky="e")
        self.entry_xml_folder = tk.Entry(frm, width=50)
        self.entry_xml_folder.grid(row=1, column=1)
        tk.Button(frm, text="Browse...", command=self.browse_xml_folder).grid(row=1, column=2)

        # Process buttons
        tk.Button(frm, text="Process and Save Text Output", command=self.run_processing, bg="#4CAF50", fg="white").grid(row=2, column=0, columnspan=3, pady=10)
        tk.Button(frm, text="Text to Excel", command=self.text_to_excel, bg="#2196F3", fg="white").grid(row=3, column=0, columnspan=3, pady=10)

    def browse_text_file(self):
        path = filedialog.askopenfilename(title="Select Mission Structure File", filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        if path:
            self.entry_text_file.delete(0, tk.END)
            self.entry_text_file.insert(0, path)

    def browse_xml_folder(self):
        path = filedialog.askdirectory(title="Select XML Data Folder")
        if path:
            self.entry_xml_folder.delete(0, tk.END)
            self.entry_xml_folder.insert(0, path)

    def run_processing(self):
        text_file = self.entry_text_file.get()
        xml_folder = self.entry_xml_folder.get()
        if not text_file or not xml_folder:
            messagebox.showerror("Error", "Please select both the mission structure file and the XML data folder.")
            return
        try:
            key_map, translation_map = build_key_and_translation_maps(xml_folder)
            output_path = filedialog.asksaveasfilename(title="Save Output File", defaultextension=".txt", initialfile="output_with_keys.txt", filetypes=[("Text Files", "*.txt")])
            if not output_path:
                return
            process_and_write_text(text_file, key_map, translation_map, output_path)
            messagebox.showinfo("Success", f"Processing complete!\nOutput saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

    def text_to_excel(self):
        text_file = self.entry_text_file.get()
        xml_folder = self.entry_xml_folder.get()
        if not text_file or not xml_folder:
            messagebox.showerror("Error", "Please select both the mission structure file and the XML data folder.")
            return
        try:
            key_map, translation_map = build_key_and_translation_maps(xml_folder)
            structure = parse_text_structure(text_file)
            output_path = filedialog.asksaveasfilename(title="Save Excel File As", defaultextension=".xlsx", initialfile="missions.xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not output_path:
                return
            save_structure_to_excel(structure, key_map, translation_map, output_path)
            messagebox.showinfo("Success", f"Excel file created:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MissionKeyMapperApp(root)
    root.mainloop()