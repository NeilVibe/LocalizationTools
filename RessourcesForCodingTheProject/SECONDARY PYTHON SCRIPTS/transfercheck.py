#!/usr/bin/env python3
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
import openpyxl

# ----------------- HELPERS -----------------
def normalize_text(txt):
    import html
    if not txt:
        return ""
    txt = html.unescape(str(txt))
    txt = re.sub(r'\s+', ' ', txt.strip())
    return txt

def robust_parse_xml(path):
    try:
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        tree = etree.parse(path, parser)
        return tree, tree.getroot()
    except Exception as e:
        print(f"[WARN] Cannot parse {path}: {e}")
        return None, None

def ask_logic_choice():
    """Show a simple button-based choice dialog for logic selection."""
    choice_win = tk.Toplevel()
    choice_win.title("Logic Choice")
    choice_win.geometry("250x120")
    choice_win.resizable(False, False)
    choice_var = tk.StringVar(value="")

    tk.Label(choice_win, text="Select logic to check:", pady=10).pack()

    def select_choice(val):
        choice_var.set(val)
        choice_win.destroy()

    tk.Button(choice_win, text="KR+ID", width=15, command=lambda: select_choice("KR+ID")).pack(pady=5)
    tk.Button(choice_win, text="KR ONLY", width=15, command=lambda: select_choice("KR ONLY")).pack(pady=5)

    choice_win.grab_set()
    choice_win.wait_window()
    return choice_var.get()

# ----------------- MAIN FUNCTION -----------------
def transfer_check():
    root = tk.Tk()
    root.withdraw()

    excel_path = filedialog.askopenfilename(title="Select Excel File",
                                            filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        return

    xml_path = filedialog.askopenfilename(title="Select Target XML File",
                                          filetypes=[("XML Files", "*.xml")])
    if not xml_path:
        return

    logic_choice = ask_logic_choice()
    if not logic_choice:
        return

    # Load XML into memory
    tree, xml_root = robust_parse_xml(xml_path)
    if tree is None:
        messagebox.showerror("Error", "Failed to parse XML file.")
        return

    # Build lookup maps from XML
    str_map = {}       # (StringId, StrOrigin) -> Str
    kr_only_map = {}   # StrOrigin -> Str
    desc_map = {}      # (StringId, DescOrigin) -> Desc

    for loc in xml_root.iter("LocStr"):
        sid = normalize_text(loc.get("StringId", ""))
        str_origin = normalize_text(loc.get("StrOrigin", ""))
        desc_origin = normalize_text(loc.get("DescOrigin", ""))
        str_val = normalize_text(loc.get("Str", ""))
        desc_val = normalize_text(loc.get("Desc", ""))

        if sid and str_origin:
            str_map[(sid, str_origin)] = str_val
        if str_origin:
            kr_only_map[str_origin] = str_val
        if sid and desc_origin:
            desc_map[(sid, desc_origin)] = desc_val

    # Load Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    transferable = []
    non_transferable = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        str_origin = normalize_text(row[0]) if len(row) > 0 else ""
        str_val    = normalize_text(row[1]) if len(row) > 1 else ""
        string_id  = normalize_text(row[2]) if len(row) > 2 else ""
        desc_origin = normalize_text(row[3]) if len(row) > 3 else ""
        desc_val    = normalize_text(row[4]) if len(row) > 4 else ""

        can_transfer = False

        # Check STR logic
        if logic_choice == "KR+ID" and string_id and str_origin:
            if (string_id, str_origin) in str_map:
                can_transfer = True
        elif logic_choice == "KR ONLY" and str_origin:
            if str_origin in kr_only_map:
                can_transfer = True

        # Check DESC logic (always Desc+ID)
        if string_id and desc_origin:
            if (string_id, desc_origin) in desc_map:
                can_transfer = True

        if can_transfer:
            transferable.append((str_origin, str_val, string_id, desc_origin, desc_val))
        else:
            non_transferable.append((str_origin, str_val, string_id, desc_origin, desc_val))

    # Save outputs
    base_dir = os.path.dirname(excel_path)
    out_transferable = os.path.join(base_dir, "transferable.xlsx")
    out_nontransferable = os.path.join(base_dir, "non_transferable.xlsx")

    def save_excel(path, rows):
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(["StrOrigin", "Str", "StringId", "DescOrigin", "Desc"])
        for r in rows:
            ws_out.append(r)
        wb_out.save(path)

    save_excel(out_transferable, transferable)
    save_excel(out_nontransferable, non_transferable)

    messagebox.showinfo("Transfer Check Completed",
                        f"Transferable rows: {len(transferable)}\n"
                        f"Non-transferable rows: {len(non_transferable)}\n\n"
                        f"Saved to:\n{out_transferable}\n{out_nontransferable}")

if __name__ == "__main__":
    transfer_check()