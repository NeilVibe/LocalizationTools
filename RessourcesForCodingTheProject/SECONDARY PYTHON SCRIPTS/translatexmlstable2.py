#!/usr/bin/env python3
import os
import re
import time
import html
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
import threading
import openpyxl

# ---- GLOBALS ----
xml_folder_path = None
xml_file_path   = None
tmx_file_path   = None

btn_xml_folder     = None
btn_tmx_file       = None
btn_simple         = None
btn_two            = None
btn_excel_simple   = None
btn_excel_2step    = None
btn_check          = None
btn_insert_excel   = None
status_label       = None
info_panel         = None
mode_var           = None

# ---- CORE NORMALIZATION FUNCTION ----
def normalize_text(txt):
    """
    Ensures consistent text normalization:
    1. Unescape HTML entities (&lt; → <, &amp; → &, etc.)
    2. Strip leading/trailing whitespace
    3. Collapse all internal whitespace (spaces, tabs, newlines) to single space
    
    This guarantees TMX and XML text are IDENTICAL for matching.
    """
    if not txt:
        return ""
    # Unescape HTML entities
    txt = html.unescape(str(txt))
    # Collapse all whitespace to single space and strip
    txt = re.sub(r'\s+', ' ', txt.strip())
    return txt

# ---- CRITICAL: TMX PREPROCESSING (FROM OLD CODE) ----
def preprocess_tmx_content(raw_content):
    """
    Pre-process TMX content BEFORE any XML parsing.
    Replace newlines inside <seg> tags with &lt;br/&gt;
    This ensures lxml treats <br/> as ESCAPED TEXT, not XML elements.
    """
    def replace_in_seg(match):
        seg_content = match.group(1)
        # Replace newlines with &lt;br/&gt;
        cleaned = seg_content.replace('\n', '&lt;br/&gt;')
        cleaned = cleaned.replace('\r\n', '&lt;br/&gt;')
        return f'<seg>{cleaned}</seg>'
    
    # Apply regex to find and clean all <seg> contents
    cleaned_content = re.sub(
        r'<seg>(.*?)</seg>', 
        replace_in_seg, 
        raw_content, 
        flags=re.DOTALL
    )
    
    return cleaned_content

# ---- HELPERS ----
def get_all_xml_files(folder):
    files = []
    for dp, _, fns in os.walk(folder):
        for fn in fns:
            if fn.lower().endswith(".xml"):
                files.append(os.path.join(dp, fn))
    return files

def get_xml_files():
    if mode_var.get() == "folder":
        return get_all_xml_files(xml_folder_path) if xml_folder_path else []
    else:
        return [xml_file_path] if xml_file_path else []

def parse_tmx(tmx_path):
    """
    Parse TMX and return two dictionaries with NORMALIZED text:
        translation_units – normal strings (context, kr, en)
        desc_units        – description strings (context, kr, en)
    
    A TU is a "description unit" if Korean starts with '&desc;' or '&amp;desc;'.
    ALL TEXT IS NORMALIZED via normalize_text() before storage.
    
    CRITICAL: Preprocesses TMX as text BEFORE parsing to handle <br/> correctly.
    """
    print(f"[TMX] Parsing TMX file: {tmx_path}")
    
    # ✅ READ TMX AS TEXT AND PREPROCESS BEFORE PARSING
    try:
        with open(tmx_path, "r", encoding="utf-8") as f:
            raw_content = f.read()
    except Exception as e:
        raise RuntimeError(f"Failed to read TMX file: {e}")
    
    # ✅ PREPROCESS: Replace <br/> with &lt;br/&gt; BEFORE XML parsing
    preprocessed_content = preprocess_tmx_content(raw_content)
    
    # ✅ NOW PARSE THE PREPROCESSED CONTENT
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                             no_network=True, recover=True)
    try:
        tree = etree.fromstring(preprocessed_content.encode('utf-8'), parser=parser)
    except Exception as e:
        raise RuntimeError(f"Failed to parse TMX: {e}")
    
    body = tree.find("body")
    if body is None:
        raise RuntimeError("TMX missing <body>")

    translation_units = []
    desc_units        = []

    for tu in body.findall("tu"):
        ctx = None
        kr = en = None
        desc_ko = None

        # Extract context
        for prop in tu.findall("prop"):
            if prop.get("type") == "x-context" and prop.text:
                ctx = prop.text.strip()
                break
        if not ctx:
            continue

        # Extract languages with NORMALIZATION
        for tuv in tu.findall("tuv"):
            lang = (tuv.get("{http://www.w3.org/XML/1998/namespace}lang")
                    or tuv.get("xml:lang")
                    or tuv.get("lang")
                    or "").lower()
            seg = tuv.find("seg")
            if seg is None:
                continue
            
            # ✅ NOW seg.text contains FULL TEXT with &lt;br/&gt; as escaped text
            raw_txt = seg.text if seg.text else ""
            
            # Check if description marker (before normalization)
            is_desc = raw_txt.strip().lower().startswith("&desc;") or raw_txt.strip().lower().startswith("&amp;desc;")
            
            # CRITICAL: Normalize text (unescapes &lt;br/&gt; to <br/>)
            txt = normalize_text(raw_txt)
            
            if lang.startswith("ko"):
                if is_desc:
                    desc_ko = txt
                else:
                    kr = txt
            elif lang.startswith("en"):
                en = txt

        # Classification
        if desc_ko:
            desc_units.append({'context': ctx, 'kr': desc_ko, 'en': en})
        elif kr or en:
            translation_units.append({'context': ctx, 'kr': kr, 'en': en})

    print(f"[TMX] Found {len(translation_units)} normal units and {len(desc_units)} description units.")
    return translation_units, desc_units

def robust_parse_xml(path):
    """Parse XML file with recovery mode."""
    try:
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        tree = etree.parse(path, parser)
        return tree, tree.getroot()
    except Exception as e:
        print(f"[WARN] Cannot parse {path}: {e}")
        return None, None

# ---- SIMPLE KR+ID TRANSLATE ----
def simple_translate_krid():
    global xml_folder_path, xml_file_path, tmx_file_path, status_label, mode_var

    if not tmx_file_path or not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select XML and TMX file/folder.")
        return

    status_label.config(text="Parsing TMX file (simple)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    # Build dictionaries with NORMALIZED keys
    perfect = {
        (u['context'], u['kr']): u['en']
        for u in tmx_trans if u.get('kr') and u.get('en')
    }
    perfect_desc = {
        (u['context'], u['kr']): u['en']
        for u in tmx_desc if u.get('kr') and u.get('en')
    }

    xml_files = get_xml_files()
    updated_str = updated_desc = processed = 0
    total_locs = sum(
        sum(1 for line in open(p, encoding='utf-8', errors='ignore') if "<LocStr" in line)
        for p in xml_files
    )
    t0 = time.time()

    for f_idx, path in enumerate(xml_files, 1):
        tree, root = robust_parse_xml(path)
        if tree is None:
            continue
        
        changed = False

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()
            # CRITICAL: Normalize XML attribute values
            origin = normalize_text(loc.get("StrOrigin", ""))
            desc_origin = normalize_text(loc.get("DescOrigin", ""))
            old_str = loc.get("Str", "")
            old_desc = loc.get("Desc", "")

            # Match Str
            key = (sid, origin)
            if key in perfect:
                new_str = perfect[key]
                if new_str and new_str != old_str:
                    loc.set("Str", new_str)
                    updated_str += 1
                    changed = True

            # Match Desc
            key_desc = (sid, desc_origin)
            if key_desc in perfect_desc:
                new_desc = perfect_desc[key_desc]
                if new_desc and new_desc != old_desc:
                    loc.set("Desc", new_desc)
                    updated_desc += 1
                    changed = True

            processed += 1

        if changed:
            tree.write(path, encoding="utf-8", xml_declaration=False, pretty_print=True)

        # Progress
        pct = 100.0 * processed / total_locs if total_locs else 0
        elapsed = time.time() - t0
        eta = (elapsed / processed) * (total_locs - processed) if processed else 0
        print(f"[PROGRESS] {processed}/{total_locs} ({pct:.2f}%) | Files {f_idx}/{len(xml_files)} | ETA {eta:.1f}s")

    status_label.config(text=f"Done! Updated Str:{updated_str}  Desc:{updated_desc}")
    messagebox.showinfo("Done",
                        f"Updated {updated_str} Str\n"
                        f"Updated {updated_desc} Desc\n"
                        f"in {len(xml_files)} files.")

# ---- 2-STEP TRANSLATE ----
def two_step_match_krid_then_kr():
    def worker():
        global xml_folder_path, xml_file_path, tmx_file_path, status_label

        if not tmx_file_path or not (xml_folder_path or xml_file_path):
            messagebox.showerror("Error", "Please select XML and TMX file/folder.")
            return

        status_label.config(text="Parsing TMX file (2-step)...")
        try:
            tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse TMX: {e}")
            return

        # Build dictionaries with NORMALIZED keys
        perfect_str = {
            (u['context'], u['kr']): u['en']
            for u in tmx_trans if u.get('kr') and u.get('en')
        }
        perfect_desc = {
            (u['context'], u['kr']): u['en']
            for u in tmx_desc if u.get('kr') and u.get('en')
        }
        kr_only = {
            u['kr']: u['en']
            for u in tmx_trans if u.get('kr') and u.get('en')
        }

        xml_files = get_xml_files()
        total_locs = sum(
            sum(1 for line in open(p, encoding='utf-8', errors='ignore') if "<LocStr" in line)
            for p in xml_files
        )
        u1 = u2 = u_desc = processed = 0
        t0 = time.time()

        for f_idx, path in enumerate(xml_files, 1):
            tree, root = robust_parse_xml(path)
            if tree is None:
                continue
            
            changed = False

            for loc in root.iter("LocStr"):
                sid = loc.get("StringId", "").strip()
                # CRITICAL: Normalize XML attribute values
                origin = normalize_text(loc.get("StrOrigin", ""))
                desc_origin = normalize_text(loc.get("DescOrigin", ""))
                old_str = loc.get("Str", "")
                old_desc = loc.get("Desc", "")

                # STEP 1: Perfect match (ID+KR)
                key = (sid, origin)
                if key in perfect_str:
                    new_str = perfect_str[key]
                    if new_str and new_str != old_str:
                        loc.set("Str", new_str)
                        u1 += 1
                        changed = True

                # Match Desc
                key_desc = (sid, desc_origin)
                if key_desc in perfect_desc:
                    new_desc = perfect_desc[key_desc]
                    if new_desc and new_desc != old_desc:
                        loc.set("Desc", new_desc)
                        u_desc += 1
                        changed = True

                # STEP 2: KR-only match (fallback)
                if key not in perfect_str and origin in kr_only:
                    new_str = kr_only[origin]
                    if new_str and new_str != old_str:
                        loc.set("Str", new_str)
                        u2 += 1
                        changed = True

                processed += 1

            if changed:
                tree.write(path, encoding="utf-8", xml_declaration=False, pretty_print=True)

            # Progress
            pct = 100.0 * processed / total_locs if total_locs else 0
            elapsed = time.time() - t0
            eta = (elapsed / processed) * (total_locs - processed) if processed else 0
            print(f"[PROGRESS] {processed}/{total_locs} ({pct:.2f}%) | Files {f_idx}/{len(xml_files)} | ETA {eta:.1f}s")

        status_label.config(text=f"Done! Perfect:{u1}  KR-only:{u2}  Desc:{u_desc}")
        messagebox.showinfo(
            "Done",
            f"Perfect (ID+KR): {u1}\n"
            f"KR-only: {u2}\n"
            f"Desc updated: {u_desc}\n"
            f"Files overwritten: {len(xml_files)}\n"
            f"Time: {time.time() - t0:.1f}s"
        )

    threading.Thread(target=worker, daemon=True).start()

# ---- EXCEL TRANSLATE – SIMPLE ----
def excel_translate_simple_krid():
    global tmx_file_path, status_label
    if not tmx_file_path:
        messagebox.showerror("Error", "Please select a TMX file first.")
        return

    status_label.config(text="Parsing TMX for Excel (Simple KR+ID)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    units = tmx_trans + tmx_desc
    perfect_dict = {(u['context'], u['kr']): u['en']
                    for u in units if u.get('context') and u.get('kr') and u.get('en')}

    if not perfect_dict:
        messagebox.showerror("Error", "No usable KR+ID translation units found in TMX.")
        return

    excel_path = filedialog.askopenfilename(title="Select Excel File",
                                            filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        status_label.config(text="Excel translation cancelled.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        updated = total = 0
        for row_idx in range(1, ws.max_row + 1):
            kr_val = ws.cell(row=row_idx, column=1).value
            strid_val = ws.cell(row=row_idx, column=3).value
            if not kr_val or not strid_val:
                continue
            total += 1

            # CRITICAL: Normalize Excel values
            key = (normalize_text(strid_val), normalize_text(kr_val))
            old_translation = str(ws.cell(row=row_idx, column=2).value or "").strip()

            if key in perfect_dict:
                new_translation = perfect_dict[key]
                if new_translation and new_translation != old_translation:
                    ws.cell(row=row_idx, column=2).value = new_translation
                    updated += 1

        wb.save(excel_path)
        status_label.config(text=f"Excel Simple completed: {updated}/{total} rows updated.")
        messagebox.showinfo("Excel Translate – Simple",
                            f"Rows processed: {total}\n"
                            f"Rows updated: {updated}\n"
                            f"File overwritten:\n{excel_path}")
        print(f"[INFO] Excel Simple: Processed {total}, Updated {updated}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel file: {e}")
        status_label.config(text="Excel translation failed.")

# ---- EXCEL TRANSLATE – 2-STEP ----
def excel_translate_2step_krid_then_kr():
    global tmx_file_path, status_label
    if not tmx_file_path:
        messagebox.showerror("Error", "Please select a TMX file first.")
        return

    status_label.config(text="Parsing TMX for Excel (2-STEP)...")
    try:
        tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return

    units = tmx_trans + tmx_desc
    perfect_dict = {(u['context'], u['kr']): u['en']
                    for u in units if u.get('context') and u.get('kr') and u.get('en')}
    kr_only_dict = {u['kr']: u['en']
                    for u in units if u.get('kr') and u.get('en')}

    if not (perfect_dict or kr_only_dict):
        messagebox.showerror("Error", "No usable translation units found in TMX.")
        return

    excel_path = filedialog.askopenfilename(title="Select Excel File",
                                            filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        status_label.config(text="Excel translation cancelled.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        upd_perfect = upd_kr = total = 0
        for row_idx in range(1, ws.max_row + 1):
            kr_val = ws.cell(row=row_idx, column=1).value
            strid_val = ws.cell(row=row_idx, column=3).value
            if not kr_val:
                continue
            total += 1

            # CRITICAL: Normalize Excel values
            kr_text = normalize_text(kr_val)
            strid_text = normalize_text(strid_val)
            old_translation = normalize_text(ws.cell(row=row_idx, column=2).value)

            updated_this_row = False

            # STEP 1: Perfect match (ID+KR)
            if strid_text and (strid_text, kr_text) in perfect_dict:
                new_translation = perfect_dict[(strid_text, kr_text)]
                if new_translation and new_translation != old_translation:
                    ws.cell(row=row_idx, column=2).value = new_translation
                    upd_perfect += 1
                    updated_this_row = True

            # STEP 2: KR-only match
            if not updated_this_row and kr_text in kr_only_dict:
                new_translation = kr_only_dict[kr_text]
                if new_translation and new_translation != old_translation:
                    ws.cell(row=row_idx, column=2).value = new_translation
                    upd_kr += 1

        wb.save(excel_path)
        status_label.config(text=f"Excel 2-STEP: Perfect {upd_perfect} + KR {upd_kr} / {total}")
        messagebox.showinfo("Excel Translate – 2-STEP",
                            f"Perfect (ID+KR) updated: {upd_perfect}\n"
                            f"KR-only updated: {upd_kr}\n"
                            f"Rows processed: {total}\n"
                            f"File overwritten:\n{excel_path}")
        print(f"[INFO] Excel 2-STEP: Processed {total}, Perfect updated {upd_perfect}, KR-only updated {upd_kr}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel file: {e}")
        status_label.config(text="Excel 2-STEP translation failed.")

# ---- INSERT FROM EXCEL ----
def insert_translations_from_excel():
    try:
        excel_path = filedialog.askopenfilename(title="Select Excel File",
                                                filetypes=[("Excel Files", "*.xlsx")])
        if not excel_path:
            return

        xml_path = filedialog.askopenfilename(title="Select XML File",
                                              filetypes=[("XML Files", "*.xml")])
        if not xml_path:
            return

        folder_path = filedialog.askdirectory(title="Select XML Folder for VoiceId Mapping")
        if not folder_path:
            return

        try:
            os.chmod(xml_path, 0o666)
        except:
            pass

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        excel_rows = []
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                excel_rows.append((str(row[0]).strip(), str(row[1]).strip()))
        print(f"[INFO] Loaded {len(excel_rows)} rows from Excel.")

        tree, root = robust_parse_xml(xml_path)
        if tree is None:
            messagebox.showerror("Error", "Failed to parse XML file.")
            return

        stringid_map = {}
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()
            if sid:
                stringid_map[sid.lower()] = loc
        print(f"[INFO] Found {len(stringid_map)} StringIds in target XML.")

        voiceid_to_stringid = {}
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        for dirpath, _, files in os.walk(folder_path):
            for fn in files:
                if fn.lower().endswith(".xml"):
                    try:
                        t = etree.parse(os.path.join(dirpath, fn), parser)
                        for loc in t.getroot().iter("LocStr"):
                            sid = loc.get("StringId", "").strip()
                            vid = loc.get("VoiceId", "").strip()
                            snd = loc.get("SoundEventName", "").strip()
                            if sid:
                                if vid:
                                    voiceid_to_stringid[vid.lower()] = sid
                                if snd:
                                    voiceid_to_stringid[snd.lower()] = sid
                    except:
                        pass
        print(f"[INFO] Built {len(voiceid_to_stringid)} VoiceId/SoundEventName mappings.")

        updated = 0
        for translation, id_or_voice in excel_rows:
            key = id_or_voice.lower()

            if key in stringid_map:
                loc = stringid_map[key]
                loc.set("Str", translation)
                updated += 1
            elif key in voiceid_to_stringid:
                sid = voiceid_to_stringid[key]
                if sid.lower() in stringid_map:
                    loc = stringid_map[sid.lower()]
                    loc.set("Str", translation)
                    updated += 1

        tree.write(xml_path, encoding="utf-8", xml_declaration=False, pretty_print=True)
        messagebox.showinfo("Done",
                            f"Inserted translations for {updated} entries.\n"
                            f"File overwritten:\n{xml_path}")
        print(f"[INFO] Updated {updated} entries in {xml_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed: {e}")

# ---- COVERAGE CHECK ----
def check_translation_coverage():
    global xml_folder_path, xml_file_path, status_label
    if not (xml_folder_path or xml_file_path):
        messagebox.showerror("Error", "Please select an XML file or folder first.")
        return
    status_label.config(text="Calculating translation coverage...")

    xml_files = get_xml_files()
    total = translated = untranslated = 0
    korean_re = re.compile(r'[\uac00-\ud7a3]')
    locstr_re = re.compile(r'<LocStr\b([^>]*)>', re.IGNORECASE | re.DOTALL)
    str_attr_re = re.compile(r'Str="([^"]*)"')

    for path in xml_files:
        try:
            with open(path, encoding="utf-8", errors="ignore") as f:
                content = f.read()
            for m in locstr_re.finditer(content):
                total += 1
                attrs = m.group(1)
                str_m = str_attr_re.search(attrs)
                val = (str_m.group(1).strip() if str_m else "")
                if not val or korean_re.search(val):
                    untranslated += 1
                else:
                    translated += 1
        except:
            pass

    coverage = (translated / total * 100) if total else 0.0
    msg = (f"Total LocStrs: {total}\nTranslated: {translated}\n"
           f"Untranslated: {untranslated}\nCoverage: {coverage:.1f}%"
           if coverage < 99.95 else
           f"Total LocStrs: {total}\nTranslated: {translated}\n"
           f"Untranslated: {untranslated}\nCoverage: {coverage:.3f}%")
    status_label.config(text=msg)
    messagebox.showinfo("Translation Coverage", msg)

# ---- GUI CALLBACKS ----
def upload_xml():
    global xml_folder_path, xml_file_path
    if mode_var.get() == "folder":
        folder = filedialog.askdirectory(title="Select XML Folder")
        if folder:
            xml_folder_path, xml_file_path = folder, None
            for p in get_all_xml_files(folder):
                try:
                    os.chmod(p, 0o666)
                except:
                    pass
            btn_xml_folder.config(bg="green")
            status_label.config(text="XML folder selected.")
    else:
        file = filedialog.askopenfilename(title="Select XML File",
                                          filetypes=[("XML Files", "*.xml")])
        if file:
            xml_file_path, xml_folder_path = file, None
            try:
                os.chmod(file, 0o666)
            except:
                pass
            btn_xml_folder.config(bg="green")
            status_label.config(text="XML file selected.")
    update_buttons()
    update_info_panel()

def upload_tmx_file():
    global tmx_file_path
    file = filedialog.askopenfilename(title="Select TMX File",
                                      filetypes=[("TMX Files", "*.tmx")])
    if file:
        tmx_file_path = file
        btn_tmx_file.config(bg="green")
        status_label.config(text="TMX file selected.")
    update_buttons()
    update_info_panel()

def update_buttons():
    if xml_folder_path or xml_file_path:
        btn_check.config(state="normal", bg="purple")
    else:
        btn_check.config(state="disabled", bg="grey")

    if tmx_file_path:
        btn_excel_simple.config(state="normal", bg="#32cd32")
        btn_excel_2step.config(state="normal", bg="#ffb347")
    else:
        btn_excel_simple.config(state="disabled", bg="grey")
        btn_excel_2step.config(state="disabled", bg="grey")

    if (xml_folder_path or xml_file_path) and tmx_file_path:
        btn_simple.config(state="normal", bg="orange")
        btn_two.config(state="normal", bg="#ffb347")
    else:
        btn_simple.config(state="disabled", bg="grey")
        btn_two.config(state="disabled", bg="grey")

def update_info_panel():
    info_panel.config(state="normal")
    info_panel.delete(1.0, tk.END)
    info_panel.insert(tk.END, "Current Selections:\n")
    info_panel.insert(tk.END, f"Mode       : {'Folder' if mode_var.get() == 'folder' else 'File'}\n")
    info_panel.insert(tk.END, f"XML Folder : {xml_folder_path if xml_folder_path else '[Not selected]'}\n")
    info_panel.insert(tk.END, f"XML File   : {xml_file_path if xml_file_path else '[Not selected]'}\n")
    info_panel.insert(tk.END, f"TMX File   : {tmx_file_path if tmx_file_path else '[Not selected]'}\n")
    info_panel.config(state="disabled")

def switch_mode():
    global xml_folder_path, xml_file_path
    xml_folder_path = None
    xml_file_path = None
    btn_xml_folder.config(bg="SystemButtonFace")
    update_buttons()
    update_info_panel()
    status_label.config(text=f"Switched to {'Folder' if mode_var.get() == 'folder' else 'File'} mode. Please select XML.")

# ---- MAIN GUI ----
def main():
    global btn_xml_folder, btn_tmx_file, btn_simple, btn_two
    global btn_excel_simple, btn_excel_2step, btn_check, btn_insert_excel
    global status_label, info_panel, mode_var

    root = tk.Tk()
    root.title("Batch XML Updater + Excel Helpers")
    root.geometry("700x780")

    mode_var = tk.StringVar(value="folder")
    mode_frame = tk.Frame(root)
    mode_frame.pack(pady=8)
    tk.Label(mode_frame, text="Mode:").pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="Folder Mode", variable=mode_var,
                   value="folder", command=switch_mode).pack(side="left", padx=4)
    tk.Radiobutton(mode_frame, text="File Mode", variable=mode_var,
                   value="file", command=switch_mode).pack(side="left", padx=4)

    btn_xml_folder = tk.Button(root, text="Upload XML Folder/File",
                               command=upload_xml, width=30, height=2)
    btn_xml_folder.pack(pady=8)

    btn_tmx_file = tk.Button(root, text="Upload TMX File",
                             command=upload_tmx_file, width=30, height=2)
    btn_tmx_file.pack(pady=8)

    btn_simple = tk.Button(root, text="KR+ID Simple Translate",
                           command=simple_translate_krid,
                           state="disabled", bg="grey", width=30, height=2)
    btn_simple.pack(pady=8)

    btn_two = tk.Button(root, text="2 STEP MATCH – KR+ID then KR",
                        command=two_step_match_krid_then_kr,
                        state="disabled", bg="grey", width=30, height=2)
    btn_two.pack(pady=8)

    btn_excel_simple = tk.Button(root,
                                 text="Excel Translate SIMPLE (KR+ID)",
                                 command=excel_translate_simple_krid,
                                 state="disabled", bg="grey", width=30, height=2)
    btn_excel_simple.pack(pady=8)

    btn_excel_2step = tk.Button(root,
                                text="Excel Translate 2-STEP (ID+KR ➜ KR)",
                                command=excel_translate_2step_krid_then_kr,
                                state="disabled", bg="grey", width=30, height=2)
    btn_excel_2step.pack(pady=8)

    btn_check = tk.Button(root, text="Check Translation Coverage",
                          command=check_translation_coverage,
                          state="disabled", bg="grey", width=30, height=2)
    btn_check.pack(pady=8)

    btn_insert_excel = tk.Button(root, text="Excel ➜ XML (StringID only)",
                                 command=insert_translations_from_excel,
                                 width=30, height=2, bg="#87ceeb")
    btn_insert_excel.pack(pady=8)

    status_label = tk.Label(root, text="Please upload XML and/or TMX.", fg="blue")
    status_label.pack(pady=12)

    info_panel = tk.Text(root, height=6, width=80, wrap="none",
                         state="disabled", bg="#f0f0f0")
    info_panel.pack(pady=8, padx=8, fill="x")

    update_info_panel()
    update_buttons()
    root.mainloop()

if __name__ == "__main__":
    main()