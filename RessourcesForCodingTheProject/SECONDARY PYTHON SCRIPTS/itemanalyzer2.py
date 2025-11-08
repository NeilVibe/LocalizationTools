#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import logging
from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURE YOUR PATHS HERE
# ─────────────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER          = r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\iteminfo"
ITEMGROUPINFO_FILE       = r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\iteminfo\itemgroupinfo.staticinfo.xml"
LANGUAGE_FOLDER          = r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
STRINGKEYTABLE_FILE      = r"F:\perforce\cd\mainline\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
OUTPUT_EXCEL             = os.path.join(os.getcwd(), "ItemData_Map.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("ItemExport")

# ─────────────────────────────────────────────────────────────────────────────
# XML UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
BAD_ENTITY_RE = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(text):
    return BAD_ENTITY_RE.sub("&amp;", text)

def parse_xml_file(path):
    """
    Read an XML file as utf-8, wrap in dummy <ROOT>, recover broken entities,
    then do a strict re‐parse. Returns the lxml root element.
    """
    log.debug(f"Parsing XML file: {path!r}")
    try:
        raw = open(path, "r", encoding="utf-8").read()
    except Exception:
        log.exception(f"Failed to read {path!r}")
        return None

    clean   = fix_bad_entities(raw)
    wrapped = "<ROOT>\n" + clean + "\n</ROOT>"

    # first pass: recover
    rec_parser = ET.XMLParser(recover=True)
    try:
        recovered = ET.fromstring(wrapped.encode("utf-8"), parser=rec_parser)
    except ET.XMLSyntaxError:
        log.exception(f"Recover‐mode XMLSyntaxError in {path}")
        return None

    # second pass: strict
    strict_parser = ET.XMLParser(recover=False)
    blob = ET.tostring(recovered, encoding="utf-8")
    try:
        final = ET.fromstring(blob, parser=strict_parser)
        return final
    except ET.XMLSyntaxError:
        log.warning(f"Strict re‐parse failed on {path}, using recovered tree")
        return recovered

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: LOAD ITEMGROUPINFO.MASTER → group_info + group_hierarchy
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_groups(path):
    log.info("Loading master ItemGroupInfo file")
    root = parse_xml_file(path)
    if root is None:
        sys.exit(1)

    group_info      = {}  # group_key -> group_name
    group_hierarchy = {}  # child_group_key -> parent_group_key

    for el in root.iter():
        tag = el.tag
        if tag == "ItemGroupInfo":
            sk = el.get("StrKey")
            gn = el.get("GroupName") or ""
            group_info[sk] = gn
            parent = el.getparent()
            if parent is not None and parent.tag == "ItemGroupInfo":
                pk = parent.get("StrKey")
                group_hierarchy[sk] = pk
                log.debug(f"Nested group: {sk} → parent {pk}")

        elif tag == "ChildGroupInfo":
            child = el.get("StrKey")
            parent = el.getparent()
            if parent is not None and parent.tag == "ItemGroupInfo":
                pk = parent.get("StrKey")
                group_hierarchy[child] = pk
                log.debug(f"ChildGroupInfo link: {child} → parent {pk}")

    log.info(f"Master groups loaded: {len(group_info)} groups, "
             f"{len(group_hierarchy)} parent→child relations")
    return group_info, group_hierarchy

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: SCAN RESOURCE FOLDER FOR EVERY ItemGroupInfo → ItemInfo
# ─────────────────────────────────────────────────────────────────────────────
def scan_resource_folder(folder):
    log.info("Scanning resource folder for items")
    group_items = {}  # group_key -> list of (item_strkey, item_name_original)

    for dp, dn, fnames in os.walk(folder):
        for fn in fnames:
            if not fn.lower().endswith(".xml"):
                continue
            full = os.path.join(dp, fn)
            root = parse_xml_file(full)
            if root is None:
                continue

            for group_el in root.iter("ItemGroupInfo"):
                gk = group_el.get("StrKey")
                if gk is None:
                    continue
                bucket = group_items.setdefault(gk, [])
                for item_el in group_el.iter("ItemInfo"):
                    isk   = item_el.get("StrKey")
                    iname = item_el.get("ItemName") or ""
                    if isk:
                        bucket.append((isk, iname))
                        log.debug(f"  {gk} → Item {isk!r} / name={iname!r}")

    total_groups = len(group_items)
    total_items  = sum(len(v) for v in group_items.values())
    log.info(f"Found {total_groups} groups with {total_items} total items")
    return group_items

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: LOAD ALL LANGUAGE FILES → code → { origin: (translation, id) }
# ─────────────────────────────────────────────────────────────────────────────
def parse_language_folder(folder):
    log.info("Scanning language folder for language files")
    languages = {}  # code -> table mapping StrOrigin -> (Str, StringId)
    for dp, dn, fnames in os.walk(folder):
        for fn in fnames:
            if not fn.lower().endswith(".xml"):
                continue
            name_no_ext = os.path.splitext(fn)[0]
            if name_no_ext.lower().startswith("languagedata_kor"):
                log.debug(f"Ignoring base language file: {fn}")
                continue
            if not name_no_ext.lower().startswith("languagedata_"):
                continue
            code = name_no_ext.split("_", 1)[1]
            path = os.path.join(dp, fn)
            log.info(f"Loading language file for '{code}' → {path}")
            root = parse_xml_file(path)
            if root is None:
                log.error(f"Failed to load language file: {path}")
                continue
            table = {}
            for el in root.iter("LocStr"):
                origin = el.get("StrOrigin")
                sid    = el.get("StringId")
                trans  = el.get("Str")
                if origin:
                    table[origin] = (trans or "", sid or "")
            languages[code] = table

    log.info(f"Language tables loaded: {len(languages)} languages")
    return languages

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3b: LOAD STRINGKEYTABLE → strkey_to_number
# ─────────────────────────────────────────────────────────────────────────────
def parse_string_key_table(path):
    log.info("Loading StringKeyTable file")
    root = parse_xml_file(path)
    if root is None:
        sys.exit(1)
    mapping = {}  # lowercase StrKey -> numeric Key
    for el in root.iter("StringKeyMap"):
        key = el.get("Key")
        sk  = el.get("StrKey")
        if sk and key:
            mapping[sk.lower()] = key
            log.debug(f"Mapping '{sk.lower()}' → {key}")
    log.info(f"StringKeyTable loaded: {len(mapping)} entries")
    return mapping

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: MERGE AND WRITE EXCEL (ONE SHEET PER LANGUAGE)
# ─────────────────────────────────────────────────────────────────────────────
def write_to_excel(group_info, group_hierarchy, group_items, languages, id_mapping, out_path):
    log.info(f"Writing results to Excel → {out_path}")
    wb = Workbook()
    wb.remove(wb.active)

    for code, lang_table in sorted(languages.items()):
        ws = wb.create_sheet(title=code)

        # HEADER
        headers = [
            "GroupKey",
            "GroupName",
            "ItemKey (name)",
            "ItemKey (number)",
            "ItemName(KOR)",
            f"ItemName({code})",
            "StringID"
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ROWS
        row_idx = 2
        for gk, items in sorted(group_items.items()):
            parent_key     = group_hierarchy.get(gk)
            group_display  = group_info.get(parent_key if parent_key else gk, "")
            for isk, kor_name in items:
                num = id_mapping.get(isk.lower(), "<MISSING>")
                trans, sid = lang_table.get(kor_name, ("<MISSING>", ""))
                ws.cell(row=row_idx, column=1, value=gk)
                ws.cell(row=row_idx, column=2, value=group_display)
                ws.cell(row=row_idx, column=3, value=isk)
                ws.cell(row=row_idx, column=4, value=num)
                ws.cell(row=row_idx, column=5, value=kor_name)
                ws.cell(row=row_idx, column=6, value=trans)
                ws.cell(row=row_idx, column=7, value=sid)
                row_idx += 1

        # AUTO-FILTER + FREEZE PANE
        ws.auto_filter.ref = f"A1:G{row_idx-1}"
        ws.freeze_panes     = "A2"

        # OPTIMAL COLUMN WIDTHS
        #   Col E (KOR name) at 30% of content width
        #   Col G (StringID) at 200% of content width
        #   All others at 70% of content width
        shrink_factors = {'E': 0.3, 'D': 2.0, 'G': 1.4}
        default_factor = 0.7
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = l
            orig_width = max_len + 2
            factor = shrink_factors.get(col_letter, default_factor)
            new_width = max(int(orig_width * factor), 5)
            ws.column_dimensions[col_letter].width = new_width

    wb.save(out_path)
    log.info("Excel file saved successfully.")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # 1) Master groups
    group_info, group_hierarchy = parse_master_groups(ITEMGROUPINFO_FILE)
    # 2) Resource items
    group_items                = scan_resource_folder(RESOURCE_FOLDER)
    # 3) Language tables
    languages                  = parse_language_folder(LANGUAGE_FOLDER)
    # 3b) StringKeyTable
    id_mapping                 = parse_string_key_table(STRINGKEYTABLE_FILE)
    # 4) Write Excel
    write_to_excel(group_info, group_hierarchy, group_items, languages, id_mapping, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()