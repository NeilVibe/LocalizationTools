import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import openpyxl
from openpyxl.styles import PatternFill
import os
import shutil
import tempfile

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f'{width}x{height}+{x}+{y}')

def clean_text(value):
    if value is None:
        return ""
    return str(value).replace('_x000D_', '').strip()

def create_temp_excel(original_file, suffix):
    temp_dir = tempfile.gettempdir()
    temp_name = f"concat_temp_{suffix}_{os.path.basename(original_file)}"
    temp_path = os.path.join(temp_dir, temp_name)
    shutil.copy2(original_file, temp_path)
    return temp_path

class SheetColumnSelector:
    def __init__(self, parent_frame, file_path, active_selectors):
        self.file_path = file_path
        self.active_selectors = active_selectors
        self.active_selectors.append(self)

        self.temp_file = create_temp_excel(file_path, "source")
        self.wb = openpyxl.load_workbook(self.temp_file, read_only=True)
        self.sheet_names = self.wb.sheetnames

        self.frame = ttk.LabelFrame(parent_frame, text=f"{os.path.basename(file_path)}")
        self.frame.pack(fill=tk.X, padx=5, pady=5, anchor="w")

        # Sheet selection
        ttk.Label(self.frame, text="Sheet:").pack(side=tk.LEFT, padx=(5, 5))
        self.sheet_var = tk.StringVar(value=self.sheet_names[0])
        self.sheet_combo = ttk.Combobox(self.frame, textvariable=self.sheet_var, values=self.sheet_names, state='readonly', width=20)
        self.sheet_combo.pack(side=tk.LEFT, padx=(5, 15))
        self.sheet_combo.bind("<<ComboboxSelected>>", self.update_columns)

        # Column selection (MULTIPLE)
        ttk.Label(self.frame, text="Columns:").pack(side=tk.LEFT)
        self.column_listbox = tk.Listbox(self.frame, selectmode=tk.MULTIPLE, height=5, exportselection=False)
        self.column_listbox.pack(side=tk.LEFT, padx=(5, 15))

        self.update_columns()

        # Delete button
        self.delete_button = ttk.Button(self.frame, text="Delete", command=self.remove_frame)
        self.delete_button.pack(side=tk.RIGHT, padx=5, pady=2)

    def update_columns(self, event=None):
        self.column_listbox.delete(0, tk.END)
        sheet = self.wb[self.sheet_var.get()]
        headers = [clean_text(cell.value) for cell in sheet[1]]
        for header in headers:
            self.column_listbox.insert(tk.END, header)

    def remove_frame(self):
        self.active_selectors.remove(self)
        self.frame.destroy()

    def get_selection(self):
        selected_indices = self.column_listbox.curselection()
        selected_columns = [self.column_listbox.get(i) for i in selected_indices]
        return {
            'file_path': self.file_path,
            'sheet': self.sheet_var.get(),
            'columns': selected_columns
        }

def start_concat(selections):
    try:
        # SAFEGUARD: Check if all selections have same number of columns
        col_counts = [len(sel['columns']) for sel in selections]
        if len(set(col_counts)) != 1:
            warn_msg = "WARNING: Not all selections have the same number of columns.\n\n"
            for sel in selections:
                warn_msg += f"{os.path.basename(sel['file_path'])} - {sel['sheet']} : {len(sel['columns'])} columns selected\n"
            warn_msg += "\nDo you still want to proceed?"
            if not messagebox.askyesno("Column Count Mismatch", warn_msg):
                return

        combined_wb = openpyxl.Workbook()
        combined_ws = combined_wb.active
        combined_ws.title = "Concatenated"

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        current_row = 1
        for idx, sel in enumerate(selections):
            temp_file = create_temp_excel(sel['file_path'], f"concat_{idx}")
            wb = openpyxl.load_workbook(temp_file, read_only=False)
            sheet = wb[sel['sheet']]

            col_indices = []
            for col_name in sel['columns']:
                found = False
                for i, cell in enumerate(sheet[1], start=1):
                    if clean_text(cell.value) == col_name:
                        col_indices.append(i)
                        found = True
                        break
                if not found:
                    messagebox.showerror("Error", f"Column '{col_name}' not found in sheet '{sel['sheet']}' of {os.path.basename(sel['file_path'])}")
                    wb.close()
                    os.remove(temp_file)
                    return

            start_row = current_row
            for r in range(2, sheet.max_row + 1):
                for c_idx, col_index in enumerate(col_indices, start=1):
                    value = clean_text(sheet.cell(row=r, column=col_index).value)
                    combined_ws.cell(row=current_row, column=c_idx).value = value
                current_row += 1

            # Highlight junction between concatenations
            if idx > 0:
                for c in range(1, len(col_indices) + 1):
                    combined_ws.cell(row=start_row, column=c).fill = yellow_fill
                    combined_ws.cell(row=start_row - 1, column=c).fill = yellow_fill

            wb.close()
            os.remove(temp_file)

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Concatenated File", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            combined_wb.save(save_path)
            messagebox.showinfo("Success", f"Concatenation completed!\nSaved to: {save_path}")
        else:
            messagebox.showwarning("Cancelled", "Save cancelled.")

    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

def create_main_gui():
    root = tk.Tk()
    root.title("Excel Multi-Column Concatenator")
    center_window(root, 900, 700)

    main_frame = ttk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    canvas = tk.Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill=tk.BOTH, expand=True)
    scrollbar.pack(side="right", fill="y")

    active_selectors = []

    def add_selector():
        file_paths = filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")])
        for file_path in file_paths:
            if file_path:
                SheetColumnSelector(scrollable_frame, file_path, active_selectors)

    def start_process():
        if not active_selectors:
            messagebox.showerror("Error", "No selections made.")
            return
        selections = [sel.get_selection() for sel in active_selectors]
        start_concat(selections)

    buttons_frame = ttk.Frame(root)
    buttons_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    add_button = ttk.Button(buttons_frame, text="Add Files", command=add_selector)
    add_button.pack(side=tk.LEFT, padx=5)

    start_button = ttk.Button(buttons_frame, text="Start Concatenation", command=start_process)
    start_button.pack(side=tk.RIGHT, padx=5)

    root.mainloop()

if __name__ == "__main__":
    create_main_gui()