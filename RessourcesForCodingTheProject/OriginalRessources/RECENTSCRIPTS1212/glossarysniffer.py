"""
Script Name: glossary_sniffer_1124.py
Created: 2025-11-24
Purpose: Extract glossary terms from XML and find them in Excel text lines
Input:
    1. XML file with StrOrigin attributes (glossary source)
    2. Excel file with lines to analyze
Output: Excel file with original lines + glossary terms found in each line
Reference: QuickSearch0818.py (Aho-Corasick glossary extraction patterns)

Usage:
    python glossary_sniffer_1124.py

    1. Run the script
    2. Select XML file (glossary source)
    3. Select Excel file (lines to analyze)
    4. Script builds glossary and searches for terms
    5. Select output location
    6. Results saved with 2 columns:
       - Column 1: Original Line
       - Column 2: Glossary Terms Found (comma-separated)

Dependencies:
    pip install openpyxl lxml pyahocorasick

Author: Generated with Claude Code
"""

import os
import re
import string
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import openpyxl
from lxml import etree
import ahocorasick

# Configuration - HARDCODED RULES (based on QuickSearch0818)
DEFAULT_LENGTH_THRESHOLD = 15  # Max chars for glossary term (Korean names/terms)
MIN_OCCURRENCE = 2  # Term must appear at least this many times to be glossary
FILTER_SENTENCES = True  # Remove entries ending with punctuation (.?!)
FILTER_PUNCTUATION = True  # Remove entries containing ANY punctuation (including …)
WORD_BOUNDARIES = True  # Only match complete words (not inside other words)

def extract_glossary_from_xml(xml_path, length_threshold=DEFAULT_LENGTH_THRESHOLD, min_occurrence=MIN_OCCURRENCE):
    """
    Extract glossary terms from XML StrOrigin attributes.

    HARDCODED FILTERING RULES (from QuickSearch0818):
    1. Length < 15 characters (Korean names/terms are usually short)
    2. No punctuation endings (.?!)
    3. No punctuation inside term (except spaces/hyphens for multi-word)
    4. Must appear >= 2 times (min_occurrence)
    5. Non-empty strings only

    Args:
        xml_path: Path to XML file
        length_threshold: Maximum character length for glossary terms
        min_occurrence: Minimum times a term must appear to be considered glossary

    Returns:
        list: Filtered glossary terms (deduplicated, sorted by frequency)
    """
    print(f"\n📖 Extracting glossary from: {Path(xml_path).name}")

    try:
        tree = etree.parse(xml_path)
        all_terms = []

        # Extract all StrOrigin attributes from LocStr elements
        for locstr in tree.xpath('//LocStr'):
            term = locstr.get('StrOrigin', '').strip()
            if term:
                all_terms.append(term)

        print(f"   Found {len(all_terms):,} total StrOrigin entries")

        # Filter to keep only glossary-worthy terms (with occurrence counting)
        glossary = filter_glossary_terms(all_terms, length_threshold, min_occurrence)

        print(f"   ✅ Filtered to {len(glossary):,} glossary terms (min {min_occurrence} occurrences)")
        return glossary

    except Exception as e:
        print(f"   ❌ Error parsing XML: {e}")
        raise

def filter_glossary_terms(terms, length_threshold, min_occurrence):
    """
    Filter terms to keep only glossary-worthy entries.

    HARDCODED FILTERING RULES (from QuickSearch0818 lines 2113-2149):
    1. Length < threshold (15 chars for Korean)
    2. Non-empty strings
    3. Does NOT end with punctuation (.?!)
    4. Does NOT contain ANY punctuation (including special ellipsis …)
    5. Must appear >= min_occurrence times (2+)

    Two-pass filtering (like QuickSearch0818):
    - Pass 1: Basic filtering + count occurrences
    - Pass 2: Apply min_occurrence filter

    Args:
        terms: List of terms to filter
        length_threshold: Maximum character length
        min_occurrence: Minimum times term must appear

    Returns:
        list: Filtered unique terms (sorted by frequency, most common first)
    """
    # Pass 1: Basic filtering + count occurrences
    count_map = {}

    for term in terms:
        # Skip empty
        if not term:
            continue

        # Skip if too long
        if len(term) >= length_threshold:
            continue

        # Skip if ends with punctuation (likely a sentence)
        if FILTER_SENTENCES and re.search(r'[.?!]\s*$', term.strip()):
            continue

        # Skip if contains ANY punctuation (except spaces and hyphens for multi-word terms)
        # Based on QuickSearch0818 pattern: line 2140
        if FILTER_PUNCTUATION:
            # Allow only alphanumeric, spaces, hyphens
            # Filter out other punctuation including special ellipsis …
            if any(ch in string.punctuation.replace('-', '').replace(' ', '') for ch in term) or '…' in term:
                continue

        # Count occurrences
        count_map[term] = count_map.get(term, 0) + 1

    # Pass 2: Apply min_occurrence filter
    filtered = [term for term, count in count_map.items() if count >= min_occurrence]

    # Sort by frequency (most common first)
    filtered.sort(key=lambda t: count_map[t], reverse=True)

    return filtered

def build_ahocorasick_automaton(glossary_terms):
    """
    Build Aho-Corasick automaton from glossary terms for fast multi-pattern matching.

    Pattern from QuickSearch0818.py lines 2217-2224

    Args:
        glossary_terms: List of terms to search for

    Returns:
        Aho-Corasick automaton object
    """
    print(f"\n🔧 Building Aho-Corasick automaton with {len(glossary_terms):,} terms...")

    automaton = ahocorasick.Automaton()

    for term in glossary_terms:
        # Add term to automaton (key=term, value=term)
        automaton.add_word(term, term)

    # Finalize the automaton
    automaton.make_automaton()

    print(f"   ✅ Automaton built successfully")
    return automaton

def search_line_for_glossary(line, automaton):
    """
    Search a single line for glossary terms using Aho-Corasick + word boundaries.

    Pattern from QuickSearch0818.py lines 2250-2260
    + Word boundary validation (only match complete words, not substrings)

    Args:
        line: Text line to search
        automaton: Aho-Corasick automaton

    Returns:
        list: Unique glossary terms found (sorted by position for consistent output)
    """
    if not line:
        return []

    matches = []
    seen = set()

    # Use Aho-Corasick to find all matches
    for end_index, term in automaton.iter(line):
        if term in seen:
            continue

        # Word boundary validation (if enabled)
        if WORD_BOUNDARIES:
            # Check if match is at word boundaries
            # Find start position
            start_index = end_index - len(term) + 1

            # Check character before match (if exists)
            if start_index > 0:
                char_before = line[start_index - 1]
                # If alphanumeric before, it's inside a word - skip
                if char_before.isalnum():
                    continue

            # Check character after match (if exists)
            if end_index + 1 < len(line):
                char_after = line[end_index + 1]
                # If alphanumeric after, it's inside a word - skip
                if char_after.isalnum():
                    continue

        matches.append(term)
        seen.add(term)

    return matches

def resolve_overlapping_matches(matches, line):
    """
    For overlapping matches, prefer longest match.

    Example: If both "Duke" and "Duke Elenor" match, keep only "Duke Elenor"

    Args:
        matches: List of matched terms
        line: Original line text

    Returns:
        list: Filtered matches with overlaps resolved
    """
    if len(matches) <= 1:
        return matches

    # Sort by length (longest first)
    sorted_matches = sorted(matches, key=len, reverse=True)

    final_matches = []
    for match in sorted_matches:
        # Check if this match is a substring of any already accepted match
        is_substring = False
        for accepted in final_matches:
            if match in accepted and match != accepted:
                is_substring = True
                break

        if not is_substring:
            final_matches.append(match)

    # Sort by appearance order in line for consistent output
    final_matches.sort(key=lambda m: line.find(m))

    return final_matches

def process_excel_lines(excel_path, automaton):
    """
    Read Excel, search each line for glossary terms.

    Args:
        excel_path: Path to input Excel file
        automaton: Aho-Corasick automaton

    Returns:
        list: Tuples of (original_line, glossary_terms_found)
    """
    print(f"\n📊 Processing Excel file: {Path(excel_path).name}")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active

        results = []
        total_matches = 0

        # Process each row (skip header if exists)
        for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Get first column value (line to analyze)
            line = str(row[0]) if row[0] else ""

            if not line or line == "None":
                results.append((line, []))
                continue

            # Search for glossary terms
            matches = search_line_for_glossary(line, automaton)

            # Resolve overlapping matches (prefer longest)
            matches = resolve_overlapping_matches(matches, line)

            results.append((line, matches))

            if matches:
                total_matches += 1

            # Progress indicator
            if idx % 100 == 0:
                print(f"   Processed {idx:,} lines ({total_matches:,} with matches)...")

        print(f"   ✅ Processed {len(results):,} total lines")
        print(f"   ✅ Found glossary terms in {total_matches:,} lines ({total_matches/len(results)*100:.1f}%)")

        wb.close()
        return results

    except Exception as e:
        print(f"   ❌ Error reading Excel: {e}")
        raise

def write_results_to_excel(results, output_path):
    """
    Write results to Excel with 2 columns:
    - Column 1: Original Line
    - Column 2: Glossary Terms Found (comma-separated)

    Args:
        results: List of (original_line, glossary_terms) tuples
        output_path: Path to save output Excel
    """
    print(f"\n💾 Saving results to: {Path(output_path).name}")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Glossary Analysis"

        # Header row
        ws.append(["Original Line", "Glossary Terms Found"])

        # Format header (bold)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for line, matches in results:
            # Join matches with comma
            glossary_str = ", ".join(matches) if matches else ""
            ws.append([line, glossary_str])

        # Auto-size columns
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 50

        wb.save(output_path)
        print(f"   ✅ Saved {len(results):,} rows")

    except Exception as e:
        print(f"   ❌ Error saving Excel: {e}")
        raise

def main():
    """
    Main execution:
    1. Pick XML glossary source
    2. Build glossary with smart filtering
    3. Build Aho-Corasick automaton
    4. Pick Excel input file
    5. Search for glossary terms in each line
    6. Save results to Excel
    """
    print("="*70)
    print("🔍 GlossarySniffer - Glossary Term Extraction & Search")
    print("="*70)

    # Tkinter setup
    root = tk.Tk()
    root.withdraw()

    try:
        # Step 1: Select XML glossary source
        print("\n📂 Step 1: Select XML glossary source file")
        xml_path = filedialog.askopenfilename(
            title="Select XML Glossary Source",
            filetypes=[
                ("XML Files", "*.xml"),
                ("All Files", "*.*")
            ]
        )

        if not xml_path:
            print("❌ No XML file selected. Exiting.")
            return

        # Step 2: Extract and filter glossary
        glossary = extract_glossary_from_xml(xml_path, DEFAULT_LENGTH_THRESHOLD)

        if not glossary:
            messagebox.showerror("Error", "No glossary terms found in XML file!")
            return

        # Show some sample terms
        print(f"\n   Sample glossary terms:")
        for term in glossary[:10]:
            print(f"      - {term}")
        if len(glossary) > 10:
            print(f"      ... and {len(glossary)-10} more")

        # Step 3: Build Aho-Corasick automaton
        automaton = build_ahocorasick_automaton(glossary)

        # Step 4: Select Excel input file
        print("\n📂 Step 2: Select Excel input file (lines to analyze)")
        excel_path = filedialog.askopenfilename(
            title="Select Excel File (Lines to Analyze)",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )

        if not excel_path:
            print("❌ No Excel file selected. Exiting.")
            return

        # Step 5: Process Excel and search for glossary terms
        results = process_excel_lines(excel_path, automaton)

        # Step 6: Select output location
        print("\n📂 Step 3: Select output location")
        output_path = filedialog.asksaveasfilename(
            title="Save Results As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{Path(excel_path).stem}_glossary_analysis.xlsx"
        )

        if not output_path:
            print("❌ No output location selected. Exiting.")
            return

        # Step 7: Write results
        write_results_to_excel(results, output_path)

        # Success!
        print("\n" + "="*70)
        print("✅ SUCCESS!")
        print("="*70)
        print(f"📊 Glossary Terms: {len(glossary):,}")
        print(f"📄 Lines Processed: {len(results):,}")
        print(f"💾 Output: {output_path}")
        print("="*70)

        messagebox.showinfo(
            "Success",
            f"Glossary analysis complete!\n\n"
            f"Glossary Terms: {len(glossary):,}\n"
            f"Lines Processed: {len(results):,}\n\n"
            f"Results saved to:\n{output_path}"
        )

    except Exception as e:
        error_msg = f"Error: {str(e)}"
        print(f"\n❌ {error_msg}")
        messagebox.showerror("Error", error_msg)
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
