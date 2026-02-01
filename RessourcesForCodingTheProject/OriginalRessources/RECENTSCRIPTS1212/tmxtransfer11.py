#!/usr/bin/env python3
import os, re, html, threading
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
import openpyxl, pandas as pd

xml_folder_path = None
xml_file_path   = None
tmx_file_path   = None

btn_upload_xml       = None
btn_upload_tmx       = None
btn_simple_translate = None
btn_two_step         = None
status_label         = None
info_panel           = None
mode_var             = None
root                 = None

def postprocess_tmx_string(xml_str):
    import re, html
    def replace_outside_ph(pattern, repl, text):
        parts = []; last = 0
        for m in re.finditer(r'<ph>.*?</ph>', text, flags=re.DOTALL):
            parts.append(re.sub(pattern, repl, text[last:m.start()]))
            parts.append(m.group(0))
            last = m.end()
        parts.append(re.sub(pattern, repl, text[last:]))
        return ''.join(parts)

    sk_pat = re.compile(r'\{Staticinfo:Knowledge:([^#}]+)#([^}]+)\}')
    seg_pat = re.compile(r'(<seg[^>]*>)(.*?)(</seg>)', re.DOTALL)

    def seg_repl(m):
        open_tag, content, close_tag = m.groups()
        counter = {'i':0}
        def sk_repl(mm):
            counter['i'] += 1
            i = counter['i']
            ident = mm.group(1)
            inner = mm.group(2)
            bpt = (f"<bpt i='{i}'>&lt;mq:rxt displaytext=&quot;"
                   f"{html.escape(ident)}&quot; val=&quot;{{Staticinfo:Knowledge:"
                   f"{html.escape(ident)}#&quot;&gt;</bpt>")
            ept = (f"<ept i='{i}'>&lt;/mq:rxt displaytext=&quot;}}&quot; "
                   f'val=&quot;}}&quot;&gt;</ept>')
            return f"{bpt}{inner}{ept}"
        new_content = replace_outside_ph(sk_pat, sk_repl, content)
        return f"{open_tag}{new_content}{close_tag}"

    xml_str = seg_pat.sub(seg_repl, xml_str)
    xml_str = replace_outside_ph(r'%(\d)#',
        lambda m: f'<ph>&lt;mq:rxt-req displaytext="Param{m.group(1)}" val="%{m.group(1)}#" /&gt;</ph>',
        xml_str)
    xml_str = replace_outside_ph(r'\{(?!Staticinfo:Knowledge:)([^}]+)\}',
        lambda m: f'<ph>&lt;mq:rxt-req displaytext="{m.group(1)}" val="{{{m.group(1)}}}" /&gt;</ph>',
        xml_str)
    xml_str = replace_outside_ph(r'\\(\w)',
        lambda m: f'<ph>&lt;mq:rxt displaytext="\\{m.group(1)}" val="\\{m.group(1)}" /&gt;</ph>',
        xml_str)
    xml_str = replace_outside_ph(r'&lt;br/&gt;',
        lambda _: '<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>',
        xml_str)
    xml_str = replace_outside_ph(r'&amp;desc;',
        lambda _: '<ph>&lt;mq:rxt-req displaytext="desc" val="&amp;desc;" /&gt;</ph>',
        xml_str)
    xml_str = re.sub(
        r'(<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>)(\s{2,})(\S)',
        lambda m: f"{m.group(1)}<ph type='fmt'>{m.group(2)}</ph>{m.group(3)}",
        xml_str
    )
    return xml_str

def preprocess_tmx_content(raw_content):
    return re.sub(
        r'<seg>(.*?)</seg>',
        lambda m: "<seg>" + m.group(1).replace('\n','&lt;br/&gt;').replace('\r\n','&lt;br/&gt;') + "</seg>",
        raw_content,
        flags=re.DOTALL
    )

def normalize_text(txt):
    if not txt:
        return ""
    txt = html.unescape(str(txt))
    return re.sub(r'\s+',' ',' '.join(txt.strip().split()))

def get_all_files(folder, ext):
    result = []
    for dp, _, fns in os.walk(folder):
        for fn in fns:
            if fn.lower().endswith(ext.lower()):
                result.append(os.path.join(dp, fn))
    return result

def get_all_xml_files(folder):
    return get_all_files(folder, ".xml")

def robust_parse_xml(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
        content = re.sub(r'^<\?xml[^>]*\?>\s*','', content, flags=re.MULTILINE)
        content = re.sub(r'<!DOCTYPE[^>]*>\s*','', content, flags=re.MULTILINE)
        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                 no_network=True, recover=True)
        root_elem = etree.fromstring(content.encode("utf-8"), parser=parser)
        return etree.ElementTree(root_elem), root_elem
    except:
        return None, None

def is_korean(text):
    return bool(text and re.search(r'[\uac00-\ud7af\u1100-\u11ff\u3130-\u318f]', text))

def is_valid_latin_description(desc):
    if not desc or not desc.strip():
        return False
    if re.search(r'[\uac00-\ud7a3]', desc):
        return False
    if not re.search(r'[A-Za-z]', desc):
        return False
    if re.fullmatch(r'[\s\W_]+', desc):
        return False
    return True

def parse_tmx(tmx_path):
    print(f"[TMX] Parsing TMX file: {tmx_path}")
    raw = open(tmx_path, "r", encoding="utf-8").read()
    pre = preprocess_tmx_content(raw)
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                             no_network=True, recover=True)
    root = etree.fromstring(pre.encode("utf-8"), parser=parser)
    body = root.find("body")
    if body is None:
        raise RuntimeError("TMX missing <body>")
    translation_units, desc_units = [], []
    for tu in body.findall("tu"):
        ctx = None; kr = en = desc_ko = None
        for prop in tu.findall("prop"):
            if prop.get("type") == "x-context" and prop.text:
                ctx = prop.text.strip(); break
        if not ctx:
            continue
        for tuv in tu.findall("tuv"):
            lang = (tuv.get("{http://www.w3.org/XML/1998/namespace}lang")
                    or tuv.get("xml:lang") or "").lower()
            seg = tuv.find("seg")
            if seg is None:
                continue
            txt = seg.text.strip() if seg.text else ""
            if lang.startswith("ko"):
                if txt.lower().startswith("&desc;") or txt.lower().startswith("&amp;desc;"):
                    desc_ko = txt
                else:
                    kr = txt
            elif lang.startswith("en"):
                en = txt
        if desc_ko:
            desc_units.append({'context':ctx,'kr':desc_ko,'en':en})
        elif kr or en:
            translation_units.append({'context':ctx,'kr':kr,'en':en})
    print(f"[TMX]   normal units: {len(translation_units)}   description units: {len(desc_units)}")
    return translation_units, desc_units

def update_buttons():
    if (xml_folder_path or xml_file_path) and tmx_file_path:
        btn_simple_translate.config(state="normal", bg="#ffb347")
        btn_two_step.config(state="normal", bg="#ffb347")
    else:
        btn_simple_translate.config(state="disabled", bg="grey")
        btn_two_step.config(state="disabled", bg="grey")

def update_info_panel():
    info_panel.config(state="normal")
    info_panel.delete("1.0", tk.END)
    info_panel.insert(tk.END, "Current Selections:\n")
    info_panel.insert(tk.END, f"Mode: {'Folder' if mode_var.get()=='folder' else 'File'}\n")
    info_panel.insert(tk.END, f"XML Folder: {xml_folder_path or '[Not selected]'}\n")
    info_panel.insert(tk.END, f"XML File:   {xml_file_path or '[Not selected]'}\n")
    info_panel.insert(tk.END, f"TMX File:   {tmx_file_path or '[Not selected]'}\n")
    info_panel.config(state="disabled")

def upload_xml():
    global xml_folder_path, xml_file_path
    if mode_var.get()=="folder":
        folder = filedialog.askdirectory(title="Select XML Folder")
        if folder:
            xml_folder_path, xml_file_path = folder, None
            for p in get_all_xml_files(folder):
                try: os.chmod(p, 0o666)
                except: pass
            btn_upload_xml.config(bg="green")
            status_label.config(text="XML folder selected.")
    else:
        file = filedialog.askopenfilename(title="Select XML File", filetypes=[("XML Files","*.xml")])
        if file:
            xml_file_path, xml_folder_path = file, None
            try: os.chmod(file, 0o666)
            except: pass
            btn_upload_xml.config(bg="green")
            status_label.config(text="XML file selected.")
    update_buttons(); update_info_panel()

def upload_tmx_file():
    global tmx_file_path
    file = filedialog.askopenfilename(title="Select TMX File", filetypes=[("TMX Files","*.tmx")])
    if file:
        tmx_file_path = file
        btn_upload_tmx.config(bg="green")
        status_label.config(text="TMX file selected.")
    update_buttons(); update_info_panel()

def switch_mode():
    global xml_folder_path, xml_file_path
    xml_folder_path = xml_file_path = None
    btn_upload_xml.config(bg="SystemButtonFace")
    status_label.config(text=f"Switched to {'Folder' if mode_var.get()=='folder' else 'File'} mode.")
    update_buttons(); update_info_panel()

def xml_to_excel():
    """
    Convert XML → Excel while PRESERVING the RAW, untouched attribute
    values of every <LocStr>.  In addition to the five standard columns
    (StrOrigin, Str, StringId, DescOrigin, DescText), this will add one
    column per any other attribute found in any <LocStr> tag.
    """
    mode = mode_var.get()

    # 1) ───────────────────────────────  get XML file(s)
    if mode == "folder":
        folder = filedialog.askdirectory(title="Select XML Folder")
        if not folder:
            return
        xml_files = get_all_files(folder, ".xml")
    else:
        file = filedialog.askopenfilename(
            title="Select XML File", filetypes=[("XML Files", "*.xml")]
        )
        if not file:
            return
        xml_files = [file]

    if not xml_files:
        messagebox.showerror("Error", "No XML files found.")
        return

    # 2) ───────────────────────────────  prepare regexes
    LOCSTR_TAG = re.compile(r"<LocStr\b([^>]*)>", re.DOTALL | re.IGNORECASE)
    ATTR_PAT   = re.compile(r'(\w+)\s*=\s*(?:"([^"]*)"|\'([^\']*)\')', re.DOTALL)

    # 3) ───────────────────────────────  discover all attribute names
    all_attrs = set()
    for xml_path in xml_files:
        try:
            raw = open(xml_path, "r", encoding="utf-8", errors="ignore").read()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read XML:\n{xml_path}\n\n{e}")
            continue
        for m in LOCSTR_TAG.finditer(raw):
            blob = m.group(1).rstrip("/").strip()
            for a in ATTR_PAT.finditer(blob):
                all_attrs.add(a.group(1))

    # 4) ───────────────────────────────  build headers
    # primary_attrs maps to columns: StrOrigin, Str, StringId, DescOrigin, DescText
    primary_attrs = ["StrOrigin", "Str", "StringId", "DescOrigin", "Desc"]
    # 'Desc' attribute goes into header 'DescText'
    header_names = ["StrOrigin", "Str", "StringId", "DescOrigin", "DescText"]
    extra_attrs = sorted(a for a in all_attrs if a not in primary_attrs)
    headers = header_names + extra_attrs

    # 5) ───────────────────────────────  extract rows
    rows = []
    for xml_path in xml_files:
        try:
            raw = open(xml_path, "r", encoding="utf-8", errors="ignore").read()
        except:
            continue
        for m in LOCSTR_TAG.finditer(raw):
            blob = m.group(1).rstrip("/").strip()
            attrs = {}
            for a in ATTR_PAT.finditer(blob):
                key = a.group(1)
                val = a.group(2) if a.group(2) is not None else a.group(3)
                attrs[key] = val
            row = [
                attrs.get("StrOrigin", ""),
                attrs.get("Str", ""),
                attrs.get("StringId", ""),
                attrs.get("DescOrigin", ""),
                attrs.get("Desc", ""),
            ]
            for extra in extra_attrs:
                row.append(attrs.get(extra, ""))
            rows.append(row)

    if not rows:
        messagebox.showerror("Error", "No <LocStr> elements found.")
        return

    # 6) ───────────────────────────────  write Excel
    if mode == "folder":
        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
        )
        if not out_path:
            return
    else:
        out_path = os.path.splitext(xml_files[0])[0] + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(out_path)

    messagebox.showinfo("Success", f"Extracted {len(rows)} rows to:\n{out_path}")

def excel_to_tmx(is_memoq=False):
    mode = mode_var.get()
    if mode=="folder":
        folder = filedialog.askdirectory(title="Select Excel Folder")
        if not folder: return
        excel_files = get_all_files(folder,".xlsx")
    else:
        file = filedialog.askopenfilename(title="Select Excel File",filetypes=[("Excel Files","*.xlsx")])
        if not file: return
        excel_files = [file]
    if not excel_files:
        messagebox.showerror("Error","No Excel files found."); return
    XML_NS = "http://www.w3.org/XML/1998/namespace"
    tmx_root = etree.Element("tmx", version="1.4")
    etree.SubElement(tmx_root,"header",
        creationtool="Excel2TMX",creationtoolversion="2.1",
        segtype="sentence",o_tmf="UTF-8",adminlang="en",
        srclang="ko",datatype="PlainText")
    body_elem = etree.SubElement(tmx_root,"body")
    memoq_ph = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
    for excel_file in excel_files:
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error",f"Failed to read Excel:\n{e}")
            continue
        basename = os.path.basename(excel_file)
        for row in ws.iter_rows(min_row=2, values_only=True):
            src = str(row[0]).strip() if row[0] else ""
            tgt = str(row[1]).strip() if len(row)>1 and row[1] else ""
            ctx = str(row[2]).strip() if len(row)>2 and row[2] else ""
            desc_src = str(row[3]).strip() if len(row)>3 and row[3] else ""
            desc_tgt = str(row[4]).strip() if len(row)>4 and row[4] else ""
            normal_ok = bool(tgt and ctx and not is_korean(tgt))
            desc_ok   = bool(desc_src and desc_tgt and ctx and not is_korean(desc_tgt))
            if not(normal_ok or desc_ok): continue
            if normal_ok:
                tu = etree.SubElement(body_elem,"tu",
                    creationid="Excel2TMX",changeid="Excel2TMX")
                tuv_ko = etree.SubElement(tu,"tuv",{f"{{{XML_NS}}}lang":"ko"})
                etree.SubElement(tuv_ko,"seg").text = src
                tuv_en = etree.SubElement(tu,"tuv",{f"{{{XML_NS}}}lang":"en-us"})
                etree.SubElement(tuv_en,"seg").text = tgt
                etree.SubElement(tu,"prop",type="x-document").text = basename
                etree.SubElement(tu,"prop",type="x-context").text  = ctx
            if desc_ok:
                tu = etree.SubElement(body_elem,"tu",
                    creationid="Excel2TMX",changeid="Excel2TMX")
                tuv_ko_d = etree.SubElement(tu,"tuv",{f"{{{XML_NS}}}lang":"ko"})
                tuv_en_d = etree.SubElement(tu,"tuv",{f"{{{XML_NS}}}lang":"en-us"})
                if is_memoq:
                    seg_ko = etree.SubElement(tuv_ko_d,"seg")
                    ph_ko  = etree.SubElement(seg_ko,"ph")
                    ph_ko.text = memoq_ph; ph_ko.tail = desc_src
                    seg_en = etree.SubElement(tuv_en_d,"seg")
                    ph_en  = etree.SubElement(seg_en,"ph")
                    ph_en.text = memoq_ph; ph_en.tail = desc_tgt
                else:
                    etree.SubElement(tuv_ko_d,"seg").text = "&desc;" + desc_src
                    etree.SubElement(tuv_en_d,"seg").text = "&desc;" + desc_tgt
                etree.SubElement(tu,"prop",type="x-document").text = basename
                etree.SubElement(tu,"prop",type="x-context").text  = ctx
    xml_bytes = etree.tostring(tmx_root, pretty_print=True,
        encoding='UTF-8', xml_declaration=True)
    xml_str = xml_bytes.decode('utf-8')
    if is_memoq:
        xml_str = postprocess_tmx_string(xml_str)
    if mode=="folder":
        out_path = filedialog.asksaveasfilename(defaultextension=".tmx",
                    filetypes=[("TMX Files","*.tmx")])
        if not out_path: return
    else:
        base = os.path.splitext(excel_files[0])[0]
        out_path = base + ("_memoq.tmx" if is_memoq else ".tmx")
    with open(out_path,"w",encoding="utf-8") as f:
        f.write(xml_str)
    messagebox.showinfo("Success",
        f"Excel → {'MemoQ ' if is_memoq else ''}TMX saved:\n{out_path}")

def tmx_to_excel():
    """
    Convert TMX → Excel while PRESERVING the RAW, untouched content
    of every <seg>.  No escaping / un-escaping / stripping of prefixes
    or white-space is performed – what is inside the <seg> element is
    what will be written to the Excel cells.
    """
    mode = mode_var.get()

    # 1) ───────────────────────────────  get TMX file(s)
    if mode == "folder":
        folder = filedialog.askdirectory(title="Select TMX Folder")
        if not folder:
            return
        tmx_files = get_all_files(folder, ".tmx")
    else:  # file
        file = filedialog.askopenfilename(
            title="Select TMX File",
            filetypes=[("TMX Files", "*.tmx")],
        )
        if not file:
            return
        tmx_files = [file]

    if not tmx_files:
        messagebox.showerror("Error", "No TMX files found.")
        return

    # 2) ───────────────────────────────  helpers
    XML_NS = "{http://www.w3.org/XML/1998/namespace}"

    MEMOQ_FLAG = '<ph>&lt;mq:rxt-req displaytext="desc" val="&amp;desc;" /&gt;</ph>'
    def _is_desc(raw: str) -> bool:
        """Detect if a RAW segment is a description."""
        chk = raw.lstrip()
        return (
            chk.startswith("&desc;")
            or chk.startswith("&amp;desc;")
            or chk.startswith(MEMOQ_FLAG[:20])  # quick check for MemoQ flag
        )

    def _extract_seg_raw(seg_elem):
        """
        Return the INNER XML (raw) of a <seg> element, with NO transformation
        whatsoever.  This is achieved by serialising the <seg> element and
        then removing the opening & closing <seg> tags.
        """
        full = etree.tostring(seg_elem, encoding="unicode", with_tail=False)
        # cleanly strip the opening and closing seg tags
        inner = re.sub(r'^<seg[^>]*>', '', full, count=1, flags=re.DOTALL)
        inner = re.sub(r'</seg>$', '', inner, count=1, flags=re.DOTALL)
        return inner

    # 3) ───────────────────────────────  collect rows
    rows_by_ctx = {}  # key = StringId / x-context
    for tmx_path in tmx_files:
        try:
            parser = etree.XMLParser(resolve_entities=False, load_dtd=False, recover=True, no_network=True)
            tree = etree.parse(tmx_path, parser)
            root_elem = tree.getroot()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse TMX:\n{tmx_path}\n\n{e}")
            continue

        for tu in root_elem.findall(".//tu"):
            ctx = "NoStringId"
            for prop in tu.findall("prop"):
                if prop.get("type") == "x-context" and prop.text:
                    ctx = prop.text
                    break

            ko_raw = en_raw = ""
            for tuv in tu.findall("tuv"):
                lang = (tuv.get(f"{XML_NS}lang") or tuv.get("xml:lang") or "").lower()
                seg = tuv.find("seg")
                if seg is None:
                    continue
                raw_content = _extract_seg_raw(seg)
                if lang.startswith("ko"):
                    ko_raw = raw_content
                elif lang.startswith("en"):
                    en_raw = raw_content

            if not (ko_raw or en_raw):
                continue

            # initialise row if new
            if ctx not in rows_by_ctx:
                rows_by_ctx[ctx] = ["", "", ctx, "", ""]

            if _is_desc(ko_raw):
                rows_by_ctx[ctx][3] = ko_raw
                rows_by_ctx[ctx][4] = en_raw
            else:
                rows_by_ctx[ctx][0] = ko_raw
                rows_by_ctx[ctx][1] = en_raw

    if not rows_by_ctx:
        messagebox.showerror("Error", "No translation units found.")
        return

    # 4) ───────────────────────────────  write Excel
    rows = [rows_by_ctx[k] for k in sorted(rows_by_ctx)]
    if mode == "folder":
        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
        )
        if not out_path:
            return
    else:
        out_path = os.path.splitext(tmx_files[0])[0] + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["StrOrigin", "Str", "StringId", "DescOrigin", "DescText"])
    for r in rows:
        ws.append(r)
    wb.save(out_path)

    messagebox.showinfo("Success", f"TMX → Excel saved:\n{out_path}")

def simple_translate_krid():
    if not tmx_file_path or not(xml_folder_path or xml_file_path):
        messagebox.showerror("Error","Please select XML and TMX."); return
    btn_simple_translate.config(state="disabled")
    btn_two_step.config(state="disabled")
    status_label.config(text="Simple translate: processing...")
    def worker():
        try:
            tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error",f"Parse TMX failed: {e}"))
            root.after(0, lambda: status_label.config(text="Error parsing TMX."))
            root.after(0, update_buttons); return
        perfect_str = {}
        for u in tmx_trans:
            if u.get('en') and u.get('context') is not None:
                perfect_str[(u['context'], normalize_text(u['kr']))] = u['en']
        perfect_desc = {}
        for u in tmx_desc:
            if u.get('en') and u.get('context') is not None:
                perfect_desc[(u['context'], normalize_text(u['kr']))] = u['en']
        xml_files = (get_all_xml_files(xml_folder_path)
                     if mode_var.get()=='folder'
                     else [xml_file_path])
        updated_s = updated_d = 0
        for path in xml_files:
            tree_obj, root_elem = robust_parse_xml(path)
            if root_elem is None: continue
            changed = False
            for loc in root_elem.iter("LocStr"):
                sid = loc.get("StringId","").strip()
                orig=normalize_text(loc.get("StrOrigin",""))
                dorig=normalize_text(loc.get("DescOrigin",""))
                old_s = loc.get("Str","")
                old_d = loc.get("Desc","")
                new_s = perfect_str.get((sid, orig))
                if new_s is not None and new_s != old_s:
                    loc.set("Str", new_s); updated_s+=1; changed=True
                new_d = perfect_desc.get((sid, dorig))
                if new_d is not None and new_d != old_d:
                    loc.set("Desc", new_d); updated_d+=1; changed=True
            if changed:
                tree_obj.write(path, encoding="utf-8",
                               xml_declaration=False, pretty_print=True)
        def on_done():
            messagebox.showinfo("Simple Translate Complete",
                f"Updated Str: {updated_s}\nUpdated Desc: {updated_d}")
            status_label.config(
                text=f"Simple translate done: Str={updated_s}, Desc={updated_d}")
            update_buttons()
        root.after(0, on_done)
    threading.Thread(target=worker, daemon=True).start()

def two_step_match_krid_then_kr():
    if not tmx_file_path or not(xml_folder_path or xml_file_path):
        messagebox.showerror("Error","Please select XML and TMX."); return
    btn_simple_translate.config(state="disabled")
    btn_two_step.config(state="disabled")
    status_label.config(text="2-step translate: processing...")
    def worker():
        try:
            tmx_trans, tmx_desc = parse_tmx(tmx_file_path)
        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error",f"Parse TMX failed: {e}"))
            root.after(0, lambda: status_label.config(text="Error parsing TMX."))
            root.after(0, update_buttons); return
        perfect_str = {
            (u['context'], normalize_text(u['kr'])): u['en']
            for u in tmx_trans if u.get('kr') and u.get('en')
        }
        perfect_desc = {
            (u['context'], normalize_text(u['kr'])): u['en']
            for u in tmx_desc if u.get('kr') and u.get('en')
        }
        kr_only = {
            normalize_text(u['kr']): u['en']
            for u in tmx_trans if u.get('kr') and u.get('en')
        }
        xml_files = (get_all_xml_files(xml_folder_path)
                     if mode_var.get()=='folder'
                     else [xml_file_path])
        upd1 = upd2 = updd = 0
        for path in xml_files:
            tree_obj, root_elem = robust_parse_xml(path)
            if root_elem is None: continue
            changed = False
            for loc in root_elem.iter("LocStr"):
                sid = loc.get("StringId","").strip()
                orig=normalize_text(loc.get("StrOrigin",""))
                dori=normalize_text(loc.get("DescOrigin",""))
                olds=loc.get("Str",""); oldd=loc.get("Desc","")
                if (sid, orig) in perfect_str:
                    new = perfect_str[(sid, orig)]
                    if new and new != olds:
                        loc.set("Str", new); upd1+=1; changed=True
                elif orig in kr_only:
                    new = kr_only[orig]
                    if new and new != olds:
                        loc.set("Str", new); upd2+=1; changed=True
                if (sid, dori) in perfect_desc:
                    new = perfect_desc[(sid, dori)]
                    if new and new != oldd:
                        loc.set("Desc", new); updd+=1; changed=True
            if changed:
                tree_obj.write(path, encoding="utf-8",
                               xml_declaration=False, pretty_print=True)
        def on_done():
            messagebox.showinfo("2-Step Translate Complete",
                f"Perfect matches: {upd1}\nKR-only matches: {upd2}\nDescription matches: {updd}")
            status_label.config(
                text=f"2-Step done: Perfect={upd1}, KR-only={upd2}, Desc={updd}")
            update_buttons()
        root.after(0, on_done)
    threading.Thread(target=worker, daemon=True).start()

def excel_to_xml():
    mode = mode_var.get()
    if mode=="folder":
        folder = filedialog.askdirectory(title="Select Excel Folder")
        if not folder: return
        excel_files = get_all_files(folder, ".xlsx")
    else:
        file = filedialog.askopenfilename(title="Select Excel File",
                  filetypes=[("Excel Files","*.xlsx")])
        if not file: return
        excel_files = [file]
    if not excel_files:
        messagebox.showerror("Error","No Excel files found."); return
    if mode=="file":
        out_file = filedialog.asksaveasfilename(defaultextension=".xml",
                    filetypes=[("XML Files","*.xml")])
        if not out_file: return
    for excel_file in excel_files:
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error",f"Failed to read Excel:\n{e}")
            continue
        root_elem = etree.Element("Root")
        for row in ws.iter_rows(min_row=2, values_only=True):
            so = str(row[0]).strip() if row[0] else ""
            st = str(row[1]).strip() if len(row)>1 and row[1] else ""
            sid= str(row[2]).strip() if len(row)>2 and row[2] else ""
            do = str(row[3]).strip() if len(row)>3 and row[3] else ""
            dt = str(row[4]).strip() if len(row)>4 and row[4] else ""
            loc = etree.SubElement(root_elem, "LocStr")
            loc.set("StrOrigin", so)
            loc.set("Str", st)
            loc.set("StringId", sid)
            if do: loc.set("DescOrigin", do)
            if dt: loc.set("Desc", dt)
        xml_str = etree.tostring(root_elem, pretty_print=True,
                     encoding="utf-8").decode("utf-8")
        if mode=="folder":
            base = os.path.splitext(os.path.basename(excel_file))[0]
            out_path = os.path.join(os.path.dirname(excel_file), base+".xml")
        else:
            out_path = out_file
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(xml_str)
    messagebox.showinfo("Success", f"Excel → XML saved for {len(excel_files)} file(s).")

def generate_stringid_from_eventname_dialogvoice():
    file = filedialog.askopenfilename(title="Select Excel File",
              filetypes=[("Excel Files","*.xlsx *.xls")])
    if not file: return
    try:
        df = pd.read_excel(file, header=None, dtype=str)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file:\n{e}")
        return
    if df.shape[1] < 2:
        messagebox.showerror("Error",
            "File must have at least two columns: DialogVoice in Col1 and EventName in Col2.")
        return
    res = []
    for _, row in df.iterrows():
        dialog = str(row[0]).lower().strip() if pd.notnull(row[0]) else ""
        event  = str(row[1]).lower().strip() if pd.notnull(row[1]) else ""
        if dialog and event and dialog in event:
            diff = event.replace(dialog,"",1)
            if diff.startswith("_"): diff = diff[1:]
            res.append(diff)
        elif not dialog and event:
            res.append(event[1:] if event.startswith("_") else event)
        else:
            res.append(event if "aidialog" in event else "")
    wb = openpyxl.Workbook(); ws = wb.active
    for i,val in enumerate(res,1): ws.cell(row=i,column=1,value=val)
    out = os.path.join(os.path.dirname(file),
           f"{os.path.splitext(os.path.basename(file))[0]}_stridgenerated.xlsx")
    wb.save(out)
    messagebox.showinfo("Done", f"String IDs saved:\n{out}")

def main():
    global root, mode_var, btn_upload_xml, btn_upload_tmx, btn_simple_translate
    global btn_two_step, status_label, info_panel

    root = tk.Tk()
    root.title("XML ⇄ Excel ⇄ TMX Toolkit (with MemoQ post-process)")
    root.geometry("720x780")

    tk.Label(root, text="XML ⇄ Excel ⇄ TMX Converter", font=("Arial",16,"bold")).pack(pady=10)

    mode_var = tk.StringVar(value="file")
    frm = tk.Frame(root); frm.pack(pady=5)
    tk.Label(frm, text="Mode:").pack(side="left", padx=4)
    tk.Radiobutton(frm, text="File", variable=mode_var, value="file", command=switch_mode).pack(side="left")
    tk.Radiobutton(frm, text="Folder", variable=mode_var, value="folder", command=switch_mode).pack(side="left")

    btn_upload_xml = tk.Button(root, text="Upload XML Folder/File", width=35, height=2, command=upload_xml)
    btn_upload_xml.pack(pady=6)
    btn_upload_tmx = tk.Button(root, text="Upload TMX File", width=35, height=2, command=upload_tmx_file)
    btn_upload_tmx.pack(pady=6)

    btn_simple_translate = tk.Button(root, text="KR+ID  SIMPLE TRANSLATE", width=35, height=2,
                                     state="disabled", bg="grey", command=simple_translate_krid)
    btn_simple_translate.pack(pady=6)
    btn_two_step = tk.Button(root, text="2-STEP MATCH  KR+ID ➜ KR", width=35, height=2,
                             state="disabled", bg="grey", command=two_step_match_krid_then_kr)
    btn_two_step.pack(pady=6)

    tk.Button(root, text="Excel → TMX", width=35, height=2, bg="#d0ffd0",
              command=lambda: excel_to_tmx(is_memoq=False)).pack(pady=6)
    tk.Button(root, text="Excel → MemoQ-TMX", width=35, height=2, bg="#ffe080",
              command=lambda: excel_to_tmx(is_memoq=True)).pack(pady=6)
    tk.Button(root, text="TMX → Excel", width=35, height=2, bg="#d0e0ff",
              command=tmx_to_excel).pack(pady=6)
    tk.Button(root, text="XML → Excel", width=35, height=2, bg="#ffcccc",
              command=xml_to_excel).pack(pady=6)
    tk.Button(root, text="Excel → XML", width=35, height=2, bg="#ffccff",
              command=excel_to_xml).pack(pady=6)
    tk.Button(root, text="Generate StringID", width=35, height=2, bg="#e0ffe0",
              command=generate_stringid_from_eventname_dialogvoice).pack(pady=6)

    status_label = tk.Label(root, text="Please select XML and TMX.", fg="blue")
    status_label.pack(pady=10)

    info_panel = tk.Text(root, height=6, width=85, wrap="none", state="disabled", bg="#f8f8f8")
    info_panel.pack(padx=8, pady=6, fill="x")

    update_info_panel(); update_buttons()
    root.mainloop()

if __name__ == "__main__":
    main()