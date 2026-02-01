#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Standalone script to annotate string IDs in an Excel file
with the export-file path (Filename/ParentFolder/ParentParentFolder)
based on the export__ folder structure from the reference script.
"""

import re
import sys
from pathlib import Path
from lxml import etree as ET
from openpyxl import load_workbook
from tkinter import Tk, filedialog

# ─────────────────────────────────────────────────────────────
# CONFIGURATION: adjust this to point at your export__ folder
# ─────────────────────────────────────────────────────────────
EXPORT_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")

# ─────────────────────────────────────────────────────────────
# XML PARSE HELPERS (same as reference script)
# ─────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)


def parse_xml_file(path: Path) -> ET._Element:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    wrapped = f"<ROOT>\n{fix_bad_entities(raw)}\n</ROOT>"
    parser = ET.XMLParser(recover=True, huge_tree=True)
    return ET.fromstring(wrapped.encode("utf-8"), parser=parser)


# ─────────────────────────────────────────────────────────────
# BUILD A MAP: StringId -> "Filename/ParentFolder/ParentParentFolder"
# ─────────────────────────────────────────────────────────────
def build_export_index(folder: Path) -> dict:
    index: dict = {}
    for xml_path in folder.rglob("*.xml"):
        try:
            root = parse_xml_file(xml_path)
        except Exception:
            continue
        # compute the description string
        rel = xml_path.relative_to(folder)
        parts = rel.parts
        if len(parts) >= 3:
            desc = f"{parts[-1]}/{parts[-2]}/{parts[-3]}"
        elif len(parts) == 2:
            desc = f"{parts[-1]}/{parts[-2]}"
        else:
            desc = parts[0]
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                index[sid] = desc
    return index


# ─────────────────────────────────────────────────────────────
# USER INTERFACE: pick Excel file and choose column header
# ─────────────────────────────────────────────────────────────
def select_excel_file() -> Path:
    root = Tk()
    root.withdraw()
    filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    path_str = filedialog.askopenfilename(title="Select Excel file", filetypes=filetypes)
    root.destroy()
    if not path_str:
        return None
    return Path(path_str)


def choose_header(ws) -> int:
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    print("Detected column headers:")
    for h in headers:
        print("  ", h)
    sel = input("Enter the exact header name containing your String IDs: ").strip()
    while sel not in headers:
        sel = input("Header not found. Enter one of the above exactly: ").strip()
    return headers.index(sel) + 1


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    excel_path = select_excel_file()
    if not excel_path or not excel_path.exists():
        print("No file selected or file doesn't exist. Exiting.")
        sys.exit(1)

    wb = load_workbook(excel_path)
    ws = wb.active

    col_idx = choose_header(ws)
    print(f"Selected column index = {col_idx}")

    print(f"[INFO] Building export index from: {EXPORT_FOLDER}")
    export_index = build_export_index(EXPORT_FOLDER)
    print(f"[INFO] {len(export_index)} StringId entries indexed.")

    # insert a new column to the right of the selected column
    insert_col = col_idx + 1
    ws.insert_cols(insert_col)
    ws.cell(1, insert_col, "FilePath")

    # fill in the file path for each StringId
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, col_idx)
        sid = str(cell.value).strip() if cell.value is not None else ""
        path_info = export_index.get(sid, "")
        ws.cell(row, insert_col, path_info)

    # save as a new workbook next to the original
    out_path = excel_path.with_name(f"{excel_path.stem}_with_paths{excel_path.suffix}")
    wb.save(out_path)
    print(f"[INFO] Annotated workbook saved to: {out_path.resolve()}")


if __name__ == "__main__":
    main()