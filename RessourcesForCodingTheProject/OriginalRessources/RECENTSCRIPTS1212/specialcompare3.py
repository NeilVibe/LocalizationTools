

#!/usr/bin/env python3
"""
XML-to-XML localisation diff → Excel report (GUI helper)

This revision uses lxml with recover=True to salvage every <LocStr> and
applies the same filter (Str non-empty & no Korean) so you get the same
“universe” size as in your big script.
"""
import os
import re
import sys
import logging
import tkinter as tk
from tkinter import filedialog, messagebox
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────── CONFIG ────────────
DEBUG = True       # set False to silence console log
HEADER_FILL = PatternFill("solid", fgColor="FFC000")
ABOUT_H1_FILL = PatternFill("solid", fgColor="4F81BD")
ABOUT_H1_FONT = Font(bold=True, color="FFFFFF")
ABOUT_NORMAL_FONT = Font(color="000000")
ABOUT_WRAP = Alignment(wrapText=True)
# ────────────────────────────────

logging.basicConfig(
    level=logging.DEBUG if DEBUG else logging.CRITICAL,
    format="%(message)s",
    stream=sys.stdout
)
log = logging.getLogger("folder-diff")

def logger(msg: str) -> None:
    if DEBUG:
        log.debug(msg)

def contains_korean(text: str) -> bool:
    """True if the string has any Hangul character."""
    return bool(re.search(r'[\uac00-\ud7a3]', text))

def levenshtein(s: str, t: str) -> int:
    """Classic Levenshtein distance (iterative, O(n·m) time, O(m) memory)."""
    if s == t:
        return 0
    n, m = len(s), len(t)
    if n == 0:
        return m
    if m == 0:
        return n
    if n > m:
        s, t, n, m = t, s, m, n
    prev = list(range(m + 1))
    for i, cs in enumerate(s, 1):
        cur = [i] + [0] * m
        for j, ct in enumerate(t, 1):
            ic = prev[j] + 1
            dc = cur[j-1] + 1
            rc = prev[j-1] + (cs != ct)
            cur[j] = min(ic, dc, rc)
        prev = cur
    return prev[m]

def parse_locstr(path: str):
    """
    Parse all <LocStr> nodes (StrOrigin, StringId, Str) using lxml in recover mode.
    This will pull in every LocStr, even from slightly malformed XML.
    """
    try:
        raw = open(path, encoding='utf-8').read()
    except Exception as exc:
        logger(f"  ⚠️  Could not read XML '{path}': {exc}")
        return []
    # fix stray ampersands
    raw = re.sub(r'&(?!lt;|gt;|amp;|apos;|quot;)', '&amp;', raw)
    wrapped = "<ROOT>\n" + raw + "\n</ROOT>"
    try:
        parser = etree.XMLParser(recover=True, remove_blank_text=True)
        root = etree.fromstring(wrapped.encode('utf-8'), parser=parser)
    except Exception as exc:
        logger(f"  ⚠️  Could not parse XML '{path}': {exc}")
        return []
    records = []
    for e in root.iterfind('.//LocStr'):
        so  = (e.get('StrOrigin') or '').strip()
        sid = (e.get('StringId')  or '').strip()
        sv  = (e.get('Str')       or '').strip()
        records.append((so, sid, sv))
    return records

def get_all_xml_files(folder: str):
    """Return list of *.xml (any depth)."""
    result = []
    for dirpath, _, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().endswith('.xml'):
                result.append(os.path.join(dirpath, fn))
    return result

def main() -> None:
    # ───────── GUI SET-UP ────────────────────────────────────────────────────
    root = tk.Tk()
    root.title("XML Compare → Excel")
    root.geometry("560x400")

    mode_var = tk.StringVar(value="file")         # file | folder
    count_mode_var = tk.StringVar(value="unique") # unique | all
    src_var = tk.StringVar()
    prev_var = tk.StringVar()
    aft_var = tk.StringVar()

    def switch_mode():
        src_var.set('')
        prev_var.set('')
        aft_var.set('')
        update_labels()
        update_run_button()

    def update_labels():
        typ = "Directory" if mode_var.get() == "folder" else "File"
        lbl_src .config(text=f"SOURCE   {typ}: {src_var.get() or '[Not selected]'}")
        lbl_prev.config(text=f"PREVIOUS {typ}: {prev_var.get() or '[Not selected]'}")
        lbl_aft .config(text=f"AFTER    {typ}: {aft_var.get() or '[Not selected]'}")

    def update_run_button():
        btn_run.config(state='normal'
                       if src_var.get() and prev_var.get() and aft_var.get()
                       else 'disabled')

    def select_path(target_var: tk.StringVar):
        p = (filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
             if mode_var.get() == "file"
             else filedialog.askdirectory())
        if p:
            target_var.set(p)
            update_labels()
            update_run_button()

    def run_compare():
        src_path  = src_var.get()
        prev_root = prev_var.get()
        aft_root  = aft_var.get()
        if not (src_path and prev_root and aft_root):
            return

        logger("\n================  RUN  ==================")
        logger(f"SOURCE   : {src_path}")
        logger(f"PREVIOUS : {prev_root}")
        logger(f"AFTER    : {aft_root}\n")

        changes = []
        files_compared = files_skipped = 0
        source_nodes_total = source_chars_total = 0
        match_both_nodes = match_both_chars = 0
        match_so_nodes   = match_so_chars   = 0

        def process_triplet(f_src: str, f_prev: str, f_aft: str):
            nonlocal files_compared
            nonlocal source_nodes_total, source_chars_total
            nonlocal match_both_nodes, match_both_chars
            nonlocal match_so_nodes,   match_so_chars

            src_nodes = parse_locstr(f_src)
            # --- USE BIG-SCRIPT FILTER: Str non-empty & no Korean
            universe = [
                (so, sid, sv) for so, sid, sv in src_nodes
                if sv and not contains_korean(sv)
            ]
            if not universe:
                return

            source_nodes_total += len(universe)
            source_chars_total += sum(len(sv) for _, _, sv in universe)

            prev_list = parse_locstr(f_prev)
            aft_list  = parse_locstr(f_aft)

            pm_both = {(so, sid): sv for so, sid, sv in prev_list}
            am_both = {(so, sid): sv for so, sid, sv in aft_list}

            pm_so = {}
            am_so = {}
            for so, sid, sv in prev_list:
                pm_so.setdefault(so, []).append((sv, sid))
            for so, sid, sv in aft_list:
                am_so.setdefault(so, []).append((sv, sid))

            for so, sid, src_val in universe:
                used_fallback = False
                if (so, sid) in pm_both and (so, sid) in am_both:
                    prev_str = pm_both[(so, sid)]
                    new_str  = am_both[(so, sid)]
                    match_both_nodes  += 1
                    match_both_chars  += len(prev_str)
                else:
                    pc = pm_so.get(so, [])
                    ac = am_so.get(so, [])
                    if len(pc) == 1 and len(ac) == 1:
                        prev_str, _ = pc[0]
                        new_str,  _ = ac[0]
                        used_fallback = True
                        match_so_nodes += 1
                        match_so_chars += len(prev_str)
                    else:
                        continue

                if prev_str != new_str:
                    changes.append({
                        "StrOrigin"   : so,
                        "PrevStr"     : prev_str,
                        "NewStr"      : new_str,
                        "CharChanges" : levenshtein(prev_str, new_str)
                    })
            files_compared += 1

        # ── run over either single files or folders ────────────────────────
        if mode_var.get() == "file":
            process_triplet(src_path, prev_root, aft_root)
        else:
            src_files  = get_all_xml_files(src_path)
            prev_files = get_all_xml_files(prev_root)
            aft_files  = get_all_xml_files(aft_root)

            logger(f"Found {len(src_files)} XMLs under SOURCE.")
            logger(f"Found {len(prev_files)} XMLs under PREVIOUS.")
            logger(f"Found {len(aft_files)} XMLs under AFTER.\n")

            prev_rel = {os.path.relpath(p, prev_root): p for p in prev_files}
            aft_rel  = {os.path.relpath(a, aft_root): a for a in aft_files}
            prev_name = {}
            for p in prev_files:
                prev_name.setdefault(os.path.basename(p), []).append(p)
            aft_name = {}
            for a in aft_files:
                aft_name.setdefault(os.path.basename(a), []).append(a)

            for f_src in src_files:
                rel = os.path.relpath(f_src, src_path)
                bn  = os.path.basename(f_src)
                f_prev = f_aft = None

                if rel in prev_rel:
                    f_prev = prev_rel[rel]
                else:
                    c = prev_name.get(bn, [])
                    if len(c)==1: f_prev = c[0]

                if rel in aft_rel:
                    f_aft = aft_rel[rel]
                else:
                    c = aft_name.get(bn, [])
                    if len(c)==1: f_aft = c[0]

                logger(f"• {rel}\n    prev → {'OK' if f_prev else 'SKIP'}  aft → {'OK' if f_aft else 'SKIP'}")
                if f_prev and f_aft:
                    process_triplet(f_src, f_prev, f_aft)
                else:
                    files_skipped += 1

            logger(f"\nCompared {files_compared} triplets, skipped {files_skipped}.")

        if not changes:
            messagebox.showinfo(
                "Result",
                "No changed nodes found.\n\n"
                f"Files compared : {files_compared}\n"
                f"Files skipped  : {files_skipped}"
            )
            return

        # ── counting mode ────────────────────────────────────────────
        mode = count_mode_var.get()
        if mode == 'unique':
            seen = set()
            filtered = []
            for c in changes:
                k = (c["StrOrigin"], c["PrevStr"], c["NewStr"])
                if k not in seen:
                    seen.add(k)
                    filtered.append(c)
            processed = filtered
        else:
            processed = changes

        total_nodes_changed = len(processed)
        total_chars_changed = sum(c["CharChanges"] for c in processed)

        out_file = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not out_file:
            return

        # ── build Workbook ────────────────────────────────────────────
        wb  = Workbook()
        ws1 = wb.active
        ws1.title = "ChangedNodes"
        ws1.append(["StrOrigin","PrevStr","NewStr"])
        for r in processed:
            ws1.append([r["StrOrigin"], r["PrevStr"], r["NewStr"]])
        for col in ("A","B","C"):
            ws1.column_dimensions[col].width = 60

        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric","Nodes","Characters"])
        for c in ("A1","B1","C1"):
            ws2[c].fill = HEADER_FILL
            ws2[c].font = Font(bold=True)
        ws2.append(["Count mode",
                    "Unique" if mode=="unique" else "All strings (duplicates included)",""])
        ws2.append(["Total nodes changed", total_nodes_changed, total_chars_changed])
        ws2.append([])
        ws2.append(["SOURCE nodes analysed", source_nodes_total, source_chars_total])
        ws2.append(["Matched by (StrOrigin + StringId)", match_both_nodes, match_both_chars])
        ws2.append(["Matched by StrOrigin only",      match_so_nodes,   match_so_chars])
        ws2.append([])
        ws2.append(["Files compared", files_compared, ""])
        ws2.append(["Files skipped",  files_skipped,  ""])
        for col in ("A","B","C"):
            ws2.column_dimensions[col].width = 45

        ws3 = wb.create_sheet("About the Results")
        ws3.column_dimensions['A'].width = 120
        about = [
            ("HOW THIS TOOL WORKS",     True),
            ("",                        False),
            ("1. Gather SOURCE keys",    False),
            ("   • Scans SOURCE XML(s) and keeps <LocStr> entries whose Str "
             "is non-empty and has NO Korean characters.", False),
            ("2. Find matching XML triplets", False),
            ("   • Folder mode: tries same relative path, then unique filename.", False),
            ("3. For every SOURCE key:", False),
            ("   a. Prefer exact match (StrOrigin + StringId).", False),
            ("   b. Else fallback on unique StrOrigin.", False),
            ("4. If PREVIOUS ≠ AFTER → record change + Levenshtein δ.", False),
            ("", False),
            ("COUNT MODES", True),
            ("Unique — identical changes listed once.", False),
            ("All strings (duplicates included) — every occurrence listed.", False),
            ("", False),
            ("RESULT SHEETS", True),
            ("ChangedNodes — one row per changed string.", False),
            ("Summary — numbers at a glance.", False),
            ("About the Results — this explanation.", False),
        ]
        for txt, hdr in about:
            ws3.append([txt])
            cell = ws3.cell(row=ws3.max_row, column=1)
            cell.alignment = ABOUT_WRAP
            if hdr:
                cell.font = ABOUT_H1_FONT
                cell.fill = ABOUT_H1_FILL
            else:
                cell.font = ABOUT_NORMAL_FONT

        wb.save(out_file)
        messagebox.showinfo(
            "Done",
            f"Excel created:\n{out_file}\n\n"
            f"Mode: {'Unique' if mode=='unique' else 'All strings (duplicates included)'}\n"
            f"Nodes changed     : {total_nodes_changed}\n"
            f"Characters changed: {total_chars_changed}"
        )

    # ───────── GUI LAYOUT ──────────────────────────────────────────────────
    frm_mode = tk.Frame(root); frm_mode.pack(pady=6)
    tk.Radiobutton(frm_mode, text="File mode",
                   variable=mode_var, value="file",
                   command=switch_mode).pack(side="left", padx=5)
    tk.Radiobutton(frm_mode, text="Folder mode",
                   variable=mode_var, value="folder",
                   command=switch_mode).pack(side="left", padx=5)

    frm_count = tk.Frame(root); frm_count.pack(pady=6)
    tk.Label(frm_count, text="Count mode:").pack(side="left", padx=(0,10))
    tk.Radiobutton(frm_count, text="Unique",
                   variable=count_mode_var, value="unique").pack(side="left", padx=5)
    tk.Radiobutton(frm_count, text="All strings (duplicates included)",
                   variable=count_mode_var, value="all").pack(side="left", padx=5)

    tk.Button(root, text="Select SOURCE",   width=34, command=lambda: select_path(src_var)).pack(pady=4)
    tk.Button(root, text="Select PREVIOUS", width=34, command=lambda: select_path(prev_var)).pack(pady=4)
    tk.Button(root, text="Select AFTER",    width=34, command=lambda: select_path(aft_var)).pack(pady=4)

    lbl_src  = tk.Label(root, text="SOURCE   File: [Not selected]", wraplength=540, justify="left"); lbl_src.pack(anchor="w")
    lbl_prev = tk.Label(root, text="PREVIOUS File: [Not selected]", wraplength=540, justify="left"); lbl_prev.pack(anchor="w")
    lbl_aft  = tk.Label(root, text="AFTER    File: [Not selected]", wraplength=540, justify="left"); lbl_aft.pack(anchor="w")

    btn_run = tk.Button(root, text="Compare & Save Excel",
                        width=34, bg="#aaffaa", state="disabled",
                        command=run_compare)
    btn_run.pack(pady=16)

    root.mainloop()

if __name__ == '__main__':
    main()
