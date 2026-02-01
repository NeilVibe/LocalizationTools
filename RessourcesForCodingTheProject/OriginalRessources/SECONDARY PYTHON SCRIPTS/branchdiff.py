import os
import re
import html
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import Dict, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from lxml import etree

# --------------------------------------------------------------------------
#                              TEXT HELPERS
# --------------------------------------------------------------------------
def normalize_text(txt: str) -> str:
    """
    1.  Un-escape HTML entities (&lt; → <, &amp; → &, ...)
    2.  Strip leading / trailing whitespace
    3.  Collapse runs of whitespace into a single space
    """
    if txt is None:
        return ""
    txt = html.unescape(str(txt))
    txt = re.sub(r"\s+", " ", txt.strip())
    return txt


# --------------------------------------------------------------------------
#                               XML PARSING
# --------------------------------------------------------------------------
def robust_parse_xml(path: str):
    """
    Returns etree.Element for the root – or None on failure.
    Ignores DTD, external entities, broken encodings ... to be as forgiving
    as possible with real-world XML files.
    """
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as fh:
            content = fh.read()
    except Exception as exc:
        print(f"[ERROR] Cannot read XML: {exc}")
        return None

    # remove XML declaration / DOCTYPE to avoid network calls
    content = re.sub(r"^\s*<\?xml[^>]*?>", "", content, flags=re.MULTILINE)
    content = re.sub(r"<!DOCTYPE[^>]*?>", "", content, flags=re.MULTILINE)

    parser = etree.XMLParser(
        resolve_entities=False,
        load_dtd=False,
        no_network=True,
        recover=True
    )
    try:
        root = etree.fromstring(content.encode("utf-8"), parser=parser)
        return root
    except Exception as exc:
        print(f"[ERROR] lxml could not parse XML: {exc}")
        return None


# --------------------------------------------------------------------------
#                               MAIN LOGIC
# --------------------------------------------------------------------------
def build_lookup_tables(xml_path: str) -> Tuple[Dict[Tuple[str, str], bool], Dict[str, bool]]:
    """
    Return
        perfect_key : { (StringID , StrOrigin) : True }
        kr_only_key : { StrOrigin              : True }
    """
    root = robust_parse_xml(xml_path)
    if root is None:
        raise RuntimeError("XML parsing failed.")

    perfect_key = {}
    kr_only_key = {}

    for loc in root.iter("LocStr"):
        sid = (loc.get("StringId") or "").strip()
        kor = normalize_text(loc.get("StrOrigin") or "")

        if not kor:
            continue

        kr_only_key[kor] = True
        if sid:
            perfect_key[(sid, kor)] = True

    print(f"[XML] Loaded {len(perfect_key):,} perfect keys   "
          f"and {len(kr_only_key):,} KR-only keys.")
    return perfect_key, kr_only_key


def analyse_excel(excel_path: str,
                  perfect_key: Dict[Tuple[str, str], bool],
                  kr_only_key: Dict[str, bool]):
    """
    Walk through the Excel rows, collect statistics and remember which
    rows were unmatched (returned as a set of row indices).
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Cannot open Excel: {exc}")

    ws = wb.active

    total_rows   = 0
    hit_perfect  = 0
    hit_kronly   = 0
    unmatched    = set()  # row numbers (1-based in Excel)
    header_row   = 1      # assume first row is header

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == header_row:
            continue  # header – skip from statistics

        total_rows += 1
        src = normalize_text(row[0]) if row and row[0] is not None else ""
        sid = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")

        if (sid, src) in perfect_key:
            hit_perfect += 1
        elif src in kr_only_key:
            hit_kronly += 1
        else:
            unmatched.add(idx)

    return {
        "total": total_rows,
        "perfect": hit_perfect,
        "kr_only": hit_kronly,
        "unmatched_rows": unmatched,
        "workbook": wb,
        "worksheet": ws
    }


def colour_unmatched_rows(workbook: openpyxl.Workbook,
                          worksheet: openpyxl.worksheet.worksheet.Worksheet,
                          unmatched_rows):
    """
    Paint every cell of unmatched rows RED background / white font
    for easy spotting.
    """
    red_fill  = PatternFill("solid", fgColor="FF6363")
    white_font = Font(color="FFFFFF")

    for row_idx in unmatched_rows:
        for cell in worksheet[row_idx]:
            cell.fill = red_fill
            cell.font = white_font


# --------------------------------------------------------------------------
#                                 GUI I/O
# --------------------------------------------------------------------------
def select_files_gui():
    """
    Returns (excel_path, xml_path) – raises SystemExit if cancelled.
    """
    root = tk.Tk()
    root.withdraw()  # hide the main window

    messagebox.showinfo("Step 1 / 2", "Choose the SOURCE Excel file.")

    excel_path = filedialog.askopenfilename(
        title="Select SOURCE Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not excel_path:
        messagebox.showerror("Cancelled", "No Excel file selected.")
        raise SystemExit

    messagebox.showinfo("Step 2 / 2", "Choose the TARGET XML file.")

    xml_path = filedialog.askopenfilename(
        title="Select TARGET XML file",
        filetypes=[("XML files", "*.xml")]
    )
    if not xml_path:
        messagebox.showerror("Cancelled", "No XML file selected.")
        raise SystemExit

    root.destroy()
    return excel_path, xml_path


# --------------------------------------------------------------------------
#                                   MAIN
# --------------------------------------------------------------------------
def main():
    excel_path, xml_path = select_files_gui()

    print("\n──────────────────────────────────────────────────────────────")
    print(f"[INFO] SOURCE Excel : {excel_path}")
    print(f"[INFO] TARGET XML   : {xml_path}")
    print("──────────────────────────────────────────────────────────────\n")

    # 1) Build lookup tables
    perfect_key, kr_only_key = build_lookup_tables(xml_path)

    # 2) Analyse Excel
    stats = analyse_excel(excel_path, perfect_key, kr_only_key)

    total      = stats["total"]
    perfect    = stats["perfect"]
    kronly     = stats["kr_only"]
    unmatched  = len(stats["unmatched_rows"])

    pct = (perfect + kronly) * 100 / total if total else 0.0

    print("\n======================  MATCH RESULT  =======================")
    print(f"Total rows analysed        : {total:,}")
    print(f"  ↳ Perfect  (KR+ID) match : {perfect:,}")
    print(f"  ↳ KR-only match          : {kronly:,}")
    print(f"  ↳ No match               : {unmatched:,}")
    print(f"-------------------------------------------------------------")
    print(f"   OVERALL HIT-RATE        : {pct:.2f} %")
    print("=============================================================\n")

    # 3) Colour unmatched rows & save
    colour_unmatched_rows(stats["workbook"], stats["worksheet"], stats["unmatched_rows"])

    base, ext = os.path.splitext(excel_path)
    out_path  = f"{base}_output.xlsx"
    stats["workbook"].save(out_path)
    print(f"[OK] Output written to  {out_path}\n")


# --------------------------------------------------------------------------
if __name__ == "__main__":
    main()