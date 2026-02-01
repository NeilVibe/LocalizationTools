#!/usr/bin/env python3
# coding: utf-8
"""
Simplified FindSequencer using VoiceRecordingSheet ordering.
No colors, reduced columns in output.
Sequencer sheet has headers from VoiceRecordingSheet (B, C, K, W).
AIDialog and NOT_FOUND have no headers.
NOT_FOUND contains original preprocessed names that had no match in W column.
"""

from __future__ import annotations
import sys
import re
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, scrolledtext
from typing import Dict, List
from openpyxl import load_workbook, Workbook

# --------------------------------------------------------------------------- #
# CONFIG
# --------------------------------------------------------------------------- #
SOURCE_FILE = Path(r"C:\Users\PEARL\Desktop\CD SCRIPTS\WorkingScripts\listID.txt")
VOICE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\editordata\VoiceRecordingSheet__")
EXCEL_OUTPUT = SOURCE_FILE.with_name("sequencer_output.xlsx")
# --------------------------------------------------------------------------- #

NAME_RE = re.compile(r"([^/\\]+?)\.[^/\\#]+(?:#[0-9]+)?$", re.IGNORECASE)

def extract_sound_event(raw_line: str) -> str | None:
    """
    Extracts the base sound event name from a line, ignoring comments (#).
    Returns None if line is empty or comment-only.
    """
    line = raw_line.strip()
    if not line or line.startswith("#"):
        return None
    line = line.split("#", 1)[0]
    m = NAME_RE.search(line)
    if m:
        return m.group(1)
    return None

def find_most_recent_excel(folder: Path) -> Path:
    files = list(folder.glob("*.xlsx"))
    if not files:
        raise RuntimeError(f"No Excel files found in {folder}")
    return max(files, key=lambda f: f.stat().st_mtime)

def regroup_by_voicerecordingsheet(textfile_path: Path, voice_folder: Path, excel_path: Path):
    # Load most recent VoiceRecordingSheet
    vr_file = find_most_recent_excel(voice_folder)
    wb_vr = load_workbook(vr_file, data_only=True)
    ws_vr = wb_vr.active

    # Read sound events from source file
    try:
        raw_lines = [ln.rstrip("\n") for ln in textfile_path.read_text(encoding="utf-8").splitlines()]
    except Exception as exc:
        raise RuntimeError(f"Cannot read {textfile_path}: {exc}")

    # Map: sound_event_name -> original line
    sound_events = {extract_sound_event(line): line
                    for line in raw_lines if extract_sound_event(line)}

    # Build mapping from VoiceRecordingSheet
    vr_rows = []
    for row in ws_vr.iter_rows(min_row=2, values_only=True):
        colA = str(row[0]).strip() if row[0] else ""
        colB = str(row[1]).strip() if row[1] else ""
        colC = str(row[2]).strip() if row[2] else ""
        colK = str(row[10]).strip() if row[10] else ""  # K = 11th column index (0-based 10)
        colW = str(row[22]).strip() if row[22] else ""  # W = 23rd column index (0-based 22)
        if colW:
            vr_rows.append((colA, colB, colC, colK, colW))

    # Match and group
    groups: Dict[str, List] = {"Sequencer": [], "AIDialog": [], "NOT_FOUND": []}
    matched_events = set()

    for colA, colB, colC, colK, colW in vr_rows:
        if colW in sound_events:
            matched_events.add(colW)
            if colA.lower() == "sequencer":
                groups["Sequencer"].append((colB, colC, colK, colW))
            elif colA.lower() == "aidialog":
                groups["AIDialog"].append(colW)
            else:
                groups["NOT_FOUND"].append(colW)  # non-sequencer, non-aidialog match

    # Add remaining unmatched from source file (preprocessed names)
    for sen in sound_events:
        if sen not in matched_events:
            groups["NOT_FOUND"].append(sen)

    # Excel output
    wb_out = Workbook()
    ws_seq = wb_out.active
    ws_seq.title = "Sequencers"
    ws_aidialog = wb_out.create_sheet("AIDialog")
    ws_unknown = wb_out.create_sheet("UNKNOWN")

    # Headers for Sequencer from VR sheet
    header_B = ws_vr.cell(row=1, column=2).value or "B"
    header_C = ws_vr.cell(row=1, column=3).value or "C"
    header_K = ws_vr.cell(row=1, column=11).value or "K"
    header_W = ws_vr.cell(row=1, column=23).value or "W"
    ws_seq.append([header_B, header_C, header_K, header_W])

    # Write Sequencer: B, C, K, W
    for b, c, k, w in groups["Sequencer"]:
        ws_seq.append([b, c, k, w])

    # Write AIDialog: only W
    for w in groups["AIDialog"]:
        ws_aidialog.append([w])

    # Write NOT_FOUND: original preprocessed names that had no match
    for name in groups["NOT_FOUND"]:
        ws_unknown.append([name])

    wb_out.save(excel_path)

###############################################################################
# GUI
###############################################################################
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sequencer Finder (VoiceRecordingSheet)")
        self.resizable(False, False)
        tk.Label(self, text=f"Source file:\n{SOURCE_FILE}", justify=tk.LEFT).pack(anchor="w", padx=10, pady=5)
        tk.Label(self, text=f"VoiceRecordingSheet folder:\n{VOICE_FOLDER}", justify=tk.LEFT).pack(anchor="w", padx=10)
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=15, fill="x", expand=True)
        tk.Button(btn_frame, text="Find Sequencer (Excel)", width=22,
                  command=self.run_sequencer_excel).pack(side="left", padx=20)
        self.log = scrolledtext.ScrolledText(self, height=6, state="disabled")
        self.log.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.update_idletasks()
        self.geometry(f"{self.winfo_reqwidth()}x{self.winfo_reqheight()}")

    def _log(self, msg: str):
        self.log["state"] = "normal"
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log["state"] = "disabled"

    def run_sequencer_excel(self):
        try:
            regroup_by_voicerecordingsheet(SOURCE_FILE, VOICE_FOLDER, EXCEL_OUTPUT)
            self._log(f"✓ Sequencer grouping Excel saved to {EXCEL_OUTPUT}")
            messagebox.showinfo("Done", f"Sequencer grouping saved to:\n{EXCEL_OUTPUT}")
        except Exception as exc:
            self._log(f"ERROR: {exc}")
            messagebox.showerror("Error", str(exc))

# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    if not SOURCE_FILE.exists():
        messagebox.showerror("Error", f"Source file not found:\n{SOURCE_FILE}")
        sys.exit(1)
    if not VOICE_FOLDER.exists():
        messagebox.showerror("Error", f"VoiceRecordingSheet folder not found:\n{VOICE_FOLDER}")
        sys.exit(1)
    App().mainloop()