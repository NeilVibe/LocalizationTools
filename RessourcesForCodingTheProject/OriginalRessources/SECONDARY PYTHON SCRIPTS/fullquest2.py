
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Quest-data extractor (Main / Faction / Challenge)
──────────────────────────────────────────────────────────────────────────────
• One ENG column for the English sheet.  
• For every other language sheets columns are:  
  [Source, ENG, <LANG>, StringKey, Teleport/Command]  
  
• Teleport logic (NPC / Quest locations)

  1.  Read every FactionNode in
        “...\GameData\StaticInfo\factioninfo”
      – use its  WorldPosition  as the authoritative teleport location.

  2.  While scanning AutoGenInfo__\*  FactionNodeSpawn records,
      map  FactionNodeKey  as well as StrKey  to the same teleport text
      that was found in step 1.

  3.  If a FactionNodeSpawn’s StrKey cannot be resolved to a FactionNode,
      the legacy fallback (MinPos / MaxPos / Spline-box) is attempted.

  Result:  Only sane coordinates end up in the Excel – no more gigantic
  placeholder numbers.

• Challenge quests:
    – For most groups: extract the most general ItemKey from EventCondition
      expressions and print it as “/create item <ItemKey>”.
    – BUT for the two special groups
          GroupName="어비스" (ABYSS)   and   GroupName="탐험" (EXPLORATION)
      the script falls back to teleport logic exactly like Main/Faction
      quests and ignores ItemKey extraction.

Full drop-in script below. Save as e.g.  quest_lqa_extract.py  and run.
"""

from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Iterable

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION – FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
QUESTGROUPINFO_FILE = Path(
    r"F:\perforce\cd\cd_delta\resource\GameData\staticinfo_quest\questgroupinfo.staticinfo.xml"
)
SCENARIO_FOLDER = Path(
    r"F:\perforce\cd\cd_delta\resource\GameData\staticinfo_quest\scenario"
)
FACTION_FOLDER = Path(
    r"F:\perforce\cd\cd_delta\resource\GameData\staticinfo_quest\quest\faction"
)
CHALLENGE_FOLDER = Path(
    r"F:\perforce\cd\cd_delta\resource\GameData\staticinfo_quest\Challenge"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_delta\resource\GameData\stringtable\loc"
)
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\cd_delta\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)
#   ──  AutoGenInfo__        (FactionNodeSpawn  ↔  FactionNodeKey)
TELEPORT_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\AutoGenInfo__"
)
#   ──  factioninfo          (FactionNode  ↔  WorldPosition)
FACTIONINFO_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\factioninfo"
)
OUTPUT_FOLDER = Path.cwd() / "QuestData_Map_All"
LOG_FILE      = Path.cwd() / "quest_scan.log"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING  – FILE + CONSOLE
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("QuestLQA")
log.setLevel(logging.DEBUG)
_file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
_file_handler.setFormatter(
    logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s",
                      datefmt="%H:%M:%S")
)
_file_handler.setLevel(logging.DEBUG)
_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)
log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ─────────────────────────────────────────────────────────────────────────────
# XML-SANITISATION UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')
def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)

def preprocess_newlines_in_tags(raw_content: str) -> str:
    def repl(match: re.Match) -> str:
        inner = match.group(1)
        inner = inner.replace("\n", "&lt;br/&gt;").replace("\\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw_content, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = preprocess_newlines_in_tags(raw)
    raw = re.sub(
        r'="([^"]*<[^"]*)"', lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw
    )
    raw = re.sub(
        r'="([^"]*&[^ltgapoqu][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )
    tag_stack: List[str] = []
    tag_open_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close_re = re.compile(r"</([A-Za-z0-9_]+)>")
    fixed_lines: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        m_open = tag_open_re.match(stripped)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed_lines.append(line)
            continue
        m_close = tag_close_re.match(stripped)
        if m_close:
            if tag_stack and tag_stack[-1] == m_close.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            else:
                fixed_lines.append(line)
            continue
        fixed_lines.append(line)
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Failed to read %s", path)
        return None
    fixed = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{fixed}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict XML parse failed (%s). Retrying with recovery ...", path.name)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"), parser=ET.XMLParser(recover=True, huge_tree=True)
            )
        except ET.XMLSyntaxError:
            log.exception("XML recovery parse also failed for %s", path)
            return None

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE  LOADING
# ─────────────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn

def parse_language_folder(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    log.info("Scanning language folder: %s", folder)
    languages: Dict[str, Dict[str, Tuple[str, str]]] = {}
    paths = list(iter_xml_files(folder))
    for idx, path in enumerate(paths, 1):
        name_no_ext = path.stem.lower()
        if (not name_no_ext.startswith("languagedata_")
                or name_no_ext.startswith("languagedata_kor")):
            continue
        code = name_no_ext.split("_", 1)[1]
        log.info("Loading language file [%s] → %s", code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            log.error("Failed to load language file: %s", path)
            continue
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            sid    = el.get("StringId")  or ""
            trans  = el.get("Str")       or ""
            if origin:
                tbl[origin] = (trans, sid)
        languages[code.lower()] = tbl
    log.info("Language tables loaded: %d", len(languages))
    return languages

# ─────────────────────────────────────────────────────────────────────────────
# STRING-KEY TABLE
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot proceed without a valid StringKeyTable")
    mapping: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        key = el.get("Key") or ""
        sk  = el.get("StrKey") or ""
        if key and sk:
            mapping[sk.lower()] = key
    log.info("StringKeyTable loaded. Entries: %d", len(mapping))
    return mapping

# ─────────────────────────────────────────────────────────────────────────────
# QUEST-GROUP META FOR CHALLENGE
# ─────────────────────────────────────────────────────────────────────────────
def parse_group_meta(path: Path) -> Dict[str, str]:
    log.info("Parsing QuestGroup meta for Challenge groups: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse group meta")
    meta: Dict[str, str] = {}
    for el in root.iter("QuestGroup"):
        code = (el.get("StrKey") or "").lower()
        name = el.get("Name") or ""
        if code:
            meta[code] = name
    log.info("Challenge group meta loaded: %d", len(meta))
    return meta

# ─────────────────────────────────────────────────────────────────────────────
# TELEPORT DATA  (NEW ‑ WORLD-POSITION BASED)
# ─────────────────────────────────────────────────────────────────────────────
_BIG_COORD_THRESHOLD = 1_000_000.0      # Ignore obviously invalid coords
_PLACEHOLDER_LIMIT   = 1_000_000_000.0  # Numbers bigger than this are placeholders

def _valid_coords(vals: List[float]) -> bool:
    return all(abs(v) < _BIG_COORD_THRESHOLD for v in vals)

def _parse_vec3(s: str) -> Optional[Tuple[float, float, float]]:
    """
    Robust vec3 parser – accepts  'x,y,z',  'x, y, z',  'x y z'
    """
    s = s.replace(",", " ")
    parts = [p.strip() for p in s.split() if p.strip()]
    try:
        nums = [float(p) for p in parts]
        if len(nums) == 3:
            return tuple(nums)  # type: ignore[return-value]
    except Exception:
        pass
    return None

def _midpoint(a: Tuple[float, float, float], b: Tuple[float, float, float]) -> Tuple[float, float, float]:
    return tuple((x + y) / 2.0 for x, y in zip(a, b))  # type: ignore[arg-type]

def _coords_from_spline(el: ET._Element) -> Optional[Tuple[float, float, float]]:
    points: List[Tuple[float, float, float]] = []
    for sp in el.iter("SplinePoint"):
        pos_raw = sp.get("Position")
        if not pos_raw:
            continue
        vec = _parse_vec3(pos_raw)
        if vec and _valid_coords(list(vec)):
            points.append(vec)
    if not points:
        return None
    xs, ys, zs = zip(*points)
    min_vec = (min(xs), min(ys), min(zs))
    max_vec = (max(xs), max(ys), max(zs))
    return _midpoint(min_vec, max_vec)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 – WORLD-POSITION TABLE FROM  factioninfo
# ─────────────────────────────────────────────────────────────────────────────
def _build_factionnode_position_table(folder: Path) -> Dict[str, Tuple[float, float, float]]:
    """
    Returns:  StrKey.lower()  →  (x, y, z)
    """
    log.info("Scanning faction-node positions in: %s", folder)
    pos_tbl: Dict[str, Tuple[float, float, float]] = {}
    for idx, path in enumerate(iter_xml_files(folder), 1):
        log.debug("FactionInfo XML (%d): %s", idx, path.name)
        root = parse_xml_file(path)
        if root is None:
            continue
        for node in root.iter("FactionNode"):
            sk = node.get("StrKey") or ""
            wp_raw = (
                node.get("WorldPosition")
                or node.get("WorldPos")
                or node.get("Position")
                or ""
            )
            if not (sk and wp_raw):
                continue
            vec = _parse_vec3(wp_raw)
            if vec and _valid_coords(list(vec)):
                pos_tbl[sk.lower()] = vec
    log.info("FactionNode positions parsed: %d", len(pos_tbl))
    return pos_tbl

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 – TELEPORT TABLE  (AUTO-GEN + FACTIONINFO)
# ─────────────────────────────────────────────────────────────────────────────
def parse_teleport_data(
    auto_gen_folder: Path,
    factioninfo_folder: Path
) -> Dict[str, str]:
    """
    Builds a map:

        •  StrKey.lower()         → '/teleport x y z'
        •  FactionNodeKey.lower() → '/teleport x y z'
    """
    faction_pos = _build_factionnode_position_table(factioninfo_folder)
    table: Dict[str, str] = {}

    xml_files = list(iter_xml_files(auto_gen_folder))
    log.info("Parsing AutoGenInfo__ for FactionNodeSpawn: %d files", len(xml_files))

    for idx, path in enumerate(xml_files, 1):
        log.debug("AutoGenInfo XML (%d/%d): %s", idx, len(xml_files), path.name)
        root = parse_xml_file(path)
        if root is None:
            continue

        for el in root.iter("FactionNodeSpawn"):
            str_key = el.get("StrKey") or ""
            node_id = el.get("FactionNodeKey") or ""

            chosen_vec: Optional[Tuple[float, float, float]] = None

            # Preferred: lookup via faction info table
            if str_key and str_key.lower() in faction_pos:
                chosen_vec = faction_pos[str_key.lower()]
            else:
                min_vec = _parse_vec3(el.get("MinPos") or "")
                max_vec = _parse_vec3(el.get("MaxPos") or "")
                if min_vec and max_vec and _valid_coords(list(min_vec)) and _valid_coords(list(max_vec)):
                    chosen_vec = _midpoint(min_vec, max_vec)
                else:
                    chosen_vec = _coords_from_spline(el)

            if not chosen_vec:
                log.debug("No usable coords for node %s (file: %s)", str_key or node_id, path.name)
                continue

            cmd = "/teleport " + " ".join(f"{c:g}" for c in chosen_vec)

            if str_key:
                table[str_key.lower()] = cmd
            if node_id:
                table[node_id.lower()] = cmd

    log.info("Teleport table built. Entries: %d", len(table))
    return table

# ─────────────────────────────────────────────────────────────────────────────
# QUEST-GROUP INFO (Main)
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_quests(path: Path) -> Tuple[List[str], Dict[str, str]]:
    log.info("Parsing QuestGroupInfo: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Failed to parse QuestGroupInfo")
    quest_order: List[str] = []
    quest_groups: Dict[str, str] = {}
    seen: set[str] = set()
    for el in root.iter():
        grp = el.get("Group")
        if grp:
            qkey = el.tag
            if qkey not in seen:
                quest_order.append(qkey)
                quest_groups[qkey] = grp
                seen.add(qkey)
    log.info(
        "QuestGroupInfo parsed – Quests: %d | Unique Groups: %d",
        len(quest_order),
        len(set(quest_groups.values())),
    )
    return quest_order, quest_groups

# ─────────────────────────────────────────────────────────────────────────────
# GENERIC  UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
RowItem = Tuple[int, str, str, str, str, bool, bool, str]

_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),
    1: PatternFill("solid", fgColor="D9E1F2"),
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FCE4D6"),
    4: PatternFill("solid", fgColor="9C27B0"),
}
_icon_fill     = PatternFill("solid", fgColor="9BC2E6")
_light_yellow  = PatternFill("solid", fgColor="FFFDEB")
_border_template = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
bold_font   = Font(bold=True)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")

def _translate(text: str, lang_tbl: Dict[str, Tuple[str, str]]) -> str:
    return lang_tbl.get(text, ("", ""))[0]

def _tp_lookup(tele_tbl: Dict[str, str], *keys: str) -> str:
    for k in keys:
        if k:
            t = tele_tbl.get(k.lower())
            if t:
                return t
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# CHALLENGE  ITEM / TELEPORT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
_itemkey_re = re.compile(r'ItemKey\(\s*([A-Za-z0-9_]+)\s*\)')

# Groups that MUST use teleport instead of item commands
_CHALLENGE_TP_GROUP_NAMES = {"어비스", "탐험"}

def _extract_item_command(element: ET._Element) -> str:
    """
    Searches for the right-most ItemKey(...) in any EventCondition attribute
    and returns '/create item <ItemKey>' or '' if none found.
    """
    last_key: Optional[str] = None
    for sub in element.iter():
        for attr_name, attr_val in sub.attrib.items():
            if "EventCondition" in attr_name or attr_name == "EventCondition":
                matches = _itemkey_re.findall(attr_val)
                if matches:
                    last_key = matches[-1]
    if last_key:
        return f"/create item {last_key}"
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# ROW BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
def build_rows_main(
    scenario_folder: Path,
    quest_order: List[str],
    quest_groups: Dict[str, str],
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str],
    teleport_tbl: Dict[str, str],
) -> List[RowItem]:
    log.info("Building rows for Main Quests ...")
    quest_elements: Dict[str, ET._Element] = {}
    xml_files = list(scenario_folder.rglob("*.xml"))
    for idx, xml_path in enumerate(xml_files, 1):
        log.debug("MainQuest XML (%d/%d): %s", idx, len(xml_files), xml_path.name)
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        for q in root.iter():
            if q.get("Name") and q.get("StartMission"):
                quest_elements[q.tag] = q

    rows: List[RowItem] = []
    current_group = None
    for quest_key in quest_order:
        q_el = quest_elements.get(quest_key)
        if q_el is None:
            continue
        grp_code = quest_groups.get(quest_key, "")
        if grp_code != current_group:
            current_group = grp_code
            grp_kor = grp_code
            rows.append((
                0,
                grp_kor,
                _translate(grp_kor, eng_tbl),
                _translate(grp_kor, lang_tbl),
                "",
                False,
                True,
                ""
            ))
        quest_kor = q_el.get("Name") or ""
        sk_raw = quest_key
        sk_map = id_table.get(sk_raw.lower(), "")
        tp_cmd = _tp_lookup(teleport_tbl, sk_raw, sk_map)
        rows.append((
            1,
            quest_kor,
            _translate(quest_kor, eng_tbl),
            _translate(quest_kor, lang_tbl),
            sk_map,
            bool(q_el.get("StageIcon")),
            True,
            tp_cmd
        ))
        for mission in q_el.findall("Mission"):
            m_kor = mission.get("Name") or ""
            m_sk_raw = mission.get("StrKey") or ""
            m_key    = id_table.get(m_sk_raw.lower(), m_sk_raw)
            tp_cmd   = _tp_lookup(teleport_tbl, m_sk_raw, m_key)
            rows.append((
                2,
                m_kor,
                _translate(m_kor, eng_tbl),
                _translate(m_kor, lang_tbl),
                m_key,
                False,
                True,
                tp_cmd
            ))
            for sub in mission.findall("SubMission"):
                s_kor    = sub.get("Name") or ""
                s_sk_raw = sub.get("StrKey") or ""
                s_key    = id_table.get(s_sk_raw.lower(), s_sk_raw) if s_sk_raw else ""
                tp_cmd   = _tp_lookup(teleport_tbl, s_sk_raw, s_key)
                rows.append((
                    3,
                    s_kor,
                    _translate(s_kor, eng_tbl),
                    _translate(s_kor, lang_tbl),
                    s_key,
                    False,
                    True,
                    tp_cmd
                ))
    log.info("Main Quest rows built: %d", len(rows))
    return rows

def build_rows_simple(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str],
    teleport_tbl: Dict[str, str],
) -> List[RowItem]:
    log.info("Building rows for folder: %s", folder.name)
    rows: List[RowItem] = []
    xml_files = list(folder.rglob("*.xml"))
    for idx, xml_path in enumerate(xml_files, 1):
        log.debug("SimpleQuest XML (%d/%d): %s", idx, len(xml_files), xml_path.name)
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        for q_el in root.iter():
            if not (q_el.get("Name") and q_el.get("StartMission")):
                continue
            quest_kor = q_el.get("Name") or ""
            sk_raw    = q_el.tag
            sk_map    = id_table.get(sk_raw.lower(), "")
            tp_cmd    = _tp_lookup(teleport_tbl, sk_raw, sk_map)
            rows.append((
                0,
                quest_kor,
                _translate(quest_kor, eng_tbl),
                _translate(quest_kor, lang_tbl),
                sk_map,
                bool(q_el.get("StageIcon")),
                True,
                tp_cmd
            ))
            for mission in q_el.findall("Mission"):
                m_kor    = mission.get("Name") or ""
                m_sk_raw = mission.get("StrKey") or ""
                m_key    = id_table.get(m_sk_raw.lower(), m_sk_raw)
                tp_cmd   = _tp_lookup(teleport_tbl, m_sk_raw, m_key)
                rows.append((
                    1,
                    m_kor,
                    _translate(m_kor, eng_tbl),
                    _translate(m_kor, lang_tbl),
                    m_key,
                    False,
                    True,
                    tp_cmd
                ))
                for sub in mission.findall("SubMission"):
                    s_kor    = sub.get("Name") or ""
                    s_sk_raw = sub.get("StrKey") or ""
                    s_key    = id_table.get(s_sk_raw.lower(), s_sk_raw) if s_sk_raw else ""
                    tp_cmd   = _tp_lookup(teleport_tbl, s_sk_raw, s_key)
                    rows.append((
                        2,
                        s_kor,
                        _translate(s_kor, eng_tbl),
                        _translate(s_kor, lang_tbl),
                        s_key,
                        False,
                        True,
                        tp_cmd
                    ))
    log.info("Rows built for %s: %d", folder.name, len(rows))
    return rows

def build_rows_challenge(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str],
    teleport_tbl: Dict[str, str],
    group_meta: Dict[str, str],
) -> List[RowItem]:
    """
    Builds rows for Challenge quests.

    – Default behaviour: use item-creation command extracted from EventCondition.
    – For groups with Korean names in _CHALLENGE_TP_GROUP_NAMES, use teleport
      coordinates (same lookup as Main/Faction quests). If no teleport found,
      leave command blank.
    """
    log.info("Building rows for Challenge Quests ...")
    rows: List[RowItem] = []
    xml_files = list(folder.rglob("*.xml"))
    for idx, xml_path in enumerate(xml_files, 1):
        log.debug("Challenge XML (%d/%d): %s", idx, len(xml_files), xml_path.name)
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        for q_el in root:
            if not (q_el.get("Name") and q_el.get("StartMission")):
                continue

            grp_code = (q_el.get("Group") or "").lower()
            grp_name = group_meta.get(grp_code, grp_code)
            use_tp   = grp_name in _CHALLENGE_TP_GROUP_NAMES

            # Insert group header row
            rows.append((
                4,
                grp_name,
                _translate(grp_name, eng_tbl),
                _translate(grp_name, lang_tbl),
                "",
                False,
                True,
                ""
            ))

            # Quest row
            quest_kor  = q_el.get("Name") or ""
            sk_raw     = q_el.tag
            sk_map     = id_table.get(sk_raw.lower(), "")
            command_q  = (_tp_lookup(teleport_tbl, sk_raw, sk_map)
                          if use_tp else _extract_item_command(q_el))
            rows.append((
                0,
                quest_kor,
                _translate(quest_kor, eng_tbl),
                _translate(quest_kor, lang_tbl),
                sk_map,
                bool(q_el.get("StageImage")),
                True,
                command_q
            ))

            # Missions & SubMissions
            for mission in q_el.findall("Mission"):
                m_kor    = mission.get("Name") or ""
                m_sk_raw = mission.get("StrKey") or ""
                m_key    = id_table.get(m_sk_raw.lower(), m_sk_raw)
                command_m = (_tp_lookup(teleport_tbl, m_sk_raw, m_key)
                             if use_tp else _extract_item_command(mission))
                rows.append((
                    1,
                    m_kor,
                    _translate(m_kor, eng_tbl),
                    _translate(m_kor, lang_tbl),
                    m_key,
                    False,
                    True,
                    command_m
                ))
                for sub in mission.findall("SubMission"):
                    s_kor    = sub.get("Name") or ""
                    s_sk_raw = sub.get("StrKey") or ""
                    s_key    = id_table.get(s_sk_raw.lower(), s_sk_raw) if s_sk_raw else ""
                    command_s = (_tp_lookup(teleport_tbl, s_sk_raw, s_key)
                                 if use_tp else _extract_item_command(sub))
                    rows.append((
                        2,
                        s_kor,
                        _translate(s_kor, eng_tbl),
                        _translate(s_kor, lang_tbl),
                        s_key,
                        False,
                        True,
                        command_s
                    ))
    log.info("Challenge Quest rows built: %d", len(rows))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL  RENDERING
# ─────────────────────────────────────────────────────────────────────────────
def autofit(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len * 1.15 + 2, 80)

def write_sheet(wb: Workbook, title: str, rows: List[RowItem], tgt_code: str) -> None:
    log.info("Creating sheet: %s  (rows: %d)", title, len(rows))
    ws = wb.create_sheet(title=title[:31])
    is_eng = tgt_code.lower().startswith("eng")
    if is_eng:
        headers = ["Source Text", "ENG", "StringKey", "Command"]
    else:
        headers = ["Source Text", "ENG", tgt_code.upper(), "StringKey", "Command"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = _border_template
    indent_threshold = headers.index("StringKey") + 1
    r = 2
    for depth, kor, eng, loc, skey, is_icon, is_name_attr, command in rows:
        fg = _depth_fill.get(depth)
        if fg is None and is_icon:
            fg = _icon_fill
        vals = (kor, eng, skey, command) if is_eng else (kor, eng, loc, skey, command)
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(r, col_idx, val)
            indent = depth if col_idx < indent_threshold else 0
            cell.alignment = Alignment(indent=indent, wrap_text=True)
            cell.border = _border_template
            cell.fill = fg if fg else _light_yellow
            if fg or is_name_attr:
                cell.font = bold_font
        r += 1
    ws.freeze_panes = "A2"
    autofit(ws)
    log.debug("Sheet '%s' completed", title)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("Quest LQA extraction started")
    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # 1. Load languages
    all_languages = parse_language_folder(LANGUAGE_FOLDER)
    if not all_languages:
        sys.exit("No language files could be loaded!")

    eng_code = next((c for c in all_languages if c.lower().startswith("eng")), None)
    if not eng_code:
        sys.exit("English language file (languagedata_ENG*.xml) not found.")
    eng_tbl = all_languages[eng_code]

    # 2. Load StringKey table
    id_table = load_string_key_table(STRINGKEYTABLE_FILE)

    # 3. Parse master quest groups and challenge group meta
    quest_order, quest_groups = parse_master_quests(QUESTGROUPINFO_FILE)
    group_meta = parse_group_meta(QUESTGROUPINFO_FILE)

    # 4. Parse teleport data (for Main / Faction / Challenge:ABYSS,EXPLORATION)
    teleport_table = parse_teleport_data(TELEPORT_FOLDER, FACTIONINFO_FOLDER)

    # 5. Generate Excel per language
    for idx, (code, lang_tbl) in enumerate(all_languages.items(), 1):
        log.info("(%d/%d) Processing language: %s", idx, len(all_languages), code.upper())

        rows_main      = build_rows_main(
            SCENARIO_FOLDER, quest_order, quest_groups,
            lang_tbl, eng_tbl, id_table, teleport_table
        )
        rows_faction   = build_rows_simple(
            FACTION_FOLDER, lang_tbl, eng_tbl,
            id_table, teleport_table
        )
        rows_challenge = build_rows_challenge(
            CHALLENGE_FOLDER, lang_tbl, eng_tbl,
            id_table, teleport_table, group_meta
        )

        wb = Workbook()
        wb.remove(wb.active)
        write_sheet(wb, "Main Quest",      rows_main,      code)
        write_sheet(wb, "Faction Quest",   rows_faction,   code)
        write_sheet(wb, "Challenge Quest", rows_challenge, code)

        out_path = OUTPUT_FOLDER / f"Quest_LQA_{code.upper()}.xlsx"
        wb.save(out_path)
        log.info("Excel saved: %s", out_path)

    log.info("All done! Extractions completed for %d language(s).", len(all_languages))

if __name__ == "__main__":
    main()
