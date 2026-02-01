#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Quest-data extractor (Main / Faction / Challenge).

 • Uses the same aggressive XML-sanitising / recovery technique as the
   “Knowledge-LQA” script.
 • Produces ONE worksheet per quest-type:
        ①  Main Quest      – data from <...\scenario>
        ②  Faction Quest   – data from <...\quest\faction>
        ③  Challenge Quest – data from <...\Challenge>
 • Rows are rendered with depth-colours, bolding, indentation, borders, light
   fallback fill, exactly like the Knowledge-LQA script:
        depth = 0 → yellow       (chapter / quest-group)
        depth = 1 → light-blue   (quest)
        depth = 2 → light-green  (mission)
        depth = 3 → light-orange (sub-mission)
        blue override when         StageIcon is present but no depth-colour.
 • Columns:  A = Original (KOR)   B = ENG translation   C = StringKey / ID
 • Quest-group information (chapters) comes from QuestGroupInfo ONLY for
   “Main Quest” sheet – faction / challenge sheets rely solely on their own
   XML hierarchy.
"""

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION –- FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
QUESTGROUPINFO_FILE = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\questgroupinfo.staticinfo.xml"
)
SCENARIO_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\scenario"
)
FACTION_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\quest\faction"
)
CHALLENGE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\Challenge"
)

LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
)                       # ENG only
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\mainline\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)

OUTPUT_EXCEL = Path.cwd() / "Quest_LQA_ENG.xlsx"
LOG_FILE     = Path.cwd() / "quest_scan.log"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")]
)
log = logging.getLogger("QuestLQA")

# ─────────────────────────────────────────────────────────────────────────────
# XML-SANITISATION UTILITIES  (identical to Knowledge-LQA implementation)
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)

def preprocess_newlines_in_tags(raw_content: str) -> str:
    def repl(match: re.Match) -> str:
        inner = match.group(1)
        inner = inner.replace("\n", "&lt;br/&gt;").replace("\\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw_content, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = preprocess_newlines_in_tags(raw)
    # escape stray "<" or "&" occurring inside attribute values
    raw = re.sub(
        r'="([^"]*<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw
    )
    raw = re.sub(
        r'="([^"]*&[^ltgapoqu][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw
    )
    # quick & dirty unbalanced-tag fixer
    tag_stack: List[str] = []
    tag_open_re  = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed_lines: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        m_open = tag_open_re.match(stripped)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed_lines.append(line)
            continue
        m_close = tag_close_re.match(stripped)
        if m_close:
            if tag_stack and tag_stack[-1] == m_close.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                # wrong closing tag – pop until match
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(
                    line.replace("</>", f"</{tag_stack.pop()}>")
                )
            else:
                fixed_lines.append(line)
            continue
        fixed_lines.append(line)
    # close whatever is still open
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    """
    Fully sanitise the XML text, wrap in <ROOT>, parse strictly (recover=False);
    if strict parse fails, fall back to recover=True.
    """
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Failed to read %s", path)
        return None
    fixed = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{fixed}\n</ROOT>"
    try:
        return ET.fromstring(
            wrapped.encode("utf-8"),
            parser=ET.XMLParser(recover=False, huge_tree=True)
        )
    except ET.XMLSyntaxError:
        log.exception("Strict parse failed for %s – retrying with recovery", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True)
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse also failed for %s", path)
            return None

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE – ENG ONLY  (StrOrigin → (ENG, StringId))
# ─────────────────────────────────────────────────────────────────────────────
def load_eng_language(folder: Path) -> Dict[str, Tuple[str, str]]:
    for xml_path in folder.rglob("LanguageData_eng.xml"):
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        table: Dict[str, Tuple[str, str]] = {}
        for loc in root.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            trans  = loc.get("Str") or ""
            sid    = loc.get("StringId") or ""
            if origin:
                table[origin] = (trans, sid)
        log.info("Loaded ENG language table: %d entries", len(table))
        return table
    log.warning("No LanguageData_eng.xml found under %s – translations blank", folder)
    return {}

# ─────────────────────────────────────────────────────────────────────────────
# STRINGKEY TABLE  (lower-case StrKey → numeric Key)
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot proceed without StringKeyTable")
    mapping: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        key = el.get("Key") or ""
        sk  = el.get("StrKey") or ""
        if key and sk:
            mapping[sk.lower()] = key
    log.info("Loaded StringKeyTable: %d entries", len(mapping))
    return mapping

# ─────────────────────────────────────────────────────────────────────────────
# QUEST-GROUP INFO  (only for Main Quest sheet)
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_quests(path: Path) -> Tuple[List[str], Dict[str, str]]:
    """
    Returns:
        quest_order  – list of quest element names in original order
        quest_group  – dict quest_name → group-title (chapter)
    """
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Failed to parse QuestGroupInfo")
    quest_order: List[str] = []
    quest_groups: Dict[str, str] = {}
    seen: set[str] = set()
    for el in root.iter():
        grp = el.get("Group")
        if grp is not None:
            qkey = el.tag
            if qkey not in seen:
                quest_order.append(qkey)
                quest_groups[qkey] = grp
                seen.add(qkey)
    log.info("QuestGroupInfo: %d quests across %d groups",
             len(quest_order), len(set(quest_groups.values())))
    return quest_order, quest_groups

# ─────────────────────────────────────────────────────────────────────────────
# GENERIC QUEST SCANNER  (build RowItem list)
# ─────────────────────────────────────────────────────────────────────────────
RowItem = Tuple[int, str, str, str, bool, bool]        # depth, kor, eng, strKey, is_icon, is_name_attr

_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),  # group / chapter
    1: PatternFill("solid", fgColor="D9E1F2"),  # quest
    2: PatternFill("solid", fgColor="E2EFDA"),  # mission
    3: PatternFill("solid", fgColor="FCE4D6"),  # sub-mission
}
_icon_fill       = PatternFill("solid", fgColor="9BC2E6")
_light_yellow    = PatternFill("solid", fgColor="FFFDEB")
_border_template = Border(
    left  = Side(style="thin"),
    right = Side(style="thin"),
    top   = Side(style="thin"),
    bottom= Side(style="thin")
)
bold_font        = Font(bold=True)
header_font      = Font(bold=True, color="FFFFFF")
header_fill      = PatternFill("solid", fgColor="4F81BD")

def _translate(text: str, eng_tbl: Dict[str, Tuple[str, str]]) -> str:
    return eng_tbl.get(text, ("", ""))[0]

def _str_id(text: str, eng_tbl: Dict[str, Tuple[str, str]]) -> str:
    return eng_tbl.get(text, ("", ""))[1]

def build_rows_main(
    scenario_folder: Path,
    quest_order: List[str],
    quest_groups: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str]
) -> List[RowItem]:
    """
    Build nested rows for MAIN QUESTS only (with chapter / group row).
    """
    # First, map quest_key → <Element>
    quest_elements: Dict[str, ET._Element] = {}
    for xml_path in scenario_folder.rglob("*.xml"):
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        for q in root.iter():
            if q.get("Name") and q.get("StartMission"):
                quest_elements[q.tag] = q

    rows: List[RowItem] = []
    current_group = None
    for quest_key in quest_order:
        q_el = quest_elements.get(quest_key)
        if q_el is None:
            log.warning("Quest %s declared in QuestGroupInfo but not found in scenario XMLs", quest_key)
            continue

        # group row (depth 0)
        grp_name = quest_groups.get(quest_key, "")
        if grp_name != current_group:
            current_group = grp_name
            rows.append((
                0,
                grp_name,
                _translate(grp_name, eng_tbl),
                "",            # no StrKey
                False,         # icon?
                True           # bold
            ))

        # quest row (depth 1)
        quest_name_kor = q_el.get("Name") or ""
        rows.append((
            1,
            quest_name_kor,
            _translate(quest_name_kor, eng_tbl),
            id_table.get(quest_key.lower(), ""),
            bool(q_el.get("StageIcon")),
            True
        ))

        # missions (depth 2)
        for mission in q_el.findall("Mission"):
            m_name_kor = mission.get("Name") or ""
            m_str_key  = mission.get("StrKey") or ""
            rows.append((
                2,
                m_name_kor,
                _translate(m_name_kor, eng_tbl),
                id_table.get(m_str_key.lower(), m_str_key),
                False,
                True
            ))
            # sub-missions (depth 3)
            for sub in mission.findall("SubMission"):
                s_name = sub.get("Name") or ""
                s_str  = sub.get("StrKey") or ""      # rarely present
                rows.append((
                    3,
                    s_name,
                    _translate(s_name, eng_tbl),
                    id_table.get(s_str.lower(), s_str) if s_str else "",
                    False,
                    True
                ))
    log.info("MainQuest rows: %d", len(rows))
    return rows

def build_rows_simple(
    folder: Path,
    eng_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str]
) -> List[RowItem]:
    """
    Build rows for Faction / Challenge folders (no external chapter info).
    depth 0 = quest
    depth 1 = mission
    depth 2 = sub-mission
    """
    rows: List[RowItem] = []
    for xml_path in folder.rglob("*.xml"):
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        for q_el in root.iter():
            if not (q_el.get("Name") and q_el.get("StartMission")):
                continue
            quest_kor = q_el.get("Name") or ""
            rows.append((
                0,
                quest_kor,
                _translate(quest_kor, eng_tbl),
                id_table.get(q_el.tag.lower(), ""),
                bool(q_el.get("StageIcon")),
                True
            ))
            for mission in q_el.findall("Mission"):
                m_name = mission.get("Name") or ""
                m_key  = mission.get("StrKey") or ""
                rows.append((
                    1,
                    m_name,
                    _translate(m_name, eng_tbl),
                    id_table.get(m_key.lower(), m_key),
                    False,
                    True
                ))
                for sub in mission.findall("SubMission"):
                    s_name = sub.get("Name") or ""
                    s_key  = sub.get("StrKey") or ""
                    rows.append((
                        2,
                        s_name,
                        _translate(s_name, eng_tbl),
                        id_table.get(s_key.lower(), s_key) if s_key else "",
                        False,
                        True
                    ))
    log.info("%s rows: %d", folder.name, len(rows))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL RENDERING  (identical look to Knowledge-LQA, plus StrKey column)
# ─────────────────────────────────────────────────────────────────────────────
def autofit(ws) -> None:
    for col in ws.iter_cols(min_col=1, max_col=3):
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len * 1.15 + 2, 80)

def write_sheet(
    wb: Workbook,
    title: str,
    rows: List[RowItem]
) -> None:
    ws = wb.create_sheet(title=title[:31])
    # header
    ws.cell(1, 1, "Original").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).border = _border_template
    ws.cell(1, 2, "ENG Translation").font = header_font
    ws.cell(1, 2).fill = header_fill
    ws.cell(1, 2).border = _border_template
    ws.cell(1, 3, "StringKey / ID").font = header_font
    ws.cell(1, 3).fill = header_fill
    ws.cell(1, 3).border = _border_template

    r = 2
    for depth, kor, eng, skey, is_icon, is_name_attr in rows:
        fg = _depth_fill.get(depth)
        if fg is None and is_icon:
            fg = _icon_fill
        # column A – original
        ca = ws.cell(r, 1, kor)
        ca.alignment = Alignment(indent=depth, wrap_text=True)
        ca.border    = _border_template
        ca.fill      = fg if fg else _light_yellow
        if fg or is_name_attr:
            ca.font = bold_font
        # column B – ENG
        cb = ws.cell(r, 2, eng)
        cb.alignment = Alignment(indent=depth, wrap_text=True)
        cb.border    = _border_template
        cb.fill      = fg if fg else _light_yellow
        if fg or is_name_attr:
            cb.font = bold_font
        # column C – StrKey / ID
        cc = ws.cell(r, 3, skey)
        cc.fill      = fg if fg else _light_yellow
        cc.alignment = Alignment(wrap_text=True)
        cc.border    = _border_template
        if fg or is_name_attr:
            cc.font = bold_font
        r += 1
    # cosmetics
    ws.freeze_panes = "A2"
    autofit(ws)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("Quest LQA extraction started")
    eng_tbl   = load_eng_language(LANGUAGE_FOLDER)
    id_table  = load_string_key_table(STRINGKEYTABLE_FILE)

    # MAIN QUEST
    quest_order, quest_groups = parse_master_quests(QUESTGROUPINFO_FILE)
    rows_main = build_rows_main(
        SCENARIO_FOLDER,
        quest_order,
        quest_groups,
        eng_tbl,
        id_table
    )

    # FACTION
    rows_faction = build_rows_simple(
        FACTION_FOLDER,
        eng_tbl,
        id_table
    )

    # CHALLENGE
    rows_challenge = build_rows_simple(
        CHALLENGE_FOLDER,
        eng_tbl,
        id_table
    )

    # WRITE EXCEL
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Main Quest",      rows_main)
    write_sheet(wb, "Faction Quest",   rows_faction)
    write_sheet(wb, "Challenge Quest", rows_challenge)
    wb.save(OUTPUT_EXCEL)
    log.info("Excel saved → %s", OUTPUT_EXCEL)

if __name__ == "__main__":
    main()