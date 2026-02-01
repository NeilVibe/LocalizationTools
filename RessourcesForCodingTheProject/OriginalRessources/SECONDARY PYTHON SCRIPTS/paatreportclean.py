import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def process_file():
    file_path = filedialog.askopenfilename(
        title="Select a text file",
        filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
    )
    if not file_path:
        return  # User cancelled

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Regex to match "not a real issue" case-insensitive
        pattern = re.compile(r'not a real issue', re.IGNORECASE)
        filtered_rows = []

        for line in lines:
            # Remove leading/trailing whitespace and newlines
            line = line.strip()
            if not line:
                continue  # Skip empty lines

            # Split by pipe, but keep empty fields (e.g., "||" -> ['', ''])
            # Remove leading/trailing pipes before splitting
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            fields = [field.strip() for field in line.split('|')]

            # Ensure at least 6 columns (pad with empty strings if needed)
            while len(fields) < 6:
                fields.append('')

            # If the 5th column contains "not a real issue", skip
            if pattern.search(fields[4]):
                continue

            # Prepare row: [STRING ID, KR, TRANSLATION, FLAGTYPE, CORRECTION]
            row = [
                fields[0],
                fields[1],
                fields[2],
                fields[3],
                fields[5]
            ]
            filtered_rows.append(row)

        if not filtered_rows:
            messagebox.showinfo("No Data", "No rows found after filtering.")
            return

        # Prepare Excel output path
        base, ext = os.path.splitext(file_path)
        excel_path = f"{base}_clean.xlsx"

        # Write to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Data"

        # Header with bold font and light blue fill
        header = ["STRING ID", "KR", "TRANSLATION", "FLAGTYPE", "CORRECTION"]
        ws.append(header)
        bold_font = Font(bold=True)
        light_blue_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = bold_font
            cell.fill = light_blue_fill

        for row in filtered_rows:
            ws.append(row)

        # Set larger column widths for a normal Excel window
        ws.column_dimensions['A'].width = 25      # STRING ID
        ws.column_dimensions['B'].width = 45      # KR
        ws.column_dimensions['C'].width = 60      # TRANSLATION
        ws.column_dimensions['D'].width = 18      # FLAGTYPE
        ws.column_dimensions['E'].width = 60      # CORRECTION

        wb.save(excel_path)

        messagebox.showinfo("Success", f"Filtered Excel file saved as:\n{excel_path}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

root = tk.Tk()
root.title("Text File to Excel Filter")

frame = tk.Frame(root, padx=20, pady=20)
frame.pack()

btn = tk.Button(frame, text="Select and Filter Text File", command=process_file, width=30, height=2)
btn.pack()

root.mainloop()