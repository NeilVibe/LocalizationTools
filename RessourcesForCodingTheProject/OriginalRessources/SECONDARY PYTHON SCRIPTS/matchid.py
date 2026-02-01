import os
import difflib
import openpyxl
from lxml import etree
from tkinter import filedialog, Tk, messagebox

def read_excel_data(path):
    print(f"[STEP] Reading Excel file: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
        row_count += 1
        if row[0]:
            col1 = str(row[0]).strip()
            col2 = str(row[1]).strip() if row[1] else ""
            col3 = str(row[2]).strip() if row[2] else ""
            col4 = str(row[3]).strip() if row[3] else ""
            rows.append((col1, col2, col3, col4))
            print(f"  [Excel Row {row_count}] Col1='{col1}' | Col2='{col2}' | Col3='{col3}' | Col4='{col4}'")
        else:
            print(f"  [Excel Row {row_count}] Skipped (missing Col1)")
    print(f"[INFO] Total valid Excel rows loaded: {len(rows)}")
    return rows

def read_xml_locstr(path):
    print(f"[STEP] Reading XML file: {path}")
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False, no_network=True, recover=True)
    try:
        tree = etree.parse(path, parser)
    except Exception as e:
        print(f"[ERROR] Failed to parse XML: {e}")
        return []
    root = tree.getroot()
    entries = []
    count = 0
    for loc in root.iter("LocStr"):
        count += 1
        entry = {
            "StrOrigin": loc.get("StrOrigin", "").strip(),
            "StringId": loc.get("StringId", "").strip(),
            "Str": loc.get("Str", "").strip(),
            "Desc": loc.get("Desc", "").strip()
        }
        entries.append(entry)
        print(f"  [XML Entry {count}] StringId='{entry['StringId']}' Str='{entry['Str']}' StrOrigin='{entry['StrOrigin']}' Desc='{entry['Desc']}'")
    print(f"[INFO] Total XML LocStr entries loaded: {len(entries)}")
    return entries

def find_voiceids(folder_path, string_ids):
    print(f"[STEP] Searching VoiceId in folder: {folder_path}")
    voice_map = {}
    parser = etree.XMLParser(resolve_entities=False, load_dtd=False, no_network=True, recover=True)
    file_count = 0
    for root_dir, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(".xml"):
                file_count += 1
                file_path = os.path.join(root_dir, file)
                try:
                    tree = etree.parse(file_path, parser)
                except Exception as e:
                    print(f"    [ERROR] Failed to parse XML: {e}")
                    continue
                for loc in tree.getroot().iter("LocStr"):
                    sid = loc.get("StringId", "").strip()
                    if sid in string_ids:
                        voice_id = loc.get("VoiceId", "").strip()
                        if voice_id:
                            voice_map.setdefault(sid, set()).add(voice_id)
    print(f"[INFO] VoiceId search complete. Found {len(voice_map)} StringIds with VoiceIds.")
    return voice_map

def match_excel_to_xml(excel_rows, xml_entries, voice_map):
    print("[STEP] Matching Excel rows to XML entries...")
    matches = []
    used_string_ids = set()

    for idx, (col1, col2, col3, col4) in enumerate(excel_rows, start=1):
        print(f"  [MATCH {idx}] Col1='{col1}' | Col2='{col2}' | Col3='{col3}' | Col4='{col4}'")
        search_terms = [t.strip().lower() for t in (col2, col3, col4) if t]

        potential_matches = [
            e for e in xml_entries
            if any(term in (e["Str"].strip().lower(), e["StrOrigin"].strip().lower(), e["Desc"].strip().lower())
                   for term in search_terms)
        ]

        potential_matches = [e for e in potential_matches if e["StringId"] not in used_string_ids]

        if not potential_matches:
            print("    -> No match found.")
            matches.append((col1, "NO_MATCH", ""))
            continue

        if len(potential_matches) == 1:
            matched_id = potential_matches[0]["StringId"]
            used_string_ids.add(matched_id)
            voice_ids = voice_map.get(matched_id, [""])
            voice_id = list(voice_ids)[0] if voice_ids else ""
            print(f"    -> Unique match: StringId='{matched_id}' VoiceId='{voice_id}'")
            matches.append((col1, matched_id, voice_id))
        else:
            print(f"    -> {len(potential_matches)} duplicates found, resolving by VoiceId similarity...")
            best_entry = None
            best_score = -1
            for e in potential_matches:
                matched_id = e["StringId"]
                voice_ids = voice_map.get(matched_id, [""])
                for vid in voice_ids:
                    score = difflib.SequenceMatcher(None, col1.lower(), vid.lower()).ratio()
                    print(f"       Compare Col1='{col1}' with VoiceId='{vid}' -> Score={score:.4f}")
                    if score > best_score:
                        best_score = score
                        best_entry = (matched_id, vid)
            if best_entry:
                used_string_ids.add(best_entry[0])
                print(f"    -> Selected StringId='{best_entry[0]}' VoiceId='{best_entry[1]}' (Score={best_score:.4f})")
                matches.append((col1, best_entry[0], best_entry[1]))
            else:
                matches.append((col1, "NO_MATCH", ""))

    print(f"[INFO] Matching complete. Total matches: {len(matches)}")
    return matches

def save_matches_to_excel(matches, output_path):
    print(f"[STEP] Saving results to Excel: {output_path}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Excel_Col1", "Matched_StringId", "Mapped_VoiceId"])
    for idx, (col1, matched_id, voice_id) in enumerate(matches, start=1):
        ws.append([col1, matched_id, voice_id])
        print(f"  [WRITE {idx}] {col1} | {matched_id} | {voice_id}")
    wb.save(output_path)
    print("[INFO] Excel file saved successfully.")

def main():
    root = Tk()
    root.withdraw()

    excel_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        messagebox.showinfo("Cancelled", "No Excel file selected.")
        return
    xml_path = filedialog.askopenfilename(title="Select XML File", filetypes=[("XML Files", "*.xml")])
    if not xml_path:
        messagebox.showinfo("Cancelled", "No XML file selected.")
        return

    excel_rows = read_excel_data(excel_path)
    xml_entries = read_xml_locstr(xml_path)

    folder_path = filedialog.askdirectory(title="Select Folder Containing XML Files for VoiceId Mapping")
    if not folder_path:
        messagebox.showinfo("Cancelled", "No folder selected.")
        return

    all_string_ids = {e["StringId"] for e in xml_entries}
    voice_map = find_voiceids(folder_path, all_string_ids)

    matches = match_excel_to_xml(excel_rows, xml_entries, voice_map)

    output_path = os.path.join(os.path.dirname(excel_path), "matched_with_voiceid.xlsx")
    save_matches_to_excel(matches, output_path)
    print(f"[DONE] Process complete. Output saved to: {output_path}")

if __name__ == "__main__":
    main()