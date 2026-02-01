import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def parse_key_file(key_file_path):
    key_map = {}
    key_pattern = re.compile(r'<StringKeyMap Key="(\d+)" StrKey="([^"]+)"')
    with open(key_file_path, encoding='utf-8') as f:
        for line in f:
            match = key_pattern.search(line)
            if match:
                key, strkey = match.groups()
                key_map[strkey.lower()] = key
    return key_map

def process_full_data(full_data_path, key_map, output_path):
    mission_pattern = re.compile(r'(\s*\*\s*)([A-Za-z0-9_]+)')
    with open(full_data_path, encoding='utf-8') as infile, open(output_path, 'w', encoding='utf-8') as outfile:
        for line in infile:
            m = mission_pattern.match(line)
            if m:
                prefix, mission_name = m.groups()
                key = key_map.get(mission_name.lower())
                if key:
                    line = f"{prefix}{mission_name} (Key = {key})\n"
            outfile.write(line)

def select_file(title, filetypes):
    return filedialog.askopenfilename(title=title, filetypes=filetypes)

def save_file(title, default_name):
    return filedialog.asksaveasfilename(title=title, defaultextension=".txt", initialfile=default_name, filetypes=[("Text Files", "*.txt")])

def run_processing():
    full_data_path = entry_full_data.get()
    key_file_path = entry_key_file.get()
    if not full_data_path or not key_file_path:
        messagebox.showerror("Error", "Please select both files.")
        return

    try:
        key_map = parse_key_file(key_file_path)
        output_path = save_file("Save Output File", "output_with_keys.txt")
        if not output_path:
            return
        process_full_data(full_data_path, key_map, output_path)
        messagebox.showinfo("Success", f"Processing complete!\nOutput saved to:\n{output_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

def browse_full_data():
    path = select_file("Select Full Data File", [("Text Files", "*.txt"), ("All Files", "*.*")])
    if path:
        entry_full_data.delete(0, tk.END)
        entry_full_data.insert(0, path)

def browse_key_file():
    path = select_file("Select Mission Key File", [("Text Files", "*.xml"), ("All Files", "*.*")])
    if path:
        entry_key_file.delete(0, tk.END)
        entry_key_file.insert(0, path)

# --- NEW: Text to Excel Logic ---

def parse_text_to_excel_structure(filepath):
    """
    Parses the input text file and returns a list of chapters in order:
    [
        {
            "chapter": chapter_name,
            "quests": [
                {
                    "quest": quest_name,
                    "missions": [ (mission, key), ... ]
                },
                ...
            ]
        },
        ...
    ]
    """
    chapter_re = re.compile(r'<([^ >]+) Group="([^"]+)">')
    value_re = re.compile(r'^\s*Value:\s*([^\n]+)')
    mission_re = re.compile(r'^\s*\*\s*([^\(]+)\(Key\s*=\s*(\d+)\)')
    sep_re = re.compile(r'^-+$')

    structure = []
    current_chapter = None
    current_quest = None

    with open(filepath, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\n')
            chapter_match = chapter_re.match(line)
            if chapter_match:
                # New chapter
                quest_name, chapter = chapter_match.groups()
                current_chapter = {
                    "chapter": chapter.strip(),
                    "quests": []
                }
                structure.append(current_chapter)
                current_quest = None
                continue

            value_match = value_re.match(line)
            if value_match:
                # New quest
                quest_name = value_match.group(1).strip()
                current_quest = {
                    "quest": quest_name,
                    "missions": []
                }
                if current_chapter is not None:
                    current_chapter["quests"].append(current_quest)
                continue

            mission_match = mission_re.match(line)
            if mission_match and current_chapter and current_quest:
                mission_name = mission_match.group(1).strip()
                key_number = mission_match.group(2).strip()
                current_quest["missions"].append((mission_name, key_number))
                continue

            # If a separator line, just skip
            if sep_re.match(line):
                continue

    return structure

def save_structure_to_excel(structure, output_path):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Missions"

    # Header
    ws.append(["CHAPTER", "QUEST", "MISSION", "KEY NUMBER"])
    header_font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    chapter_rows = {}  # chapter_name: [start_row, end_row]

    for chapter in structure:
        chapter_name = chapter["chapter"]
        quests = chapter["quests"]
        chapter_start_row = row_idx

        for quest in quests:
            quest_name = quest["quest"]
            missions = quest["missions"]
            quest_start_row = row_idx
            if missions:
                for mission, key in missions:
                    ws.cell(row=row_idx, column=1, value=chapter_name)
                    ws.cell(row=row_idx, column=2, value=quest_name)
                    ws.cell(row=row_idx, column=3, value=mission)
                    ws.cell(row=row_idx, column=4, value=key)
                    row_idx += 1
            else:
                # Quest with no missions
                ws.cell(row=row_idx, column=1, value=chapter_name)
                ws.cell(row=row_idx, column=2, value=quest_name)
                row_idx += 1
            # Merge quest cells if multiple missions
            if missions and len(missions) > 1:
                ws.merge_cells(start_row=quest_start_row, start_column=2, end_row=quest_start_row+len(missions)-1, end_column=2)
                ws[get_column_letter(2) + str(quest_start_row)].alignment = Alignment(vertical="center")

        chapter_end_row = row_idx - 1
        if chapter_name in chapter_rows:
            # If chapter appears multiple times, extend the end_row
            chapter_rows[chapter_name][1] = chapter_end_row
        else:
            chapter_rows[chapter_name] = [chapter_start_row, chapter_end_row]

    # Merge chapter cells for all identical values
    for chapter_name, (start_row, end_row) in chapter_rows.items():
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws[get_column_letter(1) + str(start_row)].alignment = Alignment(vertical="center")

    # Adjust column widths
    for col, width in zip(range(1, 5), [20, 40, 40, 15]):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)

def text_to_excel():
    input_path = filedialog.askopenfilename(
        title="Select Text File to Convert to Excel",
        filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
    )
    if not input_path:
        return

    try:
        structure = parse_text_to_excel_structure(input_path)
        output_path = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            initialfile="missions.xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not output_path:
            return
        save_structure_to_excel(structure, output_path)
        messagebox.showinfo("Success", f"Excel file created:\n{output_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

# --- GUI ---
root = tk.Tk()
root.title("Mission Key Mapper")

frm = tk.Frame(root, padx=10, pady=10)
frm.pack()

tk.Label(frm, text="Full Data File:").grid(row=0, column=0, sticky="e")
entry_full_data = tk.Entry(frm, width=50)
entry_full_data.grid(row=0, column=1)
tk.Button(frm, text="Browse...", command=browse_full_data).grid(row=0, column=2)

tk.Label(frm, text="Mission Key File:").grid(row=1, column=0, sticky="e")
entry_key_file = tk.Entry(frm, width=50)
entry_key_file.grid(row=1, column=1)
tk.Button(frm, text="Browse...", command=browse_key_file).grid(row=1, column=2)

tk.Button(frm, text="Process and Save Output", command=run_processing, bg="#4CAF50", fg="white").grid(row=2, column=0, columnspan=3, pady=10)

# --- NEW BUTTON ---
tk.Button(frm, text="Text to Excel", command=text_to_excel, bg="#2196F3", fg="white").grid(row=3, column=0, columnspan=3, pady=10)

root.mainloop()