import os
import threading
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import pandas as pd
import shutil
import re
from openpyxl import load_workbook, Workbook
from copy import copy

def filter_curse_words(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        mask = ~df.apply(lambda row: row.astype(str).str.contains('No curse word detected', na=False, regex=False)).any(axis=1)
        filtered_df = df[mask]
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_cursewordslist{ext}")
        filtered_df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def filter_underscore_col1(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        mask_no_underscore = ~df[0].astype(str).str.contains('_', na=False, regex=False)
        filtered_df = df[mask_no_underscore]
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_underscore_col1_cleaned{ext}")
        filtered_df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def filter_col1_contains_korean(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        korean_regex = re.compile(r'[\uac00-\ud7af\u1100-\u11ff\u3130-\u318f]')
        mask_no_korean = ~df[0].astype(str).apply(lambda x: bool(korean_regex.search(x)) if pd.notnull(x) else False)
        filtered_df = df[mask_no_korean]
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_col1_no_korean{ext}")
        filtered_df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def fragment_excel(input_file, num_fragments):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        total_rows = len(df)
        if num_fragments < 1 or num_fragments > total_rows:
            return "Invalid number of fragments."
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        folder_name = os.path.join(outdir, f"{base}_fragmented")
        if os.path.exists(folder_name):
            shutil.rmtree(folder_name)
        os.makedirs(folder_name)
        rows_per_fragment = total_rows // num_fragments
        extra = total_rows % num_fragments
        start = 0
        for i in range(num_fragments):
            end = start + rows_per_fragment + (1 if i < extra else 0)
            frag_df = df.iloc[start:end]
            frag_file = os.path.join(folder_name, f"{base}_{i+1}{ext}")
            frag_df.to_excel(frag_file, index=False, header=False)
            start = end
        return f"Fragmented into {num_fragments} files in folder '{folder_name}'."
    except Exception as e:
        return f"Error: {e}"

def unify_excels(file_paths):
    try:
        if not file_paths or len(file_paths) < 2:
            return "Please select at least two files to unify."
        dfs = [pd.read_excel(file, header=None, dtype=str) for file in file_paths]
        unified_df = pd.concat(dfs, ignore_index=True)
        first_file = file_paths[0]
        base, ext = os.path.splitext(os.path.basename(first_file))
        outdir = os.path.dirname(first_file)
        output_file = os.path.join(outdir, f"{base}_unified{ext}")
        unified_df.to_excel(output_file, index=False, header=False)
        return f"Unified file saved as: {output_file}"
    except Exception as e:
        return f"Error: {e}"

def keep_unique_first_column(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        unique_df = df.drop_duplicates(subset=[0], keep='first')
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_uniquefirstcolumn{ext}")
        unique_df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def add_korean_character_col1(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        df[0] = df[0].apply(lambda x: f"ㄴ{x}" if pd.notnull(x) else x)
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_addkoreancol1{ext}")
        df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def concat_col2_col1_above_col1(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        if df.shape[1] < 2:
            return "File must have at least two columns."
        df[0] = df.apply(lambda row: f"{row[1] if pd.notnull(row[1]) else ''}\n{row[0] if pd.notnull(row[0]) else ''}", axis=1)
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_col2col1{ext}")
        df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def extract_rows_col2_empty_or_korean(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        if df.shape[1] < 2:
            return "File must have at least two columns."
        mask_empty = df[1].isnull() | (df[1].astype(str).str.strip() == "")
        korean_regex = re.compile(r'[\uac00-\ud7af\u1100-\u11ff\u3130-\u318f]')
        mask_korean = df[1].astype(str).apply(lambda x: bool(korean_regex.search(x)) if pd.notnull(x) else False)
        mask = mask_empty | mask_korean
        filtered_df = df[mask]
        count = filtered_df.shape[0]
        if count == 0:
            messagebox.showinfo("No Rows Found", "No empty or Korean rows found in Col2!")
            return "No empty or Korean rows found in Col2!"
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_col2empty_or_korean{ext}")
        filtered_df.to_excel(output_file, index=False, header=False)
        messagebox.showinfo("Extraction Complete", f"Extracted {count} rows where Col2 is empty or contains Korean.\nSaved to:\n{output_file}")
        return f"Extracted {count} rows to: {output_file}"
    except Exception as e:
        messagebox.showerror("Error", f"Error: {e}")
        return f"Error: {e}"

def eventname_minus_dialogvoice(input_file):
    try:
        df = pd.read_excel(input_file, header=None, dtype=str)
        if df.shape[1] < 2:
            return "File must have at least two columns."
        results = []
        for _, row in df.iterrows():
            col1 = str(row[0]).lower().strip() if pd.notnull(row[0]) else ""
            col2 = str(row[1]).lower().strip() if pd.notnull(row[1]) else ""

            if col1 and col2 and col1 in col2:
                diff = col2.replace(col1, "", 1)
                if diff.startswith("_"):
                    diff = diff[1:]
                results.append(diff)
            elif not col1 and col2:
                results.append(col2[1:] if col2.startswith("_") else col2)
            else:
                results.append(col2 if "aidialog" in col2 else "")

        result_df = pd.DataFrame(results)
        base, ext = os.path.splitext(os.path.basename(input_file))
        outdir = os.path.dirname(input_file)
        output_file = os.path.join(outdir, f"{base}_eventname_minus_dialogvoice{ext}")
        result_df.to_excel(output_file, index=False, header=False)
        return output_file
    except Exception as e:
        return f"Error: {e}"

def vlookup_doublecol_regex_preserve_format():
    ref_file = filedialog.askopenfilename(title="Select Reference Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not ref_file:
        return
    target_file = filedialog.askopenfilename(title="Select Target Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not target_file:
        return
    try:
        ref_wb = load_workbook(ref_file)
        ref_ws = ref_wb.active
        ref_pairs = set()
        for r in range(1, ref_ws.max_row + 1):
            col1_val = str(ref_ws.cell(row=r, column=1).value).strip().lower() if ref_ws.cell(row=r, column=1).value else ""
            col2_val = str(ref_ws.cell(row=r, column=2).value).strip().lower() if ref_ws.cell(row=r, column=2).value else ""
            if col1_val and col2_val:
                ref_pairs.add((col1_val, col2_val))
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active
        new_wb = Workbook()
        new_ws = new_wb.active
        new_row = 1
        for r in range(1, target_ws.max_row + 1):
            row_values_lower = [
                str(target_ws.cell(row=r, column=c).value).strip().lower()
                if target_ws.cell(row=r, column=c).value else ""
                for c in range(1, target_ws.max_column + 1)
            ]
            match_found = False
            for ref_col1, ref_col2 in ref_pairs:
                if ref_col1 in row_values_lower and ref_col2 in row_values_lower:
                    match_found = True
                    break
            if match_found:
                for c in range(1, target_ws.max_column + 1):
                    cell = target_ws.cell(row=r, column=c)
                    new_cell = new_ws.cell(row=new_row, column=c, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                new_row += 1
        base = os.path.splitext(os.path.basename(target_file))[0]
        outdir = os.path.dirname(target_file)
        output_file = os.path.join(outdir, f"{base}_vlookup_doublecol_anywhere_preserve.xlsx")
        new_wb.save(output_file)
        messagebox.showinfo("Done", f"Matched rows saved to:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"Error: {e}")

# ------------------- GUI ACTIONS -------------------
def run_filter(): threading.Thread(target=lambda: _run_task(filter_curse_words, "Filtering...")).start()
def run_filter_underscore_col1(): threading.Thread(target=lambda: _run_task(filter_underscore_col1, "Filtering underscores in first column...")).start()
def run_filter_col1_no_korean(): threading.Thread(target=lambda: _run_task(filter_col1_contains_korean, "Filtering rows where Col1 contains Korean...")).start()
def run_fragment():
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path: return
    try:
        df = pd.read_excel(file_path, header=None, dtype=str)
        total_rows = len(df)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file: {e}")
        return
    num_fragments = simpledialog.askinteger("Fragment Excel", f"How many fragments? (max {total_rows})", minvalue=1, maxvalue=total_rows)
    if not num_fragments: return
    threading.Thread(target=lambda: _run_task(lambda f=file_path: fragment_excel(f, num_fragments), "Fragmenting...")).start()
def run_unify():
    file_paths = filedialog.askopenfilenames(title="Select Excel Files to Unify", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_paths or len(file_paths) < 2:
        status_var.set("Please select at least two files to unify.")
        return
    threading.Thread(target=lambda: _run_task(lambda: unify_excels(file_paths), "Unifying...")).start()
def run_unique_first_column(): threading.Thread(target=lambda: _run_task(keep_unique_first_column, "Processing unique terms in first column...")).start()
def run_add_korean_col1(): threading.Thread(target=lambda: _run_task(add_korean_character_col1, "Adding 'ㄴ' to first column...")).start()
def run_concat_col2_col1_above_col1(): threading.Thread(target=lambda: _run_task(concat_col2_col1_above_col1, "Concatenating Col2+Col1 above Col1...")).start()
def run_extract_rows_col2_empty_or_korean(): threading.Thread(target=lambda: _run_task(extract_rows_col2_empty_or_korean, "Extracting rows where Col2 is empty or contains Korean...")).start()
def run_eventname_minus_dialogvoice(): threading.Thread(target=lambda: _run_task(eventname_minus_dialogvoice, "Processing EventName Minus DialogVoice...")).start()

def _run_task(func, msg):
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path: return
    disable_all_buttons()
    status_var.set(msg)
    result = func(file_path)
    status_var.set(f"Done! Output: {result}")
    enable_all_buttons()

def disable_all_buttons():
    for btn in buttons: btn.config(state=tk.DISABLED)
def enable_all_buttons():
    for btn in buttons: btn.config(state=tk.NORMAL)

# GUI setup
root = tk.Tk()
root.title("Excel Toolset")
root.geometry("470x800")
root.resizable(False, False)

frame = tk.Frame(root, padx=20, pady=20)
frame.pack(expand=True, fill=tk.BOTH)

buttons = []
def add_btn(text, cmd):
    b = tk.Button(frame, text=text, width=45, height=2, command=cmd)
    b.pack(pady=5)
    buttons.append(b)

add_btn("Filter 'No curse word detected' Rows", run_filter)
add_btn("Filter Underscore in Col1", run_filter_underscore_col1)
add_btn("Filter Col1 Contains Korean", run_filter_col1_no_korean)
add_btn("Fragment Excel File", run_fragment)
add_btn("Unify Fragmented Excel Files", run_unify)
add_btn("Keep Only Unique Terms in Col1", run_unique_first_column)
add_btn("Add 'ㄴ' to First Column", run_add_korean_col1)
add_btn("Concat Col2+Col1 Above Col1", run_concat_col2_col1_above_col1)
add_btn("Extract if Col2 Empty OR Contains Korean", run_extract_rows_col2_empty_or_korean)
add_btn("EventName Minus DialogVoice", run_eventname_minus_dialogvoice)
add_btn("VLOOKUP DOUBLE COL", vlookup_doublecol_regex_preserve_format)

status_var = tk.StringVar()
status_label = tk.Label(frame, textvariable=status_var, fg="blue", anchor="w")
status_label.pack(fill=tk.X, pady=(10, 0))

root.mainloop()