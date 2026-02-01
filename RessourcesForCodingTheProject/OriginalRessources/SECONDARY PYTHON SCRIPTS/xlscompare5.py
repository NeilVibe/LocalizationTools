#!/usr/bin/env python3
"""
Excel Comparison Tool
Standalone script for:
1. Order By Similarity - Compare and reorder Excel rows by similarity
2. Compare Before vs After - Compare two tabs and categorize differences
"""

import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
import faiss
import os
import traceback
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import shutil
import tempfile

def log(msg):
    """Print with timestamp"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {msg}")

print("=" * 60)
log("Excel Comparison Tool - Initializing...")
print("=" * 60)

# Load the KRTransformer model
log("Loading KRTransformer model...")
local_model_path = r"KRTransformer"
try:
    model = SentenceTransformer(local_model_path)
    log(f"✓ Model loaded successfully (embedding dim = {model.get_sentence_embedding_dimension()})")
except Exception as e:
    log(f"✗ FATAL ERROR: Failed to load model from '{local_model_path}'")
    log(f"Error details: {e}")
    traceback.print_exc()
    log("Please ensure KRTransformer model is in the correct directory")
    exit(1)


def clean_text(text):
    """Clean text by removing unwanted characters"""
    if text is None or pd.isna(text):
        return ""
    if not isinstance(text, str):
        text = str(text)
    return text.replace('_x000D_', '').strip()


def normalize_column_d(value):
    """
    Normalize column D (StartFrame) values for comparison.
    Rules:
    - Values above 15000 → ""
    - Values below 0 (negative) → ""
    Returns: Empty string if invalid, otherwise the original value
    """
    if pd.isna(value) or value == "":
        return ""
    
    try:
        num_value = float(value)
        if num_value > 15000 or num_value < 0:
            return ""
        return value
    except (ValueError, TypeError):
        return value


def calculate_similarity(text1, text2):
    """Calculate cosine similarity between two texts using embeddings"""
    try:
        if not text1 or not text2:
            return 0.0
        
        # Generate embeddings
        emb1 = model.encode([text1], convert_to_numpy=True)
        emb2 = model.encode([text2], convert_to_numpy=True)
        
        # Normalize embeddings
        faiss.normalize_L2(emb1)
        faiss.normalize_L2(emb2)
        
        # Calculate cosine similarity
        similarity = np.dot(emb1[0], emb2[0])
        
        return similarity
    except Exception as e:
        log(f"✗ ERROR in calculate_similarity: {e}")
        traceback.print_exc()
        return 0.0


def order_by_similarity():
    """
    Function 1: Order By Similarity
    - Read column headers
    - Compare B+D (indices 1,3) vs C+E (indices 2,4)
    - Extract rows containing "{" to a separate sheet
    - Create similarity-ordered sheet with differences highlighted
    - Write similarity score to column G with header "Similarity Score"
    - Generate a comprehensive report summary sheet
    """
    print("\n" + "=" * 60)
    log("ORDER BY SIMILARITY - Started")
    print("=" * 60)
    
    # Select input file
    log("Opening file dialog...")
    file_path = filedialog.askopenfilename(
        title="Select Excel File for Similarity Ordering",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not file_path:
        log("✗ Operation cancelled by user (no file selected)")
        return
    
    log(f"File selected: {file_path}")
    
    try:
        log("Step 1/9: Reading Excel file with headers...")
        
        # Read Excel file WITH headers
        df = pd.read_excel(file_path)
        log(f"  → Raw data shape: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Store the header row
        headers = df.columns.tolist()
        log(f"  → Headers found: {headers}")
        
        log("Step 2/9: Removing empty rows...")
        df = df.dropna(how='all')  # Remove completely empty rows
        log(f"  → After removing empty rows: {len(df):,} rows")
        
        # Ensure we have at least 6 columns (A B C D E F = indices 0-5)
        if df.shape[1] < 6:
            log(f"✗ ERROR: Excel file must have at least 6 columns, but only has {df.shape[1]}")
            log(f"  Columns found: {df.shape[1]}")
            log(f"  Required: 6 (A, B, C, D, E, F)")
            return
        
        log(f"  → Columns available: {df.shape[1]}")
        log(f"  → Using columns: B (1), C (2), D (3), E (4)")
        log(f"  → Comparing: B+D vs C+E")
        
        log("Step 3/9: Checking for rows with '{' character...")
        
        # Check each row for the presence of "{" character
        special_char_mask = df.astype(str).apply(lambda row: row.str.contains('{', regex=False).any(), axis=1)
        df_with_special = df[special_char_mask].copy()
        df_without_special = df[~special_char_mask].copy()
        
        log(f"  → Rows with '{{' character: {len(df_with_special):,}")
        log(f"  → Rows without '{{' character: {len(df_without_special):,}")
        
        log("Step 4/9: Calculating similarities for rows without '{' character...")
        
        # Calculate similarities for rows WITHOUT "{" character
        similarities = []
        total_rows = len(df_without_special)
        
        log(f"  → Processing {total_rows:,} rows...")
        
        for idx, (orig_idx, row) in enumerate(df_without_special.iterrows()):
            # Concatenate B+D (columns 1,3)
            text_bd = clean_text(row.iloc[1]) + " " + clean_text(row.iloc[3])
            
            # Concatenate C+E (columns 2,4)
            text_ce = clean_text(row.iloc[2]) + " " + clean_text(row.iloc[4])
            
            # Debug: Show first few concatenations
            if idx < 3:
                log(f"  → Row {idx}: B+D = '{text_bd[:50]}...', C+E = '{text_ce[:50]}...'")
            
            # Calculate similarity
            sim_score = calculate_similarity(text_bd, text_ce)
            
            # Convert to percentage (0-100)
            similarity_pct = sim_score * 100
            similarities.append(similarity_pct)
            
            # Progress updates
            if (idx + 1) % 100 == 0:
                progress_pct = ((idx + 1) / total_rows) * 100
                log(f"  Progress: {idx + 1:,} / {total_rows:,} ({progress_pct:.1f}%)")
            elif (idx + 1) % 10 == 0 and total_rows < 100:
                log(f"  Progress: {idx + 1:,} / {total_rows:,}")
        
        avg_similarity = 0
        min_similarity = 0
        max_similarity = 0
        top_similarity = 0
        bottom_similarity = 0
        
        if similarities:
            log(f"  ✓ Completed similarity calculations for all {total_rows:,} rows")
            min_similarity = min(similarities)
            max_similarity = max(similarities)
            avg_similarity = sum(similarities)/len(similarities)
            log(f"  → Similarity range: {min_similarity:.2f}% - {max_similarity:.2f}%")
            log(f"  → Average similarity: {avg_similarity:.2f}%")
        
        log("Step 5/9: Adding similarity column...")
        # Add similarity column to index 6 (column G)
        df_without_special['Similarity Score'] = similarities
        log(f"  ✓ Similarity scores added to column G")
        
        log("Step 6/9: Sorting by similarity (descending)...")
        df_sorted = df_without_special.sort_values(by='Similarity Score', ascending=False).reset_index(drop=True)
        log(f"  ✓ Data sorted from most similar to least similar")
        if len(df_sorted) > 0:
            top_similarity = df_sorted['Similarity Score'].iloc[0]
            bottom_similarity = df_sorted['Similarity Score'].iloc[-1]
            log(f"  → Top similarity: {top_similarity:.2f}%")
            log(f"  → Bottom similarity: {bottom_similarity:.2f}%")
        
        log("Step 6.5/9: Calculating differences AFTER sorting for correct highlighting...")
        
        # NOW calculate which columns differ - AFTER sorting!
        diff_indices = []
        for idx, row in df_sorted.iterrows():
            diff_cols = []
            # Compare B vs C
            if clean_text(row.iloc[1]) != clean_text(row.iloc[2]):
                diff_cols.extend([1, 2])
            # Compare D vs E
            if clean_text(row.iloc[3]) != clean_text(row.iloc[4]):
                diff_cols.extend([3, 4])
            diff_indices.append(diff_cols)
        
        log(f"  ✓ Calculated differences for {len(diff_indices):,} sorted rows")
        
        log("Step 7/9: Creating comprehensive report summary...")
        
        # Create report data
        report_data = []
        report_data.append(['ORDER BY SIMILARITY - ANALYSIS REPORT', ''])
        report_data.append(['─' * 50, ''])  # Using box drawing character instead of =
        report_data.append(['', ''])
        report_data.append(['Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        report_data.append(['Input File:', os.path.basename(file_path)])
        report_data.append(['', ''])
        report_data.append(['PROCESSING SUMMARY', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Total Rows Processed:', len(df)])
        report_data.append(['Rows with "{" Character:', len(df_with_special)])
        report_data.append(['Rows Analyzed for Similarity:', len(df_without_special)])
        report_data.append(['', ''])
        report_data.append(['COLUMN CONFIGURATION', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Comparison Logic:', 'B+D vs C+E'])
        report_data.append(['Column B (Index 1):', headers[1] if len(headers) > 1 else 'Column B'])
        report_data.append(['Column C (Index 2):', headers[2] if len(headers) > 2 else 'Column C'])
        report_data.append(['Column D (Index 3):', headers[3] if len(headers) > 3 else 'Column D'])
        report_data.append(['Column E (Index 4):', headers[4] if len(headers) > 4 else 'Column E'])
        report_data.append(['Similarity Score Column:', 'G (Similarity Score)'])
        report_data.append(['', ''])
        
        if similarities:
            report_data.append(['SIMILARITY STATISTICS', ''])
            report_data.append(['-' * 50, ''])
            report_data.append(['Minimum Similarity:', f'{min_similarity:.2f}%'])
            report_data.append(['Maximum Similarity:', f'{max_similarity:.2f}%'])
            report_data.append(['Average Similarity:', f'{avg_similarity:.2f}%'])
            report_data.append(['Top Row Similarity:', f'{top_similarity:.2f}%'])
            report_data.append(['Bottom Row Similarity:', f'{bottom_similarity:.2f}%'])
            report_data.append(['', ''])
        
        report_data.append(['OUTPUT SHEETS', ''])
        report_data.append(['-' * 50, ''])
        if len(df_with_special) > 0:
            report_data.append(['Sheet: Special Characters', f'{len(df_with_special)} rows'])
            report_data.append(['  Description:', 'Rows containing "{" character'])
        else:
            report_data.append(['Sheet: Special Characters', 'Not created (no rows with "{")'])
        
        if len(df_sorted) > 0:
            report_data.append(['Sheet: Ordered By Similarity', f'{len(df_sorted)} rows'])
            report_data.append(['  Description:', 'Rows sorted by similarity (high to low)'])
            report_data.append(['  Highlighting:', 'Orange = cells that differ between B/C or D/E'])
        else:
            report_data.append(['Sheet: Ordered By Similarity', 'Not created (no rows to analyze)'])
        
        report_data.append(['', ''])
        report_data.append(['NOTES', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['- Rows with "{" character were separated for special handling', ''])
        report_data.append(['- Similarity scores range from 0% (completely different) to 100% (identical)', ''])
        report_data.append(['- Light orange highlighting shows cells that differ between compared columns', ''])
        report_data.append(['- Results are sorted from most similar to least similar', ''])
        
        log(f"  ✓ Report summary created with {len(report_data)} rows of information")
        
        # Generate output file path
        base_dir = os.path.dirname(file_path)
        output_path = os.path.join(base_dir, "OrderBySimilarity.xlsx")
        
        log("Step 8/9: Saving results to Excel with multiple sheets...")
        log(f"  → Output file: {output_path}")
        
        # Create a temporary file first
        temp_file = None
        try:
            with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
                temp_file = tmp.name
            
            # Save to temporary file
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                # Sheet 1: Report Summary
                df_report = pd.DataFrame(report_data, columns=['Metric', 'Value'])
                df_report.to_excel(writer, sheet_name='Report Summary', index=False, header=False)
                log(f"  ✓ Sheet 'Report Summary' created")
                
                # Sheet 2: Rows with "{" character
                if len(df_with_special) > 0:
                    df_with_special.to_excel(writer, sheet_name='Special Characters', index=False)
                    log(f"  ✓ Sheet 'Special Characters' created with {len(df_with_special):,} rows")
                else:
                    log(f"  ℹ No rows with '{{' character found, skipping special sheet")
                
                # Sheet 3: Ordered by similarity
                if len(df_sorted) > 0:
                    df_sorted.to_excel(writer, sheet_name='Ordered By Similarity', index=False)
                    log(f"  ✓ Sheet 'Ordered By Similarity' created with {len(df_sorted):,} rows")
                else:
                    log(f"  ℹ No rows to sort by similarity")
            
            log("Step 9/9: Applying formatting and highlighting differences...")
            
            # Apply formatting using openpyxl
            wb = load_workbook(temp_file)
            
            # Define colors
            orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            bold_font = Font(bold=True)
            
            # Format the "Report Summary" sheet
            if 'Report Summary' in wb.sheetnames:
                ws = wb['Report Summary']
                log(f"  → Formatting 'Report Summary' sheet...")
                
                # Set column widths
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 40
                
                # Format title row (row 1)
                ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
                
                # Format section headers
                section_headers = []
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value and isinstance(cell_value, str):
                        if cell_value in ['PROCESSING SUMMARY', 'COLUMN CONFIGURATION', 
                                         'SIMILARITY STATISTICS', 'OUTPUT SHEETS', 'NOTES']:
                            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color="1F4E78")
                            section_headers.append(row_idx)
                
                log(f"    ✓ Report summary formatted")
            
            # Format the "Ordered By Similarity" sheet
            if 'Ordered By Similarity' in wb.sheetnames and len(df_sorted) > 0:
                ws = wb['Ordered By Similarity']
                log(f"  → Applying formatting to 'Ordered By Similarity'...")
                
                # Set column widths for better readability
                for col_idx in range(df_sorted.shape[1]):
                    ws.column_dimensions[chr(65 + col_idx)].width = 20
                
                # Format header row
                for col_idx in range(1, df_sorted.shape[1] + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Highlight differences (skip header row, start from row 2)
                # NOW this matches the sorted order!
                highlighted_count = 0
                for row_idx, diff_cols in enumerate(diff_indices, start=2):
                    if diff_cols:  # Only if there are differences
                        for col_idx in diff_cols:
                            cell = ws.cell(row=row_idx, column=col_idx + 1)
                            cell.fill = orange_fill
                        highlighted_count += 1
                
                log(f"    ✓ Highlighted differences in {highlighted_count} rows")
            
            # Format the "Special Characters" sheet
            if 'Special Characters' in wb.sheetnames and len(df_with_special) > 0:
                ws = wb['Special Characters']
                log(f"  → Formatting 'Special Characters' sheet...")
                
                # Set column widths
                for col_idx in range(df_with_special.shape[1]):
                    ws.column_dimensions[chr(65 + col_idx)].width = 20
                
                # Format header row
                for col_idx in range(1, df_with_special.shape[1] + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                
                log(f"    ✓ Special Characters sheet formatted")
            
            # Save the formatted workbook
            wb.save(output_path)
            wb.close()
            log(f"  ✓ File saved successfully with formatting")
            
        except Exception as e:
            log(f"✗ ERROR while saving and formatting:")
            log(f"  Error type: {type(e).__name__}")
            log(f"  Error message: {str(e)}")
            traceback.print_exc()
            return
        finally:
            # Clean up temporary file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                    log(f"  ✓ Temporary file removed")
                except Exception as e:
                    log(f"  ⚠ Warning: Could not remove temp file: {e}")
        
        print("=" * 60)
        log("✓ ORDER BY SIMILARITY - Completed Successfully")
        log(f"  Input file: {file_path}")
        log(f"  Output file: {output_path}")
        log(f"  Total rows processed: {len(df):,}")
        log(f"  Rows with '{{' character: {len(df_with_special):,}")
        log(f"  Rows ordered by similarity: {len(df_sorted):,}")
        if len(df_sorted) > 0:
            log(f"  Similarity range: {df_sorted['Similarity Score'].min():.2f}% - {df_sorted['Similarity Score'].max():.2f}%")
        print("=" * 60)
        
    except Exception as e:
        log(f"✗ FATAL ERROR in order_by_similarity:")
        log(f"  Error type: {type(e).__name__}")
        log(f"  Error message: {str(e)}")
        log("  Full traceback:")
        traceback.print_exc()


def compare_before_after():
    """
    Function 2: Compare Before vs After
    - FOCUS ON AFTER SHEET: Check what needs attention in AFTER
    - Categorize rows:
      * Perfect Matches: A,B,C,D all match with BEFORE
      * Need To ReRecord: 
        - NEW rows (no match in BEFORE at all)
        - Modified rows (B,C match but A or D differ) - CATEGORIZED INTO 3 SUB-TYPES
      * Need To Inspect: A,D match but B or C differ (show all 4 columns)
    - Generate comprehensive report summary sheet
    """
    print("\n" + "=" * 60)
    log("COMPARE BEFORE VS AFTER - Started")
    print("=" * 60)
    
    # Select input file
    log("Opening file dialog...")
    file_path = filedialog.askopenfilename(
        title="Select Excel File (must have 'Before' and 'After' tabs)",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not file_path:
        log("✗ Operation cancelled by user (no file selected)")
        return
    
    log(f"File selected: {file_path}")
    
    temp_file = None
    wb = None
    
    try:
        log("Step 1/8: Reading 'Before' and 'After' tabs...")
        
        # Check if file has both tabs
        xl_file = pd.ExcelFile(file_path)
        if 'Before' not in xl_file.sheet_names:
            log("✗ ERROR: 'Before' tab not found in Excel file")
            log(f"  Available tabs: {xl_file.sheet_names}")
            return
        if 'After' not in xl_file.sheet_names:
            log("✗ ERROR: 'After' tab not found in Excel file")
            log(f"  Available tabs: {xl_file.sheet_names}")
            return
        
        # Read both tabs without headers
        df_before = pd.read_excel(file_path, sheet_name='Before', header=None)
        df_after = pd.read_excel(file_path, sheet_name='After', header=None)
        
        log(f"  → BEFORE tab: {df_before.shape[0]} rows × {df_before.shape[1]} columns")
        log(f"  → AFTER tab: {df_after.shape[0]} rows × {df_after.shape[1]} columns")
        
        # Verify both tabs have at least 4 columns
        if df_before.shape[1] < 4 or df_after.shape[1] < 4:
            log(f"✗ ERROR: Both tabs must have at least 4 columns")
            log(f"  BEFORE columns: {df_before.shape[1]}, AFTER columns: {df_after.shape[1]}")
            return
        
        log("Step 2/8: Cleaning data...")
        
        # Clean all text data
        for col in range(4):
            df_before[col] = df_before[col].apply(clean_text)
            df_after[col] = df_after[col].apply(clean_text)
        
        # PREPROCESS column D (index 3): normalize invalid values to empty string
        df_before[3] = df_before[3].apply(normalize_column_d)
        df_after[3] = df_after[3].apply(normalize_column_d)
        
        log(f"  ✓ Data cleaned and column D preprocessed")
        
        log("Step 3/8: Matching AFTER rows with BEFORE...")
        
        # Create composite keys for matching (B+C)
        df_before['_key'] = df_before[1] + '||' + df_before[2]
        df_after['_key'] = df_after[1] + '||' + df_after[2]
        
        # Create lookup dictionary for BEFORE rows
        before_dict = {}
        for idx, row in df_before.iterrows():
            key = row['_key']
            if key not in before_dict:
                before_dict[key] = []
            before_dict[key].append(row)
        
        log(f"  → Created lookup with {len(before_dict):,} unique B+C combinations in BEFORE")
        
        # Categories for AFTER rows
        perfect_matches = []
        new_rows = []
        modified_rows_a_only = []
        modified_rows_d_only = []
        modified_rows_both = []
        inspect_rows = []
        
        # Track which columns differ for highlighting
        modified_diffs_a_only = []
        modified_diffs_d_only = []
        modified_diffs_both = []
        inspect_diffs = []
        
        # Track matched BEFORE rows
        matched_before = set()
        
        log("Step 4/8: Categorizing AFTER rows...")
        
        for idx, after_row in df_after.iterrows():
            key = after_row['_key']
            
            if key not in before_dict:
                # NEW ROW: No match in BEFORE at all
                new_rows.append(after_row[:4].tolist())
            else:
                # Found matching B+C in BEFORE
                before_matches = before_dict[key]
                
                # Mark these BEFORE rows as matched
                for before_row in before_matches:
                    matched_before.add(tuple(before_row[:4].tolist()))
                
                # Check if A and D also match
                found_perfect = False
                found_inspect = False
                a_matches = []
                d_matches = []
                
                for before_row in before_matches:
                    a_match = (after_row[0] == before_row[0])
                    d_match = (after_row[3] == before_row[3])
                    
                    if a_match and d_match:
                        # PERFECT MATCH: All 4 columns match
                        perfect_matches.append(after_row[:4].tolist() + before_row[:4].tolist())
                        found_perfect = True
                        break
                    elif not a_match and not d_match:
                        # Need To Inspect: B+C match, but both A and D differ
                        a_matches.append(before_row[0])
                        d_matches.append(before_row[3])
                    else:
                        a_matches.append(before_row[0])
                        d_matches.append(before_row[3])
                
                if found_perfect:
                    continue
                
                # Check if any BEFORE row has matching A or D
                best_before = before_matches[0]  # Use first match as representative
                a_match = (after_row[0] == best_before[0])
                d_match = (after_row[3] == best_before[3])
                
                if not a_match and not d_match:
                    # Need To Inspect: A,D both differ
                    inspect_rows.append(after_row[:4].tolist() + best_before[:4].tolist())
                    # Highlight B and C columns (indices 1, 2)
                    inspect_diffs.append([1, 2])
                else:
                    # Need To ReRecord (Modified): B+C match, but A or D differ
                    if not a_match and d_match:
                        # Only A differs
                        modified_rows_a_only.append(after_row[:4].tolist())
                        modified_diffs_a_only.append([0])  # Highlight column A
                    elif a_match and not d_match:
                        # Only D differs
                        modified_rows_d_only.append(after_row[:4].tolist())
                        modified_diffs_d_only.append([3])  # Highlight column D
                    else:
                        # Both A and D differ (this shouldn't happen based on logic above)
                        modified_rows_both.append(after_row[:4].tolist())
                        modified_diffs_both.append([0, 3])  # Highlight columns A and D
        
        # Count statistics
        perfect_count = len(perfect_matches)
        new_count = len(new_rows)
        modified_count_a_only = len(modified_rows_a_only)
        modified_count_d_only = len(modified_rows_d_only)
        modified_count_both = len(modified_rows_both)
        total_modified = modified_count_a_only + modified_count_d_only + modified_count_both
        inspect_count = len(inspect_rows)
        unmatched_before_count = len(df_before) - len(matched_before)
        
        log(f"  ✓ Categorization complete:")
        log(f"    → Perfect Matches: {perfect_count:,}")
        log(f"    → Need To ReRecord (NEW): {new_count:,}")
        log(f"    → Need To ReRecord (MODIFIED): {total_modified:,}")
        log(f"      - CastingKey mismatch only: {modified_count_a_only:,}")
        log(f"      - StartFrame mismatch only: {modified_count_d_only:,}")
        log(f"      - Both mismatch: {modified_count_both:,}")
        log(f"    → Need To Inspect: {inspect_count:,}")
        
        log("Step 5/8: Creating comprehensive report summary...")
        
        # Create report data
        report_data = []
        report_data.append(['COMPARE BEFORE VS AFTER - ANALYSIS REPORT', ''])
        report_data.append(['─' * 50, ''])  # Using box drawing character instead of =
        report_data.append(['', ''])
        report_data.append(['Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        report_data.append(['Input File:', os.path.basename(file_path)])
        report_data.append(['', ''])
        report_data.append(['INPUT DATA SUMMARY', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Total Rows in BEFORE:', len(df_before)])
        report_data.append(['Total Rows in AFTER:', len(df_after)])
        report_data.append(['', ''])
        report_data.append(['AFTER SHEET ANALYSIS (What needs attention)', ''])
        report_data.append(['─' * 50, ''])  # Using box drawing character instead of =
        report_data.append(['', ''])
        
        # Perfect Matches
        perfect_pct = (perfect_count / len(df_after) * 100) if len(df_after) > 0 else 0
        report_data.append(['✓ PERFECT MATCHES', f'{perfect_count:,} rows ({perfect_pct:.1f}%)'])
        report_data.append(['  Description:', 'All 4 columns (A,B,C,D) match with BEFORE'])
        report_data.append(['  Action Required:', 'None - These rows are identical'])
        report_data.append(['', ''])
        
        # Need To ReRecord
        rerecord_total = new_count + total_modified
        rerecord_pct = (rerecord_total / len(df_after) * 100) if len(df_after) > 0 else 0
        report_data.append(['⚠ NEED TO RERECORD', f'{rerecord_total:,} rows ({rerecord_pct:.1f}%)'])
        report_data.append(['', ''])
        report_data.append(['  NEW Rows:', f'{new_count:,} rows'])
        report_data.append(['    Description:', 'No matching B+C combination found in BEFORE'])
        report_data.append(['    Action Required:', 'These are completely new entries'])
        report_data.append(['', ''])
        report_data.append(['  MODIFIED Rows:', f'{total_modified:,} rows'])
        report_data.append(['    Description:', 'B+C match with BEFORE, but A or D differ'])
        report_data.append(['', ''])
        report_data.append(['    CastingKey mismatch only:', f'{modified_count_a_only:,} rows'])
        report_data.append(['      Detail:', 'Column A differs, D matches'])
        report_data.append(['', ''])
        report_data.append(['    StartFrame mismatch only:', f'{modified_count_d_only:,} rows'])
        report_data.append(['      Detail:', 'Column D differs, A matches'])
        report_data.append(['', ''])
        report_data.append(['    Both CastingKey & StartFrame mismatch:', f'{modified_count_both:,} rows'])
        report_data.append(['      Detail:', 'Both columns A and D differ'])
        report_data.append(['', ''])
        
        # Need To Inspect
        inspect_pct = (inspect_count / len(df_after) * 100) if len(df_after) > 0 else 0
        report_data.append(['🔍 NEED TO INSPECT', f'{inspect_count:,} rows ({inspect_pct:.1f}%)'])
        report_data.append(['  Description:', 'A,D match with BEFORE, but B or C differ'])
        report_data.append(['  Action Required:', 'Manual review needed - unexpected changes'])
        report_data.append(['', ''])
        
        # Unmatched BEFORE rows
        report_data.append(['UNMATCHED BEFORE ROWS', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Rows in BEFORE not found in AFTER:', f'{unmatched_before_count:,} rows'])
        report_data.append(['Description:', 'These exist in BEFORE but not in AFTER'])
        report_data.append(['Note:', 'These are ignored as per requirements'])
        report_data.append(['', ''])
        
        # Output sheets summary
        report_data.append(['OUTPUT SHEETS', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Sheet: Need To ReRecord', f'{rerecord_total:,} rows'])
        report_data.append(['  Contains:', 'NEW rows + MODIFIED rows (categorized)'])
        report_data.append(['  Highlighting:', 'Orange = columns that differ'])
        report_data.append(['', ''])
        report_data.append(['Sheet: Need To Inspect', f'{inspect_count:,} rows'])
        report_data.append(['  Contains:', 'Rows where A,D match but B or C differ'])
        report_data.append(['  Format:', 'Shows both AFTER and BEFORE columns'])
        report_data.append(['  Highlighting:', 'Orange = differing columns'])
        report_data.append(['', ''])
        report_data.append(['Sheet: Perfect Matches', f'{perfect_count:,} rows'])
        report_data.append(['  Contains:', 'Rows that match perfectly (A,B,C,D)'])
        report_data.append(['  Format:', 'Shows both AFTER and BEFORE columns'])
        report_data.append(['', ''])
        
        # Percentage breakdown
        report_data.append(['PERCENTAGE BREAKDOWN', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['Perfect Matches:', f'{perfect_pct:.1f}%'])
        report_data.append(['Need To ReRecord:', f'{rerecord_pct:.1f}%'])
        report_data.append(['Need To Inspect:', f'{inspect_pct:.1f}%'])
        report_data.append(['Total:', '100.0%'])
        report_data.append(['', ''])
        
        # Key notes
        report_data.append(['KEY NOTES', ''])
        report_data.append(['-' * 50, ''])
        report_data.append(['- Analysis focuses on AFTER sheet (what needs attention)', ''])
        report_data.append(['- Matching is based on B+C columns (StrOrigin + Desc)', ''])
        report_data.append(['- NEW rows have no B+C match in BEFORE', ''])
        report_data.append(['- MODIFIED rows have B+C match but A or D differ', ''])
        report_data.append(['- INSPECT rows have unexpected differences (A,D match, B or C differ)', ''])
        report_data.append(['- Orange highlighting shows cells that differ', ''])
        
        log(f"  ✓ Report summary created with {len(report_data)} rows of information")
        
        log("Step 6/8: Creating output Excel file...")
        
        # Generate output path
        base_name = os.path.splitext(file_path)[0]
        output_path = f"{base_name}_comparison.xlsx"
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
            temp_file = tmp.name
        
        # Create Excel writer
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Sheet 1: Report Summary
            df_report = pd.DataFrame(report_data, columns=['Metric', 'Value'])
            df_report.to_excel(writer, sheet_name='Report Summary', index=False, header=False)
            log(f"  ✓ Sheet 'Report Summary' created")
            
            # Sheet 2: Need To ReRecord
            rerecord_data = []
            rerecord_labels = []
            
            if new_rows:
                rerecord_data.extend(new_rows)
                rerecord_labels.extend(['NEW'] * len(new_rows))
            
            if modified_rows_a_only:
                if rerecord_data:
                    rerecord_data.append([''] * 4)  # Empty separator row
                    rerecord_labels.append('')
                rerecord_data.append(['--- MODIFIED ROWS (CastingKey mismatch only) ---'] + [''] * 3)
                rerecord_labels.append('')
                rerecord_data.extend(modified_rows_a_only)
                rerecord_labels.extend(['MODIFIED'] * len(modified_rows_a_only))
            
            if modified_rows_d_only:
                if rerecord_data:
                    rerecord_data.append([''] * 4)  # Empty separator row
                    rerecord_labels.append('')
                rerecord_data.append(['--- MODIFIED ROWS (StartFrame mismatch only) ---'] + [''] * 3)
                rerecord_labels.append('')
                rerecord_data.extend(modified_rows_d_only)
                rerecord_labels.extend(['MODIFIED'] * len(modified_rows_d_only))
            
            if modified_rows_both:
                if rerecord_data:
                    rerecord_data.append([''] * 4)  # Empty separator row
                    rerecord_labels.append('')
                rerecord_data.append(['--- MODIFIED ROWS (Both CastingKey and StartFrame mismatch) ---'] + [''] * 3)
                rerecord_labels.append('')
                rerecord_data.extend(modified_rows_both)
                rerecord_labels.extend(['MODIFIED'] * len(modified_rows_both))
            
            if rerecord_data:
                df_rerecord = pd.DataFrame(rerecord_data, columns=['CastingKey', 'StrOrigin', 'Desc', 'StartFrame'])
                df_rerecord.to_excel(writer, sheet_name='Need To ReRecord', index=False)
                log(f"  ✓ Sheet 'Need To ReRecord' created with {new_count + total_modified:,} rows")
            
            # Sheet 3: Need To Inspect
            if inspect_rows:
                df_inspect = pd.DataFrame(
                    inspect_rows,
                    columns=['AFTER_CastingKey', 'AFTER_StrOrigin', 'AFTER_Desc', 'AFTER_StartFrame',
                            'BEFORE_CastingKey', 'BEFORE_StrOrigin', 'BEFORE_Desc', 'BEFORE_StartFrame']
                )
                df_inspect.to_excel(writer, sheet_name='Need To Inspect', index=False)
                log(f"  ✓ Sheet 'Need To Inspect' created with {inspect_count:,} rows")
            
            # Sheet 4: Perfect Matches
            if perfect_matches:
                df_perfect = pd.DataFrame(
                    perfect_matches,
                    columns=['AFTER_CastingKey', 'AFTER_StrOrigin', 'AFTER_Desc', 'AFTER_StartFrame',
                            'BEFORE_CastingKey', 'BEFORE_StrOrigin', 'BEFORE_Desc', 'BEFORE_StartFrame']
                )
                df_perfect.to_excel(writer, sheet_name='Perfect Matches', index=False)
                log(f"  ✓ Sheet 'Perfect Matches' created with {perfect_count:,} rows")
        
        log("Step 7/8: Applying formatting...")
        
        # Apply formatting using openpyxl
        try:
            wb = load_workbook(temp_file)
            
            # Define colors
            orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            red_font = Font(color="FF0000", bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Format the "Report Summary" sheet
            if 'Report Summary' in wb.sheetnames:
                ws = wb['Report Summary']
                log(f"  → Formatting 'Report Summary' sheet...")
                
                # Set column widths
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 40
                
                # Format title row (row 1)
                ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
                
                # Format section headers and key items
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value and isinstance(cell_value, str):
                        # Main section headers
                        if any(header in cell_value for header in ['INPUT DATA SUMMARY', 'AFTER SHEET ANALYSIS', 
                                                                   'UNMATCHED BEFORE ROWS', 'OUTPUT SHEETS', 
                                                                   'PERCENTAGE BREAKDOWN', 'KEY NOTES']):
                            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color="1F4E78")
                        # Sub-section headers with symbols
                        elif cell_value.startswith('✓') or cell_value.startswith('⚠') or cell_value.startswith('🔍'):
                            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=10, color="C00000")
                
                log(f"    ✓ Report summary formatted")
            
            # Format "Need To ReRecord" sheet
            if 'Need To ReRecord' in wb.sheetnames:
                ws = wb['Need To ReRecord']
                log(f"  → Applying formatting to 'Need To ReRecord'...")
                
                # Set column widths
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 35
                ws.column_dimensions['D'].width = 15
                
                # Format header row
                for col_idx in range(1, 5):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Find section headers and apply formatting
                current_row = 2  # Start after header row
                
                # Format NEW rows section
                if new_rows:
                    # No highlighting needed for NEW rows
                    current_row += len(new_rows)
                
                # Skip empty separator row if exists
                if new_rows and modified_rows_a_only:
                    current_row += 1
                
                # Format MODIFIED ROWS (A only) section
                if modified_rows_a_only:
                    # Format section header
                    ws.cell(row=current_row, column=1).font = red_font
                    current_row += 1
                    
                    # Highlight MODIFIED rows (A only)
                    for row_idx in range(len(modified_rows_a_only)):
                        actual_row = current_row + row_idx
                        diff_cols = modified_diffs_a_only[row_idx]
                        for col_idx in diff_cols:
                            cell = ws.cell(row=actual_row, column=col_idx + 1)
                            cell.fill = orange_fill
                    
                    current_row += len(modified_rows_a_only)
                
                # Skip empty separator row if exists
                if modified_rows_a_only and modified_rows_d_only:
                    current_row += 1
                
                # Format MODIFIED ROWS (D only) section
                if modified_rows_d_only:
                    # Format section header
                    ws.cell(row=current_row, column=1).font = red_font
                    current_row += 1
                    
                    # Highlight MODIFIED rows (D only)
                    for row_idx in range(len(modified_rows_d_only)):
                        actual_row = current_row + row_idx
                        diff_cols = modified_diffs_d_only[row_idx]
                        for col_idx in diff_cols:
                            cell = ws.cell(row=actual_row, column=col_idx + 1)
                            cell.fill = orange_fill
                    
                    current_row += len(modified_rows_d_only)
                
                # Skip empty separator row if exists
                if modified_rows_d_only and modified_rows_both:
                    current_row += 1
                
                # Format MODIFIED ROWS (Both) section
                if modified_rows_both:
                    # Format section header
                    ws.cell(row=current_row, column=1).font = red_font
                    current_row += 1
                    
                    # Highlight MODIFIED rows (Both)
                    for row_idx in range(len(modified_rows_both)):
                        actual_row = current_row + row_idx
                        diff_cols = modified_diffs_both[row_idx]
                        for col_idx in diff_cols:
                            cell = ws.cell(row=actual_row, column=col_idx + 1)
                            cell.fill = orange_fill
                
                log(f"    ✓ Formatted {len(new_rows)} NEW rows and {total_modified} MODIFIED rows (categorized)")
            
            # Format "Need To Inspect" sheet
            if 'Need To Inspect' in wb.sheetnames:
                ws = wb['Need To Inspect']
                log(f"  → Applying formatting to 'Need To Inspect'...")
                
                # Set column widths for better readability
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 35
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 45
                ws.column_dimensions['G'].width = 35
                ws.column_dimensions['H'].width = 15
                
                # Format header row
                for col_idx in range(1, 9):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Highlight differences (only StrOrigin and Desc columns that differ)
                for row_idx, diff_cols in enumerate(inspect_diffs, start=2):
                    for col_idx in diff_cols:
                        cell = ws.cell(row=row_idx, column=col_idx + 1)
                        cell.fill = orange_fill
                log(f"    ✓ Highlighted {len(inspect_diffs)} rows")
            
            # Format "Perfect Matches" sheet
            if 'Perfect Matches' in wb.sheetnames:
                ws = wb['Perfect Matches']
                log(f"  → Setting column widths for 'Perfect Matches'...")
                
                # Set column widths for better readability
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 35
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 45
                ws.column_dimensions['G'].width = 35
                ws.column_dimensions['H'].width = 15
                
                # Format header row
                for col_idx in range(1, 9):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
            
            # CRITICAL: Save and close the workbook
            wb.save(output_path)
            wb.close()
            log(f"  ✓ Workbook saved and closed successfully")
            
        except Exception as e:
            log(f"✗ ERROR while applying formatting:")
            log(f"  Error type: {type(e).__name__}")
            log(f"  Error message: {str(e)}")
            traceback.print_exc()
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return
        
        # Clean up temporary file
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
                log(f"  ✓ Temporary file removed")
            except Exception as e:
                log(f"  ⚠ Warning: Could not remove temp file: {e}")
        
        print("=" * 60)
        log("✓ COMPARE BEFORE VS AFTER - Completed Successfully")
        log(f"  Input file: {file_path}")
        log(f"  Output file: {output_path}")
        log(f"")
        log(f"  === AFTER SHEET ANALYSIS (what needs attention) ===")
        log(f"  Total rows in AFTER: {len(df_after):,}")
        log(f"  ├─ Perfect Matches: {perfect_count:,} ({perfect_count/len(df_after)*100:.1f}%)")
        log(f"  ├─ Need To ReRecord: {new_count + total_modified:,} ({(new_count + total_modified)/len(df_after)*100:.1f}%)")
        log(f"  │  ├─ NEW rows: {new_count:,}")
        log(f"  │  └─ Modified rows: {total_modified:,}")
        log(f"  │     ├─ CastingKey mismatch only: {modified_count_a_only:,}")
        log(f"  │     ├─ StartFrame mismatch only: {modified_count_d_only:,}")
        log(f"  │     └─ Both CastingKey and StartFrame mismatch: {modified_count_both:,}")
        log(f"  └─ Need To Inspect: {inspect_count:,} ({inspect_count/len(df_after)*100:.1f}%)")
        log(f"")
        log(f"  Rows in BEFORE that were never matched: {len(df_before) - len(matched_before):,}")
        log(f"  (These exist in BEFORE but not in AFTER - ignored as requested)")
        print("=" * 60)
        
    except Exception as e:
        log(f"✗ FATAL ERROR in compare_before_after:")
        log(f"  Error type: {type(e).__name__}")
        log(f"  Error message: {str(e)}")
        log("  Full traceback:")
        traceback.print_exc()
    finally:
        # Always clean up temp file, even if there was an error
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass


# ========== GUI ==========
def create_gui():
    """Create the main GUI window"""
    log("Creating GUI...")
    
    root = tk.Tk()
    root.title("Excel Comparison Tool")
    root.geometry("400x250")
    
    # Center window
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (400 // 2)
    y = (screen_height // 2) - (250 // 2)
    root.geometry(f"400x250+{x}+{y}")
    
    log(f"  → Window size: 400x250")
    log(f"  → Window position: ({x}, {y})")
    
    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(expand=True, fill=tk.BOTH)
    
    # Title
    title = tk.Label(
        frame, 
        text="Excel Comparison Tool",
        font=("Arial", 14, "bold")
    )
    title.pack(pady=(0, 20))
    
    # Button 1: Order By Similarity
    btn1 = tk.Button(
        frame,
        text="Order By Similarity",
        command=order_by_similarity,
        width=25,
        height=2,
        font=("Arial", 10)
    )
    btn1.pack(pady=5)
    
    # Description 1
    desc1 = tk.Label(
        frame,
        text="Compare B+D vs C+E, reorder by similarity",
        font=("Arial", 8),
        fg="gray"
    )
    desc1.pack(pady=(0, 15))
    
    # Button 2: Compare Before vs After
    btn2 = tk.Button(
        frame,
        text="Compare Before vs After",
        command=compare_before_after,
        width=25,
        height=2,
        font=("Arial", 10)
    )
    btn2.pack(pady=5)
    
    # Description 2
    desc2 = tk.Label(
        frame,
        text="Compare two tabs and categorize differences",
        font=("Arial", 8),
        fg="gray"
    )
    desc2.pack()
    
    log("✓ GUI initialized")
    print("\n" + "=" * 60)
    log("GUI Ready - Awaiting user action...")
    print("=" * 60 + "\n")
    
    root.mainloop()


if __name__ == "__main__":
    try:
        create_gui()
    except Exception as e:
        log(f"✗ FATAL ERROR in main:")
        log(f"  Error type: {type(e).__name__}")
        log(f"  Error message: {str(e)}")
        log("  Full traceback:")
        traceback.print_exc()
