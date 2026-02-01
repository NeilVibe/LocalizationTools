import os
import openpyxl
from tkinter import filedialog, Tk

def collect_file_names(folder_path):
    """Recursively collect all file names (without extension) from folder."""
    base_ids = set()
    for root, _, files in os.walk(folder_path):
        for f in files:
            name, _ = os.path.splitext(f)
            if name:
                base_ids.add(name.strip())
    return sorted(base_ids)

def read_excel_all_rows(path):
    """Read all rows from all sheets into a list of lists."""
    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # skip empty rows
                all_rows.append(list(row))
    return all_rows

def match_files_to_excel_rows(file_names, excel_rows):
    """Match file names to any cell in Excel rows (case-insensitive)."""
    matches = []
    for fname in file_names:
        fname_lower = fname.lower()
        matched_row = None
        for row in excel_rows:
            for cell in row:
                if cell is not None and fname_lower == str(cell).strip().lower():
                    matched_row = row
                    break
            if matched_row:
                break
        if matched_row:
            matches.append([fname] + list(matched_row))
        else:
            matches.append([fname, "NO_MATCH"])
    return matches

def save_matches_to_excel(matches, output_path):
    """Save matches to a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matches"
    for row in matches:
        ws.append(row)
    wb.save(output_path)

def main():
    Tk().withdraw()
    folder_path = filedialog.askdirectory(title="Select Folder to Scan for Files")
    if not folder_path:
        print("No folder selected.")
        return
    excel_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        print("No Excel file selected.")
        return

    print("[INFO] Collecting file names...")
    file_names = collect_file_names(folder_path)
    print(f"[INFO] Found {len(file_names)} unique file names.")

    print("[INFO] Reading Excel file...")
    excel_rows = read_excel_all_rows(excel_path)
    print(f"[INFO] Loaded {len(excel_rows)} rows from Excel.")

    print("[INFO] Matching...")
    matches = match_files_to_excel_rows(file_names, excel_rows)

    output_path = os.path.join(folder_path, "matched_file_rows.xlsx")
    save_matches_to_excel(matches, output_path)
    print(f"[INFO] Output saved to: {output_path}")

if __name__ == "__main__":
    main()