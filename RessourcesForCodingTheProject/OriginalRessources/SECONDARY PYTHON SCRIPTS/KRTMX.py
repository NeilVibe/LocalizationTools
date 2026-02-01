import os
import sys
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from lxml import etree

def get_all_xml_files(input_folder):
    xml_files = []
    for dirpath, _, filenames in os.walk(input_folder):
        for file in filenames:
            if file.lower().endswith(".xml"):
                xml_files.append(os.path.join(dirpath, file))
    return xml_files

def parse_xml_file(file_path):
    try:
        txt = open(file_path, encoding='utf-8').read()
    except Exception as e:
        print(f"[ERROR] Error reading {file_path!r}: {e}")
        return None
    txt = re.sub(r'&(?!lt;|gt;|amp;|apos;|quot;)', '&amp;', txt)
    wrapped = "<ROOT>\n" + txt + "\n</ROOT>"
    rec_parser = etree.XMLParser(recover=True)
    try:
        recovered = etree.fromstring(wrapped.encode('utf-8'), parser=rec_parser)
    except etree.XMLSyntaxError as e:
        print(f"[ERROR] Fatal parse error (recover mode) on {file_path!r}: {e}")
        return None
    strict_parser = etree.XMLParser(recover=False)
    blob = etree.tostring(recovered, encoding='utf-8')
    try:
        return etree.fromstring(blob, parser=strict_parser)
    except etree.XMLSyntaxError:
        return recovered

def replace_newlines_text(text):
    if text is None:
        return text
    cleaned = text.replace('\n', '&lt;br/&gt;')
    cleaned = cleaned.replace('\\n', '&lt;br/&gt;')
    return cleaned

def combine_xmls_to_tmx_ko2ko(input_folder, output_file):
    xml_files = get_all_xml_files(input_folder)
    total_files = len(xml_files)
    print(f"[TMX] Total XML files found: {total_files}")
    file_counter = 0

    tmx = etree.Element("tmx", version="1.4")
    header_attr = {
        "creationtool": "KoreanToKoreanTMX",
        "creationtoolversion": "1.0",
        "segtype": "sentence",
        "o-tmf": "UTF-8",
        "adminlang": "ko",
        "srclang": "ko",
        "datatype": "PlainText"
    }
    etree.SubElement(tmx, "header", header_attr)
    body = etree.SubElement(tmx, "body")
    tu_count = 0
    XML_NS = "http://www.w3.org/XML/1998/namespace"

    for xml_file_path in xml_files:
        file_counter += 1
        print(f"[TMX] Processing file {file_counter} of {total_files}: {xml_file_path}")
        try:
            if os.path.getsize(xml_file_path) == 0:
                continue
        except Exception as e:
            print(f"[ERROR] Error getting file size for {xml_file_path}: {e}")
            continue
        xml_root = parse_xml_file(xml_file_path)
        if xml_root is None:
            continue

        doc_name = os.path.basename(xml_file_path)

        for loc in xml_root.iter("LocStr"):
            korean_text = replace_newlines_text((loc.get("StrOrigin") or "").strip())
            target_text = replace_newlines_text((loc.get("Str") or "").strip())
            string_id   = (loc.get("StringId")  or "").strip()

            # Only require both fields to be non-empty
            if not target_text:
                continue
            if not korean_text:
                continue

            tu = etree.SubElement(body, "tu", {
                "creationid": "KoreanToKoreanTMX",
                "changeid":   "KoreanToKoreanTMX"
            })
            tuv_ko_src = etree.SubElement(tu, "tuv", {f"{{{XML_NS}}}lang": "ko"})
            seg_ko_src = etree.SubElement(tuv_ko_src, "seg")
            seg_ko_src.text = korean_text
            tuv_ko_tgt = etree.SubElement(tu, "tuv", {f"{{{XML_NS}}}lang": "ko"})
            seg_ko_tgt = etree.SubElement(tuv_ko_tgt, "seg")
            seg_ko_tgt.text = target_text
            prop_doc = etree.SubElement(tu, "prop", {"type": "x-document"})
            prop_doc.text = doc_name
            prop_ctx = etree.SubElement(tu, "prop", {"type": "x-context"})
            prop_ctx.text = string_id or "NoStringId"
            tu_count += 1

    if tu_count == 0:
        msg = "No valid translation units found under the selected folder."
        print(msg)
        return False

    xml_decl = '<?xml version="1.0" encoding="UTF-8"?>\n'
    doctype  = '<!DOCTYPE tmx SYSTEM "tmx14.dtd">\n'
    xml_bytes = etree.tostring(tmx, pretty_print=True, encoding='UTF-8')
    xml_str   = xml_bytes.decode('utf-8')
    final_content = xml_decl + doctype + xml_str
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(final_content)
        print(f"[TMX] Successfully combined {tu_count} units into:\n{output_file}")
        return True
    except Exception as e:
        print(f"[TMX] Error writing TMX {output_file}: {e}")
        return False

def parse_tmx(tmx_path):
    print(f"[TMX] Parsing TMX file: {tmx_path}")
    tree = etree.parse(tmx_path)
    root = tree.getroot()
    body = root.find("body")
    if body is None:
        raise Exception("TMX missing <body>")
    units = []
    tu_list = body.findall("tu")
    total_tu = len(tu_list)
    print(f"[TMX] Found {total_tu} translation units.")
    for idx, tu in enumerate(tu_list):
        # --- FIX: Get all <tuv> in order ---
        tuvs = tu.findall("tuv")
        kr = None
        en = None
        if len(tuvs) >= 2:
            # Assume first is source, second is target
            seg1 = tuvs[0].find("seg")
            seg2 = tuvs[1].find("seg")
            if seg1 is not None and seg1.text is not None:
                kr = seg1.text.strip()
            if seg2 is not None and seg2.text is not None:
                en = seg2.text.strip()
        else:
            # fallback: old logic
            for tuv in tuvs:
                lang = (tuv.get("{http://www.w3.org/XML/1998/namespace}lang") or
                        tuv.get("xml:lang") or
                        tuv.get("lang") or "")
                seg = tuv.find("seg")
                if seg is None or seg.text is None:
                    continue
                if lang.lower().startswith("ko"):
                    if kr is None:
                        kr = seg.text.strip()
                    else:
                        en = seg.text.strip()
                elif lang.lower().startswith("en"):
                    en = seg.text.strip()
        ctx = None
        for prop in tu.findall("prop"):
            if prop.get("type") == "x-context" and prop.text:
                ctx = prop.text.strip()
        if kr:
            units.append({'kr': kr, 'en': en, 'context': ctx})
        if (idx + 1) % 1000 == 0 or (idx + 1) == total_tu:
            print(f"[TMX] Processed {idx+1}/{total_tu} TUs...", flush=True)
    print(f"[TMX] Finished parsing TMX. {len(units)} valid units with KR.\n")
    return units

def build_perfect_match_kr_dict(bank):
    # Only use KR as key, value is EN (or target KO)
    return {k[1]: v['en'] for k, v in bank.items() if v['en'] is not None}

def excel_translate_perfect_kr_no_iskorean():
    import openpyxl
    # 1. Ask for TMX file
    tmx_path = filedialog.askopenfilename(
        title="Select TMX File",
        filetypes=[("TMX Files", "*.tmx"), ("All Files", "*.*")]
    )
    if not tmx_path:
        messagebox.showinfo("Info", "No TMX file selected. Operation cancelled.")
        return

    # 2. Parse TMX and build perfect KR dict
    try:
        tmx_units = parse_tmx(tmx_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse TMX: {e}")
        return
    if not tmx_units:
        messagebox.showerror("Error", "No valid translation units found in TMX.")
        return

    bank = {}
    for unit in tmx_units:
        key = (unit['context'], unit['kr'])
        bank[key] = {'en': unit['en'], 'context': unit['context'], 'kr': unit['kr']}
    perfect_kr_dict = build_perfect_match_kr_dict(bank)

    # 3. Ask for Excel file
    excel_path = filedialog.askopenfilename(
        title="Select Excel File to Translate",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not excel_path:
        messagebox.showinfo("Info", "No Excel file selected. Operation cancelled.")
        return

    # 4. Open Excel, process, and save
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        updated = 0
        total = 0
        for row in ws.iter_rows(min_row=1, max_col=2):
            kr_cell = row[0]
            en_cell = row[1] if len(row) > 1 else None
            kr_val = kr_cell.value
            if kr_val is None:
                continue
            total += 1
            kr_val_str = str(kr_val).strip()
            if kr_val_str in perfect_kr_dict:
                if en_cell is None:
                    en_cell = ws.cell(row=kr_cell.row, column=2)
                en_cell.value = perfect_kr_dict[kr_val_str]
                updated += 1
        wb.save(excel_path)
        messagebox.showinfo("Excel Translate", f"Excel translation complete!\n\n{updated} of {total} rows updated.\nFile overwritten:\n{excel_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process Excel file: {e}")
        return

class Ko2KoTMXGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Korean-to-Korean XML to TMX Converter")
        self.geometry("600x270")
        self.resizable(False, False)
        self.selected_folder = tk.StringVar()
        self.folder_chosen   = False
        self.create_widgets()

    def create_widgets(self):
        pad = {'padx': 10, 'pady': 10}
        frm = tk.Frame(self)
        frm.pack(fill="x", **pad)
        btn = tk.Button(frm, text="Select Folder", command=self.select_folder)
        btn.grid(row=0, column=0, sticky="w")
        self.select_folder_button = btn
        lbl = tk.Label(frm, text="No folder selected", wraplength=500)
        lbl.grid(row=0, column=1, sticky="w", padx=10)
        self.folder_label = lbl
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill="x", **pad)
        tk.Button(btn_frame, text="Convert to NORMAL TMX (Korean→Korean)", command=self.convert_normal).pack(pady=10)
        # --- New Excel Translate Button ---
        tk.Button(
            btn_frame,
            text="Excel Translate (PERFECT MATCH ONLY KR, NO isKorean)",
            command=excel_translate_perfect_kr_no_iskorean,
            bg="#32cd32", fg="black", height=2
        ).pack(pady=5)

    def select_folder(self):
        fld = filedialog.askdirectory(
            title="Select Parent Folder Containing XML Files"
        )
        if fld:
            self.selected_folder.set(fld)
            self.folder_label.config(text=fld)
            self.select_folder_button.config(bg="green")
            self.folder_chosen = True

    def convert_normal(self):
        if not self.folder_chosen:
            messagebox.showerror("Error", "Please select a folder first.")
            return
        out = filedialog.asksaveasfilename(
            title="Save Combined NORMAL TMX File As",
            defaultextension=".tmx",
            filetypes=[("TMX Files", "*.tmx"), ("All Files", "*.*")]
        )
        if not out:
            messagebox.showinfo("Info", "No output file selected. Conversion cancelled.")
            return
        threading.Thread(
            target=self.threaded_convert,
            args=(out,),
            daemon=True
        ).start()

    def threaded_convert(self, out_file):
        ok = combine_xmls_to_tmx_ko2ko(self.selected_folder.get(), out_file)
        if ok:
            self.after(0, lambda: messagebox.showinfo(
                "Success",
                f"Combined NORMAL TMX (Korean→Korean) created:\n{out_file}"
            ))
        else:
            self.after(0, lambda: messagebox.showerror(
                "Error",
                f"Failed to create the TMX."
            ))

def main():
    app = Ko2KoTMXGUI()
    app.mainloop()
    sys.exit(0)

if __name__ == "__main__":
    main()