import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import xml.etree.ElementTree as ET
import pandas as pd
import json
import os
import sys
import shutil
import codecs
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

def load_files(file_type="XML"):
    """Open file dialogs to select files"""
    root = tk.Tk()
    root.withdraw()
    
    # Select first file (XML or JSON)
    if file_type == "XML":
        file_path = filedialog.askopenfilename(
            title=f"Select {file_type} file",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
    else:  # JSON
        file_path = filedialog.askopenfilename(
            title=f"Select {file_type} file",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
    
    if not file_path:
        print(f"No {file_type} file selected")
        return None, None
    
    # Select Excel file
    excel_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
    )
    if not excel_path:
        print("No Excel file selected")
        return None, None
    
    return file_path, excel_path

def load_single_file():
    """Open file dialog to select a single XML or JSON file"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="Select XML or JSON file to extract",
        filetypes=[
            ("Supported files", "*.xml;*.json"),
            ("XML files", "*.xml"),
            ("JSON files", "*.json"),
            ("All files", "*.*")
        ]
    )
    
    if not file_path:
        print("No file selected")
        return None
    
    return file_path

def clean_xml_content(xml_path):
    """Clean XML content to remove BOM and other issues"""
    with open(xml_path, 'rb') as f:
        raw_content = f.read()
    
    # Remove BOM if present
    if raw_content.startswith(codecs.BOM_UTF8):
        raw_content = raw_content[len(codecs.BOM_UTF8):]
    elif raw_content.startswith(codecs.BOM_UTF16_LE) or raw_content.startswith(codecs.BOM_UTF16_BE):
        raw_content = raw_content[2:]
    
    # Decode to string
    try:
        content = raw_content.decode('utf-8')
    except:
        try:
            content = raw_content.decode('utf-8-sig')
        except:
            content = raw_content.decode('latin-1')
    
    # Remove any leading characters before <?xml
    if '<?xml' in content:
        xml_start = content.index('<?xml')
        if xml_start > 0:
            content = content[xml_start:]
    
    return content

def indent_xml(elem, level=0):
    """Manually indent XML for Python versions < 3.9"""
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

def get_language_mapping():
    """Define the mapping between Excel columns and locale codes"""
    return {
        'EN': 'en-US',
        'FR': 'fr-FR',
        'DE': 'de-DE', 
        'SPA-ES': 'es-ES',
        'SPA-MX': 'es-MX',
        'IT': 'it-IT',
        'PT': 'pt-PT',
        'PT-BR': 'pt-BR',
        'RU': 'ru-RU',
        'CN': 'zh-CN',  # Simplified Chinese
        'zh-Hans': 'zh-Hans',  # Alternative Simplified Chinese
        'TW': 'zh-TW',  # Traditional Chinese
        'zh-Hant': 'zh-Hant',  # Alternative Traditional Chinese
        'JP': 'ja-JP',
        'PL': 'pl-PL',
        'TR': 'tr-TR',
        'es-419': 'es-419'  # Latin American Spanish
    }

def get_json_language_mapping():
    """Define the mapping for JSON files (includes additional codes)"""
    base_mapping = get_language_mapping()
    # Add any additional mappings specific to JSON
    base_mapping.update({
        'PT-BR': 'pt-BR',
        'es-419': 'es-419',
        'zh-Hans': 'zh-Hans',
        'zh-Hant': 'zh-Hant'
    })
    return base_mapping

def identify_language_columns(df):
    """Identify which columns contain which languages based on headers"""
    headers = df.iloc[0]
    
    language_columns = {}
    lang_map = get_language_mapping()
    
    # Scan through columns to find language pairs
    for col_idx in range(2, len(df.columns), 2):  # Start from column C (index 2), step by 2
        if col_idx < len(headers):
            header_val = str(headers.iloc[col_idx]).strip().upper() if not pd.isna(headers.iloc[col_idx]) else ""
            
            # Match header with language codes
            for lang_key, locale in lang_map.items():
                if lang_key in header_val or header_val in lang_key:
                    language_columns[locale] = (col_idx, col_idx + 1)
                    print(f"Found {lang_key} ({locale}) at columns {col_idx}, {col_idx + 1}")
                    break
    
    return language_columns

def generate_length_report(xml_path, output_dir):
    """Generate a report of all text lines longer than 100 characters"""
    
    print("\nAnalyzing text lengths in XML file...")
    
    # Read the XML file directly to handle encoding properly
    with open(xml_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Parse the XML
    try:
        root = ET.fromstring(content)
    except:
        # Try with cleaned content if direct parsing fails
        content = clean_xml_content(xml_path)
        root = ET.fromstring(content)
    
    namespace_uri = 'http://config.mgt.xboxlive.com/schema/localization/1'
    namespace = {'ns': namespace_uri}
    
    long_texts = []
    total_texts = 0
    
    # Scan all LocalizedString elements
    for localized_string in root.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        for value in localized_string.findall('ns:Value', namespace):
            locale = value.get('locale')
            text = value.text if value.text else ""
            total_texts += 1
            
            char_count = len(text)
            
            # Check if text is longer than 100 characters
            if char_count > 100:
                long_texts.append({
                    'id': string_id,
                    'locale': locale,
                    'length': char_count,
                    'text': text
                })
                print(f"Found long text: {locale} - {char_count} chars - {string_id[:30]}...")
    
    print(f"\nTotal texts analyzed: {total_texts}")
    print(f"Texts exceeding 100 characters: {len(long_texts)}")
    
    # Sort by length (longest first)
    long_texts.sort(key=lambda x: x['length'], reverse=True)
    
    # Generate report file
    report_path = os.path.join(output_dir, "text_length_report.txt")
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("TEXT LENGTH REDUCTION REQUEST\n")
        f.write("="*80 + "\n\n")
        f.write(f"Total texts analyzed: {total_texts}\n")
        f.write(f"Total texts exceeding 100 characters: {len(long_texts)}\n")
        f.write(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        # Group by locale
        locales = {}
        for item in long_texts:
            if item['locale'] not in locales:
                locales[item['locale']] = []
            locales[item['locale']].append(item)
        
        for locale in sorted(locales.keys()):
            f.write(f"\n{'='*60}\n")
            f.write(f"LANGUAGE: {locale}\n")
            f.write(f"{'='*60}\n\n")
            
            for item in locales[locale]:
                f.write(f"ID: {item['id']}\n")
                f.write(f"LENGTH: {item['length']} characters\n")
                f.write(f"TEXT: {item['text']}\n")
                f.write("-"*40 + "\n\n")
    
    print(f"\nLength report generated: {report_path}")
    
    return report_path, len(long_texts)

def create_localized_file(root_elem, locale, translations, namespace_uri, output_dir):
    """Create a localized XML file for a specific language pair (KR + target language)"""
    
    ET.register_namespace('', namespace_uri)
    
    # Create new root element with namespace
    new_root = ET.Element('Localization')
    new_root.set('xmlns', namespace_uri)
    
    # Add DevDisplayLocale
    dev_locale = ET.SubElement(new_root, 'DevDisplayLocale')
    dev_locale.set('locale', locale)
    
    # Add all LocalizedString elements with only KR and target language
    namespace = {'ns': namespace_uri}
    for localized_string in root_elem.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        # Create new LocalizedString element
        new_localized = ET.SubElement(new_root, 'LocalizedString')
        new_localized.set('id', string_id)
        
        # Add Korean value
        for value in localized_string.findall('ns:Value', namespace):
            if value.get('locale') == 'ko-KR':
                kr_value = ET.SubElement(new_localized, 'Value')
                kr_value.set('locale', 'ko-KR')
                kr_value.text = value.text
                break
        
        # Add target language value
        if string_id in translations and locale in translations[string_id]:
            target_value = ET.SubElement(new_localized, 'Value')
            target_value.set('locale', locale)
            target_value.text = translations[string_id][locale]
        else:
            # Check if value already exists in original XML
            for value in localized_string.findall('ns:Value', namespace):
                if value.get('locale') == locale:
                    target_value = ET.SubElement(new_localized, 'Value')
                    target_value.set('locale', locale)
                    target_value.text = value.text
                    break
    
    # Create tree and save
    new_tree = ET.ElementTree(new_root)
    
    # Format the XML
    if sys.version_info >= (3, 9):
        ET.indent(new_tree, space="  ")
    else:
        indent_xml(new_root)
    
    # Save file
    filename = f"localization_{locale.replace('-', '_')}.xml"
    filepath = os.path.join(output_dir, filename)
    new_tree.write(filepath, encoding='utf-8', xml_declaration=True)
    
    return filepath

def process_xml_translation_mapping(xml_path, excel_path):
    """Main function to process the XML translation mapping"""
    
    # Clean and parse XML
    print("Cleaning XML content...")
    xml_content = clean_xml_content(xml_path)
    
    # Parse the cleaned XML content
    try:
        root = ET.fromstring(xml_content)
        tree = ET.ElementTree(root)
    except ET.ParseError as e:
        print(f"XML Parse Error: {e}")
        raise
    
    # Get namespace from root
    namespace_uri = 'http://config.mgt.xboxlive.com/schema/localization/1'
    namespace = {'ns': namespace_uri}
    
    # Register namespace to avoid ns0 prefix
    ET.register_namespace('', namespace_uri)
    
    # Read Excel file
    df = pd.read_excel(excel_path, header=None)
    
    print(f"Excel shape: {df.shape}")
    
    # Identify language columns
    language_columns = identify_language_columns(df)
    
    # Create dictionaries for mapping
    korean_to_id = {}
    id_to_korean = {}
    
    # Collect all Korean texts and their corresponding IDs from XML
    for localized_string in root.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        for value in localized_string.findall('ns:Value', namespace):
            if value.get('locale') == 'ko-KR':
                korean_text = value.text
                if korean_text:
                    korean_text = korean_text.strip()
                    korean_to_id[korean_text] = string_id
                    id_to_korean[string_id] = korean_text
                break
    
    print(f"\nFound {len(korean_to_id)} Korean strings in XML")
    
    # Process Excel data
    translation_updates = {}
    processed_pairs = 0
    
    for index in range(1, len(df)):  # Start from row 2 (index 1) to skip headers
        row = df.iloc[index]
        
        # Skip if both Korean columns are empty
        if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
            continue
        
        # Get Korean texts from first two columns
        kr_text_1 = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else None
        kr_text_2 = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else None
        
        # Process each language
        for locale, (col1_idx, col2_idx) in language_columns.items():
            # Get translations
            trans_1 = str(row.iloc[col1_idx]).strip() if not pd.isna(row.iloc[col1_idx]) else None
            trans_2 = str(row.iloc[col2_idx]).strip() if not pd.isna(row.iloc[col2_idx]) else None
            
            # Map first Korean text to first translation
            if kr_text_1 and trans_1 and kr_text_1 in korean_to_id:
                string_id = korean_to_id[kr_text_1]
                if string_id not in translation_updates:
                    translation_updates[string_id] = {}
                translation_updates[string_id][locale] = trans_1
                processed_pairs += 1
            
            # Map second Korean text to second translation
            if kr_text_2 and trans_2 and kr_text_2 in korean_to_id:
                string_id = korean_to_id[kr_text_2]
                if string_id not in translation_updates:
                    translation_updates[string_id] = {}
                translation_updates[string_id][locale] = trans_2
                processed_pairs += 1
    
    print(f"\nProcessed {processed_pairs} translation pairs")
    print(f"Unique IDs to update: {len(translation_updates)}")
    
    # Apply updates to main XML (master file)
    updates_count = 0
    for localized_string in root.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        if string_id in translation_updates:
            for locale, translation in translation_updates[string_id].items():
                # Check if this locale already exists
                locale_exists = False
                for value in localized_string.findall('ns:Value', namespace):
                    if value.get('locale') == locale:
                        value.text = translation
                        locale_exists = True
                        updates_count += 1
                        break
                
                # If locale doesn't exist, create new Value element
                if not locale_exists:
                    new_value = ET.SubElement(localized_string, 'Value')
                    new_value.set('locale', locale)
                    new_value.text = translation
                    updates_count += 1
    
    # Create output directory for localized files
    script_dir = os.path.dirname(os.path.abspath(__file__))
    local_dir = os.path.join(script_dir, "xml_result")
    
    # Clean and recreate the directory
    if os.path.exists(local_dir):
        shutil.rmtree(local_dir)
    os.makedirs(local_dir)
    
    # Save master file with all languages
    master_output_path = xml_path.replace('.xml', '_master.xml')
    
    # Format the XML
    if sys.version_info >= (3, 9):
        ET.indent(tree, space="  ")
    else:
        indent_xml(root)
    
    tree.write(master_output_path, encoding='utf-8', xml_declaration=True)
    
    print(f"\nMaster file saved to: {master_output_path}")
    
    # Generate length report using the master file
    report_path, long_text_count = generate_length_report(master_output_path, local_dir)
    
    # Create individual localized files for each language pair
    all_locales = list(get_language_mapping().values())
    created_files = []
    
    for locale in all_locales:
        filepath = create_localized_file(root, locale, translation_updates, namespace_uri, local_dir)
        created_files.append(filepath)
        print(f"Created localized file: {os.path.basename(filepath)}")
    
    return master_output_path, local_dir, report_path, long_text_count, updates_count

def update_json_translations(json_obj, korean_mapping, language_columns, df):
    """Recursively update JSON object with translations based on Korean text mapping"""
    updates_count = 0
    
    if isinstance(json_obj, dict):
        for key, value in json_obj.items():
            if key == "name" or key == "description":
                if isinstance(value, dict) and "ko-KR" in value:
                    korean_text = value["ko-KR"].strip()
                    
                    # Find matching Korean text in Excel
                    for index in range(1, len(df)):
                        row = df.iloc[index]
                        
                        kr_text_1 = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else None
                        kr_text_2 = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else None
                        
                        # Check if either Korean text matches
                        if kr_text_1 == korean_text or kr_text_2 == korean_text:
                            # Update translations for all languages
                            for locale, (col1_idx, col2_idx) in language_columns.items():
                                # Map the appropriate column based on which Korean text matched
                                if kr_text_1 == korean_text:
                                    trans = str(row.iloc[col1_idx]).strip() if not pd.isna(row.iloc[col1_idx]) else None
                                elif kr_text_2 == korean_text:
                                    trans = str(row.iloc[col2_idx]).strip() if not pd.isna(row.iloc[col2_idx]) else None
                                else:
                                    trans = None
                                
                                if trans:
                                    # Handle different locale formats
                                    json_locale = locale
                                    if locale == "zh-CN":
                                        json_locale = "zh-Hans"
                                    elif locale == "zh-TW":
                                        json_locale = "zh-Hant"
                                    elif locale == "es-MX":
                                        json_locale = "es-419"
                                    
                                    if json_locale in value or json_locale == locale:
                                        value[json_locale] = trans
                                        updates_count += 1
                            break
            else:
                # Recurse into nested objects
                count = update_json_translations(value, korean_mapping, language_columns, df)
                updates_count += count
    elif isinstance(json_obj, list):
        for item in json_obj:
            count = update_json_translations(item, korean_mapping, language_columns, df)
            updates_count += count
    
    return updates_count

def process_json_translation_mapping(json_path, excel_path):
    """Main function to process the JSON translation mapping"""
    
    # Read JSON file
    print("Reading JSON file...")
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    # Read Excel file
    df = pd.read_excel(excel_path, header=None)
    print(f"Excel shape: {df.shape}")
    
    # Identify language columns
    language_columns = identify_language_columns(df)
    
    # Create mapping of Korean texts
    korean_mapping = {}
    
    # Process the JSON and update translations
    print("\nProcessing translations...")
    updates_count = update_json_translations(json_data, korean_mapping, language_columns, df)
    
    print(f"\nTotal updates made: {updates_count}")
    
    # Create output directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_dir = os.path.join(script_dir, "json_result")
    
    if os.path.exists(json_dir):
        shutil.rmtree(json_dir)
    os.makedirs(json_dir)
    
    # Save updated JSON
    output_path = os.path.join(json_dir, "updated_" + os.path.basename(json_path))
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    
    print(f"Updated JSON saved to: {output_path}")
    
    # Create individual language JSON files
    languages = ["en-US", "fr-FR", "de-DE", "es-ES", "es-419", "it-IT", "pt-BR", 
                 "ru-RU", "zh-Hans", "zh-Hant", "ja-JP", "pl-PL", "tr-TR"]
    
    for lang in languages:
        lang_data = create_language_specific_json(json_data, lang)
        lang_filename = f"localization_{lang.replace('-', '_')}.json"
        lang_path = os.path.join(json_dir, lang_filename)
        
        with open(lang_path, 'w', encoding='utf-8') as f:
            json.dump(lang_data, f, ensure_ascii=False, indent=2)
        
        print(f"Created language file: {lang_filename}")
    
    return output_path, json_dir, updates_count

def create_language_specific_json(json_data, target_locale):
    """Create a JSON with only Korean and target language"""
    import copy
    result = copy.deepcopy(json_data)
    
    def filter_languages(obj):
        if isinstance(obj, dict):
            for key, value in obj.items():
                if key == "name" or key == "description":
                    if isinstance(value, dict):
                        # Keep only Korean and target language
                        filtered = {}
                        if "ko-KR" in value:
                            filtered["ko-KR"] = value["ko-KR"]
                        if target_locale in value:
                            filtered[target_locale] = value[target_locale]
                        obj[key] = filtered
                else:
                    filter_languages(value)
        elif isinstance(obj, list):
            for item in obj:
                filter_languages(item)
    
    filter_languages(result)
    return result

def extract_xml_to_excel(xml_path):
    """Extract translations from XML to Excel format - ONLY strings with Korean"""
    print(f"Extracting translations from XML: {xml_path}")
    
    # Clean and parse XML
    xml_content = clean_xml_content(xml_path)
    
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"XML Parse Error: {e}")
        raise
    
    namespace_uri = 'http://config.mgt.xboxlive.com/schema/localization/1'
    namespace = {'ns': namespace_uri}
    
    # Collect ONLY translations that have Korean
    translations_with_korean = {}
    skipped_count = 0
    
    for localized_string in root.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        # First check if Korean exists
        has_korean = False
        korean_text = None
        
        for value in localized_string.findall('ns:Value', namespace):
            if value.get('locale') == 'ko-KR':
                korean_text = value.text if value.text else ""
                if korean_text:  # Only proceed if Korean text exists and is not empty
                    has_korean = True
                break
        
        # Only collect this string if it has Korean
        if has_korean:
            translations_with_korean[string_id] = {'ko-KR': korean_text}
            
            # Now get all other translations for this string
            for value in localized_string.findall('ns:Value', namespace):
                locale = value.get('locale')
                if locale != 'ko-KR':  # Skip Korean since we already have it
                    text = value.text if value.text else ""
                    translations_with_korean[string_id][locale] = text
        else:
            skipped_count += 1
            print(f"Skipped string without Korean: {string_id}")
    
    print(f"\nTotal strings with Korean: {len(translations_with_korean)}")
    print(f"Skipped strings without Korean: {skipped_count}")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    
    # Define language order
    languages = ['en-US', 'fr-FR', 'de-DE', 'es-ES', 'es-MX', 'it-IT', 
                 'pt-PT', 'pt-BR', 'ru-RU', 'zh-CN', 'zh-TW', 'ja-JP', 
                 'pl-PL', 'tr-TR']
    
    # Create header row
    headers = ['KR', 'KR']  # Two Korean columns
    for lang in languages:
        # Map back to simpler names for headers
        if lang == 'en-US':
            headers.extend(['EN', 'EN'])
        elif lang == 'fr-FR':
            headers.extend(['FR', 'FR'])
        elif lang == 'de-DE':
            headers.extend(['DE', 'DE'])
        elif lang == 'es-ES':
            headers.extend(['SPA-ES', 'SPA-ES'])
        elif lang == 'es-MX':
            headers.extend(['SPA-MX', 'SPA-MX'])
        elif lang == 'it-IT':
            headers.extend(['IT', 'IT'])
        elif lang == 'pt-PT':
            headers.extend(['PT', 'PT'])
        elif lang == 'pt-BR':
            headers.extend(['PT-BR', 'PT-BR'])
        elif lang == 'ru-RU':
            headers.extend(['RU', 'RU'])
        elif lang == 'zh-CN':
            headers.extend(['CN', 'CN'])
        elif lang == 'zh-TW':
            headers.extend(['TW', 'TW'])
        elif lang == 'ja-JP':
            headers.extend(['JP', 'JP'])
        elif lang == 'pl-PL':
            headers.extend(['PL', 'PL'])
        elif lang == 'tr-TR':
            headers.extend(['TR', 'TR'])
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add translations
    row_num = 2
    pair_count = 0
    
    # Convert to list for pairing
    string_ids = list(translations_with_korean.keys())
    
    # Process in pairs
    for i in range(0, len(string_ids), 2):
        # First item of pair
        if i < len(string_ids):
            string_id1 = string_ids[i]
            korean_text1 = translations_with_korean[string_id1].get('ko-KR', '')
            ws.cell(row=row_num, column=1, value=korean_text1)
            
            # Add translations for first item
            col = 3
            for lang in languages:
                trans = translations_with_korean[string_id1].get(lang, '')
                ws.cell(row=row_num, column=col, value=trans)
                col += 2
        
        # Second item of pair
        if i + 1 < len(string_ids):
            string_id2 = string_ids[i + 1]
            korean_text2 = translations_with_korean[string_id2].get('ko-KR', '')
            ws.cell(row=row_num, column=2, value=korean_text2)
            
            # Add translations for second item
            col = 4
            for lang in languages:
                trans = translations_with_korean[string_id2].get(lang, '')
                ws.cell(row=row_num, column=col, value=trans)
                col += 2
        
        row_num += 1
        pair_count += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Save Excel file
    output_dir = os.path.dirname(xml_path)
    output_filename = os.path.splitext(os.path.basename(xml_path))[0] + "_translation_bank.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb.save(output_path)
    
    print(f"Excel file created: {output_path}")
    print(f"Total translation pairs: {pair_count}")
    print(f"Total strings with Korean: {len(translations_with_korean)}")
    
    return output_path, pair_count, len(translations_with_korean), skipped_count

def extract_json_to_excel(json_path):
    """Extract translations from JSON to Excel format - ONLY items with Korean"""
    print(f"Extracting translations from JSON: {json_path}")
    
    # Read JSON file
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    # Collect ONLY translations that have Korean
    translations_with_korean = []
    skipped_count = 0
    
    def extract_translations(obj, path=""):
        nonlocal skipped_count
        if isinstance(obj, dict):
            for key, value in obj.items():
                if key in ["name", "description"]:
                    if isinstance(value, dict) and "ko-KR" in value:
                        korean_text = value.get("ko-KR", "")
                        if korean_text:  # Only add if Korean text exists and is not empty
                            translations_with_korean.append({
                                'id': path + "/" + key,
                                'type': key,
                                'translations': value
                            })
                        else:
                            skipped_count += 1
                            print(f"Skipped empty Korean in: {path}/{key}")
                    elif isinstance(value, dict):
                        skipped_count += 1
                        print(f"Skipped item without Korean: {path}/{key}")
                else:
                    extract_translations(value, path + "/" + key if path else key)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                extract_translations(item, path + f"[{i}]")
    
    extract_translations(json_data)
    
    print(f"\nTotal items with Korean: {len(translations_with_korean)}")
    print(f"Skipped items without Korean: {skipped_count}")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    
    # Define language order for JSON (includes additional languages)
    languages = ['en-US', 'fr-FR', 'de-DE', 'es-ES', 'es-419', 'it-IT', 
                 'pt-BR', 'ru-RU', 'zh-Hans', 'zh-Hant', 'ja-JP', 
                 'pl-PL', 'tr-TR']
    
    # Create header row
    headers = ['KR', 'KR']  # Two Korean columns
    for lang in languages:
        # Map to simpler names for headers
        if lang == 'en-US':
            headers.extend(['EN', 'EN'])
        elif lang == 'fr-FR':
            headers.extend(['FR', 'FR'])
        elif lang == 'de-DE':
            headers.extend(['DE', 'DE'])
        elif lang == 'es-ES':
            headers.extend(['SPA-ES', 'SPA-ES'])
        elif lang == 'es-419':
            headers.extend(['SPA-MX', 'SPA-MX'])
        elif lang == 'it-IT':
            headers.extend(['IT', 'IT'])
        elif lang == 'pt-BR':
            headers.extend(['PT-BR', 'PT-BR'])
        elif lang == 'ru-RU':
            headers.extend(['RU', 'RU'])
        elif lang == 'zh-Hans':
            headers.extend(['CN', 'CN'])
        elif lang == 'zh-Hant':
            headers.extend(['TW', 'TW'])
        elif lang == 'ja-JP':
            headers.extend(['JP', 'JP'])
        elif lang == 'pl-PL':
            headers.extend(['PL', 'PL'])
        elif lang == 'tr-TR':
            headers.extend(['TR', 'TR'])
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add translations
    row_num = 2
    pair_count = 0
    
    # Process in pairs
    for i in range(0, len(translations_with_korean), 2):
        # First item of pair
        if i < len(translations_with_korean):
            trans1 = translations_with_korean[i]
            korean_text1 = trans1['translations'].get('ko-KR', '')
            ws.cell(row=row_num, column=1, value=korean_text1)
            
            # Add translations for first item
            col = 3
            for lang in languages:
                text = trans1['translations'].get(lang, '')
                ws.cell(row=row_num, column=col, value=text)
                col += 2
        
        # Second item of pair
        if i + 1 < len(translations_with_korean):
            trans2 = translations_with_korean[i + 1]
            korean_text2 = trans2['translations'].get('ko-KR', '')
            ws.cell(row=row_num, column=2, value=korean_text2)
            
            # Add translations for second item
            col = 4
            for lang in languages:
                text = trans2['translations'].get(lang, '')
                ws.cell(row=row_num, column=col, value=text)
                col += 2
        
        row_num += 1
        pair_count += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Save Excel file
    output_dir = os.path.dirname(json_path)
    output_filename = os.path.splitext(os.path.basename(json_path))[0] + "_translation_bank.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb.save(output_path)
    
    print(f"Excel file created: {output_path}")
    print(f"Total translation pairs: {pair_count}")
    print(f"Total items with Korean: {len(translations_with_korean)}")
    
    return output_path, pair_count, len(translations_with_korean), skipped_count

def extract_long_xml_to_excel(xml_path):
    """Extract ONLY translations above 100 characters from XML to Excel format"""
    print(f"Extracting long translations (>100 chars) from XML: {xml_path}")
    
    # Clean and parse XML
    xml_content = clean_xml_content(xml_path)
    
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"XML Parse Error: {e}")
        raise
    
    namespace_uri = 'http://config.mgt.xboxlive.com/schema/localization/1'
    namespace = {'ns': namespace_uri}
    
    # Collect ONLY translations that have Korean AND have at least one text >100 chars
    long_translations = {}
    skipped_short = 0
    skipped_no_korean = 0
    
    for localized_string in root.findall('ns:LocalizedString', namespace):
        string_id = localized_string.get('id')
        
        # First check if Korean exists
        has_korean = False
        korean_text = None
        
        for value in localized_string.findall('ns:Value', namespace):
            if value.get('locale') == 'ko-KR':
                korean_text = value.text if value.text else ""
                if korean_text:
                    has_korean = True
                break
        
        if not has_korean:
            skipped_no_korean += 1
            continue
        
        # Check if any translation is >100 characters
        has_long_text = False
        temp_translations = {'ko-KR': korean_text}
        
        for value in localized_string.findall('ns:Value', namespace):
            locale = value.get('locale')
            text = value.text if value.text else ""
            temp_translations[locale] = text
            
            if len(text) > 100:
                has_long_text = True
        
        # Only include if it has at least one long text
        if has_long_text:
            long_translations[string_id] = temp_translations
        else:
            skipped_short += 1
    
    print(f"\nTotal strings with long text (>100 chars): {len(long_translations)}")
    print(f"Skipped strings without Korean: {skipped_no_korean}")
    print(f"Skipped strings with all texts ≤100 chars: {skipped_short}")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Long Translations"
    
    # Define language order
    languages = ['en-US', 'fr-FR', 'de-DE', 'es-ES', 'es-MX', 'it-IT', 
                 'pt-PT', 'pt-BR', 'ru-RU', 'zh-CN', 'zh-TW', 'ja-JP', 
                 'pl-PL', 'tr-TR']
    
    # Create header row
    headers = ['KR', 'KR']
    for lang in languages:
        if lang == 'en-US':
            headers.extend(['EN', 'EN'])
        elif lang == 'fr-FR':
            headers.extend(['FR', 'FR'])
        elif lang == 'de-DE':
            headers.extend(['DE', 'DE'])
        elif lang == 'es-ES':
            headers.extend(['SPA-ES', 'SPA-ES'])
        elif lang == 'es-MX':
            headers.extend(['SPA-MX', 'SPA-MX'])
        elif lang == 'it-IT':
            headers.extend(['IT', 'IT'])
        elif lang == 'pt-PT':
            headers.extend(['PT', 'PT'])
        elif lang == 'pt-BR':
            headers.extend(['PT-BR', 'PT-BR'])
        elif lang == 'ru-RU':
            headers.extend(['RU', 'RU'])
        elif lang == 'zh-CN':
            headers.extend(['CN', 'CN'])
        elif lang == 'zh-TW':
            headers.extend(['TW', 'TW'])
        elif lang == 'ja-JP':
            headers.extend(['JP', 'JP'])
        elif lang == 'pl-PL':
            headers.extend(['PL', 'PL'])
        elif lang == 'tr-TR':
            headers.extend(['TR', 'TR'])
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red for long text
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add translations
    row_num = 2
    pair_count = 0
    
    # Convert to list for pairing
    string_ids = list(long_translations.keys())
    
    # Process in pairs
    for i in range(0, len(string_ids), 2):
        # First item of pair
        if i < len(string_ids):
            string_id1 = string_ids[i]
            korean_text1 = long_translations[string_id1].get('ko-KR', '')
            ws.cell(row=row_num, column=1, value=korean_text1)
            
            # Add translations for first item
            col = 3
            for lang in languages:
                trans = long_translations[string_id1].get(lang, '')
                cell = ws.cell(row=row_num, column=col, value=trans)
                # Highlight cells with text >100 chars
                if len(trans) > 100:
                    cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                col += 2
        
        # Second item of pair
        if i + 1 < len(string_ids):
            string_id2 = string_ids[i + 1]
            korean_text2 = long_translations[string_id2].get('ko-KR', '')
            ws.cell(row=row_num, column=2, value=korean_text2)
            
            # Add translations for second item
            col = 4
            for lang in languages:
                trans = long_translations[string_id2].get(lang, '')
                cell = ws.cell(row=row_num, column=col, value=trans)
                # Highlight cells with text >100 chars
                if len(trans) > 100:
                    cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                col += 2
        
        row_num += 1
        pair_count += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Save Excel file
    output_dir = os.path.dirname(xml_path)
    output_filename = os.path.splitext(os.path.basename(xml_path))[0] + "_long_text_bank.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb.save(output_path)
    
    print(f"Excel file created: {output_path}")
    print(f"Total translation pairs: {pair_count}")
    print(f"Total strings with long text: {len(long_translations)}")
    
    return output_path, pair_count, len(long_translations), skipped_short, skipped_no_korean

def extract_long_json_to_excel(json_path):
    """Extract ONLY translations above 100 characters from JSON to Excel format"""
    print(f"Extracting long translations (>100 chars) from JSON: {json_path}")
    
    # Read JSON file
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    # Collect ONLY translations that have Korean AND have at least one text >100 chars
    long_translations = []
    skipped_short = 0
    skipped_no_korean = 0
    
    def extract_long_translations(obj, path=""):
        nonlocal skipped_short, skipped_no_korean
        if isinstance(obj, dict):
            for key, value in obj.items():
                if key in ["name", "description"]:
                    if isinstance(value, dict) and "ko-KR" in value:
                        korean_text = value.get("ko-KR", "")
                        if not korean_text:
                            skipped_no_korean += 1
                            continue
                        
                        # Check if any translation is >100 characters
                        has_long_text = False
                        for locale, text in value.items():
                            if len(text) > 100:
                                has_long_text = True
                                break
                        
                        if has_long_text:
                            long_translations.append({
                                'id': path + "/" + key,
                                'type': key,
                                'translations': value
                            })
                        else:
                            skipped_short += 1
                    elif isinstance(value, dict):
                        skipped_no_korean += 1
                else:
                    extract_long_translations(value, path + "/" + key if path else key)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                extract_long_translations(item, path + f"[{i}]")
    
    extract_long_translations(json_data)
    
    print(f"\nTotal items with long text (>100 chars): {len(long_translations)}")
    print(f"Skipped items without Korean: {skipped_no_korean}")
    print(f"Skipped items with all texts ≤100 chars: {skipped_short}")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Long Translations"
    
    # Define language order for JSON
    languages = ['en-US', 'fr-FR', 'de-DE', 'es-ES', 'es-419', 'it-IT', 
                 'pt-BR', 'ru-RU', 'zh-Hans', 'zh-Hant', 'ja-JP', 
                 'pl-PL', 'tr-TR']
    
    # Create header row
    headers = ['KR', 'KR']
    for lang in languages:
        if lang == 'en-US':
            headers.extend(['EN', 'EN'])
        elif lang == 'fr-FR':
            headers.extend(['FR', 'FR'])
        elif lang == 'de-DE':
            headers.extend(['DE', 'DE'])
        elif lang == 'es-ES':
            headers.extend(['SPA-ES', 'SPA-ES'])
        elif lang == 'es-419':
            headers.extend(['SPA-MX', 'SPA-MX'])
        elif lang == 'it-IT':
            headers.extend(['IT', 'IT'])
        elif lang == 'pt-BR':
            headers.extend(['PT-BR', 'PT-BR'])
        elif lang == 'ru-RU':
            headers.extend(['RU', 'RU'])
        elif lang == 'zh-Hans':
            headers.extend(['CN', 'CN'])
        elif lang == 'zh-Hant':
            headers.extend(['TW', 'TW'])
        elif lang == 'ja-JP':
            headers.extend(['JP', 'JP'])
        elif lang == 'pl-PL':
            headers.extend(['PL', 'PL'])
        elif lang == 'tr-TR':
            headers.extend(['TR', 'TR'])
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red for long text
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add translations
    row_num = 2
    pair_count = 0
    
    # Process in pairs
    for i in range(0, len(long_translations), 2):
        # First item of pair
        if i < len(long_translations):
            trans1 = long_translations[i]
            korean_text1 = trans1['translations'].get('ko-KR', '')
            ws.cell(row=row_num, column=1, value=korean_text1)
            
            # Add translations for first item
            col = 3
            for lang in languages:
                text = trans1['translations'].get(lang, '')
                cell = ws.cell(row=row_num, column=col, value=text)
                # Highlight cells with text >100 chars
                if len(text) > 100:
                    cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                col += 2
        
        # Second item of pair
        if i + 1 < len(long_translations):
            trans2 = long_translations[i + 1]
            korean_text2 = trans2['translations'].get('ko-KR', '')
            ws.cell(row=row_num, column=2, value=korean_text2)
            
            # Add translations for second item
            col = 4
            for lang in languages:
                text = trans2['translations'].get(lang, '')
                cell = ws.cell(row=row_num, column=col, value=text)
                # Highlight cells with text >100 chars
                if len(text) > 100:
                    cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                col += 2
        
        row_num += 1
        pair_count += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Save Excel file
    output_dir = os.path.dirname(json_path)
    output_filename = os.path.splitext(os.path.basename(json_path))[0] + "_long_text_bank.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb.save(output_path)
    
    print(f"Excel file created: {output_path}")
    print(f"Total translation pairs: {pair_count}")
    print(f"Total items with long text: {len(long_translations)}")
    
    return output_path, pair_count, len(long_translations), skipped_short, skipped_no_korean

def extract_long_texts():
    """Auto-detect file type and extract only long texts (>100 chars) to Excel"""
    try:
        file_path = load_single_file()
        
        if not file_path:
            return
        
        # Detect file type
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xml':
            print("Detected XML file - extracting long texts")
            output_path, pair_count, total_long, skipped_short, skipped_no_kr = extract_long_xml_to_excel(file_path)
            
            root = tk.Tk()
            root.withdraw()
            message = (f"Long Text Bank Created (>100 chars)!\n\n"
                      f"Output file: {output_path}\n\n"
                      f"Total translation pairs: {pair_count}\n"
                      f"Total strings with long text: {total_long}\n"
                      f"Skipped (all texts ≤100 chars): {skipped_short}\n"
                      f"Skipped (no Korean): {skipped_no_kr}\n\n"
                      f"Cells with text >100 chars are highlighted.\n"
                      f"You can now edit this Excel file and re-import it.")
            messagebox.showinfo("Long Text Extract Success", message)
            
        elif file_ext == '.json':
            print("Detected JSON file - extracting long texts")
            output_path, pair_count, total_long, skipped_short, skipped_no_kr = extract_long_json_to_excel(file_path)
            
            root = tk.Tk()
            root.withdraw()
            message = (f"Long Text Bank Created (>100 chars)!\n\n"
                      f"Output file: {output_path}\n\n"
                      f"Total translation pairs: {pair_count}\n"
                      f"Total items with long text: {total_long}\n"
                      f"Skipped (all texts ≤100 chars): {skipped_short}\n"
                      f"Skipped (no Korean): {skipped_no_kr}\n\n"
                      f"Cells with text >100 chars are highlighted.\n"
                      f"You can now edit this Excel file and re-import it.")
            messagebox.showinfo("Long Text Extract Success", message)
            
        else:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", "Unsupported file type. Please select an XML or JSON file.")
            
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

def extract_to_excel():
    """Auto-detect file type and extract to Excel"""
    try:
        file_path = load_single_file()
        
        if not file_path:
            return
        
        # Detect file type
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xml':
            print("Detected XML file")
            output_path, pair_count, total_strings, skipped = extract_xml_to_excel(file_path)
            
            root = tk.Tk()
            root.withdraw()
            message = (f"XML Translation Bank Created!\n\n"
                      f"Output file: {output_path}\n\n"
                      f"Total translation pairs: {pair_count}\n"
                      f"Total strings with Korean: {total_strings}\n"
                      f"Skipped (no Korean): {skipped}\n\n"
                      f"You can now edit this Excel file and re-import it.")
            messagebox.showinfo("Extract Success", message)
            
        elif file_ext == '.json':
            print("Detected JSON file")
            output_path, pair_count, total_strings, skipped = extract_json_to_excel(file_path)
            
            root = tk.Tk()
            root.withdraw()
            message = (f"JSON Translation Bank Created!\n\n"
                      f"Output file: {output_path}\n\n"
                      f"Total translation pairs: {pair_count}\n"
                      f"Total items with Korean: {total_strings}\n"
                      f"Skipped (no Korean): {skipped}\n\n"
                      f"You can now edit this Excel file and re-import it.")
            messagebox.showinfo("Extract Success", message)
            
        else:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", "Unsupported file type. Please select an XML or JSON file.")
            
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

def process_xml():
    """Handle XML processing"""
    try:
        xml_path, excel_path = load_files("XML")
        
        if xml_path and excel_path:
            print(f"XML file: {xml_path}")
            print(f"Excel file: {excel_path}")
            print("\nProcessing XML translations...")
            
            master_path, local_dir, report_path, long_text_count, updates_count = process_xml_translation_mapping(xml_path, excel_path)
            
            # Show success message
            root = tk.Tk()
            root.withdraw()
            
            message = (f"XML Translation mapping completed!\n\n"
                      f"Master file: {master_path}\n"
                      f"Localized files: {local_dir}\n"
                      f"Length report: {report_path}\n\n"
                      f"Total updates: {updates_count}\n"
                      f"Found {long_text_count} texts exceeding 100 characters.")
            
            messagebox.showinfo("XML Processing Success", message)
            
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

def process_json():
    """Handle JSON processing"""
    try:
        json_path, excel_path = load_files("JSON")
        
        if json_path and excel_path:
            print(f"JSON file: {json_path}")
            print(f"Excel file: {excel_path}")
            print("\nProcessing JSON translations...")
            
            output_path, json_dir, updates_count = process_json_translation_mapping(json_path, excel_path)
            
            # Show success message
            root = tk.Tk()
            root.withdraw()
            
            message = (f"JSON Translation mapping completed!\n\n"
                      f"Updated file: {output_path}\n"
                      f"Output directory: {json_dir}\n\n"
                      f"Total updates: {updates_count}")
            
            messagebox.showinfo("JSON Processing Success", message)
            
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

def create_main_window():
    """Create the main GUI window"""
    root = tk.Tk()
    root.title("Translation Mapper")
    root.geometry("450x400")  # Increased height for new button
    root.resizable(False, False)
    
    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")
    
    # Create main frame
    main_frame = ttk.Frame(root, padding="20")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    # Title
    title_label = ttk.Label(main_frame, text="Translation Mapper", font=('Arial', 16, 'bold'))
    title_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
    
    # Import Section
    import_label = ttk.Label(main_frame, text="Import Translations (Excel → XML/JSON):", font=('Arial', 11, 'bold'))
    import_label.grid(row=1, column=0, columnspan=2, pady=(10, 10), sticky='w')
    
    # XML Button
    xml_button = ttk.Button(
        main_frame, 
        text="Excel → XML",
        command=lambda: [root.withdraw(), process_xml(), root.quit()],
        width=20
    )
    xml_button.grid(row=2, column=0, padx=10, pady=5)
    
    # JSON Button
    json_button = ttk.Button(
        main_frame, 
        text="Excel → JSON",
        command=lambda: [root.withdraw(), process_json(), root.quit()],
        width=20
    )
    json_button.grid(row=2, column=1, padx=10, pady=5)
    
    # Info labels for import
    xml_info = ttk.Label(main_frame, text="Update XML localization\nwith Excel translations", font=('Arial', 8))
    xml_info.grid(row=3, column=0, pady=(0, 10))
    
    json_info = ttk.Label(main_frame, text="Update JSON trophy\nwith Excel translations", font=('Arial', 8))
    json_info.grid(row=3, column=1, pady=(0, 10))
    
    # Separator
    separator = ttk.Separator(main_frame, orient='horizontal')
    separator.grid(row=4, column=0, columnspan=2, sticky='ew', pady=10)
    
    # Export Section
    export_label = ttk.Label(main_frame, text="Export Translations (XML/JSON → Excel):", font=('Arial', 11, 'bold'))
    export_label.grid(row=5, column=0, columnspan=2, pady=(10, 10), sticky='w')
    
    # Extract ALL to Excel Button
    extract_button = ttk.Button(
        main_frame, 
        text="Extract ALL to Excel Bank",
        command=lambda: [root.withdraw(), extract_to_excel(), root.quit()],
        width=43
    )
    extract_button.grid(row=6, column=0, columnspan=2, pady=5)
    
    extract_info = ttk.Label(main_frame, text="Create editable Excel translation bank from XML/JSON\n(All texts with Korean)", font=('Arial', 8))
    extract_info.grid(row=7, column=0, columnspan=2, pady=(0, 5))
    
    # NEW: Extract Long Texts Button
    extract_long_button = ttk.Button(
        main_frame, 
        text="Extract LONG TEXTS (>100 chars) to Excel",
        command=lambda: [root.withdraw(), extract_long_texts(), root.quit()],
        width=43,
        style='Danger.TButton'  # Red style for emphasis
    )
    extract_long_button.grid(row=8, column=0, columnspan=2, pady=5)
    
    # Create custom style for the long text button
    style = ttk.Style()
    style.configure('Danger.TButton', foreground='darkred')
    
    extract_long_info = ttk.Label(main_frame, text="Extract ONLY texts exceeding 100 characters\n(For text reduction/editing, maintains same format for re-import)", font=('Arial', 8), foreground='darkred')
    extract_long_info.grid(row=9, column=0, columnspan=2, pady=(0, 10))
    
    # Exit button
    exit_button = ttk.Button(main_frame, text="Exit", command=root.quit, width=20)
    exit_button.grid(row=10, column=0, columnspan=2, pady=(10, 0))
    
    root.mainloop()

def main():
    """Main execution function"""
    print(f"Python version: {sys.version}")
    print("\nTranslation Mapper Started")
    print("=" * 50)
    create_main_window()

if __name__ == "__main__":
    main()