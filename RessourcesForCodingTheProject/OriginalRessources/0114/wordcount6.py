
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ============================================================
#  Translation Utilities – Word-Count & StrOrigin Comparison
#  (2024-06 – updated with “Non-Priority” split, granular
#   Sequencer/Dialog handling, colours, “None” category)
#
#  2024-06-XX UPDATE
#  ------------------------------------------------------------
#  1.  “Sequencer (Total)” renamed to “Sequencer (Grand Total)”
#  2.  Quest-Groups are now rolled-up:
#        • All  Sequencer/Other/QuestGroup_*  entries are merged and
#          reported as:  “Sequencer/Other/QuestGroup_XX”
#  3.  cd_seq_memory files are now rolled-up:
#        • All  Sequencer/Sequencer/cd_seq_memory_*  entries are merged and
#          reported as: “Sequencer/Sequencer/cd_seq_memory_XX”
#  4.  Detailed Summary sheet no longer lists individual QuestGroup*
#      or cd_seq_memory_* rows – only the consolidated rows.
# ============================================================

import os
import re
import sys
import copy
import threading
import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
from pathlib import Path
from datetime import datetime
import xml.etree.ElementTree as ET

from lxml import etree as LET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ----------------------------------------------------------------------
#  HELPERS TO HANDLE “FROZEN” EXECUTABLES  (PyInstaller, cx_Freeze …)
# ----------------------------------------------------------------------
def get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


# ----------------------------------------------------------------------
#  CONSTANTS  (adjust as needed)
# ----------------------------------------------------------------------
LANGUAGE_FOLDER     = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
EXPORT_FOLDER       = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
DIFF_OUTPUT_FOLDER  = Path.cwd() / "diff_entries_output"
MISSING_XML_FILE    = Path.cwd() / "Missing_From_Export.xml"

BASE_DIR      = get_base_dir()
BEFORE_FOLDER  = BASE_DIR / "BEFORE"
AFTER_FOLDER   = BASE_DIR / "AFTER"

# ----------------------------------------------------------------------
#  RULE-SETS
# ----------------------------------------------------------------------
NON_PRIORITY_FOLDERS = {"ItemGroup", "Gimmick", "MultiChange"}   # → “Non-Priority”
DIALOG_SUBS          = ["AIDialog", "NarrationDialog",
                        "QuestDialog", "StageCloseDialog"]
SEQUENCER_TOPS       = ["Faction", "Main", "Other", "Sequencer"]

# ------------- colour map (big categories)
BIG_CAT_COLOURS = {
    "Dialog"       : "C6EFCE",  # light-green
    "Sequencer"    : "FFE599",  # light-orange
    "System"       : "D9D2E9",  # light-purple
    "World"        : "F8CBAD",  # light-red
    "Platform"     : "A9D08E",  # light-teal / green-gray
    "None"         : "DDD9C4",  # light-brown
    "Non-Priority" : "D9D9D9"   # light-grey
}
BIG_CAT_FILLS = {k: PatternFill("solid", fgColor=v) for k, v in BIG_CAT_COLOURS.items()}

_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')
korean_re      = re.compile(r'[\uac00-\ud7a3]')

def fix_bad_entities(x: str) -> str:
    return _bad_entity_re.sub("&amp;", x)

def count_words(txt: str) -> int:
    return len([w for w in re.split(r"\s+", txt.strip()) if w])

def parse_xml_file(p: Path) -> LET._Element:
    raw     = p.read_text(encoding="utf-8", errors="ignore")
    wrapped = f"<ROOT>
{fix_bad_entities(raw)}
</ROOT>"
    parser  = LET.XMLParser(recover=True, huge_tree=True)
    return LET.fromstring(wrapped.encode("utf-8"), parser=parser)

def is_korean(txt: str) -> bool:
    return bool(korean_re.search(txt))

def iter_language_files(folder: Path):
    for r, _, fs in os.walk(folder):
        for f in fs:
            lf = f.lower()
            if lf == "languagedata.xml":
                continue                         #  ←  DROP generic file
            if lf.startswith("languagedata_") and lf.endswith(".xml"):
                yield Path(r) / f

# ----------------------------------------------------------------------
#  BUILD EXPORT INDEX  (StringId → refined category)
# ----------------------------------------------------------------------
def build_export_index(export_folder: Path):
    """
    • Dialog & Sequencer keep 2-level (and sometimes 3-level) detail
    • System keeps 2-level so we can spot ItemGroup/Gimmick/MultiChange
    • Everything else → 1st folder name
    """
    index = {}
    for xml in export_folder.rglob("*.xml"):
        try:
            root = parse_xml_file(xml)
        except Exception:
            continue

        parts = xml.relative_to(export_folder).parts
        if not parts:
            continue

        p0 = parts[0].lower()

        # -------------------- Dialog  (Dialog/<sub>)
        if p0 == "dialog" and len(parts) > 1:
            cat = f"Dialog/{parts[1]}"

        # -------------------- Sequencer
        elif p0 == "sequencer" and len(parts) > 1:
            p1 = parts[1].lower()

            #  Sequencer/Other/QuestGroup_Xx
            if p1 == "other" and len(parts) > 2 and parts[2].lower().startswith("questgroup_"):
                cat = f"Sequencer/Other/{parts[2]}"

            #  Sequencer/Sequencer/cd_seq_memory_xx
            elif p1 == "sequencer" and len(parts) > 2 and Path(parts[2]).stem.lower().startswith("cd_seq_memory_"):
                cat = f"Sequencer/Sequencer/{Path(parts[2]).stem}"

            else:
                cat = f"Sequencer/{parts[1]}"

        # -------------------- System (System/<sub>)
        elif p0 == "system" and len(parts) > 1:
            cat = f"System/{parts[1]}"

        # -------------------- default → top folder
        else:
            cat = parts[0]

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                index[sid] = cat
    return index

# ----------------------------------------------------------------------
#  STYLE HELPERS  (Excel)
# ----------------------------------------------------------------------
def style_header(ws, row_idx, headers):
    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(bold=True, color="FFFFFF")
    thick = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    for col, txt in enumerate(headers, 1):
        c = ws.cell(row_idx, col, txt)
        c.fill, c.font, c.border = fill, font, thick
        ws.column_dimensions[c.column_letter].width = max(18, len(str(txt)) + 6)

def style_separator(ws, row_idx, num_cols):
    fill = PatternFill("solid", fgColor="FFFF00")
    for col in range(1, num_cols + 1):
        ws.cell(row_idx, col).fill = fill


# ----------------------------------------------------------------------
#  GUI  APPLICATION
# ----------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Translation Utilities – WordCount & StrOrigin")
        self.root.geometry("1010x780")

        # -------- buttons
        fbtn = tk.Frame(root); fbtn.pack(pady=10)
        tk.Button(fbtn, text="Word-Count Process", width=25,
                  command=self.start_word_count).pack(side=tk.LEFT, padx=10)
        tk.Button(fbtn, text="StrOrigin Comparison", width=25,
                  command=self.start_strorigin_comparison).pack(side=tk.LEFT, padx=10)

        # -------- progress + log
        self.progress = ttk.Progressbar(root, mode="determinate")
        self.progress.pack(fill=tk.X, padx=15, pady=5)

        tk.Label(root, text="Log / Output:").pack(anchor=tk.W, padx=15)
        self.out = scrolledtext.ScrolledText(root, height=32)
        self.out.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

    # -------------------------------------------------- tiny logger
    def log(self, *msg):
        self.out.insert(tk.END, " ".join(str(x) for x in msg) + "
")
        self.out.see(tk.END)
        self.root.update_idletasks()

    # ==============================================================
    #  1)  WORD-COUNT  PROCESS
    # ==============================================================
    def start_word_count(self):
        threading.Thread(target=self.run_word_count, daemon=True).start()

    def run_word_count(self):
        self.out.delete(1.0, tk.END)
        ts0 = datetime.now()
        self.log(f"[{ts0:%Y-%m-%d %H:%M:%S}] Word-count process started")
        self.log(f"LANGUAGE_FOLDER = {LANGUAGE_FOLDER}")
        self.log(f"EXPORT_FOLDER   = {EXPORT_FOLDER}
")
        self.progress['value'] = 0

        # ---------- build export StringId → category map
        self.log("Building export index …")
        export_index = build_export_index(EXPORT_FOLDER)
        self.log(f"Indexed {len(export_index):,} StringIds
")

        files = list(iter_language_files(LANGUAGE_FOLDER))
        if not files:
            messagebox.showerror("Word-count", "No LanguageData_*.xml files found.")
            return

        DIFF_OUTPUT_FOLDER.mkdir(exist_ok=True)

        per_file_rows   = []      # for Sheet-1
        category_stats  = {}      # lang → cat → totals
        global_missing  = []      # nodes missing from export

        total_files = len(files)

        for idx, xml_path in enumerate(files, 1):
            self.progress['value'] = idx / total_files * 100
            self.log(f"({idx}/{total_files}) {xml_path.name}")

            lang = xml_path.stem.split("_", 1)[1].upper()
            if lang == "KOR":                      # skip source Korean
                continue

            try:
                tree = parse_xml_file(xml_path)
            except Exception:
                continue

            total_words = completed_origin_words = translated_words = 0
            diff_nodes  = []

            for loc in tree.iter("LocStr"):
                origin = (loc.get("StrOrigin") or "").strip()
                if not origin:
                    continue
                wc_origin = count_words(origin)
                total_words += wc_origin

                sid = loc.get("StringId")
                cat = export_index.get(sid, "Unknown")

                # ---------- remap System/ItemGroup → Non-Priority, etc.
                if cat.startswith("System/") and cat.split("/", 1)[1] in NON_PRIORITY_FOLDERS:
                    cat = "Non-Priority"

                # ---------- register missing SIDs
                if sid not in export_index:
                    diff_nodes.append(loc)
                    global_missing.append(copy.deepcopy(loc))

                # ---------- update stats
                lang_map = category_stats.setdefault(lang, {})
                st = lang_map.setdefault(cat, {
                    "total_words": 0,
                    "completed_origin_words": 0,
                    "translated_words": 0,
                    "completed_nodes": 0,
                    "total_nodes": 0
                })
                st["total_words"]  += wc_origin
                st["total_nodes"]  += 1

                trans = (loc.get("Str") or "").strip()
                if trans and not is_korean(trans):
                    wc_trans = count_words(trans)
                    completed_origin_words += wc_origin
                    translated_words += wc_trans

                    st["completed_nodes"]         += 1
                    st["completed_origin_words"]  += wc_origin
                    st["translated_words"]        += wc_trans

            per_file_rows.append((lang, xml_path.name,
                                  total_words, completed_origin_words,
                                  translated_words))

            # ---------- write DIFF xml per file
            if diff_nodes:
                root_diff = LET.Element("ROOT")
                for d in diff_nodes:
                    root_diff.append(d)
                out_path = DIFF_OUTPUT_FOLDER / f"{xml_path.stem}_DIFF.xml"
                LET.ElementTree(root_diff).write(
                    str(out_path), encoding="utf-8", xml_declaration=True)
                self.log(f"   → diff written : {out_path}")

        # ---------- global missing xml
        if global_missing:
            root_all = LET.Element("ROOT")
            for n in global_missing:
                root_all.append(n)
            LET.ElementTree(root_all).write(str(MISSING_XML_FILE),
                                            encoding="utf-8",
                                            xml_declaration=True)
            self.log(f"
Global missing-from-export XML → {MISSING_XML_FILE}")

        # ===========================================================
        #  EXCEL  REPORT
        # ===========================================================
        self.progress['value'] = 95
        self.log("
Generating Excel summary …")
        wb = Workbook()

        # -------------------------------------------------- Sheet-1  (Per-File)
        ws1 = wb.active
        ws1.title = "Per-File"

        hdr1 = ["Language", "File",
                "Total Origin Words",
                "Completed Origin Words",
                "Translated Words",
                "Coverage %"]
        style_header(ws1, 1, hdr1)

        r = 2
        for lang, fn, tw, cow, tws in sorted(per_file_rows):
            cov = round(100 * cow / tw, 2) if tw else 0
            ws1.cell(r, 1, lang)
            ws1.cell(r, 2, fn)
            ws1.cell(r, 3, tw)
            ws1.cell(r, 4, cow)
            ws1.cell(r, 5, tws)
            ws1.cell(r, 6, cov)
            r += 1

        # autosize
        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = max(
                18, max(len(str(c.value or "")) for c in col) + 4)

        # -------------------------------------------------- Sheet-2  (Detailed)
        ws2 = wb.create_sheet("Detailed Summary")

        hdr2 = ["Category",
                "Total Origin Words",
                "Completed Origin Words",
                "Translated Words",
                "Coverage %"]
        row_idx = 1

        # ------------- helper to aggregate & write a row
        def agg_add(a, b):
            for k in ("total_words", "completed_origin_words", "translated_words"):
                a[k] += b[k]

        def big_category(title: str) -> str:
            """
            Extracts the top-level category string to decide the colour.
            """
            if title.startswith("Dialog"):
                return "Dialog"
            if title.startswith("Sequencer"):
                return "Sequencer"
            if title.startswith("System"):
                return "System"
            if title.startswith("World"):
                return "World"
            if title.startswith("Platform"):
                return "Platform"
            if title.startswith("None"):
                return "None"
            if title.startswith("Non-Priority"):
                return "Non-Priority"
            return ""

        def write_row(title, stats):
            nonlocal row_idx
            tw  = stats["total_words"]
            cow = stats["completed_origin_words"]
            tws = stats["translated_words"]
            cov = round(100 * cow / tw, 2) if tw else 0
            for col, val in enumerate([title, tw, cow, tws, cov], 1):
                cell = ws2.cell(row_idx, col, val)
                cat = big_category(title)
                if cat in BIG_CAT_FILLS:
                    cell.fill = BIG_CAT_FILLS[cat]
            row_idx += 1

        # ------------- iterate each language
        for lang in sorted(category_stats):
            ws2.cell(row_idx, 1, f"Language: {lang}").font = Font(bold=True, size=12)
            row_idx += 1
            style_header(ws2, row_idx, hdr2)
            row_idx += 1

            lang_stats = category_stats[lang]

            # ------------------------------------------------------------------
            #  A)  Dialog  (Grand + subs)
            dialog_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("Dialog/"):
                    agg_add(dialog_total, st)
            if dialog_total["total_words"]:
                write_row("Dialog (Grand Total)", dialog_total)
                for sub in DIALOG_SUBS:
                    key = f"Dialog/{sub}"
                    if key in lang_stats:
                        write_row(key, lang_stats[key])

            # ------------------------------------------------------------------
            #  B)  Sequencer
            seq_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("Sequencer/"):
                    agg_add(seq_total, st)
            if seq_total["total_words"]:
                write_row("Sequencer (Grand Total)", seq_total)

                # ----------- top-level subs
                for top in SEQUENCER_TOPS:
                    pref = f"Sequencer/{top}"
                    sub_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
                    for cat, st in lang_stats.items():
                        if cat == pref or cat.startswith(pref + "/"):
                            agg_add(sub_total, st)
                    if not sub_total["total_words"]:
                        continue
                    title = pref if top not in {"Other", "Sequencer"} else f"{pref} (Total)"
                    write_row(title, sub_total)

                    # ---- NEW: Consolidated QuestGroup_*  (Sequencer/Other/QuestGroup_XX)
                    if top == "Other":
                        quest_pref = "Sequencer/Other/QuestGroup_"
                        quest_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
                        for cat, st in lang_stats.items():
                            if cat.startswith(quest_pref):
                                agg_add(quest_total, st)
                        if quest_total["total_words"]:
                            write_row("Sequencer/Other/QuestGroup_XX", quest_total)

                    # ---- NEW: Consolidated cd_seq_memory_*  (Sequencer/Sequencer/cd_seq_memory_XX)
                    if top == "Sequencer":
                        mem_pref = "Sequencer/Sequencer/cd_seq_memory_"
                        mem_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
                        for cat, st in lang_stats.items():
                            if cat.startswith(mem_pref):
                                agg_add(mem_total, st)
                        if mem_total["total_words"]:
                            write_row("Sequencer/Sequencer/cd_seq_memory_XX", mem_total)

            # ------------------------------------------------------------------
            #  C)  System  (cleaned)
            sys_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("System/") and cat.split("/",1)[1] not in NON_PRIORITY_FOLDERS:
                    agg_add(sys_total, st)
            if sys_total["total_words"]:
                write_row("System", sys_total)

            # ------------------------------------------------------------------
            #  D)  World
            if "World" in lang_stats:
                write_row("World", lang_stats["World"])

            # ------------------------------------------------------------------
            #  E)  Platform
            if "Platform" in lang_stats:
                write_row("Platform", lang_stats["Platform"])

            # ------------------------------------------------------------------
            #  F)  None  (merge ‘None’ + ‘Unknown’)
            none_total = {"total_words":0,"completed_origin_words":0,"translated_words":0}
            if "None" in lang_stats:
                agg_add(none_total, lang_stats["None"])
            if "Unknown" in lang_stats:
                agg_add(none_total, lang_stats["Unknown"])
            if none_total["total_words"]:
                write_row("None", none_total)

            # ------------------------------------------------------------------
            #  G)  Non-Priority
            if "Non-Priority" in lang_stats:
                write_row("Non-Priority", lang_stats["Non-Priority"])

            # --- visual separator
            style_separator(ws2, row_idx, len(hdr2))
            row_idx += 2

        # autosize
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = max(
                18, max(len(str(c.value or "")) for c in col) + 4)

        # -------------------------------------------------- save workbook
        ts_name = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xlsx = Path.cwd() / f"Translation_Coverage_SUMMARY_{ts_name}.xlsx"
        wb.save(out_xlsx)
        self.log(f"
Excel saved → {out_xlsx}")
        self.progress['value'] = 100
        messagebox.showinfo("Word-count", f"Done!
Report:
{out_xlsx}")

    # ==============================================================
    #  2)  STRORIGIN  COMPARISON  (unchanged – kept from original)
    # ==============================================================

    def start_strorigin_comparison(self):
        before = BEFORE_FOLDER
        after  = AFTER_FOLDER

        if not before.is_dir() or not after.is_dir():
            messagebox.showerror(
                "StrOrigin",
                "Required folders not found:

"
                f"BEFORE : {before}
"
                f"AFTER  : {after}

"
                "Both folders must sit next to the executable."
            )
            return

        common = sorted({d.name for d in before.iterdir() if d.is_dir()} &
                        {d.name for d in after.iterdir()  if d.is_dir()})

        if not common:
            messagebox.showerror("StrOrigin",
                                 "No matching top-level sub-folders found "
                                 "under BEFORE / AFTER.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("StrOrigin Comparison – options")
        dlg.grab_set()

        tk.Label(dlg,
                 text=("Select the sub-folders to compare.
"
                       "Tick “Granular” to compare their sub-folders individually:")
                 ).pack(padx=10, pady=8)

        rows = []
        for name in common:
            fr = tk.Frame(dlg); fr.pack(fill=tk.X, padx=15, pady=2)
            sel  = tk.BooleanVar(value=True)
            gran = tk.BooleanVar(value=False)
            tk.Checkbutton(fr, text=name, variable=sel)\
                .pack(side=tk.LEFT)
            tk.Label(fr, text="Granular").pack(side=tk.LEFT, padx=(25,0))
            tk.Checkbutton(fr, variable=gran).pack(side=tk.LEFT)
            rows.append((name, sel, gran))

        ret = {"sel": None}
        def on_ok():
            picked = [(n, g.get()) for n, s, g in rows if s.get()]
            if not picked:
                messagebox.showerror("StrOrigin", "Select at least one folder.")
                return
            ret["sel"] = picked
            dlg.destroy()
        tk.Button(dlg, text="OK",     width=10, command=on_ok)\
            .pack(side=tk.LEFT, padx=15, pady=14)
        tk.Button(dlg, text="Cancel", width=10, command=dlg.destroy)\
            .pack(side=tk.RIGHT, padx=15, pady=14)

        self.root.wait_window(dlg)
        if not ret["sel"]:
            return

        todo = [(before / n, after / n, g) for n, g in ret["sel"]]
        threading.Thread(target=self.run_strorigin_report,
                         args=(todo,), daemon=True).start()

    # -------------------------------------------------- helpers (StrOrigin)
    def run_strorigin_report(self, selections):
        #  [identical to original implementation – colour tweaks not required]
        self.out.delete(1.0, tk.END)
        self.progress['value'] = 0
        self.log(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] StrOrigin comparison started")
        total_sel = len(selections)

        global_change = {}
        global_add    = {}
        summary_rows  = []

        files_added_list    = []
        files_modified_list = []

        for idx, (bdir, adir, gran) in enumerate(selections, 1):
            self.progress['value'] = idx / total_sel * 100
            label = f"{bdir.name}"
            self.log(f"
[{idx}/{total_sel}] {label}")

            if not gran:
                stats, chg, add, added, modified = self.compare_pair_collect_files(
                    bdir, adir, label, global_change, global_add)
                summary_rows.append(stats)
                files_added_list.extend(added)
                files_modified_list.extend(modified)
            else:
                subs_b = {d.name for d in bdir.iterdir() if d.is_dir()}
                subs_a = {d.name for d in adir.iterdir() if d.is_dir()}
                common = sorted(subs_b & subs_a)
                if not common:
                    self.log("  (no common sub-folders)")
                for sub in common:
                    self.log(f"  · {sub}")
                    stats, chg, add, added, modified = self.compare_pair_collect_files(
                        bdir / sub, adir / sub, f"{label}/{sub}",
                        global_change, global_add)
                    summary_rows.append(stats)
                    files_added_list.extend(added)
                    files_modified_list.extend(modified)

        # -------- Excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        hdr = ["Folder", "Files Compared", "Files Modified", "Files Added",
               "Files Deleted", "StrOrigin Changes (Rows)", "Changed Words",
               "StrOrigin Additions (Rows)", "Added Words", "Grand Total Words"]
        header_fill = PatternFill("solid", fgColor="A7C7E7")
        medium = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        for c, h in enumerate(hdr, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = medium

        thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        palette = ["F8CBAD", "C6E0B4", "FFD966",
                   "D9D2E9", "F4B183", "A9D08E", "FFE699"]
        parent_colors, idx_pal = {}, 0

        r = 2
        for row in summary_rows:
            parent = row[0].split("/",1)[0]
            if parent not in parent_colors:
                parent_colors[parent] = palette[idx_pal % len(palette)]
                idx_pal += 1
            fill = PatternFill("solid", fgColor=parent_colors[parent])
            for c, v in enumerate(row, 1):
                cell = ws.cell(r, c, v)
                cell.fill = fill
                cell.border = thin
            r += 1
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                15, max(len(str(c.value or "")) for c in col) + 4)

        def make_top(name, data):
            w = wb.create_sheet(name)
            w.cell(1,1,"File"); w.cell(1,2,"Count")
            w.cell(1,1).font = w.cell(1,2).font = Font(bold=True)
            for i,(f,cnt) in enumerate(sorted(data.items(),
                                              key=lambda x:x[1],
                                              reverse=True)[:50],2):
                w.cell(i,1,f); w.cell(i,2,cnt)
            for col in w.columns:
                w.column_dimensions[col[0].column_letter].width = max(
                    15, max(len(str(c.value or "")) for c in col)+4)
        make_top("Top-50 Changes", global_change)
        make_top("Top-50 Additions", global_add)

        ws_add = wb.create_sheet("Files Added")
        ws_add.cell(1,1,"File Path").font = Font(bold=True)
        for i,f in enumerate(sorted(files_added_list),2):
            ws_add.cell(i,1,f)
        ws_add.column_dimensions['A'].width = max(
            25, max((len(f) for f in files_added_list), default=12)+4)

        ws_mod = wb.create_sheet("Files Modified")
        ws_mod.cell(1,1,"File Path").font = Font(bold=True)
        for i,f in enumerate(sorted(files_modified_list),2):
            ws_mod.cell(i,1,f)
        ws_mod.column_dimensions['A'].width = max(
            25, max((len(f) for f in files_modified_list), default=12)+4)

        out = Path.cwd() / f"StrOrigin_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(out)
        self.log(f"
Excel saved → {out}")
        self.progress['value'] = 100
        messagebox.showinfo("StrOrigin", f"Done!
Report:
{out}")

    # --------------------------------------------------
    #  UNCHANGED compare_pair_collect_files
    def compare_pair_collect_files(self, before_dir: Path, after_dir: Path,
                                   label: str,
                                   global_ch: dict,
                                   global_add: dict):
        def collect(folder: Path):
            mapping = {}
            for r, _, fs in os.walk(folder):
                for f in fs:
                    if not f.lower().endswith(".xml"):
                        continue
                    fp  = Path(r) / f
                    rel = fp.relative_to(folder).as_posix()
                    try:
                        rt = ET.parse(fp).getroot()
                        mp = {}
                        for loc in rt.findall(".//LocStr"):
                            sid = loc.get("StringId")
                            so  = loc.get("StrOrigin","")
                            if sid:
                                mp[sid] = so
                        if mp:
                            mapping[rel] = mp
                    except Exception:
                        pass
            return mapping

        bmap = collect(before_dir)
        amap = collect(after_dir)

        bf, af = set(bmap), set(amap)
        common = bf & af

        files_added   = list(af - bf)
        files_deleted = len(bf - af)
        files_modified = 0

        tot_change_rows = tot_change_words = 0
        tot_add_rows    = tot_add_words    = 0

        change_per_file = {}
        add_per_file    = {}
        added_files_list, modified_files_list = [], []

        for rel in common:
            s = bmap[rel]; t = amap[rel]
            changes = [sid for sid in s if sid in t and s[sid] != t[sid]]
            adds    = [sid for sid in t if sid not in s]

            if changes:
                files_modified += 1
                change_per_file[rel] = len(changes)
                tot_change_rows  += len(changes)
                tot_change_words += sum(count_words(t[sid]) for sid in changes)
                modified_files_list.append(f"{label}/{rel}")

            if adds:
                add_per_file[rel] = len(adds)
                tot_add_rows  += len(adds)
                tot_add_words += sum(count_words(t[sid]) for sid in adds)

        for rel in af - bf:
            rows = len(amap[rel])
            if rows:
                add_per_file[rel] = rows
                tot_add_rows  += rows
                tot_add_words += sum(count_words(v) for v in amap[rel].values())
                added_files_list.append(f"{label}/{rel}")

        # ---- global dicts (for Top-50)
        for f, c in change_per_file.items():
            global_ch[f"{label}/{f}"] = global_ch.get(f"{label}/{f}", 0) + c
        for f, c in add_per_file.items():
            global_add[f"{label}/{f}"] = global_add.get(f"{label}/{f}", 0) + c

        self.log(f"    Files compared : {len(common)}")
        self.log(f"    Files modified : {files_modified}")
        self.log(f"    Files added    : {len(files_added)}")
        self.log(f"    Files deleted  : {files_deleted}")
        self.log(f"    StrOrigin changes   : {tot_change_rows} rows / {tot_change_words} words")
        self.log(f"    StrOrigin additions : {tot_add_rows} rows / {tot_add_words} words")

        stats_row = [label, len(common), files_modified, len(files_added),
                     files_deleted, tot_change_rows, tot_change_words,
                     tot_add_rows, tot_add_words,
                     tot_change_words + tot_add_words]
        return stats_row, change_per_file, add_per_file, added_files_list, modified_files_list

# ----------------------------------------------------------------------
def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()