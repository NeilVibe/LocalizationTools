"""
Faction Generator for DataListGenerator
========================================
Extracts faction names from factioninfo XML files.
"""

from pathlib import Path
from typing import List, Set

from openpyxl.styles import PatternFill

from .base import BaseGenerator, DataEntry
from ..utils.xml_parser import iter_xml_files, parse_xml_file


class FactionGenerator(BaseGenerator):
    """Generator for faction names from factioninfo XML files."""

    name = "Faction"
    output_filename = "FactionList.xlsx"
    sheet_name = "Faction List"
    headers = ["Faction Name", "Type", "Source File"]

    def __init__(self, xml_folder: Path):
        super().__init__(xml_folder)
        self.type_colors = {
            "FactionGroup": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "Faction": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "FactionNode": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        }

    def _extract_faction_nodes_recursive(
        self,
        elem,
        source_file: str,
        entries: List[DataEntry],
        seen_names: Set[str]
    ) -> None:
        """Recursively extract FactionNode names from an element and its children."""
        name = elem.get("Name") or ""

        if name and name not in seen_names:
            entries.append(DataEntry(
                name=name,
                entry_type="FactionNode",
                source_file=source_file
            ))
            seen_names.add(name)

        for child in elem:
            if child.tag == "FactionNode":
                self._extract_faction_nodes_recursive(child, source_file, entries, seen_names)

    def extract(self) -> List[DataEntry]:
        """Extract faction names from all XML files in the folder."""
        entries: List[DataEntry] = []
        seen_names: Set[str] = set()

        xml_files = iter_xml_files(self.xml_folder)

        if not xml_files:
            print(f"  No XML files found in {self.xml_folder}")
            return entries

        print(f"  Found {len(xml_files)} XML files to parse...")

        for path in xml_files:
            root = parse_xml_file(path)
            if root is None:
                continue

            source_file = path.name

            # Parse FactionGroups
            for fg_elem in root.iter("FactionGroup"):
                group_name = fg_elem.get("GroupName") or ""

                if group_name and group_name not in seen_names:
                    entries.append(DataEntry(
                        name=group_name,
                        entry_type="FactionGroup",
                        source_file=source_file
                    ))
                    seen_names.add(group_name)

                for f_elem in fg_elem.iter("Faction"):
                    faction_name = f_elem.get("Name") or ""

                    if faction_name and faction_name not in seen_names:
                        entries.append(DataEntry(
                            name=faction_name,
                            entry_type="Faction",
                            source_file=source_file
                        ))
                        seen_names.add(faction_name)

                    for node_elem in f_elem:
                        if node_elem.tag == "FactionNode":
                            self._extract_faction_nodes_recursive(
                                node_elem, source_file, entries, seen_names
                            )

            # Parse standalone Factions (not inside FactionGroup)
            for f_elem in root:
                if f_elem.tag == "Faction":
                    faction_name = f_elem.get("Name") or ""

                    if faction_name and faction_name not in seen_names:
                        entries.append(DataEntry(
                            name=faction_name,
                            entry_type="Faction",
                            source_file=source_file
                        ))
                        seen_names.add(faction_name)

                    for node_elem in f_elem:
                        if node_elem.tag == "FactionNode":
                            self._extract_faction_nodes_recursive(
                                node_elem, source_file, entries, seen_names
                            )

        return entries
