"""
GUI Application for DataListGenerator
======================================
Main GUI window with generator selection and translation features.
Styled similar to QACompiler Suite GUI.
"""

from pathlib import Path
from typing import List, Tuple, Dict, Optional
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import (
    FACTIONINFO_FOLDER, SKILLINFO_FOLDER, LOC_FOLDER,
    OUTPUT_FOLDER, TRANSLATIONS_FOLDER
)
from generators import FactionGenerator, SkillGenerator, DataEntry
from utils.excel_writer import (
    THIN_BORDER, HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT,
    YELLOW_FILL, BOLD_FONT, auto_fit_columns
)
from translation_utils import (
    discover_language_files,
    load_language_table
)


# =============================================================================
# GUI CONFIGURATION
# =============================================================================

WINDOW_TITLE = "Data List Generator v3.0"
WINDOW_WIDTH = 800
WINDOW_HEIGHT = 750

# Available generators (add new ones here)
GENERATORS = {
    "Faction": {
        "class": FactionGenerator,
        "folder": FACTIONINFO_FOLDER,
        "desc": "FactionGroup, Faction, FactionNode"
    },
    "Skill": {
        "class": SkillGenerator,
        "folder": SKILLINFO_FOLDER,
        "desc": "SkillInfo/@SkillName"
    },
}


class DataToolGUI:
    """Main GUI for Data List Generator Tool Suite."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title(WINDOW_TITLE)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.resizable(True, True)

        # Generator checkbox variables
        self.generator_vars: Dict[str, tk.BooleanVar] = {}

        # Filter option
        self.var_filter_glossary = tk.BooleanVar(value=False)
        self.filter_file_path: Optional[Path] = None

        self._create_widgets()

    def _create_widgets(self):
        """Build the GUI."""
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title = tk.Label(
            main_frame,
            text="Data List Generator",
            font=("Arial", 18, "bold")
        )
        title.pack(pady=(0, 15))

        # =================================================================
        # Section 1: Generate Data Lists
        # =================================================================
        section1_frame = ttk.LabelFrame(
            main_frame,
            text="1. Generate Data Lists",
            padding=10
        )
        section1_frame.pack(fill=tk.X, pady=5)

        # Description
        ttk.Label(
            section1_frame,
            text="Select which data lists to generate from XML files:"
        ).pack(anchor=tk.W, pady=(0, 5))

        # Checkboxes in a grid (2 columns)
        checkbox_frame = ttk.Frame(section1_frame)
        checkbox_frame.pack(fill=tk.X, pady=5)

        for i, (name, info) in enumerate(GENERATORS.items()):
            var = tk.BooleanVar(value=True)
            self.generator_vars[name] = var

            cb_text = f"{name} ({info['desc']})"
            cb = ttk.Checkbutton(checkbox_frame, text=cb_text, variable=var)
            cb.grid(row=i // 2, column=i % 2, sticky="w", padx=10, pady=2)

        # Select All / Deselect All / Generate buttons
        btn_frame = ttk.Frame(section1_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            btn_frame,
            text="Select All",
            command=self._select_all,
            width=12
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            btn_frame,
            text="Deselect All",
            command=self._deselect_all,
            width=12
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            btn_frame,
            text="Generate Selected",
            command=self._do_generate_lists,
            width=20
        ).pack(side=tk.RIGHT, padx=5)

        # =================================================================
        # Section 2: Filter Glossary (Optional)
        # =================================================================
        section2_frame = ttk.LabelFrame(
            main_frame,
            text="2. Filter Glossary (Optional)",
            padding=10
        )
        section2_frame.pack(fill=tk.X, pady=5)

        ttk.Label(
            section2_frame,
            text="Remove generated list items from another Excel glossary file.\n"
                 "Uses exact match on Column A of the selected file."
        ).pack(anchor=tk.W, pady=(0, 5))

        filter_row = ttk.Frame(section2_frame)
        filter_row.pack(fill=tk.X, pady=5)

        ttk.Checkbutton(
            filter_row,
            text="Enable filtering",
            variable=self.var_filter_glossary
        ).pack(side=tk.LEFT, padx=5)

        self.filter_file_label = ttk.Label(filter_row, text="No file selected")
        self.filter_file_label.pack(side=tk.LEFT, padx=10)

        ttk.Button(
            filter_row,
            text="Select File...",
            command=self._select_filter_file,
            width=15
        ).pack(side=tk.RIGHT, padx=5)

        # =================================================================
        # Section 3: List Concatenator + Translator
        # =================================================================
        section3_frame = ttk.LabelFrame(
            main_frame,
            text="3. List Concatenator + Translator",
            padding=10
        )
        section3_frame.pack(fill=tk.X, pady=5)

        ttk.Label(
            section3_frame,
            text="Combine Excel lists from a folder (reads Column A).\n"
                 "Translate to all available LOC languages with StrOrigin matching."
        ).pack(anchor=tk.W, pady=(0, 5))

        ttk.Button(
            section3_frame,
            text="Concatenate & Translate",
            command=self._do_concat_translate,
            width=30
        ).pack(pady=10, ipady=8)

        # =================================================================
        # Progress and Status
        # =================================================================
        self.progress_var = tk.DoubleVar()
        self.progress_label = ttk.Label(main_frame, text="")
        self.progress_label.pack(pady=(15, 0))

        self.progress_bar = ttk.Progressbar(
            main_frame,
            variable=self.progress_var,
            maximum=100
        )
        self.progress_bar.pack(fill=tk.X, pady=5)

        # Status log
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="5")
        status_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.status_text = tk.Text(status_frame, height=12, width=80, state=tk.DISABLED)
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=scrollbar.set)

        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Paths info
        paths_frame = ttk.LabelFrame(main_frame, text="Configured Paths", padding="5")
        paths_frame.pack(fill=tk.X)

        paths_text = "\n".join([
            f"{name}: {info['folder']}"
            for name, info in GENERATORS.items()
        ])
        paths_text += f"\nLOC: {LOC_FOLDER}"

        ttk.Label(
            paths_frame,
            text=paths_text,
            font=("Courier", 8)
        ).pack(anchor=tk.W)

    # =========================================================================
    # Helper Methods
    # =========================================================================

    def _select_all(self):
        """Select all generator checkboxes."""
        for var in self.generator_vars.values():
            var.set(True)

    def _deselect_all(self):
        """Deselect all generator checkboxes."""
        for var in self.generator_vars.values():
            var.set(False)

    def _get_selected_generators(self) -> List[str]:
        """Get list of selected generator names."""
        return [name for name, var in self.generator_vars.items() if var.get()]

    def _select_filter_file(self):
        """Open file dialog to select glossary file to filter."""
        file_path = filedialog.askopenfilename(
            title="Select Excel Glossary to Filter",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.filter_file_path = Path(file_path)
            self.filter_file_label.config(text=self.filter_file_path.name)
            self.var_filter_glossary.set(True)

    def _log(self, message: str):
        """Add message to status log."""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _update_progress(self, current: int, total: int, message: str):
        """Update progress bar and label."""
        percent = (current / total * 100) if total > 0 else 0
        self.progress_var.set(percent)
        self.progress_label.config(text=message)
        self.root.update_idletasks()

    def _clear_status(self):
        """Clear status log."""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.progress_var.set(0)
        self.progress_label.config(text="")

    # =========================================================================
    # Section 1: Generate Data Lists
    # =========================================================================

    def _do_generate_lists(self):
        """Execute Section 1: Generate selected data lists."""
        self._clear_status()
        self._log("=" * 50)
        self._log("DATA LIST GENERATOR")
        self._log("=" * 50)

        selected = self._get_selected_generators()

        if not selected:
            self._log("ERROR: No generators selected!")
            messagebox.showwarning("Warning", "Please select at least one data type to generate.")
            return

        self._log(f"Selected: {', '.join(selected)}")

        # Ensure output folder exists
        OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

        all_entries: List[DataEntry] = []
        generators_run = []
        total_steps = len(selected) + (1 if self.var_filter_glossary.get() else 0)
        current_step = 0

        # Run each selected generator
        for name in selected:
            current_step += 1
            self._update_progress(current_step, total_steps + 1, f"Generating {name} List...")

            info = GENERATORS[name]
            folder = info["folder"]

            if not folder.exists():
                self._log(f"\nWARNING: {name} folder not found: {folder}")
                continue

            generator = info["class"](folder)
            entries = generator.run(OUTPUT_FOLDER)
            all_entries.extend(entries)
            generators_run.append((name, len(entries)))
            self._log(f"  Generated {generator.output_filename} with {len(entries)} entries")

        # Optional: Filter glossary file
        if self.var_filter_glossary.get() and self.filter_file_path and all_entries:
            current_step += 1
            self._update_progress(current_step, total_steps + 1, "Filtering glossary...")
            self._do_filter_glossary(all_entries)

        # Complete
        self._update_progress(100, 100, "Complete!")
        self._log("\n" + "=" * 50)
        self._log("COMPLETE!")
        self._log("=" * 50)

        # Show completion message
        if generators_run:
            message = "Data List Generator Complete!\n\n"
            for name, count in generators_run:
                message += f"{name}: {count} entries\n"
            message += f"\nOutput folder: {OUTPUT_FOLDER}"
            messagebox.showinfo("Complete", message)
        else:
            messagebox.showwarning("Warning", "No data could be generated. Check folder paths.")

    def _do_filter_glossary(self, entries: List[DataEntry]):
        """Filter the selected glossary file by removing entries that match generated lists."""
        if not self.filter_file_path or not self.filter_file_path.exists():
            self._log("  Filter file not found, skipping...")
            return

        self._log(f"\nFiltering {self.filter_file_path.name}...")

        # Build set of names to filter
        filter_names = {entry.name for entry in entries}

        try:
            wb = load_workbook(self.filter_file_path)

            total_removed = 0
            removed_details: List[Tuple[int, str]] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_to_delete = []

                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value is None:
                        continue

                    col_a_value = str(cell_value).strip()

                    # Exact match on Column A
                    if col_a_value in filter_names:
                        rows_to_delete.append((row_idx, col_a_value))

                # Delete rows in reverse order to maintain row indices
                for row_idx, matched in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)
                    removed_details.append((row_idx, matched))
                    total_removed += 1

            # Save filtered file
            output_name = f"Filtered_{self.filter_file_path.name}"
            filtered_path = OUTPUT_FOLDER / output_name
            wb.save(filtered_path)

            self._log(f"  Removed {total_removed} rows")
            self._log(f"  Saved: {filtered_path.name}")

            # Generate report if rows were removed
            if removed_details:
                report_path = OUTPUT_FOLDER / f"FilterReport_{self.filter_file_path.stem}.txt"
                with open(report_path, 'w', encoding='utf-8') as f:
                    f.write(f"Filter Report\n")
                    f.write(f"Input: {self.filter_file_path.name}\n")
                    f.write(f"Total rows removed: {total_removed}\n\n")
                    f.write("Removed rows (row number, matched value):\n")
                    for row_num, value in removed_details[:100]:
                        f.write(f"  Row {row_num}: {value}\n")
                    if len(removed_details) > 100:
                        f.write(f"  ... and {len(removed_details) - 100} more\n")
                self._log(f"  Report: {report_path.name}")

        except Exception as e:
            self._log(f"  ERROR filtering: {e}")

    # =========================================================================
    # Section 3: List Concatenator + Translator
    # =========================================================================

    def _do_concat_translate(self):
        """Execute Section 3: Concatenate lists and translate to all languages."""
        self._clear_status()
        self._log("=" * 50)
        self._log("LIST CONCATENATOR + TRANSLATOR")
        self._log("=" * 50)

        # Select folder with Excel files
        folder_path = filedialog.askdirectory(
            title="Select folder containing Excel files"
        )

        if not folder_path:
            self._log("Cancelled - no folder selected")
            return

        folder_path = Path(folder_path)
        self._log(f"Selected folder: {folder_path}")

        # Check for Excel files
        excel_files = list(folder_path.glob("*.xlsx"))
        excel_files = [f for f in excel_files if not f.name.startswith("~$")]

        if not excel_files:
            self._log("ERROR: No Excel files found in folder!")
            messagebox.showerror("Error", "No Excel files found in the selected folder!")
            return

        self._log(f"Found {len(excel_files)} Excel files")

        # Step 1: Concatenate lists
        self._log("\nStep 1: Concatenating Excel lists...")
        self._update_progress(10, 100, "Concatenating lists...")

        items_with_category, sections = self._concatenate_excel_lists(folder_path)

        if not items_with_category:
            self._log("ERROR: No items found in Excel files!")
            messagebox.showerror("Error", "No items found in the Excel files!")
            return

        self._log(f"  Total items: {len(items_with_category)}")
        self._log(f"  Unique items: {len(set(item for item, _ in items_with_category))}")
        for filename, items in sections:
            self._log(f"    [{filename}]: {len(items)} items")

        # Step 2: Load language files
        self._log("\nStep 2: Loading language data...")
        self._update_progress(20, 100, "Loading language data...")

        if not LOC_FOLDER.exists():
            self._log(f"ERROR: LOC folder not found: {LOC_FOLDER}")
            messagebox.showerror(
                "Error",
                f"LOC folder not found!\n\n{LOC_FOLDER}\n\nCheck settings.json"
            )
            return

        lang_files = discover_language_files(LOC_FOLDER)

        if not lang_files:
            self._log("ERROR: No language files found!")
            messagebox.showerror("Error", "No language files found in LOC folder!")
            return

        self._log(f"Found {len(lang_files)} language files")

        # Step 3: Generate translations
        self._log("\nStep 3: Generating translations...")

        TRANSLATIONS_FOLDER.mkdir(parents=True, exist_ok=True)

        # Load English table first (for European languages)
        eng_table = None
        if "eng" in lang_files:
            self._log("  Loading English table for European languages...")
            eng_table = load_language_table(lang_files["eng"])

        total_langs = len(lang_files)
        translated_langs = []
        total_translated = 0
        total_missing = 0

        for idx, (lang_code, xml_path) in enumerate(lang_files.items()):
            progress_pct = 20 + int(75 * (idx + 1) / total_langs)
            self._update_progress(progress_pct, 100, f"Translating to {lang_code.upper()}...")

            lang_table = load_language_table(xml_path)

            if not lang_table:
                self._log(f"  Skipping {lang_code.upper()} - no data loaded")
                continue

            output_path = TRANSLATIONS_FOLDER / f"Translated_{lang_code.upper()}.xlsx"
            stats = self._generate_translation_excel(
                items_with_category, lang_code, lang_table, output_path, eng_table
            )

            translated_langs.append(lang_code.upper())
            total_translated += stats["translated"]
            total_missing += stats["missing"]

            self._log(f"  {lang_code.upper()}: {stats['translated']} translated, {stats['missing']} missing")

        # Complete
        self._update_progress(100, 100, "Complete!")
        self._log("\n" + "=" * 50)
        self._log("COMPLETE!")
        self._log("=" * 50)
        self._log(f"Generated {len(translated_langs)} translation files")
        self._log(f"Output folder: {TRANSLATIONS_FOLDER}")

        # Show completion message
        unique_count = len(set(item for item, _ in items_with_category))
        avg_translated = total_translated / len(translated_langs) if translated_langs else 0
        avg_missing = total_missing / len(translated_langs) if translated_langs else 0

        message = f"List Concatenator + Translator Complete!\n\n"
        message += f"Source files: {len(sections)}\n"
        message += f"Total items: {len(items_with_category)}\n"
        message += f"Unique items: {unique_count}\n\n"
        message += f"Languages translated: {len(translated_langs)}\n"
        message += f"Avg. translated per lang: {int(avg_translated)}\n"
        message += f"Avg. missing per lang: {int(avg_missing)}\n\n"
        message += f"Output: {TRANSLATIONS_FOLDER}"

        messagebox.showinfo("Complete", message)

    def _read_first_column_from_excel(self, file_path: Path) -> List[str]:
        """Read first column from an Excel file, skipping header row."""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            values = []

            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value is not None:
                    values.append(str(cell_value).strip())

            wb.close()
            return values
        except Exception as e:
            print(f"  Error reading {file_path.name}: {e}")
            return []

    def _concatenate_excel_lists(
        self,
        folder_path: Path
    ) -> Tuple[List[Tuple[str, str]], List[Tuple[str, List[str]]]]:
        """Read all Excel files in folder and concatenate their first columns."""
        excel_files = sorted(folder_path.glob("*.xlsx"))

        if not excel_files:
            return [], []

        wb = Workbook()
        ws = wb.active
        ws.title = "Combined List"

        headers = ["Item", "Category"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        items_with_category: List[Tuple[str, str]] = []
        sections: List[Tuple[str, List[str]]] = []
        current_row = 2

        for excel_file in excel_files:
            if excel_file.name.startswith("~$"):
                continue

            file_name = excel_file.stem
            items = self._read_first_column_from_excel(excel_file)

            if not items:
                continue

            sections.append((file_name, items))

            # Section header (yellow)
            header_cell = ws.cell(row=current_row, column=1, value=f"[{file_name}]")
            header_cell.fill = YELLOW_FILL
            header_cell.font = BOLD_FONT
            header_cell.border = THIN_BORDER
            ws.cell(row=current_row, column=2, value="").border = THIN_BORDER
            current_row += 1

            # Items
            for item in items:
                ws.cell(row=current_row, column=1, value=item).border = THIN_BORDER
                ws.cell(row=current_row, column=2, value=file_name).border = THIN_BORDER
                items_with_category.append((item, file_name))
                current_row += 1

        auto_fit_columns(ws, max_width=60)

        combined_path = OUTPUT_FOLDER / "CombinedList.xlsx"
        combined_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(combined_path)
        self._log(f"  Saved: {combined_path}")

        return items_with_category, sections

    def _generate_translation_excel(
        self,
        items_with_category: List[Tuple[str, str]],
        lang_code: str,
        lang_table: Dict[str, str],
        output_path: Path,
        eng_table: Optional[Dict[str, str]] = None
    ) -> Dict[str, int]:
        """Generate a translation Excel file for a single language."""
        ASIAN_LANGUAGES = {"zho-cn", "zho-tw", "jpn", "kor"}
        include_english = lang_code.lower() not in ASIAN_LANGUAGES and eng_table is not None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Translation_{lang_code.upper()}"

        if include_english:
            headers = ["SourceText (Korean)", "English", f"Translation ({lang_code.upper()})", "Category"]
        else:
            headers = ["SourceText (Korean)", f"Translation ({lang_code.upper()})", "Category"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        translated_count = 0
        missing_count = 0

        # Deduplicate
        seen = set()
        unique_items = []
        for item, category in items_with_category:
            if item and item not in seen:
                seen.add(item)
                unique_items.append((item, category))

        for row_idx, (korean, category) in enumerate(unique_items, start=2):
            col_offset = 0

            ws.cell(row=row_idx, column=1, value=korean).border = THIN_BORDER

            if include_english:
                eng_translation = eng_table.get(korean, "")
                eng_is_untranslated = (
                    not eng_translation or
                    eng_translation == korean or
                    eng_translation.strip() == korean.strip()
                )
                if eng_is_untranslated:
                    eng_cell = ws.cell(row=row_idx, column=2, value="NO_TRANSLATION")
                    eng_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                else:
                    eng_cell = ws.cell(row=row_idx, column=2, value=eng_translation)
                eng_cell.border = THIN_BORDER
                col_offset = 1

            translation_col = 2 + col_offset
            translation = lang_table.get(korean, "")

            is_untranslated = (
                not translation or
                translation == korean or
                translation.strip() == korean.strip()
            )

            if is_untranslated:
                target_cell = ws.cell(row=row_idx, column=translation_col, value="NO_TRANSLATION")
                target_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                missing_count += 1
            else:
                target_cell = ws.cell(row=row_idx, column=translation_col, value=translation)
                translated_count += 1

            target_cell.border = THIN_BORDER

            category_col = 3 + col_offset
            ws.cell(row=row_idx, column=category_col, value=category).border = THIN_BORDER

        auto_fit_columns(ws, max_width=60)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return {"translated": translated_count, "missing": missing_count}

    def run(self):
        """Start the GUI."""
        self.root.mainloop()
