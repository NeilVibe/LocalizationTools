"""
Unify Translations Module

Reads a reference Excel (TMX Clean output) and a LineCheck Excel (QuickCheck output),
matches by StrOrigin, and produces a merge-ready Excel with unified corrections.

Reference: StrOrigin | Correction | StringID | ... | ChangeDate
LineCheck:  Source (KR) | [Category] | Translation 1 | SID 1 | Translation 2 | SID 2 | ...
Output:     StrOrigin | Correction | StringID | Category
"""
from __future__ import annotations

import html
import logging
import os
import re
from typing import Dict, List, Optional, Tuple, Callable
from dataclasses import dataclass

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None


# Pre-compiled regex patterns for normalization (avoid re-compile per call)
_RE_UNICODE_WS = re.compile(r'[\u00A0\u1680\u180E\u2000-\u200A\u202F\u205F\u3000\uFEFF]+')
_RE_ZERO_WIDTH = re.compile(r'[\u200B-\u200F\u202A-\u202E]+')
_RE_APOSTROPHE = re.compile(r'[\u2019\u2018\u02BC\u2032\u0060\u00B4]')
_RE_WHITESPACE = re.compile(r'\s+')


def _normalize(text: str) -> str:
    """Normalize text for matching — unescape HTML, collapse whitespace, strip."""
    if not text:
        return ""
    text = html.unescape(str(text))
    text = _RE_UNICODE_WS.sub(' ', text)
    text = _RE_ZERO_WIDTH.sub('', text)
    text = _RE_APOSTROPHE.sub("'", text)
    return _RE_WHITESPACE.sub(' ', text.strip())


@dataclass
class UnifyResult:
    """Single row in the unified output."""
    str_origin: str
    correction: str
    string_id: str
    category: str


def read_reference_excel(
    path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, Tuple[str, str]]:
    """
    Read reference Excel (TMX Clean output).

    Returns: {normalized_StrOrigin: (Correction, ChangeDate)}
    If duplicate StrOrigin, keeps the one with latest ChangeDate.
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required")

    file_size_mb = os.path.getsize(path) / (1024 * 1024)
    if progress_callback:
        progress_callback(f"[1/4] Reading reference file ({file_size_mb:.1f} MB)...")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError(f"Reference file is empty: {os.path.basename(path)}")

    # Detect columns by header
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    if progress_callback:
        progress_callback(f"  Headers found: {[str(c).strip() for c in rows[0] if c]}")

    str_origin_idx = None
    correction_idx = None
    changedate_idx = None

    for i, h in enumerate(header):
        if h == "strorigin":
            str_origin_idx = i
        elif h == "correction":
            correction_idx = i
        elif h == "changedate":
            changedate_idx = i

    # Validation
    missing = []
    if str_origin_idx is None:
        missing.append("StrOrigin")
    if correction_idx is None:
        missing.append("Correction")
    if missing:
        raise ValueError(f"Reference file missing required headers: {', '.join(missing)}")

    if progress_callback:
        has_date = "YES" if changedate_idx is not None else "NO"
        progress_callback(f"  StrOrigin=col {str_origin_idx}, Correction=col {correction_idx}, ChangeDate={has_date}")

    # Build lookup
    lookup: Dict[str, Tuple[str, str]] = {}
    skipped_empty = 0
    dupes_updated = 0
    total_rows = len(rows) - 1

    for row in rows[1:]:
        if not row or len(row) <= max(str_origin_idx, correction_idx):
            skipped_empty += 1
            continue

        str_origin_val = str(row[str_origin_idx] or "").strip()
        correction_val = str(row[correction_idx] or "").strip()
        changedate_val = ""
        if changedate_idx is not None and len(row) > changedate_idx:
            changedate_val = str(row[changedate_idx] or "").strip()

        if not str_origin_val or not correction_val:
            skipped_empty += 1
            continue

        key = _normalize(str_origin_val)

        if key in lookup:
            existing_date = lookup[key][1]
            if changedate_val > existing_date:
                lookup[key] = (correction_val, changedate_val)
                dupes_updated += 1
        else:
            lookup[key] = (correction_val, changedate_val)

    if progress_callback:
        progress_callback(f"  Read {total_rows} rows → {len(lookup)} unique StrOrigin entries")
        if skipped_empty:
            progress_callback(f"  Skipped {skipped_empty} empty/incomplete rows")
        if dupes_updated:
            progress_callback(f"  {dupes_updated} duplicate StrOrigins resolved by ChangeDate")

    return lookup


def read_linecheck_excel(
    path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> List[dict]:
    """
    Read LineCheck Excel output.

    Returns list of dicts:
        {'source': str, 'category': str, 'translations': [(translation, string_id), ...]}
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required")

    file_size_mb = os.path.getsize(path) / (1024 * 1024)
    if progress_callback:
        progress_callback(f"[2/4] Reading LineCheck file ({file_size_mb:.1f} MB)...")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError(f"LineCheck file is empty: {os.path.basename(path)}")

    # Parse header to find columns dynamically
    header = [str(c).strip() if c else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]

    if progress_callback:
        progress_callback(f"  Headers found: {[h for h in header if h]}")

    # Find Source column
    source_idx = None
    for i, h in enumerate(header_lower):
        if h.startswith("source"):
            source_idx = i
            break

    if source_idx is None:
        raise ValueError("LineCheck file missing 'Source (KR)' header")

    # Find Category column (optional)
    category_idx = None
    for i, h in enumerate(header_lower):
        if h == "category":
            category_idx = i
            break

    # Find Translation N / StringID N pairs
    trans_pairs: List[Tuple[int, int]] = []
    for i, h in enumerate(header):
        if h.startswith("Translation "):
            if i + 1 < len(header) and header[i + 1].startswith("StringID "):
                trans_pairs.append((i, i + 1))

    if not trans_pairs:
        raise ValueError("LineCheck file missing Translation/StringID column pairs")

    if progress_callback:
        has_cat = "YES" if category_idx is not None else "NO"
        progress_callback(f"  Source=col {source_idx}, Category={has_cat}, {len(trans_pairs)} translation pairs")

    # Parse data rows
    results = []
    total_string_ids = 0
    for row in rows[1:]:
        if not row:
            continue

        source = str(row[source_idx] or "").strip() if len(row) > source_idx else ""
        if not source:
            continue

        if source.lower().startswith("total:"):
            continue

        category = ""
        if category_idx is not None and len(row) > category_idx:
            category = str(row[category_idx] or "").strip()

        translations = []
        for trans_col, sid_col in trans_pairs:
            trans_val = str(row[trans_col] or "").strip() if len(row) > trans_col else ""
            sid_val = str(row[sid_col] or "").strip() if len(row) > sid_col else ""
            if trans_val and sid_val:
                translations.append((trans_val, sid_val))

        if translations:
            results.append({
                'source': source,
                'category': category,
                'translations': translations,
            })
            total_string_ids += len(translations)

    if progress_callback:
        progress_callback(f"  {len(results)} inconsistency groups, {total_string_ids} total StringIDs")

    return results


@dataclass
class SelectionEntry:
    """One row in the selection log — one per unique StrOrigin."""
    str_origin: str
    correction: str
    category: str
    method: str  # "Reference", "Tiebreaker (minority)", "Tiebreaker (linebreak)"


def _pick_best_translation(translations: List[Tuple[str, str]]) -> Tuple[str, str]:
    """
    Tiebreaker when reference correction doesn't help.

    Uses the LineCheck group translations to pick the best one:
    1. Minority wins (odd split) — the mass-confirmed majority is likely wrong
    2. Even split — fewer <br/> linebreaks wins (cleaner translation)

    Args:
        translations: [(translation_text, string_id), ...]

    Returns:
        (best_translation_text, method_used)
    """
    from collections import Counter

    # Count how many StringIDs have each translation
    trans_counts = Counter(t for t, _ in translations)

    if len(trans_counts) < 2:
        return translations[0][0], "Tiebreaker (single variant)"

    sorted_variants = trans_counts.most_common()

    least_common = sorted_variants[-1]
    most_common = sorted_variants[0]

    if least_common[1] != most_common[1]:
        return least_common[0], "Tiebreaker (minority)"

    # Even split — pick the one with fewer <br/> tags
    def br_count(text: str) -> int:
        return text.lower().count('<br/>') + text.lower().count('<br />')

    candidates = [variant for variant, _ in sorted_variants]
    candidates.sort(key=br_count)
    return candidates[0], "Tiebreaker (linebreak)"


def unify(
    reference: Dict[str, Tuple[str, str]],
    linecheck_groups: List[dict],
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Tuple[List[UnifyResult], List[SelectionEntry], int, int]:
    """
    Match LineCheck groups against reference to produce unified corrections.

    Tiebreaker logic when reference correction matches none of the translations:
    1. Minority translation wins (odd split — mass-confirmed majority is likely wrong)
    2. Even split — fewer <br/> linebreaks wins (cleaner translation)

    Returns: (results, selections, matched_ref, matched_tiebreak)
    """
    if progress_callback:
        progress_callback("[3/4] Matching LineCheck groups against reference...")

    results: List[UnifyResult] = []
    selections: List[SelectionEntry] = []
    matched_ref = 0
    matched_tiebreak = 0
    total = len(linecheck_groups)

    for idx, group in enumerate(linecheck_groups):
        source = group['source']
        category = group['category']
        translations = group['translations']
        key = _normalize(source)

        correction = None
        method = ""

        # Try reference lookup first
        if key in reference:
            ref_correction, _ = reference[key]
            ref_norm = _normalize(ref_correction)
            trans_norms = {_normalize(t): t for t, _ in translations}
            if ref_norm in trans_norms:
                correction = ref_correction
                method = "Reference"
                matched_ref += 1
            else:
                correction, method = _pick_best_translation(translations)
                matched_tiebreak += 1
        else:
            correction, method = _pick_best_translation(translations)
            matched_tiebreak += 1

        selections.append(SelectionEntry(
            str_origin=source,
            correction=correction,
            category=category,
            method=method,
        ))

        for trans_val, sid_val in translations:
            results.append(UnifyResult(
                str_origin=source,
                correction=correction,
                string_id=sid_val,
                category=category,
            ))

        if progress_callback and (idx + 1) % 50 == 0:
            progress_callback(f"  Processed {idx + 1}/{total} groups...")

    if progress_callback:
        progress_callback(f"  Processed: {total} groups")
        if matched_ref:
            progress_callback(f"    Reference match: {matched_ref}")
        if matched_tiebreak:
            progress_callback(f"    Tiebreaker (minority/linebreak): {matched_tiebreak}")
        progress_callback(f"  Output: {len(results)} rows, {len(selections)} unique selections")

    return results, selections, matched_ref, matched_tiebreak


def write_unified_excel(
    results: List[UnifyResult],
    output_path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> bool:
    """Write unified results to merge-ready Excel."""
    if xlsxwriter is None:
        raise ImportError("xlsxwriter is required")

    if progress_callback:
        progress_callback(f"[4/4] Writing {len(results)} rows to {os.path.basename(output_path)}...")

    try:
        wb = xlsxwriter.Workbook(output_path)
        ws = wb.add_worksheet('Unified')

        hdr_fmt = wb.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'valign': 'vcenter', 'text_wrap': True,
        })
        cell_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'text_wrap': True,
        })
        sid_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'num_format': '@',
        })

        headers = ['StrOrigin', 'Correction', 'StringID', 'Category']
        widths = [50, 50, 22, 15]

        for col, h in enumerate(headers):
            ws.write(0, col, h, hdr_fmt)
            ws.set_column(col, col, widths[col])

        for idx, r in enumerate(results, 1):
            ws.write(idx, 0, r.str_origin, cell_fmt)
            ws.write(idx, 1, r.correction, cell_fmt)
            ws.write_string(idx, 2, str(r.string_id), sid_fmt)
            ws.write(idx, 3, r.category, cell_fmt)

        data_rows = len(results)
        if data_rows > 0:
            ws.autofilter(0, 0, data_rows, len(headers) - 1)
            ws.freeze_panes(1, 0)

        wb.close()

        if progress_callback:
            # Count unique StrOrigins for summary
            unique_sources = len(set(r.str_origin for r in results))
            progress_callback(f"  Saved: {data_rows} rows ({unique_sources} unique sources)")

        return True

    except Exception as e:
        logger.error("Failed to write unified Excel %s: %s", output_path, e)
        if progress_callback:
            progress_callback(f"ERROR writing output: {e}")
        return False


def write_selection_log(
    selections: List[SelectionEntry],
    output_path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> bool:
    """Write selection log — one row per unique StrOrigin showing the chosen correction."""
    if xlsxwriter is None:
        raise ImportError("xlsxwriter is required")

    try:
        wb = xlsxwriter.Workbook(output_path)
        ws = wb.add_worksheet('Selections')

        hdr_fmt = wb.add_format({
            'bold': True, 'bg_color': '#27AE60', 'font_color': 'white',
            'border': 1, 'valign': 'vcenter', 'text_wrap': True,
        })
        cell_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'text_wrap': True,
        })
        method_ref_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'font_color': '#008000', 'bold': True,
        })
        method_tie_fmt = wb.add_format({
            'border': 1, 'valign': 'vcenter', 'font_color': '#E67E22', 'bold': True,
        })

        headers = ['StrOrigin', 'Correction', 'Category', 'Method']
        widths = [50, 50, 15, 25]

        for col, h in enumerate(headers):
            ws.write(0, col, h, hdr_fmt)
            ws.set_column(col, col, widths[col])

        for idx, s in enumerate(selections, 1):
            ws.write(idx, 0, s.str_origin, cell_fmt)
            ws.write(idx, 1, s.correction, cell_fmt)
            ws.write(idx, 2, s.category, cell_fmt)
            fmt = method_ref_fmt if s.method == "Reference" else method_tie_fmt
            ws.write(idx, 3, s.method, fmt)

        data_rows = len(selections)
        if data_rows > 0:
            ws.autofilter(0, 0, data_rows, len(headers) - 1)
            ws.freeze_panes(1, 0)

        wb.close()

        if progress_callback:
            ref_count = sum(1 for s in selections if s.method == "Reference")
            tie_count = data_rows - ref_count
            progress_callback(f"  Selection log: {data_rows} entries ({ref_count} ref, {tie_count} tiebreaker)")

        return True

    except Exception as e:
        logger.error("Failed to write selection log %s: %s", output_path, e)
        if progress_callback:
            progress_callback(f"ERROR writing selection log: {e}")
        return False


def run_unify(
    reference_path: str,
    linecheck_path: str,
    output_path: Optional[str] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Optional[str]:
    """
    Full pipeline: read reference + linecheck, match, write output.

    Returns output path on success, None on failure.
    """
    if output_path is None:
        base = os.path.splitext(linecheck_path)[0]
        output_path = f"{base}_Unified.xlsx"

    try:
        reference = read_reference_excel(reference_path, progress_callback)
        linecheck_groups = read_linecheck_excel(linecheck_path, progress_callback)
        results, selections, matched_ref, matched_tiebreak = unify(reference, linecheck_groups, progress_callback)

        if not results:
            if progress_callback:
                progress_callback("No matches found — nothing to output.")
            return None

        if progress_callback:
            progress_callback("[4/4] Writing output files...")

        ok = write_unified_excel(results, output_path, progress_callback)

        # Write selection log (second output)
        selection_path = output_path.replace("_Unified.xlsx", "_Selections.xlsx")
        if selection_path == output_path:
            base = os.path.splitext(output_path)[0]
            selection_path = f"{base}_Selections.xlsx"
        write_selection_log(selections, selection_path, progress_callback)

        if ok and progress_callback:
            progress_callback("-" * 40)
            progress_callback(f"DONE — 2 files generated:")
            progress_callback(f"  1. {os.path.basename(output_path)} ({len(results)} rows)")
            progress_callback(f"  2. {os.path.basename(selection_path)} ({len(selections)} selections)")
            progress_callback(f"  Reference match: {matched_ref}, Tiebreaker: {matched_tiebreak}")

        if ok:
            return output_path
        return None

    except Exception as e:
        logger.exception("Unify pipeline failed")
        if progress_callback:
            progress_callback(f"ERROR: {e}")
        return None
