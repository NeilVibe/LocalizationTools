"""
TMX Tools — Cleaning, postprocessing, conversion and Excel export for TMX files.

Ported from:
  - QuickStandaloneScripts/tmx_cleaner.py (cleaning + Excel pipeline)
  - QuickStandaloneScripts/tmxconvert41.py (postprocessor + conversion + batch)

Provides:
  - clean_segment() / clean_tmx_string() — strip ALL CAT tool markup
  - postprocess_tmx_string() — wrap plain text back into MemoQ bpt/ept/ph format
  - combine_xmls_to_tmx() / batch_tmx_from_folders() — XML→TMX conversion
  - clean_and_convert_to_excel() — TMX→clean→dedup→Excel pipeline
"""
from __future__ import annotations

import html
import logging
import os
import re
from datetime import datetime
from pathlib import Path

from lxml import etree

from .korean_detection import is_korean_text
from .source_scanner import scan_source_for_languages

logger = logging.getLogger(__name__)

# QuickTranslate suffix -> MemoQ BCP-47 language codes
SUFFIX_TO_BCP47 = {
    "ENG": "en-US",    "FRE": "fr-FR",    "GER": "de-DE",
    "ITA": "it-IT",    "JPN": "ja-JP",    "KOR": "ko",
    "POL": "pl-PL",    "POR-BR": "pt-BR", "RUS": "ru-RU",
    "SPA-ES": "es-ES", "SPA-MX": "es-MX", "TUR": "tr-TR",
    "ZHO-CN": "zh-CN", "ZHO-TW": "zh-TW",
}

# =============================================================================
# COMPILED REGEX PATTERNS  (from tmx_cleaner.py)
# =============================================================================

# Segment boundary
SEG_RE = re.compile(r'(<seg[^>]*>)(.*?)(</seg>)', re.DOTALL | re.IGNORECASE)

# Zero-width characters (ZWSP, ZWNJ, ZWJ, Word Joiner, BOM)
ZERO_WIDTH_RE = re.compile(r'[\u200b\u200c\u200d\u2060\ufeff]')

# --- MemoQ Staticinfo bpt/ept pairs ---
# Pattern A: &quot; escaped quotes
MEMOQ_BPT_EPT_ESCAPED_RE = re.compile(
    r'<bpt\b[^>]*>'
    r'&lt;mq:rxt(?:-req)?\s+displaytext=&quot;[^&]*&quot;\s+'
    r'val=&quot;\{Static[Ii]nfo:(\w+):([^#}]+)#&quot;&gt;</bpt>'
    r'(.*?)'
    r'<ept\b[^>]*>.*?</ept>',
    flags=re.DOTALL | re.IGNORECASE
)
# Pattern B: plain " quotes
MEMOQ_BPT_EPT_PLAIN_RE = re.compile(
    r'<bpt\b[^>]*>'
    r'&lt;mq:rxt(?:-req)?\s+displaytext="[^"]*"\s+'
    r'val="\{Static[Ii]nfo:(\w+):([^#}]+)#"&gt;</bpt>'
    r'(.*?)'
    r'<ept\b[^>]*>.*?</ept>',
    flags=re.DOTALL | re.IGNORECASE
)

# --- Generic <bpt>...<ept> pairs (Trados bold/italic/etc.) ---
GENERIC_BPT_EPT_RE = re.compile(
    r'<bpt\b[^>]*>[^<]*</bpt>(.*?)<ept\b[^>]*>[^<]*</ept>',
    flags=re.DOTALL | re.IGNORECASE
)

# --- Non-fmt <ph>...</ph> (captures inner content) ---
PH_RE = re.compile(
    r'<ph\b(?![^>]*\btype=[\'"]fmt[\'"])[^>]*>(.*?)</ph>',
    flags=re.DOTALL | re.IGNORECASE
)

# --- Self-closing <ph .../> ---
PH_SELFCLOSE_RE = re.compile(
    r'<ph\b[^>]*/\s*>',
    flags=re.IGNORECASE
)

# --- <ph type='fmt'> (MemoQ formatting whitespace) ---
PH_FMT_RE = re.compile(
    r"<ph\b[^>]*\btype=['\"]fmt['\"][^>]*>.*?</ph>",
    flags=re.DOTALL | re.IGNORECASE
)

# --- <it> tags (Trados isolated tags) ---
IT_RE = re.compile(
    r'<it\b[^>]*>.*?</it>',
    flags=re.DOTALL | re.IGNORECASE
)

# --- <x/> self-closing tags ---
X_RE = re.compile(
    r'<x\b[^>]*/?>',
    flags=re.IGNORECASE
)

# --- <g>text</g> tags (keep inner text) ---
G_RE = re.compile(
    r'<g\b[^>]*>(.*?)</g>',
    flags=re.DOTALL | re.IGNORECASE
)

# --- Formatting content detection inside decoded <ph> ---
FORMATTING_CONTENT_RE = re.compile(
    r'^<(?:cf\b|/cf|b>|/b>|i>|/i>|u>|/u>|sub>|/sub>|sup>|/sup>)',
    flags=re.IGNORECASE
)

# --- BR tag normalization ---
BR_VARIANTS_RE = re.compile(
    r'<\s*br\s*/?\s*>|'           # <br>, <br/>, <br />, etc.
    r'<\s*br\s*/(?=\s|[^>\s])|'   # <br/ (malformed, no >)
    r'<\s*br(?=\s+[^/>\s])|'      # <br (malformed)
    r'br\s*/\s*>|'                # br/> missing opening <
    r'&lt;\s*br\s*/?\s*&gt;|'     # &lt;br/&gt; variants
    r'&amp;lt;br\s*/&amp;gt;',    # double-escaped variants
    flags=re.IGNORECASE
)


# =============================================================================
# CORE CLEANING FUNCTION
# =============================================================================

def clean_segment(content: str) -> str:
    """
    Clean a single <seg> content string.
    Strips all CAT tool markup, returns plain text.
    """
    # 1) Strip zero-width characters
    content = ZERO_WIDTH_RE.sub('', content)

    # 2) MemoQ StaticInfo bpt/ept → {StaticInfo:Category:ID#inner}
    def _staticinfo_repl(m):
        category, ident, inner = m.group(1).strip(), m.group(2).strip(), m.group(3)
        return f'{{StaticInfo:{category}:{ident}#{inner}}}'
    content = MEMOQ_BPT_EPT_ESCAPED_RE.sub(_staticinfo_repl, content)
    content = MEMOQ_BPT_EPT_PLAIN_RE.sub(_staticinfo_repl, content)

    # 3) Generic <bpt>...<ept> pairs → keep inner text
    content = GENERIC_BPT_EPT_RE.sub(r'\1', content)

    # 4) Non-fmt <ph>...</ph> → smart 5-priority extraction
    def _ph_repl(m):
        ph_inner = m.group(1)
        if not ph_inner or ph_inner.isspace():
            return ''
        decoded = html.unescape(ph_inner)

        # P1: extract val="..."
        vm = re.search(r'val="([^"]+)"', decoded)
        if vm:
            val = vm.group(1)
            val = val.replace('&', '&amp;')
            return val

        # P2: formatting content (cf, b, i, etc.) → remove
        if FORMATTING_CONTENT_RE.search(decoded):
            return ''

        # P3: displaytext with no val → extract displaytext
        dm = re.search(r'displaytext="([^"]+)"', decoded)
        if dm:
            return dm.group(1)

        # P4: structural XML tag (NOT <br/>) → remove
        stripped = decoded.strip()
        if (stripped.startswith('<') and stripped.endswith('>')
                and not re.match(r'<\s*br\s*/?\s*>', stripped, re.IGNORECASE)):
            return ''

        # P5: keep raw inner text as-is
        return ph_inner

    content = PH_RE.sub(_ph_repl, content)

    # 5) Self-closing <ph .../> → remove
    content = PH_SELFCLOSE_RE.sub('', content)

    # 6) <ph type='fmt'> → remove
    content = PH_FMT_RE.sub('', content)

    # 7) <it> tags → remove
    content = IT_RE.sub('', content)

    # 8) <x/> tags → remove
    content = X_RE.sub('', content)

    # 9) <g>text</g> → keep inner text
    content = G_RE.sub(r'\1', content)

    # 10) Normalize newlines → &lt;br/&gt;
    content = re.sub(r'\r\n|\r|\n', '&lt;br/&gt;', content)
    content = content.replace('\\n', '&lt;br/&gt;')

    # 11) Normalize all <br> variants → &lt;br/&gt;
    content = BR_VARIANTS_RE.sub('&lt;br/&gt;', content)

    return content


def clean_tmx_string(tmx_text: str) -> str:
    """Clean all <seg> contents in a TMX string."""
    def _seg_repl(m):
        open_tag, content, close_tag = m.groups()
        return f"{open_tag}{clean_segment(content)}{close_tag}"
    return SEG_RE.sub(_seg_repl, tmx_text)


# =============================================================================
# MEMOQ POSTPROCESSOR  (from tmxconvert41.py, with fixes)
# =============================================================================

def postprocess_tmx_string(xml_str: str) -> str:
    """
    Re-wrap plain text back into MemoQ bpt/ept/ph format.

    1. {StaticInfo:Category:ID#inner} → <bpt>/<ept> pairs (per-segment counter)
    2. Existing placeholder conversions → <ph>
    3. Spaces after <ph> br → fmt
    4. Trim trailing whitespace at end of each <seg>
    """

    # Helper: apply regex only outside already-made <ph>...</ph> tags
    def replace_outside_ph(pattern, repl, text):
        parts = []
        last = 0
        for m in re.finditer(r'<ph>.*?</ph>', text, flags=re.DOTALL):
            gap = text[last:m.start()]
            parts.append(re.sub(pattern, repl, gap))
            parts.append(m.group(0))
            last = m.end()
        parts.append(re.sub(pattern, repl, text[last:]))
        return ''.join(parts)

    # -------------------------------------------------------
    # 1. StaticInfo → <bpt>/<ept> PAIRS  (Fix 1: any category, 3 groups)
    # -------------------------------------------------------
    sk_pattern = re.compile(r'\{Static[Ii]nfo:(\w+):([^#}]+)#([^}]+)\}')

    seg_pattern = re.compile(r'(<seg[^>]*>)(.*?)(</seg>)', re.DOTALL)

    def seg_repl(seg_match):
        open_tag, content, close_tag = seg_match.groups()
        counter = {'i': 0}

        def sk_repl(m):
            counter['i'] += 1
            i = counter['i']
            category = m.group(1)
            ident = m.group(2)
            inner_txt = m.group(3)

            # Fix 2: mq:rxt-req in BOTH bpt AND ept
            # Fix 3: use captured category instead of hardcoded Knowledge
            bpt = (
                f"<bpt i='{i}'>&lt;mq:rxt-req displaytext=&quot;"
                f"{html.escape(ident)}&quot; "
                f"val=&quot;{{StaticInfo:{category}:{html.escape(ident)}#&quot;&gt;</bpt>"
            )
            ept = (
                f"<ept i='{i}'>&lt;/mq:rxt-req displaytext=&quot;}}&quot; "
                f"val=&quot;}}&quot;&gt;</ept>"
            )
            return f"{bpt}{inner_txt}{ept}"

        new_content = replace_outside_ph(sk_pattern, sk_repl, content)
        return f"{open_tag}{new_content}{close_tag}"

    xml_str = seg_pattern.sub(seg_repl, xml_str)

    # -------------------------
    # 2. <ph> PLACEHOLDERS
    # -------------------------
    # 2-A  %N#
    def _repl_pct(m):
        num = m.group(1)
        return f'<ph>&lt;mq:rxt-req displaytext="Param{num}" val="%{num}#" /&gt;</ph>'
    xml_str = replace_outside_ph(r'%(\d)#', _repl_pct, xml_str)

    # 2-B  {...} — exclude StaticInfo so it's not double-processed  (Fix 4)
    def _repl_br(m):
        inner = m.group(1)
        return f'<ph>&lt;mq:rxt-req displaytext="{inner}" val="{{{inner}}}" /&gt;</ph>'
    xml_str = replace_outside_ph(r'\{(?!Static[Ii]nfo:\w+:)([^}]+)\}', _repl_br, xml_str)

    # 2-C  \X
    def _repl_bs(m):
        c = m.group(1)
        return f'<ph>&lt;mq:rxt displaytext="\\{c}" val="\\{c}" /&gt;</ph>'
    xml_str = replace_outside_ph(r'\\(\w)', _repl_bs, xml_str)

    # 2-D  &lt;br/&gt;
    xml_str = replace_outside_ph(
        r'&lt;br/&gt;',
        lambda _: '<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>',
        xml_str
    )

    # 2-E  &amp;desc;
    xml_str = replace_outside_ph(
        r'&amp;desc;',
        lambda _: '<ph>&lt;mq:rxt-req displaytext="desc" val="&amp;desc;" /&gt;</ph>',
        xml_str
    )

    # -------------------------------------------------
    # 3. Clean-up: spaces after <ph> br → fmt
    # -------------------------------------------------
    def _repl_spaces(m):
        br_ph, spaces, nxt = m.groups()
        return f"{br_ph}<ph type='fmt'>{spaces}</ph>{nxt}"

    xml_str = re.sub(
        r'(<ph>&lt;mq:rxt displaytext="br" val="&lt;br/&gt;" /&gt;</ph>)(\s{2,})(\S)',
        _repl_spaces,
        xml_str
    )

    # -------------------------------------------------
    # 4. Trim trailing whitespace at end of <seg>  (Fix 5: rstrip only)
    # -------------------------------------------------
    def trim_trailing_ws_in_seg(m):
        open_tag, content, close_tag = m.groups()
        return f"{open_tag}{content.rstrip()}{close_tag}"

    xml_str = re.sub(r'(<seg[^>]*>)(.*?)(</seg>)', trim_trailing_ws_in_seg, xml_str, flags=re.DOTALL)

    return xml_str


# =============================================================================
# CONVERSION HELPERS  (from tmxconvert41.py)
# =============================================================================

def replace_newlines_text(text: str | None) -> str:
    """Replace newlines with proper XML representation."""
    if text is None:
        return ''
    cleaned = text.replace('\n', '&lt;br/&gt;')
    cleaned = cleaned.replace('\\n', '&lt;br/&gt;')
    return cleaned


def _fix_bad_entities(xml_text: str) -> str:
    return re.sub(r'&(?!lt;|gt;|amp;|apos;|quot;)', '&amp;', xml_text)


def get_all_xml_files(input_folder: str) -> list[str]:
    xml_files = []
    for dirpath, _, filenames in os.walk(input_folder):
        for file in filenames:
            if file.lower().endswith(".xml"):
                xml_files.append(os.path.join(dirpath, file))
    return xml_files


def parse_xml_file(file_path: str):
    """Parse an XML file with entity recovery and ROOT wrapper."""
    try:
        txt = open(file_path, encoding='utf-8').read()
    except Exception as e:
        logger.error(f"Error reading {file_path!r}: {e}")
        return None
    txt = re.sub(r'&(?!lt;|gt;|amp;|apos;|quot;)', '&amp;', txt)
    wrapped = "<ROOT>\n" + txt + "\n</ROOT>"
    rec_parser = etree.XMLParser(recover=True)
    try:
        recovered = etree.fromstring(wrapped.encode('utf-8'), parser=rec_parser)
    except etree.XMLSyntaxError as e:
        logger.error(f"Fatal parse error (recover mode) on {file_path!r}: {e}")
        return None
    strict_parser = etree.XMLParser(recover=False)
    blob = etree.tostring(recovered, encoding='utf-8')
    try:
        return etree.fromstring(blob, parser=strict_parser)
    except etree.XMLSyntaxError:
        return recovered


# =============================================================================
# XML → TMX CONVERSION  (from tmxconvert41.py)
# =============================================================================

def combine_xmls_to_tmx(
    input_folder: str,
    output_file: str,
    target_language: str,
    postprocess: bool = True,
    file_list: list[str] | None = None,
    creation_id: str | None = None,
) -> bool:
    """
    Combine all XML files in input_folder into a single TMX file.

    Main string TUs: dedup by (StrOrigin, StringId), keep newest.
    Description TUs: keep all (no dedup).
    """
    try:
        xml_files = file_list if file_list else get_all_xml_files(input_folder)
        total_files = len(xml_files)
        logger.info(f"[TMX] Found {total_files} XML file(s) to process")

        if not xml_files:
            logger.warning("[TMX] No XML files found — nothing to convert")
            return False

        # Build TMX skeleton
        tmx = etree.Element("tmx", version="1.4")
        header_attr = {
            "creationtool":        "CombinedXMLConverter",
            "creationtoolversion": "1.1",
            "segtype":             "sentence",
            "o-tmf":               "UTF-8",
            "adminlang":           "en",
            "srclang":             "ko",
            "datatype":            "PlainText"
        }
        etree.SubElement(tmx, "header", **header_attr)
        body = etree.SubElement(tmx, "body")

        XML_NS = "http://www.w3.org/XML/1998/namespace"
        target_lang_code = target_language.lower()

        main_tu_map = {}   # (korean_text, string_id) → (mtime, tu_element)
        desc_tu_list = []

        total_entries = 0
        total_desc = 0
        skipped_korean = 0

        for idx, xml_file_path in enumerate(xml_files, 1):
            logger.info(f"  [{idx}/{total_files}] {os.path.basename(xml_file_path)}")

            try:
                if os.path.getsize(xml_file_path) == 0:
                    continue
            except OSError:
                continue

            xml_root = parse_xml_file(xml_file_path)
            if xml_root is None:
                continue

            doc_name = os.path.basename(xml_file_path)
            file_mtime = os.path.getmtime(xml_file_path)

            for loc in xml_root.iter("LocStr"):
                string_id   = (loc.get("StringId") or "").strip()
                korean_text = replace_newlines_text(loc.get("StrOrigin", ""))
                target_text = replace_newlines_text(loc.get("Str", ""))
                desc_origin = replace_newlines_text(loc.get("DescOrigin", "").strip())
                desc_text   = replace_newlines_text(loc.get("Desc",       "").strip())

                # Clean up legacy "&desc;" markers
                korean_text = korean_text.replace("&amp;desc;", "").replace("&desc;", "")
                target_text = target_text.replace("&amp;desc;", "").replace("&desc;", "")

                # MAIN TRANSLATION UNIT
                if (target_text
                        and string_id
                        and not is_korean_text(target_text)):
                    total_entries += 1
                    tu = etree.Element(
                        "tu",
                        creationid=creation_id or "CombinedConversion",
                        changeid=creation_id or "CombinedConversion"
                    )

                    tuv_ko = etree.SubElement(
                        tu, "tuv", **{f"{{{XML_NS}}}lang": "ko"}
                    )
                    etree.SubElement(tuv_ko, "seg").text = korean_text or ""

                    tuv_tgt = etree.SubElement(
                        tu, "tuv", **{f"{{{XML_NS}}}lang": target_lang_code}
                    )
                    etree.SubElement(tuv_tgt, "seg").text = target_text

                    etree.SubElement(tu, "prop", type="x-document").text = doc_name
                    etree.SubElement(tu, "prop", type="x-context").text  = string_id

                    key = (korean_text, string_id)
                    if key not in main_tu_map or file_mtime > main_tu_map[key][0]:
                        main_tu_map[key] = (file_mtime, tu)

                elif target_text and is_korean_text(target_text):
                    skipped_korean += 1

                # DESCRIPTION TRANSLATION UNIT
                if (string_id
                        and desc_origin
                        and desc_text
                        and not is_korean_text(desc_text)):

                    if (desc_origin.strip() in ("&desc;", "&amp;desc;")
                            and desc_text.strip() in ("&desc;", "&amp;desc;")):
                        continue

                    if (not desc_origin.startswith("&desc;")
                            and not desc_origin.startswith("&amp;desc;")):
                        desc_origin = "&desc;" + desc_origin

                    desc_tu = etree.Element(
                        "tu",
                        creationid=creation_id or "PAAT",
                        changeid=creation_id or "PAAT"
                    )

                    tuv_ko_d = etree.SubElement(
                        desc_tu, "tuv", **{f"{{{XML_NS}}}lang": "ko"}
                    )
                    tuv_tgt_d = etree.SubElement(
                        desc_tu, "tuv", **{f"{{{XML_NS}}}lang": target_lang_code}
                    )

                    if postprocess:
                        seg_ko = etree.SubElement(tuv_ko_d, "seg")
                        ph_ko = etree.SubElement(seg_ko, "ph")
                        ph_ko.text = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
                        ph_ko.tail = desc_origin.replace("&amp;desc;", "").replace("&desc;", "")

                        seg_tgt = etree.SubElement(tuv_tgt_d, "seg")
                        ph_tgt = etree.SubElement(seg_tgt, "ph")
                        ph_tgt.text = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
                        ph_tgt.tail = desc_text.replace("&amp;desc;", "").replace("&desc;", "")
                    else:
                        etree.SubElement(tuv_ko_d,  "seg").text = desc_origin
                        etree.SubElement(tuv_tgt_d, "seg").text = desc_text

                    etree.SubElement(desc_tu, "prop", type="x-document").text = doc_name
                    etree.SubElement(desc_tu, "prop", type="x-context").text  = string_id
                    desc_tu_list.append(desc_tu)
                    total_desc += 1

        # Dedup summary
        deduped = total_entries - len(main_tu_map)
        logger.info(f"[TMX] Scanned {total_entries} TUs from {total_files} files")
        if deduped > 0:
            logger.info(f"  Dedup: {deduped} duplicate(s) removed -> {len(main_tu_map)} unique main TUs")
        if total_desc:
            logger.info(f"  Descriptions: {total_desc} desc TU(s)")
        if skipped_korean:
            logger.info(f"  Skipped: {skipped_korean} untranslated (Korean) entries")

        # WRITE OUT TMX
        for _, tu in main_tu_map.values():
            body.append(tu)
        for tu in desc_tu_list:
            body.append(tu)

        total_tus = len(main_tu_map) + len(desc_tu_list)
        logger.info(f"[TMX] Writing {total_tus} TU(s) to TMX...")

        xml_bytes = etree.tostring(
            tmx, pretty_print=True, encoding='UTF-8', xml_declaration=True
        )
        xml_str = xml_bytes.decode('utf-8')

        if postprocess:
            logger.info("[TMX] Postprocessing: adding MemoQ bpt/ept/ph markup...")
            xml_str = postprocess_tmx_string(xml_str)

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(xml_str)

        logger.info(f"[TMX] Done -> {os.path.basename(output_file)} ({total_tus} TUs)")
        return True

    except Exception as e:
        logger.error(f"[TMX] Error during combination: {e}", exc_info=True)
        return False


def batch_tmx_from_folders(
    input_folders: list[str],
    output_dir: str,
    target_language: str,
    postprocess: bool = False,
) -> list[tuple[str, str, bool]]:
    """
    For each folder in input_folders, create a TMX in output_dir.

    Args:
        postprocess: False = NORMAL TMX, True = MemoQ-TMX (with mq:rxt placeholders)

    Returns list of (input_folder, output_file, success) tuples.
    """
    mode = "MemoQ-TMX" if postprocess else "NORMAL TMX"
    results = []
    for folder in input_folders:
        folder_name = os.path.basename(os.path.normpath(folder))
        out_file = os.path.join(output_dir, f"{folder_name}.tmx")
        logger.info(f"[BATCH-{mode}] Processing: {folder} -> {out_file}")
        ok = combine_xmls_to_tmx(folder, out_file, target_language, postprocess=postprocess)
        results.append((folder, out_file, ok))
    return results


def convert_to_memoq_tmx(input_path: str) -> list[tuple[str, str, bool]]:
    """
    Auto-detect languages from input path, create one MemoQ TMX per language.

    Uses scan_source_for_languages() -- same logic as QuickTranslate Tab 1.
    Skips KOR (source language -- Korean target produces empty TMX since
    is_korean_text filters all TUs).

    Returns list of (language_code, output_file, success) tuples.
    """
    path = Path(input_path)
    logger.info("=" * 60)
    logger.info("[MemoQ-TMX] Scanning for languages: %s", path.name)
    scan = scan_source_for_languages(path)

    if not scan.lang_files:
        logger.warning("[MemoQ-TMX] No language files detected in: %s", input_path)
        if scan.unrecognized:
            logger.warning("  %d unrecognized items (no language suffix)",
                           len(scan.unrecognized))
        return []

    # Show detected languages
    lang_list = [f"{k.upper()} ({len(v)} files)" for k, v in sorted(scan.lang_files.items())]
    logger.info("  Detected: %s", ", ".join(lang_list))
    if scan.unrecognized:
        logger.warning("  %d unrecognized items skipped", len(scan.unrecognized))

    creation_id = f"QT_{datetime.now().strftime('%Y%m%d_%H%M')}"

    # Determine output base name
    if path.is_file():
        base_name = path.stem
        output_dir = str(path.parent)
    else:
        base_name = path.name
        output_dir = str(path)

    results = []
    langs_to_convert = [
        (k, v) for k, v in scan.lang_files.items() if k.upper() != "KOR"
    ]
    if len(langs_to_convert) < len(scan.lang_files):
        logger.info("  Skipping KOR (source language)")

    for i, (lang_code, files) in enumerate(langs_to_convert, 1):
        upper = lang_code.upper()
        bcp47 = SUFFIX_TO_BCP47.get(upper, lang_code.lower())
        out_file = os.path.join(output_dir, f"{base_name}_{upper}.tmx")

        # Convert Path objects to strings for combine_xmls_to_tmx
        file_list = [str(f) for f in files]

        logger.info("-" * 40)
        logger.info("[MemoQ-TMX] (%d/%d) %s — %d file(s), target=%s",
                     i, len(langs_to_convert), upper, len(file_list), bcp47)

        ok = combine_xmls_to_tmx(
            input_folder=output_dir,
            output_file=out_file,
            target_language=bcp47,
            postprocess=True,
            file_list=file_list,
            creation_id=creation_id,
        )
        results.append((upper, out_file, ok))

    # Final summary
    logger.info("=" * 60)
    ok_count = sum(1 for _, _, ok in results if ok)
    fail_count = len(results) - ok_count
    logger.info("[MemoQ-TMX] Complete: %d language(s) converted, %d failed",
                ok_count, fail_count)
    for lang, out, ok in results:
        status = "OK" if ok else "FAILED"
        logger.info("  %s: %s -> %s", lang, status, os.path.basename(out))

    return results


# =============================================================================
# FILE I/O HELPERS  (from tmx_cleaner.py)
# =============================================================================

def detect_encoding(raw_bytes: bytes) -> str:
    """Detect encoding from BOM or default to utf-8."""
    if raw_bytes.startswith(b'\xff\xfe') or raw_bytes.startswith(b'\xfe\xff'):
        return 'utf-16'
    if raw_bytes.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    return 'utf-8'


def read_file(path: str) -> str:
    """Read a file with auto-detected encoding."""
    with open(path, 'rb') as f:
        raw = f.read()
    enc = detect_encoding(raw)
    try:
        return raw.decode(enc)
    except UnicodeDecodeError:
        return raw.decode('utf-8', errors='replace')


# =============================================================================
# TMX → EXCEL PIPELINE  (from tmx_cleaner.py, simplified output)
# =============================================================================

def parse_tmx_to_rows(fpath: str) -> list[dict]:
    """
    Parse a TMX file into a list of row dicts.
    Cleans all <seg> contents during extraction.
    """
    raw_text = read_file(fpath)

    body = re.sub(r'^<\?xml[^>]*\?>\s*', '', raw_text, flags=re.MULTILINE)
    body = re.sub(r'<!DOCTYPE[^>]*>\s*', '', body, flags=re.IGNORECASE)

    parser = etree.XMLParser(remove_blank_text=True, recover=True)
    try:
        root = etree.fromstring(body.encode('utf-8'), parser=parser)
    except Exception as e:
        logger.error(f"Cannot parse {fpath}: {e}")
        return []

    XML_NS = "http://www.w3.org/XML/1998/namespace"
    rows = []

    for tu in root.iter('tu'):
        row = {
            'changedate': tu.get('changedate', ''),
            'creationdate': tu.get('creationdate', ''),
            'creationid': tu.get('creationid', ''),
            'changeid': tu.get('changeid', ''),
            'client': '',
            'project': '',
            'domain': '',
            'subject': '',
            'corrected': '',
            'aligned': '',
            'x_document': '',
            'x_context': '',
            'ko_seg': '',
            'tgt_seg': '',
            'tgt_lang': '',
            'is_desc': False,
        }

        for prop in tu.findall('prop'):
            ptype = prop.get('type', '')
            pval = (prop.text or '').strip()
            if ptype == 'client':
                row['client'] = pval
            elif ptype == 'project':
                row['project'] = pval
            elif ptype == 'domain':
                row['domain'] = pval
            elif ptype == 'subject':
                row['subject'] = pval
            elif ptype == 'corrected':
                row['corrected'] = pval
            elif ptype == 'aligned':
                row['aligned'] = pval
            elif ptype == 'x-document':
                row['x_document'] = pval
            elif ptype == 'x-context':
                row['x_context'] = pval

        for tuv in tu.findall('tuv'):
            lang = tuv.get(f'{{{XML_NS}}}lang', tuv.get('lang', '')).lower()
            seg = tuv.find('seg')
            if seg is None:
                continue

            # ── Step 1: Detect &desc; BEFORE serialization ──
            # Use seg.text (Python string, already unescaped by lxml)
            # like tmxtransfer11.py does. This avoids double-encoding
            # issues from etree.tostring().
            raw_text = (seg.text or '').strip()
            is_desc_seg = (
                raw_text.lower().startswith('&desc;')
                or raw_text.lower().startswith('&amp;desc;')
            )

            # Also check for MemoQ <ph> desc tag (seg.text is empty,
            # desc marker is inside <ph> child)
            if not is_desc_seg and len(seg) > 0:
                ph = seg.find('ph')
                if ph is not None:
                    ph_text = (ph.text or '').strip().lower()
                    if 'desc' in ph_text and 'rxt-req' in ph_text:
                        is_desc_seg = True

            if is_desc_seg:
                row['is_desc'] = True

            # ── Step 2: Clean CAT markup (uses tostring for full content) ──
            seg_text = etree.tostring(seg, encoding='unicode', method='xml')
            seg_text = re.sub(r'^<seg[^>]*>', '', seg_text)
            seg_text = re.sub(r'</seg>$', '', seg_text)
            seg_text = seg_text.strip()

            cleaned = clean_segment(seg_text)

            # ── Step 3: Strip &desc; prefix from cleaned text ──
            if is_desc_seg:
                cleaned = re.sub(
                    r'^(&amp;amp;desc;|&amp;desc;|&desc;)', '',
                    cleaned, flags=re.IGNORECASE
                ).strip()

            if lang.startswith('ko'):
                row['ko_seg'] = cleaned
            else:
                row['tgt_seg'] = cleaned
                row['tgt_lang'] = lang

        rows.append(row)

    return rows


def dedup_rows(rows: list[dict]) -> list[dict]:
    """
    Deduplicate by (x_context, ko_seg) and merge main + desc TUs by context.

    Main TUs: dedup by (x_context, ko_seg), keep latest.
    Desc TUs: merged into same row via desc_origin/desc_text fields.
    """
    empty_ctx = sum(1 for r in rows if not r['x_context'])
    if rows and empty_ctx > len(rows) * 0.5:
        logger.warning(f"{empty_ctx}/{len(rows)} rows have no x-context — dedup is by KO text only")

    # Separate main and desc TUs
    main_rows = [r for r in rows if not r.get('is_desc')]
    desc_rows = [r for r in rows if r.get('is_desc')]

    # Dedup main TUs by (x_context, ko_seg), keep latest
    seen: dict[tuple, dict] = {}
    for row in main_rows:
        key = (row['x_context'], row['ko_seg'])
        if key in seen:
            if row['changedate'] > seen[key]['changedate']:
                seen[key] = row
        else:
            seen[key] = row

    # Initialize desc fields on all main rows
    result = list(seen.values())
    for r in result:
        r.setdefault('desc_origin', '')
        r.setdefault('desc_text', '')

    # Build context→row lookup for merging desc into main
    ctx_lookup: dict[str, dict] = {}
    for r in result:
        ctx = r['x_context']
        if ctx:
            ctx_lookup[ctx] = r

    # Merge desc TUs into their matching main row by x_context
    merged_desc = 0
    orphan_desc = 0
    for d in desc_rows:
        ctx = d['x_context']
        if ctx and ctx in ctx_lookup:
            ctx_lookup[ctx]['desc_origin'] = d['ko_seg']
            ctx_lookup[ctx]['desc_text'] = d['tgt_seg']
            merged_desc += 1
        else:
            orphan_desc += 1

    if merged_desc:
        logger.info(f"  Merged {merged_desc} description(s) into main rows")
    if orphan_desc:
        logger.warning(f"  {orphan_desc} orphan description(s) with no matching main TU")

    return result


def write_excel(rows: list[dict], output_path: str) -> None:
    """Write rows to Excel using xlsxwriter — 5-column format with Desc support."""
    import xlsxwriter

    wb = xlsxwriter.Workbook(output_path)
    ws = wb.add_worksheet('TMX Clean')

    # 5-column format: DescOrigin + Desc (TM reference, not correction)
    # User manually adds DescCorrection column when they want to correct descriptions
    headers = ['StrOrigin', 'Correction', 'StringID', 'DescOrigin', 'Desc']
    keys = ['ko_seg', 'tgt_seg', 'x_context', 'desc_origin', 'desc_text']

    hdr_fmt = wb.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1})
    cell_fmt = wb.add_format({'text_wrap': True, 'valign': 'top', 'border': 1})
    stringid_fmt = wb.add_format({'text_wrap': True, 'valign': 'top', 'border': 1, 'num_format': '@'})

    for col, h in enumerate(headers):
        ws.write(0, col, h, hdr_fmt)

    for row_idx, row in enumerate(rows, 1):
        for col, key in enumerate(keys):
            fmt = stringid_fmt if key == 'x_context' else cell_fmt
            ws.write(row_idx, col, str(row.get(key, '')), fmt)

    # Column widths
    ws.set_column(0, 0, 60)   # StrOrigin
    ws.set_column(1, 1, 60)   # Correction
    ws.set_column(2, 2, 30)   # StringID
    ws.set_column(3, 3, 50)   # DescOrigin
    ws.set_column(4, 4, 50)   # DescText

    ws.autofilter(0, 0, len(rows), len(headers) - 1)

    # Count desc rows for logging
    desc_count = sum(1 for r in rows if r.get('desc_origin'))
    wb.close()
    logger.info(f"[EXCEL] Written {len(rows)} rows ({desc_count} with descriptions) to: {output_path}")


def clean_and_convert_to_excel(fpath: str) -> str:
    """
    Full pipeline: TMX → clean → dedup → Excel.
    Output file: same name as input but _clean.xlsx.
    Returns the output path.
    """
    logger.info("=" * 60)
    logger.info("[TMX Cleaner] %s", os.path.basename(fpath))
    logger.info("-" * 40)

    logger.info("[1/3] Parsing TMX and cleaning CAT markup...")
    rows = parse_tmx_to_rows(fpath)
    if not rows:
        logger.warning("  No TU entries found — file may be empty or malformed")
        raise ValueError(f"No TU entries found in {os.path.basename(fpath)}")
    # Count languages found
    langs = set(r.get('tgt_lang', '') for r in rows if r.get('tgt_lang'))
    logger.info(f"  Parsed {len(rows)} TU entries")
    if langs:
        logger.info(f"  Target language(s): {', '.join(sorted(langs))}")

    logger.info("[2/3] Deduplicating by (x-context + KO text), keeping latest...")
    before = len(rows)
    rows = dedup_rows(rows)
    dupes = before - len(rows)
    if dupes:
        logger.info(f"  {dupes} duplicate(s) removed -> {len(rows)} unique rows")
    else:
        logger.info(f"  No duplicates found ({len(rows)} rows)")

    base = os.path.splitext(fpath)[0]
    output_path = f"{base}_clean.xlsx"

    logger.info("[3/3] Writing Excel (5 columns: StrOrigin, Correction, StringID, DescOrigin, DescText)...")
    write_excel(rows, output_path)

    logger.info("=" * 60)
    logger.info("[TMX Cleaner] Done -> %s (%d rows)", os.path.basename(output_path), len(rows))
    return output_path


# =============================================================================
# EXCEL → TMX PIPELINE (cross-format: reads 5-column Excel, creates MemoQ TMX)
# =============================================================================

def excel_to_memoq_tmx(
    excel_path: str,
    output_path: str | None = None,
    target_language: str = "en-us",
    postprocess: bool = True,
    creation_id: str | None = None,
) -> str:
    """
    Convert a 5-column Excel file to MemoQ-compatible TMX.

    Excel format (matching tmxconvert41 legacy):
        Col 0: StrOrigin    — Korean source text
        Col 1: Correction   — Translation (Str)
        Col 2: StringID     — Context ID (x-context)
        Col 3: DescOrigin   — Korean description (optional)
        Col 4: DescText     — Translated description (optional)

    Also supports 3-column format (StrOrigin, Correction, StringID) — no desc.

    Returns the output TMX path.
    """
    import openpyxl

    logger.info("=" * 60)
    logger.info("[Excel→TMX] %s", os.path.basename(excel_path))
    logger.info("-" * 40)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_data = list(ws.iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()

    if not rows_data:
        raise ValueError(f"No data rows in {os.path.basename(excel_path)}")

    logger.info(f"  Read {len(rows_data)} data rows")

    # Detect column count (3 or 5)
    sample_cols = max(len(r) for r in rows_data if r)
    has_desc = sample_cols >= 5
    if has_desc:
        logger.info("  5-column format detected (with DescOrigin/DescText)")
    else:
        logger.info("  3-column format detected (no descriptions)")

    # Build TMX skeleton
    XML_NS = "http://www.w3.org/XML/1998/namespace"
    tmx = etree.Element("tmx", version="1.4")
    header_attr = {
        "creationtool":        "Excel2TMX",
        "creationtoolversion": "2.1",
        "segtype":             "sentence",
        "o-tmf":               "UTF-8",
        "adminlang":           "en",
        "srclang":             "ko",
        "datatype":            "PlainText",
    }
    etree.SubElement(tmx, "header", **header_attr)
    body = etree.SubElement(tmx, "body")

    cid = creation_id or f"Excel2TMX_{datetime.now().strftime('%Y%m%d_%H%M')}"
    basename = os.path.basename(excel_path)
    target_lang_code = target_language.lower()

    memoq_ph = '<mq:rxt-req displaytext="desc" val="&amp;desc;" />'
    total_main = 0
    total_desc = 0

    for row in rows_data:
        src = str(row[0]).strip() if row[0] else ""
        tgt = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        ctx = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        desc_src = str(row[3]).strip() if has_desc and len(row) > 3 and row[3] else ""
        desc_tgt = str(row[4]).strip() if has_desc and len(row) > 4 and row[4] else ""

        # Main translation TU
        normal_ok = bool(tgt and ctx and not is_korean_text(tgt))
        if normal_ok:
            tu = etree.SubElement(body, "tu", creationid=cid, changeid=cid)
            tuv_ko = etree.SubElement(tu, "tuv", **{f"{{{XML_NS}}}lang": "ko"})
            etree.SubElement(tuv_ko, "seg").text = src
            tuv_tgt = etree.SubElement(tu, "tuv", **{f"{{{XML_NS}}}lang": target_lang_code})
            etree.SubElement(tuv_tgt, "seg").text = tgt
            etree.SubElement(tu, "prop", type="x-document").text = basename
            etree.SubElement(tu, "prop", type="x-context").text = ctx
            total_main += 1

        # Description TU
        desc_ok = bool(desc_src and desc_tgt and ctx and not is_korean_text(desc_tgt))
        if desc_ok:
            tu_d = etree.SubElement(body, "tu", creationid=cid, changeid=cid)
            tuv_ko_d = etree.SubElement(tu_d, "tuv", **{f"{{{XML_NS}}}lang": "ko"})
            tuv_tgt_d = etree.SubElement(tu_d, "tuv", **{f"{{{XML_NS}}}lang": target_lang_code})

            if postprocess:
                # MemoQ format: <ph> tag for &desc; marker
                seg_ko = etree.SubElement(tuv_ko_d, "seg")
                ph_ko = etree.SubElement(seg_ko, "ph")
                ph_ko.text = memoq_ph
                ph_ko.tail = desc_src

                seg_tgt = etree.SubElement(tuv_tgt_d, "seg")
                ph_tgt = etree.SubElement(seg_tgt, "ph")
                ph_tgt.text = memoq_ph
                ph_tgt.tail = desc_tgt
            else:
                # Plain format: &desc; prefix
                etree.SubElement(tuv_ko_d, "seg").text = "&desc;" + desc_src
                etree.SubElement(tuv_tgt_d, "seg").text = "&desc;" + desc_tgt

            etree.SubElement(tu_d, "prop", type="x-document").text = basename
            etree.SubElement(tu_d, "prop", type="x-context").text = ctx
            total_desc += 1

    total_tus = total_main + total_desc
    logger.info(f"  Created {total_tus} TU(s): {total_main} main + {total_desc} desc")

    # Serialize
    xml_bytes = etree.tostring(tmx, pretty_print=True, encoding='UTF-8', xml_declaration=True)
    xml_str = xml_bytes.decode('utf-8')

    if postprocess:
        logger.info("  Postprocessing: adding MemoQ bpt/ept/ph markup...")
        xml_str = postprocess_tmx_string(xml_str)

    # Output path
    if output_path is None:
        output_path = os.path.splitext(excel_path)[0] + ".tmx"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(xml_str)

    logger.info("=" * 60)
    logger.info("[Excel→TMX] Done -> %s (%d TUs)", os.path.basename(output_path), total_tus)
    return output_path
