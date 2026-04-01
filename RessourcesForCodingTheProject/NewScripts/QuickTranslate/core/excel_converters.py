"""
Excel Converters Module

Standalone Excel-to-text conversion utilities for the Other Tools tab:
  - Excel -> Wiki (Confluence table, active sheet only)
  - Excel -> Markdown table (active sheet only)
  - Excel -> HTML (entire workbook with tabs, sort/filter/resize)

Supports both FILE mode (single file) and FOLDER mode (all .xlsx/.xls in folder).
"""
from __future__ import annotations

import logging
import os
from typing import Callable, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# Visible indent character for plain-text output
INDENT_CHAR = '\u2506'  # ┆


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_excel_files(folder: str) -> List[str]:
    """Find all .xlsx/.xls files in folder (non-recursive, skip temp files)."""
    result = []
    for f in os.listdir(folder):
        if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            result.append(os.path.join(folder, f))
    return sorted(result)


def _get_indent(cell) -> int:
    """Safely extract indent level from a cell's alignment."""
    align = cell.alignment
    if align is None:
        return 0
    try:
        return int(align.indent) if align.indent else 0
    except (ValueError, TypeError):
        return 0


def _plain_cell(cell_value) -> str:
    """Convert cell value to plain text (for Wiki/Markdown).

    Newlines become spaces. No HTML escaping.
    """
    if cell_value is None:
        return ''
    s = str(cell_value)
    return s.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')


def _html_cell(cell_value) -> str:
    """Convert cell value to safe HTML (entity-escaped, newlines to <br/>)."""
    if cell_value is None:
        return ''
    s = str(cell_value)
    return (
        s.replace('&', '&amp;')
         .replace('<', '&lt;')
         .replace('>', '&gt;')
         .replace('\r\n', '<br/>')
         .replace('\r', '<br/>')
         .replace('\n', '<br/>')
    )


def _add_indent(text: str, indent_level: int) -> str:
    """Prefix text with visible indent chars for Wiki/Markdown output."""
    return INDENT_CHAR * indent_level + ' ' + text if indent_level else text


def _add_indent_css(indent_level: int) -> str:
    """Return CSS padding for HTML output."""
    return f'padding-left:{indent_level * 10}px;' if indent_level else ''


def _get_cell_bgcolor(cell) -> str:
    """Return background-color CSS if cell has a solid RGB fill colour."""
    fill = cell.fill
    if not fill or fill.fill_type != 'solid':
        return ''
    fg = getattr(fill, 'fgColor', None)
    if not fg or fg.type != 'rgb':
        return ''
    rgb = fg.rgb
    if not rgb or not isinstance(rgb, str):
        return ''
    if len(rgb) == 8:  # ARGB -> strip alpha
        rgb = rgb[2:]
    # Skip default/transparent fills
    if rgb.upper() in ('000000', 'FFFFFF', '00000000', 'FFFFFFFF'):
        return ''
    return f'background-color:#{rgb};'


def _excel_chars_to_px(chars: float) -> int:
    """Rough Excel char-width to pixel conversion."""
    return int(chars * 7 + 10)


def _calc_column_widths(ws) -> List[int]:
    """Decide pixel width for every column from workbook dims or content."""
    widths = []
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        excel_w = dim.width if dim and dim.width else None

        if excel_w:
            chars = excel_w
        else:
            longest = 0
            for cell in ws[letter]:
                longest = max(longest, len(str(cell.value or '')))
            chars = longest + 2

        px = _excel_chars_to_px(chars)
        px = max(60, min(px, 600))
        widths.append(px)
    return widths


def _is_empty_sheet(ws) -> bool:
    """Check if a worksheet has no data."""
    return not ws.max_row or ws.max_row < 1 or not ws.max_column or ws.max_column < 1


# ---------------------------------------------------------------------------
# Confluence Wiki converter (active sheet)
# ---------------------------------------------------------------------------

def _sheet_to_wiki(ws) -> str:
    """Convert active sheet to Confluence wiki table markup."""
    if _is_empty_sheet(ws):
        return ''
    headers = '||' + '||'.join(_plain_cell(h.value) for h in ws[1]) + '||\n'
    rows = ''
    for row in ws.iter_rows(min_row=2):
        cells = []
        for cell in row:
            val = _plain_cell(cell.value)
            val = _add_indent(val, _get_indent(cell))
            cells.append(val)
        rows += '|' + '|'.join(cells) + '|\n'
    return headers + rows


# ---------------------------------------------------------------------------
# Markdown converter (active sheet)
# ---------------------------------------------------------------------------

def _sheet_to_markdown(ws) -> str:
    """Convert active sheet to Markdown table."""
    if _is_empty_sheet(ws):
        return ''
    headers = '| ' + ' | '.join(_plain_cell(h.value) for h in ws[1]) + ' |\n'
    separator = '| ' + ' | '.join(['---'] * ws.max_column) + ' |\n'
    rows = ''
    for row in ws.iter_rows(min_row=2):
        cells = []
        for cell in row:
            val = _plain_cell(cell.value)
            val = _add_indent(val, _get_indent(cell))
            cells.append(val)
        rows += '| ' + ' | '.join(cells) + ' |\n'
    return headers + separator + rows


# ---------------------------------------------------------------------------
# HTML converter (entire workbook with tabs)
# Note: Requires full openpyxl mode (not read_only) for cell styles/alignment.
# ---------------------------------------------------------------------------

def _build_table_html(ws) -> str:
    """Build a <table> block with colours, indentation, sortable headers."""
    if _is_empty_sheet(ws):
        return '<p><em>Empty sheet</em></p>'

    out = ['<table>']

    col_widths = _calc_column_widths(ws)
    out.append('<colgroup>')
    for px in col_widths:
        out.append(f'<col style="width:{px}px;">')
    out.append('</colgroup>')

    header_cells = ''.join(
        f'<th class="sortable" onclick="headerClick(event,this)" '
        f'title="Click to sort | Ctrl+Click to set width" '
        f'style="{_get_cell_bgcolor(c)}">{_html_cell(c.value)}</th>'
        for c in ws[1]
    )
    out.append(f'<thead><tr>{header_cells}</tr></thead><tbody>')

    for row in ws.iter_rows(min_row=2):
        tds = []
        for cell in row:
            style = _add_indent_css(_get_indent(cell)) + _get_cell_bgcolor(cell)
            tds.append(f'<td style="{style}">{_html_cell(cell.value)}</td>')
        out.append(f'<tr>{"".join(tds)}</tr>')

    out.append('</tbody></table>')
    return '\n'.join(out)


_HTML_JS = r'''
<script>
function showSheet(i){
  document.querySelectorAll('.sheet-tabs li')
    .forEach((li,idx)=>li.classList.toggle('active',idx===i));
  document.querySelectorAll('.sheet')
    .forEach((s,idx)=>s.classList.toggle('active',idx===i));
}
function sortTable(th){
  const table=th.closest('table'),tbody=table.querySelector('tbody');
  const idx=Array.from(th.parentNode.children).indexOf(th);
  const asc=!th.classList.contains('asc');
  table.querySelectorAll('th.sortable').forEach(h=>h.classList.remove('asc','desc'));
  th.classList.add(asc?'asc':'desc');
  const rows=Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a,b)=>{
    const va=a.children[idx].innerText.trim(),vb=b.children[idx].innerText.trim();
    const na=parseFloat(va.replace(/[^0-9.\-]/g,'')),nb=parseFloat(vb.replace(/[^0-9.\-]/g,''));
    if(!isNaN(na)&&!isNaN(nb)) return asc?na-nb:nb-na;
    return asc?va.localeCompare(vb):vb.localeCompare(va);
  });
  rows.forEach(r=>tbody.appendChild(r));
}
document.querySelectorAll('.filter-box').forEach(inp=>{
  inp.addEventListener('input',function(){
    const val=this.value.toLowerCase();
    const tbody=this.nextElementSibling.querySelector('tbody');
    Array.from(tbody.rows).forEach(r=>{
      r.style.display=r.innerText.toLowerCase().includes(val)?'':'none';
    });
  });
});
function headerClick(e,th){
  if(e.ctrlKey){
    e.preventDefault();e.stopPropagation();
    const nw=prompt('Width (px):',th.offsetWidth);
    if(nw===null)return;
    const w=parseInt(nw,10);if(isNaN(w)||w<=0)return;
    const table=th.closest('table'),idx=Array.from(th.parentNode.children).indexOf(th);
    const col=table.querySelectorAll('col')[idx];
    if(col) col.style.width=w+'px';
    table.querySelectorAll('tr').forEach(tr=>{
      const c=tr.children[idx];if(c) c.style.width=w+'px';
    });
  }else{sortTable(th);}
}
</script>
'''

_HTML_CSS = '''<style>
body{font-family:Arial,Helvetica,sans-serif;font-size:14px;}
ul.sheet-tabs{list-style:none;padding:0;margin:0 0 10px 0;}
ul.sheet-tabs li{display:inline-block;padding:4px 10px;border:1px solid #ccc;
  border-bottom:none;cursor:pointer;background:#f8f8f8;}
ul.sheet-tabs li.active{background:#fff;font-weight:bold;}
.sheet{display:none;border:1px solid #ccc;padding:10px;}
.sheet.active{display:block;}
.filter-box{margin-bottom:6px;padding:4px 8px;width:250px;font-size:13px;}
table{border-collapse:collapse;width:100%;}
th,td{border:1px solid #ccc;padding:4px;vertical-align:top;word-break:break-word;}
th.sortable{cursor:pointer;position:relative;}
th.sortable:after{content:"\\2195";font-size:11px;margin-left:4px;color:#666;}
th.sortable.asc:after{content:"\\25B2";}
th.sortable.desc:after{content:"\\25BC";}
</style>'''


def _workbook_to_html(wb) -> str:
    """Convert entire workbook to self-contained HTML with sheet tabs."""
    lines = [
        '<!DOCTYPE html>',
        '<html lang="en"><head><meta charset="UTF-8">',
        '<title>Excel \u2192 HTML</title>',
        _HTML_CSS,
        '</head><body>',
        '<p style="font-size:13px;margin:6px 0 12px 0;">'
        'Click a column header to sort. Ctrl+Click a header to set width.</p>',
        '<ul class="sheet-tabs">',
    ]

    for idx, ws in enumerate(wb.worksheets):
        active = ' active' if idx == 0 else ''
        lines.append(
            f'<li class="{active}" onclick="showSheet({idx})">'
            f'{_html_cell(ws.title)}</li>')
    lines.append('</ul>')

    for idx, ws in enumerate(wb.worksheets):
        cls = 'sheet active' if idx == 0 else 'sheet'
        lines.append(f'<div class="{cls}" id="sheet{idx}">')
        lines.append(
            '<input type="text" class="filter-box" '
            'placeholder="Type to filter rows\u2026" />')
        lines.append(_build_table_html(ws))
        lines.append('</div>')

    lines.append(_HTML_JS)
    lines.append('</body></html>')
    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Public API (called from GUI)
# ---------------------------------------------------------------------------

# Conversion specs: (sheet converter func, uses whole workbook, output suffix)
_CONVERTERS = {
    'wiki':     (_sheet_to_wiki,     False, '_wiki.txt'),
    'markdown': (_sheet_to_markdown, False, '_markdown.txt'),
    'html':     (_workbook_to_html,  True,  '.html'),
}


def _convert_excel(
    fmt: str,
    input_path: str,
    is_folder: bool = False,
    progress_callback: Optional[Callable] = None,
) -> dict:
    """Shared conversion driver for all three formats.

    Args:
        fmt: One of 'wiki', 'markdown', 'html'.
        input_path: Folder path (is_folder=True) or single file path.
        is_folder: True = process all Excel files in folder.
        progress_callback: Optional callable(str) for log messages.

    Returns:
        dict with 'files', 'output_files'
    """
    if openpyxl is None:
        raise ImportError("openpyxl required for Excel reading")

    convert_fn, use_workbook, suffix = _CONVERTERS[fmt]

    files = _get_excel_files(input_path) if is_folder else [input_path]
    if not files:
        raise FileNotFoundError("No Excel files found")

    output_files = []
    for fpath in files:
        if progress_callback:
            progress_callback(f"Converting {os.path.basename(fpath)}...")
        wb = openpyxl.load_workbook(fpath)
        try:
            text = convert_fn(wb) if use_workbook else convert_fn(wb.active)
        finally:
            wb.close()
        out = os.path.splitext(fpath)[0] + suffix
        with open(out, 'w', encoding='utf-8') as fh:
            fh.write(text)
        output_files.append(out)

    return {'files': len(files), 'output_files': output_files}


def excel_to_wiki(input_path: str, is_folder: bool = False,
                  progress_callback: Optional[Callable] = None) -> dict:
    """Convert Excel to Confluence Wiki table(s)."""
    return _convert_excel('wiki', input_path, is_folder, progress_callback)


def excel_to_markdown(input_path: str, is_folder: bool = False,
                      progress_callback: Optional[Callable] = None) -> dict:
    """Convert Excel to Markdown table(s)."""
    return _convert_excel('markdown', input_path, is_folder, progress_callback)


def excel_to_html(input_path: str, is_folder: bool = False,
                  progress_callback: Optional[Callable] = None) -> dict:
    """Convert Excel to HTML with tabs, sort, filter, colour, indent."""
    return _convert_excel('html', input_path, is_folder, progress_callback)
