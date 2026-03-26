"""
Quick Converters Module

Standalone file/folder conversion utilities for the Other Tools tab:
  - XML → Excel
  - Excel → XML
  - Concatenate XML (folder → single XML)

TMX ↔ Excel conversions reuse existing functions in tmx_tools.py.
"""
from __future__ import annotations

import copy
import logging
import os
import re
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    from lxml import etree
except ImportError:
    etree = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_all_files(folder: str, ext: str) -> List[str]:
    """Recursively find all files with given extension in folder."""
    result = []
    for root, _dirs, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(ext.lower()):
                result.append(os.path.join(root, f))
    return sorted(result)


# Regex for raw attribute extraction (preserves original values exactly)
_RE_LOCSTR_TAG = re.compile(r"<LocStr\b([^>]*)>", re.DOTALL | re.IGNORECASE)
_RE_ATTR = re.compile(r'(\w+)\s*=\s*(?:"([^"]*)"|\'([^\']*)\')', re.DOTALL)

# Standard columns in order
_PRIMARY_ATTRS = ["StrOrigin", "Str", "StringId", "DescOrigin", "Desc"]
_HEADER_NAMES = ["StrOrigin", "Str", "StringId", "DescOrigin", "DescText"]


# ---------------------------------------------------------------------------
# XML → Excel
# ---------------------------------------------------------------------------

def xml_to_excel(
    input_path: str,
    output_path: str,
    is_folder: bool = True,
    progress_callback=None,
) -> dict:
    """Convert XML file(s) to Excel.

    Extracts all <LocStr> attributes. Standard 5 columns + any extra attributes.

    Args:
        input_path: Folder path (is_folder=True) or single file path
        output_path: Output .xlsx file path
        is_folder: True = scan folder recursively, False = single file
        progress_callback: Optional callable(str) for log messages

    Returns:
        dict with keys: 'rows', 'files', 'output'
    """
    if xlsxwriter is None:
        raise ImportError("xlsxwriter required for Excel output")

    # Get XML files
    if is_folder:
        xml_files = _get_all_files(input_path, ".xml")
    else:
        xml_files = [input_path]

    if not xml_files:
        raise FileNotFoundError("No XML files found")

    if progress_callback:
        progress_callback(f"Found {len(xml_files)} XML file(s)")

    # Pass 1: discover all attribute names
    all_attrs = set()
    for xml_path in xml_files:
        try:
            with open(xml_path, "r", encoding="utf-8", errors="ignore") as f:
                raw = f.read()
        except Exception:
            continue
        for m in _RE_LOCSTR_TAG.finditer(raw):
            blob = m.group(1).rstrip("/").strip()
            for a in _RE_ATTR.finditer(blob):
                all_attrs.add(a.group(1))

    extra_attrs = sorted(a for a in all_attrs if a not in _PRIMARY_ATTRS)
    headers = _HEADER_NAMES + extra_attrs

    # Pass 2: extract rows
    rows = []
    for i, xml_path in enumerate(xml_files):
        try:
            with open(xml_path, "r", encoding="utf-8", errors="ignore") as f:
                raw = f.read()
        except Exception:
            continue
        for m in _RE_LOCSTR_TAG.finditer(raw):
            blob = m.group(1).rstrip("/").strip()
            attrs = {}
            for a in _RE_ATTR.finditer(blob):
                key = a.group(1)
                val = a.group(2) if a.group(2) is not None else a.group(3)
                attrs[key] = val
            row = [
                attrs.get("StrOrigin", ""),
                attrs.get("Str", ""),
                attrs.get("StringId", ""),
                attrs.get("DescOrigin", ""),
                attrs.get("Desc", ""),
            ]
            for extra in extra_attrs:
                row.append(attrs.get(extra, ""))
            rows.append(row)

        if progress_callback and (i + 1) % 10 == 0:
            progress_callback(f"Processed {i + 1}/{len(xml_files)} files")

    if not rows:
        raise ValueError("No <LocStr> elements found in input")

    # Write Excel
    wb = xlsxwriter.Workbook(output_path)
    ws = wb.add_worksheet("XMLData")

    header_fmt = wb.add_format({
        "bold": True, "bg_color": "#4472C4", "font_color": "#FFFFFF",
        "border": 1,
    })
    text_fmt = wb.add_format({"num_format": "@"})

    for col, h in enumerate(headers):
        ws.write(0, col, h, header_fmt)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row):
            # StringId column = text format
            if c_idx == 2:
                ws.write_string(r_idx, c_idx, str(val), text_fmt)
            else:
                ws.write(r_idx, c_idx, val)

    ws.set_column(0, 0, 40)  # StrOrigin
    ws.set_column(1, 1, 40)  # Str
    ws.set_column(2, 2, 22)  # StringId
    ws.set_column(3, 3, 40)  # DescOrigin
    ws.set_column(4, 4, 40)  # DescText
    ws.autofilter(0, 0, len(rows), len(headers) - 1)
    ws.freeze_panes(1, 0)

    wb.close()

    if progress_callback:
        progress_callback(f"XML→Excel: {len(rows)} rows from {len(xml_files)} file(s) → {os.path.basename(output_path)}")

    return {"rows": len(rows), "files": len(xml_files), "output": output_path}


# ---------------------------------------------------------------------------
# Excel → XML
# ---------------------------------------------------------------------------

def excel_to_xml(
    input_path: str,
    output_path: str,
    progress_callback=None,
) -> dict:
    """Convert Excel file back to XML format.

    Reads columns: StrOrigin, Str, StringId, DescOrigin, DescText.
    Falls back to positional columns if headers not found.

    Returns:
        dict with keys: 'rows', 'output'
    """
    if etree is None:
        raise ImportError("lxml required for XML output")
    if openpyxl is None:
        raise ImportError("openpyxl required for Excel reading")

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not all_rows:
        raise ValueError("Empty Excel file")

    # Detect headers
    header = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    header_lower = [h.lower() for h in header]

    # Map column indices
    col_map = {}
    for target, candidates in [
        ("StrOrigin", ["strorigin", "str_origin"]),
        ("Str", ["str", "correction"]),
        ("StringId", ["stringid", "string_id"]),
        ("DescOrigin", ["descorigin", "desc_origin"]),
        ("Desc", ["desctext", "desc_text", "desc"]),
    ]:
        for cand in candidates:
            if cand in header_lower:
                col_map[target] = header_lower.index(cand)
                break

    # Fallback to positional
    data_start = 1
    if "StrOrigin" not in col_map:
        col_map = {"StrOrigin": 0, "Str": 1, "StringId": 2, "DescOrigin": 3, "Desc": 4}
        data_start = 0  # No header row

    root_elem = etree.Element("Root")
    count = 0

    for row in all_rows[data_start:]:
        if not row:
            continue

        def _get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        so = _get("StrOrigin")
        st = _get("Str")
        sid = _get("StringId")
        do = _get("DescOrigin")
        dt = _get("Desc")

        if not so and not st:
            continue

        loc = etree.SubElement(root_elem, "LocStr")
        loc.set("StrOrigin", so)
        loc.set("Str", st)
        loc.set("StringId", sid)
        if do:
            loc.set("DescOrigin", do)
        if dt:
            loc.set("Desc", dt)
        count += 1

    if count == 0:
        raise ValueError("No data rows found in Excel")

    xml_bytes = etree.tostring(root_elem, pretty_print=True, encoding="utf-8",
                               xml_declaration=True)
    with open(output_path, "wb") as f:
        f.write(xml_bytes)

    if progress_callback:
        progress_callback(f"Excel→XML: {count} entries → {os.path.basename(output_path)}")

    return {"rows": count, "output": output_path}


# ---------------------------------------------------------------------------
# Concatenate XML
# ---------------------------------------------------------------------------

def concatenate_xmls(
    input_folder: str,
    output_path: str,
    progress_callback=None,
) -> dict:
    """Merge all <LocStr> elements from all XML files in folder into one XML.

    Recursively walks input folder, finds all .xml files, deep-copies every
    <LocStr> element into a single <root> element.

    Returns:
        dict with keys: 'total_locs', 'files', 'output'
    """
    if etree is None:
        raise ImportError("lxml required for XML processing")

    xml_files = _get_all_files(input_folder, ".xml")
    if not xml_files:
        raise FileNotFoundError("No XML files found in folder")

    if progress_callback:
        progress_callback(f"Found {len(xml_files)} XML file(s) to concatenate")

    parser = etree.XMLParser(recover=True, resolve_entities=False)
    root = etree.Element("root", FileName=os.path.basename(output_path))
    total_locs = 0

    for i, xml_file in enumerate(xml_files):
        try:
            tree = etree.parse(xml_file, parser)
            xml_root = tree.getroot()
            for loc in xml_root.iter("LocStr"):
                root.append(copy.deepcopy(loc))
                total_locs += 1
        except Exception as exc:
            logger.warning("Failed to parse %s: %s", xml_file, exc)

        if progress_callback and (i + 1) % 10 == 0:
            progress_callback(f"Concatenated {i + 1}/{len(xml_files)} files ({total_locs} LocStrs)")

    if total_locs == 0:
        raise ValueError("No <LocStr> elements found in any XML file")

    xml_bytes = etree.tostring(root, pretty_print=True, encoding="utf-8",
                               xml_declaration=False)
    with open(output_path, "wb") as f:
        f.write(xml_bytes)

    if progress_callback:
        progress_callback(f"Concatenate: {total_locs} LocStrs from {len(xml_files)} files → {os.path.basename(output_path)}")

    return {"total_locs": total_locs, "files": len(xml_files), "output": output_path}
