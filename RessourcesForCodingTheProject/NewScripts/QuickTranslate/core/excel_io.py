"""
Excel I/O Operations.

Read input Excel files and write output Excel files.
Uses patterns from LanguageDataExporter for robustness.
"""

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, numbers

from .korean_detection import is_korean_text
from .text_utils import normalize_text, normalize_nospace, is_formula_text

# ---------------------------------------------------------------------------
# Formula safeguard — detect Excel formulas, error values, and non-str types
# ---------------------------------------------------------------------------

try:
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
    _FORMULA_TYPES = (ArrayFormula, DataTableFormula)
except ImportError:
    _FORMULA_TYPES = ()

def _is_bad_cell_type(value) -> Optional[str]:
    """Check raw cell.value type. Returns reason string if bad, None if OK."""
    if value is None:
        return None
    if isinstance(value, str):
        return None
    if _FORMULA_TYPES and isinstance(value, _FORMULA_TYPES):
        return f'Excel formula object ({type(value).__name__})'
    if isinstance(value, bool):
        return f'Boolean value ({value})'
    if isinstance(value, (int, float)):
        return f'Numeric value ({value})'
    return f'Unexpected type ({type(value).__name__}: {str(value)[:50]})'


def _is_bad_cell_text(text: str) -> Optional[str]:
    """Check string cell value. Delegates to shared is_formula_text()."""
    return is_formula_text(text)


def _detect_column_indices(ws) -> Dict[str, int]:
    """
    Detect column indices from header row (CASE INSENSITIVE).

    From LanguageDataExporter - allows flexible column ordering.

    Args:
        ws: Worksheet to scan

    Returns:
        Dict mapping lowercase header name to column index (1-based)
    """
    indices = {}
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if first_row and first_row[0]:
        for col, header in enumerate(first_row[0], 1):
            if header:
                # Store lowercase key for case-insensitive lookup
                indices[str(header).strip().lower()] = col
    return indices


def detect_excel_columns(excel_path: Path) -> Dict[str, bool]:
    """
    Lightweight header-only scan: detect which columns exist in an Excel file.

    Opens the file, reads ONLY the first row, checks for known column names.
    Does NOT read any data rows.

    Returns:
        Dict with keys: has_stringid, has_strorigin, has_correction,
                        has_eventname, has_dialogvoice, has_desc, has_descorigin
    """
    result = {
        "has_stringid": False,
        "has_strorigin": False,
        "has_correction": False,
        "has_eventname": False,
        "has_dialogvoice": False,
        "has_desc": False,
        "has_descorigin": False,
    }

    wb = load_workbook(excel_path, read_only=True)
    try:
        ws = wb.active
        col_indices = _detect_column_indices(ws)

        result["has_stringid"] = (
            "stringid" in col_indices or "string_id" in col_indices
        )
        result["has_strorigin"] = (
            "strorigin" in col_indices or "str_origin" in col_indices
        )
        result["has_correction"] = (
            "correction" in col_indices or "corrected" in col_indices
        )
        result["has_eventname"] = (
            "eventname" in col_indices or "event_name" in col_indices
            or "soundeventname" in col_indices
        )
        result["has_dialogvoice"] = (
            "dialogvoice" in col_indices or "dialog_voice" in col_indices
        )
        result["has_descorigin"] = (
            "descorigin" in col_indices or "desc_origin" in col_indices
        )
        result["has_desc"] = (
            "desc" in col_indices or "desctext" in col_indices
            or "desc_text" in col_indices or "desccorrection" in col_indices
        )

        return result
    finally:
        wb.close()


def read_korean_input(excel_path: Path) -> List[str]:
    """
    Read Korean text from Column 1 of input Excel file.

    Args:
        excel_path: Path to input Excel file

    Returns:
        List of Korean text strings (trimmed, non-empty)
    """
    wb = load_workbook(excel_path, read_only=True)
    try:
        ws = wb.active
        korean_texts = []
        for row in ws.iter_rows(min_row=1, max_col=1):
            cell_value = row[0].value
            if cell_value:
                korean_texts.append(str(cell_value).strip())
        return korean_texts
    finally:
        wb.close()


def read_corrections_from_excel(
    excel_path: Path,
    has_header: bool = True,
    formula_report: Optional[list] = None,
) -> List[Dict]:
    """
    Read corrections from Excel file for transfer mode.

    Uses case-insensitive column detection from headers. NO positional fallbacks.
    Validates that required columns exist before processing any rows.

    Supported column combinations:
    - StringID + Correction (+ optional StrOrigin): direct StringID mode
    - EventName + Correction (+ optional DialogVoice, StrOrigin): waterfall resolution
    - Both StringID and EventName: per-row priority (StringID wins when present)

    Args:
        excel_path: Path to input Excel file
        has_header: If True, detect columns from header row (case-insensitive)
        formula_report: Optional list to collect formula/error skips.
            Pass [] to collect; each entry is a dict with row, column,
            string_id, reason. Pass None (default) to skip silently.

    Returns:
        List of correction dicts with keys: string_id, str_origin, corrected

    Raises:
        ValueError: If required columns are not found in headers
    """
    wb = load_workbook(excel_path, read_only=True)
    try:
        ws = wb.active
        corrections = []
        start_row = 2 if has_header else 1

        # Detect ALL columns from header row (case-insensitive) — NO positional fallbacks
        stringid_col = None
        strorigin_col = None
        correction_col = None
        eventname_col = None
        dialogvoice_col = None
        desc_col = None
        descorigin_col = None

        if has_header:
            col_indices = _detect_column_indices(ws)

            stringid_col = col_indices.get("stringid", col_indices.get("string_id", None))
            strorigin_col = col_indices.get("strorigin", col_indices.get("str_origin", None))
            correction_col = col_indices.get("correction", col_indices.get("corrected", None))
            eventname_col = col_indices.get("eventname", col_indices.get("event_name",
                            col_indices.get("soundeventname", None)))
            dialogvoice_col = col_indices.get("dialogvoice", col_indices.get("dialog_voice", None))
            # Desc/DescOrigin columns (optional — voice direction descriptions)
            descorigin_col = col_indices.get("descorigin", col_indices.get("desc_origin", None))
            desc_col = col_indices.get("desc", col_indices.get("desctext",
                       col_indices.get("desc_text", col_indices.get("desccorrection", None))))

            # ── Column validation: fail fast with clear errors ──
            if correction_col is None:
                raise ValueError(
                    f"Missing required 'Correction' column in {excel_path.name}. "
                    f"Found headers: {list(col_indices.keys())}"
                )

            if stringid_col is None and eventname_col is None:
                raise ValueError(
                    f"Missing both 'StringID' and 'EventName' columns in {excel_path.name}. "
                    f"Need at least one. Found headers: {list(col_indices.keys())}"
                )

            # Only log when EventName resolution is involved (non-obvious behavior)
            if stringid_col and eventname_col:
                logger.info(
                    "EventName column detected alongside StringID. "
                    "Per-row priority: StringID when present, EventName as fallback."
                )
            elif eventname_col and dialogvoice_col:
                logger.info(
                    "No StringID column — resolving EventName → StringID "
                    "via DialogVoice stripping + keyword extraction."
                )
            elif eventname_col:
                logger.info(
                    "No StringID column — resolving EventName → StringID "
                    "via keyword extraction + export file lookup."
                )

        for row in ws.iter_rows(min_row=start_row):
            try:
                string_id = row[stringid_col - 1].value if stringid_col is not None and stringid_col <= len(row) else None
                str_origin = row[strorigin_col - 1].value if strorigin_col is not None and strorigin_col <= len(row) else None
                corrected = row[correction_col - 1].value if correction_col is not None and correction_col <= len(row) else None

                # Formula safeguard Layer 1: type check BEFORE str coercion
                bad_type = _is_bad_cell_type(corrected)
                if bad_type:
                    logger.debug("Row %s Correction skipped: %s", row[0].row, bad_type)
                    if formula_report is not None:
                        formula_report.append({
                            'row': row[0].row, 'column': 'Correction',
                            'string_id': str(string_id or ''), 'reason': bad_type,
                        })
                    continue

                eventname = None
                if eventname_col is not None:
                    eventname = row[eventname_col - 1].value if eventname_col <= len(row) else None
                dialogvoice = None
                if dialogvoice_col is not None:
                    dialogvoice = row[dialogvoice_col - 1].value if dialogvoice_col <= len(row) else None

                # Accept row if EITHER string_id or eventname is present
                # Use 'is not None' to handle numeric 0 as a valid StringID
                has_id = string_id is not None and str(string_id).strip()
                has_eventname = eventname is not None and str(eventname).strip()

                if not (has_id or has_eventname) or not corrected:
                    continue

                corrected_str = str(corrected).strip()

                # Formula safeguard Layer 2: string check AFTER str coercion
                bad_text = _is_bad_cell_text(corrected_str)
                if bad_text:
                    logger.debug("Row %s Correction skipped: %s", row[0].row, bad_text)
                    if formula_report is not None:
                        formula_report.append({
                            'row': row[0].row, 'column': 'Correction',
                            'string_id': str(string_id or ''), 'reason': bad_text,
                        })
                    continue

                # Skip entries where the "correction" is still Korean (untranslated)
                if is_korean_text(corrected_str):
                    continue

                entry = {
                    "string_id": str(string_id).strip() if has_id else "",
                    "str_origin": normalize_text(str_origin) if str_origin else "",
                    "corrected": corrected_str,  # Preserve linebreaks! Don't normalize output text
                }

                # Read Desc/DescOrigin if columns exist (voice direction descriptions)
                if descorigin_col is not None:
                    do_val = row[descorigin_col - 1].value if descorigin_col <= len(row) else None
                    if do_val is not None and str(do_val).strip():
                        entry["desc_origin"] = str(do_val).strip()
                if desc_col is not None:
                    d_val = row[desc_col - 1].value if desc_col <= len(row) else None
                    # Formula safeguard: neutralize bad Desc values (don't skip row)
                    bad = _is_bad_cell_type(d_val) or (isinstance(d_val, str) and _is_bad_cell_text(d_val.strip()))
                    if bad:
                        logger.debug("Row %s Desc skipped: %s", row[0].row, bad)
                        if formula_report is not None:
                            formula_report.append({
                                'row': row[0].row, 'column': 'Desc',
                                'string_id': str(string_id or ''), 'reason': bad,
                            })
                        d_val = None  # Neutralize — prevents downstream str(d_val)
                    if d_val is not None and str(d_val).strip():
                        desc_str = str(d_val).strip()
                        if not is_korean_text(desc_str):
                            entry["desc_corrected"] = desc_str

                # Per-row priority: StringID takes precedence over EventName
                # EventName is fallback when StringID is empty for that row
                if not has_id and has_eventname:
                    entry["_source_eventname"] = str(eventname).strip()
                    # Attach DialogVoice for waterfall Step 1 resolution
                    if dialogvoice is not None:
                        entry["_source_dialogvoice"] = str(dialogvoice).strip()

                # Flag when no EventName column exists — transfer pipeline uses
                # this to auto-detect EventNames mixed into StringID column
                if not eventname_col and has_id:
                    entry["_no_eventname_col"] = True

                # Always stash original EventName/DialogVoice as recovery metadata
                # Used by EventName recovery pass if StringID fails to match
                if has_eventname:
                    entry["_original_eventname"] = str(eventname).strip()
                if dialogvoice is not None:
                    entry["_original_dialogvoice"] = str(dialogvoice).strip()

                corrections.append(entry)
            except (IndexError, AttributeError):
                continue

        return corrections
    finally:
        wb.close()


def get_ordered_languages(available_langs: List[str], language_order: List[str] = None) -> List[str]:
    """
    Get languages in preferred order.

    Args:
        available_langs: List of available language codes
        language_order: Preferred order (optional, uses default if None)

    Returns:
        Ordered list of language codes (excludes KOR)
    """
    if language_order is None:
        language_order = [
            "kor", "eng", "fre", "ger", "spa", "por", "ita", "rus",
            "tur", "pol", "zho-cn", "zho-tw", "jpn", "tha", "vie", "ind", "msa"
        ]

    ordered = []
    for lang in language_order:
        if lang in available_langs and lang != "kor":
            ordered.append(lang)

    for lang in available_langs:
        if lang not in ordered and lang != "kor":
            ordered.append(lang)

    return ordered


def write_output_excel(
    output_path: Path,
    korean_inputs: List[str],
    matches_per_input: List[List[str]],
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str],
    language_names: Dict[str, str] = None,
    stats: Dict[str, int] = None,
    match_type: str = "substring",
):
    """
    Write output Excel file with translations and summary.

    Args:
        output_path: Path to output Excel file
        korean_inputs: List of Korean input texts
        matches_per_input: List of StringID lists for each input
        translation_lookup: Dict mapping lang_code to {StringID: translation}
        available_langs: List of available language codes
        language_names: Optional mapping of lang_code to display name
        stats: Optional statistics dict for Summary sheet
        match_type: Match type used (for Summary sheet)
    """
    if language_names is None:
        language_names = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    ordered_langs = get_ordered_languages(available_langs)

    # Header row - add Status and StringID columns
    headers = ["KOR (Input)", "Status", "StringID"] + [language_names.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    from .matching import format_multiple_matches

    for row_idx, (korean_text, string_ids) in enumerate(zip(korean_inputs, matches_per_input), start=2):
        ws.cell(row=row_idx, column=1, value=korean_text)

        # Status column
        if not string_ids:
            status = "NOT FOUND"
            status_cell = ws.cell(row=row_idx, column=2, value=status)
            status_cell.font = Font(color="FF0000")  # Red
        elif len(string_ids) == 1:
            status = "MATCHED"
            status_cell = ws.cell(row=row_idx, column=2, value=status)
            status_cell.font = Font(color="008000")  # Green
        else:
            status = f"MULTI ({len(string_ids)})"
            status_cell = ws.cell(row=row_idx, column=2, value=status)
            status_cell.font = Font(color="FF8C00")  # Orange

        # StringID column - force TEXT format to prevent scientific notation
        if string_ids:
            if len(string_ids) == 1:
                sid_cell = ws.cell(row=row_idx, column=3, value=string_ids[0])
                sid_cell.number_format = numbers.FORMAT_TEXT  # Prevent scientific notation
            else:
                sid_text = "\n".join(f"{i+1}. {sid}" for i, sid in enumerate(string_ids))
                sid_cell = ws.cell(row=row_idx, column=3, value=sid_text)
                sid_cell.alignment = Alignment(wrap_text=True, vertical='top')
                sid_cell.number_format = numbers.FORMAT_TEXT

        # Translation columns
        for col_idx, lang_code in enumerate(ordered_langs, start=4):
            if not string_ids:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue

            translations = []
            for sid in string_ids:
                trans = translation_lookup.get(lang_code, {}).get(sid, "")
                if trans:
                    translations.append(trans)

            cell_value = format_multiple_matches(translations)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if "\n" in cell_value:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 40  # KOR Input
    ws.column_dimensions['B'].width = 12  # Status
    ws.column_dimensions['C'].width = 30  # StringID
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    # Create Summary sheet
    _create_summary_sheet(wb, stats, match_type, len(korean_inputs), matches_per_input)

    wb.save(output_path)


def _create_summary_sheet(
    wb: Workbook,
    stats: Dict[str, int],
    match_type: str,
    total_inputs: int,
    matches_per_input: List[List[str]],
):
    """Create Summary sheet with statistics."""
    ws = wb.create_sheet(title="Summary", index=0)

    # Calculate stats if not provided
    if stats is None:
        stats = {
            "total": total_inputs,
            "matched": sum(1 for m in matches_per_input if len(m) == 1),
            "no_match": sum(1 for m in matches_per_input if len(m) == 0),
            "multi_match": sum(1 for m in matches_per_input if len(m) > 1),
            "total_matches": sum(len(m) for m in matches_per_input),
        }

    # Title
    ws.cell(row=1, column=1, value="QuickTranslate Report Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    # Match type
    ws.cell(row=3, column=1, value="Match Type:")
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2, value=match_type.upper())

    # Statistics table
    stat_rows = [
        ("Total Inputs", stats.get("total", total_inputs)),
        ("Matched (1 match)", stats.get("matched", 0)),
        ("Multiple Matches", stats.get("multi_match", 0)),
        ("Not Found", stats.get("no_match", 0)),
        ("Empty/Skipped", stats.get("empty_input", 0) + stats.get("skipped", 0)),
        ("Total StringIDs Found", stats.get("total_matches", 0)),
    ]

    ws.cell(row=5, column=1, value="Statistic")
    ws.cell(row=5, column=2, value="Count")
    ws.cell(row=5, column=3, value="Percentage")
    for col in range(1, 4):
        ws.cell(row=5, column=col).font = Font(bold=True)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

    total = stats.get("total", total_inputs) or 1  # Avoid division by zero

    for row_idx, (label, value) in enumerate(stat_rows, start=6):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label not in ("Total Inputs", "Total StringIDs Found"):
            pct = (value / total * 100) if total > 0 else 0
            ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

    # Color coding
    ws.cell(row=6, column=2).font = Font(bold=True)  # Total
    ws.cell(row=7, column=2).font = Font(color="008000")  # Matched - green
    ws.cell(row=8, column=2).font = Font(color="FF8C00")  # Multi - orange
    ws.cell(row=9, column=2).font = Font(color="FF0000")  # Not found - red

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def write_stringid_lookup_excel(
    output_path: Path,
    string_id: str,
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str],
    language_names: Dict[str, str] = None,
):
    """
    Write output Excel for a single StringID lookup.

    Output format: StringID | ENG | FRE | GER | ...
    """
    if language_names is None:
        language_names = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "StringID Lookup"

    ordered_langs = get_ordered_languages(available_langs)

    # Header row
    headers = ["StringID"] + [language_names.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data row - force TEXT format for StringID (prevent scientific notation)
    sid_cell = ws.cell(row=2, column=1, value=string_id)
    sid_cell.number_format = numbers.FORMAT_TEXT

    for col_idx, lang_code in enumerate(ordered_langs, start=2):
        trans = translation_lookup.get(lang_code, {}).get(string_id, "")
        cell = ws.cell(row=2, column=col_idx, value=trans)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    wb.save(output_path)


def write_folder_translation_excel(
    output_path: Path,
    string_map: Dict[str, str],
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str],
    language_names: Dict[str, str] = None,
):
    """
    Write Excel with one sheet per language.

    Columns: StrOrigin | English | Translation | StringID

    "NO TRANSLATION" if translation is empty or contains Korean characters.
    """
    if language_names is None:
        language_names = {}

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ordered_langs = get_ordered_languages(available_langs)

    # Get English lookup for reference
    eng_lookup = translation_lookup.get("eng", {})

    for lang_code in ordered_langs:
        lang_name = language_names.get(lang_code, lang_code.upper())
        ws = wb.create_sheet(title=lang_name)
        lang_lookup = translation_lookup.get(lang_code, {})

        # Header row
        headers = ["StrOrigin", "English", lang_name, "StringID"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row_idx = 2
        for string_id, str_origin in string_map.items():
            # StrOrigin
            ws.cell(row=row_idx, column=1, value=str_origin)

            # English
            eng_trans = eng_lookup.get(string_id, "")
            if not eng_trans or is_korean_text(eng_trans):
                eng_trans = "NO TRANSLATION"
            ws.cell(row=row_idx, column=2, value=eng_trans)

            # Target language translation
            lang_trans = lang_lookup.get(string_id, "")
            # Only mark as NO TRANSLATION if empty or if non-KOR column has Korean text
            if not lang_trans or (lang_code != "kor" and is_korean_text(lang_trans)):
                lang_trans = "NO TRANSLATION"
            cell = ws.cell(row=row_idx, column=3, value=lang_trans)
            if "\n" in lang_trans:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            # StringID - force TEXT format (prevent scientific notation)
            sid_cell = ws.cell(row=row_idx, column=4, value=string_id)
            sid_cell.number_format = numbers.FORMAT_TEXT

            row_idx += 1

        # Column widths
        ws.column_dimensions['A'].width = 50  # StrOrigin
        ws.column_dimensions['B'].width = 40  # English
        ws.column_dimensions['C'].width = 40  # Translation
        ws.column_dimensions['D'].width = 25  # StringID

    wb.save(output_path)


def write_reverse_lookup_excel(
    output_path: Path,
    input_texts: List[str],
    stringid_map: Dict[str, str],  # input_text -> StringID
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str],
    language_names: Dict[str, str] = None,
):
    """
    Write Excel with all languages in columns.

    Columns: Input | KOR | ENG | FRE | GER | ...
    """
    if language_names is None:
        language_names = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reverse Lookup"

    # Order: KOR first, then others
    ordered_langs = ["kor"] + get_ordered_languages(available_langs)

    # Header row
    headers = ["Input"] + [language_names.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, input_text in enumerate(input_texts, start=2):
        # Write input text
        ws.cell(row=row_idx, column=1, value=input_text)

        string_id = stringid_map.get(input_text)
        if not string_id:
            # NOT FOUND - write "NOT FOUND" in KOR column, leave rest empty
            ws.cell(row=row_idx, column=2, value="NOT FOUND")
            continue

        # For each language, get translation
        for col_idx, lang_code in enumerate(ordered_langs, start=2):
            trans = translation_lookup.get(lang_code, {}).get(string_id, "")
            # Only mark as NO TRANSLATION if empty or if non-KOR column has Korean text
            if not trans or (lang_code != "kor" and is_korean_text(trans)):
                trans = "NO TRANSLATION"
            cell = ws.cell(row=row_idx, column=col_idx, value=trans)
            if "\n" in trans:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 50  # Input
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    wb.save(output_path)


# =============================================================================
# Excel Target Merge Functions (for TRANSFER mode)
# =============================================================================


def _convert_linebreaks_for_excel(txt: str) -> str:
    """
    Normalize linebreak variants to <br/> for Excel cells.

    Unified format: <br/> everywhere (XML and Excel).
    This prevents conversion bugs when corrections are transferred back to XML.
    Normalizes escaped/variant forms to the canonical <br/>.
    """
    if not txt:
        return txt
    txt = txt.replace('&lt;br/&gt;', '<br/>')
    txt = txt.replace('&lt;br /&gt;', '<br/>')
    txt = txt.replace('<br />', '<br/>')
    txt = txt.replace('\\n', '<br/>')
    txt = txt.replace('\n', '<br/>')
    return txt


def _insert_str_column(ws, strorigin_col: int) -> int:
    """
    Insert a 'Str' column right after the StrOrigin column, shifting existing columns right.

    Args:
        ws: openpyxl worksheet (read-write mode)
        strorigin_col: 1-based column index of StrOrigin

    Returns:
        1-based column index of the newly inserted Str column
    """
    str_col = strorigin_col + 1
    ws.insert_cols(str_col)
    ws.cell(row=1, column=str_col, value="Str")
    ws.cell(row=1, column=str_col).font = Font(bold=True)
    logger.info(f"Auto-inserted 'Str' column at position {str_col} (after StrOrigin at {strorigin_col})")
    return str_col


def merge_corrections_to_excel(
    excel_path: Path,
    corrections: List[Dict],
    match_mode: str = "strict",
    dry_run: bool = False,
    only_untranslated: bool = False,
    stringid_to_category: Optional[Dict[str, str]] = None,
    stringid_to_subfolder: Optional[Dict[str, str]] = None,
    stringid_all_categories: bool = False,
) -> Dict:
    """
    Merge corrections into a target Excel file.

    Supports match modes: "strict", "strorigin_only", "stringid_only".
    Auto-creates Str column if missing (inserts after StrOrigin).
    Same safeguards as XML merge: golden rule, Korean rejection, linebreak conversion.

    Args:
        excel_path: Path to target Excel file
        corrections: List of correction dicts with string_id, str_origin, corrected
        match_mode: "strict", "strorigin_only", or "stringid_only"
        dry_run: If True, don't write changes
        only_untranslated: If True, skip entries that already have non-Korean Str
        stringid_to_category: Category mapping (for stringid_only mode)
        stringid_to_subfolder: Subfolder mapping (for stringid_only exclusions)

    Returns:
        Dict with stats matching XML merge format
    """
    result = {
        "matched": 0,
        "updated": 0,
        "not_found": 0,
        "strorigin_mismatch": 0,
        "skipped_translated": 0,
        "skipped_non_script": 0,
        "skipped_excluded": 0,
        "errors": [],
        "details": [],
    }

    if not corrections:
        return result

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        col_indices = _detect_column_indices(ws)

        stringid_col = col_indices.get("stringid", col_indices.get("string_id"))
        strorigin_col = col_indices.get("strorigin", col_indices.get("str_origin"))
        str_col = col_indices.get("str")
        desc_col = col_indices.get("desc", col_indices.get("desctext",
                   col_indices.get("desc_text", col_indices.get("desccorrection"))))
        descorigin_col = col_indices.get("descorigin", col_indices.get("desc_origin"))

        if not stringid_col:
            result["errors"].append(f"Missing required column in {excel_path.name}: StringID")
            wb.close()
            return result

        # StrOrigin required for strict and strorigin_only, optional for stringid_only
        if not strorigin_col and match_mode != "stringid_only":
            result["errors"].append(f"Missing required column in {excel_path.name}: StrOrigin")
            wb.close()
            return result

        # Auto-create Str column if missing
        if not str_col:
            if strorigin_col:
                insert_after = strorigin_col
            else:
                insert_after = stringid_col
            if not dry_run:
                str_col = _insert_str_column(ws, insert_after)
                # Fix column references shifted by the insert
                if stringid_col >= str_col:
                    stringid_col += 1
                if strorigin_col and strorigin_col >= str_col:
                    strorigin_col += 1
                if descorigin_col and descorigin_col >= str_col:
                    descorigin_col += 1
                if desc_col and desc_col >= str_col:
                    desc_col += 1
            else:
                str_col = insert_after + 1
                logger.info("Dry run: would insert Str column")

        # Read all target entries
        target_entries = []
        target_stringids = set()
        target_strorigin_map = {}

        for row_idx in range(2, ws.max_row + 1):
            sid_val = ws.cell(row=row_idx, column=stringid_col).value
            so_val = ws.cell(row=row_idx, column=strorigin_col).value if strorigin_col else None
            str_val = ws.cell(row=row_idx, column=str_col).value if str_col else None
            do_val = ws.cell(row=row_idx, column=descorigin_col).value if descorigin_col else None
            dv_val = ws.cell(row=row_idx, column=desc_col).value if desc_col else None

            sid = str(sid_val).strip() if sid_val is not None else ""
            so = str(so_val).strip() if so_val is not None else ""
            sv = str(str_val).strip() if str_val is not None else ""
            do_str = str(do_val).strip() if do_val is not None else ""
            dv_str = str(dv_val).strip() if dv_val is not None else ""

            if not sid:
                continue

            target_stringids.add(sid.lower())
            if sid.lower() not in target_strorigin_map:
                target_strorigin_map[sid.lower()] = so
            target_entries.append({
                "string_id": sid,
                "str_origin": so,
                "str_value": sv,
                "desc_origin": do_str,
                "desc_value": dv_str,
                "row": row_idx,
            })

        # Auto-create Desc column if DescOrigin exists but Desc doesn't
        # Only for modes that write Desc (strict, stringid_only — NOT strorigin_only)
        if descorigin_col and not desc_col and not dry_run and match_mode != "strorigin_only":
            # Insert Desc column after DescOrigin
            desc_col = descorigin_col + 1
            ws.insert_cols(desc_col)
            ws.cell(row=1, column=desc_col, value="Desc")
            ws.cell(row=1, column=desc_col).font = Font(bold=True)
            logger.info(f"Auto-inserted 'Desc' column at position {desc_col} (after DescOrigin at {descorigin_col})")
            # Fix shifted column references
            if stringid_col >= desc_col:
                stringid_col += 1
            if strorigin_col and strorigin_col >= desc_col:
                strorigin_col += 1
            if str_col and str_col >= desc_col:
                str_col += 1

        # Dispatch to mode-specific merge
        if match_mode == "strict":
            result = _merge_excel_strict(
                ws, str_col, target_entries, target_stringids,
                target_strorigin_map, corrections, only_untranslated, result,
                desc_col=desc_col, descorigin_col=descorigin_col,
            )
        elif match_mode == "strorigin_only":
            result = _merge_excel_strorigin_only(
                ws, str_col, target_entries, corrections,
                only_untranslated, result,
                stringid_to_category=stringid_to_category,
            )
        elif match_mode == "stringid_only":
            result = _merge_excel_stringid_only(
                ws, str_col, target_entries, corrections,
                stringid_to_category, stringid_to_subfolder,
                only_untranslated, result,
                desc_col=desc_col, descorigin_col=descorigin_col,
                stringid_all_categories=stringid_all_categories,
            )
        else:
            result["errors"].append(f"Unsupported match mode for Excel: {match_mode}")
            wb.close()
            return result

        # Save if changes were made
        if result["updated"] > 0 and not dry_run:
            import os
            import stat
            try:
                current_mode = os.stat(excel_path).st_mode
                if not current_mode & stat.S_IWRITE:
                    os.chmod(excel_path, current_mode | stat.S_IWRITE)
            except Exception as e:
                logger.warning(f"Could not make {excel_path.name} writable: {e}")
            wb.save(excel_path)
            logger.info(f"Saved {excel_path.name}: {result['updated']} entries updated")

        wb.close()

    except Exception as e:
        result["errors"].append(str(e))
        logger.error(f"Error merging to {excel_path}: {e}")

    return result


def _merge_excel_strict(
    ws, str_col: int, target_entries: List[Dict],
    target_stringids: set, target_strorigin_map: Dict,
    corrections: List[Dict], only_untranslated: bool,
    result: Dict,
    desc_col: Optional[int] = None,
    descorigin_col: Optional[int] = None,
) -> Dict:
    """Strict matching: StringID + normalized StrOrigin."""
    target_lookup = defaultdict(list)
    target_lookup_nospace = defaultdict(list)
    for entry in target_entries:
        if not entry["str_origin"].strip():
            continue
        sid_lower = entry["string_id"].lower()
        norm_origin = normalize_text(entry["str_origin"])
        nospace_origin = normalize_nospace(norm_origin)
        target_lookup[(sid_lower, norm_origin)].append(entry)
        target_lookup_nospace[(sid_lower, nospace_origin)].append(entry)

    for c in corrections:
        sid_lower = c["string_id"].lower()
        origin_norm = normalize_text(c.get("str_origin", ""))
        origin_nospace = normalize_nospace(origin_norm)

        matching_entries = target_lookup.get((sid_lower, origin_norm), [])
        if not matching_entries:
            matching_entries = target_lookup_nospace.get((sid_lower, origin_nospace), [])

        if matching_entries:
            for target_entry in matching_entries:
                result["matched"] += 1
                old_str = target_entry["str_value"]

                if only_untranslated and old_str and not is_korean_text(old_str):
                    result["skipped_translated"] += 1
                    result["details"].append({
                        "string_id": c["string_id"], "status": "SKIPPED_TRANSLATED",
                        "old": c.get("str_origin", ""), "new": c.get("corrected", ""),
                    })
                    continue

                new_str = _convert_linebreaks_for_excel(c["corrected"])
                if new_str != old_str:
                    ws.cell(row=target_entry["row"], column=str_col, value=new_str)
                    if "<br/>" in new_str:
                        ws.cell(row=target_entry["row"], column=str_col).alignment = Alignment(wrap_text=True, vertical='top')
                    result["updated"] += 1
                    result["details"].append({
                        "string_id": c["string_id"], "status": "UPDATED",
                        "old": old_str, "new": new_str,
                    })
                else:
                    result["details"].append({
                        "string_id": c["string_id"], "status": "UNCHANGED",
                        "old": old_str, "new": "(same)",
                    })

                # Write Desc if correction has desc_corrected and target has DescOrigin
                if desc_col and c.get("desc_corrected"):
                    target_do = target_entry.get("desc_origin", "")
                    if target_do:
                        new_desc = _convert_linebreaks_for_excel(c["desc_corrected"])
                        old_desc = target_entry.get("desc_value", "")
                        if new_desc != old_desc:
                            ws.cell(row=target_entry["row"], column=desc_col, value=new_desc)
                            if "<br/>" in new_desc:
                                ws.cell(row=target_entry["row"], column=desc_col).alignment = Alignment(wrap_text=True, vertical='top')
                            result["desc_updated"] = result.get("desc_updated", 0) + 1
        else:
            if sid_lower in target_stringids:
                result["strorigin_mismatch"] += 1
                result["details"].append({
                    "string_id": c["string_id"], "status": "STRORIGIN_MISMATCH",
                    "old": c.get("str_origin", ""), "new": c["corrected"],
                    "target_strorigin": target_strorigin_map.get(sid_lower, ""),
                })
            else:
                result["not_found"] += 1
                result["details"].append({
                    "string_id": c["string_id"], "status": "NOT_FOUND",
                    "old": c.get("str_origin", ""), "new": c["corrected"],
                })

    return result


def _merge_excel_strorigin_only(
    ws, str_col: int, target_entries: List[Dict],
    corrections: List[Dict], only_untranslated: bool,
    result: Dict,
    stringid_to_category: Optional[Dict[str, str]] = None,
) -> Dict:
    """StrOrigin-only matching. Skips SCRIPT (Dialog/Sequencer) corrections when category data available."""
    from .text_utils import normalize_for_matching
    import config as _cfg

    # SAFETY: Filter out SCRIPT category corrections (Dialog/Sequencer)
    if stringid_to_category:
        ci_category = {k.lower(): v for k, v in stringid_to_category.items()}
        filtered = []
        for c in corrections:
            sid = c.get("string_id", "").lower()
            category = ci_category.get(sid, "")
            if category in _cfg.SCRIPT_CATEGORIES:
                result.setdefault("skipped_script", 0)
                result["skipped_script"] += 1
                result["details"].append({
                    "string_id": c.get("string_id", ""),
                    "status": "SKIPPED_SCRIPT",
                    "old": f"Category: {category}",
                    "new": c.get("corrected", ""),
                })
                continue
            filtered.append(c)
        corrections = filtered

    target_by_origin = {}
    target_by_origin_nospace = {}
    for entry in target_entries:
        if not entry["str_origin"].strip():
            continue
        norm = normalize_for_matching(entry["str_origin"])
        nospace = normalize_nospace(norm)
        target_by_origin.setdefault(norm, []).append(entry)
        target_by_origin_nospace.setdefault(nospace, []).append(entry)

    correction_lookup = {}
    conflicting = 0
    for c in corrections:
        origin_norm = normalize_for_matching(c.get("str_origin", ""))
        if not origin_norm:
            continue
        if origin_norm in correction_lookup:
            old_corrected = correction_lookup[origin_norm].get("corrected", "")
            if old_corrected != c.get("corrected", ""):
                conflicting += 1
                logger.debug(
                    f"StrOrigin-only Excel: conflicting correction for StrOrigin "
                    f"'{c.get('str_origin', '')}' — overwriting with last"
                )
        correction_lookup[origin_norm] = c
    if conflicting:
        logger.warning(f"StrOrigin-only Excel: {conflicting} conflicting corrections (last wins)")

    for origin_norm, c in correction_lookup.items():
        origin_nospace = normalize_nospace(origin_norm)
        entries = target_by_origin.get(origin_norm, [])
        if not entries:
            entries = target_by_origin_nospace.get(origin_nospace, [])

        if entries:
            for target_entry in entries:
                result["matched"] += 1
                old_str = target_entry["str_value"]

                if only_untranslated and old_str and not is_korean_text(old_str):
                    result["skipped_translated"] += 1
                    result["details"].append({
                        "string_id": target_entry["string_id"], "status": "SKIPPED_TRANSLATED",
                        "old": c.get("str_origin", ""), "new": c.get("corrected", ""),
                    })
                    continue

                new_str = _convert_linebreaks_for_excel(c["corrected"])
                if new_str != old_str:
                    ws.cell(row=target_entry["row"], column=str_col, value=new_str)
                    if "<br/>" in new_str:
                        ws.cell(row=target_entry["row"], column=str_col).alignment = Alignment(wrap_text=True, vertical='top')
                    result["updated"] += 1
                    result["details"].append({
                        "string_id": target_entry["string_id"], "status": "UPDATED",
                        "old": old_str, "new": new_str,
                    })
                else:
                    result["details"].append({
                        "string_id": target_entry["string_id"], "status": "UNCHANGED",
                        "old": old_str, "new": "(same)",
                    })
        else:
            result["not_found"] += 1
            result["details"].append({
                "string_id": c.get("string_id", ""), "status": "NOT_FOUND",
                "old": c.get("str_origin", ""), "new": c["corrected"],
            })

    return result


def _merge_excel_stringid_only(
    ws, str_col: int, target_entries: List[Dict],
    corrections: List[Dict],
    stringid_to_category: Optional[Dict[str, str]],
    stringid_to_subfolder: Optional[Dict[str, str]],
    only_untranslated: bool, result: Dict,
    desc_col: Optional[int] = None,
    descorigin_col: Optional[int] = None,
    stringid_all_categories: bool = False,
) -> Dict:
    """StringID-only matching. SCRIPT types only by default, ALL when stringid_all_categories=True."""
    import config as _cfg
    SCRIPT_CATEGORIES = _cfg.SCRIPT_CATEGORIES
    SCRIPT_EXCLUDE_SUBFOLDERS = _cfg.SCRIPT_EXCLUDE_SUBFOLDERS

    ci_category = {k.lower(): v for k, v in stringid_to_category.items()} if stringid_to_category else {}
    ci_subfolder = {k.lower(): v for k, v in stringid_to_subfolder.items()} if stringid_to_subfolder else {}

    # Lazy EventName resolver — only loaded if needed
    _en_resolver_loaded = False
    _en_mapping = None
    _en_extract_fn = None
    _eventname_resolved_count = 0

    def _try_resolve_as_eventname(eventname_str):
        nonlocal _en_resolver_loaded, _en_mapping, _en_extract_fn
        if not _en_resolver_loaded:
            _en_resolver_loaded = True
            try:
                from .eventname_resolver import get_eventname_mapping, extract_stringid_from_dialog_keyword
                _en_mapping = get_eventname_mapping(_cfg.EXPORT_FOLDER)
                _en_extract_fn = extract_stringid_from_dialog_keyword
            except Exception:
                _en_mapping = None
        if not _en_mapping:
            return None
        extracted = _en_extract_fn(eventname_str)
        if extracted and extracted.lower() != eventname_str.lower():
            return extracted
        data = _en_mapping.get(eventname_str.lower())
        if data:
            return data["stringid"]
        return None

    # Build list-based lookup: one StringID can map to MULTIPLE target rows
    # Note: do NOT filter out empty str_origin — StringID-only mode matches by ID alone,
    # and the target may legitimately have no StrOrigin column (3D.C3 fix).
    target_by_sid = defaultdict(list)
    for entry in target_entries:
        target_by_sid[entry["string_id"].lower()].append(entry)

    for c in corrections:
        sid = c["string_id"]
        sid_lower = sid.lower()
        category = ci_category.get(sid_lower, "Uncategorized")
        subfolder = ci_subfolder.get(sid_lower, "")

        if not stringid_all_categories:
            if category not in SCRIPT_CATEGORIES:
                # Multi-pass: try resolving as EventName before skipping
                resolved_sid = _try_resolve_as_eventname(sid)
                if resolved_sid:
                    resolved_lower = resolved_sid.lower()
                    resolved_category = ci_category.get(resolved_lower, "Uncategorized")
                    if resolved_category in SCRIPT_CATEGORIES:
                        _eventname_resolved_count += 1
                        logger.debug(f"EventName '{sid}' resolved to SCRIPT StringID '{resolved_sid}' (category={resolved_category})")
                        sid = resolved_sid
                        sid_lower = resolved_lower
                        category = resolved_category
                        subfolder = ci_subfolder.get(resolved_lower, "")
                        c = dict(c)
                        c["string_id"] = resolved_sid
                    else:
                        result["skipped_non_script"] += 1
                        result["details"].append({
                            "string_id": sid, "status": "SKIPPED_NON_SCRIPT",
                            "old": f"Category: {resolved_category} (resolved from EventName)", "new": c["corrected"],
                        })
                        continue
                else:
                    result["skipped_non_script"] += 1
                    result["details"].append({
                        "string_id": sid, "status": "SKIPPED_NON_SCRIPT",
                        "old": f"Category: {category}", "new": c["corrected"],
                    })
                    continue

            if subfolder.lower() in {s.lower() for s in SCRIPT_EXCLUDE_SUBFOLDERS}:
                result["skipped_excluded"] += 1
                result["details"].append({
                    "string_id": sid, "status": "SKIPPED_EXCLUDED",
                    "old": f"Subfolder: {subfolder}", "new": c["corrected"],
                })
                continue

        matching_entries = target_by_sid.get(sid_lower, [])
        if matching_entries:
            for target_entry in matching_entries:
                result["matched"] += 1
                old_str = target_entry["str_value"]

                if only_untranslated and old_str and not is_korean_text(old_str):
                    result["skipped_translated"] += 1
                    result["details"].append({
                        "string_id": sid, "status": "SKIPPED_TRANSLATED",
                        "old": c.get("str_origin", ""), "new": c.get("corrected", ""),
                    })
                    continue

                new_str = _convert_linebreaks_for_excel(c["corrected"])
                if new_str != old_str:
                    ws.cell(row=target_entry["row"], column=str_col, value=new_str)
                    if "<br/>" in new_str:
                        ws.cell(row=target_entry["row"], column=str_col).alignment = Alignment(wrap_text=True, vertical='top')
                    result["updated"] += 1
                    result["details"].append({
                        "string_id": sid, "status": "UPDATED",
                        "old": old_str, "new": new_str,
                    })
                else:
                    result["details"].append({
                        "string_id": sid, "status": "UNCHANGED",
                        "old": old_str, "new": "(same)",
                    })

                # Write Desc if correction has desc_corrected and target has DescOrigin
                if desc_col and c.get("desc_corrected"):
                    target_do = target_entry.get("desc_origin", "")
                    if target_do:
                        new_desc = _convert_linebreaks_for_excel(c["desc_corrected"])
                        old_desc = target_entry.get("desc_value", "")
                        if new_desc != old_desc:
                            ws.cell(row=target_entry["row"], column=desc_col, value=new_desc)
                            if "<br/>" in new_desc:
                                ws.cell(row=target_entry["row"], column=desc_col).alignment = Alignment(wrap_text=True, vertical='top')
                            result["desc_updated"] = result.get("desc_updated", 0) + 1
        else:
            result["not_found"] += 1
            result["details"].append({
                "string_id": sid, "status": "NOT_FOUND",
                "old": c.get("str_origin", ""), "new": c["corrected"],
            })

    if _eventname_resolved_count:
        logger.info(f"EventName resolution: {_eventname_resolved_count} EventNames resolved to SCRIPT StringIDs")

    return result
