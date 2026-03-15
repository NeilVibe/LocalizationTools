"""
XML Post-Processing Pipeline.

Runs cleanup passes on XML files AFTER transfer operations.
Each cleanup function is independent and can be added/removed easily.

Current cleanup steps (run in order, applied to both Str and Desc):
  1. cleanup_wrong_newlines    - Normalize ALL newlines to <br/>
  2. cleanup_empty_strorigin   - Clear Str/Desc when StrOrigin/DescOrigin is empty
  3. cleanup_no_translation    - Replace "no translation" Str/Desc with StrOrigin/DescOrigin
  4. cleanup_apostrophes       - Normalize curly/fancy apostrophes to ASCII apostrophe
  5. cleanup_invisible_chars   - Auto-fix invisible Unicode characters (NBSP→space, delete zero-width)
  6. cleanup_hyphens           - Normalize Unicode hyphen lookalikes to ASCII hyphen-minus
  7. cleanup_ellipsis          - Normalize Unicode ellipsis (…) to three dots (non-CJK only)
  8. cleanup_double_escaped    - Decode double-escaped entities (&lt;/&gt;/&quot;/&apos;/&amp;ENTITY;)

Usage:
    from core.postprocess import run_all_postprocess
    fixed = run_all_postprocess(xml_path)
"""

from __future__ import annotations

import os
import re
import stat
import logging
import unicodedata
from pathlib import Path

# Try lxml first (more robust), fallback to standard library
try:
    from lxml import etree
    USING_LXML = True
except ImportError:
    from xml.etree import ElementTree as etree
    USING_LXML = False

from .xml_parser import LOCSTR_TAGS, STR_ATTRS, STRORIGIN_ATTRS, DESC_ATTRS, DESCORIGIN_ATTRS
from .text_utils import _SAFE_INVISIBLE_DELETE, _SAFE_INVISIBLE_NAMES, _GREY_ZONE_CHARS

logger = logging.getLogger(__name__)

# Languages that NEVER get ellipsis normalization (CJK — ellipsis is intentional)
_CJK_SKIP_LANGS = frozenset({
    "KOR", "JPN", "ZHO-CN", "ZHO_CN", "ZHO-TW", "ZHO_TW",
})

# --- Compiled regex patterns for newline normalization ---

# <br> tag variants: <br>, <br/>, <br />, <BR>, <BR/>, </br>, </BR>, etc.
_BR_TAG_RE = re.compile(r'</?[Bb][Rr]\s*/?>')

# HTML-escaped <br> variants: &lt;br/&gt;, &lt;br&gt;, &lt;/br&gt;, &lt;BR/&gt;, etc.
_BR_ESCAPED_RE = re.compile(r'&lt;/?[Bb][Rr]\s*/?&gt;')

# CRLF as XML entity combo: &#13;&#10; or &#xD;&#xA; (with optional whitespace between)
_CRLF_ENTITY_RE = re.compile(r'&#[Xx]0?[Dd];\s*&#[Xx]0?[Aa];|&#13;\s*&#10;')

# Individual hex entity references for LF and CR
_HEX_LF_RE = re.compile(r'&#[Xx]0?[Aa];')
_HEX_CR_RE = re.compile(r'&#[Xx]0?[Dd];')


def _normalize_newlines(text: str) -> str:
    """
    Normalize ALL newline representations to <br/>.

    <br/> is the ONLY correct newline format in XML language data.
    This function handles every known newline variant:

    Escaped HTML tags:
        &lt;br/&gt;, &lt;br&gt;, &lt;BR/&gt;, etc.

    Wrong <br> tag variants:
        <br>, <BR/>, <BR>, <br />, <BR />, <Br/>, <bR/>, mixed case

    XML numeric entity references:
        &#10;, &#13;, &#xA;, &#xa;, &#x0A;, &#xD;, &#xd;, &#x0D;

    Actual control characters:
        \\r\\n (CRLF), \\r (CR), \\n (LF)

    Literal escape text:
        Backslash-n, backslash-r, backslash-r-backslash-n

    Unicode line break characters:
        NEL (U+0085), Line Separator (U+2028), Paragraph Separator (U+2029),
        Vertical Tab (U+000B), Form Feed (U+000C)

    Args:
        text: Attribute value (after XML parser decoding)

    Returns:
        Text with all newlines normalized to <br/>
    """
    if not text:
        return text

    # Step 0: Decode double-escaped &amp;lt;br/&amp;gt; patterns first
    # In memory after lxml: &amp;lt;br/&amp;gt; (from &amp;amp;lt;br/&amp;amp;gt; on disk)
    text = re.sub(r'&amp;lt;/?[Bb][Rr]\s*/?&amp;gt;', '<br/>', text)

    # Step 1: Unescape HTML-escaped <br> variants (from double-escaping)
    # &lt;br/&gt; → <br/>,  &lt;br&gt; → <br/>,  &lt;BR/&gt; → <br/>
    text = _BR_ESCAPED_RE.sub('<br/>', text)

    # Step 2: Normalize all <br> tag variants to <br/>
    # <br> → <br/>,  <BR/> → <br/>,  <br /> → <br/>,  <Br> → <br/>
    # Note: <br/> → <br/> (no-op, correct form stays correct)
    text = _BR_TAG_RE.sub('<br/>', text)

    # Step 3: Replace CRLF entity combos FIRST (before individual entities)
    # &#13;&#10; → <br/>,  &#xD;&#xA; → <br/>
    text = _CRLF_ENTITY_RE.sub('<br/>', text)

    # Step 4: Replace individual XML entity references
    # &#10; → <br/>,  &#13; → <br/>
    text = text.replace('&#10;', '<br/>')
    text = text.replace('&#13;', '<br/>')
    # &#xA; &#xa; &#x0A; &#x0a; → <br/>
    text = _HEX_LF_RE.sub('<br/>', text)
    # &#xD; &#xd; &#x0D; &#x0d; → <br/>
    text = _HEX_CR_RE.sub('<br/>', text)

    # Step 5: Replace actual control characters (CRLF combo first!)
    text = text.replace('\r\n', '<br/>')
    text = text.replace('\r', '<br/>')
    text = text.replace('\n', '<br/>')

    # Step 6: Replace literal escape sequences (backslash + letter as text)
    text = text.replace('\\r\\n', '<br/>')
    text = text.replace('\\n', '<br/>')
    text = text.replace('\\r', '<br/>')

    # Step 7: Unicode line break characters
    text = text.replace('\u0085', '<br/>')   # NEL (Next Line)
    text = text.replace('\u2028', '<br/>')   # Line Separator
    text = text.replace('\u2029', '<br/>')   # Paragraph Separator
    text = text.replace('\x0b', '<br/>')     # Vertical Tab
    text = text.replace('\x0c', '<br/>')     # Form Feed

    return text


# --- Apostrophe normalization table ---

_APOSTROPHE_TABLE = str.maketrans({
    '\u2018': "'",  # LEFT SINGLE QUOTATION MARK
    '\u2019': "'",  # RIGHT SINGLE QUOTATION MARK
    '\u00B4': "'",  # ACUTE ACCENT
    '\u02BC': "'",  # MODIFIER LETTER APOSTROPHE
    '\u201B': "'",  # SINGLE HIGH-REVERSED-9 QUOTATION MARK
    '\uFF07': "'",  # FULLWIDTH APOSTROPHE
})


def _normalize_apostrophes(text: str) -> str:
    """Normalize curly/fancy apostrophes to standard ASCII apostrophe."""
    return text.translate(_APOSTROPHE_TABLE)


# --- Hyphen normalization table ---

_HYPHEN_TABLE = str.maketrans({
    '\u2010': "-",  # HYPHEN (visually identical to ASCII hyphen-minus)
    '\u2011': "-",  # NON-BREAKING HYPHEN (same visual, no game engine respects non-break)
})


def _normalize_hyphens(text: str) -> str:
    """Normalize Unicode hyphen lookalikes to standard ASCII hyphen-minus."""
    return text.translate(_HYPHEN_TABLE)


def _normalize_ellipsis(text: str) -> str:
    """Replace Unicode horizontal ellipsis (U+2026) with three ASCII dots."""
    return text.replace('\u2026', '...')


# --- Step 8: Safe double-escaped entity decode ---
# Explicit allowlist approach: every decode pattern is complete and atomic.
_SAFE_AMP_ENTITIES_RE = re.compile(
    r'&amp;(desc|nbsp|ensp|emsp|thinsp|hellip|bull|middot|lrm|rlm);',
    re.IGNORECASE,
)


def _decode_safe_entities(text: str) -> str:
    """Safely decode double-escaped XML entities.

    Decodes (explicit allowlist — no guessing):
      - &lt; / &gt;    → < / >    (double-escaped on disk → decoded in memory)
      - &amp;desc;     → &desc;   (and other known named entities)
      - &quot;         → "        (standalone, no pairing risk)
      - &apos;         → '        (standalone, no pairing risk)

    NOT decoded (too dangerous):
      - &amp;          → NOT blindly decoded (could create broken entities)

    Every pattern is complete and atomic. No partial decode possible.
    """
    # 1. Decode known &amp;ENTITY; compounds (explicit allowlist)
    text = _SAFE_AMP_ENTITIES_RE.sub(lambda m: f'&{m.group(1)};', text)
    # 2. Decode double-escaped &lt; and &gt; (safe: these can ONLY exist in
    #    memory from &amp;lt; / &amp;gt; on disk. Step 1 already converted
    #    br-tag cases, so remaining ones are standalone.)
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    # 3. Decode &quot; and &apos; (standalone, no pairing risk)
    text = text.replace('&quot;', '"')
    text = text.replace('&apos;', "'")
    return text


def _extract_language_from_path(xml_path: Path) -> str:
    """Extract language code from filename like 'languagedata_fre.xml' or 'languagedata_zho-cn.xml'."""
    stem = xml_path.stem.lower()
    if stem.startswith("languagedata_"):
        return stem[len("languagedata_"):].upper()
    return ""


def _parse_xml(xml_path: Path):
    """Parse XML file, returns (tree, root). Shared by all postprocess functions."""
    if USING_LXML:
        parser = etree.XMLParser(
            resolve_entities=False, load_dtd=False,
            no_network=True, recover=True,
        )
        tree = etree.parse(str(xml_path), parser)
    else:
        tree = etree.parse(str(xml_path))
    return tree, tree.getroot()


def _write_xml(tree, xml_path: Path):
    """Write XML tree back to file. Shared by all postprocess functions."""
    try:
        current_mode = os.stat(xml_path).st_mode
        if not current_mode & stat.S_IWRITE:
            os.chmod(xml_path, current_mode | stat.S_IWRITE)
    except Exception:
        pass

    if USING_LXML:
        tree.write(str(xml_path), encoding="utf-8", xml_declaration=False, pretty_print=True)
    else:
        tree.write(str(xml_path), encoding="utf-8", xml_declaration=False)


def _iter_locstr(root):
    """Iterate all LocStr elements (case-insensitive tag matching)."""
    elements = []
    for tag in LOCSTR_TAGS:
        elements.extend(root.iter(tag))
    return elements


def _get_attr(loc, attr_variants):
    """Get attribute name and value from element, trying multiple case variants.

    NOTE: This is intentionally NOT the centralized get_attr() from xml_parser.py.
    Unlike the centralized version (which returns just the value string), this
    returns a (attr_name, value) tuple because postprocess needs the actual
    attribute name to write back the corrected value to the same key.
    """
    for attr_name in attr_variants:
        val = loc.get(attr_name)
        if val is not None:
            return attr_name, val
    return None, None


# ─── Cleanup Step 1: Normalize newlines ─────────────────────────────────────


def cleanup_wrong_newlines_on_tree(root) -> int:
    """
    Normalize all wrong newline representations to <br/> in parsed tree.

    Modifies Str and Desc attributes. StrOrigin/DescOrigin (Korean source) are
    never touched — they must remain exactly as the source data provides them.

    Args:
        root: Parsed XML root element

    Returns:
        Number of attributes fixed
    """
    fixed = 0
    for loc in _iter_locstr(root):
        attr_name, val = _get_attr(loc, STR_ATTRS)
        if val is not None:
            normalized = _normalize_newlines(val)
            if normalized != val:
                loc.set(attr_name, normalized)
                fixed += 1
        # Also normalize newlines in Desc attribute
        desc_name, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _normalize_newlines(desc_val)
            if desc_normalized != desc_val:
                loc.set(desc_name, desc_normalized)
                fixed += 1
    return fixed


# ─── Cleanup Step 2: Empty StrOrigin enforcement ────────────────────────────


def cleanup_empty_strorigin_on_tree(root) -> int:
    """
    Clear Str/Desc on any element where StrOrigin/DescOrigin is empty.

    Golden rule: if StrOrigin is empty, Str MUST be empty too.
    Same applies to Desc/DescOrigin.

    Args:
        root: Parsed XML root element

    Returns:
        Number of entries cleaned (Str or Desc cleared)
    """
    cleaned = 0
    for loc in _iter_locstr(root):
        _, origin = _get_attr(loc, STRORIGIN_ATTRS)
        origin = (origin or "").strip()

        str_attr, str_val = _get_attr(loc, STR_ATTRS)
        str_val = (str_val or "").strip()

        if not origin and str_val:
            loc.set(str_attr or "Str", "")
            cleaned += 1

        # Golden rule for Desc/DescOrigin: if DescOrigin is empty, Desc MUST be empty
        _, desc_origin = _get_attr(loc, DESCORIGIN_ATTRS)
        desc_origin = (desc_origin or "").strip()

        desc_attr, desc_val = _get_attr(loc, DESC_ATTRS)
        desc_val = (desc_val or "").strip()

        if not desc_origin and desc_val:
            loc.set(desc_attr or "Desc", "")
            cleaned += 1
    return cleaned


# ─── Cleanup Step 3: "No translation" replacement ─────────────────────────────

# Regex to collapse whitespace for normalization
_WHITESPACE_RE = re.compile(r'\s+')


def cleanup_no_translation_on_tree(root) -> int:
    """
    Replace Str/Desc with StrOrigin/DescOrigin when value is exactly "no translation".

    Matching is case-insensitive with whitespace normalization (trim + collapse).
    So "  NO   TRANSLATION  " matches, but "some no translation text" does NOT.
    Applies to both Str and Desc attributes independently.

    Args:
        root: Parsed XML root element

    Returns:
        Number of entries fixed (Str/Desc replaced with origin)
    """
    fixed = 0
    for loc in _iter_locstr(root):
        str_attr, str_val = _get_attr(loc, STR_ATTRS)
        if str_val is not None:
            # Normalize: trim + collapse internal whitespace + lowercase
            normalized = _WHITESPACE_RE.sub(' ', str_val.strip()).lower()
            if normalized == 'no translation':
                # Get StrOrigin to copy
                _, origin = _get_attr(loc, STRORIGIN_ATTRS)
                origin = (origin or "").strip()

                if origin:
                    loc.set(str_attr, origin)
                    fixed += 1
                else:
                    # StrOrigin is also empty — clear Str (golden rule)
                    loc.set(str_attr, "")
                    fixed += 1

        # Same logic for Desc: if "no translation", replace with DescOrigin
        desc_attr, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _WHITESPACE_RE.sub(' ', desc_val.strip()).lower()
            if desc_normalized == 'no translation':
                _, desc_origin = _get_attr(loc, DESCORIGIN_ATTRS)
                desc_origin = (desc_origin or "").strip()

                if desc_origin:
                    loc.set(desc_attr, desc_origin)
                    fixed += 1
                else:
                    loc.set(desc_attr, "")
                    fixed += 1

    return fixed


# ─── Cleanup Step 4: Apostrophe normalization ────────────────────────────────


def cleanup_apostrophes_on_tree(root) -> int:
    """
    Normalize curly/fancy apostrophes to standard ASCII apostrophe in parsed tree.

    Modifies Str and Desc attributes ONLY. StrOrigin/DescOrigin (Korean source) are
    never touched — they must remain exactly as the source data provides them.

    Args:
        root: Parsed XML root element

    Returns:
        Number of attributes fixed
    """
    fixed = 0
    for loc in _iter_locstr(root):
        attr_name, val = _get_attr(loc, STR_ATTRS)
        if val is not None:
            normalized = _normalize_apostrophes(val)
            if normalized != val:
                loc.set(attr_name, normalized)
                fixed += 1
        # Also normalize apostrophes in Desc attribute
        desc_name, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _normalize_apostrophes(desc_val)
            if desc_normalized != desc_val:
                loc.set(desc_name, desc_normalized)
                fixed += 1
    return fixed


# ─── Cleanup Step 6: Hyphen normalization ─────────────────────────────────────


def cleanup_hyphens_on_tree(root) -> int:
    """
    Normalize Unicode hyphen lookalikes to standard ASCII hyphen-minus in parsed tree.

    Only normalizes U+2010 (Hyphen) and U+2011 (Non-breaking hyphen) — these are
    visually identical to ASCII hyphen-minus and never intentional in game text.
    CJK hyphens, en/em dashes, minus signs are NOT touched.

    Modifies Str and Desc attributes ONLY. StrOrigin/DescOrigin are never touched.

    Args:
        root: Parsed XML root element

    Returns:
        Number of attributes fixed
    """
    fixed = 0
    for loc in _iter_locstr(root):
        attr_name, val = _get_attr(loc, STR_ATTRS)
        if val is not None:
            normalized = _normalize_hyphens(val)
            if normalized != val:
                loc.set(attr_name, normalized)
                fixed += 1
        desc_name, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _normalize_hyphens(desc_val)
            if desc_normalized != desc_val:
                loc.set(desc_name, desc_normalized)
                fixed += 1
    return fixed


# ─── Cleanup Step 7: Ellipsis normalization (non-CJK only) ────────────────


def cleanup_ellipsis_on_tree(root) -> int:
    """
    Normalize Unicode ellipsis (U+2026 …) to three ASCII dots (...) in parsed tree.

    Modifies Str and Desc attributes ONLY. StrOrigin/DescOrigin are never touched.
    Caller is responsible for skipping CJK languages (check _CJK_SKIP_LANGS).

    Args:
        root: Parsed XML root element

    Returns:
        Number of attributes fixed
    """
    fixed = 0
    for loc in _iter_locstr(root):
        attr_name, val = _get_attr(loc, STR_ATTRS)
        if val is not None:
            normalized = _normalize_ellipsis(val)
            if normalized != val:
                loc.set(attr_name, normalized)
                fixed += 1
        desc_name, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _normalize_ellipsis(desc_val)
            if desc_normalized != desc_val:
                loc.set(desc_name, desc_normalized)
                fixed += 1
    return fixed


# ─── Cleanup Step 8: Safe double-escaped entity decode ─────────────────────


def cleanup_double_escaped_on_tree(root) -> int:
    """
    Safe decode of double-escaped entities (&quot; &apos; &amp;ENTITY;) in parsed tree.

    Modifies Str and Desc attributes ONLY. StrOrigin/DescOrigin are never touched.

    Args:
        root: Parsed XML root element

    Returns:
        Number of attributes fixed
    """
    fixed = 0
    for loc in _iter_locstr(root):
        attr_name, val = _get_attr(loc, STR_ATTRS)
        if val is not None:
            decoded = _decode_safe_entities(val)
            if decoded != val:
                loc.set(attr_name, decoded)
                fixed += 1
        desc_name, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_decoded = _decode_safe_entities(desc_val)
            if desc_decoded != desc_val:
                loc.set(desc_name, desc_decoded)
                fixed += 1
    return fixed


# ─── Public API: Standalone functions (backward-compatible) ──────────────────


def cleanup_wrong_newlines(xml_path: Path, dry_run: bool = False) -> int:
    """
    Post-process: normalize all wrong newline representations to <br/>.

    <br/> is the ONLY correct newline format in XML language data.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Number of attributes fixed
    """
    try:
        tree, root = _parse_xml(xml_path)
        fixed = cleanup_wrong_newlines_on_tree(root)

        if fixed > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.info(f"Normalized {fixed} wrong newlines in {xml_path.name}")

        return fixed
    except Exception as e:
        logger.error(f"Error normalizing newlines in {xml_path}: {e}")
        return 0


def cleanup_empty_strorigin(xml_path: Path, dry_run: bool = False) -> int:
    """
    Post-process: clear Str on any element where StrOrigin is empty.

    Golden rule: if StrOrigin is empty, Str MUST be empty too.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Number of entries cleaned (Str cleared)
    """
    try:
        tree, root = _parse_xml(xml_path)
        cleaned = cleanup_empty_strorigin_on_tree(root)

        if cleaned > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.info(f"Cleaned {cleaned} entries with empty StrOrigin in {xml_path.name}")

        return cleaned
    except Exception as e:
        logger.error(f"Error cleaning empty StrOrigin in {xml_path}: {e}")
        return 0


def cleanup_no_translation(xml_path: Path, dry_run: bool = False) -> int:
    """
    Post-process: replace Str with StrOrigin when Str is exactly "no translation".

    Case-insensitive with whitespace normalization.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Number of entries fixed
    """
    try:
        tree, root = _parse_xml(xml_path)
        fixed = cleanup_no_translation_on_tree(root)

        if fixed > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.info(
                f"Replaced {fixed} 'no translation' entries with StrOrigin "
                f"in {xml_path.name}"
            )

        return fixed
    except Exception as e:
        logger.error(f"Error cleaning 'no translation' in {xml_path}: {e}")
        return 0


def cleanup_apostrophes(xml_path: Path, dry_run: bool = False) -> int:
    """
    Post-process: normalize curly/fancy apostrophes to ASCII apostrophe.

    Modifies Str and Desc attributes only.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Number of attributes fixed
    """
    try:
        tree, root = _parse_xml(xml_path)
        fixed = cleanup_apostrophes_on_tree(root)

        if fixed > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.debug("Normalized apostrophes in %d entries in %s", fixed, xml_path.name)

        return fixed
    except Exception as e:
        logger.error(f"Error normalizing apostrophes in {xml_path}: {e}")
        return 0


def cleanup_ellipsis(xml_path: Path, dry_run: bool = False) -> int:
    """
    Post-process: normalize Unicode ellipsis to three ASCII dots (non-CJK only).

    Skips CJK languages (KOR, JPN, ZHO-CN, ZHO-TW) silently.
    Modifies Str and Desc attributes only.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Number of attributes fixed (0 for CJK languages)
    """
    lang = _extract_language_from_path(xml_path)
    if lang in _CJK_SKIP_LANGS:
        return 0

    try:
        tree, root = _parse_xml(xml_path)
        fixed = cleanup_ellipsis_on_tree(root)

        if fixed > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.debug("Normalized ellipsis in %d entries in %s", fixed, xml_path.name)

        return fixed
    except Exception as e:
        logger.error(f"Error normalizing ellipsis in {xml_path}: {e}")
        return 0


# ─── Cleanup Step 5: Invisible character cleanup ──────────────────────────


def _cleanup_invisible_chars(text):
    """Clean invisible characters from a single text value.

    Bucket 1 (Zs spaces): Replace with regular space U+0020
    Bucket 2 (safe invisible): Delete entirely

    Returns:
        (cleaned_text, spaces_count, deleted_count, detail_counts)
        detail_counts is a dict {char_name: count}
    """
    if not text:
        return text, 0, 0, {}

    chars = []
    spaces_count = 0
    deleted_count = 0
    detail_counts = {}

    for ch in text:
        if ch in _SAFE_INVISIBLE_DELETE:
            # Bucket 2: delete
            deleted_count += 1
            name = _SAFE_INVISIBLE_NAMES.get(ch, f'U+{ord(ch):04X}')
            detail_counts[name] = detail_counts.get(name, 0) + 1
        elif ch != ' ' and ch != '\u3000' and unicodedata.category(ch) == 'Zs':
            # Bucket 1: Zs space → regular space (U+3000 ideographic space excluded — CJK intentional)
            chars.append(' ')
            spaces_count += 1
            # Try to get a readable name
            try:
                name = unicodedata.name(ch, f'U+{ord(ch):04X}')
            except ValueError:
                name = f'U+{ord(ch):04X}'
            detail_counts[name] = detail_counts.get(name, 0) + 1
        else:
            chars.append(ch)

    if spaces_count == 0 and deleted_count == 0:
        return text, 0, 0, {}

    return ''.join(chars), spaces_count, deleted_count, detail_counts


def cleanup_invisible_chars_on_tree(root) -> dict:
    """
    Clean invisible Unicode characters from Str and Desc attributes in parsed tree.

    Bucket 1 (Zs spaces like NBSP, en-space, em-space): replaced with regular space.
    Bucket 2 (safe invisible like zero-width space, BOM, bidi marks): deleted.
    Grey zone (ZWNJ, ZWJ): detected and warned, but NOT modified.

    StrOrigin/DescOrigin are NEVER touched.

    Args:
        root: Parsed XML root element

    Returns:
        {"spaces_normalized": int, "invisibles_removed": int,
         "invisible_detail": {name: count}, "grey_zone_detected": {name: count}}
    """
    result = {
        "spaces_normalized": 0,
        "invisibles_removed": 0,
        "invisible_detail": {},
        "grey_zone_detected": {},
    }

    for loc in _iter_locstr(root):
        for attr_variants in (STR_ATTRS, DESC_ATTRS):
            attr_name, val = _get_attr(loc, attr_variants)
            if val is None:
                continue

            cleaned, sp, dl, detail = _cleanup_invisible_chars(val)
            if sp or dl:
                loc.set(attr_name, cleaned)
                result["spaces_normalized"] += sp
                result["invisibles_removed"] += dl
                for name, cnt in detail.items():
                    result["invisible_detail"][name] = result["invisible_detail"].get(name, 0) + cnt

            # Grey zone detection (no modification)
            for ch, name in _GREY_ZONE_CHARS.items():
                count = val.count(ch)
                if count:
                    result["grey_zone_detected"][name] = result["grey_zone_detected"].get(name, 0) + count

    return result


def cleanup_invisible_chars(xml_path, dry_run=False):
    """
    Post-process: clean invisible Unicode characters from Str/Desc attributes.

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Dict with cleanup counts
    """
    try:
        tree, root = _parse_xml(xml_path)
        result = cleanup_invisible_chars_on_tree(root)

        total = result["spaces_normalized"] + result["invisibles_removed"]
        if total > 0 and not dry_run:
            _write_xml(tree, xml_path)
            logger.info(
                f"Post-process: cleaned {total} invisible chars in {xml_path.name} "
                f"({result['spaces_normalized']} spaces, {result['invisibles_removed']} deleted)"
            )

        return result
    except Exception as e:
        logger.error(f"Error cleaning invisible chars in {xml_path}: {e}")
        return {"spaces_normalized": 0, "invisibles_removed": 0, "invisible_detail": {}, "grey_zone_detected": {}}


# ─── Combined single-pass cleanup (for fast folder merge) ─────────────────


def run_all_postprocess_on_tree(root, language: str = "") -> dict:
    """
    Run ALL postprocess cleanup steps in a SINGLE iteration over LocStr elements.

    Combines newline normalization, empty StrOrigin enforcement,
    "no translation" replacement, apostrophe normalization, and ellipsis
    normalization (non-CJK only) into one pass.
    Used by _fast_folder_merge() to avoid re-parsing and re-iterating the tree.

    Args:
        root: Parsed XML root element (modified in-place)
        language: Language code (e.g. "FRE", "KOR"). CJK languages skip ellipsis step.

    Returns:
        {"changed": bool, "newlines_fixed": int, "empty_strorigin_cleaned": int,
         "no_translation_replaced": int, "apostrophes_normalized": int,
         "ellipsis_normalized": int}
    """
    result = {
        "changed": False,
        "newlines_fixed": 0,
        "empty_strorigin_cleaned": 0,
        "no_translation_replaced": 0,
        "apostrophes_normalized": 0,
        "hyphens_normalized": 0,
        "ellipsis_normalized": 0,
        "entities_decoded": 0,
        "spaces_normalized": 0,
        "invisibles_removed": 0,
        "invisible_detail": {},
        "grey_zone_detected": {},
    }

    skip_ellipsis = language.upper() in _CJK_SKIP_LANGS if language else False

    for loc in _iter_locstr(root):
        # --- Step 1: Normalize newlines in Str ---
        str_attr, str_val = _get_attr(loc, STR_ATTRS)
        if str_val is not None:
            normalized = _normalize_newlines(str_val)
            if normalized != str_val:
                loc.set(str_attr, normalized)
                result["newlines_fixed"] += 1
                result["changed"] = True
                str_val = normalized  # use cleaned value for later steps

        # --- Step 1b: Normalize newlines in Desc ---
        desc_attr, desc_val = _get_attr(loc, DESC_ATTRS)
        if desc_val is not None:
            desc_normalized = _normalize_newlines(desc_val)
            if desc_normalized != desc_val:
                loc.set(desc_attr, desc_normalized)
                result["newlines_fixed"] += 1
                result["changed"] = True
                desc_val = desc_normalized

        # --- Step 2: Empty StrOrigin enforcement ---
        _, origin = _get_attr(loc, STRORIGIN_ATTRS)
        origin_stripped = (origin or "").strip()
        str_val_stripped = (str_val or "").strip() if str_val is not None else ""

        if not origin_stripped and str_val_stripped:
            loc.set(str_attr or "Str", "")
            result["empty_strorigin_cleaned"] += 1
            result["changed"] = True
            str_val = ""  # cleared

        _, desc_origin = _get_attr(loc, DESCORIGIN_ATTRS)
        desc_origin_stripped = (desc_origin or "").strip()
        desc_val_stripped = (desc_val or "").strip() if desc_val is not None else ""

        if not desc_origin_stripped and desc_val_stripped:
            loc.set(desc_attr or "Desc", "")
            result["empty_strorigin_cleaned"] += 1
            result["changed"] = True
            desc_val = ""

        # --- Step 3: "no translation" replacement ---
        if str_val:
            check = _WHITESPACE_RE.sub(' ', str_val.strip()).lower()
            if check == 'no translation':
                if origin_stripped:
                    loc.set(str_attr if str_attr else "Str", origin_stripped)
                else:
                    loc.set(str_attr if str_attr else "Str", "")
                result["no_translation_replaced"] += 1
                result["changed"] = True

        if desc_val:
            desc_check = _WHITESPACE_RE.sub(' ', desc_val.strip()).lower()
            if desc_check == 'no translation':
                if desc_origin_stripped:
                    loc.set(desc_attr if desc_attr else "Desc", desc_origin_stripped)
                else:
                    loc.set(desc_attr if desc_attr else "Desc", "")
                result["no_translation_replaced"] += 1
                result["changed"] = True

        # --- Step 4: Normalize apostrophes in Str ---
        # Re-read current Str value (may have been modified by earlier steps)
        str_attr_4, str_val_4 = _get_attr(loc, STR_ATTRS)
        if str_val_4 is not None:
            apo_normalized = _normalize_apostrophes(str_val_4)
            if apo_normalized != str_val_4:
                loc.set(str_attr_4, apo_normalized)
                result["apostrophes_normalized"] += 1
                result["changed"] = True

        # --- Step 4b: Normalize apostrophes in Desc ---
        desc_attr_4, desc_val_4 = _get_attr(loc, DESC_ATTRS)
        if desc_val_4 is not None:
            desc_apo_normalized = _normalize_apostrophes(desc_val_4)
            if desc_apo_normalized != desc_val_4:
                loc.set(desc_attr_4, desc_apo_normalized)
                result["apostrophes_normalized"] += 1
                result["changed"] = True

        # --- Step 6: Normalize hyphens in Str ---
        str_attr_6, str_val_6 = _get_attr(loc, STR_ATTRS)
        if str_val_6 is not None:
            hyp_normalized = _normalize_hyphens(str_val_6)
            if hyp_normalized != str_val_6:
                loc.set(str_attr_6, hyp_normalized)
                result["hyphens_normalized"] += 1
                result["changed"] = True

        # --- Step 6b: Normalize hyphens in Desc ---
        desc_attr_6, desc_val_6 = _get_attr(loc, DESC_ATTRS)
        if desc_val_6 is not None:
            desc_hyp_normalized = _normalize_hyphens(desc_val_6)
            if desc_hyp_normalized != desc_val_6:
                loc.set(desc_attr_6, desc_hyp_normalized)
                result["hyphens_normalized"] += 1
                result["changed"] = True

        # --- Step 7: Normalize ellipsis in Str (non-CJK only) ---
        if not skip_ellipsis:
            str_attr_7, str_val_7 = _get_attr(loc, STR_ATTRS)
            if str_val_7 is not None:
                ell_normalized = _normalize_ellipsis(str_val_7)
                if ell_normalized != str_val_7:
                    loc.set(str_attr_7, ell_normalized)
                    result["ellipsis_normalized"] += 1
                    result["changed"] = True

            # --- Step 7b: Normalize ellipsis in Desc ---
            desc_attr_7, desc_val_7 = _get_attr(loc, DESC_ATTRS)
            if desc_val_7 is not None:
                desc_ell_normalized = _normalize_ellipsis(desc_val_7)
                if desc_ell_normalized != desc_val_7:
                    loc.set(desc_attr_7, desc_ell_normalized)
                    result["ellipsis_normalized"] += 1
                    result["changed"] = True

        # --- Step 8: Safe decode of double-escaped entities in Str ---
        str_attr_8, str_val_8 = _get_attr(loc, STR_ATTRS)
        if str_val_8 is not None:
            ent_decoded = _decode_safe_entities(str_val_8)
            if ent_decoded != str_val_8:
                loc.set(str_attr_8, ent_decoded)
                result["entities_decoded"] += 1
                result["changed"] = True

        # --- Step 8b: Safe decode of double-escaped entities in Desc ---
        desc_attr_8, desc_val_8 = _get_attr(loc, DESC_ATTRS)
        if desc_val_8 is not None:
            desc_ent_decoded = _decode_safe_entities(desc_val_8)
            if desc_ent_decoded != desc_val_8:
                loc.set(desc_attr_8, desc_ent_decoded)
                result["entities_decoded"] += 1
                result["changed"] = True

        # --- Step 5: Invisible character cleanup (Str) ---
        str_attr_5, str_val_5 = _get_attr(loc, STR_ATTRS)
        if str_val_5 is not None:
            cleaned, sp, dl, detail = _cleanup_invisible_chars(str_val_5)
            if sp or dl:
                loc.set(str_attr_5, cleaned)
                result["spaces_normalized"] += sp
                result["invisibles_removed"] += dl
                result["changed"] = True
                for name, cnt in detail.items():
                    result["invisible_detail"][name] = result["invisible_detail"].get(name, 0) + cnt
            # Grey zone detection (no modification)
            for ch, name in _GREY_ZONE_CHARS.items():
                count = str_val_5.count(ch)
                if count:
                    result["grey_zone_detected"][name] = result["grey_zone_detected"].get(name, 0) + count

        # --- Step 5b: Invisible character cleanup (Desc) ---
        desc_attr_5, desc_val_5 = _get_attr(loc, DESC_ATTRS)
        if desc_val_5 is not None:
            d_cleaned, d_sp, d_dl, d_detail = _cleanup_invisible_chars(desc_val_5)
            if d_sp or d_dl:
                loc.set(desc_attr_5, d_cleaned)
                result["spaces_normalized"] += d_sp
                result["invisibles_removed"] += d_dl
                result["changed"] = True
                for name, cnt in d_detail.items():
                    result["invisible_detail"][name] = result["invisible_detail"].get(name, 0) + cnt
            for ch, name in _GREY_ZONE_CHARS.items():
                count = desc_val_5.count(ch)
                if count:
                    result["grey_zone_detected"][name] = result["grey_zone_detected"].get(name, 0) + count

    return result


# ─── Unified Runner ─────────────────────────────────────────────────────────


def run_all_postprocess(xml_path: Path, dry_run: bool = False) -> dict:
    """
    Run ALL postprocess cleanup steps on an XML file.

    Single parse, all cleanups, single write. Efficient and atomic.

    Current steps (in order, applied to both Str and Desc):
      1. cleanup_wrong_newlines    - Normalize ALL newlines to <br/>
      2. cleanup_empty_strorigin   - Clear Str/Desc when StrOrigin/DescOrigin is empty
      3. cleanup_no_translation    - Replace "no translation" Str/Desc with origin
      4. cleanup_apostrophes       - Normalize curly/fancy apostrophes to ASCII
      5. cleanup_invisible_chars   - Auto-fix invisible Unicode chars (NBSP→space, delete zero-width)
      6. cleanup_hyphens           - Normalize Unicode hyphen lookalikes to ASCII hyphen-minus
      7. cleanup_ellipsis          - Normalize Unicode ellipsis to three dots (non-CJK only)

    Args:
        xml_path: Path to XML file
        dry_run: If True, count but don't write

    Returns:
        Dict with counts per cleanup step, e.g.:
        {"newlines_fixed": 5, "empty_strorigin_cleaned": 2, "total_fixes": 7}
    """
    result = {
        "newlines_fixed": 0,
        "empty_strorigin_cleaned": 0,
        "no_translation_replaced": 0,
        "apostrophes_normalized": 0,
        "hyphens_normalized": 0,
        "ellipsis_normalized": 0,
        "entities_decoded": 0,
        "spaces_normalized": 0,
        "invisibles_removed": 0,
        "invisible_detail": {},
        "grey_zone_detected": {},
        "total_fixes": 0,
    }

    try:
        tree, root = _parse_xml(xml_path)

        # Step 1: Normalize newlines (run FIRST so text is clean for later steps)
        result["newlines_fixed"] = cleanup_wrong_newlines_on_tree(root)

        # Step 2: Empty StrOrigin enforcement
        result["empty_strorigin_cleaned"] = cleanup_empty_strorigin_on_tree(root)

        # Step 3: Replace "no translation" with StrOrigin
        result["no_translation_replaced"] = cleanup_no_translation_on_tree(root)

        # Step 4: Normalize curly/fancy apostrophes to ASCII
        result["apostrophes_normalized"] = cleanup_apostrophes_on_tree(root)

        # Step 5: Clean invisible Unicode characters
        invis_result = cleanup_invisible_chars_on_tree(root)
        result["spaces_normalized"] = invis_result["spaces_normalized"]
        result["invisibles_removed"] = invis_result["invisibles_removed"]
        result["invisible_detail"] = invis_result["invisible_detail"]
        result["grey_zone_detected"] = invis_result["grey_zone_detected"]

        # Step 6: Normalize Unicode hyphen lookalikes to ASCII hyphen-minus
        result["hyphens_normalized"] = cleanup_hyphens_on_tree(root)

        # Step 7: Normalize Unicode ellipsis to three dots (non-CJK only)
        lang = _extract_language_from_path(xml_path)
        if lang not in _CJK_SKIP_LANGS:
            result["ellipsis_normalized"] = cleanup_ellipsis_on_tree(root)

        # Step 8: Safe decode of double-escaped entities
        result["entities_decoded"] = cleanup_double_escaped_on_tree(root)

        result["total_fixes"] = (
            result["newlines_fixed"]
            + result["empty_strorigin_cleaned"]
            + result["no_translation_replaced"]
            + result["apostrophes_normalized"]
            + result["hyphens_normalized"]
            + result["ellipsis_normalized"]
            + result["entities_decoded"]
            + result["spaces_normalized"]
            + result["invisibles_removed"]
        )

        # Write once if anything changed
        if result["total_fixes"] > 0 and not dry_run:
            _write_xml(tree, xml_path)

            if result["newlines_fixed"] > 0:
                logger.info(
                    f"Post-process: normalized {result['newlines_fixed']} "
                    f"wrong newlines in {xml_path.name}"
                )
            if result["empty_strorigin_cleaned"] > 0:
                logger.info(
                    f"Post-process: cleared Str on {result['empty_strorigin_cleaned']} "
                    f"entries with empty StrOrigin in {xml_path.name}"
                )
            if result["no_translation_replaced"] > 0:
                logger.info(
                    f"Post-process: replaced {result['no_translation_replaced']} "
                    f"'no translation' entries with StrOrigin in {xml_path.name}"
                )
            if result["apostrophes_normalized"] > 0:
                logger.debug(
                    "Normalized apostrophes in %d entries in %s",
                    result["apostrophes_normalized"], xml_path.name
                )
            if result["hyphens_normalized"] > 0:
                logger.debug(
                    "Normalized hyphens in %d entries in %s",
                    result["hyphens_normalized"], xml_path.name
                )
            if result["ellipsis_normalized"] > 0:
                logger.debug(
                    "Normalized ellipsis in %d entries in %s",
                    result["ellipsis_normalized"], xml_path.name
                )
            if result["entities_decoded"] > 0:
                logger.debug(
                    "Decoded double-escaped entities in %d entries in %s",
                    result["entities_decoded"], xml_path.name
                )
            invis_total = result["spaces_normalized"] + result["invisibles_removed"]
            if invis_total > 0:
                logger.info(
                    f"Post-process: cleaned {invis_total} invisible chars in {xml_path.name} "
                    f"({result['spaces_normalized']} spaces, {result['invisibles_removed']} deleted)"
                )

        return result
    except Exception as e:
        logger.error(f"Error in postprocess for {xml_path}: {e}")
        return result


# ─── Excel Pre-Process ────────────────────────────────────────────────────────


# Columns to clean in Excel (case-insensitive matching)
_EXCEL_CLEAN_COLUMNS = frozenset({
    "str", "correction", "desc", "description",
})

# Columns to NEVER touch (source data)
_EXCEL_SKIP_COLUMNS = frozenset({
    "stringid", "strorigin", "descorigin", "key", "id",
})


def run_preprocess_excel(xlsx_path: Path, dry_run: bool = False) -> dict:
    """
    Run cleanup on an Excel file — same normalizations as XML postprocess.

    Cleans Str/Correction/Desc columns (NOT StrOrigin/DescOrigin/StringID).

    Steps applied to each cell:
      1. Normalize linebreaks (including double-escaped &amp;lt;br/&amp;gt;)
      2. Normalize apostrophes (curly → ASCII)
      3. Clean invisible characters (NBSP → space, delete zero-width)
      4. Normalize hyphens (Unicode → ASCII hyphen-minus)
      5. Decode double-escaped entities (&amp;desc; → &desc;, &quot; → ", etc.)

    Args:
        xlsx_path: Path to Excel file
        dry_run: If True, count but don't write

    Returns:
        Dict with counts per cleanup step + total_fixes
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required for Excel preprocess — install with: pip install openpyxl")
        return {"total_fixes": 0, "error": "openpyxl not installed"}

    result = {
        "linebreaks_fixed": 0,
        "apostrophes_normalized": 0,
        "invisibles_cleaned": 0,
        "hyphens_normalized": 0,
        "entities_decoded": 0,
        "total_fixes": 0,
    }

    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
    except Exception as e:
        logger.error(f"Cannot open Excel file {xlsx_path}: {e}")
        result["error"] = f"Cannot open: {e}"
        return result

    try:
        for ws in wb.worksheets:
            # Find header row and identify cleanable columns
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
            if not header_row:
                continue

            clean_cols = []  # list of (col_index, col_name)
            for cell in header_row:
                if cell.value and isinstance(cell.value, str):
                    name_lower = cell.value.strip().lower()
                    if name_lower in _EXCEL_CLEAN_COLUMNS and name_lower not in _EXCEL_SKIP_COLUMNS:
                        clean_cols.append((cell.column, cell.value.strip()))

            if not clean_cols:
                continue

            # Process data rows
            for row in ws.iter_rows(min_row=2):
                for col_idx, col_name in clean_cols:
                    # Guard against ragged rows (fewer cells than header)
                    if col_idx - 1 >= len(row):
                        continue
                    cell = row[col_idx - 1]
                    if cell.value is None or not isinstance(cell.value, str):
                        continue

                    original = cell.value
                    text = original

                    try:
                        # Step 1: Normalize linebreaks (double-escaped + single-escaped + control chars)
                        text = re.sub(r'&amp;lt;/?[Bb][Rr]\s*/?&amp;gt;', '<br/>', text)
                        text = text.replace('&lt;br/&gt;', '<br/>')
                        text = text.replace('&lt;br /&gt;', '<br/>')
                        text = text.replace('<br />', '<br/>')
                        text = text.replace('\r\n', '<br/>')
                        text = text.replace('\r', '<br/>')
                        text = text.replace('\n', '<br/>')
                        text = text.replace('\\r\\n', '<br/>')
                        text = text.replace('\\n', '<br/>')
                        text = text.replace('\\r', '<br/>')
                        if text != original:
                            result["linebreaks_fixed"] += 1

                        # Step 2: Normalize apostrophes
                        apo = _normalize_apostrophes(text)
                        if apo != text:
                            result["apostrophes_normalized"] += 1
                        text = apo

                        # Step 3: Clean invisible characters
                        cleaned, sp, dl, _ = _cleanup_invisible_chars(text)
                        if sp or dl:
                            result["invisibles_cleaned"] += sp + dl
                        text = cleaned

                        # Step 4: Normalize hyphens
                        hyp = _normalize_hyphens(text)
                        if hyp != text:
                            result["hyphens_normalized"] += 1
                        text = hyp

                        # Step 5: Decode double-escaped entities
                        ent = _decode_safe_entities(text)
                        if ent != text:
                            result["entities_decoded"] += 1
                        text = ent

                        if text != original:
                            cell.value = text
                    except Exception as e:
                        logger.warning(f"Skipping cell row={cell.row} col={col_name}: {e}")
                        continue

        result["total_fixes"] = (
            result["linebreaks_fixed"]
            + result["apostrophes_normalized"]
            + result["invisibles_cleaned"]
            + result["hyphens_normalized"]
            + result["entities_decoded"]
        )

        if result["total_fixes"] > 0 and not dry_run:
            try:
                wb.save(str(xlsx_path))
            except PermissionError:
                logger.error(f"Cannot save {xlsx_path.name}: file is locked (close Excel first)")
                result["total_fixes"] = 0
                result["error"] = f"File locked: close {xlsx_path.name} in Excel and retry"
                return result
            except Exception as e:
                logger.error(f"Failed to save {xlsx_path.name}: {e}")
                result["total_fixes"] = 0
                result["error"] = f"Save failed: {e}"
                return result
            logger.info(
                f"Pre-process Excel: fixed {result['total_fixes']} issues in {xlsx_path.name} "
                f"(linebreaks={result['linebreaks_fixed']}, apostrophes={result['apostrophes_normalized']}, "
                f"invisibles={result['invisibles_cleaned']}, hyphens={result['hyphens_normalized']}, "
                f"entities={result['entities_decoded']})"
            )

        return result
    except Exception as e:
        logger.error(f"Error in Excel preprocess for {xlsx_path}: {e}")
        result["error"] = str(e)
        return result
    finally:
        wb.close()
