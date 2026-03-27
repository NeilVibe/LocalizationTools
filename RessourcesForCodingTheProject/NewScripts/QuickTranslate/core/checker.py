"""
Pre-Submission Quality Checks.

Korean detection and pattern code mismatch checking for languagedata files.
Supports both XML and Excel source files (mixed sources per language).
Scans Source folder, groups by language, writes per-language results.

Output format: pure LocStr elements in <root>, same format as source XML.
"""
from __future__ import annotations

import logging
import re
import threading
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

from .xml_parser import (
    parse_xml_file, iter_locstr_elements,
    get_attr as _get_attr,
    STRINGID_ATTRS as _STRINGID_ATTRS,
    STRORIGIN_ATTRS as _STRORIGIN_ATTRS,
    STR_ATTRS as _STR_ATTRS,
    DESC_ATTRS as _DESC_ATTRS,
    DESCORIGIN_ATTRS as _DESCORIGIN_ATTRS,
)
from .korean_detection import is_korean_text
from .source_scanner import scan_source_for_languages
from .text_utils import is_formula_text, is_text_integrity_issue

logger = logging.getLogger(__name__)


def should_skip_locstr(elem, skip_staticinfo_knowledge: bool = True) -> bool:
    """
    Return True if LocStr should be skipped (staticinfo:knowledge filter).

    When skip_staticinfo_knowledge is True (default), any LocStr where Str OR
    StrOrigin contains 'staticinfo:knowledge' (case insensitive) is skipped
    because pattern codes in those entries cause false positives.

    This filter is for Pattern Check ONLY.
    Korean Check never calls this function (Korean = always scan everything).
    Quality Check currently uses this too, but may be removed in the future
    so that only Pattern Check has the toggle.
    """
    if not skip_staticinfo_knowledge:
        return False
    str_text = _get_attr(elem, _STR_ATTRS)
    strorigin_text = _get_attr(elem, _STRORIGIN_ATTRS)
    return ("staticinfo:knowledge" in str_text.lower()
            or "staticinfo:knowledge" in strorigin_text.lower())


def extract_code_patterns(text: str) -> Set[str]:
    """Extract {code} patterns from text."""
    return set(re.findall(r'\{.*?\}', text))


def normalize_staticinfo_pattern(code: str) -> str:
    """
    Normalize staticinfo patterns by stripping variable parts after #.

    {Staticinfo:Knowledge#123} -> {Staticinfo:Knowledge#}

    Ported from checkpatternerror.py battle-tested logic.
    """
    if re.search(r'\{[^{}]*Staticinfo:[^{}]*#', code, re.I):
        return code.split('#', 1)[0] + '#}'
    return code


def normalize_pattern_set(raw_set: Set[str]) -> Set[str]:
    """Normalize all patterns in a set."""
    return {normalize_staticinfo_pattern(p) for p in raw_set}


# Regex to find any <br...> tag variant (we then check if it's exactly <br/>)
_BR_TAG_RE = re.compile(r'<br\s*/?\s*>', re.IGNORECASE)


    # NOTE: _has_wrong_newlines was removed — all newline variants are auto-fixed
    # by transfer + postprocess. Only broken linebreaks (is_broken_linebreak) matter.


def _has_unbalanced_brackets(text: str) -> bool:
    """
    Check if curly brackets in Str are properly paired and nested.

    Uses a stack approach: '{' pushes, '}' pops. Catches:
    - Missing closing bracket: {code without }
    - Missing opening bracket: } without {
    - Wrong nesting: }text{

    Only checks Str (translation) — StrOrigin is assumed correct.
    """
    if not text:
        return False
    depth = 0
    for ch in text:
        if ch == '{':
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth < 0:
                return True  # closing without opening
    return depth != 0  # unclosed opening brackets





def _escape_attr_value(value: str) -> str:
    """Escape attribute value for XML output."""
    return (value
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;'))


def _elem_to_locstr_line(elem) -> str:
    """Reconstruct a LocStr element as an XML string from its attributes."""
    attribs = []
    for key, value in elem.attrib.items():
        attribs.append(f'{key}="{_escape_attr_value(value)}"')
    attrib_str = ' '.join(attribs)
    return f'  <LocStr {attrib_str} />'


def check_formula_text_in_file(xml_path: Path) -> list:
    """
    Scan one XML file for formula-like text in Str and Desc values.

    Catches Excel formulas (=VLOOKUP, +SUM, etc.), error values (#REF!, #N/A),
    and openpyxl repr leaks that may have been written into XML previously.

    Args:
        xml_path: Path to XML file

    Returns:
        List of matching LocStr elements (lxml elements with original attributes).
    """
    findings = []
    try:
        root = parse_xml_file(xml_path)
        for elem in iter_locstr_elements(root):
            str_text = _get_attr(elem, _STR_ATTRS).strip()
            desc_text = _get_attr(elem, _DESC_ATTRS).strip()

            if str_text and is_formula_text(str_text):
                findings.append(elem)
                continue
            if desc_text and is_formula_text(desc_text):
                findings.append(elem)
    except Exception as e:
        logger.warning(f"Failed to parse {xml_path.name}: {e}")

    return findings


def check_text_integrity_in_file(xml_path: Path) -> list:
    """
    Scan one XML file for text integrity issues in Str and Desc values.

    Catches broken <br/> tags, encoding artifacts, control characters,
    and zero-width/invisible characters.

    Args:
        xml_path: Path to XML file

    Returns:
        List of matching LocStr elements (lxml elements with original attributes).
    """
    findings = []
    try:
        root = parse_xml_file(xml_path)
        for elem in iter_locstr_elements(root):
            str_text = _get_attr(elem, _STR_ATTRS).strip()
            desc_text = _get_attr(elem, _DESC_ATTRS).strip()
            str_origin = _get_attr(elem, _STRORIGIN_ATTRS).strip()
            desc_origin = _get_attr(elem, _DESCORIGIN_ATTRS).strip()

            if str_text and is_text_integrity_issue(str_text, from_xml=True, source_text=str_origin):
                findings.append(elem)
                continue
            if desc_text and is_text_integrity_issue(desc_text, from_xml=True, source_text=desc_origin):
                findings.append(elem)
    except Exception as e:
        logger.warning(f"Failed to parse {xml_path.name}: {e}")

    return findings


def check_korean_in_file(xml_path: Path) -> list:
    """
    Scan one XML file for Korean characters in Str values.

    Args:
        xml_path: Path to XML file

    Returns:
        List of matching LocStr elements (lxml elements with original attributes).
    """
    findings = []
    try:
        root = parse_xml_file(xml_path)
        for elem in iter_locstr_elements(root):
            str_text = _get_attr(elem, _STR_ATTRS).strip()
            if not str_text:
                continue

            if is_korean_text(str_text):
                findings.append(elem)
    except Exception as e:
        logger.warning(f"Failed to parse {xml_path.name}: {e}")

    return findings


def check_patterns_in_file(
    xml_path: Path,
    skip_staticinfo_knowledge: bool = True,
) -> Tuple[list, list, list]:
    """
    Scan one XML file for pattern code mismatches,
    unbalanced brackets, and empty Str with StrOrigin.

    Compares normalized {code} pattern sets. If they differ, it's a pattern error.
    Also checks for wrong newline representations (only <br/> is correct).
    Also checks for unbalanced curly brackets in Str (missing { or }).
    Also flags entries where Str is empty but StrOrigin exists (untranslated).

    Args:
        xml_path: Path to XML file
        skip_staticinfo_knowledge: If True, skip entries containing 'staticinfo:knowledge'

    Returns:
        Tuple of (pattern_errors, bracket_errors, empty_str_errors).
    """
    pattern_errors = []
    bracket_errors = []
    empty_str_errors = []
    try:
        root = parse_xml_file(xml_path)
        for elem in iter_locstr_elements(root):
            str_text = _get_attr(elem, _STR_ATTRS).strip()
            if not str_text:
                # Empty Str check runs on ALL entries (no staticinfo skip) —
                # an untranslated entry is always worth flagging.
                strorigin_for_empty = _get_attr(elem, _STRORIGIN_ATTRS).strip()
                if strorigin_for_empty:
                    empty_str_errors.append(elem)
                continue

            # Bracket check runs on ALL entries — unbalanced brackets are
            # always critical, even in staticinfo:knowledge entries.
            if _has_unbalanced_brackets(str_text):
                bracket_errors.append(elem)

            # Pattern check respects the staticinfo skip filter
            if should_skip_locstr(elem, skip_staticinfo_knowledge):
                continue

            strorigin_text = _get_attr(elem, _STRORIGIN_ATTRS).strip()
            if not strorigin_text:
                continue

            # Pattern mismatch check
            origin_patterns = normalize_pattern_set(extract_code_patterns(strorigin_text))
            str_patterns = normalize_pattern_set(extract_code_patterns(str_text))

            if origin_patterns != str_patterns:
                pattern_errors.append(elem)

            # NOTE: Wrong newlines are NOT checked here — all newline variants
            # are auto-fixed by transfer + postprocess. Only truly broken
            # linebreaks (malformed tags) are caught by the integrity check.
    except Exception as e:
        logger.warning(f"Failed to parse {xml_path.name}: {e}")

    return pattern_errors, bracket_errors, empty_str_errors


# ─── XML Health Check (auto-validation on upload) ─────────────────────────


def check_xml_health(xml_path: Path) -> dict:
    """Full health check on an XML languagedata file.

    Runs automatically when source/target folder is selected.
    Catches damage already present in the XML — not just corrections.

    Returns dict with categorized findings:
        empty_str: [(stringid, str_origin_preview)]  — Str empty but StrOrigin exists
        integrity_str: [(stringid, reason)]            — Str has integrity issue
        integrity_desc: [(stringid, reason)]           — Desc has integrity issue
        no_translation: [(stringid, str_origin_preview)] — Str is "no translation"
        formula_str: [(stringid, reason)]              — Str has formula/garbage text
    """
    from .text_utils import _SAFE_INVISIBLE_DELETE

    result = {
        "empty_str": [],
        "integrity_str": [],
        "integrity_desc": [],
        "no_translation": [],
        "formula_str": [],
    }

    try:
        root = parse_xml_file(xml_path)
        for elem in iter_locstr_elements(root):
            sid = _get_attr(elem, _STRINGID_ATTRS).strip() or "(no ID)"
            str_text = _get_attr(elem, _STR_ATTRS)
            str_origin = _get_attr(elem, _STRORIGIN_ATTRS).strip()
            desc_text = _get_attr(elem, _DESC_ATTRS)
            desc_origin = _get_attr(elem, _DESCORIGIN_ATTRS).strip()

            # --- Str checks ---
            str_stripped = (str_text or "").strip()

            if not str_stripped and str_origin:
                # Empty after basic strip — flag immediately
                result["empty_str"].append((sid, str_origin[:60]))
                continue

            if str_stripped:
                # Check if Str is only invisible chars (would become empty after cleanup)
                visible = str_stripped
                for ch in _SAFE_INVISIBLE_DELETE:
                    visible = visible.replace(ch, '')
                visible = visible.strip()
                if not visible and str_origin:
                    result["empty_str"].append((sid, str_origin[:60]))
                    continue

                # "no translation" in Str
                if ' '.join(str_stripped.split()).lower() == 'no translation':
                    result["no_translation"].append((sid, str_origin[:60] if str_origin else ""))
                    continue

                # Formula/garbage text in Str
                formula_reason = is_formula_text(str_stripped)
                if formula_reason:
                    result["formula_str"].append((sid, formula_reason))
                    continue

                # Text integrity (tabs, U+FFFC, control chars, broken br, etc.)
                reason = is_text_integrity_issue(str_stripped, from_xml=True,
                                                 source_text=str_origin)
                if reason and not reason.startswith('Warning:'):
                    result["integrity_str"].append((sid, reason))

            # --- Desc checks ---
            desc_stripped = (desc_text or "").strip()
            if desc_stripped:
                reason = is_text_integrity_issue(desc_stripped, from_xml=True,
                                                 source_text=desc_origin)
                if reason and not reason.startswith('Warning:'):
                    result["integrity_desc"].append((sid, reason))

    except Exception as e:
        logger.warning("Health check failed for %s: %s", xml_path.name, e)

    return result


# ─── Excel Support ─────────────────────────────────────────────────────────


class _CheckRow:
    """Lightweight stand-in for an lxml element, used for Excel-sourced findings.

    Compatible with get_attr() and _elem_to_locstr_line() so all existing
    report generation code works unchanged.
    """
    def __init__(self, string_id: str = "", str_origin: str = "",
                 str_val: str = "", desc_val: str = ""):
        self.attrib = {}
        if string_id:
            self.attrib["StringId"] = str(string_id)
        if str_origin:
            self.attrib["StrOrigin"] = str(str_origin)
        if str_val:
            self.attrib["Str"] = str(str_val)
        if desc_val:
            self.attrib["Desc"] = str(desc_val)

    def get(self, key, default=None):
        return self.attrib.get(key, default)


    # NOTE: _has_wrong_newlines_excel was removed — all newline variants
    # are auto-fixed by transfer + postprocess. Only broken linebreaks
    # (malformed tags caught by is_broken_linebreak) matter.


def _resolve_excel_col(col_indices: dict, *names: str) -> Optional[int]:
    """Find a column index by trying multiple name variants."""
    for name in names:
        if name in col_indices:
            return col_indices[name]
    return None


def check_patterns_in_excel(
    xlsx_path: Path,
    skip_staticinfo_knowledge: bool = True,
) -> Tuple[list, list, list, list, list]:
    """
    Run all pattern checks on one Excel file.

    Same checks as check_patterns_in_file but reads from Excel columns:
    StringID, StrOrigin, Str/Correction, Desc.

    Returns:
        Tuple of (pattern_errors, bracket_errors,
                  empty_str_errors, formula_findings, integrity_findings).
    """
    pattern_errors = []
    bracket_errors = []
    empty_str_errors = []
    formula_findings = []
    integrity_findings = []

    _empty = (pattern_errors, bracket_errors, empty_str_errors, formula_findings, integrity_findings)

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error(
            "openpyxl is required for Excel pattern checks but is not installed. "
            "Excel file %s was NOT checked. Install with: pip install openpyxl",
            xlsx_path.name
        )
        return _empty

    try:
        wb = load_workbook(str(xlsx_path), read_only=True)
    except Exception as e:
        logger.error(
            f"Failed to open Excel file {xlsx_path.name}: {e}. "
            f"Check that the file is not corrupted, password-protected, or open in another program."
        )
        return _empty

    try:
        ws = wb.active
        if ws is None:
            logger.warning(f"No active worksheet in {xlsx_path.name} — skipping Excel pattern check")
            return _empty

        from .excel_io import _detect_column_indices
        col_indices = _detect_column_indices(ws)

        # Find columns
        sid_col = _resolve_excel_col(col_indices, "stringid", "string_id")
        strorigin_col = _resolve_excel_col(col_indices, "strorigin", "str_origin")
        str_col = _resolve_excel_col(col_indices, "str", "correction", "corrected")
        desc_col = _resolve_excel_col(col_indices, "desc", "description", "desccorrection")

        if not str_col:
            logger.warning(
                f"No Str/Correction column found in {xlsx_path.name} — "
                f"available columns: {list(col_indices.keys())}. Skipping pattern check."
            )
            return _empty

        for row in ws.iter_rows(min_row=2, values_only=False):
            # Extract cell values safely
            def _cell_str(col_idx):
                if col_idx is None or col_idx - 1 >= len(row):
                    return ""
                val = row[col_idx - 1].value
                return str(val).strip() if val is not None else ""

            string_id = _cell_str(sid_col)
            str_origin = _cell_str(strorigin_col)
            str_val = _cell_str(str_col)
            desc_val = _cell_str(desc_col)

            # Skip completely empty rows
            if not str_val and not str_origin:
                continue

            row_obj = _CheckRow(string_id, str_origin, str_val, desc_val)

            # --- Empty Str check ---
            if not str_val:
                if str_origin:
                    empty_str_errors.append(row_obj)
                continue

            # --- Bracket check (always runs — critical) ---
            if _has_unbalanced_brackets(str_val):
                bracket_errors.append(row_obj)

            # --- staticinfo:knowledge filter ---
            if skip_staticinfo_knowledge:
                if ("staticinfo:knowledge" in str_val.lower()
                        or "staticinfo:knowledge" in str_origin.lower()):
                    continue

            # --- Pattern mismatch check ---
            if str_origin:
                origin_patterns = normalize_pattern_set(extract_code_patterns(str_origin))
                str_patterns = normalize_pattern_set(extract_code_patterns(str_val))
                if origin_patterns != str_patterns:
                    pattern_errors.append(row_obj)

            # NOTE: Wrong newlines are NOT checked — all newline variants
            # are auto-fixed by transfer + postprocess.

            # --- Formula text check ---
            formula_reason = is_formula_text(str_val) or is_formula_text(desc_val)
            if formula_reason:
                formula_findings.append(row_obj)

            # --- Text integrity check (from_xml=False for Excel) ---
            # No normalization needed — is_text_integrity_issue does NOT
            # flag bare \n or \r (excluded from _CONTROL_CHARS_RE).
            # Broken linebreaks and unfixable entities are still caught.
            # Pass str_origin so lone bracket check can compare source vs translation.
            # Note: Desc uses None (no DescOrigin column in QA check path).
            integrity_reason = (
                is_text_integrity_issue(str_val, from_xml=False, source_text=str_origin)
                or (desc_val and is_text_integrity_issue(desc_val, from_xml=False))
            )
            if integrity_reason:
                integrity_findings.append(row_obj)

    except Exception as e:
        partial = (len(pattern_errors) + len(bracket_errors)
                   + len(empty_str_errors) + len(formula_findings) + len(integrity_findings))
        logger.error(
            f"Error checking Excel {xlsx_path.name}: {e}. "
            f"Check was INCOMPLETE — {partial} findings collected before error.",
            exc_info=True
        )
    finally:
        wb.close()

    return pattern_errors, bracket_errors, empty_str_errors, formula_findings, integrity_findings


def iter_source_files(source_folder: Path) -> Tuple[Dict[str, List[Path]], Dict[str, List[Path]]]:
    """
    Discover all XML and Excel files in Source folder, grouped by language.

    Returns:
        Tuple of (xml_by_lang, xlsx_by_lang) dicts.
        Example: ({"FRE": [xml1, xml2]}, {"FRE": [xlsx1]})
    """
    scan_result = scan_source_for_languages(source_folder)

    xml_by_lang = {}
    xlsx_by_lang = {}
    for lang, files in scan_result.lang_files.items():
        xml_files = [f for f in files if f.suffix.lower() == ".xml"]
        xlsx_files = [f for f in files if f.suffix.lower() == ".xlsx"]
        xls_files = [f for f in files if f.suffix.lower() == ".xls"]
        if xls_files:
            logger.warning(f"Legacy .xls files for {lang} not supported (use .xlsx): "
                           f"{[f.name for f in xls_files]}")
        if xml_files:
            xml_by_lang[lang] = xml_files
        if xlsx_files:
            xlsx_by_lang[lang] = xlsx_files

    return xml_by_lang, xlsx_by_lang


def iter_source_xml_files(source_folder: Path) -> Dict[str, List[Path]]:
    """
    Discover all XML files in Source folder, grouped by language.

    Uses the existing source_scanner to detect language codes from
    folder/file naming conventions.

    Args:
        source_folder: Path to Source folder

    Returns:
        Dict mapping language code to list of XML file paths.
        Example: {"ENG": [path1, path2], "FRE": [path3], ...}
    """
    scan_result = scan_source_for_languages(source_folder)

    # Filter to XML files only
    xml_by_lang = {}
    for lang, files in scan_result.lang_files.items():
        xml_files = [f for f in files if f.suffix.lower() == ".xml"]
        if xml_files:
            xml_by_lang[lang] = xml_files

    return xml_by_lang


def _write_check_results_excel(path: Path, sheets_data: list):
    """Write unified check results Excel with multiple tabs.

    Args:
        path: Output .xlsx file path
        sheets_data: List of (sheet_name, headers, rows) tuples.
                     Only sheets with rows should be included.
                     Each row is a tuple matching the headers.
    """
    import xlsxwriter
    wb = xlsxwriter.Workbook(str(path))
    header_fmt = wb.add_format({'bold': True, 'bg_color': '#D9E1F2'})
    for sheet_name, headers, rows in sheets_data:
        ws = wb.add_worksheet(sheet_name)
        for col, h in enumerate(headers):
            ws.write(0, col, h, header_fmt)
        for r, row_data in enumerate(rows, 1):
            for col, val in enumerate(row_data):
                ws.write(r, col, val or "")
        for col in range(len(headers)):
            ws.set_column(col, col, 40)
    wb.close()


def _write_results_xml(output_path: Path, elements: list):
    """
    Write LocStr elements to XML file in pure source format.

    Output: <root> with LocStr elements preserving original attributes.
    Same format as missing_translation_finder and checkpatternerror.py.
    """
    lines = ['<?xml version="1.0" encoding="utf-8"?>']
    lines.append('<root>')

    for elem in elements:
        lines.append(_elem_to_locstr_line(elem))

    lines.append('</root>')

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text('\n'.join(lines), encoding='utf-8')


def run_korean_check(
    source_folder: Path,
    output_folder: Path,
    progress_callback: Optional[Callable[[str], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict[str, int]:
    """
    Run Korean character check on all languages in Source folder.

    Scans all languagedata XML files, finds Korean in Str values,
    writes per-language result XMLs to output_folder/Korean/.

    Args:
        source_folder: Path to Source folder with language subfolders
        output_folder: Path to CheckResults folder
        progress_callback: Optional callback for progress updates
        cancel_event: Optional threading.Event to support cancellation

    Returns:
        Summary dict: {"FRE": 5, "GER": 0, ...} (finding counts per language)
    """
    xml_by_lang = iter_source_xml_files(source_folder)
    if not xml_by_lang:
        if progress_callback:
            progress_callback("No XML files found in Source folder")
        return {}

    korean_dir = output_folder / "Korean"
    korean_dir.mkdir(parents=True, exist_ok=True)

    languages = sorted(xml_by_lang.keys())
    summary = {}

    for i, lang in enumerate(languages):
        if cancel_event and cancel_event.is_set():
            raise InterruptedError("Operation cancelled by user")
        xml_files = xml_by_lang[lang]
        if progress_callback:
            progress_callback(f"Checking Korean... ({i + 1}/{len(languages)} languages: {lang})")

        all_findings = []
        for xml_path in xml_files:
            if cancel_event and cancel_event.is_set():
                raise InterruptedError("Operation cancelled by user")
            all_findings.extend(check_korean_in_file(xml_path))

        summary[lang] = len(all_findings)

        if all_findings:
            out_path = korean_dir / f"korean_findings_{lang}.xml"
            _write_results_xml(out_path, all_findings)
            logger.info(f"Korean check {lang}: {len(all_findings)} findings in {len(xml_files)} files -> {out_path.name}")
        else:
            logger.info(f"Korean check {lang}: clean ({len(xml_files)} files)")

    return summary


def run_pattern_check(
    source_folder: Path,
    output_folder: Path,
    progress_callback: Optional[Callable[[str], None]] = None,
    skip_staticinfo_knowledge: bool = True,
    cancel_event: Optional[threading.Event] = None,
) -> Dict[str, Tuple[int, int, int, int, int, int]]:
    """
    Run pattern mismatch + bracket + empty Str + formula + integrity check.

    Only flags issues that CANNOT be auto-fixed by transfer + postprocess.
    Wrong newlines are NOT checked — all newline variants are auto-fixable.

    Checks:
    - Pattern code mismatch between StrOrigin and Str
    - Unbalanced curly brackets (critical)
    - Empty Str with StrOrigin (untranslated)
    - Formula-like text in Str/Desc (leaked Excel formulas/errors)
    - Text integrity (broken linebreaks, encoding artifacts, unfixable entities)

    Lone bracket warnings (matching source) are counted separately and only
    appear in the LowImpactWarnings Excel tab, not in the TextIntegrity folder.

    Writes per-language result XMLs to:
    - output_folder/PatternErrors/pattern_errors_{LANG}.xml  (all issues combined)
    - output_folder/MissingBrackets/MissingBrackets_{LANG}.xml  (critical bracket issues only)
    - output_folder/EmptyStr/EmptyStr_{LANG}.xml  (empty Str with StrOrigin)
    - output_folder/FormulaText/FormulaText_{LANG}.xml  (formula-like text in Str/Desc)
    - output_folder/TextIntegrity/TextIntegrity_{LANG}.xml  (real integrity issues only)

    Args:
        source_folder: Path to Source folder with language subfolders
        output_folder: Path to CheckResults folder
        progress_callback: Optional callback for progress updates
        skip_staticinfo_knowledge: If True, skip entries containing 'staticinfo:knowledge'
        cancel_event: Optional threading.Event to support cancellation

    Returns:
        Summary dict: {"FRE": (pattern, bracket, empty_str, formula_text, integrity, warnings), ...}
    """
    xml_by_lang, xlsx_by_lang = iter_source_files(source_folder)
    all_langs = sorted(set(list(xml_by_lang.keys()) + list(xlsx_by_lang.keys())))
    if not all_langs:
        if progress_callback:
            progress_callback("No XML or Excel files found in Source folder")
        return {}

    pattern_dir = output_folder / "PatternErrors"
    pattern_dir.mkdir(parents=True, exist_ok=True)
    bracket_dir = output_folder / "MissingBrackets"
    bracket_dir.mkdir(parents=True, exist_ok=True)
    empty_dir = output_folder / "EmptyStr"
    empty_dir.mkdir(parents=True, exist_ok=True)
    formula_dir = output_folder / "FormulaText"
    formula_dir.mkdir(parents=True, exist_ok=True)
    integrity_dir = output_folder / "TextIntegrity"
    integrity_dir.mkdir(parents=True, exist_ok=True)

    languages = all_langs
    summary = {}

    for i, lang in enumerate(languages):
        if cancel_event and cancel_event.is_set():
            raise InterruptedError("Operation cancelled by user")
        xml_files = xml_by_lang.get(lang, [])
        xlsx_files = xlsx_by_lang.get(lang, [])
        file_count = len(xml_files) + len(xlsx_files)
        if progress_callback:
            progress_callback(f"Checking patterns... ({i + 1}/{len(languages)} languages: {lang})")

        all_pattern_errors = []
        all_bracket_errors = []
        all_empty_str = []
        all_formula_text = []
        all_integrity = []

        # --- Process XML files ---
        for xml_path in xml_files:
            if cancel_event and cancel_event.is_set():
                raise InterruptedError("Operation cancelled by user")
            p_errors, b_errors, e_errors = check_patterns_in_file(xml_path, skip_staticinfo_knowledge)
            all_pattern_errors.extend(p_errors)
            all_bracket_errors.extend(b_errors)
            all_empty_str.extend(e_errors)

            # Formula text check — catches leaked Excel formulas/errors in Str/Desc
            formula = check_formula_text_in_file(xml_path)
            all_formula_text.extend(formula)

            # Text integrity check — broken linebreaks, encoding artifacts, bad chars
            integrity = check_text_integrity_in_file(xml_path)
            all_integrity.extend(integrity)

        # --- Process Excel files ---
        for xlsx_path in xlsx_files:
            if cancel_event and cancel_event.is_set():
                raise InterruptedError("Operation cancelled by user")
            xl_pattern, xl_bracket, xl_empty, xl_formula, xl_integrity = \
                check_patterns_in_excel(xlsx_path, skip_staticinfo_knowledge)
            all_pattern_errors.extend(xl_pattern)
            all_bracket_errors.extend(xl_bracket)
            all_empty_str.extend(xl_empty)
            all_formula_text.extend(xl_formula)
            all_integrity.extend(xl_integrity)

        # Split integrity into real issues vs low-impact warnings (lone brackets matching source)
        real_integrity = []
        warning_integrity = []
        for elem in all_integrity:
            str_val = _get_attr(elem, _STR_ATTRS)
            desc_val = _get_attr(elem, _DESC_ATTRS)
            str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
            desc_origin = _get_attr(elem, _DESCORIGIN_ATTRS)
            reason = (is_text_integrity_issue(str_val, from_xml=True, source_text=str_origin)
                      or is_text_integrity_issue(desc_val, from_xml=True, source_text=desc_origin)
                      or "")
            if reason.startswith('Warning:'):
                warning_integrity.append(elem)
            else:
                real_integrity.append(elem)

        # Summary: (pattern, bracket, empty_str, formula_text, integrity, warnings)
        summary[lang] = (len(all_pattern_errors), len(all_bracket_errors),
                         len(all_empty_str), len(all_formula_text),
                         len(real_integrity), len(warning_integrity))

        # Combine all error lists for main XML output (deduplicate by element identity)
        seen = set()
        combined = []
        for elem in all_pattern_errors + all_bracket_errors + all_empty_str + all_formula_text + real_integrity:
            eid = id(elem)
            if eid not in seen:
                seen.add(eid)
                combined.append(elem)

        if combined:
            out_path = pattern_dir / f"pattern_errors_{lang}.xml"
            _write_results_xml(out_path, combined)
            p_count = len(all_pattern_errors)
            b_count = len(all_bracket_errors)
            e_count = len(all_empty_str)
            f_count = len(all_formula_text)
            i_count = len(real_integrity)
            logger.info(f"Pattern check {lang}: {p_count} pattern + {b_count} bracket + {e_count} empty + {f_count} formula + {i_count} integrity errors in {file_count} files -> {out_path.name}")
        else:
            logger.info(f"Pattern check {lang}: clean ({file_count} files)")

        # Write separate critical bracket file
        if all_bracket_errors:
            bracket_path = bracket_dir / f"MissingBrackets_{lang}.xml"
            _write_results_xml(bracket_path, all_bracket_errors)
            logger.info(f"Bracket check {lang}: {len(all_bracket_errors)} CRITICAL entries -> {bracket_path.name}")

        # Write separate empty Str file
        if all_empty_str:
            empty_path = empty_dir / f"EmptyStr_{lang}.xml"
            _write_results_xml(empty_path, all_empty_str)
            logger.info(f"Empty Str {lang}: {len(all_empty_str)} entries -> {empty_path.name}")

        # Write separate formula text XML
        if all_formula_text:
            formula_path = formula_dir / f"FormulaText_{lang}.xml"
            _write_results_xml(formula_path, all_formula_text)
            logger.info(f"Formula text {lang}: {len(all_formula_text)} CRITICAL entries -> {formula_path.name}")

        # Write separate text integrity XML (real issues only, no lone bracket warnings)
        if real_integrity:
            integrity_path = integrity_dir / f"TextIntegrity_{lang}.xml"
            _write_results_xml(integrity_path, real_integrity)
            logger.info(f"Text integrity {lang}: {len(real_integrity)} entries -> {integrity_path.name}")

        # --- Unified Excel report: one file per language, tabs only if findings ---
        elem_headers = ("StringID", "StrOrigin", "Str", "Reason")
        sheets = []

        # Tab 1: Critical — formula text + broken/truncated linebreak integrity
        critical_rows = []
        secondary_rows = []
        for elem in all_formula_text:
            sid = _get_attr(elem, _STRINGID_ATTRS)
            str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
            str_val = _get_attr(elem, _STR_ATTRS)
            desc_val = _get_attr(elem, _DESC_ATTRS)
            reason = is_formula_text(str_val) or is_formula_text(desc_val) or "Unknown"
            critical_rows.append((sid, str_origin, str_val, reason))
        for elem in real_integrity:
            str_val = _get_attr(elem, _STR_ATTRS)
            desc_val = _get_attr(elem, _DESC_ATTRS)
            str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
            desc_origin = _get_attr(elem, _DESCORIGIN_ATTRS)
            reason = (is_text_integrity_issue(str_val, from_xml=True, source_text=str_origin)
                      or is_text_integrity_issue(desc_val, from_xml=True, source_text=desc_origin)
                      or "Unknown")
            sid = _get_attr(elem, _STRINGID_ATTRS)
            if reason.startswith('Broken') or reason.startswith('Truncated'):
                critical_rows.append((sid, str_origin, str_val, reason))
            else:
                secondary_rows.append((sid, str_origin, str_val, reason))
        if critical_rows:
            sheets.append(("Critical", elem_headers, critical_rows))

        # Tab 2: Secondary — encoding artifacts, invisible chars, control chars
        if secondary_rows:
            sheets.append(("Secondary", elem_headers, secondary_rows))

        # Tab 2b: LowImpactWarnings — lone brackets matching source (low-impact)
        if warning_integrity:
            lone_bracket_rows = []
            for elem in warning_integrity:
                sid = _get_attr(elem, _STRINGID_ATTRS)
                str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
                str_val = _get_attr(elem, _STR_ATTRS)
                desc_val = _get_attr(elem, _DESC_ATTRS)
                desc_origin = _get_attr(elem, _DESCORIGIN_ATTRS)
                reason = (is_text_integrity_issue(str_val, from_xml=True, source_text=str_origin)
                          or is_text_integrity_issue(desc_val, from_xml=True, source_text=desc_origin)
                          or "Warning: Unknown")
                lone_bracket_rows.append((sid, str_origin, str_val, reason))
            sheets.append(("LowImpactWarnings", elem_headers, lone_bracket_rows))

        # Tab 3: PatternErrors
        if all_pattern_errors:
            rows = []
            for elem in all_pattern_errors:
                sid = _get_attr(elem, _STRINGID_ATTRS)
                str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
                str_val = _get_attr(elem, _STR_ATTRS)
                rows.append((sid, str_origin, str_val, "Pattern mismatch"))
            sheets.append(("PatternErrors", elem_headers, rows))

        # NOTE: WrongNewlines tab removed — all newline variants are auto-fixed
        # by transfer + postprocess. Only broken linebreaks (integrity check) matter.

        # Tab 4: MissingBrackets (was Tab 5)
        if all_bracket_errors:
            rows = []
            for elem in all_bracket_errors:
                sid = _get_attr(elem, _STRINGID_ATTRS)
                str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
                str_val = _get_attr(elem, _STR_ATTRS)
                rows.append((sid, str_origin, str_val, "Unbalanced brackets"))
            sheets.append(("MissingBrackets", elem_headers, rows))

        # Tab 6: EmptyStr
        if all_empty_str:
            rows = []
            for elem in all_empty_str:
                sid = _get_attr(elem, _STRINGID_ATTRS)
                str_origin = _get_attr(elem, _STRORIGIN_ATTRS)
                str_val = _get_attr(elem, _STR_ATTRS)
                rows.append((sid, str_origin, str_val, "Empty translation"))
            sheets.append(("EmptyStr", elem_headers, rows))


        if sheets:
            excel_path = output_folder / f"CodeFormatIssue_{lang}.xlsx"
            _write_check_results_excel(excel_path, sheets)
            tab_names = [s[0] for s in sheets]
            logger.info(f"Check results {lang}: Excel report with tabs [{', '.join(tab_names)}] -> {excel_path.name}")

    return summary
