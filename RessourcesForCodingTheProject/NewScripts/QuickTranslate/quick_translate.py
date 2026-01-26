"""
QuickTranslate - Find translations for Korean text by matching StrOrigin.

Matches Korean input text against StrOrigin attribute in Sequencer XML files
and outputs an Excel file with translations in all available languages.

Usage:
    python quick_translate.py

Features:
    - Substring matching: finds all entries where input is contained in StrOrigin
    - Multiple matches: formats as "1. XXX\n2. YYY" in same cell
    - Auto-discovers all language files
    - Branch selection: mainline or cd_lambda
    - Direct StringID lookup
    - Reverse Lookup: input text in any language -> find all translations
"""

import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

# Try lxml first (more robust), fallback to standard library
try:
    from lxml import etree as ET
    USING_LXML = True
except ImportError:
    from xml.etree import ElementTree as ET
    USING_LXML = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

import config

# =============================================================================
# KOREAN DETECTION
# =============================================================================

# Korean Hangul syllables range
KOREAN_REGEX = re.compile(r'[\uac00-\ud7a3]')


def is_korean_text(text: str) -> bool:
    """Check if text contains Korean characters (indicates untranslated)."""
    if not text:
        return False
    return bool(KOREAN_REGEX.search(text))

# =============================================================================
# BRANCH CONFIGURATION
# =============================================================================

BRANCHES = {
    "mainline": {
        "loc": Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"),
        "export": Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__"),
    },
    "cd_lambda": {
        "loc": Path(r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"),
        "export": Path(r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\export__"),
    },
}

# =============================================================================
# XML SANITIZATION (from LanguageDataExporter - battle-tested)
# =============================================================================

_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def _fix_bad_entities(txt: str) -> str:
    """Fix unescaped ampersands."""
    return _bad_entity_re.sub("&amp;", txt)


def _preprocess_newlines(raw: str) -> str:
    """Handle newlines in seg elements."""
    def repl(m: re.Match) -> str:
        inner = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw, flags=re.DOTALL)


def sanitize_xml_content(raw: str) -> str:
    """
    Sanitize XML content to handle common issues.
    """
    raw = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', raw)
    raw = _fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"',
                 lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                 lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', raw)
    return raw


def parse_xml_file(xml_path: Path) -> ET.Element:
    """Parse XML file with sanitization and recovery mode."""
    content = xml_path.read_text(encoding='utf-8')
    content = sanitize_xml_content(content)

    if USING_LXML:
        wrapped = f"<ROOT>\n{content}\n</ROOT>"
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True)
            )
        except ET.XMLSyntaxError:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True)
            )
    else:
        return ET.fromstring(content)


# =============================================================================
# SEQUENCER STRORIGIN INDEX
# =============================================================================

def build_sequencer_strorigin_index(sequencer_folder: Path, progress_callback=None) -> Dict[str, str]:
    """
    Scan Sequencer/*.loc.xml files and build StringID->StrOrigin mapping.
    """
    if not sequencer_folder.exists():
        return {}

    index = {}
    xml_files = list(sequencer_folder.rglob("*.loc.xml"))
    total = len(xml_files)

    for i, xml_file in enumerate(xml_files):
        if progress_callback:
            progress_callback(f"Indexing Sequencer... {i+1}/{total}")
        try:
            root = parse_xml_file(xml_file)
            for elem in root.iter('LocStr'):
                string_id = (elem.get('StringId') or elem.get('StringID') or
                            elem.get('stringid') or elem.get('STRINGID'))
                str_origin = (elem.get('StrOrigin') or elem.get('Strorigin') or
                             elem.get('strorigin') or elem.get('STRORIGIN') or '')
                if string_id and str_origin:
                    index[string_id] = str_origin
        except Exception:
            continue

    return index


def scan_folder_for_strings(folder: Path, progress_callback=None) -> Dict[str, str]:
    """
    Recursively scan folder for XML files and extract StringID -> StrOrigin mapping.

    Scans ALL .xml files (not just .loc.xml) to maximize coverage.
    Returns {StringID: StrOrigin} dict.
    """
    if not folder.exists():
        return {}

    string_map = {}
    xml_files = list(folder.rglob("*.xml"))
    total = len(xml_files)

    for i, xml_file in enumerate(xml_files):
        if progress_callback:
            progress_callback(f"Scanning folder... {i+1}/{total}")
        try:
            root = parse_xml_file(xml_file)
            for elem in root.iter('LocStr'):
                string_id = (elem.get('StringId') or elem.get('StringID') or
                            elem.get('stringid') or elem.get('STRINGID'))
                str_origin = (elem.get('StrOrigin') or elem.get('Strorigin') or
                             elem.get('strorigin') or elem.get('STRORIGIN') or '')
                if string_id and str_origin:
                    string_map[string_id] = str_origin
        except Exception:
            continue

    return string_map


# =============================================================================
# LANGUAGE FILE DISCOVERY AND PARSING
# =============================================================================

def discover_language_files(loc_folder: Path) -> Dict[str, Path]:
    """Find all languagedata_*.xml files."""
    if not loc_folder.exists():
        return {}

    lang_files = {}
    for xml_file in loc_folder.glob("languagedata_*.xml"):
        match = re.match(r'languagedata_(.+)\.xml', xml_file.name, re.IGNORECASE)
        if match:
            lang_code = match.group(1).lower()
            lang_files[lang_code] = xml_file

    return lang_files


def build_translation_lookup(lang_files: Dict[str, Path], progress_callback=None) -> Dict[str, Dict[str, str]]:
    """Parse all language files and build lookup."""
    lookup = {}
    total = len(lang_files)

    for i, (lang_code, xml_path) in enumerate(lang_files.items()):
        if progress_callback:
            progress_callback(f"Loading {lang_code.upper()}... {i+1}/{total}")
        lookup[lang_code] = {}

        try:
            root = parse_xml_file(xml_path)
            for elem in root.iter('LocStr'):
                string_id = (elem.get('StringId') or elem.get('StringID') or
                            elem.get('stringid') or elem.get('STRINGID'))
                str_value = (elem.get('Str') or elem.get('str') or elem.get('STR') or '')
                if string_id:
                    lookup[lang_code][string_id] = str_value
        except Exception:
            continue

    return lookup


def build_reverse_lookup(translation_lookup: Dict[str, Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    """
    Build reverse lookup: {lang_code: {translation_text: StringID}}.

    This allows finding StringID from any translation text.
    """
    reverse = {}
    for lang_code, id_to_text in translation_lookup.items():
        reverse[lang_code] = {}
        for string_id, text in id_to_text.items():
            if text and text.strip():
                # Store text -> StringID (trimmed for matching)
                reverse[lang_code][text.strip()] = string_id
    return reverse


def read_text_file_lines(file_path: Path) -> List[str]:
    """Read text file and return trimmed non-empty lines."""
    lines = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            stripped = line.strip()
            if stripped:
                lines.append(stripped)
    return lines


def find_stringid_from_text(
    text: str,
    reverse_lookup: Dict[str, Dict[str, str]]
) -> Optional[tuple]:
    """
    Find StringID by searching all languages for the text.

    Returns (StringID, detected_lang_code) or None if not found.
    """
    for lang_code, text_to_id in reverse_lookup.items():
        if text in text_to_id:
            return (text_to_id[text], lang_code)
    return None


# =============================================================================
# MATCHING LOGIC
# =============================================================================

def find_matches(korean_input: str, strorigin_index: Dict[str, str]) -> List[str]:
    """Find all StringIDs where korean_input is a substring of StrOrigin."""
    if not korean_input.strip():
        return []

    matches = []
    search_text = korean_input.strip()

    for string_id, str_origin in strorigin_index.items():
        if search_text in str_origin:
            matches.append(string_id)

    return matches


def format_multiple_matches(translations: List[str]) -> str:
    """Format multiple matches as numbered list."""
    translations = [t for t in translations if t and t.strip()]
    if not translations:
        return ""
    if len(translations) == 1:
        return translations[0]
    return "\n".join(f"{i+1}. {t}" for i, t in enumerate(translations))


# =============================================================================
# EXCEL I/O
# =============================================================================

def read_korean_input(excel_path: Path) -> List[str]:
    """Read Korean text from Column 1 of input Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    korean_texts = []
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell_value = row[0].value
        if cell_value:
            korean_texts.append(str(cell_value).strip())

    wb.close()
    return korean_texts


def get_ordered_languages(available_langs: List[str]) -> List[str]:
    """Get languages in preferred order."""
    ordered = []
    for lang in config.LANGUAGE_ORDER:
        if lang in available_langs and lang != "kor":
            ordered.append(lang)
    for lang in available_langs:
        if lang not in ordered and lang != "kor":
            ordered.append(lang)
    return ordered


def write_output_excel(
    output_path: Path,
    korean_inputs: List[str],
    matches_per_input: List[List[str]],
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str]
):
    """Write output Excel file with translations."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    ordered_langs = get_ordered_languages(available_langs)

    # Header row
    headers = ["KOR (Input)"] + [config.LANGUAGE_NAMES.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, (korean_text, string_ids) in enumerate(zip(korean_inputs, matches_per_input), start=2):
        ws.cell(row=row_idx, column=1, value=korean_text)

        for col_idx, lang_code in enumerate(ordered_langs, start=2):
            if not string_ids:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue

            translations = []
            for sid in string_ids:
                trans = translation_lookup.get(lang_code, {}).get(sid, "")
                if trans:
                    translations.append(trans)

            cell_value = format_multiple_matches(translations)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if "\n" in cell_value:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    wb.save(output_path)


def write_stringid_lookup_excel(
    output_path: Path,
    string_id: str,
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str]
):
    """
    Write output Excel for a single StringID lookup.

    Output format: StringID | ENG | FRE | GER | ...
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "StringID Lookup"

    ordered_langs = get_ordered_languages(available_langs)

    # Header row
    headers = ["StringID"] + [config.LANGUAGE_NAMES.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data row
    ws.cell(row=2, column=1, value=string_id)

    for col_idx, lang_code in enumerate(ordered_langs, start=2):
        trans = translation_lookup.get(lang_code, {}).get(string_id, "")
        cell = ws.cell(row=2, column=col_idx, value=trans)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    wb.save(output_path)


def write_folder_translation_excel(
    output_path: Path,
    string_map: Dict[str, str],
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str]
):
    """
    Write Excel with one sheet per language.
    Columns: StrOrigin | English | Translation | StringID

    - "NO TRANSLATION" if translation is empty or contains Korean characters
    """
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ordered_langs = get_ordered_languages(available_langs)

    # Get English lookup for reference
    eng_lookup = translation_lookup.get("eng", {})

    for lang_code in ordered_langs:
        lang_name = config.LANGUAGE_NAMES.get(lang_code, lang_code.upper())
        ws = wb.create_sheet(title=lang_name)
        lang_lookup = translation_lookup.get(lang_code, {})

        # Header row
        headers = ["StrOrigin", "English", lang_name, "StringID"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row_idx = 2
        for string_id, str_origin in string_map.items():
            # StrOrigin
            ws.cell(row=row_idx, column=1, value=str_origin)

            # English
            eng_trans = eng_lookup.get(string_id, "")
            if not eng_trans or is_korean_text(eng_trans):
                eng_trans = "NO TRANSLATION"
            ws.cell(row=row_idx, column=2, value=eng_trans)

            # Target language translation
            lang_trans = lang_lookup.get(string_id, "")
            if not lang_trans or is_korean_text(lang_trans):
                lang_trans = "NO TRANSLATION"
            cell = ws.cell(row=row_idx, column=3, value=lang_trans)
            if "\n" in lang_trans:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            # StringID
            ws.cell(row=row_idx, column=4, value=string_id)

            row_idx += 1

        # Column widths
        ws.column_dimensions['A'].width = 50  # StrOrigin
        ws.column_dimensions['B'].width = 40  # English
        ws.column_dimensions['C'].width = 40  # Translation
        ws.column_dimensions['D'].width = 25  # StringID

    wb.save(output_path)


def write_reverse_lookup_excel(
    output_path: Path,
    input_texts: List[str],
    stringid_map: Dict[str, str],  # input_text -> StringID
    translation_lookup: Dict[str, Dict[str, str]],
    available_langs: List[str]
):
    """
    Write Excel with all languages in columns.
    Columns: Input | KOR | ENG | FRE | GER | ...
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reverse Lookup"

    # Order: KOR first, then others
    ordered_langs = ["kor"] + get_ordered_languages(available_langs)

    # Header row
    headers = ["Input"] + [config.LANGUAGE_NAMES.get(lang, lang.upper()) for lang in ordered_langs]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, input_text in enumerate(input_texts, start=2):
        # Write input text
        ws.cell(row=row_idx, column=1, value=input_text)

        string_id = stringid_map.get(input_text)
        if not string_id:
            # NOT FOUND - write "NOT FOUND" in KOR column, leave rest empty
            ws.cell(row=row_idx, column=2, value="NOT FOUND")
            continue

        # For each language, get translation from translation_lookup[lang][string_id]
        for col_idx, lang_code in enumerate(ordered_langs, start=2):
            trans = translation_lookup.get(lang_code, {}).get(string_id, "")
            if not trans or is_korean_text(trans):
                trans = "NO TRANSLATION"
            cell = ws.cell(row=row_idx, column=col_idx, value=trans)
            if "\n" in trans:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 50  # Input
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    wb.save(output_path)


# =============================================================================
# GUI APPLICATION
# =============================================================================

class QuickTranslateApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("QuickTranslate")
        self.root.geometry("500x560")
        self.root.resizable(False, False)
        self.root.configure(bg='#f0f0f0')

        # Variables
        self.input_path = tk.StringVar()
        self.string_id_input = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.reverse_file_path = tk.StringVar()
        self.selected_branch = tk.StringVar(value="mainline")
        self.status_text = tk.StringVar(value="Ready")

        # Cached data
        self.strorigin_index: Optional[Dict[str, str]] = None
        self.translation_lookup: Optional[Dict[str, Dict[str, str]]] = None
        self.available_langs: Optional[List[str]] = None
        self.cached_branch: Optional[str] = None

        self._create_ui()

    def _create_ui(self):
        # Main container
        main = tk.Frame(self.root, bg='#f0f0f0', padx=20, pady=15)
        main.pack(fill=tk.BOTH, expand=True)

        # Title
        title = tk.Label(main, text="QuickTranslate", font=('Segoe UI', 16, 'bold'),
                        bg='#f0f0f0', fg='#333')
        title.pack(pady=(0, 12))

        # Branch selection section (moved to top)
        branch_frame = tk.LabelFrame(main, text="Branch", font=('Segoe UI', 9),
                                     bg='#f0f0f0', fg='#555', padx=10, pady=6)
        branch_frame.pack(fill=tk.X, pady=(0, 10))

        branch_inner = tk.Frame(branch_frame, bg='#f0f0f0')
        branch_inner.pack()

        for branch_name in BRANCHES.keys():
            btn = tk.Radiobutton(
                branch_inner,
                text=branch_name,
                variable=self.selected_branch,
                value=branch_name,
                font=('Segoe UI', 10),
                bg='#f0f0f0',
                activebackground='#f0f0f0',
                cursor='hand2',
                padx=15
            )
            btn.pack(side=tk.LEFT, padx=5)

        # ===== Korean Match Section =====
        korean_frame = tk.LabelFrame(main, text="Korean Text Match (from Excel)", font=('Segoe UI', 9),
                                     bg='#f0f0f0', fg='#555', padx=10, pady=8)
        korean_frame.pack(fill=tk.X, pady=(0, 10))

        input_inner = tk.Frame(korean_frame, bg='#f0f0f0')
        input_inner.pack(fill=tk.X)

        self.input_entry = tk.Entry(input_inner, textvariable=self.input_path,
                                    font=('Segoe UI', 9), relief='solid', bd=1)
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        browse_btn = tk.Button(input_inner, text="Browse", command=self._browse_input,
                              font=('Segoe UI', 9), bg='#e0e0e0', relief='solid', bd=1,
                              padx=10, cursor='hand2')
        browse_btn.pack(side=tk.LEFT, padx=(6, 0))

        self.translate_btn = tk.Button(input_inner, text="Translate", command=self._translate,
                                       font=('Segoe UI', 9, 'bold'), bg='#4a90d9', fg='white',
                                       relief='flat', padx=10, cursor='hand2')
        self.translate_btn.pack(side=tk.LEFT, padx=(6, 0))

        # ===== StringID Lookup Section =====
        stringid_frame = tk.LabelFrame(main, text="StringID Lookup", font=('Segoe UI', 9),
                                       bg='#f0f0f0', fg='#555', padx=10, pady=8)
        stringid_frame.pack(fill=tk.X, pady=(0, 10))

        stringid_inner = tk.Frame(stringid_frame, bg='#f0f0f0')
        stringid_inner.pack(fill=tk.X)

        self.stringid_entry = tk.Entry(stringid_inner, textvariable=self.string_id_input,
                                       font=('Segoe UI', 9), relief='solid', bd=1)
        self.stringid_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        self.lookup_btn = tk.Button(stringid_inner, text="Lookup", command=self._lookup_stringid,
                                    font=('Segoe UI', 9, 'bold'), bg='#5cb85c', fg='white',
                                    relief='flat', padx=14, cursor='hand2')
        self.lookup_btn.pack(side=tk.LEFT, padx=(6, 0))

        # ===== Translate Folder Section =====
        folder_frame = tk.LabelFrame(main, text="Translate Folder (XML -> Multi-Language Excel)",
                                     font=('Segoe UI', 9), bg='#f0f0f0', fg='#555', padx=10, pady=8)
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        folder_inner = tk.Frame(folder_frame, bg='#f0f0f0')
        folder_inner.pack(fill=tk.X)

        self.folder_entry = tk.Entry(folder_inner, textvariable=self.folder_path,
                                     font=('Segoe UI', 9), relief='solid', bd=1)
        self.folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        folder_browse_btn = tk.Button(folder_inner, text="Browse", command=self._browse_folder,
                                      font=('Segoe UI', 9), bg='#e0e0e0', relief='solid', bd=1,
                                      padx=10, cursor='hand2')
        folder_browse_btn.pack(side=tk.LEFT, padx=(6, 0))

        self.export_btn = tk.Button(folder_inner, text="Export", command=self._translate_folder,
                                    font=('Segoe UI', 9, 'bold'), bg='#5cb85c', fg='white',
                                    relief='flat', padx=14, cursor='hand2')
        self.export_btn.pack(side=tk.LEFT, padx=(6, 0))

        # ===== Reverse Lookup Section =====
        reverse_frame = tk.LabelFrame(main, text="Reverse Lookup (Any Language -> All Languages)",
                                      font=('Segoe UI', 9), bg='#f0f0f0', fg='#555', padx=10, pady=8)
        reverse_frame.pack(fill=tk.X, pady=(0, 10))

        reverse_inner = tk.Frame(reverse_frame, bg='#f0f0f0')
        reverse_inner.pack(fill=tk.X)

        self.reverse_entry = tk.Entry(reverse_inner, textvariable=self.reverse_file_path,
                                      font=('Segoe UI', 9), relief='solid', bd=1)
        self.reverse_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        reverse_browse_btn = tk.Button(reverse_inner, text="Browse", command=self._browse_reverse_file,
                                       font=('Segoe UI', 9), bg='#e0e0e0', relief='solid', bd=1,
                                       padx=10, cursor='hand2')
        reverse_browse_btn.pack(side=tk.LEFT, padx=(6, 0))

        self.reverse_btn = tk.Button(reverse_inner, text="Find All", command=self._reverse_lookup,
                                     font=('Segoe UI', 9, 'bold'), bg='#d9534f', fg='white',
                                     relief='flat', padx=10, cursor='hand2')
        self.reverse_btn.pack(side=tk.LEFT, padx=(6, 0))

        # Status bar
        status_frame = tk.Frame(main, bg='#e8e8e8', relief='solid', bd=1)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(10, 0))

        status_label = tk.Label(status_frame, textvariable=self.status_text,
                               font=('Segoe UI', 9), bg='#e8e8e8', fg='#666',
                               anchor='w', padx=8, pady=4)
        status_label.pack(fill=tk.X)

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.input_path.set(path)

    def _browse_folder(self):
        """Open folder selection dialog."""
        path = filedialog.askdirectory(title="Select Folder Containing XML Files")
        if path:
            self.folder_path.set(path)

    def _browse_reverse_file(self):
        """Open file selection dialog for text file."""
        path = filedialog.askopenfilename(
            title="Select Text File with List",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if path:
            self.reverse_file_path.set(path)

    def _update_status(self, text: str):
        self.status_text.set(text)
        self.root.update()

    def _load_data_if_needed(self, need_sequencer: bool = True) -> bool:
        """Load data if not cached or branch changed."""
        branch = self.selected_branch.get()

        if self.cached_branch == branch and self.translation_lookup is not None:
            if not need_sequencer or self.strorigin_index is not None:
                return True  # Already loaded

        paths = BRANCHES.get(branch)
        if not paths:
            messagebox.showerror("Error", f"Unknown branch: {branch}")
            return False

        loc_folder = paths["loc"]
        export_folder = paths["export"]
        sequencer_folder = export_folder / "Sequencer"

        # Validate paths
        if not loc_folder.exists():
            messagebox.showerror("Error", f"LOC folder not found:\n{loc_folder}")
            return False

        if need_sequencer:
            if not sequencer_folder.exists():
                messagebox.showerror("Error", f"Sequencer folder not found:\n{sequencer_folder}")
                return False

            # Build Sequencer index
            self.strorigin_index = build_sequencer_strorigin_index(
                sequencer_folder, self._update_status
            )

            if not self.strorigin_index:
                messagebox.showerror("Error", "No Sequencer data found.")
                return False

        # Load language files
        lang_files = discover_language_files(loc_folder)
        if not lang_files:
            messagebox.showerror("Error", "No language files found.")
            return False

        self.translation_lookup = build_translation_lookup(lang_files, self._update_status)
        self.available_langs = list(lang_files.keys())
        self.cached_branch = branch

        return True

    def _translate(self):
        """Run Korean text translation and generate output."""
        # Validate input
        if not self.input_path.get():
            messagebox.showwarning("Warning", "Please select an input Excel file.")
            return

        input_file = Path(self.input_path.get())
        if not input_file.exists():
            messagebox.showerror("Error", f"Input file not found:\n{input_file}")
            return

        # Disable all buttons during processing
        self.translate_btn.config(state='disabled')
        self.lookup_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.reverse_btn.config(state='disabled')

        try:
            # Load data if needed
            if not self._load_data_if_needed(need_sequencer=True):
                return

            # Read input
            self._update_status("Reading input file...")
            korean_inputs = read_korean_input(input_file)

            if not korean_inputs:
                messagebox.showwarning("Warning", "No text found in input file.")
                return

            # Find matches
            self._update_status("Finding matches...")
            matches_per_input = []
            total_matches = 0

            for korean_text in korean_inputs:
                matches = find_matches(korean_text, self.strorigin_index)
                matches_per_input.append(matches)
                total_matches += len(matches)

            # Write output
            self._update_status("Writing output...")
            config.ensure_output_folder()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = config.OUTPUT_FOLDER / f"QuickTranslate_{timestamp}.xlsx"

            write_output_excel(
                output_path,
                korean_inputs,
                matches_per_input,
                self.translation_lookup,
                self.available_langs
            )

            self._update_status(f"Done! {len(korean_inputs)} inputs, {total_matches} matches")
            messagebox.showinfo("Success", f"Output saved to:\n{output_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Translation failed:\n{e}")
            self._update_status(f"Error: {e}")

        finally:
            self.translate_btn.config(state='normal')
            self.lookup_btn.config(state='normal')
            self.export_btn.config(state='normal')
            self.reverse_btn.config(state='normal')

    def _lookup_stringid(self):
        """Look up a single StringID and generate output."""
        string_id = self.string_id_input.get().strip()

        if not string_id:
            messagebox.showwarning("Warning", "Please enter a StringID.")
            return

        # Disable all buttons during processing
        self.translate_btn.config(state='disabled')
        self.lookup_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.reverse_btn.config(state='disabled')

        try:
            # Load data if needed (don't need Sequencer for direct StringID lookup)
            if not self._load_data_if_needed(need_sequencer=False):
                return

            # Check if StringID exists in any language
            found = False
            for lang_code, lookup in self.translation_lookup.items():
                if string_id in lookup:
                    found = True
                    break

            if not found:
                messagebox.showwarning("Warning", f"StringID not found: {string_id}")
                self._update_status(f"StringID not found: {string_id}")
                return

            # Write output
            self._update_status("Writing output...")
            config.ensure_output_folder()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = config.OUTPUT_FOLDER / f"StringID_{string_id}_{timestamp}.xlsx"

            write_stringid_lookup_excel(
                output_path,
                string_id,
                self.translation_lookup,
                self.available_langs
            )

            self._update_status(f"Done! Lookup for {string_id}")
            messagebox.showinfo("Success", f"Output saved to:\n{output_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Lookup failed:\n{e}")
            self._update_status(f"Error: {e}")

        finally:
            self.translate_btn.config(state='normal')
            self.lookup_btn.config(state='normal')
            self.export_btn.config(state='normal')
            self.reverse_btn.config(state='normal')

    def _translate_folder(self):
        """Process folder and generate multi-language Excel."""
        # Validate folder path
        if not self.folder_path.get():
            messagebox.showwarning("Warning", "Please select a folder containing XML files.")
            return

        folder = Path(self.folder_path.get())
        if not folder.exists():
            messagebox.showerror("Error", f"Folder not found:\n{folder}")
            return

        # Disable all buttons during processing
        self.translate_btn.config(state='disabled')
        self.lookup_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.reverse_btn.config(state='disabled')

        try:
            # Scan folder for StringID -> StrOrigin mapping
            self._update_status("Scanning folder for XML files...")
            string_map = scan_folder_for_strings(folder, self._update_status)

            if not string_map:
                messagebox.showwarning("Warning", "No StringID/StrOrigin pairs found in folder.")
                self._update_status("No data found")
                return

            # Load language files (don't need Sequencer for this)
            if not self._load_data_if_needed(need_sequencer=False):
                return

            # Write output Excel
            self._update_status("Writing multi-language Excel...")
            config.ensure_output_folder()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            folder_name = folder.name
            output_path = config.OUTPUT_FOLDER / f"FolderTranslate_{folder_name}_{timestamp}.xlsx"

            write_folder_translation_excel(
                output_path,
                string_map,
                self.translation_lookup,
                self.available_langs
            )

            lang_count = len(get_ordered_languages(self.available_langs))
            self._update_status(f"Done! {len(string_map)} strings, {lang_count} languages")
            messagebox.showinfo("Success", f"Output saved to:\n{output_path}\n\n"
                               f"Strings: {len(string_map)}\nLanguage sheets: {lang_count}")

        except Exception as e:
            messagebox.showerror("Error", f"Folder translation failed:\n{e}")
            self._update_status(f"Error: {e}")

        finally:
            self.translate_btn.config(state='normal')
            self.lookup_btn.config(state='normal')
            self.export_btn.config(state='normal')
            self.reverse_btn.config(state='normal')

    def _reverse_lookup(self):
        """Perform reverse lookup from any language to all languages."""
        # Validate file path
        if not self.reverse_file_path.get():
            messagebox.showwarning("Warning", "Please select a text file with list of strings.")
            return

        file_path = Path(self.reverse_file_path.get())
        if not file_path.exists():
            messagebox.showerror("Error", f"File not found:\n{file_path}")
            return

        # Disable all buttons during processing
        self.translate_btn.config(state='disabled')
        self.lookup_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.reverse_btn.config(state='disabled')

        try:
            # Load data if needed (don't need Sequencer for reverse lookup)
            if not self._load_data_if_needed(need_sequencer=False):
                return

            # Read input file
            self._update_status("Reading input file...")
            input_texts = read_text_file_lines(file_path)

            if not input_texts:
                messagebox.showwarning("Warning", "No text found in input file.")
                self._update_status("No input text")
                return

            # Build reverse lookup
            self._update_status("Building reverse lookup...")
            reverse_lookup = build_reverse_lookup(self.translation_lookup)

            # Find StringID for each input text
            self._update_status("Finding StringIDs...")
            stringid_map = {}
            not_found = []
            detected_langs = set()

            for text in input_texts:
                result = find_stringid_from_text(text, reverse_lookup)
                if result:
                    string_id, lang = result
                    stringid_map[text] = string_id
                    detected_langs.add(lang)
                else:
                    not_found.append(text)

            # Write output Excel
            self._update_status("Writing output...")
            config.ensure_output_folder()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = config.OUTPUT_FOLDER / f"ReverseLookup_{timestamp}.xlsx"

            write_reverse_lookup_excel(
                output_path,
                input_texts,
                stringid_map,
                self.translation_lookup,
                self.available_langs
            )

            # Build summary
            found_count = len(stringid_map)
            total_count = len(input_texts)
            detected_str = ", ".join(sorted(detected_langs)) if detected_langs else "N/A"

            self._update_status(f"Done! {found_count}/{total_count} found")
            messagebox.showinfo("Success",
                f"Output saved to:\n{output_path}\n\n"
                f"Found: {found_count}/{total_count}\n"
                f"Detected languages: {detected_str}")

        except Exception as e:
            messagebox.showerror("Error", f"Reverse lookup failed:\n{e}")
            self._update_status(f"Error: {e}")

        finally:
            self.translate_btn.config(state='normal')
            self.lookup_btn.config(state='normal')
            self.export_btn.config(state='normal')
            self.reverse_btn.config(state='normal')


# =============================================================================
# MAIN
# =============================================================================

def main():
    root = tk.Tk()
    app = QuickTranslateApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
