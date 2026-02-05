#!/usr/bin/env python3
"""
EventName to StringID Converter
===============================
Standalone tool to convert EventNames to StringIDs using export XML files.

Features:
- File dialog to select input Excel (EventName in first column)
- Automatically scans mainline export folder for StringID mappings
- Creates output Excel with EventName | StringID
- Highlights unfound items in light orange

Usage:
1. Run the script
2. Select your Excel file with EventNames in first column
3. Output file created automatically next to input file
"""

import os
import sys
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from tkinter import Tk, filedialog, messagebox

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).parent

# Default export path (F: drive)
DEFAULT_EXPORT_PATH = r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__"


def load_settings() -> dict:
    """Load settings from settings.json if it exists."""
    settings_file = SCRIPT_DIR / "settings.json"

    if not settings_file.exists():
        return {}

    try:
        with open(settings_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Could not load settings.json: {e}")
        return {}


def get_export_folder() -> Path:
    """Get the export folder path, applying drive letter from settings if configured."""
    settings = load_settings()
    drive_letter = settings.get('drive_letter', 'F')

    path_str = DEFAULT_EXPORT_PATH
    if path_str.startswith("F:") or path_str.startswith("f:"):
        path_str = f"{drive_letter.upper()}:{path_str[2:]}"

    return Path(path_str)


# =============================================================================
# STYLING
# =============================================================================

# Light orange for unfound items
UNFOUND_FILL = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
UNFOUND_FONT = Font(color="CC6600")

# Header styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# Found items styling
FOUND_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FOUND_FONT = Font(color="2E7D32")

# Border
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# XML PARSING - Build EventName to StringID mapping
# =============================================================================

def build_eventname_to_stringid_mapping(export_folder: Path) -> dict:
    """
    Scan all XML files in export folder and build EventName -> StringID mapping.

    Export XML structure:
    <ExportTable>
        <Item>
            <StringID>12345</StringID>
            <SoundEventName>Event_Name_Here</SoundEventName>
            ...
        </Item>
    </ExportTable>

    Returns:
        Dict mapping EventName -> StringID
    """
    mapping = {}
    xml_files = []

    print(f"\nScanning export folder: {export_folder}")

    if not export_folder.exists():
        print(f"ERROR: Export folder does not exist: {export_folder}")
        return mapping

    # Find all XML files recursively
    for xml_path in export_folder.rglob("*.xml"):
        xml_files.append(xml_path)

    print(f"Found {len(xml_files)} XML files to scan...")

    # Parse each XML file
    processed = 0
    errors = 0

    for xml_path in xml_files:
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()

            # Look for Item elements with StringID and SoundEventName
            for item in root.iter('Item'):
                stringid_elem = item.find('StringID')
                eventname_elem = item.find('SoundEventName')

                if stringid_elem is not None and eventname_elem is not None:
                    stringid = stringid_elem.text.strip() if stringid_elem.text else ""
                    eventname = eventname_elem.text.strip() if eventname_elem.text else ""

                    if eventname and stringid:
                        # Store mapping (first occurrence wins)
                        if eventname not in mapping:
                            mapping[eventname] = stringid

            processed += 1

            # Progress indicator
            if processed % 100 == 0:
                print(f"  Processed {processed}/{len(xml_files)} files...")

        except ET.ParseError as e:
            errors += 1
            # Skip malformed XML files silently
        except Exception as e:
            errors += 1
            print(f"  Warning: Error parsing {xml_path.name}: {e}")

    print(f"Scan complete: {len(mapping)} EventName mappings found")
    if errors > 0:
        print(f"  ({errors} files skipped due to parse errors)")

    return mapping


# =============================================================================
# MAIN CONVERSION
# =============================================================================

def select_input_file() -> Path:
    """Open file dialog to select input Excel file."""
    root = Tk()
    root.withdraw()  # Hide the main window
    root.attributes('-topmost', True)  # Bring dialog to front

    file_path = filedialog.askopenfilename(
        title="Select Excel file with EventNames",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )

    root.destroy()

    if not file_path:
        return None

    return Path(file_path)


def convert_eventnames_to_stringids(input_path: Path, mapping: dict) -> Path:
    """
    Convert EventNames in input Excel to StringIDs.

    Args:
        input_path: Path to input Excel (EventNames in column A)
        mapping: Dict of EventName -> StringID

    Returns:
        Path to output Excel file
    """
    print(f"\nProcessing: {input_path.name}")

    # Load input workbook
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Converted"

    # Write headers
    out_ws.cell(1, 1, "EventName")
    out_ws.cell(1, 2, "StringID")
    out_ws.cell(1, 3, "Status")

    for col in range(1, 4):
        cell = out_ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Process each row
    found_count = 0
    unfound_count = 0

    for row in range(1, ws.max_row + 1):
        eventname_cell = ws.cell(row, 1)
        eventname = str(eventname_cell.value or "").strip()

        if not eventname:
            continue

        # Skip header row if present
        if row == 1 and eventname.upper() in ("EVENTNAME", "EVENT_NAME", "EVENT NAME", "NAME"):
            continue

        out_row = out_ws.max_row + 1

        # Write EventName
        out_ws.cell(out_row, 1, eventname)
        out_ws.cell(out_row, 1).border = THIN_BORDER

        # Look up StringID
        stringid = mapping.get(eventname, "")

        if stringid:
            # Found
            out_ws.cell(out_row, 2, stringid)
            out_ws.cell(out_row, 3, "FOUND")

            out_ws.cell(out_row, 1).fill = FOUND_FILL
            out_ws.cell(out_row, 2).fill = FOUND_FILL
            out_ws.cell(out_row, 2).font = FOUND_FONT
            out_ws.cell(out_row, 3).fill = FOUND_FILL
            out_ws.cell(out_row, 3).font = FOUND_FONT

            found_count += 1
        else:
            # Not found - highlight in light orange
            out_ws.cell(out_row, 2, "NOT FOUND")
            out_ws.cell(out_row, 3, "MISSING")

            out_ws.cell(out_row, 1).fill = UNFOUND_FILL
            out_ws.cell(out_row, 1).font = UNFOUND_FONT
            out_ws.cell(out_row, 2).fill = UNFOUND_FILL
            out_ws.cell(out_row, 2).font = UNFOUND_FONT
            out_ws.cell(out_row, 3).fill = UNFOUND_FILL
            out_ws.cell(out_row, 3).font = UNFOUND_FONT

            unfound_count += 1

        out_ws.cell(out_row, 2).border = THIN_BORDER
        out_ws.cell(out_row, 3).border = THIN_BORDER

    # Set column widths
    out_ws.column_dimensions['A'].width = 50
    out_ws.column_dimensions['B'].width = 20
    out_ws.column_dimensions['C'].width = 12

    # Generate output path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = input_path.parent / f"{input_path.stem}_StringIDs_{timestamp}.xlsx"

    # Save
    out_wb.save(output_path)

    print(f"\nResults:")
    print(f"  Found:     {found_count}")
    print(f"  Not Found: {unfound_count} (highlighted in orange)")
    print(f"\nOutput saved: {output_path}")

    return output_path


def main():
    """Main entry point."""
    print("=" * 60)
    print("EventName to StringID Converter")
    print("=" * 60)

    # Get export folder
    export_folder = get_export_folder()
    print(f"\nExport folder: {export_folder}")

    if not export_folder.exists():
        print(f"\nERROR: Export folder does not exist!")
        print(f"Expected: {export_folder}")
        print(f"\nCreate a settings.json file with your drive letter:")
        print('{"drive_letter": "D"}')
        messagebox.showerror("Error", f"Export folder not found:\n{export_folder}")
        return

    # Build mapping from export XMLs
    print("\nBuilding EventName to StringID mapping...")
    mapping = build_eventname_to_stringid_mapping(export_folder)

    if not mapping:
        print("\nERROR: No EventName mappings found in export folder!")
        messagebox.showerror("Error", "No EventName mappings found in export folder!")
        return

    # Select input file
    print("\nPlease select your Excel file with EventNames...")
    input_path = select_input_file()

    if not input_path:
        print("No file selected. Exiting.")
        return

    # Convert
    try:
        output_path = convert_eventnames_to_stringids(input_path, mapping)

        # Show success message
        messagebox.showinfo(
            "Conversion Complete",
            f"Output saved to:\n{output_path.name}\n\nLocation: {output_path.parent}"
        )

        # Open output folder
        os.startfile(output_path.parent)

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Conversion failed:\n{e}")


if __name__ == "__main__":
    main()
