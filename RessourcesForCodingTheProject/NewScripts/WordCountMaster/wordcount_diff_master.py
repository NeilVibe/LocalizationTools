#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script Name: wordcount_diff_master.py
Created: 2025-11-18
Updated: 2025-11-18 v2.0 (SIMPLIFIED - Smart weekly/monthly categorization)
Purpose: Track translation word count changes by comparing TODAY's data against any past date
Input: XML translation files from loc and export folders
Output:
    - wordcount_history.json (historical data)
    - WordCountAnalysis_YYYYMMDD_HHMMSS.xlsx (diff report with 4 sheets)
Reference: wordcount1.py (from SECONDARY PYTHON SCRIPTS)

Key Design Decisions (v2.0):
    1. SIMPLIFIED LOGIC - No daily diffs, only Weekly and Monthly
       - Always compares TODAY's current data vs selected PAST date
       - Smart categorization: closer to 7 days → Weekly, closer to 30 days → Monthly

    2. SMART CATEGORIZATION
       Example: Compare today (2025-11-18) vs past date (2025-11-10):
       - Days difference: 8 days
       - |8 - 7| = 1, |8 - 30| = 22
       - 8 is closer to 7 than 30 → Shows in "Weekly" sheets
       - Title: "Period: 2025-11-18 to 2025-11-10 (8 days)"

    3. WORD METRICS ONLY - Removed node metrics (Total Nodes, Completed Nodes)
       Keep: Total Words, Completed Words, Word Coverage %

    4. DETAILED SHEET STRUCTURE - Matches wordcount1.py format
       For each language:
       - "Language: ENG" title row (bold)
       - Header row: Category | Total Words | Completed Words | Coverage % | diffs...
       - Category data rows (Faction, Main, Sequencer + Other, System, World, etc.)
       - Yellow separator row
       Then next language block

    5. CATEGORY STRUCTURE - Matches wordcount1.py
       - Sequencer/Faction → "Faction"
       - Sequencer/Main → "Main"
       - Sequencer/Sequencer + Sequencer/Other → "Sequencer + Other"
       - Other top-level folders → their own categories

    6. NO PLATFORM/NONE GROUPS - Simplified from original
       Removed platform grouping and "None" group handling

Usage:
    python wordcount_diff_master.py

    1. Run the script (processes TODAY's data automatically)
    2. Enter a PAST date to compare against (YYYY-MM-DD)
    3. Script compares TODAY vs PAST DATE
    4. Script categorizes as Weekly or Monthly based on days difference
    5. Script updates JSON history with today's data
    6. Script generates Excel report with 4 sheets

Dependencies:
    pip install lxml openpyxl

Excel Report Structure (4 sheets):
    1. Weekly Diff - Full Summary (language-level comparison if days ≈ 7)
    2. Monthly Diff - Full Summary (language-level comparison if days ≈ 30)
    3. Weekly Diff - Detailed (category-level tables per language if days ≈ 7)
    4. Monthly Diff - Detailed (category-level tables per language if days ≈ 30)

    Note: Only 2 sheets will have data (either Weekly or Monthly pair),
          the other pair will show "N/A - Select appropriate comparison period"
"""

import os
import re
import json
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Set, Optional

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
LANGUAGE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
EXPORT_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
HISTORY_JSON = Path.cwd() / "wordcount_history.json"
OUTPUT_EXCEL_PREFIX = "WordCountAnalysis"

# ─────────────────────────────────────────────────────────────
# XML PARSE HELPERS (from wordcount1.py)
# ─────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(xml_text: str) -> str:
    """Fix malformed XML entities"""
    return _bad_entity_re.sub("&amp;", xml_text)

def parse_xml_file(path: Path) -> ET._Element:
    """Parse XML file with error recovery"""
    raw = path.read_text(encoding="utf-8", errors="ignore")
    wrapped = f"<ROOT>\n{fix_bad_entities(raw)}\n</ROOT>"
    parser = ET.XMLParser(recover=True, huge_tree=True)
    return ET.fromstring(wrapped.encode("utf-8"), parser=parser)

# ─────────────────────────────────────────────────────────────
# FILE ITERATION (from wordcount1.py)
# ─────────────────────────────────────────────────────────────
def iter_language_files(folder: Path):
    """Iterate through all language XML files"""
    for dirpath, _, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().startswith("languagedata_") and fn.lower().endswith(".xml"):
                yield Path(dirpath) / fn

# ─────────────────────────────────────────────────────────────
# WORD-COUNT & TRANSLATION DETECTION (from wordcount1.py)
# ─────────────────────────────────────────────────────────────
korean_re = re.compile(r'[\uac00-\ud7a3]')

def count_words(text: str) -> int:
    """Count words in text"""
    return len([w for w in re.split(r'\s+', text.strip()) if w])

def is_korean(text: str) -> bool:
    """Check if text contains Korean characters"""
    return bool(korean_re.search(text))

def analyse_file(path: Path) -> Tuple[int, int]:
    """
    Analyse a single language file

    Returns:
        (total_words, completed_words)
    """
    root = parse_xml_file(path)
    total_words = completed_words = 0

    for loc in root.iter("LocStr"):
        origin = (loc.get("StrOrigin") or "").strip()
        if not origin:
            continue
        origin_wc = count_words(origin)
        total_words += origin_wc

        trans = (loc.get("Str") or "").strip()
        if trans and not is_korean(trans):
            completed_words += origin_wc

    return total_words, completed_words

def collect_completed_ids_for_paths(xml_paths: List[Path]) -> Set[str]:
    """
    Collect all completed string IDs from a list of XML files

    Returns:
        Set of completed StringIDs
    """
    completed: Set[str] = set()
    for xml_path in xml_paths:
        root = parse_xml_file(xml_path)
        for loc in root.iter("LocStr"):
            origin = (loc.get("StrOrigin") or "").strip()
            if not origin:
                continue
            trans = (loc.get("Str") or "").strip()
            if trans and not is_korean(trans):
                sid = loc.get("StringId")
                if sid:
                    completed.add(sid)
    return completed

def analyse_export_file(path: Path, completed_ids: Set[str]) -> Tuple[int, int]:
    """
    Analyse export file using completed IDs

    Returns:
        (total_words, completed_words)
    """
    root = parse_xml_file(path)
    tw = cw = 0

    for loc in root.iter("LocStr"):
        origin = (loc.get("StrOrigin") or "").strip()
        if not origin:
            continue
        origin_wc = count_words(origin)
        tw += origin_wc

        sid = loc.get("StringId")
        if sid in completed_ids:
            cw += origin_wc

    return tw, cw

# ─────────────────────────────────────────────────────────────
# DATA COLLECTION (NEW - Modified from wordcount1.py)
# ─────────────────────────────────────────────────────────────
def collect_language_data(lang_code: str, xml_paths: List[Path]) -> dict:
    """
    Collect word count data for a single language

    Returns:
        {
            "full_summary": {
                "total_words": int,
                "completed_words": int,
                "word_coverage_pct": float
            },
            "detailed_categories": {
                "category_name": {
                    "total_words": int,
                    "completed_words": int,
                    "word_coverage_pct": float
                },
                ...
            }
        }
    """
    # Collect full summary from language files
    total_words = completed_words = 0
    for xml_path in xml_paths:
        tw, cw = analyse_file(xml_path)
        total_words += tw
        completed_words += cw

    word_coverage_pct = (completed_words / total_words * 100) if total_words else 0.0

    full_summary = {
        "total_words": total_words,
        "completed_words": completed_words,
        "word_coverage_pct": word_coverage_pct
    }

    # Collect detailed categories from export folder
    print(f"      Collecting categories for {lang_code}...")
    completed_ids = collect_completed_ids_for_paths(xml_paths)

    categories: Dict[str, List[Path]] = {}
    for child in EXPORT_FOLDER.iterdir():
        if not child.is_dir():
            continue
        if child.name == "Sequencer":
            # Special handling for Sequencer folder
            for name in ("Faction", "Main"):
                sub = child / name
                if sub.is_dir():
                    categories[name] = [sub]
            group_paths = []
            for name in ("Sequencer", "Other"):
                sub = child / name
                if sub.is_dir():
                    group_paths.append(sub)
            if group_paths:
                categories["Sequencer + Other"] = group_paths
        else:
            categories[child.name] = [child]

    detailed_categories = {}
    for cat, paths in categories.items():
        cat_total_words = cat_completed_words = 0
        for base in paths:
            for xml_path in base.rglob("*.xml"):
                tw, cw = analyse_export_file(xml_path, completed_ids)
                cat_total_words += tw
                cat_completed_words += cw

        cat_coverage_pct = (cat_completed_words / cat_total_words * 100) if cat_total_words else 0.0

        detailed_categories[cat] = {
            "total_words": cat_total_words,
            "completed_words": cat_completed_words,
            "word_coverage_pct": cat_coverage_pct
        }

    return {
        "full_summary": full_summary,
        "detailed_categories": detailed_categories
    }

def collect_all_languages_data() -> Dict[str, dict]:
    """
    Scan all language files and collect data

    Returns:
        {
            "ENG": {full_summary: {...}, detailed_categories: {...}},
            "FRA": {...},
            ...
        }
    """
    print(f"  Scanning language folder: {LANGUAGE_FOLDER}")

    xmls_by_lang: Dict[str, List[Path]] = {}

    for xml_path in iter_language_files(LANGUAGE_FOLDER):
        stem = xml_path.stem
        parts = stem.split("_", 1)
        if len(parts) != 2:
            continue
        lang_code = parts[1].upper()

        # Filter out KOR
        if lang_code == "KOR":
            continue

        xmls_by_lang.setdefault(lang_code, []).append(xml_path)

    print(f"  Found {len(xmls_by_lang)} languages (excluding KOR)")

    # Collect data for each language
    all_languages_data = {}
    for lang_code in sorted(xmls_by_lang.keys()):
        print(f"    Processing {lang_code}...")
        lang_data = collect_language_data(lang_code, xmls_by_lang[lang_code])
        all_languages_data[lang_code] = lang_data

        # Print summary
        fs = lang_data["full_summary"]
        print(f"      Full: {fs['completed_words']:,}/{fs['total_words']:,} words ({fs['word_coverage_pct']:.2f}%)")

    return all_languages_data

# ─────────────────────────────────────────────────────────────
# USER INPUT
# ─────────────────────────────────────────────────────────────
def get_comparison_date_from_user() -> str:
    """
    Prompt user for a PAST date to compare against today

    Returns:
        Date string in format: YYYY-MM-DD
    """
    today = datetime.now().strftime("%Y-%m-%d")

    while True:
        print("\n" + "="*60)
        print(f"TODAY'S DATE: {today}")
        print("="*60)
        print("Enter a PAST date to compare against:")
        print("Format: YYYY-MM-DD (e.g., 2025-11-10)")
        print("\nExamples:")
        print("  - Enter date ~7 days ago  → Shows in Weekly sheets")
        print("  - Enter date ~30 days ago → Shows in Monthly sheets")
        print("="*60)
        date_str = input("Past date: ").strip()

        # Validate date format
        try:
            past_date = datetime.strptime(date_str, "%Y-%m-%d")
            today_date = datetime.strptime(today, "%Y-%m-%d")

            if past_date >= today_date:
                print("Error: Past date must be BEFORE today!")
                continue

            return date_str
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD")

# ─────────────────────────────────────────────────────────────
# JSON HISTORY MANAGEMENT
# ─────────────────────────────────────────────────────────────
def load_history() -> dict:
    """Load existing history or create new"""
    if HISTORY_JSON.exists():
        with open(HISTORY_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "runs": [],
        "metadata": {
            "total_runs": 0,
            "first_run": None,
            "last_run": None
        }
    }

def save_history(history: dict) -> None:
    """Save history to JSON file"""
    with open(HISTORY_JSON, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

def append_run_to_history(history: dict, run_data: dict) -> dict:
    """Add new run to history"""
    run_id = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    run_entry = {
        "run_id": run_id,
        "data_date": run_data["data_date"],
        "run_timestamp": datetime.now().isoformat(),
        "languages": run_data["languages"]
    }

    history["runs"].append(run_entry)
    history["metadata"]["total_runs"] = len(history["runs"])
    history["metadata"]["last_run"] = run_id
    if not history["metadata"]["first_run"]:
        history["metadata"]["first_run"] = run_id

    return history

# ─────────────────────────────────────────────────────────────
# DIFF CALCULATION ENGINE (SIMPLIFIED v2.0)
# ─────────────────────────────────────────────────────────────
def determine_period_category(today_date_str: str, past_date_str: str) -> Tuple[str, int, str]:
    """
    Determine if comparison is Weekly or Monthly based on days difference

    Args:
        today_date_str: Today's date (YYYY-MM-DD)
        past_date_str: Past date to compare against (YYYY-MM-DD)

    Returns:
        Tuple of (category, days_diff, period_title)
        - category: "weekly" or "monthly"
        - days_diff: Number of days between dates
        - period_title: e.g., "Period: 2025-11-18 to 2025-11-10 (8 days)"

    Examples:
        - 8 days apart: |8-7|=1 < |8-30|=22 → "weekly"
        - 13 days apart: |13-7|=6 < |13-30|=17 → "weekly"
        - 25 days apart: |25-7|=18 > |25-30|=5 → "monthly"
        - 39 days apart: |39-7|=32 > |39-30|=9 → "monthly"
    """
    today = datetime.strptime(today_date_str, "%Y-%m-%d")
    past = datetime.strptime(past_date_str, "%Y-%m-%d")

    days_diff = (today - past).days

    # Determine which is closer: 7 or 30?
    dist_to_7 = abs(days_diff - 7)
    dist_to_30 = abs(days_diff - 30)

    if dist_to_7 < dist_to_30:
        category = "weekly"
    else:
        category = "monthly"

    period_title = f"Period: {today_date_str} to {past_date_str} ({days_diff} days)"

    return category, days_diff, period_title

def find_past_run_in_history(history: dict, past_date: str) -> Optional[dict]:
    """
    Find a run in history matching the past date

    Args:
        history: History dict
        past_date: Past date to find (YYYY-MM-DD)

    Returns:
        Run dict or None if not found
    """
    for run in history["runs"]:
        if run["data_date"] == past_date:
            return run
    return None

def calculate_diff(current_value: float, previous_value: float) -> dict:
    """
    Calculate diff metrics

    Returns:
        {
            "net_change": float,
            "percent_change": float
        }
    """
    net_change = current_value - previous_value

    if previous_value == 0:
        percent_change = 0.0
    else:
        percent_change = (net_change / previous_value) * 100

    return {
        "net_change": net_change,
        "percent_change": percent_change
    }

def calculate_all_diffs(current_run: dict, history: dict, data_date: str) -> dict:
    """
    Calculate daily, weekly, and monthly diffs

    Returns:
        {
            "daily": {...},
            "weekly": {...},
            "monthly": {...}
        }
    """
    diffs = {}

    for period, days_back in [("daily", 1), ("weekly", 7), ("monthly", 30)]:
        comparison_run = find_comparison_run(history, data_date, days_back)

        if not comparison_run:
            diffs[period] = None
            continue

        period_diffs = {}

        for lang, lang_data in current_run["languages"].items():
            if lang not in comparison_run["languages"]:
                continue

            prev_lang = comparison_run["languages"][lang]

            # Full summary diffs
            full_diffs = {}
            for metric in ["total_words", "completed_words", "word_coverage_pct"]:
                current = lang_data["full_summary"][metric]
                previous = prev_lang["full_summary"][metric]
                full_diffs[metric] = calculate_diff(current, previous)

            # Detailed category diffs
            category_diffs = {}
            for cat, cat_data in lang_data["detailed_categories"].items():
                if cat not in prev_lang["detailed_categories"]:
                    continue

                prev_cat = prev_lang["detailed_categories"][cat]
                cat_diffs = {}

                for metric in ["total_words", "completed_words", "word_coverage_pct"]:
                    current = cat_data[metric]
                    previous = prev_cat[metric]
                    cat_diffs[metric] = calculate_diff(current, previous)

                category_diffs[cat] = cat_diffs

            period_diffs[lang] = {
                "full_summary": full_diffs,
                "detailed_categories": category_diffs,
                "comparison_date": comparison_run["data_date"]
            }

        diffs[period] = period_diffs

    return diffs

# ─────────────────────────────────────────────────────────────
# EXCEL UTILITIES
# ─────────────────────────────────────────────────────────────
def style_header(ws, row: int, headers: List[str]) -> None:
    """Style header row"""
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 2)

def style_separator(ws, row: int, num_cols: int) -> None:
    """Style separator row"""
    sep_fill = PatternFill("solid", fgColor="FFFF00")  # Yellow
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col, "")
        cell.fill = sep_fill

def apply_diff_color(cell):
    """Apply color based on diff value"""
    value = cell.value
    if value is None:
        return

    if value > 0:
        cell.font = Font(color="00B050")  # Green
    elif value < 0:
        cell.font = Font(color="FF0000")  # Red
    else:
        cell.font = Font(color="808080")  # Gray

# ─────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATION (V2.0)
# ─────────────────────────────────────────────────────────────
def create_full_summary_sheet_v2(wb: Workbook, sheet_name: str,
                                  current_data: dict, diff_data: Optional[dict],
                                  period_title: str, is_active: bool) -> None:
    """
    Create a full summary sheet with diff columns - V2.0

    Args:
        wb: Workbook
        sheet_name: Name of sheet
        current_data: Current language data
        diff_data: Diff data (or None)
        period_title: e.g., "Period: 2025-11-18 to 2025-11-10 (8 days)"
        is_active: True if this sheet should have data, False for N/A
    """
    ws = wb.create_sheet(sheet_name)

    if not is_active:
        # Show N/A message
        ws.cell(1, 1, "N/A - Select appropriate comparison period").font = Font(bold=True, size=14)
        if "Weekly" in sheet_name:
            ws.cell(2, 1, "This sheet is for comparisons around 7 days apart.")
            ws.cell(3, 1, "Your selected period was categorized as MONTHLY.")
        else:
            ws.cell(2, 1, "This sheet is for comparisons around 30 days apart.")
            ws.cell(3, 1, "Your selected period was categorized as WEEKLY.")
        return

    # Add period title row
    title_cell = ws.cell(1, 1, period_title)
    title_cell.font = Font(bold=True, size=14, color="4F81BD")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # Headers start at row 3
    headers = [
        "Language",
        "Total Words", "Completed Words", "Coverage %",
        "Total Δ", "Total Δ%",
        "Completed Δ", "Completed Δ%",
        "Coverage Δ", "Coverage Δ%"
    ]

    style_header(ws, 3, headers)

    row = 4
    for lang in sorted(current_data.keys()):
        lang_data = current_data[lang]["full_summary"]

        ws.cell(row, 1, lang)
        ws.cell(row, 2, lang_data["total_words"])
        ws.cell(row, 3, lang_data["completed_words"])
        ws.cell(row, 4, round(lang_data["word_coverage_pct"], 2))

        if diff_data and lang in diff_data:
            diffs = diff_data[lang]["full_summary"]

            ws.cell(row, 5, round(diffs["total_words"]["net_change"], 0))
            ws.cell(row, 6, round(diffs["total_words"]["percent_change"], 2))
            ws.cell(row, 7, round(diffs["completed_words"]["net_change"], 0))
            ws.cell(row, 8, round(diffs["completed_words"]["percent_change"], 2))
            ws.cell(row, 9, round(diffs["word_coverage_pct"]["net_change"], 2))
            ws.cell(row, 10, round(diffs["word_coverage_pct"]["percent_change"], 2))

            # Apply coloring
            for col in [5, 6, 7, 8, 9, 10]:
                cell = ws.cell(row, col)
                apply_diff_color(cell)

        row += 1

def create_detailed_sheet_v2(wb: Workbook, sheet_name: str,
                              current_data: dict, diff_data: Optional[dict],
                              period_title: str, is_active: bool) -> None:
    """
    Create detailed category sheet with diffs - V2.0

    Structure per language (matching wordcount1.py):
    - Period title row
    - Language title row (bold): "Language: ENG"
    - Header row: Category | Total Words | ...
    - Category data rows
    - Yellow separator row
    - Repeat for next language

    Args:
        wb: Workbook
        sheet_name: Name of sheet
        current_data: Current language data
        diff_data: Diff data (or None)
        period_title: e.g., "Period: 2025-11-18 to 2025-11-10 (8 days)"
        is_active: True if this sheet should have data, False for N/A
    """
    ws = wb.create_sheet(sheet_name)

    if not is_active:
        # Show N/A message
        ws.cell(1, 1, "N/A - Select appropriate comparison period").font = Font(bold=True, size=14)
        if "Weekly" in sheet_name:
            ws.cell(2, 1, "This sheet is for comparisons around 7 days apart.")
            ws.cell(3, 1, "Your selected period was categorized as MONTHLY.")
        else:
            ws.cell(2, 1, "This sheet is for comparisons around 30 days apart.")
            ws.cell(3, 1, "Your selected period was categorized as WEEKLY.")
        return

    # Add period title row
    title_cell = ws.cell(1, 1, period_title)
    title_cell.font = Font(bold=True, size=14, color="4F81BD")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = [
        "Category",
        "Total Words", "Completed Words", "Coverage %",
        "Total Δ", "Total Δ%",
        "Completed Δ", "Completed Δ%"
    ]

    current_row = 3

    for lang in sorted(current_data.keys()):
        lang_data = current_data[lang]["detailed_categories"]

        # Language title row (bold)
        lang_cell = ws.cell(current_row, 1, f"Language: {lang}")
        lang_cell.font = Font(bold=True, size=12)
        current_row += 1

        # Header row
        style_header(ws, current_row, headers)
        current_row += 1

        # Data rows for each category
        for cat in sorted(lang_data.keys()):
            cat_data = lang_data[cat]

            ws.cell(current_row, 1, cat)
            ws.cell(current_row, 2, cat_data["total_words"])
            ws.cell(current_row, 3, cat_data["completed_words"])
            ws.cell(current_row, 4, round(cat_data["word_coverage_pct"], 2))

            if diff_data and lang in diff_data:
                cat_diffs = diff_data[lang]["detailed_categories"].get(cat)
                if cat_diffs:
                    ws.cell(current_row, 5, round(cat_diffs["total_words"]["net_change"], 0))
                    ws.cell(current_row, 6, round(cat_diffs["total_words"]["percent_change"], 2))
                    ws.cell(current_row, 7, round(cat_diffs["completed_words"]["net_change"], 0))
                    ws.cell(current_row, 8, round(cat_diffs["completed_words"]["percent_change"], 2))

                    # Apply coloring
                    for col in [5, 6, 7, 8]:
                        cell = ws.cell(current_row, col)
                        apply_diff_color(cell)

            current_row += 1

        # Add yellow separator between languages
        style_separator(ws, current_row, len(headers))
        current_row += 1

# ─────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATION (OLD V1.0 - DEPRECATED)
# ─────────────────────────────────────────────────────────────
def create_full_summary_sheet(wb: Workbook, sheet_name: str,
                               current_data: dict, diff_data: Optional[dict]) -> None:
    """
    Create a full summary sheet with diff columns
    """
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Language",
        "Total Words", "Completed Words", "Coverage %",
        "Total Δ", "Total Δ%",
        "Completed Δ", "Completed Δ%",
        "Coverage Δ", "Coverage Δ%"
    ]

    style_header(ws, 1, headers)

    row = 2
    for lang in sorted(current_data.keys()):
        lang_data = current_data[lang]["full_summary"]

        ws.cell(row, 1, lang)
        ws.cell(row, 2, lang_data["total_words"])
        ws.cell(row, 3, lang_data["completed_words"])
        ws.cell(row, 4, round(lang_data["word_coverage_pct"], 2))

        if diff_data and lang in diff_data:
            diffs = diff_data[lang]["full_summary"]

            # Total words diff
            ws.cell(row, 5, round(diffs["total_words"]["net_change"], 0))
            ws.cell(row, 6, round(diffs["total_words"]["percent_change"], 2))

            # Completed words diff
            ws.cell(row, 7, round(diffs["completed_words"]["net_change"], 0))
            ws.cell(row, 8, round(diffs["completed_words"]["percent_change"], 2))

            # Coverage diff
            ws.cell(row, 9, round(diffs["word_coverage_pct"]["net_change"], 2))
            ws.cell(row, 10, round(diffs["word_coverage_pct"]["percent_change"], 2))

            # Apply coloring
            for col in [5, 6, 7, 8, 9, 10]:
                cell = ws.cell(row, col)
                apply_diff_color(cell)

        row += 1

def create_detailed_sheet(wb: Workbook, sheet_name: str,
                          current_data: dict, diff_data: Optional[dict]) -> None:
    """
    Create detailed category sheet with diffs

    Structure per language (matching wordcount1.py):
    - Language title row (bold): "Language: ENG"
    - Header row: Category | Total Words | Completed Words | Coverage % | diffs...
    - Category data rows
    - Yellow separator row
    - Repeat for next language
    """
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Category",
        "Total Words", "Completed Words", "Coverage %",
        "Total Δ", "Total Δ%",
        "Completed Δ", "Completed Δ%"
    ]

    current_row = 1

    for lang in sorted(current_data.keys()):
        lang_data = current_data[lang]["detailed_categories"]

        # Language title row (bold)
        lang_cell = ws.cell(current_row, 1, f"Language: {lang}")
        lang_cell.font = Font(bold=True, size=12)
        current_row += 1

        # Header row
        style_header(ws, current_row, headers)
        current_row += 1

        # Data rows for each category
        for cat in sorted(lang_data.keys()):
            cat_data = lang_data[cat]

            ws.cell(current_row, 1, cat)
            ws.cell(current_row, 2, cat_data["total_words"])
            ws.cell(current_row, 3, cat_data["completed_words"])
            ws.cell(current_row, 4, round(cat_data["word_coverage_pct"], 2))

            if diff_data and lang in diff_data:
                cat_diffs = diff_data[lang]["detailed_categories"].get(cat)
                if cat_diffs:
                    ws.cell(current_row, 5, round(cat_diffs["total_words"]["net_change"], 0))
                    ws.cell(current_row, 6, round(cat_diffs["total_words"]["percent_change"], 2))
                    ws.cell(current_row, 7, round(cat_diffs["completed_words"]["net_change"], 0))
                    ws.cell(current_row, 8, round(cat_diffs["completed_words"]["percent_change"], 2))

                    # Apply coloring
                    for col in [5, 6, 7, 8]:
                        cell = ws.cell(current_row, col)
                        apply_diff_color(cell)

            current_row += 1

        # Add yellow separator between languages
        style_separator(ws, current_row, len(headers))
        current_row += 1

def delete_old_reports() -> None:
    """Delete all existing WordCountAnalysis_*.xlsx files"""
    for file in Path.cwd().glob(f"{OUTPUT_EXCEL_PREFIX}_*.xlsx"):
        try:
            file.unlink()
            print(f"    Deleted old report: {file.name}")
        except Exception as e:
            print(f"    Warning: Could not delete {file.name}: {e}")

def generate_excel_report_v2(current_data: dict, diffs: Optional[dict],
                             category: str, period_title: str, run_id: str) -> Path:
    """
    Generate complete Excel report - V2.0 (4 sheets)

    Creates 4 sheets:
    1. Weekly Diff - Full Summary (has data if category="weekly", else N/A)
    2. Monthly Diff - Full Summary (has data if category="monthly", else N/A)
    3. Weekly Diff - Detailed (has data if category="weekly", else N/A)
    4. Monthly Diff - Detailed (has data if category="monthly", else N/A)

    Args:
        current_data: Current language data
        diffs: Diff data (or None if no past data)
        category: "weekly" or "monthly"
        period_title: e.g., "Period: 2025-11-18 to 2025-11-10 (8 days)"
        run_id: Timestamp for filename
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Determine which sheets get data
    weekly_diffs = diffs if category == "weekly" else None
    monthly_diffs = diffs if category == "monthly" else None

    # Create all 4 sheets
    create_full_summary_sheet_v2(wb, "Weekly Diff - Full",
                                  current_data, weekly_diffs, period_title, category == "weekly")
    create_full_summary_sheet_v2(wb, "Monthly Diff - Full",
                                  current_data, monthly_diffs, period_title, category == "monthly")

    create_detailed_sheet_v2(wb, "Weekly Diff - Detailed",
                              current_data, weekly_diffs, period_title, category == "weekly")
    create_detailed_sheet_v2(wb, "Monthly Diff - Detailed",
                              current_data, monthly_diffs, period_title, category == "monthly")

    # Save with timestamped filename
    output_path = Path.cwd() / f"{OUTPUT_EXCEL_PREFIX}_{run_id}.xlsx"
    wb.save(output_path)

    return output_path

# ─────────────────────────────────────────────────────────────
# MAIN EXECUTION
# ─────────────────────────────────────────────────────────────
def main() -> None:
    """Main execution - V2.0 Simplified Workflow"""
    print("="*60)
    print("WordCount Diff Master Report Generator v2.0")
    print("="*60)

    # Step 1: Get today's date and process current data
    today = datetime.now().strftime("%Y-%m-%d")
    print(f"\nTODAY: {today}")
    print("Processing current XML files...")

    # Step 2: Scan and collect TODAY's word count data
    print("\n[1/5] Scanning language files...")
    current_languages = collect_all_languages_data()
    print(f"Processed {len(current_languages)} languages")

    # Step 3: Get past date from user for comparison
    past_date = get_comparison_date_from_user()

    # Step 4: Determine period category (weekly or monthly)
    category, days_diff, period_title = determine_period_category(today, past_date)
    print(f"\n[2/5] Period Analysis...")
    print(f"  {period_title}")
    print(f"  Category: {category.upper()}")

    # Step 5: Load history and find past run
    print(f"\n[3/5] Loading history...")
    history = load_history()
    print(f"  Found {len(history['runs'])} previous runs")

    past_run = find_past_run_in_history(history, past_date)
    if not past_run:
        print(f"\n  WARNING: No data found for {past_date}")
        print(f"  Available dates in history:")
        for run in history["runs"][-10:]:  # Show last 10
            print(f"    - {run['data_date']}")
        print(f"\n  Continuing without diff data...")
        diffs = None
    else:
        print(f"  Found comparison data for {past_date}")

        # Step 6: Calculate diffs
        print(f"\n[4/5] Calculating diffs...")
        diffs = {}
        for lang in current_languages.keys():
            if lang not in past_run["languages"]:
                continue

            current_lang = current_languages[lang]
            past_lang = past_run["languages"][lang]

            # Full summary diffs
            full_diffs = {}
            for metric in ["total_words", "completed_words", "word_coverage_pct"]:
                current_val = current_lang["full_summary"][metric]
                past_val = past_lang["full_summary"][metric]
                full_diffs[metric] = calculate_diff(current_val, past_val)

            # Detailed category diffs
            category_diffs = {}
            for cat in current_lang["detailed_categories"].keys():
                if cat not in past_lang["detailed_categories"]:
                    continue

                current_cat = current_lang["detailed_categories"][cat]
                past_cat = past_lang["detailed_categories"][cat]
                cat_diffs = {}

                for metric in ["total_words", "completed_words", "word_coverage_pct"]:
                    current_val = current_cat[metric]
                    past_val = past_cat[metric]
                    cat_diffs[metric] = calculate_diff(current_val, past_val)

                category_diffs[cat] = cat_diffs

            diffs[lang] = {
                "full_summary": full_diffs,
                "detailed_categories": category_diffs
            }

        print(f"  Calculated diffs for {len(diffs)} languages")

    # Step 7: Update history with today's data
    print(f"\n[5/5] Updating history...")
    current_run = {
        "data_date": today,  # Always today in V2.0
        "languages": current_languages
    }
    history = append_run_to_history(history, current_run)
    save_history(history)
    print(f"  History updated (total runs: {history['metadata']['total_runs']})")

    # Step 8: Delete old reports
    print(f"\nCleaning up old reports...")
    delete_old_reports()

    # Step 9: Generate new Excel report
    print(f"\nGenerating Excel report...")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = generate_excel_report_v2(current_languages, diffs, category, period_title, run_id)
    print(f"  Report saved: {output_path.name}")

    # Final summary
    print("\n" + "="*60)
    print("COMPLETE")
    print("="*60)
    print(f"Report: {output_path.resolve()}")
    print(f"History: {HISTORY_JSON.resolve()}")
    print(f"Period: {period_title}")
    print(f"Category: {category.upper()}")
    print("="*60)

if __name__ == "__main__":
    main()
