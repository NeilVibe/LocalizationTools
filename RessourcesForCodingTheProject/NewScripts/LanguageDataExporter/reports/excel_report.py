"""
Excel Report Writer for Word Count Reports.

Generates a SINGLE unified Excel file with:
- General Summary sheet: Overview totals per language
- Detailed Summary sheet: All language tables stacked vertically
"""

import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .report_generator import WordCountReport, LanguageReport

logger = logging.getLogger(__name__)

# Import CATEGORY_COLORS from config
try:
    from config import CATEGORY_COLORS
except ImportError:
    CATEGORY_COLORS = {
        "Sequencer": "FFE599",
        "AIDialog": "C6EFCE",
        "QuestDialog": "C6EFCE",
        "NarrationDialog": "C6EFCE",
        "Item": "D9D2E9",
        "Quest": "D9D2E9",
        "Character": "F8CBAD",
        "Gimmick": "D9D2E9",
        "Skill": "D9D2E9",
        "Knowledge": "D9D2E9",
        "Faction": "D9D2E9",
        "UI": "A9D08E",
        "Region": "F8CBAD",
        "System_Misc": "D9D9D9",
        "Uncategorized": "DDD9C4",
    }


def _get_category_fill(category: str) -> PatternFill:
    """Get PatternFill for a category."""
    color = CATEGORY_COLORS.get(category, "FFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# =============================================================================
# STYLING CONSTANTS
# =============================================================================

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

DATA_ALIGNMENT = Alignment(horizontal="right", vertical="center")
CATEGORY_ALIGNMENT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

UNTRANSLATED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")


class ExcelReportWriter:
    """
    Writes word count reports to a SINGLE unified Excel file.

    Structure:
    - Sheet 1: General Summary (one row per language with totals)
    - Sheet 2: Detailed Summary (all language tables stacked vertically)
    """

    def __init__(self, output_path: Path):
        self.output_path = output_path

    def write_report(self, report: WordCountReport) -> bool:
        """Write complete word count report to Excel."""
        try:
            wb = Workbook()

            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Create General Summary sheet
            self._write_general_summary(wb, report)

            # Create Detailed Summary sheet (all languages stacked)
            self._write_detailed_summary(wb, report)

            # Ensure output directory exists
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            wb.save(self.output_path)

            logger.info(f"Generated unified report: {self.output_path.name}")
            return True

        except Exception as e:
            logger.error(f"Error writing Excel report: {e}")
            return False

    def _write_general_summary(self, wb: Workbook, report: WordCountReport):
        """
        Write General Summary sheet.

        Columns: Language | Total Strings | Korean Words | Translation Count | Untranslated Strings | Untranslated KR Words | Count Type
        """
        ws = wb.create_sheet("General Summary")

        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value="Word Count Summary - All Languages")
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Headers (row 3)
        headers = ["Language", "Total Strings", "Korean Words", "Translation Count",
                   "Untranslated Strings", "Untranslated KR Words", "Count Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Data rows
        row = 4

        for lang_code in sorted(report.languages.keys()):
            lang_report = report.languages[lang_code]
            wc = lang_report.word_count

            ws.cell(row=row, column=1, value=lang_report.display_name).alignment = CATEGORY_ALIGNMENT
            ws.cell(row=row, column=2, value=wc.total_strings).alignment = DATA_ALIGNMENT
            ws.cell(row=row, column=3, value=wc.total_korean_words).alignment = DATA_ALIGNMENT
            ws.cell(row=row, column=4, value=wc.total_translation_count).alignment = DATA_ALIGNMENT
            ws.cell(row=row, column=5, value=wc.total_untranslated_strings).alignment = DATA_ALIGNMENT
            ws.cell(row=row, column=6, value=wc.total_untranslated_korean_words).alignment = DATA_ALIGNMENT
            ws.cell(row=row, column=7, value=lang_report.count_type).alignment = DATA_ALIGNMENT

            # Highlight untranslated columns if > 0
            if wc.total_untranslated_strings > 0:
                ws.cell(row=row, column=5).fill = UNTRANSLATED_FILL
            if wc.total_untranslated_korean_words > 0:
                ws.cell(row=row, column=6).fill = UNTRANSLATED_FILL

            # Apply borders and number format
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
                if col in (2, 3, 4, 5, 6):
                    ws.cell(row=row, column=col).number_format = "#,##0"

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 12

        # Freeze header
        ws.freeze_panes = 'A4'

    def _write_detailed_summary(self, wb: Workbook, report: WordCountReport):
        """
        Write Detailed Summary sheet with all language tables stacked vertically.

        Each table has:
        - Title row with language name
        - Headers: Category | Strings | Korean Words | Translation Words/Chars | Untranslated Strings | Untranslated KR Words
        - Data rows per category
        - Total row
        - Blank rows before next table
        """
        ws = wb.create_sheet("Detailed Summary")

        categories = report.get_sorted_categories()
        row = 1

        for lang_code in sorted(report.languages.keys()):
            lang_report = report.languages[lang_code]
            wc = lang_report.word_count
            count_label = lang_report.count_type

            # === Language Title ===
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            title_cell = ws.cell(row=row, column=1, value=f"{lang_report.display_name} ({count_label})")
            title_cell.font = TITLE_FONT
            title_cell.fill = TITLE_FILL
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

            # === Headers ===
            headers = ["Category", "Strings", "Korean Words", f"Translation {count_label}",
                       "Untranslated Strings", "Untranslated KR Words"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            row += 1

            # === Data rows per category ===
            for category in categories:
                cat_count = wc.categories.get(category)
                cat_fill = _get_category_fill(category)

                if cat_count:
                    ws.cell(row=row, column=1, value=category).alignment = CATEGORY_ALIGNMENT
                    ws.cell(row=row, column=2, value=cat_count.total_strings).alignment = DATA_ALIGNMENT
                    ws.cell(row=row, column=3, value=cat_count.korean_words).alignment = DATA_ALIGNMENT
                    ws.cell(row=row, column=4, value=cat_count.translation_count).alignment = DATA_ALIGNMENT
                    ws.cell(row=row, column=5, value=cat_count.untranslated_strings).alignment = DATA_ALIGNMENT
                    ws.cell(row=row, column=6, value=cat_count.untranslated_korean_words).alignment = DATA_ALIGNMENT

                    # Highlight untranslated columns
                    if cat_count.untranslated_strings > 0:
                        ws.cell(row=row, column=5).fill = UNTRANSLATED_FILL
                    else:
                        ws.cell(row=row, column=5).fill = cat_fill
                    if cat_count.untranslated_korean_words > 0:
                        ws.cell(row=row, column=6).fill = UNTRANSLATED_FILL
                    else:
                        ws.cell(row=row, column=6).fill = cat_fill
                else:
                    ws.cell(row=row, column=1, value=category).alignment = CATEGORY_ALIGNMENT
                    for col in range(2, 7):
                        ws.cell(row=row, column=col, value=0).alignment = DATA_ALIGNMENT
                        ws.cell(row=row, column=col).fill = cat_fill

                # Apply borders and category color
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = THIN_BORDER
                    if col < 5:
                        cell.fill = cat_fill
                    if col > 1:
                        cell.number_format = "#,##0"

                row += 1

            # === Total row ===
            ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            ws.cell(row=row, column=1).alignment = CATEGORY_ALIGNMENT

            totals = [
                wc.total_strings,
                wc.total_korean_words,
                wc.total_translation_count,
                wc.total_untranslated_strings,
                wc.total_untranslated_korean_words
            ]

            for col, total in enumerate(totals, 2):
                cell = ws.cell(row=row, column=col, value=total)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.alignment = DATA_ALIGNMENT
                cell.number_format = "#,##0"

                # Highlight untranslated totals if > 0
                if col == 5 and total > 0:
                    cell.fill = UNTRANSLATED_FILL
                if col == 6 and total > 0:
                    cell.fill = UNTRANSLATED_FILL

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER

            row += 1

            # === Blank rows before next table ===
            row += 2

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 22

        # Freeze first column
        ws.freeze_panes = 'B1'
