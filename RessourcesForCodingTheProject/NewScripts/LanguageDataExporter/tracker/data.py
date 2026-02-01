"""
Weekly Data Manager for Correction Progress Tracker.

Handles _WEEKLY_DATA sheet with REPLACE mode to prevent duplicates.
Key = (week_start, language) - same key overwrites existing row.

Schema tracks merge results: Corrections, Success, Fail per language per week.
"""

import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

# _WEEKLY_DATA schema (NEW: merge-based tracking with per-category breakdown)
DATA_SHEET_NAME = "_WEEKLY_DATA"
DATA_HEADERS = [
    "WeekStart",      # Monday date (YYYY-MM-DD) from file mod time
    "Language",       # Language code (ENG, FRE, etc.)
    "Category",       # Content category (Sequencer, Item, Quest, etc.)
    "MergeDate",      # When merge was performed (YYYY-MM-DD HH:MM:SS)
    "Corrections",    # Count of rows with Correction values
    "Success",        # Successfully merged (matched in LOCDEV)
    "Fail",           # Failed to match in LOCDEV
    "Timestamp",      # When this record was written
]


def get_week_start(date: datetime) -> str:
    """
    Convert date to week start (Monday) in YYYY-MM-DD format.

    Args:
        date: Any datetime

    Returns:
        String date of that week's Monday (YYYY-MM-DD)
    """
    monday = date - timedelta(days=date.weekday())
    return monday.strftime("%Y-%m-%d")


class WeeklyDataManager:
    """
    Manages _WEEKLY_DATA sheet with REPLACE mode.

    Key: (week_start, language, category)
    Same key overwrites existing row instead of appending.
    """

    def __init__(self, tracker_path: Path):
        """
        Initialize with tracker file path.

        Args:
            tracker_path: Path to Correction_ProgressTracker.xlsx
        """
        self.tracker_path = tracker_path

    def _get_or_create_workbook(self) -> Workbook:
        """Load existing workbook or create new one."""
        if self.tracker_path.exists():
            return load_workbook(self.tracker_path)
        else:
            wb = Workbook()
            # Remove default sheet, we'll create our own
            wb.remove(wb.active)
            return wb

    def _migrate_old_schema(self, ws) -> None:
        """
        Migrate old schema (without Category) to new schema (with Category).

        Old: WeekStart | Language | MergeDate | Corrections | Success | Fail | Timestamp
        New: WeekStart | Language | Category | MergeDate | Corrections | Success | Fail | Timestamp

        Inserts Category column at position 3, shifts other columns right.
        """
        logger.info("Migrating tracker schema: adding Category column...")

        # Insert new column at position 3
        ws.insert_cols(3)

        # Set header for new column
        ws.cell(row=1, column=3, value="Category")
        ws.cell(row=1, column=3).font = Font(bold=True)

        # Fill existing data rows with "Uncategorized"
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=3, value="Uncategorized")

        logger.info(f"Migrated {ws.max_row - 1} rows to new schema with Category column")

    def _get_or_create_data_sheet(self, wb: Workbook):
        """Get or create _WEEKLY_DATA sheet with headers."""
        if DATA_SHEET_NAME in wb.sheetnames:
            ws = wb[DATA_SHEET_NAME]

            # Check if migration needed (old schema without Category)
            header_col3 = ws.cell(row=1, column=3).value
            if header_col3 != "Category":
                self._migrate_old_schema(ws)

            return ws

        # Create sheet at the end
        ws = wb.create_sheet(DATA_SHEET_NAME)

        # Write headers
        for col, header in enumerate(DATA_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Hide the sheet (raw data) - but only if other visible sheets exist
        # Excel/openpyxl requires at least one visible sheet in workbook
        visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state != 'hidden']
        if len(visible_sheets) > 1:
            ws.sheet_state = 'hidden'
        # Otherwise leave visible - will be hidden after WEEKLY/TOTAL sheets are created

        return ws

    def _build_key_index(self, ws) -> Dict[Tuple[str, str, str], int]:
        """
        Build index of existing rows by key.

        Returns:
            Dict mapping (week_start, language, category) to row number
        """
        index = {}
        for row in range(2, ws.max_row + 1):
            week_start = ws.cell(row=row, column=1).value
            language = ws.cell(row=row, column=2).value
            category = ws.cell(row=row, column=3).value
            if week_start and language and category:
                index[(str(week_start), str(language), str(category))] = row
        return index

    def write_data(self, records: List[Dict]) -> int:
        """
        Write records to _WEEKLY_DATA with REPLACE mode.

        Args:
            records: List of dicts with keys:
                - WeekStart: Monday date (YYYY-MM-DD)
                - Language: Language code
                - Category: Content category (Sequencer, Item, Quest, etc.)
                - MergeDate: When merge was performed
                - Corrections: Count of rows with Correction values
                - Success: Successfully merged count
                - Fail: Failed to merge count

        Returns:
            Number of records written/updated
        """
        if not records:
            return 0

        wb = self._get_or_create_workbook()
        ws = self._get_or_create_data_sheet(wb)

        # Build key index for REPLACE mode
        key_index = self._build_key_index(ws)

        # Find next available row for new records
        next_row = ws.max_row + 1 if ws.max_row > 1 else 2

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        written = 0

        for record in records:
            week_start = str(record.get("WeekStart", ""))
            language = str(record.get("Language", ""))
            category = str(record.get("Category", "Uncategorized"))

            if not all([week_start, language, category]):
                continue

            key = (week_start, language, category)

            # REPLACE mode: use existing row or new row
            if key in key_index:
                row = key_index[key]
            else:
                row = next_row
                next_row += 1

            # Write record (column order matches DATA_HEADERS)
            ws.cell(row=row, column=1, value=week_start)
            ws.cell(row=row, column=2, value=language)
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=record.get("MergeDate", timestamp))
            ws.cell(row=row, column=5, value=record.get("Corrections", 0))
            ws.cell(row=row, column=6, value=record.get("Success", 0))
            ws.cell(row=row, column=7, value=record.get("Fail", 0))
            ws.cell(row=row, column=8, value=timestamp)

            written += 1

        # Save and close workbook
        self.tracker_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.tracker_path)
        wb.close()

        logger.info(f"Wrote {written} records to {DATA_SHEET_NAME}")
        return written

    def _detect_schema_version(self, ws) -> bool:
        """
        Detect if the schema has the Category column.

        Returns:
            True if new schema (with Category), False if old schema
        """
        # Check header row for "Category" in column 3
        header_col3 = ws.cell(row=1, column=3).value
        return header_col3 == "Category"

    def read_all_data(self) -> List[Dict]:
        """
        Read all records from _WEEKLY_DATA.

        Handles both old schema (without Category) and new schema (with Category).

        Returns:
            List of dicts with keys matching DATA_HEADERS
        """
        if not self.tracker_path.exists():
            return []

        wb = load_workbook(self.tracker_path, read_only=True)

        if DATA_SHEET_NAME not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[DATA_SHEET_NAME]
        has_category = self._detect_schema_version(ws)
        records = []

        for row in range(2, ws.max_row + 1):
            if has_category:
                # New schema: 8 columns with Category
                record = {
                    "WeekStart": ws.cell(row=row, column=1).value,
                    "Language": ws.cell(row=row, column=2).value,
                    "Category": ws.cell(row=row, column=3).value or "Uncategorized",
                    "MergeDate": ws.cell(row=row, column=4).value,
                    "Corrections": ws.cell(row=row, column=5).value or 0,
                    "Success": ws.cell(row=row, column=6).value or 0,
                    "Fail": ws.cell(row=row, column=7).value or 0,
                    "Timestamp": ws.cell(row=row, column=8).value,
                }
            else:
                # Old schema: 7 columns without Category
                record = {
                    "WeekStart": ws.cell(row=row, column=1).value,
                    "Language": ws.cell(row=row, column=2).value,
                    "Category": "Uncategorized",  # Default for old data
                    "MergeDate": ws.cell(row=row, column=3).value,
                    "Corrections": ws.cell(row=row, column=4).value or 0,
                    "Success": ws.cell(row=row, column=5).value or 0,
                    "Fail": ws.cell(row=row, column=6).value or 0,
                    "Timestamp": ws.cell(row=row, column=7).value,
                }

            if record["WeekStart"] and record["Language"]:
                records.append(record)

        wb.close()
        return records

    def get_latest_week_data(self) -> Dict[str, Dict]:
        """
        Get data for the most recent week only.

        Returns:
            Dict[language, Dict] with latest week data per language
            Each dict has: Corrections, Success, Fail, MergeDate, by_category
        """
        all_data = self.read_all_data()
        if not all_data:
            return {}

        # Find latest week
        weeks = set(r["WeekStart"] for r in all_data)
        if not weeks:
            return {}
        latest_week = max(weeks)

        # Filter to latest week and aggregate by language
        result = {}
        for record in all_data:
            if record["WeekStart"] != latest_week:
                continue

            lang = record["Language"]
            category = record["Category"]

            if lang not in result:
                result[lang] = {
                    "Corrections": 0,
                    "Success": 0,
                    "Fail": 0,
                    "MergeDate": record["MergeDate"],
                    "by_category": {},
                }

            # Aggregate totals
            result[lang]["Corrections"] += record["Corrections"]
            result[lang]["Success"] += record["Success"]
            result[lang]["Fail"] += record["Fail"]

            # Store per-category breakdown
            result[lang]["by_category"][category] = {
                "Corrections": record["Corrections"],
                "Success": record["Success"],
                "Fail": record["Fail"],
            }

        return result

    def get_weekly_summary(self) -> List[Dict]:
        """
        Get week-by-week summary per language (aggregated from per-category data).

        Returns:
            List of dicts with: WeekStart, Language, Corrections, Success, Fail, SuccessRate, by_category
        """
        all_data = self.read_all_data()
        if not all_data:
            return []

        # Aggregate by (week, language) since data is per-category
        aggregated = {}
        for record in all_data:
            key = (record["WeekStart"], record["Language"])
            category = record["Category"]

            if key not in aggregated:
                aggregated[key] = {
                    "WeekStart": record["WeekStart"],
                    "Language": record["Language"],
                    "Corrections": 0,
                    "Success": 0,
                    "Fail": 0,
                    "by_category": {},
                }

            aggregated[key]["Corrections"] += record["Corrections"]
            aggregated[key]["Success"] += record["Success"]
            aggregated[key]["Fail"] += record["Fail"]
            aggregated[key]["by_category"][category] = {
                "Corrections": record["Corrections"],
                "Success": record["Success"],
                "Fail": record["Fail"],
            }

        # Build summary list
        summary = []
        for key in sorted(aggregated.keys()):
            data = aggregated[key]
            corrections = data["Corrections"]
            success = data["Success"]
            fail = data["Fail"]
            success_rate = (success / corrections * 100) if corrections > 0 else 0

            summary.append({
                "WeekStart": data["WeekStart"],
                "Language": data["Language"],
                "Corrections": corrections,
                "Success": success,
                "Fail": fail,
                "SuccessRate": round(success_rate, 1),
                "by_category": data["by_category"],
            })

        return summary
