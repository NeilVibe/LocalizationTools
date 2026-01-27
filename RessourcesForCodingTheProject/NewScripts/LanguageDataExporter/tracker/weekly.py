"""
WEEKLY sheet builder for Correction Progress Tracker.

Shows week-over-week merge results per language with:
- Language, Week, Corrections, Success, Fail, Success %
- Professional formatting with alternating rows, thick borders, color-coded rates
"""

import logging
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

WEEKLY_SHEET_NAME = "WEEKLY"

# Column configuration
WEEKLY_HEADERS = ["Language", "Week", "Corrections", "Success", "Fail", "Success %"]
WEEKLY_WIDTHS = [14, 14, 14, 12, 10, 12]

# =============================================================================
# STYLING CONSTANTS (Professional Excel Theme)
# =============================================================================

# Colors (Microsoft Office Blue palette)
BLUE_DARK = "1F4E79"
BLUE_HEADER = "4472C4"
BLUE_LIGHT = "D9E2F3"
GREEN_SUCCESS = "C6EFCE"
GREEN_DARK = "63BE7B"
YELLOW_WARN = "FFEB9C"
RED_FAIL = "FFC7CE"
GRAY_ALT = "F2F2F2"

# Title styling
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Header styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Data styling
DATA_FONT = Font(size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Alternating row fills
ROW_FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_FILL_ALT = PatternFill(start_color=GRAY_ALT, end_color=GRAY_ALT, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='medium', color=BLUE_DARK)  # Thick bottom border
)

TITLE_BORDER = Border(
    left=Side(style='medium', color=BLUE_DARK),
    right=Side(style='medium', color=BLUE_DARK),
    top=Side(style='medium', color=BLUE_DARK),
    bottom=Side(style='medium', color=BLUE_DARK)
)


def get_success_rate_fill(rate: float) -> PatternFill:
    """Return fill color based on success rate (0-100)."""
    if rate >= 95:
        return PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
    elif rate >= 80:
        return PatternFill(start_color=GREEN_SUCCESS, end_color=GREEN_SUCCESS, fill_type="solid")
    elif rate >= 60:
        return PatternFill(start_color=YELLOW_WARN, end_color=YELLOW_WARN, fill_type="solid")
    else:
        return PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type="solid")


def build_weekly_sheet(wb: Workbook, weekly_data: List[Dict]) -> None:
    """
    Build the WEEKLY sheet with week-over-week merge results.

    Args:
        wb: Workbook to add sheet to
        weekly_data: List of dicts from WeeklyDataManager.get_weekly_summary()
    """
    # Create or get sheet
    if WEEKLY_SHEET_NAME in wb.sheetnames:
        del wb[WEEKLY_SHEET_NAME]

    ws = wb.create_sheet(WEEKLY_SHEET_NAME, 0)  # First position

    # ==========================================================================
    # TITLE ROW
    # ==========================================================================
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="WEEKLY MERGE RESULTS")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = TITLE_ALIGNMENT
    title_cell.border = TITLE_BORDER

    # Apply border to merged cells
    for col in range(2, 7):
        ws.cell(row=1, column=col).border = TITLE_BORDER

    # Empty row 2 for spacing
    ws.row_dimensions[2].height = 8

    # ==========================================================================
    # HEADER ROW (row 3)
    # ==========================================================================
    header_row = 3
    for col, header in enumerate(WEEKLY_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    # Set column widths
    for col, width in enumerate(WEEKLY_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ==========================================================================
    # DATA ROWS
    # ==========================================================================
    if not weekly_data:
        ws.cell(row=4, column=1, value="No data yet. Run 'Merge to LOCDEV' to collect data.")
        ws.merge_cells('A4:F4')
        return

    data_row = header_row + 1
    for idx, record in enumerate(weekly_data):
        # Alternating row color
        row_fill = ROW_FILL_ALT if (idx % 2 == 0) else ROW_FILL_WHITE

        # Language
        cell = ws.cell(row=data_row, column=1, value=record["Language"])
        cell.font = Font(bold=True, size=10)
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER

        # Week
        cell = ws.cell(row=data_row, column=2, value=record["WeekStart"])
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER

        # Corrections (with thousands separator)
        cell = ws.cell(row=data_row, column=3, value=record["Corrections"])
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Success (with thousands separator)
        cell = ws.cell(row=data_row, column=4, value=record["Success"])
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Fail (with thousands separator)
        cell = ws.cell(row=data_row, column=5, value=record["Fail"])
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Success % with color-coded fill
        success_rate = record["SuccessRate"]
        cell = ws.cell(row=data_row, column=6, value=success_rate / 100)
        cell.font = Font(bold=True, size=10)
        cell.fill = get_success_rate_fill(success_rate)
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '0.0%'

        data_row += 1

    # ==========================================================================
    # FINAL TOUCHES
    # ==========================================================================

    # Freeze header
    ws.freeze_panes = 'A4'

    # Set row height for data rows
    for row in range(4, data_row):
        ws.row_dimensions[row].height = 18

    logger.info(f"Built WEEKLY sheet with {len(weekly_data)} records")
