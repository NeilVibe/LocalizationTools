"""
Main Correction Tracker class.

Orchestrates data collection from merge results and report generation.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Callable

from openpyxl import Workbook, load_workbook

from .data import WeeklyDataManager, get_week_start
from .weekly import build_weekly_sheet
from .total import build_total_sheet

logger = logging.getLogger(__name__)


class CorrectionTracker:
    """
    Main tracker class for merge progress.

    Usage:
        tracker = CorrectionTracker(tracker_path, categories)
        tracker.update_and_rebuild_from_merge(merge_results)
    """

    def __init__(self, tracker_path: Path, categories: List[str]):
        """
        Initialize tracker.

        Args:
            tracker_path: Path to Correction_ProgressTracker.xlsx
            categories: List of category names for reports
        """
        self.tracker_path = tracker_path
        self.categories = categories
        self.data_manager = WeeklyDataManager(tracker_path)

    def update_from_merge_results(
        self,
        merge_results: Dict,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> int:
        """
        Update tracker data from merge results.

        Args:
            merge_results: Dict from merge_all_corrections() with:
                - file_results: Dict[lang, Dict] with corrections, success, fail
                - file_mod_times: Dict[lang, datetime]
            progress_callback: Optional progress callback

        Returns:
            Number of records written
        """
        file_results = merge_results.get("file_results", {})
        file_mod_times = merge_results.get("file_mod_times", {})

        if not file_results:
            logger.warning("No merge results to record")
            return 0

        if progress_callback:
            progress_callback(10, "Building tracker records...")

        # Build records for _WEEKLY_DATA
        records = []
        now = datetime.now()
        merge_date = now.strftime("%Y-%m-%d %H:%M:%S")

        for lang, result in file_results.items():
            # Determine week start from file mod time or current date
            if lang in file_mod_times:
                week_start = get_week_start(file_mod_times[lang])
            else:
                week_start = get_week_start(now)

            records.append({
                "WeekStart": week_start,
                "Language": lang,
                "MergeDate": merge_date,
                "Corrections": result.get("corrections", 0),
                "Success": result.get("success", 0),
                "Fail": result.get("fail", 0),
            })

        if progress_callback:
            progress_callback(50, f"Writing {len(records)} records...")

        # Write to _WEEKLY_DATA with REPLACE mode
        written = self.data_manager.write_data(records)

        if progress_callback:
            progress_callback(100, f"Wrote {written} records")

        return written

    def rebuild_report(self) -> bool:
        """
        Rebuild WEEKLY and TOTAL sheets from _WEEKLY_DATA.

        Returns:
            True if successful
        """
        try:
            # Load or create workbook
            if self.tracker_path.exists():
                wb = load_workbook(self.tracker_path)
            else:
                wb = Workbook()
                wb.remove(wb.active)

            # Get data for reports
            weekly_summary = self.data_manager.get_weekly_summary()
            latest_data = self.data_manager.get_latest_week_data()

            # Build sheets
            build_weekly_sheet(wb, weekly_summary)
            build_total_sheet(wb, latest_data)

            # Ensure _WEEKLY_DATA exists (for new files)
            self.data_manager._get_or_create_data_sheet(wb)

            # Save
            wb.save(self.tracker_path)

            logger.info(f"Rebuilt tracker report: {self.tracker_path}")
            return True

        except Exception as e:
            logger.error(f"Error rebuilding tracker report: {e}")
            return False

    def update_and_rebuild_from_merge(
        self,
        merge_results: Dict,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> bool:
        """
        Combined update and rebuild operation from merge results.

        Args:
            merge_results: Results from merge_all_corrections()
            progress_callback: Optional progress callback

        Returns:
            True if successful
        """
        try:
            if progress_callback:
                progress_callback(10, "Updating tracker data...")

            written = self.update_from_merge_results(merge_results)

            if progress_callback:
                progress_callback(50, f"Wrote {written} records, rebuilding report...")

            success = self.rebuild_report()

            if progress_callback:
                status = "Tracker updated" if success else "Tracker update failed"
                progress_callback(100, status)

            return success

        except Exception as e:
            logger.error(f"Error in update_and_rebuild_from_merge: {e}")
            if progress_callback:
                progress_callback(100, f"Error: {e}")
            return False

    def get_summary_text(self) -> str:
        """
        Get a text summary of current tracker state.

        Returns:
            Multi-line summary string
        """
        latest = self.data_manager.get_latest_week_data()
        if not latest:
            return "No data in tracker yet."

        lines = ["=== MERGE PROGRESS SUMMARY ===", ""]

        total_corrections = 0
        total_success = 0
        total_fail = 0

        for lang in sorted(latest.keys()):
            lang_data = latest[lang]
            corrections = lang_data.get("Corrections", 0)
            success = lang_data.get("Success", 0)
            fail = lang_data.get("Fail", 0)
            success_rate = (success / corrections * 100) if corrections > 0 else 0

            total_corrections += corrections
            total_success += success
            total_fail += fail

            lines.append(f"{lang}: {success:,}/{corrections:,} ({success_rate:.1f}% success)")

        lines.append("")
        overall_rate = (total_success / total_corrections * 100) if total_corrections > 0 else 0
        lines.append(f"TOTAL: {total_success:,}/{total_corrections:,} ({overall_rate:.1f}% success)")

        return "\n".join(lines)
