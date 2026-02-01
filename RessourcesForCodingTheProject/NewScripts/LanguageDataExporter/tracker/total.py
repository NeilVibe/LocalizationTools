"""
TOTAL sheet builder for Correction Progress Tracker.

Contains summary table showing per-language merge results:
- Language, Corrections, Success, Fail, Success %
- Professional formatting with color-coded success rates, bold totals
"""

import logging
from typing import Dict, List, Set

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

TOTAL_SHEET_NAME = "TOTAL"

# =============================================================================
# STYLING CONSTANTS (Professional Excel Theme)
# =============================================================================

# Colors (Microsoft Office Blue palette)
BLUE_DARK = "1F4E79"
BLUE_HEADER = "4472C4"
BLUE_LIGHT = "D9E2F3"
GREEN_SUCCESS = "C6EFCE"
GREEN_DARK = "63BE7B"
YELLOW_WARN = "FFEB9C"
RED_FAIL = "FFC7CE"
GRAY_ALT = "F2F2F2"

# Title styling
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Header styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Data styling
DATA_FONT = Font(size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Total row styling
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")

# Alternating row fills
ROW_FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_FILL_ALT = PatternFill(start_color=GRAY_ALT, end_color=GRAY_ALT, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='medium', color=BLUE_DARK)
)

TITLE_BORDER = Border(
    left=Side(style='medium', color=BLUE_DARK),
    right=Side(style='medium', color=BLUE_DARK),
    top=Side(style='medium', color=BLUE_DARK),
    bottom=Side(style='medium', color=BLUE_DARK)
)

TOTAL_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='double', color=BLUE_DARK),  # Double line above total
    bottom=Side(style='medium', color=BLUE_DARK)
)


def get_success_rate_fill(rate: float) -> PatternFill:
    """Return fill color based on success rate (0.0-1.0)."""
    rate_pct = rate * 100
    if rate_pct >= 95:
        return PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
    elif rate_pct >= 80:
        return PatternFill(start_color=GREEN_SUCCESS, end_color=GREEN_SUCCESS, fill_type="solid")
    elif rate_pct >= 60:
        return PatternFill(start_color=YELLOW_WARN, end_color=YELLOW_WARN, fill_type="solid")
    else:
        return PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type="solid")


def build_total_sheet(
    wb: Workbook,
    latest_data: Dict[str, Dict],
    categories: List[str] = None
) -> None:
    """
    Build the TOTAL sheet with summary table and per-category breakdown.

    Args:
        wb: Workbook to add sheet to
        latest_data: Dict[language, Dict] from get_latest_week_data()
            Each dict has: Corrections, Success, Fail, MergeDate, by_category
        categories: Not used in new schema (kept for API compatibility)
    """
    # Create or get sheet
    if TOTAL_SHEET_NAME in wb.sheetnames:
        del wb[TOTAL_SHEET_NAME]

    ws = wb.create_sheet(TOTAL_SHEET_NAME, 1)  # Second position

    if not latest_data:
        ws.cell(row=1, column=1, value="No data yet. Run 'Merge to LOCDEV' to collect data.")
        return

    # Get sorted language list
    languages = sorted(latest_data.keys())

    # Collect all categories across all languages
    all_categories: Set[str] = set()
    for lang_data in latest_data.values():
        by_category = lang_data.get("by_category", {})
        all_categories.update(by_category.keys())
    sorted_categories = sorted(all_categories)

    # ==========================================================================
    # TITLE ROW
    # ==========================================================================
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="MERGE SUMMARY - LATEST WEEK")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = TITLE_ALIGNMENT
    title_cell.border = TITLE_BORDER

    # Apply border to merged cells
    for col in range(2, 6):
        ws.cell(row=1, column=col).border = TITLE_BORDER

    # Empty row 2 for spacing
    ws.row_dimensions[2].height = 8

    # ==========================================================================
    # HEADER ROW (row 3)
    # ==========================================================================
    header_row = 3
    headers = ["Language", "Corrections", "Success", "Fail", "Success %"]
    widths = [14, 14, 12, 10, 14]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    # Set column widths
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ==========================================================================
    # DATA ROWS
    # ==========================================================================
    data_row = header_row + 1
    total_corrections = 0
    total_success = 0
    total_fail = 0

    for idx, lang in enumerate(languages):
        lang_data = latest_data.get(lang, {})

        corrections = lang_data.get("Corrections", 0)
        success = lang_data.get("Success", 0)
        fail = lang_data.get("Fail", 0)
        success_rate = (success / corrections) if corrections > 0 else 0

        total_corrections += corrections
        total_success += success
        total_fail += fail

        # Alternating row color
        row_fill = ROW_FILL_ALT if (idx % 2 == 0) else ROW_FILL_WHITE

        # Language (bold)
        cell = ws.cell(row=data_row, column=1, value=lang)
        cell.font = Font(bold=True, size=10)
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER

        # Corrections
        cell = ws.cell(row=data_row, column=2, value=corrections)
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Success
        cell = ws.cell(row=data_row, column=3, value=success)
        cell.font = DATA_FONT
        cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Fail
        cell = ws.cell(row=data_row, column=4, value=fail)
        cell.font = DATA_FONT
        if fail > 0:
            cell.fill = PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type="solid")
        else:
            cell.fill = row_fill
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

        # Success % with color gradient
        cell = ws.cell(row=data_row, column=5, value=success_rate)
        cell.font = Font(bold=True, size=10)
        cell.fill = get_success_rate_fill(success_rate)
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        cell.number_format = '0.0%'

        data_row += 1

    # ==========================================================================
    # TOTAL ROW
    # ==========================================================================
    total_success_rate = (total_success / total_corrections) if total_corrections > 0 else 0

    # Language column - "TOTAL"
    cell = ws.cell(row=data_row, column=1, value="TOTAL")
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.alignment = DATA_ALIGNMENT
    cell.border = TOTAL_BORDER

    # Corrections total
    cell = ws.cell(row=data_row, column=2, value=total_corrections)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.alignment = DATA_ALIGNMENT
    cell.border = TOTAL_BORDER
    cell.number_format = '#,##0'

    # Success total
    cell = ws.cell(row=data_row, column=3, value=total_success)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.alignment = DATA_ALIGNMENT
    cell.border = TOTAL_BORDER
    cell.number_format = '#,##0'

    # Fail total
    cell = ws.cell(row=data_row, column=4, value=total_fail)
    cell.font = TOTAL_FONT
    if total_fail > 0:
        cell.fill = PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type="solid")
    else:
        cell.fill = TOTAL_FILL
    cell.alignment = DATA_ALIGNMENT
    cell.border = TOTAL_BORDER
    cell.number_format = '#,##0'

    # Success % total with color
    cell = ws.cell(row=data_row, column=5, value=total_success_rate)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = get_success_rate_fill(total_success_rate)
    cell.alignment = DATA_ALIGNMENT
    cell.border = TOTAL_BORDER
    cell.number_format = '0.0%'

    data_row += 1

    # ==========================================================================
    # CATEGORY BREAKDOWN SECTION (if categories exist)
    # ==========================================================================
    if sorted_categories:
        # Spacing
        data_row += 2

        # Category section title
        ws.merge_cells(f'A{data_row}:E{data_row}')
        cat_title = ws.cell(row=data_row, column=1, value="BREAKDOWN BY CATEGORY")
        cat_title.font = TITLE_FONT
        cat_title.fill = TITLE_FILL
        cat_title.alignment = TITLE_ALIGNMENT
        cat_title.border = TITLE_BORDER
        for col in range(2, 6):
            ws.cell(row=data_row, column=col).border = TITLE_BORDER
        ws.row_dimensions[data_row].height = 28
        data_row += 1

        # Spacing
        ws.row_dimensions[data_row].height = 8
        data_row += 1

        # Category header row
        cat_headers = ["Category", "Corrections", "Success", "Fail", "Success %"]
        for col, header in enumerate(cat_headers, 1):
            cell = ws.cell(row=data_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER
        ws.row_dimensions[data_row].height = 22
        data_row += 1

        # Aggregate by category across all languages
        cat_totals = {}
        for lang_data in latest_data.values():
            by_category = lang_data.get("by_category", {})
            for cat, cat_stats in by_category.items():
                if cat not in cat_totals:
                    cat_totals[cat] = {"Corrections": 0, "Success": 0, "Fail": 0}
                cat_totals[cat]["Corrections"] += cat_stats.get("Corrections", 0)
                cat_totals[cat]["Success"] += cat_stats.get("Success", 0)
                cat_totals[cat]["Fail"] += cat_stats.get("Fail", 0)

        # Category data rows
        for idx, cat in enumerate(sorted_categories):
            cat_data = cat_totals.get(cat, {})
            corrections = cat_data.get("Corrections", 0)
            success = cat_data.get("Success", 0)
            fail = cat_data.get("Fail", 0)
            success_rate = (success / corrections) if corrections > 0 else 0

            # Alternating row color
            row_fill = ROW_FILL_ALT if (idx % 2 == 0) else ROW_FILL_WHITE

            # Category name
            cell = ws.cell(row=data_row, column=1, value=cat)
            cell.font = Font(bold=True, size=10)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER

            # Corrections
            cell = ws.cell(row=data_row, column=2, value=corrections)
            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            cell.number_format = '#,##0'

            # Success
            cell = ws.cell(row=data_row, column=3, value=success)
            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            cell.number_format = '#,##0'

            # Fail
            cell = ws.cell(row=data_row, column=4, value=fail)
            cell.font = DATA_FONT
            if fail > 0:
                cell.fill = PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type="solid")
            else:
                cell.fill = row_fill
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            cell.number_format = '#,##0'

            # Success % with color gradient
            cell = ws.cell(row=data_row, column=5, value=success_rate)
            cell.font = Font(bold=True, size=10)
            cell.fill = get_success_rate_fill(success_rate)
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            cell.number_format = '0.0%'

            ws.row_dimensions[data_row].height = 20
            data_row += 1

    # ==========================================================================
    # FINAL TOUCHES
    # ==========================================================================

    # Freeze header row
    ws.freeze_panes = 'A4'

    # Set row heights for main section
    ws.row_dimensions[1].height = 28  # Title row
    ws.row_dimensions[3].height = 22  # Header row

    logger.info(f"Built TOTAL sheet with {len(languages)} languages, {len(sorted_categories)} categories")
