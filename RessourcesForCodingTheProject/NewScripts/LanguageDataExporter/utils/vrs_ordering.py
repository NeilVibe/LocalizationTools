"""
VoiceRecordingSheet Ordering Module.

Provides chronological story ordering for STORY strings based on
the VoiceRecordingSheet EventName column.

The VoiceRecordingSheet is the master list of all voiced lines in the game,
ordered by their appearance in the story. We use this to sort STORY category
strings (Sequencer, AIDialog, QuestDialog, NarrationDialog) so LQA reviewers
see them in the order players experience them.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# VoiceRecordingSheet column indices (0-based)
COL_TYPE = 0       # Column A: Type (Sequencer, AIDialog, etc.)
COL_EVENTNAME = 22  # Column W: EventName (SoundEventName in XML)


def find_most_recent_excel(folder: Path) -> Optional[Path]:
    """
    Find the most recently modified Excel file in a folder.

    Args:
        folder: Path to folder containing Excel files

    Returns:
        Path to most recent .xlsx file, or None if not found
    """
    if not folder.exists():
        logger.warning(f"VoiceRecordingSheet folder not found: {folder}")
        return None

    files = list(folder.glob("*.xlsx"))
    if not files:
        logger.warning(f"No Excel files found in {folder}")
        return None

    most_recent = max(files, key=lambda f: f.stat().st_mtime)
    logger.info(f"Using VoiceRecordingSheet: {most_recent.name}")
    return most_recent


def load_vrs_order(vrs_folder: Path) -> Tuple[Dict[str, int], int]:
    """
    Load EventName ordering from VoiceRecordingSheet.

    Args:
        vrs_folder: Path to VoiceRecordingSheet folder

    Returns:
        Tuple of:
        - Dict mapping EventName (lowercase) to row position (0-based)
        - Total number of events loaded
    """
    vrs_file = find_most_recent_excel(vrs_folder)
    if not vrs_file:
        return {}, 0

    try:
        wb = load_workbook(vrs_file, data_only=True, read_only=True)
        ws = wb.active

        event_order: Dict[str, int] = {}
        position = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Get EventName from column W (index 22)
            if len(row) > COL_EVENTNAME:
                event_name = row[COL_EVENTNAME]
                if event_name:
                    event_name_lower = str(event_name).strip().lower()
                    if event_name_lower and event_name_lower not in event_order:
                        event_order[event_name_lower] = position
                        position += 1

        wb.close()
        logger.info(f"Loaded {len(event_order)} EventNames from VoiceRecordingSheet")
        return event_order, len(event_order)

    except Exception as e:
        logger.error(f"Error loading VoiceRecordingSheet: {e}")
        return {}, 0


def load_vrs_with_categories(vrs_folder: Path) -> Dict[str, Tuple[int, str]]:
    """
    Load EventName ordering with category info from VoiceRecordingSheet.

    Args:
        vrs_folder: Path to VoiceRecordingSheet folder

    Returns:
        Dict mapping EventName (lowercase) to (position, category)
        where category is from Column A (Sequencer, AIDialog, etc.)
    """
    vrs_file = find_most_recent_excel(vrs_folder)
    if not vrs_file:
        return {}

    try:
        wb = load_workbook(vrs_file, data_only=True, read_only=True)
        ws = wb.active

        event_data: Dict[str, Tuple[int, str]] = {}
        position = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > COL_EVENTNAME:
                event_name = row[COL_EVENTNAME]
                category = row[COL_TYPE] if row[COL_TYPE] else ""

                if event_name:
                    event_name_lower = str(event_name).strip().lower()
                    category_str = str(category).strip()

                    if event_name_lower and event_name_lower not in event_data:
                        event_data[event_name_lower] = (position, category_str)
                        position += 1

        wb.close()
        logger.info(f"Loaded {len(event_data)} EventNames with categories")
        return event_data

    except Exception as e:
        logger.error(f"Error loading VoiceRecordingSheet: {e}")
        return {}


class VRSOrderer:
    """
    Orders strings based on VoiceRecordingSheet EventName sequence.

    Usage:
        orderer = VRSOrderer(vrs_folder)
        orderer.load()
        sorted_entries = orderer.sort_entries(entries, stringid_to_soundevent)
    """

    def __init__(self, vrs_folder: Path):
        """
        Initialize VRS orderer.

        Args:
            vrs_folder: Path to VoiceRecordingSheet folder
        """
        self.vrs_folder = vrs_folder
        self.event_order: Dict[str, int] = {}
        self.total_events = 0
        self.loaded = False

    def load(self) -> bool:
        """
        Load VoiceRecordingSheet ordering.

        Returns:
            True if loaded successfully
        """
        self.event_order, self.total_events = load_vrs_order(self.vrs_folder)
        self.loaded = self.total_events > 0
        return self.loaded

    def get_position(self, event_name: str) -> int:
        """
        Get position for an EventName.

        Args:
            event_name: The SoundEventName/EventName to look up

        Returns:
            Position (0-based), or sys.maxsize if not found
        """
        if not event_name:
            return float('inf')
        return self.event_order.get(event_name.lower().strip(), float('inf'))

    def sort_entries(
        self,
        entries: List[Dict],
        stringid_to_soundevent: Dict[str, str]
    ) -> List[Dict]:
        """
        Sort entries by VoiceRecordingSheet order.

        Args:
            entries: List of {"string_id", "str_origin", "str"} dicts
            stringid_to_soundevent: Mapping of StringId to SoundEventName

        Returns:
            Sorted list of entries
        """
        def sort_key(entry):
            string_id = entry.get("string_id", "")
            sound_event = stringid_to_soundevent.get(string_id, "")
            return self.get_position(sound_event)

        return sorted(entries, key=sort_key)

    def sort_by_soundevent(
        self,
        items: List[Tuple[str, any]],
    ) -> List[Tuple[str, any]]:
        """
        Sort list of (SoundEventName, data) tuples by VRS order.

        Args:
            items: List of (event_name, data) tuples

        Returns:
            Sorted list
        """
        return sorted(items, key=lambda x: self.get_position(x[0]))
