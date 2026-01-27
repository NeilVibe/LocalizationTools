"""
Submit Preparer for LQA submission.

Transforms Excel files in ToSubmit folder for final submission:
- If Correction column has value, replace Str with Correction
- Output only: StrOrigin, Str, StringID
- Creates backup before overwriting
"""

import logging
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Callable

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import TYPE_STRING

logger = logging.getLogger(__name__)

# Output columns for submission (final structure)
OUTPUT_COLUMNS = ["StrOrigin", "Str", "StringID"]

# Excel styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def discover_submit_files(submit_folder: Path) -> List[Tuple[Path, str]]:
    """
    Find languagedata_*.xlsx files in ToSubmit folder.

    Args:
        submit_folder: Path to ToSubmit folder

    Returns:
        List of (file_path, language_code) tuples
    """
    if not submit_folder.exists():
        logger.warning(f"ToSubmit folder not found: {submit_folder}")
        return []

    files = []
    for xlsx_file in submit_folder.glob("languagedata_*.xlsx"):
        # Extract language code from filename (e.g., "languagedata_FRE.xlsx" -> "FRE")
        name = xlsx_file.stem  # "languagedata_FRE"
        if name.startswith("languagedata_"):
            lang_code = name.replace("languagedata_", "")
            files.append((xlsx_file, lang_code))

    # Also support LanguageData_*.xlsx pattern (case variations)
    for xlsx_file in submit_folder.glob("LanguageData_*.xlsx"):
        name = xlsx_file.stem
        if name.startswith("LanguageData_"):
            lang_code = name.replace("LanguageData_", "")
            # Avoid duplicates
            if not any(f[1].upper() == lang_code.upper() for f in files):
                files.append((xlsx_file, lang_code))

    logger.info(f"Discovered {len(files)} files in ToSubmit folder")
    return sorted(files, key=lambda x: x[1])


def create_backup(submit_folder: Path, files: List[Path]) -> Optional[Path]:
    """
    Create timestamped backup before overwriting.

    Args:
        submit_folder: Path to ToSubmit folder
        files: List of file paths to backup

    Returns:
        Path to backup folder, or None if no files to backup
    """
    if not files:
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_folder = submit_folder / f"backup_{timestamp}"
    backup_folder.mkdir(parents=True, exist_ok=True)

    for file_path in files:
        if file_path.exists():
            dest = backup_folder / file_path.name
            shutil.copy2(file_path, dest)
            logger.info(f"Backed up: {file_path.name}")

    logger.info(f"Backup created: {backup_folder}")
    return backup_folder


def _detect_column_indices(ws) -> Dict[str, int]:
    """
    Detect column indices from header row.

    Returns dict mapping column name to 1-based index.
    Handles both EU and Asian column structures.
    Works with both normal and read-only worksheets.
    """
    indices = {}
    # Handle read-only worksheets where max_column may be None
    # Iterate over first row directly instead of using max_column
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if first_row and first_row[0]:
        for col, header in enumerate(first_row[0], 1):
            if header:
                indices[str(header)] = col
    return indices


def prepare_file_for_submit(input_path: Path, output_path: Path) -> Dict:
    """
    Transform single file - apply corrections, extract columns.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output Excel file (can be same for overwrite)

    Returns:
        Dict with statistics: {
            "total_rows": int,
            "corrections_applied": int,
            "error": str or None
        }
    """
    result = {
        "total_rows": 0,
        "corrections_applied": 0,
        "error": None
    }

    try:
        # Load input workbook
        wb_in = load_workbook(input_path)
        ws_in = wb_in.active

        # Detect column indices
        col_indices = _detect_column_indices(ws_in)

        # Required columns
        str_origin_col = col_indices.get("StrOrigin")
        str_col = col_indices.get("Str")
        correction_col = col_indices.get("Correction")
        stringid_col = col_indices.get("StringID")

        if not all([str_origin_col, str_col, stringid_col]):
            result["error"] = f"Missing required columns. Found: {list(col_indices.keys())}"
            return result

        # Create output workbook
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = ws_in.title

        # Write headers
        for col, header in enumerate(OUTPUT_COLUMNS, 1):
            cell = ws_out.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Process data rows
        out_row = 2
        for in_row in range(2, ws_in.max_row + 1):
            str_origin = ws_in.cell(row=in_row, column=str_origin_col).value
            str_value = ws_in.cell(row=in_row, column=str_col).value
            correction = ws_in.cell(row=in_row, column=correction_col).value if correction_col else None
            string_id = ws_in.cell(row=in_row, column=stringid_col).value

            # Skip empty rows
            if not string_id:
                continue

            result["total_rows"] += 1

            # Apply correction if present
            final_str = str_value
            if correction and str(correction).strip():
                final_str = correction
                result["corrections_applied"] += 1

            # Write output row: StrOrigin, Str, StringID
            ws_out.cell(row=out_row, column=1, value=str_origin).alignment = CELL_ALIGNMENT
            ws_out.cell(row=out_row, column=2, value=final_str).alignment = CELL_ALIGNMENT

            # Force StringID to TEXT format (prevent scientific notation like 1.23E+12)
            stringid_cell = ws_out.cell(row=out_row, column=3, value=str(string_id) if string_id else string_id)
            stringid_cell.alignment = CELL_ALIGNMENT
            stringid_cell.data_type = TYPE_STRING

            for col in range(1, 4):
                ws_out.cell(row=out_row, column=col).border = THIN_BORDER

            out_row += 1

        # Set column widths
        ws_out.column_dimensions['A'].width = 45  # StrOrigin
        ws_out.column_dimensions['B'].width = 45  # Str
        ws_out.column_dimensions['C'].width = 15  # StringID

        # Freeze header row
        ws_out.freeze_panes = 'A2'

        # Add auto-filter
        ws_out.auto_filter.ref = ws_out.dimensions

        # Save output
        wb_out.save(output_path)

        logger.info(f"Prepared {output_path.name}: {result['total_rows']} rows, "
                   f"{result['corrections_applied']} corrections applied")

    except Exception as e:
        result["error"] = str(e)
        logger.error(f"Error preparing {input_path.name}: {e}")

    return result


def prepare_all_for_submit(
    submit_folder: Path,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> Dict:
    """
    Process all files in ToSubmit folder.

    Args:
        submit_folder: Path to ToSubmit folder
        progress_callback: Optional callback(progress_percent, status_message)

    Returns:
        Dict with overall results: {
            "files_processed": int,
            "total_corrections": int,
            "backup_folder": Path or None,
            "errors": List[str],
            "file_results": Dict[str, Dict]  # per-file stats
        }
    """
    results = {
        "files_processed": 0,
        "total_corrections": 0,
        "backup_folder": None,
        "errors": [],
        "file_results": {}
    }

    # Discover files
    files = discover_submit_files(submit_folder)
    if not files:
        results["errors"].append("No languagedata_*.xlsx files found in ToSubmit folder")
        return results

    if progress_callback:
        progress_callback(5, f"Found {len(files)} files to process")

    # Create backup
    file_paths = [f[0] for f in files]
    results["backup_folder"] = create_backup(submit_folder, file_paths)

    if progress_callback:
        progress_callback(10, "Backup created")

    # Process each file
    total_files = len(files)
    for idx, (file_path, lang_code) in enumerate(files):
        if progress_callback:
            progress = 10 + int((idx / total_files) * 85)
            progress_callback(progress, f"Processing {lang_code}...")

        # Process file (overwrite in place)
        file_result = prepare_file_for_submit(file_path, file_path)
        results["file_results"][lang_code] = file_result

        if file_result["error"]:
            results["errors"].append(f"{lang_code}: {file_result['error']}")
        else:
            results["files_processed"] += 1
            results["total_corrections"] += file_result["corrections_applied"]

    if progress_callback:
        progress_callback(100, f"Completed: {results['files_processed']} files processed")

    logger.info(f"Submit preparation complete: {results['files_processed']} files, "
               f"{results['total_corrections']} total corrections")

    return results


# Pre-compiled Korean pattern for performance
KOREAN_PATTERN = re.compile(r'[\uAC00-\uD7AF]+')


def collect_correction_stats(submit_folder: Path) -> Dict:
    """
    Collect correction statistics from files in ToSubmit folder.

    Used by tracker to gather data BEFORE preparation (when Correction column exists).

    Returns:
        Dict[lang_code, Dict[category, Dict]] with:
            - corrected: count of rows with Correction value
            - pending: count of rows without Correction value
            - kr_words: count of Korean words in corrected+pending
    """
    stats = {}

    files = discover_submit_files(submit_folder)

    for file_path, lang_code in files:
        try:
            # Don't use read_only=True - we need full worksheet access
            wb = load_workbook(file_path)
            ws = wb.active

            col_indices = _detect_column_indices(ws)
            str_col = col_indices.get("Str")
            correction_col = col_indices.get("Correction")
            category_col = col_indices.get("Category")
            str_origin_col = col_indices.get("StrOrigin")

            if not all([str_col, category_col, str_origin_col]):
                logger.warning(f"Skipping {lang_code}: missing required columns. Found: {list(col_indices.keys())}")
                wb.close()
                continue

            lang_stats = {}

            for row in range(2, ws.max_row + 1):
                category = ws.cell(row=row, column=category_col).value or "Uncategorized"
                str_origin = ws.cell(row=row, column=str_origin_col).value or ""
                correction = ws.cell(row=row, column=correction_col).value if correction_col else None

                if category not in lang_stats:
                    lang_stats[category] = {
                        "corrected": 0,
                        "pending": 0,
                        "kr_words": 0
                    }

                # Count Korean words in StrOrigin (source language)
                korean_matches = KOREAN_PATTERN.findall(str(str_origin))
                kr_word_count = len(korean_matches) if korean_matches else 0

                lang_stats[category]["kr_words"] += kr_word_count

                if correction and str(correction).strip():
                    lang_stats[category]["corrected"] += 1
                else:
                    lang_stats[category]["pending"] += 1

            stats[lang_code] = lang_stats
            wb.close()

        except Exception as e:
            logger.error(f"Error collecting stats for {lang_code}: {e}")

    return stats
