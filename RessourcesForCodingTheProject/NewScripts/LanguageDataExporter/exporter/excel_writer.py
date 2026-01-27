"""
Excel Writer for categorized language data.

Generates Excel files with openpyxl, styled with headers and proper column widths.
STORY category strings are ordered by VoiceRecordingSheet EventName for
chronological story order.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import TYPE_STRING
from openpyxl.worksheet.protection import SheetProtection

if TYPE_CHECKING:
    from utils.vrs_ordering import VRSOrderer

logger = logging.getLogger(__name__)

# Import STORY_CATEGORIES from config
try:
    from config import STORY_CATEGORIES
except ImportError:
    STORY_CATEGORIES = ["Sequencer", "AIDialog", "QuestDialog", "NarrationDialog"]

# Excel styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths
DEFAULT_WIDTHS = {
    "StrOrigin": 45,
    "ENG": 45,
    "Str": 45,
    "Correction": 45,    # NEW column for LQA corrections
    "Category": 20,
    "StringID": 15,      # Moved to end
}


def _apply_header_style(ws, row: int = 1):
    """Apply styling to header row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _set_column_widths(ws, headers: List[str]):
    """Set column widths based on header names."""
    for idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(idx)
        width = DEFAULT_WIDTHS.get(header, 20)
        ws.column_dimensions[col_letter].width = width


def _sort_entries_for_output(
    entries: List[Dict],
    category_index: Dict[str, str],
    default_category: str,
    vrs_orderer: Optional["VRSOrderer"] = None,
    stringid_to_soundevent: Optional[Dict[str, str]] = None,
) -> List[Dict]:
    """
    Sort entries for Excel output.

    - STORY categories (Sequencer, AIDialog, QuestDialog, NarrationDialog)
      are sorted by VoiceRecordingSheet EventName for chronological story order
    - Other categories are grouped together alphabetically

    Args:
        entries: Raw language entries
        category_index: {StringID: Category} mapping
        default_category: Fallback category
        vrs_orderer: VRSOrderer for story ordering
        stringid_to_soundevent: {StringID: SoundEventName} mapping

    Returns:
        Sorted list of entries
    """
    if not vrs_orderer or not vrs_orderer.loaded or not stringid_to_soundevent:
        # No VRS ordering available - just sort by category name
        return sorted(entries, key=lambda e: category_index.get(e.get("string_id", ""), default_category))

    # Separate STORY and non-STORY entries
    story_entries = []
    other_entries = []

    for entry in entries:
        string_id = entry.get("string_id", "")
        category = category_index.get(string_id, default_category)

        if category in STORY_CATEGORIES:
            story_entries.append(entry)
        else:
            other_entries.append(entry)

    # Sort STORY entries by VRS order (chronological story order)
    def story_sort_key(entry):
        string_id = entry.get("string_id", "")
        category = category_index.get(string_id, default_category)
        sound_event = stringid_to_soundevent.get(string_id, "")
        vrs_position = vrs_orderer.get_position(sound_event)
        # Sort by: category order, then VRS position
        cat_order = STORY_CATEGORIES.index(category) if category in STORY_CATEGORIES else 999
        return (cat_order, vrs_position)

    sorted_story = sorted(story_entries, key=story_sort_key)

    # Sort other entries by category name
    sorted_other = sorted(other_entries, key=lambda e: category_index.get(e.get("string_id", ""), default_category))

    # STORY entries first, then other categories
    return sorted_story + sorted_other


def write_language_excel(
    lang_code: str,
    lang_data: List[Dict],
    eng_lookup: Dict[str, str],
    category_index: Dict[str, str],
    output_path: Path,
    include_english: bool = True,
    default_category: str = "Uncategorized",
    vrs_orderer: Optional["VRSOrderer"] = None,
    stringid_to_soundevent: Optional[Dict[str, str]] = None,
    excluded_categories: Optional[set] = None,
    protect_sheet: bool = True,
) -> bool:
    """
    Write Excel file for one language.

    STORY category strings (Sequencer, AIDialog, QuestDialog, NarrationDialog)
    are sorted by VoiceRecordingSheet EventName for chronological story order.

    Args:
        lang_code: Language code (e.g., "eng", "fre")
        lang_data: List of dicts with str_origin, str, string_id
        eng_lookup: {StringID: English translation} for cross-reference
        category_index: {StringID: Category} mapping
        output_path: Path to output Excel file
        include_english: Whether to include English column
        default_category: Category for unmapped StringIDs
        vrs_orderer: VRSOrderer for story ordering (optional)
        stringid_to_soundevent: {StringID: SoundEventName} mapping (optional)
        excluded_categories: Set of category names to exclude (optional)
        protect_sheet: If True, lock all columns except Correction (default: True)

    Returns:
        True if successful, False otherwise

    Column structure:
    - European: StrOrigin | ENG | Str | Correction | Category | StringID
    - Asian:    StrOrigin | Str | Correction | Category | StringID

    Sheet Protection:
    - When protect_sheet=True, only the Correction column is editable
    - All other columns are locked/read-only
    - QA testers can only modify the Correction column
    """
    try:
        # Filter out excluded categories BEFORE processing
        if excluded_categories:
            original_count = len(lang_data)
            lang_data = [
                entry for entry in lang_data
                if category_index.get(entry.get('string_id', ''), default_category) not in excluded_categories
            ]
            filtered_count = original_count - len(lang_data)
            if filtered_count > 0:
                logger.info(f"Filtered out {filtered_count} entries from excluded categories: {excluded_categories}")
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = lang_code.upper()

        # Define headers based on language type
        # New column order: Correction column added, StringID moved to end
        if include_english:
            headers = ["StrOrigin", "ENG", "Str", "Correction", "Category", "StringID"]
        else:
            headers = ["StrOrigin", "Str", "Correction", "Category", "StringID"]

        # Write header row
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Apply header styling
        _apply_header_style(ws)
        _set_column_widths(ws, headers)

        # Sort data: STORY categories by VRS order, others alphabetically by category
        sorted_data = _sort_entries_for_output(
            lang_data,
            category_index,
            default_category,
            vrs_orderer,
            stringid_to_soundevent,
        )

        # Write data rows
        row_num = 2
        for entry in sorted_data:
            string_id = entry.get('string_id', '')
            str_origin = entry.get('str_origin', '')
            str_value = entry.get('str', '')

            # Get category
            category = category_index.get(string_id, default_category)

            # Get English translation if needed
            english = eng_lookup.get(string_id, '') if include_english else None

            # Build row data (Correction column empty, to be filled during LQA)
            if include_english:
                row_data = [str_origin, english, str_value, "", category, string_id]
            else:
                row_data = [str_origin, str_value, "", category, string_id]

            # Write cells
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

                # Force StringID to TEXT format (prevent scientific notation like 1.23E+12)
                if headers[col - 1] == "StringID" and value is not None:
                    cell.value = str(value)
                    cell.data_type = TYPE_STRING

            row_num += 1

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        # =====================================================================
        # SHEET PROTECTION - Only Correction column is editable
        # =====================================================================
        # How Excel protection works:
        # 1. Each cell has a "locked" property (default: True)
        # 2. When sheet protection is ENABLED, locked cells become read-only
        # 3. Cells with locked=False remain editable even when sheet is protected
        #
        # Our approach:
        # 1. Lock ALL cells first (headers + all data)
        # 2. Explicitly unlock ONLY the Correction column data cells
        # 3. Enable sheet protection to enforce the locking
        # =====================================================================
        if protect_sheet:
            # Find Correction column index (1-based)
            correction_col_idx = headers.index("Correction") + 1 if "Correction" in headers else None

            if correction_col_idx is None:
                logger.warning("protect_sheet=True but no Correction column found - entire sheet will be read-only")

            # STEP 1: Lock ALL cells (headers and data)
            # row_num has been incremented past the last data row, so max_row is row_num - 1
            for row in ws.iter_rows(min_row=1, max_row=row_num - 1):
                for cell in row:
                    cell.protection = Protection(locked=True, hidden=False)

            # STEP 2: Unlock ONLY the Correction column (data rows, not header)
            if correction_col_idx:
                for row in range(2, row_num):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row, column=correction_col_idx)
                    cell.protection = Protection(locked=False, hidden=False)

            # STEP 3: Enable sheet protection
            # In SheetProtection: True = ALLOW the action, False = DISALLOW
            ws.protection = SheetProtection(
                sheet=True,              # Enable protection
                objects=False,           # Disallow: edit objects
                scenarios=False,         # Disallow: edit scenarios
                formatCells=True,        # Allow: format cells
                formatColumns=True,      # Allow: format columns
                formatRows=True,         # Allow: format rows
                insertColumns=False,     # Disallow: insert columns
                insertRows=False,        # Disallow: insert rows
                insertHyperlinks=False,  # Disallow: insert hyperlinks
                deleteColumns=False,     # Disallow: delete columns
                deleteRows=False,        # Disallow: delete rows
                selectLockedCells=True,  # Allow: select locked cells
                sort=True,               # Allow: sort
                autoFilter=True,         # Allow: use auto-filter
                pivotTables=False,       # Disallow: pivot tables
                selectUnlockedCells=True,  # Allow: select unlocked cells (Correction)
            )
            logger.info(f"Sheet protection enabled - only Correction column (col {correction_col_idx}) is editable")

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(output_path)

        logger.info(f"Generated {output_path.name} with {row_num - 2} rows")
        return True

    except Exception as e:
        logger.error(f"Error writing Excel for {lang_code}: {e}")
        return False


def write_summary_excel(
    language_stats: Dict[str, Dict],
    category_stats: Dict[str, int],
    output_path: Path
) -> bool:
    """
    Write a summary Excel file with statistics.

    Args:
        language_stats: {lang_code: {"rows": N, "file": "path"}}
        category_stats: {category: count}
        output_path: Path to output Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        wb = Workbook()

        # Languages sheet
        ws_lang = wb.active
        ws_lang.title = "Languages"

        headers = ["Language", "Rows", "File"]
        for col, header in enumerate(headers, 1):
            ws_lang.cell(row=1, column=col, value=header)
        _apply_header_style(ws_lang)

        row = 2
        for lang, stats in sorted(language_stats.items()):
            ws_lang.cell(row=row, column=1, value=lang.upper())
            ws_lang.cell(row=row, column=2, value=stats.get("rows", 0))
            ws_lang.cell(row=row, column=3, value=stats.get("file", ""))
            row += 1

        # Categories sheet
        ws_cat = wb.create_sheet("Categories")

        headers = ["Category", "StringIDs"]
        for col, header in enumerate(headers, 1):
            ws_cat.cell(row=1, column=col, value=header)
        _apply_header_style(ws_cat)

        row = 2
        for category, count in sorted(category_stats.items(), key=lambda x: -x[1]):
            ws_cat.cell(row=row, column=1, value=category)
            ws_cat.cell(row=row, column=2, value=count)
            row += 1

        # Set column widths
        ws_lang.column_dimensions['A'].width = 15
        ws_lang.column_dimensions['B'].width = 10
        ws_lang.column_dimensions['C'].width = 40

        ws_cat.column_dimensions['A'].width = 25
        ws_cat.column_dimensions['B'].width = 15

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Generated summary: {output_path.name}")
        return True

    except Exception as e:
        logger.error(f"Error writing summary Excel: {e}")
        return False
