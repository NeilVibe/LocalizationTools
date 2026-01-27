"""
Code Pattern Analyzer for LanguageDataExporter.

Extracts {code} patterns from languagedata XMLs, clusters by similarity,
and generates a report showing TOP 3 categories per cluster.
"""

import re
import logging
from pathlib import Path
from typing import Dict, List, Set, Tuple
from collections import Counter, defaultdict
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Category colors (matching config.py style)
CATEGORY_COLORS = {
    "Sequencer": "FFE599",
    "AIDialog": "C6EFCE",
    "QuestDialog": "C6EFCE",
    "NarrationDialog": "C6EFCE",
    "Item": "D9D2E9",
    "Quest": "D9D2E9",
    "Character": "F8CBAD",
    "Gimmick": "D9D2E9",
    "Skill": "D9D2E9",
    "Knowledge": "D9D2E9",
    "Faction": "D9D2E9",
    "UI": "A9D08E",
    "Region": "F8CBAD",
    "System_Misc": "D9D9D9",
    "Uncategorized": "DDD9C4",
}

# Similarity threshold for clustering
DEFAULT_THRESHOLD = 0.8


def extract_code_patterns(text: str) -> Set[str]:
    """
    Extract all {code} patterns from text.

    Args:
        text: Input text to search

    Returns:
        Set of patterns found (e.g., {'{name}', '{lvl}'})
    """
    if not text:
        return set()
    return set(re.findall(r'\{[^{}]+\}', text))


def similarity(a: str, b: str) -> float:
    """
    Return similarity ratio between two strings.

    Uses difflib.SequenceMatcher for comparison.

    Args:
        a: First string
        b: Second string

    Returns:
        Similarity ratio from 0.0 to 1.0
    """
    return SequenceMatcher(None, a, b).ratio()


def cluster_patterns(patterns: List[str], threshold: float = DEFAULT_THRESHOLD) -> List[List[str]]:
    """
    Cluster patterns by similarity using greedy algorithm.

    For each pattern, find existing cluster where first pattern is >= threshold similar.
    If none found, create new cluster.

    Args:
        patterns: List of unique patterns
        threshold: Similarity threshold (0.0 to 1.0)

    Returns:
        List of clusters, each cluster is a list of similar patterns
    """
    clusters = []
    for pattern in sorted(patterns):  # Sort for deterministic results
        added = False
        for cluster in clusters:
            if similarity(pattern, cluster[0]) >= threshold:
                cluster.append(pattern)
                added = True
                break
        if not added:
            clusters.append([pattern])
    return clusters


def scan_language_file(
    xml_path: Path,
    category_index: Dict[str, str],
    default_category: str = "Uncategorized",
) -> Dict[str, List[Tuple[str, str]]]:
    """
    Scan a languagedata XML and extract pattern -> [(StringID, Category)] mapping.

    Args:
        xml_path: Path to languagedata_*.xml
        category_index: {StringID: Category} mapping
        default_category: Fallback category

    Returns:
        {pattern: [(string_id, category), ...]}
    """
    from .xml_parser import parse_language_file

    pattern_occurrences = defaultdict(list)

    entries = parse_language_file(xml_path)
    for entry in entries:
        string_id = entry.get('string_id', '')
        str_origin = entry.get('str_origin', '')
        str_value = entry.get('str', '')

        category = category_index.get(string_id, default_category)

        # Extract patterns from both StrOrigin and Str
        patterns = extract_code_patterns(str_origin) | extract_code_patterns(str_value)

        for pattern in patterns:
            pattern_occurrences[pattern].append((string_id, category))

    return dict(pattern_occurrences)


def analyze_patterns(
    loc_folder: Path,
    category_index: Dict[str, str],
    threshold: float = DEFAULT_THRESHOLD,
    default_category: str = "Uncategorized",
) -> Dict:
    """
    Full pattern analysis: scan XMLs, cluster, calculate top categories.

    Args:
        loc_folder: Path to LOC folder with languagedata_*.xml files
        category_index: {StringID: Category} mapping
        threshold: Similarity threshold for clustering (0.0 to 1.0)
        default_category: Fallback category for unmapped StringIDs

    Returns:
        {
            "clusters": [
                {
                    "representative": "{ItemName}",
                    "patterns": ["{ItemName}", "{itemname}", ...],
                    "total_count": 15234,
                    "top_categories": [("Item", 72.3), ("Quest", 15.1), ("UI", 8.2)],
                    "pattern_details": {"{ItemName}": {"Item": 100, "Quest": 50, ...}, ...}
                },
                ...
            ],
            "total_patterns": 1234,
            "total_clusters": 56,
        }
    """
    # Step 1: Scan all language files and collect pattern occurrences
    all_occurrences = defaultdict(list)

    # Scan ENG for efficiency (or first available if no ENG)
    lang_files = list(loc_folder.glob("languagedata_eng.xml"))
    if not lang_files:
        lang_files = list(loc_folder.glob("languagedata_*.xml"))[:1]

    if not lang_files:
        logger.error(f"No languagedata XML files found in {loc_folder}")
        return {"clusters": [], "total_patterns": 0, "total_clusters": 0}

    logger.info(f"Scanning {len(lang_files)} language file(s) for patterns...")

    for xml_path in lang_files:
        occurrences = scan_language_file(xml_path, category_index, default_category)
        for pattern, occ_list in occurrences.items():
            all_occurrences[pattern].extend(occ_list)

    logger.info(f"Found {len(all_occurrences)} unique patterns")

    # Step 2: Cluster patterns by similarity
    unique_patterns = list(all_occurrences.keys())
    clusters_raw = cluster_patterns(unique_patterns, threshold)

    logger.info(f"Clustered into {len(clusters_raw)} groups (threshold={threshold})")

    # Step 3: Build cluster results with top categories
    clusters = []
    for cluster_patterns_list in clusters_raw:
        # Aggregate all occurrences for this cluster
        category_counter = Counter()
        pattern_details = {}

        for pattern in cluster_patterns_list:
            pattern_category_count = Counter()
            for string_id, category in all_occurrences[pattern]:
                category_counter[category] += 1
                pattern_category_count[category] += 1
            pattern_details[pattern] = dict(pattern_category_count)

        total_count = sum(category_counter.values())

        # Get top 3 categories with percentages
        top_categories = []
        for cat, count in category_counter.most_common(3):
            pct = (count / total_count * 100) if total_count > 0 else 0
            top_categories.append((cat, round(pct, 1)))

        clusters.append({
            "representative": cluster_patterns_list[0],
            "patterns": cluster_patterns_list,
            "total_count": total_count,
            "top_categories": top_categories,
            "pattern_details": pattern_details,
        })

    # Sort clusters by total count (descending)
    clusters.sort(key=lambda c: c["total_count"], reverse=True)

    return {
        "clusters": clusters,
        "total_patterns": len(unique_patterns),
        "total_clusters": len(clusters),
    }


def generate_pattern_report(
    analysis_result: Dict,
    output_path: Path,
) -> bool:
    """
    Generate Excel report from analysis results.

    Creates two sheets:
    1. Cluster Summary - overview with top 3 categories per cluster
    2. Cluster Details - all patterns with per-category counts

    Args:
        analysis_result: Result from analyze_patterns()
        output_path: Path to output Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        wb = Workbook()

        # === STYLES ===
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_bottom = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='medium')
        )

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Alternating row colors
        alt_fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        alt_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Number alignment
        num_align = Alignment(horizontal='center')
        text_align = Alignment(horizontal='left', wrap_text=True)

        # =======================================================================
        # Sheet 1: Cluster Summary
        # =======================================================================
        ws_summary = wb.active
        ws_summary.title = "Cluster Summary"

        headers = ["#", "Representative Pattern", "Variants", "Total Uses",
                   "Top 1 Category", "Top 2 Category", "Top 3 Category"]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thick_bottom
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, cluster in enumerate(analysis_result["clusters"], 2):
            is_even = (row_idx % 2 == 0)
            row_fill = alt_fill_light if is_even else alt_fill_white

            # Cluster number
            cell = ws_summary.cell(row=row_idx, column=1, value=row_idx - 1)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = num_align
            cell.font = Font(bold=True)

            # Representative pattern
            cell = ws_summary.cell(row=row_idx, column=2, value=cluster["representative"])
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = text_align
            cell.font = Font(name='Consolas', size=10)

            # Variants count
            cell = ws_summary.cell(row=row_idx, column=3, value=len(cluster["patterns"]))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = num_align

            # Total uses
            cell = ws_summary.cell(row=row_idx, column=4, value=cluster["total_count"])
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = num_align
            cell.number_format = '#,##0'

            # Top 3 categories with colors
            for i, (cat, pct) in enumerate(cluster["top_categories"]):
                cell = ws_summary.cell(row=row_idx, column=5 + i, value=f"{cat} ({pct}%)")
                cat_color = CATEGORY_COLORS.get(cat, "FFFFFF")
                cell.fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True if pct > 50 else False)

            # Fill empty category cells
            for i in range(len(cluster["top_categories"]), 3):
                cell = ws_summary.cell(row=row_idx, column=5 + i, value="")
                cell.fill = row_fill
                cell.border = thin_border

        # Column widths
        ws_summary.column_dimensions['A'].width = 6
        ws_summary.column_dimensions['B'].width = 45
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 22
        ws_summary.column_dimensions['F'].width = 22
        ws_summary.column_dimensions['G'].width = 22

        # Freeze header row
        ws_summary.freeze_panes = 'A2'

        # =======================================================================
        # Sheet 2: Cluster Details
        # =======================================================================
        ws_details = wb.create_sheet("Cluster Details")

        detail_headers = ["Cluster", "Pattern", "Category", "Count"]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thick_bottom
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row = 2
        prev_cluster = None
        for cluster_idx, cluster in enumerate(analysis_result["clusters"], 1):
            cluster_start_row = row

            for pattern, cat_counts in cluster["pattern_details"].items():
                for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
                    # Cluster number (only show on first row of cluster)
                    cell = ws_details.cell(row=row, column=1, value=cluster_idx if row == cluster_start_row else "")
                    cell.border = thin_border
                    cell.alignment = num_align
                    if row == cluster_start_row:
                        cell.font = Font(bold=True, size=11)

                    # Pattern
                    cell = ws_details.cell(row=row, column=2, value=pattern)
                    cell.border = thin_border
                    cell.font = Font(name='Consolas', size=10)

                    # Category with color
                    cell = ws_details.cell(row=row, column=3, value=cat)
                    cat_color = CATEGORY_COLORS.get(cat, "FFFFFF")
                    cell.fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

                    # Count
                    cell = ws_details.cell(row=row, column=4, value=count)
                    cell.border = thin_border
                    cell.alignment = num_align
                    cell.number_format = '#,##0'

                    row += 1

            # Add thick bottom border after each cluster group
            if row > cluster_start_row:
                for col in range(1, 5):
                    ws_details.cell(row=row - 1, column=col).border = thick_bottom

        # Column widths
        ws_details.column_dimensions['A'].width = 10
        ws_details.column_dimensions['B'].width = 50
        ws_details.column_dimensions['C'].width = 18
        ws_details.column_dimensions['D'].width = 12

        # Freeze header row
        ws_details.freeze_panes = 'A2'

        # =======================================================================
        # Sheet 3: Statistics
        # =======================================================================
        ws_stats = wb.create_sheet("Statistics")

        stats_data = [
            ("Total Unique Patterns", analysis_result["total_patterns"]),
            ("Total Clusters", analysis_result["total_clusters"]),
            ("Similarity Threshold", "80%"),
            ("", ""),
            ("Top 10 Clusters by Usage", ""),
        ]

        for row_idx, (label, value) in enumerate(stats_data, 1):
            cell = ws_stats.cell(row=row_idx, column=1, value=label)
            cell.font = Font(bold=True)
            cell = ws_stats.cell(row=row_idx, column=2, value=value)
            if isinstance(value, int):
                cell.number_format = '#,##0'

        # Top 10 clusters mini-table
        row_offset = 7
        mini_headers = ["#", "Pattern", "Uses"]
        for col, h in enumerate(mini_headers, 1):
            cell = ws_stats.cell(row=row_offset, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thick_bottom

        for i, cluster in enumerate(analysis_result["clusters"][:10], 1):
            row_idx = row_offset + i
            ws_stats.cell(row=row_idx, column=1, value=i).border = thin_border
            cell = ws_stats.cell(row=row_idx, column=2, value=cluster["representative"])
            cell.font = Font(name='Consolas', size=10)
            cell.border = thin_border
            cell = ws_stats.cell(row=row_idx, column=3, value=cluster["total_count"])
            cell.number_format = '#,##0'
            cell.border = thin_border

        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 45
        ws_stats.column_dimensions['C'].width = 12

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Generated pattern report: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Error generating pattern report: {e}")
        return False
