"""
LOCDEV Merger - Merge corrections from Excel back to LOCDEV XML files.

QA testers return Excel files with corrections in "Correction" column.
This module reads ONLY rows with corrections and merges them into LOCDEV
languagedata XML files using STRICT matching: StrOrigin + StringID must BOTH match.

Also tracks SUCCESS/FAIL for each correction to update the Progress Tracker.
"""

import html
import os
import re
import stat
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from lxml import etree

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def normalize_text(txt: Optional[str]) -> str:
    """
    Normalize text for consistent matching.

    Ensures TMX/Excel and XML text are IDENTICAL for matching:
    1. Unescape HTML entities (&lt; -> <, &amp; -> &, etc.)
    2. Strip leading/trailing whitespace
    3. Collapse all internal whitespace (spaces, tabs, newlines) to single space
    4. Remove &desc; markers (legacy description prefix)

    This is the same normalization used in translatexmlstable7.py and tmxtransfer11.py.

    Args:
        txt: Text to normalize

    Returns:
        Normalized text string
    """
    if not txt:
        return ""
    # Unescape HTML entities
    txt = html.unescape(str(txt))
    # Collapse all whitespace to single space and strip
    txt = re.sub(r'\s+', ' ', txt.strip())
    # Remove &desc; markers (case-insensitive)
    if txt.lower().startswith("&desc;"):
        txt = txt[6:].lstrip()
    elif txt.lower().startswith("&amp;desc;"):
        txt = txt[10:].lstrip()
    return txt


def normalize_nospace(txt: str) -> str:
    """
    Remove ALL whitespace from text for fallback matching.

    Used when exact match fails but text differs only in whitespace placement.
    Same pattern as translatexmlstable7.py nospace fallback.

    Args:
        txt: Already normalized text

    Returns:
        Text with all whitespace removed
    """
    return re.sub(r'\s+', '', txt)


def _detect_column_indices(ws) -> Dict[str, int]:
    """
    Detect column indices from header row (CASE INSENSITIVE).

    Returns dict mapping LOWERCASE column name to 1-based index.
    This ensures headers like "StringID", "STRINGID", "stringid" all work.
    """
    indices = {}
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if first_row and first_row[0]:
        for col, header in enumerate(first_row[0], 1):
            if header:
                # Store lowercase key for case-insensitive lookup
                indices[str(header).strip().lower()] = col
    return indices


def parse_corrections_from_excel(excel_path: Path) -> List[Dict]:
    """
    Parse corrections from an Excel file.

    Only extracts rows where the Correction column has a non-empty value.

    Args:
        excel_path: Path to the Excel file

    Returns:
        List of dicts with keys: string_id, str_origin, corrected, category
    """
    corrections = []

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        col_indices = _detect_column_indices(ws)

        # Use lowercase keys for case-insensitive matching
        str_origin_col = col_indices.get("strorigin")
        correction_col = col_indices.get("correction")
        stringid_col = col_indices.get("stringid")
        category_col = col_indices.get("category")  # May be None if not present

        if not all([str_origin_col, correction_col, stringid_col]):
            logger.warning(
                f"Missing required columns in {excel_path.name}. "
                f"Found: {list(col_indices.keys())}"
            )
            wb.close()
            return corrections

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Access by 0-based index (subtract 1 from column number)
            str_origin = row[str_origin_col - 1] if str_origin_col <= len(row) else None
            correction = row[correction_col - 1] if correction_col <= len(row) else None
            string_id = row[stringid_col - 1] if stringid_col <= len(row) else None
            category = None
            if category_col and category_col <= len(row):
                category = row[category_col - 1]

            # Only process rows with a correction value
            if correction and str(correction).strip() and string_id:
                corrections.append({
                    "string_id": str(string_id).strip(),
                    "str_origin": str(str_origin or "").strip(),
                    "corrected": str(correction).strip(),
                    "category": str(category).strip() if category else "Uncategorized",
                })

        wb.close()
        logger.info(f"Parsed {len(corrections)} corrections from {excel_path.name}")

    except Exception as e:
        logger.error(f"Error parsing corrections from {excel_path}: {e}")

    return corrections


def merge_corrections_to_locdev(
    xml_path: Path,
    corrections: List[Dict],
    dry_run: bool = False,
) -> Dict:
    """
    Merge corrections into a LOCDEV XML file.

    Uses STRICT matching: (StringID, normalized_StrOrigin) tuple key.
    Both must match for the correction to be applied.

    Args:
        xml_path: Path to the LOCDEV XML file
        corrections: List of correction dicts from parse_corrections_from_excel
        dry_run: If True, don't write changes, just return what would change

    Returns:
        Dict with statistics: {
            "matched": int,
            "updated": int,
            "not_found": int,
            "errors": List[str],
            "by_category": Dict[str, Dict]  # Per-category breakdown
        }
    """
    result = {
        "matched": 0,
        "updated": 0,
        "not_found": 0,
        "errors": [],
        "by_category": {},  # {category: {matched, updated, not_found}}
    }

    if not corrections:
        return result

    # Build lookup: (StringID, normalized_StrOrigin) -> (corrected_text, category)
    correction_lookup = {}
    correction_lookup_nospace = {}  # Fallback for whitespace variations
    # Track which corrections matched (by index)
    correction_matched = [False] * len(corrections)

    for i, c in enumerate(corrections):
        sid = c["string_id"]
        origin_norm = normalize_text(c["str_origin"])
        origin_nospace = normalize_nospace(origin_norm)
        category = c.get("category", "Uncategorized")

        correction_lookup[(sid, origin_norm)] = (c["corrected"], category, i)
        correction_lookup_nospace[(sid, origin_nospace)] = (c["corrected"], category, i)

    # Initialize category stats
    categories_seen = set(c.get("category", "Uncategorized") for c in corrections)
    for cat in categories_seen:
        result["by_category"][cat] = {"matched": 0, "updated": 0, "not_found": 0}

    try:
        # Parse XML with recovery mode
        parser = etree.XMLParser(
            resolve_entities=False,
            load_dtd=False,
            no_network=True,
            recover=True,
        )
        tree = etree.parse(str(xml_path), parser)
        root = tree.getroot()

        changed = False

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()
            orig = normalize_text(loc.get("StrOrigin", ""))
            orig_nospace = normalize_nospace(orig)
            key = (sid, orig)
            key_nospace = (sid, orig_nospace)

            # Try exact match first, then nospace fallback
            match_data = None
            if key in correction_lookup:
                match_data = correction_lookup[key]
            elif key_nospace in correction_lookup_nospace:
                match_data = correction_lookup_nospace[key_nospace]
                logger.debug(f"Matched via nospace fallback: StringId={sid}")

            if match_data is not None:
                new_str, category, idx = match_data
                correction_matched[idx] = True
                result["matched"] += 1
                result["by_category"][category]["matched"] += 1
                old_str = loc.get("Str", "")

                if new_str != old_str:
                    if not dry_run:
                        loc.set("Str", new_str)
                    result["updated"] += 1
                    result["by_category"][category]["updated"] += 1
                    changed = True
                    logger.debug(f"Updated StringId={sid}: '{old_str}' -> '{new_str}'")

        # Count corrections that didn't match any XML entry (per category)
        for i, c in enumerate(corrections):
            if not correction_matched[i]:
                category = c.get("category", "Uncategorized")
                result["by_category"][category]["not_found"] += 1

        result["not_found"] = len(corrections) - result["matched"]

        if changed and not dry_run:
            # Make file writable if read-only (LOC folder files are often read-only)
            try:
                current_mode = os.stat(xml_path).st_mode
                if not current_mode & stat.S_IWRITE:
                    os.chmod(xml_path, current_mode | stat.S_IWRITE)
                    logger.debug(f"Made {xml_path.name} writable")
            except Exception as e:
                logger.warning(f"Could not make {xml_path.name} writable: {e}")

            tree.write(str(xml_path), encoding="utf-8", xml_declaration=False, pretty_print=True)
            logger.info(f"Saved {xml_path.name}: {result['updated']} entries updated")

    except Exception as e:
        result["errors"].append(str(e))
        logger.error(f"Error merging to {xml_path}: {e}")

    return result


def get_file_mod_time(file_path: Path) -> Optional[datetime]:
    """
    Get the modification time of a file.

    Args:
        file_path: Path to the file

    Returns:
        datetime of last modification, or None if file doesn't exist
    """
    try:
        if file_path.exists():
            mtime = os.path.getmtime(file_path)
            return datetime.fromtimestamp(mtime)
    except Exception as e:
        logger.warning(f"Could not get mod time for {file_path}: {e}")
    return None


def merge_all_corrections(
    submit_folder: Path,
    locdev_folder: Path,
    dry_run: bool = False,
) -> Dict:
    """
    Process all Excel files in ToSubmit folder and merge corrections to LOCDEV.

    Matches Excel files to LOCDEV XML files by language code:
    - languagedata_FRE.xlsx -> languagedata_fre.xml (or languagedata_FRE.xml)
    - LanguageData_GER.xlsx -> languagedata_ger.xml

    Args:
        submit_folder: Path to ToSubmit folder containing Excel files
        locdev_folder: Path to LOCDEV folder containing XML files
        dry_run: If True, don't write changes, just report what would change

    Returns:
        Dict with overall results: {
            "files_processed": int,
            "total_corrections": int,  # Total rows with Correction values
            "total_success": int,      # Successfully merged (matched + updated)
            "total_fail": int,         # Failed to match in LOCDEV
            "errors": List[str],
            "file_results": Dict[str, Dict],  # Per-language results
            "file_mod_times": Dict[str, datetime],  # Per-language file mod times
        }
    """
    results = {
        "files_processed": 0,
        "total_corrections": 0,
        "total_success": 0,
        "total_fail": 0,
        "errors": [],
        "file_results": {},
        "file_mod_times": {},
    }

    if not submit_folder.exists():
        results["errors"].append(f"ToSubmit folder not found: {submit_folder}")
        return results

    if not locdev_folder.exists():
        results["errors"].append(f"LOCDEV folder not found: {locdev_folder}")
        return results

    # Build dynamic language code mapping from LOCDEV folder
    locdev_lang_files = {}  # lang_code (lowercase) -> xml_path
    for xml_file in locdev_folder.glob("languagedata_*.xml"):
        name = xml_file.stem.lower()
        if name.startswith("languagedata_"):
            lang = name[13:]  # After "languagedata_"
            locdev_lang_files[lang.lower()] = xml_file

    if not locdev_lang_files:
        results["errors"].append(f"No languagedata_*.xml files found in LOCDEV: {locdev_folder}")
        return results

    logger.info(f"Found {len(locdev_lang_files)} language files in LOCDEV: {sorted(locdev_lang_files.keys())}")

    # Find Excel files in ToSubmit folder using suffix-based matching
    from .submit_preparer import discover_submit_files
    submit_files = discover_submit_files(submit_folder, locdev_folder)

    if not submit_files:
        results["errors"].append(f"No matching Excel files found in ToSubmit: {submit_folder}")
        return results

    for excel_path, lang_code in submit_files:
        # Get file modification time for weekly tracking
        file_mod_time = get_file_mod_time(excel_path)
        if file_mod_time:
            results["file_mod_times"][lang_code] = file_mod_time

        # Find corresponding LOCDEV XML file (case-insensitive)
        lang_lower = lang_code.lower()
        xml_path = locdev_lang_files.get(lang_lower)

        if not xml_path:
            results["errors"].append(f"No LOCDEV XML found for {excel_path.name} (lang={lang_code})")
            continue

        # Parse corrections from Excel
        corrections = parse_corrections_from_excel(excel_path)

        if not corrections:
            logger.info(f"No corrections found in {excel_path.name}")
            results["file_results"][lang_code] = {
                "corrections": 0,
                "success": 0,
                "fail": 0,
                "matched": 0,
                "updated": 0,
                "not_found": 0,
                "errors": ["No corrections found"],
            }
            continue

        # Merge corrections to LOCDEV
        file_result = merge_corrections_to_locdev(xml_path, corrections, dry_run)

        # Calculate success/fail counts
        # Success = corrections that matched in LOCDEV (whether updated or already same)
        # Fail = corrections that did NOT match in LOCDEV
        corrections_count = len(corrections)
        success_count = file_result["matched"]
        fail_count = file_result["not_found"]

        # Add to file result
        file_result["corrections"] = corrections_count
        file_result["success"] = success_count
        file_result["fail"] = fail_count

        # Add per-category corrections count
        category_counts = {}
        for c in corrections:
            cat = c.get("category", "Uncategorized")
            category_counts[cat] = category_counts.get(cat, 0) + 1
        for cat, stats in file_result.get("by_category", {}).items():
            stats["corrections"] = category_counts.get(cat, 0)

        results["file_results"][lang_code] = file_result

        results["files_processed"] += 1
        results["total_corrections"] += corrections_count
        results["total_success"] += success_count
        results["total_fail"] += fail_count
        results["errors"].extend(file_result["errors"])

        logger.info(
            f"Processed {lang_code}: corrections={corrections_count}, "
            f"success={success_count}, fail={fail_count}"
        )

    return results


# =============================================================================
# STRINGID-ONLY MATCHING (SCRIPT STRINGS)
# =============================================================================

# SCRIPT CATEGORIES - top-level folders where StringID-only matching is safe
# These are STORY-type strings that can be matched by StringID alone
SCRIPT_CATEGORIES = {"Dialog", "Sequencer"}

# Subfolders explicitly excluded from StringID-only transfer
SCRIPT_EXCLUDE_SUBFOLDERS = {"NarrationDialog"}


def merge_corrections_stringid_only_script(
    xml_path: Path,
    corrections: List[Dict],
    stringid_to_toplevel: Dict[str, str],
    stringid_to_subfolder: Optional[Dict[str, str]] = None,
    dry_run: bool = False,
) -> Dict:
    """
    Merge corrections using StringID-ONLY matching.
    ONLY applies to SCRIPT TYPE strings (Dialog/Sequencer).

    This is useful when source text changed but StringID is still valid.
    Non-SCRIPT strings are skipped for safety.

    Args:
        xml_path: Path to LOCDEV XML file
        corrections: List of correction dicts with string_id, corrected keys
        stringid_to_toplevel: Pre-built StringID→TopLevelFolder mapping from EXPORT
        stringid_to_subfolder: Pre-built StringID→Subfolder mapping (for exclusions)
        dry_run: If True, don't write changes

    Returns:
        Dict with: updated, skipped_non_script, skipped_excluded, not_found, errors
    """
    result = {
        "matched": 0,
        "updated": 0,
        "skipped_non_script": 0,
        "skipped_excluded": 0,
        "not_found": 0,
        "errors": [],
        "by_category": {},
    }

    if not corrections:
        return result

    # Filter corrections to SCRIPT TYPE only (and not in excluded subfolders)
    script_corrections = []

    for c in corrections:
        sid = c["string_id"]
        toplevel = stringid_to_toplevel.get(sid, "Uncategorized")
        subfolder = stringid_to_subfolder.get(sid, "") if stringid_to_subfolder else ""

        # Check if in SCRIPT categories (Dialog/Sequencer)
        if toplevel not in SCRIPT_CATEGORIES:
            result["skipped_non_script"] += 1
            logger.debug(f"Skipped non-SCRIPT StringID={sid} (toplevel={toplevel})")
            continue

        # Check if subfolder is in exclusion list
        if subfolder in SCRIPT_EXCLUDE_SUBFOLDERS:
            result["skipped_excluded"] += 1
            logger.debug(f"Skipped excluded subfolder StringID={sid} (subfolder={subfolder})")
            continue

        script_corrections.append({
            **c,
            "category": toplevel,
        })

    if not script_corrections:
        logger.info(f"No SCRIPT corrections to apply to {xml_path.name}")
        return result

    # Build StringID-only lookup (NOT tuple!)
    correction_lookup = {}
    correction_matched = {}

    for c in script_corrections:
        sid = c["string_id"]
        correction_lookup[sid] = c["corrected"]
        correction_matched[sid] = False

        # Initialize category stats
        category = c.get("category", "Uncategorized")
        if category not in result["by_category"]:
            result["by_category"][category] = {"matched": 0, "updated": 0, "not_found": 0}

    try:
        # Parse XML with recovery mode
        parser = etree.XMLParser(
            resolve_entities=False,
            load_dtd=False,
            no_network=True,
            recover=True,
        )
        tree = etree.parse(str(xml_path), parser)
        root = tree.getroot()

        changed = False

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId", "").strip()

            # StringID-only matching
            if sid in correction_lookup:
                new_str = correction_lookup[sid]
                correction_matched[sid] = True

                # Find category for stats
                category = stringid_to_category.get(sid, "Uncategorized")

                result["matched"] += 1
                if category in result["by_category"]:
                    result["by_category"][category]["matched"] += 1

                old_str = loc.get("Str", "")

                if new_str != old_str:
                    if not dry_run:
                        loc.set("Str", new_str)
                    result["updated"] += 1
                    if category in result["by_category"]:
                        result["by_category"][category]["updated"] += 1
                    changed = True
                    logger.debug(f"Updated StringId={sid}: '{old_str}' -> '{new_str}'")

        # Count corrections that didn't match any XML entry
        for sid, matched in correction_matched.items():
            if not matched:
                category = stringid_to_category.get(sid, "Uncategorized")
                result["not_found"] += 1
                if category in result["by_category"]:
                    result["by_category"][category]["not_found"] += 1

        if changed and not dry_run:
            # Make file writable if read-only
            try:
                current_mode = os.stat(xml_path).st_mode
                if not current_mode & stat.S_IWRITE:
                    os.chmod(xml_path, current_mode | stat.S_IWRITE)
                    logger.debug(f"Made {xml_path.name} writable")
            except Exception as e:
                logger.warning(f"Could not make {xml_path.name} writable: {e}")

            tree.write(str(xml_path), encoding="utf-8", xml_declaration=False, pretty_print=True)
            logger.info(f"Saved {xml_path.name}: {result['updated']} entries updated (StringID-only)")

    except Exception as e:
        result["errors"].append(str(e))
        logger.error(f"Error merging to {xml_path}: {e}")

    return result


def merge_all_corrections_stringid_only_script(
    submit_folder: Path,
    locdev_folder: Path,
    export_folder: Path,
    dry_run: bool = False,
) -> Dict:
    """
    Process all Excel files using StringID-only matching for SCRIPT strings.

    This function:
    1. Builds StringID→TopLevel and StringID→Subfolder indexes from EXPORT folder
    2. Parses corrections from Excel files in ToSubmit folder
    3. Filters to only SCRIPT-type strings (Dialog/Sequencer), excluding NarrationDialog
    4. Applies corrections using StringID-only matching

    Args:
        submit_folder: Path to ToSubmit folder containing Excel files
        locdev_folder: Path to LOCDEV folder containing XML files
        export_folder: Path to EXPORT folder for category detection
        dry_run: If True, don't write changes

    Returns:
        Dict with overall results
    """
    from .category_mapper import build_stringid_to_toplevel, build_stringid_to_subfolder

    results = {
        "files_processed": 0,
        "total_corrections": 0,
        "total_script_corrections": 0,
        "total_skipped_non_script": 0,
        "total_skipped_excluded": 0,
        "total_success": 0,
        "total_fail": 0,
        "errors": [],
        "file_results": {},
    }

    if not submit_folder.exists():
        results["errors"].append(f"ToSubmit folder not found: {submit_folder}")
        return results

    if not locdev_folder.exists():
        results["errors"].append(f"LOCDEV folder not found: {locdev_folder}")
        return results

    if not export_folder.exists():
        results["errors"].append(f"EXPORT folder not found: {export_folder}")
        return results

    # Build StringID→TopLevel and StringID→Subfolder indexes from EXPORT folder
    logger.info("Building StringID→TopLevel index from EXPORT folder...")
    stringid_to_toplevel = build_stringid_to_toplevel(export_folder)

    if not stringid_to_toplevel:
        results["errors"].append("Failed to build StringID→TopLevel index from EXPORT folder")
        return results

    logger.info(f"Built top-level index with {len(stringid_to_toplevel)} StringIDs")

    logger.info("Building StringID→Subfolder index from EXPORT folder...")
    stringid_to_subfolder = build_stringid_to_subfolder(export_folder)
    logger.info(f"Built subfolder index with {len(stringid_to_subfolder)} StringIDs")

    # Build dynamic language code mapping from LOCDEV folder
    locdev_lang_files = {}  # lang_code (lowercase) -> xml_path
    for xml_file in locdev_folder.glob("languagedata_*.xml"):
        name = xml_file.stem.lower()
        if name.startswith("languagedata_"):
            lang = name[13:]  # After "languagedata_"
            locdev_lang_files[lang.lower()] = xml_file

    if not locdev_lang_files:
        results["errors"].append(f"No languagedata_*.xml files found in LOCDEV: {locdev_folder}")
        return results

    logger.info(f"Found {len(locdev_lang_files)} language files in LOCDEV: {sorted(locdev_lang_files.keys())}")

    # Find Excel files in ToSubmit folder using suffix-based matching
    from .submit_preparer import discover_submit_files
    submit_files = discover_submit_files(submit_folder, locdev_folder)

    if not submit_files:
        results["errors"].append(f"No matching Excel files found in ToSubmit: {submit_folder}")
        return results

    for excel_path, lang_code in submit_files:
        # Find corresponding LOCDEV XML file (case-insensitive)
        lang_lower = lang_code.lower()
        xml_path = locdev_lang_files.get(lang_lower)

        if not xml_path:
            results["errors"].append(f"No LOCDEV XML found for {excel_path.name} (lang={lang_code})")
            continue

        # Parse corrections from Excel
        corrections = parse_corrections_from_excel(excel_path)

        if not corrections:
            logger.info(f"No corrections found in {excel_path.name}")
            results["file_results"][lang_code] = {
                "corrections": 0,
                "script_corrections": 0,
                "skipped_non_script": 0,
                "success": 0,
                "fail": 0,
                "errors": ["No corrections found"],
            }
            continue

        # Merge corrections using StringID-only matching for SCRIPT strings
        file_result = merge_corrections_stringid_only_script(
            xml_path, corrections, stringid_to_toplevel, stringid_to_subfolder, dry_run
        )

        # Calculate counts
        corrections_count = len(corrections)
        skipped_total = file_result["skipped_non_script"] + file_result.get("skipped_excluded", 0)
        script_count = corrections_count - skipped_total
        success_count = file_result["matched"]
        fail_count = file_result["not_found"]

        # Add to file result
        file_result["corrections"] = corrections_count
        file_result["script_corrections"] = script_count
        file_result["success"] = success_count
        file_result["fail"] = fail_count

        results["file_results"][lang_code] = file_result

        results["files_processed"] += 1
        results["total_corrections"] += corrections_count
        results["total_script_corrections"] += script_count
        results["total_skipped_non_script"] += file_result["skipped_non_script"]
        results["total_skipped_excluded"] += file_result.get("skipped_excluded", 0)
        results["total_success"] += success_count
        results["total_fail"] += fail_count
        results["errors"].extend(file_result["errors"])

        logger.info(
            f"Processed {lang_code}: corrections={corrections_count}, "
            f"script={script_count}, skipped_non_script={file_result['skipped_non_script']}, "
            f"skipped_excluded={file_result.get('skipped_excluded', 0)}, "
            f"success={success_count}, fail={fail_count}"
        )

    return results


def print_stringid_only_report(results: Dict) -> None:
    """
    Print a formatted terminal report of StringID-only merge results.

    Args:
        results: The dict returned by merge_all_corrections_stringid_only_script()
    """
    # Box drawing characters
    H = "═"
    V = "║"
    TL = "╔"
    TR = "╗"
    BL = "╚"
    BR = "╝"
    LT = "╠"
    RT = "╣"

    width = 80

    print()
    print(TL + H * (width - 2) + TR)
    title = "STRINGID-ONLY MERGE REPORT (SCRIPT STRINGS)"
    print(V + title.center(width - 2) + V)
    print(LT + H * (width - 2) + RT)

    # Column widths
    lang_w = 10
    corr_w = 10
    script_w = 10
    skip_w = 10
    succ_w = 10
    fail_w = 10
    rate_w = 10

    # Header
    header = (
        f"{V} {'LANG':<{lang_w}} {'TOTAL':>{corr_w}} {'SCRIPT':>{script_w}} "
        f"{'SKIPPED':>{skip_w}} {'SUCCESS':>{succ_w}} {'FAIL':>{fail_w}} {'RATE':>{rate_w}} {V}"
    )
    print(header)
    print(LT + H * (width - 2) + RT)

    # Per-language rows
    file_results = results.get("file_results", {})
    for lang_code in sorted(file_results.keys()):
        file_result = file_results[lang_code]
        corrections = file_result.get("corrections", 0)
        script = file_result.get("script_corrections", 0)
        skipped = file_result.get("skipped_non_script", 0)
        success = file_result.get("success", 0)
        fail = file_result.get("fail", 0)
        rate = (success / script * 100) if script > 0 else 0.0

        # Status indicator
        if rate >= 95:
            status = "●"
        elif rate >= 80:
            status = "◐"
        else:
            status = "○"

        row = (
            f"{V} {lang_code.upper():<{lang_w}} {corrections:>{corr_w},} {script:>{script_w},} "
            f"{skipped:>{skip_w},} {success:>{succ_w},} {fail:>{fail_w},} {rate:>{rate_w - 2}.1f}% {status}{V}"
        )
        print(row)

    # Separator before totals
    print(LT + H * (width - 2) + RT)

    # Totals
    total_corr = results.get("total_corrections", 0)
    total_script = results.get("total_script_corrections", 0)
    total_skip = results.get("total_skipped_non_script", 0)
    total_succ = results.get("total_success", 0)
    total_fail = results.get("total_fail", 0)
    total_rate = (total_succ / total_script * 100) if total_script > 0 else 0.0

    total_row = (
        f"{V} {'TOTAL':<{lang_w}} {total_corr:>{corr_w},} {total_script:>{script_w},} "
        f"{total_skip:>{skip_w},} {total_succ:>{succ_w},} {total_fail:>{fail_w},} {total_rate:>{rate_w - 2}.1f}%  {V}"
    )
    print(total_row)
    print(BL + H * (width - 2) + BR)

    # Errors section
    errors = results.get("errors", [])
    if errors:
        print()
        print("ERRORS:")
        for error in errors[:5]:
            print(f"  × {error}")
        if len(errors) > 5:
            print(f"  ... and {len(errors) - 5} more errors")

    # Legend
    print()
    print("Legend: ● ≥95% success  ◐ ≥80% success  ○ <80% success")
    print("SCRIPT categories: Sequencer, AIDialog, QuestDialog, NarrationDialog")
    print()


def print_merge_report(results: Dict) -> None:
    """
    Print a formatted terminal report of merge results.

    Args:
        results: The dict returned by merge_all_corrections()
    """
    # Box drawing characters for nice borders
    H = "═"  # Horizontal
    V = "║"  # Vertical
    TL = "╔"  # Top-left
    TR = "╗"  # Top-right
    BL = "╚"  # Bottom-left
    BR = "╝"  # Bottom-right
    LT = "╠"  # Left-T
    RT = "╣"  # Right-T
    TT = "╦"  # Top-T
    BT = "╩"  # Bottom-T
    X = "╬"   # Cross

    width = 72

    print()
    print(TL + H * (width - 2) + TR)
    title = "LOCDEV MERGE REPORT"
    print(V + title.center(width - 2) + V)
    print(LT + H * (width - 2) + RT)

    # Column widths
    lang_w = 14
    corr_w = 14
    succ_w = 12
    fail_w = 10
    rate_w = 12

    # Header
    header = f"{V} {'LANGUAGE':<{lang_w}} {'CORRECTIONS':>{corr_w}} {'SUCCESS':>{succ_w}} {'FAIL':>{fail_w}} {'RATE':>{rate_w}} {V}"
    print(header)
    print(LT + H * (width - 2) + RT)

    # Per-language rows
    file_results = results.get("file_results", {})
    for lang_code in sorted(file_results.keys()):
        file_result = file_results[lang_code]
        corrections = file_result.get("corrections", 0)
        success = file_result.get("success", 0)
        fail = file_result.get("fail", 0)
        rate = (success / corrections * 100) if corrections > 0 else 0.0

        # Color indicators
        if rate >= 95:
            status = "●"  # Green indicator
        elif rate >= 80:
            status = "◐"  # Half indicator
        else:
            status = "○"  # Empty indicator

        row = f"{V} {lang_code.upper():<{lang_w}} {corrections:>{corr_w},} {success:>{succ_w},} {fail:>{fail_w},} {rate:>{rate_w - 2}.1f}% {status} {V}"
        print(row)

    # Separator before totals
    print(LT + H * (width - 2) + RT)

    # Totals
    total_corr = results.get("total_corrections", 0)
    total_succ = results.get("total_success", 0)
    total_fail = results.get("total_fail", 0)
    total_rate = (total_succ / total_corr * 100) if total_corr > 0 else 0.0

    total_row = f"{V} {'TOTAL':<{lang_w}} {total_corr:>{corr_w},} {total_succ:>{succ_w},} {total_fail:>{fail_w},} {total_rate:>{rate_w - 2}.1f}%   {V}"
    print(total_row)
    print(BL + H * (width - 2) + BR)

    # Errors section (if any)
    errors = results.get("errors", [])
    if errors:
        print()
        print("ERRORS:")
        for error in errors[:5]:  # Show first 5 errors
            print(f"  × {error}")
        if len(errors) > 5:
            print(f"  ... and {len(errors) - 5} more errors")

    # Legend
    print()
    print("Legend: ● ≥95% success  ◐ ≥80% success  ○ <80% success")
    print()
