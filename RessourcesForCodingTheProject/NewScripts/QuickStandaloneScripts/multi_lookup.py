#!/usr/bin/env python3
# coding: utf-8
"""
MultiLookup v1.2
==================
Standalone GUI tool for Excel-to-Excel lookup transfer with multi-file
support, normalized key matching, and source-to-target routing.

Build lookup dictionaries from SOURCE Excel files, then write matched
values into TARGET Excel files — with per-column control over which
source provides the value.

Features:
  - Multi-file source/target with per-file sheet + column config
  - Normalized key matching (strip, collapse spaces, remove _x000D_, case-insensitive)
  - Composite keys (optional KEY Col 2 for matching on 2 columns)
  - Source-to-target routing: each target WRITE column picks its source
  - First-wins dedup for duplicate normalized keys
  - Save as copy (default) or overwrite original
  - PanedWindow layout: controls left (75%), log right (25%)

Usage: python multi_lookup.py
"""
from __future__ import annotations

import json
import re
import sys
import logging
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Try openpyxl (required for both read and write on existing files)
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# =============================================================================
# LOGGING
# =============================================================================

logger = logging.getLogger("MultiLookup")
logger.setLevel(logging.DEBUG)

# =============================================================================
# CONSTANTS
# =============================================================================

if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent

SETTINGS_FILE = SCRIPT_DIR / "multi_lookup_settings.json"

SAVE_MODES = ["Save as _lookup copy", "Overwrite original"]

COMPOSITE_SEP = "|||"

ALL_MERGED = "ALL"

# =============================================================================
# SETTINGS PERSISTENCE
# =============================================================================

def _load_settings() -> dict:
    """Load persisted settings from JSON."""
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_settings(settings: dict):
    """Save settings to JSON."""
    try:
        SETTINGS_FILE.write_text(
            json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8",
        )
    except Exception as e:
        logger.warning("Failed to save settings: %s", e)


# =============================================================================
# NORMALIZATION
# =============================================================================

def normalize_key(value) -> str:
    """Normalize a cell value for key matching.

    Strip whitespace, remove _x000D_, collapse multiple spaces, lowercase.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = re.sub(r'_x000D_', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s.lower()


def clean_value(value) -> Optional[str]:
    """Clean a value for transfer. Strip whitespace, remove _x000D_.
    NOT lowercased — preserves original casing."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r'_x000D_', '', s, flags=re.IGNORECASE)
    return s


# =============================================================================
# EXCEL INTROSPECTION
# =============================================================================

def read_sheets(file_path: Path) -> List[str]:
    """Read sheet names from an Excel file."""
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        logger.error("Failed to read sheets from %s: %s", file_path, e)
        return []


def read_headers(file_path: Path, sheet_name: str) -> List[str]:
    """Read column headers (row 1) from a specific sheet."""
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb[sheet_name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    headers.append(str(cell_val).strip())
                else:
                    headers.append("")
        wb.close()
        return headers
    except Exception as e:
        logger.error("Failed to read headers from %s [%s]: %s", file_path, sheet_name, e)
        return []


# =============================================================================
# CORE LOGIC
# =============================================================================

def _make_composite_key(row, key_idx: int, key2_idx: int) -> str:
    """Build a (possibly composite) normalized key from a row."""
    if key_idx >= len(row):
        return ""
    nk = normalize_key(row[key_idx])
    if not nk:
        return ""
    if key2_idx >= 0:
        raw2 = row[key2_idx] if key2_idx < len(row) else None
        nk2 = normalize_key(raw2)
        nk = nk + COMPOSITE_SEP + nk2
    return nk


def build_source_dicts(
    entries: List[dict],
    log_fn=None,
) -> Tuple[Dict[str, Dict[str, str]], int]:
    """Build per-source lookup dicts keyed by source_key.

    Each entry: {source_key, path, sheet, key_col_idx, value_col_idx, key_col2_idx}

    Returns:
        (source_dicts: {source_key: {norm_key: clean_value}}, total_duplicates)
    """
    source_dicts: Dict[str, Dict[str, str]] = {}
    total_duplicates = 0

    for entry in entries:
        file_path = Path(entry["path"])
        sheet = entry["sheet"]
        key_idx = entry["key_col_idx"]
        val_idx = entry["value_col_idx"]
        key2_idx = entry.get("key_col2_idx", -1)
        src_key = entry["source_key"]

        if log_fn:
            log_fn(f"  Reading: {file_path.name} [{sheet}]", "info")

        wb = None
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            ws = wb[sheet]
        except Exception as e:
            if log_fn:
                log_fn(f"    ERROR: {e}", "error")
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            continue

        lookup: Dict[str, str] = {}
        row_count = 0
        added = 0
        duplicates = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            nk = _make_composite_key(row, key_idx, key2_idx)
            if not nk:
                continue

            raw_val = row[val_idx] if val_idx < len(row) else None
            cv = clean_value(raw_val)

            if nk in lookup:
                duplicates += 1
            else:
                lookup[nk] = cv if cv is not None else ""
                added += 1

        wb.close()
        total_duplicates += duplicates
        source_dicts[src_key] = lookup

        if log_fn:
            log_fn(f"    {row_count} rows scanned, {added} keys added", "info")

    return source_dicts, total_duplicates


def build_merged_dict(source_dicts: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    """Merge all per-source dicts. First-wins dedup across sources."""
    merged: Dict[str, str] = {}
    for d in source_dicts.values():
        for k, v in d.items():
            if k not in merged:
                merged[k] = v
    return merged


def transfer_to_targets(
    source_dicts: Dict[str, Dict[str, str]],
    merged_dict: Optional[Dict[str, str]],
    target_entries: List[dict],
    save_mode: str,
    log_fn=None,
) -> Tuple[int, int, int]:
    """Transfer values from source dicts into target files via write mappings.

    Each target entry: {path, sheet, key_col_idx, key_col2_idx,
                        write_mappings: [{col_idx, source_key}]}

    Returns:
        (total_matches, total_rows, files_saved)
    """
    total_matches = 0
    total_rows = 0
    files_saved = 0

    for entry in target_entries:
        file_path = Path(entry["path"])
        sheet = entry["sheet"]
        key_idx = entry["key_col_idx"]
        key2_idx = entry.get("key_col2_idx", -1)
        mappings = entry.get("write_mappings", [])

        if not mappings:
            if log_fn:
                log_fn(f"\n  SKIP: {file_path.name} [{sheet}] — no write mappings", "warning")
            continue

        if log_fn:
            log_fn(f"\n  Processing: {file_path.name} [{sheet}]", "info")

        # Resolve source dicts for each mapping
        resolved: List[Tuple[int, Optional[Dict[str, str]], str]] = []
        for m in mappings:
            col_idx = m["col_idx"]
            sk = m["source_key"]
            if sk == ALL_MERGED:
                resolved.append((col_idx, merged_dict, "ALL (merged)"))
            elif sk in source_dicts:
                resolved.append((col_idx, source_dicts[sk], sk.split("|")[-1] if "|" in sk else sk))
            else:
                if log_fn:
                    src_name = sk.split("|")[0].rsplit("/", 1)[-1] if "|" in sk else sk
                    log_fn(f"    WARNING: Source not found: {src_name} — skipping column", "warning")
                resolved.append((col_idx, None, sk))

        try:
            wb = openpyxl.load_workbook(str(file_path))
            ws = wb[sheet]
        except PermissionError:
            if log_fn:
                log_fn(f"    ERROR: File locked — close it in Excel first", "error")
            continue
        except Exception as e:
            if log_fn:
                log_fn(f"    ERROR: {e}", "error")
            continue

        # Per-mapping match counters
        match_counts = [0] * len(resolved)
        rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            rows += 1
            if key_idx >= len(row):
                continue
            raw_key = row[key_idx].value
            nk = normalize_key(raw_key)
            if not nk:
                continue
            if key2_idx >= 0:
                raw_key2 = row[key2_idx].value if key2_idx < len(row) else None
                nk2 = normalize_key(raw_key2)
                nk = nk + COMPOSITE_SEP + nk2

            for mi, (col_idx, lookup, _label) in enumerate(resolved):
                if lookup is None:
                    continue
                if nk in lookup:
                    if col_idx < len(row):
                        row[col_idx].value = lookup[nk]
                    else:
                        ws.cell(row=row_idx, column=col_idx + 1, value=lookup[nk])
                    match_counts[mi] += 1

        total_rows += rows
        total_matches += sum(match_counts)

        # Log per-mapping stats
        if log_fn:
            for mi, (col_idx, _lookup, label) in enumerate(resolved):
                if _lookup is not None:
                    col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                    log_fn(f"    {col_letter} <- {label}: {match_counts[mi]}/{rows} matched",
                           "success" if match_counts[mi] > 0 else "info")

        # Save
        try:
            if save_mode == SAVE_MODES[0]:
                stem = file_path.stem
                out_path = file_path.parent / f"{stem}_lookup{file_path.suffix}"
            else:
                out_path = file_path

            wb.save(str(out_path))
            wb.close()
            files_saved += 1

            if log_fn:
                log_fn(f"    Saved: {out_path.name}", "success")

        except PermissionError:
            if log_fn:
                log_fn(f"    ERROR: Cannot save — file locked. Close it in Excel first.", "error")
            wb.close()
        except Exception as e:
            if log_fn:
                log_fn(f"    ERROR saving: {e}", "error")
            wb.close()

    return total_matches, total_rows, files_saved


# =============================================================================
# DATA MODELS
# =============================================================================

class WriteMapping:
    """One write column in a target file, linked to a specific source."""

    __slots__ = ("col_name", "col_idx", "source_key")

    def __init__(self):
        self.col_name: str = ""     # "2: ENG_Text" (display string)
        self.col_idx: int = -1      # 2 (0-based column index)
        self.source_key: str = ALL_MERGED  # "ALL" or "path|sheet"


class FileEntry:
    """Stores config for one file in the source or target list.

    Source entries: use col2/col2_idx for VALUE column. write_mappings empty.
    Target entries: use write_mappings for WRITE columns. col2/col2_idx unused.
    """

    __slots__ = ("path", "sheets", "headers", "sheet", "col1", "col2",
                 "col1_idx", "col2_idx", "col1b", "col1b_idx",
                 "write_mappings")

    def __init__(self, path: str):
        self.path = path
        self.sheets: List[str] = []
        self.headers: List[str] = []
        self.sheet: str = ""
        self.col1: str = ""   # KEY column
        self.col2: str = ""   # VALUE (source only)
        self.col1_idx: int = -1
        self.col2_idx: int = -1
        self.col1b: str = ""  # KEY Col 2 (optional composite key)
        self.col1b_idx: int = -1
        self.write_mappings: list = []  # List[WriteMapping], target only

    @property
    def source_key(self) -> str:
        """Stable identity for this source entry: 'path|sheet'"""
        return f"{self.path}|{self.sheet}"

    @property
    def display(self) -> str:
        name = Path(self.path).name
        if self.sheet:
            return f"{name} [{self.sheet}]"
        return name

    def to_source_dict(self) -> dict:
        return {
            "source_key": self.source_key,
            "path": self.path,
            "sheet": self.sheet,
            "key_col_idx": self.col1_idx,
            "value_col_idx": self.col2_idx,
            "key_col2_idx": self.col1b_idx,
        }

    def to_target_dict(self) -> dict:
        return {
            "path": self.path,
            "sheet": self.sheet,
            "key_col_idx": self.col1_idx,
            "key_col2_idx": self.col1b_idx,
            "write_mappings": [
                {"col_idx": wm.col_idx, "source_key": wm.source_key}
                for wm in self.write_mappings
                if wm.col_idx >= 0
            ],
        }


# =============================================================================
# GUI APPLICATION
# =============================================================================

class MultiLookupApp:
    """GUI for Excel-to-Excel lookup transfer with source-to-target routing."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MultiLookup v1.2")
        self.root.geometry("1400x850")
        self.root.resizable(True, True)

        self.source_entries: List[FileEntry] = []
        self.target_entries: List[FileEntry] = []

        self._current_src_idx: int = -1
        self._current_tgt_idx: int = -1
        self._ignore_events: bool = False

        # Widget references for dynamic write mapping rows: {id(mapping): (frame, col_cb, src_cb, num_label)}
        self._mapping_widgets: Dict[int, tuple] = {}
        # Mapping from source display string → source_key
        self._source_display_to_key: Dict[str, str] = {}
        # Current source selector values list
        self._source_selector_values: List[str] = []

        self._build_ui()
        self._load_persisted_settings()
        self._check_openpyxl()

    def _check_openpyxl(self):
        if not HAS_OPENPYXL:
            self._log("ERROR: openpyxl is required. Install with: pip install openpyxl", "error")
            self.transfer_btn.config(state="disabled")

    # -----------------------------------------------------------------
    # UI BUILDING
    # -----------------------------------------------------------------

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        # Title
        ttk.Label(
            main, text="MultiLookup v1.2",
            font=("Segoe UI", 14, "bold"),
        ).pack(pady=(0, 2))
        ttk.Label(
            main,
            text="Excel-to-Excel lookup transfer with source-to-target routing",
            font=("Segoe UI", 9),
        ).pack(pady=(0, 10))

        # === Horizontal Split: Controls (left) | Log (right) ===
        self._paned = tk.PanedWindow(main, orient=tk.HORIZONTAL, sashwidth=6,
                                     bg='#cccccc', sashrelief='raised')
        self._paned.pack(fill="both", expand=True)

        # --- Left pane ---
        left_pane = ttk.Frame(self._paned)

        self._build_source_section(left_pane)
        self._build_target_section(left_pane)

        # Transfer button
        self.transfer_btn = ttk.Button(
            left_pane, text="===== TRANSFER =====", width=30,
            command=self._run_transfer,
        )
        self.transfer_btn.pack(pady=8)

        # --- Right pane: log ---
        right_pane = ttk.Frame(self._paned)

        f_log = ttk.LabelFrame(right_pane, text="Log", padding=5)
        f_log.pack(fill="both", expand=True)
        self.log = scrolledtext.ScrolledText(
            f_log, height=12, font=("Consolas", 9), wrap="word",
        )
        self.log.pack(fill="both", expand=True)

        ttk.Button(
            f_log, text="Clear Log", width=12,
            command=lambda: self.log.delete("1.0", "end"),
        ).pack(pady=(4, 0))

        self.log.tag_config("info", foreground="black")
        self.log.tag_config("success", foreground="green")
        self.log.tag_config("warning", foreground="orange")
        self.log.tag_config("error", foreground="red")
        self.log.tag_config("header", foreground="blue", font=("Consolas", 9, "bold"))

        # Add panes
        self._paned.add(left_pane, minsize=900)
        self._paned.add(right_pane, minsize=200)

        self.root.after(200, self._set_initial_sash)

    def _set_initial_sash(self):
        """Set PanedWindow sash to ~75% left / 25% right split."""
        total_width = self._paned.winfo_width()
        if total_width > 1:
            self._paned.sash_place(0, int(total_width * 0.75), 0)

    def _build_file_list_and_keys(self, parent, title: str, is_source: bool):
        """Shared helper: button row + listbox + sheet/KEY/KEY2 comboboxes."""
        pad = {"padx": 8, "pady": 4}
        lf = ttk.LabelFrame(parent, text=title, padding=5)
        lf.pack(fill="both", expand=True, **pad)

        # Buttons row
        btn_row = ttk.Frame(lf)
        btn_row.pack(fill="x", pady=(0, 4))

        ttk.Button(
            btn_row, text="Add Files...",
            command=lambda: self._add_files(is_source),
        ).pack(side="left", padx=(0, 3))
        ttk.Button(
            btn_row, text="Remove",
            command=lambda: self._remove_selected(is_source),
        ).pack(side="left", padx=(0, 3))
        ttk.Button(
            btn_row, text="Clear All",
            command=lambda: self._clear_files(is_source),
        ).pack(side="left", padx=(0, 3))

        # Status label
        status_var = tk.StringVar(value="No files")
        ttk.Label(btn_row, textvariable=status_var, font=("Segoe UI", 8)).pack(
            side="right", padx=5,
        )

        # Listbox
        lb_frame = ttk.Frame(lf)
        lb_frame.pack(fill="both", expand=True, pady=(0, 4))
        lb = tk.Listbox(lb_frame, height=8, font=("Consolas", 9))
        lb.pack(fill="both", side="left", expand=True)
        sb = ttk.Scrollbar(lb_frame, orient="vertical", command=lb.yview)
        sb.pack(side="right", fill="y")
        lb.config(yscrollcommand=sb.set)

        # Config frame: Sheet + KEY + KEY Col 2
        cfg = ttk.Frame(lf)
        cfg.pack(fill="x", pady=(0, 4))

        ttk.Label(cfg, text="Sheet:").grid(row=0, column=0, sticky="w", padx=(0, 5))
        sheet_cb = ttk.Combobox(cfg, state="readonly", width=25)
        sheet_cb.grid(row=0, column=1, sticky="w", padx=(0, 15))

        ttk.Label(cfg, text="KEY Column:").grid(row=0, column=2, sticky="w", padx=(0, 5))
        key_cb = ttk.Combobox(cfg, state="readonly", width=25)
        key_cb.grid(row=0, column=3, sticky="w", padx=(0, 15))

        ttk.Label(cfg, text="KEY Col 2:").grid(row=0, column=4, sticky="w", padx=(0, 5))
        key2_cb = ttk.Combobox(cfg, state="readonly", width=25)
        key2_cb.grid(row=0, column=5, sticky="w", padx=(0, 5))
        ttk.Label(cfg, text="(optional)", font=("Segoe UI", 8)).grid(
            row=0, column=6, sticky="w", padx=(0, 3))
        ttk.Button(
            cfg, text="Clr", width=3,
            command=lambda: self._clear_key2(is_source),
        ).grid(row=0, column=7, sticky="w")

        # Bind events
        lb.bind("<<ListboxSelect>>", lambda e: self._on_listbox_select(is_source))
        sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._on_sheet_select(is_source))
        key_cb.bind("<<ComboboxSelected>>", lambda e: self._on_col_select(is_source, is_key=True))
        key2_cb.bind("<<ComboboxSelected>>", lambda e: self._on_col_select(is_source, is_key2=True))

        return lf, lb, sheet_cb, key_cb, key2_cb, status_var

    def _build_source_section(self, parent):
        """Build SOURCE section: file list + keys + VALUE column."""
        lf, lb, sheet_cb, key_cb, key2_cb, status_var = self._build_file_list_and_keys(
            parent, "SOURCE Files (lookup dictionary)", is_source=True,
        )

        # VALUE column combobox (source-specific)
        val_frame = ttk.Frame(lf)
        val_frame.pack(fill="x", pady=(0, 4))
        ttk.Label(val_frame, text="VALUE Column:").pack(side="left", padx=(0, 5))
        val_cb = ttk.Combobox(val_frame, state="readonly", width=25)
        val_cb.pack(side="left")
        val_cb.bind("<<ComboboxSelected>>", lambda e: self._on_col_select(True, is_key=False))

        self.src_lb = lb
        self.src_sheet_cb = sheet_cb
        self.src_key_cb = key_cb
        self.src_key2_cb = key2_cb
        self.src_val_cb = val_cb
        self.src_status = status_var

    def _build_target_section(self, parent):
        """Build TARGET section: file list + keys + write mappings."""
        lf, lb, sheet_cb, key_cb, key2_cb, status_var = self._build_file_list_and_keys(
            parent, "TARGET Files (write into)", is_source=False,
        )

        # Write Mappings area
        wm_outer = ttk.LabelFrame(lf, text="Write Mappings", padding=3)
        wm_outer.pack(fill="x", pady=(0, 4))

        # Header row
        hdr = ttk.Frame(wm_outer)
        hdr.pack(fill="x")
        ttk.Label(hdr, text="#", width=3, font=("Consolas", 8)).pack(side="left", padx=(0, 5))
        ttk.Label(hdr, text="Write Into Column", width=28, font=("Segoe UI", 8)).pack(side="left", padx=(0, 5))
        ttk.Label(hdr, text="Value From Source", width=32, font=("Segoe UI", 8)).pack(side="left", padx=(0, 5))

        # Container for dynamic mapping rows
        self._wm_container = ttk.Frame(wm_outer)
        self._wm_container.pack(fill="x")

        # Add / buttons
        wm_btn_row = ttk.Frame(wm_outer)
        wm_btn_row.pack(fill="x", pady=(4, 0))
        ttk.Button(
            wm_btn_row, text="+ Add Write Column", width=20,
            command=self._add_write_mapping,
        ).pack(side="left", padx=(0, 5))

        # Save mode
        save_frame = ttk.Frame(lf)
        save_frame.pack(fill="x", pady=(0, 2))
        ttk.Label(save_frame, text="Save Mode:").pack(side="left", padx=(0, 5))
        save_cb = ttk.Combobox(
            save_frame, state="readonly", values=SAVE_MODES, width=30,
        )
        save_cb.set(SAVE_MODES[0])
        save_cb.pack(side="left")

        self.tgt_lb = lb
        self.tgt_sheet_cb = sheet_cb
        self.tgt_key_cb = key_cb
        self.tgt_key2_cb = key2_cb
        self.tgt_status = status_var
        self.save_mode_cb = save_cb

    # -----------------------------------------------------------------
    # SOURCE DISPLAY NUMBERING
    # -----------------------------------------------------------------

    def _numbered_source_display(self, idx: int, entry: FileEntry) -> str:
        return f"[{idx + 1}] {entry.display}"

    def _update_source_listbox(self):
        """Refresh source listbox with [N] prefixes."""
        self.src_lb.delete(0, "end")
        for i, entry in enumerate(self.source_entries):
            self.src_lb.insert("end", self._numbered_source_display(i, entry))
        self._update_status(True)

    def _rebuild_source_selectors(self):
        """Rebuild source selector dropdown values for all target write mappings."""
        values = ["ALL (merged)"]
        key_map = {"ALL (merged)": ALL_MERGED}
        for i, entry in enumerate(self.source_entries):
            display = f"[{i + 1}] {entry.display}"
            values.append(display)
            key_map[display] = entry.source_key

        self._source_selector_values = values
        self._source_display_to_key = key_map

        # Update any visible write mapping source comboboxes
        for _mid, (frame, col_cb, src_cb, _lbl) in self._mapping_widgets.items():
            current = src_cb.get()
            src_cb["values"] = values
            if current in values:
                src_cb.set(current)
            else:
                # Source was removed — fall back to ALL
                src_cb.set("ALL (merged)")

    def _source_key_to_display(self, source_key: str) -> str:
        """Convert a source_key to its display string for the combobox."""
        if source_key == ALL_MERGED:
            return "ALL (merged)"
        for display, key in self._source_display_to_key.items():
            if key == source_key:
                return display
        return "ALL (merged)"

    # -----------------------------------------------------------------
    # WRITE MAPPING ROWS (dynamic)
    # -----------------------------------------------------------------

    def _add_write_mapping(self, mapping: Optional[WriteMapping] = None):
        """Add a write mapping row to the current target entry."""
        tgt_entry = self._get_selected_entry(False)
        if not tgt_entry:
            return

        if mapping is None:
            mapping = WriteMapping()
            tgt_entry.write_mappings.append(mapping)

        row_num = len([w for w in self._mapping_widgets.values()]) + 1

        row_frame = ttk.Frame(self._wm_container)
        row_frame.pack(fill="x", pady=1)

        num_label = ttk.Label(row_frame, text=str(row_num), width=3, font=("Consolas", 9))
        num_label.pack(side="left", padx=(0, 5))

        # Column combobox — populated from target headers
        col_display = [f"{i}: {h}" for i, h in enumerate(tgt_entry.headers)] if tgt_entry.headers else []
        col_cb = ttk.Combobox(row_frame, state="readonly", width=25, values=col_display)
        col_cb.pack(side="left", padx=(0, 5))

        # Source selector combobox
        src_cb = ttk.Combobox(row_frame, state="readonly", width=30,
                              values=self._source_selector_values)
        src_cb.pack(side="left", padx=(0, 5))

        # [X] remove button
        ttk.Button(
            row_frame, text="X", width=2,
            command=lambda m=mapping, f=row_frame: self._remove_write_mapping(m, f),
        ).pack(side="left")

        # Set values
        self._ignore_events = True
        if mapping.col_name and mapping.col_name in col_display:
            col_cb.set(mapping.col_name)
        src_display = self._source_key_to_display(mapping.source_key)
        src_cb.set(src_display)
        self._ignore_events = False

        # Bind events
        col_cb.bind("<<ComboboxSelected>>",
                     lambda e, m=mapping, cb=col_cb: self._on_wm_col_select(m, cb))
        src_cb.bind("<<ComboboxSelected>>",
                     lambda e, m=mapping, cb=src_cb: self._on_wm_src_select(m, cb))

        self._mapping_widgets[id(mapping)] = (row_frame, col_cb, src_cb, num_label)

    def _remove_write_mapping(self, mapping: WriteMapping, frame: tk.Widget):
        """Remove a specific write mapping row."""
        tgt_entry = self._get_selected_entry(False)
        if tgt_entry and mapping in tgt_entry.write_mappings:
            tgt_entry.write_mappings.remove(mapping)

        mid = id(mapping)
        if mid in self._mapping_widgets:
            del self._mapping_widgets[mid]

        frame.destroy()
        self._renumber_mapping_rows()
        self._update_status(False)

    def _renumber_mapping_rows(self):
        """Update row numbers after removal."""
        for i, (_mid, (_frame, _col_cb, _src_cb, num_label)) in enumerate(self._mapping_widgets.items()):
            num_label.config(text=str(i + 1))

    def _clear_mapping_widgets(self):
        """Destroy all write mapping row widgets."""
        for _mid, (frame, _col_cb, _src_cb, _lbl) in self._mapping_widgets.items():
            frame.destroy()
        self._mapping_widgets.clear()

    def _rebuild_mapping_widgets(self, entry: FileEntry):
        """Rebuild write mapping widgets for a target entry."""
        self._clear_mapping_widgets()
        for mapping in entry.write_mappings:
            self._add_write_mapping(mapping)

    def _on_wm_col_select(self, mapping: WriteMapping, cb: ttk.Combobox):
        if self._ignore_events:
            return
        val = cb.get()
        if val:
            mapping.col_name = val
            try:
                mapping.col_idx = int(val.split(":")[0])
            except (ValueError, IndexError):
                pass
        self._update_status(False)

    def _on_wm_src_select(self, mapping: WriteMapping, cb: ttk.Combobox):
        if self._ignore_events:
            return
        val = cb.get()
        mapping.source_key = self._source_display_to_key.get(val, ALL_MERGED)

    # -----------------------------------------------------------------
    # FILE MANAGEMENT
    # -----------------------------------------------------------------

    def _get_entries(self, is_source: bool) -> List[FileEntry]:
        return self.source_entries if is_source else self.target_entries

    def _get_listbox(self, is_source: bool) -> tk.Listbox:
        return self.src_lb if is_source else self.tgt_lb

    def _get_status(self, is_source: bool) -> tk.StringVar:
        return self.src_status if is_source else self.tgt_status

    def _get_current_idx(self, is_source: bool) -> int:
        return self._current_src_idx if is_source else self._current_tgt_idx

    def _set_current_idx(self, is_source: bool, idx: int):
        if is_source:
            self._current_src_idx = idx
        else:
            self._current_tgt_idx = idx

    def _update_listbox(self, is_source: bool):
        """Refresh listbox display from entries."""
        if is_source:
            self._update_source_listbox()
        else:
            lb = self.tgt_lb
            lb.delete(0, "end")
            for entry in self.target_entries:
                lb.insert("end", entry.display)
            self._update_status(False)

    def _update_status(self, is_source: bool):
        entries = self._get_entries(is_source)
        status = self._get_status(is_source)
        n = len(entries)
        if is_source:
            configured = sum(1 for e in entries if e.col1_idx >= 0 and e.col2_idx >= 0)
        else:
            configured = sum(
                1 for e in entries
                if e.col1_idx >= 0
                and len(e.write_mappings) > 0
                and all(wm.col_idx >= 0 for wm in e.write_mappings)
            )
        if n == 0:
            status.set("No files")
        else:
            status.set(f"{n} file(s), {configured} configured")

    def _save_current_config(self, is_source: bool):
        """Save current combobox values back to the previously-tracked FileEntry."""
        idx = self._get_current_idx(is_source)
        entries = self._get_entries(is_source)
        if idx < 0 or idx >= len(entries):
            return
        entry = entries[idx]

        if is_source:
            sheet_cb = self.src_sheet_cb
            key_cb = self.src_key_cb
            key2_cb = self.src_key2_cb
        else:
            sheet_cb = self.tgt_sheet_cb
            key_cb = self.tgt_key_cb
            key2_cb = self.tgt_key2_cb

        # Save sheet
        sheet_val = sheet_cb.get()
        if sheet_val and sheet_val in entry.sheets:
            entry.sheet = sheet_val

        # Save key column
        key_val = key_cb.get()
        if key_val:
            entry.col1 = key_val
            try:
                entry.col1_idx = int(key_val.split(":")[0])
            except (ValueError, IndexError):
                pass

        # Save key col 2
        key2_val = key2_cb.get()
        if key2_val:
            entry.col1b = key2_val
            try:
                entry.col1b_idx = int(key2_val.split(":")[0])
            except (ValueError, IndexError):
                pass
        else:
            entry.col1b = ""
            entry.col1b_idx = -1

        if is_source:
            # Save value column
            val_val = self.src_val_cb.get()
            if val_val:
                entry.col2 = val_val
                try:
                    entry.col2_idx = int(val_val.split(":")[0])
                except (ValueError, IndexError):
                    pass
        else:
            # Save write mappings from widgets
            for mapping in entry.write_mappings:
                mid = id(mapping)
                if mid in self._mapping_widgets:
                    _frame, col_cb, src_cb, _lbl = self._mapping_widgets[mid]
                    col_val = col_cb.get()
                    if col_val:
                        mapping.col_name = col_val
                        try:
                            mapping.col_idx = int(col_val.split(":")[0])
                        except (ValueError, IndexError):
                            pass
                    src_val = src_cb.get()
                    mapping.source_key = self._source_display_to_key.get(src_val, ALL_MERGED)

    def _add_files(self, is_source: bool):
        paths = filedialog.askopenfilenames(
            title="Select Excel files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not paths:
            return

        self._save_current_config(is_source)

        entries = self._get_entries(is_source)
        existing_paths = {e.path for e in entries}

        for p in paths:
            if p in existing_paths:
                continue
            entry = FileEntry(p)
            entry.sheets = read_sheets(Path(p))
            if entry.sheets:
                entry.sheet = entry.sheets[0]
                entry.headers = read_headers(Path(p), entry.sheet)
            entries.append(entry)

        self._update_listbox(is_source)

        if is_source:
            self._rebuild_source_selectors()

        # Auto-select last added
        lb = self._get_listbox(is_source)
        if entries:
            lb.selection_clear(0, "end")
            lb.selection_set(len(entries) - 1)
            self._set_current_idx(is_source, len(entries) - 1)
            self._on_listbox_select(is_source)

    def _remove_selected(self, is_source: bool):
        lb = self._get_listbox(is_source)
        sel = lb.curselection()
        if not sel:
            return
        entries = self._get_entries(is_source)
        idx = sel[0]
        entries.pop(idx)
        self._set_current_idx(is_source, -1)
        self._update_listbox(is_source)
        self._clear_config(is_source)

        if is_source:
            self._rebuild_source_selectors()

    def _clear_files(self, is_source: bool):
        entries = self._get_entries(is_source)
        entries.clear()
        self._set_current_idx(is_source, -1)
        self._update_listbox(is_source)
        self._clear_config(is_source)

        if is_source:
            self._rebuild_source_selectors()

    def _clear_config(self, is_source: bool):
        """Clear the config comboboxes."""
        if is_source:
            self.src_sheet_cb.set("")
            self.src_sheet_cb["values"] = []
            self.src_key_cb.set("")
            self.src_key_cb["values"] = []
            self.src_key2_cb.set("")
            self.src_key2_cb["values"] = []
            self.src_val_cb.set("")
            self.src_val_cb["values"] = []
        else:
            self.tgt_sheet_cb.set("")
            self.tgt_sheet_cb["values"] = []
            self.tgt_key_cb.set("")
            self.tgt_key_cb["values"] = []
            self.tgt_key2_cb.set("")
            self.tgt_key2_cb["values"] = []
            self._clear_mapping_widgets()

    def _clear_key2(self, is_source: bool):
        """Clear the KEY Col 2 combobox and reset entry."""
        entry = self._get_selected_entry(is_source)
        if entry:
            entry.col1b = ""
            entry.col1b_idx = -1
        key2_cb = self.src_key2_cb if is_source else self.tgt_key2_cb
        key2_cb.set("")

    # -----------------------------------------------------------------
    # SELECTION EVENTS
    # -----------------------------------------------------------------

    def _get_selected_entry(self, is_source: bool) -> Optional[FileEntry]:
        lb = self._get_listbox(is_source)
        sel = lb.curselection()
        if not sel:
            return None
        entries = self._get_entries(is_source)
        idx = sel[0]
        if idx < len(entries):
            return entries[idx]
        return None

    def _on_listbox_select(self, is_source: bool):
        if self._ignore_events:
            return

        lb = self._get_listbox(is_source)
        sel = lb.curselection()
        if not sel:
            return
        new_idx = sel[0]
        entries = self._get_entries(is_source)
        if new_idx >= len(entries):
            return

        # Save previous config before switching
        self._save_current_config(is_source)

        entry = entries[new_idx]
        self._set_current_idx(is_source, new_idx)

        if is_source:
            sheet_cb = self.src_sheet_cb
            key_cb = self.src_key_cb
            key2_cb = self.src_key2_cb
        else:
            sheet_cb = self.tgt_sheet_cb
            key_cb = self.tgt_key_cb
            key2_cb = self.tgt_key2_cb

        self._ignore_events = True

        # Populate sheet
        sheet_cb["values"] = entry.sheets
        if entry.sheet:
            sheet_cb.set(entry.sheet)
        elif entry.sheets:
            sheet_cb.set(entry.sheets[0])

        # Populate column comboboxes
        col_display = [f"{i}: {h}" for i, h in enumerate(entry.headers)] if entry.headers else []
        key_cb["values"] = col_display
        key2_cb["values"] = col_display

        # Restore KEY
        if entry.col1 and entry.col1 in col_display:
            key_cb.set(entry.col1)
        else:
            key_cb.set("")

        # Restore KEY Col 2
        if entry.col1b and entry.col1b in col_display:
            key2_cb.set(entry.col1b)
        else:
            key2_cb.set("")

        if is_source:
            # Restore VALUE column
            self.src_val_cb["values"] = col_display
            if entry.col2 and entry.col2 in col_display:
                self.src_val_cb.set(entry.col2)
            else:
                self.src_val_cb.set("")
        else:
            # Rebuild write mapping widgets for this target
            self._rebuild_mapping_widgets(entry)

        self._ignore_events = False

    def _on_sheet_select(self, is_source: bool):
        if self._ignore_events:
            return

        entry = self._get_selected_entry(is_source)
        if not entry:
            return

        if is_source:
            sheet_cb = self.src_sheet_cb
            key_cb = self.src_key_cb
            key2_cb = self.src_key2_cb
        else:
            sheet_cb = self.tgt_sheet_cb
            key_cb = self.tgt_key_cb
            key2_cb = self.tgt_key2_cb

        new_sheet = sheet_cb.get()
        if new_sheet == entry.sheet:
            return

        entry.sheet = new_sheet
        entry.headers = read_headers(Path(entry.path), new_sheet)
        entry.col1 = ""
        entry.col1_idx = -1
        entry.col1b = ""
        entry.col1b_idx = -1

        col_display = [f"{i}: {h}" for i, h in enumerate(entry.headers)] if entry.headers else []
        key_cb["values"] = col_display
        key_cb.set("")
        key2_cb["values"] = col_display
        key2_cb.set("")

        if is_source:
            entry.col2 = ""
            entry.col2_idx = -1
            self.src_val_cb["values"] = col_display
            self.src_val_cb.set("")
            # Source key changed — rebuild source selectors
            self._rebuild_source_selectors()
        else:
            # Clear write mappings — headers changed, old columns invalid
            entry.write_mappings.clear()
            self._clear_mapping_widgets()

        # Targeted listbox single-item update
        lb = self._get_listbox(is_source)
        sel = lb.curselection()
        if sel:
            idx = sel[0]
            self._ignore_events = True
            lb.delete(idx)
            if is_source:
                lb.insert(idx, self._numbered_source_display(idx, entry))
            else:
                lb.insert(idx, entry.display)
            lb.selection_set(idx)
            self._ignore_events = False

    def _on_col_select(self, is_source: bool, is_key: bool = False, is_key2: bool = False):
        if self._ignore_events:
            return

        entry = self._get_selected_entry(is_source)
        if not entry:
            return

        if is_key2:
            cb = self.src_key2_cb if is_source else self.tgt_key2_cb
        elif is_key:
            cb = self.src_key_cb if is_source else self.tgt_key_cb
        else:
            # VALUE column — source only, targets use write mappings
            if not is_source:
                return
            cb = self.src_val_cb

        val = cb.get()
        if not val:
            return

        try:
            col_idx = int(val.split(":")[0])
        except (ValueError, IndexError):
            return

        if is_key2:
            entry.col1b = val
            entry.col1b_idx = col_idx
        elif is_key:
            entry.col1 = val
            entry.col1_idx = col_idx
        else:
            entry.col2 = val
            entry.col2_idx = col_idx

        self._update_status(is_source)

    # -----------------------------------------------------------------
    # LOGGING
    # -----------------------------------------------------------------

    def _log(self, msg: str, tag: str = "info"):
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.root.update_idletasks()

    # -----------------------------------------------------------------
    # TRANSFER LOGIC
    # -----------------------------------------------------------------

    def _run_transfer(self):
        # Save current configs
        self._save_current_config(True)
        self._save_current_config(False)

        self.log.delete("1.0", "end")

        # Validate sources
        if not self.source_entries:
            messagebox.showerror("Error", "No SOURCE files added.")
            return

        unconfigured_src = [
            e for e in self.source_entries if e.col1_idx < 0 or e.col2_idx < 0
        ]
        if unconfigured_src:
            names = ", ".join(Path(e.path).name for e in unconfigured_src)
            messagebox.showerror(
                "Error",
                f"Source files not fully configured (select Sheet + KEY + VALUE):\n{names}",
            )
            return

        # Validate targets
        if not self.target_entries:
            messagebox.showerror("Error", "No TARGET files added.")
            return

        problems = []
        for e in self.target_entries:
            name = Path(e.path).name
            if e.col1_idx < 0:
                problems.append(f"{name}: KEY column not set")
            elif len(e.write_mappings) == 0:
                problems.append(f"{name}: No write mappings")
            else:
                bad = [wm for wm in e.write_mappings if wm.col_idx < 0]
                if bad:
                    problems.append(f"{name}: {len(bad)} write mapping(s) with no column selected")

        if problems:
            messagebox.showerror(
                "Error",
                "Target files not fully configured:\n" + "\n".join(problems),
            )
            return

        # Warn about duplicate write columns (last-wins)
        for e in self.target_entries:
            col_counts: Dict[int, int] = {}
            for wm in e.write_mappings:
                if wm.col_idx >= 0:
                    col_counts[wm.col_idx] = col_counts.get(wm.col_idx, 0) + 1
            dupes = [idx for idx, cnt in col_counts.items() if cnt > 1]
            if dupes:
                name = Path(e.path).name
                cols = ", ".join(chr(65 + c) if c < 26 else f"Col{c}" for c in dupes)
                self._log(f"WARNING: {name} has duplicate write column(s): {cols} — last mapping wins", "warning")

        save_mode = self.save_mode_cb.get() if self.save_mode_cb else SAVE_MODES[0]

        self.transfer_btn.config(state="disabled")

        try:
            self._execute_transfer(save_mode)
        except Exception as e:
            self._log(f"\nERROR: {e}", "error")
            logger.exception("MultiLookup transfer failed")
        finally:
            self.transfer_btn.config(state="normal")
            self._persist_settings()

    def _execute_transfer(self, save_mode: str):
        t0 = time.time()

        self._log("=" * 60, "header")
        self._log("  MULTILOOKUP TRANSFER v1.2", "header")
        self._log("=" * 60, "header")

        # Step 1: Build per-source dictionaries
        self._log(f"\n{'─' * 60}", "info")
        self._log("BUILDING SOURCE DICTIONARIES...", "header")
        self._log(f"{'─' * 60}", "info")

        src_dicts_input = [e.to_source_dict() for e in self.source_entries]
        source_dicts, dup_count = build_source_dicts(src_dicts_input, log_fn=self._log)

        if not source_dicts:
            self._log("\nERROR: No source dictionaries built — nothing to transfer.", "error")
            return

        total_keys = sum(len(d) for d in source_dicts.values())
        self._log(f"\n  {len(source_dicts)} source dict(s), {total_keys:,} total keys", "success")
        if dup_count > 0:
            self._log(
                f"  WARNING: {dup_count:,} duplicate keys (first value wins per source)",
                "warning",
            )

        # Build merged dict only if needed
        needs_merged = any(
            wm.source_key == ALL_MERGED
            for e in self.target_entries
            for wm in e.write_mappings
        )
        merged_dict = build_merged_dict(source_dicts) if needs_merged else None
        if merged_dict is not None:
            self._log(f"  Merged dict: {len(merged_dict):,} unique keys", "info")

        # Step 2: Transfer to targets
        self._log(f"\n{'─' * 60}", "info")
        self._log(f"TRANSFERRING TO TARGETS ({save_mode})...", "header")
        self._log(f"{'─' * 60}", "info")

        tgt_dicts = [e.to_target_dict() for e in self.target_entries]
        total_matches, total_rows, files_saved = transfer_to_targets(
            source_dicts, merged_dict, tgt_dicts, save_mode, log_fn=self._log,
        )

        # Step 3: Summary
        elapsed = time.time() - t0
        self._log(f"\n{'=' * 60}", "header")
        self._log("  DONE", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Sources: {len(self.source_entries)} file(s), {total_keys:,} keys", "info")
        self._log(f"  Targets: {total_matches:,}/{total_rows:,} writes across {len(self.target_entries)} file(s)", "info")
        self._log(f"  Files saved: {files_saved}", "success")
        self._log(f"  Time: {elapsed:.1f}s", "info")

        if total_matches == 0:
            self._log(
                "\n  WARNING: 0 matches — check KEY columns and source routing.",
                "warning",
            )

    # -----------------------------------------------------------------
    # SETTINGS
    # -----------------------------------------------------------------

    def _persist_settings(self):
        """Save current file paths and column selections."""
        self._save_current_config(True)
        self._save_current_config(False)

        settings = _load_settings()

        src_list = []
        for e in self.source_entries:
            src_list.append({
                "path": e.path,
                "sheet": e.sheet,
                "col1": e.col1,
                "col2": e.col2,
                "col1b": e.col1b,
            })

        tgt_list = []
        for e in self.target_entries:
            tgt_list.append({
                "path": e.path,
                "sheet": e.sheet,
                "col1": e.col1,
                "col1b": e.col1b,
                "write_mappings": [
                    {"col_name": wm.col_name, "source_key": wm.source_key}
                    for wm in e.write_mappings
                ],
            })

        settings["source_entries"] = src_list
        settings["target_entries"] = tgt_list
        if self.save_mode_cb:
            settings["save_mode"] = self.save_mode_cb.get()

        _save_settings(settings)

    def _load_persisted_settings(self):
        """Restore file entries from saved settings."""
        settings = _load_settings()

        if self.save_mode_cb and "save_mode" in settings:
            mode = settings["save_mode"]
            if mode in SAVE_MODES:
                self.save_mode_cb.set(mode)

        # Load sources first (needed for source selector rebuild)
        for item in settings.get("source_entries", []):
            p = item.get("path", "")
            if not p or not Path(p).exists():
                continue
            entry = FileEntry(p)
            entry.sheets = read_sheets(Path(p))

            sheet = item.get("sheet", "")
            if sheet and sheet in entry.sheets:
                entry.sheet = sheet
            elif entry.sheets:
                entry.sheet = entry.sheets[0]

            if entry.sheet:
                entry.headers = read_headers(Path(p), entry.sheet)

            col_display = [f"{i}: {h}" for i, h in enumerate(entry.headers)]

            col1 = item.get("col1", "")
            if col1 and col1 in col_display:
                entry.col1 = col1
                try:
                    entry.col1_idx = int(col1.split(":")[0])
                except (ValueError, IndexError):
                    pass

            col2 = item.get("col2", "")
            if col2 and col2 in col_display:
                entry.col2 = col2
                try:
                    entry.col2_idx = int(col2.split(":")[0])
                except (ValueError, IndexError):
                    pass

            col1b = item.get("col1b", "")
            if col1b and col1b in col_display:
                entry.col1b = col1b
                try:
                    entry.col1b_idx = int(col1b.split(":")[0])
                except (ValueError, IndexError):
                    pass

            self.source_entries.append(entry)

        self._update_listbox(True)
        self._rebuild_source_selectors()

        # Load targets
        for item in settings.get("target_entries", []):
            p = item.get("path", "")
            if not p or not Path(p).exists():
                continue
            entry = FileEntry(p)
            entry.sheets = read_sheets(Path(p))

            sheet = item.get("sheet", "")
            if sheet and sheet in entry.sheets:
                entry.sheet = sheet
            elif entry.sheets:
                entry.sheet = entry.sheets[0]

            if entry.sheet:
                entry.headers = read_headers(Path(p), entry.sheet)

            col_display = [f"{i}: {h}" for i, h in enumerate(entry.headers)]

            col1 = item.get("col1", "")
            if col1 and col1 in col_display:
                entry.col1 = col1
                try:
                    entry.col1_idx = int(col1.split(":")[0])
                except (ValueError, IndexError):
                    pass

            col1b = item.get("col1b", "")
            if col1b and col1b in col_display:
                entry.col1b = col1b
                try:
                    entry.col1b_idx = int(col1b.split(":")[0])
                except (ValueError, IndexError):
                    pass

            # Load write mappings (v1.2 format)
            if "write_mappings" in item:
                for wm_data in item["write_mappings"]:
                    wm = WriteMapping()
                    col_name = wm_data.get("col_name", "")
                    wm.source_key = wm_data.get("source_key", ALL_MERGED)
                    if col_name and col_name in col_display:
                        wm.col_name = col_name
                        try:
                            wm.col_idx = int(col_name.split(":")[0])
                        except (ValueError, IndexError):
                            pass
                        entry.write_mappings.append(wm)
            elif "col2" in item and item["col2"]:
                # Backward compat: v1.1 format — migrate col2 → single WriteMapping
                col2 = item["col2"]
                if col2 in col_display:
                    wm = WriteMapping()
                    wm.col_name = col2
                    wm.source_key = ALL_MERGED
                    try:
                        wm.col_idx = int(col2.split(":")[0])
                    except (ValueError, IndexError):
                        pass
                    entry.write_mappings.append(wm)

            self.target_entries.append(entry)

        self._update_listbox(False)

    def run(self):
        self.root.mainloop()


# =============================================================================
# MAIN
# =============================================================================

def main():
    app = MultiLookupApp()
    app.run()


if __name__ == "__main__":
    main()
