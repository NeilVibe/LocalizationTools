#!/usr/bin/env python3
# coding: utf-8
"""
Script Long String Extractor v2.0
==================================
Standalone GUI tool that extracts LocStr entries from XML/Excel files
that are SCRIPT TYPE (Dialog/Sequencer) AND above a character length threshold.

Inputs:
  - Export Folder: export__ folder to build StringID -> Category mapping
  - Source Folder: Folder with languagedata_*.xml or .xlsx files to extract from
  - Min Length: Minimum character count threshold for Str content

Output (per language, in script directory):
  - Excel: StringID, StrOrigin, Str, CharCount
  - XML: Raw LocStr elements under <root>

Usage: python script_long_string_extractor.py
"""

import os
import re
import sys
import html
import logging
import datetime as _dt
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Try lxml first (more robust), fallback to standard library
try:
    from lxml import etree
    USING_LXML = True
except ImportError:
    from xml.etree import ElementTree as etree
    USING_LXML = False

# Try openpyxl for reading Excel
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Try xlsxwriter for writing Excel
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

# =============================================================================
# LOGGING
# =============================================================================

logger = logging.getLogger("ScriptLongStringExtractor")
logger.setLevel(logging.DEBUG)

# =============================================================================
# CONFIGURATION
# =============================================================================

WINDOW_TITLE = "Script Long String Extractor v2.0"
WINDOW_WIDTH = 820
WINDOW_HEIGHT = 720
ENTRY_WIDTH = 55

SCRIPT_CATEGORIES = {"Sequencer", "Dialog"}

# Subfolders to EXCLUDE from extraction (case-insensitive)
EXCLUDE_SUBFOLDERS = {"narrationdialog"}

# Detect if running as PyInstaller bundle
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent

OUTPUT_DIR = SCRIPT_DIR / "Extraction_Output"

# LocStr tag and attribute variants
LOCSTR_TAGS = ['LocStr', 'locstr', 'LOCSTR', 'LOCStr', 'Locstr']
STRINGID_ATTRS = ('StringId', 'StringID', 'stringid', 'STRINGID', 'Stringid', 'stringId')
STRORIGIN_ATTRS = ('StrOrigin', 'Strorigin', 'strorigin', 'STRORIGIN')
STR_ATTRS = ('Str', 'str', 'STR')

# Excel header variants
STRINGID_HEADERS = {'stringid', 'string_id', 'string id'}
STRORIGIN_HEADERS = {'strorigin', 'str_origin', 'str origin'}
STR_HEADERS = {'str'}

DEFAULT_MIN_LENGTH = 50


# =============================================================================
# HELPERS
# =============================================================================

def get_attr(elem, attr_variants):
    """Get attribute value from element, trying multiple case variants."""
    for attr_name in attr_variants:
        val = elem.get(attr_name)
        if val is not None:
            return attr_name, val
    return None, None


def visible_char_count(text: str) -> int:
    """Count visible characters: strip markup tags, HTML entities, code blocks."""
    if not text:
        return 0
    t = text
    # Remove <br/> newline tags
    t = re.sub(r'<br\s*/?>', '', t, flags=re.IGNORECASE)
    # Remove Scale tags
    t = re.sub(r'<Scale[^>]*>|</Scale>', '', t)
    # Remove color tags
    t = re.sub(r'<color[^>]*>|</color>', '', t)
    # Remove PAColor and PAOldColor tags
    t = re.sub(r'<PAColor[^>]*>|<PAOldColor>', '', t)
    # Remove Style tags
    t = re.sub(r'<Style:[^>]*>', '', t)
    # Remove literal \n
    t = re.sub(r'\\n', '', t)
    # Remove content in curly braces (code blocks)
    t = re.sub(r'\{[^}]*\}', '', t)
    # Unescape HTML entities
    t = html.unescape(t)
    # Collapse whitespace and strip
    t = re.sub(r'\s+', ' ', t)
    t = t.strip()
    return len(t)


def extract_language_from_filename(filename: str) -> str:
    """Extract language code from languagedata_*.xml filename."""
    m = re.match(r'languagedata_(.+)\.xml', filename, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return ""


# =============================================================================
# CATEGORY MAPPING (from export folder)
# =============================================================================

def build_stringid_to_category(export_folder: Path, progress_fn=None) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Build StringID -> Category and StringID -> Subfolder mappings from export folder.

    Returns:
        (category_map, subfolder_map) where both are {stringid: name}
    """
    category_map = {}
    subfolder_map = {}
    if not export_folder.exists():
        return category_map, subfolder_map

    categories = [d for d in export_folder.iterdir() if d.is_dir()]

    for cat_folder in categories:
        cat_name = cat_folder.name
        xml_files = list(cat_folder.rglob("*.loc.xml"))

        if progress_fn:
            progress_fn(f"Indexing {cat_name} ({len(xml_files)} files)...")

        for xml_file in xml_files:
            # Determine subfolder relative to category folder
            rel = xml_file.relative_to(cat_folder)
            subfolder = rel.parts[0] if len(rel.parts) > 1 else ""

            try:
                root = _parse_xml_root(xml_file)
                for elem in _iter_locstr(root):
                    _, sid = get_attr(elem, STRINGID_ATTRS)
                    if sid and sid.strip():
                        sid_clean = sid.strip()
                        category_map[sid_clean] = cat_name
                        subfolder_map[sid_clean] = subfolder
            except Exception as e:
                logger.warning("Failed to parse %s: %s", xml_file, e)
                continue

    return category_map, subfolder_map


# =============================================================================
# XML PARSING
# =============================================================================

def _parse_xml_root(xml_path: Path):
    """Parse XML, return root element."""
    if USING_LXML:
        parser = etree.XMLParser(
            resolve_entities=False, load_dtd=False,
            no_network=True, recover=True,
        )
        tree = etree.parse(str(xml_path), parser)
    else:
        tree = etree.parse(str(xml_path))
    return tree.getroot()


def _iter_locstr(root):
    """Iterate all LocStr elements (case-insensitive tag matching)."""
    elements = []
    for tag in LOCSTR_TAGS:
        elements.extend(root.iter(tag))
    return elements


# =============================================================================
# EXTRACTION LOGIC
# =============================================================================

def extract_from_xml(
    xml_path: Path,
    ci_category: Dict[str, str],
    ci_subfolder: Dict[str, str],
    min_length: int,
) -> List[Dict]:
    """
    Extract LocStr entries from XML that are SCRIPT type + above length threshold.

    Excludes entries in EXCLUDE_SUBFOLDERS (e.g. NarrationDialog).

    Returns list of dicts: {string_id, str_origin, str_value, category, char_count, raw_attribs, language}
    """
    results = []
    language = extract_language_from_filename(xml_path.name)
    root = _parse_xml_root(xml_path)

    for elem in _iter_locstr(root):
        _, sid = get_attr(elem, STRINGID_ATTRS)
        _, so = get_attr(elem, STRORIGIN_ATTRS)
        _, sv = get_attr(elem, STR_ATTRS)

        sid = (sid or "").strip()
        so = (so or "").strip()
        sv = (sv or "").strip()

        if not sid or not sv:
            continue

        sid_lower = sid.lower()

        # Check SCRIPT category
        category = ci_category.get(sid_lower, "Uncategorized")
        if category not in SCRIPT_CATEGORIES:
            continue

        # Check excluded subfolders (e.g. NarrationDialog)
        subfolder = ci_subfolder.get(sid_lower, "")
        if subfolder.lower() in EXCLUDE_SUBFOLDERS:
            continue

        # Check character length
        char_count = visible_char_count(sv)
        if char_count < min_length:
            continue

        results.append({
            "string_id": sid,
            "str_origin": so,
            "str_value": sv,
            "category": category,
            "char_count": char_count,
            "raw_attribs": dict(elem.attrib),
            "language": language,
        })

    return results


def extract_from_excel(
    xlsx_path: Path,
    ci_category: Dict[str, str],
    ci_subfolder: Dict[str, str],
    min_length: int,
) -> List[Dict]:
    """
    Extract entries from Excel that are SCRIPT type + above length threshold.

    Excludes entries in EXCLUDE_SUBFOLDERS (e.g. NarrationDialog).
    Expects columns: StringID, StrOrigin, Str
    Language determined from filename if possible, otherwise "EXCEL".
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for Excel: pip install openpyxl")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    results = []

    # Try to extract language from filename (e.g. corrections_ENG.xlsx)
    lang_match = re.search(r'[_-]([a-zA-Z]{2,6})\.xlsx$', xlsx_path.name, re.IGNORECASE)
    language = lang_match.group(1).upper() if lang_match else "EXCEL"

    # Detect headers
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False))):
        val = str(cell.value or "").strip().lower()
        if val in STRINGID_HEADERS:
            headers['stringid'] = col_idx
        elif val in STRORIGIN_HEADERS:
            headers['strorigin'] = col_idx
        elif val in STR_HEADERS:
            headers['str'] = col_idx

    if 'stringid' not in headers:
        wb.close()
        raise ValueError(f"No StringID column found in {xlsx_path.name}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[headers['stringid']] or "").strip() if 'stringid' in headers else ""
        so = str(row[headers['strorigin']] or "").strip() if 'strorigin' in headers else ""
        sv = str(row[headers['str']] or "").strip() if 'str' in headers else ""

        if not sid or not sv:
            continue

        sid_lower = sid.lower()

        category = ci_category.get(sid_lower, "Uncategorized")
        if category not in SCRIPT_CATEGORIES:
            continue

        # Check excluded subfolders (e.g. NarrationDialog)
        subfolder = ci_subfolder.get(sid_lower, "")
        if subfolder.lower() in EXCLUDE_SUBFOLDERS:
            continue

        char_count = visible_char_count(sv)
        if char_count < min_length:
            continue

        results.append({
            "string_id": sid,
            "str_origin": so,
            "str_value": sv,
            "category": category,
            "char_count": char_count,
            "raw_attribs": {"StringID": sid, "StrOrigin": so, "Str": sv},
            "language": language,
        })

    wb.close()
    return results


# =============================================================================
# REPORT GENERATION
# =============================================================================

def write_excel_report(entries: List[Dict], output_path: Path) -> bool:
    """Write extraction results to Excel (StringID, StrOrigin, Str, CharCount)."""
    if not HAS_XLSXWRITER:
        logger.warning("xlsxwriter not available")
        return False

    if not entries:
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = xlsxwriter.Workbook(str(output_path))
    ws = wb.add_worksheet("SCRIPT Long Strings")

    # Formats
    header_fmt = wb.add_format({
        'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#2E4057',
        'border': 1, 'text_wrap': True,
    })
    cell_fmt = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    num_fmt = wb.add_format({'border': 1, 'align': 'center', 'valign': 'top'})

    headers = ["StringID", "StrOrigin", "Str", "CharCount"]
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_fmt)

    # Sort by char_count descending
    entries_sorted = sorted(entries, key=lambda e: e["char_count"], reverse=True)

    for row, e in enumerate(entries_sorted, 1):
        ws.write(row, 0, e["string_id"], cell_fmt)
        ws.write(row, 1, e["str_origin"], cell_fmt)
        ws.write(row, 2, e["str_value"], cell_fmt)
        ws.write(row, 3, e["char_count"], num_fmt)

    # Column widths
    ws.set_column(0, 0, 35)   # StringID
    ws.set_column(1, 1, 45)   # StrOrigin
    ws.set_column(2, 2, 60)   # Str
    ws.set_column(3, 3, 12)   # CharCount

    ws.autofilter(0, 0, len(entries_sorted), len(headers) - 1)
    ws.freeze_panes(1, 0)

    wb.close()
    return True


def write_xml_report(entries: List[Dict], output_path: Path) -> bool:
    """Write raw LocStr elements under <root> to XML file."""
    if not entries:
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = ['<?xml version="1.0" encoding="utf-8"?>', '<root>']

    for e in sorted(entries, key=lambda x: x["char_count"], reverse=True):
        attribs = e.get("raw_attribs", {})
        if not attribs:
            # Fallback: build from entry data
            attribs = {"StringID": e["string_id"], "StrOrigin": e["str_origin"], "Str": e["str_value"]}

        # Build attribute string preserving original attribute names/values
        attr_parts = []
        for k, v in attribs.items():
            # Escape XML special chars in attribute values
            escaped = str(v).replace('&', '&amp;').replace('"', '&quot;').replace('<', '&lt;').replace('>', '&gt;')
            attr_parts.append(f'{k}="{escaped}"')

        lines.append(f'  <LocStr {" ".join(attr_parts)} />')

    lines.append('</root>')
    lines.append('')  # trailing newline

    output_path.write_text('\n'.join(lines), encoding='utf-8')
    return True


# =============================================================================
# GUI APPLICATION
# =============================================================================

class ScriptLongStringExtractorGUI:
    """GUI for extracting SCRIPT-type long strings from XML/Excel."""

    def __init__(self, root):
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.resizable(True, True)

        self.export_var = tk.StringVar()
        self.source_var = tk.StringVar()
        self.min_length_var = tk.IntVar(value=DEFAULT_MIN_LENGTH)

        self._build_ui()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        # Title
        ttk.Label(main, text="Script Long String Extractor", font=("Segoe UI", 14, "bold")).pack(pady=(0, 5))
        ttk.Label(
            main,
            text="Extract SCRIPT-type (Dialog/Sequencer) entries above a character length threshold",
            font=("Segoe UI", 9),
        ).pack(pady=(0, 10))

        # Export folder frame
        exp_frame = ttk.LabelFrame(main, text="Export Folder (export__ — for SCRIPT category detection)", padding=5)
        exp_frame.pack(fill=tk.X, pady=5)

        exp_row = ttk.Frame(exp_frame)
        exp_row.pack(fill=tk.X)
        ttk.Entry(exp_row, textvariable=self.export_var, width=ENTRY_WIDTH).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(exp_row, text="Browse Folder", width=15, command=self._browse_export).pack(side=tk.RIGHT, padx=(5, 0))

        # Source folder frame
        src_frame = ttk.LabelFrame(main, text="Source Folder (XML or Excel files to extract from)", padding=5)
        src_frame.pack(fill=tk.X, pady=5)

        src_row = ttk.Frame(src_frame)
        src_row.pack(fill=tk.X)
        ttk.Entry(src_row, textvariable=self.source_var, width=ENTRY_WIDTH).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(src_row, text="Browse Folder", width=15, command=self._browse_source).pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Label(src_frame, text="Accepts: languagedata_*.xml or .xlsx with StringID + StrOrigin + Str columns", font=("Segoe UI", 8)).pack(anchor=tk.W, pady=(3, 0))

        # Settings frame
        settings_frame = ttk.LabelFrame(main, text="Settings", padding=5)
        settings_frame.pack(fill=tk.X, pady=5)

        len_row = ttk.Frame(settings_frame)
        len_row.pack(fill=tk.X)
        ttk.Label(len_row, text="Minimum character length:").pack(side=tk.LEFT)
        self.length_spin = ttk.Spinbox(
            len_row, from_=1, to=10000, textvariable=self.min_length_var, width=8,
        )
        self.length_spin.pack(side=tk.LEFT, padx=(5, 10))
        ttk.Label(len_row, text="(only entries with Str >= this many visible chars)", font=("Segoe UI", 8)).pack(side=tk.LEFT)

        # Output info
        out_frame = ttk.LabelFrame(main, text="Output", padding=5)
        out_frame.pack(fill=tk.X, pady=5)
        ttk.Label(out_frame, text=f"Output directory: {OUTPUT_DIR}", font=("Segoe UI", 8)).pack(anchor=tk.W)
        ttk.Label(out_frame, text="Per language: one Excel (.xlsx) + one XML (.xml)", font=("Segoe UI", 8)).pack(anchor=tk.W)

        # Action buttons
        btn_frame = ttk.Frame(main)
        btn_frame.pack(pady=10)

        self.extract_btn = ttk.Button(
            btn_frame, text="EXTRACT LONG SCRIPT STRINGS", width=35,
            command=self._run_extract,
        )
        self.extract_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(btn_frame, text="Clear All", width=12, command=self._clear_all).pack(side=tk.LEFT, padx=5)

        # Log area
        log_frame = ttk.LabelFrame(main, text="Log", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log = scrolledtext.ScrolledText(log_frame, height=15, font=("Consolas", 9), wrap=tk.WORD)
        self.log.pack(fill=tk.BOTH, expand=True)

        self.log.tag_config("info", foreground="black")
        self.log.tag_config("success", foreground="green")
        self.log.tag_config("warning", foreground="orange")
        self.log.tag_config("error", foreground="red")
        self.log.tag_config("header", foreground="blue", font=("Consolas", 9, "bold"))

    def _log(self, msg: str, tag: str = "info"):
        self.log.insert(tk.END, msg + "\n", tag)
        self.log.see(tk.END)
        self.root.update_idletasks()

    def _browse_export(self):
        path = filedialog.askdirectory(title="Select Export Folder (export__)")
        if path:
            self.export_var.set(path)

    def _browse_source(self):
        path = filedialog.askdirectory(title="Select Source Folder")
        if path:
            self.source_var.set(path)

    def _clear_all(self):
        self.export_var.set("")
        self.source_var.set("")
        self.min_length_var.set(DEFAULT_MIN_LENGTH)
        self.log.delete("1.0", tk.END)

    def _run_extract(self):
        export_path = self.export_var.get().strip()
        source_path = self.source_var.get().strip()

        if not export_path or not source_path:
            messagebox.showwarning("Missing Paths", "Please set both Export and Source folders.")
            return

        export_folder = Path(export_path)
        source_folder = Path(source_path)

        if not export_folder.is_dir():
            messagebox.showerror("Error", f"Export folder not found:\n{export_folder}")
            return
        if not source_folder.is_dir():
            messagebox.showerror("Error", f"Source folder not found:\n{source_folder}")
            return

        min_length = self.min_length_var.get()
        if min_length < 1:
            messagebox.showerror("Error", "Minimum length must be >= 1")
            return

        if not HAS_XLSXWRITER:
            messagebox.showerror("Error", "xlsxwriter is required for report output.\npip install xlsxwriter")
            return

        self.log.delete("1.0", tk.END)
        self.extract_btn.config(state=tk.DISABLED)

        try:
            self._execute_extract(export_folder, source_folder, min_length)
        except Exception as e:
            self._log(f"\nERROR: {e}", "error")
            logger.exception("Extraction failed")
        finally:
            self.extract_btn.config(state=tk.NORMAL)

    def _execute_extract(self, export_folder: Path, source_folder: Path, min_length: int):
        self._log("=" * 60, "header")
        self._log("  SCRIPT LONG STRING EXTRACTOR v2.0", "header")
        self._log("=" * 60, "header")

        # Step 1: Build category mapping
        self._log(f"\nExport folder: {export_folder}", "info")
        self._log("Building StringID -> Category mapping...", "info")

        category_map, subfolder_map = build_stringid_to_category(export_folder, progress_fn=self._log)
        ci_category = {k.lower(): v for k, v in category_map.items()}
        ci_subfolder = {k.lower(): v for k, v in subfolder_map.items()}

        script_count = sum(1 for v in category_map.values() if v in SCRIPT_CATEGORIES)
        self._log(f"  Total StringIDs indexed: {len(category_map)}", "info")
        self._log(f"  SCRIPT StringIDs (Dialog/Sequencer): {script_count}", "info")
        self._log(f"  Excluded subfolders: {', '.join(sorted(EXCLUDE_SUBFOLDERS)) or 'none'}", "info")

        # Step 2: Find source files
        self._log(f"\nSource folder: {source_folder}", "info")
        self._log(f"Min length threshold: {min_length} chars", "info")

        xml_files = sorted(source_folder.rglob("languagedata_*.xml"))
        xlsx_files = sorted(source_folder.rglob("*.xlsx"))
        # Filter out temp files
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

        self._log(f"  XML files: {len(xml_files)}", "info")
        self._log(f"  Excel files: {len(xlsx_files)}", "info")

        if not xml_files and not xlsx_files:
            self._log("\nNo source files found.", "warning")
            return

        # Step 3: Extract
        self._log(f"\n{'─' * 60}", "info")
        self._log("EXTRACTING...", "header")
        self._log(f"{'─' * 60}", "info")

        all_entries = []

        for xml_file in xml_files:
            try:
                entries = extract_from_xml(xml_file, ci_category, ci_subfolder, min_length)
                if entries:
                    lang = extract_language_from_filename(xml_file.name) or "UNKNOWN"
                    self._log(f"  {xml_file.name} [{lang}]: {len(entries)} entries", "success")
                all_entries.extend(entries)
            except Exception as e:
                self._log(f"  {xml_file.name}: ERROR - {e}", "error")

        for xlsx_file in xlsx_files:
            try:
                entries = extract_from_excel(xlsx_file, ci_category, ci_subfolder, min_length)
                if entries:
                    self._log(f"  {xlsx_file.name}: {len(entries)} entries", "success")
                all_entries.extend(entries)
            except Exception as e:
                self._log(f"  {xlsx_file.name}: ERROR - {e}", "error")

        # Step 4: Summary
        self._log(f"\n{'=' * 60}", "header")
        self._log("  SUMMARY", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Total SCRIPT entries found: {len(all_entries)}", "success" if all_entries else "info")

        if not all_entries:
            self._log("\nNo entries matched the criteria.", "warning")
            return

        # Group by language
        by_language: Dict[str, List[Dict]] = {}
        for e in all_entries:
            lang = e.get("language", "UNKNOWN") or "UNKNOWN"
            by_language.setdefault(lang, []).append(e)

        self._log(f"  Languages: {len(by_language)}", "info")
        for lang in sorted(by_language.keys()):
            entries_for_lang = by_language[lang]
            lengths = [e["char_count"] for e in entries_for_lang]
            self._log(f"    {lang}: {len(entries_for_lang)} entries (chars: {min(lengths)}-{max(lengths)})", "info")

        # Category breakdown
        cat_counts = {}
        for e in all_entries:
            cat_counts[e["category"]] = cat_counts.get(e["category"], 0) + 1
        for cat, cnt in sorted(cat_counts.items()):
            self._log(f"    {cat}: {cnt}", "info")

        # Step 5: Write per-language reports
        self._log(f"\n{'─' * 60}", "info")
        self._log("WRITING OUTPUT...", "header")
        self._log(f"{'─' * 60}", "info")

        timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = OUTPUT_DIR / f"extraction_{min_length}chars_{timestamp}"

        total_xlsx = 0
        total_xml = 0

        for lang in sorted(by_language.keys()):
            entries_for_lang = by_language[lang]

            # Excel output
            xlsx_path = output_dir / f"{lang}_script_long_strings.xlsx"
            if write_excel_report(entries_for_lang, xlsx_path):
                total_xlsx += 1
                self._log(f"  {xlsx_path.name}: {len(entries_for_lang)} entries", "success")

            # XML output (raw LocStr under <root>)
            xml_path = output_dir / f"{lang}_script_long_strings.xml"
            if write_xml_report(entries_for_lang, xml_path):
                total_xml += 1
                self._log(f"  {xml_path.name}: {len(entries_for_lang)} entries", "success")

        self._log(f"\n{'=' * 60}", "header")
        self._log("  DONE", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Output: {output_dir}", "success")
        self._log(f"  Excel files: {total_xlsx}", "info")
        self._log(f"  XML files:   {total_xml}", "info")
        self._log(f"  Total entries: {len(all_entries)} across {len(by_language)} language(s)", "info")


# =============================================================================
# MAIN
# =============================================================================

def main():
    root = tk.Tk()
    ScriptLongStringExtractorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
