#!/usr/bin/env python3
# coding: utf-8
"""
String Eraser XML v1.0
=======================
Standalone GUI tool that removes LocStr nodes from Target XML files
for entries that match Source by StringID + StrOrigin.

Based on TFM LITE's string_eraser logic, adapted for XML languagedata files.

Source: Folder containing Excel (.xlsx) or XML files with StringID + StrOrigin
Target: Folder containing languagedata_*.xml files

For each <LocStr> in Target where:
  - StringID matches a Source row (case-insensitive)
  - StrOrigin matches a Source row (normalized)
→ Remove the entire <LocStr> node from the XML tree

Usage: python string_eraser_xml.py
"""

import os
import re
import sys
import html
import stat
import logging
import datetime as _dt
from pathlib import Path
from typing import Dict, Set, Tuple, List

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Try lxml first (more robust), fallback to standard library
try:
    from lxml import etree
    USING_LXML = True
except ImportError:
    from xml.etree import ElementTree as etree
    USING_LXML = False

# Try openpyxl for Excel support
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# =============================================================================
# LOGGING
# =============================================================================

logger = logging.getLogger("StringEraserXML")
logger.setLevel(logging.DEBUG)

# =============================================================================
# CONFIGURATION
# =============================================================================

WINDOW_TITLE = "String Eraser XML v1.0"
WINDOW_WIDTH = 780
WINDOW_HEIGHT = 680
BUTTON_WIDTH = 20
ENTRY_WIDTH = 55

# LocStr tag and attribute variants (case-insensitive handling)
LOCSTR_TAGS = ['LocStr', 'locstr', 'LOCSTR', 'LOCStr', 'Locstr']
STRINGID_ATTRS = ('StringId', 'StringID', 'stringid', 'STRINGID', 'Stringid', 'stringId')
STRORIGIN_ATTRS = ('StrOrigin', 'Strorigin', 'strorigin', 'STRORIGIN')
STR_ATTRS = ('Str', 'str', 'STR')

# Excel header variants (case-insensitive detection)
STRINGID_HEADERS = {'stringid', 'string_id', 'string id'}
STRORIGIN_HEADERS = {'strorigin', 'str_origin', 'str origin'}


# =============================================================================
# TEXT NORMALIZATION (matches QuickTranslate's logic)
# =============================================================================

def normalize_text(txt: str) -> str:
    """Normalize text for matching: HTML unescape, whitespace collapse, &desc; removal."""
    if not txt:
        return ""
    txt = html.unescape(str(txt))
    txt = re.sub(r'\s+', ' ', txt.strip())
    if txt.lower().startswith("&desc;"):
        txt = txt[6:].lstrip()
    elif txt.lower().startswith("&amp;desc;"):
        txt = txt[10:].lstrip()
    return txt


def normalize_nospace(txt: str) -> str:
    """Remove ALL whitespace for fallback matching."""
    return re.sub(r'\s+', '', txt)


# =============================================================================
# XML HELPERS
# =============================================================================

def get_attr(elem, attr_variants):
    """Get attribute value from element, trying multiple case variants."""
    for attr_name in attr_variants:
        val = elem.get(attr_name)
        if val is not None:
            return attr_name, val
    return None, None


def parse_xml(xml_path: Path):
    """Parse XML file, returns (tree, root)."""
    if USING_LXML:
        parser = etree.XMLParser(
            resolve_entities=False, load_dtd=False,
            no_network=True, recover=True,
        )
        tree = etree.parse(str(xml_path), parser)
    else:
        tree = etree.parse(str(xml_path))
    return tree, tree.getroot()


def iter_locstr(root):
    """Iterate all LocStr elements (case-insensitive tag matching)."""
    elements = []
    for tag in LOCSTR_TAGS:
        elements.extend(root.iter(tag))
    return elements


def write_xml(tree, xml_path: Path):
    """Write XML tree back to file."""
    try:
        current_mode = os.stat(xml_path).st_mode
        if not current_mode & stat.S_IWRITE:
            os.chmod(xml_path, current_mode | stat.S_IWRITE)
    except Exception:
        pass

    if USING_LXML:
        tree.write(str(xml_path), encoding="utf-8", xml_declaration=False, pretty_print=True)
    else:
        tree.write(str(xml_path), encoding="utf-8", xml_declaration=False)


# =============================================================================
# SOURCE PARSING
# =============================================================================

def load_source_keys_from_excel(xlsx_path: Path) -> Set[Tuple[str, str]]:
    """
    Read Excel file and extract (StringID_lower, normalized_StrOrigin) tuples.

    Returns:
        Set of (stringid_lower, normalized_strorigin) keys
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel files: pip install openpyxl")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    keys = set()
    nospace_keys = set()

    # Detect headers from first row
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False))):
        val = str(cell.value or "").strip().lower()
        if val in STRINGID_HEADERS:
            headers['stringid'] = col_idx
        elif val in STRORIGIN_HEADERS:
            headers['strorigin'] = col_idx

    if 'stringid' not in headers or 'strorigin' not in headers:
        wb.close()
        raise ValueError(
            f"Excel file must have StringID and StrOrigin columns.\n"
            f"Found headers: {[str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]}"
        )

    sid_col = headers['stringid']
    so_col = headers['strorigin']

    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_col] or "").strip()
        so = str(row[so_col] or "").strip()
        if sid and so:
            norm_so = normalize_text(so)
            keys.add((sid.lower(), norm_so))
            nospace_keys.add((sid.lower(), normalize_nospace(norm_so)))

    wb.close()
    return keys, nospace_keys


def load_source_keys_from_xml(xml_path: Path) -> Tuple[Set[Tuple[str, str]], Set[Tuple[str, str]]]:
    """
    Read XML file and extract (StringID_lower, normalized_StrOrigin) tuples.

    Returns:
        Tuple of (keys_set, nospace_keys_set)
    """
    _, root = parse_xml(xml_path)
    keys = set()
    nospace_keys = set()

    for loc in iter_locstr(root):
        _, sid = get_attr(loc, STRINGID_ATTRS)
        _, so = get_attr(loc, STRORIGIN_ATTRS)

        sid = (sid or "").strip()
        so = (so or "").strip()

        if sid and so:
            norm_so = normalize_text(so)
            keys.add((sid.lower(), norm_so))
            nospace_keys.add((sid.lower(), normalize_nospace(norm_so)))

    return keys, nospace_keys


def load_source_keys(source_folder: Path) -> Tuple[Set[Tuple[str, str]], Set[Tuple[str, str]]]:
    """
    Scan source folder for Excel/XML files and build the erase key set.

    Returns:
        Tuple of (keys_set, nospace_keys_set) — each is a set of (sid_lower, normalized_strorigin)
    """
    all_keys = set()
    all_nospace_keys = set()
    files_processed = 0

    for f in sorted(source_folder.rglob("*")):
        if f.suffix.lower() == '.xlsx' and not f.name.startswith('~$'):
            keys, nospace = load_source_keys_from_excel(f)
            all_keys.update(keys)
            all_nospace_keys.update(nospace)
            files_processed += 1
        elif f.suffix.lower() in ('.xml', '.loc.xml'):
            keys, nospace = load_source_keys_from_xml(f)
            all_keys.update(keys)
            all_nospace_keys.update(nospace)
            files_processed += 1

    return all_keys, all_nospace_keys, files_processed


# =============================================================================
# ERASE LOGIC
# =============================================================================

def erase_matching_strings(
    target_xml: Path,
    keys: Set[Tuple[str, str]],
    nospace_keys: Set[Tuple[str, str]],
) -> Tuple[int, int, List[Dict]]:
    """
    Remove entire LocStr nodes matching the source keys from the XML tree.

    Match logic (same as QuickTranslate's 2-step cascade):
      Step 1: (sid.lower(), normalize_text(StrOrigin)) exact match
      Step 2: (sid.lower(), normalize_nospace(normalize_text(StrOrigin))) fallback

    Args:
        target_xml: Path to target languagedata_*.xml
        keys: Set of (sid_lower, normalized_strorigin) from source
        nospace_keys: Set of (sid_lower, nospace_strorigin) fallback keys

    Returns:
        (erased_count, skipped_already_empty, details_list)
    """
    tree, root = parse_xml(target_xml)
    erased = 0
    already_empty = 0
    details = []
    changed = False

    # Collect elements to remove (can't modify tree during iteration)
    to_remove = []

    for loc in iter_locstr(root):
        _, sid = get_attr(loc, STRINGID_ATTRS)
        _, so_raw = get_attr(loc, STRORIGIN_ATTRS)

        sid = (sid or "").strip()
        so_raw = (so_raw or "")

        if not sid:
            continue

        so_norm = normalize_text(so_raw)
        sid_lower = sid.lower()

        # 2-step cascade match
        key = (sid_lower, so_norm)
        key_nospace = (sid_lower, normalize_nospace(so_norm))

        matched = key in keys or key_nospace in nospace_keys
        if not matched:
            continue

        # Found a match
        str_attr, str_val = get_attr(loc, STR_ATTRS)

        if str_attr is None:
            already_empty += 1
            details.append({"stringid": sid, "status": "NO_STR_ATTR", "old": ""})
            continue

        if not str_val or not str_val.strip():
            already_empty += 1
            details.append({"stringid": sid, "status": "ALREADY_EMPTY", "old": ""})
            continue

        to_remove.append((loc, sid, str_val))

    # Remove matched LocStr nodes from the tree
    for loc, sid, old_val in to_remove:
        parent = loc.getparent() if USING_LXML else None
        if parent is not None:
            parent.remove(loc)
        else:
            # stdlib fallback: can't easily get parent, clear Str instead
            str_attr, _ = get_attr(loc, STR_ATTRS)
            if str_attr:
                loc.set(str_attr, "")
        erased += 1
        changed = True
        details.append({"stringid": sid, "status": "ERASED", "old": old_val})

    if changed:
        write_xml(tree, target_xml)

    return erased, already_empty, details


# =============================================================================
# GUI APPLICATION
# =============================================================================

class StringEraserGUI:
    """GUI for erasing XML strings matching source StringID + StrOrigin."""

    def __init__(self, root):
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.resizable(True, True)

        self.source_var = tk.StringVar()
        self.target_var = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        """Build the GUI layout."""
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        # Title
        ttk.Label(main, text="String Eraser XML", font=("Segoe UI", 14, "bold")).pack(pady=(0, 5))
        ttk.Label(
            main,
            text="Remove LocStr nodes from Target XML where StringID + StrOrigin match Source",
            font=("Segoe UI", 9),
        ).pack(pady=(0, 10))

        # Source frame
        src_frame = ttk.LabelFrame(main, text="Source (Excel/XML with StringID + StrOrigin)", padding=5)
        src_frame.pack(fill=tk.X, pady=5)

        src_row = ttk.Frame(src_frame)
        src_row.pack(fill=tk.X)
        ttk.Entry(src_row, textvariable=self.source_var, width=ENTRY_WIDTH).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(src_row, text="Browse Folder", width=15, command=self._browse_source).pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Label(src_frame, text="Accepts: .xlsx files (StringID + StrOrigin columns) or .xml files", font=("Segoe UI", 8)).pack(anchor=tk.W, pady=(3, 0))

        # Target frame
        tgt_frame = ttk.LabelFrame(main, text="Target (languagedata_*.xml files)", padding=5)
        tgt_frame.pack(fill=tk.X, pady=5)

        tgt_row = ttk.Frame(tgt_frame)
        tgt_row.pack(fill=tk.X)
        ttk.Entry(tgt_row, textvariable=self.target_var, width=ENTRY_WIDTH).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(tgt_row, text="Browse Folder", width=15, command=self._browse_target).pack(side=tk.RIGHT, padx=(5, 0))

        ttk.Label(tgt_frame, text="Scans for languagedata_*.xml files recursively", font=("Segoe UI", 8)).pack(anchor=tk.W, pady=(3, 0))

        # Action buttons
        btn_frame = ttk.Frame(main)
        btn_frame.pack(pady=10)

        self.erase_btn = ttk.Button(
            btn_frame, text="ERASE MATCHING STRINGS", width=30,
            command=self._run_erase,
        )
        self.erase_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(btn_frame, text="Clear All", width=12, command=self._clear_all).pack(side=tk.LEFT, padx=5)

        # Log area
        log_frame = ttk.LabelFrame(main, text="Log", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log = scrolledtext.ScrolledText(log_frame, height=20, font=("Consolas", 9), wrap=tk.WORD)
        self.log.pack(fill=tk.BOTH, expand=True)

        # Tags for colored log output
        self.log.tag_config("info", foreground="black")
        self.log.tag_config("success", foreground="green")
        self.log.tag_config("warning", foreground="orange")
        self.log.tag_config("error", foreground="red")
        self.log.tag_config("header", foreground="blue", font=("Consolas", 9, "bold"))

    def _log(self, msg: str, tag: str = "info"):
        """Append message to log area."""
        self.log.insert(tk.END, msg + "\n", tag)
        self.log.see(tk.END)
        self.root.update_idletasks()

    def _browse_source(self):
        path = filedialog.askdirectory(title="Select Source Folder")
        if path:
            self.source_var.set(path)

    def _browse_target(self):
        path = filedialog.askdirectory(title="Select Target Folder (languagedata_*.xml)")
        if path:
            self.target_var.set(path)

    def _clear_all(self):
        self.source_var.set("")
        self.target_var.set("")
        self.log.delete("1.0", tk.END)

    def _run_erase(self):
        """Execute the erase operation."""
        source_path = self.source_var.get().strip()
        target_path = self.target_var.get().strip()

        if not source_path or not target_path:
            messagebox.showwarning("Missing Paths", "Please set both Source and Target folders.")
            return

        source_folder = Path(source_path)
        target_folder = Path(target_path)

        if not source_folder.is_dir():
            messagebox.showerror("Error", f"Source folder not found:\n{source_folder}")
            return
        if not target_folder.is_dir():
            messagebox.showerror("Error", f"Target folder not found:\n{target_folder}")
            return

        # Confirm
        result = messagebox.askyesno(
            "Confirm Erase",
            "This will REMOVE entire LocStr nodes from Target XML files\n"
            "for all entries matching Source StringID + StrOrigin.\n\n"
            "This operation modifies files directly.\n"
            "Make sure your Target files are backed up or in Perforce.\n\n"
            "Continue?",
        )
        if not result:
            return

        self.log.delete("1.0", tk.END)
        self.erase_btn.config(state=tk.DISABLED)

        try:
            self._execute_erase(source_folder, target_folder)
        except Exception as e:
            self._log(f"\nERROR: {e}", "error")
            logger.exception("Erase failed")
        finally:
            self.erase_btn.config(state=tk.NORMAL)

    def _execute_erase(self, source_folder: Path, target_folder: Path):
        """Core erase logic."""
        self._log("=" * 60, "header")
        self._log("  STRING ERASER XML", "header")
        self._log("=" * 60, "header")

        # Step 1: Load source keys
        self._log(f"\nSource: {source_folder}", "info")
        self._log("Loading source keys (StringID + StrOrigin)...", "info")

        try:
            keys, nospace_keys, files_processed = load_source_keys(source_folder)
        except Exception as e:
            self._log(f"Failed to load source: {e}", "error")
            return

        self._log(f"  Files processed: {files_processed}", "info")
        self._log(f"  Unique erase keys: {len(keys)}", "info")

        if not keys:
            self._log("\nNo keys found in source. Nothing to erase.", "warning")
            return

        # Step 2: Find target XML files
        self._log(f"\nTarget: {target_folder}", "info")
        target_files = sorted(target_folder.rglob("languagedata_*.xml"))

        if not target_files:
            self._log("No languagedata_*.xml files found in target folder.", "warning")
            return

        self._log(f"  Target XML files: {len(target_files)}", "info")

        # Step 3: Erase matching strings
        self._log(f"\n{'─' * 60}", "info")
        self._log("ERASING...", "header")
        self._log(f"{'─' * 60}", "info")

        total_erased = 0
        total_already_empty = 0
        total_files_modified = 0
        all_details = []

        for xml_file in target_files:
            lang = xml_file.stem.replace("languagedata_", "").upper()

            try:
                erased, already_empty, details = erase_matching_strings(xml_file, keys, nospace_keys)
            except Exception as e:
                self._log(f"  {lang}: ERROR - {e}", "error")
                continue

            if erased > 0:
                self._log(f"  {lang}: {erased} erased, {already_empty} already empty", "success")
                total_files_modified += 1
            elif already_empty > 0:
                self._log(f"  {lang}: 0 erased, {already_empty} already empty", "info")
            # Skip files with 0 matches entirely

            total_erased += erased
            total_already_empty += already_empty
            all_details.extend(details)

        # Step 4: Summary
        self._log(f"\n{'=' * 60}", "header")
        self._log("  SUMMARY", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Source keys:        {len(keys)}", "info")
        self._log(f"  Target files:       {len(target_files)}", "info")
        self._log(f"  Files modified:     {total_files_modified}", "info")
        self._log(f"  Strings erased:     {total_erased}", "success" if total_erased > 0 else "info")
        self._log(f"  Already empty:      {total_already_empty}", "info")

        # Step 5: Write erase report
        if all_details:
            report_dir = source_folder / "Erase_Reports"
            report_dir.mkdir(exist_ok=True)
            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = report_dir / f"erase_report_{ts}.txt"

            with open(report_path, "w", encoding="utf-8") as f:
                f.write(f"String Eraser XML Report\n")
                f.write(f"Date: {_dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Source: {source_folder}\n")
                f.write(f"Target: {target_folder}\n")
                f.write(f"Keys: {len(keys)} | Erased: {total_erased} | Already empty: {total_already_empty}\n")
                f.write(f"\n{'StringID':<40} {'Status':<20} {'Old Value'}\n")
                f.write(f"{'-'*40} {'-'*20} {'-'*40}\n")
                for d in all_details:
                    old = d['old'][:60] + "..." if len(d['old']) > 60 else d['old']
                    f.write(f"{d['stringid']:<40} {d['status']:<20} {old}\n")

            self._log(f"\n  Report: {report_path}", "info")

        if total_erased > 0:
            self._log(f"\nDone! {total_erased} strings erased.", "success")
        else:
            self._log("\nDone. No strings were erased (all already empty or no matches).", "info")


# =============================================================================
# ENTRY POINT
# =============================================================================

def main():
    root = tk.Tk()
    StringEraserGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
