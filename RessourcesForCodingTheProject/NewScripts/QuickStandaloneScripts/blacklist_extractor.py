#!/usr/bin/env python3
# coding: utf-8
"""
BlacklistExtractor v1.0
========================
Standalone GUI tool that extracts LocStr entries from languagedata XML files
whose Str value contains any blacklisted term for that language.

Inputs:
  - LOC Folder: folder with languagedata_*.xml (for language suffix validation)
  - Source: Excel file(s) with one column per language (header = language suffix)
  - Each cell = one blacklisted term

Output (per language, in Blacklist_Output/ next to script):
  - Excel (.xlsx): StringID, StrOrigin, Str, MatchedTerm
  - XML (.xml): Raw LocStr elements under <root>

Usage: python blacklist_extractor.py
"""

import json
import re as _re
import sys
import time
import logging
import datetime as _dt
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Try lxml first (preserves attribute order, recovery mode), fallback to stdlib
try:
    from lxml import etree
    USING_LXML = True
except ImportError:
    from xml.etree import ElementTree as etree
    USING_LXML = False

# Try openpyxl for reading Excel
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Try xlsxwriter for writing Excel
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

# =============================================================================
# LOGGING
# =============================================================================

logger = logging.getLogger("BlacklistExtractor")
logger.setLevel(logging.DEBUG)

# =============================================================================
# CONSTANTS
# =============================================================================

# Detect script directory (supports PyInstaller frozen exe)
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent

SETTINGS_FILE = SCRIPT_DIR / "blacklist_extractor_settings.json"
OUTPUT_DIR = SCRIPT_DIR / "Blacklist_Output"

# LocStr tag and attribute variants (case-insensitive matching)
LOCSTR_TAGS = ['LocStr', 'locstr', 'LOCSTR', 'LOCStr', 'Locstr']
STRINGID_ATTRS = ('StringId', 'StringID', 'stringid', 'STRINGID', 'Stringid', 'stringId')
STRORIGIN_ATTRS = ('StrOrigin', 'Strorigin', 'strorigin', 'STRORIGIN')
STR_ATTRS = ('Str', 'str', 'STR')


# =============================================================================
# SETTINGS PERSISTENCE
# =============================================================================

def _load_settings() -> dict:
    """Load persisted settings from JSON."""
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_settings(settings: dict):
    """Save settings to JSON."""
    try:
        SETTINGS_FILE.write_text(
            json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8",
        )
    except Exception as e:
        logger.warning("Failed to save settings: %s", e)


# =============================================================================
# LANGUAGE DETECTION (standalone — no QuickTranslate dependency)
# =============================================================================

def discover_valid_codes(loc_folder: Path) -> Dict[str, Path]:
    """
    Discover valid language codes from LOC folder by scanning languagedata_*.xml.

    Returns:
        {uppercase_code: Path_to_xml} e.g. {'ENG': Path('languagedata_ENG.xml')}
    """
    codes = {}
    if loc_folder.exists() and loc_folder.is_dir():
        for xml_file in sorted(loc_folder.glob("languagedata_*.xml")):
            match = _re.match(r'languagedata_(.+)\.xml', xml_file.name, _re.IGNORECASE)
            if match:
                code = match.group(1).upper()
                codes[code] = xml_file
    return codes


# =============================================================================
# XML PARSING
# =============================================================================

def _read_xml_raw(xml_path: Path) -> Optional[str]:
    """Read XML file with encoding fallback (latin-1 always succeeds)."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return xml_path.read_text(encoding=enc)
        except (UnicodeDecodeError, ValueError):
            continue
    return None  # Unreachable — latin-1 decodes any byte sequence


# Fix bare & that aren't valid XML entities
_bad_amp = _re.compile(r'&(?!(?:amp|lt|gt|apos|quot|#\d+|#x[0-9a-fA-F]+);)')
# Fix malformed self-closing tags like </>
_bad_selfclose = _re.compile(r'</>')
# Fix unescaped < inside attribute values, but PRESERVE <br/> tags
# Matches < that is NOT part of <br/> or <br /> patterns
_attr_dangerous_lt = _re.compile(r'<(?![bB][rR]\s*/?>)')


def _escape_attr_lt(m):
    """Escape < inside attribute values, preserving <br/> tags."""
    content = m.group(1)
    content = _attr_dangerous_lt.sub('&lt;', content)
    return '="' + content + '"'


# Matches attribute values containing <
_attr_with_lt = _re.compile(r'="([^"]*<[^"]*)"')


def sanitize_xml(raw: str) -> str:
    """Fix common XML issues in game data files before parsing.
    PRESERVES <br/> tags inside attribute values."""
    raw = _bad_selfclose.sub('', raw)
    raw = _attr_with_lt.sub(_escape_attr_lt, raw)
    raw = _bad_amp.sub('&amp;', raw)
    return raw


def _parse_root(raw: str):
    """Parse sanitized XML string and return root element."""
    sanitized = sanitize_xml(raw)
    try:
        if USING_LXML:
            parser = etree.XMLParser(
                recover=True, encoding="utf-8",
                resolve_entities=False, load_dtd=False, no_network=True,
            )
            return etree.fromstring(sanitized.encode("utf-8"), parser)
        else:
            return etree.fromstring(sanitized)
    except Exception as e:
        logger.error("XML parse error: %s", e)
        return None


def get_attr_value(attrs: dict, attr_variants: tuple) -> str:
    """Get attribute value from dict, trying multiple case variants."""
    for attr_name in attr_variants:
        if attr_name in attrs:
            return attrs[attr_name]
    return ""


def iter_locstr(root) -> list:
    """Iterate all LocStr elements (case-insensitive tag matching)."""
    elements = []
    for tag in LOCSTR_TAGS:
        elements.extend(root.iter(tag))
    return elements


# =============================================================================
# BLACKLIST READING (Excel)
# =============================================================================

def read_blacklist_from_excel(
    excel_path: Path, valid_codes: Set[str],
) -> Tuple[Dict[str, List[str]], List[str]]:
    """
    Read blacklist terms from Excel file. Each column header = language suffix.

    Args:
        excel_path: Path to .xlsx file
        valid_codes: Set of valid uppercase language codes

    Returns:
        ({lang_code: [terms]}, [warnings])
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for reading Excel: pip install openpyxl")

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    warnings = []

    if len(wb.sheetnames) > 1:
        warnings.append(f"Workbook has {len(wb.sheetnames)} sheets — only reading active sheet '{ws.title}'")

    # Read header row to detect language columns
    lang_columns: Dict[int, str] = {}  # {col_index: lang_code}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if not header_row:
        wb.close()
        return {}, ["Empty file or no header row"]

    for col_idx, cell in enumerate(header_row):
        val = str(cell.value or "").strip().upper()
        if not val:
            continue
        if val in valid_codes:
            lang_columns[col_idx] = val
        else:
            warnings.append(f"Column '{cell.value}' is not a recognized language suffix — skipped")

    if not lang_columns:
        wb.close()
        return {}, warnings + ["No language columns detected in header"]

    # Read terms from each language column
    blacklist: Dict[str, List[str]] = {lang: [] for lang in lang_columns.values()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        for col_idx, lang in lang_columns.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    term = str(val).strip()
                    if term:
                        blacklist[lang].append(term)

    wb.close()

    # Deduplicate per language (preserve order)
    for lang in blacklist:
        seen = set()
        unique = []
        for term in blacklist[lang]:
            lower = term.lower()
            if lower not in seen:
                seen.add(lower)
                unique.append(term)
        blacklist[lang] = unique

    return blacklist, warnings


# =============================================================================
# BLACKLIST SEARCH
# =============================================================================

def search_languagedata(
    xml_path: Path, terms: List[str], progress_fn=None,
) -> Tuple[List[Dict], float]:
    """
    Search a languagedata XML file for LocStr entries containing any blacklisted term.

    Substring match: term.lower() in str_value.lower()

    Returns:
        (results_list, elapsed_seconds)
    """
    t0 = time.time()

    raw = _read_xml_raw(xml_path)
    if raw is None:
        return [], 0.0

    root = _parse_root(raw)
    if root is None:
        logger.error("Empty or invalid XML: %s", xml_path)
        return [], 0.0

    # Pre-lowercase all terms for matching
    terms_lower = [(t, t.lower()) for t in terms]

    results = []
    elements = iter_locstr(root)
    total = len(elements)

    if progress_fn:
        progress_fn(f"    Parsing done — {total} entries to scan...")

    # Dynamic progress interval: ~20 updates per language, min 1000
    progress_interval = max(1000, total // 20)
    hits = 0

    for idx, elem in enumerate(elements):
        attrs = dict(elem.attrib)
        sid = get_attr_value(attrs, STRINGID_ATTRS)
        so = get_attr_value(attrs, STRORIGIN_ATTRS)
        sv = get_attr_value(attrs, STR_ATTRS)

        if not sid or not sv:
            continue

        sv_lower = sv.lower()

        # Check ALL blacklist terms — one row per match for easy Excel filtering
        for term_orig, term_lower in terms_lower:
            if term_lower in sv_lower:
                results.append({
                    "string_id": sid,
                    "str_origin": so,
                    "str_value": sv,
                    "matched_term": term_orig,
                    "raw_attribs": attrs,
                })
                hits += 1

        if progress_fn and idx > 0 and idx % progress_interval == 0:
            pct = (idx / total) * 100
            elapsed = time.time() - t0
            progress_fn(f"    {pct:5.1f}% ({idx}/{total}) — {hits} hits so far [{elapsed:.1f}s]")

    elapsed = time.time() - t0
    return results, elapsed


# =============================================================================
# REPORT GENERATION
# =============================================================================

def _xml_escape_attr(value: str) -> str:
    """Escape a string for use in an XML attribute value.
    PRESERVES <br/> tags — the standard newline format in game XML data."""
    # Protect <br/> variants before escaping
    value = _re.sub(r'<br\s*/?>', '\x00BR\x00', value, flags=_re.IGNORECASE)
    value = (
        value
        .replace('&', '&amp;')
        .replace('"', '&quot;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
    )
    # Restore <br/> tags
    value = value.replace('\x00BR\x00', '<br/>')
    return value


def write_excel_report(entries: List[Dict], output_path: Path) -> bool:
    """Write blacklist results to Excel (StringID, StrOrigin, Str, MatchedTerm)."""
    if not HAS_XLSXWRITER:
        logger.warning("xlsxwriter not available — skipping Excel output")
        return False

    if not entries:
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = xlsxwriter.Workbook(str(output_path))
    ws = wb.add_worksheet("Blacklist Matches")

    # Formats
    header_fmt = wb.add_format({
        'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#8B0000',
        'border': 1, 'text_wrap': True,
    })
    cell_fmt = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    term_fmt = wb.add_format({
        'border': 1, 'text_wrap': True, 'valign': 'top',
        'font_color': '#8B0000', 'bold': True,
    })

    headers = ["StringID", "StrOrigin", "Str", "MatchedTerm"]
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_fmt)

    # Sort by matched_term then string_id
    entries_sorted = sorted(entries, key=lambda e: (e["matched_term"].lower(), e["string_id"].lower()))

    for row, e in enumerate(entries_sorted, 1):
        ws.write(row, 0, e["string_id"], cell_fmt)
        ws.write(row, 1, e["str_origin"], cell_fmt)
        ws.write(row, 2, e["str_value"], cell_fmt)
        ws.write(row, 3, e["matched_term"], term_fmt)

    # Column widths
    ws.set_column(0, 0, 35)   # StringID
    ws.set_column(1, 1, 45)   # StrOrigin
    ws.set_column(2, 2, 60)   # Str
    ws.set_column(3, 3, 25)   # MatchedTerm

    ws.autofilter(0, 0, len(entries_sorted), len(headers) - 1)
    ws.freeze_panes(1, 0)

    wb.close()
    return True


def write_xml_report(entries: List[Dict], output_path: Path) -> int:
    """Write raw LocStr elements under <root> to XML file.
    Deduplicates by StringID (same entry may match multiple terms).
    Returns number of unique entries written."""
    if not entries:
        return 0

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Deduplicate by StringID — one LocStr per entry in XML
    seen_sids = set()
    unique_entries = []
    for e in entries:
        sid = e["string_id"]
        if sid not in seen_sids:
            seen_sids.add(sid)
            unique_entries.append(e)

    lines = ['<?xml version="1.0" encoding="utf-8"?>', '<root>']

    for e in sorted(unique_entries, key=lambda x: x["string_id"].lower()):
        attribs = e.get("raw_attribs", {})
        if not attribs:
            attribs = {"StringID": e["string_id"], "StrOrigin": e["str_origin"], "Str": e["str_value"]}

        attr_parts = []
        for k, v in attribs.items():
            escaped = _xml_escape_attr(str(v))
            attr_parts.append(f'{k}="{escaped}"')

        lines.append(f'  <LocStr {" ".join(attr_parts)} />')

    lines.append('</root>')
    lines.append('')

    output_path.write_text('\n'.join(lines), encoding='utf-8')
    return len(unique_entries)


# =============================================================================
# GUI APPLICATION
# =============================================================================

class BlacklistExtractorApp:
    """GUI for extracting LocStr entries that contain blacklisted terms."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("BlacklistExtractor v1.0")
        self.root.geometry("800x700")
        self.root.resizable(True, True)

        self._build_ui()
        self._load_persisted_settings()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill="both", expand=True)
        pad = {"padx": 8, "pady": 4}

        # Title
        ttk.Label(
            main, text="BlacklistExtractor",
            font=("Segoe UI", 14, "bold"),
        ).pack(pady=(0, 2))
        ttk.Label(
            main,
            text="Extract LocStr entries containing blacklisted terms (per language)",
            font=("Segoe UI", 9),
        ).pack(pady=(0, 10))

        # LOC folder (for language suffix validation)
        self.loc_var = tk.StringVar()
        f_loc = ttk.LabelFrame(main, text="LOC Folder (languagedata_*.xml — for language detection)", padding=5)
        f_loc.pack(fill="x", **pad)
        loc_row = ttk.Frame(f_loc)
        loc_row.pack(fill="x")
        ttk.Entry(loc_row, textvariable=self.loc_var, width=60).pack(
            side="left", fill="x", expand=True, padx=(0, 5),
        )
        ttk.Button(
            loc_row, text="Browse...",
            command=lambda: self._browse_folder(self.loc_var, "LOC Folder"),
        ).pack(side="right", padx=(0, 3))
        ttk.Button(loc_row, text="Save", command=self._save_loc_setting).pack(side="right")

        # Source file/folder
        self.source_var = tk.StringVar()
        f_src = ttk.LabelFrame(main, text="SOURCE (Excel file or folder with Excel files)", padding=5)
        f_src.pack(fill="x", **pad)
        src_row = ttk.Frame(f_src)
        src_row.pack(fill="x")
        ttk.Entry(src_row, textvariable=self.source_var, width=60).pack(
            side="left", fill="x", expand=True, padx=(0, 5),
        )
        ttk.Button(
            src_row, text="File...",
            command=self._browse_source_file,
        ).pack(side="right", padx=(0, 3))
        ttk.Button(
            src_row, text="Folder...",
            command=lambda: self._browse_folder(self.source_var, "Source Folder"),
        ).pack(side="right", padx=(0, 3))
        ttk.Label(
            f_src,
            text="Excel columns: one per language (header = language suffix: ENG, FRE, GER, etc.)",
            font=("Segoe UI", 8),
        ).pack(anchor="w", pady=(3, 0))

        # Output info
        out_frame = ttk.LabelFrame(main, text="Output", padding=5)
        out_frame.pack(fill="x", **pad)
        ttk.Label(out_frame, text=f"Output directory: {OUTPUT_DIR}", font=("Segoe UI", 8)).pack(anchor="w")
        ttk.Label(out_frame, text="Per language: one Excel (.xlsx) + one XML (.xml)", font=("Segoe UI", 8)).pack(anchor="w")

        # Action buttons
        btn_frame = ttk.Frame(main)
        btn_frame.pack(pady=10)
        self.run_btn = ttk.Button(btn_frame, text="EXTRACT BLACKLISTED STRINGS", width=35, command=self._run)
        self.run_btn.pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Clear All", width=12, command=self._clear_all).pack(side="left", padx=5)

        # Log
        f_log = ttk.LabelFrame(main, text="Log", padding=5)
        f_log.pack(fill="both", expand=True, **pad)
        self.log = scrolledtext.ScrolledText(f_log, height=18, font=("Consolas", 9), wrap="word")
        self.log.pack(fill="both", expand=True)

        self.log.tag_config("info", foreground="black")
        self.log.tag_config("success", foreground="green")
        self.log.tag_config("warning", foreground="orange")
        self.log.tag_config("error", foreground="red")
        self.log.tag_config("header", foreground="blue", font=("Consolas", 9, "bold"))

    # -----------------------------------------------------------------
    # HELPERS
    # -----------------------------------------------------------------

    def _log(self, msg: str, tag: str = "info"):
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.root.update_idletasks()

    def _browse_folder(self, var: tk.StringVar, title: str):
        path = filedialog.askdirectory(title=f"Select {title}")
        if path:
            var.set(path)

    def _browse_source_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.source_var.set(path)

    def _clear_all(self):
        self.source_var.set("")
        self.log.delete("1.0", "end")

    def _save_loc_setting(self):
        loc = self.loc_var.get().strip()
        if not loc:
            return
        settings = _load_settings()
        settings["loc_folder"] = loc
        _save_settings(settings)
        messagebox.showinfo("Saved", f"LOC folder saved:\n{loc}")

    def _load_persisted_settings(self):
        settings = _load_settings()
        loc = settings.get("loc_folder", "")
        if loc:
            self.loc_var.set(loc)

    # -----------------------------------------------------------------
    # MAIN LOGIC
    # -----------------------------------------------------------------

    def _run(self):
        self.log.delete("1.0", "end")

        loc = self.loc_var.get().strip()
        source = self.source_var.get().strip()

        # --- Validate inputs ---
        if not loc:
            messagebox.showerror("Error", "Please select the LOC folder (for language detection).")
            return
        loc_path = Path(loc)
        if not loc_path.is_dir():
            messagebox.showerror("Error", f"LOC folder not found:\n{loc_path}")
            return

        if not source:
            messagebox.showerror("Error", "Please select a SOURCE Excel file or folder.")
            return
        source_path = Path(source)
        if not source_path.exists():
            messagebox.showerror("Error", f"SOURCE not found:\n{source_path}")
            return

        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "openpyxl is required for reading Excel.\npip install openpyxl")
            return

        self.run_btn.config(state="disabled")

        try:
            self._execute(loc_path, source_path)
        except Exception as e:
            self._log(f"\nERROR: {e}", "error")
            logger.exception("BlacklistExtractor failed")
        finally:
            self.run_btn.config(state="normal")

    def _execute(self, loc_path: Path, source_path: Path):
        self._log("=" * 60, "header")
        self._log("  BLACKLIST EXTRACTOR v1.0", "header")
        self._log("=" * 60, "header")

        # Step 1: Discover languages from LOC
        self._log(f"\nLOC folder: {loc_path}", "info")
        self._log("Discovering languages...", "info")

        lang_xmls = discover_valid_codes(loc_path)
        if not lang_xmls:
            self._log("No languagedata_*.xml files found in LOC folder.", "error")
            return

        valid_codes = set(lang_xmls.keys())
        self._log(f"  {len(valid_codes)} languages: {', '.join(sorted(valid_codes))}", "info")

        # Step 2: Collect Excel files from source
        self._log(f"\nSource: {source_path}", "info")

        excel_files: List[Path] = []
        if source_path.is_file():
            if source_path.suffix.lower() == ".xlsx":
                excel_files.append(source_path)
            else:
                self._log(f"Not an Excel file: {source_path.name}", "error")
                return
        elif source_path.is_dir():
            excel_files = sorted([
                f for f in source_path.iterdir()
                if f.is_file() and f.suffix.lower() == ".xlsx" and not f.name.startswith("~$")
            ])
        else:
            self._log("Source is neither a file nor a folder.", "error")
            return

        if not excel_files:
            self._log("No .xlsx files found in source.", "warning")
            return

        self._log(f"  Excel files: {len(excel_files)}", "info")
        for f in excel_files:
            self._log(f"    {f.name}", "info")

        # Step 3: Read blacklists from all Excel files
        self._log(f"\n{'─' * 60}", "info")
        self._log("READING BLACKLISTS...", "header")
        self._log(f"{'─' * 60}", "info")

        combined_blacklist: Dict[str, List[str]] = {}  # {lang: [terms]}

        for excel_file in excel_files:
            self._log(f"\n  {excel_file.name}:", "info")
            try:
                blacklist, warnings = read_blacklist_from_excel(excel_file, valid_codes)

                for w in warnings:
                    self._log(f"    WARNING: {w}", "warning")

                for lang, terms in sorted(blacklist.items()):
                    if not terms:
                        continue
                    self._log(f"    {lang}: {len(terms)} terms", "success")
                    if lang not in combined_blacklist:
                        combined_blacklist[lang] = []
                    combined_blacklist[lang].extend(terms)

            except Exception as e:
                self._log(f"    ERROR: {e}", "error")

        if not combined_blacklist:
            self._log("\nNo blacklist terms found in any Excel file.", "warning")
            return

        # Deduplicate across files
        for lang in combined_blacklist:
            seen = set()
            unique = []
            for term in combined_blacklist[lang]:
                lower = term.lower()
                if lower not in seen:
                    seen.add(lower)
                    unique.append(term)
            combined_blacklist[lang] = unique

        self._log(f"\nCombined blacklist:", "info")
        total_terms = 0
        for lang in sorted(combined_blacklist.keys()):
            terms = combined_blacklist[lang]
            total_terms += len(terms)
            preview = ", ".join(terms[:5])
            if len(terms) > 5:
                preview += f", ... ({len(terms)} total)"
            self._log(f"  {lang}: {preview}", "info")
        self._log(f"  Total: {total_terms} terms across {len(combined_blacklist)} languages", "info")

        # Step 4: Search languagedata for each language
        self._log(f"\n{'─' * 60}", "info")
        self._log("SEARCHING LANGUAGE DATA...", "header")
        self._log(f"{'─' * 60}", "info")

        all_results: Dict[str, List[Dict]] = {}  # {lang: [matches]}
        total_matches = 0
        total_elapsed = 0.0
        langs_to_search = sorted(combined_blacklist.keys())
        num_langs = len(langs_to_search)

        for lang_idx, lang in enumerate(langs_to_search, 1):
            terms = combined_blacklist[lang]
            xml_path = lang_xmls.get(lang)

            if not xml_path:
                self._log(f"\n  [{lang_idx}/{num_langs}] {lang}: No languagedata file found — skipped", "warning")
                continue

            self._log(f"\n  [{lang_idx}/{num_langs}] {lang} — {xml_path.name} — {len(terms)} blacklist terms", "info")

            matches, elapsed = search_languagedata(
                xml_path, terms,
                progress_fn=lambda msg: self._log(msg, "info"),
            )
            total_elapsed += elapsed

            if matches:
                all_results[lang] = matches
                total_matches += len(matches)
                unique_sids = len(set(e["string_id"] for e in matches))
                self._log(
                    f"  [{lang}] {unique_sids} entries, {len(matches)} term hits [{elapsed:.1f}s]",
                    "success",
                )
            else:
                self._log(f"  [{lang}] No matches [{elapsed:.1f}s]", "info")

        # Step 5: Summary
        self._log(f"\n{'=' * 60}", "header")
        self._log("  SUMMARY", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Total matches: {total_matches}", "success" if total_matches else "info")
        self._log(f"  Search time: {total_elapsed:.1f}s across {num_langs} languages", "info")

        if not all_results:
            self._log("\nNo blacklisted entries found in any language.", "info")
            return

        for lang in sorted(all_results.keys()):
            entries = all_results[lang]
            unique_terms = set(e["matched_term"].lower() for e in entries)
            unique_sids = set(e["string_id"] for e in entries)
            self._log(
                f"  {lang}: {len(unique_sids)} entries, {len(entries)} term hits "
                f"({len(unique_terms)} unique terms matched)", "info",
            )

        # Step 6: Write per-language reports
        self._log(f"\n{'─' * 60}", "info")
        self._log("WRITING OUTPUT...", "header")
        self._log(f"{'─' * 60}", "info")

        timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = OUTPUT_DIR / f"blacklist_{timestamp}"

        total_xlsx = 0
        total_xml = 0

        for lang in sorted(all_results.keys()):
            entries = all_results[lang]

            # Excel output (all rows — one row per term match for filtering)
            xlsx_path = output_dir / f"BLACKLIST_{lang}.xlsx"
            if write_excel_report(entries, xlsx_path):
                total_xlsx += 1
                self._log(f"  {xlsx_path.name}: {len(entries)} rows", "success")

            # XML output (deduplicated by StringID — clean LocStr list)
            xml_path = output_dir / f"BLACKLIST_{lang}.xml"
            xml_count = write_xml_report(entries, xml_path)
            if xml_count:
                total_xml += 1
                self._log(f"  {xml_path.name}: {xml_count} unique entries", "success")

        self._log(f"\n{'=' * 60}", "header")
        self._log("  DONE", "header")
        self._log(f"{'=' * 60}", "header")
        self._log(f"  Output: {output_dir}", "success")
        self._log(f"  Excel files: {total_xlsx}", "info")
        self._log(f"  XML files:   {total_xml}", "info")
        self._log(f"  Total entries: {total_matches} across {len(all_results)} language(s)", "info")

    def run(self):
        self.root.mainloop()


# =============================================================================
# MAIN
# =============================================================================

def main():
    app = BlacklistExtractorApp()
    app.run()


if __name__ == "__main__":
    main()
