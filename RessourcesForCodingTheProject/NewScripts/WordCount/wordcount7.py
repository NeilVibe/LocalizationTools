#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ============================================================
#  Translation Utilities – Word-Count with Unused String Detection
#  (2026-02 – Enhanced with __backup folder scanning for unused strings)
#
#  NEW FEATURES:
#  ------------------------------------------------------------
#  1.  Scans __backup folders in editordata and sequencer paths
#  2.  Parses .seqc files to extract DialogStrKey values
#  3.  Cross-references with export to identify unused strings
#  4.  "Unused Strings" column added to output
#  5.  "Coverage %" renamed to "True Coverage %" (excludes unused)
# ============================================================

import os
import re
import sys
import copy
import threading
import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
from pathlib import Path
from datetime import datetime
from typing import Set, List, Optional, Iterator, Callable, Dict, Any, Tuple
import xml.etree.ElementTree as ET

from lxml import etree as LET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ----------------------------------------------------------------------
#  HELPERS TO HANDLE "FROZEN" EXECUTABLES  (PyInstaller, cx_Freeze …)
# ----------------------------------------------------------------------
def get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


# ----------------------------------------------------------------------
#  CONSTANTS  (adjust as needed)
# ----------------------------------------------------------------------
LANGUAGE_FOLDER     = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")
EXPORT_FOLDER       = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")
DIFF_OUTPUT_FOLDER  = Path.cwd() / "diff_entries_output"
MISSING_XML_FILE    = Path.cwd() / "Missing_From_Export.xml"

# Backup folder paths for unused string detection
BACKUP_SCAN_ROOTS = [
    Path(r"F:\perforce\cd\mainline\resource\editordata"),
    Path(r"F:\perforce\cd\mainline\resource\sequencer"),
]

BASE_DIR      = get_base_dir()
BEFORE_FOLDER = BASE_DIR / "BEFORE"
AFTER_FOLDER  = BASE_DIR / "AFTER"

# ----------------------------------------------------------------------
#  RULE-SETS
# ----------------------------------------------------------------------
NON_PRIORITY_FOLDERS = {"ItemGroup", "Gimmick", "MultiChange"}   # → "Non-Priority"
DIALOG_SUBS          = ["AIDialog", "NarrationDialog",
                        "QuestDialog", "StageCloseDialog"]
SEQUENCER_TOPS       = ["Faction", "Main", "Other", "Sequencer"]

# ------------- colour map (big categories)
BIG_CAT_COLOURS = {
    "Dialog"       : "C6EFCE",  # light-green
    "Sequencer"    : "FFE599",  # light-orange
    "System"       : "D9D2E9",  # light-purple
    "World"        : "F8CBAD",  # light-red
    "Platform"     : "A9D08E",  # light-teal / green-gray
    "None"         : "DDD9C4",  # light-brown
    "Non-Priority" : "D9D9D9"   # light-grey
}
BIG_CAT_FILLS = {k: PatternFill("solid", fgColor=v) for k, v in BIG_CAT_COLOURS.items()}

# Regex patterns
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')
korean_re      = re.compile(r'[\uac00-\ud7a3]')
_ctrl_char_re  = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Tag stack repair patterns
_TAG_OPEN_RE  = re.compile(r"^<([A-Za-z_][\w.-]*)")
_TAG_CLOSE_RE = re.compile(r"^</([A-Za-z_][\w.-]*)>")


# ----------------------------------------------------------------------
#  XML SANITIZATION (5-pass pipeline from battle-tested QACompilerNEW)
# ----------------------------------------------------------------------
def fix_bad_entities(x: str) -> str:
    """Pass 1: Fix malformed entities like &unknown; → &amp;unknown;"""
    return _bad_entity_re.sub("&amp;", x)


def remove_control_chars(x: str) -> str:
    """Remove invalid XML control characters."""
    return _ctrl_char_re.sub("", x)


def escape_lt_in_attrs(x: str) -> str:
    """Pass 2: Escape < characters inside attribute values."""
    return re.sub(r'="([^"]*<[^"]*)"',
                  lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', x)


def escape_amp_in_attrs(x: str) -> str:
    """Pass 3: Escape & in attribute values (when not a valid entity)."""
    return re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                  lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', x)


def repair_tag_stack(raw: str) -> str:
    """
    Pass 4 & 5: Tag stack repair for malformed XML.
    Handles:
    - Empty close tags </>
    - Mismatched close tags
    - Unclosed tags at end
    """
    lines = raw.splitlines()
    out = []
    stack = []

    for line in lines:
        s = line.strip()

        # Handle empty close tag </>
        if "</>" in s:
            if stack:
                tag = stack.pop()
                s = s.replace("</>", f"</{tag}>")
                line = line.replace("</>", f"</{tag}>")
            out.append(line)
            continue

        # Check for opening tag
        m_open = _TAG_OPEN_RE.match(s)
        if m_open and not s.endswith("/>") and not s.startswith("</"):
            stack.append(m_open.group(1))

        # Check for closing tag
        m_close = _TAG_CLOSE_RE.match(s)
        if m_close:
            want = m_close.group(1)
            if stack and stack[-1] == want:
                stack.pop()

        out.append(line)

    # Auto-close remaining unclosed tags
    while stack:
        out.append(f"</{stack.pop()}>")

    return "\n".join(out)


def sanitize_xml_content(raw: str) -> str:
    """Full 5-pass XML sanitization pipeline."""
    raw = remove_control_chars(raw)
    raw = fix_bad_entities(raw)
    raw = escape_lt_in_attrs(raw)
    raw = escape_amp_in_attrs(raw)
    raw = repair_tag_stack(raw)
    return raw


def parse_xml_file(p: Path) -> Optional[LET._Element]:
    """
    Robust XML parsing with:
    - 5-pass sanitization
    - Virtual ROOT wrap
    - Recovery mode fallback
    - XXE protection (no external entities, no network, no DTD)
    """
    try:
        raw = p.read_text(encoding="utf-8", errors="ignore")
        sanitized = sanitize_xml_content(raw)
        wrapped = f"<ROOT>\n{sanitized}\n</ROOT>"

        # Try strict parsing first (with XXE protection)
        try:
            parser = LET.XMLParser(
                huge_tree=True,
                resolve_entities=False,
                no_network=True,
                dtd_validation=False,
                load_dtd=False
            )
            return LET.fromstring(wrapped.encode("utf-8"), parser=parser)
        except LET.XMLSyntaxError:
            # Fallback to recovery mode (with XXE protection)
            parser = LET.XMLParser(
                recover=True,
                huge_tree=True,
                resolve_entities=False,
                no_network=True,
                dtd_validation=False,
                load_dtd=False
            )
            return LET.fromstring(wrapped.encode("utf-8"), parser=parser)
    except Exception:
        return None


def count_words(txt: str) -> int:
    return len([w for w in re.split(r"\s+", txt.strip()) if w])


def is_korean(txt: str) -> bool:
    return bool(korean_re.search(txt))


def iter_language_files(folder: Path) -> Iterator[Path]:
    for r, _, fs in os.walk(folder):
        for f in fs:
            lf = f.lower()
            if lf == "languagedata.xml":
                continue                         #  ←  DROP generic file
            if lf.startswith("languagedata_") and lf.endswith(".xml"):
                yield Path(r) / f


# ----------------------------------------------------------------------
#  BACKUP FOLDER SCANNING FOR UNUSED STRINGS
# ----------------------------------------------------------------------
def find_backup_folders(root_paths: List[Path]) -> List[Path]:
    """
    Recursively find all __backup folders (case-insensitive) within root paths.
    """
    backup_folders = []

    for root_path in root_paths:
        if not root_path.exists():
            continue

        for dirpath, dirnames, _ in os.walk(root_path):
            for dirname in dirnames:
                if dirname.lower() == "__backup":
                    backup_folders.append(Path(dirpath) / dirname)

    return backup_folders


def extract_dialog_str_keys_from_seqc(seqc_path: Path) -> Set[str]:
    """
    Parse a .seqc file (XML) and extract all DialogStrKey attribute values.

    Example:
    <Subtitle Speaker="Boss_NorthernWarrior_55028"
              DialogStr="..."
              DialogStrKey="norhernwarrior_9000_boss_00_00007"/>

    Returns: Set of DialogStrKey values found in the file.
    """
    keys = set()

    root = parse_xml_file(seqc_path)
    if root is None:
        return keys

    # Search all elements for DialogStrKey attribute
    for elem in root.iter():
        dialog_str_key = elem.get("DialogStrKey")
        if dialog_str_key:
            keys.add(dialog_str_key.strip())

    return keys


def scan_backup_seqc_files(backup_folders: List[Path], log_func: Optional[Callable[[str], None]] = None) -> Set[str]:
    """
    Scan all .seqc files in backup folders and extract DialogStrKey values.

    Returns: Set of all DialogStrKey values found (these are "unused" strings).
    """
    all_keys = set()
    seqc_count = 0

    for backup_folder in backup_folders:
        for seqc_file in backup_folder.rglob("*.seqc"):
            if not seqc_file.is_file():
                continue

            keys = extract_dialog_str_keys_from_seqc(seqc_file)
            if keys:
                all_keys.update(keys)
                seqc_count += 1

    if log_func:
        log_func(f"Scanned {seqc_count} .seqc files in backup folders")
        log_func(f"Found {len(all_keys)} unique DialogStrKey values (unused strings)")

    return all_keys


def identify_unused_string_ids(
    unused_dialog_keys: Set[str],
    export_folder: Path,
    log_func: Optional[Callable[[str], None]] = None
) -> Set[str]:
    """
    Cross-reference DialogStrKey values from backup folders with StringIds
    in the export folder.

    Returns: Set of StringId values that are "unused" (found in backup .seqc files).
    """
    unused_sids = set()

    for xml_file in export_folder.rglob("*.xml"):
        root = parse_xml_file(xml_file)
        if root is None:
            continue

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid and sid in unused_dialog_keys:
                unused_sids.add(sid)

    if log_func:
        log_func(f"Matched {len(unused_sids)} StringIds as unused (in backup folders)")

    return unused_sids


# ----------------------------------------------------------------------
#  BUILD EXPORT INDEX  (StringId → refined category)
# ----------------------------------------------------------------------
def build_export_index(export_folder: Path) -> Dict[str, str]:
    """
    • Dialog & Sequencer keep 2-level (and sometimes 3-level) detail
    • System keeps 2-level so we can spot ItemGroup/Gimmick/MultiChange
    • Everything else → 1st folder name
    """
    index: Dict[str, str] = {}
    for xml in export_folder.rglob("*.xml"):
        try:
            root = parse_xml_file(xml)
            if root is None:
                continue
        except Exception:
            continue

        parts = xml.relative_to(export_folder).parts
        if not parts:
            continue

        p0 = parts[0].lower()

        # -------------------- Dialog  (Dialog/<sub>)
        if p0 == "dialog" and len(parts) > 1:
            cat = f"Dialog/{parts[1]}"

        # -------------------- Sequencer
        elif p0 == "sequencer" and len(parts) > 1:
            p1 = parts[1].lower()

            #  Sequencer/Other/QuestGroup_Xx
            if p1 == "other" and len(parts) > 2 and parts[2].lower().startswith("questgroup_"):
                cat = f"Sequencer/Other/{parts[2]}"

            #  Sequencer/Sequencer/cd_seq_memory_xx
            elif p1 == "sequencer" and len(parts) > 2 and Path(parts[2]).stem.lower().startswith("cd_seq_memory_"):
                cat = f"Sequencer/Sequencer/{Path(parts[2]).stem}"

            else:
                cat = f"Sequencer/{parts[1]}"

        # -------------------- System (System/<sub>)
        elif p0 == "system" and len(parts) > 1:
            cat = f"System/{parts[1]}"

        # -------------------- default → top folder
        else:
            cat = parts[0]

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                index[sid] = cat
    return index


# ----------------------------------------------------------------------
#  STYLE HELPERS  (Excel)
# ----------------------------------------------------------------------
def style_header(ws: Any, row_idx: int, headers: List[str]) -> None:
    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(bold=True, color="FFFFFF")
    thick = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    for col, txt in enumerate(headers, 1):
        c = ws.cell(row_idx, col, txt)
        c.fill, c.font, c.border = fill, font, thick
        ws.column_dimensions[c.column_letter].width = max(18, len(str(txt)) + 6)


def style_separator(ws: Any, row_idx: int, num_cols: int) -> None:
    fill = PatternFill("solid", fgColor="FFFF00")
    for col in range(1, num_cols + 1):
        ws.cell(row_idx, col).fill = fill


# ----------------------------------------------------------------------
#  GUI  APPLICATION
# ----------------------------------------------------------------------
class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Translation Utilities – WordCount with Unused Detection")
        self.root.geometry("1100x820")

        # -------- buttons
        fbtn = tk.Frame(root); fbtn.pack(pady=10)
        tk.Button(fbtn, text="Word-Count Process", width=25,
                  command=self.start_word_count).pack(side=tk.LEFT, padx=10)
        tk.Button(fbtn, text="StrOrigin Comparison", width=25,
                  command=self.start_strorigin_comparison).pack(side=tk.LEFT, padx=10)

        # -------- progress + log
        self.progress = ttk.Progressbar(root, mode="determinate")
        self.progress.pack(fill=tk.X, padx=15, pady=5)

        tk.Label(root, text="Log / Output:").pack(anchor=tk.W, padx=15)
        self.out = scrolledtext.ScrolledText(root, height=36)
        self.out.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

    # -------------------------------------------------- tiny logger
    def log(self, *msg: Any) -> None:
        self.out.insert(tk.END, " ".join(str(x) for x in msg) + "\n")
        self.out.see(tk.END)
        self.root.update_idletasks()

    # ==============================================================
    #  1)  WORD-COUNT  PROCESS  (Enhanced with Unused String Detection)
    # ==============================================================
    def start_word_count(self) -> None:
        threading.Thread(target=self.run_word_count, daemon=True).start()

    def run_word_count(self) -> None:
        self.out.delete(1.0, tk.END)
        ts0 = datetime.now()
        self.log(f"[{ts0:%Y-%m-%d %H:%M:%S}] Word-count process started")
        self.log(f"LANGUAGE_FOLDER = {LANGUAGE_FOLDER}")
        self.log(f"EXPORT_FOLDER   = {EXPORT_FOLDER}\n")
        self.progress['value'] = 0

        # ---------- Phase 1: Scan backup folders for unused strings
        self.log("=" * 60)
        self.log("PHASE 1: Scanning __backup folders for unused strings...")
        self.log("=" * 60)

        backup_folders = find_backup_folders(BACKUP_SCAN_ROOTS)
        self.log(f"Found {len(backup_folders)} __backup folders:")
        for bf in backup_folders[:10]:  # Show first 10
            self.log(f"  • {bf}")
        if len(backup_folders) > 10:
            self.log(f"  ... and {len(backup_folders) - 10} more")

        # Extract DialogStrKey values from .seqc files in backup folders
        unused_dialog_keys = scan_backup_seqc_files(backup_folders, self.log)

        # Cross-reference with export to get unused StringIds
        self.log("\nCross-referencing with export folder...")
        unused_sids = identify_unused_string_ids(unused_dialog_keys, EXPORT_FOLDER, self.log)
        self.log(f"\nTotal unused StringIds identified: {len(unused_sids)}\n")

        # ---------- Phase 2: Build export index
        self.log("=" * 60)
        self.log("PHASE 2: Building export index...")
        self.log("=" * 60)
        export_index = build_export_index(EXPORT_FOLDER)
        self.log(f"Indexed {len(export_index):,} StringIds\n")

        files = list(iter_language_files(LANGUAGE_FOLDER))
        if not files:
            messagebox.showerror("Word-count", "No LanguageData_*.xml files found.")
            return

        DIFF_OUTPUT_FOLDER.mkdir(exist_ok=True)

        per_file_rows   = []      # for Sheet-1
        category_stats  = {}      # lang → cat → totals
        global_missing  = []      # nodes missing from export

        total_files = len(files)

        # ---------- Phase 3: Process language files
        self.log("=" * 60)
        self.log("PHASE 3: Processing language files...")
        self.log("=" * 60)

        for idx, xml_path in enumerate(files, 1):
            self.progress['value'] = idx / total_files * 80  # Reserve 20% for final processing
            self.log(f"({idx}/{total_files}) {xml_path.name}")

            lang = xml_path.stem.split("_", 1)[1].upper()
            if lang == "KOR":                      # skip source Korean
                continue

            try:
                tree = parse_xml_file(xml_path)
                if tree is None:
                    continue
            except Exception:
                continue

            total_words = completed_origin_words = translated_words = 0
            unused_words = 0  # NEW: Track unused string words
            diff_nodes  = []

            for loc in tree.iter("LocStr"):
                origin = (loc.get("StrOrigin") or "").strip()
                if not origin:
                    continue
                wc_origin = count_words(origin)
                total_words += wc_origin

                sid = loc.get("StringId")
                cat = export_index.get(sid, "Unknown")

                # ---------- Check if this is an unused string
                is_unused = sid in unused_sids

                # ---------- remap System/ItemGroup → Non-Priority, etc.
                if cat.startswith("System/") and cat.split("/", 1)[1] in NON_PRIORITY_FOLDERS:
                    cat = "Non-Priority"

                # ---------- register missing SIDs
                if sid not in export_index:
                    diff_nodes.append(loc)
                    global_missing.append(copy.deepcopy(loc))

                # ---------- update stats
                lang_map = category_stats.setdefault(lang, {})
                st = lang_map.setdefault(cat, {
                    "total_words": 0,
                    "completed_origin_words": 0,
                    "translated_words": 0,
                    "completed_nodes": 0,
                    "total_nodes": 0,
                    "unused_words": 0,          # NEW
                    "unused_nodes": 0           # NEW
                })
                st["total_words"]  += wc_origin
                st["total_nodes"]  += 1

                # Track unused strings
                if is_unused:
                    st["unused_words"] += wc_origin
                    st["unused_nodes"] += 1
                    unused_words += wc_origin

                trans = (loc.get("Str") or "").strip()
                if trans and not is_korean(trans):
                    wc_trans = count_words(trans)
                    completed_origin_words += wc_origin
                    translated_words += wc_trans

                    st["completed_nodes"]         += 1
                    st["completed_origin_words"]  += wc_origin
                    st["translated_words"]        += wc_trans

            per_file_rows.append((lang, xml_path.name,
                                  total_words, completed_origin_words,
                                  translated_words, unused_words))

            # ---------- write DIFF xml per file
            if diff_nodes:
                root_diff = LET.Element("ROOT")
                for d in diff_nodes:
                    root_diff.append(d)
                out_path = DIFF_OUTPUT_FOLDER / f"{xml_path.stem}_DIFF.xml"
                LET.ElementTree(root_diff).write(
                    str(out_path), encoding="utf-8", xml_declaration=True)
                self.log(f"   → diff written : {out_path}")

        # ---------- global missing xml
        if global_missing:
            root_all = LET.Element("ROOT")
            for n in global_missing:
                root_all.append(n)
            LET.ElementTree(root_all).write(str(MISSING_XML_FILE),
                                            encoding="utf-8",
                                            xml_declaration=True)
            self.log(f"\nGlobal missing-from-export XML → {MISSING_XML_FILE}")

        # ===========================================================
        #  EXCEL  REPORT
        # ===========================================================
        self.progress['value'] = 85
        self.log("\n" + "=" * 60)
        self.log("PHASE 4: Generating Excel summary...")
        self.log("=" * 60)
        wb = Workbook()

        # -------------------------------------------------- Sheet-1  (Per-File)
        ws1 = wb.active
        ws1.title = "Per-File"

        # UPDATED headers with Unused Strings and True Coverage
        hdr1 = ["Language", "File",
                "Total Origin Words",
                "Unused Words",              # NEW
                "Effective Words",           # NEW (Total - Unused)
                "Completed Origin Words",
                "Translated Words",
                "Coverage %",                # Original (includes unused)
                "True Coverage %"]           # NEW (excludes unused)
        style_header(ws1, 1, hdr1)

        r = 2
        for lang, fn, tw, cow, tws, uw in sorted(per_file_rows):
            effective_words = tw - uw  # Words that actually need translation
            cov = round(100 * cow / tw, 2) if tw else 0
            true_cov = round(100 * cow / effective_words, 2) if effective_words else 0

            ws1.cell(r, 1, lang)
            ws1.cell(r, 2, fn)
            ws1.cell(r, 3, tw)
            ws1.cell(r, 4, uw)               # Unused Words
            ws1.cell(r, 5, effective_words)  # Effective Words
            ws1.cell(r, 6, cow)
            ws1.cell(r, 7, tws)
            ws1.cell(r, 8, cov)
            ws1.cell(r, 9, true_cov)         # True Coverage
            r += 1

        # autosize
        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = max(
                18, max(len(str(c.value or "")) for c in col) + 4)

        # -------------------------------------------------- Sheet-2  (Detailed)
        ws2 = wb.create_sheet("Detailed Summary")

        # UPDATED headers
        hdr2 = ["Category",
                "Total Origin Words",
                "Unused Words",              # NEW
                "Effective Words",           # NEW
                "Completed Origin Words",
                "Translated Words",
                "Coverage %",
                "True Coverage %"]           # NEW
        row_idx = 1

        # ------------- helper to aggregate & write a row
        def agg_add(a: Dict[str, int], b: Dict[str, int]) -> None:
            for k in ("total_words", "completed_origin_words", "translated_words", "unused_words"):
                a[k] = a.get(k, 0) + b.get(k, 0)

        def big_category(title: str) -> str:
            """
            Extracts the top-level category string to decide the colour.
            """
            if title.startswith("Dialog"):
                return "Dialog"
            if title.startswith("Sequencer"):
                return "Sequencer"
            if title.startswith("System"):
                return "System"
            if title.startswith("World"):
                return "World"
            if title.startswith("Platform"):
                return "Platform"
            if title.startswith("None"):
                return "None"
            if title.startswith("Non-Priority"):
                return "Non-Priority"
            return ""

        def write_row(title: str, stats: Dict[str, int]) -> None:
            nonlocal row_idx
            tw  = stats.get("total_words", 0)
            cow = stats.get("completed_origin_words", 0)
            tws = stats.get("translated_words", 0)
            uw  = stats.get("unused_words", 0)
            eff = tw - uw  # Effective words
            cov = round(100 * cow / tw, 2) if tw else 0
            true_cov = round(100 * cow / eff, 2) if eff else 0

            for col, val in enumerate([title, tw, uw, eff, cow, tws, cov, true_cov], 1):
                cell = ws2.cell(row_idx, col, val)
                cat = big_category(title)
                if cat in BIG_CAT_FILLS:
                    cell.fill = BIG_CAT_FILLS[cat]
            row_idx += 1

        # ------------- iterate each language
        for lang in sorted(category_stats):
            ws2.cell(row_idx, 1, f"Language: {lang}").font = Font(bold=True, size=12)
            row_idx += 1
            style_header(ws2, row_idx, hdr2)
            row_idx += 1

            lang_stats = category_stats[lang]

            # ------------------------------------------------------------------
            #  A)  Dialog  (Grand + subs)
            dialog_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("Dialog/"):
                    agg_add(dialog_total, st)
            if dialog_total["total_words"]:
                write_row("Dialog (Grand Total)", dialog_total)
                for sub in DIALOG_SUBS:
                    key = f"Dialog/{sub}"
                    if key in lang_stats:
                        write_row(key, lang_stats[key])

            # ------------------------------------------------------------------
            #  B)  Sequencer
            seq_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("Sequencer/"):
                    agg_add(seq_total, st)
            if seq_total["total_words"]:
                write_row("Sequencer (Grand Total)", seq_total)

                # ----------- top-level subs
                for top in SEQUENCER_TOPS:
                    pref = f"Sequencer/{top}"
                    sub_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
                    for cat, st in lang_stats.items():
                        if cat == pref or cat.startswith(pref + "/"):
                            agg_add(sub_total, st)
                    if not sub_total["total_words"]:
                        continue
                    title = pref if top not in {"Other", "Sequencer"} else f"{pref} (Total)"
                    write_row(title, sub_total)

                    # ---- Consolidated QuestGroup_*  (Sequencer/Other/QuestGroup_XX)
                    if top == "Other":
                        quest_pref = "Sequencer/Other/QuestGroup_"
                        quest_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
                        for cat, st in lang_stats.items():
                            if cat.startswith(quest_pref):
                                agg_add(quest_total, st)
                        if quest_total["total_words"]:
                            write_row("Sequencer/Other/QuestGroup_XX", quest_total)

                    # ---- Consolidated cd_seq_memory_*  (Sequencer/Sequencer/cd_seq_memory_XX)
                    if top == "Sequencer":
                        mem_pref = "Sequencer/Sequencer/cd_seq_memory_"
                        mem_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
                        for cat, st in lang_stats.items():
                            if cat.startswith(mem_pref):
                                agg_add(mem_total, st)
                        if mem_total["total_words"]:
                            write_row("Sequencer/Sequencer/cd_seq_memory_XX", mem_total)

            # ------------------------------------------------------------------
            #  C)  System  (cleaned)
            sys_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
            for cat, st in lang_stats.items():
                if cat.startswith("System/") and cat.split("/",1)[1] not in NON_PRIORITY_FOLDERS:
                    agg_add(sys_total, st)
            if sys_total["total_words"]:
                write_row("System", sys_total)

            # ------------------------------------------------------------------
            #  D)  World
            if "World" in lang_stats:
                write_row("World", lang_stats["World"])

            # ------------------------------------------------------------------
            #  E)  Platform
            if "Platform" in lang_stats:
                write_row("Platform", lang_stats["Platform"])

            # ------------------------------------------------------------------
            #  F)  None  (merge 'None' + 'Unknown')
            none_total = {"total_words":0,"completed_origin_words":0,"translated_words":0,"unused_words":0}
            if "None" in lang_stats:
                agg_add(none_total, lang_stats["None"])
            if "Unknown" in lang_stats:
                agg_add(none_total, lang_stats["Unknown"])
            if none_total["total_words"]:
                write_row("None", none_total)

            # ------------------------------------------------------------------
            #  G)  Non-Priority
            if "Non-Priority" in lang_stats:
                write_row("Non-Priority", lang_stats["Non-Priority"])

            # --- visual separator
            style_separator(ws2, row_idx, len(hdr2))
            row_idx += 2

        # autosize
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = max(
                18, max(len(str(c.value or "")) for c in col) + 4)

        # -------------------------------------------------- Sheet-3  (Unused Strings Detail)
        ws3 = wb.create_sheet("Unused Strings")
        hdr3 = ["DialogStrKey", "Category", "Source Backup Folder"]
        style_header(ws3, 1, hdr3)

        # Write unused string details (use unused_sids for proper category matching)
        r = 2
        for key in sorted(unused_sids)[:5000]:  # Limit to 5000 for Excel
            cat = export_index.get(key, "Unknown")
            ws3.cell(r, 1, key)
            ws3.cell(r, 2, cat)
            ws3.cell(r, 3, "__backup")
            r += 1

        # autosize
        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = max(
                25, max(len(str(c.value or "")) for c in col) + 4)

        self.log(f"Added {min(len(unused_sids), 5000)} unused strings to Sheet-3")

        # -------------------------------------------------- save workbook
        ts_name = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xlsx = Path.cwd() / f"Translation_Coverage_SUMMARY_{ts_name}.xlsx"
        wb.save(out_xlsx)
        self.log(f"\nExcel saved → {out_xlsx}")
        self.progress['value'] = 100

        # Summary
        self.log("\n" + "=" * 60)
        self.log("SUMMARY")
        self.log("=" * 60)
        self.log(f"Total backup folders scanned: {len(backup_folders)}")
        self.log(f"Total .seqc DialogStrKey values: {len(unused_dialog_keys)}")
        self.log(f"Matched unused StringIds: {len(unused_sids)}")
        self.log(f"Language files processed: {len(files)}")

        messagebox.showinfo("Word-count", f"Done!\nReport:\n{out_xlsx}")

    # ==============================================================
    #  2)  STRORIGIN  COMPARISON  (unchanged from wordcount6)
    # ==============================================================

    def start_strorigin_comparison(self) -> None:
        before = BEFORE_FOLDER
        after  = AFTER_FOLDER

        if not before.is_dir() or not after.is_dir():
            messagebox.showerror(
                "StrOrigin",
                "Required folders not found:\n\n"
                f"BEFORE : {before}\n"
                f"AFTER  : {after}\n\n"
                "Both folders must sit next to the executable."
            )
            return

        common = sorted({d.name for d in before.iterdir() if d.is_dir()} &
                        {d.name for d in after.iterdir()  if d.is_dir()})

        if not common:
            messagebox.showerror("StrOrigin",
                                 "No matching top-level sub-folders found "
                                 "under BEFORE / AFTER.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("StrOrigin Comparison – options")
        dlg.grab_set()

        tk.Label(dlg,
                 text=("Select the sub-folders to compare.\n"
                       "Tick 'Granular' to compare their sub-folders individually:")
                 ).pack(padx=10, pady=8)

        rows = []
        for name in common:
            fr = tk.Frame(dlg); fr.pack(fill=tk.X, padx=15, pady=2)
            sel  = tk.BooleanVar(value=True)
            gran = tk.BooleanVar(value=False)
            tk.Checkbutton(fr, text=name, variable=sel)\
                .pack(side=tk.LEFT)
            tk.Label(fr, text="Granular").pack(side=tk.LEFT, padx=(25,0))
            tk.Checkbutton(fr, variable=gran).pack(side=tk.LEFT)
            rows.append((name, sel, gran))

        ret: Dict[str, Any] = {"sel": None}
        def on_ok() -> None:
            picked = [(n, g.get()) for n, s, g in rows if s.get()]
            if not picked:
                messagebox.showerror("StrOrigin", "Select at least one folder.")
                return
            ret["sel"] = picked
            dlg.destroy()
        tk.Button(dlg, text="OK",     width=10, command=on_ok)\
            .pack(side=tk.LEFT, padx=15, pady=14)
        tk.Button(dlg, text="Cancel", width=10, command=dlg.destroy)\
            .pack(side=tk.RIGHT, padx=15, pady=14)

        self.root.wait_window(dlg)
        if not ret["sel"]:
            return

        todo = [(before / n, after / n, g) for n, g in ret["sel"]]
        threading.Thread(target=self.run_strorigin_report,
                         args=(todo,), daemon=True).start()

    # -------------------------------------------------- helpers (StrOrigin)
    def run_strorigin_report(self, selections: List[Tuple[Path, Path, bool]]) -> None:
        self.out.delete(1.0, tk.END)
        self.progress['value'] = 0
        self.log(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] StrOrigin comparison started")
        total_sel = len(selections)

        global_change = {}
        global_add    = {}
        summary_rows  = []

        files_added_list    = []
        files_modified_list = []

        for idx, (bdir, adir, gran) in enumerate(selections, 1):
            self.progress['value'] = idx / total_sel * 100
            label = f"{bdir.name}"
            self.log(f"\n[{idx}/{total_sel}] {label}")

            if not gran:
                stats, chg, add, added, modified = self.compare_pair_collect_files(
                    bdir, adir, label, global_change, global_add)
                summary_rows.append(stats)
                files_added_list.extend(added)
                files_modified_list.extend(modified)
            else:
                subs_b = {d.name for d in bdir.iterdir() if d.is_dir()}
                subs_a = {d.name for d in adir.iterdir() if d.is_dir()}
                common = sorted(subs_b & subs_a)
                if not common:
                    self.log("  (no common sub-folders)")
                for sub in common:
                    self.log(f"  · {sub}")
                    stats, chg, add, added, modified = self.compare_pair_collect_files(
                        bdir / sub, adir / sub, f"{label}/{sub}",
                        global_change, global_add)
                    summary_rows.append(stats)
                    files_added_list.extend(added)
                    files_modified_list.extend(modified)

        # -------- Excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        hdr = ["Folder", "Files Compared", "Files Modified", "Files Added",
               "Files Deleted", "StrOrigin Changes (Rows)", "Changed Words",
               "StrOrigin Additions (Rows)", "Added Words", "Grand Total Words"]
        header_fill = PatternFill("solid", fgColor="A7C7E7")
        medium = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        for c, h in enumerate(hdr, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = medium

        thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        palette = ["F8CBAD", "C6E0B4", "FFD966",
                   "D9D2E9", "F4B183", "A9D08E", "FFE699"]
        parent_colors, idx_pal = {}, 0

        r = 2
        for row in summary_rows:
            parent = row[0].split("/",1)[0]
            if parent not in parent_colors:
                parent_colors[parent] = palette[idx_pal % len(palette)]
                idx_pal += 1
            fill = PatternFill("solid", fgColor=parent_colors[parent])
            for c, v in enumerate(row, 1):
                cell = ws.cell(r, c, v)
                cell.fill = fill
                cell.border = thin
            r += 1
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                15, max(len(str(c.value or "")) for c in col) + 4)

        def make_top(name: str, data: Dict[str, int]) -> None:
            w = wb.create_sheet(name)
            w.cell(1,1,"File"); w.cell(1,2,"Count")
            w.cell(1,1).font = w.cell(1,2).font = Font(bold=True)
            for i,(f,cnt) in enumerate(sorted(data.items(),
                                              key=lambda x:x[1],
                                              reverse=True)[:50],2):
                w.cell(i,1,f); w.cell(i,2,cnt)
            for col in w.columns:
                w.column_dimensions[col[0].column_letter].width = max(
                    15, max(len(str(c.value or "")) for c in col)+4)
        make_top("Top-50 Changes", global_change)
        make_top("Top-50 Additions", global_add)

        ws_add = wb.create_sheet("Files Added")
        ws_add.cell(1,1,"File Path").font = Font(bold=True)
        for i,f in enumerate(sorted(files_added_list),2):
            ws_add.cell(i,1,f)
        ws_add.column_dimensions['A'].width = max(
            25, max((len(f) for f in files_added_list), default=12)+4)

        ws_mod = wb.create_sheet("Files Modified")
        ws_mod.cell(1,1,"File Path").font = Font(bold=True)
        for i,f in enumerate(sorted(files_modified_list),2):
            ws_mod.cell(i,1,f)
        ws_mod.column_dimensions['A'].width = max(
            25, max((len(f) for f in files_modified_list), default=12)+4)

        out = Path.cwd() / f"StrOrigin_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(out)
        self.log(f"\nExcel saved → {out}")
        self.progress['value'] = 100
        messagebox.showinfo("StrOrigin", f"Done!\nReport:\n{out}")

    # --------------------------------------------------
    def compare_pair_collect_files(
        self,
        before_dir: Path,
        after_dir: Path,
        label: str,
        global_ch: Dict[str, int],
        global_add: Dict[str, int]
    ) -> Tuple[List[Any], Dict[str, int], Dict[str, int], List[str], List[str]]:
        def collect(folder: Path) -> Dict[str, Dict[str, str]]:
            mapping: Dict[str, Dict[str, str]] = {}
            for r, _, fs in os.walk(folder):
                for f in fs:
                    if not f.lower().endswith(".xml"):
                        continue
                    fp  = Path(r) / f
                    rel = fp.relative_to(folder).as_posix()
                    try:
                        rt = ET.parse(fp).getroot()
                        mp: Dict[str, str] = {}
                        for loc in rt.findall(".//LocStr"):
                            sid = loc.get("StringId")
                            so  = loc.get("StrOrigin","")
                            if sid:
                                mp[sid] = so
                        if mp:
                            mapping[rel] = mp
                    except Exception:
                        pass
            return mapping

        bmap = collect(before_dir)
        amap = collect(after_dir)

        bf, af = set(bmap), set(amap)
        common = bf & af

        files_added   = list(af - bf)
        files_deleted = len(bf - af)
        files_modified = 0

        tot_change_rows = tot_change_words = 0
        tot_add_rows    = tot_add_words    = 0

        change_per_file = {}
        add_per_file    = {}
        added_files_list, modified_files_list = [], []

        for rel in common:
            s = bmap[rel]; t = amap[rel]
            changes = [sid for sid in s if sid in t and s[sid] != t[sid]]
            adds    = [sid for sid in t if sid not in s]

            if changes:
                files_modified += 1
                change_per_file[rel] = len(changes)
                tot_change_rows  += len(changes)
                tot_change_words += sum(count_words(t[sid]) for sid in changes)
                modified_files_list.append(f"{label}/{rel}")

            if adds:
                add_per_file[rel] = len(adds)
                tot_add_rows  += len(adds)
                tot_add_words += sum(count_words(t[sid]) for sid in adds)

        for rel in af - bf:
            rows = len(amap[rel])
            if rows:
                add_per_file[rel] = rows
                tot_add_rows  += rows
                tot_add_words += sum(count_words(v) for v in amap[rel].values())
                added_files_list.append(f"{label}/{rel}")

        # ---- global dicts (for Top-50)
        for f, c in change_per_file.items():
            global_ch[f"{label}/{f}"] = global_ch.get(f"{label}/{f}", 0) + c
        for f, c in add_per_file.items():
            global_add[f"{label}/{f}"] = global_add.get(f"{label}/{f}", 0) + c

        self.log(f"    Files compared : {len(common)}")
        self.log(f"    Files modified : {files_modified}")
        self.log(f"    Files added    : {len(files_added)}")
        self.log(f"    Files deleted  : {files_deleted}")
        self.log(f"    StrOrigin changes   : {tot_change_rows} rows / {tot_change_words} words")
        self.log(f"    StrOrigin additions : {tot_add_rows} rows / {tot_add_words} words")

        stats_row = [label, len(common), files_modified, len(files_added),
                     files_deleted, tot_change_rows, tot_change_words,
                     tot_add_rows, tot_add_words,
                     tot_change_words + tot_add_words]
        return stats_row, change_per_file, add_per_file, added_files_list, modified_files_list


# ----------------------------------------------------------------------
def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
