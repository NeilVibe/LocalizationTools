#!/usr/bin/env python3
"""
EventName to StringID Converter
===============================
Simple file dialog tool - select Excel with EventNames, get StringIDs back.

Uses the same XML parsing approach as findeventname8.py:
- Scans ALL XML files in export folder
- Looks for SoundEventName and StringId ATTRIBUTES on any element
"""

import sys
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).parent

# Default export path (F: drive)
DEFAULT_EXPORT_PATH = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")


def get_export_folder() -> Path:
    """Get export folder, checking settings.json for drive letter override."""
    import json
    settings_file = SCRIPT_DIR / "settings.json"

    if settings_file.exists():
        try:
            with open(settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                drive = settings.get('drive_letter', 'F')
                path_str = str(DEFAULT_EXPORT_PATH)
                if path_str[0].upper() == 'F':
                    path_str = f"{drive.upper()}{path_str[1:]}"
                return Path(path_str)
        except:
            pass

    return DEFAULT_EXPORT_PATH


# =============================================================================
# STYLING
# =============================================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FOUND_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FOUND_FONT = Font(color="2E7D32")
UNFOUND_FILL = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
UNFOUND_FONT = Font(color="CC6600")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# =============================================================================
# XML PARSING - Same approach as findeventname8.py
# =============================================================================

def robust_parse_xml(path: Path):
    """Parse XML ignoring BOM/DTD/encoding issues."""
    import re
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as fh:
            content = fh.read()

        content = re.sub(r'^<\?xml[^>]*\?>\s*', '', content, flags=re.MULTILINE)
        content = re.sub(r'<!DOCTYPE[^>]*>\s*', '', content, flags=re.MULTILINE)

        parser = etree.XMLParser(resolve_entities=False, load_dtd=False,
                                  no_network=True, recover=True)
        root = etree.fromstring(content.encode("utf-8"), parser=parser)
        return root
    except:
        return None


def build_eventname_to_stringid_mapping(export_folder: Path) -> dict:
    """
    Scan ALL XML files in export folder.
    Look for SoundEventName and StringId ATTRIBUTES on any element.

    Returns: {eventname_lowercase: stringid}
    """
    mapping = {}

    if not export_folder.exists():
        print(f"ERROR: Export folder not found: {export_folder}")
        return mapping

    xml_files = list(export_folder.rglob("*.xml"))
    print(f"Scanning {len(xml_files)} XML files...")

    for idx, xml_path in enumerate(xml_files, 1):
        if idx % 500 == 0:
            print(f"  Processed {idx}/{len(xml_files)} files...")

        root = robust_parse_xml(xml_path)
        if root is None:
            continue

        # Iterate ALL elements, get ATTRIBUTES
        for node in root.iter():
            # Try multiple attribute name variations (case-insensitive)
            se = (node.get("SoundEventName") or node.get("soundeventname") or
                  node.get("EventName") or node.get("eventname") or "").strip()
            sid = (node.get("StringId") or node.get("StringID") or
                   node.get("stringid") or "").strip()

            if se and sid:
                mapping[se.lower()] = sid

    print(f"Found {len(mapping)} EventName -> StringID mappings")
    return mapping


# =============================================================================
# MAIN
# =============================================================================

def select_input_file() -> Path:
    """Open file dialog to select input Excel."""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    file_path = filedialog.askopenfilename(
        title="Select Excel file with EventNames",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    root.destroy()
    return Path(file_path) if file_path else None


def convert_eventnames(input_path: Path, mapping: dict) -> Path:
    """Convert EventNames to StringIDs."""
    print(f"\nProcessing: {input_path.name}")

    wb = load_workbook(input_path)
    ws = wb.active

    # Create output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Converted"

    # Headers
    out_ws.cell(1, 1, "EventName")
    out_ws.cell(1, 2, "StringID")
    out_ws.cell(1, 3, "Status")
    for col in range(1, 4):
        c = out_ws.cell(1, col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    found = 0
    not_found = 0

    # Process all cells in column 1
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if not val:
            continue

        eventname = str(val).strip()
        if not eventname:
            continue

        # Skip header
        if row == 1 and eventname.upper() in ("EVENTNAME", "EVENT_NAME", "SOUNDEVENTNAME"):
            continue

        out_row = out_ws.max_row + 1
        out_ws.cell(out_row, 1, eventname)
        out_ws.cell(out_row, 1).border = THIN_BORDER

        # Lookup (case-insensitive)
        stringid = mapping.get(eventname.lower())

        if stringid:
            out_ws.cell(out_row, 2, stringid)
            out_ws.cell(out_row, 3, "FOUND")
            for col in range(1, 4):
                out_ws.cell(out_row, col).fill = FOUND_FILL
            out_ws.cell(out_row, 2).font = FOUND_FONT
            out_ws.cell(out_row, 3).font = FOUND_FONT
            found += 1
        else:
            out_ws.cell(out_row, 2, "NOT FOUND")
            out_ws.cell(out_row, 3, "MISSING")
            for col in range(1, 4):
                out_ws.cell(out_row, col).fill = UNFOUND_FILL
                out_ws.cell(out_row, col).font = UNFOUND_FONT
            not_found += 1

        out_ws.cell(out_row, 2).border = THIN_BORDER
        out_ws.cell(out_row, 3).border = THIN_BORDER

    # Column widths
    out_ws.column_dimensions['A'].width = 50
    out_ws.column_dimensions['B'].width = 20
    out_ws.column_dimensions['C'].width = 12

    # Save
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = input_path.parent / f"{input_path.stem}_StringIDs_{timestamp}.xlsx"
    out_wb.save(output_path)

    print(f"\nResults: Found={found}, Not Found={not_found}")
    print(f"Output: {output_path}")

    return output_path


def main():
    print("=" * 60)
    print("EventName to StringID Converter")
    print("=" * 60)

    export_folder = get_export_folder()
    print(f"Export folder: {export_folder}")

    if not export_folder.exists():
        msg = f"Export folder not found:\n{export_folder}\n\nEdit settings.json to set your drive letter."
        print(f"ERROR: {msg}")
        messagebox.showerror("Error", msg)
        return

    # Build mapping
    print("\nBuilding EventName -> StringID mapping...")
    mapping = build_eventname_to_stringid_mapping(export_folder)

    if not mapping:
        messagebox.showerror("Error", "No mappings found in export folder!")
        return

    # Select file
    print("\nSelect your Excel file...")
    input_path = select_input_file()

    if not input_path:
        print("No file selected.")
        return

    # Convert
    try:
        output_path = convert_eventnames(input_path, mapping)
        messagebox.showinfo("Done", f"Output saved:\n{output_path.name}")

        import os
        os.startfile(output_path.parent)
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    main()
