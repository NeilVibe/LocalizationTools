"""
Tracker Facial Sheet Module
=============================
Build Facial sheet from _FACIAL_DATA with Daily + Total + Category sections.

Completely separate from the standard _DAILY_DATA schema which uses
ISSUE/NO ISSUE/BLOCKED/KOREAN columns that don't apply to Face.

Face-specific schema uses: NoIssue, Mismatch, Missing.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List
from collections import defaultdict
import logging

from datetime import datetime, timedelta

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import TRACKER_STYLES

logger = logging.getLogger(__name__)


# =============================================================================
# DATE COMPARISON HELPER
# =============================================================================

def _parse_date_for_comparison(d):
    """
    Parse date value for proper comparison (not string comparison!).

    String comparison fails: "2024-02" > "2024-12" = True (WRONG!)
    This function converts to datetime for proper comparison.
    """
    if d is None:
        return datetime.min
    if isinstance(d, datetime):
        return d
    if isinstance(d, str):
        try:
            return datetime.strptime(d[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            try:
                return datetime.strptime(d[:10], "%d/%m/%Y")
            except (ValueError, TypeError):
                return datetime.min
    try:
        return datetime(1899, 12, 30) + timedelta(days=int(d))
    except Exception:
        return datetime.min


# =============================================================================
# _FACIAL_DATA SCHEMA
# =============================================================================

FACIAL_DATA_HEADERS = [
    "Date",        # 1
    "User",        # 2
    "Group",       # 3
    "TotalRows",   # 4
    "Done",        # 5
    "NoIssue",     # 6
    "Mismatch",    # 7
    "Missing",     # 8
    "Lang",        # 9
]


# =============================================================================
# DATA OPERATIONS
# =============================================================================

def update_facial_data_sheet(wb: openpyxl.Workbook, face_entries: List[Dict]) -> None:
    """
    Update hidden _FACIAL_DATA sheet with Face daily entries.

    Each entry contains per-group breakdowns which get expanded into
    one row per (date, user, group) combination.

    Args:
        wb: Tracker workbook
        face_entries: List of dicts from process_face_category() with:
            {date, user, category, lang, total_rows, done, no_issue, mismatch, missing,
             groups: {group: {total, done, no_issue, mismatch, missing}}}

    Mode: REPLACE - same (date, user, group) overwrites existing row
    """
    # Ensure _FACIAL_DATA sheet exists
    if "_FACIAL_DATA" not in wb.sheetnames:
        ws = wb.create_sheet("_FACIAL_DATA")
        ws.sheet_state = 'hidden'
    else:
        ws = wb["_FACIAL_DATA"]

    # Ensure headers (text format for clean Excel output)
    if ws.cell(1, 1).value != "Date" or ws.max_column < len(FACIAL_DATA_HEADERS):
        for col, header in enumerate(FACIAL_DATA_HEADERS, 1):
            ws.cell(1, col, header).number_format = '@'

    # Build index of existing rows: (date, user, group, lang) -> row_number
    # Normalize date to string to prevent type mismatch (datetime vs str duplicates)
    existing = {}
    for row in range(2, (ws.max_row or 1) + 1):
        date_val = ws.cell(row, 1).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_val = date_val.strftime("%Y-%m-%d")
        else:
            date_val = str(date_val)
        key = (
            date_val,
            ws.cell(row, 2).value,  # User
            ws.cell(row, 3).value,  # Group
            ws.cell(row, 9).value or "EN",  # Lang
        )
        existing[key] = row

    next_row = (ws.max_row or 1) + 1

    for entry in face_entries:
        date = entry["date"]
        user = entry["user"]
        lang = entry.get("lang", "EN")
        groups = entry.get("groups", {})

        if not groups:
            # No group breakdown: write single row with "All" as group
            key = (date, user, "All", lang)
            row = existing.get(key, next_row)
            if row == next_row:
                existing[key] = row
                next_row += 1

            ws.cell(row, 1, date).number_format = '@'
            ws.cell(row, 2, user).number_format = '@'
            ws.cell(row, 3, "All").number_format = '@'
            ws.cell(row, 4, entry.get("total_rows", 0))
            ws.cell(row, 5, entry.get("done", 0))
            ws.cell(row, 6, entry.get("no_issue", 0))
            ws.cell(row, 7, entry.get("mismatch", 0))
            ws.cell(row, 8, entry.get("missing", 0))
            ws.cell(row, 9, lang).number_format = '@'
        else:
            # Write one row per group
            for group, gcounts in groups.items():
                key = (date, user, group, lang)
                row = existing.get(key, next_row)
                if row == next_row:
                    existing[key] = row
                    next_row += 1

                ws.cell(row, 1, date).number_format = '@'
                ws.cell(row, 2, user).number_format = '@'
                ws.cell(row, 3, group).number_format = '@'
                ws.cell(row, 4, gcounts.get("total", 0))
                ws.cell(row, 5, gcounts.get("done", 0))
                ws.cell(row, 6, gcounts.get("no_issue", 0))
                ws.cell(row, 7, gcounts.get("mismatch", 0))
                ws.cell(row, 8, gcounts.get("missing", 0))
                ws.cell(row, 9, lang).number_format = '@'


def read_facial_data(wb: openpyxl.Workbook) -> Dict:
    """
    Read all data from _FACIAL_DATA sheet.

    Returns:
        Dict with:
            - raw_rows: List of row dicts
            - users: Set of all usernames
            - groups: Set of all groups
            - dates: Sorted list of dates
            - langs: Dict of {user: lang}
    """
    if "_FACIAL_DATA" not in wb.sheetnames:
        return {"raw_rows": [], "users": set(), "groups": set(), "dates": [], "langs": {}}

    ws = wb["_FACIAL_DATA"]

    raw_rows = []
    users = set()
    groups = set()
    dates_set = set()
    langs = {}

    for row in range(2, (ws.max_row or 1) + 1):
        date = ws.cell(row, 1).value
        user = ws.cell(row, 2).value
        group = ws.cell(row, 3).value or "Unknown"
        total_rows = int(ws.cell(row, 4).value or 0)
        done = int(ws.cell(row, 5).value or 0)
        no_issue = int(ws.cell(row, 6).value or 0)
        mismatch = int(ws.cell(row, 7).value or 0)
        missing = int(ws.cell(row, 8).value or 0)
        lang = ws.cell(row, 9).value or "EN"

        if date and user:
            raw_rows.append({
                "date": str(date),
                "user": user,
                "group": group,
                "total_rows": total_rows,
                "done": done,
                "no_issue": no_issue,
                "mismatch": mismatch,
                "missing": missing,
                "lang": lang,
            })
            users.add(user)
            groups.add(group)
            dates_set.add(str(date))
            langs[user] = lang

    return {
        "raw_rows": raw_rows,
        "users": users,
        "groups": groups,
        "dates": sorted(dates_set, key=_parse_date_for_comparison),
        "langs": langs,
    }


# =============================================================================
# STYLES
# =============================================================================

def get_facial_styles():
    """Get all styles used in Facial sheet."""
    thin = Side(style='medium', color=TRACKER_STYLES["border_color"])
    thick = Side(style='thick', color='000000')

    return {
        "title_fill": PatternFill(start_color=TRACKER_STYLES["title_color"],
                                  end_color=TRACKER_STYLES["title_color"], fill_type="solid"),
        "header_fill": PatternFill(start_color=TRACKER_STYLES["header_color"],
                                   end_color=TRACKER_STYLES["header_color"], fill_type="solid"),
        "subheader_fill": PatternFill(start_color=TRACKER_STYLES["subheader_color"],
                                      end_color=TRACKER_STYLES["subheader_color"], fill_type="solid"),
        "alt_fill": PatternFill(start_color=TRACKER_STYLES["alt_row_color"],
                                end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid"),
        "total_fill": PatternFill(start_color=TRACKER_STYLES["total_row_color"],
                                  end_color=TRACKER_STYLES["total_row_color"], fill_type="solid"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "thick_top": Border(left=thin, right=thin, top=thick, bottom=thin),
        "thick_side": thick,
        "thin_side": thin,
        "center": Alignment(horizontal='center', vertical='center', wrap_text=True),
        "bold": Font(bold=True),
        "white_bold": Font(bold=True, color="FFFFFF"),
    }


# =============================================================================
# SECTION BUILDERS
# =============================================================================

def _build_facial_daily_section(ws, start_row: int, facial_data: Dict, styles: Dict) -> int:
    """
    Build FACIAL DAILY TABLE section.

    Layout:
    | FACIAL DAILY TABLE                                               |
    |           | User1                    | User2                     |
    | Date      | Done | Mismatch | Missing | Done | Mismatch | Missing|
    | 01/15     |  50  |    3     |    2    |  45  |    5     |    1   |

    Returns:
        Next row after section
    """
    dates = facial_data["dates"]
    users = sorted(facial_data["users"])
    raw_rows = facial_data["raw_rows"]

    if not users or not dates:
        return start_row

    current_row = start_row
    cols_per_user = 3  # Done, Mismatch, Missing
    total_cols = 1 + len(users) * cols_per_user

    # Aggregate data: {date: {user: {done, mismatch, missing}}}
    daily_agg = defaultdict(lambda: defaultdict(lambda: {"done": 0, "mismatch": 0, "missing": 0}))
    for r in raw_rows:
        daily_agg[r["date"]][r["user"]]["done"] += r["done"]
        daily_agg[r["date"]][r["user"]]["mismatch"] += r["mismatch"]
        daily_agg[r["date"]][r["user"]]["missing"] += r["missing"]

    # Title row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_cell = ws.cell(current_row, 1, "FACIAL DAILY TABLE")
    title_cell.fill = styles["title_fill"]
    title_cell.font = styles["white_bold"]
    title_cell.alignment = styles["center"]
    current_row += 1

    # User names row (merged across their columns)
    ws.cell(current_row, 1, "")
    col = 2
    for user in users:
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + cols_per_user - 1)
        cell = ws.cell(current_row, col, user)
        cell.fill = styles["header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                            bottom=styles["thin_side"], right=styles["thin_side"])
        for c in range(col + 1, col + cols_per_user):
            ws.cell(current_row, c).fill = styles["header_fill"]
            ws.cell(current_row, c).border = styles["border"]
        col += cols_per_user
    current_row += 1

    # Sub-headers row
    date_cell = ws.cell(current_row, 1, "Date")
    date_cell.fill = styles["subheader_fill"]
    date_cell.font = styles["bold"]
    date_cell.alignment = styles["center"]
    date_cell.border = styles["border"]

    col = 2
    for user in users:
        for i, header in enumerate(["Done", "Mismatch", "Missing"]):
            cell = ws.cell(current_row, col + i, header)
            cell.fill = styles["subheader_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            if i == 0:
                cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                                    bottom=styles["thin_side"], right=styles["thin_side"])
            else:
                cell.border = styles["border"]
        col += cols_per_user
    current_row += 1

    # Data rows
    for idx, date in enumerate(dates):
        # Format date as MM/DD
        if isinstance(date, str) and len(date) >= 10:
            display_date = date[5:7] + "/" + date[8:10]
        else:
            display_date = str(date)

        date_cell = ws.cell(current_row, 1, display_date)
        date_cell.alignment = styles["center"]
        date_cell.border = styles["border"]
        if idx % 2 == 1:
            date_cell.fill = styles["alt_fill"]

        col = 2
        for user in users:
            user_data = daily_agg.get(date, {}).get(user, {"done": 0, "mismatch": 0, "missing": 0})
            values = [user_data["done"], user_data["mismatch"], user_data["missing"]]
            for i, val in enumerate(values):
                display_val = val if val > 0 else "--"
                cell = ws.cell(current_row, col + i, display_val)
                cell.alignment = styles["center"]
                if i == 0:
                    cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                                        bottom=styles["thin_side"], right=styles["thin_side"])
                else:
                    cell.border = styles["border"]
                if idx % 2 == 1:
                    cell.fill = styles["alt_fill"]
            col += cols_per_user

        current_row += 1

    return current_row + 1  # Spacing


def _build_facial_total_section(ws, start_row: int, facial_data: Dict, styles: Dict, latest_rows: Dict) -> int:
    """
    Build FACIAL TOTAL TABLE section — separate tables for EN and CN.

    Layout (repeated for each language):
    | EN FACIAL TOTAL TABLE                                              |
    | User   | Total | Done | NoIssue | Mismatch | Missing | Done%      |
    | User1  | 100   |  60  |   55    |    3     |    2    | 60%        |
    | TOTAL  |       |  120 |  110    |    5     |    5    | 120.0%     |

    Returns:
        Next row after section
    """
    users = sorted(facial_data["users"])
    langs = facial_data["langs"]

    if not users:
        return start_row

    current_row = start_row
    headers = ["User", "Total", "Done", "NoIssue", "Mismatch", "Missing", "Done%"]
    total_cols = len(headers)

    # Aggregate per user from latest-only rows
    user_totals = defaultdict(lambda: {"total": 0, "done": 0, "no_issue": 0, "mismatch": 0, "missing": 0})
    for r in latest_rows.values():
        user_totals[r["user"]]["total"] += r["total_rows"]
        user_totals[r["user"]]["done"] += r["done"]
        user_totals[r["user"]]["no_issue"] += r["no_issue"]
        user_totals[r["user"]]["mismatch"] += r["mismatch"]
        user_totals[r["user"]]["missing"] += r["missing"]

    # Split users by language
    en_users = sorted([u for u in users if langs.get(u, "EN") == "EN"])
    cn_users = sorted([u for u in users if langs.get(u, "EN") == "CN"])

    for lang_label, lang_users, title_fill in [
        ("EN", en_users, styles.get("en_title_fill", styles["title_fill"])),
        ("CN", cn_users, styles.get("cn_title_fill", styles["title_fill"])),
    ]:
        if not lang_users:
            continue

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, f"{lang_label} FACIAL TOTAL TABLE")
        title_cell.fill = title_fill
        title_cell.font = styles["white_bold"]
        title_cell.alignment = styles["center"]
        current_row += 1

        # Headers
        for i, header in enumerate(headers, 1):
            cell = ws.cell(current_row, i, header)
            cell.fill = styles["header_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

        # User rows
        for idx, user in enumerate(lang_users):
            ut = user_totals[user]
            done_pct = f"{(ut['done'] / ut['total'] * 100):.1f}%" if ut["total"] > 0 else "0%"

            values = [user, ut["total"], ut["done"],
                      ut["no_issue"], ut["mismatch"], ut["missing"], done_pct]

            for i, val in enumerate(values, 1):
                cell = ws.cell(current_row, i, val)
                cell.alignment = styles["center"]
                cell.border = styles["border"]
                if idx % 2 == 1:
                    cell.fill = styles["alt_fill"]

            current_row += 1

        # TOTAL row (sum of individual Done% values)
        grand = {"done": 0, "no_issue": 0, "mismatch": 0, "missing": 0}
        grand_done_pct_sum = 0.0
        for user in lang_users:
            ut = user_totals[user]
            grand["done"] += ut["done"]
            grand["no_issue"] += ut["no_issue"]
            grand["mismatch"] += ut["mismatch"]
            grand["missing"] += ut["missing"]
            grand_done_pct_sum += (ut["done"] / ut["total"] * 100) if ut["total"] > 0 else 0

        grand_pct = f"{grand_done_pct_sum:.1f}%"
        total_values = ["TOTAL", "", grand["done"],
                        grand["no_issue"], grand["mismatch"], grand["missing"], grand_pct]

        for i, val in enumerate(total_values, 1):
            cell = ws.cell(current_row, i, val)
            cell.fill = styles["total_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

        current_row += 1  # Spacing between EN and CN tables

    return current_row


def _build_facial_category_section(ws, start_row: int, facial_data: Dict, styles: Dict, latest_rows: Dict) -> int:
    """
    Build FACIAL CATEGORY TABLE sections — one per language.

    Layout (repeated per language):
    | FACIAL CATEGORY TABLE (EN)                                        |
    | Group       | Total | Done | NoIssue | Mismatch | Missing | Done% |
    | GroupA      |  50   |  40  |   35    |    3     |    2    | 80%   |

    Returns:
        Next row after all language sections
    """
    if not latest_rows:
        return start_row

    # Discover languages present in data
    langs = sorted(set(r["lang"] for r in latest_rows.values()))
    if not langs:
        return start_row

    current_row = start_row
    headers = ["Group", "Total", "Done", "NoIssue", "Mismatch", "Missing", "Done%"]
    total_cols = len(headers)

    for lang in langs:
        # Filter latest_rows for this language
        lang_rows = {k: r for k, r in latest_rows.items() if r["lang"] == lang}
        if not lang_rows:
            continue

        # Per group: aggregate across ALL users (sum latest entry per user)
        # latest_rows already has the latest row per (user, group, lang),
        # so summing across users for each group gives the full picture.
        group_totals = defaultdict(lambda: {"total": 0, "done": 0, "no_issue": 0, "mismatch": 0, "missing": 0})
        for r in lang_rows.values():
            g = r["group"]
            group_totals[g]["total"] += r["total_rows"]
            group_totals[g]["done"] += r["done"]
            group_totals[g]["no_issue"] += r["no_issue"]
            group_totals[g]["mismatch"] += r["mismatch"]
            group_totals[g]["missing"] += r["missing"]

        lang_groups = sorted(group_totals.keys())

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, f"FACIAL CATEGORY TABLE ({lang})")
        title_cell.fill = styles["title_fill"]
        title_cell.font = styles["white_bold"]
        title_cell.alignment = styles["center"]
        current_row += 1

        # Headers
        for i, header in enumerate(headers, 1):
            cell = ws.cell(current_row, i, header)
            cell.fill = styles["header_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

        # Group rows
        for idx, group in enumerate(lang_groups):
            gt = group_totals[group]
            done_pct = f"{(gt['done'] / gt['total'] * 100):.1f}%" if gt["total"] > 0 else "0%"

            values = [group, gt["total"], gt["done"], gt["no_issue"],
                      gt["mismatch"], gt["missing"], done_pct]

            for i, val in enumerate(values, 1):
                cell = ws.cell(current_row, i, val)
                cell.alignment = styles["center"]
                cell.border = styles["border"]
                if idx % 2 == 1:
                    cell.fill = styles["alt_fill"]

            current_row += 1

        current_row += 1  # Spacing between language tables

    return current_row


# =============================================================================
# AUTOFIT HELPERS
# =============================================================================

def _autofit_columns(ws, min_width: int = 8, max_width: int = 50):
    """Autofit column widths based on content."""
    for col_idx in range(1, (ws.max_column or 1) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, min((ws.max_row or 1) + 1, 100)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)

        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# =============================================================================
# MAIN BUILD FUNCTION
# =============================================================================

def build_facial_sheet(wb: openpyxl.Workbook) -> None:
    """
    Build Facial sheet from _FACIAL_DATA.

    Three sections:
    1. FACIAL DAILY TABLE — per-user daily Done/Mismatch/Missing
    2. FACIAL TOTAL TABLE — per-user cumulative with Done%
    3. FACIAL CATEGORY TABLE — per-group breakdown with Done%
    """
    # Delete and recreate (preserve sheet position)
    facial_idx = None
    if "Facial" in wb.sheetnames:
        facial_idx = wb.sheetnames.index("Facial")
        del wb["Facial"]
    ws = wb.create_sheet("Facial")
    if facial_idx is not None:
        wb.move_sheet(ws, offset=facial_idx - (len(wb.sheetnames) - 1))

    # Read _FACIAL_DATA
    facial_data = read_facial_data(wb)

    if not facial_data["users"]:
        ws.cell(1, 1, "No facial data yet")
        return

    styles = get_facial_styles()

    # Pre-compute latest rows per (user, group, lang) once for both Total and Category sections
    # Filters to latest date per (user, group, lang) to avoid double-counting
    # Same pattern as total.py read_latest_data_for_total()
    latest_rows = {}  # (user, group, lang) -> row with latest date
    for r in facial_data["raw_rows"]:
        key = (r["user"], r["group"], r["lang"])
        if key not in latest_rows or _parse_date_for_comparison(r["date"]) > _parse_date_for_comparison(latest_rows[key]["date"]):
            latest_rows[key] = r

    # Build 3 sections
    current_row = 1
    current_row = _build_facial_daily_section(ws, current_row, facial_data, styles)
    current_row = _build_facial_total_section(ws, current_row, facial_data, styles, latest_rows)
    current_row = _build_facial_category_section(ws, current_row, facial_data, styles, latest_rows)

    # Autofit
    _autofit_columns(ws)

    logger.info(f"[FACIAL] Built Facial sheet: {len(facial_data['dates'])} dates, "
                f"{len(facial_data['users'])} users, {len(facial_data['groups'])} groups")
