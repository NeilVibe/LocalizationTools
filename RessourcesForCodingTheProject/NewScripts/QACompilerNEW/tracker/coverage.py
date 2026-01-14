#!/usr/bin/env python3
"""
Standalone Coverage Calculator
==============================
Post-process coverage calculation for monolith datasheet outputs.

Run this AFTER generating all datasheets with the monolith generators.
Reads ENG Excel files from output folders in the same directory.

Usage:
    python coverage_standalone.py
"""

import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Set, Tuple, List
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from lxml import etree as ET

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).parent

# Output folders (relative to script location)
OUTPUT_FOLDERS = {
    "Character":  SCRIPT_DIR / "Character_LQA_All",
    "Quest":      SCRIPT_DIR / "QuestData_Map_All",
    "Item":       SCRIPT_DIR / "ItemData_Map_All" / "Item_Full_LQA",
    "Knowledge":  SCRIPT_DIR / "Knowledge_LQA_All",
    "Skill":      SCRIPT_DIR / "Skill_LQA_All",
    "Region":     SCRIPT_DIR / "Region_LQA_v3",
    "Gimmick":    SCRIPT_DIR / "Gimmick_LQA_Output",
    "Help":       SCRIPT_DIR / "GameAdvice_LQA_All",
}

# ENG file patterns per category
ENG_FILE_PATTERNS = {
    "Character":  "Character_LQA_ENG.xlsx",
    "Quest":      "Quest_LQA_ENG.xlsx",
    "Item":       "*_ENG*.xlsx",  # Multiple files
    "Knowledge":  "Knowledge_LQA_ENG.xlsx",
    "Skill":      "LQA_Skill_ENG.xlsx",
    "Region":     "Region_LQA_ENG.xlsx",
    "Gimmick":    "Gimmick_LQA_ENG.xlsx",
    "Help":       "LQA_GameAdvice_ENG.xlsx",
}

# Korean column index (1-based) per category - column A = 1
# Can be int for single column, or tuple for multiple columns
KOREAN_COLUMN = {
    "Character":  1,            # Original (KR)
    "Quest":      1,            # Original
    "Item":       (6, 8),       # ItemName(KOR)=col F, ItemDesc(KOR)=col H
    "Knowledge":  1,            # Original (KR)
    "Skill":      1,            # Original (KR)
    "Region":     1,            # Original (KR)
    "Gimmick":    (6, 9, 11),   # GimmickName(KOR)=6, ItemName(KOR)=9, ItemDesc(KOR)=11
    "Help":       1,            # Original (KR)
}

# English/Translation column index (1-based) per category
TRANSLATION_COLUMN = {
    "Character":  2,            # English (ENG)
    "Quest":      2,            # ENG
    "Item":       (7, 9),       # ItemName(ENG)=col G, ItemDesc(ENG)=col I
    "Knowledge":  2,            # English (ENG)
    "Skill":      2,            # English (ENG)
    "Region":     2,            # English (ENG)
    "Gimmick":    (7, 10, 12),  # GimmickName(ENG)=7, ItemName(ENG)=10, ItemDesc(ENG)=12
    "Help":       2,            # English (ENG)
}

# Language data folder
LANGUAGE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")

# Export folder (for analyzing unconsumed strings)
EXPORT_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")

# Additional export paths for word count (inherently tested but not in Excel)
EXPORT_LOOKAT_FOLDER = EXPORT_FOLDER / "System" / "LookAt"    # Additional for Item
EXPORT_QUEST_FOLDER = EXPORT_FOLDER / "System" / "Quest"      # Additional for Quest

# Voice Recording Sheet folder
VOICE_RECORDING_FOLDER = Path(r"F:\perforce\cd\mainline\resource\editordata\VoiceRecordingSheet__")

# Non-priority folders (under System/) - excluded from CLEAN report
NON_PRIORITY_FOLDERS = {"ItemGroup", "Gimmick", "MultiChange"}

# Output folder for missing strings XML files
MISSING_STRINGS_OUTPUT = SCRIPT_DIR / "Missing_Strings"

# Threshold for generating XML output (percentage of CLEAN total)
XML_OUTPUT_THRESHOLD_PCT = 1.0

# =============================================================================
# HELPERS
# =============================================================================

_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)


def normalize_placeholders(text: str) -> str:
    """Normalize text: remove placeholder suffixes, collapse whitespace."""
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text


def count_korean_words(text: str) -> int:
    """Count words in Korean text (whitespace-separated tokens)."""
    if not text:
        return 0
    tokens = text.split()
    return len(tokens) if tokens else (1 if text.strip() else 0)


def count_words_in_set(strings: Set[str]) -> int:
    """Count total words in a set of strings."""
    return sum(count_korean_words(s) for s in strings)


# =============================================================================
# XML PARSING (from monolith)
# =============================================================================

_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def _fix_bad_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)


def _preprocess_newlines(raw: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = _fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"',
                 lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                 lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', raw)
    tag_open = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    stack: List[str] = []
    out: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            stack.append(mo.group(1))
            out.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if stack and stack[-1] == mc.group(1):
                stack.pop()
                out.append(line)
            else:
                out.append(stack and f"</{stack.pop()}>" or line)
            continue
        if stripped.startswith("</>"):
            out.append(stack and line.replace("</>", f"</{stack.pop()}>") or line)
            continue
        out.append(line)
    while stack:
        out.append(f"</{stack.pop()}>")
    return "\n".join(out)


def parse_xml_file(path: Path):
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        try:
            return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(recover=True, huge_tree=True))
        except ET.XMLSyntaxError:
            return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_master_language_data(language_folder: Path) -> Tuple[Set[str], int]:
    """Load all StrOrigin values from language data.

    Returns:
        (master_strings set, word_count)
    """
    print(f"Loading master language data from: {language_folder}")

    master_strings: Set[str] = set()

    if not language_folder.exists():
        print(f"  ERROR: Language folder not found")
        return master_strings, 0

    lang_files = sorted(language_folder.glob("languagedata_*.xml"))
    if not lang_files:
        print(f"  ERROR: No language data files found")
        return master_strings, 0

    # Prefer English
    target_file = None
    for f in lang_files:
        if "eng" in f.stem.lower():
            target_file = f
            break
    if target_file is None:
        target_file = lang_files[0]

    print(f"  Using: {target_file.name}")

    root = parse_xml_file(target_file)
    if root is None:
        print(f"  ERROR: Failed to parse language file")
        return master_strings, 0

    for loc in root.iter("LocStr"):
        origin = loc.get("StrOrigin") or ""
        if origin:
            normalized = normalize_placeholders(origin)
            if normalized:
                master_strings.add(normalized)

    total_words = count_words_in_set(master_strings)
    print(f"  Loaded {len(master_strings):,} unique strings ({total_words:,} words)")

    return master_strings, total_words


def load_voice_recording_sheet(folder: Path) -> Set[str]:
    """Load StrOrigin values from VoiceRecordingSheet Excel file."""
    print(f"Loading VoiceRecordingSheet from: {folder}")

    voice_strings: Set[str] = set()

    if not folder.exists():
        print(f"  WARNING: VoiceRecordingSheet folder not found")
        return voice_strings

    excel_files = sorted(folder.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not excel_files:
        print(f"  WARNING: No Excel files found")
        return voice_strings

    target_file = excel_files[0]
    print(f"  Using: {target_file.name}")

    try:
        wb = load_workbook(target_file, read_only=True, data_only=True)

        for sheet in wb.worksheets:
            header_row = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = row
                break

            if header_row is None:
                continue

            str_origin_col = None
            for idx, header in enumerate(header_row):
                if header and "StrOrigin" in str(header):
                    str_origin_col = idx
                    break

            if str_origin_col is None:
                continue

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > str_origin_col:
                    value = row[str_origin_col]
                    if value:
                        normalized = normalize_placeholders(str(value))
                        if normalized:
                            voice_strings.add(normalized)

        wb.close()

    except Exception as e:
        print(f"  ERROR: {e}")

    print(f"  Loaded {len(voice_strings):,} strings from VoiceRecordingSheet")
    return voice_strings


def load_korean_from_excel(xlsx_path: Path, korean_cols = 1) -> Set[str]:
    """Load Korean strings from specified column(s) of Excel file.

    Args:
        xlsx_path: Path to Excel file
        korean_cols: int for single column, or tuple of ints for multiple columns (1-based)
    """
    strings: Set[str] = set()

    # Normalize to tuple
    if isinstance(korean_cols, int):
        cols = (korean_cols,)
    else:
        cols = korean_cols

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if not row:
                    continue
                for col in cols:
                    if len(row) >= col:
                        value = row[col - 1]  # Convert to 0-based
                        if value:
                            normalized = normalize_placeholders(str(value))
                            if normalized:
                                strings.add(normalized)

        wb.close()

    except Exception as e:
        print(f"    ERROR loading {xlsx_path.name}: {e}")

    return strings


def collect_category_strings() -> Dict[str, Set[str]]:
    """Collect Korean strings from all category output folders."""
    print()
    print("=" * 70)
    print("Collecting Korean strings from datasheet outputs")
    print("=" * 70)

    category_strings: Dict[str, Set[str]] = {}

    for category, folder in OUTPUT_FOLDERS.items():
        print(f"\n{category}:")

        if not folder.exists():
            print(f"  SKIP: Folder not found: {folder.name}")
            continue

        pattern = ENG_FILE_PATTERNS.get(category, "*.xlsx")
        korean_col = KOREAN_COLUMN.get(category, 1)

        # Find matching files
        if "*" in pattern:
            files = list(folder.glob(pattern))
        else:
            files = [folder / pattern] if (folder / pattern).exists() else []

        if not files:
            print(f"  SKIP: No ENG files found matching {pattern}")
            continue

        category_set: Set[str] = set()

        for xlsx_file in files:
            strings = load_korean_from_excel(xlsx_file, korean_col)
            category_set.update(strings)
            print(f"  {xlsx_file.name}: {len(strings):,} strings")

        category_strings[category] = category_set
        print(f"  TOTAL: {len(category_set):,} unique strings")

    return category_strings


def load_korean_strings_from_datasheets(output_folder: Path) -> Dict[str, Set[str]]:
    """Load Korean strings from datasheet Excel files in output folder.

    This is a wrapper for GUI compatibility - scans subfolders for ENG Excel files.

    Args:
        output_folder: Path to GeneratedDatasheets folder

    Returns:
        Dict mapping category name -> set of Korean strings
    """
    print(f"\nLoading Korean strings from: {output_folder}")

    category_strings: Dict[str, Set[str]] = {}

    if not output_folder.exists():
        print(f"  ERROR: Output folder not found")
        return category_strings

    # Scan subfolders matching our category patterns
    for category, folder_name in [
        ("Character", "Character_LQA_All"),
        ("Quest", "QuestData_Map_All"),
        ("Item", "ItemData_Map_All"),
        ("Knowledge", "Knowledge_LQA_All"),
        ("Skill", "Skill_LQA_All"),
        ("Region", "Region_LQA_v3"),
        ("Gimmick", "Gimmick_LQA_Output"),
        ("Help", "GameAdvice_LQA_All"),
    ]:
        folder = output_folder / folder_name
        if not folder.exists():
            # Try direct subfolders
            for subfolder in output_folder.iterdir():
                if subfolder.is_dir() and category.lower() in subfolder.name.lower():
                    folder = subfolder
                    break

        if not folder.exists():
            continue

        pattern = ENG_FILE_PATTERNS.get(category, "*.xlsx")
        korean_col = KOREAN_COLUMN.get(category, 1)

        # Find ENG files
        if "*" in pattern:
            files = list(folder.glob(pattern))
        else:
            files = [folder / pattern] if (folder / pattern).exists() else []

        if not files:
            # Try any ENG xlsx
            files = list(folder.glob("*ENG*.xlsx"))

        if not files:
            continue

        category_set: Set[str] = set()
        for xlsx_file in files:
            strings = load_korean_from_excel(xlsx_file, korean_col)
            category_set.update(strings)
            print(f"  {category}/{xlsx_file.name}: {len(strings):,} strings")

        if category_set:
            category_strings[category] = category_set

    total = sum(len(s) for s in category_strings.values())
    print(f"  TOTAL: {total:,} unique strings across {len(category_strings)} categories")

    return category_strings


# =============================================================================
# WORD COUNT TABLE (Korean + Translation from Excel + Additional exports)
# =============================================================================

@dataclass
class CategoryWordCount:
    """Word counts for a category."""
    korean_words: int = 0
    translation_words: int = 0


def load_wordcount_from_excel(xlsx_path: Path, korean_cols, translation_cols) -> CategoryWordCount:
    """Load word counts from specified columns of Excel file.

    Args:
        xlsx_path: Path to Excel file
        korean_cols: int or tuple of ints for Korean columns (1-based)
        translation_cols: int or tuple of ints for Translation columns (1-based)

    Returns:
        CategoryWordCount with Korean and Translation word totals
    """
    # Normalize to tuples
    kr_cols = (korean_cols,) if isinstance(korean_cols, int) else korean_cols
    tr_cols = (translation_cols,) if isinstance(translation_cols, int) else translation_cols

    korean_words = 0
    translation_words = 0

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Count Korean words
                for col in kr_cols:
                    if len(row) >= col and row[col - 1]:
                        korean_words += count_korean_words(str(row[col - 1]))
                # Count Translation words
                for col in tr_cols:
                    if len(row) >= col and row[col - 1]:
                        translation_words += count_korean_words(str(row[col - 1]))

        wb.close()

    except Exception as e:
        print(f"    ERROR loading {xlsx_path.name}: {e}")

    return CategoryWordCount(korean_words=korean_words, translation_words=translation_words)


def collect_category_wordcounts() -> Dict[str, CategoryWordCount]:
    """Collect word counts (Korean + Translation) from all category Excel files."""
    print()
    print("=" * 70)
    print("Collecting word counts from datasheet outputs")
    print("=" * 70)

    category_counts: Dict[str, CategoryWordCount] = {}

    for category, folder in OUTPUT_FOLDERS.items():
        print(f"\n{category}:")

        if not folder.exists():
            print(f"  SKIP: Folder not found: {folder.name}")
            continue

        pattern = ENG_FILE_PATTERNS.get(category, "*.xlsx")
        korean_col = KOREAN_COLUMN.get(category, 1)
        translation_col = TRANSLATION_COLUMN.get(category, 2)

        # Find matching files
        if "*" in pattern:
            files = list(folder.glob(pattern))
        else:
            files = [folder / pattern] if (folder / pattern).exists() else []

        if not files:
            print(f"  SKIP: No ENG files found matching {pattern}")
            continue

        total_kr = 0
        total_tr = 0

        for xlsx_file in files:
            wc = load_wordcount_from_excel(xlsx_file, korean_col, translation_col)
            total_kr += wc.korean_words
            total_tr += wc.translation_words
            print(f"  {xlsx_file.name}: {wc.korean_words:,} KR / {wc.translation_words:,} TR words")

        category_counts[category] = CategoryWordCount(korean_words=total_kr, translation_words=total_tr)
        print(f"  TOTAL: {total_kr:,} KR / {total_tr:,} TR words")

    return category_counts


def load_additional_wordcount(export_folder: Path) -> CategoryWordCount:
    """Load word counts from export XML folder by mapping to LOC ENG.

    Args:
        export_folder: Path to export subfolder (e.g., System/Quest)

    Returns:
        CategoryWordCount with Korean and Translation word totals
    """
    if not export_folder.exists():
        return CategoryWordCount()

    korean_words = 0
    translation_words = 0

    # Build StringID set from export XML
    string_ids: Set[str] = set()
    for xml_file in export_folder.rglob("*.xml"):
        if xml_file.name.lower() == "languagedata.xml":
            continue
        root = parse_xml_file(xml_file)
        if root is None:
            continue
        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                string_ids.add(sid)

    if not string_ids:
        return CategoryWordCount()

    # Map StringIDs to LOC ENG to get Korean + Translation text
    lang_files = sorted(LANGUAGE_FOLDER.glob("languagedata_*.xml"))
    target_file = None
    for f in lang_files:
        if "eng" in f.stem.lower():
            target_file = f
            break
    if target_file is None and lang_files:
        target_file = lang_files[0]

    if target_file is None:
        return CategoryWordCount()

    root = parse_xml_file(target_file)
    if root is None:
        return CategoryWordCount()

    for loc in root.iter("LocStr"):
        sid = loc.get("StringId")
        if sid and sid in string_ids:
            origin = loc.get("StrOrigin") or ""
            trans = loc.get("Str") or ""
            if origin:
                korean_words += count_korean_words(origin)
            if trans:
                translation_words += count_korean_words(trans)

    return CategoryWordCount(korean_words=korean_words, translation_words=translation_words)


def print_wordcount_table(
    category_counts: Dict[str, CategoryWordCount],
    quest_additional: CategoryWordCount = None,
    item_additional: CategoryWordCount = None,
) -> None:
    """Print word count table with category totals and sub-categories."""
    width = 80

    print()
    print("=" * width)
    print("                       WORD COUNT BY CATEGORY")
    print("=" * width)
    print()
    print(f"{'Category':<30} {'Korean Words':>20} {'Translation Words':>20}")
    print("-" * width)

    grand_total_kr = 0
    grand_total_tr = 0

    category_order = ["Character", "Quest", "Item", "Knowledge", "Skill", "Region", "Gimmick", "Help"]

    for category in category_order:
        if category not in category_counts:
            continue

        wc = category_counts[category]
        cat_kr = wc.korean_words
        cat_tr = wc.translation_words

        # Add additional for Quest
        if category == "Quest" and quest_additional:
            cat_kr += quest_additional.korean_words
            cat_tr += quest_additional.translation_words
            print(f"{category + ' (TOTAL)':<30} {cat_kr:>20,} {cat_tr:>20,}")
            print(f"  {'├─ Quest (main)':<27} {wc.korean_words:>20,} {wc.translation_words:>20,}")
            print(f"  {'└─ Additional (System/Quest)':<27} {quest_additional.korean_words:>20,} {quest_additional.translation_words:>20,}")
        # Add additional for Item
        elif category == "Item" and item_additional:
            cat_kr += item_additional.korean_words
            cat_tr += item_additional.translation_words
            print(f"{category + ' (TOTAL)':<30} {cat_kr:>20,} {cat_tr:>20,}")
            print(f"  {'├─ Item (main)':<27} {wc.korean_words:>20,} {wc.translation_words:>20,}")
            print(f"  {'└─ Additional (LookAt)':<27} {item_additional.korean_words:>20,} {item_additional.translation_words:>20,}")
        else:
            print(f"{category:<30} {cat_kr:>20,} {cat_tr:>20,}")

        grand_total_kr += cat_kr
        grand_total_tr += cat_tr

    print("-" * width)
    print(f"{'GRAND TOTAL':<30} {grand_total_kr:>20,} {grand_total_tr:>20,}")
    print("=" * width)
    print()


def export_wordcount_to_excel(
    category_counts: Dict[str, CategoryWordCount],
    quest_additional: CategoryWordCount = None,
    item_additional: CategoryWordCount = None,
    output_folder: Path = None,
) -> Path:
    """Export word count table to a nicely formatted Excel file.

    Args:
        category_counts: Dict of category -> CategoryWordCount
        quest_additional: Additional word count for Quest (System/Quest)
        item_additional: Additional word count for Item (LookAt)
        output_folder: Output folder (default: script directory)

    Returns:
        Path to generated Excel file
    """
    if output_folder is None:
        output_folder = SCRIPT_DIR

    wb = Workbook()
    ws = wb.active
    ws.title = "Word Count by Category"

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", fgColor="FFC000")
    total_font = Font(bold=True, size=11)
    subtotal_fill = PatternFill("solid", fgColor="E2EFDA")
    subtotal_font = Font(bold=True, size=10)
    sub_item_fill = PatternFill("solid", fgColor="F2F2F2")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.cell(1, 1, "WORD COUNT BY CATEGORY").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.merge_cells('A1:C1')
    ws.merge_cells('A2:C2')

    # Headers
    headers = ["Category", "Korean Words", "Translation Words"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    row = 5
    grand_total_kr = 0
    grand_total_tr = 0

    category_order = ["Character", "Quest", "Item", "Knowledge", "Skill", "Region", "Gimmick", "Help"]

    for category in category_order:
        if category not in category_counts:
            continue

        wc = category_counts[category]
        cat_kr = wc.korean_words
        cat_tr = wc.translation_words

        # Quest with additional
        if category == "Quest" and quest_additional:
            cat_kr += quest_additional.korean_words
            cat_tr += quest_additional.translation_words

            # Category total row (green)
            for col in range(1, 4):
                ws.cell(row, col).fill = subtotal_fill
                ws.cell(row, col).font = subtotal_font
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, f"{category} (TOTAL)")
            ws.cell(row, 2, cat_kr).number_format = '#,##0'
            ws.cell(row, 3, cat_tr).number_format = '#,##0'
            row += 1

            # Main sub-item (gray)
            for col in range(1, 4):
                ws.cell(row, col).fill = sub_item_fill
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, "  ├─ Quest (main)")
            ws.cell(row, 2, wc.korean_words).number_format = '#,##0'
            ws.cell(row, 3, wc.translation_words).number_format = '#,##0'
            row += 1

            # Additional sub-item (gray)
            for col in range(1, 4):
                ws.cell(row, col).fill = sub_item_fill
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, "  └─ Additional (System/Quest)")
            ws.cell(row, 2, quest_additional.korean_words).number_format = '#,##0'
            ws.cell(row, 3, quest_additional.translation_words).number_format = '#,##0'
            row += 1

        # Item with additional
        elif category == "Item" and item_additional:
            cat_kr += item_additional.korean_words
            cat_tr += item_additional.translation_words

            # Category total row (green)
            for col in range(1, 4):
                ws.cell(row, col).fill = subtotal_fill
                ws.cell(row, col).font = subtotal_font
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, f"{category} (TOTAL)")
            ws.cell(row, 2, cat_kr).number_format = '#,##0'
            ws.cell(row, 3, cat_tr).number_format = '#,##0'
            row += 1

            # Main sub-item (gray)
            for col in range(1, 4):
                ws.cell(row, col).fill = sub_item_fill
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, "  ├─ Item (main)")
            ws.cell(row, 2, wc.korean_words).number_format = '#,##0'
            ws.cell(row, 3, wc.translation_words).number_format = '#,##0'
            row += 1

            # Additional sub-item (gray)
            for col in range(1, 4):
                ws.cell(row, col).fill = sub_item_fill
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, "  └─ Additional (LookAt)")
            ws.cell(row, 2, item_additional.korean_words).number_format = '#,##0'
            ws.cell(row, 3, item_additional.translation_words).number_format = '#,##0'
            row += 1

        else:
            # Regular category row
            for col in range(1, 4):
                ws.cell(row, col).border = thin_border
            ws.cell(row, 1, category)
            ws.cell(row, 2, cat_kr).number_format = '#,##0'
            ws.cell(row, 3, cat_tr).number_format = '#,##0'
            row += 1

        grand_total_kr += cat_kr
        grand_total_tr += cat_tr

    # Grand total row (orange/yellow)
    row += 1
    for col in range(1, 4):
        ws.cell(row, col).fill = total_fill
        ws.cell(row, col).font = total_font
        ws.cell(row, col).border = thin_border
    ws.cell(row, 1, "GRAND TOTAL")
    ws.cell(row, 2, grand_total_kr).number_format = '#,##0'
    ws.cell(row, 3, grand_total_tr).number_format = '#,##0'

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / f"WordCount_Report_{timestamp}.xlsx"
    wb.save(output_file)

    print(f"\n📊 Excel report saved: {output_file}")
    return output_file


# =============================================================================
# COVERAGE CALCULATION
# =============================================================================

@dataclass
class CategoryCoverage:
    name: str
    unique_strings: int = 0
    word_count: int = 0
    sub_categories: List["CategoryCoverage"] = field(default_factory=list)


@dataclass
class CoverageReport:
    categories: List[CategoryCoverage] = field(default_factory=list)
    total_master_strings: int = 0
    total_master_words: int = 0
    total_covered_strings: int = 0
    total_covered_words: int = 0


def calculate_coverage(
    master_strings: Set[str],
    master_word_count: int,
    category_strings: Dict[str, Set[str]],
    voice_sheet_strings: Set[str],
) -> Tuple[CoverageReport, Set[str]]:
    """Calculate coverage using consume technique.

    Returns:
        (CoverageReport, remaining unconsumed strings)
    """
    print()
    print("Calculating coverage with consume technique...")

    report = CoverageReport(
        total_master_strings=len(master_strings),
        total_master_words=master_word_count,
    )

    remaining = master_strings.copy()

    # Process order
    category_order = ["Character", "Quest", "Item", "Knowledge", "Skill", "Region", "Gimmick", "Help"]

    for category_name in category_order:
        if category_name not in category_strings:
            continue

        cat_strings = category_strings[category_name]

        # Consume: intersection with remaining
        consumed = cat_strings & remaining
        remaining -= consumed

        word_count = count_words_in_set(consumed)

        cat_coverage = CategoryCoverage(
            name=category_name,
            unique_strings=len(consumed),
            word_count=word_count,
        )

        # Quest: add voice sheet as sub-category
        if category_name == "Quest" and voice_sheet_strings:
            voice_consumed = voice_sheet_strings & remaining
            remaining -= voice_consumed
            voice_word_count = count_words_in_set(voice_consumed)

            voice_coverage = CategoryCoverage(
                name="Voice Sheet",
                unique_strings=len(voice_consumed),
                word_count=voice_word_count,
            )
            cat_coverage.sub_categories.append(voice_coverage)

            report.total_covered_strings += len(voice_consumed)
            report.total_covered_words += voice_word_count

        report.categories.append(cat_coverage)
        report.total_covered_strings += len(consumed)
        report.total_covered_words += word_count

    return report, remaining


def print_coverage_report(report: CoverageReport) -> None:
    """Print formatted coverage report."""
    width = 75

    print()
    print("=" * width)
    print("                    LANGUAGE DATA COVERAGE REPORT")
    print("=" * width)
    print()
    print("NOTE: All counts are UNIQUE strings (duplicates removed via normalization)")
    print()

    print(f"{'Category':<20} {'Unique Strings':>15} {'Words Covered':>15} {'% Coverage':>12}")
    print("-" * width)

    for cat in report.categories:
        pct = (cat.unique_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
        print(f"{cat.name:<20} {cat.unique_strings:>15,} {cat.word_count:>15,} {pct:>11.1f}%")

        for sub in cat.sub_categories:
            sub_pct = (sub.unique_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
            print(f"  {'└─ ' + sub.name:<17} {sub.unique_strings:>15,} {sub.word_count:>15,} {sub_pct:>11.1f}%")

    print()
    print("=" * width)

    string_pct = (report.total_covered_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
    word_pct = (report.total_covered_words / report.total_master_words * 100) if report.total_master_words > 0 else 0

    print(f"TOTAL COVERAGE:  {report.total_covered_strings:,} / {report.total_master_strings:,} strings ({string_pct:.1f}%)")
    print(f"                 {report.total_covered_words:,} / {report.total_master_words:,} words ({word_pct:.1f}%)")
    print("=" * width)
    print()


# =============================================================================
# UNCONSUMED STRINGS ANALYSIS
# =============================================================================

def build_export_index(export_folder: Path, depth: int = 2) -> Dict[str, str]:
    """Build index mapping Korean string → category (folder path with subfolder).

    Args:
        export_folder: Path to export folder
        depth: How many folder levels to include (default 2 = TopFolder/SubFolder)

    Returns:
        Dict mapping normalized Korean string → category path (e.g., "World/NPC", "System/UI")
    """
    print(f"\nBuilding export index from: {export_folder}")
    print(f"  Category depth: {depth} levels")

    idx: Dict[str, str] = {}

    if not export_folder.exists():
        print(f"  WARNING: Export folder not found")
        return idx

    xml_count = 0
    for xml in export_folder.rglob("*.xml"):
        # Skip languagedata.xml files - they contain ALL strings, not categorized
        if xml.name.lower() == "languagedata.xml":
            continue

        try:
            root = parse_xml_file(xml)
            if root is None:
                continue
        except Exception:
            continue

        xml_count += 1
        parts = xml.relative_to(export_folder).parts

        # Build category from folder path up to specified depth
        # e.g., depth=2 gives "World/NPC" or "Dialog/MainQuest"
        if parts:
            # Take up to 'depth' folder parts (excluding the filename)
            folder_parts = parts[:-1] if len(parts) > 1 else parts[:1]
            cat_parts = folder_parts[:depth]
            cat = "/".join(cat_parts) if cat_parts else parts[0]
        else:
            cat = "Unknown"

        # Remap System/ItemGroup, System/Gimmick, System/MultiChange → "Non-Priority"
        if cat.startswith("System/"):
            subfolder = cat.split("/", 1)[1] if "/" in cat else ""
            if subfolder in NON_PRIORITY_FOLDERS:
                cat = "Non-Priority"

        # Index by Korean string (StrOrigin), not StringId
        for loc in root.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            if origin:
                normalized = normalize_placeholders(origin)
                if normalized:
                    idx[normalized] = cat

    print(f"  Indexed {len(idx):,} Korean strings from {xml_count:,} XML files")
    return idx


@dataclass
class UnconsumedAnalysis:
    """Analysis of unconsumed strings by category."""
    category_breakdown: Dict[str, int] = field(default_factory=dict)
    category_words: Dict[str, int] = field(default_factory=dict)
    category_strings: Dict[str, Set[str]] = field(default_factory=dict)  # Actual strings per category
    total_strings: int = 0
    total_words: int = 0


def analyze_unconsumed(
    remaining: Set[str],
    export_index: Dict[str, str],
) -> UnconsumedAnalysis:
    """Analyze unconsumed strings by looking up their category in export.

    Args:
        remaining: Set of unconsumed Korean strings
        export_index: Mapping from Korean string → category (from export folder)

    Returns:
        UnconsumedAnalysis with breakdown by category
    """
    print("\nAnalyzing unconsumed strings...")

    analysis = UnconsumedAnalysis(
        total_strings=len(remaining),
        total_words=count_words_in_set(remaining),
    )

    category_strings: Dict[str, Set[str]] = {}

    for origin in remaining:
        cat = export_index.get(origin, "Not in Export")

        if cat not in category_strings:
            category_strings[cat] = set()
        category_strings[cat].add(origin)

    # Calculate counts and words per category
    for cat, strings in category_strings.items():
        analysis.category_breakdown[cat] = len(strings)
        analysis.category_words[cat] = count_words_in_set(strings)
        analysis.category_strings[cat] = strings

    return analysis


def load_languagedata_nodes(export_folder: Path) -> Dict[str, ET._Element]:
    """Load languagedata.xml and build Korean string → LocStr element mapping.

    Args:
        export_folder: Path to export folder containing languagedata.xml

    Returns:
        Dict mapping normalized Korean string → LocStr XML element
    """
    print("\nLoading languagedata.xml for node extraction...")

    nodes: Dict[str, ET._Element] = {}

    # Find languagedata.xml in export folder
    lang_file = export_folder / "languagedata.xml"
    if not lang_file.exists():
        # Try to find it recursively
        lang_files = list(export_folder.rglob("languagedata.xml"))
        if lang_files:
            lang_file = lang_files[0]
        else:
            print("  WARNING: languagedata.xml not found in export folder")
            return nodes

    print(f"  Using: {lang_file}")

    root = parse_xml_file(lang_file)
    if root is None:
        print("  ERROR: Failed to parse languagedata.xml")
        return nodes

    for loc in root.iter("LocStr"):
        origin = loc.get("StrOrigin") or ""
        if origin:
            normalized = normalize_placeholders(origin)
            if normalized:
                nodes[normalized] = loc

    print(f"  Loaded {len(nodes):,} LocStr nodes")
    return nodes


def generate_missing_xml_files(
    analysis: UnconsumedAnalysis,
    lang_nodes: Dict[str, ET._Element],
    output_folder: Path,
    threshold_pct: float = 1.0,
) -> int:
    """Generate XML files for categories above threshold percentage.

    Args:
        analysis: Unconsumed strings analysis
        lang_nodes: Mapping from Korean string → LocStr element
        output_folder: Output folder path
        threshold_pct: Minimum percentage of CLEAN total to generate file

    Returns:
        Number of files generated
    """
    print(f"\nGenerating XML files for categories >= {threshold_pct}% of missing strings...")

    # Categories to exclude (same as CLEAN report)
    excluded_cats = {"Not in Export", "Non-Priority"}

    # Calculate CLEAN total
    clean_total = sum(
        count for cat, count in analysis.category_breakdown.items()
        if cat not in excluded_cats
    )

    if clean_total == 0:
        print("  No CLEAN strings to process")
        return 0

    # Create output folder
    output_folder.mkdir(parents=True, exist_ok=True)

    files_created = 0

    for cat, strings in analysis.category_strings.items():
        if cat in excluded_cats:
            continue

        count = len(strings)
        pct = (count / clean_total * 100) if clean_total > 0 else 0

        if pct < threshold_pct:
            continue

        # Build XML with concatenated LocStr nodes
        root = ET.Element("root")

        nodes_added = 0
        for korean_str in sorted(strings):
            node = lang_nodes.get(korean_str)
            if node is not None:
                # Deep copy the node to avoid modifying original
                root.append(node)
                nodes_added += 1

        if nodes_added == 0:
            continue

        # Generate filename from category (replace / with _)
        safe_cat = cat.replace("/", "_").replace("\\", "_")
        out_file = output_folder / f"{safe_cat}.xml"

        # Write XML file
        tree = ET.ElementTree(root)
        tree.write(str(out_file), encoding="utf-8", xml_declaration=True, pretty_print=True)

        print(f"  {safe_cat}.xml: {nodes_added:,} nodes ({pct:.1f}%)")
        files_created += 1

    print(f"  Generated {files_created} XML files in: {output_folder.name}/")
    return files_created


def print_unconsumed_report(analysis: UnconsumedAnalysis) -> None:
    """Print formatted report of unconsumed strings by subfolder.

    Shows TWO reports:
    - RAW REPORT: All unconsumed strings
    - CLEAN REPORT: Excludes "Not in Export", "Non-Priority"
    """
    width = 90

    # Categories to exclude from CLEAN report
    excluded_cats = {"Not in Export", "Non-Priority"}

    # Calculate CLEAN totals
    clean_strings = 0
    clean_words = 0
    for cat, count in analysis.category_breakdown.items():
        if cat not in excluded_cats:
            clean_strings += count
            clean_words += analysis.category_words.get(cat, 0)

    # Sort by count descending
    sorted_cats = sorted(
        analysis.category_breakdown.items(),
        key=lambda x: x[1],
        reverse=True
    )

    # =========================================================================
    # RAW REPORT
    # =========================================================================
    print()
    print("=" * width)
    print("              UNCONSUMED STRINGS - RAW REPORT (All Categories)")
    print("=" * width)
    print()
    print("NOTE: All counts are UNIQUE strings (no duplicates)")
    print()

    print(f"Total unconsumed (RAW): {analysis.total_strings:,} unique strings ({analysis.total_words:,} words)")
    print()

    print(f"{'Folder/Subfolder':<45} {'Strings':>12} {'Words':>12} {'% of RAW':>15}")
    print("-" * width)

    for cat, count in sorted_cats:
        words = analysis.category_words.get(cat, 0)
        pct = (count / analysis.total_strings * 100) if analysis.total_strings > 0 else 0
        # Truncate long paths
        display_cat = cat if len(cat) <= 44 else cat[:41] + "..."
        # Mark excluded categories
        marker = " [EXCLUDED]" if cat in excluded_cats else ""
        print(f"{display_cat:<45} {count:>12,} {words:>12,} {pct:>14.1f}%{marker}")

    print("=" * width)

    # =========================================================================
    # CLEAN REPORT
    # =========================================================================
    print()
    print("=" * width)
    print("              UNCONSUMED STRINGS - CLEAN REPORT (Priority Only)")
    print("=" * width)
    print()
    print("EXCLUDED: Not in Export, Non-Priority")
    print("          (Non-Priority = System/ItemGroup, System/Gimmick, System/MultiChange)")
    print()

    print(f"Total unconsumed (CLEAN): {clean_strings:,} unique strings ({clean_words:,} words)")
    print()

    print(f"{'Folder/Subfolder':<45} {'Strings':>12} {'Words':>12} {'% Missing':>15}")
    print("-" * width)

    for cat, count in sorted_cats:
        if cat in excluded_cats:
            continue
        words = analysis.category_words.get(cat, 0)
        pct = (count / clean_strings * 100) if clean_strings > 0 else 0
        # Truncate long paths
        display_cat = cat if len(cat) <= 44 else cat[:41] + "..."
        print(f"{display_cat:<45} {count:>12,} {words:>12,} {pct:>14.1f}%")

    print("=" * width)
    print()


# =============================================================================
# MAIN
# =============================================================================

def run_coverage_analysis(
    language_folder: Path = None,
    voice_sheet_folder: Path = None,
    category_strings: Dict[str, Set[str]] = None,
) -> CoverageReport:
    """Run coverage analysis with provided data.

    Args:
        language_folder: Path to language data folder (default: LANGUAGE_FOLDER)
        voice_sheet_folder: Path to voice recording sheet folder (default: VOICE_RECORDING_FOLDER)
        category_strings: Pre-loaded category strings (if None, loads from OUTPUT_FOLDERS)

    Returns:
        CoverageReport with coverage results
    """
    # Use defaults if not provided
    lang_folder = language_folder or LANGUAGE_FOLDER
    voice_folder = voice_sheet_folder or VOICE_RECORDING_FOLDER

    # Load master language data
    master_strings, master_word_count = load_master_language_data(lang_folder)
    if not master_strings:
        return CoverageReport()

    # Load voice recording sheet
    voice_strings = load_voice_recording_sheet(voice_folder)

    # Use provided category strings or collect from folders
    if category_strings is None:
        category_strings = collect_category_strings()

    if not category_strings:
        return CoverageReport()

    # Calculate coverage
    report, remaining = calculate_coverage(
        master_strings,
        master_word_count,
        category_strings,
        voice_strings,
    )

    # Print report
    print_coverage_report(report)

    return report


def main():
    print()
    print("=" * 70)
    print("         STANDALONE COVERAGE CALCULATOR FOR QA DATASHEETS")
    print("=" * 70)
    print()
    print(f"Script location: {SCRIPT_DIR}")
    print()

    # 1. Load master language data (Korean strings only)
    master_strings, master_word_count = load_master_language_data(LANGUAGE_FOLDER)
    if not master_strings:
        print("ERROR: No master language data - cannot calculate coverage")
        sys.exit(1)

    # 2. Load voice recording sheet
    voice_strings = load_voice_recording_sheet(VOICE_RECORDING_FOLDER)

    # 3. Collect Korean strings from all category outputs
    category_strings = collect_category_strings()

    if not category_strings:
        print("\nERROR: No category data found - make sure datasheets are generated first")
        sys.exit(1)

    # 4. Calculate coverage (returns report + remaining unconsumed)
    report, remaining = calculate_coverage(
        master_strings,
        master_word_count,
        category_strings,
        voice_strings,
    )

    # 5. Print coverage report
    print_coverage_report(report)

    # 6. Analyze unconsumed strings (if any)
    if remaining:
        # Build export index to categorize unconsumed strings
        export_index = build_export_index(EXPORT_FOLDER)

        # Analyze where unconsumed strings come from
        unconsumed_analysis = analyze_unconsumed(remaining, export_index)

        # Print unconsumed report
        print_unconsumed_report(unconsumed_analysis)

        # Calculate and print CLEAN COVERAGE SUMMARY
        excluded_cats = {"Not in Export", "Non-Priority"}
        excluded_count = sum(
            count for cat, count in unconsumed_analysis.category_breakdown.items()
            if cat in excluded_cats
        )
        excluded_words = sum(
            unconsumed_analysis.category_words.get(cat, 0)
            for cat in excluded_cats
            if cat in unconsumed_analysis.category_breakdown
        )

        # CLEAN total = master total - excluded unconsumed
        clean_master_total = report.total_master_strings - excluded_count
        clean_master_words = report.total_master_words - excluded_words

        print("=" * 75)
        print("                         COVERAGE SUMMARY")
        print("=" * 75)
        print()

        raw_pct = (report.total_covered_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
        clean_pct = (report.total_covered_strings / clean_master_total * 100) if clean_master_total > 0 else 0

        print(f"RAW COVERAGE:   {report.total_covered_strings:,} / {report.total_master_strings:,} strings ({raw_pct:.1f}%)")
        print(f"CLEAN COVERAGE: {report.total_covered_strings:,} / {clean_master_total:,} strings ({clean_pct:.1f}%)")
        print()
        print(f"  (CLEAN excludes {excluded_count:,} strings from Non-Priority + Not in Export)")
        print("=" * 75)

        # 7. Generate XML files for missing strings (categories >= 1%)
        lang_nodes = load_languagedata_nodes(EXPORT_FOLDER)
        if lang_nodes:
            generate_missing_xml_files(
                unconsumed_analysis,
                lang_nodes,
                MISSING_STRINGS_OUTPUT,
                XML_OUTPUT_THRESHOLD_PCT,
            )
    else:
        print("\n*** 100% COVERAGE - No unconsumed strings! ***\n")

    # ==========================================================================
    # WORD COUNT TABLE (Korean + Translation from Excel + Additional exports)
    # ==========================================================================
    print()
    print("=" * 70)
    print("         WORD COUNT TABLE (Korean + Translation)")
    print("=" * 70)

    # Collect word counts from Excel
    category_wordcounts = collect_category_wordcounts()

    # Load additional word counts from export folders
    print("\nLoading additional word counts from export folders...")
    quest_additional = load_additional_wordcount(EXPORT_QUEST_FOLDER)
    print(f"  System/Quest: {quest_additional.korean_words:,} KR / {quest_additional.translation_words:,} TR words")
    item_additional = load_additional_wordcount(EXPORT_LOOKAT_FOLDER)
    print(f"  System/LookAt: {item_additional.korean_words:,} KR / {item_additional.translation_words:,} TR words")

    # Print word count table
    print_wordcount_table(category_wordcounts, quest_additional, item_additional)

    # Export word count table to Excel
    export_wordcount_to_excel(category_wordcounts, quest_additional, item_additional)

    print("\nDone!")


if __name__ == "__main__":
    main()
