"""
Masterfile Pending Issue Reader
===============================
Given Masterfolder_EN and Masterfolder_CN paths, extract per-tester per-category
status counts from TESTER_STATUS_{user} + STATUS_{user} column pairs.

Single responsibility: read masterfiles -> return structured counts.
"""
from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent.parent))

# Master filename -> category mapping (reverse of CATEGORY_TO_MASTER)
# Master_Script.xlsx contains Sequencer + Dialog sheets -> category "Script"
# Master_System.xlsx contains System + Help sheets -> category "System"
# Master_Item.xlsx contains Item + Gimmick sheets -> category "Item"
# All others: Master_Quest.xlsx -> "Quest", etc.
_MASTER_FILE_TO_CATEGORY = {
    "master_script": "Script",
    "master_system": "System",
    "master_item": "Item",
    "master_quest": "Quest",
    "master_knowledge": "Knowledge",
    "master_region": "Region",
    "master_character": "Character",
    "master_skill": "Skill",
    "master_itemknowledgecluster": "ItemKnowledgeCluster",
    "master_face": "Face",
    "master_contents": "Contents",
}

_TESTER_STATUS_PREFIX = "TESTER_STATUS_"
_STATUS_PREFIX = "STATUS_"

# Sheets to skip
_SKIP_SHEETS = {"STATUS"}

# Categories to EXCLUDE from active pending (same logic as total.py zeroing)
# Script types have rows without STATUS that are NOT pending work.
# Face uses non-standard masterfile format.
_EXCLUDED_CATEGORIES = {"Script", "Face"}


def _extract_category_from_filename(filename: str) -> Optional[str]:
    """Extract category from Master_XXX.xlsx filename."""
    stem = Path(filename).stem.lower()
    return _MASTER_FILE_TO_CATEGORY.get(stem)


def _discover_latest_masterfiles(folder: Path) -> Dict[str, Path]:
    """
    Recursively find Master_*.xlsx files, group by category,
    keep only the latest mtime per category.

    Returns: {category: path}
    """
    if not folder.exists():
        return {}

    candidates: Dict[str, List[Tuple[float, Path]]] = defaultdict(list)

    for path in folder.rglob("Master_*.xlsx"):
        # Skip temp files
        if path.name.startswith("~"):
            continue

        category = _extract_category_from_filename(path.name)
        if category is None:
            logger.debug(f"Skipping unrecognized masterfile: {path.name}")
            continue

        mtime = path.stat().st_mtime
        candidates[category].append((mtime, path))

    # Keep latest per category
    result = {}
    for category, entries in candidates.items():
        entries.sort(key=lambda x: x[0], reverse=True)
        result[category] = entries[0][1]

    return result


def _classify_status(raw_status: Optional[str]) -> str:
    """
    Classify a manager STATUS_{user} value.

    Returns one of: 'pending', 'fixed', 'reported', 'checking', 'nonissue'
    """
    if raw_status is None or str(raw_status).strip() == "":
        return "pending"

    normalized = str(raw_status).strip().upper()

    if normalized == "FIXED":
        return "fixed"
    if normalized == "REPORTED":
        return "reported"
    if normalized == "CHECKING":
        return "checking"
    if normalized in ("NON-ISSUE", "NON ISSUE"):
        return "nonissue"

    # Unknown status -> treat as pending
    return "pending"


def _empty_counts() -> Dict[str, int]:
    """Return a fresh zero-initialized status count dict."""
    return {
        "active_issues": 0,
        "pending": 0,
        "fixed": 0,
        "reported": 0,
        "checking": 0,
        "nonissue": 0,
    }


def _read_master_statuses(
    master_path: Path, category: str, debug_log: List[str] = None
) -> Dict[str, Dict[str, int]]:
    """
    Read TESTER_STATUS_{user} + STATUS_{user} columns from all non-STATUS sheets.

    Returns: {username: {active_issues, pending, fixed, reported, checking, nonissue}}
    """
    result: Dict[str, Dict[str, int]] = {}
    if debug_log is None:
        debug_log = []

    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)
    except Exception as exc:
        debug_log.append(f"  ERROR: Failed to open {master_path}: {exc}")
        return result

    debug_log.append(f"\n  FILE: {master_path.name} (category={category})")
    debug_log.append(f"  SHEETS: {[ws.title for ws in wb.worksheets]}")

    try:
        for ws in wb.worksheets:
            if ws.title in _SKIP_SHEETS:
                debug_log.append(f"    SKIP sheet '{ws.title}' (in _SKIP_SHEETS)")
                continue

            # Read headers from first row using iter_rows
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h) if h is not None else "" for h in row]
                break

            if not headers:
                debug_log.append(f"    SKIP sheet '{ws.title}' (no headers)")
                continue

            # Find TESTER_STATUS_{user}, STATUS_{user}, and comment columns per user
            user_columns: Dict[str, Tuple[int, Optional[int]]] = {}
            # comment_columns[username] = list of 0-based indices for COMMENT/MEMO/SCREENSHOT
            user_comment_cols: Dict[str, List[int]] = {}

            for idx, hdr in enumerate(headers):
                hdr_upper = hdr.upper()
                if hdr_upper.startswith(_TESTER_STATUS_PREFIX.upper()):
                    username = hdr[len(_TESTER_STATUS_PREFIX):].strip()
                    if username:
                        status_col_name = f"{_STATUS_PREFIX}{username}"
                        status_idx = None
                        for sidx, shdr in enumerate(headers):
                            if shdr.upper() == status_col_name.upper():
                                status_idx = sidx
                                break
                        user_columns[username] = (idx, status_idx)

            # Discover comment columns by header name (not position)
            # Matches COMMENT_{user}, MEMO_{user}, SCREENSHOT_{user}
            _COMMENT_PREFIXES = ("COMMENT_", "MEMO_", "SCREENSHOT_")
            for idx, hdr in enumerate(headers):
                hdr_upper = hdr.upper()
                for prefix in _COMMENT_PREFIXES:
                    if hdr_upper.startswith(prefix.upper()):
                        uname = hdr[len(prefix):].strip()
                        if uname in user_columns:
                            if uname not in user_comment_cols:
                                user_comment_cols[uname] = []
                            user_comment_cols[uname].append(idx)

            if not user_columns:
                debug_log.append(f"    SKIP sheet '{ws.title}' (no TESTER_STATUS_ columns)")
                continue

            debug_log.append(f"\n    SHEET '{ws.title}':")
            for u, (ts, ms) in user_columns.items():
                debug_log.append(f"      TESTER: {u} → TESTER_STATUS col={ts}, STATUS col={ms}")

            # Process data rows — with full status value tracking
            row_count = 0
            status_value_tracker: Dict[str, Dict[str, int]] = {}  # user -> {status_value: count}
            sample_rows: Dict[str, List[str]] = {}  # user -> first 5 sample rows
            phantom_tracker: Dict[str, int] = {}  # user -> count of phantom issues (ISSUE without comment)
            # Track unrecognized STATUS values that fall through to "pending"
            fallthrough_tracker: Dict[str, Dict[str, int]] = {}  # user -> {raw_value: count}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_count += 1

                for username, (ts_idx, s_idx) in user_columns.items():
                    tester_status = row[ts_idx] if ts_idx < len(row) else None
                    if tester_status is None:
                        continue
                    ts_str = str(tester_status).strip().upper()
                    if ts_str != "ISSUE":
                        continue

                    # Check named comment columns (COMMENT/MEMO/SCREENSHOT for this user)
                    # ISSUE without any comment = phantom issue → treat as nonissue.
                    comment_cols = user_comment_cols.get(username, [])
                    has_comment = False
                    for c_idx in comment_cols:
                        if c_idx < len(row) and row[c_idx] is not None:
                            val = str(row[c_idx]).strip()
                            if val:
                                has_comment = True
                                break
                    # If no named comment columns found at all, treat as real issue
                    # (can't detect phantom without comment columns)
                    if not comment_cols:
                        has_comment = True

                    if not has_comment:
                        # Phantom issue: ISSUE mark with no comment → auto-nonissue
                        phantom_tracker[username] = phantom_tracker.get(username, 0) + 1
                        # Still count as active_issues for tracking, but classify as nonissue
                        if username not in result:
                            result[username] = _empty_counts()
                        result[username]["active_issues"] += 1
                        result[username]["nonissue"] += 1
                        continue

                    # This row is a REAL active issue for this user (has comment)
                    if username not in result:
                        result[username] = _empty_counts()

                    counts = result[username]
                    counts["active_issues"] += 1

                    # Get manager status
                    manager_status = None
                    if s_idx is not None and s_idx < len(row):
                        manager_status = row[s_idx]

                    # Track raw STATUS value for debug
                    raw_ms = str(manager_status).strip() if manager_status else "(empty)"
                    if username not in status_value_tracker:
                        status_value_tracker[username] = {}
                    status_value_tracker[username][raw_ms] = status_value_tracker[username].get(raw_ms, 0) + 1

                    # Sample first 5 rows per user
                    if username not in sample_rows:
                        sample_rows[username] = []
                    if len(sample_rows[username]) < 5:
                        stringid = row[0] if len(row) > 0 else "?"
                        sample_rows[username].append(f"StringID={stringid} TESTER_STATUS={ts_str} STATUS={raw_ms}")

                    classification = _classify_status(manager_status)

                    # Track unrecognized STATUS values that became "pending"
                    if classification == "pending" and manager_status is not None and str(manager_status).strip() != "":
                        raw_val = str(manager_status).strip()
                        if username not in fallthrough_tracker:
                            fallthrough_tracker[username] = {}
                        fallthrough_tracker[username][raw_val] = fallthrough_tracker[username].get(raw_val, 0) + 1

                    if classification == "checking":
                        counts["checking"] += 1
                        counts["pending"] += 1
                    else:
                        counts[classification] += 1

            debug_log.append(f"      ROWS scanned: {row_count}")
            # Log phantom issues (ISSUE without comment → auto-nonissue)
            if phantom_tracker:
                total_phantoms = sum(phantom_tracker.values())
                debug_log.append(f"      PHANTOM ISSUES (ISSUE without comment → auto-nonissue): {total_phantoms}")
                for u in sorted(phantom_tracker.keys()):
                    debug_log.append(f"        {u}: {phantom_tracker[u]} phantom issues")
            # Log unrecognized STATUS values that fell through to "pending"
            if fallthrough_tracker:
                total_fallthrough = sum(sum(v.values()) for v in fallthrough_tracker.values())
                debug_log.append(f"      *** UNRECOGNIZED STATUS VALUES (treated as PENDING): {total_fallthrough} rows ***")
                for u in sorted(fallthrough_tracker.keys()):
                    for val, cnt in sorted(fallthrough_tracker[u].items(), key=lambda x: -x[1]):
                        debug_log.append(f"        {u}: '{val}' × {cnt} (NOT in whitelist → counted as pending!)")

            for u in sorted(status_value_tracker.keys()):
                debug_log.append(f"      {u} STATUS values (for REAL ISSUE rows only):")
                for sv, cnt in sorted(status_value_tracker[u].items(), key=lambda x: -x[1]):
                    debug_log.append(f"        '{sv}': {cnt} rows")
                if u in sample_rows:
                    debug_log.append(f"      {u} SAMPLE rows:")
                    for s in sample_rows[u]:
                        debug_log.append(f"        {s}")

    finally:
        wb.close()

    return result


def build_pending_from_masterfiles(
    masterfolder_en: Path,
    masterfolder_cn: Path,
) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    Build pending issue counts from masterfiles in EN and CN folders.

    Returns:
        {username: {category: {active_issues, pending, fixed, reported, checking, nonissue}}}
    """
    debug_log: List[str] = []
    debug_log.append("=" * 80)
    debug_log.append("MASTERFILE PENDING — FULL DEBUG LOG")
    debug_log.append("=" * 80)

    # Discover latest masterfiles from both folders (dedup within each folder)
    debug_log.append(f"\nEN FOLDER: {masterfolder_en}")
    debug_log.append(f"CN FOLDER: {masterfolder_cn}")

    en_masters = _discover_latest_masterfiles(masterfolder_en)
    cn_masters = _discover_latest_masterfiles(masterfolder_cn)

    debug_log.append(f"\nEN masterfiles found ({len(en_masters)}):")
    for cat, path in sorted(en_masters.items()):
        debug_log.append(f"  {cat}: {path.name}")
    debug_log.append(f"CN masterfiles found ({len(cn_masters)}):")
    for cat, path in sorted(cn_masters.items()):
        debug_log.append(f"  {cat}: {path.name}")

    # Collect all masterfiles to read
    masters_to_read: List[Tuple[str, Path]] = []
    for cat, path in en_masters.items():
        masters_to_read.append((cat, path))
    for cat, path in cn_masters.items():
        masters_to_read.append((cat, path))

    debug_log.append(f"\nTOTAL masterfiles to process: {len(masters_to_read)}")
    debug_log.append(f"EXCLUDED categories: {_EXCLUDED_CATEGORIES}")

    # Read statuses from each masterfile and merge
    result: Dict[str, Dict[str, Dict[str, int]]] = {}

    for category, master_path in masters_to_read:
        if category in _EXCLUDED_CATEGORIES:
            debug_log.append(f"\n  SKIP: {category} ({master_path.name}) — excluded")
            continue
        user_counts = _read_master_statuses(master_path, category, debug_log)

        for username, counts in user_counts.items():
            if username not in result:
                result[username] = {}
            if category in result[username]:
                existing = result[username][category]
                debug_log.append(f"\n  MERGE: {username}/{category} — adding to existing (EN+CN overlap?)")
                debug_log.append(f"    BEFORE: {existing}")
                debug_log.append(f"    ADDING: {counts}")
                for key in counts:
                    existing[key] += counts[key]
                debug_log.append(f"    AFTER:  {existing}")
            else:
                result[username][category] = counts

    # SUMMARY TABLE
    debug_log.append("\n" + "=" * 80)
    debug_log.append("SUMMARY — PER USER PER CATEGORY")
    debug_log.append("=" * 80)
    grand_active = 0
    grand_pending = 0
    for username in sorted(result.keys()):
        user_active = 0
        user_pending = 0
        for cat in sorted(result[username].keys()):
            c = result[username][cat]
            debug_log.append(f"  {username:20s} | {cat:25s} | issues={c['active_issues']:4d} pending={c['pending']:4d} fixed={c['fixed']:4d} reported={c['reported']:4d} checking={c['checking']:4d} nonissue={c['nonissue']:4d}")
            user_active += c["active_issues"]
            user_pending += c["pending"]
        debug_log.append(f"  {username:20s} | {'--- USER TOTAL ---':25s} | issues={user_active:4d} pending={user_pending:4d}")
        grand_active += user_active
        grand_pending += user_pending
    debug_log.append(f"\n  {'=== GRAND TOTAL ===':20s} | {'ALL':25s} | issues={grand_active:4d} pending={grand_pending:4d}")

    # PENDING BREAKDOWN — where does "pending" come from?
    grand_checking_as_pending = sum(
        c["checking"] for cats in result.values() for c in cats.values()
    )
    grand_truly_pending = grand_pending - grand_checking_as_pending
    debug_log.append(f"\n  PENDING BREAKDOWN:")
    debug_log.append(f"    Total pending:             {grand_pending}")
    debug_log.append(f"    - Truly pending (no status): {grand_truly_pending}")
    debug_log.append(f"    - CHECKING (also pending):   {grand_checking_as_pending}")
    debug_log.append(f"  (If 'truly pending' is too high, check UNRECOGNIZED STATUS VALUES above)")
    debug_log.append("=" * 80)

    # Write to file AND print — use SCRIPT_DIR/logs/ for PyInstaller compat
    from config import SCRIPT_DIR
    log_dir = SCRIPT_DIR / "logs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / "MASTERFILE_PENDING_DEBUG.log"
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(debug_log))
        print(f"\n[DEBUG] Full log written to: {log_path}")
    except Exception as e:
        print(f"\n[DEBUG] Could not write log to {log_path}: {e}")
    # Also print to console
    print("\n".join(debug_log))

    return result


if __name__ == "__main__":
    """Standalone debug: run ONLY masterfile pending analysis."""
    import argparse
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from config import TRACKER_UPDATE_MASTER_EN, TRACKER_UPDATE_MASTER_CN, MASTER_FOLDER_EN, MASTER_FOLDER_CN

    parser = argparse.ArgumentParser(description="Debug masterfile pending counts")
    parser.add_argument("--en", type=str, default=None, help="EN masterfile folder path")
    parser.add_argument("--cn", type=str, default=None, help="CN masterfile folder path")
    parser.add_argument("--use-main", action="store_true", help="Use Masterfolder_EN/CN instead of TrackerUpdateFolder")
    args = parser.parse_args()

    if args.en and args.cn:
        en_path = Path(args.en)
        cn_path = Path(args.cn)
    elif args.use_main:
        en_path = MASTER_FOLDER_EN
        cn_path = MASTER_FOLDER_CN
    else:
        en_path = TRACKER_UPDATE_MASTER_EN
        cn_path = TRACKER_UPDATE_MASTER_CN

    print(f"EN folder: {en_path}")
    print(f"CN folder: {cn_path}")
    print(f"EN exists: {en_path.exists()}")
    print(f"CN exists: {cn_path.exists()}")

    result = build_pending_from_masterfiles(en_path, cn_path)
    print(f"\nDone. Check logs/MASTERFILE_PENDING_DEBUG.log for full details.")
