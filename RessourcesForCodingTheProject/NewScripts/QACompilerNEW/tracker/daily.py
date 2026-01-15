"""
Tracker Daily Sheet Module
==========================
Build DAILY sheet from _DAILY_DATA with EN/CN sections.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Set
from collections import defaultdict

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import TRACKER_STYLES, load_tester_mapping
from tracker.data import read_daily_data, compute_daily_deltas


# =============================================================================
# STYLES
# =============================================================================

def get_daily_styles():
    """Get all styles used in DAILY sheet."""
    thin = Side(style='thin', color=TRACKER_STYLES["border_color"])
    thick = Side(style='medium', color='000000')  # Bold black line

    return {
        "en_title_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "cn_title_fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "manager_title_fill": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "header_fill": PatternFill(start_color=TRACKER_STYLES["header_color"],
                                   end_color=TRACKER_STYLES["header_color"], fill_type="solid"),
        "subheader_fill": PatternFill(start_color=TRACKER_STYLES["subheader_color"],
                                      end_color=TRACKER_STYLES["subheader_color"], fill_type="solid"),
        "alt_fill": PatternFill(start_color=TRACKER_STYLES["alt_row_color"],
                                end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid"),
        # Standard thin border
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        # Thick borders for section separation
        "thick_top": Border(left=thin, right=thin, top=thick, bottom=thin),
        "thick_bottom": Border(left=thin, right=thin, top=thin, bottom=thick),
        "thick_all": Border(left=thick, right=thick, top=thick, bottom=thick),
        "thick_side": thick,
        "thin_side": thin,
        "center": Alignment(horizontal='center', vertical='center', wrap_text=True),
        "bold": Font(bold=True),
        "white_bold": Font(bold=True, color="FFFFFF"),
    }


def autofit_columns(ws, min_width: int = 8, max_width: int = 50):
    """Autofit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)

        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def autofit_row_heights(ws, default_height: int = 15, line_height: int = 15):
    """Autofit row heights based on content with word wrap."""
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                text = str(cell.value)
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 10

                lines = text.count('\n') + 1
                chars_per_line = max(int(col_width * 1.2), 10)
                wrapped_lines = max(1, len(text) // chars_per_line + 1)
                max_lines = max(max_lines, lines, wrapped_lines)

        ws.row_dimensions[row_idx].height = max(default_height, max_lines * line_height)


# Tester columns: Done, Issues, NoIssue, Blocked, Korean, Words/Chars
TESTER_HEADERS = ["Done", "Issues", "NoIssue", "Blocked", "Korean"]
COLS_PER_USER = len(TESTER_HEADERS) + 1  # +1 for Words/Chars


# =============================================================================
# DAILY SECTION BUILDERS
# =============================================================================

def build_daily_section(
    ws,
    start_row: int,
    section_title: str,
    title_fill: PatternFill,
    users_list: List[str],
    is_english: bool,
    dates: List[str],
    daily_delta: Dict,
    styles: Dict
) -> int:
    """
    Build a daily stats section for EN or CN users.

    Args:
        ws: Worksheet to write to
        start_row: Starting row
        section_title: Title text (e.g., "EN DAILY STATS")
        title_fill: Fill color for title
        users_list: List of usernames for this section
        is_english: True for EN (shows "Words"), False for CN (shows "Chars")
        dates: Sorted list of all dates
        daily_delta: Computed daily deltas from compute_daily_deltas()
        styles: Style dict from get_daily_styles()

    Returns:
        Next row after section
    """
    if not users_list:
        return start_row

    current_row = start_row
    word_label = "Words" if is_english else "Chars"

    # Calculate total columns: Date(1) + users * cols_per_user
    total_cols = 1 + len(users_list) * COLS_PER_USER

    default_data = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    }

    # Title row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_cell = ws.cell(current_row, 1, section_title)
    title_cell.fill = title_fill
    title_cell.font = styles["white_bold"]
    title_cell.alignment = styles["center"]
    current_row += 1

    # User names row (merged across their columns)
    ws.cell(current_row, 1, "")  # Empty cell above Date
    col = 2
    for user in users_list:
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + COLS_PER_USER - 1)
        cell = ws.cell(current_row, col, user)
        cell.fill = styles["header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                            bottom=styles["thin_side"], right=styles["thin_side"])
        # Style merged cells
        for c in range(col + 1, col + COLS_PER_USER):
            ws.cell(current_row, c).fill = styles["header_fill"]
            ws.cell(current_row, c).border = styles["border"]
        col += COLS_PER_USER
    current_row += 1

    # Sub-headers row: Date, then per user columns
    date_cell = ws.cell(current_row, 1, "Date")
    date_cell.fill = styles["subheader_fill"]
    date_cell.font = styles["bold"]
    date_cell.alignment = styles["center"]
    date_cell.border = styles["border"]

    col = 2
    for user in users_list:
        user_headers = TESTER_HEADERS + [word_label]
        for i, header in enumerate(user_headers):
            cell = ws.cell(current_row, col + i, header)
            cell.fill = styles["subheader_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            if i == 0:
                cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                                    bottom=styles["thin_side"], right=styles["thin_side"])
            else:
                cell.border = styles["border"]
        col += COLS_PER_USER
    current_row += 1

    # Data rows
    for idx, date in enumerate(dates):
        # Date column - format as MM/DD
        if isinstance(date, str) and len(date) >= 10:
            display_date = date[5:7] + "/" + date[8:10]
        else:
            display_date = str(date)

        date_cell = ws.cell(current_row, 1, display_date)
        date_cell.alignment = styles["center"]
        date_cell.border = styles["border"]
        if idx % 2 == 1:
            date_cell.fill = styles["alt_fill"]

        col = 2
        for user in users_list:
            user_data = daily_delta.get(date, {}).get(user, default_data.copy())
            done_val = user_data["done"]
            issues_val = user_data["issues"]
            no_issue_val = user_data["no_issue"]
            blocked_val = user_data["blocked"]
            korean_val = user_data["korean"]
            word_val = user_data["word_count"]

            values = [done_val, issues_val, no_issue_val, blocked_val, korean_val, word_val]
            for i, val in enumerate(values):
                display_val = val if val > 0 else "--"
                cell = ws.cell(current_row, col + i, display_val)
                cell.alignment = styles["center"]
                if i == 0:
                    cell.border = Border(left=styles["thick_side"], top=styles["thin_side"],
                                        bottom=styles["thin_side"], right=styles["thin_side"])
                else:
                    cell.border = styles["border"]
                if idx % 2 == 1:
                    cell.fill = styles["alt_fill"]
            col += COLS_PER_USER

        current_row += 1

    return current_row + 1  # Extra row for spacing


def build_manager_section(
    ws,
    start_row: int,
    dates: List[str],
    daily_delta: Dict,
    users: Set[str],
    styles: Dict
) -> int:
    """
    Build manager stats section (aggregated across all users).

    Args:
        ws: Worksheet to write to
        start_row: Starting row
        dates: Sorted list of all dates
        daily_delta: Computed daily deltas
        users: Set of all usernames
        styles: Style dict from get_daily_styles()

    Returns:
        Next row after section
    """
    current_row = start_row
    manager_headers = ["Date", "Fixed", "Reported", "NonIssue", "Checking", "Pending"]
    total_cols = len(manager_headers)

    default_data = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    }

    # Title row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_cell = ws.cell(current_row, 1, "MANAGER STATS")
    title_cell.fill = styles["manager_title_fill"]
    title_cell.font = styles["bold"]
    title_cell.alignment = styles["center"]
    current_row += 1

    # Headers row
    for i, header in enumerate(manager_headers, 1):
        cell = ws.cell(current_row, i, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    current_row += 1

    # Data rows
    for idx, date in enumerate(dates):
        if isinstance(date, str) and len(date) >= 10:
            display_date = date[5:7] + "/" + date[8:10]
        else:
            display_date = str(date)

        # Aggregate manager stats across ALL users for this date
        day_fixed = sum(daily_delta.get(date, {}).get(u, default_data.copy())["fixed"] for u in users)
        day_reported = sum(daily_delta.get(date, {}).get(u, default_data.copy())["reported"] for u in users)
        day_checking = sum(daily_delta.get(date, {}).get(u, default_data.copy())["checking"] for u in users)
        day_nonissue = sum(daily_delta.get(date, {}).get(u, default_data.copy())["nonissue"] for u in users)
        day_issues = sum(daily_delta.get(date, {}).get(u, default_data.copy())["issues"] for u in users)
        day_pending = max(0, day_issues - day_fixed - day_reported - day_checking - day_nonissue)

        row_values = [display_date, day_fixed, day_reported, day_nonissue, day_checking, day_pending]
        for i, val in enumerate(row_values, 1):
            display_val = val if (i == 1 or val > 0) else "--"
            cell = ws.cell(current_row, i, display_val)
            cell.alignment = styles["center"]
            cell.border = styles["border"]
            if idx % 2 == 1:
                cell.fill = styles["alt_fill"]

        current_row += 1

    return current_row + 1


# =============================================================================
# MAIN BUILD FUNCTION
# =============================================================================

def build_daily_sheet(wb: openpyxl.Workbook) -> None:
    """
    Build DAILY sheet from _DAILY_DATA.

    Separate EN and CN sections with full status breakdown:
    - Done, Issues, No Issue, Blocked, Korean, Words/Chars
    Plus Manager Stats (Fixed, Reported, Checking, Pending).
    """
    # Delete and recreate sheet to handle merged cells properly
    if "DAILY" in wb.sheetnames:
        del wb["DAILY"]
    ws = wb.create_sheet("DAILY", 0)

    # Load tester mapping for EN/CN separation
    tester_mapping = load_tester_mapping()

    # Read raw data from _DAILY_DATA
    data_result = read_daily_data(wb)
    raw_data = data_result["raw_data"]
    users = data_result["users"]
    categories = data_result["categories"]
    dates = data_result["dates"]

    if not users:
        ws.cell(1, 1, "No data yet")
        return

    # Separate EN and CN users
    en_users = sorted([u for u in users if tester_mapping.get(u, "EN") == "EN"])
    cn_users = sorted([u for u in users if tester_mapping.get(u) == "CN"])

    # Compute daily deltas (the fixed algorithm that prevents cross-category contamination)
    daily_delta = compute_daily_deltas(raw_data, users, categories, dates)

    # Get styles
    styles = get_daily_styles()

    # Build sections
    current_row = 1

    # EN DAILY STATS
    if en_users:
        current_row = build_daily_section(
            ws, current_row, "EN DAILY STATS", styles["en_title_fill"],
            en_users, is_english=True, dates=dates, daily_delta=daily_delta, styles=styles
        )

    # CN DAILY STATS
    if cn_users:
        current_row = build_daily_section(
            ws, current_row, "CN DAILY STATS", styles["cn_title_fill"],
            cn_users, is_english=False, dates=dates, daily_delta=daily_delta, styles=styles
        )

    # MANAGER STATS
    current_row = build_manager_section(ws, current_row, dates, daily_delta, users, styles)

    # Autofit columns and rows
    autofit_columns(ws, min_width=8, max_width=40)
    autofit_row_heights(ws)
