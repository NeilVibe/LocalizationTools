"""
Tracker Total Sheet Module
==========================
Build TOTAL sheet with tester stats, category breakdown, and rankings.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Set, Tuple
from collections import defaultdict

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import TRACKER_STYLES, CATEGORIES, load_tester_mapping


# =============================================================================
# STYLES
# =============================================================================

def get_total_styles():
    """Get all styles used in TOTAL sheet."""
    thin = Side(style='thin', color=TRACKER_STYLES["border_color"])
    thick = Side(style='medium', color='000000')  # Bold black line

    return {
        "title_fill": PatternFill(start_color=TRACKER_STYLES["title_color"],
                                  end_color=TRACKER_STYLES["title_color"], fill_type="solid"),
        "en_title_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "cn_title_fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "header_fill": PatternFill(start_color=TRACKER_STYLES["header_color"],
                                   end_color=TRACKER_STYLES["header_color"], fill_type="solid"),
        "manager_header_fill": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "alt_fill": PatternFill(start_color=TRACKER_STYLES["alt_row_color"],
                                end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid"),
        "total_fill": PatternFill(start_color=TRACKER_STYLES["total_row_color"],
                                  end_color=TRACKER_STYLES["total_row_color"], fill_type="solid"),
        # Standard thin border
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        # Thick borders for section separation
        "thick_top": Border(left=thin, right=thin, top=thick, bottom=thin),
        "thick_bottom": Border(left=thin, right=thin, top=thin, bottom=thick),
        "thick_all": Border(left=thick, right=thick, top=thick, bottom=thick),
        "thick_left": Border(left=thick, right=thin, top=thin, bottom=thin),
        "thick_right": Border(left=thin, right=thick, top=thin, bottom=thin),
        "thick_top_left": Border(left=thick, right=thin, top=thick, bottom=thin),
        "thick_top_right": Border(left=thin, right=thick, top=thick, bottom=thin),
        "thick_bottom_left": Border(left=thick, right=thin, top=thin, bottom=thick),
        "thick_bottom_right": Border(left=thin, right=thick, top=thin, bottom=thick),
        "center": Alignment(horizontal='center', vertical='center', wrap_text=True),
        "bold": Font(bold=True),
        "white_bold": Font(bold=True, color="FFFFFF"),
    }


def autofit_columns(ws, min_width: int = 8, max_width: int = 50):
    """
    Autofit column widths based on content.

    Args:
        ws: Worksheet
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, min(ws.max_row + 1, 100)):  # Sample first 100 rows
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)

        # Add padding and apply constraints
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def autofit_row_heights(ws, default_height: int = 15, line_height: int = 15):
    """
    Autofit row heights based on content with word wrap.

    Args:
        ws: Worksheet
        default_height: Default row height
        line_height: Height per line of text
    """
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                # Count lines based on newlines and column width
                text = str(cell.value)
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 10

                lines = text.count('\n') + 1
                # Estimate wrapped lines
                chars_per_line = max(int(col_width * 1.2), 10)
                wrapped_lines = max(1, len(text) // chars_per_line + 1)
                max_lines = max(max_lines, lines, wrapped_lines)

        ws.row_dimensions[row_idx].height = max(default_height, max_lines * line_height)


# Headers for tester stats
TESTER_HEADERS = ["User", "Done", "Issues", "No Issue", "Blocked", "Korean"]
MANAGER_HEADERS = ["Fixed", "Reported", "Checking", "Pending"]


# =============================================================================
# DATA READING
# =============================================================================

def read_latest_data_for_total(wb: openpyxl.Workbook) -> Tuple[Dict, Dict]:
    """
    Read _DAILY_DATA and extract latest data per (user, category).

    IMPORTANT: _DAILY_DATA stores CUMULATIVE values per (date, user, category).
    We must only use the LATEST date's data for each (user, category) to avoid double-counting.

    Returns:
        Tuple of (latest_data, user_data):
        - latest_data: (user, category) -> {date, total_rows, done, issues, ...}
        - user_data: user -> aggregated stats across all categories
    """
    data_ws = wb["_DAILY_DATA"]

    # First pass: find the latest date for each (user, category)
    latest_data = {}  # (user, category) -> row data

    for row in range(2, data_ws.max_row + 1):
        date = data_ws.cell(row, 1).value
        user = data_ws.cell(row, 2).value
        category = data_ws.cell(row, 3).value

        if not user or not date:
            continue

        key = (user, category)

        # Keep the row with the latest date for each (user, category)
        if key not in latest_data or str(date) > str(latest_data[key]["date"]):
            latest_data[key] = {
                "date": date,
                "category": category,
                "total_rows": data_ws.cell(row, 4).value or 0,
                "done": data_ws.cell(row, 5).value or 0,
                "issues": data_ws.cell(row, 6).value or 0,
                "no_issue": data_ws.cell(row, 7).value or 0,
                "blocked": data_ws.cell(row, 8).value or 0,
                "fixed": data_ws.cell(row, 9).value or 0,
                "reported": data_ws.cell(row, 10).value or 0,
                "checking": data_ws.cell(row, 11).value or 0,
                "nonissue": data_ws.cell(row, 12).value or 0,
                "word_count": data_ws.cell(row, 13).value or 0,
                "korean": data_ws.cell(row, 14).value or 0
            }

    # Second pass: aggregate latest data by user (sum across categories)
    user_data = defaultdict(lambda: {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "korean": 0
    })

    for (user, category), data in latest_data.items():
        user_data[user]["total_rows"] += data["total_rows"]
        user_data[user]["done"] += data["done"]
        user_data[user]["issues"] += data["issues"]
        user_data[user]["no_issue"] += data["no_issue"]
        user_data[user]["blocked"] += data["blocked"]
        user_data[user]["fixed"] += data["fixed"]
        user_data[user]["reported"] += data["reported"]
        user_data[user]["checking"] += data["checking"]
        user_data[user]["nonissue"] += data["nonissue"]
        user_data[user]["korean"] += data.get("korean", 0)

    return latest_data, dict(user_data)


# =============================================================================
# TESTER STATS SECTION
# =============================================================================

def build_tester_section(
    ws,
    start_row: int,
    section_title: str,
    section_fill: PatternFill,
    users_list: List[str],
    user_data: Dict,
    styles: Dict,
    user_row_tracker: List = None
) -> Tuple[int, Dict]:
    """
    Build a tester section (EN or CN) with title, headers, data rows, and subtotal.

    Args:
        ws: Worksheet to write to
        start_row: Starting row
        section_title: Title text (e.g., "EN TESTER STATS")
        section_fill: Fill color for title
        users_list: List of usernames for this section
        user_data: Dict of user -> stats
        styles: Style dict
        user_row_tracker: Optional list to append (row_num, user) tuples

    Returns:
        Tuple of (next_row, section_total_dict)
    """
    if not users_list:
        return start_row, None

    total_cols = len(TESTER_HEADERS) + len(MANAGER_HEADERS)
    current_row = start_row

    # Section Title
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_cell = ws.cell(current_row, 1, section_title)
    title_cell.fill = section_fill
    title_cell.font = styles["white_bold"]
    title_cell.alignment = styles["center"]
    current_row += 1

    # Headers (with thick top border for separation)
    for col, header in enumerate(TESTER_HEADERS, 1):
        cell = ws.cell(current_row, col, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = styles["thick_top"]

    manager_start_col = len(TESTER_HEADERS) + 1
    for col, header in enumerate(MANAGER_HEADERS, manager_start_col):
        cell = ws.cell(current_row, col, header)
        cell.fill = styles["manager_header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = styles["thick_top"]
    current_row += 1

    # Data rows
    section_total = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "pending": 0, "nonissue": 0
    }

    for idx, user in enumerate(users_list):
        data = user_data[user]
        total_rows = data["total_rows"]
        done = data["done"]
        issues = data["issues"]
        no_issue = data["no_issue"]
        blocked = data["blocked"]
        korean = data.get("korean", 0)
        fixed = data["fixed"]
        reported = data["reported"]
        checking = data["checking"]
        nonissue = data["nonissue"]
        # Pending = Issues - Fixed - Reported - Checking - NonIssue
        pending = max(0, issues - fixed - reported - checking - nonissue)

        # Accumulate section totals
        section_total["total_rows"] += total_rows
        section_total["done"] += done
        section_total["issues"] += issues
        section_total["no_issue"] += no_issue
        section_total["blocked"] += blocked
        section_total["korean"] += korean
        section_total["fixed"] += fixed
        section_total["reported"] += reported
        section_total["checking"] += checking
        section_total["pending"] += pending
        section_total["nonissue"] += nonissue

        # Row data: User, Done, Issues, No Issue, Blocked, Korean, Fixed, Reported, Checking, Pending
        row_data = [user, done, issues, no_issue, blocked, korean, fixed, reported, checking, pending]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(current_row, col, value)
            cell.alignment = styles["center"]
            cell.border = styles["border"]
            if idx % 2 == 1:
                cell.fill = styles["alt_fill"]

        # Track this row for chart references
        if user_row_tracker is not None:
            user_row_tracker.append((current_row, user))

        current_row += 1

    # Section subtotal row (with thick bottom border for separation)
    st = section_total
    subtotal_data = [
        "SUBTOTAL", st["done"], st["issues"], st["no_issue"], st["blocked"], st["korean"],
        st["fixed"], st["reported"], st["checking"], st["pending"]
    ]

    for col, value in enumerate(subtotal_data, 1):
        cell = ws.cell(current_row, col, value)
        cell.fill = styles["total_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = styles["thick_bottom"]
    current_row += 1

    return current_row, section_total


# =============================================================================
# RANKING TABLE
# =============================================================================

def build_ranking_table(
    ws,
    start_row: int,
    user_data: Dict,
    tester_mapping: Dict,
    styles: Dict
) -> int:
    """
    Build Ranking Tables with weighted scoring - separate EN and CN rankings.

    Score = 70% Done + 30% Actual Issues (SCALAR values, not percentages)
    - Actual Issues = Issues - NonIssue

    Args:
        ws: Worksheet to write to
        start_row: Row to start building
        user_data: Dict of user -> stats
        tester_mapping: Dict of username -> language
        styles: Style dict

    Returns:
        Next row after section
    """
    if not user_data:
        return start_row

    # Calculate scores for each user (SCALAR based, not percentage)
    en_scores = []
    cn_scores = []
    for user, data in user_data.items():
        done = data.get("done", 0)
        issues = data.get("issues", 0)
        nonissue = data.get("nonissue", 0)

        # Actual Issues (scalar) = Issues - NonIssue
        actual_issues = max(0, issues - nonissue)

        # Weighted score: 70% Done + 30% Actual Issues (SCALAR values)
        score = round(0.7 * done + 0.3 * actual_issues, 1)

        lang = tester_mapping.get(user, "EN")
        user_score = {
            "user": user,
            "lang": lang,
            "done": done,
            "actual_issues": actual_issues,
            "score": score
        }

        # Split by language
        if lang == "CN":
            cn_scores.append(user_score)
        else:
            en_scores.append(user_score)

    # Sort each list by score descending
    en_scores.sort(key=lambda x: (-x["score"], -x["done"], x["user"]))
    cn_scores.sort(key=lambda x: (-x["score"], -x["done"], x["user"]))

    current_row = start_row
    total_cols = 5  # Rank, User, Done, Actual Issues, Score
    headers = ["Rank", "User", "Done", "Actual Issues", "Score"]

    def build_ranking_section(scores_list, title, title_fill):
        """Build a single ranking section (EN or CN)."""
        nonlocal current_row

        if not scores_list:
            return

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, title)
        title_cell.fill = title_fill
        title_cell.font = styles["white_bold"]
        title_cell.alignment = styles["center"]
        current_row += 1

        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.fill = styles["header_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

        # Data rows
        for idx, user_score in enumerate(scores_list):
            rank = idx + 1

            # Rank cell (no medal colors per boss request)
            rank_cell = ws.cell(current_row, 1, rank)
            rank_cell.alignment = styles["center"]
            rank_cell.border = styles["border"]
            if rank <= 3:
                rank_cell.font = styles["bold"]
            if idx % 2 == 1:
                rank_cell.fill = styles["alt_fill"]

            # User
            user_cell = ws.cell(current_row, 2, user_score["user"])
            user_cell.alignment = styles["center"]
            user_cell.border = styles["border"]
            if rank <= 3:
                user_cell.font = styles["bold"]
            if idx % 2 == 1 and rank > 3:
                user_cell.fill = styles["alt_fill"]

            # Done (scalar)
            done_cell = ws.cell(current_row, 3, user_score["done"])
            done_cell.alignment = styles["center"]
            done_cell.border = styles["border"]
            if idx % 2 == 1 and rank > 3:
                done_cell.fill = styles["alt_fill"]

            # Actual Issues (scalar)
            ai_cell = ws.cell(current_row, 4, user_score["actual_issues"])
            ai_cell.alignment = styles["center"]
            ai_cell.border = styles["border"]
            if idx % 2 == 1 and rank > 3:
                ai_cell.fill = styles["alt_fill"]

            # Score
            score_cell = ws.cell(current_row, 5, user_score["score"])
            score_cell.number_format = '0.0'
            score_cell.alignment = styles["center"]
            score_cell.border = styles["border"]
            score_cell.font = styles["bold"]
            if idx % 2 == 1 and rank > 3:
                score_cell.fill = styles["alt_fill"]

            current_row += 1

    # Build EN Ranking
    build_ranking_section(en_scores, "EN RANKING (70% Done + 30% Actual Issues)", styles["en_title_fill"])

    if en_scores and cn_scores:
        current_row += 1  # Gap between tables

    # Build CN Ranking
    build_ranking_section(cn_scores, "CN RANKING (70% Done + 30% Actual Issues)", styles["cn_title_fill"])

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 6   # Rank
    ws.column_dimensions[get_column_letter(2)].width = 12  # User
    ws.column_dimensions[get_column_letter(3)].width = 12  # Done
    ws.column_dimensions[get_column_letter(4)].width = 14  # Actual Issues
    ws.column_dimensions[get_column_letter(5)].width = 8   # Score

    return current_row


# =============================================================================
# CATEGORY BREAKDOWN TABLE
# =============================================================================

def build_category_breakdown_section(
    ws,
    start_row: int,
    latest_data: Dict,
    users_list: List[str],
    is_english: bool,
    tester_mapping: Dict,
    styles: Dict
) -> int:
    """
    Build Category Breakdown pivot table for EN or CN testers.

    Shows per-category completion % and word/character count for each user.

    Args:
        ws: Worksheet to write to
        start_row: Row to start building
        latest_data: Dict of (user, category) -> {done, total_rows, word_count, ...}
        users_list: List of usernames for this section (EN or CN)
        is_english: True for EN (shows "Words"), False for CN (shows "Chars")
        tester_mapping: Dict of username -> language
        styles: Style dict

    Returns:
        Next row after section
    """
    if not users_list:
        return start_row

    # Build pivot: user -> {category -> {pct, count}}
    pivot = {}
    for (user, category), data in latest_data.items():
        if user not in users_list:
            continue
        if user not in pivot:
            pivot[user] = {}
        total_rows = data["total_rows"]
        done = data["done"]
        pct = round(done / total_rows * 100, 1) if total_rows > 0 else 0
        pivot[user][category] = {
            "pct": pct,
            "count": data.get("word_count", 0)
        }

    # Categories to show (use global CATEGORIES)
    categories = CATEGORIES

    # Calculate column count: User + (Done%/Count) * categories + Total
    count_label = "Words" if is_english else "Chars"
    total_cols = 1 + len(categories) * 2 + 1  # User + categories*2 + Total

    current_row = start_row

    # Title row
    title_text = f"EN CATEGORY BREAKDOWN ({count_label})" if is_english else f"CN CATEGORY BREAKDOWN ({count_label})"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_fill = styles["en_title_fill"] if is_english else styles["cn_title_fill"]
    title_cell = ws.cell(current_row, 1, title_text)
    title_cell.fill = title_fill
    title_cell.font = styles["white_bold"]
    title_cell.alignment = styles["center"]
    current_row += 1

    # Header row 1: Category names (merged across 2 columns each)
    ws.cell(current_row, 1, "User").fill = styles["header_fill"]
    ws.cell(current_row, 1).font = styles["bold"]
    ws.cell(current_row, 1).alignment = styles["center"]
    ws.cell(current_row, 1).border = styles["border"]

    col = 2
    for cat in categories:
        # Merge 2 cells for category name
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
        cell = ws.cell(current_row, col, cat)
        cell.fill = styles["header_fill"]
        cell.font = styles["bold"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
        ws.cell(current_row, col + 1).border = styles["border"]
        col += 2

    # Total column header
    ws.cell(current_row, col, f"Total {count_label}").fill = styles["header_fill"]
    ws.cell(current_row, col).font = styles["bold"]
    ws.cell(current_row, col).alignment = styles["center"]
    ws.cell(current_row, col).border = styles["border"]
    current_row += 1

    # Header row 2: Done% / Words|Chars sub-headers
    ws.cell(current_row, 1, "").border = styles["border"]  # Empty under "User"

    col = 2
    for cat in categories:
        cell1 = ws.cell(current_row, col, "Done%")
        cell1.fill = styles["header_fill"]
        cell1.alignment = styles["center"]
        cell1.border = styles["border"]
        cell2 = ws.cell(current_row, col + 1, count_label)
        cell2.fill = styles["header_fill"]
        cell2.alignment = styles["center"]
        cell2.border = styles["border"]
        col += 2

    ws.cell(current_row, col, "").border = styles["border"]  # Empty under Total
    current_row += 1

    # Data rows
    category_totals = {cat: {"done": 0, "total_rows": 0, "count": 0} for cat in categories}
    grand_total_count = 0

    for idx, user in enumerate(sorted(users_list)):
        user_total_count = 0

        # User name
        cell = ws.cell(current_row, 1, user)
        cell.alignment = styles["center"]
        cell.border = styles["border"]
        if idx % 2 == 1:
            cell.fill = styles["alt_fill"]

        col = 2
        for cat in categories:
            user_cat_data = pivot.get(user, {}).get(cat, None)

            if user_cat_data:
                pct = user_cat_data["pct"]
                count = user_cat_data["count"]
                user_total_count += count

                # Aggregate for category totals
                for (u, c), data in latest_data.items():
                    if u == user and c == cat:
                        category_totals[cat]["done"] += data["done"]
                        category_totals[cat]["total_rows"] += data["total_rows"]
                        category_totals[cat]["count"] += data.get("word_count", 0)
                        break

                # Done%
                cell1 = ws.cell(current_row, col, pct)
                cell1.number_format = '0.0"%"'
                cell1.alignment = styles["center"]
                cell1.border = styles["border"]
                if idx % 2 == 1:
                    cell1.fill = styles["alt_fill"]

                # Word/Char count
                cell2 = ws.cell(current_row, col + 1, count)
                cell2.number_format = '#,##0'
                cell2.alignment = styles["center"]
                cell2.border = styles["border"]
                if idx % 2 == 1:
                    cell2.fill = styles["alt_fill"]
            else:
                # No data for this category - show "-"
                cell1 = ws.cell(current_row, col, "-")
                cell1.alignment = styles["center"]
                cell1.border = styles["border"]
                if idx % 2 == 1:
                    cell1.fill = styles["alt_fill"]

                cell2 = ws.cell(current_row, col + 1, "-")
                cell2.alignment = styles["center"]
                cell2.border = styles["border"]
                if idx % 2 == 1:
                    cell2.fill = styles["alt_fill"]

            col += 2

        # Total count for this user
        grand_total_count += user_total_count
        cell = ws.cell(current_row, col, user_total_count)
        cell.number_format = '#,##0'
        cell.alignment = styles["center"]
        cell.border = styles["border"]
        if idx % 2 == 1:
            cell.fill = styles["alt_fill"]

        current_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12  # User column
    col_letter_idx = 2
    for cat in categories:
        ws.column_dimensions[get_column_letter(col_letter_idx)].width = 8  # Done%
        ws.column_dimensions[get_column_letter(col_letter_idx + 1)].width = 10  # Words/Chars
        col_letter_idx += 2
    ws.column_dimensions[get_column_letter(col_letter_idx)].width = 12  # Total

    return current_row


# =============================================================================
# MAIN BUILD FUNCTION
# =============================================================================

def build_total_sheet(wb: openpyxl.Workbook) -> None:
    """
    Build TOTAL sheet from _DAILY_DATA.

    Aggregates by user across all dates and categories.
    Includes both tester stats and manager stats.
    Separates EN and CN testers into distinct sections.
    Includes Category Breakdown and Ranking tables.
    """
    # Delete and recreate sheet to handle merged cells properly
    if "TOTAL" in wb.sheetnames:
        del wb["TOTAL"]
    ws = wb.create_sheet("TOTAL", 1)

    # Load tester mapping to separate EN/CN
    tester_mapping = load_tester_mapping()

    # Read latest data
    latest_data, user_data = read_latest_data_for_total(wb)

    if not user_data:
        ws.cell(1, 1, "No data yet")
        return

    # Separate users by language
    en_users = sorted([u for u in user_data.keys() if tester_mapping.get(u, "EN") == "EN"])
    cn_users = sorted([u for u in user_data.keys() if tester_mapping.get(u) == "CN"])

    # Get styles
    styles = get_total_styles()

    # Build EN section
    current_row = 1
    en_total = None
    user_data_rows = []  # Track (row_num, user) for potential chart references
    if en_users:
        current_row, en_total = build_tester_section(
            ws, current_row, "EN TESTER STATS", styles["en_title_fill"],
            en_users, user_data, styles, user_data_rows
        )
        current_row += 1  # Empty row between sections

    # Build CN section
    cn_total = None
    if cn_users:
        current_row, cn_total = build_tester_section(
            ws, current_row, "CN TESTER STATS", styles["cn_title_fill"],
            cn_users, user_data, styles, user_data_rows
        )
        current_row += 1  # Empty row before grand total

    # Grand total row (combines EN + CN)
    total_cols = len(TESTER_HEADERS) + len(MANAGER_HEADERS)
    grand_total = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "pending": 0, "nonissue": 0
    }
    for t in [en_total, cn_total]:
        if t:
            for key in grand_total:
                grand_total[key] += t[key]

    if grand_total["total_rows"] > 0 or grand_total["done"] > 0:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        gt_title = ws.cell(current_row, 1, "GRAND TOTAL (ALL LANGUAGES)")
        gt_title.fill = styles["title_fill"]
        gt_title.font = Font(bold=True, size=12)
        gt_title.alignment = styles["center"]
        current_row += 1

        gt = grand_total
        total_row_data = [
            "TOTAL", gt["done"], gt["issues"], gt["no_issue"], gt["blocked"], gt["korean"],
            gt["fixed"], gt["reported"], gt["checking"], gt["pending"]
        ]

        for col, value in enumerate(total_row_data, 1):
            cell = ws.cell(current_row, col, value)
            cell.fill = styles["total_fill"]
            cell.font = styles["bold"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

    # Set column widths
    PADDING = 2
    all_headers = TESTER_HEADERS + MANAGER_HEADERS
    for col, header in enumerate(all_headers, 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = len(header) + PADDING

    breakdown_start_row = current_row + 2

    # Add Category Breakdown Tables (EN and CN)
    if en_users:
        breakdown_start_row = build_category_breakdown_section(
            ws, breakdown_start_row, latest_data, en_users,
            is_english=True, tester_mapping=tester_mapping, styles=styles
        )
        breakdown_start_row += 2  # Gap between tables

    if cn_users:
        breakdown_start_row = build_category_breakdown_section(
            ws, breakdown_start_row, latest_data, cn_users,
            is_english=False, tester_mapping=tester_mapping, styles=styles
        )
        breakdown_start_row += 2  # Gap before ranking

    # Add Ranking Table
    build_ranking_table(ws, breakdown_start_row, user_data, tester_mapping, styles)

    # Autofit columns and rows
    autofit_columns(ws, min_width=10, max_width=40)
    autofit_row_heights(ws)
