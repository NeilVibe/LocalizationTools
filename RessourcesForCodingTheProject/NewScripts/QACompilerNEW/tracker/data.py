"""
Tracker Data Module
===================
Core tracker operations: create/load tracker, update _DAILY_DATA sheet.
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import TRACKER_PATH, get_target_master_category
from core.excel_ops import safe_load_workbook


# =============================================================================
# TRACKER INITIALIZATION
# =============================================================================

def get_or_create_tracker(tracker_path: Path = None) -> Tuple[openpyxl.Workbook, Path]:
    """
    Load existing tracker or create new one with sheets.

    Args:
        tracker_path: Optional custom tracker path (defaults to TRACKER_PATH)

    Returns:
        Tuple of (workbook, path)
    """
    if tracker_path is None:
        tracker_path = TRACKER_PATH

    if tracker_path.exists():
        wb = safe_load_workbook(tracker_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        # Create sheets in order
        wb.create_sheet("DAILY", 0)
        wb.create_sheet("TOTAL", 1)
        wb.create_sheet("_DAILY_DATA", 2)
        # Remove default sheet
        wb.remove(default_sheet)
        # Hide data sheet
        wb["_DAILY_DATA"].sheet_state = 'hidden'

    return wb, tracker_path


# =============================================================================
# _DAILY_DATA SHEET OPERATIONS
# =============================================================================

# Schema for _DAILY_DATA sheet
DAILY_DATA_HEADERS = [
    "Date",        # 1
    "User",        # 2
    "Category",    # 3
    "TotalRows",   # 4
    "Done",        # 5
    "Issues",      # 6
    "NoIssue",     # 7
    "Blocked",     # 8
    "Fixed",       # 9
    "Reported",    # 10
    "Checking",    # 11
    "NonIssue",    # 12
    "WordCount",   # 13
    "Korean",      # 14
]


def update_daily_data_sheet(
    wb: openpyxl.Workbook,
    daily_entries: List[Dict],
    manager_stats: Dict = None,
    manager_dates: Dict = None
) -> None:
    """
    Update hidden _DAILY_DATA sheet with new entries including manager stats.

    Args:
        wb: Tracker workbook
        daily_entries: List of dicts with:
            {date, user, category, total_rows, done, issues, no_issue, blocked,
             word_count, korean}
        manager_stats: Optional dict of:
            {category: {user: {fixed, reported, checking, nonissue}}}
        manager_dates: Optional dict of:
            {(category, user): file_date} - dates from master file mtime

    Mode: REPLACE - same (date, user, category) overwrites existing row
    """
    if manager_stats is None:
        manager_stats = {}
    if manager_dates is None:
        manager_dates = {}

    ws = wb["_DAILY_DATA"]

    # Ensure headers exist (14 columns total)
    if ws.cell(1, 1).value != "Date" or ws.max_column < 14:
        for col, header in enumerate(DAILY_DATA_HEADERS, 1):
            ws.cell(1, col, header)

    # Build index of existing rows: (date, user, category) -> row_number
    existing = {}
    for row in range(2, ws.max_row + 1):
        key = (
            ws.cell(row, 1).value,  # Date
            ws.cell(row, 2).value,  # User
            ws.cell(row, 3).value   # Category
        )
        if key[0] is not None:  # Only index non-empty rows
            existing[key] = row

    # Update or insert entries from daily_entries (tester stats)
    # Concise log - append to same file
    from pathlib import Path
    log_path = Path(__file__).parent.parent / "MANAGER_STATS_DEBUG.log"
    log_lines = ["\n=== LOOKUP PHASE ==="]
    log_lines.append(f"daily_entries: {len(daily_entries)}, manager_stats categories: {list(manager_stats.keys())}")

    lookup_hits = 0
    lookup_misses = []

    for entry in daily_entries:
        key = (entry["date"], entry["user"], entry["category"])
        row = existing.get(key) or ws.max_row + 1
        existing[key] = row

        category = entry["category"]
        user = entry["user"]
        # Map folder category to target master category for lookup (e.g., Sequencer→Script, Help→System)
        lookup_category = get_target_master_category(category)
        category_stats = manager_stats.get(lookup_category, {})
        user_manager_stats = category_stats.get(user, {"fixed": 0, "reported": 0, "checking": 0, "nonissue": 0})

        has_stats = any(user_manager_stats[k] > 0 for k in ["fixed", "reported", "checking", "nonissue"])
        if has_stats:
            lookup_hits += 1
        else:
            reason = "NO_CAT" if lookup_category not in manager_stats else ("NO_USER" if user not in category_stats else "ZERO")
            lookup_misses.append(f"{user}/{category}:{reason}")

        # Write row data according to schema
        ws.cell(row, 1, entry["date"])
        ws.cell(row, 2, entry["user"])
        ws.cell(row, 3, entry["category"])
        ws.cell(row, 4, entry.get("total_rows", 0))      # TotalRows
        ws.cell(row, 5, entry["done"])                   # Done
        ws.cell(row, 6, entry["issues"])                 # Issues
        ws.cell(row, 7, entry["no_issue"])               # NoIssue
        ws.cell(row, 8, entry["blocked"])                # Blocked
        ws.cell(row, 9, user_manager_stats["fixed"])     # Fixed
        ws.cell(row, 10, user_manager_stats["reported"]) # Reported
        ws.cell(row, 11, user_manager_stats["checking"]) # Checking
        ws.cell(row, 12, user_manager_stats["nonissue"]) # NonIssue
        ws.cell(row, 13, entry.get("word_count", 0))     # WordCount
        ws.cell(row, 14, entry.get("korean", 0))         # Korean

    # Write lookup summary to log
    log_lines.append(f"HITS: {lookup_hits}, MISSES: {len(lookup_misses)}")
    if lookup_misses:
        log_lines.append(f"MISS DETAILS: {lookup_misses[:10]}{'...' if len(lookup_misses) > 10 else ''}")
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines) + "\n")
    except:
        pass

    # Manager stats row handling
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    normal_compilation_mode = len(daily_entries) > 0

    # Build existing row index
    existing_date_user_cat = {}
    actual_max_row = ws.max_row
    for row in range(2, ws.max_row + 1):
        row_date = ws.cell(row, 1).value
        row_user = ws.cell(row, 2).value
        row_category = ws.cell(row, 3).value
        if row_date and row_user and row_category:
            existing_date_user_cat[(str(row_date), row_user, row_category)] = row
            actual_max_row = max(actual_max_row, row)

    # Process manager stats rows
    rows_created = 0
    rows_updated = 0
    for category, users in manager_stats.items():
        for user, stats in users.items():
            file_date = manager_dates.get((category, user), today)
            key = (str(file_date), user, category)

            if key in existing_date_user_cat:
                found_row = existing_date_user_cat[key]
                ws.cell(found_row, 9, stats["fixed"])
                ws.cell(found_row, 10, stats["reported"])
                ws.cell(found_row, 11, stats["checking"])
                ws.cell(found_row, 12, stats["nonissue"])
                rows_updated += 1
            elif not normal_compilation_mode:
                actual_max_row += 1
                new_row = actual_max_row
                ws.cell(new_row, 1, file_date)
                ws.cell(new_row, 2, user)
                ws.cell(new_row, 3, category)
                for c in [4,5,6,7,8,13,14]: ws.cell(new_row, c, 0)
                ws.cell(new_row, 9, stats["fixed"])
                ws.cell(new_row, 10, stats["reported"])
                ws.cell(new_row, 11, stats["checking"])
                ws.cell(new_row, 12, stats["nonissue"])
                existing_date_user_cat[key] = new_row
                rows_created += 1


def read_daily_data(wb: openpyxl.Workbook) -> Dict:
    """
    Read all data from _DAILY_DATA sheet.

    Returns:
        Dict with structure: raw_data[date][user][category] = {
            total_rows, done, issues, no_issue, blocked, korean,
            fixed, reported, checking, nonissue, word_count
        }
    """
    from collections import defaultdict

    ws = wb["_DAILY_DATA"]

    # Structure: raw_data[date][user][category] = {...stats...}
    raw_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    })))
    users = set()
    categories = set()

    for row in range(2, ws.max_row + 1):
        date = ws.cell(row, 1).value
        user = ws.cell(row, 2).value
        category = ws.cell(row, 3).value or "Unknown"
        total_rows = ws.cell(row, 4).value or 0
        done = ws.cell(row, 5).value or 0
        issues = ws.cell(row, 6).value or 0
        no_issue = ws.cell(row, 7).value or 0
        blocked = ws.cell(row, 8).value or 0
        fixed = ws.cell(row, 9).value or 0
        reported = ws.cell(row, 10).value or 0
        checking = ws.cell(row, 11).value or 0
        nonissue = ws.cell(row, 12).value or 0
        word_count = ws.cell(row, 13).value or 0
        korean = ws.cell(row, 14).value or 0

        if date and user:
            # Store per-category data
            raw_data[date][user][category]["total_rows"] = total_rows
            raw_data[date][user][category]["done"] = done
            raw_data[date][user][category]["issues"] = issues
            raw_data[date][user][category]["no_issue"] = no_issue
            raw_data[date][user][category]["blocked"] = blocked
            raw_data[date][user][category]["korean"] = korean
            raw_data[date][user][category]["fixed"] = fixed
            raw_data[date][user][category]["reported"] = reported
            raw_data[date][user][category]["checking"] = checking
            raw_data[date][user][category]["nonissue"] = nonissue
            raw_data[date][user][category]["word_count"] = word_count
            users.add(user)
            categories.add(category)

    return {
        "raw_data": dict(raw_data),
        "users": users,
        "categories": categories,
        "dates": sorted(raw_data.keys())
    }


def compute_daily_deltas(raw_data: Dict, users: set, categories: set, dates: list) -> Dict:
    """
    Compute daily deltas per (user, category), then aggregate.

    This prevents cross-category contamination in delta calculation.
    The fix for the Quest delta bug.

    Args:
        raw_data: Dict from read_daily_data
        users: Set of all usernames
        categories: Set of all categories
        dates: Sorted list of all dates

    Returns:
        Dict: daily_delta[date][user] = {done, issues, no_issue, blocked, korean,
                                         fixed, reported, checking, nonissue, word_count}
    """
    from collections import defaultdict

    default_data = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    }

    daily_delta = defaultdict(lambda: defaultdict(lambda: default_data.copy()))

    # For each (user, category), find dates where that combo has data and compute deltas
    for user in users:
        for category in categories:
            # Get all dates where this user+category has data
            user_cat_dates = sorted([
                d for d in dates
                if category in raw_data.get(d, {}).get(user, {})
            ])

            for i, date in enumerate(user_cat_dates):
                current = raw_data[date][user][category]

                if i == 0:
                    prev = default_data.copy()
                else:
                    prev_date = user_cat_dates[i - 1]
                    prev = raw_data[prev_date][user][category]

                # Calculate delta for this category (ensure non-negative)
                cat_delta_total_rows = current["total_rows"]  # total_rows is not cumulative
                cat_delta_done = max(0, current["done"] - prev["done"])
                cat_delta_issues = max(0, current["issues"] - prev["issues"])
                cat_delta_no_issue = max(0, current["no_issue"] - prev["no_issue"])
                cat_delta_blocked = max(0, current["blocked"] - prev["blocked"])
                cat_delta_korean = max(0, current["korean"] - prev["korean"])
                cat_delta_fixed = max(0, current["fixed"] - prev["fixed"])
                cat_delta_reported = max(0, current["reported"] - prev["reported"])
                cat_delta_checking = max(0, current["checking"] - prev["checking"])
                cat_delta_nonissue = max(0, current["nonissue"] - prev["nonissue"])
                cat_delta_word_count = max(0, current["word_count"] - prev["word_count"])

                # Aggregate this category's delta into the (date, user) totals
                daily_delta[date][user]["total_rows"] += cat_delta_total_rows
                daily_delta[date][user]["done"] += cat_delta_done
                daily_delta[date][user]["issues"] += cat_delta_issues
                daily_delta[date][user]["no_issue"] += cat_delta_no_issue
                daily_delta[date][user]["blocked"] += cat_delta_blocked
                daily_delta[date][user]["korean"] += cat_delta_korean
                daily_delta[date][user]["fixed"] += cat_delta_fixed
                daily_delta[date][user]["reported"] += cat_delta_reported
                daily_delta[date][user]["checking"] += cat_delta_checking
                daily_delta[date][user]["nonissue"] += cat_delta_nonissue
                daily_delta[date][user]["word_count"] += cat_delta_word_count

    return dict(daily_delta)
