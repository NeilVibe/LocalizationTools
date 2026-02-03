"""
Face Category Processor
========================
Custom processing pipeline for Face (Facial animation) QA category.

Unlike standard categories, Face does NOT build a traditional Master file.
Instead it produces:
- MasterMismatch_{lang}.xlsx  — EventNames with MISMATCH status (deduped)
- MasterMissing_{lang}.xlsx   — EventNames with MISSING status (deduped)
- MasterConflict_{lang}.xlsx  — EventNames that appear in BOTH (also put in MISSING)

Each file preserves history: each compilation adds a NEW tab named by date (MMDD)
from the latest QA file modification time. Old tabs are kept for reference.

Returns daily_entries with Face-specific schema for the Facial tracker tab.
"""

from pathlib import Path
from typing import Dict, List
from collections import defaultdict
from datetime import datetime

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import FACE_COLS
from core.excel_ops import safe_load_workbook, build_column_map

import logging
logger = logging.getLogger(__name__)

# Module-level style constants (created once, reused across all calls)
_FACE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FACE_HEADER_FONT = Font(bold=True, color="FFFFFF")
_FACE_HEADER_ALIGN = Alignment(horizontal='center')
_CONFLICT_HEADER_FILL = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")


def process_face_category(
    qa_folders: List[Dict],
    master_folder: Path,
    lang_label: str,
    tester_mapping: Dict
) -> List[Dict]:
    """
    Custom Face category processing.

    Instead of building a standard Master_Face.xlsx:
    1. Reads all QA files, collects EventName + STATUS + Group
    2. Separates MISMATCH vs MISSING (ignores NO ISSUE)
    3. Deduplicates by EventName
    4. Detects conflicts (same EventName in both) -> flags in conflict file, puts in MISSING
    5. Outputs 2-3 Excel files in master_folder
    6. Returns daily_entries for Facial tracker

    Args:
        qa_folders: List of folder dicts from discovery
        master_folder: Target Master folder (EN or CN)
        lang_label: "EN" or "CN"
        tester_mapping: Dict mapping tester names to language codes

    Returns:
        List of daily_entry dicts for Facial tracker
    """
    logger.info("=" * 50)
    logger.info(f"Processing: Face [{lang_label}] ({len(qa_folders)} folders)")
    logger.info("=" * 50)
    logger.info("  [FACE] Custom processing pipeline (no standard master)")

    # Collectors
    mismatch_events = {}   # {eventname: set_of_groups}
    missing_events = {}    # {eventname: set_of_groups}

    # Track latest QA file mtime for date-tab naming
    latest_mtime = 0.0

    # Per-tester stats for tracker
    tester_stats = {}  # {username: {total, no_issue, mismatch, missing, groups: {group: {total, no_issue, mismatch, missing}}}}

    for qf in qa_folders:
        username = qf["username"]
        xlsx_path = qf["xlsx_path"]

        logger.info(f"  Processing: {username}")

        # Track latest file mtime across all QA files
        file_mtime = xlsx_path.stat().st_mtime
        if file_mtime > latest_mtime:
            latest_mtime = file_mtime

        # Initialize tester stats
        if username not in tester_stats:
            tester_stats[username] = {
                "total": 0,
                "no_issue": 0,
                "mismatch": 0,
                "missing": 0,
                "groups": defaultdict(lambda: {"total": 0, "no_issue": 0, "mismatch": 0, "missing": 0}),
                "file_date": datetime.fromtimestamp(xlsx_path.stat().st_mtime).strftime("%Y-%m-%d"),
            }

        try:
            wb = safe_load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    if sheet_name == "STATUS":
                        continue

                    ws = wb[sheet_name]

                    if ws.max_row is None or ws.max_row < 2:
                        continue
                    if ws.max_column is None or ws.max_column < 1:
                        continue

                    # Find columns by header name (single scan, O(1) lookups)
                    col_map = build_column_map(ws)
                    eventname_col = col_map.get(FACE_COLS["eventname"].upper())
                    group_col = col_map.get(FACE_COLS["group"].upper())
                    status_col = col_map.get(FACE_COLS["status"].upper())

                    if not eventname_col or not status_col:
                        logger.warning(f"    Sheet '{sheet_name}': Missing EventName or STATUS column, skipping")
                        continue

                    # Pre-compute 0-based indices for tuple access
                    eventname_idx = eventname_col - 1
                    status_idx = status_col - 1
                    group_idx = (group_col - 1) if group_col else None

                    sheet_rows = 0
                    for row_tuple in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
                        eventname_val = row_tuple[eventname_idx]
                        if not eventname_val:
                            continue

                        eventname = str(eventname_val).strip()
                        if not eventname:
                            continue

                        status_val = row_tuple[status_idx]
                        status = str(status_val).strip().upper() if status_val else ""

                        group_val = row_tuple[group_idx] if group_idx is not None else None
                        group = str(group_val).strip() if group_val else "Unknown"

                        # Count total rows (any row with an EventName)
                        tester_stats[username]["total"] += 1
                        tester_stats[username]["groups"][group]["total"] += 1
                        sheet_rows += 1

                        if status == "NO ISSUE":
                            tester_stats[username]["no_issue"] += 1
                            tester_stats[username]["groups"][group]["no_issue"] += 1
                        elif status == "MISMATCH":
                            tester_stats[username]["mismatch"] += 1
                            tester_stats[username]["groups"][group]["mismatch"] += 1
                            if eventname not in mismatch_events:
                                mismatch_events[eventname] = set()
                            mismatch_events[eventname].add(group)
                        elif status == "MISSING":
                            tester_stats[username]["missing"] += 1
                            tester_stats[username]["groups"][group]["missing"] += 1
                            if eventname not in missing_events:
                                missing_events[eventname] = set()
                            missing_events[eventname].add(group)
                        # Rows without status or with unknown status: counted in total but not in done

                    logger.info(f"    {sheet_name}: {sheet_rows} rows")
            finally:
                wb.close()

        except Exception as e:
            logger.error(f"Error processing {xlsx_path}: {e}", exc_info=True)

    # Detect conflicts: EventNames in BOTH mismatch and missing
    conflict_events = set(mismatch_events.keys()) & set(missing_events.keys())

    # Move conflicts to MISSING (merge groups from mismatch into missing)
    for evt in conflict_events:
        missing_events[evt] |= mismatch_events.pop(evt)

    # Summary
    logger.info("  [FACE SUMMARY]")
    logger.info(f"  MISMATCH events: {len(mismatch_events)}")
    logger.info(f"  MISSING events:  {len(missing_events)}")
    logger.info(f"  CONFLICT events: {len(conflict_events)} (moved to MISSING)")

    # Compute date tab name from latest QA file mtime (MMDD format)
    date_tab = datetime.fromtimestamp(latest_mtime).strftime("%m%d") if latest_mtime > 0 else datetime.now().strftime("%m%d")
    logger.info(f"  Date tab: {date_tab}")

    # Write output files (load existing to preserve old date tabs)
    _write_face_output(master_folder / f"MasterMismatch_{lang_label}.xlsx", mismatch_events, "MISMATCH", date_tab)
    _write_face_output(master_folder / f"MasterMissing_{lang_label}.xlsx", missing_events, "MISSING", date_tab)
    if conflict_events:
        _write_face_conflict(master_folder / f"MasterConflict_{lang_label}.xlsx", conflict_events, date_tab)

    # Build daily entries for Facial tracker
    daily_entries = _build_face_daily_entries(tester_stats, lang_label)

    return daily_entries


def _write_face_output(output_path: Path, events: Dict[str, set], label: str, date_tab: str) -> None:
    """
    Write Face output Excel file with EventName list.

    Loads existing file to preserve old date tabs, then adds a new tab
    named by date (MMDD). If a tab with the same date exists, it is replaced.

    Args:
        output_path: Path to write the xlsx file
        events: Dict of {eventname: set_of_groups}
        label: "MISMATCH" or "MISSING" for logging
        date_tab: Tab name in MMDD format (e.g., "0204")
    """
    # Load existing file or create new
    if output_path.exists():
        wb = load_workbook(output_path)
        # Remove same-date tab if re-running on same day
        if date_tab in wb.sheetnames:
            del wb[date_tab]
        ws = wb.create_sheet(date_tab)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = date_tab

    try:
        # Header (using module-level style constants)
        ws.cell(row=1, column=1, value="EventName")
        ws.cell(row=1, column=1).fill = _FACE_HEADER_FILL
        ws.cell(row=1, column=1).font = _FACE_HEADER_FONT
        ws.cell(row=1, column=1).alignment = _FACE_HEADER_ALIGN

        # Data rows (sorted for consistency)
        for idx, eventname in enumerate(sorted(events.keys()), 2):
            ws.cell(row=idx, column=1, value=eventname)

        # Autofit column width
        max_len = max((len(str(e)) for e in events), default=10)
        ws.column_dimensions['A'].width = max(max_len + 2, 15)

        wb.save(output_path)
        tab_count = len(wb.sheetnames)
        logger.info(f"  Saved: {output_path.name} tab '{date_tab}' ({len(events)} events, {tab_count} total tabs)")
    finally:
        wb.close()


def _write_face_conflict(output_path: Path, conflict_events: set, date_tab: str) -> None:
    """
    Write Face conflict Excel file.

    Loads existing file to preserve old date tabs, then adds a new tab
    named by date (MMDD). If a tab with the same date exists, it is replaced.

    Args:
        output_path: Path to write the xlsx file
        conflict_events: Set of EventNames that appeared in both MISMATCH and MISSING
        date_tab: Tab name in MMDD format (e.g., "0204")
    """
    # Load existing file or create new
    if output_path.exists():
        wb = load_workbook(output_path)
        if date_tab in wb.sheetnames:
            del wb[date_tab]
        ws = wb.create_sheet(date_tab)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = date_tab

    try:
        # Header (using module-level style constants)
        ws.cell(row=1, column=1, value="EventName")
        ws.cell(row=1, column=1).fill = _CONFLICT_HEADER_FILL
        ws.cell(row=1, column=1).font = _FACE_HEADER_FONT
        ws.cell(row=1, column=1).alignment = _FACE_HEADER_ALIGN

        ws.cell(row=1, column=2, value="Note")
        ws.cell(row=1, column=2).fill = _CONFLICT_HEADER_FILL
        ws.cell(row=1, column=2).font = _FACE_HEADER_FONT
        ws.cell(row=1, column=2).alignment = _FACE_HEADER_ALIGN

        # Data rows
        for idx, eventname in enumerate(sorted(conflict_events), 2):
            ws.cell(row=idx, column=1, value=eventname)
            ws.cell(row=idx, column=2, value="In both MISMATCH and MISSING (placed in MISSING)")

        # Autofit
        max_len = max((len(str(e)) for e in conflict_events), default=10)
        ws.column_dimensions['A'].width = max(max_len + 2, 15)
        ws.column_dimensions['B'].width = 50

        wb.save(output_path)
        tab_count = len(wb.sheetnames)
        logger.info(f"  Saved: {output_path.name} tab '{date_tab}' ({len(conflict_events)} conflicts, {tab_count} total tabs)")
    finally:
        wb.close()


def _build_face_daily_entries(tester_stats: Dict, lang_label: str) -> List[Dict]:
    """
    Build daily entries for the Facial tracker from tester stats.

    Args:
        tester_stats: Per-tester stats from processing
        lang_label: "EN" or "CN"

    Returns:
        List of daily_entry dicts with Face-specific schema
    """
    entries = []

    for username, stats in tester_stats.items():
        done = stats["no_issue"] + stats["mismatch"] + stats["missing"]

        # Convert groups defaultdict to regular dict
        groups = {}
        for group, gcounts in stats["groups"].items():
            groups[group] = {
                "total": gcounts["total"],
                "done": gcounts["no_issue"] + gcounts["mismatch"] + gcounts["missing"],
                "no_issue": gcounts["no_issue"],
                "mismatch": gcounts["mismatch"],
                "missing": gcounts["missing"],
            }

        entry = {
            "date": stats["file_date"],
            "user": username,
            "category": "Face",
            "lang": lang_label,
            "total_rows": stats["total"],
            "done": done,
            "no_issue": stats["no_issue"],
            "mismatch": stats["mismatch"],
            "missing": stats["missing"],
            "groups": groups,
        }
        logger.info(f"    Face daily_entry: {entry['date']} | {entry['user']} | done={entry['done']}, mismatch={entry['mismatch']}, missing={entry['missing']}")
        entries.append(entry)

    return entries
