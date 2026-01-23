"""
Sheet Processing Module
=======================
Core logic for processing QA sheets and compiling to master.
"""

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import (
    CATEGORIES, TRANSLATION_COLS, TRACKER_STYLES, SCRIPT_TYPE_CATEGORIES,
    get_target_master_category, load_tester_mapping
)
from core.excel_ops import (
    safe_load_workbook, find_column_by_header,
    get_or_create_master, copy_images_with_unique_names,
    ensure_master_folders, THIN_BORDER, style_header_cell,
    add_manager_dropdown
)
from core.discovery import discover_qa_folders, group_folders_by_language
from core.matching import (
    build_master_index, find_matching_row_in_master, extract_qa_row_data
)


# =============================================================================
# SCRIPT DEBUG LOGGING (shared log file with compiler.py)
# =============================================================================

_SCRIPT_DEBUG_FILE = Path(__file__).parent.parent / "SCRIPT_DEBUG.log"
_SCRIPT_DEBUG_LINES = []


def _script_debug_log(msg: str):
    """Add message to Script debug log."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    _SCRIPT_DEBUG_LINES.append(f"[{timestamp}] {msg}")


def _script_debug_flush():
    """Flush Script debug log to file."""
    global _SCRIPT_DEBUG_LINES
    if not _SCRIPT_DEBUG_LINES:
        return
    try:
        mode = "a" if _SCRIPT_DEBUG_FILE.exists() else "w"
        with open(_SCRIPT_DEBUG_FILE, mode, encoding="utf-8") as f:
            f.write("\n".join(_SCRIPT_DEBUG_LINES) + "\n")
        _SCRIPT_DEBUG_LINES = []
    except Exception as e:
        print(f"[SCRIPT DEBUG ERROR] {e}")


# =============================================================================
# COMMENT FORMATTING
# =============================================================================

def sanitize_for_excel(text: str) -> str:
    """
    Sanitize text for Excel to prevent formula injection.

    Prefixes dangerous characters with a single quote.
    """
    if not text:
        return text
    text = str(text)
    # Characters that could trigger formula execution
    if text.startswith(('=', '+', '-', '@', '\t', '\r')):
        return "'" + text
    return text


def extract_comment_text(comment: str) -> str:
    """
    Extract the original comment text (before metadata).

    Format: "comment text\n---\n..." -> "comment text"
    """
    if not comment:
        return ""
    parts = str(comment).split("---", 1)
    return parts[0].strip()


def format_comment(new_comment, string_id=None, existing_comment=None, file_mod_time=None):
    """
    Format comment with StringID and file modification time.

    Format (with stringid):
        <comment text>
        ---
        stringid:
        <stringid value>
        (updated: YYMMDD HHMM)

    Format (without stringid):
        <comment text>
        ---
        (updated: YYMMDD HHMM)

    REPLACE MODE: If comment text differs, REPLACE entirely (no append).
    DUPLICATE CHECK: Split on '---' delimiter to extract original comment for comparison.
    """
    if not new_comment or str(new_comment).strip() == "":
        return existing_comment  # Return existing if no new comment

    new_text = str(new_comment).strip()

    # Check if this exact comment text already exists (avoid re-update on re-run)
    if existing_comment:
        existing_str = str(existing_comment)
        # Split on --- delimiter to extract original comment
        if "\n---\n" in existing_str:
            existing_original = existing_str.split("\n---\n")[0].strip()
            if existing_original == new_text:
                return existing_comment  # Duplicate, skip

    # Format timestamp from file modification time (or fallback to now)
    if file_mod_time:
        timestamp = file_mod_time.strftime("%y%m%d %H%M")
    else:
        timestamp = datetime.now().strftime("%y%m%d %H%M")

    # Build formatted comment: text + delimiter + metadata
    if string_id and str(string_id).strip():
        # With stringid: comment -> --- -> stringid: -> value -> (updated)
        return f"{new_text}\n---\nstringid:\n{str(string_id).strip()}\n(updated: {timestamp})"
    else:
        # Without stringid: comment -> --- -> (updated)
        return f"{new_text}\n---\n(updated: {timestamp})"


# =============================================================================
# USER COLUMN MANAGEMENT
# =============================================================================

def get_or_create_user_comment_column(ws, username: str) -> int:
    """
    Find or create COMMENT_{username} column.

    Returns: Column index (1-based)
    """
    target_header = f"COMMENT_{username}"

    # Search for existing column
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().upper() == target_header.upper():
            return col

    # Create new column at the end
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=target_header)
    # Style: Light blue fill with MEDIUM blue borders (match monolith)
    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )
    print(f"    Created column: {target_header} at {get_column_letter(new_col)} (styled)")
    return new_col


def get_or_create_tester_status_column(ws, username: str, after_col: int) -> int:
    """
    Find or create TESTER_STATUS_{username} column (hidden).

    Returns: Column index (1-based)
    """
    target_header = f"TESTER_STATUS_{username}"

    # Search for existing column
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().upper() == target_header.upper():
            return col

    # Create new column after COMMENT
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=target_header)
    style_header_cell(cell)

    # Hide this column (internal use only)
    ws.column_dimensions[get_column_letter(new_col)].hidden = True
    print(f"    Created column: {target_header} at {get_column_letter(new_col)} (tester status - HIDDEN)")
    return new_col


def get_or_create_user_status_column(ws, username: str, after_col: int) -> int:
    """
    Find or create STATUS_{username} column with manager dropdown.

    Returns: Column index (1-based)
    """
    target_header = f"STATUS_{username}"

    # Search for existing column
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().upper() == target_header.upper():
            return col

    # Create new column
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=target_header)
    # Style: Light green fill with MEDIUM green borders (match monolith)
    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='medium', color='228B22'),
        right=Side(style='medium', color='228B22'),
        top=Side(style='medium', color='228B22'),
        bottom=Side(style='medium', color='228B22')
    )

    # Add dropdown with 50-row buffer (match monolith)
    last_row = max(ws.max_row, 10) + 50
    add_manager_dropdown(ws, new_col, end_row=last_row)
    print(f"    Created column: {target_header} at {get_column_letter(new_col)} (manager status - dropdown)")
    return new_col


def get_or_create_user_manager_comment_column(ws, username: str, after_col: int) -> int:
    """
    Find or create MANAGER_COMMENT_{username} column.

    This column is paired with STATUS_{username} for manager notes.
    Returns: Column index (1-based)
    """
    target_header = f"MANAGER_COMMENT_{username}"

    # Search for existing column
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().upper() == target_header.upper():
            return col

    # Create new column
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=target_header)
    # Style: Light green fill with MEDIUM green borders (match manager status)
    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='medium', color='228B22'),
        right=Side(style='medium', color='228B22'),
        top=Side(style='medium', color='228B22'),
        bottom=Side(style='medium', color='228B22')
    )
    print(f"    Created column: {target_header} at {get_column_letter(new_col)} (manager comment)")
    return new_col


def get_or_create_user_screenshot_column(ws, username: str, after_col: int) -> int:
    """
    Find or create SCREENSHOT_{username} column.

    Returns: Column index (1-based)
    """
    target_header = f"SCREENSHOT_{username}"

    # Search for existing column
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().upper() == target_header.upper():
            return col

    # Create new column
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=target_header)
    # Style: Light blue fill with MEDIUM blue borders (match monolith)
    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )
    print(f"    Created column: {target_header} at {get_column_letter(new_col)} (styled)")
    return new_col


# =============================================================================
# USER COLUMN HELPERS
# =============================================================================

def find_or_create_user_columns(ws, username: str) -> Tuple[int, int, int, int, int]:
    """
    Find or create all user columns in order:
    COMMENT_{user} -> TESTER_STATUS_{user} (hidden) -> STATUS_{user} -> MANAGER_COMMENT_{user} -> SCREENSHOT_{user}

    Args:
        ws: Worksheet
        username: Tester username

    Returns:
        Tuple of (comment_col, tester_status_col, status_col, manager_comment_col, screenshot_col)
    """
    comment_col = get_or_create_user_comment_column(ws, username)
    tester_status_col = get_or_create_tester_status_column(ws, username, comment_col)
    status_col = get_or_create_user_status_column(ws, username, comment_col)
    manager_comment_col = get_or_create_user_manager_comment_column(ws, username, status_col)
    screenshot_col = get_or_create_user_screenshot_column(ws, username, manager_comment_col)
    return comment_col, tester_status_col, status_col, manager_comment_col, screenshot_col


def add_user_columns(ws, username: str) -> Tuple[int, int, int, int, int]:
    """Alias for find_or_create_user_columns for backwards compatibility."""
    return find_or_create_user_columns(ws, username)

# =============================================================================
# SHEET PROCESSING
# =============================================================================

def process_sheet(
    master_ws,
    qa_ws,
    username: str,
    category: str,
    is_english: bool = True,
    image_mapping: Dict = None,
    xlsx_path: Path = None,
    manager_status: Dict = None
) -> Dict:
    """
    Process a single sheet: copy COMMENT and SCREENSHOT from QA to master.

    Uses CONTENT-BASED MATCHING to find correct master row for each QA row.
    This enables processing QA files with mixed old/new structure.

    Matching Strategy:
    - Standard (Quest, Knowledge, etc.): STRINGID + Translation, fallback to Translation only
    - Item: ItemName + ItemDesc + STRINGID, fallback to ItemName + ItemDesc
    - Contents: INSTRUCTIONS column

    Args:
        master_ws: Master worksheet
        qa_ws: QA worksheet
        username: User identifier
        category: Category name
        is_english: Whether file is English (affects column selection)
        image_mapping: Dict mapping original_name -> new_name
        xlsx_path: Path to QA xlsx file (for modification time)
        manager_status: Dict for preserving manager status

    Returns:
        Dict with {comments, screenshots, stats, manager_restored, match_stats}
    """
    if image_mapping is None:
        image_mapping = {}
    if manager_status is None:
        manager_status = {}

    # Get file modification time
    file_mod_time = None
    if xlsx_path:
        file_mod_time = datetime.fromtimestamp(xlsx_path.stat().st_mtime)

    # Detect if Script category (uses MEMO instead of COMMENT, no SCREENSHOT)
    is_script = category.lower() in SCRIPT_TYPE_CATEGORIES

    # DEBUG: Log manager_status for Script categories
    # NOTE: manager_status here is already sheet-level: {(stringid, comment): {username: info}}
    if is_script:
        file_name = xlsx_path.name if xlsx_path else "unknown"
        _script_debug_log(f"[PROCESS] {category}/{file_name}/{username}")
        _script_debug_log(f"  manager_status keys (sheet-level): {len(manager_status)}")
        if manager_status:
            sample_keys = list(manager_status.keys())[:3]
            _script_debug_log(f"  Sample keys: {sample_keys}")

    # Find columns in QA worksheet
    qa_status_col = find_column_by_header(qa_ws, "STATUS")
    # Script uses MEMO column instead of COMMENT
    if is_script:
        qa_comment_col = find_column_by_header(qa_ws, "MEMO")
        qa_screenshot_col = None  # Script has no SCREENSHOT
        _script_debug_log(f"  QA MEMO column: {qa_comment_col}")
    else:
        qa_comment_col = find_column_by_header(qa_ws, "COMMENT")
        qa_screenshot_col = find_column_by_header(qa_ws, "SCREENSHOT")
    qa_stringid_col = find_column_by_header(qa_ws, "STRINGID")
    # For Script, also check for EventName as STRINGID
    if is_script and not qa_stringid_col:
        qa_stringid_col = find_column_by_header(qa_ws, "EventName")

    # Find or create user columns in master
    # Note: Script category has NO SCREENSHOT column
    master_comment_col = get_or_create_user_comment_column(master_ws, username)
    master_tester_status_col = get_or_create_tester_status_column(master_ws, username, master_comment_col)
    master_user_status_col = get_or_create_user_status_column(master_ws, username, master_comment_col)
    master_manager_comment_col = get_or_create_user_manager_comment_column(master_ws, username, master_user_status_col)
    if not is_script:
        master_screenshot_col = get_or_create_user_screenshot_column(master_ws, username, master_manager_comment_col)
    else:
        master_screenshot_col = None  # Script has no screenshot column

    # Find STRINGID column in master (for manager status lookup)
    # Script-type categories use EventName instead of STRINGID
    master_stringid_col = find_column_by_header(master_ws, "STRINGID")
    if not master_stringid_col and is_script:
        master_stringid_col = find_column_by_header(master_ws, "EventName")

    result = {
        "comments": 0,
        "screenshots": 0,
        "stats": {"issue": 0, "no_issue": 0, "blocked": 0, "korean": 0, "total": 0},
        "manager_restored": 0,
        "match_stats": {"exact": 0, "fallback": 0, "unmatched": 0}
    }

    # Build master index for O(1) content-based matching
    master_index = build_master_index(master_ws, category, is_english)

    # OPTIMIZATION: For Script-type categories, pre-filter to only rows WITH status
    # This dramatically speeds up processing for large files (Sequencer/Dialog can have 10,000+ rows)
    rows_to_process = []
    if is_script and qa_status_col:
        # SIMPLE APPROACH: If STATUS has ANY value (not empty), include the row
        # Accept both "NON-ISSUE" (Script-type) and "NO ISSUE" (other categories)
        for qa_row in range(2, qa_ws.max_row + 1):
            status_val = qa_ws.cell(row=qa_row, column=qa_status_col).value
            if status_val and str(status_val).strip():  # Any non-empty value
                rows_to_process.append(qa_row)
        print(f"      [OPTIMIZATION] {len(rows_to_process)} rows with STATUS (skipping {qa_ws.max_row - 1 - len(rows_to_process)} empty rows)")
    else:
        # Non-Script categories: process all rows
        rows_to_process = list(range(2, qa_ws.max_row + 1))

    # Process each row using CONTENT-BASED MATCHING
    for qa_row in rows_to_process:
        # Skip empty rows - check column 1 (first column always has data if row is valid)
        first_col_value = qa_ws.cell(row=qa_row, column=1).value
        if first_col_value is None or str(first_col_value).strip() == "":
            continue

        # Extract QA row data for matching
        qa_row_data = extract_qa_row_data(qa_ws, qa_row, category, is_english)

        # Find matching master row using content-based matching
        master_row, match_type = find_matching_row_in_master(qa_row_data, master_index, category)

        if master_row is None:
            # No match found - log and skip
            result["match_stats"]["unmatched"] += 1
            continue

        # Only count rows that successfully matched (for accurate tracker stats)
        result["stats"]["total"] += 1

        # Track match type
        if match_type == "exact":
            result["match_stats"]["exact"] += 1
        else:
            result["match_stats"]["fallback"] += 1

        # Get QA STATUS
        should_compile_comment = False
        status_type = None
        if qa_status_col:
            qa_status = qa_ws.cell(row=qa_row, column=qa_status_col).value
            if qa_status:
                status_upper = str(qa_status).strip().upper()
                if status_upper == "ISSUE":
                    result["stats"]["issue"] += 1
                    should_compile_comment = True
                    status_type = "ISSUE"
                elif status_upper in ("NO ISSUE", "NON-ISSUE"):
                    # Accept both "NO ISSUE" (standard) and "NON-ISSUE" (Script-type)
                    result["stats"]["no_issue"] += 1
                    should_compile_comment = True
                    status_type = "NO ISSUE"  # Normalize to "NO ISSUE" for consistency
                elif status_upper == "BLOCKED":
                    result["stats"]["blocked"] += 1
                    should_compile_comment = True
                    status_type = "BLOCKED"
                elif status_upper == "KOREAN":
                    result["stats"]["korean"] += 1
                    should_compile_comment = True
                    status_type = "KOREAN"

        # Write TESTER STATUS
        if status_type:
            tester_status_cell = master_ws.cell(row=master_row, column=master_tester_status_col)
            tester_status_cell.value = status_type
            tester_status_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Process COMMENT
        if qa_comment_col and should_compile_comment:
            qa_comment = qa_ws.cell(row=qa_row, column=qa_comment_col).value
            if qa_comment and str(qa_comment).strip():
                string_id = None
                if qa_stringid_col:
                    string_id = qa_ws.cell(row=qa_row, column=qa_stringid_col).value

                existing = master_ws.cell(row=master_row, column=master_comment_col).value
                new_value = format_comment(qa_comment, string_id, existing, file_mod_time)

                if new_value != existing:
                    cell = master_ws.cell(row=master_row, column=master_comment_col)
                    cell.value = sanitize_for_excel(new_value)

                    # Style based on status
                    if status_type == "ISSUE":
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif status_type == "BLOCKED":
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                        cell.font = Font(bold=True, color="FF8C00")
                    elif status_type == "KOREAN":
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        cell.font = Font(bold=True, color="800080")
                    elif status_type in ("NO ISSUE", "NON-ISSUE"):
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        cell.font = Font(bold=True, color="2E7D32")

                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin', color='87CEEB'),
                        right=Side(style='thin', color='87CEEB'),
                        top=Side(style='thin', color='87CEEB'),
                        bottom=Side(style='thin', color='87CEEB')
                    )
                    result["comments"] += 1

        # Process SCREENSHOT (not for Script category)
        if qa_screenshot_col and master_screenshot_col:
            qa_screenshot_cell = qa_ws.cell(row=qa_row, column=qa_screenshot_col)
            screenshot_value = qa_screenshot_cell.value
            screenshot_hyperlink = qa_screenshot_cell.hyperlink

            if screenshot_value and str(screenshot_value).strip():
                master_screenshot_cell = master_ws.cell(row=master_row, column=master_screenshot_col)

                new_screenshot_value = None
                new_screenshot_target = None
                is_warning = False

                if screenshot_hyperlink and screenshot_hyperlink.target:
                    original_target = screenshot_hyperlink.target
                    original_name = os.path.basename(original_target)

                    if original_name in image_mapping:
                        new_name = image_mapping[original_name]
                        new_screenshot_value = new_name
                        new_screenshot_target = f"Images/{new_name}"
                    else:
                        # Case-insensitive match
                        matched_name = None
                        for img_name in image_mapping.keys():
                            if img_name.lower() == original_name.lower():
                                matched_name = img_name
                                break
                        if matched_name:
                            new_name = image_mapping[matched_name]
                            new_screenshot_value = new_name
                            new_screenshot_target = f"Images/{new_name}"
                        else:
                            new_screenshot_value = original_name
                            new_screenshot_target = f"Images/{original_name}"
                            is_warning = True
                else:
                    # No hyperlink - just value
                    original_name = str(screenshot_value).strip()
                    if original_name in image_mapping:
                        new_name = image_mapping[original_name]
                        new_screenshot_value = new_name
                        new_screenshot_target = f"Images/{new_name}"
                    else:
                        # Case-insensitive match (match monolith)
                        matched_name = None
                        for img_name in image_mapping.keys():
                            if img_name.lower() == original_name.lower():
                                matched_name = img_name
                                break
                        if matched_name:
                            new_name = image_mapping[matched_name]
                            new_screenshot_value = new_name
                            new_screenshot_target = f"Images/{new_name}"
                        else:
                            new_screenshot_value = original_name
                            new_screenshot_target = f"Images/{original_name}"
                            is_warning = True

                existing_hyperlink = master_screenshot_cell.hyperlink.target if master_screenshot_cell.hyperlink else None
                existing_screenshot = master_screenshot_cell.value
                needs_update = (new_screenshot_value != existing_screenshot) or (new_screenshot_target != existing_hyperlink)

                if needs_update:
                    master_screenshot_cell.value = sanitize_for_excel(new_screenshot_value)
                    if new_screenshot_target:
                        master_screenshot_cell.hyperlink = new_screenshot_target

                    master_screenshot_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    master_screenshot_cell.alignment = Alignment(horizontal='left', vertical='center')
                    master_screenshot_cell.border = Border(
                        left=Side(style='thin', color='87CEEB'),
                        right=Side(style='thin', color='87CEEB'),
                        top=Side(style='thin', color='87CEEB'),
                        bottom=Side(style='thin', color='87CEEB')
                    )

                    if new_screenshot_target:
                        if is_warning:
                            master_screenshot_cell.font = Font(color="FF6600", underline="single")
                        else:
                            master_screenshot_cell.font = Font(color="0000FF", underline="single")

                    result["screenshots"] += 1

        # Apply manager STATUS and MANAGER_COMMENT
        # Key = (stringid, tester_comment_text) - Manager status is paired with tester's comment
        # STRINGID from MASTER row (reliable), comment from QA file (what tester wrote)

        # Get tester's comment from QA file for lookup
        tester_comment_for_lookup = ""
        if qa_comment_col:
            qa_comment_raw = qa_ws.cell(row=qa_row, column=qa_comment_col).value
            if qa_comment_raw and str(qa_comment_raw).strip():
                tester_comment_for_lookup = extract_comment_text(qa_comment_raw)

        # DEBUG: Log lookup attempts for Script
        if is_script and qa_row <= 5:  # First 5 rows only
            _script_debug_log(f"  Row {qa_row}: tester_comment='{tester_comment_for_lookup[:30] if tester_comment_for_lookup else 'EMPTY'}...'")

        if manager_status and tester_comment_for_lookup:
            # Get STRINGID from MASTER row (not QA file - QA file might have empty STRINGID!)
            master_stringid = ""
            if master_stringid_col:
                stringid_val = master_ws.cell(row=master_row, column=master_stringid_col).value
                if stringid_val:
                    master_stringid = str(stringid_val).strip()

            # Primary key: (stringid, tester_comment_text)
            manager_key = (master_stringid, tester_comment_for_lookup)

            # DEBUG: Log lookup for Script
            if is_script and qa_row <= 5:
                found = manager_key in manager_status
                key_str = str(manager_key)[:80]
                _script_debug_log(f"  Row {qa_row}: key={key_str}, found={found}")
                _script_debug_flush()

            # Fallback: ("", tester_comment_text) if exact match fails (STRINGID changed)
            if manager_key not in manager_status:
                fallback_key = ("", tester_comment_for_lookup)
                if fallback_key in manager_status:
                    manager_key = fallback_key

            if manager_key in manager_status:
                user_data = manager_status[manager_key]
                if username in user_data:
                    manager_info = user_data[username]
                    # Handle both old format (string) and new format (dict)
                    if isinstance(manager_info, dict):
                        status_value = manager_info.get("status")
                        manager_comment_value = manager_info.get("manager_comment")
                    else:
                        # Backwards compatibility: old format was just the status string
                        status_value = manager_info
                        manager_comment_value = None

                    # Apply status
                    if status_value:
                        status_cell = master_ws.cell(row=master_row, column=master_user_status_col)
                        status_cell.value = status_value
                        status_cell.alignment = Alignment(horizontal='center', vertical='center')
                        if status_value == "FIXED":
                            status_cell.font = Font(bold=True, color="228B22")
                        elif status_value == "REPORTED":
                            status_cell.font = Font(bold=True, color="FF8C00")
                        elif status_value == "CHECKING":
                            status_cell.font = Font(bold=True, color="0000FF")
                        elif status_value in ("NON-ISSUE", "NON ISSUE"):
                            status_cell.font = Font(bold=True, color="808080")
                        result["manager_restored"] += 1

                    # Apply manager comment
                    if manager_comment_value:
                        manager_comment_cell = master_ws.cell(row=master_row, column=master_manager_comment_col)
                        manager_comment_cell.value = manager_comment_value
                        manager_comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
                        # Light green background to match manager status theme
                        manager_comment_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        manager_comment_cell.border = Border(
                            left=Side(style='thin', color='228B22'),
                            right=Side(style='thin', color='228B22'),
                            top=Side(style='thin', color='228B22'),
                            bottom=Side(style='thin', color='228B22')
                        )

    return result


# =============================================================================
# STATUS SHEET
# =============================================================================

def update_status_sheet(wb, users, user_stats):
    """
    Create/update STATUS sheet with completion tracking and detailed stats.

    Args:
        wb: Master workbook
        users: Set of usernames
        user_stats: {username: {total: n, issue: n, no_issue: n, blocked: n}}
    """
    # Create or clear STATUS sheet
    if "STATUS" in wb.sheetnames:
        del wb["STATUS"]

    ws = wb.create_sheet("STATUS", 0)  # Insert at position 0 (first tab)

    # Styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold/Yellow
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center')

    # Headers
    headers = ["User", "Completion %", "Total Rows", "ISSUE #", "NO ISSUE %", "BLOCKED %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    # Data rows
    for row_idx, user in enumerate(sorted(users), 2):
        stats = user_stats.get(user, {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0})

        total = stats["total"]
        issue = stats["issue"]
        no_issue = stats["no_issue"]
        blocked = stats["blocked"]
        completed = issue + no_issue + blocked

        # Calculate percentages
        completion_pct = round(completed / total * 100, 1) if total > 0 else 0
        no_issue_pct = round(no_issue / total * 100, 1) if total > 0 else 0
        blocked_pct = round(blocked / total * 100, 1) if total > 0 else 0

        # Write data
        row_data = [
            user,
            f"{completion_pct}%",
            total,
            issue,  # Raw number for ISSUE
            f"{no_issue_pct}%",
            f"{blocked_pct}%"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.alignment = center

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    print(f"  Updated STATUS sheet with {len(users)} users (first tab, yellow header)")


# =============================================================================
# WORD/CHARACTER COUNTING
# =============================================================================

def contains_korean(text) -> bool:
    """
    Check if text contains Korean characters (Hangul).

    Korean Unicode ranges:
    - Hangul Syllables: U+AC00 to U+D7AF (most common, 11,172 chars)
    - Hangul Jamo: U+1100 to U+11FF (archaic/combining)
    - Hangul Compatibility Jamo: U+3130 to U+318F

    Args:
        text: Text to check

    Returns:
        True if text contains Korean characters, False otherwise
    """
    if not text:
        return False
    for char in str(text):
        # Hangul Syllables (most common)
        if '\uAC00' <= char <= '\uD7AF':
            return True
        # Hangul Jamo
        if '\u1100' <= char <= '\u11FF':
            return True
        # Hangul Compatibility Jamo
        if '\u3130' <= char <= '\u318F':
            return True
    return False


def count_words_english(text) -> int:
    """
    Count words in English text.

    Splits by whitespace and counts tokens.
    Returns 0 if text contains Korean (untranslated).

    Args:
        text: Text to count words in

    Returns:
        int: Word count (0 if Korean detected or empty)
    """
    if not text or contains_korean(text):
        return 0
    return len(str(text).split())


def count_chars_chinese(text) -> int:
    """
    Count characters in Chinese text (excluding whitespace).

    For CJK languages, character count is more meaningful than word count.
    Returns 0 if text contains Korean (untranslated).

    Args:
        text: Text to count characters in

    Returns:
        int: Character count excluding whitespace (0 if Korean detected or empty)
    """
    if not text or contains_korean(text):
        return 0
    # Remove all whitespace characters
    cleaned = str(text).replace(" ", "").replace("\n", "").replace("\t", "").replace("\r", "")
    return len(cleaned)


# =============================================================================
# ROW/SHEET/COLUMN HIDING
# =============================================================================

def hide_empty_comment_rows(wb, context_rows: int = 1, debug: bool = False) -> tuple:
    """
    Post-process: Hide rows/sheets/columns based on tester and manager status.

    This allows focusing on ISSUE rows that need attention while preserving all data.
    Rows are hidden (not deleted) so they can be unhidden in Excel if needed.

    Features:
    - Sheet hiding: Sheets with NO comments at all are hidden entirely
    - Column hiding: COMMENT_{User} columns with no data in a sheet are hidden
      (also hides paired SCREENSHOT_{User}, STATUS_{User}, and TESTER_STATUS_{User} columns)
    - TESTER_STATUS columns: Always hidden (internal use for filtering)
    - Row hiding based on TESTER STATUS:
      - ISSUE rows: Visible (these are the active issues)
      - BLOCKED/KOREAN/NO ISSUE rows: Hidden
    - Row hiding based on MANAGER STATUS:
      - FIXED/NON-ISSUE: Hidden (issue resolved)
      - REPORTED/CHECKING: Visible (still being tracked)
      - Empty: Visible (pending manager review)

    Args:
        wb: Master workbook
        context_rows: Number of rows above/below visible rows to keep visible (default: 1)
        debug: If True, print debug info about what's being detected

    Returns: Tuple of (rows_hidden, sheets_hidden, hidden_columns_total)
    """
    hidden_rows = 0
    hidden_sheets = []
    hidden_columns_total = 0

    # === PHASE 0: RESET all sheets to visible first (fixes re-run UNHIDE bug) ===
    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue
        wb[sheet_name].sheet_state = 'visible'
        if debug:
            print(f"    [DEBUG] Reset sheet '{sheet_name}' to visible")

    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue

        ws = wb[sheet_name]

        # CRITICAL: Clear any AutoFilter from QA files before applying our logic
        if ws.auto_filter.ref:
            ws.auto_filter.ref = None
            if debug:
                print(f"    [DEBUG] Cleared AutoFilter from sheet: {sheet_name}")

        # Find all COMMENT_{User} columns (case-insensitive)
        comment_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).upper().startswith("COMMENT_"):
                comment_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found comment column: {header} at col {col}")

        if not comment_cols:
            if debug:
                print(f"    [DEBUG] No COMMENT_ columns found in {sheet_name}")
            continue

        # === RESET all COMMENT_*, SCREENSHOT_*, STATUS_* columns to visible first ===
        # NOTE: TESTER_STATUS_* columns STAY HIDDEN (they're internal for filtering only)
        # Uses case-insensitive matching for consistency with hide logic
        for col in comment_cols:
            header = ws.cell(row=1, column=col).value
            username = str(header).replace("COMMENT_", "") if header else ""
            username_upper = username.upper()
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].hidden = False

            # Also unhide paired SCREENSHOT_{User}, STATUS_{User}, and MANAGER_COMMENT_{User} columns (case-insensitive)
            for search_col in range(1, ws.max_column + 1):
                search_header = ws.cell(row=1, column=search_col).value
                if search_header:
                    search_header_upper = str(search_header).upper()
                    if search_header_upper in [f"SCREENSHOT_{username_upper}", f"STATUS_{username_upper}", f"MANAGER_COMMENT_{username_upper}"]:
                        search_col_letter = get_column_letter(search_col)
                        ws.column_dimensions[search_col_letter].hidden = False
                    # TESTER_STATUS always stays hidden
                    elif search_header_upper == f"TESTER_STATUS_{username_upper}":
                        search_col_letter = get_column_letter(search_col)
                        ws.column_dimensions[search_col_letter].hidden = True
            if debug:
                print(f"    [DEBUG] Reset column group for user '{username}' to visible (TESTER_STATUS stays hidden)")

        # First pass: Find rows that have comments AND track which columns have comments
        rows_with_comments = set()
        cols_with_comments = set()

        for row in range(2, ws.max_row + 1):
            for col in comment_cols:
                value = ws.cell(row=row, column=col).value
                if value is not None and str(value).strip():
                    rows_with_comments.add(row)
                    cols_with_comments.add(col)
                    if debug and row <= 10:
                        print(f"    [DEBUG] Row {row} has comment in col {col}: {repr(str(value)[:30])}")

        # If NO comments in entire sheet, hide the sheet tab
        if not rows_with_comments:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has {len(comment_cols)} COMMENT_ columns but ALL are empty - HIDING SHEET")
            ws.sheet_state = 'hidden'
            hidden_sheets.append(sheet_name)
            continue
        else:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has {len(rows_with_comments)} rows with comments - keeping visible")

        # Column hiding: Hide COMMENT_{User} columns that are entirely empty in this sheet
        hidden_cols_this_sheet = 0
        for col in comment_cols:
            if col not in cols_with_comments:
                header = ws.cell(row=1, column=col).value
                username = str(header).replace("COMMENT_", "") if header else ""

                # Hide the COMMENT_{User} column
                col_letter = ws.cell(row=1, column=col).column_letter
                ws.column_dimensions[col_letter].hidden = True
                hidden_cols_this_sheet += 1

                # Find and hide paired columns (case-insensitive match)
                for search_col in range(1, ws.max_column + 1):
                    search_header = ws.cell(row=1, column=search_col).value
                    if search_header:
                        search_header_upper = str(search_header).upper()
                        username_upper = username.upper()
                        if search_header_upper in [f"SCREENSHOT_{username_upper}", f"STATUS_{username_upper}", f"TESTER_STATUS_{username_upper}", f"MANAGER_COMMENT_{username_upper}"]:
                            search_col_letter = ws.cell(row=1, column=search_col).column_letter
                            ws.column_dimensions[search_col_letter].hidden = True
                            hidden_cols_this_sheet += 1

                if debug:
                    print(f"    [DEBUG] Hidden empty column group for user: {username}")

        hidden_columns_total += hidden_cols_this_sheet

        # === HIDE EMPTY SCREENSHOT_{User} COLUMNS (case-insensitive) ===
        # Even if user has comments, hide SCREENSHOT column if ALL cells are empty
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).upper().startswith("SCREENSHOT_"):
                # Check if column is already hidden
                col_letter = get_column_letter(col)
                if ws.column_dimensions[col_letter].hidden:
                    continue  # Skip already hidden columns

                # Check if any cell has content
                has_content = False
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_content = True
                        break

                if not has_content:
                    ws.column_dimensions[col_letter].hidden = True
                    hidden_columns_total += 1
                    if debug:
                        print(f"    [DEBUG] Hidden empty SCREENSHOT column: {header}")

        # === FIND TESTER_STATUS_{User} columns for tester status hiding (case-insensitive) ===
        tester_status_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).upper().startswith("TESTER_STATUS_"):
                tester_status_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found tester status column: {header} at col {col}")

        # Find rows to hide due to tester status (BLOCKED, KOREAN, NO ISSUE)
        rows_non_issue_by_tester = set()
        rows_with_issue_status = set()
        # Accept both "NO ISSUE" (standard) and "NON-ISSUE" (Script-type)
        TESTER_HIDE_STATUSES = {"BLOCKED", "KOREAN", "NO ISSUE", "NON-ISSUE"}
        for row in range(2, ws.max_row + 1):
            has_issue = False
            has_any_status = False
            for col in tester_status_cols:
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    has_any_status = True
                    status_upper = str(value).strip().upper()
                    if status_upper == "ISSUE":
                        has_issue = True
                        rows_with_issue_status.add(row)
                        break
            if has_any_status and not has_issue:
                rows_non_issue_by_tester.add(row)
                if debug and row <= 20:
                    print(f"    [DEBUG] Row {row} has tester status but no ISSUE - will hide")

        if debug:
            if rows_with_issue_status:
                print(f"    [DEBUG] {len(rows_with_issue_status)} rows with ISSUE tester status (will show)")
            if rows_non_issue_by_tester:
                print(f"    [DEBUG] {len(rows_non_issue_by_tester)} rows with only non-ISSUE tester status (will hide)")

        # === FIND STATUS_{User} columns for manager status hiding (case-insensitive) ===
        manager_status_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_upper = str(header).upper()
                if header_upper.startswith("STATUS_") and not header_upper.startswith("TESTER_STATUS_"):
                    manager_status_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found manager status column: {header} at col {col}")

        # Find rows to hide due to manager status (FIXED, NON-ISSUE only)
        # Accept both "NON-ISSUE" (hyphen) and "NON ISSUE" (space) for backwards compatibility
        rows_resolved_by_manager = set()
        MANAGER_HIDE_STATUSES = {"FIXED", "NON-ISSUE", "NON ISSUE"}
        for row in range(2, ws.max_row + 1):
            for col in manager_status_cols:
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip().upper() in MANAGER_HIDE_STATUSES:
                    rows_resolved_by_manager.add(row)
                    if debug and row <= 20:
                        print(f"    [DEBUG] Row {row} has manager status '{value}' - will hide")
                    break

        if debug and rows_resolved_by_manager:
            print(f"    [DEBUG] {len(rows_resolved_by_manager)} rows resolved by manager (FIXED/NON-ISSUE)")

        # === COMBINE HIDING RULES ===
        rows_to_hide = rows_non_issue_by_tester | rows_resolved_by_manager
        rows_to_show = (rows_with_comments - rows_non_issue_by_tester) - rows_resolved_by_manager

        # Add context rows
        for row in list(rows_to_show):
            for offset in range(1, context_rows + 1):
                if row - offset >= 2 and row - offset not in rows_to_hide:
                    rows_to_show.add(row - offset)
            for offset in range(1, context_rows + 1):
                if row + offset <= ws.max_row and row + offset not in rows_to_hide:
                    rows_to_show.add(row + offset)

        # If NO rows to show after filtering, hide entire sheet
        if not rows_to_show:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has comments but NONE are visible ISSUE rows - HIDING SHEET")
            ws.sheet_state = 'hidden'
            hidden_sheets.append(sheet_name)
            continue

        # Pass 1: UNHIDE all rows first
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].hidden = False

        # Pass 2: Hide rows not in the show set
        for row in range(2, ws.max_row + 1):
            if row not in rows_to_show:
                ws.row_dimensions[row].hidden = True
                hidden_rows += 1

    return hidden_rows, hidden_sheets, hidden_columns_total


def autofit_rows_with_wordwrap(wb, default_row_height: int = 15, chars_per_line: int = 50):
    """
    Apply word wrap to all cells and autofit column widths + row heights based on content.

    Features:
    - Auto-width for COMMENT_{User} columns (min 40, max 80)
    - Fixed widths for other user columns (SCREENSHOT, STATUS, TESTER_STATUS)
    - Auto-height for all rows based on content

    Args:
        wb: Workbook
        default_row_height: Default height for single-line rows
        chars_per_line: Estimated characters per line (for height calculation)
    """
    # Column width settings
    COMMENT_MIN_WIDTH = 40
    COMMENT_MAX_WIDTH = 80
    SCREENSHOT_WIDTH = 25
    STATUS_WIDTH = 12

    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue

        ws = wb[sheet_name]

        # === PHASE 1: Auto-fit column widths ===
        # Autofit ALL columns - hiding happens AFTER this, so no need to check hidden state
        # Bonus: if user unhides a column in Excel later, it already has proper width
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not header:
                continue

            header_str = str(header)
            header_upper = header_str.upper()  # Case-insensitive comparison
            col_letter = get_column_letter(col)

            # COMMENT_{User} columns: auto-width based on content
            if header_upper.startswith("COMMENT_"):
                max_content_len = len(header_str)  # Start with header length
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # Get longest line in cell (for multi-line content)
                        lines = str(cell_value).split('\n')
                        longest_line = max(len(line) for line in lines) if lines else 0
                        max_content_len = max(max_content_len, longest_line)

                # Apply width with min/max bounds
                width = min(max(max_content_len * 1.1 + 2, COMMENT_MIN_WIDTH), COMMENT_MAX_WIDTH)
                ws.column_dimensions[col_letter].width = width

            # SCREENSHOT_{User} columns: fixed width
            elif header_upper.startswith("SCREENSHOT_"):
                ws.column_dimensions[col_letter].width = SCREENSHOT_WIDTH

            # STATUS_{User} columns: fixed width
            elif header_upper.startswith("STATUS_") and not header_upper.startswith("TESTER_STATUS_"):
                ws.column_dimensions[col_letter].width = STATUS_WIDTH

            # TESTER_STATUS_{User} columns: fixed width (hidden anyway)
            elif header_upper.startswith("TESTER_STATUS_"):
                ws.column_dimensions[col_letter].width = STATUS_WIDTH

        # === PHASE 2: Auto-fit row heights ===
        for row in range(1, ws.max_row + 1):
            max_lines = 1

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                # Apply word wrap to all cells
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Calculate lines needed based on content
                if cell.value:
                    content = str(cell.value)
                    explicit_lines = content.count('\n') + 1

                    # Get column width for wrap calculation
                    col_letter = get_column_letter(col)
                    col_width = ws.column_dimensions[col_letter].width or chars_per_line
                    effective_chars_per_line = max(int(col_width * 0.9), 10)

                    longest_line = max(len(line) for line in content.split('\n')) if content else 0
                    wrapped_lines = max(1, (longest_line // effective_chars_per_line) + 1)
                    total_lines = explicit_lines + wrapped_lines - 1
                    max_lines = max(max_lines, total_lines)

            # Set row height based on content (15 points per line is standard)
            calculated_height = max_lines * default_row_height
            # Cap at reasonable max (300 points)
            ws.row_dimensions[row].height = min(calculated_height, 300)
