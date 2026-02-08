"""
Excel Operations Module
=======================
Workbook loading, saving, and manipulation utilities.
"""

import os
import re
import shutil
import zipfile
import tempfile
from copy import copy
from datetime import datetime as _datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import (
    MASTER_FOLDER_EN, MASTER_FOLDER_CN,
    IMAGES_FOLDER_EN, IMAGES_FOLDER_CN,
    CATEGORY_TO_MASTER, STATUS_OPTIONS, MANAGER_STATUS_OPTIONS,
    TRACKER_STYLES, get_target_master_category, SCRIPT_TYPE_CATEGORIES,
    TRANSLATION_COLS
)


# =============================================================================
# WORKBOOK LOADING
# =============================================================================

def repair_excel_filters(filepath: Path) -> Optional[Path]:
    """
    Repair Excel file by stripping corrupted auto-filter XML.

    Excel files are ZIP archives. This extracts, removes autoFilter
    elements from worksheet XML, and repacks to a temp file.

    Args:
        filepath: Path to corrupted Excel file

    Returns:
        Path to repaired temp file, or None if repair failed
    """
    try:
        # Create temp file for repaired Excel
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)

        with zipfile.ZipFile(filepath, 'r') as zip_in:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for item in zip_in.namelist():
                    data = zip_in.read(item)

                    # Strip autoFilter from worksheet XML files
                    if item.startswith('xl/worksheets/') and item.endswith('.xml'):
                        content = data.decode('utf-8')
                        # Remove autoFilter elements (including corrupted ones)
                        content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                        content = re.sub(r'<autoFilter[^/]*/>', '', content)
                        data = content.encode('utf-8')

                    zip_out.writestr(item, data)

        return Path(temp_path)
    except Exception as e:
        print(f"    Repair failed: {e}")
        return None


def safe_load_workbook(filepath: Path, **kwargs) -> Workbook:
    """
    Safely load an Excel workbook, handling common corruption issues.

    ROBUST: Auto-repairs corrupted auto-filters by stripping them from XML.

    Args:
        filepath: Path to the Excel file
        **kwargs: Additional arguments passed to openpyxl.load_workbook

    Returns:
        Workbook object

    Raises:
        Exception for unrecoverable errors
    """
    try:
        return openpyxl.load_workbook(filepath, **kwargs)
    except ValueError as e:
        error_msg = str(e)

        # Handle corrupted filter values - AUTO REPAIR
        if "numerical or a string containing a wildcard" in error_msg or "could not read worksheets" in error_msg:
            print(f"    WARNING: Corrupted filters detected, auto-repairing...")

            repaired_path = repair_excel_filters(filepath)
            if repaired_path:
                try:
                    wb = openpyxl.load_workbook(repaired_path, **kwargs)
                    try:
                        os.unlink(repaired_path)
                    except:
                        pass
                    print(f"    SUCCESS: File repaired and loaded (filters stripped)")
                    return wb
                except Exception as repair_error:
                    print(f"    Repair load failed: {repair_error}")
                    try:
                        os.unlink(repaired_path)
                    except:
                        pass

            print(f"    ERROR: Auto-repair failed for '{filepath.name}'")
            print(f"           Manual fix: Open in Excel → Data → Clear → Save")
            raise ValueError(f"Corrupted Excel file: {filepath.name}. Auto-repair failed.")

        raise
    except Exception as e:
        print(f"    ERROR loading '{filepath}': {e}")
        raise


# =============================================================================
# COLUMN OPERATIONS
# =============================================================================

def find_column_by_header(ws, header_name: str, case_insensitive: bool = True) -> Optional[int]:
    """
    Find column index by header name.

    Args:
        ws: Worksheet
        header_name: Header to search for
        case_insensitive: Match case-insensitively

    Returns:
        Column index (1-based) or None if not found
    """
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header_str = str(header).strip()
            if case_insensitive:
                if header_str.upper() == header_name.upper():
                    return col
            else:
                if header_str == header_name:
                    return col
    return None


def build_column_map(ws) -> Dict[str, int]:
    """
    Build a dict mapping header names to column indices (1-based).

    Scans row 1 once and returns {HEADER_UPPER: col_index}.
    First occurrence wins (matches find_column_by_header behavior).
    Keys are uppercased for case-insensitive matching.

    Uses iter_rows(values_only=True) which is safe for both normal
    and read_only worksheets (ws.cell() is unreliable in read_only mode).

    Args:
        ws: Worksheet (normal or read_only)

    Returns:
        Dict mapping uppercase header name to column index
    """
    col_map = {}
    header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_tuple = next(header_iter, None)
    if not header_tuple:
        return col_map
    for idx, header_val in enumerate(header_tuple):
        if header_val:
            key = str(header_val).strip().upper()
            if key not in col_map:
                col_map[key] = idx + 1  # 1-based
    return col_map


def find_column_by_prefix(ws, prefix: str) -> Optional[int]:
    """
    Find column index by header prefix.

    Args:
        ws: Worksheet
        prefix: Header prefix to match

    Returns:
        Column index (1-based) or None if not found
    """
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).upper().startswith(prefix.upper()):
            return col
    return None


def preload_worksheet_data(ws) -> Tuple[Dict[str, int], List[tuple]]:
    """
    FAST: Pre-load entire worksheet into memory as tuples.

    Uses iter_rows(values_only=True) which is 10-50x faster than ws.cell().
    Call ONCE per worksheet, then access data by tuple index.

    Args:
        ws: Worksheet to preload

    Returns:
        Tuple of (col_index_map, row_tuples) where:
        - col_index_map: {HEADER_UPPER: 0-based index into tuple}
        - row_tuples: List of tuples, index 0 = row 2 (first data row)

    Usage:
        col_idx, rows = preload_worksheet_data(ws)
        for row_tuple in rows:
            status = row_tuple[col_idx.get('STATUS')] if 'STATUS' in col_idx else None
            comment = row_tuple[col_idx.get('COMMENT')] if 'COMMENT' in col_idx else None
    """
    # Read ALL data in one call (10-50x faster than cell-by-cell)
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return {}, []

    # Build column index from header row (row 0 = Excel row 1)
    header_row = all_rows[0]
    col_idx = {}
    for i, header in enumerate(header_row):
        if header:
            key = str(header).strip().upper()
            if key not in col_idx:  # First occurrence wins
                col_idx[key] = i

    # Data rows (skip header)
    data_rows = all_rows[1:] if len(all_rows) > 1 else []

    return col_idx, data_rows


def get_tuple_value(row_tuple: tuple, col_idx: Dict[str, int], header: str, default=None):
    """
    Safely get value from preloaded tuple by header name.

    Args:
        row_tuple: Preloaded row tuple
        col_idx: Column index map from preload_worksheet_data()
        header: Header name (case-insensitive)
        default: Default value if not found

    Returns:
        Value from tuple or default
    """
    idx = col_idx.get(header.upper())
    if idx is not None and idx < len(row_tuple):
        return row_tuple[idx]
    return default


# =============================================================================
# TESTER DATA PRESERVATION (for master rebuild)
# =============================================================================

def _parse_updated_timestamp(comment_value):
    """Extract (updated: YYMMDD HHMM) timestamp from comment metadata.

    Returns datetime for comparison, or datetime.min if not found.
    """
    if not comment_value:
        return _datetime.min
    match = re.search(r'\(updated:\s*(\d{6}\s+\d{4})\)', str(comment_value))
    if match:
        try:
            return _datetime.strptime(match.group(1), "%y%m%d %H%M")
        except ValueError:
            pass
    return _datetime.min


def extract_tester_data_from_master(
    master_path: Path,
    category: str,
    is_english: bool
) -> Dict[str, Dict[tuple, Dict[str, Dict]]]:
    """
    Extract all tester column data from existing master before rebuild.

    This enables preserving tester work (comments, status, screenshots) when
    rebuilding the master from a new template. Data is keyed by content so it
    can be restored to matching rows in the new master.

    Args:
        master_path: Path to existing master file
        category: Category name (Quest, Item, Script, etc.)
        is_english: Whether this is EN (True) or CN (False) master

    Returns:
        Dict with structure:
        {sheet_name: {content_key: {username: {
            "comment": value,
            "tester_status": value,
            "manager_status": value,
            "manager_comment": value,
            "screenshot": value,
            "screenshot_hyperlink": target
        }}}}
    """
    if not master_path.exists():
        return {}

    extracted = {}
    category_lower = category.lower()
    is_script = category_lower in SCRIPT_TYPE_CATEGORIES

    try:
        # Use read_only=True for performance (3-5x faster)
        wb = safe_load_workbook(master_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name == "STATUS":
                continue

            ws = wb[sheet_name]

            # Check for empty sheet
            if ws.max_row is None or ws.max_row < 2:
                continue
            if ws.max_column is None or ws.max_column < 1:
                continue

            # Build column map from headers using streaming (iter_rows)
            header_iter = ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column, values_only=True)
            header_tuple = next(header_iter, None)
            if not header_tuple:
                continue

            # Find all user columns and content columns
            col_map = {}  # header_upper -> 0-based index
            for idx, header_val in enumerate(header_tuple):
                if header_val:
                    col_map[str(header_val).strip().upper()] = idx

            # Identify user columns by pattern
            user_columns = {}  # username -> {comment_idx, tester_status_idx, status_idx, manager_comment_idx, screenshot_idx}
            for header_upper, idx in col_map.items():
                if header_upper.startswith("COMMENT_"):
                    username = header_upper[8:]
                    if username not in user_columns:
                        user_columns[username] = {}
                    user_columns[username]["comment_idx"] = idx
                elif header_upper.startswith("TESTER_STATUS_"):
                    username = header_upper[14:]
                    if username not in user_columns:
                        user_columns[username] = {}
                    user_columns[username]["tester_status_idx"] = idx
                elif header_upper.startswith("STATUS_"):
                    username = header_upper[7:]
                    if username not in user_columns:
                        user_columns[username] = {}
                    user_columns[username]["status_idx"] = idx
                elif header_upper.startswith("MANAGER_COMMENT_"):
                    username = header_upper[16:]
                    if username not in user_columns:
                        user_columns[username] = {}
                    user_columns[username]["manager_comment_idx"] = idx
                elif header_upper.startswith("SCREENSHOT_"):
                    username = header_upper[11:]
                    if username not in user_columns:
                        user_columns[username] = {}
                    user_columns[username]["screenshot_idx"] = idx

            if not user_columns:
                continue

            # Find content columns for building keys
            # Standard: STRINGID + Translation column
            # Item: ItemName + ItemDesc + STRINGID
            # Script: EventName + Text
            # Contents: INSTRUCTIONS

            # Import sanitization to match build_master_index() key format
            from core.matching import sanitize_stringid_for_match

            stringid_idx = col_map.get("STRINGID")
            eventname_idx = col_map.get("EVENTNAME")
            text_idx = col_map.get("TEXT") or col_map.get("TRANSLATION")
            instructions_idx = col_map.get("INSTRUCTIONS")

            # For Item category
            itemname_eng_idx = col_map.get("ITEMNAME(ENG)")
            itemname_loc_idx = col_map.get("ITEMNAME(LOC)")
            itemdesc_eng_idx = col_map.get("ITEMDESC(ENG)")
            itemdesc_loc_idx = col_map.get("ITEMDESC(LOC)")

            # For standard categories (Quest, Knowledge, etc.)
            # Translation column positions from config
            from config import TRANSLATION_COLS, ITEM_DESC_COLS
            trans_col_config = TRANSLATION_COLS.get(category, {"eng": 2, "other": 3})
            trans_col_idx = (trans_col_config["eng"] if is_english else trans_col_config["other"]) - 1  # 0-based

            sheet_data = {}

            # Stream through all data rows
            for row_tuple in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
                # Build content key based on category
                content_key = None

                if category_lower == "contents":
                    # Contents: INSTRUCTIONS column
                    if instructions_idx is not None and instructions_idx < len(row_tuple):
                        instructions = str(row_tuple[instructions_idx] or "").strip()
                        if instructions:
                            content_key = (instructions,)

                elif category_lower == "item":
                    # Item: (ItemName, ItemDesc, STRINGID)
                    name_idx = itemname_eng_idx if is_english else itemname_loc_idx
                    desc_idx = itemdesc_eng_idx if is_english else itemdesc_loc_idx
                    if name_idx is not None and name_idx < len(row_tuple):
                        item_name = str(row_tuple[name_idx] or "").strip()
                        item_desc = ""
                        if desc_idx is not None and desc_idx < len(row_tuple):
                            item_desc = str(row_tuple[desc_idx] or "").strip()
                        stringid = sanitize_stringid_for_match(row_tuple[stringid_idx]) if stringid_idx is not None and stringid_idx < len(row_tuple) else ""
                        if item_name:
                            content_key = (item_name, item_desc, stringid)

                elif is_script:
                    # Script: (Text, EventName) - matches index order in matching.py
                    eventname = ""
                    if eventname_idx is not None and eventname_idx < len(row_tuple):
                        eventname = str(row_tuple[eventname_idx] or "").strip()
                    text = ""
                    if text_idx is not None and text_idx < len(row_tuple):
                        text = str(row_tuple[text_idx] or "").strip()
                    if eventname or text:
                        content_key = (text, eventname)  # Order: (text, eventname) to match build_master_index()

                else:
                    # Standard: (STRINGID, Translation)
                    stringid = sanitize_stringid_for_match(row_tuple[stringid_idx]) if stringid_idx is not None and stringid_idx < len(row_tuple) else ""
                    translation = ""
                    if trans_col_idx < len(row_tuple):
                        translation = str(row_tuple[trans_col_idx] or "").strip()
                    if translation:
                        content_key = (stringid, translation)

                if content_key is None:
                    continue

                # Extract tester data for each user
                for username, cols in user_columns.items():
                    user_data = {}
                    has_data = False

                    # Comment
                    c_idx = cols.get("comment_idx")
                    if c_idx is not None and c_idx < len(row_tuple):
                        val = row_tuple[c_idx]
                        if val is not None and str(val).strip():
                            user_data["comment"] = val
                            has_data = True

                    # Tester status
                    ts_idx = cols.get("tester_status_idx")
                    if ts_idx is not None and ts_idx < len(row_tuple):
                        val = row_tuple[ts_idx]
                        if val is not None and str(val).strip():
                            user_data["tester_status"] = val
                            has_data = True

                    # Manager status (STATUS_{user})
                    ms_idx = cols.get("status_idx")
                    if ms_idx is not None and ms_idx < len(row_tuple):
                        val = row_tuple[ms_idx]
                        if val is not None and str(val).strip():
                            user_data["manager_status"] = val
                            has_data = True

                    # Manager comment
                    mc_idx = cols.get("manager_comment_idx")
                    if mc_idx is not None and mc_idx < len(row_tuple):
                        val = row_tuple[mc_idx]
                        if val is not None and str(val).strip():
                            user_data["manager_comment"] = val
                            has_data = True

                    # Screenshot (not for Script)
                    # Note: In read_only mode, hyperlinks are not available
                    # We store the value; hyperlink will be reconstructed
                    if not is_script:
                        sc_idx = cols.get("screenshot_idx")
                        if sc_idx is not None and sc_idx < len(row_tuple):
                            val = row_tuple[sc_idx]
                            if val is not None and str(val).strip():
                                user_data["screenshot"] = val
                                # Hyperlink target follows standard pattern
                                user_data["screenshot_hyperlink"] = f"Images/{val}"
                                has_data = True

                    if has_data:
                        if content_key not in sheet_data:
                            sheet_data[content_key] = {}
                        if username in sheet_data[content_key]:
                            # Duplicate row — merge: prefer entry with MORE data (especially manager fields)
                            existing = sheet_data[content_key][username]
                            existing_has_manager = "manager_status" in existing or "manager_comment" in existing
                            new_has_manager = "manager_status" in user_data or "manager_comment" in user_data

                            if new_has_manager and not existing_has_manager:
                                # New row has manager data that existing lacks — replace
                                sheet_data[content_key][username] = user_data
                            elif not new_has_manager and existing_has_manager:
                                # Existing has manager data — keep it, but merge any new fields
                                for k, v in user_data.items():
                                    if k not in existing:
                                        existing[k] = v
                            else:
                                # Both have (or lack) manager data — keep most recently updated
                                existing_ts = _parse_updated_timestamp(existing.get("comment"))
                                new_ts = _parse_updated_timestamp(user_data.get("comment"))
                                if new_ts > existing_ts:
                                    # Merge manager data from existing into new before replacing
                                    for k in ("manager_status", "manager_comment"):
                                        if k in existing and k not in user_data:
                                            user_data[k] = existing[k]
                                    sheet_data[content_key][username] = user_data
                        else:
                            sheet_data[content_key][username] = user_data

            if sheet_data:
                extracted[sheet_name] = sheet_data

        wb.close()

    except Exception as e:
        print(f"    WARNING: Failed to extract tester data from {master_path.name}: {e}")
        return {}

    # Summary
    total_entries = sum(
        sum(len(users) for users in sheet.values())
        for sheet in extracted.values()
    )
    if total_entries > 0:
        print(f"    Extracted {total_entries} tester data entries from {len(extracted)} sheets")

    return extracted


def restore_tester_data_to_master(
    master_wb: Workbook,
    tester_data: Dict[str, Dict[tuple, Dict[str, Dict]]],
    category: str,
    is_english: bool
) -> Dict:
    """
    Restore tester data to newly created master using content-based matching.

    Args:
        master_wb: New master workbook (already created from template)
        tester_data: Data from extract_tester_data_from_master()
        category: Category name
        is_english: Whether this is EN or CN master

    Returns:
        Dict with {"restored": count, "orphaned": count, "sheets_processed": count}
    """
    if not tester_data:
        return {"restored": 0, "orphaned": 0, "sheets_processed": 0}

    # Import matching functions
    from core.matching import build_master_index, sanitize_stringid_for_match

    # Import processing functions for column creation
    from core.processing import (
        find_or_create_user_columns,
        COMMENT_FILL_ISSUE, COMMENT_FONT_ISSUE,
        COMMENT_FILL_BLOCKED, COMMENT_FONT_BLOCKED,
        COMMENT_FILL_KOREAN, COMMENT_FONT_KOREAN,
        COMMENT_FILL_NO_ISSUE, COMMENT_FONT_NO_ISSUE,
        COMMENT_ALIGNMENT, COMMENT_BORDER,
        SCREENSHOT_FILL, SCREENSHOT_ALIGNMENT, SCREENSHOT_BORDER,
        SCREENSHOT_FONT_NORMAL,
        MANAGER_STATUS_ALIGNMENT,
        MANAGER_FONT_FIXED, MANAGER_FONT_REPORTED,
        MANAGER_FONT_CHECKING, MANAGER_FONT_NONISSUE,
        MANAGER_COMMENT_FILL, MANAGER_COMMENT_BORDER,
        sanitize_for_excel
    )

    category_lower = category.lower()
    is_script = category_lower in SCRIPT_TYPE_CATEGORIES

    stats = {"restored": 0, "orphaned": 0, "sheets_processed": 0}

    for sheet_name, sheet_data in tester_data.items():
        if sheet_name not in master_wb.sheetnames:
            # Sheet no longer exists in new master - all data orphaned
            orphaned_count = sum(len(users) for users in sheet_data.values())
            stats["orphaned"] += orphaned_count
            print(f"      {sheet_name}: REMOVED (orphaned {orphaned_count} entries)")
            continue

        master_ws = master_wb[sheet_name]
        stats["sheets_processed"] += 1

        # Build master index for content-based matching
        master_index = build_master_index(master_ws, category, is_english)

        # Collect all usernames from extracted data
        all_users = set()
        for users in sheet_data.values():
            all_users.update(users.keys())

        # Create user columns for ALL users (including absent testers)
        user_cols = {}  # username -> (comment_col, tester_status_col, status_col, manager_comment_col, screenshot_col)
        for username in all_users:
            user_cols[username] = find_or_create_user_columns(master_ws, username)

        sheet_restored = 0
        sheet_orphaned = 0

        # === TWO-PASS RESTORE ===
        # Pass 1: Primary matches (exact key) → consume all true duplicates
        # Pass 2: Fallback matches (partial key) → first unconsumed only
        # This ensures primary matches always win over fallback.

        # Build primary match results: content_key -> [matching_rows]
        primary_matched = {}   # content_key -> list of rows
        fallback_keys = []     # content_keys that need fallback

        for content_key in sheet_data:
            matching_rows = []

            if category_lower == "contents":
                instructions = content_key[0] if content_key else ""
                if instructions and instructions in master_index.get("all_primary", {}):
                    matching_rows = list(master_index["all_primary"][instructions])
                elif instructions and instructions in master_index["primary"]:
                    matching_rows.append(master_index["primary"][instructions])

            elif category_lower == "item":
                if len(content_key) >= 2:
                    stringid = content_key[2] if len(content_key) > 2 else ""
                    if stringid:
                        pk = (content_key[0], content_key[1], stringid)
                        if pk in master_index.get("all_primary", {}):
                            matching_rows = list(master_index["all_primary"][pk])

            elif is_script:
                if len(content_key) >= 2:
                    text, eventname = content_key[0], content_key[1]
                    if text and eventname:
                        pk = (text, eventname)
                        if pk in master_index.get("all_primary", {}):
                            matching_rows = list(master_index["all_primary"][pk])

            else:
                if len(content_key) >= 2:
                    stringid, translation = content_key[0], content_key[1]
                    if stringid and translation:
                        pk = (stringid, translation)
                        if pk in master_index.get("all_primary", {}):
                            matching_rows = list(master_index["all_primary"][pk])

            if matching_rows:
                primary_matched[content_key] = matching_rows
                for mr in matching_rows:
                    master_index["consumed"].add(mr)
            else:
                fallback_keys.append(content_key)

        # Pass 2: Fallback for unmatched keys (stringid changed, etc.)
        fallback_matched = {}
        for content_key in fallback_keys:
            matching_rows = []

            if category_lower == "contents":
                pass  # Contents has no fallback

            elif category_lower == "item":
                if len(content_key) >= 2:
                    fallback_key = (content_key[0], content_key[1])
                    if fallback_key in master_index["fallback"]:
                        for row in master_index["fallback"][fallback_key]:
                            if row not in master_index["consumed"]:
                                matching_rows.append(row)
                                break

            elif is_script:
                if len(content_key) >= 2:
                    eventname = content_key[1]
                    if eventname and eventname in master_index["fallback"]:
                        for row in master_index["fallback"][eventname]:
                            if row not in master_index["consumed"]:
                                matching_rows.append(row)
                                break

            else:
                if len(content_key) >= 2:
                    translation = content_key[1]
                    if translation and translation in master_index["fallback"]:
                        for row in master_index["fallback"][translation]:
                            if row not in master_index["consumed"]:
                                matching_rows.append(row)
                                break

            if matching_rows:
                fallback_matched[content_key] = matching_rows
                for mr in matching_rows:
                    master_index["consumed"].add(mr)

        # Combine results and restore
        all_matched = {**primary_matched, **fallback_matched}
        unmatched = set(sheet_data.keys()) - set(all_matched.keys())
        sheet_orphaned += sum(len(sheet_data[k]) for k in unmatched)

        for content_key, matching_rows in all_matched.items():
            user_data_dict = sheet_data[content_key]
            sheet_restored += len(user_data_dict)  # Count per user-entry (consistent with orphaned)

            # Restore data to ALL matching rows (replicate across true duplicates)
            for master_row in matching_rows:
                for username, user_data in user_data_dict.items():
                    cols = user_cols.get(username)
                    if not cols:
                        continue

                    comment_col, tester_status_col, status_col, manager_comment_col, screenshot_col = cols

                    # Restore comment
                    if "comment" in user_data:
                        cell = master_ws.cell(row=master_row, column=comment_col)
                        cell.value = sanitize_for_excel(user_data["comment"])

                        # Apply styling based on tester status
                        tester_status = user_data.get("tester_status", "")
                        status_upper = str(tester_status).strip().upper() if tester_status else ""
                        if status_upper == "ISSUE":
                            cell.fill = COMMENT_FILL_ISSUE
                            cell.font = COMMENT_FONT_ISSUE
                        elif status_upper == "BLOCKED":
                            cell.fill = COMMENT_FILL_BLOCKED
                            cell.font = COMMENT_FONT_BLOCKED
                        elif status_upper == "KOREAN":
                            cell.fill = COMMENT_FILL_KOREAN
                            cell.font = COMMENT_FONT_KOREAN
                        elif status_upper in ("NO ISSUE", "NON-ISSUE", "NON ISSUE"):
                            cell.fill = COMMENT_FILL_NO_ISSUE
                            cell.font = COMMENT_FONT_NO_ISSUE
                        cell.alignment = COMMENT_ALIGNMENT
                        cell.border = COMMENT_BORDER

                    # Restore tester status
                    if "tester_status" in user_data:
                        cell = master_ws.cell(row=master_row, column=tester_status_col)
                        cell.value = user_data["tester_status"]
                        cell.alignment = MANAGER_STATUS_ALIGNMENT

                    # Restore manager status
                    if "manager_status" in user_data:
                        cell = master_ws.cell(row=master_row, column=status_col)
                        status_value = user_data["manager_status"]
                        cell.value = status_value
                        cell.alignment = MANAGER_STATUS_ALIGNMENT

                        # Apply font color
                        status_upper = str(status_value).strip().upper()
                        if status_upper == "FIXED":
                            cell.font = MANAGER_FONT_FIXED
                        elif status_upper == "REPORTED":
                            cell.font = MANAGER_FONT_REPORTED
                        elif status_upper == "CHECKING":
                            cell.font = MANAGER_FONT_CHECKING
                        elif status_upper in ("NON-ISSUE", "NON ISSUE"):
                            cell.font = MANAGER_FONT_NONISSUE

                    # Restore manager comment
                    if "manager_comment" in user_data:
                        cell = master_ws.cell(row=master_row, column=manager_comment_col)
                        cell.value = user_data["manager_comment"]
                        cell.alignment = COMMENT_ALIGNMENT
                        cell.fill = MANAGER_COMMENT_FILL
                        cell.border = MANAGER_COMMENT_BORDER

                    # Restore screenshot (not for Script)
                    if not is_script and screenshot_col and "screenshot" in user_data:
                        cell = master_ws.cell(row=master_row, column=screenshot_col)
                        cell.value = sanitize_for_excel(user_data["screenshot"])
                        if "screenshot_hyperlink" in user_data:
                            cell.hyperlink = user_data["screenshot_hyperlink"]
                        cell.fill = SCREENSHOT_FILL
                        cell.alignment = SCREENSHOT_ALIGNMENT
                        cell.border = SCREENSHOT_BORDER
                        cell.font = SCREENSHOT_FONT_NORMAL

        stats["restored"] += sheet_restored
        stats["orphaned"] += sheet_orphaned

        if sheet_restored > 0 or sheet_orphaned > 0:
            print(f"      {sheet_name}: restored {sheet_restored}, orphaned {sheet_orphaned}")

    return stats


# =============================================================================
# DUPLICATE ROW POST-PROCESSING
# =============================================================================

def replicate_duplicate_row_data(master_wb, category: str, is_english: bool) -> int:
    """
    Post-process: for every group of duplicate rows (same STRINGID + Translation),
    find the row with the MOST data per user column and replicate it to ALL duplicates.

    This is a brute-force safety net that runs after all matching/restoration.
    Scans every sheet, groups rows by content key, and fills gaps.

    Returns total number of cells filled.
    """
    from core.matching import sanitize_stringid_for_match, get_translation_column
    from config import TRANSLATION_COLS, ITEM_DESC_COLS, SCRIPT_TYPE_CATEGORIES

    category_lower = category.lower()
    is_script = category_lower in SCRIPT_TYPE_CATEGORIES
    total_filled = 0

    for sheet_name in master_wb.sheetnames:
        if sheet_name == "STATUS":
            continue

        ws = master_wb[sheet_name]
        if ws.max_row is None or ws.max_row < 3:
            continue

        # Build header map (1-based column index)
        headers = {}
        for col in range(1, (ws.max_column or 0) + 1):
            val = ws.cell(1, col).value
            if val:
                headers[col] = str(val).strip().upper()

        # Find content key columns
        stringid_col = None
        trans_col = None
        eventname_col = None
        instructions_col = None
        itemname_col = None
        itemdesc_col = None

        for col, header in headers.items():
            if header == "STRINGID":
                stringid_col = col
            elif header == "TRANSLATION" or header == "TEXT":
                trans_col = col
            elif header == "EVENTNAME":
                eventname_col = col
            elif header == "INSTRUCTIONS":
                instructions_col = col
            elif header in ("ITEMNAME(ENG)", "ITEMNAME(LOC)", "ITEMNAME"):
                if itemname_col is None:
                    itemname_col = col
            elif header in ("ITEMDESC(ENG)", "ITEMDESC(LOC)", "ITEMDESC"):
                if itemdesc_col is None:
                    itemdesc_col = col

        # Find user data columns (COMMENT_, STATUS_, MANAGER_COMMENT_, SCREENSHOT_, TESTER_STATUS_)
        user_cols = []
        for col, header in headers.items():
            if any(header.startswith(prefix) for prefix in
                   ("COMMENT_", "STATUS_", "MANAGER_COMMENT_", "SCREENSHOT_", "TESTER_STATUS_")):
                user_cols.append(col)

        if not user_cols:
            continue

        # Group rows by content key
        row_groups = defaultdict(list)  # content_key -> [row_num, ...]

        for row in range(2, ws.max_row + 1):
            if category_lower == "contents":
                if instructions_col:
                    key_val = str(ws.cell(row, instructions_col).value or "").strip()
                    if key_val:
                        row_groups[key_val].append(row)
            elif category_lower == "item":
                name = str(ws.cell(row, itemname_col).value or "").strip() if itemname_col else ""
                desc = str(ws.cell(row, itemdesc_col).value or "").strip() if itemdesc_col else ""
                sid = sanitize_stringid_for_match(ws.cell(row, stringid_col).value) if stringid_col else ""
                if name:
                    row_groups[(name, desc, sid)].append(row)
            elif is_script:
                text = str(ws.cell(row, trans_col).value or "").strip() if trans_col else ""
                ename = str(ws.cell(row, eventname_col).value or "").strip() if eventname_col else ""
                if text or ename:
                    row_groups[(text, ename)].append(row)
            else:
                sid = sanitize_stringid_for_match(ws.cell(row, stringid_col).value) if stringid_col else ""
                trans = str(ws.cell(row, trans_col).value or "").strip() if trans_col else ""
                if trans:
                    row_groups[(sid, trans)].append(row)

        # For each group with duplicates, replicate data
        for key, rows in row_groups.items():
            if len(rows) < 2:
                continue

            # For each user column, find the best value across all rows in the group
            for col in user_cols:
                # Collect all values for this column across the group
                best_value = None
                best_row = None
                for row in rows:
                    val = ws.cell(row, col).value
                    if val is not None and str(val).strip():
                        best_value = val
                        best_row = row
                        break  # Take the first non-empty value

                if best_value is None:
                    continue

                # Copy to all rows that are empty for this column
                for row in rows:
                    cell = ws.cell(row, col)
                    if cell.value is None or not str(cell.value).strip():
                        cell.value = best_value
                        # Copy formatting from source cell
                        src = ws.cell(best_row, col)
                        cell.font = src.font.copy() if src.font else None
                        cell.fill = src.fill.copy() if src.fill else None
                        cell.alignment = src.alignment.copy() if src.alignment else None
                        cell.border = src.border.copy() if src.border else None
                        # Copy hyperlink if present
                        if src.hyperlink:
                            cell.hyperlink = src.hyperlink.target
                        total_filled += 1

    if total_filled > 0:
        print(f"    Duplicate post-process: replicated {total_filled} cells across duplicate rows")

    return total_filled


# =============================================================================
# MASTER FILE OPERATIONS
# =============================================================================

def get_or_create_master(
    category: str,
    master_folder: Path,
    template_file: Path = None,
    rebuild: bool = True,
    is_english: bool = True
) -> Tuple[Optional[Workbook], Path]:
    """
    Get or create master workbook for a category.

    When rebuilding, preserves ALL tester data (comments, status, screenshots)
    from the old master by extracting it before deletion and restoring it to
    matching rows in the new master. This prevents data loss when templates change.

    Args:
        category: Category name (Quest, Knowledge, etc.)
        master_folder: Target Master folder (EN or CN)
        template_file: Path to first QA file to use as template
        rebuild: If True, delete old and create fresh. If False, append to existing.
        is_english: If True, EN master (affects column detection for matching)

    Returns:
        Tuple of (Workbook, master_path)
    """
    target_category = get_target_master_category(category)
    master_path = master_folder / f"Master_{target_category}.xlsx"

    # If not rebuilding and master exists, load it and add new sheets
    if not rebuild and master_path.exists():
        print(f"  Loading existing master: {master_path.name} (appending {category} sheets)")
        wb = safe_load_workbook(master_path)

        # Copy sheets from template that don't exist in master yet
        if template_file:
            template_wb = safe_load_workbook(template_file)
            sheets_added = []

            for sheet_name in template_wb.sheetnames:
                if sheet_name == "STATUS":
                    continue
                if sheet_name not in wb.sheetnames:
                    # Copy sheet from template to master
                    source_ws = template_wb[sheet_name]
                    target_ws = wb.create_sheet(sheet_name)

                    # Copy all cells
                    for row in source_ws.iter_rows():
                        for cell in row:
                            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.alignment = copy(cell.alignment)

                    # Copy column widths
                    for col_letter, col_dim in source_ws.column_dimensions.items():
                        target_ws.column_dimensions[col_letter].width = col_dim.width

                    # Copy row heights
                    for row_num, row_dim in source_ws.row_dimensions.items():
                        target_ws.row_dimensions[row_num].height = row_dim.height

                    # Clean the new sheet (delete tester columns)
                    # MEMO is Script category's equivalent of COMMENT
                    cols_to_delete = []
                    for h in ["STATUS", "COMMENT", "MEMO", "SCREENSHOT"]:
                        col = find_column_by_header(target_ws, h)
                        if col:
                            cols_to_delete.append(col)
                    cols_to_delete.sort(reverse=True)
                    for col in cols_to_delete:
                        target_ws.delete_cols(col)

                    sheets_added.append(sheet_name)

            template_wb.close()

            if sheets_added:
                print(f"    Added {len(sheets_added)} new sheets from {category}: {', '.join(sheets_added)}")

        return wb, master_path

    # Rebuild mode: extract tester data BEFORE deletion, then create fresh
    extracted_data = {}
    preserved_sheets = {}  # Sheets from other categories to preserve through rebuild
    if rebuild and master_path.exists():
        # Extract tester data BEFORE deletion
        print(f"  Extracting tester data from: {master_path.name}")
        extracted_data = extract_tester_data_from_master(master_path, category, is_english)

        # Preserve sheets that belong to OTHER categories in the same worker group.
        # E.g., when Item rebuilds Master_Item.xlsx, Gimmick sheets must survive.
        # When Sequencer rebuilds Master_Script.xlsx, Dialog sheets must survive.
        if template_file:
            template_wb = safe_load_workbook(template_file, read_only=True)
            template_sheet_names = set(template_wb.sheetnames)
            template_wb.close()

            old_wb = safe_load_workbook(master_path)
            for sn in old_wb.sheetnames:
                if sn in template_sheet_names or sn == "STATUS":
                    continue
                # This sheet belongs to another category - preserve it verbatim
                old_ws = old_wb[sn]
                sheet_info = {"cells": [], "col_widths": {}, "row_heights": {}}
                for row in old_ws.iter_rows():
                    for cell in row:
                        ci = {"row": cell.row, "col": cell.column, "value": cell.value}
                        if cell.has_style:
                            ci["font"] = copy(cell.font)
                            ci["border"] = copy(cell.border)
                            ci["fill"] = copy(cell.fill)
                            ci["number_format"] = cell.number_format
                            ci["alignment"] = copy(cell.alignment)
                        sheet_info["cells"].append(ci)
                for col_letter, col_dim in old_ws.column_dimensions.items():
                    sheet_info["col_widths"][col_letter] = col_dim.width
                for row_num, row_dim in old_ws.row_dimensions.items():
                    sheet_info["row_heights"][row_num] = row_dim.height
                preserved_sheets[sn] = sheet_info
            old_wb.close()

            if preserved_sheets:
                print(f"    Preserving {len(preserved_sheets)} sheets from other categories: {', '.join(preserved_sheets.keys())}")

        print(f"  Deleting old master: {master_path.name} (rebuilding fresh)")
        master_path.unlink()

    if template_file:
        print(f"  Creating new master from: {template_file.name}")
        wb = safe_load_workbook(template_file)

        # Delete tester columns (MEMO is Script category's equivalent of COMMENT)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Clear AutoFilter
            if ws.auto_filter.ref:
                ws.auto_filter.ref = None

            cols_to_delete = []
            for h in ["STATUS", "COMMENT", "MEMO", "SCREENSHOT"]:
                col = find_column_by_header(ws, h)
                if col:
                    cols_to_delete.append(col)

            cols_to_delete.sort(reverse=True)
            for col in cols_to_delete:
                ws.delete_cols(col)
                print(f"    Deleted column {col} from {sheet_name}")

        print(f"    Master cleaned: tester columns removed")

        # Restore tester data to new master (if any was extracted)
        # NOTE: This runs BEFORE preserved sheets are added, so restore only
        # targets template sheets. Entries from other categories are harmlessly
        # orphaned since those sheets aren't in the workbook yet.
        if extracted_data:
            print(f"  Restoring tester data to new master...")
            restore_stats = restore_tester_data_to_master(wb, extracted_data, category, is_english)
            print(f"    Restored: {restore_stats['restored']} cells, Orphaned: {restore_stats['orphaned']} rows")

        # Restore preserved sheets from other categories (verbatim, with all data)
        # These sheets already contain their complete data including any tester columns
        # from previous compilations, so no cleaning or restoration needed.
        if preserved_sheets:
            for sn, sheet_info in preserved_sheets.items():
                if sn not in wb.sheetnames:
                    new_ws = wb.create_sheet(sn)
                    for ci in sheet_info["cells"]:
                        new_cell = new_ws.cell(row=ci["row"], column=ci["col"], value=ci["value"])
                        if "font" in ci:
                            new_cell.font = ci["font"]
                            new_cell.border = ci["border"]
                            new_cell.fill = ci["fill"]
                            new_cell.number_format = ci["number_format"]
                            new_cell.alignment = ci["alignment"]
                    for col_letter, width in sheet_info["col_widths"].items():
                        new_ws.column_dimensions[col_letter].width = width
                    for row_num, height in sheet_info["row_heights"].items():
                        new_ws.row_dimensions[row_num].height = height
            print(f"    Restored {len(preserved_sheets)} preserved sheets: {', '.join(preserved_sheets.keys())}")

        return wb, master_path
    else:
        print(f"  ERROR: No template file for {category}")
        return None, master_path


# =============================================================================
# FOLDER OPERATIONS
# =============================================================================

def ensure_master_folders():
    """Create both EN and CN master folders if they don't exist."""
    MASTER_FOLDER_EN.mkdir(exist_ok=True)
    MASTER_FOLDER_CN.mkdir(exist_ok=True)
    IMAGES_FOLDER_EN.mkdir(exist_ok=True)
    IMAGES_FOLDER_CN.mkdir(exist_ok=True)


def get_master_folder(username: str, tester_mapping: Dict[str, str]) -> Tuple[Path, Path]:
    """Get the correct master folder based on tester's language."""
    lang = tester_mapping.get(username, "EN")
    if lang == "CN":
        return MASTER_FOLDER_CN, IMAGES_FOLDER_CN
    return MASTER_FOLDER_EN, IMAGES_FOLDER_EN


# =============================================================================
# IMAGE OPERATIONS
# =============================================================================

IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}


def copy_images_with_unique_names(
    qa_folder_info: Dict,
    images_folder: Path,
    skip_images: set = None
) -> Dict[str, str]:
    """
    Copy images from QA folder to Images/ with original names.

    Optimization: Skip images in skip_images set (e.g., FIXED screenshots).

    Args:
        qa_folder_info: Dict with {folder_path, username, category, images}
        images_folder: Target Images folder
        skip_images: Set of image filenames to skip (case-insensitive)

    Returns:
        Dict mapping original_name -> original_name
    """
    images = qa_folder_info["images"]
    image_mapping = {}
    skipped = 0

    # Normalize skip list for case-insensitive matching
    skip_set = {s.lower() for s in (skip_images or set())}

    for img_path in images:
        original_name = img_path.name

        # Skip FIXED images
        if original_name.lower() in skip_set:
            skipped += 1
            continue

        dest_path = images_folder / original_name
        shutil.copy2(img_path, dest_path)
        image_mapping[original_name] = original_name

    if image_mapping or skipped:
        msg = f"    Copied {len(image_mapping)} images to Images/"
        if skipped:
            msg += f" (skipped {skipped} FIXED)"
        print(msg)

    return image_mapping


# =============================================================================
# STYLING HELPERS
# =============================================================================

# Standard border
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header styles
HEADER_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def style_header_cell(cell):
    """Apply standard header styling to a cell."""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def add_status_dropdown(ws, col: int, start_row: int = 2, end_row: int = None):
    """Add STATUS dropdown validation to a column."""
    if end_row is None:
        end_row = ws.max_row

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True
    )
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)

    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col))


def add_manager_dropdown(ws, col: int, start_row: int = 2, end_row: int = None):
    """Add manager status dropdown validation to a column."""
    if end_row is None:
        end_row = ws.max_row

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(MANAGER_STATUS_OPTIONS)}"',
        allow_blank=True
    )
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)

    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col))


# =============================================================================
# WORKSHEET SORTING
# =============================================================================

def sort_worksheet_az(ws, sort_column: int = 1):
    """
    Sort worksheet rows A-Z by specified column.

    Used for EN Item category to match tester's A-Z sorted files.
    Preserves header row (row 1), sorts data rows (row 2+).

    Args:
        ws: Worksheet to sort
        sort_column: Column index to sort by (1-based, default=1 for column A)
    """
    # CRITICAL: Clear any AutoFilter before sorting (testers may have filters applied)
    if ws.auto_filter.ref:
        ws.auto_filter.ref = None

    # Get all data rows (skip header)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append({
                'value': cell.value,
                'font': copy(cell.font) if cell.font else None,
                'fill': copy(cell.fill) if cell.fill else None,
                'border': copy(cell.border) if cell.border else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
                'hyperlink': cell.hyperlink.target if cell.hyperlink else None
            })
        # Store sort key (specified column value) and row data
        sort_key = ws.cell(row=row, column=sort_column).value
        sort_key = str(sort_key).lower() if sort_key else ""
        data_rows.append((sort_key, row_data))

    # Sort by specified column (A-Z)
    data_rows.sort(key=lambda x: x[0])

    # Write back sorted data
    for new_row_idx, (sort_key, row_data) in enumerate(data_rows, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            cell = ws.cell(row=new_row_idx, column=col_idx)
            cell.value = cell_data['value']
            if cell_data['font']:
                cell.font = cell_data['font']
            if cell_data['fill']:
                cell.fill = cell_data['fill']
            if cell_data['border']:
                cell.border = cell_data['border']
            if cell_data['alignment']:
                cell.alignment = cell_data['alignment']
            if cell_data['hyperlink']:
                cell.hyperlink = cell_data['hyperlink']
