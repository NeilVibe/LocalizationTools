"""
Compiler Module
===============
Main compilation orchestration: QAfolder -> Masterfolder_EN/CN + Tracker

Handles:
- Discovery of QA folders
- Language-based routing (EN/CN)
- Category processing with clustering
- Master file generation
- Progress tracker updates
"""

import shutil
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
from datetime import datetime

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

# =============================================================================
# GRANULAR DEBUG LOGGING
# =============================================================================

_COMPILER_LOG_FILE = Path(__file__).parent.parent / "COMPILER_DEBUG.log"
_COMPILER_LOG_ENABLED = False  # Set to True for verbose logging
_COMPILER_LOG_LINES = []  # Buffer for batch writing


def _compiler_log(msg: str, level: str = "INFO"):
    """Add message to compiler log buffer."""
    if not _COMPILER_LOG_ENABLED:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    _COMPILER_LOG_LINES.append(f"[{timestamp}] [{level}] {msg}")


def _compiler_log_flush(header: str = None):
    """Flush log buffer to file."""
    global _COMPILER_LOG_LINES
    if not _COMPILER_LOG_LINES:
        return
    try:
        mode = "a" if _COMPILER_LOG_FILE.exists() else "w"
        with open(_COMPILER_LOG_FILE, mode, encoding="utf-8") as f:
            if header:
                f.write(f"\n{'='*60}\n{header}\n{'='*60}\n")
            f.write("\n".join(_COMPILER_LOG_LINES) + "\n")
        _COMPILER_LOG_LINES = []
    except Exception as e:
        print(f"[COMPILER LOG ERROR] {e}")


def _compiler_log_clear():
    """Clear log file for fresh run."""
    global _COMPILER_LOG_LINES
    _COMPILER_LOG_LINES = []
    try:
        with open(_COMPILER_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(f"=== COMPILER DEBUG LOG === {datetime.now().isoformat()}\n\n")
    except:
        pass


# =============================================================================
# SCRIPT DEBUG LOGGING (for investigating Script manager status issue)
# =============================================================================

_SCRIPT_DEBUG_FILE = Path(__file__).parent.parent / "SCRIPT_DEBUG.log"
_SCRIPT_DEBUG_LINES = []


def _script_debug_log(msg: str):
    """Add message to Script debug log."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    _SCRIPT_DEBUG_LINES.append(f"[{timestamp}] {msg}")


def _script_debug_flush():
    """Flush Script debug log to file."""
    global _SCRIPT_DEBUG_LINES
    if not _SCRIPT_DEBUG_LINES:
        return
    try:
        mode = "a" if _SCRIPT_DEBUG_FILE.exists() else "w"
        with open(_SCRIPT_DEBUG_FILE, mode, encoding="utf-8") as f:
            f.write("\n".join(_SCRIPT_DEBUG_LINES) + "\n")
        _SCRIPT_DEBUG_LINES = []
    except Exception as e:
        print(f"[SCRIPT DEBUG ERROR] {e}")


def _script_debug_clear():
    """Clear Script debug log for fresh run."""
    global _SCRIPT_DEBUG_LINES
    _SCRIPT_DEBUG_LINES = []
    try:
        with open(_SCRIPT_DEBUG_FILE, "w", encoding="utf-8") as f:
            f.write(f"=== SCRIPT DEBUG LOG === {datetime.now().isoformat()}\n")
            f.write(f"Investigating why Script manager status is empty\n\n")
    except Exception:
        pass


from config import (
    CATEGORIES, CATEGORY_TO_MASTER, SCRIPT_TYPE_CATEGORIES,
    SCRIPT_COLS, FACE_TYPE_CATEGORIES,
    MASTER_FOLDER_EN, MASTER_FOLDER_CN,
    IMAGES_FOLDER_EN, IMAGES_FOLDER_CN,
    TRACKER_PATH,
    load_tester_mapping, ensure_folders_exist,
    get_target_master_category,
    WORKER_GROUPS, MAX_PARALLEL_WORKERS
)
from core.discovery import discover_qa_folders, group_folders_by_language
from core.excel_ops import (
    safe_load_workbook, ensure_master_folders,
    get_or_create_master,
    copy_images_with_unique_names,
    build_column_map,
    preload_worksheet_data,
    replicate_duplicate_row_data,
    beautify_master_sheet,
    reapply_manager_dropdowns,
    final_column_sweep,
)
from core.processing import (
    process_sheet, update_status_sheet,
    hide_empty_comment_rows, autofit_rows_with_wordwrap,
    count_words_english, count_chars_chinese
)
from core.matching import (
    build_master_index, clone_with_fresh_consumed, clear_master_index_cache
)


# Valid manager status values
VALID_MANAGER_STATUS = {"FIXED", "REPORTED", "CHECKING", "NON-ISSUE", "NON ISSUE"}

# Valid tester status values (rows with these are "done")
# NOTE: Script-type categories (Sequencer/Dialog) use "NON-ISSUE" (with hyphen)
# while other categories use "NO ISSUE" (with space). Accept both.
VALID_TESTER_STATUS = {"ISSUE", "NO ISSUE", "NON-ISSUE", "NON ISSUE", "BLOCKED", "KOREAN"}


# =============================================================================
# CONSOLIDATED MASTER DATA COLLECTION (Phase A optimization)
# =============================================================================
# Replaces separate functions that each opened all master files independently:
#   - collect_fixed_screenshots()  -> fixed_screenshots set
#   - collect_manager_stats_for_tracker() -> manager_stats dict
# Now: ONE pass per master file with read_only=True for both data structures.
# Manager status preservation is handled by System 1 (excel_ops.py) during rebuild.

def _extract_category_from_filename(filename: str) -> str:
    """Extract category from master filename (e.g., 'Master_Quest.xlsx' -> 'Quest').

    Returns empty string if not a valid Master_*.xlsx filename.
    """
    stem = Path(filename).stem
    if stem.startswith("Master_"):
        return stem[7:]  # strip 'Master_'
    return ""


# Categories to EXCLUDE from manager stats aggregation (Script-type + Face)
_MANAGER_STATS_EXCLUDED_CATEGORIES = (
    SCRIPT_TYPE_CATEGORIES | FACE_TYPE_CATEGORIES |
    {"Script", "Face"}  # Also exclude the master-level names
)


def _scan_master_file_streaming(
    master_path: Path,
    file_mtime: float,
    content_index: dict,
    fixed_screenshots: set,
    tester_mapping: dict,
    manager_dates: dict,
    skip_manager_stats: bool,
    log_lines: list,
):
    """Scan a single master file using read_only + iter_rows (streaming).

    Extracts fixed_screenshots and populates content_index for manager stats.
    Content dedup key: (StringID, Translation, Comment) per user.
    If same key seen in multiple files, latest mtime wins.

    Args:
        master_path: Path to the .xlsx file.
        file_mtime: File modification time (epoch).
        content_index: Shared dict {(username, stringid, translation, comment): (mtime, status)}.
        fixed_screenshots: Set to add FIXED screenshot names to.
        tester_mapping: Tester -> language mapping.
        manager_dates: Dict {(category, username) -> "YYYY-MM-DD"} (latest wins).
        skip_manager_stats: If True, only collect screenshots (for Script/Face types).
        log_lines: List to append debug log messages to.
    """
    category = _extract_category_from_filename(master_path.name)
    file_date = datetime.fromtimestamp(file_mtime).strftime("%Y-%m-%d")

    try:
        wb = safe_load_workbook(master_path, read_only=True, data_only=True)
    except Exception as e:
        log_lines.append(f"[ERROR] Failed to open {master_path}: {e}")
        print(f"\n  WARN: Error opening {master_path.name}: {e}")
        return

    try:
        log_lines.append(f"")
        log_lines.append(f"{'~'*60}")
        log_lines.append(f"MASTER FILE: {master_path.name} (category={category}, skip_stats={skip_manager_stats})")
        log_lines.append(f"{'~'*60}")
        log_lines.append(f"Sheets: {wb.sheetnames}")

        sheets_processed = 0
        total_rows_scanned = 0

        for sheet_name in wb.sheetnames:
            if sheet_name == "STATUS":
                continue

            ws = wb[sheet_name]

            # In read_only mode, max_row/max_column can be None for empty sheets
            if ws.max_row is None or ws.max_row < 2:
                continue
            if ws.max_column is None or ws.max_column < 1:
                continue

            # === HEADER SCAN via iter_rows (streaming, not random access) ===
            header_iter = ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column, values_only=True)
            header_tuple = next(header_iter, None)
            if not header_tuple:
                continue

            stringid_idx = None     # 0-based index
            translation_idx = None  # 0-based index
            comment_cols = {}       # username -> 0-based idx
            status_cols = {}        # username -> 0-based idx (STATUS_{User} = manager status)
            screenshot_cols = {}    # username -> 0-based idx

            for col_idx, header_val in enumerate(header_tuple):
                if not header_val:
                    continue
                header_str = str(header_val)
                header_upper = header_str.upper()

                if header_upper.startswith("STATUS_") and not header_upper.startswith("TESTER_STATUS_"):
                    status_cols[header_str[7:]] = col_idx
                elif header_upper.startswith("COMMENT_"):
                    comment_cols[header_str[8:]] = col_idx
                elif header_upper.startswith("SCREENSHOT_"):
                    screenshot_cols[header_str[11:]] = col_idx
                elif header_upper == "STRINGID":
                    stringid_idx = col_idx
                elif header_upper == "EVENTNAME" and stringid_idx is None:
                    stringid_idx = col_idx
                # Translation column detection: match common patterns
                elif translation_idx is None:
                    if (header_upper == "TEXT" or header_upper == "TRANSLATION"
                            or header_upper.startswith("TRANSLATION (")
                            or header_upper.startswith("ENGLISH")):
                        translation_idx = col_idx

            if not status_cols:
                continue

            log_lines.append(f"  [{sheet_name}] STATUS cols: {list(status_cols.keys())}, "
                             f"stringid={stringid_idx}, trans={translation_idx}")

            # === SINGLE-PASS ROW SCAN via iter_rows (streaming tuples) ===
            status_entries_count = 0
            row_count = 0

            for row_tuple in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
                row_count += 1

                # Pre-extract shared row values (stringid, translation)
                stringid_val = ""
                if stringid_idx is not None and stringid_idx < len(row_tuple):
                    raw = row_tuple[stringid_idx]
                    if raw:
                        stringid_val = str(raw).strip()

                translation_val = ""
                if translation_idx is not None and translation_idx < len(row_tuple):
                    raw = row_tuple[translation_idx]
                    if raw:
                        translation_val = str(raw).strip()

                for username, status_idx in status_cols.items():
                    status_value = row_tuple[status_idx] if status_idx < len(row_tuple) else None
                    status_str = str(status_value).strip() if status_value else ""
                    status_upper = status_str.upper()

                    has_status = status_upper in VALID_MANAGER_STATUS

                    # --- DATA 1: fixed_screenshots (always collected) ---
                    if status_upper == "FIXED":
                        sc_idx = screenshot_cols.get(username)
                        if sc_idx is not None and sc_idx < len(row_tuple):
                            sc_val = row_tuple[sc_idx]
                            if sc_val and str(sc_val).strip():
                                fixed_screenshots.add(str(sc_val).strip())

                    # --- DATA 2: manager_stats via content_index (skipped for Script/Face) ---
                    if skip_manager_stats or not has_status:
                        continue

                    status_entries_count += 1

                    # Extract comment for content dedup key
                    c_idx = comment_cols.get(username)
                    comment_val = ""
                    if c_idx is not None and c_idx < len(row_tuple):
                        raw_c = row_tuple[c_idx]
                        if raw_c:
                            comment_val = str(raw_c).strip()

                    # Content dedup key: (username, stringid, translation, comment)
                    content_key = (username, stringid_val, translation_val, comment_val)

                    # Latest mtime wins
                    existing = content_index.get(content_key)
                    if existing is None or file_mtime > existing[0]:
                        content_index[content_key] = (file_mtime, status_upper, category, tester_mapping.get(username, "EN"))

            log_lines.append(f"    {status_entries_count} status entries, {row_count} rows")
            sheets_processed += 1
            total_rows_scanned += row_count

            # manager_dates: keep latest mtime per (category, username)
            if not skip_manager_stats:
                for username in status_cols.keys():
                    existing_date = manager_dates.get((category, username))
                    if existing_date is None or file_date > existing_date:
                        manager_dates[(category, username)] = file_date

        print(f" {sheets_processed} sheets, {total_rows_scanned} rows")
    finally:
        wb.close()


def collect_all_master_data(tester_mapping: Dict = None, log_callback=None):
    """
    Single-pass collection of master file data for compilation.

    NEW STRATEGY (full rewrite):
    - Glob ALL .xlsx files in each master folder (no category-based discovery)
    - Content dedup key: (StringID, Translation, Comment) per user — identical = same entry
    - Filter out Script types (Sequencer, Dialog, Face) from manager stats
    - File mtime as date — latest mtime wins for duplicate content
    - Per-user stats: count FIXED/REPORTED/CHECKING/NONISSUE from STATUS_{username} columns

    Opens each master file ONCE with read_only=True, extracting:
    1. fixed_screenshots (EN/CN) - for skipping FIXED images during copy
    2. manager_stats - for tracker (FIXED/REPORTED/CHECKING/NON-ISSUE counts)

    Args:
        tester_mapping: Dict mapping tester names to language codes (EN/CN).
                        If None, loaded from file.
        log_callback: Optional callback(message, tag) for GUI logging.

    Returns:
        Tuple of (fixed_screenshots_en, fixed_screenshots_cn, manager_stats, manager_dates)
    """
    def _log(msg, tag='info'):
        if log_callback:
            log_callback(msg, tag)

    if tester_mapping is None:
        tester_mapping = load_tester_mapping()

    fixed_screenshots_en = set()
    fixed_screenshots_cn = set()
    # Content index for dedup: {(username, stringid, translation, comment): (mtime, status_upper, category, lang)}
    content_index = {}
    manager_dates = {}  # (category, username) -> "YYYY-MM-DD" from master file mtime

    # LOG FILE for manager stats debug
    log_path = Path(__file__).parent.parent / "MANAGER_STATS_DEBUG.log"
    L = []

    L.append(f"{'='*80}")
    L.append(f"CONSOLIDATED MASTER DATA COLLECTION (glob-all strategy)")
    L.append(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L.append(f"{'='*80}")
    L.append(f"")
    L.append(f"[CONFIG] MASTER_FOLDER_EN: {MASTER_FOLDER_EN}")
    L.append(f"[CONFIG] MASTER_FOLDER_CN: {MASTER_FOLDER_CN}")
    L.append(f"[CONFIG] Tester mapping: {len(tester_mapping)} entries")
    L.append(f"[CONFIG] Excluded categories for stats: {_MANAGER_STATS_EXCLUDED_CATEGORIES}")
    L.append(f"")

    for master_folder in [MASTER_FOLDER_EN, MASTER_FOLDER_CN]:
        is_en = "EN" in str(master_folder)
        folder_label = "EN" if is_en else "CN"
        fixed_screenshots = fixed_screenshots_en if is_en else fixed_screenshots_cn

        print(f"  [{folder_label}] Scanning master folder: {master_folder}")
        L.append(f"{'='*80}")
        L.append(f"PROCESSING FOLDER: {master_folder} [{folder_label}]")
        L.append(f"{'='*80}")

        if not master_folder.exists():
            print(f"  [{folder_label}] Folder does not exist - skipping")
            L.append(f"Folder does not exist - skipping")
            continue

        # Glob ALL .xlsx files — no category filtering at discovery time
        all_xlsx = sorted([
            f for f in master_folder.glob("*.xlsx")
            if not f.name.startswith("~")
        ])

        L.append(f"Found {len(all_xlsx)} .xlsx files")
        print(f"  [{folder_label}] Found {len(all_xlsx)} .xlsx files")

        for master_path in all_xlsx:
            category = _extract_category_from_filename(master_path.name)
            # Determine if this category should be excluded from manager stats
            skip_stats = category.lower() in {c.lower() for c in _MANAGER_STATS_EXCLUDED_CATEGORIES}

            file_mtime = master_path.stat().st_mtime
            print(f"    Opening: {master_path.name}...", end="", flush=True)

            _scan_master_file_streaming(
                master_path=master_path,
                file_mtime=file_mtime,
                content_index=content_index,
                fixed_screenshots=fixed_screenshots,
                tester_mapping=tester_mapping,
                manager_dates=manager_dates,
                skip_manager_stats=skip_stats,
                log_lines=L,
            )

    # Phase 2: Aggregate counts from content_index
    L.append(f"")
    L.append(f"{'='*80}")
    L.append(f"CONTENT INDEX AGGREGATION")
    L.append(f"{'='*80}")
    L.append(f"Unique content entries: {len(content_index)}")

    manager_stats_agg = defaultdict(lambda: defaultdict(
        lambda: {"fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "lang": "EN"}
    ))

    for (username, _sid, _trans, _comment), (mtime, status_upper, category, lang) in content_index.items():
        if status_upper == "FIXED":
            manager_stats_agg[category][username]["fixed"] += 1
        elif status_upper == "REPORTED":
            manager_stats_agg[category][username]["reported"] += 1
        elif status_upper == "CHECKING":
            manager_stats_agg[category][username]["checking"] += 1
        elif status_upper in ("NON-ISSUE", "NON ISSUE"):
            manager_stats_agg[category][username]["nonissue"] += 1
        manager_stats_agg[category][username]["lang"] = lang

    # Convert to regular dicts
    manager_stats_result = {}
    for cat, users_dd in manager_stats_agg.items():
        manager_stats_result[cat] = {}
        for user, stats_dd in users_dd.items():
            manager_stats_result[cat][user] = dict(stats_dd)

    # Log final summary
    L.append(f"")
    L.append(f"{'='*80}")
    L.append(f"FINAL SUMMARY")
    L.append(f"{'='*80}")
    L.append(f"fixed_screenshots_en: {len(fixed_screenshots_en)}")
    L.append(f"fixed_screenshots_cn: {len(fixed_screenshots_cn)}")
    L.append(f"manager_stats categories: {list(manager_stats_result.keys())}")
    for cat, users in manager_stats_result.items():
        L.append(f"  {cat}:")
        for user, stats in users.items():
            L.append(f"    {user}: F={stats['fixed']} R={stats['reported']} C={stats['checking']} N={stats['nonissue']} lang={stats['lang']}")
    L.append(f"{'='*80}")

    # Write log file
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(L))
        print(f"[LOG] Master data collection ({len(L)} lines): {log_path}")
    except Exception as e:
        print(f"[LOG ERROR] {e}")

    return (fixed_screenshots_en, fixed_screenshots_cn,
            manager_stats_result, manager_dates)


# =============================================================================
# SCRIPT-TYPE PREPROCESSING (Sequencer/Dialog optimization)
# =============================================================================

def preprocess_script_category(
    qa_folders: List[Dict],
    is_english: bool = True
) -> Dict:
    """
    Preprocess Script-type category files to build global universe of rows WITH status.

    Phase B enhancement: Also stores full_row data and sheet headers so that
    build_master_from_universe() can create the master workbook directly in memory
    without needing create_filtered_script_template().

    Args:
        qa_folders: List of folder dicts from discovery (all for this category)
        is_english: Whether these are English files

    Returns:
        Dict with:
            - "rows": {(eventname, text): row_data} - all unique rows with status
            - "row_count": int - total rows with status
            - "source_files": int - number of files scanned
            - "errors": List of error messages encountered
            - "headers": {sheet_name: [header1, header2, ...]} - per-sheet headers
            - "num_columns": {sheet_name: int} - column count per sheet
    """
    import traceback
    # SCRIPT_COLS already imported from config at module level

    universe = {}  # (eventname, text) -> row_data
    headers = {}   # sheet_name -> [header values]
    num_columns = {}  # sheet_name -> int
    source_files = 0
    errors = []  # Track errors for debugging

    print(f"  [PREPROCESS] Scanning {len(qa_folders)} QA files for rows with STATUS...")

    if not qa_folders:
        print(f"    [WARN] No QA folders provided")
        return {"rows": {}, "row_count": 0, "source_files": 0, "errors": ["No QA folders provided"],
                "headers": {}, "num_columns": {}}

    for qf_idx, qf in enumerate(qa_folders, 1):
        xlsx_path = qf.get("xlsx_path")
        username = qf.get("username", "unknown")

        if xlsx_path is None:
            errors.append(f"Missing xlsx_path for user {username}")
            continue

        source_files += 1

        try:
            # Validate file exists
            if not xlsx_path.exists():
                errors.append(f"File not found: {xlsx_path}")
                print(f"    [WARN] File not found: {xlsx_path}")
                continue

            # read_only=True is 3-5x faster for large files
            print(f"    [{qf_idx}/{len(qa_folders)}] {username}...", end="", flush=True)
            wb = safe_load_workbook(xlsx_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                if sheet_name == "STATUS":
                    continue

                try:
                    ws = wb[sheet_name]

                    # Check for empty sheet
                    if ws.max_row is None or ws.max_row < 2:
                        continue
                    if ws.max_column is None or ws.max_column < 1:
                        continue

                    # build_column_map uses iter_rows internally (safe for read_only mode)
                    col_map = build_column_map(ws)

                    status_col = col_map.get("STATUS")
                    # Translation: try "Text" first, then "Translation"
                    text_col = col_map.get("TEXT")
                    if text_col is None:
                        text_col = col_map.get("TRANSLATION")
                    # Unique ID: try "EventName" first, then "STRINGID"
                    eventname_col = col_map.get("EVENTNAME")
                    if eventname_col is None:
                        eventname_col = col_map.get("STRINGID")
                    # Comment: try "MEMO" first, then "COMMENT"
                    memo_col = col_map.get("MEMO")
                    if memo_col is None:
                        memo_col = col_map.get("COMMENT")

                    if not status_col or not text_col or not eventname_col:
                        # Not a script sheet - skip silently
                        continue

                    total_cols = ws.max_column

                    # Collect sheet structure from first file encountered per sheet
                    # (Phase B: needed for build_master_from_universe)
                    # Use iter_rows for streaming header read (not ws.cell!)
                    if sheet_name not in headers:
                        header_iter = ws.iter_rows(min_row=1, max_row=1, max_col=total_cols, values_only=True)
                        header_tuple = next(header_iter, None)
                        headers[sheet_name] = list(header_tuple) if header_tuple else []
                        num_columns[sheet_name] = total_cols

                    # Scan rows using iter_rows for batch reading (Phase C3)
                    status_idx = status_col - 1
                    eventname_idx = eventname_col - 1
                    text_idx = text_col - 1

                    for row_idx, row_tuple in enumerate(
                        ws.iter_rows(min_row=2, max_col=total_cols, values_only=True),
                        start=2
                    ):
                        try:
                            # Pad row_tuple if shorter than total_cols (can happen with sparse data)
                            if len(row_tuple) <= status_idx:
                                continue

                            status_val = row_tuple[status_idx]
                            if not status_val:
                                continue

                            status_str = str(status_val).strip()
                            if not status_str:
                                continue

                            # This row has a status value - add to universe
                            eventname = str(row_tuple[eventname_idx] or "").strip() if eventname_idx < len(row_tuple) else ""
                            text = str(row_tuple[text_idx] or "").strip() if text_idx < len(row_tuple) else ""

                            if not eventname and not text:
                                continue

                            key = (eventname, text)

                            # Store row data with full_row (Phase B: for direct master building)
                            # Convert tuple to list for storage
                            full_row = list(row_tuple)

                            if key not in universe:
                                universe[key] = {
                                    "eventname": eventname,
                                    "text": text,
                                    "sheet": sheet_name,
                                    "sources": [(username, xlsx_path, sheet_name, row_idx)],
                                    "full_row": full_row,
                                }
                            else:
                                # Track additional sources for debugging
                                universe[key]["sources"].append((username, xlsx_path, sheet_name, row_idx))

                        except Exception as row_err:
                            err_msg = f"Error reading row {row_idx} in {xlsx_path.name}/{sheet_name}: {row_err}"
                            errors.append(err_msg)

                except Exception as sheet_err:
                    err_msg = f"Error processing sheet '{sheet_name}' in {xlsx_path.name}: {sheet_err}"
                    errors.append(err_msg)
                    print(f"    [WARN] {err_msg}")

            file_rows = sum(1 for _ in universe.values() if any(s[0] == username for s in _.get("sources", [])))
            print(f" {len(wb.sheetnames)} sheets")
            wb.close()

        except Exception as e:
            err_msg = f"Error preprocessing {xlsx_path.name}: {type(e).__name__}: {e}"
            errors.append(err_msg)
            print(f"\n    [ERROR] {err_msg}")
            traceback.print_exc()

    print(f"  [PREPROCESS] Found {len(universe)} unique rows with STATUS from {source_files} files")
    if errors:
        print(f"  [PREPROCESS] Encountered {len(errors)} error(s) during preprocessing")

    return {
        "rows": universe,
        "row_count": len(universe),
        "source_files": source_files,
        "errors": errors,
        "headers": headers,
        "num_columns": num_columns,
    }


def build_master_from_universe(
    category: str,
    universe: Dict,
    master_folder: Path
):
    """
    Build master workbook directly from universe data (Phase B optimization).

    Creates the master workbook in memory from preprocessed data, eliminating
    the create_filtered_script_template() bottleneck which:
    - Opened template in FULL mode
    - Scanned ALL 10K rows to build index
    - Opened source files AGAIN
    - Wrote temp file to disk
    - Temp file immediately re-loaded by get_or_create_master()

    This function replaces that entire pipeline with direct in-memory construction.

    Args:
        category: Category name (e.g., "Sequencer")
        universe: Dict from preprocess_script_category() with "rows", "headers", etc.
        master_folder: Target Master folder (EN or CN)

    Returns:
        Tuple of (Workbook, master_path) - same interface as get_or_create_master()
    """
    from openpyxl import Workbook as NewWorkbook

    target_category = get_target_master_category(category)
    master_path = master_folder / f"Master_{target_category}.xlsx"

    # Backup old master if it exists (crash-safe rebuild)
    if master_path.exists():
        backup_path = master_path.with_suffix(".xlsx.bak")
        print(f"  Backing up old master: {master_path.name} -> {backup_path.name}")
        shutil.move(str(master_path), str(backup_path))

    wb = NewWorkbook()
    # Remove default "Sheet"
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    rows_data = universe.get("rows", {})
    sheet_headers = universe.get("headers", {})
    sheet_num_cols = universe.get("num_columns", {})

    # Columns to DELETE from master (same as get_or_create_master):
    # STATUS, COMMENT, SCREENSHOT are tester columns replaced by user-specific columns
    # NOTE: STRINGID is a DATA column used for matching -- it must be PRESERVED
    COLS_TO_SKIP = {"STATUS", "COMMENT", "SCREENSHOT"}

    # Group rows by sheet name
    rows_by_sheet = {}
    for key, data in rows_data.items():
        sheet = data.get("sheet", "Script")
        if sheet not in rows_by_sheet:
            rows_by_sheet[sheet] = []
        rows_by_sheet[sheet].append(data)

    # Build each sheet
    total_rows_written = 0
    for sheet_name in sheet_headers:
        raw_headers = sheet_headers[sheet_name]
        n_cols = sheet_num_cols.get(sheet_name, len(raw_headers))

        # Determine which column indices to KEEP (skip STATUS/COMMENT/SCREENSHOT/STRINGID)
        cols_to_keep = []  # list of (original_0based_idx, header_value)
        for idx, header_val in enumerate(raw_headers):
            header_upper = str(header_val).strip().upper() if header_val else ""
            if header_upper not in COLS_TO_SKIP:
                cols_to_keep.append((idx, header_val))

        if not cols_to_keep:
            continue

        ws = wb.create_sheet(sheet_name)

        # Write filtered headers
        for new_col, (orig_idx, header_val) in enumerate(cols_to_keep, 1):
            ws.cell(row=1, column=new_col, value=header_val)

        # Write data rows for this sheet
        sheet_rows = rows_by_sheet.get(sheet_name, [])
        dst_row = 2
        for row_data in sheet_rows:
            full_row = row_data.get("full_row")
            if full_row:
                for new_col, (orig_idx, _) in enumerate(cols_to_keep, 1):
                    if orig_idx < len(full_row):
                        ws.cell(row=dst_row, column=new_col, value=full_row[orig_idx])
                dst_row += 1
                total_rows_written += 1

    print(f"  [BUILD] Created master with {total_rows_written} rows across {len(rows_by_sheet)} sheets (in-memory)")

    # Clear AutoFilter on all sheets (safety measure)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.auto_filter.ref:
            ws.auto_filter.ref = None

    return wb, master_path


def build_prefiltered_rows(universe: Dict, xlsx_path: Path, sheet_name: str, username: str) -> Optional[List[int]]:
    """
    Extract prefiltered row numbers from universe for a specific user/file/sheet.

    Phase C2 optimization: Avoids re-scanning STATUS column in process_sheet()
    by providing row numbers already known from preprocessing.

    Args:
        universe: Dict from preprocess_script_category()
        xlsx_path: Path to the QA file
        sheet_name: Sheet name to filter for
        username: Username to filter for

    Returns:
        Sorted list of row numbers, or None if not available
    """
    rows_data = universe.get("rows", {})
    if not rows_data:
        return None

    row_numbers = []
    xlsx_str = str(xlsx_path)

    for key, data in rows_data.items():
        for src_user, src_path, src_sheet, src_row in data.get("sources", []):
            if src_user == username and str(src_path) == xlsx_str and src_sheet == sheet_name:
                row_numbers.append(src_row)

    if not row_numbers:
        return None

    return sorted(set(row_numbers))


# =============================================================================
# CATEGORY PROCESSING
# =============================================================================

def process_category(
    category: str,
    qa_folders: List[Dict],
    master_folder: Path,
    images_folder: Path,
    lang_label: str,
    rebuild: bool = True,
    fixed_screenshots: set = None,
    accumulated_users: set = None,
    accumulated_stats: Dict = None,
    deferred_save: bool = False,
    log_callback=None
) -> Tuple[List[Dict], set, Dict, Optional[object], Optional[Path]]:
    """
    Process all QA folders for one category.

    Args:
        category: Category name (Quest, Knowledge, etc.)
        qa_folders: List of folder dicts from discovery
        master_folder: Target Master folder (EN or CN)
        images_folder: Target Images folder
        lang_label: "EN" or "CN"
        rebuild: If True, rebuild master. If False, append (for clustering)
        fixed_screenshots: Set of screenshot filenames to skip (FIXED optimization)
        accumulated_users: Set of users accumulated across categories sharing same master
        accumulated_stats: Dict of user stats accumulated across categories sharing same master
        deferred_save: If True, skip autofit/save and return workbook for caller to finalize

    Returns:
        Tuple of (daily_entries, accumulated_users, accumulated_stats, master_wb, master_path)
        - daily_entries: List of daily_entry dicts for tracker
        - accumulated_users: Updated set of all users for this master
        - accumulated_stats: Updated dict of user stats for this master
        - master_wb: The master workbook (None if deferred_save=False, i.e., already saved)
        - master_path: Path to master file
    """
    def _log(msg, tag='info'):
        if log_callback:
            log_callback(msg, tag)

    if fixed_screenshots is None:
        fixed_screenshots = set()
    if accumulated_users is None:
        accumulated_users = set()
    if accumulated_stats is None:
        accumulated_stats = defaultdict(lambda: {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0})

    target_master = get_target_master_category(category)
    cluster_info = f" -> Master_{target_master}" if target_master != category else ""

    # SCRIPT DEBUG: Log when processing Script-type categories
    is_script_category = category.lower() in SCRIPT_TYPE_CATEGORIES
    if is_script_category:
        _script_debug_log(f"")
        _script_debug_log(f"{'='*60}")
        _script_debug_log(f"[PROCESS_CATEGORY] {category} [{lang_label}]")
        _script_debug_log(f"{'='*60}")
        _script_debug_log(f"  target_master: {target_master}")
        _script_debug_log(f"  qa_folders count: {len(qa_folders)}")
        for qf in qa_folders:
            _script_debug_log(f"    - {qf.get('username')}: {qf.get('xlsx_path')}")
        _script_debug_flush()

    _log(f"Processing: {category} [{lang_label}] ({len(qa_folders)} folders){cluster_info}", 'header')

    daily_entries = []

    # Determine if EN or CN for word counting
    is_english = (lang_label == "EN")

    # Get or create master — pick template with MOST DATA ROWS across all sheets.
    # Using most-recent-mtime was wrong: if the newest file happened to be empty or
    # have fewer sheets, the master inherited that sparse structure and all other
    # testers' rows became unmatched → silently dropped → master appeared empty.
    def _count_data_rows(xlsx_path):
        """Count total data rows (excluding headers) across all sheets."""
        try:
            wb = safe_load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            _log(f"  WARN: Cannot count rows in {xlsx_path.name}: {type(e).__name__}: {e}", 'warning')
            return 0
        total = sum(
            (ws.max_row or 0) - 1
            for sn in wb.sheetnames if sn != "STATUS"
            for ws in [wb[sn]]
            if ws.max_row and ws.max_row > 1
        )
        wb.close()
        return total

    # Pick template with most data rows; break ties by mtime (newest wins)
    best_rows, _, best_folder = max(
        ((_count_data_rows(qf["xlsx_path"]), qf["xlsx_path"].stat().st_mtime, qf)
         for qf in qa_folders),
        key=lambda x: (x[0], x[1])
    )
    template_xlsx = best_folder["xlsx_path"]
    template_user = best_folder["username"]
    _log(f"  Template: {template_user} ({best_rows} rows, most complete)")
    if best_rows == 0:
        _log(f"  ERROR: All QA files returned 0 data rows — cannot select template!", 'error')
        return daily_entries, accumulated_users, dict(accumulated_stats) if accumulated_stats else {}, None, None

    # Use unified template-based master creation for ALL categories
    # This preserves ALL template rows (including REPORTED issues that may not have
    # STATUS in current QA files). The old Script-specific "universe" method that
    # rebuilt from scratch was causing data loss when testers cleared STATUS.
    #
    # NOTE: Script-type categories (Sequencer/Dialog) previously used a separate
    # preprocess_script_category() + build_master_from_universe() pipeline that
    # only kept rows WITH STATUS. This optimization caused REPORTED rows to be lost
    # when testers removed STATUS from their QA files. Now all categories use the
    # unified method which preserves data integrity while still being fast.
    master_wb, master_path = get_or_create_master(
        category, master_folder, template_xlsx,
        rebuild=rebuild,
        is_english=is_english
    )

    if master_wb is None:
        return daily_entries, accumulated_users, dict(accumulated_stats) if accumulated_stats else {}, None, None

    # Track stats - use accumulated data if provided (for clustered categories)
    all_users = accumulated_users
    user_stats = accumulated_stats
    user_wordcount = defaultdict(int)  # username -> word count (EN) or char count (CN)
    user_file_dates = {}  # username -> file modification date
    total_images = 0
    total_screenshots = 0

    # Get translation column for word counting (is_english already defined above)
    # All categories use header-name detection — no position fallback
    is_script_category = category.lower() in SCRIPT_TYPE_CATEGORIES

    # ==========================================================================
    # PERFORMANCE OPTIMIZATION: Build master indexes ONCE per sheet
    # ==========================================================================
    # Instead of rebuilding the index for each user (O(users × rows)),
    # build once and clone with fresh consumed set (O(rows) + O(users))
    # This gives 10x speedup for 10 users on same master sheet.
    master_indexes = {}  # sheet_name -> master_index
    for sheet_name in master_wb.sheetnames:
        if sheet_name == "STATUS":
            continue
        master_ws = master_wb[sheet_name]
        if master_ws.max_row and master_ws.max_row > 1:
            master_indexes[sheet_name] = build_master_index(master_ws, category, is_english)

    # Process each QA folder
    for qf in qa_folders:
        try:
            username = qf["username"]
            xlsx_path = qf["xlsx_path"]
            all_users.add(username)

            # Ensure user exists in user_stats (may be a regular dict from previous category)
            # This handles the case where accumulated_stats was converted from defaultdict to dict
            if username not in user_stats:
                user_stats[username] = {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0}

            # Get file modification date for tracker
            file_mod_time = __import__('datetime').datetime.fromtimestamp(xlsx_path.stat().st_mtime)
            user_file_dates[username] = file_mod_time.strftime("%Y-%m-%d")

            _log(f"  {username}")

            # Copy images FIRST to get mapping for screenshot links
            # Script-type categories (Sequencer/Dialog) have NO screenshots - skip image copying
            if category.lower() in SCRIPT_TYPE_CATEGORIES:
                image_mapping = {}
            else:
                image_mapping = copy_images_with_unique_names(qf, images_folder, skip_images=fixed_screenshots)
                total_images += len(image_mapping)

            # Load QA workbook
            # Script categories (Sequencer/Dialog) have NO screenshots, so ws.cell() is never
            # called on QA workbook in the hot loop → safe to use read_only for 3-5x faster load.
            # Non-Script categories need standard mode for screenshot hyperlink reads (ws.cell).
            if category.lower() in SCRIPT_TYPE_CATEGORIES:
                qa_wb = safe_load_workbook(xlsx_path, read_only=True, data_only=True)
            else:
                qa_wb = safe_load_workbook(xlsx_path)

            # Process each sheet
            for sheet_name in qa_wb.sheetnames:
                if sheet_name == "STATUS":
                    continue

                qa_ws = qa_wb[sheet_name]

                # Check if sheet exists in master
                if sheet_name not in master_wb.sheetnames:
                    _log(f"    WARN: Sheet '{sheet_name}' not in master", 'warning')
                    continue

                master_ws = master_wb[sheet_name]

                # Process the sheet (creates user columns internally)
                # NOTE: prefiltered_rows optimization removed - all categories now scan
                # all rows for STATUS, which is slightly slower but preserves data integrity
                # Uses content-based matching for robust row matching
                qa_rows = qa_ws.max_row - 1 if qa_ws.max_row and qa_ws.max_row > 1 else 0

                # OPTIMIZATION: Use pre-built index with fresh consumed set for each user
                # This avoids rebuilding the index for each user (10x speedup for 10 users)
                user_index = None
                if sheet_name in master_indexes:
                    user_index = clone_with_fresh_consumed(master_indexes[sheet_name])

                result = process_sheet(
                    master_ws, qa_ws, username, category,
                    is_english=is_english,
                    image_mapping=image_mapping,
                    xlsx_path=xlsx_path,
                    master_index=user_index
                )

                # Accumulate stats from result["stats"]
                # Defensive check: ensure user exists in user_stats (fixes KeyError for clustered categories)
                if username not in user_stats:
                    user_stats[username] = {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0}
                stats = result.get("stats", {})
                for key in ["issue", "no_issue", "blocked", "korean"]:
                    user_stats[username][key] += stats.get(key, 0)
                user_stats[username]["total"] += stats.get("total", 0)
                total_screenshots += result.get("screenshots", 0)

                # Log unmatched/appended rows
                match_stats = result.get("match_stats", {})
                appended = match_stats.get("appended", 0)
                unmatched = match_stats.get("unmatched", 0)
                if unmatched > 0:
                    _log(f"    INFO: {sheet_name}: {unmatched} rows not in template, {appended} appended to master")

                # Merge appended rows back to the source index so the NEXT user
                # can find them via clone_with_fresh_consumed().  Without this,
                # N testers with the same unmatched row would each append a duplicate.
                if appended > 0 and user_index and sheet_name in master_indexes:
                    src = master_indexes[sheet_name]
                    for key, row in user_index["primary"].items():
                        if key not in src["primary"]:
                            src["primary"][key] = row
                    for index_key in ("all_primary", "fallback"):
                        for key, rows in user_index[index_key].items():
                            existing = src[index_key].get(key, [])
                            for r in rows:
                                if r not in existing:
                                    src[index_key][key].append(r)

                # Count words (EN) or characters (CN) from translation column
                # ONLY count rows where STATUS is filled (DONE rows)
                # FAST: Reuse preloaded data from process_sheet (avoids redundant 3rd preload)
                wc_col_idx, wc_data_rows = result["_preloaded"]

                # Get column indices from preloaded header map
                wc_status_idx = wc_col_idx.get("STATUS")

                # All categories: find translation column by header name
                from core.matching import find_translation_col_in_headers
                wc_trans_idx = find_translation_col_in_headers(wc_col_idx, is_english)
                if wc_trans_idx is None:
                    # Skip word counting for this sheet - no translation column found
                    continue

                wc_label = "words" if is_english else "chars"
                wc_before = user_wordcount[username]

                # FAST: Iterate through preloaded tuples (no ws.cell() calls!)
                for row_tuple in wc_data_rows:
                    if wc_status_idx is not None:
                        status_val = row_tuple[wc_status_idx] if wc_status_idx < len(row_tuple) else None
                        # Accept all variants: "NO ISSUE", "NON-ISSUE", "NON ISSUE"
                        if not status_val or str(status_val).strip().upper() not in ["ISSUE", "NO ISSUE", "NON-ISSUE", "NON ISSUE", "BLOCKED", "KOREAN"]:
                            continue  # Skip rows not marked as done

                    cell_value = row_tuple[wc_trans_idx] if wc_trans_idx < len(row_tuple) else None
                    if is_english:
                        user_wordcount[username] += count_words_english(cell_value)
                    else:
                        user_wordcount[username] += count_chars_chinese(cell_value)
                wc_added = user_wordcount[username] - wc_before

            qa_wb.close()

            # Per-user stats summary
            us = user_stats[username]
            done = us['issue'] + us['no_issue'] + us['blocked'] + us['korean']
            _log(f"    {done}/{us['total']} done, {us['issue']} issues")
        except Exception as e:
            _log(f"  ERROR: {qf.get('username', '?')}: {e}", 'error')
            # Close workbook if it was opened before the error
            try:
                qa_wb.close()
            except (NameError, Exception):
                pass

    if total_images > 0:
        _log(f"  Images: {total_images} copied")

    # Create daily entry for tracker
    for username in all_users:
        # Defensive check: ensure user exists (fixes potential KeyError for accumulated users)
        if username not in user_stats:
            user_stats[username] = {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0}
        stats = user_stats[username]
        file_mod_date = user_file_dates.get(username, __import__('datetime').datetime.now().strftime("%Y-%m-%d"))
        entry = {
            "date": file_mod_date,
            "user": username,
            "category": category,
            "lang": lang_label,
            "total_rows": stats["total"],
            "done": stats["issue"] + stats["no_issue"] + stats["blocked"] + stats["korean"],
            "issues": stats["issue"],
            "no_issue": stats["no_issue"],
            "blocked": stats["blocked"],
            "korean": stats["korean"],
            "word_count": user_wordcount[username],
        }
        daily_entries.append(entry)

        # SCRIPT DEBUG: Detailed logging for Script-type categories
        if is_script_category:
            _script_debug_log(f"")
            _script_debug_log(f"[DAILY_ENTRY CREATED] {category}/{username}")
            _script_debug_log(f"  date: {file_mod_date}")
            _script_debug_log(f"  category: {category} (lookup will use: {target_master})")
            _script_debug_log(f"  total_rows: {stats['total']}")
            _script_debug_log(f"  done: {entry['done']} (issue={stats['issue']} + no_issue={stats['no_issue']} + blocked={stats['blocked']} + korean={stats['korean']})")
            _script_debug_log(f"  issues: {stats['issue']} <-- THIS IS THE TESTER ISSUE COUNT")
            _script_debug_log(f"  word_count: {user_wordcount[username]}")
            _script_debug_flush()

    # NOTE: STATUS sheet update is DEFERRED - caller will call update_status_sheet()
    # after all categories sharing this master have been processed. This prevents
    # clustered categories (e.g., Sequencer + Dialog -> Master_Script) from
    # overwriting each other's users in the STATUS sheet.

    if deferred_save:
        # DEFERRED SAVE: Return workbook for caller to finalize (autofit + save)
        # This enables running autofit ONCE per master after ALL categories processed
        return daily_entries, all_users, dict(user_stats), master_wb, master_path
    else:
        # Post-process: replicate data across duplicate rows
        replicate_duplicate_row_data(master_wb, category, is_english)

        # IMMEDIATE SAVE: Apply word wrap and autofit FIRST (before hiding)
        # This way all columns get proper widths, even if hidden later
        # Bonus: if user unhides a column in Excel, it already looks good
        print(f"\n  Formatting: autofit columns and row heights...")
        sheet_data_cache = autofit_rows_with_wordwrap(master_wb)

        # THEN hide empty rows/sheets/columns (reuses preloaded data)
        print(f"  Optimizing: hiding empty rows/sheets/columns...")
        hidden_rows, hidden_sheets, hidden_columns = hide_empty_comment_rows(master_wb, preloaded_sheets=sheet_data_cache)

        # Save master
        print(f"  Saving master file...", end="", flush=True)
        master_wb.save(master_path)
        print(f" done")
        # Clean up .bak backup now that new master is safely saved
        # (unless high orphan rate flagged it for preservation)
        backup_path = master_path.with_suffix(".xlsx.bak")
        if backup_path.exists():
            if getattr(master_wb, '_qacompiler_keep_backup', False):
                print(f"  ⚠ Backup PRESERVED (high orphan rate): {backup_path.name}")
            else:
                backup_path.unlink()
                print(f"  Cleaned up backup: {backup_path.name}")
        print(f"\n  Saved: {master_path}")
        if hidden_sheets:
            print(f"  Hidden sheets (no comments): {', '.join(hidden_sheets)}")
        if hidden_columns > 0:
            print(f"  Hidden: {hidden_columns} empty user columns (unhide in Excel if needed)")
        if hidden_rows > 0:
            print(f"  Hidden: {hidden_rows} rows with no comments (unhide in Excel if needed)")
        if total_images > 0:
            print(f"  Images: {total_images} copied to Images/, {total_screenshots} hyperlinks updated")

        return daily_entries, all_users, dict(user_stats), None, master_path


# =============================================================================
# MAIN COMPILER
# =============================================================================

def run_compiler(log_callback=None, progress_callback=None):
    """
    Main compiler entry point.

    Discovers QA folders, routes by language, processes categories,
    and updates the progress tracker.

    Args:
        log_callback: Optional callable(msg, tag) for GUI logging.
                      Tags: 'header', 'info', 'success', 'warning', 'error'
        progress_callback: Optional callable(pct) for GUI progress bar (0-100).
    """
    # GUI log wrappers
    def _log(msg, tag='info'):
        if log_callback:
            log_callback(msg, tag)

    def _progress(pct):
        if progress_callback:
            progress_callback(pct)

    # Start timing for final report
    compilation_start_time = time.time()

    # Clear Script debug log for fresh run
    _script_debug_clear()
    _script_debug_log("=== STARTING FULL COMPILATION ===")
    _script_debug_flush()

    # Clear master index cache for fresh run
    clear_master_index_cache()

    _log("=== Build Master Files ===", 'header')
    _progress(0)

    # Load tester mapping
    tester_mapping = load_tester_mapping()

    # Ensure folders exist
    ensure_master_folders()

    # Preprocess: Collect master data in single pass (Phase A optimization)
    # Opens each master file ONCE with read_only=True
    # Manager status preservation is handled by System 1 in excel_ops.py during rebuild
    _log("Collecting master data...")
    collect_start = time.time()
    (fixed_screenshots_en, fixed_screenshots_cn,
     manager_stats, manager_dates) = collect_all_master_data(tester_mapping, log_callback=log_callback)
    collect_elapsed = time.time() - collect_start

    _log(f"Master data collected ({collect_elapsed:.1f}s)")
    total_fixed = len(fixed_screenshots_en) + len(fixed_screenshots_cn)
    if total_fixed > 0:
        _log(f"Found {total_fixed} FIXED screenshots to skip")
    _progress(2)

    # Discover QA folders
    qa_folders = discover_qa_folders()

    if not qa_folders:
        _log("No valid QA folders found", 'error')
        return

    _log(f"Found {len(qa_folders)} QA folders")
    _progress(5)

    # Group by category AND language
    by_category_en, by_category_cn = group_folders_by_language(qa_folders, tester_mapping)
    _log(f"  EN: {sum(len(v) for v in by_category_en.values())} folders, CN: {sum(len(v) for v in by_category_cn.values())} folders")

    # ==========================================================================
    # EARLY OUTPUT: Generate MasterSubmitScript FIRST (quick - just ISSUE rows)
    # ==========================================================================
    # This runs BEFORE heavy Master file processing so output is available early
    _log("STEP 1: Generating Submit Scripts", 'header')

    from core.export_index import get_soundevent_mapping
    from core.submit_script import collect_issue_rows, generate_master_submit_script, generate_conflict_file

    export_mapping = get_soundevent_mapping()

    # EN: Combine Sequencer + Dialog folders
    en_script_folders = []
    for cat in ["Sequencer", "Dialog"]:
        if cat in by_category_en:
            en_script_folders.extend(by_category_en[cat])

    if en_script_folders:
        en_issues, en_conflicts = collect_issue_rows(en_script_folders, export_mapping)
        if en_issues:
            generate_master_submit_script(
                en_issues,
                MASTER_FOLDER_EN / "MasterSubmitScript_EN.xlsx",
                "EN"
            )
            _log(f"  EN: {len(en_issues)} ISSUE rows -> MasterSubmitScript_EN.xlsx", 'success')
        else:
            _log("  EN: No ISSUE rows found")
        if en_conflicts:
            generate_conflict_file(
                en_conflicts,
                MASTER_FOLDER_EN / "MasterSubmitScript_Conflicts_EN.xlsx",
                "EN"
            )
            _log(f"  EN: {len(en_conflicts)} script conflicts", 'warning')
    else:
        _log("  EN: No Script category files to process")

    # CN: Same pattern
    cn_script_folders = []
    for cat in ["Sequencer", "Dialog"]:
        if cat in by_category_cn:
            cn_script_folders.extend(by_category_cn[cat])

    if cn_script_folders:
        cn_issues, cn_conflicts = collect_issue_rows(cn_script_folders, export_mapping)
        if cn_issues:
            generate_master_submit_script(
                cn_issues,
                MASTER_FOLDER_CN / "MasterSubmitScript_CN.xlsx",
                "CN"
            )
            _log(f"  CN: {len(cn_issues)} ISSUE rows -> MasterSubmitScript_CN.xlsx", 'success')
        else:
            _log("  CN: No ISSUE rows found")
        if cn_conflicts:
            generate_conflict_file(
                cn_conflicts,
                MASTER_FOLDER_CN / "MasterSubmitScript_Conflicts_CN.xlsx",
                "CN"
            )
            _log(f"  CN: {len(cn_conflicts)} script conflicts", 'warning')
    else:
        _log("  CN: No Script category files to process")

    # ==========================================================================
    # STEP 1b: MasterSubmitDatasheet (Non-Script categories)
    # ==========================================================================
    _progress(10)
    _log("STEP 2: Generating Submit Datasheets", 'header')

    from core.submit_datasheet import (
        collect_datasheet_issue_rows,
        generate_master_submit_datasheet,
        generate_datasheet_conflict_file,
    )

    SCRIPT_CATEGORIES = {"Sequencer", "Dialog"}

    # EN: Collect all non-script QA folders
    en_datasheet_folders = []
    for cat, folders in by_category_en.items():
        if cat not in SCRIPT_CATEGORIES:
            en_datasheet_folders.extend(folders)

    if en_datasheet_folders:
        en_ds_issues, en_ds_conflicts = collect_datasheet_issue_rows(en_datasheet_folders)
        if en_ds_issues:
            generate_master_submit_datasheet(
                en_ds_issues,
                MASTER_FOLDER_EN / "MasterSubmitDatasheet_EN.xlsx",
                "EN"
            )
            _log(f"  EN: {len(en_ds_issues)} ISSUE rows -> MasterSubmitDatasheet_EN.xlsx", 'success')
        else:
            _log("  EN: No ISSUE rows found")
        if en_ds_conflicts:
            generate_datasheet_conflict_file(
                en_ds_conflicts,
                MASTER_FOLDER_EN / "MasterSubmitDatasheet_Conflicts_EN.xlsx",
                "EN"
            )
            _log(f"  EN: {len(en_ds_conflicts)} datasheet conflicts", 'warning')
    else:
        _log("  EN: No non-Script category files to process")

    # CN: Same pattern
    cn_datasheet_folders = []
    for cat, folders in by_category_cn.items():
        if cat not in SCRIPT_CATEGORIES:
            cn_datasheet_folders.extend(folders)

    if cn_datasheet_folders:
        cn_ds_issues, cn_ds_conflicts = collect_datasheet_issue_rows(cn_datasheet_folders)
        if cn_ds_issues:
            generate_master_submit_datasheet(
                cn_ds_issues,
                MASTER_FOLDER_CN / "MasterSubmitDatasheet_CN.xlsx",
                "CN"
            )
            _log(f"  CN: {len(cn_ds_issues)} ISSUE rows -> MasterSubmitDatasheet_CN.xlsx", 'success')
        else:
            _log("  CN: No ISSUE rows found")
        if cn_ds_conflicts:
            generate_datasheet_conflict_file(
                cn_ds_conflicts,
                MASTER_FOLDER_CN / "MasterSubmitDatasheet_Conflicts_CN.xlsx",
                "CN"
            )
            _log(f"  CN: {len(cn_ds_conflicts)} datasheet conflicts", 'warning')
    else:
        _log("  CN: No non-Script category files to process")

    # ==========================================================================
    # STEP 2: Process Master Files (Heavy Processing)
    # ==========================================================================
    _progress(15)
    _log("STEP 3: Building Master Files", 'header')

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    # Thread-safe result collection
    results_lock = threading.Lock()

    def process_worker_group(group_name, categories, lang_label, by_category,
                             master_folder, images_folder,
                             fixed_screenshots, tester_mapping_ref):
        """
        Process a worker group (categories sharing same master file).
        Categories within the group are processed SEQUENTIALLY (shared master).
        Returns (group_name, lang, daily_entries, master_status_data).
        """
        from core.face_processor import process_face_category

        daily_entries = []
        processed_masters = set()
        master_status_data = {}

        # Filter to categories that exist in this language
        active_categories = [c for c in categories if c in by_category]
        if not active_categories:
            return (group_name, lang_label, [], {})

        for category in active_categories:
            # Face category: custom processing pipeline
            if category.lower() in FACE_TYPE_CATEGORIES:
                entries = process_face_category(
                    by_category[category], master_folder, lang_label, tester_mapping_ref
                )
                daily_entries.extend(entries)
                continue

            target_master = get_target_master_category(category)
            rebuild = target_master not in processed_masters
            processed_masters.add(target_master)

            # When a second category shares the same master (e.g., Gimmick after Item),
            # save the in-memory workbook to disk so get_or_create_master can find it.
            # With deferred_save=True, the first category's workbook is only in memory.
            if not rebuild:
                prev_data = master_status_data.get(target_master)
                if prev_data and prev_data["workbook"] is not None and prev_data["path"] is not None:
                    prev_data["workbook"].save(prev_data["path"])

            # Get or initialize accumulated data for this master
            if target_master not in master_status_data:
                master_status_data[target_master] = {
                    "users": set(),
                    "stats": defaultdict(lambda: {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0}),
                    "workbook": None,
                    "path": None,
                    "category": category,
                    "is_english": lang_label == "EN",
                }
            data = master_status_data[target_master]
            acc_users = data["users"]
            acc_stats = data["stats"]

            entries, acc_users, acc_stats, master_wb, master_path = process_category(
                category, by_category[category],
                master_folder, images_folder, lang_label,
                rebuild=rebuild,
                fixed_screenshots=fixed_screenshots,
                accumulated_users=acc_users,
                accumulated_stats=acc_stats,
                deferred_save=True,
                log_callback=log_callback
            )
            daily_entries.extend(entries)
            master_status_data[target_master]["users"] = acc_users
            master_status_data[target_master]["stats"] = acc_stats
            if master_wb is not None:
                master_status_data[target_master]["workbook"] = master_wb
                master_status_data[target_master]["path"] = master_path

        return (group_name, lang_label, daily_entries, master_status_data)

    # ==========================================================================
    # PARALLEL PROCESSING: Worker groups across EN + CN (up to MAX_PARALLEL_WORKERS)
    # ==========================================================================
    all_daily_entries = []
    master_status_data_en = {}
    master_status_data_cn = {}

    # Submit all worker groups for both languages
    with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS, thread_name_prefix="wg") as executor:
        futures = []

        # Submit EN worker groups
        for group_name, categories in WORKER_GROUPS.items():
            if any(c in by_category_en for c in categories):
                future = executor.submit(
                    process_worker_group,
                    group_name, categories, "EN", by_category_en,
                    MASTER_FOLDER_EN, IMAGES_FOLDER_EN,
                    fixed_screenshots_en, tester_mapping
                )
                futures.append(future)

        # Submit CN worker groups
        for group_name, categories in WORKER_GROUPS.items():
            if any(c in by_category_cn for c in categories):
                future = executor.submit(
                    process_worker_group,
                    group_name, categories, "CN", by_category_cn,
                    MASTER_FOLDER_CN, IMAGES_FOLDER_CN,
                    fixed_screenshots_cn, tester_mapping
                )
                futures.append(future)

        total_futures = len(futures)

        # Collect results as they complete
        completed = 0
        for future in as_completed(futures):
            try:
                group_name, lang, entries, master_data = future.result()
                completed += 1

                with results_lock:
                    all_daily_entries.extend(entries)
                    if lang == "EN":
                        master_status_data_en.update(master_data)
                    else:
                        master_status_data_cn.update(master_data)

                if entries:
                    _log(f"  [{lang}/{group_name}] Complete ({completed}/{total_futures})")
                # Progress: 15% to 85% based on worker completion
                worker_pct = 15 + int(70 * completed / total_futures)
                _progress(worker_pct)
            except Exception as e:
                _log(f"Worker failed: {group_name}: {e}", 'error')
                raise

    _log("All workers completed", 'success')
    _progress(85)

    # ==========================================================================
    # FINAL PASS: STATUS sheet + autofit + hide + save (PARALLEL)
    # ==========================================================================
    # This is the DEFERRED SAVE optimization: instead of autofit+save after each
    # category, we do it once per master after ALL categories are processed.
    # Now also parallelized for additional speedup.
    _log("Finalizing master files...", 'header')
    _progress(90)

    def finalize_master(target_master, data, lang_label):
        """Finalize and save a single master file. Thread-safe."""
        users = data["users"]
        stats = data["stats"]
        wb = data["workbook"]
        path = data["path"]
        cat = data.get("category", "Quest")
        is_eng = data.get("is_english", True)

        if wb is None or path is None:
            return None

        # 0. Post-process: replicate data across duplicate rows
        replicate_duplicate_row_data(wb, cat, is_eng)

        # 1. Re-apply manager dropdowns (before autofit)
        for sheet_name in wb.sheetnames:
            if sheet_name == "STATUS":
                continue
            ws = wb[sheet_name]
            if ws.max_row and ws.max_row >= 2:
                reapply_manager_dropdowns(ws)

        # 2. Update STATUS sheet
        if users:
            update_status_sheet(wb, users, dict(stats))

        # 3. Autofit (returns preloaded sheet data for reuse)
        sheet_data_cache = autofit_rows_with_wordwrap(wb)

        # 4. Hide empty rows/sheets/columns (reuses preloaded data — avoids redundant load)
        hidden_rows, hidden_sheets, hidden_columns = hide_empty_comment_rows(wb, preloaded_sheets=sheet_data_cache)

        # 5. Beautify — after autofit so header colors/alignment aren't stomped
        for sheet_name in wb.sheetnames:
            if sheet_name == "STATUS":
                continue
            ws = wb[sheet_name]
            if ws.max_row and ws.max_row >= 2:
                beautify_master_sheet(ws)

        # 6. FINAL COLUMN SWEEP — absolute last step before save
        # Hides user column blocks where ALL visible rows are empty
        # (catches columns where all comments are on hidden/resolved rows)
        final_column_sweep(wb)

        # 7. Save
        wb.save(path)

        # Clean up .bak backup now that new master is safely saved
        # (unless high orphan rate flagged it for preservation)
        backup_path = path.with_suffix(".xlsx.bak")
        if backup_path.exists():
            if getattr(wb, '_qacompiler_keep_backup', False):
                print(f"  ⚠ Backup PRESERVED (high orphan rate): {backup_path.name}")
            else:
                backup_path.unlink()

        return {
            "lang": lang_label,
            "master": target_master,
            "users": len(users),
            "hidden_rows": hidden_rows,
            "hidden_sheets": hidden_sheets,
            "hidden_columns": hidden_columns,
        }

    # Parallel finalize + save
    finalize_futures = []
    with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS, thread_name_prefix="save") as executor:
        for target_master, data in master_status_data_en.items():
            future = executor.submit(finalize_master, target_master, data, "EN")
            finalize_futures.append(future)

        for target_master, data in master_status_data_cn.items():
            future = executor.submit(finalize_master, target_master, data, "CN")
            finalize_futures.append(future)

        for future in as_completed(finalize_futures):
            try:
                result = future.result()
                if result:
                    lang = result["lang"]
                    master = result["master"]
                    users = result["users"]
                    hidden = result.get("hidden_rows", 0)
                    _log(f"  Master_{master} [{lang}]: {users} users, {hidden} rows hidden, saved", 'success')
            except Exception as e:
                _log(f"ERROR: Failed to finalize: {e}", 'error')
                import traceback
                traceback.print_exc()
                raise

    # Show skipped categories
    for category in CATEGORIES:
        if category not in by_category_en and category not in by_category_cn:
            _log(f"  Skipped: {category} (no folders)")

    # Update Progress Tracker
    if all_daily_entries:
        _log("Updating Tracker...", 'header')
        _progress(92)

        # Import tracker modules
        from tracker.data import get_or_create_tracker, update_daily_data_sheet
        from tracker.daily import build_daily_sheet
        from tracker.total import build_total_sheet

        # manager_stats already collected by collect_all_master_data() above
        tracker_wb, tracker_path = get_or_create_tracker()

        # Separate Face entries (different schema) from standard entries
        standard_entries = [e for e in all_daily_entries if e.get("category") != "Face"]
        face_entries = [e for e in all_daily_entries if e.get("category") == "Face"]

        # Update standard tracker tabs
        if standard_entries:
            update_daily_data_sheet(tracker_wb, standard_entries, manager_stats, manager_dates)
        build_daily_sheet(tracker_wb)
        build_total_sheet(tracker_wb)

        # Update Facial tracker tab (if Face entries exist)
        if face_entries:
            from tracker.facial import update_facial_data_sheet, build_facial_sheet
            update_facial_data_sheet(tracker_wb, face_entries)
            build_facial_sheet(tracker_wb)

        # Remove deprecated GRAPHS sheet
        if "GRAPHS" in tracker_wb.sheetnames:
            del tracker_wb["GRAPHS"]

        tracker_wb.save(tracker_path)
        _log(f"Tracker saved: {len(standard_entries)} standard + {len(face_entries)} face entries", 'success')
        _progress(95)

    # =========================================================================
    # FINAL SUMMARY
    # =========================================================================
    elapsed = time.time() - compilation_start_time
    minutes = int(elapsed // 60)
    seconds = elapsed % 60

    _log("=== COMPILATION COMPLETE ===", 'header')

    if minutes > 0:
        _log(f"Time: {minutes}m {seconds:.1f}s")
    else:
        _log(f"Time: {seconds:.1f}s")

    if all_daily_entries:
        testers_en = set()
        testers_cn = set()
        total_rows = 0
        total_done = 0
        total_issues = 0
        total_blocked = 0
        total_korean = 0

        for entry in all_daily_entries:
            lang = entry.get("lang", "?")
            user = entry.get("user", "?")
            if lang == "EN":
                testers_en.add(user)
            else:
                testers_cn.add(user)
            total_rows += entry.get("total_rows", 0)
            total_done += entry.get("done", 0)
            total_issues += entry.get("issues", 0)
            total_blocked += entry.get("blocked", 0)
            total_korean += entry.get("korean", 0)

        _log(f"Testers: {len(testers_en)} EN + {len(testers_cn)} CN ({len(testers_en | testers_cn)} total)")
        _log(f"Rows: {total_rows:,} processed, {total_done:,} done")
        if total_issues > 0:
            _log(f"Issues: {total_issues:,} ISSUE, {total_blocked:,} BLOCKED", 'warning')

        try:
            finalize_count = len(finalize_futures)
        except NameError:
            finalize_count = 0
        if finalize_count > 0:
            _log(f"Master files saved: {finalize_count}", 'success')

    _progress(100)
