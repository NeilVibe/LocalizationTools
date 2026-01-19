#!/usr/bin/env python3
"""
Debug script for Progress Tracker date issues.

Shows:
1. File modification dates for all QA files
2. Tester -> Language mapping
3. What dates would be recorded
4. Why EN/CN might differ
"""

import os
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from config import (
    QA_FOLDER, CATEGORIES, TRACKER_PATH,
    load_tester_mapping
)

def main():
    print("=" * 70)
    print("PROGRESS TRACKER DEBUG")
    print("=" * 70)
    print(f"Today's date: {datetime.now().strftime('%Y-%m-%d')}")
    print(f"QA Folder: {QA_FOLDER}")
    print(f"Tracker: {TRACKER_PATH}")
    print()

    # Load tester mapping
    print("=" * 70)
    print("1. TESTER -> LANGUAGE MAPPING")
    print("=" * 70)
    tester_mapping = load_tester_mapping()

    if not tester_mapping:
        print("  WARNING: No tester mapping found!")
        print(f"  Expected file: {QA_FOLDER.parent / 'languageTOtester_list.txt'}")
    else:
        en_testers = [t for t, lang in tester_mapping.items() if lang == "EN"]
        cn_testers = [t for t, lang in tester_mapping.items() if lang == "CN"]
        print(f"  EN testers ({len(en_testers)}): {', '.join(en_testers[:5])}{'...' if len(en_testers) > 5 else ''}")
        print(f"  CN testers ({len(cn_testers)}): {', '.join(cn_testers[:5])}{'...' if len(cn_testers) > 5 else ''}")
    print()

    # Scan QA folders
    print("=" * 70)
    print("2. QA FOLDER FILE DATES")
    print("=" * 70)

    if not QA_FOLDER.exists():
        print(f"  ERROR: QA folder not found: {QA_FOLDER}")
        return

    folders = list(QA_FOLDER.iterdir())
    if not folders:
        print("  No folders found in QAfolder/")
        return

    # Collect data
    en_files = []
    cn_files = []
    unknown_files = []

    today = datetime.now().strftime("%Y-%m-%d")

    print(f"\n  {'Folder':<40} {'File Date':<12} {'Lang':<4} {'Today?':<6}")
    print("  " + "-" * 66)

    for folder in sorted(folders):
        if not folder.is_dir():
            continue

        # Parse folder name
        parts = folder.name.rsplit("_", 1)
        if len(parts) != 2:
            continue

        username, category = parts
        if category not in CATEGORIES:
            continue

        # Find xlsx file
        xlsx_files = list(folder.glob("*.xlsx"))
        if not xlsx_files:
            continue

        xlsx_path = xlsx_files[0]

        # Get file modification date
        file_mtime = datetime.fromtimestamp(xlsx_path.stat().st_mtime)
        file_date = file_mtime.strftime("%Y-%m-%d")
        file_time = file_mtime.strftime("%H:%M:%S")

        # Determine language
        lang = tester_mapping.get(username, "EN")  # Default to EN like the code does
        is_today = "YES" if file_date == today else "NO"

        # Color coding hint
        marker = "*" if file_date == today else " "

        print(f"  {folder.name:<40} {file_date} {file_time}  {lang:<4} {is_today:<6} {marker}")

        entry = {
            "folder": folder.name,
            "username": username,
            "category": category,
            "file_date": file_date,
            "file_time": file_time,
            "xlsx_path": xlsx_path,
        }

        if lang == "EN":
            en_files.append(entry)
        elif lang == "CN":
            cn_files.append(entry)
        else:
            unknown_files.append(entry)

    # Summary
    print()
    print("=" * 70)
    print("3. SUMMARY BY LANGUAGE")
    print("=" * 70)

    en_today = [f for f in en_files if f["file_date"] == today]
    cn_today = [f for f in cn_files if f["file_date"] == today]

    print(f"\n  EN FILES:")
    print(f"    Total: {len(en_files)}")
    print(f"    Modified today: {len(en_today)}")
    if en_today:
        for f in en_today:
            print(f"      - {f['folder']} ({f['file_time']})")
    else:
        print(f"    >>> NO EN FILES MODIFIED TODAY <<<")
        if en_files:
            # Show most recent EN files
            en_sorted = sorted(en_files, key=lambda x: x["file_date"], reverse=True)
            print(f"    Most recent EN file dates:")
            for f in en_sorted[:3]:
                print(f"      - {f['folder']}: {f['file_date']}")

    print(f"\n  CN FILES:")
    print(f"    Total: {len(cn_files)}")
    print(f"    Modified today: {len(cn_today)}")
    if cn_today:
        for f in cn_today:
            print(f"      - {f['folder']} ({f['file_time']})")
    else:
        print(f"    >>> NO CN FILES MODIFIED TODAY <<<")

    # Check tracker if exists
    print()
    print("=" * 70)
    print("4. EXISTING TRACKER DATA")
    print("=" * 70)

    if TRACKER_PATH.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(TRACKER_PATH, data_only=True)

            if "_DAILY_DATA" in wb.sheetnames:
                ws = wb["_DAILY_DATA"]

                # Read dates from tracker
                tracker_dates = set()
                tracker_users = defaultdict(set)  # date -> set of users

                for row in range(2, ws.max_row + 1):
                    date = ws.cell(row, 1).value
                    user = ws.cell(row, 2).value
                    if date and user:
                        date_str = str(date)[:10] if date else ""
                        tracker_dates.add(date_str)
                        tracker_users[date_str].add(user)

                print(f"  Dates in tracker: {sorted(tracker_dates)[-5:] if tracker_dates else 'None'}")
                print(f"  Today ({today}) in tracker: {'YES' if today in tracker_dates else 'NO'}")

                if today in tracker_users:
                    users_today = tracker_users[today]
                    en_in_tracker = [u for u in users_today if tester_mapping.get(u, "EN") == "EN"]
                    cn_in_tracker = [u for u in users_today if tester_mapping.get(u) == "CN"]
                    print(f"  Users for today: EN={len(en_in_tracker)}, CN={len(cn_in_tracker)}")
                else:
                    print(f"  No users recorded for today yet")

            wb.close()
        except Exception as e:
            print(f"  Error reading tracker: {e}")
    else:
        print(f"  Tracker file not found: {TRACKER_PATH}")

    # Diagnosis
    print()
    print("=" * 70)
    print("5. DIAGNOSIS")
    print("=" * 70)

    if not en_today and en_files:
        print("""
  ISSUE: EN files exist but none modified today.

  The tracker uses FILE MODIFICATION DATE, not compile date.
  EN testers' Excel files haven't been saved today.

  SOLUTIONS:
  1. Have EN testers save their files (even without changes)
  2. Or touch the files: touch QAfolder/*_*/file.xlsx
  3. Or modify compiler.py to use current date instead of file mtime
""")
    elif not en_files:
        print("""
  ISSUE: No EN files found in QAfolder.

  Either:
  1. No EN tester folders exist
  2. EN testers are mapped to CN in languageTOtester_list.txt
  3. Folder naming doesn't match {Username}_{Category} pattern
""")
    elif en_today:
        print(f"""
  EN files look OK - {len(en_today)} files modified today.

  If tracker still not showing, try:
  1. Run the build: python main.py --build
  2. Check for errors in the build output
""")

    print()


if __name__ == "__main__":
    main()
