#!/usr/bin/env python3
"""
System Sheet Localizer - Standalone Tool

Creates localized versions of a System datasheet for all languages.
Uses a 2-step matching process:
  1. StringID → Target Language (DIRECT lookup in each language's data)
  2. English → Korean → Target Language (fallback if no StringID match)

System Excel Column Layout:
  - Column A (1): CONTENT - The text to localize
  - Column B (2): STATUS
  - Column C (3): COMMENT
  - Column D (4): STRINGID - Unique identifier for direct lookup
  - Column E (5): SCREENSHOT

Usage:
  python system_localizer.py
  (Opens file dialogs for input selection)

Output:
  Creates System_LQA_All/ folder with Excel files for each language.
"""

import os
import sys
import logging
from pathlib import Path
from typing import Dict, Tuple, Optional, List
from copy import copy

# GUI imports
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    print("ERROR: tkinter not available. Please install python3-tk")
    sys.exit(1)

# Excel imports
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# XML imports
try:
    from lxml import etree as ET
except ImportError:
    print("ERROR: lxml not installed. Run: pip install lxml")
    sys.exit(1)

# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

# =============================================================================
# LANGUAGE DATA LOADING
# =============================================================================

def iter_xml_files(folder: Path, pattern: str = "*.xml"):
    """Iterate over XML files in folder (non-recursive)."""
    if not folder.exists():
        return
    for f in folder.glob(pattern):
        if f.is_file():
            yield f


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    """Parse XML file, return root element or None on error."""
    try:
        tree = ET.parse(str(path), ET.XMLParser(recover=True))
        return tree.getroot()
    except Exception as e:
        log.warning("Failed to parse %s: %s", path.name, e)
        return None


def load_all_language_data(folder: Path) -> Tuple[
    Dict[str, Dict[str, Tuple[str, str]]],  # lang_code → {korean: (translation, stringid)}
    Dict[str, Dict[str, str]],               # lang_code → {stringid: translation} (DIRECT lookup)
    Dict[str, str]                           # english → korean (for fallback)
]:
    """
    Load all language tables and build lookup indexes.

    Returns:
        lang_tables: {lang_code: {korean: (translation, stringid)}} - for text matching
        sid_tables: {lang_code: {stringid: translation}} - for DIRECT StringID lookup
        eng_to_korean: {english_text: korean_text} - for fallback matching
    """
    lang_tables: Dict[str, Dict[str, Tuple[str, str]]] = {}
    sid_tables: Dict[str, Dict[str, str]] = {}  # DIRECT: StringID → Translation per language
    eng_to_korean: Dict[str, str] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}
        sid_tbl: Dict[str, str] = {}  # StringID → Translation for THIS language

        for loc in root_el.iter("LocStr"):
            korean = loc.get("StrOrigin") or ""
            translation = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not korean:
                continue

            # Store in language table (korean → translation)
            tbl[korean] = (translation, sid)

            # Store DIRECT StringID → Translation for this language
            if sid and translation:
                sid_tbl[sid] = translation

            # Build English → Korean reverse lookup (for fallback)
            if lang == "eng" and translation:
                eng_to_korean[translation] = korean

        lang_tables[lang] = tbl
        sid_tables[lang] = sid_tbl
        log.info("  Loaded %s: %d entries, %d StringIDs", lang, len(tbl), len(sid_tbl))

    return lang_tables, sid_tables, eng_to_korean


# =============================================================================
# EXCEL PROCESSING
# =============================================================================

def find_columns(ws, headers: List[str]) -> Dict[str, int]:
    """
    Find column indices for given header names (case-insensitive).
    Searches first 5 rows for headers.

    Returns:
        {header_name_lower: column_index} (1-based)
    """
    columns: Dict[str, int] = {}
    headers_lower = [h.lower() for h in headers]

    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val:
                val_lower = str(val).strip().lower()
                for h, h_lower in zip(headers, headers_lower):
                    if h_lower in val_lower or val_lower in h_lower:
                        if h_lower not in columns:
                            columns[h_lower] = col
                            log.info("    Found '%s' at column %d (row %d)", h, col, row)

    return columns


def copy_cell_style(src_cell, dst_cell):
    """Copy style from source cell to destination cell."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def localize_system_sheet(
    input_path: Path,
    lang_folder: Path,
    output_folder: Path,
    progress_callback=None
) -> Dict[str, any]:
    """
    Create localized versions of a System datasheet for all languages.

    Args:
        input_path: Path to the English System Excel file
        lang_folder: Path to language data folder
        output_folder: Path to output folder
        progress_callback: Optional callback(current, total, message)

    Returns:
        Result dict with stats and errors
    """
    result = {
        "success": True,
        "files_created": 0,
        "languages": [],
        "errors": [],
        "stats": {}
    }

    # 1. Load language data
    log.info("Loading language data from: %s", lang_folder)
    if progress_callback:
        progress_callback(0, 100, "Loading language data...")

    lang_tables, sid_tables, eng_to_korean = load_all_language_data(lang_folder)

    if "eng" not in lang_tables:
        result["success"] = False
        result["errors"].append("English language data not found!")
        return result

    log.info("Loaded %d languages, %d English→Korean mappings",
             len(lang_tables), len(eng_to_korean))

    # 2. Load input Excel
    log.info("Loading input file: %s", input_path)
    if progress_callback:
        progress_callback(10, 100, "Loading input Excel...")

    try:
        wb_in = load_workbook(input_path)
    except Exception as e:
        result["success"] = False
        result["errors"].append(f"Failed to load Excel: {e}")
        return result

    # 3. Create output folder
    output_folder.mkdir(parents=True, exist_ok=True)

    # 4. Process each language
    languages = [code for code in lang_tables.keys() if code != "kor"]
    total_langs = len(languages)

    for idx, lang_code in enumerate(languages):
        lang_tbl = lang_tables[lang_code]
        sid_tbl = sid_tables.get(lang_code, {})  # DIRECT StringID → Translation for this language
        progress_pct = 10 + int(80 * (idx + 1) / total_langs)

        if progress_callback:
            progress_callback(progress_pct, 100, f"Processing {lang_code.upper()}...")

        log.info("Processing language: %s (%d/%d)", lang_code.upper(), idx + 1, total_langs)

        # Create new workbook by copying input
        wb_out = Workbook()
        wb_out.remove(wb_out.active)

        lang_stats = {"total_rows": 0, "matched_by_sid": 0, "matched_by_text": 0, "no_match": 0}

        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            ws_out = wb_out.create_sheet(sheet_name)

            # System datasheet layout (hardcoded):
            # Column A (1): CONTENT - text to localize
            # Column B (2): STATUS
            # Column C (3): COMMENT
            # Column D (4): STRINGID - for lookup
            # Column E (5): SCREENSHOT
            CONTENT_COL = 1
            STRINGID_COL = 4

            log.info("  Sheet '%s': CONTENT=col %d, STRINGID=col %d",
                     sheet_name, CONTENT_COL, STRINGID_COL)

            # Copy all cells, replacing translation column
            for row in range(1, ws_in.max_row + 1):
                for col in range(1, ws_in.max_column + 1):
                    src_cell = ws_in.cell(row, col)
                    dst_cell = ws_out.cell(row, col)

                    # Copy style
                    copy_cell_style(src_cell, dst_cell)

                    # Copy value (may be replaced below)
                    dst_cell.value = src_cell.value

                # Skip header row(s)
                if row <= 1:
                    continue

                # Get STRINGID and current CONTENT for matching
                string_id = ws_in.cell(row, STRINGID_COL).value
                current_content = ws_in.cell(row, CONTENT_COL).value

                translation = None
                match_type = None

                # Step 1: DIRECT StringID → Target Language
                if string_id and str(string_id).strip():
                    sid_clean = str(string_id).strip()
                    translation = sid_tbl.get(sid_clean)
                    if translation:
                        match_type = "sid"

                # Step 2: Fallback - English (CONTENT) → Korean → Target Language
                if not translation and current_content and str(current_content).strip():
                    eng_clean = str(current_content).strip()
                    korean = eng_to_korean.get(eng_clean)
                    if korean and korean in lang_tbl:
                        translation, _ = lang_tbl[korean]
                        if translation:
                            match_type = "text"

                # Update CONTENT column with translation
                if translation:
                    ws_out.cell(row, CONTENT_COL).value = translation
                    if match_type == "sid":
                        lang_stats["matched_by_sid"] += 1
                    else:
                        lang_stats["matched_by_text"] += 1
                else:
                    lang_stats["no_match"] += 1

                lang_stats["total_rows"] += 1

            # Copy column widths
            for col in range(1, ws_in.max_column + 1):
                col_letter = get_column_letter(col)
                if ws_in.column_dimensions[col_letter].width:
                    ws_out.column_dimensions[col_letter].width = ws_in.column_dimensions[col_letter].width

            # Copy row heights
            for row in range(1, ws_in.max_row + 1):
                if ws_in.row_dimensions[row].height:
                    ws_out.row_dimensions[row].height = ws_in.row_dimensions[row].height

            # Add STATUS dropdown (Column B) - same as main QA compiler
            STATUS_COL = 2  # Column B
            last_row = max(ws_out.max_row, 10) + 50  # Buffer for future rows
            status_dv = DataValidation(
                type="list",
                formula1='"ISSUE,NO ISSUE,BLOCKED,KOREAN"',
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid Status",
                error="Please select: ISSUE, NO ISSUE, BLOCKED, or KOREAN"
            )
            ws_out.add_data_validation(status_dv)
            status_dv.add(f"B2:B{last_row}")

        # Save output file
        output_name = f"System_{lang_code.upper()}.xlsx"
        output_path = output_folder / output_name

        try:
            wb_out.save(output_path)
            result["files_created"] += 1
            result["languages"].append(lang_code.upper())
            result["stats"][lang_code] = lang_stats
            log.info("  Saved: %s (SID: %d, Text: %d, NoMatch: %d)",
                     output_name, lang_stats["matched_by_sid"],
                     lang_stats["matched_by_text"], lang_stats["no_match"])
        except Exception as e:
            result["errors"].append(f"Failed to save {output_name}: {e}")

    # 5. Also copy the original English file for completeness
    try:
        import shutil
        eng_output = output_folder / "System_ENG.xlsx"
        shutil.copy2(input_path, eng_output)
        log.info("Copied original English file")
    except Exception as e:
        log.warning("Failed to copy English file: %s", e)

    if progress_callback:
        progress_callback(100, 100, "Complete!")

    return result


# =============================================================================
# GUI
# =============================================================================

class SystemLocalizerGUI:
    """Simple GUI for System Sheet Localizer."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("System Sheet Localizer")
        self.root.geometry("600x400")
        self.root.resizable(True, True)

        self.input_path = tk.StringVar()
        self.lang_folder = tk.StringVar()
        self.output_folder = tk.StringVar(value="System_LQA_All")

        self._build_ui()

    def _build_ui(self):
        """Build the GUI."""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title = ttk.Label(main_frame, text="System Sheet Localizer",
                          font=("Helvetica", 14, "bold"))
        title.pack(pady=(0, 10))

        desc = ttk.Label(main_frame, text="Create localized versions of System datasheet for all languages.\n"
                                          "Columns: CONTENT (A) | STATUS (B) | COMMENT (C) | STRINGID (D) | SCREENSHOT (E)\n"
                                          "Match: StringID→Translation (direct) or CONTENT→Korean→Translation (fallback)",
                         justify=tk.CENTER)
        desc.pack(pady=(0, 15))

        # Input file
        input_frame = ttk.LabelFrame(main_frame, text="1. Select System Excel File", padding="5")
        input_frame.pack(fill=tk.X, pady=5)

        ttk.Entry(input_frame, textvariable=self.input_path, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(input_frame, text="Browse...", command=self._browse_input).pack(side=tk.RIGHT, padx=(5, 0))

        # Language folder
        lang_frame = ttk.LabelFrame(main_frame, text="2. Select Language Data Folder", padding="5")
        lang_frame.pack(fill=tk.X, pady=5)

        ttk.Entry(lang_frame, textvariable=self.lang_folder, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(lang_frame, text="Browse...", command=self._browse_lang).pack(side=tk.RIGHT, padx=(5, 0))

        # Output info
        output_frame = ttk.LabelFrame(main_frame, text="3. Output", padding="5")
        output_frame.pack(fill=tk.X, pady=5)

        ttk.Label(output_frame, text="Output folder: System_LQA_All/ (created next to input file)").pack(anchor=tk.W)

        # Progress
        self.progress_var = tk.DoubleVar()
        self.progress_label = ttk.Label(main_frame, text="")
        self.progress_label.pack(pady=(10, 0))

        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=5)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        self.run_btn = ttk.Button(btn_frame, text="Generate All Languages", command=self._run)
        self.run_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(btn_frame, text="Exit", command=self.root.quit).pack(side=tk.LEFT, padx=5)

        # Status
        self.status_text = tk.Text(main_frame, height=8, width=70, state=tk.DISABLED)
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

    def _browse_input(self):
        """Browse for input Excel file."""
        path = filedialog.askopenfilename(
            title="Select System Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.input_path.set(path)

    def _browse_lang(self):
        """Browse for language data folder."""
        path = filedialog.askdirectory(title="Select Language Data Folder (containing LanguageData_*.xml)")
        if path:
            self.lang_folder.set(path)

    def _update_progress(self, current, total, message):
        """Update progress bar and label."""
        self.progress_var.set(current)
        self.progress_label.config(text=message)
        self.root.update_idletasks()

    def _log_status(self, message):
        """Add message to status text."""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _run(self):
        """Run the localization."""
        # Validate inputs
        if not self.input_path.get():
            messagebox.showerror("Error", "Please select an input Excel file.")
            return

        if not self.lang_folder.get():
            messagebox.showerror("Error", "Please select a language data folder.")
            return

        input_path = Path(self.input_path.get())
        lang_folder = Path(self.lang_folder.get())

        if not input_path.exists():
            messagebox.showerror("Error", f"Input file not found: {input_path}")
            return

        if not lang_folder.exists():
            messagebox.showerror("Error", f"Language folder not found: {lang_folder}")
            return

        # Output folder next to input file
        output_folder = input_path.parent / "System_LQA_All"

        # Clear status
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.config(state=tk.DISABLED)

        # Disable button during processing
        self.run_btn.config(state=tk.DISABLED)

        try:
            self._log_status(f"Input: {input_path.name}")
            self._log_status(f"Language folder: {lang_folder}")
            self._log_status(f"Output: {output_folder}")
            self._log_status("-" * 50)

            result = localize_system_sheet(
                input_path,
                lang_folder,
                output_folder,
                progress_callback=self._update_progress
            )

            # Show results
            self._log_status("-" * 50)
            if result["success"]:
                self._log_status(f"SUCCESS! Created {result['files_created']} files")
                self._log_status(f"Languages: {', '.join(result['languages'])}")

                # Show stats summary
                total_matched = sum(s.get("matched_by_sid", 0) + s.get("matched_by_text", 0)
                                    for s in result["stats"].values())
                total_no_match = sum(s.get("no_match", 0) for s in result["stats"].values())
                self._log_status(f"Total matched: {total_matched}, No match: {total_no_match}")

                messagebox.showinfo("Success",
                                    f"Created {result['files_created']} localized files!\n\n"
                                    f"Output folder:\n{output_folder}")
            else:
                self._log_status("FAILED!")
                for err in result["errors"]:
                    self._log_status(f"  ERROR: {err}")
                messagebox.showerror("Error", "\n".join(result["errors"]))

        except Exception as e:
            self._log_status(f"EXCEPTION: {e}")
            messagebox.showerror("Error", f"An error occurred:\n{e}")

        finally:
            self.run_btn.config(state=tk.NORMAL)

    def run(self):
        """Start the GUI."""
        self.root.mainloop()


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    log.info("System Sheet Localizer - Starting...")

    # Check for command line args (for CLI mode)
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        # CLI mode (for scripting)
        if len(sys.argv) < 4:
            print("Usage: python system_localizer.py --cli <input.xlsx> <lang_folder>")
            sys.exit(1)

        input_path = Path(sys.argv[2])
        lang_folder = Path(sys.argv[3])
        output_folder = input_path.parent / "System_LQA_All"

        result = localize_system_sheet(input_path, lang_folder, output_folder)

        if result["success"]:
            print(f"SUCCESS! Created {result['files_created']} files in {output_folder}")
        else:
            print("FAILED!")
            for err in result["errors"]:
                print(f"  ERROR: {err}")
            sys.exit(1)
    else:
        # GUI mode
        app = SystemLocalizerGUI()
        app.run()


if __name__ == "__main__":
    main()
