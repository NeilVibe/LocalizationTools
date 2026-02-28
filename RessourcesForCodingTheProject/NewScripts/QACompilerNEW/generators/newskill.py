"""
NEW Skill Datasheet Generator
===============================
Row-per-text skill datasheet with two comparison tabs:
  Tab 1: SkillGroup (from skillgroupinfo.staticinfo.xml)
  Tab 2: SkillTree (from SkillTreeInfo.staticinfo.xml)

Each skill produces up to 6 rows:
  1. SkillData   — SkillName
  2. SkillData   — SkillDesc
  3. KnowledgeData  — Name  (Pass 1: LearnKnowledgeKey direct lookup)
  4. KnowledgeData  — Desc  (Pass 1)
  5. KnowledgeData2 — Name  (Pass 2: identical SkillName == KnowledgeInfo.Name)
  6. KnowledgeData2 — Desc  (Pass 2)

Key features:
- Two data sources for grouping: SkillGroupInfo and SkillTreeInfo
- Reuses knowledge loading from newitem.py
- Gold header rows for groups/trees
- Alternating fill per skill, depth-based coloring for SkillTree tab
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT
from generators.base import (
    get_logger,
    parse_xml_file,
    load_language_tables,
    normalize_placeholders,
    autofit_worksheet,
    THIN_BORDER,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
    add_status_dropdown,
)
from generators.newitem import _find_knowledge_key, load_knowledge_data

log = get_logger("NewSkillGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class NewSkillEntry:
    """Complete data for a single skill with knowledge fields."""
    key: str                        # SkillInfo Key
    strkey: str                     # SkillInfo StrKey
    skill_name_kor: str             # SkillInfo.SkillName
    skill_desc_kor: str             # SkillInfo.SkillDesc
    learn_knowledge_key: str        # SkillInfo.LearnKnowledgeKey
    knowledge_name_kor: str         # Pass 1: KnowledgeInfo.Name
    knowledge_desc_kor: str         # Pass 1: KnowledgeInfo.Desc
    knowledge2_name_kor: str        # Pass 2: identical name match
    knowledge2_desc_kor: str        # Pass 2: desc
    source_file: str                # skillinfo_pc filename
    knowledge_source_file: str      # Knowledge XML filename
    knowledge2_source_file: str = ""  # Pass 2 source file


@dataclass
class SkillGroupEntry:
    """One SkillGroupInfo element with its child skill keys."""
    strkey: str
    group_name_kor: str
    skill_keys: List[str] = field(default_factory=list)


@dataclass
class SkillNodeEntry:
    """A single SkillNode within a SkillTreeInfo."""
    node_id: str
    skill_key: str
    parent_id: str
    children: List["SkillNodeEntry"] = field(default_factory=list)


@dataclass
class SkillTreeEntry:
    """One SkillTreeInfo element with its node hierarchy."""
    key: str
    strkey: str
    character_key: str
    ui_page_name_kor: str
    nodes: List[SkillNodeEntry] = field(default_factory=list)


# =============================================================================
# SKILL EXTRACTION
# =============================================================================

def scan_skills_with_knowledge(
    skill_file: Path,
    knowledge_map: Dict[str, Tuple[str, str, str]],
    knowledge_name_index: Dict[str, List[Tuple[str, str, str]]],
) -> Dict[str, NewSkillEntry]:
    """Parse skillinfo_pc.staticinfo.xml and build skill lookup with knowledge.

    Pass 1: LearnKnowledgeKey -> knowledge_map (direct key lookup)
    Pass 2: SkillName -> knowledge_name_index (identical name match)

    IMPORTANT: Lookup is keyed by StrKey (e.g. "Skill_Wrestle_AirBodySlam"),
    NOT by the numeric Key (e.g. "15004"). SkillGroupInfo and SkillTreeInfo
    reference skills using StrKey values in their child elements.

    Returns:
        skill_lookup: {StrKey: NewSkillEntry} in document order
    """
    log.info("Scanning skills with knowledge: %s", skill_file.name)
    skill_lookup: Dict[str, NewSkillEntry] = {}
    pass2_hits = 0

    root = parse_xml_file(skill_file)
    if root is None:
        log.error("Cannot parse skill file: %s", skill_file)
        return skill_lookup

    source_file = skill_file.name

    for el in root.iter("SkillInfo"):
        key = el.get("Key") or ""
        strkey = el.get("StrKey") or ""
        skill_name = el.get("SkillName") or ""
        skill_desc = el.get("SkillDesc") or ""
        learn_knowledge_key = el.get("LearnKnowledgeKey") or ""

        if not strkey or not skill_name:
            continue

        # Dedup by StrKey
        if strkey in skill_lookup:
            continue

        # Pass 1: Resolve knowledge data via LearnKnowledgeKey
        knowledge_name = ""
        knowledge_desc = ""
        knowledge_source_file = ""
        pass1_strkey = ""
        if learn_knowledge_key and learn_knowledge_key in knowledge_map:
            knowledge_name, knowledge_desc, knowledge_source_file = knowledge_map[learn_knowledge_key]
            pass1_strkey = learn_knowledge_key

        # Pass 2: Identical name match (SkillName == KnowledgeInfo.Name)
        knowledge2_name = ""
        knowledge2_desc = ""
        knowledge2_source_file = ""
        if skill_name and skill_name in knowledge_name_index:
            for kn_strkey, kn_desc, kn_src in knowledge_name_index[skill_name]:
                if kn_strkey != pass1_strkey:
                    knowledge2_name = skill_name
                    knowledge2_desc = kn_desc
                    knowledge2_source_file = kn_src
                    pass2_hits += 1
                    break

        # Collect Korean strings for coverage tracking
        _collect_korean_string(skill_name)
        _collect_korean_string(skill_desc)
        _collect_korean_string(knowledge_name)
        _collect_korean_string(knowledge_desc)
        _collect_korean_string(knowledge2_name)
        _collect_korean_string(knowledge2_desc)

        skill_lookup[strkey] = NewSkillEntry(
            key=key,
            strkey=strkey,
            skill_name_kor=skill_name,
            skill_desc_kor=skill_desc,
            learn_knowledge_key=learn_knowledge_key,
            knowledge_name_kor=knowledge_name,
            knowledge_desc_kor=knowledge_desc,
            knowledge2_name_kor=knowledge2_name,
            knowledge2_desc_kor=knowledge2_desc,
            source_file=source_file,
            knowledge_source_file=knowledge_source_file,
            knowledge2_source_file=knowledge2_source_file,
        )

    log.info("Skills scanned: %d (Pass 2 hits: %d)", len(skill_lookup), pass2_hits)
    return skill_lookup


# =============================================================================
# SKILLGROUPINFO PARSING (Tab 1)
# =============================================================================

def parse_skill_groups(group_file: Path) -> List[SkillGroupEntry]:
    """Parse skillgroupinfo.staticinfo.xml.

    Each SkillGroupInfo has:
      - StrKey, GroupName
      - Child <SkillInfo Key="..."/> elements
    """
    log.info("Parsing skill groups: %s", group_file.name)
    groups: List[SkillGroupEntry] = []

    root = parse_xml_file(group_file)
    if root is None:
        log.error("Cannot parse skill group file: %s", group_file)
        return groups

    for sg_el in root.iter("SkillGroupInfo"):
        strkey = sg_el.get("StrKey") or ""
        group_name = sg_el.get("GroupName") or ""

        if not strkey:
            continue

        # Collect child SkillInfo Key references
        skill_keys: List[str] = []
        for child_skill in sg_el.findall("SkillInfo"):
            sk = child_skill.get("Key") or ""
            if sk:
                skill_keys.append(sk)

        _collect_korean_string(group_name)

        groups.append(SkillGroupEntry(
            strkey=strkey,
            group_name_kor=group_name,
            skill_keys=skill_keys,
        ))

    log.info("Skill groups parsed: %d groups, %d total skill refs",
             len(groups), sum(len(g.skill_keys) for g in groups))
    return groups


# =============================================================================
# SKILLTREEINFO PARSING (Tab 2)
# =============================================================================

def parse_skill_trees(tree_file: Path) -> List[SkillTreeEntry]:
    """Parse SkillTreeInfo.staticinfo.xml.

    Each SkillTreeInfo has:
      - Key, StrKey, CharacterKey, UIPageName
      - Child <SkillNode> elements with Id, SkillKey, ParentId
    """
    log.info("Parsing skill trees: %s", tree_file.name)
    trees: List[SkillTreeEntry] = []

    root = parse_xml_file(tree_file)
    if root is None:
        log.error("Cannot parse skill tree file: %s", tree_file)
        return trees

    for st_el in root.iter("SkillTreeInfo"):
        key = st_el.get("Key") or ""
        strkey = st_el.get("StrKey") or ""
        character_key = st_el.get("CharacterKey") or ""
        ui_page_name = st_el.get("UIPageName") or ""

        if not key:
            continue

        _collect_korean_string(ui_page_name)

        # Collect all SkillNode elements
        all_nodes: Dict[str, SkillNodeEntry] = {}
        for sn_el in st_el.iter("SkillNode"):
            node_id = sn_el.get("Id") or ""
            skill_key = sn_el.get("SkillKey") or ""
            parent_id = sn_el.get("ParentId") or ""

            if not node_id:
                continue

            all_nodes[node_id] = SkillNodeEntry(
                node_id=node_id,
                skill_key=skill_key,
                parent_id=parent_id,
            )

        # Build parent-child hierarchy
        root_nodes: List[SkillNodeEntry] = []
        for node in all_nodes.values():
            if node.parent_id and node.parent_id in all_nodes:
                all_nodes[node.parent_id].children.append(node)
            else:
                root_nodes.append(node)

        trees.append(SkillTreeEntry(
            key=key,
            strkey=strkey,
            character_key=character_key,
            ui_page_name_kor=ui_page_name,
            nodes=root_nodes,
        ))

    total_nodes = sum(len(t.nodes) for t in trees)
    log.info("Skill trees parsed: %d trees, %d root nodes", len(trees), total_nodes)
    return trees


# =============================================================================
# EXCEL WRITER (2-tab format)
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_gold_fill = PatternFill("solid", fgColor="FFD700")
_gold_font = Font(bold=True, size=11)
_fill_a = PatternFill("solid", fgColor="E2EFDA")  # Light green
_fill_b = PatternFill("solid", fgColor="FCE4D6")  # Light orange
_blue_fill = PatternFill("solid", fgColor="B4C6E7")  # Light blue (depth 1 in tree)


def _write_skill_rows(
    ws,
    excel_row: int,
    entry: NewSkillEntry,
    current_fill: PatternFill,
    pre: Dict[Tuple[str, str], Tuple[str, str]],
    group_info: str,
) -> int:
    """Write 1-6 rows for a single skill. Returns next excel_row."""

    def _write_row(data_type: str, kor_text: str, trans: str, sid: str) -> int:
        nonlocal excel_row
        vals = [data_type, group_info, kor_text, trans, "", "", "", sid]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(excel_row, ci, val)
            cell.fill = current_fill
            cell.border = THIN_BORDER
            if ci == 5:  # STATUS
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # STRINGID as text format
        ws.cell(excel_row, 8).number_format = '@'
        excel_row += 1
        return excel_row

    # 1. SkillData -- SkillName (always)
    t, s = pre.get((entry.strkey, "skill_name"), ("", ""))
    excel_row = _write_row("SkillData", entry.skill_name_kor, t, s)

    # 2. SkillData -- SkillDesc (always, even if empty)
    t, s = pre.get((entry.strkey, "skill_desc"), ("", ""))
    excel_row = _write_row("SkillData", entry.skill_desc_kor, t, s)

    # 3. KnowledgeData -- Name (skip if empty)
    if entry.knowledge_name_kor:
        t, s = pre.get((entry.strkey, "knowledge_name"), ("", ""))
        excel_row = _write_row("KnowledgeData", entry.knowledge_name_kor, t, s)

    # 4. KnowledgeData -- Desc (skip if empty)
    if entry.knowledge_desc_kor:
        t, s = pre.get((entry.strkey, "knowledge_desc"), ("", ""))
        excel_row = _write_row("KnowledgeData", entry.knowledge_desc_kor, t, s)

    # 5. KnowledgeData2 -- Name (skip if empty)
    if entry.knowledge2_name_kor:
        t, s = pre.get((entry.strkey, "knowledge2_name"), ("", ""))
        excel_row = _write_row("KnowledgeData2", entry.knowledge2_name_kor, t, s)

    # 6. KnowledgeData2 -- Desc (skip if empty)
    if entry.knowledge2_desc_kor:
        t, s = pre.get((entry.strkey, "knowledge2_desc"), ("", ""))
        excel_row = _write_row("KnowledgeData2", entry.knowledge2_desc_kor, t, s)

    return excel_row


def _write_header_row(ws, excel_row: int, data_type: str, group_info: str, kor_text: str) -> int:
    """Write a gold header row. Returns next excel_row."""
    vals = [data_type, group_info, kor_text, "", "", "", "", ""]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(excel_row, ci, val)
        cell.fill = _gold_fill
        cell.font = _gold_font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return excel_row + 1


def write_newskill_excel(
    skill_lookup: Dict[str, NewSkillEntry],
    skill_groups: List[SkillGroupEntry],
    skill_trees: List[SkillTreeEntry],
    lang_tbl: Dict[str, List[Tuple[str, str]]],
    lang_code: str,
    export_index: Dict[str, Set[str]],
    output_path: Path,
) -> None:
    """Write NewSkill Excel with 2 tabs: SkillGroup and SkillTree.

    8 columns: DataType | GroupInfo | SourceText (KR) | Translation | STATUS | COMMENT | SCREENSHOT | STRINGID
    """
    wb = Workbook()
    wb.remove(wb.active)
    code = lang_code.upper()

    headers = [
        "DataType",
        "GroupInfo",
        "SourceText (KR)",
        f"Translation ({code})",
        "STATUS",
        "COMMENT",
        "SCREENSHOT",
        "STRINGID",
    ]

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)

    # ------------------------------------------------------------------
    # PRE-RESOLVE: consume StringIDs in DOCUMENT ORDER (skill_lookup
    # preserves insertion order = XML scan order of skillinfo_pc).
    # Both tabs reference the same skills, so pre-resolve once.
    # ------------------------------------------------------------------
    pre: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for sk, entry in skill_lookup.items():
        pre[(sk, "skill_name")] = resolve_translation(
            entry.skill_name_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        pre[(sk, "skill_desc")] = resolve_translation(
            entry.skill_desc_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        if entry.knowledge_name_kor:
            pre[(sk, "knowledge_name")] = resolve_translation(
                entry.knowledge_name_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge_desc_kor:
            pre[(sk, "knowledge_desc")] = resolve_translation(
                entry.knowledge_desc_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge2_name_kor:
            pre[(sk, "knowledge2_name")] = resolve_translation(
                entry.knowledge2_name_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)
        if entry.knowledge2_desc_kor:
            pre[(sk, "knowledge2_desc")] = resolve_translation(
                entry.knowledge2_desc_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)

    if consumer.warnings:
        log.warning("StringID overruns during pre-resolve: %d", consumer.warnings)

    def _write_column_headers(ws):
        for col_idx, txt in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, txt)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    def _finalize_sheet(ws, excel_row):
        if excel_row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
        ws.freeze_panes = "A2"
        add_status_dropdown(ws, col=5)
        autofit_worksheet(ws)

    # ==================================================================
    # TAB 1: SkillGroup
    # ==================================================================
    ws1 = wb.create_sheet("SkillGroup")
    _write_column_headers(ws1)
    excel_row = 2
    current_fill = _fill_a

    for group in skill_groups:
        # Gold header row for the group
        excel_row = _write_header_row(
            ws1, excel_row, "SkillGroupHeader", group.strkey, group.group_name_kor)

        last_key = None
        for sk in group.skill_keys:
            entry = skill_lookup.get(sk)
            if not entry:
                continue

            # Alternate fill per skill
            if last_key is not None and sk != last_key:
                current_fill = _fill_b if current_fill == _fill_a else _fill_a
            last_key = sk

            excel_row = _write_skill_rows(
                ws1, excel_row, entry, current_fill, pre, group.group_name_kor)

    _finalize_sheet(ws1, excel_row)
    log.info("  Tab 'SkillGroup': %d rows", excel_row - 2)

    # ==================================================================
    # TAB 2: SkillTree
    # ==================================================================
    ws2 = wb.create_sheet("SkillTree")
    _write_column_headers(ws2)
    excel_row = 2

    # Depth-based fills for tree hierarchy
    depth_fills = [
        _blue_fill,   # Depth 0 (root nodes)
        _fill_a,      # Depth 1 (light green)
        _fill_b,      # Depth 2 (light orange)
    ]

    def _get_tree_fill(depth: int) -> PatternFill:
        if depth < len(depth_fills):
            return depth_fills[depth]
        return depth_fills[-1]

    def _write_tree_nodes(nodes: List[SkillNodeEntry], depth: int, group_info: str) -> None:
        nonlocal excel_row
        for node in nodes:
            entry = skill_lookup.get(node.skill_key)
            if entry:
                fill = _get_tree_fill(depth)
                excel_row = _write_skill_rows(
                    ws2, excel_row, entry, fill, pre, group_info)

            # Recurse into children
            if node.children:
                _write_tree_nodes(node.children, depth + 1, group_info)

    for tree in skill_trees:
        # Gold header row for the tree
        group_info = f"{tree.character_key} | {tree.ui_page_name_kor}" if tree.character_key else tree.key
        excel_row = _write_header_row(
            ws2, excel_row, "SkillTreeHeader", group_info, tree.ui_page_name_kor)

        _write_tree_nodes(tree.nodes, 0, group_info)

    _finalize_sheet(ws2, excel_row)
    log.info("  Tab 'SkillTree': %d rows", excel_row - 2)

    wb.save(output_path)
    log.info("NewSkill Excel saved: %s", output_path.name)


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_newskill_datasheets() -> Dict:
    """
    Generate NewSkill datasheets for all languages.

    Pipeline:
    1. Load language tables
    2. Load knowledge data (Name + Desc + source_file)
    3. Parse skillinfo_pc.staticinfo.xml -> skill_lookup
    4. Parse skillgroupinfo.staticinfo.xml -> skill_groups
    5. Parse SkillTreeInfo.staticinfo.xml -> skill_trees
    6. Get EXPORT index
    7. For each language: write 2-tab Excel

    Returns:
        Dict with results
    """
    result = {
        "category": "NewSkill",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()

    log.info("=" * 70)
    log.info("NEW Skill Datasheet Generator")
    log.info("=" * 70)

    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "NewSkillData_Map_All"
    output_folder.mkdir(exist_ok=True)

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    knowledge_folder = RESOURCE_FOLDER / "knowledgeinfo"
    skill_file = RESOURCE_FOLDER / "skillinfo" / "skillinfo_pc.staticinfo.xml"
    group_file = RESOURCE_FOLDER / "skillinfo" / "skillgroupinfo.staticinfo.xml"
    tree_file = RESOURCE_FOLDER / "skillinfo" / "SkillTreeInfo.staticinfo.xml"

    if not skill_file.exists():
        result["errors"].append(f"Skill file not found: {skill_file}")
        log.error("Skill file not found: %s", skill_file)
        return result

    try:
        # 1. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        # 2. Load knowledge data (map + name index for Pass 2)
        knowledge_map, knowledge_name_index = load_knowledge_data(knowledge_folder)

        # 3. Parse skillinfo_pc -> skill_lookup
        skill_lookup = scan_skills_with_knowledge(
            skill_file, knowledge_map, knowledge_name_index)

        if not skill_lookup:
            result["errors"].append("No skill data found!")
            log.warning("No skill data found!")
            return result

        # 4. Parse skillgroupinfo -> skill_groups
        skill_groups: List[SkillGroupEntry] = []
        if group_file.exists():
            skill_groups = parse_skill_groups(group_file)
        else:
            log.warning("SkillGroupInfo file not found: %s", group_file)

        # 5. Parse SkillTreeInfo -> skill_trees
        skill_trees: List[SkillTreeEntry] = []
        if tree_file.exists():
            skill_trees = parse_skill_trees(tree_file)
        else:
            log.warning("SkillTreeInfo file not found: %s", tree_file)

        # 6. Get EXPORT index
        export_index = get_export_index()

        # 7. Generate Excel per language
        log.info("Processing languages...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())
            excel_path = output_folder / f"NewSkill_LQA_{code.upper()}.xlsx"
            write_newskill_excel(
                skill_lookup, skill_groups, skill_trees,
                tbl, code, export_index, excel_path,
            )
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in NewSkill generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_newskill_datasheets()
    print(f"\nResult: {result}")
