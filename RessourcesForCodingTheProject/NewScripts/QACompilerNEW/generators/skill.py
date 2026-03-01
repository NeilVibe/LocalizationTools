"""
Skill Datasheet Generator
=========================
Row-per-text skill datasheet ordered by UIPositionXY (screen reading order).

Each skill with LearnKnowledgeKey produces up to 6 rows:
  1. SkillData   — SkillName
  2. SkillData   — SkillDesc
  3. KnowledgeData  — Name  (Pass 1: LearnKnowledgeKey direct lookup)
  4. KnowledgeData  — Desc  (Pass 1)
  5. KnowledgeData2 — Name  (Pass 2: identical SkillName == KnowledgeInfo.Name)
  6. KnowledgeData2 — Desc  (Pass 2)

Skills are grouped by SkillTreeInfo and sorted by UIPositionXY within each tree
(reading order: Y ascending, X ascending on the skill tree UI page).
Skills not in any tree are collected at the end under "Other Skills".

Key features:
- Gold header rows per skill tree (character + weapon page)
- UIPositionXY ordering within each tree for tester-friendly reading order
- Alternating fill per skill
- Reuses knowledge loading from newitem.py
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT
from generators.base import (
    get_logger,
    parse_xml_file,
    load_language_tables,
    normalize_placeholders,
    br_to_newline,
    autofit_worksheet,
    THIN_BORDER,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
    add_status_dropdown,
    iter_xml_files,
)
from generators.newitem import load_knowledge_data

log = get_logger("SkillGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class SkillEntry:
    """Complete data for a single skill with knowledge fields."""
    key: str                        # SkillInfo Key
    strkey: str                     # SkillInfo StrKey
    skill_name_kor: str             # SkillInfo.SkillName
    skill_desc_kor: str             # SkillInfo.SkillDesc
    learn_knowledge_key: str        # SkillInfo.LearnKnowledgeKey
    knowledge_name_kor: str         # Pass 1: KnowledgeInfo.Name
    knowledge_desc_kor: str         # Pass 1: KnowledgeInfo.Desc
    knowledge2_name_kor: str        # Pass 2: identical name match
    knowledge2_desc_kor: str        # Pass 2: desc
    source_file: str                # skillinfo_pc filename
    knowledge_source_file: str      # Knowledge XML filename
    knowledge2_source_file: str = ""  # Pass 2 source file


@dataclass
class KnowledgeChild:
    """A child KnowledgeInfo nested under a parent KnowledgeInfo element."""
    strkey: str        # Child KnowledgeInfo StrKey
    name_kor: str      # Child KnowledgeInfo Name
    desc_kor: str      # Child KnowledgeInfo Desc
    source_file: str   # Knowledge XML filename


@dataclass
class SkillNodeEntry:
    """A single SkillNode within a SkillTreeInfo."""
    node_id: str
    skill_key: str
    parent_id: str
    ui_xy: Optional[Tuple[int, int]] = None  # (x, y) from UIPositionXY
    ui_position: str = ""                     # grid "col_row" (fallback only)


@dataclass
class SkillTreeEntry:
    """One SkillTreeInfo element with its node hierarchy."""
    key: str
    strkey: str
    character_key: str
    ui_page_name_kor: str
    nodes: List[SkillNodeEntry] = field(default_factory=list)


def _parse_grid_sort_key(ui_position: str) -> Tuple[int, int]:
    """Parse UIPosition grid "col_row" into (row, col) for fallback sorting."""
    if ui_position and "_" in ui_position:
        parts = ui_position.split("_")
        try:
            return (int(parts[1]), int(parts[0]))  # (row, col)
        except (ValueError, IndexError):
            pass
    return (9999, 9999)


# =============================================================================
# SKILL EXTRACTION
# =============================================================================

def scan_skills_with_knowledge(
    skill_file: Path,
    knowledge_map: Dict[str, Tuple[str, str, str]],
    knowledge_name_index: Dict[str, List[Tuple[str, str, str]]],
) -> Dict[str, SkillEntry]:
    """Parse skillinfo_pc.staticinfo.xml and build skill lookup with knowledge.

    Pass 1: LearnKnowledgeKey -> knowledge_map (direct key lookup)
    Pass 2: SkillName -> knowledge_name_index (identical name match)

    IMPORTANT: Lookup is keyed by StrKey (e.g. "Skill_Wrestle_AirBodySlam"),
    NOT by the numeric Key (e.g. "15004"). SkillTreeInfo references skills
    using StrKey values in their child SkillNode elements.

    Returns:
        skill_lookup: {StrKey: SkillEntry} in document order
    """
    log.info("Scanning skills with knowledge: %s", skill_file.name)
    skill_lookup: Dict[str, SkillEntry] = {}
    pass2_hits = 0

    root = parse_xml_file(skill_file)
    if root is None:
        log.error("Cannot parse skill file: %s", skill_file)
        return skill_lookup

    source_file = skill_file.name

    for el in root.iter("SkillInfo"):
        key = el.get("Key") or ""
        strkey = el.get("StrKey") or ""
        skill_name = el.get("SkillName") or ""
        skill_desc = el.get("SkillDesc") or ""
        learn_knowledge_key = el.get("LearnKnowledgeKey") or ""

        if not strkey or not skill_name:
            continue

        # Dedup by StrKey (case-insensitive — tree SkillKey may differ in case)
        lookup_key = strkey.lower()
        if lookup_key in skill_lookup:
            continue

        # Pass 1: Resolve knowledge data via LearnKnowledgeKey
        knowledge_name = ""
        knowledge_desc = ""
        knowledge_source_file = ""
        pass1_strkey = ""
        if learn_knowledge_key and learn_knowledge_key.lower() in knowledge_map:
            knowledge_name, knowledge_desc, knowledge_source_file = knowledge_map[learn_knowledge_key.lower()]
            pass1_strkey = learn_knowledge_key.lower()

        # Pass 2: Identical name match (SkillName == KnowledgeInfo.Name)
        knowledge2_name = ""
        knowledge2_desc = ""
        knowledge2_source_file = ""
        if skill_name and skill_name in knowledge_name_index:
            for kn_strkey, kn_desc, kn_src in knowledge_name_index[skill_name]:
                if kn_strkey != pass1_strkey:
                    knowledge2_name = skill_name
                    knowledge2_desc = kn_desc
                    knowledge2_source_file = kn_src
                    pass2_hits += 1
                    break

        # Collect Korean strings for coverage tracking
        _collect_korean_string(skill_name)
        _collect_korean_string(skill_desc)
        _collect_korean_string(knowledge_name)
        _collect_korean_string(knowledge_desc)
        _collect_korean_string(knowledge2_name)
        _collect_korean_string(knowledge2_desc)

        skill_lookup[lookup_key] = SkillEntry(
            key=key,
            strkey=strkey,
            skill_name_kor=skill_name,
            skill_desc_kor=skill_desc,
            learn_knowledge_key=learn_knowledge_key,
            knowledge_name_kor=knowledge_name,
            knowledge_desc_kor=knowledge_desc,
            knowledge2_name_kor=knowledge2_name,
            knowledge2_desc_kor=knowledge2_desc,
            source_file=source_file,
            knowledge_source_file=knowledge_source_file,
            knowledge2_source_file=knowledge2_source_file,
        )

    log.info("Skills scanned: %d (Pass 2 hits: %d)", len(skill_lookup), pass2_hits)
    return skill_lookup


# =============================================================================
# KNOWLEDGE CHILDREN MAP (sub-skill nesting)
# =============================================================================

def build_knowledge_children_map(
    knowledge_folder: Path,
) -> Dict[str, List[KnowledgeChild]]:
    """Build a map of parent KnowledgeInfo StrKey -> direct child KnowledgeInfo elements.

    Parses knowledge XMLs and checks DIRECT children only (not recursive .iter())
    for nested KnowledgeInfo elements.

    Returns:
        {parent_strkey.lower(): [KnowledgeChild, ...]}
    """
    children_map: Dict[str, List[KnowledgeChild]] = {}

    if not knowledge_folder.exists():
        log.warning("Knowledge folder does not exist for children map: %s", knowledge_folder)
        return children_map

    file_count = 0
    child_count = 0
    for path in iter_xml_files(knowledge_folder):
        root = parse_xml_file(path)
        if root is None:
            continue
        file_count += 1
        source_file = path.name

        for el in root.iter("KnowledgeInfo"):
            parent_strkey = el.get("StrKey") or ""
            if not parent_strkey:
                continue

            # Direct children only — NOT el.iter() which would recurse
            kids: List[KnowledgeChild] = []
            for ch in el:
                if ch.tag == "KnowledgeInfo":
                    ch_strkey = ch.get("StrKey") or ""
                    ch_name = ch.get("Name") or ""
                    ch_desc = ch.get("Desc") or ""
                    if ch_strkey:
                        kids.append(KnowledgeChild(
                            strkey=ch_strkey,
                            name_kor=ch_name,
                            desc_kor=ch_desc,
                            source_file=source_file,
                        ))
                        # Collect Korean strings for coverage
                        _collect_korean_string(ch_name)
                        _collect_korean_string(ch_desc)

            if kids:
                children_map[parent_strkey.lower()] = kids
                child_count += len(kids)

    log.info("Knowledge children map: %d parents with %d children from %d files",
             len(children_map), child_count, file_count)
    return children_map


def build_skill_by_knowledge_map(
    skill_lookup: Dict[str, SkillEntry],
) -> Dict[str, str]:
    """Reverse map: {learn_knowledge_key.lower(): skill_strkey.lower()}.

    Used to check if a child KnowledgeInfo has a corresponding SkillInfo.
    """
    result: Dict[str, str] = {}
    for sk, entry in skill_lookup.items():
        if entry.learn_knowledge_key:
            result[entry.learn_knowledge_key.lower()] = sk
    return result


def _discover_claimed_skills(
    skill_lookup: Dict[str, SkillEntry],
    knowledge_children_map: Dict[str, List[KnowledgeChild]],
    skill_by_knowledge: Dict[str, str],
) -> Set[str]:
    """Pre-compute all skill keys that will appear as sub-skills of another skill.

    These get skipped during tree walking and orphan calculation to prevent
    double output. Recursive — handles multi-level nesting.

    Returns:
        Set of skill_strkey.lower() values that are claimed as sub-skills.
    """
    claimed: Set[str] = set()
    visited: Set[str] = set()

    def _walk(knowledge_key: str) -> None:
        """Recursively discover claimed skills starting from a knowledge key."""
        kk_lower = knowledge_key.lower()
        if kk_lower in visited:
            return
        visited.add(kk_lower)

        children = knowledge_children_map.get(kk_lower, [])
        for child in children:
            # If child KnowledgeInfo has a SkillInfo -> that skill is claimed
            sub_skill_key = skill_by_knowledge.get(child.strkey.lower())
            if sub_skill_key and sub_skill_key in skill_lookup:
                claimed.add(sub_skill_key)
                # Recurse: the sub-skill may have its own children
                sub_entry = skill_lookup[sub_skill_key]
                if sub_entry.learn_knowledge_key:
                    _walk(sub_entry.learn_knowledge_key)
            # Also recurse into knowledge-only children (they may have sub-children)
            _walk(child.strkey)

    # Start from every skill's LearnKnowledgeKey
    for sk, entry in skill_lookup.items():
        if entry.learn_knowledge_key:
            _walk(entry.learn_knowledge_key)

    if claimed:
        log.info("Claimed sub-skills (will nest under parents): %d", len(claimed))
    return claimed


# =============================================================================
# SKILLTREEINFO PARSING (with UIPosition)
# =============================================================================

def parse_skill_trees(tree_file: Path) -> List[SkillTreeEntry]:
    """Parse SkillTreeInfo.staticinfo.xml.

    Each SkillTreeInfo has:
      - Key, StrKey, CharacterKey, UIPageName
      - Child <SkillNode> elements with Id, SkillKey, ParentId, UIPositionXY, UIPosition

    UIPositionXY format: "X Y" pixel coordinates (e.g. "640 30").
    Used for sorting skills in screen reading order (Y ascending, X ascending).
    UIPosition grid format "col_row" is kept as fallback for nodes missing XY.
    """
    log.info("Parsing skill trees: %s", tree_file.name)
    trees: List[SkillTreeEntry] = []

    root = parse_xml_file(tree_file)
    if root is None:
        log.error("Cannot parse skill tree file: %s", tree_file)
        return trees

    for st_el in root.iter("SkillTreeInfo"):
        key = st_el.get("Key") or ""
        strkey = st_el.get("StrKey") or ""
        character_key = st_el.get("CharacterKey") or ""
        ui_page_name = st_el.get("UIPageName") or ""

        if not key:
            continue

        _collect_korean_string(ui_page_name)

        # Collect all SkillNode elements (flat list, no hierarchy needed)
        all_nodes: List[SkillNodeEntry] = []
        for sn_el in st_el.iter("SkillNode"):
            node_id = sn_el.get("Id") or ""
            skill_key = sn_el.get("SkillKey") or ""
            parent_id = sn_el.get("ParentId") or ""

            if not node_id:
                continue

            # Parse UIPositionXY "X Y" → (x, y) pixel coordinates
            ui_xy_str = sn_el.get("UIPositionXY") or ""
            ui_xy = None
            if ui_xy_str:
                xy_parts = ui_xy_str.split()
                try:
                    ui_xy = (int(xy_parts[0]), int(xy_parts[1]))
                except (ValueError, IndexError):
                    pass

            ui_position = sn_el.get("UIPosition") or ""

            all_nodes.append(SkillNodeEntry(
                node_id=node_id,
                skill_key=skill_key,
                parent_id=parent_id,
                ui_xy=ui_xy,
                ui_position=ui_position,
            ))

        # Sort by UIPositionXY (Y asc, X asc); grid fallback for nodes missing XY
        with_xy = [n for n in all_nodes if n.ui_xy is not None]
        without_xy = [n for n in all_nodes if n.ui_xy is None]
        with_xy.sort(key=lambda n: (n.ui_xy[1], n.ui_xy[0]))
        without_xy.sort(key=lambda n: _parse_grid_sort_key(n.ui_position))
        all_nodes = with_xy + without_xy

        trees.append(SkillTreeEntry(
            key=key,
            strkey=strkey,
            character_key=character_key,
            ui_page_name_kor=ui_page_name,
            nodes=all_nodes,
        ))

    total_nodes = sum(len(t.nodes) for t in trees)
    log.info("Skill trees parsed: %d trees, %d total nodes", len(trees), total_nodes)
    return trees


# =============================================================================
# EXCEL WRITER (single sheet: Skills)
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_gold_fill = PatternFill("solid", fgColor="FFD700")
_gold_font = Font(bold=True, size=11)
_fill_a = PatternFill("solid", fgColor="E2EFDA")  # Light green
_fill_b = PatternFill("solid", fgColor="FCE4D6")  # Light orange


def _write_skill_rows(
    ws,
    excel_row: int,
    entry: SkillEntry,
    current_fill: PatternFill,
    pre: Dict[Tuple[str, str], Tuple[str, str]],
    group_info: str,
    knowledge_children_map: Optional[Dict[str, List[KnowledgeChild]]] = None,
    skill_by_knowledge: Optional[Dict[str, str]] = None,
    skill_lookup: Optional[Dict[str, SkillEntry]] = None,
    claimed_skills: Optional[Set[str]] = None,
    _visited: Optional[Set[str]] = None,
    depth: int = 0,
) -> int:
    """Write 1-6 rows for a single skill, then sub-skills. Returns next excel_row.

    Sub-skills are written immediately after the parent skill's rows:
    - If a child KnowledgeInfo has a SkillInfo -> recursive call (SkillData rows)
    - If a child KnowledgeInfo has no SkillInfo -> SubKnowledgeData rows (Name + Desc)

    depth controls Excel cell indentation (0 = top-level, 1+ = sub-skills).
    """
    if _visited is None:
        _visited = set()

    # Cycle protection
    sk_lower = entry.strkey.lower()
    if sk_lower in _visited:
        return excel_row
    _visited.add(sk_lower)

    def _write_row(data_type: str, kor_text: str, trans: str, sid: str, indent: int = depth) -> int:
        nonlocal excel_row
        vals = [data_type, group_info, br_to_newline(kor_text), br_to_newline(trans), "", "", "", sid]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(excel_row, ci, val)
            cell.fill = current_fill
            cell.border = THIN_BORDER
            if ci == 5:  # STATUS
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (3, 4):  # SourceText + Translation — apply indent
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True, indent=indent)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # STRINGID as text format
        ws.cell(excel_row, 8).number_format = '@'
        excel_row += 1
        return excel_row

    # 1. SkillData -- SkillName (always)
    t, s = pre.get((sk_lower, "skill_name"), ("", ""))
    excel_row = _write_row("SkillData", entry.skill_name_kor, t, s)

    # 2. SkillData -- SkillDesc (always, even if empty)
    t, s = pre.get((sk_lower, "skill_desc"), ("", ""))
    excel_row = _write_row("SkillData", entry.skill_desc_kor, t, s)

    # 3. KnowledgeData -- Name (skip if empty)
    if entry.knowledge_name_kor:
        t, s = pre.get((sk_lower, "knowledge_name"), ("", ""))
        excel_row = _write_row("KnowledgeData", entry.knowledge_name_kor, t, s)

    # 4. KnowledgeData -- Desc (skip if empty)
    if entry.knowledge_desc_kor:
        t, s = pre.get((sk_lower, "knowledge_desc"), ("", ""))
        excel_row = _write_row("KnowledgeData", entry.knowledge_desc_kor, t, s)

    # 5. KnowledgeData2 -- Name (skip if empty)
    if entry.knowledge2_name_kor:
        t, s = pre.get((sk_lower, "knowledge2_name"), ("", ""))
        excel_row = _write_row("KnowledgeData2", entry.knowledge2_name_kor, t, s)

    # 6. KnowledgeData2 -- Desc (skip if empty)
    if entry.knowledge2_desc_kor:
        t, s = pre.get((sk_lower, "knowledge2_desc"), ("", ""))
        excel_row = _write_row("KnowledgeData2", entry.knowledge2_desc_kor, t, s)

    # ------------------------------------------------------------------
    # SUB-SKILLS: Write children of this skill's LearnKnowledgeKey
    # ------------------------------------------------------------------
    if (knowledge_children_map and skill_by_knowledge is not None
            and skill_lookup and entry.learn_knowledge_key):
        children = knowledge_children_map.get(entry.learn_knowledge_key.lower(), [])
        for child in children:
            sub_skill_key = skill_by_knowledge.get(child.strkey.lower())
            if sub_skill_key and sub_skill_key in skill_lookup:
                # Child has a SkillInfo -> recursive call (same fill, same group, depth+1)
                sub_entry = skill_lookup[sub_skill_key]
                excel_row = _write_skill_rows(
                    ws, excel_row, sub_entry, current_fill, pre, group_info,
                    knowledge_children_map, skill_by_knowledge,
                    skill_lookup, claimed_skills, _visited,
                    depth=depth + 1,
                )
            else:
                # Knowledge-only child -> SubKnowledgeData rows (depth+1)
                if child.name_kor:
                    t, s = pre.get((child.strkey.lower(), "kchild_name"), ("", ""))
                    excel_row = _write_row("SubKnowledgeData", child.name_kor, t, s, indent=depth + 1)
                if child.desc_kor:
                    t, s = pre.get((child.strkey.lower(), "kchild_desc"), ("", ""))
                    excel_row = _write_row("SubKnowledgeData", child.desc_kor, t, s, indent=depth + 1)

    return excel_row


def _write_header_row(ws, excel_row: int, data_type: str, group_info: str, kor_text: str) -> int:
    """Write a gold header row. Returns next excel_row."""
    vals = [data_type, group_info, br_to_newline(kor_text), "", "", "", "", ""]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(excel_row, ci, val)
        cell.fill = _gold_fill
        cell.font = _gold_font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return excel_row + 1


def write_skill_excel(
    skill_lookup: Dict[str, SkillEntry],
    skill_trees: List[SkillTreeEntry],
    lang_tbl: Dict[str, List[Tuple[str, str]]],
    lang_code: str,
    export_index: Dict[str, Set[str]],
    output_path: Path,
    knowledge_children_map: Optional[Dict[str, List[KnowledgeChild]]] = None,
) -> None:
    """Write Skill Excel with single "Skills" sheet, UIPositionXY ordered.

    8 columns: DataType | GroupInfo | SourceText (KR) | Translation | STATUS | COMMENT | SCREENSHOT | STRINGID

    Skills are grouped by SkillTreeInfo with gold headers. Within each tree,
    skills are sorted by UIPositionXY pixel coordinates (Y asc, X asc = screen reading order).
    Skills not in any tree appear at the end under "Other Skills".

    Sub-skills are nested under their parent skill via knowledge children mapping.
    """
    if knowledge_children_map is None:
        knowledge_children_map = {}

    # Build sub-skill support structures
    skill_by_knowledge = build_skill_by_knowledge_map(skill_lookup)
    claimed_skills = _discover_claimed_skills(
        skill_lookup, knowledge_children_map, skill_by_knowledge)

    wb = Workbook()
    wb.remove(wb.active)
    code = lang_code.upper()

    headers = [
        "DataType",
        "GroupInfo",
        "SourceText (KR)",
        f"Translation ({code})",
        "STATUS",
        "COMMENT",
        "SCREENSHOT",
        "STRINGID",
    ]

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)

    # ------------------------------------------------------------------
    # PRE-RESOLVE: consume StringIDs in DOCUMENT ORDER (skill_lookup
    # preserves insertion order = XML scan order of skillinfo_pc).
    # ------------------------------------------------------------------
    pre: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for sk, entry in skill_lookup.items():
        pre[(sk, "skill_name")] = resolve_translation(
            entry.skill_name_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        pre[(sk, "skill_desc")] = resolve_translation(
            entry.skill_desc_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        if entry.knowledge_name_kor:
            pre[(sk, "knowledge_name")] = resolve_translation(
                entry.knowledge_name_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge_desc_kor:
            pre[(sk, "knowledge_desc")] = resolve_translation(
                entry.knowledge_desc_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge2_name_kor:
            pre[(sk, "knowledge2_name")] = resolve_translation(
                entry.knowledge2_name_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)
        if entry.knowledge2_desc_kor:
            pre[(sk, "knowledge2_desc")] = resolve_translation(
                entry.knowledge2_desc_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)

    # Pre-resolve knowledge-only children (children without a SkillInfo)
    seen_children: set = set()
    for children in knowledge_children_map.values():
        for child in children:
            ck_lower = child.strkey.lower()
            if ck_lower in seen_children:
                continue  # Dedup: same child under multiple parents
            seen_children.add(ck_lower)
            if ck_lower not in skill_by_knowledge:
                # Knowledge-only child — resolve its Name and Desc
                if child.name_kor:
                    pre[(ck_lower, "kchild_name")] = resolve_translation(
                        child.name_kor, lang_tbl, child.source_file, export_index, consumer=consumer)
                if child.desc_kor:
                    pre[(ck_lower, "kchild_desc")] = resolve_translation(
                        child.desc_kor, lang_tbl, child.source_file, export_index, consumer=consumer)

    if consumer.warnings:
        log.warning("StringID overruns during pre-resolve: %d", consumer.warnings)

    def _write_column_headers(ws):
        for col_idx, txt in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, txt)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    def _finalize_sheet(ws, excel_row):
        if excel_row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
        ws.freeze_panes = "A2"
        add_status_dropdown(ws, col=5)
        autofit_worksheet(ws)

    # ==================================================================
    # Single sheet: Skills (UIPosition ordered)
    # ==================================================================
    ws = wb.create_sheet("Skills")
    _write_column_headers(ws)
    excel_row = 2
    current_fill = _fill_a

    # Track which skills have been written (to find orphans)
    written_skills: Set[str] = set()
    # Shared visited set — prevents diamond-shaped duplicates where two
    # top-level skills share a common sub-skill
    visited_skills: Set[str] = set()

    # Write skills grouped by SkillTreeInfo, ordered by UIPosition
    for tree in skill_trees:
        # Gold header row for the tree
        group_info = f"{tree.character_key} | {tree.ui_page_name_kor}" if tree.character_key else tree.key
        excel_row = _write_header_row(
            ws, excel_row, "SkillTreeHeader", group_info, tree.ui_page_name_kor)

        # Nodes are already sorted by (Y, X) from parse_skill_trees
        for node in tree.nodes:
            sk_lower = node.skill_key.lower()
            entry = skill_lookup.get(sk_lower)
            if not entry:
                continue
            if not entry.learn_knowledge_key:
                continue  # Only include skills with LearnKnowledgeKey
            if sk_lower in claimed_skills:
                continue  # Skip — will appear nested under its parent

            # Alternate fill per skill
            current_fill = _fill_b if current_fill == _fill_a else _fill_a

            excel_row = _write_skill_rows(
                ws, excel_row, entry, current_fill, pre, group_info,
                knowledge_children_map, skill_by_knowledge,
                skill_lookup, claimed_skills, visited_skills,
            )
            written_skills.add(sk_lower)

    # Orphaned skills: have LearnKnowledgeKey but aren't in any tree
    # Use visited_skills (true record of what was written) instead of
    # written_skills to catch sub-skills that were recursively written
    orphans = [
        entry for sk, entry in skill_lookup.items()
        if entry.learn_knowledge_key
        and sk not in visited_skills
        and sk not in claimed_skills
    ]

    if orphans:
        excel_row = _write_header_row(
            ws, excel_row, "SkillTreeHeader", "Other Skills", "기타 스킬")

        for entry in orphans:
            current_fill = _fill_b if current_fill == _fill_a else _fill_a
            excel_row = _write_skill_rows(
                ws, excel_row, entry, current_fill, pre, "Other Skills",
                knowledge_children_map, skill_by_knowledge,
                skill_lookup, claimed_skills, visited_skills,
            )

    _finalize_sheet(ws, excel_row)
    log.info("  Sheet 'Skills': %d rows", excel_row - 2)

    wb.save(output_path)
    log.info("Skill Excel saved: %s", output_path.name)


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_skill_datasheets() -> Dict:
    """
    Generate Skill datasheets for all languages.

    Pipeline:
    1. Load language tables
    2. Load knowledge data (Name + Desc + source_file)
    3. Parse skillinfo_pc.staticinfo.xml -> skill_lookup
    4. Parse SkillTreeInfo.staticinfo.xml -> skill_trees (with UIPosition)
    5. Get EXPORT index
    6. For each language: write single-sheet Excel (UIPosition ordered)

    Returns:
        Dict with results
    """
    result = {
        "category": "Skill",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()

    log.info("=" * 70)
    log.info("Skill Datasheet Generator")
    log.info("=" * 70)

    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "Skill_LQA_All"
    output_folder.mkdir(exist_ok=True)

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    knowledge_folder = RESOURCE_FOLDER / "knowledgeinfo"
    skill_file = RESOURCE_FOLDER / "skillinfo" / "skillinfo_pc.staticinfo.xml"
    tree_file = RESOURCE_FOLDER / "skillinfo" / "SkillTreeInfo.staticinfo.xml"

    if not skill_file.exists():
        result["errors"].append(f"Skill file not found: {skill_file}")
        log.error("Skill file not found: %s", skill_file)
        return result

    try:
        # 1. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        # 2. Load knowledge data (map + name index for Pass 2)
        knowledge_map, knowledge_name_index = load_knowledge_data(knowledge_folder)

        # 2b. Build knowledge children map (for sub-skill nesting)
        knowledge_children_map = build_knowledge_children_map(knowledge_folder)

        # 3. Parse skillinfo_pc -> skill_lookup
        skill_lookup = scan_skills_with_knowledge(
            skill_file, knowledge_map, knowledge_name_index)

        if not skill_lookup:
            result["errors"].append("No skill data found!")
            log.warning("No skill data found!")
            return result

        # 4. Parse SkillTreeInfo -> skill_trees (with UIPosition)
        skill_trees: List[SkillTreeEntry] = []
        if tree_file.exists():
            skill_trees = parse_skill_trees(tree_file)
        else:
            log.warning("SkillTreeInfo file not found: %s", tree_file)

        # 5. Get EXPORT index
        export_index = get_export_index()

        # 6. Generate Excel per language
        log.info("Processing languages...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())
            excel_path = output_folder / f"Skill_LQA_{code.upper()}.xlsx"
            write_skill_excel(
                skill_lookup, skill_trees,
                tbl, code, export_index, excel_path,
                knowledge_children_map,
            )
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Skill generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_skill_datasheets()
    print(f"\nResult: {result}")
