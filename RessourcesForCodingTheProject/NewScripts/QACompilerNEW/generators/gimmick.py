"""
Gimmick Datasheet Generator (Revamped)
=======================================
Extracts Gimmick data from StaticInfo/gimmickinfo/ with folder-based organization.

Structure:
  ONE TAB = ONE FOLDER (e.g., "Background", "Item")
  Within each tab:
    Depth 0: GimmickGroupInfo (Name or StrKey) — filtered out if all children have no GimmickName
    Depth 1: GimmickInfo entries — filtered out if no GimmickName
    Depth 2: SealData.Desc (description row)

Columns:
  DataType | GroupInfo | GimmickName(KOR) | GimmickName(LOC) | SealDesc(KOR) | SealDesc(LOC) |
  Command | WorldPosition | STATUS | COMMENT | STRINGID | SCREENSHOT

Command format: /create gimmick {GimmickInfo.StrKey}
"""

from __future__ import annotations

import re
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    load_language_tables,
    normalize_placeholders,
    br_to_newline,
    autofit_worksheet,
    THIN_BORDER,
    get_first_translation,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
)

log = get_logger("GimmickGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class GimmickInfoEntry:
    """A single GimmickInfo element with optional SealData description."""
    strkey: str
    gimmick_name: str           # GimmickName attribute (KOR)
    seal_desc: str = ""         # SealData.Desc (KOR, optional)
    source_file: str = ""
    order_index: int = 0


@dataclass
class GimmickGroupEntry:
    """A GimmickGroupInfo element containing GimmickInfo entries."""
    strkey: str
    group_name: str             # GimmickGroupInfo name (GimmickName or StrKey fallback)
    source_file: str = ""
    gimmick_infos: List[GimmickInfoEntry] = field(default_factory=list)


@dataclass
class GimmickFileData:
    """All gimmick data from a single XML file."""
    filename: str               # e.g., "GimmickInfo_Background_Door.StaticInfo.xml"
    groups: List[GimmickGroupEntry] = field(default_factory=list)


# =============================================================================
# STYLING
# =============================================================================

STYLES = {
    "GroupInfo": {
        "fill": PatternFill("solid", fgColor="4472C4"),  # Dark blue
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "row_height": 35,
    },
    "GimmickInfo": {
        "fill": PatternFill("solid", fgColor="B4C6E7"),  # Light blue
        "font": Font(bold=True),
        "row_height": None,
    },
    "Description": {
        "fill": PatternFill("solid", fgColor="F5F5F5"),  # Light gray
        "font": Font(italic=False),
        "row_height": None,
    },
}

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_default_font = Font()
_default_fill = PatternFill("solid", fgColor="FFFFFF")


def get_style(style_type: str) -> Tuple[PatternFill, Font, Optional[float]]:
    """Get style for a row type. Returns (fill, font, row_height)."""
    style = STYLES.get(style_type, {})
    return (
        style.get("fill", _default_fill),
        style.get("font", _default_font),
        style.get("row_height"),
    )


# =============================================================================
# GIMMICK EXTRACTION
# =============================================================================

def _extract_gimmick_infos(
    parent_el,
    source_file: str,
    global_seen: Set[str],
    counter: List[int],
) -> List[GimmickInfoEntry]:
    """Recursively extract GimmickInfo elements from a parent element.

    Walks through GimmickAttributeGroup nesting transparently —
    only GimmickGroupInfo and GimmickInfo matter for output.
    """
    entries: List[GimmickInfoEntry] = []

    for child in parent_el:
        if child.tag == "GimmickInfo":
            strkey = child.get("StrKey") or ""
            if not strkey:
                continue

            # Dedup
            if strkey.lower() in global_seen:
                continue
            global_seen.add(strkey.lower())

            gimmick_name = child.get("GimmickName") or ""
            # Filter: skip entries with no GimmickName (no source text = nothing to translate)
            if not gimmick_name:
                continue

            # Look for SealData.Desc
            seal_desc = ""
            seal_el = child.find("SealData")
            if seal_el is not None:
                seal_desc = seal_el.get("Desc") or ""

            # Collect Korean strings
            _collect_korean_string(gimmick_name)
            if seal_desc:
                _collect_korean_string(seal_desc)

            counter[0] += 1
            entries.append(GimmickInfoEntry(
                strkey=strkey,
                gimmick_name=gimmick_name,
                seal_desc=seal_desc,
                source_file=source_file,
                order_index=counter[0],
            ))

        elif child.tag == "GimmickAttributeGroup":
            # Recurse into GimmickAttributeGroup (transparent nesting)
            entries.extend(_extract_gimmick_infos(child, source_file, global_seen, counter))

    return entries


def index_gimmick_folder(
    gimmick_folder: Path,
    global_seen: Set[str],
) -> Dict[str, List[GimmickFileData]]:
    """
    Index all gimmick data organized by subfolder.

    Returns:
        Dict mapping folder_name → [GimmickFileData, ...]
    """
    log.info("Indexing gimmick data from: %s", gimmick_folder)

    if not gimmick_folder.exists():
        log.warning("Gimmick folder not found: %s", gimmick_folder)
        return {}

    folder_data: Dict[str, List[GimmickFileData]] = OrderedDict()
    total_groups = 0
    total_gimmicks = 0
    counter: List[int] = [0]  # Mutable counter for ordering

    # Walk subfolders in sorted order
    for subfolder in sorted(gimmick_folder.iterdir()):
        if not subfolder.is_dir():
            continue

        folder_name = subfolder.name
        file_list: List[GimmickFileData] = []

        # Parse all XML files in this subfolder
        for xml_path in sorted(subfolder.iterdir()):
            if not xml_path.is_file():
                continue
            if not xml_path.name.lower().endswith(".xml"):
                continue

            root_el = parse_xml_file(xml_path)
            if root_el is None:
                continue

            source_file = xml_path.name

            file_data = GimmickFileData(
                filename=source_file,
            )

            # Find GimmickGroupInfo elements
            for group_el in root_el.iter("GimmickGroupInfo"):
                group_strkey = group_el.get("StrKey") or ""
                # Use GimmickGroupInfo's name attributes, fallback to StrKey
                group_name = group_el.get("GimmickName") or group_el.get("Name") or group_strkey

                if not group_name:
                    continue

                _collect_korean_string(group_name)

                group_entry = GimmickGroupEntry(
                    strkey=group_strkey,
                    group_name=group_name,
                    source_file=source_file,
                )

                # Extract GimmickInfo entries (recursing through GimmickAttributeGroup)
                group_entry.gimmick_infos = _extract_gimmick_infos(
                    group_el, source_file, global_seen, counter
                )

                if group_entry.gimmick_infos:
                    file_data.groups.append(group_entry)
                    total_groups += 1
                    total_gimmicks += len(group_entry.gimmick_infos)

            if file_data.groups:
                file_list.append(file_data)

        if file_list:
            folder_data[folder_name] = file_list

    log.info("  → %d folders, %d groups, %d gimmicks",
             len(folder_data), total_groups, total_gimmicks)
    return folder_data


# =============================================================================
# ROW GENERATION
# =============================================================================

# Row format: (depth, text, style_type, is_description, source_file, data_type, command, group_name)
# command: /create gimmick X (on GimmickInfo + Desc rows)
# group_name: repeated on EVERY row within a group for Excel filtering
RowItem = Tuple[int, str, str, bool, str, str, str, str]


def emit_gimmick_rows(file_list: List[GimmickFileData]) -> List[RowItem]:
    """Generate rows for one folder tab.

    No FileRoot rows — tab name is sufficient.
    Groups with zero valid children are excluded entirely.
    """
    rows: List[RowItem] = []

    for file_data in file_list:
        for group in file_data.groups:
            gname = group.group_name

            # Build children rows first to check if any exist
            child_rows: List[RowItem] = []
            for ginfo in group.gimmick_infos:
                command = f"/create gimmick {ginfo.strkey}"
                child_rows.append((1, ginfo.gimmick_name, "GimmickInfo", False, ginfo.source_file, "GimmickInfo", command, gname))

                if ginfo.seal_desc:
                    child_rows.append((2, ginfo.seal_desc, "Description", True, ginfo.source_file, "SealData.Desc", command, gname))

            # Only emit group + children if at least one child exists
            if child_rows:
                rows.append((0, gname, "GroupInfo", False, group.source_file, "GroupInfo", "", gname))
                rows.extend(child_rows)

    return rows


# =============================================================================
# EXCEL WRITER
# =============================================================================

def write_sheet_content(
    sheet,
    rows: List[RowItem],
    is_eng: bool,
    eng_tbl: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Optional[Dict[str, List[Tuple[str, str]]]],
    lang_code: str,
    export_index: Optional[Dict[str, Set[str]]] = None,
    consumer: Optional[StringIdConsumer] = None,
    eng_consumer: Optional[StringIdConsumer] = None,
) -> None:
    """Write rows to a sheet with proper formatting."""

    # Headers
    headers = ["DataType", "GroupInfo", "Original (KR)", "English (ENG)"]
    if not is_eng:
        headers.append(f"Translation ({lang_code.upper()})")

    headers.extend(["Command", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"])

    for col, txt in enumerate(headers, start=1):
        cell = sheet.cell(1, col, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    sheet.row_dimensions[1].height = 25

    # Data rows
    r_idx = 2

    for (depth, text, style_type, is_desc, source_file, data_type, command, group_name) in rows:
        fill, font, row_height = get_style(style_type)

        # Translate
        if source_file and export_index:
            trans_eng, sid_eng = resolve_translation(text, eng_tbl, source_file, export_index, consumer=eng_consumer)
        else:
            trans_eng, sid_eng = get_first_translation(eng_tbl, text)

        trans_other = sid_other = ""
        if not is_eng and lang_tbl:
            if source_file and export_index:
                trans_other, sid_other = resolve_translation(text, lang_tbl, source_file, export_index, consumer=consumer)
            else:
                trans_other, sid_other = get_first_translation(lang_tbl, text)

        col = 1

        # DataType
        c = sheet.cell(r_idx, col, data_type)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        col += 1

        # GroupInfo (repeated every row within a group for easy Excel filtering)
        c = sheet.cell(r_idx, col, group_name)
        c.fill = fill; c.font = font
        c.alignment = Alignment(vertical="center")
        c.border = THIN_BORDER
        col += 1

        # Original (KR)
        c = sheet.cell(r_idx, col, br_to_newline(text))
        c.fill = fill; c.font = font
        c.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
        c.border = THIN_BORDER
        col += 1

        # English (ENG)
        c = sheet.cell(r_idx, col, br_to_newline(trans_eng))
        c.fill = fill; c.font = font
        c.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
        c.border = THIN_BORDER
        col += 1

        # Target language
        if not is_eng:
            c = sheet.cell(r_idx, col, br_to_newline(trans_other))
            c.fill = fill; c.font = font
            c.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
            c.border = THIN_BORDER
            col += 1

        # Command
        c = sheet.cell(r_idx, col, command)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN_BORDER
        col += 1

        # STATUS
        c = sheet.cell(r_idx, col, "")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        col += 1

        # COMMENT
        c = sheet.cell(r_idx, col, "")
        c.fill = fill
        c.border = THIN_BORDER
        col += 1

        # STRINGID
        sid = sid_other if not is_eng else sid_eng
        c = sheet.cell(r_idx, col, sid)
        c.fill = fill; c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.number_format = '@'
        col += 1

        # SCREENSHOT
        c = sheet.cell(r_idx, col, "")
        c.fill = fill
        c.border = THIN_BORDER

        if row_height:
            sheet.row_dimensions[r_idx].height = row_height

        r_idx += 1

    last_row = r_idx - 1

    # Column widths
    base_widths = [18, 25, 40, 80]
    if not is_eng:
        base_widths.append(80)
    base_widths.extend([40, 15, 70, 25, 25])
    for idx, w in enumerate(base_widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w

    # Hide English column for non-English workbooks
    if not is_eng:
        sheet.column_dimensions["D"].hidden = True

    # Status dropdown
    status_col_idx = headers.index("STATUS") + 1
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    col_letter = get_column_letter(status_col_idx)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")
    sheet.add_data_validation(dv)

    # Freeze + auto-filter
    last_col = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"

    # Auto-fit
    autofit_worksheet(sheet)


# =============================================================================
# TAB NAME HELPER
# =============================================================================

def _sanitize_tab_name(name: str) -> str:
    """Sanitize folder name for Excel tab (max 31 chars, no special chars)."""
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31]


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_gimmick_datasheets() -> Dict:
    """
    Generate Gimmick datasheets for all languages.

    Returns:
        Dict with results: {
            "category": "Gimmick",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Gimmick",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("Gimmick Datasheet Generator (Revamped)")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "Gimmick_LQA_Output"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        # 1. Index gimmick data by folder
        gimmick_folder = RESOURCE_FOLDER / "gimmickinfo"
        global_seen: Set[str] = set()
        folder_data = index_gimmick_folder(gimmick_folder, global_seen)

        if not folder_data:
            result["errors"].append("No valid gimmick data found!")
            log.warning("No valid gimmick data found!")
            return result

        # 2. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)
        eng_tbl = lang_tables.get("eng", {})

        if not eng_tbl:
            log.warning("English language table not found!")

        # 3. Get EXPORT index
        export_index = get_export_index()

        # 4. Generate workbooks for each language
        log.info("Processing languages...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())

            wb = Workbook()
            wb.remove(wb.active)

            is_eng = code.lower() == "eng"

            # Order-based StringID consumer (fresh per language)
            ordered_idx = get_ordered_export_index()
            consumer = StringIdConsumer(ordered_idx)
            eng_consumer = StringIdConsumer(ordered_idx) if is_eng else None

            # One tab per folder
            for folder_name, file_list in folder_data.items():
                rows = emit_gimmick_rows(file_list)
                if not rows:
                    continue

                tab_name = _sanitize_tab_name(folder_name)

                sheet = wb.create_sheet(title=tab_name)
                write_sheet_content(
                    sheet, rows, is_eng, eng_tbl,
                    tbl if not is_eng else None,
                    code, export_index, consumer, eng_consumer,
                )

                log.info("    Tab '%s': %d rows", tab_name, len(rows))

            # Save
            if wb.worksheets:
                out_path = output_folder / f"Gimmick_LQA_{code.upper()}.xlsx"
                wb.save(out_path)
                log.info("  Saved: %s", out_path.name)
                result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Gimmick generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_gimmick_datasheets()
    print(f"\nResult: {result}")
