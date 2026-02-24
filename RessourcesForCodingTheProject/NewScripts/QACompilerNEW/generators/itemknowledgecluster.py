"""
ItemKnowledgeCluster Datasheet Generator
==========================================
Mega datasheet that clusters all item-knowledge relationships using
3-pass matching: KnowledgeKey → exact name → fuzzy difflib.

Output: Single mega-sheet per language with clusters ordered by
greedy nearest-neighbor similarity for gradient adjacency.

Columns: DataType | SourceText (KR) | Translation | STATUS | COMMENT | SCREENSHOT | STRINGID
"""

import difflib
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    autofit_worksheet,
    THIN_BORDER,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
    add_status_dropdown,
)
from generators.newitem import _find_knowledge_key

log = get_logger("ItemKnowledgeCluster")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class ClusterRow:
    """Single row in a cluster."""
    data_type: str       # "ItemData", "KnowledgeData", "KnowledgeMatch-Exact",
                         # "KnowledgeMatch-Fuzzy(85%)", "ItemMatch-Fuzzy(82%)"
    kor_text: str
    source_file: str
    fuzzy_score: float = 0.0


@dataclass
class ItemCluster:
    """All rows belonging to one item (anchor)."""
    item_strkey: str
    item_name: str       # Anchor for inter-cluster ordering
    rows: List[ClusterRow] = field(default_factory=list)


# =============================================================================
# STATICINFO SCANNING
# =============================================================================

@dataclass
class ItemRecord:
    """Raw item data from XML."""
    strkey: str
    name: str
    desc: str
    knowledge_key: str
    source_file: str


@dataclass
class KnowledgeRecord:
    """Raw knowledge data from XML."""
    strkey: str
    name: str
    desc: str
    source_file: str


def scan_all_staticinfo(
    folder: Path,
) -> Tuple[List[ItemRecord], List[KnowledgeRecord]]:
    """Scan entire RESOURCE_FOLDER for all ItemInfo + KnowledgeInfo elements.

    Returns:
        (items, knowledges) - raw records from XML
    """
    log.info("Scanning all StaticInfo in %s", folder)
    items: List[ItemRecord] = []
    knowledges: List[KnowledgeRecord] = []
    seen_item_keys: Set[str] = set()
    seen_knowledge_keys: Set[str] = set()

    if not folder.exists():
        log.warning("Folder does not exist: %s", folder)
        return items, knowledges

    # Count files first for progress
    all_paths = list(iter_xml_files(folder))
    total_files = len(all_paths)
    log.info("  Found %d XML files to scan", total_files)

    file_count = 0
    last_pct = -1
    for path in all_paths:
        root = parse_xml_file(path)
        if root is None:
            continue
        file_count += 1
        src = path.name

        # Progress every 10%
        pct = (file_count * 100) // total_files if total_files else 100
        if pct // 10 > last_pct // 10:
            last_pct = pct
            log.info("  Scanning: %d%% (%d/%d files, %d items, %d knowledges)",
                     pct, file_count, total_files, len(items), len(knowledges))

        for el in root.iter("ItemInfo"):
            sk = el.get("StrKey") or ""
            if not sk or sk in seen_item_keys:
                continue
            seen_item_keys.add(sk)
            items.append(ItemRecord(
                strkey=sk,
                name=el.get("ItemName") or "",
                desc=el.get("ItemDesc") or "",
                knowledge_key=_find_knowledge_key(el),
                source_file=src,
            ))

        for el in root.iter("KnowledgeInfo"):
            sk = el.get("StrKey") or ""
            if not sk or sk in seen_knowledge_keys:
                continue
            seen_knowledge_keys.add(sk)
            knowledges.append(KnowledgeRecord(
                strkey=sk,
                name=el.get("Name") or "",
                desc=el.get("Desc") or "",
                source_file=src,
            ))

    log.info("Scan complete: %d files, %d items, %d knowledges",
             file_count, len(items), len(knowledges))
    return items, knowledges


# =============================================================================
# FUZZY MATCHING
# =============================================================================

def _extract_trigrams(name: str) -> Set[str]:
    """Extract character trigrams from a name for fuzzy pre-filtering."""
    if len(name) < 3:
        return {name}
    return {name[i:i+3] for i in range(len(name) - 2)}


@dataclass
class FuzzyIndex:
    """Pre-built index for fast fuzzy candidate lookup.

    Three layers of filtering (cheapest first):
    1. Length bounds: impossible to reach threshold if lengths too different
    2. Trigram overlap: cheap set intersection, high correlation with SequenceMatcher
    3. SequenceMatcher cascade: real_quick_ratio → quick_ratio → ratio
    """
    length_buckets: Dict[int, List[str]] = field(default_factory=dict)
    trigram_index: Dict[str, Set[str]] = field(default_factory=dict)  # {trigram: {names...}}
    name_trigrams: Dict[str, Set[str]] = field(default_factory=dict)  # {name: {trigrams...}}


def _build_fuzzy_index(
    all_name_index: Dict[str, List[Tuple[str, str, str, str]]],
) -> FuzzyIndex:
    """Pre-build multi-layer index for fast fuzzy matching.

    Layer 1 - Length buckets:
        SequenceMatcher ratio can't exceed 2*min(a,b)/(a+b).
        For T=0.80: candidate length must be within [0.667x, 1.5x] of anchor.

    Layer 2 - Trigram index:
        Names sharing >= 30% of trigrams with anchor are likely matches.
        Names sharing < 30% almost never reach 80% SequenceMatcher ratio.
        This eliminates ~95% of length-valid candidates.
    """
    idx = FuzzyIndex()

    # Length buckets
    idx.length_buckets = defaultdict(list)
    for name in all_name_index:
        idx.length_buckets[len(name)].append(name)

    # Trigram index (inverted: trigram → set of names that contain it)
    idx.trigram_index = defaultdict(set)
    for name in all_name_index:
        trigrams = _extract_trigrams(name)
        idx.name_trigrams[name] = trigrams
        for tri in trigrams:
            idx.trigram_index[tri].add(name)

    return idx


def find_fuzzy_matches(
    anchor_name: str,
    all_name_index: Dict[str, List[Tuple[str, str, str, str]]],
    exclude_strkeys: Set[str],
    fuzzy_index: FuzzyIndex,
    threshold: float = 0.80,
    max_matches: int = 5,
) -> List[Tuple[str, str, str, str, float, str]]:
    """Find fuzzy matches using 3-layer pre-filtering.

    Layer 1: Length bounds (O(1) per candidate — skip impossible lengths)
    Layer 2: Trigram overlap (O(T) set intersection — skip dissimilar names)
    Layer 3: SequenceMatcher cascade (real_quick → quick → full ratio)

    Args:
        anchor_name: The name to match against
        all_name_index: {name: [(strkey, desc, source_file, tag), ...]}
        exclude_strkeys: StrKeys to skip (already matched in Pass 1/2)
        fuzzy_index: Pre-built FuzzyIndex from _build_fuzzy_index()
        threshold: Minimum similarity ratio (0.0 - 1.0)
        max_matches: Maximum fuzzy matches to return

    Returns:
        List of (strkey, name, desc, source_file, score, tag) sorted by score desc
    """
    if not anchor_name:
        return []

    anchor_len = len(anchor_name)

    # Layer 1: Length bounds
    min_len = max(1, int(anchor_len * threshold / (2.0 - threshold)))
    max_len = int(anchor_len * (2.0 - threshold) / threshold) + 1

    length_valid: Set[str] = set()
    for blen in range(min_len, max_len + 1):
        if blen in fuzzy_index.length_buckets:
            length_valid.update(fuzzy_index.length_buckets[blen])

    if not length_valid:
        return []

    # Layer 2: Trigram overlap — only check names sharing enough trigrams
    anchor_trigrams = _extract_trigrams(anchor_name)
    if not anchor_trigrams:
        return []

    # Collect candidates that share at least 1 trigram (fast via inverted index)
    trigram_candidates: Dict[str, int] = defaultdict(int)
    for tri in anchor_trigrams:
        for name in fuzzy_index.trigram_index.get(tri, set()):
            if name in length_valid:
                trigram_candidates[name] += 1

    # Filter: need >= 30% trigram overlap to have a chance at 80% ratio
    min_shared = max(1, int(len(anchor_trigrams) * 0.30))
    candidates = [name for name, count in trigram_candidates.items()
                  if count >= min_shared and name != anchor_name]

    # Layer 3: SequenceMatcher cascade
    matches: List[Tuple[str, str, str, str, float, str]] = []
    sm = difflib.SequenceMatcher(None, anchor_name, "")

    for name in candidates:
        sm.set_seq2(name)
        if sm.real_quick_ratio() < threshold:
            continue
        if sm.quick_ratio() < threshold:
            continue
        ratio = sm.ratio()
        if ratio >= threshold:
            for strkey, desc, source_file, tag in all_name_index[name]:
                if strkey not in exclude_strkeys:
                    matches.append((strkey, name, desc, source_file, ratio, tag))

    matches.sort(key=lambda x: -x[4])
    return matches[:max_matches]


# =============================================================================
# CLUSTER BUILDER
# =============================================================================

def build_clusters(
    items: List[ItemRecord],
    knowledges: List[KnowledgeRecord],
) -> List[ItemCluster]:
    """Build item clusters with 3-pass knowledge matching.

    Pass 1: KnowledgeKey direct lookup
    Pass 2: Exact name match (ItemName == KnowledgeInfo.Name)
    Pass 3: Fuzzy difflib matching against ALL names (Knowledge + Item)
    """
    total_items = len(items)
    total_knowledges = len(knowledges)
    log.info("Building clusters for %d items against %d knowledges...", total_items, total_knowledges)

    # Build knowledge maps
    log.info("  Building knowledge lookup maps...")
    knowledge_by_key: Dict[str, KnowledgeRecord] = {}
    knowledge_by_name: Dict[str, List[KnowledgeRecord]] = defaultdict(list)
    for kr in knowledges:
        if kr.strkey:
            knowledge_by_key[kr.strkey] = kr
        if kr.name:
            knowledge_by_name[kr.name].append(kr)
    log.info("  Knowledge by key: %d | Knowledge by name: %d unique names",
             len(knowledge_by_key), len(knowledge_by_name))

    # Build combined name index for fuzzy matching
    # {name: [(strkey, desc, source_file, tag), ...]}
    all_name_index: Dict[str, List[Tuple[str, str, str, str]]] = defaultdict(list)
    for kr in knowledges:
        if kr.name:
            all_name_index[kr.name].append((kr.strkey, kr.desc, kr.source_file, "Knowledge"))
    for ir in items:
        if ir.name:
            all_name_index[ir.name].append((ir.strkey, ir.desc, ir.source_file, "Item"))
    log.info("  Combined name index: %d unique names (for fuzzy matching)", len(all_name_index))

    # Pre-build multi-layer fuzzy index (length buckets + trigram inverted index)
    log.info("  Building fuzzy index (length buckets + trigram index)...")
    fuzzy_index = _build_fuzzy_index(all_name_index)
    log.info("  Length buckets: %d distinct lengths | Trigrams: %d unique",
             len(fuzzy_index.length_buckets), len(fuzzy_index.trigram_index))

    clusters: List[ItemCluster] = []
    pass1_count = pass2_count = pass3_count = dedup_removed = 0
    last_pct = -1
    t0 = time.time()

    for item_idx, item in enumerate(items):
        # Progress every 5%
        pct = ((item_idx + 1) * 100) // total_items if total_items else 100
        if pct // 5 > last_pct // 5:
            last_pct = pct
            elapsed = time.time() - t0
            rate = (item_idx + 1) / elapsed if elapsed > 0 else 0
            remaining = (total_items - item_idx - 1) / rate if rate > 0 else 0
            log.info("  Clustering: %d%% (%d/%d) | P1:%d P2:%d P3:%d | %.0fs elapsed, ~%.0fs remaining",
                     pct, item_idx + 1, total_items, pass1_count, pass2_count, pass3_count,
                     elapsed, remaining)

        cluster = ItemCluster(item_strkey=item.strkey, item_name=item.name)
        exclude_strkeys: Set[str] = {item.strkey}

        # ItemData rows (always)
        if item.name:
            cluster.rows.append(ClusterRow("ItemData", item.name, item.source_file))
            _collect_korean_string(item.name)
        if item.desc:
            cluster.rows.append(ClusterRow("ItemData", item.desc, item.source_file))
            _collect_korean_string(item.desc)

        # Pass 1: KnowledgeKey lookup
        if item.knowledge_key and item.knowledge_key in knowledge_by_key:
            kr = knowledge_by_key[item.knowledge_key]
            exclude_strkeys.add(kr.strkey)
            if kr.name:
                cluster.rows.append(ClusterRow("KnowledgeData", kr.name, kr.source_file))
                _collect_korean_string(kr.name)
            if kr.desc:
                cluster.rows.append(ClusterRow("KnowledgeData", kr.desc, kr.source_file))
                _collect_korean_string(kr.desc)
            pass1_count += 1

        # Pass 2: Exact name match
        if item.name and item.name in knowledge_by_name:
            for kr in knowledge_by_name[item.name]:
                if kr.strkey not in exclude_strkeys:
                    exclude_strkeys.add(kr.strkey)
                    if kr.name:
                        cluster.rows.append(ClusterRow("KnowledgeMatch-Exact", kr.name, kr.source_file))
                        _collect_korean_string(kr.name)
                    if kr.desc:
                        cluster.rows.append(ClusterRow("KnowledgeMatch-Exact", kr.desc, kr.source_file))
                        _collect_korean_string(kr.desc)
                    pass2_count += 1
                    break  # Take first non-duplicate match

        # Pass 3: Fuzzy matching across ALL names
        fuzzy_hits = find_fuzzy_matches(item.name, all_name_index, exclude_strkeys, fuzzy_index)
        for strkey, name, desc, source_file, score, src_tag in fuzzy_hits:
            score_pct = int(score * 100)
            tag = "KnowledgeMatch-Fuzzy" if src_tag == "Knowledge" else "ItemMatch-Fuzzy"
            dtype = f"{tag}({score_pct}%)"

            if name:
                cluster.rows.append(ClusterRow(dtype, name, source_file, score))
                _collect_korean_string(name)
            if desc:
                cluster.rows.append(ClusterRow(dtype, desc, source_file, score))
                _collect_korean_string(desc)
            exclude_strkeys.add(strkey)
            pass3_count += 1

        # Deduplicate rows where kor_text is identical within the same cluster.
        # Insertion order = priority: ItemData > KnowledgeData > KnowledgeMatch-Exact > Fuzzy
        if cluster.rows:
            seen_texts: Set[str] = set()
            deduped: List[ClusterRow] = []
            for row in cluster.rows:
                if row.kor_text not in seen_texts:
                    seen_texts.add(row.kor_text)
                    deduped.append(row)
                else:
                    dedup_removed += 1
            cluster.rows = deduped

        if cluster.rows:
            clusters.append(cluster)

    log.info("Clusters built: %d total | P1(KnowledgeKey): %d | P2(ExactName): %d | P3(Fuzzy): %d | Deduped: %d rows removed",
             len(clusters), pass1_count, pass2_count, pass3_count, dedup_removed)
    return clusters


# =============================================================================
# INTER-CLUSTER ORDERING
# =============================================================================

def order_clusters_by_similarity(clusters: List[ItemCluster]) -> List[ItemCluster]:
    """Order clusters so related items appear adjacent.

    Uses alphabetical sort on item_name — Korean names with shared prefixes
    naturally cluster together (에세리온 단검, 에세리온 도끼, 에세리온 장검).
    O(N log N) instead of O(N^2) greedy nearest-neighbor.
    """
    if len(clusters) <= 1:
        return clusters

    log.info("Ordering %d clusters alphabetically by item name...", len(clusters))
    t0 = time.time()

    ordered = sorted(clusters, key=lambda c: c.item_name or "")

    elapsed = time.time() - t0
    log.info("Cluster ordering complete in %.1fs", elapsed)
    return ordered


# =============================================================================
# EXCEL WRITER
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_fill_a = PatternFill("solid", fgColor="E2EFDA")  # Light green
_fill_b = PatternFill("solid", fgColor="FCE4D6")  # Light orange


def write_cluster_excel(
    clusters: List[ItemCluster],
    lang_tbl: Dict[str, List[Tuple[str, str]]],
    lang_code: str,
    export_index: Dict[str, Set[str]],
    output_path: Path,
) -> None:
    """Write ItemKnowledgeCluster Excel with one mega-sheet.

    7 columns: DataType | SourceText (KR) | Translation | STATUS | COMMENT | SCREENSHOT | STRINGID
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ItemKnowledgeCluster"
    code = lang_code.upper()

    headers = [
        "DataType",
        "SourceText (KR)",
        f"Translation ({code})",
        "STATUS",
        "COMMENT",
        "SCREENSHOT",
        "STRINGID",
    ]

    # Write header row
    for col_idx, txt in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    excel_row = 2
    current_fill = _fill_a
    total_clusters = len(clusters)
    last_pct = -1

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)

    # Global dedup: track (kor_text, stringid) across all clusters
    # ItemData rows are ALWAYS kept (cluster anchors). Knowledge-side rows
    # (KnowledgeData, KnowledgeMatch-*, ItemMatch-Fuzzy) are deduplicated.
    global_seen: Set[Tuple[str, str]] = set()
    global_seen_texts: Set[str] = set()
    global_dedup_count = 0

    for cl_idx, cluster in enumerate(clusters):
        if not cluster.rows:
            continue

        # Progress every 10%
        pct = ((cl_idx + 1) * 100) // total_clusters if total_clusters else 100
        if pct // 10 > last_pct // 10:
            last_pct = pct
            log.info("  Writing Excel: %d%% (%d/%d clusters, %d rows so far)",
                     pct, cl_idx + 1, total_clusters, excel_row - 2)

        # Alternate fill per cluster for visual grouping
        current_fill = _fill_b if current_fill == _fill_a else _fill_a

        for crow in cluster.rows:
            is_anchor = crow.data_type == "ItemData"

            # Pre-dedup: skip non-anchor rows whose text we've already seen
            # (same text in same export file -> same SID, so text-only check suffices)
            if not is_anchor and crow.kor_text in global_seen_texts:
                global_dedup_count += 1
                continue

            # NOW consume (only for rows that pass dedup)
            trans, sid = resolve_translation(
                crow.kor_text, lang_tbl, crow.source_file, export_index,
                consumer=consumer,
            )

            global_seen_texts.add(crow.kor_text)

            # Full dedup key for global tracking (text + sid)
            dedup_key = (crow.kor_text, sid)
            global_seen.add(dedup_key)

            vals = [crow.data_type, crow.kor_text, trans, "", "", "", sid]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(excel_row, ci, val)
                cell.fill = current_fill
                cell.border = THIN_BORDER
                if ci == 4:  # STATUS
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # STRINGID as text format (prevent scientific notation)
            ws.cell(excel_row, 7).number_format = '@'
            excel_row += 1

    if global_dedup_count:
        log.info("Global dedup: %d knowledge rows removed across clusters", global_dedup_count)
    if consumer.warnings:
        log.warning("StringID overruns: %d (data had more duplicates than export)", consumer.warnings)

    # Sheet cosmetics
    if excel_row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
    ws.freeze_panes = "A2"

    add_status_dropdown(ws, col=4)
    autofit_worksheet(ws)

    wb.save(output_path)
    log.info("ItemKnowledgeCluster Excel saved: %s (%d rows)", output_path.name, excel_row - 2)


# =============================================================================
# MAIN GENERATOR
# =============================================================================

def generate_itemknowledgecluster_datasheets() -> Dict:
    """Generate ItemKnowledgeCluster mega datasheets for all languages.

    Pipeline:
    1. Load language tables
    2. Scan all StaticInfo for ItemInfo + KnowledgeInfo
    3. Build clusters (3-pass matching)
    4. Order clusters by name similarity
    5. Get EXPORT index
    6. For each language: write mega-sheet Excel

    Returns:
        Dict with results
    """
    result = {
        "category": "ItemKnowledgeCluster",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()

    log.info("=" * 70)
    log.info("ItemKnowledgeCluster Datasheet Generator")
    log.info("=" * 70)

    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "ItemKnowledgeCluster"
    output_folder.mkdir(exist_ok=True)

    item_folder = RESOURCE_FOLDER / "iteminfo"
    if not item_folder.exists():
        item_folder = RESOURCE_FOLDER

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        gen_start = time.time()

        # 1. Load language tables
        log.info("[1/6] Loading language tables...")
        lang_tables = load_language_tables(LANGUAGE_FOLDER)
        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        # 2. Scan all StaticInfo (items from iteminfo, knowledge from knowledgeinfo)
        log.info("[2/6] Scanning StaticInfo XML files...")
        t0 = time.time()
        items_raw, knowledges_raw = scan_all_staticinfo(item_folder)

        # Also scan knowledgeinfo folder separately
        knowledge_folder = RESOURCE_FOLDER / "knowledgeinfo"
        if knowledge_folder.exists():
            log.info("  Also scanning knowledgeinfo folder...")
            _, extra_knowledges = scan_all_staticinfo(knowledge_folder)
            # Merge, avoiding duplicates
            seen_keys = {k.strkey for k in knowledges_raw}
            added = 0
            for kr in extra_knowledges:
                if kr.strkey not in seen_keys:
                    knowledges_raw.append(kr)
                    seen_keys.add(kr.strkey)
                    added += 1
            log.info("  Merged knowledgeinfo: +%d new records = %d total knowledge",
                     added, len(knowledges_raw))
        log.info("  Scan complete in %.1fs: %d items, %d knowledges",
                 time.time() - t0, len(items_raw), len(knowledges_raw))

        if not items_raw:
            result["errors"].append("No item data found!")
            log.warning("No item data found!")
            return result

        # 3. Build clusters (3-pass matching)
        log.info("[3/6] Building clusters (3-pass matching)...")
        t0 = time.time()
        clusters = build_clusters(items_raw, knowledges_raw)
        log.info("  Clustering complete in %.1fs", time.time() - t0)

        # 4. Order clusters by name similarity
        log.info("[4/6] Ordering clusters by name similarity...")
        t0 = time.time()
        clusters = order_clusters_by_similarity(clusters)
        log.info("  Ordering complete in %.1fs", time.time() - t0)

        # 5. Get EXPORT index
        log.info("[5/6] Loading EXPORT index...")
        export_index = get_export_index()

        # 6. Generate Excel per language
        log.info("[6/6] Writing Excel files...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())
            excel_path = output_folder / f"ItemKnowledgeCluster_LQA_{code.upper()}.xlsx"
            write_cluster_excel(clusters, tbl, code, export_index, excel_path)
            result["files_created"] += 1

        total_elapsed = time.time() - gen_start
        log.info("=" * 70)
        log.info("Done! Total time: %.1fs | Output: %s", total_elapsed, output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in ItemKnowledgeCluster generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_itemknowledgecluster_datasheets()
    print(f"\nResult: {result}")
