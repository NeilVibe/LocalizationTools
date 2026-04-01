"""
InputMap Datasheet Generator
=============================
Extracts input bindings from inputmap.xml and creates a reference Excel file
mapping actions to their GamePad and Keyboard/Mouse key bindings with translations.

Structure (inputmap.xml):
- <Input Name="...">: Individual input binding definitions
  - <GamePad Key="..." Method="..."/>
  - <KeyboardMouse Key="..." Method="..."/>
- <InputGroup Name="..." LayerName="...">: Named groups containing Input elements
- <SubGroup Name="..." UIName="..." UIDesc="...">: Translatable action groups
  - <Input Name="..."/>  (references to Input definitions)
- <CustomizationGroup Name="..." UIName="..." CheckType="...">: UI categories
  - <SubGroup Name="..."/>  (references to SubGroup definitions)
- <CustomizationPreset Name="..." UIName="..." Path="..."/>
- <AllowCustomizeInputKey>: Allowed rebindable keys

Output (2 sheets):
- Sheet 1 "Actions": SubGroups with Korean UIName/UIDesc, translations, and key bindings
  - CustomizationGroup headers (depth 0, gold)
  - SubGroup rows (depth 1, blue) with UIName, UIDesc, linked keys
- Sheet 2 "Input Bindings": Complete reference of all Input -> Key mappings
  - InputGroup headers (depth 0, gold)
  - Input rows (depth 1) with GamePad/KB keys
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from config import INPUTMAP_FILE, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    load_language_tables,
    normalize_placeholders,
    resolve_staticinfo_codes,
    autofit_worksheet,
    contains_korean,
    br_to_newline,
    THIN_BORDER,
    get_first_translation,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
)

log = get_logger("InputMapGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class InputBinding:
    """A single input binding (GamePad or KeyboardMouse child)."""
    device: str          # "GamePad" or "KeyboardMouse"
    key: str             # Key value (e.g., "buttonLB buttonRT")
    method: str          # Method (e.g., "press", "downonce", "repeat", "analog")
    name: str = ""       # Override name (e.g., "OverrideKey")


@dataclass
class InputEntry:
    """A single <Input> element with its bindings."""
    name: str                                    # Input Name attribute
    bindings: List[InputBinding] = field(default_factory=list)
    # Extra attributes
    keep_triggered: str = ""
    enable_block_trigger: str = ""
    consume: str = ""
    do_not_consume: str = ""
    ignore_grant_time: str = ""
    cool_time: str = ""
    group_name: str = ""                         # Parent InputGroup Name (if any)
    group_layer: str = ""                        # Parent InputGroup LayerName

    @property
    def gamepad_keys(self) -> str:
        """Combined GamePad keys string."""
        parts = []
        for b in self.bindings:
            if b.device == "GamePad":
                parts.append(b.key)
        return " + ".join(parts) if parts else ""

    @property
    def gamepad_method(self) -> str:
        """GamePad method (usually same for all)."""
        for b in self.bindings:
            if b.device == "GamePad":
                return b.method
        return ""

    @property
    def kb_keys(self) -> str:
        """Combined KeyboardMouse keys string."""
        parts = []
        for b in self.bindings:
            if b.device == "KeyboardMouse":
                parts.append(b.key)
        return " / ".join(parts) if parts else ""

    @property
    def kb_method(self) -> str:
        """KeyboardMouse method."""
        for b in self.bindings:
            if b.device == "KeyboardMouse":
                return b.method
        return ""

    @property
    def extra_attrs(self) -> str:
        """Combined extra attributes as readable string."""
        parts = []
        if self.keep_triggered:
            parts.append(f"KeepTriggered={self.keep_triggered}")
        if self.enable_block_trigger:
            parts.append(f"BlockTrigger={self.enable_block_trigger}")
        if self.consume:
            parts.append(f"Consume={self.consume}")
        if self.cool_time:
            parts.append(f"CoolTime={self.cool_time}")
        if self.ignore_grant_time:
            parts.append(f"IgnoreGrantTime={self.ignore_grant_time}")
        return ", ".join(parts)


@dataclass
class SubGroupEntry:
    """A <SubGroup> with UIName/UIDesc and input references."""
    name: str
    ui_name: str         # Korean UI label
    ui_desc: str         # Korean UI description (may contain {StaticInfo:...})
    ui_desc_resolved: str = ""  # UIDesc with StaticInfo codes resolved to Korean
    input_names: List[str] = field(default_factory=list)  # Referenced Input Names


@dataclass
class CustomizationGroupEntry:
    """A <CustomizationGroup> containing SubGroup references."""
    name: str
    ui_name: str         # Korean UI label (e.g., "기본", "전투")
    check_type: str      # e.g., "Action"
    subgroup_names: List[str] = field(default_factory=list)


@dataclass
class CustomizationPresetEntry:
    """A <CustomizationPreset>."""
    name: str
    ui_name: str
    path: str


# =============================================================================
# STYLING
# =============================================================================

# Depth 0: CustomizationGroup / InputGroup header (gold)
_depth0_fill = PatternFill("solid", fgColor="FFD700")
_depth0_font = Font(bold=True, size=12)
_depth0_row_height = 30

# Depth 1: SubGroup / Input entry (light blue)
_depth1_fill = PatternFill("solid", fgColor="B4C6E7")
_depth1_font = Font(bold=True, size=10)
_depth1_row_height = 22

# Depth 2: Sub-input or detail (light green)
_depth2_fill = PatternFill("solid", fgColor="E2EFDA")
_depth2_font = Font(size=10)
_depth2_row_height = None

# Ungrouped inputs (very light gray)
_ungrouped_fill = PatternFill("solid", fgColor="F2F2F2")
_ungrouped_font = Font(size=10)

# Headers
_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")

# Key binding highlight (dark slate for key columns)
_key_font = Font(name="Consolas", size=10)


def _get_style_for_depth(depth: int) -> Tuple[PatternFill, Font, Optional[float]]:
    if depth == 0:
        return _depth0_fill, _depth0_font, _depth0_row_height
    elif depth == 1:
        return _depth1_fill, _depth1_font, _depth1_row_height
    else:
        return _depth2_fill, _depth2_font, _depth2_row_height


# =============================================================================
# EXTRACTION
# =============================================================================

def extract_inputmap_data(file_path: Path) -> Tuple[
    Dict[str, InputEntry],
    List[SubGroupEntry],
    List[CustomizationGroupEntry],
    List[CustomizationPresetEntry],
    List[str],  # InputGroup names in order
]:
    """
    Parse inputmap.xml and extract all structures.

    The file uses </>  shorthand closings (handled by sanitize_xml).
    Uses parse_xml_file() for proper XML sanitization + virtual ROOT wrapper.

    Returns:
        (inputs_by_name, subgroups, customization_groups, presets, group_order)
    """
    log.info("Parsing InputMap: %s", file_path)

    root = parse_xml_file(file_path)
    if root is None:
        log.error("Failed to parse InputMap file: %s", file_path)
        return {}, [], [], [], []

    inputs_by_name: Dict[str, InputEntry] = {}
    subgroups: List[SubGroupEntry] = []
    customization_groups: List[CustomizationGroupEntry] = []
    presets: List[CustomizationPresetEntry] = []
    group_order: List[str] = []

    def _parse_input_element(input_el, group_name: str = "", group_layer: str = "") -> Optional[InputEntry]:
        """Parse a single <Input> element with its child bindings."""
        name = input_el.get("Name") or ""
        if not name:
            return None

        entry = InputEntry(
            name=name,
            group_name=group_name,
            group_layer=group_layer,
            keep_triggered=input_el.get("KeepTriggered") or "",
            enable_block_trigger=input_el.get("EnableBlockTrigger") or "",
            consume=input_el.get("Consume") or "",
            do_not_consume=input_el.get("DoNotConsume") or "",
            ignore_grant_time=input_el.get("IgnoreGrantTime") or "",
            cool_time=input_el.get("CoolTime") or "",
        )

        # Collect child bindings
        for child in input_el:
            if child.tag in ("GamePad", "KeyboardMouse"):
                binding = InputBinding(
                    device=child.tag,
                    key=child.get("Key") or "",
                    method=child.get("Method") or "",
                    name=child.get("Name") or "",
                )
                if binding.key:
                    entry.bindings.append(binding)

        return entry

    # -------------------------------------------------------------------------
    # Pass 1: Collect ALL Input definitions (top-level + inside InputGroups)
    # -------------------------------------------------------------------------

    # Top-level inputs (direct children of ROOT, not inside InputGroup)
    for el in root:
        if el.tag == "Input":
            entry = _parse_input_element(el)
            if entry:
                inputs_by_name[entry.name] = entry

        elif el.tag == "InputGroup":
            gname = el.get("Name") or ""
            glayer = el.get("LayerName") or ""
            if gname and gname not in group_order:
                group_order.append(gname)

            for child_el in el:
                if child_el.tag == "Input":
                    entry = _parse_input_element(child_el, group_name=gname, group_layer=glayer)
                    if entry:
                        inputs_by_name[entry.name] = entry

    # Some Input elements may be nested deeper (inside other structures)
    # Do a full-tree scan for any we missed
    for input_el in root.iter("Input"):
        name = input_el.get("Name") or ""
        if not name:
            continue
        # Only parse if it has child bindings (not a reference-only element)
        has_bindings = any(c.tag in ("GamePad", "KeyboardMouse") for c in input_el)
        if has_bindings and name not in inputs_by_name:
            # Try to find parent group
            parent = input_el.getparent()
            gname = ""
            glayer = ""
            if parent is not None and parent.tag == "InputGroup":
                gname = parent.get("Name") or ""
                glayer = parent.get("LayerName") or ""
                if gname and gname not in group_order:
                    group_order.append(gname)

            entry = _parse_input_element(input_el, group_name=gname, group_layer=glayer)
            if entry:
                inputs_by_name[entry.name] = entry

    log.info("Collected %d Input definitions across %d groups",
             len(inputs_by_name), len(group_order))

    # -------------------------------------------------------------------------
    # Pass 2: Collect SubGroups
    # -------------------------------------------------------------------------

    for sg_el in root.iter("SubGroup"):
        name = sg_el.get("Name") or ""
        ui_name = sg_el.get("UIName") or ""
        ui_desc = sg_el.get("UIDesc") or ""

        # Skip empty reference-only SubGroup elements (inside CustomizationGroup)
        # Those have no UIName — they're just name references
        if not ui_name and not ui_desc:
            # Check if this is a reference-only element (no children)
            if len(sg_el) == 0:
                continue

        # Collect Korean strings for coverage
        _collect_korean_string(ui_name)
        if ui_desc:
            resolved_desc = resolve_staticinfo_codes(ui_desc)
            _collect_korean_string(resolved_desc)
        else:
            resolved_desc = ""

        # Collect referenced Input Names
        input_names = []
        for child in sg_el:
            if child.tag == "Input":
                ref_name = child.get("Name") or ""
                if ref_name:
                    input_names.append(ref_name)

        if ui_name or input_names:
            subgroups.append(SubGroupEntry(
                name=name,
                ui_name=ui_name,
                ui_desc=ui_desc,
                ui_desc_resolved=resolved_desc,
                input_names=input_names,
            ))

    log.info("Collected %d SubGroups", len(subgroups))

    # -------------------------------------------------------------------------
    # Pass 3: Collect CustomizationGroups
    # -------------------------------------------------------------------------

    for cg_el in root.iter("CustomizationGroup"):
        name = cg_el.get("Name") or ""
        ui_name = cg_el.get("UIName") or ""
        check_type = cg_el.get("CheckType") or ""

        _collect_korean_string(ui_name)

        # Parse SubGroup references — can be inline text or child elements
        sg_names: List[str] = []

        for child in cg_el:
            if child.tag == "SubGroup":
                # SubGroup as child element with Name attribute
                sg_name_attr = child.get("Name") or ""
                if sg_name_attr:
                    sg_names.append(sg_name_attr)
                else:
                    # SubGroup with inline text list (space-separated names)
                    text = (child.text or "").strip()
                    if text:
                        sg_names.extend(text.split())

        customization_groups.append(CustomizationGroupEntry(
            name=name,
            ui_name=ui_name,
            check_type=check_type,
            subgroup_names=sg_names,
        ))

    log.info("Collected %d CustomizationGroups", len(customization_groups))

    # -------------------------------------------------------------------------
    # Pass 4: Collect CustomizationPresets
    # -------------------------------------------------------------------------

    for cp_el in root.iter("CustomizationPreset"):
        name = cp_el.get("Name") or ""
        ui_name = cp_el.get("UIName") or ""
        path = cp_el.get("Path") or ""

        _collect_korean_string(ui_name)

        presets.append(CustomizationPresetEntry(
            name=name, ui_name=ui_name, path=path
        ))

    log.info("Collected %d CustomizationPresets", len(presets))

    return inputs_by_name, subgroups, customization_groups, presets, group_order


# =============================================================================
# KEY BINDING RESOLVER
# =============================================================================

def _resolve_subgroup_keys(
    sg: SubGroupEntry,
    inputs_by_name: Dict[str, InputEntry],
) -> Tuple[str, str]:
    """
    Get combined GamePad and KB/Mouse keys for a SubGroup's linked inputs.

    Returns the keys from the PRIMARY input (first one, usually without _Start suffix).
    """
    gamepad_parts: List[str] = []
    kb_parts: List[str] = []

    for iname in sg.input_names:
        entry = inputs_by_name.get(iname)
        if entry is None:
            continue
        gp = entry.gamepad_keys
        kb = entry.kb_keys
        if gp and gp not in gamepad_parts:
            gamepad_parts.append(gp)
        if kb and kb not in kb_parts:
            kb_parts.append(kb)

    return (" | ".join(gamepad_parts), " | ".join(kb_parts))


# =============================================================================
# EXCEL WRITER — Sheet 1: Actions (SubGroups with translations)
# =============================================================================

def _write_actions_sheet(
    ws,
    subgroups: List[SubGroupEntry],
    customization_groups: List[CustomizationGroupEntry],
    inputs_by_name: Dict[str, InputEntry],
    eng_tbl: Dict,
    lang_tbl: Optional[Dict],
    lang_code: str,
    is_eng: bool,
) -> None:
    """
    Write the Actions sheet: SubGroups organized by CustomizationGroup,
    with Korean UIName/UIDesc, translations, and key binding columns.
    """
    # Build SubGroup lookup by name
    sg_by_name: Dict[str, SubGroupEntry] = {sg.name: sg for sg in subgroups}

    # ---- HEADERS ----
    headers = [
        "Action Name",
        "UIName (KR)",
        "UIDesc (KR)",
        "UIName (ENG)",
        "UIDesc (ENG)",
    ]
    if not is_eng:
        headers.append(f"UIName ({lang_code.upper()})")
        headers.append(f"UIDesc ({lang_code.upper()})")

    headers.extend([
        "GamePad Keys",
        "KB/Mouse Keys",
        "Input Names",
        "STATUS",
        "COMMENT",
    ])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28

    # ---- DATA ROWS ----
    row_idx = 2

    # Track which SubGroups we've output (to catch orphans)
    output_sg_names: Set[str] = set()

    for cg in customization_groups:
        # CustomizationGroup header row (depth 0, gold)
        fill, font, row_h = _get_style_for_depth(0)
        cg_label = f"{cg.ui_name}  ({cg.name})" if cg.ui_name else cg.name
        cell = ws.cell(row_idx, 1, cg_label)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

        # Translate the CustomizationGroup UIName
        if cg.ui_name:
            eng_tr, _, _ = get_first_translation(eng_tbl, cg.ui_name) if eng_tbl else ("", "", "")
            ws.cell(row_idx, 2, cg.ui_name).border = THIN_BORDER
            ws.cell(row_idx, 4, eng_tr).border = THIN_BORDER

        # Fill remaining cells with gold
        for col in range(2, len(headers) + 1):
            c = ws.cell(row_idx, col)
            c.fill = fill
            c.border = THIN_BORDER
        if row_h:
            ws.row_dimensions[row_idx].height = row_h
        row_idx += 1

        # SubGroup rows under this CustomizationGroup (depth 1, blue)
        for sg_name in cg.subgroup_names:
            sg = sg_by_name.get(sg_name)
            if sg is None:
                # SubGroup not found (reference-only, no full definition)
                fill1, font1, row_h1 = _get_style_for_depth(1)
                cell = ws.cell(row_idx, 1, sg_name)
                cell.fill = fill1
                cell.font = font1
                cell.border = THIN_BORDER
                for col in range(2, len(headers) + 1):
                    ws.cell(row_idx, col).border = THIN_BORDER
                    ws.cell(row_idx, col).fill = fill1
                if row_h1:
                    ws.row_dimensions[row_idx].height = row_h1
                row_idx += 1
                continue

            output_sg_names.add(sg_name)
            row_idx = _write_subgroup_row(
                ws, row_idx, sg, inputs_by_name, eng_tbl, lang_tbl,
                lang_code, is_eng, len(headers)
            )

    # Output orphan SubGroups (not in any CustomizationGroup)
    orphans = [sg for sg in subgroups if sg.name not in output_sg_names]
    if orphans:
        # Orphan header
        fill, font, row_h = _get_style_for_depth(0)
        cell = ws.cell(row_idx, 1, "Other Actions (Ungrouped)")
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        for col in range(2, len(headers) + 1):
            ws.cell(row_idx, col).fill = fill
            ws.cell(row_idx, col).border = THIN_BORDER
        if row_h:
            ws.row_dimensions[row_idx].height = row_h
        row_idx += 1

        for sg in orphans:
            row_idx = _write_subgroup_row(
                ws, row_idx, sg, inputs_by_name, eng_tbl, lang_tbl,
                lang_code, is_eng, len(headers)
            )

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 30   # Action Name
    ws.column_dimensions["B"].width = 25   # UIName (KR)
    ws.column_dimensions["C"].width = 45   # UIDesc (KR)
    ws.column_dimensions["D"].width = 25   # UIName (ENG)
    ws.column_dimensions["E"].width = 45   # UIDesc (ENG)

    col_offset = 5
    if not is_eng:
        ws.column_dimensions["F"].width = 25   # UIName (LOC)
        ws.column_dimensions["G"].width = 45   # UIDesc (LOC)
        col_offset = 7

    # GamePad, KB, Input Names, STATUS, COMMENT
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_offset + 1)].width = 30
    ws.column_dimensions[get_column_letter(col_offset + 2)].width = 30
    ws.column_dimensions[get_column_letter(col_offset + 3)].width = 40
    ws.column_dimensions[get_column_letter(col_offset + 4)].width = 12
    ws.column_dimensions[get_column_letter(col_offset + 5)].width = 30

    # STATUS dropdown
    status_col = col_offset + 4
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    dv.error = "Invalid status"
    ws.add_data_validation(dv)
    for r in range(2, row_idx):
        dv.add(ws.cell(r, status_col))


def _write_subgroup_row(
    ws, row_idx: int,
    sg: SubGroupEntry,
    inputs_by_name: Dict[str, InputEntry],
    eng_tbl: Dict,
    lang_tbl: Optional[Dict],
    lang_code: str,
    is_eng: bool,
    total_cols: int,
) -> int:
    """Write a single SubGroup row with translations and key bindings. Returns next row."""
    fill, font, row_h = _get_style_for_depth(1)

    # Resolve keys
    gp_keys, kb_keys = _resolve_subgroup_keys(sg, inputs_by_name)

    # Translate UIName
    eng_uiname = ""
    loc_uiname = ""
    if sg.ui_name and eng_tbl:
        eng_uiname, _, _ = get_first_translation(eng_tbl, sg.ui_name)
    if sg.ui_name and lang_tbl and not is_eng:
        loc_uiname, _, _ = get_first_translation(lang_tbl, sg.ui_name)

    # Translate UIDesc (resolve StaticInfo codes first)
    desc_for_translation = sg.ui_desc_resolved or sg.ui_desc
    eng_uidesc = ""
    loc_uidesc = ""
    if desc_for_translation and eng_tbl:
        eng_uidesc, _, _ = get_first_translation(eng_tbl, desc_for_translation)
    if desc_for_translation and lang_tbl and not is_eng:
        loc_uidesc, _, _ = get_first_translation(lang_tbl, desc_for_translation)

    # If UIDesc contains {StaticInfo:...} and we couldn't translate the whole thing,
    # show the resolved Korean version at minimum
    uidesc_display = sg.ui_desc
    if sg.ui_desc_resolved and sg.ui_desc_resolved != sg.ui_desc:
        uidesc_display = f"{sg.ui_desc}\n→ {sg.ui_desc_resolved}"

    # Input names as comma-separated list
    input_names_str = ", ".join(sg.input_names)

    col = 1
    # A: Action Name
    c = ws.cell(row_idx, col, sg.name)
    c.fill = fill; c.font = font; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")
    col += 1

    # B: UIName (KR)
    c = ws.cell(row_idx, col, sg.ui_name)
    c.fill = fill; c.font = Font(bold=True, size=10); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    col += 1

    # C: UIDesc (KR)
    c = ws.cell(row_idx, col, br_to_newline(uidesc_display))
    c.fill = fill; c.font = Font(size=10); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    col += 1

    # D: UIName (ENG)
    c = ws.cell(row_idx, col, eng_uiname)
    c.fill = fill; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    col += 1

    # E: UIDesc (ENG)
    c = ws.cell(row_idx, col, eng_uidesc)
    c.fill = fill; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    col += 1

    # F-G: Translation columns (non-ENG only)
    if not is_eng:
        c = ws.cell(row_idx, col, loc_uiname)
        c.fill = fill; c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        col += 1

        c = ws.cell(row_idx, col, loc_uidesc)
        c.fill = fill; c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        col += 1

    # GamePad Keys (monospace font for key display)
    c = ws.cell(row_idx, col, gp_keys)
    c.font = _key_font; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")
    col += 1

    # KB/Mouse Keys
    c = ws.cell(row_idx, col, kb_keys)
    c.font = _key_font; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")
    col += 1

    # Input Names
    c = ws.cell(row_idx, col, input_names_str)
    c.font = Font(size=9, color="666666"); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    col += 1

    # STATUS
    c = ws.cell(row_idx, col, "")
    c.border = THIN_BORDER
    col += 1

    # COMMENT
    c = ws.cell(row_idx, col, "")
    c.border = THIN_BORDER

    if row_h:
        ws.row_dimensions[row_idx].height = row_h

    return row_idx + 1


# =============================================================================
# EXCEL WRITER — Sheet 2: Input Bindings (complete reference)
# =============================================================================

def _write_bindings_sheet(
    ws,
    inputs_by_name: Dict[str, InputEntry],
    group_order: List[str],
) -> None:
    """
    Write the Input Bindings sheet: all Input -> Key mappings grouped by InputGroup.
    """
    headers = [
        "Input Name",
        "GamePad Key(s)",
        "GamePad Method",
        "KB/Mouse Key(s)",
        "KB/Mouse Method",
        "Attributes",
        "Group",
        "Layer",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 25

    row_idx = 2

    # Group inputs by their group_name
    grouped: Dict[str, List[InputEntry]] = {}
    ungrouped: List[InputEntry] = []

    for entry in inputs_by_name.values():
        if entry.group_name:
            grouped.setdefault(entry.group_name, []).append(entry)
        else:
            ungrouped.append(entry)

    # Write grouped inputs in group_order
    written_groups: Set[str] = set()

    for gname in group_order:
        entries = grouped.get(gname, [])
        if not entries:
            continue
        written_groups.add(gname)

        # Group header row (depth 0)
        fill, font, row_h = _get_style_for_depth(0)
        layer = entries[0].group_layer if entries else ""
        label = f"{gname}" + (f"  (Layer: {layer})" if layer else "")
        cell = ws.cell(row_idx, 1, label)
        cell.fill = fill; cell.font = font; cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        for c in range(2, len(headers) + 1):
            ws.cell(row_idx, c).fill = fill
            ws.cell(row_idx, c).border = THIN_BORDER
        if row_h:
            ws.row_dimensions[row_idx].height = row_h
        row_idx += 1

        # Input rows
        for entry in sorted(entries, key=lambda e: e.name):
            row_idx = _write_input_row(ws, row_idx, entry, len(headers))

    # Write groups not in group_order
    for gname, entries in sorted(grouped.items()):
        if gname in written_groups:
            continue
        fill, font, row_h = _get_style_for_depth(0)
        layer = entries[0].group_layer if entries else ""
        label = f"{gname}" + (f"  (Layer: {layer})" if layer else "")
        cell = ws.cell(row_idx, 1, label)
        cell.fill = fill; cell.font = font; cell.border = THIN_BORDER
        for c in range(2, len(headers) + 1):
            ws.cell(row_idx, c).fill = fill
            ws.cell(row_idx, c).border = THIN_BORDER
        if row_h:
            ws.row_dimensions[row_idx].height = row_h
        row_idx += 1
        for entry in sorted(entries, key=lambda e: e.name):
            row_idx = _write_input_row(ws, row_idx, entry, len(headers))

    # Write ungrouped inputs
    if ungrouped:
        fill, font, row_h = _get_style_for_depth(0)
        cell = ws.cell(row_idx, 1, "Top-Level Inputs (No Group)")
        cell.fill = fill; cell.font = font; cell.border = THIN_BORDER
        for c in range(2, len(headers) + 1):
            ws.cell(row_idx, c).fill = fill
            ws.cell(row_idx, c).border = THIN_BORDER
        if row_h:
            ws.row_dimensions[row_idx].height = row_h
        row_idx += 1
        for entry in sorted(ungrouped, key=lambda e: e.name):
            row_idx = _write_input_row(ws, row_idx, entry, len(headers))

    # Column widths
    ws.column_dimensions["A"].width = 40   # Input Name
    ws.column_dimensions["B"].width = 30   # GamePad Key(s)
    ws.column_dimensions["C"].width = 14   # GamePad Method
    ws.column_dimensions["D"].width = 30   # KB/Mouse Key(s)
    ws.column_dimensions["E"].width = 14   # KB/Mouse Method
    ws.column_dimensions["F"].width = 40   # Attributes
    ws.column_dimensions["G"].width = 25   # Group
    ws.column_dimensions["H"].width = 20   # Layer


def _write_input_row(ws, row_idx: int, entry: InputEntry, total_cols: int) -> int:
    """Write a single Input binding row. Returns next row."""
    # A: Input Name
    c = ws.cell(row_idx, 1, entry.name)
    c.font = Font(bold=True, size=10); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", indent=1)

    # B: GamePad Key(s)
    c = ws.cell(row_idx, 2, entry.gamepad_keys)
    c.font = _key_font; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")

    # C: GamePad Method
    c = ws.cell(row_idx, 3, entry.gamepad_method)
    c.border = THIN_BORDER; c.alignment = Alignment(vertical="center")

    # D: KB/Mouse Key(s)
    c = ws.cell(row_idx, 4, entry.kb_keys)
    c.font = _key_font; c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")

    # E: KB/Mouse Method
    c = ws.cell(row_idx, 5, entry.kb_method)
    c.border = THIN_BORDER; c.alignment = Alignment(vertical="center")

    # F: Attributes
    c = ws.cell(row_idx, 6, entry.extra_attrs)
    c.font = Font(size=9, color="888888"); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)

    # G: Group
    c = ws.cell(row_idx, 7, entry.group_name)
    c.font = Font(size=9, color="888888"); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")

    # H: Layer
    c = ws.cell(row_idx, 8, entry.group_layer)
    c.font = Font(size=9, color="888888"); c.border = THIN_BORDER
    c.alignment = Alignment(vertical="center")

    return row_idx + 1


# =============================================================================
# WORKBOOK WRITER (main orchestrator)
# =============================================================================

def write_workbook(
    inputs_by_name: Dict[str, InputEntry],
    subgroups: List[SubGroupEntry],
    customization_groups: List[CustomizationGroupEntry],
    presets: List[CustomizationPresetEntry],
    group_order: List[str],
    eng_tbl: Dict,
    lang_tbl: Optional[Dict],
    lang_code: str,
    out_path: Path,
) -> None:
    """Write the complete InputMap workbook (2 sheets)."""
    is_eng = lang_code.lower() == "eng"

    wb = Workbook()

    # Sheet 1: Actions
    ws_actions = wb.active
    ws_actions.title = "Actions"
    _write_actions_sheet(
        ws_actions, subgroups, customization_groups, inputs_by_name,
        eng_tbl, lang_tbl, lang_code, is_eng
    )

    # Sheet 2: Input Bindings
    ws_bindings = wb.create_sheet("Input Bindings")
    _write_bindings_sheet(ws_bindings, inputs_by_name, group_order)

    # Sheet 3: Customization Presets (small info sheet)
    if presets:
        ws_presets = wb.create_sheet("Presets")
        preset_headers = ["Preset Name", "UIName (KR)", "UIName (ENG)", "Path"]
        for col, h in enumerate(preset_headers, 1):
            cell = ws_presets.cell(1, col, h)
            cell.font = _header_font; cell.fill = _header_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for pidx, preset in enumerate(presets, start=2):
            eng_name = ""
            if preset.ui_name and eng_tbl:
                eng_name, _, _ = get_first_translation(eng_tbl, preset.ui_name)
            ws_presets.cell(pidx, 1, preset.name).border = THIN_BORDER
            ws_presets.cell(pidx, 2, preset.ui_name).border = THIN_BORDER
            ws_presets.cell(pidx, 3, eng_name).border = THIN_BORDER
            ws_presets.cell(pidx, 4, preset.path).border = THIN_BORDER

        ws_presets.column_dimensions["A"].width = 20
        ws_presets.column_dimensions["B"].width = 20
        ws_presets.column_dimensions["C"].width = 20
        ws_presets.column_dimensions["D"].width = 50

    wb.save(out_path)
    log.info("Saved: %s", out_path.name)


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_inputmap_datasheets() -> Dict:
    """
    Generate InputMap reference datasheets for all languages.

    Returns:
        Dict with results: {
            "category": "InputMap",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "InputMap",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()

    log.info("=" * 70)
    log.info("InputMap Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "InputMap_LQA_All"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    if not INPUTMAP_FILE.exists():
        result["errors"].append(f"InputMap file not found: {INPUTMAP_FILE}")
        log.error("InputMap file not found: %s", INPUTMAP_FILE)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        # 1. Extract InputMap data
        inputs_by_name, subgroups, cust_groups, presets, group_order = \
            extract_inputmap_data(INPUTMAP_FILE)

        if not inputs_by_name:
            result["errors"].append("No Input data found in inputmap.xml!")
            log.warning("No Input data found!")
            return result

        log.info("Summary: %d inputs, %d subgroups, %d customization groups, %d presets",
                 len(inputs_by_name), len(subgroups), len(cust_groups), len(presets))

        # 2. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)
        eng_tbl = lang_tables.get("eng", {})

        if not eng_tbl:
            log.warning("English language table not found — translations will be empty")

        # 3. Write English workbook
        write_workbook(
            inputs_by_name, subgroups, cust_groups, presets, group_order,
            eng_tbl, None, "eng",
            output_folder / "LQA_InputMap_ENG.xlsx",
        )
        result["files_created"] += 1

        # 4. Write other language workbooks
        for lang_code, lang_tbl in lang_tables.items():
            if lang_code.lower() == "eng":
                continue
            write_workbook(
                inputs_by_name, subgroups, cust_groups, presets, group_order,
                eng_tbl, lang_tbl, lang_code,
                output_folder / f"LQA_InputMap_{lang_code.upper()}.xlsx",
            )
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in InputMap generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_inputmap_datasheets()
    print(f"\nResult: {result}")
