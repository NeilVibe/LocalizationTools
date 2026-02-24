"""
NEW Character Datasheet Generator
==================================
Row-per-text character datasheet with knowledge passes:
  CharacterName -> KnowledgeName -> KnowledgeDesc -> Knowledge2Name -> Knowledge2Desc

Each text gets its OWN row (unlike character.py which uses a different layout).

Key features:
- Scans characterinfo_*.staticinfo.xml files
- Grouped by filename pattern (NPC, MONSTER, etc.)
- Two-pass knowledge linking (KnowledgeKey direct + name match)
- Alternating fill per character for visual grouping
- One sheet per group key
"""

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    is_good_translation,
    autofit_worksheet,
    THIN_BORDER,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
    add_status_dropdown,
)
from generators.newitem import _find_knowledge_key, load_knowledge_data

log = get_logger("NewCharacterGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DATA STRUCTURE
# =============================================================================

@dataclass
class NewCharacterEntry:
    """Complete data for a single character with knowledge fields."""
    strkey: str                    # CharacterInfo StrKey
    char_name_kor: str             # CharacterInfo.CharacterName
    knowledge_key: str             # from KnowledgeKey/RewardKnowledgeKey
    knowledge_name_kor: str        # Pass 1: direct KnowledgeKey lookup
    knowledge_desc_kor: str        # Pass 1: KnowledgeInfo.Desc
    knowledge2_name_kor: str       # Pass 2: name match
    knowledge2_desc_kor: str       # Pass 2: KnowledgeInfo.Desc
    source_file: str               # Character XML filename (EXPORT matching)
    knowledge_source_file: str     # Knowledge XML filename (EXPORT matching)
    knowledge2_source_file: str = ""   # Pass 2: source file for EXPORT
    group_key: str = ""            # e.g., "npc", "monster" (from filename)


# =============================================================================
# CHARACTER EXTRACTION
# =============================================================================

def iter_characterinfo_files(root: Path):
    """Find all characterinfo_*.staticinfo.xml files."""
    for path in iter_xml_files(root):
        fn = path.name.lower()
        if fn.startswith("characterinfo_") and fn.endswith(".staticinfo.xml"):
            yield path


def get_group_key(filename: str) -> str:
    """
    Extract group key from filename.

    characterinfo_npc.staticinfo.xml -> npc
    characterinfo_npc_shop.staticinfo.xml -> npc
    characterinfo_monster_unique.staticinfo.xml -> monster
    """
    stem = filename.lower()
    if stem.startswith("characterinfo_"):
        stem = stem[len("characterinfo_"):]
    if stem.endswith(".staticinfo.xml"):
        stem = stem[:-len(".staticinfo.xml")]

    parts = stem.split("_")
    return parts[0] if parts else "unknown"


# =============================================================================
# CHARACTER SCANNER WITH KNOWLEDGE
# =============================================================================

def scan_characters_with_knowledge(
    folder: Path,
    knowledge_map: Dict[str, Tuple[str, str, str]],
    knowledge_name_index: Dict[str, List[Tuple[str, str, str]]],
) -> Dict[str, List[NewCharacterEntry]]:
    """Scan character folder and build entries with knowledge data resolved.

    Pass 1: KnowledgeKey/RewardKnowledgeKey -> knowledge_map (direct key lookup)
    Pass 2: CharacterName -> knowledge_name_index (identical name match)

    Returns:
        groups: {group_key: [NewCharacterEntry, ...]} -- grouped by filename pattern
    """
    log.info("Scanning characters with knowledge: %s", folder)
    groups: Dict[str, List[NewCharacterEntry]] = defaultdict(list)
    seen_strkeys: Set[str] = set()
    total_chars = 0
    duplicates = 0
    pass2_hits = 0

    for path in sorted(iter_characterinfo_files(folder)):
        root = parse_xml_file(path)
        if root is None:
            continue

        source_file = path.name
        group_key = get_group_key(path.name)

        for el in root.iter("CharacterInfo"):
            strkey = el.get("StrKey") or ""
            char_name = el.get("CharacterName") or ""

            if not strkey or not char_name:
                continue

            # Dedup by StrKey across all files
            if strkey in seen_strkeys:
                duplicates += 1
                continue
            seen_strkeys.add(strkey)

            knowledge_key = _find_knowledge_key(el)

            # Pass 1: Resolve knowledge data via KnowledgeKey
            knowledge_name = ""
            knowledge_desc = ""
            knowledge_source_file = ""
            pass1_strkey = ""
            if knowledge_key and knowledge_key in knowledge_map:
                knowledge_name, knowledge_desc, knowledge_source_file = knowledge_map[knowledge_key]
                pass1_strkey = knowledge_key

            # Pass 2: Identical name match (CharacterName == KnowledgeInfo.Name)
            knowledge2_name = ""
            knowledge2_desc = ""
            knowledge2_source_file = ""
            if char_name and char_name in knowledge_name_index:
                for kn_strkey, kn_desc, kn_src in knowledge_name_index[char_name]:
                    if kn_strkey != pass1_strkey:
                        knowledge2_name = char_name
                        knowledge2_desc = kn_desc
                        knowledge2_source_file = kn_src
                        pass2_hits += 1
                        break

            # Collect Korean strings for coverage tracking
            _collect_korean_string(char_name)
            _collect_korean_string(knowledge_name)
            _collect_korean_string(knowledge_desc)
            _collect_korean_string(knowledge2_name)
            _collect_korean_string(knowledge2_desc)

            total_chars += 1
            groups[group_key].append(NewCharacterEntry(
                strkey=strkey,
                char_name_kor=char_name,
                knowledge_key=knowledge_key,
                knowledge_name_kor=knowledge_name,
                knowledge_desc_kor=knowledge_desc,
                knowledge2_name_kor=knowledge2_name,
                knowledge2_desc_kor=knowledge2_desc,
                source_file=source_file,
                knowledge_source_file=knowledge_source_file,
                knowledge2_source_file=knowledge2_source_file,
                group_key=group_key,
            ))

        log.debug("  %s -> %s: scanned", path.name, group_key.upper())

    if duplicates:
        log.info("  Skipped %d duplicate StrKeys", duplicates)
    log.info("Characters scanned: %d in %d groups (Pass 2 hits: %d)",
             total_chars, len(groups), pass2_hits)
    for key, entries in sorted(groups.items()):
        log.info("  %s: %d characters", key.upper(), len(entries))

    return dict(groups)


# =============================================================================
# EXCEL WRITER (row-per-text format)
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_fill_a = PatternFill("solid", fgColor="E2EFDA")  # Light green
_fill_b = PatternFill("solid", fgColor="FCE4D6")  # Light orange


def write_newcharacter_excel(
    groups: Dict[str, List[NewCharacterEntry]],
    lang_tbl: Dict[str, List[Tuple[str, str]]],
    lang_code: str,
    export_index: Dict[str, Set[str]],
    output_path: Path,
) -> None:
    """Write NewCharacter Excel with one row per text field.

    9 columns: DataType | Filename | SourceText (KR) | Translation | COMMAND | STATUS | COMMENT | SCREENSHOT | STRINGID

    Per-character row generation (strict order):
    1. CharacterData -- char_name_kor (always output)
    2. KnowledgeData -- knowledge_name_kor (skip if empty)  [Pass 1: KnowledgeKey]
    3. KnowledgeData -- knowledge_desc_kor (skip if empty)  [Pass 1: KnowledgeKey]
    4. KnowledgeData2 -- knowledge2_name_kor (skip if empty) [Pass 2: identical name]
    5. KnowledgeData2 -- knowledge2_desc_kor (skip if empty) [Pass 2: identical name]
    """
    wb = Workbook()
    wb.remove(wb.active)
    code = lang_code.upper()

    headers = [
        "DataType",
        "Filename",
        "SourceText (KR)",
        f"Translation ({code})",
        "COMMAND",
        "STATUS",
        "COMMENT",
        "SCREENSHOT",
        "STRINGID",
    ]

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)

    # ------------------------------------------------------------------
    # PRE-RESOLVE: consume StringIDs in DOCUMENT ORDER (before sorting).
    # The consumer's Nth call for a Korean text must match the Nth
    # occurrence in the XML data file.  Sorting entries by strkey would
    # break that order, so we resolve here first, then store results.
    # ------------------------------------------------------------------
    pre: Dict[Tuple[str, str], Tuple[str, str]] = {}  # (strkey, field) -> (trans, sid)
    for gk in sorted(groups.keys()):
        for entry in groups[gk]:  # list order = document order
            pre[(entry.strkey, "char_name")] = resolve_translation(
                entry.char_name_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
            if entry.knowledge_name_kor:
                pre[(entry.strkey, "knowledge_name")] = resolve_translation(
                    entry.knowledge_name_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
            if entry.knowledge_desc_kor:
                pre[(entry.strkey, "knowledge_desc")] = resolve_translation(
                    entry.knowledge_desc_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
            if entry.knowledge2_name_kor:
                pre[(entry.strkey, "knowledge2_name")] = resolve_translation(
                    entry.knowledge2_name_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)
            if entry.knowledge2_desc_kor:
                pre[(entry.strkey, "knowledge2_desc")] = resolve_translation(
                    entry.knowledge2_desc_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)

    if consumer.warnings:
        log.warning("StringID overruns during pre-resolve: %d", consumer.warnings)

    for group_key in sorted(groups.keys()):
        entries = groups[group_key]
        if not entries:
            continue

        # Create sheet named after group key (e.g., NPC, MONSTER)
        title = group_key.upper()[:31]
        cnt = 1
        while title in {ws.title for ws in wb.worksheets}:
            title = f"{title[:28]}_{cnt}"
            cnt += 1
        ws = wb.create_sheet(title)

        # Write header row
        for col_idx, txt in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, txt)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        excel_row = 2
        current_fill = _fill_a
        last_strkey = None

        # Sort characters by strkey within each group (display order only)
        sorted_entries = sorted(entries, key=lambda e: e.strkey)

        for entry in sorted_entries:
            # Alternate fill per character for visual grouping
            if last_strkey is not None and entry.strkey != last_strkey:
                current_fill = _fill_b if current_fill == _fill_a else _fill_a
            last_strkey = entry.strkey

            # Filename column: source XML filename
            filename = entry.source_file

            # COMMAND for CharacterData rows: /create character {strkey}
            command = f"/create character {entry.strkey}"

            def _write_row(data_type: str, kor_text: str, trans: str, sid: str, cmd: str = "") -> None:
                """Write a single data row using pre-resolved translation."""
                nonlocal excel_row
                vals = [data_type, filename, kor_text, trans, cmd, "", "", "", sid]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(excel_row, ci, val)
                    cell.fill = current_fill
                    cell.border = THIN_BORDER
                    if ci == 6:  # STATUS
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                # STRINGID as text format (prevent scientific notation)
                ws.cell(excel_row, 9).number_format = '@'
                excel_row += 1

            # 1. CharacterData -- CharacterName (always output, with COMMAND)
            t, s = pre[(entry.strkey, "char_name")]
            _write_row("CharacterData", entry.char_name_kor, t, s, command)

            # 2. KnowledgeData -- Name (Pass 1: KnowledgeKey, skip if empty)
            if entry.knowledge_name_kor:
                t, s = pre[(entry.strkey, "knowledge_name")]
                _write_row("KnowledgeData", entry.knowledge_name_kor, t, s)

            # 3. KnowledgeData -- Desc (Pass 1: KnowledgeKey, skip if empty)
            if entry.knowledge_desc_kor:
                t, s = pre[(entry.strkey, "knowledge_desc")]
                _write_row("KnowledgeData", entry.knowledge_desc_kor, t, s)

            # 4. KnowledgeData2 -- Name (Pass 2: identical name match, skip if empty)
            if entry.knowledge2_name_kor:
                t, s = pre[(entry.strkey, "knowledge2_name")]
                _write_row("KnowledgeData2", entry.knowledge2_name_kor, t, s)

            # 5. KnowledgeData2 -- Desc (Pass 2: identical name match, skip if empty)
            if entry.knowledge2_desc_kor:
                t, s = pre[(entry.strkey, "knowledge2_desc")]
                _write_row("KnowledgeData2", entry.knowledge2_desc_kor, t, s)

        # Sheet cosmetics
        if excel_row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
        ws.freeze_panes = "A2"

        # Add STATUS drop-down (column 6)
        add_status_dropdown(ws, col=6)

        # Auto-fit column widths
        autofit_worksheet(ws)

        log.info("  Sheet '%s': %d rows", title, excel_row - 2)

    wb.save(output_path)
    log.info("NewCharacter Excel saved: %s", output_path.name)


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_newcharacter_datasheets() -> Dict:
    """
    Generate NewCharacter datasheets for all languages.

    Pipeline:
    1. Load language tables
    2. Load knowledge data (Name + Desc + source_file)
    3. Scan characters with knowledge resolution
    4. Get EXPORT index (for StringID disambiguation)
    5. For each language: write row-per-text Excel

    Returns:
        Dict with results: {
            "category": "NewCharacter",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "NewCharacter",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("NEW Character Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "NewCharacterData_Map_All"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    knowledge_folder = RESOURCE_FOLDER / "knowledgeinfo"

    try:
        # 1. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        # 2. Load knowledge data (map + name index for Pass 2)
        knowledge_map, knowledge_name_index = load_knowledge_data(knowledge_folder)

        # 3. Scan characters with knowledge (Pass 1: KnowledgeKey, Pass 2: identical name)
        groups = scan_characters_with_knowledge(
            RESOURCE_FOLDER, knowledge_map, knowledge_name_index
        )

        if not groups:
            result["errors"].append("No character data found!")
            log.warning("No character data found!")
            return result

        # 4. Get EXPORT index
        export_index = get_export_index()

        # 5. Generate Excel per language
        log.info("Processing languages...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())
            excel_path = output_folder / f"NewCharacter_LQA_{code.upper()}.xlsx"
            write_newcharacter_excel(groups, tbl, code, export_index, excel_path)
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in NewCharacter generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_newcharacter_datasheets()
    print(f"\nResult: {result}")
