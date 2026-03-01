"""
NEW Region Datasheet Generator
================================
Exact copy of Region generator with ONE addition: DisplayName from RegionInfo.

Per FactionNode, outputs up to 3 rows:
  1. KnowledgeInfo.Name  (at depth)     — same as Region
  2. RegionInfo.DisplayName (at depth)   — NEW: only if different from Knowledge Name
  3. KnowledgeInfo.Desc  (at depth + 1)  — same as Region

All hierarchy, indentation, colors, and shop data are IDENTICAL to Region.

Linkage:
  FactionNode.KnowledgeKey → KnowledgeInfo.(Name, Desc)
  FactionNode.KnowledgeKey → RegionInfo.KnowledgeKey → RegionInfo.DisplayName
"""

from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
)

# Import everything we reuse from Region
from generators.region import (
    # Data classes
    FactionNodeData,
    FactionData,
    FactionGroupData,
    ShopGroup,
    # Parsing
    build_knowledge_name_lookup,
    parse_all_faction_groups,
    parse_standalone_factions,
    parse_shop_file,
    # Row helpers
    RowItem,
    emit_shop_rows,
    # Excel helpers
    get_translated_tab_name,
    write_sheet_content,
    # Constants
    EMPTY_FACTION_NAME,
    # Korean collection (shared parsers write to region's set)
    get_collected_korean_strings as get_region_korean_strings,
    reset_korean_collection as reset_region_korean_collection,
)

log = get_logger("NewRegionGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# DISPLAYNAME LOOKUP (RegionInfo.KnowledgeKey → DisplayName)
# =============================================================================

def build_displayname_lookup(folder: Path) -> Dict[str, str]:
    """
    Build lookup: RegionInfo.KnowledgeKey → RegionInfo.DisplayName

    Scans all XML files for RegionInfo elements (recursive hierarchy).
    Used to add DisplayName rows in NewRegion output.
    """
    log.info("Building RegionInfo DisplayName lookup...")

    lookup: Dict[str, str] = {}
    count = 0

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for ri in root_el.iter("RegionInfo"):
            kk = ri.get("KnowledgeKey") or ""
            display_name = ri.get("DisplayName") or ""

            if not kk or not display_name:
                continue

            # First occurrence wins (same as knowledge lookup)
            if kk.lower() not in lookup:
                lookup[kk.lower()] = display_name
                count += 1

    log.info("  → %d entries", count)
    return lookup


# =============================================================================
# ROW GENERATION (Region rows + DisplayName)
# =============================================================================

def emit_faction_node_rows(
    node: FactionNodeData,
    depth: int,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for a FactionNode and its children. Adds DisplayName row."""
    rows: List[RowItem] = []

    # 1. Name (from KnowledgeInfo.Name — same as Region)
    style = f"FactionNode_{node.node_type}" if node.node_type else "FactionNode"
    rows.append((depth, node.name, style, False, node.source_file, "KnowledgeInfo"))

    # 2. DisplayName (from RegionInfo — NEW)
    displayname = displayname_lookup.get(node.knowledge_key.lower(), "")
    if displayname and displayname != node.name:
        rows.append((depth, displayname, style, False, node.source_file, "RegionInfo.DisplayName"))
        _collect_korean_string(displayname)

    # 3. Description (from KnowledgeInfo.Desc — same as Region)
    if node.description:
        rows.append((depth + 1, node.description, "Description", True, node.source_file, "KnowledgeInfo.Desc"))

    # Children
    for child in node.children:
        rows.extend(emit_faction_node_rows(child, depth + 1, displayname_lookup))

    return rows


def emit_faction_rows(
    faction: FactionData,
    depth: int,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for a Faction and all its FactionNodes."""
    rows: List[RowItem] = []

    rows.append((depth, faction.name, "Faction", False, faction.source_file, "Faction"))

    for node in faction.nodes:
        rows.extend(emit_faction_node_rows(node, depth + 1, displayname_lookup))

    return rows


def emit_faction_group_rows(
    fg: FactionGroupData,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate all rows for a FactionGroup."""
    rows: List[RowItem] = []

    for faction in fg.factions:
        rows.extend(emit_faction_rows(faction, 0, displayname_lookup))

    return rows


def emit_standalone_faction_rows(
    standalone_factions: List[FactionData],
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for standalone Factions."""
    rows: List[RowItem] = []

    for faction in standalone_factions:
        rows.append((0, faction.name, "Faction", False, faction.source_file, "Faction"))
        for node in faction.nodes:
            rows.extend(emit_faction_node_rows(node, 1, displayname_lookup))

    return rows


# =============================================================================
# EXCEL WRITER
# =============================================================================

def write_workbook(
    faction_groups: List[FactionGroupData],
    standalone_factions: List[FactionData],
    shop_groups: List[ShopGroup],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Optional[Dict[str, List[Tuple[str, str]]]],
    lang_code: str,
    out_path: Path,
    export_index: Optional[Dict[str, Set[str]]] = None,
    displayname_lookup: Optional[Dict[str, str]] = None,
) -> bool:
    """Generate Excel workbook for a language. Returns True if saved."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    is_eng = lang_code.lower() == "eng"
    used_titles: Set[str] = set()
    dn_lookup = displayname_lookup or {}

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)
    eng_consumer = StringIdConsumer(ordered_idx) if is_eng else None

    def get_unique_title(base: str) -> str:
        title = base
        counter = 1
        while title in used_titles:
            title = f"{base[:28]}_{counter}"
            counter += 1
        used_titles.add(title)
        return title

    # FactionGroup sheets
    for fg in faction_groups:
        rows = emit_faction_group_rows(fg, dn_lookup)
        if not rows:
            continue

        title = get_translated_tab_name(fg.group_name, eng_tbl, lang_tbl, lang_code, "Group")
        title = get_unique_title(title)

        sheet = wb.create_sheet(title=title)
        write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)

        log.info("    Sheet '%s': %d rows", title, len(rows))

    # Standalone Faction sheet
    if standalone_factions:
        rows = emit_standalone_faction_rows(standalone_factions, dn_lookup)
        if rows:
            title = get_translated_tab_name(EMPTY_FACTION_NAME, eng_tbl, lang_tbl, lang_code, "No Faction")
            title = get_unique_title(title)

            sheet = wb.create_sheet(title=title)
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)

            log.info("    Sheet '%s': %d rows", title, len(rows))

    # Shop sheet (same as Region — no DisplayName needed for shop)
    if shop_groups:
        rows = emit_shop_rows(shop_groups)
        if rows:
            sheet = wb.create_sheet(title="Shop")
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)
            log.info("    Sheet 'Shop': %d rows", len(rows))

    # Save
    if wb.worksheets:
        wb.save(out_path)
        log.info("  → Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))
        return True

    return False


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_newregion_datasheets() -> Dict:
    """
    Generate NewRegion datasheets for all languages.
    Same as Region but with DisplayName rows from RegionInfo.
    """
    result = {
        "category": "NewRegion",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()
    reset_region_korean_collection()  # Shared parsers write to region's set

    log.info("=" * 70)
    log.info("NewRegion Datasheet Generator")
    log.info("=" * 70)

    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "NewRegionData_Map_All"
    output_folder.mkdir(exist_ok=True)

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        global_seen: Set[str] = set()

        # 1. Knowledge lookup (Name + Desc)
        knowledge_lookup = build_knowledge_name_lookup(RESOURCE_FOLDER)

        # 2. DisplayName lookup (RegionInfo.KnowledgeKey → DisplayName) — NEW
        displayname_lookup = build_displayname_lookup(RESOURCE_FOLDER)

        # 3. Parse FactionGroup hierarchy (reuses Region's parsing)
        faction_groups = parse_all_faction_groups(RESOURCE_FOLDER, knowledge_lookup, global_seen)

        # 4. Parse Standalone Factions
        standalone_factions = parse_standalone_factions(RESOURCE_FOLDER, knowledge_lookup, global_seen)

        # 5. Parse Shop data
        shop_file = RESOURCE_FOLDER.parent / "staticinfo_quest" / "funcnpc" / "shop_world.staticinfo.xml"
        shop_groups = parse_shop_file(shop_file, global_seen)

        # 6. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        eng_tbl = lang_tables.get("eng", {})
        export_index = get_export_index()

        # 7. Generate workbooks
        log.info("Generating Excel workbooks...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Generating: %s", idx, total, code.upper())
            out_path = output_folder / f"NewRegion_LQA_{code.upper()}.xlsx"

            if code.lower() == "eng":
                saved = write_workbook(
                    faction_groups, standalone_factions, shop_groups,
                    eng_tbl, None, code, out_path, export_index, displayname_lookup,
                )
            else:
                saved = write_workbook(
                    faction_groups, standalone_factions, shop_groups,
                    eng_tbl, tbl, code, out_path, export_index, displayname_lookup,
                )
            if saved:
                result["files_created"] += 1

        # Merge Korean strings from region's shared parsers into our collection
        _collected_korean_strings.update(get_region_korean_strings())

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in NewRegion generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    generate_newregion_datasheets()
