"""
Help (GameAdvice) Datasheet Generator
=====================================
Extracts GameAdviceGroupInfo / GameAdviceInfo entries (tutorial tips, help text).

Structure:
- GameAdviceGroupInfo (parent): StrKey, GroupName
  - GameAdviceInfo (child): StrKey, Title, Desc

Output:
- ONE Excel sheet with all data
- Parent-child indentation (depth 0 = group, depth 1 = item title, depth 2 = item desc)
- Columns: Original (KR) | English (ENG) | Translation (LOC) | STATUS | COMMENT | STRINGID | SCREENSHOT
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    autofit_worksheet,
    THIN_BORDER,
)

log = get_logger("HelpGenerator")

# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class AdviceItem:
    """Single GameAdvice item with title and description."""
    strkey: str
    title: str
    desc: str


@dataclass
class AdviceGroup:
    """Group of GameAdvice items."""
    strkey: str
    group_name: str
    items: List[AdviceItem] = field(default_factory=list)


# =============================================================================
# STYLING
# =============================================================================

_depth0_fill = PatternFill("solid", fgColor="FFD700")  # Gold for groups
_depth0_font = Font(bold=True, size=12)

_depth1_fill = PatternFill("solid", fgColor="B4C6E7")  # Light blue for titles
_depth1_font = Font(bold=True, size=11)

_depth2_fill = PatternFill("solid", fgColor="E2EFDA")  # Light green for descriptions
_depth2_font = Font(size=10)

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")


def _get_style_for_depth(depth: int) -> Tuple[PatternFill, Font]:
    """Get fill and font for a given depth level."""
    if depth == 0:
        return _depth0_fill, _depth0_font
    elif depth == 1:
        return _depth1_fill, _depth1_font
    else:
        return _depth2_fill, _depth2_font


# =============================================================================
# EXTRACTION
# =============================================================================

def extract_gameadvice_data(folder: Path) -> List[AdviceGroup]:
    """
    Scan StaticInfo folder for GameAdviceGroupInfo/GameAdviceInfo elements.

    Args:
        folder: Path to StaticInfo folder

    Returns:
        List of AdviceGroup with nested AdviceItems
    """
    groups: List[AdviceGroup] = []
    seen_group_keys: set = set()
    seen_item_keys: set = set()

    log.info("Scanning for GameAdvice data in: %s", folder)

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        # Find all GameAdviceGroupInfo elements
        for group_el in root_el.iter("GameAdviceGroupInfo"):
            strkey = group_el.get("StrKey") or ""
            group_name = group_el.get("GroupName") or ""

            # Skip duplicates
            if strkey and strkey in seen_group_keys:
                continue
            if strkey:
                seen_group_keys.add(strkey)

            group = AdviceGroup(strkey=strkey, group_name=group_name)

            # Find GameAdviceInfo children
            for item_el in group_el.iter("GameAdviceInfo"):
                item_strkey = item_el.get("StrKey") or ""
                title = item_el.get("Title") or ""
                desc = item_el.get("Desc") or ""

                # Skip duplicates
                if item_strkey and item_strkey in seen_item_keys:
                    continue
                if item_strkey:
                    seen_item_keys.add(item_strkey)

                # Skip if no content
                if not title and not desc:
                    continue

                group.items.append(AdviceItem(
                    strkey=item_strkey,
                    title=title,
                    desc=desc,
                ))

            # Only add group if it has items or a name
            if group.items or group.group_name:
                groups.append(group)

    log.info("Found %d groups with %d total items",
             len(groups), sum(len(g.items) for g in groups))

    return groups


# =============================================================================
# ROW GENERATION
# =============================================================================

# (depth, text, strkey)
RowItem = Tuple[int, str, str]


def emit_rows(groups: List[AdviceGroup]) -> List[RowItem]:
    """Generate rows with proper indentation."""
    rows: List[RowItem] = []

    for group in groups:
        # Emit group name (depth 0)
        if group.group_name:
            rows.append((0, group.group_name, group.strkey))

        # Emit items
        for item in group.items:
            # Title (depth 1)
            if item.title:
                rows.append((1, item.title, item.strkey))

            # Description (depth 2)
            if item.desc:
                rows.append((2, item.desc, item.strkey))

    # Postprocess: drop empty rows (whitespace-only text)
    rows = [(d, t, s) for (d, t, s) in rows if t and t.strip()]

    return rows


# =============================================================================
# EXCEL WRITER
# =============================================================================

def write_workbook(
    rows: List[RowItem],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with ONE sheet (GameAdvice).

    Args:
        rows: List of (depth, text, strkey) tuples
        eng_tbl: English language table
        lang_tbl: Target language table (None for ENG)
        lang_code: Language code (e.g., "eng", "fre")
        out_path: Output file path
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "GameAdvice"

    is_eng = lang_code.lower() == "eng"

    # Header row
    headers: List = []
    h1 = ws.cell(1, 1, "Original (KR)")
    h2 = ws.cell(1, 2, "English (ENG)")
    headers.extend([h1, h2])

    if not is_eng:
        h3 = ws.cell(1, 3, f"Translation ({lang_code.upper()})")
        headers.append(h3)

    start_extra_col = len(headers) + 1
    extra_names = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    for idx, name in enumerate(extra_names, start=start_extra_col):
        headers.append(ws.cell(1, idx, name))

    for hcell in headers:
        hcell.font = _header_font
        hcell.fill = _header_fill
        hcell.alignment = Alignment(horizontal="center", vertical="center")
        hcell.border = THIN_BORDER
    ws.row_dimensions[1].height = 25

    # Column widths / visibility
    ws.column_dimensions["A"].width = 50  # Original (KR)
    ws.column_dimensions["B"].width = 50  # English (ENG)
    ws.column_dimensions["B"].hidden = not is_eng  # Hide ENG for non-English

    if not is_eng:
        ws.column_dimensions["C"].width = 50  # Translation
        ws.column_dimensions["D"].width = 12  # STATUS
        ws.column_dimensions["E"].width = 40  # COMMENT
        ws.column_dimensions["F"].width = 20  # STRINGID
        ws.column_dimensions["G"].width = 20  # SCREENSHOT
    else:
        ws.column_dimensions["C"].width = 12  # STATUS
        ws.column_dimensions["D"].width = 40  # COMMENT
        ws.column_dimensions["E"].width = 20  # STRINGID
        ws.column_dimensions["F"].width = 20  # SCREENSHOT

    # Data validation for STATUS
    status_col = 4 if not is_eng else 3
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    dv.error = "Invalid status"
    dv.prompt = "Select status"
    ws.add_data_validation(dv)

    # Write data rows
    for row_idx, (depth, text, strkey) in enumerate(rows, start=2):
        normalized = normalize_placeholders(text)
        eng_tr, sid = eng_tbl.get(normalized, ("", ""))
        loc_tr = ""
        if lang_tbl:
            loc_tr, sid = lang_tbl.get(normalized, (loc_tr, sid))

        fill, font = _get_style_for_depth(depth)
        indent = depth

        # Column A: Original (KR)
        c1 = ws.cell(row_idx, 1, text)
        c1.fill = fill
        c1.font = font
        c1.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c1.border = THIN_BORDER

        # Column B: English (ENG)
        c2 = ws.cell(row_idx, 2, eng_tr)
        c2.fill = fill
        c2.font = font
        c2.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c2.border = THIN_BORDER

        col_offset = 2

        # Column C: Translation (if not ENG)
        if not is_eng:
            c3 = ws.cell(row_idx, 3, loc_tr)
            c3.fill = fill
            c3.font = font
            c3.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
            c3.border = THIN_BORDER
            col_offset = 3

        # STATUS, COMMENT, STRINGID, SCREENSHOT
        for extra_idx, val in enumerate(["", "", sid, ""], start=col_offset + 1):
            cell = ws.cell(row_idx, extra_idx, val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # STRINGID as text to prevent scientific notation
            if extra_idx == col_offset + 3:
                cell.number_format = '@'

        # Apply STATUS data validation
        dv.add(ws.cell(row_idx, status_col))

    # Auto-fit columns and rows
    autofit_worksheet(ws)

    # Save
    wb.save(out_path)
    log.info("Saved: %s (%d rows)", out_path.name, len(rows))


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_help_datasheets() -> Dict:
    """
    Generate Help/GameAdvice datasheets for all languages.

    Returns:
        Dict with results: {
            "category": "Help",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Help",
        "files_created": 0,
        "errors": [],
    }

    log.info("=" * 70)
    log.info("Help (GameAdvice) Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "Help_LQA"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        # 1. Extract GameAdvice data
        groups = extract_gameadvice_data(RESOURCE_FOLDER)
        if not groups:
            result["errors"].append("No GameAdvice data found!")
            log.warning("No GameAdvice data found!")
            return result

        # 2. Generate rows
        rows = emit_rows(groups)
        log.info("Generated %d rows", len(rows))

        # 3. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)
        eng_tbl = lang_tables.get("eng", {})

        if not eng_tbl:
            log.warning("English language table not found!")

        # 4. Write workbooks (one per language)
        # Always write English
        write_workbook(
            rows, eng_tbl, None, "eng",
            output_folder / "Help_LQA_ENG.xlsx"
        )
        result["files_created"] += 1

        # Write other languages
        for lang_code, lang_tbl in lang_tables.items():
            if lang_code.lower() == "eng":
                continue
            write_workbook(
                rows, eng_tbl, lang_tbl, lang_code,
                output_folder / f"Help_LQA_{lang_code.upper()}.xlsx"
            )
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Help generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_help_datasheets()
    print(f"\nResult: {result}")
