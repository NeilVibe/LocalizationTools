"""
Quest Datasheet Generator
=========================
Extracts quest data from multiple sources:
- Main Quests (scenario folder)
- Faction Quests (faction folder with FactionKeyList/FactionNodeKeyList lookup)
- Daily Quests (faction folder with "daily" group filter)
- Challenge Quests (challenge folder)
- Minigame Quests (contents_minigame file)

Tab Organization:
  - Main Quest (scenario-based)
  - Faction tabs (ordered by OrderByString from factioninfo)
  - Region Quest (leftover factions with *_Request StrKey)
  - Daily (leftover *_Daily + Group="daily" merged)
  - Politics (leftover factions with *_Situation StrKey)
  - Challenge Quest
  - Minigame Quest
  - Others (leftover factions without pattern match - at end)

Output per-language Excel files with:
  - One sheet per quest type/faction
  - Columns: Original | ENG | Translation | StringKey | Command | STATUS | COMMENT | STRINGID | SCREENSHOT
  - Teleport post-processing from reference file
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable

from lxml import etree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS,
    QUESTGROUPINFO_FILE, SCENARIO_FOLDER, FACTION_QUEST_FOLDER,
    CHALLENGE_FOLDER, MINIGAME_FILE, STRINGKEYTABLE_FILE,
    SEQUENCER_FOLDER, FACTIONINFO_FOLDER, TELEPORT_SOURCE_FILE,
)
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    THIN_BORDER,
    autofit_worksheet,
)

log = get_logger("QuestGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


def _get_name_and_collect(element, attr: str = "Name") -> str:
    """Get Name attribute from element and collect for coverage tracking."""
    name = element.get(attr) or ""
    if name:
        _collect_korean_string(name)
    return name


# =============================================================================
# TYPE ALIASES
# =============================================================================

# Row tuple: (depth, orig, eng, loc, stringkey, stringid, icon, bold, cmd, status, comment, screenshot)
Row = Tuple[int, str, str, str, str, str, bool, bool, str, str, str, str]

# =============================================================================
# STYLING
# =============================================================================

_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),
    1: PatternFill("solid", fgColor="D9E1F2"),
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FCE4D6"),
    4: PatternFill("solid", fgColor="9C27B0"),
}
_icon_fill = PatternFill("solid", fgColor="9BC2E6")
_light_yellow = PatternFill("solid", fgColor="FFFDEB")
_header_font = Font(bold=True)
_header_fill = PatternFill("solid", fgColor="9BC2E6")
_bold_font = Font(bold=True)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _tr(text: str, tbl: Dict[str, Tuple[str, str]]) -> str:
    """Get translation from table."""
    normalized = normalize_placeholders(text)
    return tbl.get(normalized, ("", ""))[0]


def _sid(text: str, tbl: Dict[str, Tuple[str, str]]) -> str:
    """Get StringID from table."""
    normalized = normalize_placeholders(text)
    return tbl.get(normalized, ("", ""))[1]


def _parse_vec3(s: str) -> Optional[Tuple[float, float, float]]:
    """Parse a vector3 string into a tuple of floats."""
    parts = [p for p in s.replace(",", " ").split() if p]
    if len(parts) != 3:
        return None
    try:
        return tuple(float(x) for x in parts)  # type: ignore[return-value]
    except Exception:
        return None


def _norm_seq_name(raw: str) -> str:
    """Normalize sequencer name for lookup."""
    return Path(raw).stem.lower()


def _iter_quest_xml_files(folder: Path) -> Iterable[Path]:
    """Iterate over XML files in quest folder."""
    if not folder.exists():
        return
    for path in iter_xml_files(folder, "*.xml"):
        yield path


# Regex patterns for extracting conditions
_complete_mission_re = re.compile(r'CompleteMission\(([A-Za-z0-9_]+)\)')
_complete_quest_re = re.compile(r'CompleteQuest\(([A-Za-z0-9_]+)\)')


def _extract_conditions(condition_str: str) -> Tuple[List[str], List[str]]:
    """
    Extract mission and quest prerequisites from a condition string.

    Args:
        condition_str: e.g. "CompleteMission(Mission_A) && CompleteQuest(Quest_B)"

    Returns:
        (missions, quests) - lists of prerequisite mission/quest keys
    """
    if not condition_str:
        return [], []

    missions = _complete_mission_re.findall(condition_str)
    quests = _complete_quest_re.findall(condition_str)
    return missions, quests


def _build_prereq_command(missions: List[str], quests: List[str]) -> str:
    """
    Build /complete command from prerequisite missions and quests.

    Args:
        missions: List of prerequisite mission keys
        quests: List of prerequisite quest keys

    Returns:
        Command string like "/complete mission M1 && M2" or "/complete quest Q1 && Q2"
    """
    parts = []

    if missions:
        parts.append("/complete mission " + " && ".join(missions))

    if quests:
        parts.append("/complete quest " + " && ".join(quests))

    return "\n".join(parts)


# =============================================================================
# STRING KEY TABLE LOADING
# =============================================================================

def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse StringKeyTable")
    tbl: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        num = el.get("Key") or ""
        sk  = el.get("StrKey") or ""
        if num and sk:
            tbl[sk.lower()] = num
    log.info("StringKeyTable entries: %d", len(tbl))
    return tbl


# =============================================================================
# STAGE → SEQUENCER MAP & POSITION MAP
# =============================================================================

def build_stage_to_seq_map(
    folders: List[Path]
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Build Stage→Sequencer mapping from quest folders.

    Returns:
        (stage_to_seq, name_to_seq) dictionaries
    """
    log.info("Building Stage→Sequencer map...")
    mapping: Dict[str, str] = {}
    name_map: Dict[str, str] = {}

    for root_folder in folders:
        if not root_folder.exists():
            continue
        for p in _iter_quest_xml_files(root_folder):
            rt = parse_xml_file(p)
            if rt is None:
                continue

            for st in rt.iter("Stage"):
                key = (st.get("StrKey") or "").lower()
                seq_tag = st.find("SequencerStage")
                if not key or seq_tag is None:
                    continue

                raw = seq_tag.get("Sequencer") or ""
                if not raw:
                    continue

                seq = _norm_seq_name(raw)
                mapping.setdefault(key, seq)

                # Handle mission_ prefix variations
                if not key.startswith("mission_"):
                    mapping.setdefault(f"mission_{key}", seq)
                else:
                    mapping.setdefault(key[len("mission_"):], seq)

                name = st.get("Name") or ""
                if name:
                    name_map.setdefault(name.lower(), seq)

    log.info("  Stage2Seq: %d ; Name→Seq: %d", len(mapping), len(name_map))
    return mapping, name_map


def build_seq_position_map(root_folder: Path) -> Dict[str, Tuple[float, float, float]]:
    """
    Build Sequencer→Position mapping from .seqc files.

    Returns:
        Dict mapping sequencer_name.lower() to (x, y, z) position
    """
    log.info("Scanning Sequencer folder for positions: %s", root_folder)
    seq_pos: Dict[str, Tuple[float, float, float]] = {}

    if not root_folder.exists():
        log.warning("Sequencer folder not found")
        return seq_pos

    count = 0
    for p in root_folder.rglob("*.seqc"):
        if not p.is_file():
            continue

        rt = parse_xml_file(p)
        if rt is None:
            continue

        seq_el = rt if rt.tag == "Sequencer" else next(rt.iter("Sequencer"), None)
        if seq_el is None:
            continue

        vec = _parse_vec3(seq_el.get("Position") or "")
        if vec:
            seq_pos[p.stem.lower()] = vec
            count += 1

        if count % 500 == 0:
            log.debug("  Parsed %d sequencer files...", count)

    log.info("Sequencer positions parsed: %d", len(seq_pos))
    return seq_pos


# =============================================================================
# FACTION INFO PARSING
# =============================================================================

def parse_faction_info(
    folder: Path
) -> Tuple[Dict[str, Tuple[str, str]], Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, Tuple[List[str], List[str]]]]:
    """
    Parse faction info files for quest grouping.

    Returns:
        faction_map: faction_strkey.lower() → (faction_name, orderByString) [only with OrderByString]
        quest_to_fac: questKey.lower() → faction_strkey.lower()
        node_to_fac: factionnode_strkey.lower() → faction_strkey.lower()
        all_factions: faction_strkey.lower() → faction_name [ALL factions for leftover classification]
        quest_conditions: questKey.lower() → (missions, quests) prereqs from Condition attribute
    """
    log.info("Scanning faction info folder: %s", folder)

    faction_map: Dict[str, Tuple[str, str]] = {}  # Only factions with OrderByString
    all_factions: Dict[str, str] = {}  # ALL factions: strkey → name
    quest_to_faction: Dict[str, str] = {}
    node_to_faction: Dict[str, str] = {}
    quest_conditions: Dict[str, Tuple[List[str], List[str]]] = {}  # QuestKey → (missions, quests)

    if not folder.exists():
        log.warning("Faction info folder not found")
        return faction_map, quest_to_faction, node_to_faction, all_factions, quest_conditions

    for p in iter_xml_files(folder, "*.xml"):
        rt = parse_xml_file(p)
        if rt is None:
            continue

        for fac in rt.iter("Faction"):
            fac_sk = (fac.get("StrKey") or "").lower()
            fac_name = fac.get("Name") or ""

            # Store ALL factions for leftover classification
            if fac_sk and fac_name:
                all_factions[fac_sk] = fac_name

            # Get OrderByString from Dev node
            order = None
            for dev in fac.iter("Dev"):
                order = dev.get("OrderByString")
                if order:
                    break

            if fac_sk and fac_name and order:
                faction_map[fac_sk] = (fac_name, order)

            if fac_sk:
                # Map quests to faction AND collect conditions
                for qel in fac.iter("Quest"):
                    qk = (qel.get("QuestKey") or "").lower()
                    if qk:
                        quest_to_faction[qk] = fac_sk

                        # Collect Condition attribute for prereqs
                        cond = qel.get("Condition") or ""
                        if cond:
                            missions, quests = _extract_conditions(cond)
                            if missions or quests:
                                quest_conditions[qk] = (missions, quests)

                # Map faction nodes to faction
                for fn in fac.iter("FactionNode"):
                    node_sk = (fn.get("StrKey") or "").lower()
                    if node_sk:
                        node_to_faction[node_sk] = fac_sk

    log.info("Faction infos: %d primary, %d total ; Quests→Faction: %d ; Nodes→Faction: %d ; Quest conditions: %d",
             len(faction_map), len(all_factions), len(quest_to_faction), len(node_to_faction), len(quest_conditions))
    return faction_map, quest_to_faction, node_to_faction, all_factions, quest_conditions


def parse_branch_conditions(folder: Path) -> Dict[str, Tuple[List[str], List[str]]]:
    """
    Parse Branch conditions from quest XML files.

    For each Quest Node element, collects ALL Branch conditions from nested Missions.
    All prereqs are combined into one list per Quest Node.

    Example:
        <Quest_Node_Her_HernandCastle_FinaleStory>
          <Mission StrKey="...">
            <Branch Condition="CompleteMission(Mission_A)" Execute="..."/>
            <Branch Condition="CompleteMission(Mission_B)" Execute="..."/>
          </Mission>
        </Quest_Node_Her_HernandCastle_FinaleStory>

    Returns:
        quest_node_conditions: quest_node_tag.lower() → (all_missions, all_quests)
    """
    log.info("Parsing Branch conditions from: %s", folder)

    quest_node_conditions: Dict[str, Tuple[List[str], List[str]]] = {}

    if not folder.exists():
        return quest_node_conditions

    for p in _iter_quest_xml_files(folder):
        rt = parse_xml_file(p)
        if rt is None:
            continue

        # Find Quest Node elements (top-level elements with Name and StartMission)
        for quest_node in rt:
            if not (quest_node.get("Name") and (quest_node.get("StartMission") or quest_node.get("QuestStart"))):
                continue

            quest_tag = quest_node.tag.lower()
            all_missions: List[str] = []
            all_quests: List[str] = []

            # Collect ALL Branch conditions from ALL nested elements
            for branch in quest_node.iter("Branch"):
                cond = branch.get("Condition") or ""
                if cond:
                    missions, quests = _extract_conditions(cond)
                    for m in missions:
                        if m not in all_missions:
                            all_missions.append(m)
                    for q in quests:
                        if q not in all_quests:
                            all_quests.append(q)

            if all_missions or all_quests:
                quest_node_conditions[quest_tag] = (all_missions, all_quests)

    log.info("Branch conditions: %d Quest Nodes with prereqs found", len(quest_node_conditions))
    return quest_node_conditions


# =============================================================================
# COMMAND BUILDER
# =============================================================================

def build_command(
    el: ET._Element,
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> str:
    """
    Build /complete + /teleport command for a quest element.
    """
    mission_sk = (
        (el.get("StrKey") or "")
        or (el.get("StartMission") or "")
        or (el.get("QuestStart") or "")
    ).strip()

    if not mission_sk:
        return ""

    # Determine prefix based on element type
    if el.tag == "Mission":
        prefix = "prevmission"
    elif el.tag == "SubMission":
        prefix = "prevsubmission"
    else:
        prefix = "prevquest"

    # Build /complete command
    skey = id_tbl.get(mission_sk.lower())
    if skey:
        complete = f"/complete {prefix} {skey}"
    else:
        complete = f"/complete prevquestgroup {mission_sk}"

    # Find sequencer for teleport
    seq = stage2seq.get(mission_sk.lower())

    if not seq:
        for sub in el.iter("ExecuteStage"):
            sk = (sub.get("StageKey") or "").lower()
            seq = stage2seq.get(sk)
            if seq:
                break

    if not seq:
        nm = el.get("Name") or ""
        seq = name_map.get(nm.lower())

    if not seq:
        return complete

    pos = seq_pos.get(seq)
    if not pos:
        return complete

    teleport = "/teleport " + " ".join(f"{c:g}" for c in pos)
    return f"{complete}\n{teleport}"


# =============================================================================
# QUEST GROUP INFO PARSING
# =============================================================================

def parse_master_quests(path: Path) -> Tuple[List[str], Dict[str, str]]:
    """
    Parse QuestGroupInfo for quest ordering.

    Returns:
        (quest_order, quest_to_group) - ordered quest keys and their group mappings
    """
    if not path.exists():
        return [], {}

    root = parse_xml_file(path)
    if root is None:
        return [], {}

    order: List[str] = []
    q2g: Dict[str, str] = {}
    seen: set = set()

    for el in root.iter():
        grp = el.get("Group")
        if grp:
            qk = el.tag
            if qk not in seen:
                order.append(qk)
                q2g[qk] = grp
                seen.add(qk)

    return order, q2g


def parse_group_meta(path: Path) -> Dict[str, str]:
    """
    Parse QuestGroupInfo for group name metadata.

    Returns:
        Dict mapping group_strkey.lower() to group_name
    """
    if not path.exists():
        return {}

    root = parse_xml_file(path)
    if root is None:
        return {}

    meta: Dict[str, str] = {}
    for el in root.iter("QuestGroup"):
        sk = (el.get("StrKey") or "").lower()
        nm = el.get("Name") or ""
        if sk:
            meta[sk] = nm

    return meta


# =============================================================================
# TELEPORT LOOKUP (two-pass matching)
# =============================================================================

def load_teleport_map(path: Path) -> Tuple[Dict[str, Dict[str, str]], Dict[str, str]]:
    """
    Load teleport data from reference Excel file.

    Returns:
        Tuple of:
        - per_sheet: { sheet_title[:31] : { StringKey : teleport_string } }
        - global_map: { StringKey : teleport_string } (all entries for fallback)
    """
    log.info("Loading teleport lookup from: %s", path)

    if not path.is_file():
        log.info("Teleport source file not found - skipping teleport lookup")
        return {}, {}

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log.warning("Failed to load teleport file: %s", e)
        return {}, {}

    per_sheet: Dict[str, Dict[str, str]] = {}
    global_map: Dict[str, str] = {}

    for ws in wb.worksheets:
        sheet_key = ws.title[:31]
        sub_map: Dict[str, str] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 5:
                continue

            key_cell, tele_cell = row[2], row[4]
            if key_cell is None or tele_cell in (None, ""):
                continue

            sk = str(key_cell).strip()
            teleport = str(tele_cell).strip()
            if sk and teleport:
                sub_map.setdefault(sk, teleport)
                global_map.setdefault(sk, teleport)  # Also add to global

        if sub_map:
            per_sheet[sheet_key] = sub_map

    log.info("Teleport entries loaded: %d sheets, %d total entries, %d unique StringKeys",
             len(per_sheet), sum(len(v) for v in per_sheet.values()), len(global_map))
    return per_sheet, global_map


def apply_prereq_commands(
    sheet_data: Dict[str, List[Row]],
    quest_conditions: Dict[str, Tuple[List[str], List[str]]],
    branch_conditions: Dict[str, Tuple[List[str], List[str]]],
    target_tabs: List[str]
) -> Tuple[Dict[str, List[Row]], int]:
    """
    Prepend /complete commands for prerequisite quests/missions.

    Only applies to specified tabs (Daily, Politics, Region Quest).

    Args:
        sheet_data: { sheet_name : [rows] }
        quest_conditions: QuestKey.lower() → (missions, quests) from factioninfo
        branch_conditions: MissionKey.lower() → (missions, quests) from Branch elements
        target_tabs: List of tab names to process

    Returns:
        Tuple of (updated_sheet_data, count of commands added)
    """
    updated: Dict[str, List[Row]] = {}
    count = 0

    for sheet_name, rows in sheet_data.items():
        # Only process target tabs
        if sheet_name not in target_tabs:
            updated[sheet_name] = rows
            continue

        new_rows: List[Row] = []
        for row in rows:
            # Row: (depth, orig, eng, loc, stringkey, stringid, icon, bold, cmd, status, comment, screenshot)
            depth, orig, eng, loc, sk, sid, icon, bold, cmd, status, comment, shot = row

            prereq_cmd = ""

            # Check quest conditions (from factioninfo)
            # StringKey might be the quest tag name (e.g., quest_node_her_xxx)
            sk_lower = sk.lower() if sk else ""
            if sk_lower in quest_conditions:
                missions, quests = quest_conditions[sk_lower]
                prereq_cmd = _build_prereq_command(missions, quests)

            # Check branch conditions (from Branch Execute)
            # StringKey might be the mission StrKey
            if sk_lower in branch_conditions:
                missions, quests = branch_conditions[sk_lower]
                branch_prereq = _build_prereq_command(missions, quests)
                if branch_prereq:
                    if prereq_cmd:
                        prereq_cmd = prereq_cmd + "\n" + branch_prereq
                    else:
                        prereq_cmd = branch_prereq

            # Prepend prereq command to existing command
            if prereq_cmd:
                if cmd:
                    cmd = prereq_cmd + "\n" + cmd
                else:
                    cmd = prereq_cmd
                count += 1

            new_rows.append((depth, orig, eng, loc, sk, sid, icon, bold, cmd, status, comment, shot))

        updated[sheet_name] = new_rows

    log.info("Prereq commands: %d added to tabs %s", count, target_tabs)
    return updated, count


def apply_teleport_two_pass(
    sheet_data: Dict[str, List[Row]],
    per_sheet_map: Dict[str, Dict[str, str]],
    global_map: Dict[str, str]
) -> Tuple[Dict[str, List[Row]], int, int]:
    """
    Apply teleport data with two-pass matching.

    Pass 1: Match by tab name + StringKey (strict, consume entries)
    Pass 2: Match remaining by StringKey only (lenient fallback)

    Args:
        sheet_data: { sheet_name : [rows] }
        per_sheet_map: { sheet_name : { StringKey : teleport } }
        global_map: { StringKey : teleport }

    Returns:
        Tuple of (updated_sheet_data, pass1_count, pass2_count)
    """
    consumed: set = set()  # Track consumed StringKeys
    pass1_count = 0
    pass2_count = 0

    # PASS 1: Strict match by tab name + StringKey
    for sheet_name, rows in sheet_data.items():
        sub_map = per_sheet_map.get(sheet_name, {})
        if not sub_map:
            continue

        new_rows: List[Row] = []
        for (depth, orig, eng, loc, sk, sid, icon, bold, cmd, status, comment, shot) in rows:
            new_cmd = cmd
            if sk and sk in sub_map:
                tele = sub_map[sk]
                if tele and tele not in (cmd or ""):
                    new_cmd = f"{cmd}\n{tele}" if cmd else tele
                    consumed.add(sk)
                    pass1_count += 1
            new_rows.append((depth, orig, eng, loc, sk, sid, icon, bold, new_cmd, status, comment, shot))
        sheet_data[sheet_name] = new_rows

    # PASS 2: Lenient fallback - match by StringKey only (not consumed)
    for sheet_name, rows in sheet_data.items():
        new_rows: List[Row] = []
        for (depth, orig, eng, loc, sk, sid, icon, bold, cmd, status, comment, shot) in rows:
            new_cmd = cmd
            # Only try fallback if not already has teleport and StringKey not consumed
            if sk and sk not in consumed and sk in global_map:
                tele = global_map[sk]
                if tele and tele not in (cmd or ""):
                    new_cmd = f"{cmd}\n{tele}" if cmd else tele
                    consumed.add(sk)
                    pass2_count += 1
            new_rows.append((depth, orig, eng, loc, sk, sid, icon, bold, new_cmd, status, comment, shot))
        sheet_data[sheet_name] = new_rows

    return sheet_data, pass1_count, pass2_count


# =============================================================================
# ROW BUILDERS
# =============================================================================

def build_rows_main(
    scenario_folder: Path,
    quest_order: List[str],
    quest_groups: Dict[str, str],
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    group_meta: Dict[str, str],
) -> List[Row]:
    """Build rows for Main Quest sheet."""
    log.info("Building rows for MAIN quests...")

    quests: Dict[str, ET._Element] = {}

    if scenario_folder.exists():
        for p in _iter_quest_xml_files(scenario_folder):
            rt = parse_xml_file(p)
            if rt is None:
                continue

            for q in rt.iter():
                if q.get("Name") and (q.get("StartMission") or q.get("QuestStart")):
                    quests[q.tag] = q

    rows: List[Row] = []
    current_group = None

    for qk in quest_order:
        q = quests.get(qk)
        if q is None:
            continue

        grp = quest_groups.get(qk, "")

        # Add group header row
        if grp != current_group:
            current_group = grp
            kr_group = group_meta.get(grp.lower(), grp)
            rows.append((
                0,
                kr_group,
                _tr(kr_group, eng_tbl),
                _tr(kr_group, lang_tbl),
                "",
                _sid(kr_group, eng_tbl),
                False, True,
                "", "", "", ""
            ))

        # Quest row
        kor = q.get("Name") or ""
        _collect_korean_string(kor)
        skn = id_tbl.get(qk.lower(), "")
        rows.append((
            1,
            kor,
            _tr(kor, eng_tbl),
            _tr(kor, lang_tbl),
            skn,
            _sid(kor, eng_tbl),
            bool(q.get("StageIcon")), True,
            "", "", "", ""
        ))

        # Mission rows
        for m in q.findall("Mission"):
            mk = m.get("Name") or ""
            _collect_korean_string(mk)
            msk = m.get("StrKey") or ""
            cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
            rows.append((
                2,
                mk,
                _tr(mk, eng_tbl),
                _tr(mk, lang_tbl),
                id_tbl.get(msk.lower(), msk),
                _sid(mk, eng_tbl),
                False, True,
                cmd, "", "", ""
            ))

            # SubMission rows
            for s in m.findall("SubMission"):
                sk = s.get("Name") or ""
                _collect_korean_string(sk)
                ssk = s.get("StrKey") or ""
                cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    3,
                    sk,
                    _tr(sk, eng_tbl),
                    _tr(sk, lang_tbl),
                    id_tbl.get(ssk.lower(), ssk),
                    _sid(sk, eng_tbl),
                    False, True,
                    cmd2, "", "", ""
                ))

    log.info("Main quest rows: %d", len(rows))
    return rows


def _bump_depth(rows: List[Row], inc: int = 1) -> List[Row]:
    """Increase depth of all rows by inc."""
    return [(r[0] + inc,) + r[1:] for r in rows]


def build_rows_faction(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    quest2fac: Dict[str, str],
    node2fac: Dict[str, str],
    quest_conditions: Dict[str, Tuple[List[str], List[str]]] = None,
    branch_conditions: Dict[str, Tuple[List[str], List[str]]] = None,
) -> Dict[str, List[Row]]:
    """Build rows for Faction Quest sheets.

    Args:
        quest_conditions: QuestKey.lower() → (missions, quests) from factioninfo Condition
        branch_conditions: QuestTag.lower() → (missions, quests) from Branch Conditions
    """
    log.info("Building rows for FACTION quests...")

    if quest_conditions is None:
        quest_conditions = {}
    if branch_conditions is None:
        branch_conditions = {}

    def find_faction(el: ET._Element) -> str:
        # 1. FIRST: Authoritative Quest mapping from factioninfo <Quest QuestKey="..."/>
        fk = quest2fac.get(el.tag.lower())
        if fk:
            return fk

        # 2. Fallback: FactionKeyList in quest XML (explicit declaration)
        for sub in el.iter():
            if "FactionKeyList" in sub.attrib:
                parts = [v.strip() for v in (sub.get("FactionKeyList") or "").split(",") if v.strip()]
                if parts:
                    return parts[0].lower()

        # 3. Last resort: FactionNodeKeyList (derived from node mapping)
        for sub in el.iter():
            if "FactionNodeKeyList" in sub.attrib:
                parts = [v.strip() for v in (sub.get("FactionNodeKeyList") or "").split() if v.strip()]
                if parts:
                    fk = node2fac.get(parts[0].lower())
                    if fk:
                        return fk

        return ""

    grouping: Dict[str, List[Row]] = {}

    if not folder.exists():
        return grouping

    for p in _iter_quest_xml_files(folder):
        rt = parse_xml_file(p)
        if rt is None:
            continue

        for q in rt.iter():
            if not (q.get("Name") and (q.get("StartMission") or q.get("QuestStart"))):
                continue

            fk = find_faction(q)
            if not fk:
                continue

            kor = q.get("Name") or ""
            _collect_korean_string(kor)
            rows: List[Row] = []

            # Build prereq command for quest node row
            quest_tag = q.tag.lower()
            prereq_cmd = ""

            # Check factioninfo quest conditions (keyed by QuestKey)
            if quest_tag in quest_conditions:
                missions, quests = quest_conditions[quest_tag]
                prereq_cmd = _build_prereq_command(missions, quests)

            # Check branch conditions (keyed by quest tag)
            if quest_tag in branch_conditions:
                missions, quests = branch_conditions[quest_tag]
                branch_prereq = _build_prereq_command(missions, quests)
                if branch_prereq:
                    if prereq_cmd:
                        prereq_cmd = prereq_cmd + "\n" + branch_prereq
                    else:
                        prereq_cmd = branch_prereq

            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(quest_tag, ""),
                _sid(kor, eng_tbl),
                bool(q.get("StageIcon")), True,
                prereq_cmd, "", "", ""
            ))

            for m in q.findall("Mission"):
                mk = m.get("Name") or ""
                _collect_korean_string(mk)
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))

                for s in m.findall("SubMission"):
                    sk = s.get("Name") or ""
                    _collect_korean_string(sk)
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))

            grouping.setdefault(fk, []).extend(_bump_depth(rows, 1))

    log.info("Faction grouping keys: %d", len(grouping))
    return grouping


def build_rows_daily(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> List[Row]:
    """Build rows for Daily Quest sheet."""
    log.info("Building rows for DAILY quests...")

    rows: List[Row] = []

    if not folder.exists():
        return rows

    for p in _iter_quest_xml_files(folder):
        rt = parse_xml_file(p)
        if rt is None:
            continue

        for q in rt.iter():
            if not (q.get("Name") and q.get("StartMission")):
                continue

            grp = (q.get("Group") or "").lower()
            if "daily" not in grp:
                continue

            kor = q.get("Name") or ""
            _collect_korean_string(kor)
            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(q.tag.lower(), ""),
                _sid(kor, eng_tbl),
                bool(q.get("StageIcon")), True,
                "", "", "", ""
            ))

            for m in q.findall("Mission"):
                mk = m.get("Name") or ""
                _collect_korean_string(mk)
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))

                for s in m.findall("SubMission"):
                    sk = s.get("Name") or ""
                    _collect_korean_string(sk)
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))

    log.info("Daily rows: %d", len(rows))
    return rows


_CHT_TP = {"어비스", "탐험"}
_itemkey_re = re.compile(r'ItemKey\(\s*([A-Za-z0-9_]+)\s*\)')


def _extract_item_command(el: ET._Element) -> str:
    """Extract /create item command from EventCondition."""
    last: Optional[str] = None
    for sub in el.iter():
        for n, v in sub.attrib.items():
            if "EventCondition" in n or n == "EventCondition":
                m = _itemkey_re.findall(v)
                if m:
                    last = m[-1]
    return f"/create item {last}" if last else ""


def build_rows_challenge(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    group_meta: Dict[str, str],
) -> List[Row]:
    """Build rows for Challenge Quest sheet."""
    log.info("Building rows for CHALLENGE quests...")

    rows: List[Row] = []

    if not folder.exists():
        return rows

    for p in _iter_quest_xml_files(folder):
        rt = parse_xml_file(p)
        if rt is None:
            continue

        for q in rt:
            if not (q.get("Name") and q.get("StartMission")):
                continue

            group_code = (q.get("Group") or "").lower()
            group_name = group_meta.get(group_code, group_code)

            # Group header row (depth 4 for visual distinction)
            rows.append((
                4,
                group_name, _tr(group_name, eng_tbl), _tr(group_name, lang_tbl),
                "", _sid(group_name, eng_tbl),
                False, True,
                "", "", "", ""
            ))

            kor = q.get("Name") or ""
            _collect_korean_string(kor)
            use_tp = group_name in _CHT_TP
            cmd = build_command(q, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(q)

            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(q.tag.lower(), ""),
                _sid(kor, eng_tbl),
                False, True, cmd, "", "", ""
            ))

            for m in q.findall("Mission"):
                mk = m.get("Name") or ""
                _collect_korean_string(mk)
                msk = m.get("StrKey") or ""
                cmd2 = build_command(m, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(m)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd2, "", "", ""
                ))

                for s in m.findall("SubMission"):
                    sk = s.get("Name") or ""
                    _collect_korean_string(sk)
                    ssk = s.get("StrKey") or ""
                    cmd3 = build_command(s, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(s)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd3, "", "", ""
                    ))

    log.info("Challenge rows: %d", len(rows))
    return rows


def build_rows_minigame(
    minigame_file: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> List[Row]:
    """Build rows for Minigame Quest sheet."""
    log.info("Building rows for MINIGAME quests from: %s", minigame_file)

    if not minigame_file.is_file():
        log.warning("Minigame file not found")
        return []

    rt = parse_xml_file(minigame_file)
    if rt is None:
        log.error("Failed to parse minigame file")
        return []

    rows: List[Row] = []

    for el in rt:
        tag = el.tag
        name = el.get("Name") or ""
        group = (el.get("Group") or "").lower()

        if not name:
            continue

        _collect_korean_string(name)

        # Type 1: Mission-based minigames (Challenge_Minigame)
        if "_mission" in tag.lower() and el.get("StartMission"):
            rows.append((
                0,
                name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                id_tbl.get(tag.lower(), ""),
                _sid(name, eng_tbl),
                bool(el.get("StageIcon")), True,
                "", "", "", ""
            ))

            for m in el.findall("Mission"):
                mk = m.get("Name") or ""
                _collect_korean_string(mk)
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))

                for s in m.findall("SubMission"):
                    sk = s.get("Name") or ""
                    _collect_korean_string(sk)
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))

        # Type 2: Stage-based minigames (MiniGame)
        elif group == "minigame" or el.find("Stage") is not None:
            rows.append((
                0,
                name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                id_tbl.get(tag.lower(), ""),
                _sid(name, eng_tbl),
                bool(el.get("StageIcon")), True,
                "", "", "", ""
            ))

            for stage in el.findall("Stage"):
                stage_name = stage.get("Name") or ""
                _collect_korean_string(stage_name)
                stage_sk = stage.get("StrKey") or ""

                # Build teleport command from StagePosition if available
                cmd = ""
                stage_pos_str = stage.get("StagePosition") or ""
                if stage_pos_str:
                    pos = _parse_vec3(stage_pos_str)
                    if pos:
                        cmd = "/teleport " + " ".join(f"{c:g}" for c in pos)

                # If no StagePosition, try SequencerStage
                if not cmd:
                    seq_stage = stage.find("SequencerStage")
                    if seq_stage is not None:
                        seq_name = seq_stage.get("Sequencer") or ""
                        if seq_name:
                            seq_key = _norm_seq_name(seq_name)
                            pos = seq_pos.get(seq_key)
                            if pos:
                                cmd = "/teleport " + " ".join(f"{c:g}" for c in pos)

                rows.append((
                    1,
                    stage_name, _tr(stage_name, eng_tbl), _tr(stage_name, lang_tbl),
                    id_tbl.get(stage_sk.lower(), stage_sk),
                    _sid(stage_name, eng_tbl),
                    bool(stage.get("StageIcon")), True,
                    cmd, "", "", ""
                ))

    log.info("Minigame rows: %d", len(rows))
    return rows


# =============================================================================
# EXCEL WRITING
# =============================================================================

def autofit(ws) -> None:
    """Auto-fit column widths and row heights."""
    fixed_widths = {
        "COMMENT": 50,
        "SCREENSHOT": 20,
        "STRINGID": 25,
        "StringKey": 13,
        "ENG": 40,
        "Command": 60,  # Wider for prereq + teleport commands
    }

    headers = [c.value for c in ws[1]]

    # === PHASE 1: Column widths ===
    for col_idx, col in enumerate(ws.columns, start=1):
        letter = col[0].column_letter
        header_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else None

        if header_name in fixed_widths:
            ws.column_dimensions[letter].width = fixed_widths[header_name]
            continue

        max_len = 0
        for cell in col[1:]:
            if cell.value not in (None, ""):
                for line in str(cell.value).splitlines():
                    max_len = max(max_len, len(line))

        ws.column_dimensions[letter].width = 10 if max_len == 0 else min(max_len + 2, 80)

    # === PHASE 2: Row heights (for multi-line Command cells) ===
    default_row_height = 15
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                content = str(cell.value)
                lines = content.count('\n') + 1
                max_lines = max(max_lines, lines)

        # Set row height based on number of lines
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = min(max_lines * default_row_height, 300)


def write_sheet(
    wb: Workbook,
    title: str,
    rows: List[Row],
    tgt_code: str
) -> None:
    """Write a single sheet to the workbook."""
    ws = wb.create_sheet(title=title[:31])
    is_eng = tgt_code.lower().startswith("eng")

    headers = (
        ["Original", "ENG"] +
        ([] if is_eng else [tgt_code.upper()]) +
        ["StringKey", "Command", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    )

    # Write headers
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(1, col, hdr)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    indent_stop = headers.index("StringKey") + 1

    # Write data rows
    r = 2
    for (depth, orig, eng, loc, sk, sid, icon, bold, cmd, status, comment, shot) in rows:
        if is_eng:
            vals = (orig, eng, sk, cmd, status, comment, sid, shot)
        else:
            vals = (orig, eng, loc, sk, cmd, status, comment, sid, shot)

        fg = _icon_fill if icon else _depth_fill.get(depth, _light_yellow)

        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(r, col_idx, v)
            cell.alignment = Alignment(
                indent=(depth if col_idx < indent_stop else 0),
                wrap_text=True,
                vertical="top"
            )
            cell.border = THIN_BORDER
            cell.fill = fg
            if bold:
                cell.font = _bold_font

        r += 1

    ws.freeze_panes = "A2"
    autofit(ws)

    # Hide ENG column for non-ENG workbooks
    if not is_eng:
        eng_idx = headers.index("ENG") + 1
        ws.column_dimensions[ws.cell(row=1, column=eng_idx).column_letter].hidden = True

    # Add STATUS validation
    status_col_idx = headers.index("STATUS") + 1
    col_letter = get_column_letter(status_col_idx)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
        showErrorMessage=True
    )
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    dv.add(rng)
    ws.add_data_validation(dv)

    # Force STRINGID column to text format
    stringid_col_idx = headers.index("STRINGID") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, stringid_col_idx).number_format = '@'


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_quest_datasheets() -> Dict:
    """
    Generate Quest datasheets for all languages.

    Returns:
        Dict with results: {
            "category": "Quest",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Quest",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("Quest Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "QuestData_Map_All"
    output_folder.mkdir(exist_ok=True)

    # Check critical paths
    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        # 1. Load teleport lookup (per-sheet + global for fallback)
        teleport_per_sheet, teleport_global = load_teleport_map(TELEPORT_SOURCE_FILE)

        # 2. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        eng_code = next((c for c in lang_tables if c.startswith("eng")), None)
        if eng_code is None:
            result["errors"].append("English language table not found!")
            log.error("English language table not found!")
            return result

        eng_tbl = lang_tables[eng_code]

        # 3. Load StringKey table
        id_tbl = load_string_key_table(STRINGKEYTABLE_FILE)

        # 4. Parse quest group info
        quest_order, quest_groups = parse_master_quests(QUESTGROUPINFO_FILE)
        group_meta = parse_group_meta(QUESTGROUPINFO_FILE)

        # 5. Build stage→sequencer maps
        stage2seq, name_map = build_stage_to_seq_map([
            SCENARIO_FOLDER, FACTION_QUEST_FOLDER, CHALLENGE_FOLDER
        ])
        seq_pos = build_seq_position_map(SEQUENCER_FOLDER)

        # 6. Parse faction info and conditions
        faction_info_map, quest2fac_map, node2fac_map, all_factions_map, quest_conditions = parse_faction_info(FACTIONINFO_FOLDER)

        # 7. Parse Branch conditions from quest XML
        branch_conditions = parse_branch_conditions(FACTION_QUEST_FOLDER)

        # 7. Process each language
        log.info("Generating Excel workbooks...")
        total = len(lang_tables)

        for idx, (code, lang_tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())

            # Build rows for each quest type
            rows_main = build_rows_main(
                SCENARIO_FOLDER, quest_order, quest_groups,
                lang_tbl, eng_tbl, id_tbl,
                stage2seq, name_map, seq_pos,
                group_meta
            )

            faction_groups = build_rows_faction(
                FACTION_QUEST_FOLDER, lang_tbl, eng_tbl,
                id_tbl, stage2seq, name_map, seq_pos,
                quest2fac_map, node2fac_map,
                quest_conditions, branch_conditions
            )

            rows_daily = build_rows_daily(
                FACTION_QUEST_FOLDER, lang_tbl, eng_tbl,
                id_tbl, stage2seq, name_map, seq_pos
            )

            rows_chal = build_rows_challenge(
                CHALLENGE_FOLDER, lang_tbl, eng_tbl,
                id_tbl, stage2seq, name_map, seq_pos,
                group_meta
            )

            rows_minigame = build_rows_minigame(
                MINIGAME_FILE, lang_tbl, eng_tbl,
                id_tbl, stage2seq, name_map, seq_pos
            )

            # Organize sheets
            sheet_data: Dict[str, List[Row]] = {"Main Quest": rows_main}

            def parse_order(ordstr: str) -> Tuple[int, object]:
                m = re.search(r"(\d+)", ordstr)
                if m:
                    return (0, int(m.group(1)))
                return (1, ordstr)

            sorted_factions = sorted(
                [(fk, name, ordstr) for fk, (name, ordstr) in faction_info_map.items()],
                key=lambda tup: parse_order(tup[2])
            )

            faction_sheet_ordstr: Dict[str, str] = {}

            for fk, name, ordstr in sorted_factions:
                rows = faction_groups.get(fk)
                if not rows:
                    continue

                sheet_name = ordstr[:31]
                faction_sheet_ordstr[sheet_name] = ordstr

                header_row: Row = (
                    0, name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                    "", _sid(name, eng_tbl),
                    False, True, "", "", "", ""
                )

                sheet_data.setdefault(sheet_name, []).append(header_row)
                sheet_data[sheet_name].extend(rows)

            # Handle leftover faction quests - classify by StrKey pattern
            leftover_keys = set(faction_groups) - set(faction_info_map)
            if leftover_keys:
                # Group leftovers by StrKey pattern: Request, Daily, Situation, or Others
                region_quest_rows: List[Row] = []  # *_Request
                daily_rows: List[Row] = []          # *_Daily
                politics_rows: List[Row] = []       # *_Situation
                true_others_rows: List[Row] = []    # No pattern match

                for fk in leftover_keys:
                    rows = faction_groups[fk]
                    fac_name = all_factions_map.get(fk, fk)

                    # Create header row for this faction group
                    header_row: Row = (
                        0, fac_name, _tr(fac_name, eng_tbl), _tr(fac_name, lang_tbl),
                        "", _sid(fac_name, eng_tbl),
                        False, True, "", "", "", ""
                    )

                    # Classify by StrKey pattern
                    if "_request" in fk:
                        region_quest_rows.append(header_row)
                        region_quest_rows.extend(rows)
                    elif "_daily" in fk:
                        daily_rows.append(header_row)
                        daily_rows.extend(rows)
                    elif "_situation" in fk:
                        politics_rows.append(header_row)
                        politics_rows.extend(rows)
                    else:
                        true_others_rows.append(header_row)
                        true_others_rows.extend(rows)

                # Add classified tabs
                if region_quest_rows:
                    sheet_data["Region Quest"] = region_quest_rows
                if daily_rows:
                    sheet_data["Daily"] = daily_rows
                if politics_rows:
                    sheet_data["Politics"] = politics_rows
                if true_others_rows:
                    sheet_data["Others"] = true_others_rows

            # Merge Daily Quest (Group="daily") into Daily tab
            if rows_daily:
                if "Daily" in sheet_data:
                    sheet_data["Daily"].extend(rows_daily)
                else:
                    sheet_data["Daily"] = rows_daily

            sheet_data["Challenge Quest"] = rows_chal

            if rows_minigame:
                sheet_data["Minigame Quest"] = rows_minigame

            # Apply teleport data (two-pass: strict tab match first, then global fallback)
            sheet_data, tp_pass1, tp_pass2 = apply_teleport_two_pass(
                sheet_data, teleport_per_sheet, teleport_global
            )
            if tp_pass1 or tp_pass2:
                log.info("  Teleport applied: %d by tab match, %d by fallback", tp_pass1, tp_pass2)

            # Write workbook
            wb = Workbook()
            wb.remove(wb.active)

            ordered = ["Main Quest"]
            ordered += sorted(faction_sheet_ordstr, key=lambda sn: parse_order(faction_sheet_ordstr[sn]))
            # Leftover classification tabs (from StrKey patterns)
            if "Region Quest" in sheet_data:
                ordered.append("Region Quest")
            if "Daily" in sheet_data:
                ordered.append("Daily")
            if "Politics" in sheet_data:
                ordered.append("Politics")
            # Quest type tabs
            ordered.append("Challenge Quest")
            if "Minigame Quest" in sheet_data:
                ordered.append("Minigame Quest")
            # Others at the end
            if "Others" in sheet_data:
                ordered.append("Others")

            for title in ordered:
                if title in sheet_data and sheet_data[title]:
                    write_sheet(wb, title, sheet_data[title], code)

            out_xlsx = output_folder / f"Quest_LQA_{code.upper()}.xlsx"

            if wb.worksheets:
                wb.save(out_xlsx)
                log.info("→ Saved: %s (%d sheets)", out_xlsx.name, len(wb.worksheets))
                result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Quest generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_quest_datasheets()
    print(f"\nResult: {result}")
