"""
Item Datasheet Generator
=============================
Row-per-text item datasheet with multi-pass knowledge + InspectData resolution:
  ItemName -> ItemDesc -> ChildKnowledge -> KnowledgeData -> KnowledgeData2
  -> InspectData -> InspectKnowledgeData

Each text gets its OWN row with multi-pass knowledge resolution.

Key features:
- ItemGroupInfo hierarchy with parent-child relationships
- Separate Knowledge data rows (Name + Desc from KnowledgeInfo)
- InspectData extraction (Pattern A: direct children, Pattern B: PageData books)
- InspectData -> RewardKnowledgeKey -> KnowledgeInfo chain resolution
- Depth-based clustering for group organization
- Monster item extraction to separate folder
"""

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    is_good_translation,
    br_to_newline,
    autofit_worksheet,
    THIN_BORDER,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
    get_first_translation,
    add_status_dropdown,
    _find_knowledge_key,
    load_knowledge_data,
)

log = get_logger("ItemGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# =============================================================================
# CONSTANTS
# =============================================================================

MERGE_UP_THRESHOLD = 50       # Groups with < this many items merge into parent
FOLDER_MIN_THRESHOLD = 100    # Folders with < this many items merge into Others
MAX_COMMANDS_PER_FILE = 300   # Split files if larger
MIN_FOLDER_DEPTH = 1          # BLUE = minimum folder depth (ceiling)

OTHERS_KEY = "OTHERS"
MONSTER_ITEM_KEY = "MONSTER_ITEM"
MONSTER_SUBSTRING = "mon_"


# =============================================================================
# DATA STRUCTURE
# =============================================================================

@dataclass
class ItemEntry:
    """Complete data for a single item with separate knowledge fields."""
    item_strkey: str           # ItemInfo StrKey
    item_name_kor: str         # ItemInfo.ItemName
    item_desc_kor: str         # ItemInfo.ItemDesc
    knowledge_key: str         # ItemInfo.KnowledgeKey (may be empty)
    knowledge_name_kor: str    # KnowledgeInfo.Name (empty if no key)
    knowledge_desc_kor: str    # KnowledgeInfo.Desc (empty if no key)
    group_key: str             # Parent group StrKey
    source_file: str           # Item XML filename (EXPORT matching)
    knowledge_source_file: str # Knowledge XML filename (EXPORT matching)
    knowledge2_name_kor: str = ""       # Pass 2: identical name match
    knowledge2_desc_kor: str = ""       # Pass 2: identical name match
    knowledge2_source_file: str = ""    # Pass 2: source file for EXPORT
    child_knowledge_entries: List[Tuple[str, str, str]] = field(default_factory=list)  # Pass 0: (name, desc, source_file)
    # InspectData entries: list of (desc, knowledge_name, knowledge_desc, knowledge_source_file)
    # Pattern A (recipe): 1 entry — direct <InspectData> child of <ItemInfo>
    # Pattern B (book): N entries — <PageData> -> <LeftPage>/<RightPage> -> <InspectData>
    inspect_entries: List[Tuple[str, str, str, str]] = field(default_factory=list)


# =============================================================================
# HELPER FUNCTIONS (copied from item.py for independence)
# =============================================================================




def _collect_inspect_data(
    item_element,
    knowledge_map: Dict[str, Tuple[str, str, str]],
) -> List[Tuple[str, str, str, str]]:
    """Collect all InspectData from an ItemInfo element.

    Handles two XML patterns:
      Pattern A (recipe/letter): <InspectData> is a direct child of <ItemInfo>
      Pattern B (book): <PageData> -> <LeftPage>/<RightPage> -> <InspectData>

    For each InspectData element:
      - Skip if UseLeftPageInspectData="True" (duplicate reference)
      - Skip if Desc is empty/missing
      - If RewardKnowledgeKey exists, lookup in knowledge_map for linked Name+Desc

    Returns:
        List of (desc, knowledge_name, knowledge_desc, knowledge_source_file)
        in XML document order.  Empty knowledge fields = "" when no link found.
    """
    results: List[Tuple[str, str, str, str]] = []

    def _process_inspect(el) -> None:
        """Process a single <InspectData> element."""
        # Skip duplicate references (book right pages that mirror left)
        if (el.get("UseLeftPageInspectData") or "").lower() == "true":
            return
        desc = el.get("Desc") or ""
        if not desc:
            return

        # Check for linked knowledge via RewardKnowledgeKey
        rk = el.get("RewardKnowledgeKey") or ""
        k_name = ""
        k_desc = ""
        k_src = ""
        if rk and rk.lower() in knowledge_map:
            k_name, k_desc, k_src = knowledge_map[rk.lower()]

        results.append((desc, k_name, k_desc, k_src))

    # Pattern A: direct <InspectData> children of <ItemInfo>
    for child in item_element.findall("InspectData"):
        _process_inspect(child)

    # Pattern B: <PageData> -> <LeftPage>/<RightPage> -> <InspectData>
    for page_data in item_element.findall("PageData"):
        for page in page_data:  # LeftPage, RightPage, etc.
            for inspect in page.findall("InspectData"):
                _process_inspect(inspect)

    return results


def get_depth_color(depth: int) -> str:
    """Get color name for a given depth level."""
    if depth == 0:
        return "YELLOW"
    elif depth == 1:
        return "BLUE"
    elif depth == 2:
        return "GREEN"
    else:
        return "RED"


def calc_depth(node: str, parent_of: Dict[str, str]) -> int:
    """Calculate depth of a node in the hierarchy."""
    d = 0
    visited: Set[str] = set()
    while node in parent_of and parent_of[node]:
        if node in visited:
            break
        visited.add(node)
        node = parent_of[node]
        d += 1
    return d


def sanitize_filename(name: str) -> str:
    """Sanitize name for use as filename."""
    if not name:
        return "UNKNOWN"
    clean = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', name)
    clean = clean.replace(' ', '_')
    clean = re.sub(r'_+', '_', clean)
    clean = clean.strip('_')
    if len(clean) > 50:
        clean = clean[:50]
    return clean if clean else "UNKNOWN"


def get_display_name(
    group_key: str,
    group_names: Dict[str, str],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
) -> str:
    """Get display name: ENG preferred, else raw group_key."""
    if group_key == OTHERS_KEY:
        return "Others"
    if group_key == MONSTER_ITEM_KEY:
        return "Monster_Item"

    kor_name = group_names.get(group_key, "")
    if kor_name:
        eng_name, _, _ = get_first_translation(eng_tbl, kor_name)
        if eng_name and is_good_translation(eng_name):
            return sanitize_filename(eng_name)

    if group_key.lower().startswith("itemgroup_"):
        fallback = group_key[10:]
    else:
        fallback = group_key

    return sanitize_filename(fallback)






# =============================================================================
# MASTER ITEM GROUP PARSING
# =============================================================================

def parse_master_groups(path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Parse ItemGroupInfo master file.

    Returns:
        (group_name, parent_of) dicts
    """
    log.info("Parsing ItemGroupInfo master: %s", path)
    root = parse_xml_file(path)
    if root is None:
        log.error("Cannot parse ItemGroupInfo master file")
        return {}, {}

    group_name: Dict[str, str] = {}
    parent_of: Dict[str, str] = {}
    duplicate_count = 0

    for el in root.iter("ItemGroupInfo"):
        sk = el.get("StrKey") or ""
        name = el.get("GroupName") or ""
        if sk:
            if sk not in group_name:
                group_name[sk] = name
            p = el.getparent()
            if p is not None and p.tag == "ItemGroupInfo":
                parent_key = p.get("StrKey") or ""
                if parent_key and sk not in parent_of:
                    parent_of[sk] = parent_key
                elif parent_key:
                    duplicate_count += 1

    for el in root.iter("ChildGroupInfo"):
        child = el.get("StrKey") or ""
        name = el.get("GroupName") or ""
        p = el.getparent()
        if child:
            if name and child not in group_name:
                group_name[child] = name
            if p is not None and p.tag == "ItemGroupInfo":
                parent_key = p.get("StrKey") or ""
                if parent_key and child not in parent_of:
                    parent_of[child] = parent_key
                elif parent_key:
                    duplicate_count += 1

    log.info(
        "Groups parsed: %d  |  parent links: %d  |  duplicates ignored: %d",
        len(group_name), len(parent_of), duplicate_count
    )
    return group_name, parent_of


# =============================================================================
# ITEM SCANNER WITH KNOWLEDGE
# =============================================================================

def scan_items_with_knowledge(
    folder: Path,
    knowledge_map: Dict[str, Tuple[str, str, str]],
    knowledge_name_index: Dict[str, List[Tuple[str, str, str]]],
) -> Tuple[Dict[str, List[str]], Dict[str, str], Dict[str, ItemEntry]]:
    """Scan item folder and build entries with knowledge data resolved.

    Pass 1: KnowledgeKey -> knowledge_map (direct key lookup)
    Pass 2: ItemName -> knowledge_name_index (identical name match)

    Returns:
        group_items: {group_key: [item_strkey, ...]} -- for clustering
        group_names: {group_key: GroupName} -- scanned from XML
        items: {item_strkey: ItemEntry} -- full entry for Excel writing
    """
    log.info("Scanning items with knowledge: %s", folder)
    group_items: Dict[str, List[str]] = {}
    scanned_group_names: Dict[str, str] = {}
    items: Dict[str, ItemEntry] = {}
    pass2_hits = 0

    for path in iter_xml_files(folder):
        root = parse_xml_file(path)
        if root is None:
            continue

        source_file = path.name

        for g_el in root.iter("ItemGroupInfo"):
            g_key = g_el.get("StrKey") or ""
            g_name = g_el.get("GroupName") or ""
            if not g_key:
                continue
            if g_name and g_key not in scanned_group_names:
                scanned_group_names[g_key] = g_name

            bucket = group_items.setdefault(g_key, [])
            for item in g_el.iter("ItemInfo"):
                ik = item.get("StrKey") or ""
                if not ik:
                    continue

                item_name = item.get("ItemName") or ""
                item_desc = item.get("ItemDesc") or ""

                # Pass 0: Inline Knowledge children (direct child nodes of ItemInfo)
                child_knowledge_entries = []
                for child in item.findall("Knowledge"):
                    child_name = child.get("Name") or ""
                    child_desc = child.get("Desc") or ""
                    child_strkey = (child.get("StrKey") or "").lower()
                    if child_strkey and child_strkey in knowledge_map:
                        kname, kdesc, ksrc = knowledge_map[child_strkey]
                        child_knowledge_entries.append((kname, kdesc, ksrc))
                    elif child_name or child_desc:
                        child_knowledge_entries.append((child_name, child_desc, source_file))

                knowledge_key = _find_knowledge_key(item)

                # Pass 1: Resolve knowledge data via KnowledgeKey
                knowledge_name = ""
                knowledge_desc = ""
                knowledge_source_file = ""
                pass1_strkey = ""
                if knowledge_key and knowledge_key.lower() in knowledge_map:
                    knowledge_name, knowledge_desc, knowledge_source_file = knowledge_map[knowledge_key.lower()]
                    pass1_strkey = knowledge_key.lower()

                # Pass 2: Identical name match (ItemName == KnowledgeInfo.Name)
                knowledge2_name = ""
                knowledge2_desc = ""
                knowledge2_source_file = ""
                if item_name and item_name in knowledge_name_index:
                    for kn_strkey, kn_desc, kn_src in knowledge_name_index[item_name]:
                        knowledge2_name = item_name
                        knowledge2_desc = kn_desc
                        knowledge2_source_file = kn_src
                        pass2_hits += 1
                        break

                # InspectData: collect from direct children + PageData books
                inspect_entries = _collect_inspect_data(item, knowledge_map)

                # Collect Korean strings for coverage tracking
                _collect_korean_string(item_name)
                _collect_korean_string(item_desc)
                for cname, cdesc, _ in child_knowledge_entries:
                    _collect_korean_string(cname)
                    _collect_korean_string(cdesc)
                _collect_korean_string(knowledge_name)
                _collect_korean_string(knowledge_desc)
                _collect_korean_string(knowledge2_name)
                _collect_korean_string(knowledge2_desc)
                for idesc, ikname, ikdesc, _ in inspect_entries:
                    _collect_korean_string(idesc)
                    _collect_korean_string(ikname)
                    _collect_korean_string(ikdesc)

                bucket.append(ik)
                items[ik] = ItemEntry(
                    item_strkey=ik,
                    item_name_kor=item_name,
                    item_desc_kor=item_desc,
                    knowledge_key=knowledge_key,
                    knowledge_name_kor=knowledge_name,
                    knowledge_desc_kor=knowledge_desc,
                    group_key=g_key,
                    source_file=source_file,
                    knowledge_source_file=knowledge_source_file,
                    knowledge2_name_kor=knowledge2_name,
                    knowledge2_desc_kor=knowledge2_desc,
                    knowledge2_source_file=knowledge2_source_file,
                    child_knowledge_entries=child_knowledge_entries,
                    inspect_entries=inspect_entries,
                )

    total_items = sum(len(v) for v in group_items.values())
    log.info("Items scanned: %d items in %d groups (Pass 2 hits: %d)",
             total_items, len(group_items), pass2_hits)
    return group_items, scanned_group_names, items


# =============================================================================
# CLUSTERING FUNCTIONS (adapted from item.py for List[str] group_items)
# =============================================================================

def apply_depth_based_clustering(
    group_items: Dict[str, List[str]],
    parent_of: Dict[str, str],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
) -> Dict[str, Dict[str, List[str]]]:
    """Apply depth-based clustering to merge small groups into parents."""
    log.info("Applying depth-based clustering (threshold: %d items)", MERGE_UP_THRESHOLD)

    group_item_count: Dict[str, int] = {}
    group_depth: Dict[str, int] = {}
    for group_key, items_list in group_items.items():
        group_item_count[group_key] = len(items_list)
        group_depth[group_key] = calc_depth(group_key, parent_of)

    log.info("Initial group counts by depth:")
    depth_summary: Dict[int, Tuple[int, int]] = defaultdict(lambda: (0, 0))
    for gk, count in group_item_count.items():
        d = group_depth[gk]
        g, i = depth_summary[d]
        depth_summary[d] = (g + 1, i + count)
    for d in sorted(depth_summary):
        groups, items = depth_summary[d]
        log.info("  Depth %d (%s): %d groups, %d items", d, get_depth_color(d), groups, items)

    merge_target: Dict[str, str] = {gk: gk for gk in group_items}
    groups_by_depth = sorted(group_items.keys(), key=lambda g: -group_depth.get(g, 0))
    accumulated_counts = dict(group_item_count)
    merge_count = skip_at_ceiling = skip_above_threshold = skip_no_parent = 0

    for gk in groups_by_depth:
        depth = group_depth[gk]
        count = accumulated_counts.get(gk, 0)
        if depth <= MIN_FOLDER_DEPTH:
            skip_at_ceiling += 1
            continue
        if count < MERGE_UP_THRESHOLD:
            parent = parent_of.get(gk, "")
            if parent:
                pdepth = group_depth.get(parent, 0)
                if pdepth < depth:
                    merge_target[gk] = parent
                    merge_count += 1
                    accumulated_counts[parent] = accumulated_counts.get(parent, 0) + count
                    log.debug("  MERGE: %s->%s", get_display_name(gk, group_names, eng_tbl),
                              get_display_name(parent, group_names, eng_tbl))
            else:
                skip_no_parent += 1
        else:
            skip_above_threshold += 1

    log.info("Merge: %d, at ceiling: %d, >=thr: %d, no parent: %d",
             merge_count, skip_at_ceiling, skip_above_threshold, skip_no_parent)

    def resolve(gk: str, seen: Set[str]) -> str:
        if gk in seen:
            return gk
        tgt = merge_target.get(gk, gk)
        if tgt == gk:
            return gk
        seen.add(gk)
        return resolve(tgt, seen)

    for gk in merge_target:
        merge_target[gk] = resolve(gk, set())

    def find_folder_at_blue(gk: str) -> str:
        cur = gk
        seen: Set[str] = set()
        while True:
            if cur in seen:
                return OTHERS_KEY
            seen.add(cur)
            d = calc_depth(cur, parent_of)
            if d == 1:
                return cur
            if d == 0:
                return OTHERS_KEY
            cur = parent_of.get(cur, "")
            if not cur:
                return OTHERS_KEY

    structure: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for group_key, items_list in group_items.items():
        sub = merge_target[group_key]
        folder = find_folder_at_blue(sub)
        for ik in items_list:  # List[str] — adapted from item.py's List[Tuple]
            structure[folder][sub].append(ik)

    log.info("Final structure:")
    for fk, subs in structure.items():
        log.info("  %s: %d subgroups, %d items",
                 get_display_name(fk, group_names, eng_tbl), len(subs),
                 sum(len(it) for it in subs.values()))
    return dict(structure)


def extract_monster_items(
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
) -> Dict[str, Dict[str, List[str]]]:
    """Extract mon_ items to Monster_Item folder."""
    log.info("Extracting mon_ items to Monster_Item")
    new_struct: Dict[str, Dict[str, List[str]]] = {}
    monster: Dict[str, List[str]] = {}

    for fk, subs in structure.items():
        keep: Dict[str, List[str]] = {}
        for sk, items in subs.items():
            if MONSTER_SUBSTRING in sk.lower():
                monster[sk] = items
                log.debug("  Extracted %s", sk)
            else:
                keep[sk] = items
        if keep:
            new_struct[fk] = keep

    if monster:
        new_struct[MONSTER_ITEM_KEY] = monster
        log.info("  %d subgroups moved to Monster_Item", len(monster))

    return new_struct


def consolidate_small_folders(
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
) -> Dict[str, Dict[str, List[str]]]:
    """Consolidate small folders into Others."""
    log.info("Consolidating small folders (<%d)", FOLDER_MIN_THRESHOLD)
    totals = {fk: sum(len(it) for it in subs.values()) for fk, subs in structure.items()}
    small = []
    keep = []
    special = []

    for fk, tot in totals.items():
        if fk in (OTHERS_KEY, MONSTER_ITEM_KEY):
            special.append(fk)
        elif tot < FOLDER_MIN_THRESHOLD:
            small.append(fk)
            log.info("  Small: %s (%d)->Others",
                     get_display_name(fk, group_names, eng_tbl), tot)
        else:
            keep.append(fk)

    if not small:
        log.info("  None to consolidate")
        return structure

    new_struct: Dict[str, Dict[str, List[str]]] = {}
    for fk in keep + special:
        new_struct[fk] = structure[fk]

    others_sub: Dict[str, List[str]] = {}
    if OTHERS_KEY in new_struct:
        others_sub = dict(new_struct[OTHERS_KEY])
    for fk in small:
        for sk, items in structure[fk].items():
            others_sub.setdefault(sk, []).extend(items)
    if others_sub:
        new_struct[OTHERS_KEY] = others_sub
        log.info("  Others now %d subgroups, %d items",
                 len(others_sub), sum(len(it) for it in others_sub.values()))

    return new_struct


# =============================================================================
# TEXT FILE WRITER
# =============================================================================

def write_text_files(
    out_folder: Path,
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
) -> Dict[str, List[Tuple[str, str, List[str]]]]:
    """Write text files with /create item commands."""
    out_folder.mkdir(parents=True, exist_ok=True)
    folder_files: Dict[str, List[Tuple[str, str, List[str]]]] = {}

    for fk, subs in sorted(structure.items()):
        folder_display = get_display_name(fk, group_names, eng_tbl) or fk
        folder_path = out_folder / folder_display
        folder_path.mkdir(parents=True, exist_ok=True)
        file_list: List[Tuple[str, str, List[str]]] = []

        if fk == OTHERS_KEY:
            all_items = sorted({ik for lst in subs.values() for ik in lst})
            chunks = [all_items[i:i+MAX_COMMANDS_PER_FILE]
                      for i in range(0, len(all_items), MAX_COMMANDS_PER_FILE)]
            for idx, chunk in enumerate(chunks, 1):
                fn = f"Others_{idx}.txt"
                with open(folder_path/fn, "w", encoding="utf-8") as f:
                    f.write("/reset inventory\n/expandinventory 2 300\n")
                    for ik in chunk:
                        f.write(f"/create item {ik}\n")
                    f.write("/reset alert\n")
                file_list.append((fn, OTHERS_KEY, chunk))
            log.info("Folder %s: %d files (%d items)",
                     folder_display, len(chunks), len(all_items))
        else:
            disp_map: Dict[str, List[str]] = {}
            key_map: Dict[str, str] = {}
            for sk, lst in subs.items():
                disp = get_display_name(sk, group_names, eng_tbl) or sk
                disp_map.setdefault(disp, []).extend(lst)
                key_map.setdefault(disp, sk)
            for disp, all_iks in sorted(disp_map.items()):
                chunked = [all_iks[i:i+MAX_COMMANDS_PER_FILE]
                           for i in range(0, len(all_iks), MAX_COMMANDS_PER_FILE)]
                for idx, chunk in enumerate(chunked, 1):
                    fn = disp + (f"_{idx}.txt" if len(chunked) > 1 else ".txt")
                    with open(folder_path/fn, "w", encoding="utf-8") as f:
                        f.write("/reset inventory\n/expandinventory 2 300\n")
                        for ik in sorted(set(chunk)):
                            f.write(f"/create item {ik}\n")
                        f.write("/reset alert\n")
                    file_list.append((fn, key_map[disp], chunk))
            log.info("Folder %s: %d text files", folder_display, len(file_list))

        folder_files[folder_display] = file_list
    return folder_files


# =============================================================================
# EXCEL WRITER (row-per-text format)
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_fill_a = PatternFill("solid", fgColor="E2EFDA")  # Light green
_fill_b = PatternFill("solid", fgColor="FCE4D6")  # Light orange


def write_item_excel(
    items: Dict[str, ItemEntry],
    folder_files: Dict[str, List[Tuple[str, str, List[str]]]],
    lang_tbl: Dict[str, List[Tuple[str, str]]],
    lang_code: str,
    export_index: Dict[str, Set[str]],
    output_path: Path,
) -> None:
    """Write Item Excel with one row per text field.

    8 columns: DataType | Filename | SourceText (KR) | Translation | STATUS | COMMENT | SCREENSHOT | STRINGID

    Per-item row generation (strict order):
    1.  ItemData              -- item_name_kor (always output)
    2.  ItemData              -- item_desc_kor (always output)
    2b. ChildKnowledgeData    -- inline child Name/Desc pairs (skip if none) [Pass 0]
    3.  KnowledgeData         -- knowledge_name_kor (skip if empty)  [Pass 1: KnowledgeKey]
    4.  KnowledgeData         -- knowledge_desc_kor (skip if empty)  [Pass 1: KnowledgeKey]
    5.  KnowledgeData2        -- knowledge2_name_kor (skip if empty) [Pass 2: identical name]
    6.  KnowledgeData2        -- knowledge2_desc_kor (skip if empty) [Pass 2: identical name]
    7.  InspectData           -- inspect Desc (per entry, sequential) [Pattern A or B]
    7b. InspectKnowledgeData  -- linked knowledge Name (if RewardKnowledgeKey found)
    7c. InspectKnowledgeData  -- linked knowledge Desc (if RewardKnowledgeKey found)
    """
    wb = Workbook()
    wb.remove(wb.active)
    code = lang_code.upper()

    headers = [
        "DataType",
        "Filename",
        "SourceText (KR)",
        f"Translation ({code})",
        "STATUS",
        "COMMENT",
        "SCREENSHOT",
        "STRINGID",
    ]

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)

    # ------------------------------------------------------------------
    # PRE-RESOLVE: consume StringIDs in DOCUMENT ORDER (before sorting).
    # items dict preserves insertion order (= XML scan order).  The
    # sorted write loop below would break the consumer's pointer, so we
    # resolve here first and store results keyed by (item_strkey, field).
    # ------------------------------------------------------------------
    pre: Dict[Tuple[str, str], Tuple[str, str, str]] = {}  # (ik, field) -> (trans, sid, str_origin)
    for ik, entry in items.items():  # insertion order = document order
        pre[(ik, "item_name")] = resolve_translation(
            entry.item_name_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        pre[(ik, "item_desc")] = resolve_translation(
            entry.item_desc_kor, lang_tbl, entry.source_file, export_index, consumer=consumer)
        # Pass 0: inline child knowledge (before Pass 1 to preserve document order)
        for i, (cname, cdesc, csrc) in enumerate(entry.child_knowledge_entries):
            if cname:
                pre[(ik, f"ck_{i}_name")] = resolve_translation(
                    cname, lang_tbl, csrc, export_index, consumer=consumer)
            if cdesc:
                pre[(ik, f"ck_{i}_desc")] = resolve_translation(
                    cdesc, lang_tbl, csrc, export_index, consumer=consumer)
        if entry.knowledge_name_kor:
            pre[(ik, "knowledge_name")] = resolve_translation(
                entry.knowledge_name_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge_desc_kor:
            pre[(ik, "knowledge_desc")] = resolve_translation(
                entry.knowledge_desc_kor, lang_tbl, entry.knowledge_source_file, export_index, consumer=consumer)
        if entry.knowledge2_name_kor:
            pre[(ik, "knowledge2_name")] = resolve_translation(
                entry.knowledge2_name_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)
        if entry.knowledge2_desc_kor:
            pre[(ik, "knowledge2_desc")] = resolve_translation(
                entry.knowledge2_desc_kor, lang_tbl, entry.knowledge2_source_file, export_index, consumer=consumer)
        # InspectData + InspectKnowledgeData (after all item-level knowledge)
        for i, (idesc, ikname, ikdesc, iksrc) in enumerate(entry.inspect_entries):
            pre[(ik, f"inspect_{i}_desc")] = resolve_translation(
                idesc, lang_tbl, entry.source_file, export_index, consumer=consumer)
            if ikname:
                pre[(ik, f"inspect_{i}_kname")] = resolve_translation(
                    ikname, lang_tbl, iksrc, export_index, consumer=consumer)
            if ikdesc:
                pre[(ik, f"inspect_{i}_kdesc")] = resolve_translation(
                    ikdesc, lang_tbl, iksrc, export_index, consumer=consumer)

    if consumer.warnings:
        log.warning("StringID overruns during pre-resolve: %d", consumer.warnings)

    # ------------------------------------------------------------------
    # KNOWLEDGEDATA2 DEDUP: per-cluster surgical removal of pure
    # duplicates where (SourceText, Translation, StringID) matches a
    # primary-type row within the SAME entity.
    # ------------------------------------------------------------------
    dedup_skip: Set[Tuple[str, str]] = set()  # (ik, field) pairs to skip
    dedup_count = 0
    dedup_entities = 0
    for ik, entry in items.items():
        # Collect primary triples: (kor_text, translation, stringid)
        primary_triples: Set[Tuple[str, str, str]] = set()
        # ItemData
        t, s, _ = pre[(ik, "item_name")]
        if entry.item_name_kor:
            primary_triples.add((entry.item_name_kor, t, s))
        t, s, _ = pre[(ik, "item_desc")]
        if entry.item_desc_kor:
            primary_triples.add((entry.item_desc_kor, t, s))
        # ChildKnowledgeData (Pass 0)
        for i, (cname, cdesc, _csrc) in enumerate(entry.child_knowledge_entries):
            if cname:
                t, s, _ = pre.get((ik, f"ck_{i}_name"), ("", "", ""))
                primary_triples.add((cname, t, s))
            if cdesc:
                t, s, _ = pre.get((ik, f"ck_{i}_desc"), ("", "", ""))
                primary_triples.add((cdesc, t, s))
        # KnowledgeData (Pass 1)
        if entry.knowledge_name_kor:
            t, s, _ = pre[(ik, "knowledge_name")]
            primary_triples.add((entry.knowledge_name_kor, t, s))
        if entry.knowledge_desc_kor:
            t, s, _ = pre[(ik, "knowledge_desc")]
            primary_triples.add((entry.knowledge_desc_kor, t, s))
        # InspectData + InspectKnowledgeData
        for i, (idesc, ikname, ikdesc, _iksrc) in enumerate(entry.inspect_entries):
            t, s, _ = pre.get((ik, f"inspect_{i}_desc"), ("", "", ""))
            primary_triples.add((idesc, t, s))
            if ikname:
                t, s, _ = pre.get((ik, f"inspect_{i}_kname"), ("", "", ""))
                primary_triples.add((ikname, t, s))
            if ikdesc:
                t, s, _ = pre.get((ik, f"inspect_{i}_kdesc"), ("", "", ""))
                primary_triples.add((ikdesc, t, s))

        # Check KnowledgeData2 fields against primary triples
        entity_deduped = False
        if entry.knowledge2_name_kor:
            t, s, _ = pre[(ik, "knowledge2_name")]
            if (entry.knowledge2_name_kor, t, s) in primary_triples:
                dedup_skip.add((ik, "knowledge2_name"))
                dedup_count += 1
                entity_deduped = True
                log.debug("KnowledgeData2 dedup: '%s' in %s — matches primary data",
                          entry.knowledge2_name_kor[:40], ik)
        if entry.knowledge2_desc_kor:
            t, s, _ = pre[(ik, "knowledge2_desc")]
            if (entry.knowledge2_desc_kor, t, s) in primary_triples:
                dedup_skip.add((ik, "knowledge2_desc"))
                dedup_count += 1
                entity_deduped = True
                log.debug("KnowledgeData2 dedup: '%s' in %s — matches primary data",
                          entry.knowledge2_desc_kor[:40], ik)
        if entity_deduped:
            dedup_entities += 1

    if dedup_count:
        log.info("KnowledgeData2 dedup: removed %d duplicate rows across %d entities",
                 dedup_count, dedup_entities)

    for folder_display, flist in sorted(folder_files.items()):
        if not flist:
            continue

        # Create sheet (handle name length and duplicates)
        base = folder_display[:31]
        title = base
        cnt = 1
        while title in {ws.title for ws in wb.worksheets}:
            title = f"{base[:28]}_{cnt}"
            cnt += 1
        ws = wb.create_sheet(title)

        # Write header row
        for col_idx, txt in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, txt)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        excel_row = 2
        current_fill = _fill_a
        last_ik = None

        for fn, subgroup_key, iks in flist:
            for ik in sorted(iks):
                entry = items.get(ik)
                if not entry:
                    continue

                # Alternate fill per item for visual grouping
                if last_ik is not None and ik != last_ik:
                    current_fill = _fill_b if current_fill == _fill_a else _fill_a
                last_ik = ik

                def _write_row(data_type: str, kor_text: str, trans: str, sid: str) -> None:
                    """Write a single data row using pre-resolved translation."""
                    nonlocal excel_row
                    vals = [data_type, fn, br_to_newline(kor_text), br_to_newline(trans), "", "", "", sid]
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(excel_row, ci, val)
                        cell.fill = current_fill
                        cell.border = THIN_BORDER
                        if ci == 5:  # STATUS
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    # STRINGID as text format (prevent scientific notation)
                    ws.cell(excel_row, 8).number_format = '@'
                    excel_row += 1

                # 1. ItemData -- ItemName (always output)
                t, s, so = pre[(ik, "item_name")]
                _write_row("ItemData", so if so else entry.item_name_kor, t, s)

                # 2. ItemData -- ItemDesc (always output)
                t, s, so = pre[(ik, "item_desc")]
                _write_row("ItemData", so if so else entry.item_desc_kor, t, s)

                # 2b. ChildKnowledgeData -- inline Knowledge child Name/Desc (Pass 0)
                for i, (cname, cdesc, _csrc) in enumerate(entry.child_knowledge_entries):
                    if cname:
                        t, s, so = pre.get((ik, f"ck_{i}_name"), ("", "", ""))
                        _write_row("ChildKnowledgeData", so if so else cname, t, s)
                    if cdesc:
                        t, s, so = pre.get((ik, f"ck_{i}_desc"), ("", "", ""))
                        _write_row("ChildKnowledgeData", so if so else cdesc, t, s)

                # 3. KnowledgeData -- Name (skip if empty)
                if entry.knowledge_name_kor:
                    t, s, so = pre[(ik, "knowledge_name")]
                    _write_row("KnowledgeData", so if so else entry.knowledge_name_kor, t, s)

                # 4. KnowledgeData -- Desc (skip if empty)
                if entry.knowledge_desc_kor:
                    t, s, so = pre[(ik, "knowledge_desc")]
                    _write_row("KnowledgeData", so if so else entry.knowledge_desc_kor, t, s)

                # 5. KnowledgeData2 -- Name (Pass 2: identical name match, skip if empty or deduped)
                if entry.knowledge2_name_kor and (ik, "knowledge2_name") not in dedup_skip:
                    t, s, so = pre[(ik, "knowledge2_name")]
                    _write_row("KnowledgeData2", so if so else entry.knowledge2_name_kor, t, s)

                # 6. KnowledgeData2 -- Desc (Pass 2: identical name match, skip if empty or deduped)
                if entry.knowledge2_desc_kor and (ik, "knowledge2_desc") not in dedup_skip:
                    t, s, so = pre[(ik, "knowledge2_desc")]
                    _write_row("KnowledgeData2", so if so else entry.knowledge2_desc_kor, t, s)

                # 7. InspectData + InspectKnowledgeData (Pattern A: recipe, Pattern B: book)
                for i, (idesc, ikname, ikdesc, _iksrc) in enumerate(entry.inspect_entries):
                    # InspectData -- the Desc text from <InspectData> element
                    t, s, so = pre.get((ik, f"inspect_{i}_desc"), ("", "", ""))
                    _write_row("InspectData", so if so else idesc, t, s)
                    # InspectKnowledgeData -- linked via RewardKnowledgeKey (if exists)
                    if ikname:
                        t, s, so = pre.get((ik, f"inspect_{i}_kname"), ("", "", ""))
                        _write_row("InspectKnowledgeData", so if so else ikname, t, s)
                    if ikdesc:
                        t, s, so = pre.get((ik, f"inspect_{i}_kdesc"), ("", "", ""))
                        _write_row("InspectKnowledgeData", so if so else ikdesc, t, s)

        # Sheet cosmetics
        if excel_row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"
        ws.freeze_panes = "A2"

        # Add STATUS drop-down (column 5)
        add_status_dropdown(ws, col=5)

        # Auto-fit column widths
        autofit_worksheet(ws)

        log.info("  Sheet '%s': %d rows", title, excel_row - 2)

    wb.save(output_path)
    log.info("Item Excel saved: %s", output_path.name)


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_item_datasheets() -> Dict:
    """
    Generate Item datasheets for all languages.

    Pipeline:
    1. Load language tables
    2. Load knowledge data (Name + Desc + source_file)
    3. Parse master groups (ItemGroupInfo hierarchy)
    4. Scan items with knowledge resolution
    5. Cluster (depth-based -> monster extraction -> small folder consolidation)
    6. Write text files (command files for testers)
    7. Get EXPORT index (for StringID disambiguation)
    8. For each language: write row-per-text Excel

    Returns:
        Dict with results: {
            "category": "Item",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Item",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("Item Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "ItemData_Map_All"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    item_folder = RESOURCE_FOLDER / "iteminfo"
    if not item_folder.exists():
        item_folder = RESOURCE_FOLDER  # Fallback

    knowledge_folder = RESOURCE_FOLDER / "knowledgeinfo"

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    # Create text files output folder
    textfiles_folder = output_folder / "ExecuteFiles"
    textfiles_folder.mkdir(exist_ok=True)

    try:
        # 1. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)
        eng_tbl = lang_tables.get("eng", {})

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        # 2. Load knowledge data (map + name index for Pass 2)
        knowledge_map, knowledge_name_index = load_knowledge_data(knowledge_folder)

        # 3. Parse master groups
        itemgroupinfo_file = item_folder / "itemgroupinfo.staticinfo.xml"
        if itemgroupinfo_file.exists():
            master_names, parent_of = parse_master_groups(itemgroupinfo_file)
        else:
            master_names, parent_of = {}, {}
            log.warning("ItemGroupInfo master file not found")

        # 4. Scan items with knowledge (Pass 1: KnowledgeKey, Pass 2: identical name)
        group_items, scanned_names, items = scan_items_with_knowledge(
            item_folder, knowledge_map, knowledge_name_index
        )

        all_names = {**scanned_names, **master_names}
        all_names[OTHERS_KEY] = "Others"
        all_names[MONSTER_ITEM_KEY] = "Monster_Item"

        if not items:
            result["errors"].append("No item data found!")
            log.warning("No item data found!")
            return result

        # 5. Cluster
        structure = apply_depth_based_clustering(group_items, parent_of, all_names, eng_tbl)
        structure = extract_monster_items(structure, all_names, eng_tbl)
        structure = consolidate_small_folders(structure, all_names, eng_tbl)

        # 6. Write text files (command files)
        folder_files = write_text_files(textfiles_folder, structure, all_names, eng_tbl)

        # 7. Get EXPORT index
        export_index = get_export_index()

        # 8. Generate Excel per language
        log.info("Processing languages...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Language %s", idx, total, code.upper())
            excel_path = output_folder / f"Item_LQA_{code.upper()}.xlsx"
            write_item_excel(items, folder_files, tbl, code, export_index, excel_path)
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Item generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_item_datasheets()
    print(f"\nResult: {result}")
