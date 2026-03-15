from __future__ import annotations

"""
Region Datasheet Generator
================================
Region datasheet with DisplayName from RegionInfo.

Per FactionNode, outputs up to 3 rows:
  1. KnowledgeInfo.Name  (at depth)     — same as Region
  2. RegionInfo.DisplayName (at depth)   — NEW: only if different from Knowledge Name
  3. KnowledgeInfo.Desc  (at depth + 1)  — same as Region

All hierarchy, indentation, colors, and shop data are IDENTICAL to Region.

Linkage:
  FactionNode.KnowledgeKey → KnowledgeInfo.(Name, Desc)
  FactionNode.KnowledgeKey → RegionInfo.KnowledgeKey → RegionInfo.DisplayName
"""

import re
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    is_good_translation,
    br_to_newline,
    autofit_worksheet,
    THIN_BORDER,
    get_first_translation,
    resolve_translation,
    get_export_index,
    get_ordered_export_index,
    StringIdConsumer,
)

log = get_logger("RegionGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# Empty Faction identifier (Korean)
EMPTY_FACTION_NAME = "세력 없음"

# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class FactionNodeData:
    """A single FactionNode with nested children."""
    strkey: str
    name: str              # Display name (from KnowledgeInfo.Name lookup)
    original_name: str     # Original Name attribute (fallback)
    description: str       # From KnowledgeInfo.Desc (via KnowledgeKey)
    knowledge_key: str
    node_type: str         # Main, Sub, etc.
    world_position: str = ""  # WorldPosition attribute from FactionNode
    source_file: str = ""
    children: List["FactionNodeData"] = field(default_factory=list)


@dataclass
class FactionData:
    """A Faction containing multiple FactionNodes."""
    strkey: str
    name: str              # Faction Name attribute
    knowledge_key: str
    source_file: str = ""
    nodes: List[FactionNodeData] = field(default_factory=list)


@dataclass
class FactionGroupData:
    """A FactionGroup (top-level, becomes sheet tab)."""
    strkey: str
    group_name: str        # GroupName attribute - used as sheet tab name
    knowledge_key: str
    source_file: str = ""
    factions: List[FactionData] = field(default_factory=list)


@dataclass
class ShopStage:
    """A single shop stage."""
    strkey: str
    name: str
    description: str
    category: str = ""
    npc_key: str = ""
    dev_comment: str = ""
    alias_name: str = ""
    source_file: str = ""


@dataclass
class ShopGroup:
    """A shop group."""
    tag_name: str
    name: str
    group: str
    source_file: str = ""
    stages: List[ShopStage] = field(default_factory=list)


@dataclass
class KnowledgeContentEntry:
    """A KnowledgeInfo from knowledgeinfo_contents for the Knowledge_Contents tab."""
    strkey: str
    name: str
    description: str
    world_position: str = ""    # Resolved via 3-hop chain
    source_file: str = ""


@dataclass
class KnowledgeContentGroup:
    """A KnowledgeGroupInfo from knowledgeinfo_contents."""
    strkey: str
    group_name: str             # GroupName attribute
    source_file: str = ""
    entries: List[KnowledgeContentEntry] = field(default_factory=list)


# =============================================================================
# STYLING
# =============================================================================

STYLES = {
    "Faction": {
        "fill": PatternFill("solid", fgColor="4472C4"),  # Dark blue
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "row_height": 35,
    },
    "FactionNode_Main": {
        "fill": PatternFill("solid", fgColor="5B9BD5"),  # Medium blue
        "font": Font(bold=True, size=11, color="FFFFFF"),
        "row_height": 30,
    },
    "FactionNode_Sub": {
        "fill": PatternFill("solid", fgColor="B4C6E7"),  # Light blue
        "font": Font(bold=True),
        "row_height": None,
    },
    "FactionNode": {
        "fill": PatternFill("solid", fgColor="D6DCE4"),  # Very light blue
        "font": Font(),
        "row_height": None,
    },
    "Description": {
        "fill": PatternFill("solid", fgColor="F5F5F5"),  # Light gray
        "font": Font(italic=False),
        "row_height": None,
    },
    "ShopGroup": {
        "fill": PatternFill("solid", fgColor="00B0B0"),  # Teal
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "row_height": 35,
    },
    "ShopStage": {
        "fill": PatternFill("solid", fgColor="E0F7FA"),  # Light cyan
        "font": Font(bold=True),
        "row_height": None,
    },
}

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_default_font = Font()
_default_fill = PatternFill("solid", fgColor="FFFFFF")


def get_style(style_type: str) -> Tuple[PatternFill, Font, Optional[float]]:
    """Get style for a row type. Returns (fill, font, row_height)."""
    style = STYLES.get(style_type, {})
    return (
        style.get("fill", _default_fill),
        style.get("font", _default_font),
        style.get("row_height"),
    )


# =============================================================================
# KNOWLEDGE NAME LOOKUP
# =============================================================================

def build_knowledge_name_lookup(folder: Path) -> Dict[str, Tuple[str, str]]:
    """
    Build lookup: KnowledgeInfo.StrKey → (Name, Desc)

    Used to get correct display names AND descriptions for FactionNodes
    via their KnowledgeKey linkage.
    """
    log.info("Building Knowledge Name+Desc lookup...")

    name_lookup: Dict[str, Tuple[str, str]] = {}
    duplicates = 0

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for ki in root_el.iter("KnowledgeInfo"):
            strkey = ki.get("StrKey") or ""
            name = ki.get("Name") or ""

            if not strkey or not name:
                continue

            if strkey.lower() in name_lookup:
                duplicates += 1
                continue

            desc = ki.get("Desc") or ""
            name_lookup[strkey.lower()] = (name, desc)

    log.info("  → %d entries (%d duplicates ignored)", len(name_lookup), duplicates)
    return name_lookup


# =============================================================================
# FACTION PARSING
# =============================================================================

def parse_faction_node_recursive(
    elem,
    knowledge_lookup: Dict[str, Tuple[str, str]],
    global_seen: Set[str],
    depth: int = 0,
    source_file: str = ""
) -> Optional[FactionNodeData]:
    """Parse a FactionNode element and all its nested FactionNode children."""
    strkey = elem.get("StrKey") or ""
    original_name = elem.get("Name") or ""
    knowledge_key = elem.get("KnowledgeKey") or elem.get("RewardKnowledgeKey") or ""
    description = elem.get("Desc") or ""
    node_type = elem.get("Type") or ""
    world_position = elem.get("WorldPosition") or ""

    # Resolve display name + description via KnowledgeInfo lookup
    if knowledge_key and knowledge_key.lower() in knowledge_lookup:
        ki_name, ki_desc = knowledge_lookup[knowledge_key.lower()]
        display_name = ki_name
        # Pull description from KnowledgeInfo if FactionNode has none
        if not description and ki_desc:
            description = ki_desc
    else:
        display_name = original_name

    if not display_name:
        return None

    # Duplicate check
    if strkey:
        if strkey in global_seen:
            return None
        global_seen.add(strkey)

    # Collect Korean strings for coverage tracking
    _collect_korean_string(display_name)
    _collect_korean_string(description)

    node = FactionNodeData(
        strkey=strkey,
        name=display_name,
        original_name=original_name,
        description=description,
        knowledge_key=knowledge_key,
        node_type=node_type,
        world_position=world_position,
        source_file=source_file,
    )

    # Parse nested FactionNodes
    for child_elem in elem:
        if child_elem.tag == "FactionNode":
            child_node = parse_faction_node_recursive(
                child_elem, knowledge_lookup, global_seen, depth + 1, source_file
            )
            if child_node:
                node.children.append(child_node)

    return node


def parse_faction_element(
    elem,
    knowledge_lookup: Dict[str, Tuple[str, str]],
    global_seen: Set[str],
    source_file: str = ""
) -> Optional[FactionData]:
    """Parse a Faction element and all its FactionNode children."""
    strkey = elem.get("StrKey") or ""
    name = elem.get("Name") or ""
    knowledge_key = elem.get("KnowledgeKey") or elem.get("RewardKnowledgeKey") or ""

    if not name:
        return None

    # Collect Korean string for coverage tracking
    _collect_korean_string(name)

    faction = FactionData(
        strkey=strkey,
        name=name,
        knowledge_key=knowledge_key,
        source_file=source_file,
    )

    # Parse direct FactionNode children
    for child_elem in elem:
        if child_elem.tag == "FactionNode":
            node = parse_faction_node_recursive(child_elem, knowledge_lookup, global_seen, 0, source_file)
            if node:
                faction.nodes.append(node)

    return faction


def parse_faction_group_element(
    elem,
    knowledge_lookup: Dict[str, Tuple[str, str]],
    global_seen: Set[str],
    source_file: str = ""
) -> Optional[FactionGroupData]:
    """Parse a FactionGroup element and all its Faction children."""
    strkey = elem.get("StrKey") or ""
    group_name = elem.get("GroupName") or ""
    knowledge_key = elem.get("KnowledgeKey") or elem.get("RewardKnowledgeKey") or ""

    if not group_name:
        return None

    # Collect Korean string for coverage tracking
    _collect_korean_string(group_name)

    faction_group = FactionGroupData(
        strkey=strkey,
        group_name=group_name,
        knowledge_key=knowledge_key,
        source_file=source_file,
    )

    # Parse Faction children
    for child_elem in elem:
        if child_elem.tag == "Faction":
            faction = parse_faction_element(child_elem, knowledge_lookup, global_seen, source_file)
            if faction:
                faction_group.factions.append(faction)

    return faction_group


def count_nodes_recursive(nodes: List[FactionNodeData]) -> int:
    """Count total nodes including nested children."""
    count = len(nodes)
    for node in nodes:
        count += count_nodes_recursive(node.children)
    return count


def parse_all_faction_groups(
    folder: Path,
    knowledge_lookup: Dict[str, Tuple[str, str]],
    global_seen: Set[str]
) -> List[FactionGroupData]:
    """Parse all FactionGroup elements from all XML files."""
    log.info("Parsing FactionGroup → Faction → FactionNode hierarchy...")

    faction_groups: List[FactionGroupData] = []

    total_factions = 0
    total_nodes = 0

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue
        source_file = path.name

        for fg_elem in root_el.iter("FactionGroup"):
            fg = parse_faction_group_element(fg_elem, knowledge_lookup, global_seen, source_file)
            if fg and fg.factions:
                faction_groups.append(fg)

                for faction in fg.factions:
                    total_factions += 1
                    total_nodes += count_nodes_recursive(faction.nodes)

    log.info("  → %d FactionGroups, %d Factions, %d FactionNodes",
             len(faction_groups), total_factions, total_nodes)

    return faction_groups


def parse_standalone_factions(
    folder: Path,
    knowledge_lookup: Dict[str, Tuple[str, str]],
    global_seen: Set[str],
) -> List[FactionData]:
    """Parse Faction elements that are NOT inside any FactionGroup."""
    log.info("Parsing standalone Factions (not in any FactionGroup)...")

    standalone_factions: List[FactionData] = []

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue
        source_file = path.name

        # Collect all Faction StrKeys that ARE inside FactionGroups
        factions_in_groups: Set[str] = set()
        for fg_elem in root_el.iter("FactionGroup"):
            for faction_elem in fg_elem.iter("Faction"):
                sk = faction_elem.get("StrKey") or ""
                if sk:
                    factions_in_groups.add(sk)

        # Find Factions that are direct children of ROOT
        for child in root_el:
            if child.tag == "Faction":
                strkey = child.get("StrKey") or ""

                # Skip factions already in groups (avoid double counting)
                if strkey and strkey in factions_in_groups:
                    continue

                faction = parse_faction_element(child, knowledge_lookup, global_seen, source_file)
                if faction and faction.nodes:
                    standalone_factions.append(faction)

    total_nodes = sum(count_nodes_recursive(f.nodes) for f in standalone_factions)
    log.info("  → %d standalone Factions, %d total nodes",
             len(standalone_factions), total_nodes)

    return standalone_factions


# =============================================================================
# SCENEOBJECTDATA LOOKUP (AliasName → Position for shop teleport)
# =============================================================================

def build_sceneobjectdata_lookup(folder: Path) -> Dict[str, str]:
    """
    Build lookup: SceneObjectData.AliasName → Position

    Scans normal StaticInfo XMLs for SceneObjectData elements.
    Used to resolve shop stage positions via AliasName cross-reference.

    Returns:
        Dict mapping AliasName (lowercase) to Position string "x,y,z"
    """
    log.info("Building SceneObjectData AliasName→Position lookup...")

    lookup: Dict[str, str] = {}

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for sod in root_el.iter("SceneObjectData"):
            alias = sod.get("AliasName") or ""
            position = sod.get("Position") or ""

            if not alias or not position:
                continue

            # First occurrence wins
            if alias.lower() not in lookup:
                lookup[alias.lower()] = position

    log.info("  → %d SceneObjectData entries with AliasName+Position", len(lookup))
    return lookup


# =============================================================================
# KNOWLEDGE CONTENTS PARSING (3-hop position chain)
# =============================================================================

def _build_uimaptexture_lookup(folder: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Build two lookups from UIMapTextureInfo elements:
      1. KnowledgeKey (lowercase) → UIMapTextureInfo.StrKey
      2. UIMapTextureInfo.StrKey (lowercase) → first SceneObjectData.Position

    Scans uimaptextureinfo files for KnowledgeKey mapping,
    then scans all staticinfo for LevelGimmickSceneObjectInfo to get positions.
    """
    log.info("Building UIMapTextureInfo lookups...")

    # Hop 1: KnowledgeKey → UIMapTextureInfo.StrKey
    knowledge_to_texture: Dict[str, str] = {}

    for path in iter_xml_files(folder):
        if "uimaptextureinfo" not in path.name.lower():
            continue
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for el in root_el.iter("UIMapTextureInfo"):
            kk = el.get("KnowledgeKey") or ""
            strkey = el.get("StrKey") or ""
            if kk and strkey and kk.lower() not in knowledge_to_texture:
                knowledge_to_texture[kk.lower()] = strkey

    log.info("  → UIMapTextureInfo: %d KnowledgeKey→StrKey mappings", len(knowledge_to_texture))

    # Hop 2: UIMapTextureKey → first SceneObjectData.Position
    # Collect all texture StrKeys we need to resolve
    texture_keys_needed: Set[str] = set(v.lower() for v in knowledge_to_texture.values())

    texture_to_position: Dict[str, str] = {}

    for path in iter_xml_files(folder):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for lgsoi in root_el.iter("LevelGimmickSceneObjectInfo"):
            ui_key = lgsoi.get("UIMapTextureKey") or ""
            if not ui_key or ui_key.lower() not in texture_keys_needed:
                continue
            if ui_key.lower() in texture_to_position:
                continue  # Already have position for this key (first win)

            # Get first SceneObjectData with Position
            for sod in lgsoi.iter("SceneObjectData"):
                pos = sod.get("Position") or ""
                if pos:
                    texture_to_position[ui_key.lower()] = pos
                    break

    unresolved = texture_keys_needed - set(texture_to_position.keys())
    log.info("  → LevelGimmickSceneObjectInfo: %d positions resolved", len(texture_to_position))
    if unresolved:
        log.warning("  → %d UIMapTextureKeys could not be resolved to positions", len(unresolved))
    return knowledge_to_texture, texture_to_position


def parse_knowledge_contents(
    static_folder: Path,
    global_seen: Set[str],
) -> List[KnowledgeContentGroup]:
    """
    Parse knowledgeinfo_contents.staticinfo.xml for the Knowledge_Contents tab.

    Resolves teleport positions via 3-hop chain:
      KnowledgeInfo.StrKey → UIMapTextureInfo.KnowledgeKey → StrKey
      → LevelGimmickSceneObjectInfo.UIMapTextureKey → SceneObjectData.Position
    """
    # Find the knowledge contents file
    contents_file = None
    for path in iter_xml_files(static_folder):
        if "knowledgeinfo_contents" in path.name.lower():
            contents_file = path
            break

    if contents_file is None:
        log.info("knowledgeinfo_contents file not found — skipping Knowledge_Contents tab")
        return []

    log.info("Parsing knowledge contents: %s", contents_file.name)

    root_el = parse_xml_file(contents_file)
    if root_el is None:
        log.error("Failed to parse knowledge contents file")
        return []

    # Build position resolution chain
    knowledge_to_texture, texture_to_position = _build_uimaptexture_lookup(static_folder)

    source_file = contents_file.name
    groups: List[KnowledgeContentGroup] = []

    for group_el in root_el.iter("KnowledgeGroupInfo"):
        group_name = group_el.get("GroupName") or ""
        group_strkey = group_el.get("StrKey") or ""

        if not group_name:
            continue

        _collect_korean_string(group_name)

        group = KnowledgeContentGroup(
            strkey=group_strkey,
            group_name=group_name,
            source_file=source_file,
        )

        for ki in group_el.findall("KnowledgeInfo"):
            strkey = ki.get("StrKey") or ""
            name = ki.get("Name") or ""
            desc = ki.get("Desc") or ""

            if not strkey or not name:
                continue

            # Dedup
            dedup_key = f"kc_{strkey}"
            if dedup_key in global_seen:
                continue
            global_seen.add(dedup_key)

            _collect_korean_string(name)
            if desc:
                _collect_korean_string(desc)

            # Resolve position via 3-hop chain
            world_position = ""
            texture_strkey = knowledge_to_texture.get(strkey.lower(), "")
            if texture_strkey:
                pos_csv = texture_to_position.get(texture_strkey.lower(), "")
                if pos_csv:
                    world_position = pos_csv.replace(",", " ")

            group.entries.append(KnowledgeContentEntry(
                strkey=strkey,
                name=name,
                description=desc,
                world_position=world_position,
                source_file=source_file,
            ))

        if group.entries:
            groups.append(group)

    total_entries = sum(len(g.entries) for g in groups)
    log.info("  → %d groups, %d knowledge entries", len(groups), total_entries)
    return groups


# =============================================================================
# SHOP PARSING
# =============================================================================

def parse_shop_file(shop_path: Path, global_seen: Set[str]) -> List[ShopGroup]:
    """Parse shop file for shop stage data."""
    if not shop_path.exists():
        log.warning("Shop file not found: %s", shop_path)
        return []

    log.info("Parsing shop file: %s", shop_path)

    root_el = parse_xml_file(shop_path)
    if root_el is None:
        log.error("Failed to parse shop file")
        return []

    source_file = shop_path.name
    merged_groups: OrderedDict[str, ShopGroup] = OrderedDict()

    for child in root_el:
        if not isinstance(child.tag, str):
            continue

        name = child.get("Name") or ""
        group = child.get("Group") or ""

        if not name:
            continue

        new_stages: List[ShopStage] = []
        for stage_el in child.iter("Stage"):
            strkey = stage_el.get("StrKey") or ""
            stage_name = stage_el.get("Name") or ""
            desc = stage_el.get("Desc") or ""
            category = stage_el.get("Category") or ""
            npc_key = stage_el.get("NPCShopCharacterKey") or ""
            alias_name = stage_el.get("AliasName") or ""

            if not stage_name:
                continue

            if strkey:
                if strkey in global_seen:
                    continue
                global_seen.add(strkey)

            stage = ShopStage(
                strkey=strkey,
                name=stage_name,
                description=desc,
                category=category,
                npc_key=npc_key,
                alias_name=alias_name,
                source_file=source_file,
            )
            new_stages.append(stage)

        if not new_stages:
            continue

        if name in merged_groups:
            merged_groups[name].stages.extend(new_stages)
        else:
            merged_groups[name] = ShopGroup(
                tag_name=child.tag,
                name=name,
                group=group,
                source_file=source_file,
                stages=new_stages,
            )

    shop_groups = list(merged_groups.values())
    log.info("  → %d groups, %d stages",
             len(shop_groups), sum(len(g.stages) for g in shop_groups))

    return shop_groups


# =============================================================================
# ROW GENERATION
# =============================================================================

# Row format: (depth, text, style_type, is_description, source_file, data_type, world_position)
RowItem = Tuple[int, str, str, bool, str, str, str]


def emit_shop_rows(
    shop_groups: List[ShopGroup],
    sceneobject_lookup: Optional[Dict[str, str]] = None,
) -> List[RowItem]:
    """Generate rows for Shop data with position from SceneObjectData lookup."""
    rows: List[RowItem] = []
    sod = sceneobject_lookup or {}

    for group in shop_groups:
        rows.append((0, group.name, "ShopGroup", False, group.source_file, "ShopGroup", ""))
        for stage in group.stages:
            # Resolve position via AliasName → SceneObjectData.Position
            world_position = ""
            if stage.alias_name:
                pos_csv = sod.get(stage.alias_name.lower(), "")
                if pos_csv:
                    # Convert "x,y,z" to "x y z" for /teleport command
                    world_position = pos_csv.replace(",", " ")
            rows.append((1, stage.name, "ShopStage", False, stage.source_file, "ShopStage", world_position))
            if stage.description:
                rows.append((2, stage.description, "Description", True, stage.source_file, "ShopStage.Desc", world_position))

    return rows


def emit_knowledge_content_rows(groups: List[KnowledgeContentGroup]) -> List[RowItem]:
    """Generate rows for the Knowledge_Contents tab."""
    rows: List[RowItem] = []

    for group in groups:
        # Root: KnowledgeGroupInfo.GroupName (depth 0)
        rows.append((0, group.group_name, "Faction", False, group.source_file, "KnowledgeGroupInfo", ""))

        for entry in group.entries:
            # Name row (depth 1) with teleport
            rows.append((1, entry.name, "FactionNode_Main", False, entry.source_file, "KnowledgeInfo", entry.world_position))

            # Desc row (depth 2)
            if entry.description:
                rows.append((2, entry.description, "Description", True, entry.source_file, "KnowledgeInfo.Desc", entry.world_position))

    return rows


# =============================================================================
# DISPLAYNAME LOOKUP (RegionInfo.KnowledgeKey → DisplayName)
# =============================================================================

def build_displayname_lookup(folder: Path) -> Dict[str, str]:
    """
    Build lookup: RegionInfo.KnowledgeKey → RegionInfo.DisplayName

    Scans all XML files for RegionInfo elements (recursive hierarchy).
    Used to add DisplayName rows in Region output.
    """
    log.info("Building RegionInfo DisplayName lookup...")

    lookup: Dict[str, str] = {}
    count = 0

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for ri in root_el.iter("RegionInfo"):
            kk = ri.get("KnowledgeKey") or ""
            display_name = ri.get("DisplayName") or ""

            if not kk or not display_name:
                continue

            # First occurrence wins (same as knowledge lookup)
            if kk.lower() not in lookup:
                lookup[kk.lower()] = display_name
                count += 1

    log.info("  → %d entries", count)
    return lookup


# =============================================================================
# ROW GENERATION (Region rows + DisplayName)
# =============================================================================

def emit_faction_node_rows(
    node: FactionNodeData,
    depth: int,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for a FactionNode and its children. Adds DisplayName row."""
    rows: List[RowItem] = []

    # 1. Name (from KnowledgeInfo.Name — same as Region)
    style = f"FactionNode_{node.node_type}" if node.node_type else "FactionNode"
    rows.append((depth, node.name, style, False, node.source_file, "KnowledgeInfo", node.world_position))

    # 2. DisplayName (from RegionInfo — NEW)
    displayname = displayname_lookup.get(node.knowledge_key.lower(), "")
    if displayname and displayname != node.name:
        rows.append((depth, displayname, style, False, node.source_file, "RegionInfo.DisplayName", node.world_position))
        _collect_korean_string(displayname)

    # 3. Description (from KnowledgeInfo.Desc — same as Region)
    if node.description:
        rows.append((depth + 1, node.description, "Description", True, node.source_file, "KnowledgeInfo.Desc", node.world_position))

    # Children
    for child in node.children:
        rows.extend(emit_faction_node_rows(child, depth + 1, displayname_lookup))

    return rows


def emit_faction_rows(
    faction: FactionData,
    depth: int,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for a Faction and all its FactionNodes."""
    rows: List[RowItem] = []

    rows.append((depth, faction.name, "Faction", False, faction.source_file, "Faction", ""))

    for node in faction.nodes:
        rows.extend(emit_faction_node_rows(node, depth + 1, displayname_lookup))

    return rows


def emit_faction_group_rows(
    fg: FactionGroupData,
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate all rows for a FactionGroup."""
    rows: List[RowItem] = []

    for faction in fg.factions:
        rows.extend(emit_faction_rows(faction, 0, displayname_lookup))

    return rows


def emit_standalone_faction_rows(
    standalone_factions: List[FactionData],
    displayname_lookup: Dict[str, str],
) -> List[RowItem]:
    """Generate rows for standalone Factions."""
    rows: List[RowItem] = []

    for faction in standalone_factions:
        rows.append((0, faction.name, "Faction", False, faction.source_file, "Faction", ""))
        for node in faction.nodes:
            rows.extend(emit_faction_node_rows(node, 1, displayname_lookup))

    return rows


# =============================================================================
# EXCEL HELPERS
# =============================================================================

def get_translated_tab_name(
    korean_name: str,
    eng_tbl: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Optional[Dict[str, List[Tuple[str, str]]]],
    lang_code: str,
    fallback: str = "Sheet"
) -> str:
    """Get translated sheet tab name."""
    is_eng = lang_code.lower() == "eng"

    if is_eng:
        trans, _ = get_first_translation(eng_tbl, korean_name)
        title = trans.strip() if trans else korean_name.strip()
    else:
        if lang_tbl:
            trans, _ = get_first_translation(lang_tbl, korean_name)
            if trans and is_good_translation(trans):
                title = trans.strip()
            else:
                trans_eng, _ = get_first_translation(eng_tbl, korean_name)
                title = trans_eng.strip() if trans_eng else korean_name.strip()
        else:
            trans_eng, _ = get_first_translation(eng_tbl, korean_name)
            title = trans_eng.strip() if trans_eng else korean_name.strip()

    if not title:
        title = fallback

    # Sanitize for Excel
    title = title[:31]
    title = re.sub(r"[\\/*?:\[\]]", "_", title)

    return title


def write_sheet_content(
    sheet,
    rows: List[RowItem],
    is_eng: bool,
    eng_tbl: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Optional[Dict[str, List[Tuple[str, str]]]],
    lang_code: str,
    export_index: Optional[Dict[str, Set[str]]] = None,
    consumer: Optional[StringIdConsumer] = None,
    eng_consumer: Optional[StringIdConsumer] = None,
) -> None:
    """Write rows to a sheet with proper formatting.

    Args:
        consumer: StringID consumer for target language (non-ENG workbooks).
        eng_consumer: StringID consumer for ENG workbook StringID column.
                      For non-ENG workbooks, eng is display-only so pass None.
    """

    # Headers (DataType is column 1, everything else shifts right by 1)
    headers = []
    headers.append(sheet.cell(1, 1, "DataType"))
    headers.append(sheet.cell(1, 2, "Original (KR)"))
    headers.append(sheet.cell(1, 3, "English (ENG)"))
    if not is_eng:
        headers.append(sheet.cell(1, 4, f"Translation ({lang_code.upper()})"))

    extra_headers = ["WorldPosition", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    start_col = len(headers) + 1
    for idx, name in enumerate(extra_headers, start=start_col):
        headers.append(sheet.cell(1, idx, name))

    for c in headers:
        c.font = _header_font
        c.fill = _header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    sheet.row_dimensions[1].height = 25

    # Data rows (raw data, no deduplication)
    r_idx = 2

    for (depth, text, style_type, is_desc, source_file, data_type, world_position) in rows:
        fill, font, row_height = get_style(style_type)

        # For ENG workbook: eng_consumer disambiguates StringIDs
        # For non-ENG: consumer=None (display-only), StringID comes from lang_tbl
        if source_file and export_index:
            trans_eng, sid_eng = resolve_translation(text, eng_tbl, source_file, export_index, consumer=eng_consumer)
        else:
            trans_eng, sid_eng = get_first_translation(eng_tbl, text)
        trans_other = sid_other = ""
        if not is_eng and lang_tbl:
            if source_file and export_index:
                trans_other, sid_other = resolve_translation(text, lang_tbl, source_file, export_index, consumer=consumer)
            else:
                trans_other, sid_other = get_first_translation(lang_tbl, text)

        # DataType
        c_dtype = sheet.cell(r_idx, 1, data_type)
        c_dtype.fill = fill
        c_dtype.font = font
        c_dtype.alignment = Alignment(horizontal="center", vertical="center")
        c_dtype.border = THIN_BORDER

        # Original
        c_orig = sheet.cell(r_idx, 2, br_to_newline(text))
        c_orig.fill = fill
        c_orig.font = font
        c_orig.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
        c_orig.border = THIN_BORDER

        # English
        c_eng = sheet.cell(r_idx, 3, br_to_newline(trans_eng))
        c_eng.fill = fill
        c_eng.font = font
        c_eng.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
        c_eng.border = THIN_BORDER

        # Target language
        col_offset = 3
        if not is_eng:
            c_other = sheet.cell(r_idx, 4, br_to_newline(trans_other))
            c_other.fill = fill
            c_other.font = font
            c_other.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
            c_other.border = THIN_BORDER
            col_offset = 4

        # Extra columns — format as /teleport command for easy copy-paste
        teleport_cmd = f"/teleport {world_position}" if world_position else ""
        c_worldpos = sheet.cell(r_idx, col_offset + 1, teleport_cmd)
        c_worldpos.fill = fill
        c_worldpos.font = font
        c_worldpos.alignment = Alignment(horizontal="center", vertical="center")
        c_worldpos.border = THIN_BORDER

        c_status = sheet.cell(r_idx, col_offset + 2, "")
        c_status.fill = fill
        c_status.alignment = Alignment(horizontal="center", vertical="center")
        c_status.border = THIN_BORDER

        c_comment = sheet.cell(r_idx, col_offset + 3, "")
        c_comment.fill = fill
        c_comment.border = THIN_BORDER

        c_stringid = sheet.cell(r_idx, col_offset + 4, sid_other if not is_eng else sid_eng)
        c_stringid.fill = fill
        c_stringid.font = Font(bold=True)
        c_stringid.border = THIN_BORDER
        c_stringid.number_format = '@'

        c_screenshot = sheet.cell(r_idx, col_offset + 5, "")
        c_screenshot.fill = fill
        c_screenshot.border = THIN_BORDER

        # Apply row height if specified
        if row_height:
            sheet.row_dimensions[r_idx].height = row_height

        r_idx += 1

    last_row = r_idx - 1

    # Column widths
    widths = [18, 40, 80] + ([] if is_eng else [80]) + [30, 15, 70, 25, 25]
    for idx, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w

    # Hide English column for non-English workbooks
    if not is_eng:
        sheet.column_dimensions["C"].hidden = True

    # Status dropdown
    status_col = 5 if is_eng else 6
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    col_letter = get_column_letter(status_col)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")
    sheet.add_data_validation(dv)

    # Auto-fit
    autofit_worksheet(sheet)


# =============================================================================
# EXCEL WRITER (workbook)
# =============================================================================

def write_workbook(
    faction_groups: List[FactionGroupData],
    standalone_factions: List[FactionData],
    shop_groups: List[ShopGroup],
    eng_tbl: Dict[str, List[Tuple[str, str]]],
    lang_tbl: Optional[Dict[str, List[Tuple[str, str]]]],
    lang_code: str,
    out_path: Path,
    export_index: Optional[Dict[str, Set[str]]] = None,
    displayname_lookup: Optional[Dict[str, str]] = None,
    sceneobject_lookup: Optional[Dict[str, str]] = None,
    knowledge_content_groups: Optional[List[KnowledgeContentGroup]] = None,
) -> bool:
    """Generate Excel workbook for a language. Returns True if saved."""
    wb = Workbook()
    wb.remove(wb.active)

    is_eng = lang_code.lower() == "eng"
    used_titles: Set[str] = set()
    dn_lookup = displayname_lookup or {}

    # Order-based StringID consumer (fresh per language write pass)
    ordered_idx = get_ordered_export_index()
    consumer = StringIdConsumer(ordered_idx)
    eng_consumer = StringIdConsumer(ordered_idx) if is_eng else None

    def get_unique_title(base: str) -> str:
        title = base
        counter = 1
        while title in used_titles:
            title = f"{base[:28]}_{counter}"
            counter += 1
        used_titles.add(title)
        return title

    # FactionGroup sheets
    for fg in faction_groups:
        rows = emit_faction_group_rows(fg, dn_lookup)
        if not rows:
            continue

        title = get_translated_tab_name(fg.group_name, eng_tbl, lang_tbl, lang_code, "Group")
        title = get_unique_title(title)

        sheet = wb.create_sheet(title=title)
        write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)

        log.info("    Sheet '%s': %d rows", title, len(rows))

    # Standalone Faction sheet
    if standalone_factions:
        rows = emit_standalone_faction_rows(standalone_factions, dn_lookup)
        if rows:
            title = get_translated_tab_name(EMPTY_FACTION_NAME, eng_tbl, lang_tbl, lang_code, "No Faction")
            title = get_unique_title(title)

            sheet = wb.create_sheet(title=title)
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)

            log.info("    Sheet '%s': %d rows", title, len(rows))

    # Shop sheet (no DisplayName needed for shop)
    if shop_groups:
        rows = emit_shop_rows(shop_groups, sceneobject_lookup)
        if rows:
            sheet = wb.create_sheet(title="Shop")
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)
            log.info("    Sheet 'Shop': %d rows", len(rows))

    # Knowledge_Contents sheet (3-hop position resolution)
    if knowledge_content_groups:
        rows = emit_knowledge_content_rows(knowledge_content_groups)
        if rows:
            sheet = wb.create_sheet(title="Knowledge_Contents")
            write_sheet_content(sheet, rows, is_eng, eng_tbl, lang_tbl, lang_code, export_index, consumer, eng_consumer)
            log.info("    Sheet 'Knowledge_Contents': %d rows", len(rows))

    # Save
    if wb.worksheets:
        wb.save(out_path)
        log.info("  → Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))
        return True

    return False


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_region_datasheets() -> Dict:
    """
    Generate Region datasheets for all languages.
    Includes DisplayName rows from RegionInfo.
    """
    result = {
        "category": "Region",
        "files_created": 0,
        "errors": [],
    }

    reset_korean_collection()

    log.info("=" * 70)
    log.info("Region Datasheet Generator")
    log.info("=" * 70)

    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "RegionData_Map_All"
    output_folder.mkdir(exist_ok=True)

    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        global_seen: Set[str] = set()

        # 1. Knowledge lookup (Name + Desc)
        knowledge_lookup = build_knowledge_name_lookup(RESOURCE_FOLDER)

        # 2. DisplayName lookup (RegionInfo.KnowledgeKey → DisplayName)
        displayname_lookup = build_displayname_lookup(RESOURCE_FOLDER)

        # 3. Parse FactionGroup hierarchy
        faction_groups = parse_all_faction_groups(RESOURCE_FOLDER, knowledge_lookup, global_seen)

        # 4. Parse Standalone Factions
        standalone_factions = parse_standalone_factions(RESOURCE_FOLDER, knowledge_lookup, global_seen)

        # 5. Parse Shop data
        shop_file = RESOURCE_FOLDER.parent / "staticinfo_quest" / "funcnpc" / "shop_world.staticinfo.xml"
        shop_groups = parse_shop_file(shop_file, global_seen)

        # 5.5. Build SceneObjectData lookup (AliasName → Position for shop teleport)
        sceneobject_lookup = build_sceneobjectdata_lookup(RESOURCE_FOLDER)

        # 5.6. Parse Knowledge Contents (3-hop position chain)
        knowledge_content_groups = parse_knowledge_contents(RESOURCE_FOLDER, global_seen)

        # 6. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        eng_tbl = lang_tables.get("eng", {})
        export_index = get_export_index()

        # 7. Generate workbooks
        log.info("Generating Excel workbooks...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Generating: %s", idx, total, code.upper())
            out_path = output_folder / f"Region_LQA_{code.upper()}.xlsx"

            if code.lower() == "eng":
                saved = write_workbook(
                    faction_groups, standalone_factions, shop_groups,
                    eng_tbl, None, code, out_path, export_index, displayname_lookup,
                    sceneobject_lookup, knowledge_content_groups,
                )
            else:
                saved = write_workbook(
                    faction_groups, standalone_factions, shop_groups,
                    eng_tbl, tbl, code, out_path, export_index, displayname_lookup,
                    sceneobject_lookup, knowledge_content_groups,
                )
            if saved:
                result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Region generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    generate_region_datasheets()
