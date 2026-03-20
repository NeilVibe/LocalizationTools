"""
Generators Base Module
======================
Shared functionality for all datasheet generators.

Contains:
- XML parsing and sanitization
- Language table loading (with duplicate handling)
- Korean detection
- Placeholder normalization
- Excel autofit
- Common styles
- EXPORT-based duplicate translation resolution
"""

import os
import re
import logging
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Tuple, Optional, Iterator, Set
from dataclasses import dataclass

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config as _cfg
from config import DEPTH_COLORS, STATUS_OPTIONS, EXPORT_FOLDER

# =============================================================================
# LOGGING
# =============================================================================

def get_logger(name: str) -> logging.Logger:
    """Create a logger with console output."""
    log = logging.getLogger(name)
    log.setLevel(logging.DEBUG)

    if not log.handlers:
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
        handler.setLevel(logging.INFO)
        log.addHandler(handler)

    log.propagate = False
    return log


log = get_logger("base")

# =============================================================================
# TEXT NORMALIZATION
# =============================================================================

_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_br_tag_re = re.compile(r'<br\s*/?>', flags=re.IGNORECASE)
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)
_staticinfo_code_re = re.compile(r'\{StaticInfo:[^}]+\}')


def normalize_placeholders(text: str) -> str:
    """
    Normalize text for matching:
    1) Remove '#...' suffix inside {...} placeholders
    2) Normalize <br/> tags to space (data files may use \\n, language data uses <br/>)
    3) Collapse all whitespace to single space
    4) Trim leading/trailing spaces
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _br_tag_re.sub(' ', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text


# =============================================================================
# STATICINFO CODE RESOLUTION ({StaticInfo:Category:StrKey} → Korean)
# =============================================================================

_STRKEY_LOOKUP: Optional[Dict[str, str]] = None  # StrKey.lower() → DevMemo/DevComment


def _build_strkey_lookup(resource_folder: Path) -> Dict[str, str]:
    """Scan StaticInfo XMLs for StrKey → DevMemo/DevComment mapping.

    Uses parse_xml_file() for proper XML sanitization and virtual ROOT wrapper
    (StaticInfo XMLs don't have a single root element).

    Returns:
        {strkey_lower: korean_text}
    """
    lookup: Dict[str, str] = {}
    if not resource_folder or not resource_folder.exists():
        log.warning("StrKey lookup: RESOURCE_FOLDER is None or does not exist: %s", resource_folder)
        return lookup

    file_count = 0
    for xml_path in resource_folder.rglob("*.xml"):
        root = parse_xml_file(xml_path)
        if root is None:
            continue
        file_count += 1
        for el in root.iter():
            strkey = el.get("StrKey")
            if not strkey:
                continue
            # Try DevMemo first, then DevComment
            korean = el.get("DevMemo") or el.get("DevComment") or ""
            if korean and strkey.lower() not in lookup:
                lookup[strkey.lower()] = korean

    log.info("StrKey lookup: scanned %d XML files, found %d StrKey entries", file_count, len(lookup))
    return lookup


def get_strkey_lookup() -> Dict[str, str]:
    """Lazy-load and cache the StrKey → Korean lookup."""
    global _STRKEY_LOOKUP
    if _STRKEY_LOOKUP is None:
        log.info("Building StrKey → DevMemo/DevComment lookup...")
        _STRKEY_LOOKUP = _build_strkey_lookup(_cfg.RESOURCE_FOLDER)
        log.info("Indexed %d StrKey entries", len(_STRKEY_LOOKUP))
    return _STRKEY_LOOKUP


def resolve_staticinfo_codes(text: str) -> str:
    """Resolve {StaticInfo:Category:StrKey} codes to Korean text.

    Example: '{StaticInfo:Status:Fishing}-힘겨루기' → '낚시-힘겨루기'

    Extracts the last segment after ':' as StrKey, looks up DevMemo/DevComment,
    and replaces the full code. Unknown codes are left as-is.
    """
    if '{StaticInfo:' not in text:
        return text

    lookup = get_strkey_lookup()

    def _replace_code(m: re.Match) -> str:
        code = m.group(0)  # e.g. {StaticInfo:Status:Fishing}
        # Extract last segment after last ':'
        parts = code.strip('{}').split(':')
        if len(parts) >= 3:
            strkey = parts[-1]
            korean = lookup.get(strkey.lower())
            if korean:
                log.debug("StaticInfo code resolved: %s → %s (via StrKey=%s)", code, korean, strkey)
                return korean
            else:
                log.warning("StaticInfo code NOT resolved: %s (StrKey=%s not in %d-entry lookup)", code, strkey, len(lookup))
        return code  # Unknown code — leave as-is

    result = _staticinfo_code_re.sub(_replace_code, text)
    if result != text:
        log.info("StaticInfo resolution: '%s' → '%s'", text[:80], result[:80])
    return result


def br_to_newline(text: str) -> str:
    """Convert <br/> tags to real newlines for Excel output.

    XML attributes store linebreaks as <br/> (after lxml parse).
    Excel cells need real \\n (displays as Alt+Enter).
    """
    if not text:
        return text
    return _br_tag_re.sub('\n', text)


# =============================================================================
# KOREAN DETECTION
# =============================================================================

_korean_re = re.compile(r'[\uAC00-\uD7AF\u1100-\u11FF\u3130-\u318F]')


def contains_korean(text: str) -> bool:
    """Check if text contains Korean characters."""
    return bool(_korean_re.search(text)) if text else False


def is_good_translation(text: str) -> bool:
    """Check if translation is valid (non-empty and no Korean)."""
    return bool(text) and not contains_korean(text)


# =============================================================================
# EXPORT STRINGID INDEX (for duplicate translation resolution)
# =============================================================================

_EXPORT_INDEX: Optional[Dict[str, Set[str]]] = None  # Module-level cache
_ORDERED_EXPORT_INDEX: Optional[Dict[str, Dict[str, List[str]]]] = None


def build_export_indexes(
    export_folder: Path,
) -> Tuple[Dict[str, Set[str]], Dict[str, Dict[str, List[str]]]]:
    """
    Build both flat and ordered export indexes in a single parse pass.

    The flat index maps: normalized_filename → {stringid_1, stringid_2, ...}
    The ordered index maps: normalized_filename → {normalized_kor_text → [sid1, sid2, ...]}

    The ordered list preserves XML document order, so the Nth occurrence of a
    Korean text in a file maps to the Nth StringID in the list.

    Args:
        export_folder: Path to EXPORT folder containing .loc.xml files

    Returns:
        (flat_index, ordered_index)
    """
    flat: Dict[str, Set[str]] = {}
    ordered: Dict[str, Dict[str, List[str]]] = {}

    if not export_folder.exists():
        return flat, ordered

    for path in export_folder.rglob("*.loc.xml"):
        if not path.is_file():
            continue

        filename_key = path.name.lower().replace(".loc.xml", "")

        try:
            tree = ET.parse(str(path))
            root = tree.getroot()

            sids: Set[str] = set()
            kor_map: Dict[str, List[str]] = {}

            for loc_str in root.iter("LocStr"):
                sid = loc_str.get("StringId") or ""
                origin = loc_str.get("StrOrigin") or ""
                if sid:
                    sids.add(sid)
                if origin and sid:
                    norm = normalize_placeholders(origin)
                    if norm:
                        kor_map.setdefault(norm, []).append(sid)

            if sids:
                if filename_key in flat:
                    flat[filename_key].update(sids)
                else:
                    flat[filename_key] = sids

            if kor_map:
                if filename_key in ordered:
                    # Merge: append new SIDs to existing lists
                    existing = ordered[filename_key]
                    for norm_text, sid_list in kor_map.items():
                        existing.setdefault(norm_text, []).extend(sid_list)
                else:
                    ordered[filename_key] = kor_map

        except ET.XMLSyntaxError:
            continue
        except Exception:
            continue

    return flat, ordered


def _ensure_export_indexes() -> None:
    """Lazy-build both export indexes on first access."""
    global _EXPORT_INDEX, _ORDERED_EXPORT_INDEX
    if _EXPORT_INDEX is None or _ORDERED_EXPORT_INDEX is None:
        log.info("Building EXPORT StringID indexes (flat + ordered)...")
        _EXPORT_INDEX, _ORDERED_EXPORT_INDEX = build_export_indexes(EXPORT_FOLDER)
        log.info("Indexed %d EXPORT files (flat), %d files (ordered)",
                 len(_EXPORT_INDEX), len(_ORDERED_EXPORT_INDEX))


def get_export_index() -> Dict[str, Set[str]]:
    """
    Lazy-load and cache flat EXPORT index.

    Returns:
        Dict mapping normalized filename to set of StringIDs
    """
    _ensure_export_indexes()
    return _EXPORT_INDEX  # type: ignore[return-value]


def get_ordered_export_index() -> Dict[str, Dict[str, List[str]]]:
    """
    Lazy-load and cache ordered EXPORT index.

    Returns:
        Dict mapping normalized filename → {normalized_kor → [sid1, sid2, ...]}
        Lists preserve XML document order.
    """
    _ensure_export_indexes()
    return _ORDERED_EXPORT_INDEX  # type: ignore[return-value]


# =============================================================================
# STRINGID CONSUMER (order-based disambiguation)
# =============================================================================

class StringIdConsumer:
    """Consume StringIDs from ordered export index in document order.

    For each (export_file, kor_text) pair, maintains a pointer that advances
    each time a StringID is consumed. Nth call for the same text in the same
    file returns the Nth StringID from the ordered list.

    Usage:
        consumer = StringIdConsumer(get_ordered_export_index())
        # ... for each row in document order:
        sid = consumer.consume(normalized_kor, export_key)
    """

    def __init__(self, ordered_index: Dict[str, Dict[str, List[str]]]):
        self._index = ordered_index
        self._consumed: Dict[Tuple[str, str], int] = defaultdict(int)
        self.warnings: int = 0

    def consume(self, normalized_kor: str, export_key: str) -> Optional[str]:
        """Get next StringID for this text in this file.

        Returns None if no ordered data available for this text/file combination.
        """
        file_map = self._index.get(export_key)
        if not file_map:
            return None
        sid_list = file_map.get(normalized_kor)
        if not sid_list:
            return None

        key = (export_key, normalized_kor)
        idx = self._consumed[key]
        self._consumed[key] += 1

        if idx < len(sid_list):
            return sid_list[idx]
        else:
            # More occurrences in data than export — warn, return last available
            self.warnings += 1
            log.warning(
                "StringID overrun: '%s...' in %s — occurrence %d but only %d available",
                normalized_kor[:30], export_key, idx + 1, len(sid_list),
            )
            return sid_list[-1]


def get_export_key(data_filename: str) -> str:
    """
    Convert data filename to export lookup key.

    Args:
        data_filename: e.g., "skillinfo_pc.staticinfo.xml"

    Returns:
        Normalized key: e.g., "skillinfo_pc.staticinfo"
    """
    return data_filename.lower().replace(".xml", "").replace(".loc", "")


# =============================================================================
# CASCADING NORMALIZATION
# =============================================================================
# Each level is a function that normalizes text differently.
# The cascade tries each level until candidates are found in the language table.
# Extensible — add new functions to NORMALIZATION_CASCADE to handle new patterns.

NORMALIZATION_CASCADE = [
    # Level 0+1: normalize_placeholders (strip placeholder suffixes, br→space, collapse ws)
    lambda text: normalize_placeholders(text),
    # Level 2: resolve {StaticInfo:...} codes first, then normalize
    lambda text: normalize_placeholders(resolve_staticinfo_codes(text)),
]


def _find_candidates_cascading(
    korean_text: str,
    lang_table: Dict[str, List[Tuple[str, str, str]]],
) -> Tuple[Optional[List[Tuple[str, str, str]]], str]:
    """Try each normalization level until candidates are found.

    Returns:
        (candidates, normalized_key) or (None, "") if no match at any level.
    """
    for normalize_fn in NORMALIZATION_CASCADE:
        normalized = normalize_fn(korean_text)
        if not normalized:
            continue
        candidates = lang_table.get(normalized)
        if candidates:
            return candidates, normalized
    return None, ""


def resolve_translation(
    korean_text: str,
    lang_table: Dict[str, List[Tuple[str, str, str]]],
    data_filename: str = "",
    export_index: Optional[Dict[str, Set[str]]] = None,
    consumer: Optional[StringIdConsumer] = None,
) -> Tuple[str, str, str]:
    """
    Resolve correct translation using cascading normalization + EXPORT disambiguation.

    Tries each normalization level sequentially until candidates are found in the
    language table. Then disambiguates using consumer (order-based) or export index
    (file-set-based) to find the correct match for the current data file.

    Algorithm:
    1. Cascading normalization — try each level until candidates found
    2. If only one candidate → use it
    3. If consumer provided → try order-based consumption (Nth occurrence → Nth StringID)
    4. If multiple → find StringID that exists in matching EXPORT file
    5. Fallback → first good translation (no Korean)

    Args:
        korean_text: The Korean source text to translate
        lang_table: Language table with ALL translations stored as lists
                    {normalized_korean: [(translation, stringid, str_origin), ...]}
        data_filename: Source data file name (e.g., "skillinfo_pc.staticinfo.xml")
        export_index: Optional pre-loaded EXPORT index (uses global cache if None)
        consumer: Optional StringIdConsumer for order-based disambiguation.
                  When provided, the Nth call for the same text in the same file
                  returns the Nth StringID from document order.

    Returns:
        Tuple of (translation, stringid, str_origin). Returns ("", "", "") if not found.
    """
    if not korean_text:
        return ("", "", "")

    # Cascading normalization — try each level until candidates found
    candidates, normalized = _find_candidates_cascading(korean_text, lang_table)
    if not candidates:
        return ("", "", "")

    # If only one candidate, no disambiguation needed — just return it
    if len(candidates) == 1:
        return candidates[0]

    # Multiple candidates — try order-based consumption first (most precise)
    # NOTE: consumer always uses level-0/1 normalized key (normalize_placeholders)
    # because the ordered export index was built with that normalization.
    # If cascade matched at a higher level, consumer lookup may return None — that's OK,
    # we fall through to export-index file-set matching.
    if consumer and data_filename:
        export_key = get_export_key(data_filename)
        consumer_key = normalize_placeholders(korean_text)
        target_sid = consumer.consume(consumer_key, export_key)
        if target_sid:
            # Try exact match first (SID + translation from lang_table)
            for translation, stringid, str_origin in candidates:
                if stringid == target_sid:
                    return (translation, stringid, str_origin)
            # SID not in lang_table candidates — use SID from export directly,
            # pair with best available translation
            for translation, stringid, str_origin in candidates:
                if is_good_translation(translation):
                    return (translation, target_sid, str_origin)
            return (candidates[0][0], target_sid, candidates[0][2])

    # Fall back to file-set matching using EXPORT index
    if data_filename:
        if export_index is None:
            export_index = get_export_index()

        export_key = get_export_key(data_filename)
        export_stringids = export_index.get(export_key, set())

        if export_stringids:
            # Find candidate whose StringID is in this EXPORT file
            for translation, stringid, str_origin in candidates:
                if stringid in export_stringids:
                    return (translation, stringid, str_origin)

    # Fallback: prefer good translation (no Korean)
    for translation, stringid, str_origin in candidates:
        if is_good_translation(translation):
            return (translation, stringid, str_origin)

    # Last resort: return first candidate
    return candidates[0]


# =============================================================================
# XML SANITIZATION (EXACT COPY FROM MONOLITH)
# =============================================================================

_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def _fix_bad_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)


def _preprocess_newlines(raw: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = _fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"',
                 lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                 lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', raw)
    # Tag stack repair for malformed XML
    tag_open  = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    stack: List[str] = []
    out:   List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            # Only push OPEN tags, not self-closing (<Tag .../>) ones.
            # Self-closing tags end with "/>", they don't need a closing tag.
            if not stripped.rstrip().endswith("/>"):
                stack.append(mo.group(1))
            out.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if stack and stack[-1] == mc.group(1):
                stack.pop(); out.append(line)
            else:
                out.append(stack and f"</{stack.pop()}>" or line)
            continue
        if stripped.startswith("</>"):
            out.append(stack and line.replace("</>", f"</{stack.pop()}>") or line)
            continue
        out.append(line)
    while stack:
        out.append(f"</{stack.pop()}>")
    return "\n".join(out)


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    try:
        return ET.fromstring(
            wrapped.encode("utf-8"),
            parser=ET.XMLParser(huge_tree=True)
        )
    except ET.XMLSyntaxError:
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True)
            )
        except ET.XMLSyntaxError:
            return None


def iter_xml_files(
    root: Path,
    suffixes: Tuple[str, ...] = (".xml", ".seqc")
) -> Iterator[Path]:
    """Recursively iterate over XML files in a folder."""
    if not root.exists():
        return
    for dp, _, files in os.walk(root):
        for fn in files:
            low = fn.lower()
            if any(low.endswith(suf) for suf in suffixes):
                yield Path(dp) / fn


# =============================================================================
# LANGUAGE TABLE LOADING
# =============================================================================

def load_language_tables(folder: Path) -> Dict[str, Dict[str, List[Tuple[str, str, str]]]]:
    """
    Load all non-Korean language tables with normalized placeholder keys.

    Stores ALL translations for each Korean text to support duplicate resolution.
    When the same Korean text appears in multiple files with different context-specific
    translations, we preserve all of them for later disambiguation using EXPORT mapping.

    Args:
        folder: Path to language data folder

    Returns:
        {lang_code: {normalized_korean: [(translation, stringid), ...]}}

    Note: The list preserves all translations for a given Korean text.
          Good translations (no Korean) are sorted to the front.
    """
    tables: Dict[str, Dict[str, List[Tuple[str, str, str]]]] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, List[Tuple[str, str, str]]] = {}

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            normalized_origin = normalize_placeholders(origin)

            # Store ALL translations for this Korean text
            if normalized_origin not in tbl:
                tbl[normalized_origin] = []

            # Add this translation (avoid exact duplicates)
            # origin = raw StrOrigin from language data (preserves <br/> tags)
            entry = (tr, sid, origin)
            if entry not in tbl[normalized_origin]:
                tbl[normalized_origin].append(entry)

        # Sort each list: good translations first, then bad ones
        for key in tbl:
            tbl[key].sort(key=lambda x: (0 if is_good_translation(x[0]) else 1))

        tables[lang] = tbl

    return tables


def get_first_translation(
    lang_table: Dict[str, List[Tuple[str, str, str]]],
    korean_text: str
) -> Tuple[str, str, str]:
    """
    Simple lookup that returns the first (best) translation.

    This provides backward compatibility for code that doesn't need
    context-aware duplicate resolution.

    Args:
        lang_table: Language table {normalized_korean: [(translation, stringid, str_origin), ...]}
        korean_text: Korean text to translate

    Returns:
        (translation, stringid, str_origin) or ("", "", "") if not found
    """
    candidates, _ = _find_candidates_cascading(korean_text, lang_table)
    return candidates[0] if candidates else ("", "", "")


# =============================================================================
# EXCEL STYLES
# =============================================================================

# Border style
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Header style
HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Depth-based styles
def get_depth_fill(depth: int) -> PatternFill:
    """Get fill color for a given depth level."""
    colors = DEPTH_COLORS
    color = colors[min(depth, len(colors) - 1)]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_depth_font(depth: int) -> Font:
    """Get font for a given depth level."""
    return Font(bold=(depth == 0))


# =============================================================================
# EXCEL AUTOFIT
# =============================================================================

def autofit_worksheet(
    ws,
    min_width: int = 10,
    max_width: int = 80,
    row_height_per_line: float = 15.0,
    min_row_height: float = 20.0
) -> None:
    """
    Auto-fit column widths and row heights based on cell content.

    Args:
        ws: Worksheet to autofit
        min_width: Minimum column width
        max_width: Maximum column width
        row_height_per_line: Height per line of wrapped text
        min_row_height: Minimum row height
    """
    # Calculate optimal column widths
    col_widths: Dict[str, float] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                content_len = len(str(cell.value))
                width = min(max(content_len * 1.1 + 2, min_width), max_width)
                col_widths[col_letter] = max(col_widths.get(col_letter, min_width), width)

    # Apply column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Calculate optimal row heights (considering text wrap)
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        max_lines = 1
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                col_width = col_widths.get(col_letter, min_width)
                text = str(cell.value)
                # Estimate lines needed
                chars_per_line = max(int(col_width / 1.1), 1)
                lines = max(1, len(text) // chars_per_line + text.count('\n') + 1)
                max_lines = max(max_lines, lines)

        height = max(min_row_height, max_lines * row_height_per_line)
        ws.row_dimensions[row_idx].height = height


# =============================================================================
# EXCEL HELPERS
# =============================================================================

def add_status_dropdown(ws, col: int, start_row: int = 2, end_row: int = 10000) -> None:
    """Add STATUS dropdown validation to a column."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True
    )
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)

    for row in range(start_row, min(end_row, ws.max_row + 1)):
        dv.add(ws.cell(row=row, column=col))


def create_header_row(ws, headers: List[str], row: int = 1) -> None:
    """Create a styled header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


# =============================================================================
# ROW DATA STRUCTURE
# =============================================================================

@dataclass
class RowItem:
    """Represents a row of data for the Excel output."""
    depth: int           # Indentation level (0 = parent)
    text: str            # Korean original text
    needs_translation: bool = True  # Whether this row needs translation
    strkey: str = ""     # String key for reference


def emit_rows_to_worksheet(
    ws,
    rows: List[RowItem],
    lang_table: Dict[str, List[Tuple[str, str, str]]],
    lang_code: str,
    include_translation_col: bool = True,
    data_filename: str = "",
    export_index: Optional[Dict[str, Set[str]]] = None,
    consumer: Optional[StringIdConsumer] = None,
) -> None:
    """
    Write rows to a worksheet with translations.

    Args:
        ws: Target worksheet
        rows: List of RowItem to write
        lang_table: Language table {normalized_korean: [(translation, stringid), ...]}
        lang_code: Language code (e.g., "eng", "fre")
        include_translation_col: Whether to include Translation column (False for ENG)
        data_filename: Source data file for context-aware duplicate resolution
        export_index: Optional pre-loaded EXPORT index (uses global cache if None)
        consumer: Optional StringIdConsumer for order-based StringID disambiguation
    """
    # Headers
    if include_translation_col:
        headers = ["Original (KR)", "English (ENG)", "Translation (LOC)", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    else:
        headers = ["Original (KR)", "English (ENG)", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]

    create_header_row(ws, headers)

    # Get English table for ENG column
    # (Assume lang_table includes 'eng' key from parent)

    current_row = 2
    for item in rows:
        # Use context-aware resolution if data_filename is provided
        if data_filename:
            translation, stringid, str_origin = resolve_translation(
                item.text, lang_table, data_filename, export_index,
                consumer=consumer,
            )
        else:
            # Fallback to simple first-translation lookup
            translation, stringid, str_origin = get_first_translation(lang_table, item.text)

        # Write cells
        col = 1

        # Original (KR) — use StrOrigin from language data when available
        source_text = str_origin if str_origin else item.text
        cell = ws.cell(row=current_row, column=col, value=br_to_newline(source_text))
        cell.fill = get_depth_fill(item.depth)
        cell.font = get_depth_font(item.depth)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        col += 1

        # English (ENG) - for now use translation if lang_code is eng
        eng_text = translation if lang_code == "eng" else ""
        cell = ws.cell(row=current_row, column=col, value=br_to_newline(eng_text))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        col += 1

        # Translation (LOC) - skip for ENG workbook
        if include_translation_col:
            cell = ws.cell(row=current_row, column=col, value=br_to_newline(translation) if lang_code != "eng" else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            col += 1

        # STATUS
        cell = ws.cell(row=current_row, column=col, value="")
        cell.border = THIN_BORDER
        col += 1

        # COMMENT
        cell = ws.cell(row=current_row, column=col, value="")
        cell.border = THIN_BORDER
        col += 1

        # STRINGID (as text to prevent scientific notation)
        cell = ws.cell(row=current_row, column=col, value=stringid)
        cell.number_format = '@'
        cell.border = THIN_BORDER
        col += 1

        # SCREENSHOT
        cell = ws.cell(row=current_row, column=col, value="")
        cell.border = THIN_BORDER

        current_row += 1

    # Add STATUS dropdown
    status_col = 4 if include_translation_col else 3
    add_status_dropdown(ws, status_col, start_row=2, end_row=current_row)


# =============================================================================
# KNOWLEDGE UTILITIES (shared across item, character, skill generators)
# =============================================================================

def _find_knowledge_key(item_element) -> str:
    """Search element and children for KnowledgeKey or RewardKnowledgeKey.

    Skips InspectData and PageData children — those have their own
    RewardKnowledgeKey chain handled separately by _collect_inspect_data().
    """
    for attr in ("KnowledgeKey", "RewardKnowledgeKey"):
        direct = item_element.get(attr) or ""
        if direct:
            return direct
    for child in item_element:
        if child.tag in ("InspectData", "PageData"):
            continue
        for attr in ("KnowledgeKey", "RewardKnowledgeKey"):
            kk = child.get(attr) or ""
            if kk:
                return kk
    return ""


def load_knowledge_data(
    folder: Path,
) -> Tuple[Dict[str, Tuple[str, str, str]], Dict[str, List[Tuple[str, str, str]]]]:
    """Load KnowledgeKey -> (Name, Desc, source_file) from knowledge files.

    Enhanced version that returns the Name field and tracks source_file
    for EXPORT matching.

    Returns:
        (knowledge_map, knowledge_name_index) where:
        - knowledge_map: {StrKey: (Name, Desc, source_file)}
        - knowledge_name_index: {Name: [(StrKey, Desc, source_file), ...]}
    """
    log.info("Loading knowledge data...")
    knowledge_map: Dict[str, Tuple[str, str, str]] = {}
    knowledge_name_index: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)

    if not folder.exists():
        log.warning("Knowledge folder does not exist: %s", folder)
        return knowledge_map, knowledge_name_index

    file_count = 0
    for path in iter_xml_files(folder):
        root = parse_xml_file(path)
        if root is None:
            continue
        file_count += 1

        for el in root.iter("KnowledgeInfo"):
            strkey = el.get("StrKey") or ""
            name = el.get("Name") or ""
            desc = el.get("Desc") or ""

            if strkey and strkey.lower() not in knowledge_map:
                knowledge_map[strkey.lower()] = (name, desc, path.name)
                # Build name index for Pass 2 (only first occurrence per StrKey)
                if name:
                    knowledge_name_index[name].append((strkey.lower(), desc, path.name))

    log.info("Knowledge data loaded: %d entries from %d files, %d unique names",
             len(knowledge_map), file_count, len(knowledge_name_index))
    return knowledge_map, knowledge_name_index
