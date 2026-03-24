"""
Word Count Report Generator
============================
Generates a beautiful Excel report with word counts for all generated datasheets.

Adapts to whatever datasheets were generated — scans the output folder,
detects translation columns by header name, counts words per sheet/tab,
and produces a clean summary report.

Output: GeneratedDatasheets/WordCount_Report.xlsx
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DATASHEET_OUTPUT
from core.processing import count_words_english, count_chars_chinese, contains_korean


# =============================================================================
# CJK DETECTION
# =============================================================================

# CJK language codes (character-based counting)
_CJK_CODES = {"CHS", "CHT", "CN", "JP", "JPN", "TW", "ZH", "ZHS", "ZHT", "JA"}

# CJK Unicode ranges for auto-detection
def _contains_cjk(text: str) -> bool:
    """Check if text contains CJK characters (Chinese/Japanese/Korean excluded — Korean handled separately)."""
    for ch in text:
        cp = ord(ch)
        # CJK Unified Ideographs
        if 0x4E00 <= cp <= 0x9FFF:
            return True
        # Hiragana
        if 0x3040 <= cp <= 0x309F:
            return True
        # Katakana
        if 0x30A0 <= cp <= 0x30FF:
            return True
        # CJK Extension A
        if 0x3400 <= cp <= 0x4DBF:
            return True
    return False


def _detect_counting_mode(trans_header: str, sample_texts: list) -> str:
    """
    Detect whether to count words (Latin) or characters (CJK).

    Detection order:
    1. Header contains a CJK language code (e.g., "Translation (CHS)", "CHS", "JP")
    2. Sample text contains CJK characters → character mode
    3. Default → word mode (Latin/English)

    Returns: "words" or "chars"
    """
    header_upper = trans_header.strip().upper()

    # Check if header contains a CJK language code
    # e.g., "Translation (CHS)" → extract "CHS"
    for code in _CJK_CODES:
        if code in header_upper:
            return "chars"

    # Check if header itself IS a CJK code (Quest bare codes like "CHS", "JP")
    clean = header_upper.strip("() ")
    if clean in _CJK_CODES:
        return "chars"

    # Auto-detect from sample text content
    for text in sample_texts[:10]:
        if text and _contains_cjk(str(text)):
            return "chars"

    return "words"


# =============================================================================
# STYLING CONSTANTS
# =============================================================================

# Header styles
_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
_TITLE_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_SUBHEADER_FONT = Font(bold=True, size=10, color="333333")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")

# Data styles
_DATA_FONT = Font(size=10)
_DATA_FILL_A = PatternFill("solid", fgColor="F2F7FB")
_DATA_FILL_B = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FONT = Font(bold=True, size=11, color="2F5496")
_TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
_GRAND_TOTAL_FONT = Font(bold=True, size=12, color="FFFFFF")
_GRAND_TOTAL_FILL = PatternFill("solid", fgColor="548235")

# Number format
_NUM_FORMAT = "#,##0"

# Border
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="4472C4"),
)


# =============================================================================
# TRANSLATION COLUMN DETECTION
# =============================================================================

# Known translation column header patterns (case-insensitive)
_TRANSLATION_PATTERNS = [
    "TRANSLATION",    # Translation (KR), Translation (EN), etc.
    "ENGLISH",        # English (ENG)
    "ENG",            # ENG column in Quest datasheets
    "TEXT",           # Script category
]

# Columns to SKIP (look like translation but aren't)
_SKIP_PATTERNS = [
    "SOURCETEXT",     # SourceText (KR) = Korean source, not translation
    "ORIGINAL",       # Original = Korean source
]


def _read_headers(ws) -> Dict[int, str]:
    """
    Read header row using iter_rows (streaming-safe for read_only mode).

    Returns: {1-based_col_index: HEADER_UPPER}
    """
    headers = {}
    header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_tuple = next(header_iter, None)
    if header_tuple:
        for idx, val in enumerate(header_tuple):
            if val:
                headers[idx + 1] = str(val).strip().upper()
    return headers


_KNOWN_NON_TRANS = {
    "STATUS", "COMMENT", "STRINGID", "SCREENSHOT", "COMMAND",
    "STRINGKEY", "MEMO", "ORIGINAL", "ENG", "DATATYPE", "FILENAME",
    "SOURCETEXT", "RECORDING", "DIALOGTYPE", "GROUP", "SEQUENCENAME",
    "DIALOGVOICE", "SUBTIMELINENAME", "EVENTNAME", "INSTRUCTIONS",
}


def _find_translation_col(ws) -> Optional[int]:
    """
    Find the translation column in a worksheet by header name.
    Uses iter_rows for read_only compatibility (no ws.cell calls).

    Returns 1-based column index, or None if not found.
    """
    headers = _read_headers(ws)
    if not headers:
        return None

    # Pass 1: TRANSLATION* (best match — Item, Character, Skill, Knowledge, Help, etc.)
    for col, h in headers.items():
        if h.startswith("TRANSLATION"):
            return col

    # Pass 2: ENGLISH* (e.g., "English (ENG)")
    for col, h in headers.items():
        if h.startswith("ENGLISH"):
            return col

    # Pass 3: Language column after ENG (Quest: Original | ENG | ZHO-CN | ...)
    eng_col = None
    for col, h in headers.items():
        if h == "ENG":
            eng_col = col
            break
    if eng_col is not None:
        for col in sorted(headers.keys()):
            if col > eng_col and headers[col] not in _KNOWN_NON_TRANS:
                return col

    # Pass 4: ENG column itself (Quest ENG datasheets)
    if eng_col is not None:
        return eng_col

    # Pass 5: TEXT (Script category)
    for col, h in headers.items():
        if h == "TEXT":
            return col

    return None


# =============================================================================
# WORD COUNTING
# =============================================================================

def _count_words_in_sheet(ws, trans_col: int, counting_mode: str = "words") -> Dict:
    """
    Count words (Latin) or characters (CJK) in the translation column.
    Uses iter_rows for streaming performance (instant, not minutes).

    Args:
        ws: Worksheet
        trans_col: 1-based column index of translation column
        counting_mode: "words" for Latin languages, "chars" for CJK

    Returns dict with counts and mode.
    """
    stats = {
        "total_count": 0,
        "total_rows": 0,
        "translated_rows": 0,
        "korean_rows": 0,
        "empty_rows": 0,
        "mode": counting_mode,
    }

    count_fn = count_chars_chinese if counting_mode == "chars" else count_words_english
    col_idx = trans_col - 1  # iter_rows uses 0-based tuples

    for row_tuple in ws.iter_rows(min_row=2, values_only=True):
        if col_idx >= len(row_tuple):
            stats["empty_rows"] += 1
            continue

        cell_value = row_tuple[col_idx]
        if cell_value is None or str(cell_value).strip() == "":
            stats["empty_rows"] += 1
            continue

        stats["total_rows"] += 1
        text = str(cell_value).strip()

        if contains_korean(text):
            stats["korean_rows"] += 1
            continue

        count = count_fn(text)
        stats["total_count"] += count
        stats["translated_rows"] += 1

    return stats


def _scan_datasheet(filepath: Path, log_callback=None) -> Optional[Dict]:
    """
    Scan a single datasheet Excel file and count words per sheet.

    Returns dict with:
        filename: str
        filepath: Path
        sheets: [{name, words, rows, translated, korean, empty, trans_col_header}]
        total_words: int
        total_sheets: int
    """
    def _log(msg, tag='info'):
        if log_callback:
            log_callback(msg, tag)
        print(msg)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except (PermissionError, OSError) as e:
        _log(f"    WARNING: Cannot open {filepath.name}: {e}", 'warning')
        return None
    except Exception as e:
        _log(f"    WARNING: Cannot open {filepath.name}: {e}", 'warning')
        return None

    result = {
        "filename": filepath.name,
        "filepath": filepath,
        "sheets": [],
        "total_words": 0,
        "total_sheets": 0,
    }

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == "STATUS":
                continue

            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            trans_col = _find_translation_col(ws)
            if trans_col is None:
                headers_found = _read_headers(ws)
                _log(f"    Skipped sheet '{sheet_name}': no translation column found (headers: {list(headers_found.values())[:8]})", 'warning')
                continue

            # Get header name from cached headers (no ws.cell call)
            sheet_headers = _read_headers(ws)
            trans_header = sheet_headers.get(trans_col, "")

            # Collect sample texts for CJK auto-detection using iter_rows
            sample_texts = []
            col_idx = trans_col - 1
            for row_tuple in ws.iter_rows(min_row=2, max_row=min(11, ws.max_row or 1), values_only=True):
                if col_idx < len(row_tuple) and row_tuple[col_idx]:
                    sample_texts.append(str(row_tuple[col_idx]))

            counting_mode = _detect_counting_mode(trans_header, sample_texts)
            stats = _count_words_in_sheet(ws, trans_col, counting_mode)

            result["sheets"].append({
                "name": sheet_name,
                "count": stats["total_count"],
                "mode": stats["mode"],
                "rows": stats["total_rows"],
                "translated": stats["translated_rows"],
                "korean": stats["korean_rows"],
                "empty": stats["empty_rows"],
                "trans_col_header": trans_header,
            })
            result["total_words"] += stats["total_count"]
            result["total_sheets"] += 1
    finally:
        wb.close()

    if result["total_sheets"] == 0:
        return None

    return result


# =============================================================================
# REPORT GENERATION
# =============================================================================

def generate_wordcount_report(log_callback=None) -> Optional[Path]:
    """
    Generate a Word Count Report Excel file from all generated datasheets.

    Scans GeneratedDatasheets/ for all .xlsx files, counts words in
    translation columns, and produces a clean summary report.

    Args:
        log_callback: Optional callback(message, tag) for GUI logging

    Returns:
        Path to generated report, or None if no datasheets found
    """
    def _log(msg, tag='info'):
        if log_callback:
            log_callback(msg, tag)
        print(msg)

    _log("Generating Word Count Report...")

    # Scan for all generated datasheets
    if not DATASHEET_OUTPUT.exists():
        _log("No GeneratedDatasheets folder found.", 'error')
        return None

    xlsx_files = []
    for category_dir in sorted(DATASHEET_OUTPUT.iterdir()):
        if not category_dir.is_dir():
            continue
        for f in sorted(category_dir.rglob("*.xlsx")):
            if f.name.startswith("~$") or f.name == "WordCount_Report.xlsx":
                continue
            xlsx_files.append((category_dir.name, f))

    if not xlsx_files:
        _log("No datasheet files found.", 'error')
        return None

    _log(f"  Found {len(xlsx_files)} datasheet file(s) in {len(set(d for d, _ in xlsx_files))} category folder(s)")

    # Scan each datasheet
    all_data = []  # [(category_folder, scan_result)]
    for cat_folder, filepath in xlsx_files:
        _log(f"  Scanning: {cat_folder}/{filepath.name}...")
        scan = _scan_datasheet(filepath, log_callback=log_callback)
        if scan:
            all_data.append((cat_folder, scan))

    if not all_data:
        _log("No datasheets with translation data found.", 'error')
        return None

    # Generate the report — 1 sheet per language
    report_path = DATASHEET_OUTPUT / "WordCount_Report.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group datasheets by language (extracted from filename)
    # Pattern: Category_LQA_LANG.xlsx or Quest_LQA_LANG_*.xlsx
    import re
    lang_groups = {}  # lang_code -> [(cat_folder, scan)]
    for cat_folder, scan in all_data:
        fname = scan["filename"].upper()
        # Try to extract language code from filename
        # Common patterns: _ENG.xlsx, _ZHO-CN.xlsx, _FRA.xlsx, _LQA_ENG.xlsx
        lang_match = re.search(r'_([A-Z]{2,3}(?:-[A-Z]{2})?)\.XLSX$', fname)
        if lang_match:
            lang = lang_match.group(1)
        else:
            # Fallback: use the translation column header from first sheet
            first_header = scan["sheets"][0]["trans_col_header"] if scan["sheets"] else "UNKNOWN"
            lang = first_header.upper().replace("TRANSLATION (", "").replace(")", "").replace("ENGLISH (", "").replace(")", "").strip()
            if not lang or lang == "LOC":
                lang = "UNKNOWN"
        lang_groups.setdefault(lang, []).append((cat_folder, scan))

    _log(f"  Languages detected: {', '.join(sorted(lang_groups.keys()))}")

    grand_total_count = 0
    grand_total_sheets = 0

    for lang_code in sorted(lang_groups.keys()):
        lang_data = lang_groups[lang_code]

        # Create sheet for this language (max 31 chars for sheet name)
        sheet_title = f"WordCount_{lang_code}"[:31]
        ws = wb.create_sheet(title=sheet_title)

        # --- TITLE ROW ---
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        title_cell = ws.cell(row, 1, f"WORD COUNT — {lang_code}")
        title_cell.font = _TITLE_FONT
        title_cell.fill = _TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, 8):
            ws.cell(row, c).fill = _TITLE_FILL

        row = 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=10, italic=True, color="666666")
        ws.cell(row, 1).alignment = Alignment(horizontal="center")

        row = 3  # spacer
        lang_total_count = 0

        # --- ONE TABLE PER CATEGORY ---
        for cat_idx, (cat_folder, scan) in enumerate(lang_data):
            if cat_idx > 0:
                row += 1  # empty row between category tables

            # Category header
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cat_cell = ws.cell(row, 1, f"{cat_folder} — {scan['filename']}")
            cat_cell.font = Font(bold=True, size=11, color="2F5496")
            cat_cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 8):
                ws.cell(row, c).fill = PatternFill("solid", fgColor="D6E4F0")
                ws.cell(row, c).border = _BOTTOM_BORDER
            row += 1

            # Table headers
            detail_headers = ["Sheet/Tab", "Translation Column", "Total Rows", "Translated", "Korean", "Empty", "Words / Chars"]
            for c, hdr in enumerate(detail_headers, 1):
                cell = ws.cell(row, c, hdr)
                cell.font = _SUBHEADER_FONT
                cell.fill = _SUBHEADER_FILL
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = _THIN_BORDER
            row += 1

            # Sheet rows
            for s_idx, sheet in enumerate(scan["sheets"]):
                fill = _DATA_FILL_A if s_idx % 2 == 0 else _DATA_FILL_B
                mode_label = "chars" if sheet.get("mode") == "chars" else "words"
                sheet_vals = [
                    sheet["name"],
                    f"{sheet['trans_col_header']} ({mode_label})",
                    sheet["rows"],
                    sheet["translated"],
                    sheet["korean"],
                    sheet["empty"],
                    sheet["count"],
                ]
                for c, val in enumerate(sheet_vals, 1):
                    cell = ws.cell(row, c, val)
                    cell.font = _DATA_FONT
                    cell.fill = fill
                    cell.border = _THIN_BORDER
                    if isinstance(val, int):
                        cell.number_format = _NUM_FORMAT
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                row += 1

            # Category total
            cat_total_count = scan["total_words"]
            cat_total_rows = sum(s["rows"] for s in scan["sheets"])
            cat_total_translated = sum(s["translated"] for s in scan["sheets"])
            cat_total_korean = sum(s["korean"] for s in scan["sheets"])
            cat_total_empty = sum(s["empty"] for s in scan["sheets"])
            total_vals = [
                f"TOTAL ({scan['total_sheets']} sheets)",
                "",
                cat_total_rows,
                cat_total_translated,
                cat_total_korean,
                cat_total_empty,
                cat_total_count,
            ]
            for c, val in enumerate(total_vals, 1):
                cell = ws.cell(row, c, val)
                cell.font = _TOTAL_FONT
                cell.fill = _TOTAL_FILL
                cell.border = _THIN_BORDER
                if isinstance(val, int):
                    cell.number_format = _NUM_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            row += 1
            lang_total_count += cat_total_count

        # Language grand total
        row += 1
        grand_vals = [
            f"GRAND TOTAL — {lang_code}",
            f"{len(lang_data)} file(s)",
            "", "", "", "",
            lang_total_count,
        ]
        for c, val in enumerate(grand_vals, 1):
            cell = ws.cell(row, c, val)
            cell.font = _GRAND_TOTAL_FONT
            cell.fill = _GRAND_TOTAL_FILL
            cell.border = _THIN_BORDER
            if isinstance(val, int):
                cell.number_format = _NUM_FORMAT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

        # Column widths
        col_widths = [25, 30, 12, 14, 14, 12, 16]
        for c, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.freeze_panes = "A4"

        grand_total_count += lang_total_count
        grand_total_sheets += sum(s["total_sheets"] for _, s in lang_data)

    # Save
    try:
        wb.save(report_path)
    except PermissionError:
        _log(f"  ERROR: Cannot save report — {report_path.name} is open in another program. Close it and retry.", 'error')
        return None
    except OSError as e:
        _log(f"  ERROR: Cannot save report: {e}", 'error')
        return None
    _log(f"  Word Count Report saved: {report_path.name}", 'success')
    _log(f"  {len(lang_groups)} language(s): {', '.join(sorted(lang_groups.keys()))}")
    _log(f"  Grand Total: {grand_total_count:,} words/chars across {len(all_data)} file(s), {grand_total_sheets} sheet(s)")

    return report_path


if __name__ == "__main__":
    report = generate_wordcount_report()
    if report:
        print(f"\nReport saved to: {report}")
    else:
        print("\nNo report generated.")
