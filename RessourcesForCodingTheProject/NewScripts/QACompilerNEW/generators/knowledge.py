"""
Knowledge Datasheet Generator
=============================
Extracts Knowledge data from StaticInfo XMLs with hierarchical structure.

Structure:
- ONE Excel sheet per TOP-LEVEL KnowledgeGroupInfo ("mega root" like 인물, 생태, etc.)
- Everything nested inside that mega root goes into that single sheet
- Proper indentation based on hierarchy depth

Key features:
- KnowledgeInfo nested directly inside KnowledgeGroupInfo
- Knowledge nodes in CharacterInfo that reference groups via KnowledgeGroupKey
- LevelData descriptions for detailed progression info
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import RESOURCE_FOLDER, LANGUAGE_FOLDER, DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    parse_xml_file,
    iter_xml_files,
    load_language_tables,
    normalize_placeholders,
    autofit_worksheet,
    THIN_BORDER,
)

log = get_logger("KnowledgeGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection (normalized)."""
    if text:
        normalized = normalize_placeholders(text)
        if normalized:
            _collected_korean_strings.add(normalized)


# Text to ignore
IGNORE_LIST = {
    "해제/면역 연금 포션",
}


def should_ignore(text: str) -> bool:
    """Check if text should be ignored."""
    return any(bad in text for bad in IGNORE_LIST)


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class LevelLine:
    """Level-specific data."""
    level: str
    desc: str


@dataclass
class KnowledgeItem:
    """Single knowledge item."""
    strkey: str
    name: str
    desc: str
    levels: List[LevelLine] = field(default_factory=list)


@dataclass
class GroupNode:
    """Knowledge group node with children."""
    strkey: str
    name: str
    desc: str
    icon: bool
    children: List["GroupNode"] = field(default_factory=list)
    knowledges: List[KnowledgeItem] = field(default_factory=list)


# =============================================================================
# STYLING
# =============================================================================

# Depth 0: Master group (tab name) - GOLD, BIGGEST
_depth0_fill = PatternFill("solid", fgColor="FFD700")
_depth0_font = Font(bold=True, size=14)
_depth0_row_height = 40

# Depth 1: Sub-group just below master - BRIGHT PURPLE, BIG
_depth1_fill = PatternFill("solid", fgColor="9966FF")
_depth1_font = Font(bold=True, size=12, color="FFFFFF")
_depth1_row_height = 35

# Other depths
_depth_fills: Dict[int, PatternFill] = {
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FCE4D6"),
    4: PatternFill("solid", fgColor="DDEBF7"),
    5: PatternFill("solid", fgColor="FFF2CC"),
}
_icon_fill = PatternFill("solid", fgColor="9BC2E6")
_no_colour_fill = PatternFill("solid", fgColor="FFFDEB")

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_normal_font = Font()


def _get_style_for_depth(depth: int, is_icon: bool) -> Tuple[PatternFill, Font, Optional[float]]:
    """
    Returns (fill, font, row_height) for a given depth.
    row_height is None for default height.
    """
    if depth == 0:
        return _depth0_fill, _depth0_font, _depth0_row_height
    elif depth == 1:
        return _depth1_fill, _depth1_font, _depth1_row_height
    else:
        fill = _depth_fills.get(depth)
        if fill is None and is_icon:
            fill = _icon_fill
        if fill is None:
            fill = _no_colour_fill
        return fill, _bold_font, None


# =============================================================================
# EXTRACTION HELPERS
# =============================================================================

def extract_knowledge_info(node) -> Optional[KnowledgeItem]:
    """Extract KnowledgeInfo element."""
    strkey = node.get("StrKey") or ""
    name = node.get("Name") or ""
    desc = node.get("Desc") or ""

    if not name and not desc:
        return None

    # Collect Korean strings for coverage tracking
    _collect_korean_string(name)
    _collect_korean_string(desc)

    ki = KnowledgeItem(strkey=strkey, name=name, desc=desc)

    for lvl in node.findall("LevelData"):
        lvl_no = lvl.get("Level") or ""
        lvl_desc = lvl.get("Desc") or ""
        if lvl_desc and not should_ignore(lvl_desc):
            ki.levels.append(LevelLine(level=lvl_no, desc=lvl_desc))

    return ki


def extract_character_knowledge(node) -> Optional[KnowledgeItem]:
    """Extract Knowledge element inside CharacterInfo."""
    strkey = node.get("StrKey") or ""
    name = node.get("Name") or ""
    desc = node.get("Desc") or ""

    if not name and not desc:
        return None

    # Collect Korean strings for coverage tracking
    _collect_korean_string(name)
    _collect_korean_string(desc)

    ki = KnowledgeItem(strkey=strkey, name=name, desc=desc)

    for lvl in node.findall("LevelData"):
        lvl_no = lvl.get("Level") or ""
        lvl_desc = lvl.get("Desc") or ""
        if lvl_desc and not should_ignore(lvl_desc):
            ki.levels.append(LevelLine(level=lvl_no, desc=lvl_desc))

    return ki


# =============================================================================
# BUILD HIERARCHY
# =============================================================================

def build_hierarchy(folder: Path) -> Tuple[List[GroupNode], Dict[str, GroupNode]]:
    """
    Two-phase parsing:
    1. Build the nested group hierarchy from KnowledgeGroupInfo structure
    2. Find Knowledge nodes with KnowledgeGroupKey and attach them to groups

    Returns:
        - mega_roots: List of top-level GroupNode (these become Excel tabs)
        - groups_by_key: Dict for lookup by StrKey
    """
    groups_by_key: Dict[str, GroupNode] = {}
    mega_roots_by_key: Dict[str, GroupNode] = {}
    pending_knowledges: List[Tuple[str, KnowledgeItem]] = []
    seen_knowledge_keys: set = set()

    log.info("Phase 1: Building group hierarchy...")

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        def parse_group(node) -> Optional[GroupNode]:
            """Recursively parse KnowledgeGroupInfo and its nested children."""
            if node.tag != "KnowledgeGroupInfo":
                return None

            strkey = node.get("StrKey") or ""
            name = node.get("GroupName") or ""
            desc = node.get("Desc") or ""
            icon = bool(node.get("KnowledgeGroupIcon"))

            # Collect Korean strings for coverage tracking
            _collect_korean_string(name)
            _collect_korean_string(desc)

            # Skip if already processed
            if strkey and strkey in groups_by_key:
                existing = groups_by_key[strkey]
                for ch in node:
                    if ch.tag == "KnowledgeInfo":
                        ki = extract_knowledge_info(ch)
                        if ki and ki.strkey not in seen_knowledge_keys:
                            existing.knowledges.append(ki)
                            if ki.strkey:
                                seen_knowledge_keys.add(ki.strkey)
                return None

            group = GroupNode(
                strkey=strkey,
                name=name,
                desc=desc,
                icon=icon,
            )

            if strkey:
                groups_by_key[strkey] = group

            # Process children
            for ch in node:
                if ch.tag == "KnowledgeGroupInfo":
                    child_group = parse_group(ch)
                    if child_group:
                        group.children.append(child_group)
                elif ch.tag == "KnowledgeInfo":
                    ki = extract_knowledge_info(ch)
                    if ki:
                        if ki.strkey and ki.strkey in seen_knowledge_keys:
                            continue
                        group.knowledges.append(ki)
                        if ki.strkey:
                            seen_knowledge_keys.add(ki.strkey)

            return group

        def find_top_level_groups(node):
            """Find KnowledgeGroupInfo that are direct children of root."""
            for ch in node:
                if ch.tag == "KnowledgeGroupInfo":
                    strkey = ch.get("StrKey") or ""
                    if strkey and strkey in mega_roots_by_key:
                        parse_group(ch)
                        continue

                    group = parse_group(ch)
                    if group and group.strkey:
                        mega_roots_by_key[group.strkey] = group
                else:
                    find_top_level_groups(ch)

        def find_character_knowledge(node):
            """Find Knowledge nodes inside CharacterInfo etc."""
            for ch in node:
                if ch.tag == "Knowledge":
                    group_key = ch.get("KnowledgeGroupKey") or ""
                    strkey = ch.get("StrKey") or ""
                    if group_key:
                        if strkey and strkey in seen_knowledge_keys:
                            continue
                        ki = extract_character_knowledge(ch)
                        if ki:
                            pending_knowledges.append((group_key, ki))
                            if strkey:
                                seen_knowledge_keys.add(strkey)
                if ch.tag == "CharacterInfo":
                    for inner in ch:
                        if inner.tag == "Knowledge":
                            group_key = inner.get("KnowledgeGroupKey") or ""
                            strkey = inner.get("StrKey") or ""
                            if group_key:
                                if strkey and strkey in seen_knowledge_keys:
                                    continue
                                ki = extract_character_knowledge(inner)
                                if ki:
                                    pending_knowledges.append((group_key, ki))
                                    if strkey:
                                        seen_knowledge_keys.add(strkey)
                find_character_knowledge(ch)

        find_top_level_groups(root_el)
        find_character_knowledge(root_el)

    mega_roots = list(mega_roots_by_key.values())

    log.info("Phase 1 complete: %d mega roots, %d total groups", len(mega_roots), len(groups_by_key))

    # Phase 2: Attach pending Knowledge items to their groups
    log.info("Phase 2: Attaching %d character knowledge items...", len(pending_knowledges))
    attached = 0
    for group_key, ki in pending_knowledges:
        if group_key in groups_by_key:
            groups_by_key[group_key].knowledges.append(ki)
            attached += 1

    log.info("Phase 2 complete: %d knowledge items attached", attached)

    return mega_roots, groups_by_key


# =============================================================================
# ROW GENERATION
# =============================================================================

# (depth, text, needs_translation, is_icon_group, is_name_attribute)
RowItem = Tuple[int, str, bool, bool, bool]


def emit_group_rows(group: GroupNode, depth: int) -> List[RowItem]:
    """Generate rows for a group and all its nested content."""
    rows: List[RowItem] = []

    # Emit group name
    if group.name and not should_ignore(group.name):
        rows.append((depth, group.name, True, group.icon, True))

    # Emit group description
    if group.desc and not should_ignore(group.desc):
        rows.append((depth + 1, group.desc, True, group.icon, False))

    # Emit child groups (recursively)
    for child in group.children:
        rows.extend(emit_group_rows(child, depth + 1))

    # Emit knowledge items in this group
    for kn in group.knowledges:
        rows.extend(emit_knowledge_rows(kn, depth + 1))

    return rows


def emit_knowledge_rows(kn: KnowledgeItem, depth: int) -> List[RowItem]:
    """Generate rows for a single knowledge item."""
    rows: List[RowItem] = []

    # Knowledge name
    if kn.name and not should_ignore(kn.name):
        rows.append((depth, kn.name, True, False, True))

    # Knowledge description
    if kn.desc and not should_ignore(kn.desc):
        rows.append((depth + 1, kn.desc, True, False, False))

    # Level data
    for lvl in kn.levels:
        label = f"Level {lvl.level}"
        if not should_ignore(label):
            rows.append((depth + 1, label, False, False, False))
        if lvl.desc and not should_ignore(lvl.desc):
            rows.append((depth + 2, lvl.desc, True, False, False))

    return rows


# =============================================================================
# EXCEL WRITER
# =============================================================================

def write_workbook(
    mega_roots: List[GroupNode],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with one sheet per mega root.

    Columns:
      - Original (KR)
      - English (ENG)
      - Translation (OTHER) [skipped for ENG workbook]
      - STATUS, COMMENT, STRINGID, SCREENSHOT
    """
    wb = Workbook()
    wb.remove(wb.active)

    is_eng = lang_code.lower() == "eng"

    for root_group in mega_roots:
        rows = emit_group_rows(root_group, 0)
        if not rows:
            continue

        # Sheet title
        normalized_name = normalize_placeholders(root_group.name)
        title_eng = eng_tbl.get(normalized_name, ("", ""))[0].strip()
        title_orig = root_group.name.strip() or "Sheet"
        title = (title_eng or title_orig)[:31]
        title = re.sub(r"[\\/*?:\[\]]", "_", title)

        base_title = title
        c = 1
        while title in wb.sheetnames:
            title = f"{base_title[:28]}_{c}"
            c += 1

        ws = wb.create_sheet(title)

        # Header row
        headers: List = []
        h1 = ws.cell(1, 1, "Original (KR)")
        h2 = ws.cell(1, 2, "English (ENG)")
        headers.extend([h1, h2])

        if not is_eng:
            h3 = ws.cell(1, 3, f"Translation ({lang_code.upper()})")
            headers.append(h3)

        start_extra_col = len(headers) + 1
        extra_names = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
        for idx, name in enumerate(extra_names, start=start_extra_col):
            headers.append(ws.cell(1, idx, name))

        for hcell in headers:
            hcell.font = _header_font
            hcell.fill = _header_fill
            hcell.alignment = Alignment(horizontal="center", vertical="center")
            hcell.border = THIN_BORDER
        ws.row_dimensions[1].height = 25

        # Data rows
        r_idx = 2

        for (depth, text, needs_tr, is_icon, is_name_attr) in rows:
            fill, font, rh = _get_style_for_depth(depth, is_icon)
            if depth >= 2:
                if is_name_attr:
                    font = _bold_font
                else:
                    font = _normal_font

            norm_text = normalize_placeholders(text)

            eng_tr, sid_eng = ("", "")
            if needs_tr:
                eng_tr, sid_eng = eng_tbl.get(norm_text, ("", ""))

            other_tr, sid_other = ("", "")
            if needs_tr and not is_eng and lang_tbl is not None:
                other_tr, sid_other = lang_tbl.get(norm_text, ("", ""))

            sid = sid_eng if is_eng else sid_other

            # Write core columns
            c_orig = ws.cell(r_idx, 1, text)
            c_eng = ws.cell(r_idx, 2, eng_tr)
            c_other = None
            if not is_eng:
                c_other = ws.cell(r_idx, 3, other_tr)

            # Extra columns
            col_off = 2 if is_eng else 3
            c_status = ws.cell(r_idx, col_off + 1, "")
            c_comment = ws.cell(r_idx, col_off + 2, "")
            c_stringid = ws.cell(r_idx, col_off + 3, sid_other if not is_eng else sid_eng)
            c_screenshot = ws.cell(r_idx, col_off + 4, "")

            # Styling
            c_orig.fill = fill
            c_orig.font = font
            c_orig.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
            c_orig.border = THIN_BORDER

            c_eng.fill = fill
            c_eng.font = font
            c_eng.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
            c_eng.border = THIN_BORDER

            if c_other:
                c_other.fill = fill
                c_other.font = font
                c_other.alignment = Alignment(indent=depth, wrap_text=True, vertical="center")
                c_other.border = THIN_BORDER

            c_status.fill = fill
            c_status.border = THIN_BORDER
            c_comment.fill = fill
            c_comment.border = THIN_BORDER
            c_stringid.fill = fill
            c_stringid.font = _bold_font
            c_stringid.border = THIN_BORDER
            c_stringid.number_format = '@'
            c_screenshot.fill = fill
            c_screenshot.border = THIN_BORDER

            # Apply row height for depth 0 and 1
            if rh is not None:
                ws.row_dimensions[r_idx].height = rh

            r_idx += 1

        last_row = r_idx - 1

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["B"].hidden = not is_eng

        if not is_eng:
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 11
            ws.column_dimensions["E"].width = 70
            ws.column_dimensions["F"].width = 25
            ws.column_dimensions["G"].width = 20
        else:
            ws.column_dimensions["C"].width = 11
            ws.column_dimensions["D"].width = 70
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["F"].width = 20

        # Status dropdown
        status_col_idx = 3 if is_eng else 4
        status_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_OPTIONS)}"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"${status_letter}$2:${status_letter}${last_row}")

        # Auto-fit
        autofit_worksheet(ws)

        log.info("  Sheet '%s': %d rows", title, last_row - 1)

    if wb.worksheets:
        wb.save(out_path)
        log.info("→ Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_knowledge_datasheets() -> Dict:
    """
    Generate Knowledge datasheets for all languages.

    Returns:
        Dict with results: {
            "category": "Knowledge",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Knowledge",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("Knowledge Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "Knowledge_LQA_All"
    output_folder.mkdir(exist_ok=True)

    # Check paths
    if not RESOURCE_FOLDER.exists():
        result["errors"].append(f"Resource folder not found: {RESOURCE_FOLDER}")
        log.error("Resource folder not found: %s", RESOURCE_FOLDER)
        return result

    if not LANGUAGE_FOLDER.exists():
        result["errors"].append(f"Language folder not found: {LANGUAGE_FOLDER}")
        log.error("Language folder not found: %s", LANGUAGE_FOLDER)
        return result

    try:
        # 1. Build hierarchy from StaticInfo
        mega_roots, groups_by_key = build_hierarchy(RESOURCE_FOLDER)

        if not mega_roots:
            result["errors"].append("No mega root groups found!")
            log.warning("No mega root groups found!")
            return result

        log.info("Mega roots found (these become Excel tabs):")
        for root in mega_roots:
            log.info("  • %s (%s)", root.name, root.strkey)

        # 2. Load language tables
        lang_tables = load_language_tables(LANGUAGE_FOLDER)

        if not lang_tables:
            result["errors"].append("No language tables found!")
            log.warning("No language tables found!")
            return result

        eng_tbl = lang_tables.get("eng", {})

        # 3. Generate workbooks
        log.info("Generating Excel workbooks...")
        total = len(lang_tables)

        for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
            log.info("(%d/%d) Processing language: %s", idx, total, code.upper())
            out_xlsx = output_folder / f"Knowledge_LQA_{code.upper()}.xlsx"
            if code.lower() == "eng":
                write_workbook(mega_roots, eng_tbl, None, code, out_xlsx)
            else:
                write_workbook(mega_roots, eng_tbl, tbl, code, out_xlsx)
            result["files_created"] += 1

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Knowledge generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_knowledge_datasheets()
    print(f"\nResult: {result}")
