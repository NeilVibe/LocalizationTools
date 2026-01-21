"""
Script Datasheet Generator
==========================
Processes Script/Dialog LQA files for QA workflow.

Script category handles two types of source files:
- Sequencer_LQA_LocLanguage.xlsx → "Script" tab
- Dialog_LQA_LocLanguage.xlsx → "Dialog" tab

These files are provided externally (not generated from XML) and have columns:
- EventName: StringID for matching
- Text: Translation text
- STATUS: ISSUE/NON-ISSUE status
- MEMO: Comments (maps to COMMENT in master)
- NO SCREENSHOT column

Matching Logic (DIFFERENT from standard):
- Primary: Text + EventName (both must match)
- Fallback: EventName ONLY (not Text only like other categories!)

Output:
- Master_Script.xlsx with two tabs (Script, Dialog)
- All columns visible (no base column hiding)
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DATASHEET_OUTPUT, STATUS_OPTIONS
from generators.base import (
    get_logger,
    autofit_worksheet,
    THIN_BORDER,
)

log = get_logger("ScriptGenerator")

# =============================================================================
# KOREAN STRING COLLECTION (for coverage tracking)
# =============================================================================

_collected_korean_strings: set = set()


def reset_korean_collection() -> None:
    """Reset the Korean string collection before a new run."""
    global _collected_korean_strings
    _collected_korean_strings = set()


def get_collected_korean_strings() -> set:
    """Return a copy of collected Korean strings."""
    return _collected_korean_strings.copy()


def _collect_korean_string(text: str) -> None:
    """Add a Korean string to the collection."""
    if text and text.strip():
        _collected_korean_strings.add(text.strip())


# =============================================================================
# STYLING
# =============================================================================

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_data_font = Font(size=10)


# =============================================================================
# TEMPLATE GENERATION
# =============================================================================

def create_script_template_workbook() -> Workbook:
    """
    Create a template workbook for Script category with two sheets.

    Returns:
        Workbook with Script and Dialog tabs, each with proper headers.
    """
    wb = Workbook()

    # Create Script sheet (rename active sheet)
    ws_script = wb.active
    ws_script.title = "Script"

    # Create Dialog sheet
    ws_dialog = wb.create_sheet("Dialog")

    # Define headers for Script category
    # Note: MEMO is used instead of COMMENT, no SCREENSHOT column
    headers = ["EventName", "Text", "Translation", "STATUS", "MEMO", "STRINGID"]

    for ws in [ws_script, ws_dialog]:
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Set column widths
        ws.column_dimensions["A"].width = 30  # EventName
        ws.column_dimensions["B"].width = 50  # Text (original)
        ws.column_dimensions["C"].width = 50  # Translation
        ws.column_dimensions["D"].width = 12  # STATUS
        ws.column_dimensions["E"].width = 40  # MEMO
        ws.column_dimensions["F"].width = 25  # STRINGID

        # Add STATUS dropdown
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NON-ISSUE"',  # Script uses ISSUE/NON-ISSUE
            allow_blank=True,
        )
        dv.error = "Invalid status"
        dv.prompt = "Select status"
        ws.add_data_validation(dv)

        # Add some placeholder rows for structure
        for row in range(2, 6):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row, col, "")
                cell.border = THIN_BORDER
            dv.add(ws.cell(row, 4))  # STATUS column

        ws.row_dimensions[1].height = 25

    return wb


# =============================================================================
# MAIN GENERATOR FUNCTION
# =============================================================================

def generate_script_datasheets() -> Dict:
    """
    Generate Script category template datasheets.

    Script category is unique:
    - Uses external source files (Sequencer/Dialog LQA files)
    - Has MEMO column instead of COMMENT
    - Has NO SCREENSHOT column
    - Two-tab output (Script, Dialog)

    Returns:
        Dict with results: {
            "category": "Script",
            "files_created": N,
            "errors": [...]
        }
    """
    result = {
        "category": "Script",
        "files_created": 0,
        "errors": [],
    }

    # Reset Korean string collection
    reset_korean_collection()

    log.info("=" * 70)
    log.info("Script Datasheet Generator")
    log.info("=" * 70)

    # Ensure output folder exists
    DATASHEET_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_folder = DATASHEET_OUTPUT / "Script_LQA_Template"
    output_folder.mkdir(exist_ok=True)

    try:
        # Create template workbook
        log.info("Creating Script template workbook...")
        wb = create_script_template_workbook()

        # Save template
        output_path = output_folder / "LQA_Script_Template.xlsx"
        wb.save(output_path)
        result["files_created"] += 1
        log.info("Created: %s", output_path.name)

        # Note about external source files
        log.info("")
        log.info("NOTE: Script category uses external LQA source files:")
        log.info("  - Sequencer_LQA_<Language>.xlsx → 'Script' tab")
        log.info("  - Dialog_LQA_<Language>.xlsx → 'Dialog' tab")
        log.info("")
        log.info("Testers should populate this template with data from those files.")
        log.info("The compiler will merge both tabs into Master_Script.xlsx")

        log.info("=" * 70)
        log.info("Done! Output folder: %s", output_folder)
        log.info("=" * 70)

    except Exception as e:
        result["errors"].append(f"Generator error: {e}")
        log.exception("Error in Script generator: %s", e)

    return result


# Allow standalone execution for testing
if __name__ == "__main__":
    result = generate_script_datasheets()
    print(f"\nResult: {result}")
