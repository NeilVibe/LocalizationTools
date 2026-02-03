#!/usr/bin/env python3
"""
Benchmark script for QACompiler master file building operations.

Measures performance of key operations on a 10,000-row Script-type fixture:
1. safe_load_workbook() - Excel file loading with auto-repair
2. find_column_by_header() - Header scanning (10,000 iterations)
3. Alignment object creation - 10,000 new objects vs reusing 1
4. PatternFill object creation - 10,000 new objects vs reusing 1
5. build_master_index() - Content-based index building on 10K rows

Requires: benchmark_script_10k.xlsx (run create_benchmark_fixture.py first)
"""

import gc
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

from core.excel_ops import safe_load_workbook, find_column_by_header
from core.matching import build_master_index


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "benchmark_script_10k.xlsx"

# Number of iterations for per-call benchmarks
ITERATIONS = 10_000


def fmt_ms(seconds: float) -> str:
    """Format seconds as milliseconds with 2 decimal places."""
    return f"{seconds * 1000:.2f} ms"


def fmt_us(seconds: float) -> str:
    """Format seconds as microseconds with 1 decimal place."""
    return f"{seconds * 1_000_000:.1f} us"


def benchmark_load_workbook():
    """Benchmark: Load 10K-row workbook with safe_load_workbook."""
    gc.collect()
    start = time.perf_counter()
    wb = safe_load_workbook(FIXTURE_PATH)
    elapsed = time.perf_counter() - start

    total_rows = 0
    for name in wb.sheetnames:
        total_rows += wb[name].max_row - 1  # Subtract header
    wb.close()

    return elapsed, total_rows


def benchmark_find_column_by_header(wb):
    """Benchmark: Call find_column_by_header N times on a worksheet."""
    ws = wb[wb.sheetnames[0]]

    # Warm up (ensure any lazy loading is done)
    find_column_by_header(ws, "STATUS")

    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        find_column_by_header(ws, "STATUS")
    elapsed = time.perf_counter() - start

    return elapsed


def benchmark_find_column_miss(wb):
    """Benchmark: find_column_by_header with a header that does NOT exist."""
    ws = wb[wb.sheetnames[0]]

    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        find_column_by_header(ws, "NONEXISTENT_COLUMN")
    elapsed = time.perf_counter() - start

    return elapsed


def benchmark_alignment_creation():
    """Benchmark: Creating N Alignment objects vs reusing 1."""
    # --- Create new each time ---
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        a = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elapsed_new = time.perf_counter() - start

    # --- Reuse single instance ---
    shared = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        a = shared
    elapsed_reuse = time.perf_counter() - start

    return elapsed_new, elapsed_reuse


def benchmark_patternfill_creation():
    """Benchmark: Creating N PatternFill objects vs reusing 1."""
    # --- Create new each time ---
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        f = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    elapsed_new = time.perf_counter() - start

    # --- Reuse single instance ---
    shared = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        f = shared
    elapsed_reuse = time.perf_counter() - start

    return elapsed_new, elapsed_reuse


def benchmark_full_style_set():
    """Benchmark: Creating a full style tuple (Font+Fill+Alignment+Border) N times vs reuse."""
    # --- Create new each time ---
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        font = Font(bold=True, color="000000")
        fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="medium", color="4472C4"),
            right=Side(style="medium", color="4472C4"),
            top=Side(style="medium", color="4472C4"),
            bottom=Side(style="medium", color="4472C4"),
        )
    elapsed_new = time.perf_counter() - start

    # --- Reuse single instances ---
    s_font = Font(bold=True, color="000000")
    s_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    s_align = Alignment(horizontal="center", vertical="center")
    s_border = Border(
        left=Side(style="medium", color="4472C4"),
        right=Side(style="medium", color="4472C4"),
        top=Side(style="medium", color="4472C4"),
        bottom=Side(style="medium", color="4472C4"),
    )
    gc.collect()
    start = time.perf_counter()
    for _ in range(ITERATIONS):
        font = s_font
        fill = s_fill
        align = s_align
        border = s_border
    elapsed_reuse = time.perf_counter() - start

    return elapsed_new, elapsed_reuse


def benchmark_build_master_index(wb):
    """Benchmark: Build content-based master index for the largest sheet."""
    # Use the largest sheet
    largest_sheet = None
    largest_rows = 0
    for name in wb.sheetnames:
        rows = wb[name].max_row
        if rows > largest_rows:
            largest_rows = rows
            largest_sheet = name

    ws = wb[largest_sheet]

    gc.collect()
    start = time.perf_counter()
    index = build_master_index(ws, "Sequencer", is_english=True)
    elapsed = time.perf_counter() - start

    primary_keys = len(index["primary"])
    fallback_keys = len(index["fallback"])

    return elapsed, largest_sheet, largest_rows - 1, primary_keys, fallback_keys


def print_header(title: str):
    """Print a formatted section header."""
    print()
    print(f"  {title}")
    print(f"  {'=' * len(title)}")


def print_row(label: str, value: str, note: str = ""):
    """Print a formatted result row."""
    if note:
        print(f"  {label:<45} {value:>12}  {note}")
    else:
        print(f"  {label:<45} {value:>12}")


def main():
    if not FIXTURE_PATH.exists():
        print(f"ERROR: Fixture file not found: {FIXTURE_PATH}")
        print(f"       Run create_benchmark_fixture.py first.")
        sys.exit(1)

    file_size_kb = FIXTURE_PATH.stat().st_size / 1024

    print()
    print("=" * 72)
    print("  QACompiler Benchmark - Script-Type 10K Row File")
    print("=" * 72)
    print(f"  Fixture: {FIXTURE_PATH.name} ({file_size_kb:.0f} KB)")
    print(f"  Iterations per micro-benchmark: {ITERATIONS:,}")

    # -----------------------------------------------------------------------
    # 1. Load workbook
    # -----------------------------------------------------------------------
    print_header("1. Workbook Loading (safe_load_workbook)")

    elapsed, total_rows = benchmark_load_workbook()
    print_row("Load 10K-row .xlsx", fmt_ms(elapsed), f"({total_rows:,} data rows)")

    # Load once for remaining benchmarks
    wb = safe_load_workbook(FIXTURE_PATH)

    # -----------------------------------------------------------------------
    # 2. find_column_by_header
    # -----------------------------------------------------------------------
    print_header(f"2. find_column_by_header ({ITERATIONS:,} calls)")

    elapsed_hit = benchmark_find_column_by_header(wb)
    per_call_hit = elapsed_hit / ITERATIONS
    print_row("Header found (STATUS)", fmt_ms(elapsed_hit), f"({fmt_us(per_call_hit)}/call)")

    elapsed_miss = benchmark_find_column_miss(wb)
    per_call_miss = elapsed_miss / ITERATIONS
    print_row("Header NOT found (miss)", fmt_ms(elapsed_miss), f"({fmt_us(per_call_miss)}/call)")

    # -----------------------------------------------------------------------
    # 3. Alignment creation
    # -----------------------------------------------------------------------
    print_header(f"3. Alignment Object Creation ({ITERATIONS:,}x)")

    elapsed_new, elapsed_reuse = benchmark_alignment_creation()
    speedup = elapsed_new / elapsed_reuse if elapsed_reuse > 0 else float("inf")
    print_row("Create new each time", fmt_ms(elapsed_new))
    print_row("Reuse single instance", fmt_ms(elapsed_reuse))
    print_row("Speedup (reuse vs new)", f"{speedup:.0f}x", "")

    # -----------------------------------------------------------------------
    # 4. PatternFill creation
    # -----------------------------------------------------------------------
    print_header(f"4. PatternFill Object Creation ({ITERATIONS:,}x)")

    elapsed_new, elapsed_reuse = benchmark_patternfill_creation()
    speedup = elapsed_new / elapsed_reuse if elapsed_reuse > 0 else float("inf")
    print_row("Create new each time", fmt_ms(elapsed_new))
    print_row("Reuse single instance", fmt_ms(elapsed_reuse))
    print_row("Speedup (reuse vs new)", f"{speedup:.0f}x", "")

    # -----------------------------------------------------------------------
    # 5. Full style set (Font + Fill + Alignment + Border)
    # -----------------------------------------------------------------------
    print_header(f"5. Full Style Set Creation ({ITERATIONS:,}x)")

    elapsed_new, elapsed_reuse = benchmark_full_style_set()
    speedup = elapsed_new / elapsed_reuse if elapsed_reuse > 0 else float("inf")
    print_row("Create 4 objects each time", fmt_ms(elapsed_new))
    print_row("Reuse 4 shared instances", fmt_ms(elapsed_reuse))
    print_row("Speedup (reuse vs new)", f"{speedup:.0f}x", "")

    # -----------------------------------------------------------------------
    # 6. build_master_index
    # -----------------------------------------------------------------------
    print_header("6. build_master_index (Sequencer category)")

    elapsed, sheet_name, data_rows, primary, fallback = benchmark_build_master_index(wb)
    print_row(f"Index sheet '{sheet_name}'", fmt_ms(elapsed), f"({data_rows:,} rows)")
    print_row("Primary keys indexed", str(primary))
    print_row("Fallback keys indexed", str(fallback))

    wb.close()

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    print()
    print("-" * 72)
    print("  SUMMARY")
    print("-" * 72)
    print()
    print("  Key takeaways for QACompiler optimization:")
    print("  - Workbook load time is dominated by openpyxl XML parsing")
    print("  - find_column_by_header() is very cheap per call (linear scan of ~5 cols)")
    print("  - Style object creation (Alignment, PatternFill, Font, Border) adds up")
    print("    when done per-row; reusing shared instances avoids this overhead")
    print("  - build_master_index() is O(n) and builds O(1) lookup; critical for")
    print("    10K+ row Script files where naive O(n^2) matching would be too slow")
    print()
    print("=" * 72)


if __name__ == "__main__":
    main()
