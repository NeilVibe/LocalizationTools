"""
Test that reproduces the EXACT scenario from user's logs:
- Master_Script.xlsx with sheets 'English Script', 'Work Transform'
- STATUS_{User} columns exist
- But sample values show (empty)

We test multiple hypotheses for why data might not be read.
"""
import sys
import tempfile
import shutil
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))


def test_scenario_1_data_starts_after_row_11():
    """
    Hypothesis: Data exists but starts AFTER row 11 (old sample only checked first 10 rows).
    User might have 1000+ header/empty rows before actual data.
    """
    print("\n" + "="*60)
    print("SCENARIO 1: Data starts after row 11")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        # Create Master_Script.xlsx with data starting at row 500
        wb = Workbook()
        ws = wb.active
        ws.title = "English Script"

        # Headers at row 1
        ws.cell(1, 1, "EventName")
        ws.cell(1, 2, "Text")
        ws.cell(1, 3, "Translation")
        ws.cell(1, 4, "STATUS_유지윤")
        ws.cell(1, 5, "STATUS_조서영")

        # Rows 2-499 are EMPTY
        # Data starts at row 500
        ws.cell(500, 1, "EVT001")
        ws.cell(500, 2, "Hello")
        ws.cell(500, 3, "안녕")
        ws.cell(500, 4, "FIXED")
        ws.cell(500, 5, "REPORTED")

        ws.cell(501, 1, "EVT002")
        ws.cell(501, 4, "CHECKING")
        ws.cell(501, 5, "NON-ISSUE")

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"Result: {result}")

        if "Script" in result:
            script = result["Script"]
            print(f"유지윤: {script.get('유지윤', 'NOT FOUND')}")
            print(f"조서영: {script.get('조서영', 'NOT FOUND')}")

            # Should find data even if it starts late
            assert script.get("유지윤", {}).get("fixed", 0) == 1, "Should find FIXED at row 500"
            assert script.get("조서영", {}).get("reported", 0) == 1, "Should find REPORTED at row 500"
            print("PASS: Data found even when starting at row 500")
        else:
            print("FAIL: Script category not found!")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


def test_scenario_2_lowercase_status_header():
    """
    Hypothesis: Header is 'status_유지윤' (lowercase) not 'STATUS_유지윤'.
    """
    print("\n" + "="*60)
    print("SCENARIO 2: Lowercase 'status_' header")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "English Script"

        # LOWERCASE headers
        ws.cell(1, 1, "EventName")
        ws.cell(1, 2, "status_유지윤")  # lowercase!
        ws.cell(1, 3, "status_조서영")  # lowercase!

        ws.cell(2, 1, "EVT001")
        ws.cell(2, 2, "FIXED")
        ws.cell(2, 3, "REPORTED")

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"Result: {result}")

        if "Script" in result and result["Script"]:
            script = result["Script"]
            print(f"유지윤: {script.get('유지윤', 'NOT FOUND')}")
            if script.get("유지윤", {}).get("fixed", 0) == 1:
                print("PASS: Lowercase header handled correctly")
            else:
                print("FAIL: Lowercase header NOT handled - this might be the bug!")
        else:
            print("FAIL: No Script data found - lowercase headers not detected!")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


def test_scenario_3_status_without_underscore():
    """
    Hypothesis: Header is 'STATUS유지윤' (no underscore) or 'STATUS 유지윤' (space).
    """
    print("\n" + "="*60)
    print("SCENARIO 3: STATUS without underscore")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "English Script"

        # No underscore or space instead
        ws.cell(1, 1, "EventName")
        ws.cell(1, 2, "STATUS유지윤")  # no underscore
        ws.cell(1, 3, "STATUS 조서영")  # space instead of underscore

        ws.cell(2, 1, "EVT001")
        ws.cell(2, 2, "FIXED")
        ws.cell(2, 3, "REPORTED")

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"Result: {result}")

        # These should NOT be found since we require STATUS_
        if "Script" in result and result["Script"]:
            print("UNEXPECTED: Found data with non-standard header format")
        else:
            print("EXPECTED: No data found - headers must be STATUS_{User}")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


def test_scenario_4_tester_status_prefix():
    """
    Hypothesis: Column is 'TESTER_STATUS_유지윤' not 'STATUS_유지윤'.
    This should be EXCLUDED (it's tester status, not manager status).
    """
    print("\n" + "="*60)
    print("SCENARIO 4: TESTER_STATUS_ prefix (should be excluded)")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "English Script"

        # TESTER_STATUS_ prefix - should be excluded
        ws.cell(1, 1, "EventName")
        ws.cell(1, 2, "TESTER_STATUS_유지윤")
        ws.cell(1, 3, "STATUS_조서영")  # This one should work

        ws.cell(2, 1, "EVT001")
        ws.cell(2, 2, "ISSUE")  # Tester status
        ws.cell(2, 3, "FIXED")  # Manager status

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"Result: {result}")

        if "Script" in result:
            script = result["Script"]
            if "유지윤" in script:
                print("UNEXPECTED: TESTER_STATUS_ was counted as manager status!")
            else:
                print("CORRECT: TESTER_STATUS_ excluded")
            if script.get("조서영", {}).get("fixed", 0) == 1:
                print("CORRECT: STATUS_ was counted")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


def test_scenario_5_real_user_structure():
    """
    Test with EXACT structure from user's log:
    - Sheet names: 'STATUS', 'English Script', 'Work Transform'
    - Columns: EventName, Text, Translation, STRINGID, COMMENT_유지윤, TESTER_STATUS_유지윤, STATUS_유지윤, ...
    """
    print("\n" + "="*60)
    print("SCENARIO 5: Exact user file structure")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        wb = Workbook()

        # STATUS sheet (should be skipped)
        ws_status = wb.active
        ws_status.title = "STATUS"
        ws_status.cell(1, 1, "This sheet should be skipped")

        # English Script sheet - exact structure from log
        ws = wb.create_sheet("English Script")
        # From log: rows=1778, cols=16, STATUS_ columns: ['유지윤', '조서영']
        headers = [
            "EventName", "Text", "Translation", "STRINGID",
            "COMMENT_유지윤", "TESTER_STATUS_유지윤", "STATUS_유지윤", "MANAGER_COMMENT_유지윤",
            "COMMENT_조서영", "TESTER_STATUS_조서영", "STATUS_조서영", "MANAGER_COMMENT_조서영",
            "Column13", "Column14", "Column15", "Column16"
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, h)

        # Add data rows with manager status
        # Row 2
        ws.cell(2, 1, "EVT001")
        ws.cell(2, 2, "Hello world")
        ws.cell(2, 3, "안녕 세상")
        ws.cell(2, 4, "12345")
        ws.cell(2, 5, "Typo here")  # COMMENT_유지윤
        ws.cell(2, 6, "ISSUE")      # TESTER_STATUS_유지윤
        ws.cell(2, 7, "FIXED")      # STATUS_유지윤 - MANAGER STATUS
        ws.cell(2, 8, "Fixed in build 5")  # MANAGER_COMMENT_유지윤

        # Row 3
        ws.cell(3, 1, "EVT002")
        ws.cell(3, 7, "REPORTED")   # STATUS_유지윤

        # Row 4 - 조서영 has status
        ws.cell(4, 1, "EVT003")
        ws.cell(4, 11, "CHECKING")  # STATUS_조서영

        # Row 5
        ws.cell(5, 1, "EVT004")
        ws.cell(5, 7, "NON-ISSUE")  # STATUS_유지윤
        ws.cell(5, 11, "NON ISSUE") # STATUS_조서영 (with space)

        # Work Transform sheet
        ws2 = wb.create_sheet("Work Transform")
        headers2 = [
            "EventName", "Text", "Translation", "STRINGID",
            "COMMENT_김정원", "TESTER_STATUS_김정원", "STATUS_김정원",
            "COMMENT_김찬용", "TESTER_STATUS_김찬용", "STATUS_김찬용",
            "COMMENT_황하연", "TESTER_STATUS_황하연", "STATUS_황하연",
            "Col14", "Col15", "Col16", "Col17", "Col18", "Col19", "Col20"
        ]
        for i, h in enumerate(headers2, 1):
            ws2.cell(1, i, h)

        # Add data
        ws2.cell(2, 1, "WRK001")
        ws2.cell(2, 7, "FIXED")      # STATUS_김정원
        ws2.cell(2, 10, "REPORTED")  # STATUS_김찬용
        ws2.cell(2, 13, "CHECKING")  # STATUS_황하연

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"\nResult categories: {list(result.keys())}")

        if "Script" in result:
            script = result["Script"]
            print(f"\nScript category users: {list(script.keys())}")
            for user, stats in script.items():
                print(f"  {user}: F={stats['fixed']} R={stats['reported']} C={stats['checking']} N={stats['nonissue']}")

            # Verify counts
            # 유지윤: FIXED(row2) + REPORTED(row3) + NON-ISSUE(row5) = F=1 R=1 N=1
            if script.get("유지윤", {}).get("fixed", 0) >= 1:
                print("\nPASS: 유지윤 FIXED counted correctly")
            else:
                print("\nFAIL: 유지윤 FIXED not counted!")

            # 조서영: CHECKING(row4) + NON ISSUE(row5) = C=1 N=1
            if script.get("조서영", {}).get("checking", 0) >= 1:
                print("PASS: 조서영 CHECKING counted correctly")
            else:
                print("FAIL: 조서영 CHECKING not counted!")

            # Check all expected users found
            expected_users = ["유지윤", "조서영", "김정원", "김찬용", "황하연"]
            for user in expected_users:
                if user in script:
                    print(f"PASS: {user} found in Script")
                else:
                    print(f"FAIL: {user} NOT found in Script!")
        else:
            print("\nFAIL: Script category not in results!")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


def test_scenario_6_all_empty_status():
    """
    Hypothesis: The STATUS_{User} columns exist but ALL cells are truly empty.
    This would mean manager hasn't filled in ANY status yet.
    """
    print("\n" + "="*60)
    print("SCENARIO 6: All STATUS cells empty (manager hasn't processed)")
    print("="*60)

    import config
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "English Script"

        # Headers exist
        ws.cell(1, 1, "EventName")
        ws.cell(1, 2, "COMMENT_유지윤")
        ws.cell(1, 3, "STATUS_유지윤")

        # Data rows - COMMENT exists but STATUS is empty
        ws.cell(2, 1, "EVT001")
        ws.cell(2, 2, "Found typo here")  # Tester wrote comment
        ws.cell(2, 3, None)  # Manager hasn't filled status yet

        ws.cell(3, 1, "EVT002")
        ws.cell(3, 2, "Another issue")
        ws.cell(3, 3, "")  # Empty string

        ws.cell(4, 1, "EVT003")
        ws.cell(4, 2, "Third issue")
        # Column 3 not even written - truly None

        wb.save(master_en / "Master_Script.xlsx")
        wb.close()

        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_en
        compiler_module.MASTER_FOLDER_CN = master_cn

        from core.compiler import collect_all_master_data
        (_, _, _, _, result) = collect_all_master_data(tester_mapping={})

        print(f"Result: {result}")

        if "Script" in result:
            script = result["Script"]
            user_stats = script.get("유지윤", {})
            total = user_stats.get("fixed", 0) + user_stats.get("reported", 0) + user_stats.get("checking", 0) + user_stats.get("nonissue", 0)

            if total == 0:
                print("EXPECTED: All stats are 0 because manager hasn't filled in STATUS")
                print("This is NOT a bug - the STATUS column is genuinely empty")
                print("\nThe REAL question: Should 'pending' issues (COMMENT exists, STATUS empty) be counted?")
            else:
                print(f"UNEXPECTED: Found {total} stats in empty STATUS column")
        else:
            print("Script not in results (expected - no STATUS values)")

    finally:
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_en
        compiler_module.MASTER_FOLDER_CN = original_cn
        shutil.rmtree(temp_dir)


if __name__ == "__main__":
    print("="*60)
    print("TESTING USER SCENARIOS")
    print("="*60)

    test_scenario_1_data_starts_after_row_11()
    test_scenario_2_lowercase_status_header()
    test_scenario_3_status_without_underscore()
    test_scenario_4_tester_status_prefix()
    test_scenario_5_real_user_structure()
    test_scenario_6_all_empty_status()

    print("\n" + "="*60)
    print("ALL SCENARIOS COMPLETE")
    print("="*60)
