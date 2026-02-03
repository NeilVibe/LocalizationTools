#!/usr/bin/env python3
"""
Create a realistic 10,000-row Script-type benchmark Excel fixture.

Generates benchmark_script_10k.xlsx for performance testing of:
- safe_load_workbook() loading speed
- find_column_by_header() header scanning
- Content-based matching with build_master_index()
- openpyxl styling object creation overhead

The file mimics a real Sequencer QA file with:
- 3 sheets: MainQuest_01, MainQuest_02, SubQuest_01
- Columns: KOR, Text, EventName, STATUS, MEMO
- 10,000 total rows with realistic distribution of STATUS values
- STATUS rows randomly distributed (not grouped)
"""

import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Realistic data pools
# ---------------------------------------------------------------------------

# Korean placeholder sentences (mix of lengths, uses real Hangul syllables)
KOREAN_SENTENCES = [
    "대화를 시작합니다.",
    "이 퀘스트를 완료하려면 마을로 돌아가세요.",
    "적을 처치하고 보상을 받으세요.",
    "전투가 시작됩니다. 준비하세요!",
    "당신의 도움이 필요합니다.",
    "이 지역은 위험합니다. 조심하세요.",
    "마법사의 탑으로 향하세요.",
    "감사합니다. 덕분에 마을이 평화로워졌습니다.",
    "여기서 잠시 기다려 주세요.",
    "무기를 강화할 수 있는 대장장이를 찾으세요.",
    "고대 유적에서 비밀 통로를 발견했습니다.",
    "임무를 완료하면 특별한 보상이 주어집니다.",
    "동굴 깊숙한 곳에 보물이 숨겨져 있습니다.",
    "이 편지를 마을 이장에게 전달해 주세요.",
    "전설의 검을 찾기 위한 모험이 시작됩니다.",
    "숲속에서 이상한 소리가 들립니다.",
    "왕국의 평화를 위해 싸워야 합니다.",
    "포션을 사용하여 체력을 회복하세요.",
    "다음 목표 지점까지 안전하게 이동하세요.",
    "보스 몬스터가 곧 나타날 것입니다.",
    "NPC와 대화하여 정보를 수집하세요.",
    "퀘스트 아이템을 모두 수집했습니다.",
    "파티원들과 함께 던전에 입장하세요.",
    "레벨이 부족합니다. 더 성장한 후 도전하세요.",
    "스킬을 업그레이드하여 전투력을 높이세요.",
    "마을 광장에서 축제가 열리고 있습니다.",
    "이 문을 열려면 특별한 열쇠가 필요합니다.",
    "적의 약점을 파악하여 공격하세요.",
    "동맹군이 지원을 보내왔습니다.",
    "최종 보스와의 전투가 시작됩니다!",
]

# English translation sentences (matching tone and variety)
ENGLISH_SENTENCES = [
    "The conversation begins.",
    "Return to the village to complete this quest.",
    "Defeat the enemies and claim your reward.",
    "The battle is about to begin. Prepare yourself!",
    "We need your help.",
    "This area is dangerous. Be careful.",
    "Head to the Wizard's Tower.",
    "Thank you. The village is at peace once again.",
    "Please wait here for a moment.",
    "Find the blacksmith who can upgrade your weapon.",
    "A secret passage was discovered in the ancient ruins.",
    "You will receive a special reward upon completing the mission.",
    "Treasure is hidden deep within the cave.",
    "Please deliver this letter to the village elder.",
    "The adventure to find the legendary sword begins now.",
    "Strange sounds can be heard from within the forest.",
    "You must fight for the peace of the kingdom.",
    "Use a potion to restore your health.",
    "Travel safely to the next objective point.",
    "The boss monster will appear soon.",
    "Talk to the NPC to gather information.",
    "All quest items have been collected.",
    "Enter the dungeon together with your party members.",
    "Your level is too low. Come back after growing stronger.",
    "Upgrade your skills to increase combat power.",
    "A festival is being held in the village square.",
    "A special key is required to open this door.",
    "Identify the enemy's weakness and attack.",
    "Allied reinforcements have arrived.",
    "The final battle with the boss begins!",
]

# Scene name prefixes for realistic EventNames
SCENE_PREFIXES = [
    "MainQuest_Ch01", "MainQuest_Ch02", "MainQuest_Ch03",
    "MainQuest_Prologue", "MainQuest_Epilogue",
    "SubQuest_Village", "SubQuest_Forest", "SubQuest_Dungeon",
    "SubQuest_Castle", "SubQuest_Harbor",
    "CutScene_Battle", "CutScene_Ending", "CutScene_Opening",
    "Event_Festival", "Event_Ambush", "Event_Rescue",
    "Tutorial_Basic", "Tutorial_Combat", "Tutorial_Craft",
    "Boss_Dragon", "Boss_Demon", "Boss_Lich",
]

# Realistic tester comments for ISSUE rows
ISSUE_COMMENTS = [
    "Translation doesn't match context",
    "Missing honorific in formal dialogue",
    "Incorrect weapon name translation",
    "Gender mismatch in pronoun",
    "Tone inconsistent with character personality",
    "Placeholder text left untranslated",
    "Text overflow in dialogue box",
    "Wrong NPC name used",
    "Skill name doesn't match glossary",
    "Awkward phrasing - needs natural English",
    "Location name inconsistency with map",
    "Item name doesn't match inventory",
    "Translation is too literal from Korean",
    "Cultural reference lost in translation",
    "Dialogue tag mismatch with speaker",
    "Typo in translated text",
    "Missing punctuation at end of sentence",
    "Wrong tense used in conversation",
    "Abbreviation unclear for Western audience",
    "Line break position creates reading issue",
]

# Realistic tester comments for BLOCKED rows
BLOCKED_COMMENTS = [
    "Cannot access this scene - quest prerequisite missing",
    "Game crashes before dialogue appears",
    "Cutscene skips this line too fast to verify",
    "NPC does not appear at expected location",
    "Quest chain broken - cannot progress to this point",
    "Audio plays but subtitle text is missing",
    "Scene requires co-op partner to trigger",
    "Build version too old - scene not available",
    "Loading screen freeze before this sequence",
    "Dialogue appears behind UI element - cannot read",
]


def generate_event_name(scene_prefix: str, index: int) -> str:
    """Generate a realistic EventName like EVT_MainQuest_Ch01_047."""
    return f"EVT_{scene_prefix}_{index:03d}"


def create_benchmark_fixture():
    """Create the 10,000-row benchmark fixture Excel file."""
    output_path = Path(__file__).parent / "fixtures" / "benchmark_script_10k.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet configuration: name -> approximate row count
    # Total must be 10,000
    sheet_config = {
        "MainQuest_01": 4500,
        "MainQuest_02": 3500,
        "SubQuest_01": 2000,
    }

    # STATUS distribution across all 10,000 rows
    # 200 ISSUE, 100 NO ISSUE, 50 BLOCKED, 30 KOREAN, 20 NON-ISSUE, 9600 empty
    status_pool = (
        ["ISSUE"] * 200
        + ["NO ISSUE"] * 100
        + ["BLOCKED"] * 50
        + ["KOREAN"] * 30
        + ["NON-ISSUE"] * 20
        + [""] * 9600
    )
    random.shuffle(status_pool)

    # Styling
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ["KOR", "Text", "EventName", "STATUS", "MEMO"]

    global_row_idx = 0  # Tracks position in status_pool
    total_rows_written = 0

    for sheet_name, row_count in sheet_config.items():
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Pick scene prefixes for this sheet (rotate through them)
        sheet_scenes = random.sample(
            SCENE_PREFIXES, min(len(SCENE_PREFIXES), 8)
        )

        for i in range(row_count):
            row_num = i + 2  # Excel row (1-based, skip header)
            pool_idx = global_row_idx + i

            # Deterministic but varied content selection
            kor_text = KOREAN_SENTENCES[i % len(KOREAN_SENTENCES)]
            eng_text = ENGLISH_SENTENCES[i % len(ENGLISH_SENTENCES)]

            # EventName: rotate through scenes, increment counter
            scene = sheet_scenes[i % len(sheet_scenes)]
            event_name = generate_event_name(scene, (i // len(sheet_scenes)) + 1)

            # STATUS from the pre-shuffled pool
            status = status_pool[pool_idx] if pool_idx < len(status_pool) else ""

            # MEMO: only for ISSUE and BLOCKED rows
            memo = ""
            if status == "ISSUE":
                memo = ISSUE_COMMENTS[i % len(ISSUE_COMMENTS)]
            elif status == "BLOCKED":
                memo = BLOCKED_COMMENTS[i % len(BLOCKED_COMMENTS)]

            # Write cells
            ws.cell(row=row_num, column=1, value=kor_text)
            ws.cell(row=row_num, column=2, value=eng_text)
            ws.cell(row=row_num, column=3, value=event_name)
            if status:
                ws.cell(row=row_num, column=4, value=status)
            if memo:
                ws.cell(row=row_num, column=5, value=memo)

        global_row_idx += row_count
        total_rows_written += row_count

        # Set column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 45

    # Save
    wb.save(output_path)

    # Print summary
    print(f"Created: {output_path}")
    print(f"  File size: {output_path.stat().st_size / 1024:.1f} KB")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Total data rows: {total_rows_written}")
    print()

    # Count actual status distribution
    status_counts = {}
    for s in status_pool:
        label = s if s else "(empty)"
        status_counts[label] = status_counts.get(label, 0) + 1

    print("  STATUS distribution:")
    for label in ["ISSUE", "NO ISSUE", "BLOCKED", "KOREAN", "NON-ISSUE", "(empty)"]:
        count = status_counts.get(label, 0)
        print(f"    {label:>12}: {count:,}")

    print()
    for sheet_name, row_count in sheet_config.items():
        print(f"  {sheet_name}: {row_count:,} rows")

    return output_path


if __name__ == "__main__":
    create_benchmark_fixture()
