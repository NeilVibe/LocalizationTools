#!/usr/bin/env python3
"""
Test script to diagnose Script category manager status collection.

Creates a mock Master_Script.xlsx with:
- STATUS_{username} columns with FIXED/NON-ISSUE values
- COMMENT_{username} columns with tester comments (some empty, some filled)

Then runs collect_manager_status() to see what gets captured.
"""

import sys
import os
from pathlib import Path
import tempfile
import shutil

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_mock_master_script(folder: Path) -> Path:
    """Create a mock Master_Script.xlsx with manager status data."""
    master_path = folder / "Master_Script.xlsx"

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create "English Script" sheet (typical Script sheet name)
    ws = wb.create_sheet("English Script")

    # Headers - simulate what a compiled master would have
    headers = [
        "EventName",           # Col 1
        "Text",                # Col 2
        "Translation",         # Col 3
        "STRINGID",            # Col 4
        "COMMENT_TestUser",    # Col 5 - Tester's comment
        "TESTER_STATUS_TestUser",  # Col 6 - Tester's status (ISSUE/NO ISSUE)
        "STATUS_TestUser",     # Col 7 - Manager's status (FIXED/NON-ISSUE)
        "MANAGER_COMMENT_TestUser",  # Col 8 - Manager's comment
        "COMMENT_OtherUser",   # Col 9 - Another user
        "STATUS_OtherUser",    # Col 10 - Another user's manager status
    ]

    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # Data rows - simulate various scenarios
    test_data = [
        # Row 2: Has STATUS and COMMENT - should be STORED
        {
            "EventName": "EVT_001",
            "Text": "Hello world",
            "Translation": "안녕하세요",
            "STRINGID": "10001",
            "COMMENT_TestUser": "Typo in translation\n---\nstringid:\n10001",
            "TESTER_STATUS_TestUser": "ISSUE",
            "STATUS_TestUser": "FIXED",
            "MANAGER_COMMENT_TestUser": "Fixed in build 5",
            "COMMENT_OtherUser": None,
            "STATUS_OtherUser": None,
        },
        # Row 3: Has STATUS but NO COMMENT - should be SKIPPED
        {
            "EventName": "EVT_002",
            "Text": "Goodbye",
            "Translation": "안녕히 가세요",
            "STRINGID": "10002",
            "COMMENT_TestUser": None,  # NO COMMENT!
            "TESTER_STATUS_TestUser": None,
            "STATUS_TestUser": "NON-ISSUE",  # Manager set status without tester comment
            "MANAGER_COMMENT_TestUser": "Looks fine",
            "COMMENT_OtherUser": None,
            "STATUS_OtherUser": None,
        },
        # Row 4: Has COMMENT but no STATUS - should NOT trigger (no manager status)
        {
            "EventName": "EVT_003",
            "Text": "Test line",
            "Translation": "테스트",
            "STRINGID": "10003",
            "COMMENT_TestUser": "Minor issue",
            "TESTER_STATUS_TestUser": "ISSUE",
            "STATUS_TestUser": None,  # Manager hasn't responded yet
            "MANAGER_COMMENT_TestUser": None,
            "COMMENT_OtherUser": None,
            "STATUS_OtherUser": None,
        },
        # Row 5: OtherUser has STATUS and COMMENT
        {
            "EventName": "EVT_004",
            "Text": "Another line",
            "Translation": "다른 줄",
            "STRINGID": "10004",
            "COMMENT_TestUser": None,
            "TESTER_STATUS_TestUser": None,
            "STATUS_TestUser": None,
            "MANAGER_COMMENT_TestUser": None,
            "COMMENT_OtherUser": "OtherUser found issue",
            "STATUS_OtherUser": "REPORTED",
        },
        # Row 6: STATUS with EMPTY string comment (edge case)
        {
            "EventName": "EVT_005",
            "Text": "Edge case",
            "Translation": "엣지 케이스",
            "STRINGID": "10005",
            "COMMENT_TestUser": "",  # Empty string, not None
            "TESTER_STATUS_TestUser": None,
            "STATUS_TestUser": "CHECKING",
            "MANAGER_COMMENT_TestUser": None,
            "COMMENT_OtherUser": None,
            "STATUS_OtherUser": None,
        },
    ]

    # Write data
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header)
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Save
    wb.save(master_path)
    print(f"Created mock: {master_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Rows: {ws.max_row}")
    print(f"  Headers: {headers}")

    return master_path


def run_collection_test():
    """Run the collection function against mock data."""
    # Create temp directory structure
    temp_dir = Path(tempfile.mkdtemp(prefix="qacompiler_test_"))
    print(f"\nTest directory: {temp_dir}")

    try:
        # Create mock folder structure
        master_folder_en = temp_dir / "Masterfolder_EN"
        master_folder_cn = temp_dir / "Masterfolder_CN"
        master_folder_en.mkdir()
        master_folder_cn.mkdir()

        # Create mock Master_Script.xlsx
        create_mock_master_script(master_folder_en)

        # Temporarily override config paths
        import config
        original_master_en = config.MASTER_FOLDER_EN
        original_master_cn = config.MASTER_FOLDER_CN

        config.MASTER_FOLDER_EN = master_folder_en
        config.MASTER_FOLDER_CN = master_folder_cn

        # Also update compiler module's imported references
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = master_folder_en
        compiler_module.MASTER_FOLDER_CN = master_folder_cn

        # Clear any existing debug log
        debug_log = Path(__file__).parent.parent / "SCRIPT_DEBUG.log"
        if debug_log.exists():
            debug_log.unlink()

        # Run collection
        print("\n" + "=" * 60)
        print("Running collect_all_master_data()...")
        print("=" * 60)

        from core.compiler import collect_all_master_data
        (fixed_screenshots_en, fixed_screenshots_cn, manager_stats) = collect_all_master_data(tester_mapping={})

        # Print results
        print("\n" + "=" * 60)
        print("RESULTS")
        print("=" * 60)

        print(f"\nFixed screenshots EN: {fixed_screenshots_en}")
        print(f"Fixed screenshots CN: {fixed_screenshots_cn}")
        print(f"Manager stats categories: {list(manager_stats.keys())}")
        for category, users in manager_stats.items():
            print(f"\nCategory: {category}")
            for user, stats in users.items():
                print(f"  {user}: {stats}")

        # Check debug log
        print("\n" + "=" * 60)
        print("DEBUG LOG CONTENT")
        print("=" * 60)

        if debug_log.exists():
            print(debug_log.read_text())
        else:
            print("No debug log generated!")

    finally:
        # Restore config and compiler module paths
        config.MASTER_FOLDER_EN = original_master_en
        config.MASTER_FOLDER_CN = original_master_cn
        import core.compiler as compiler_module
        compiler_module.MASTER_FOLDER_EN = original_master_en
        compiler_module.MASTER_FOLDER_CN = original_master_cn
        # Cleanup
        shutil.rmtree(temp_dir)
        print(f"\nCleaned up: {temp_dir}")


if __name__ == "__main__":
    # Change to QACompilerNEW directory for proper imports
    os.chdir(Path(__file__).parent.parent)
    run_collection_test()
