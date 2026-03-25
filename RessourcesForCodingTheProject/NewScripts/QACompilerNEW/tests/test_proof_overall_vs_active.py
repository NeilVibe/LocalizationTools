"""
PROOF TEST: Overall QA Data vs Active QA Data vs Active Manager Data
=====================================================================
Demonstrates and verifies the COMPLETE data flow for the two-tier issue system.

Scenario:
- Tester "Alice" checked 20 rows in Quest QA file
  - 10 marked ISSUE, 5 marked NO ISSUE, 3 BLOCKED, 2 KOREAN
- Game data changed: 3 of those ISSUE StringIDs were REMOVED from the game
- Masterfile was recompiled: only 7 of Alice's 10 issues still exist
- Manager reviewed the masterfile:
  - 2 marked FIXED
  - 1 marked REPORTED
  - 1 marked CHECKING (still pending)
  - 1 marked NON-ISSUE
  - 2 have empty STATUS (untouched = pending)

EXPECTED RESULTS:
  OVERALL (from QA file):
    issues = 10  (all issues Alice ever reported)

  ACTIVE (from masterfile):
    active_issues = 7  (only issues that exist in current masterfile)
    pending = 3        (2 empty + 1 CHECKING)
    fixed = 2
    reported = 1
    checking = 1       (subset of pending)
    nonissue = 1

  STALE = overall - active = 10 - 7 = 3
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
from openpyxl import Workbook


def _make_qa_file(folder: Path, username: str, category: str):
    """Create Alice's QA file with 20 rows: 10 ISSUE, 5 NO ISSUE, 3 BLOCKED, 2 KOREAN."""
    wb = Workbook()
    ws = wb.active
    ws.title = category
    ws.cell(1, 1, "STRINGID")
    ws.cell(1, 2, "TEXT")
    ws.cell(1, 3, "STATUS")

    row = 2
    # 10 ISSUE rows (StringIDs 1001-1010)
    for i in range(10):
        ws.cell(row, 1, str(1001 + i))
        ws.cell(row, 2, f"Issue text {i}")
        ws.cell(row, 3, "ISSUE")
        row += 1
    # 5 NO ISSUE rows
    for i in range(5):
        ws.cell(row, 1, str(2001 + i))
        ws.cell(row, 2, f"OK text {i}")
        ws.cell(row, 3, "NO ISSUE")
        row += 1
    # 3 BLOCKED rows
    for i in range(3):
        ws.cell(row, 1, str(3001 + i))
        ws.cell(row, 2, f"Blocked text {i}")
        ws.cell(row, 3, "BLOCKED")
        row += 1
    # 2 KOREAN rows
    for i in range(2):
        ws.cell(row, 1, str(4001 + i))
        ws.cell(row, 2, f"Korean text {i}")
        ws.cell(row, 3, "KOREAN")
        row += 1

    qa_dir = folder / f"{username}_{category}"
    qa_dir.mkdir(parents=True, exist_ok=True)
    path = qa_dir / f"{username}_{category}.xlsx"
    wb.save(path)
    wb.close()
    return path


def _make_masterfile(folder: Path, category: str):
    """Create Master_Quest.xlsx with only 7 of Alice's 10 issues still present.

    3 StringIDs (1008, 1009, 1010) were removed from game data.
    Manager has reviewed the remaining 7:
      1001: ISSUE -> FIXED
      1002: ISSUE -> FIXED
      1003: ISSUE -> REPORTED
      1004: ISSUE -> CHECKING (still pending)
      1005: ISSUE -> NON-ISSUE
      1006: ISSUE -> (empty = pending)
      1007: ISSUE -> (empty = pending)
    """
    wb = Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(category)

    # Headers
    ws.cell(1, 1, "STRINGID")
    ws.cell(1, 2, "TEXT")
    ws.cell(1, 3, "TESTER_STATUS_Alice")
    ws.cell(1, 4, "STATUS_Alice")
    ws.cell(1, 5, "COMMENT_Alice")

    # 7 surviving ISSUE rows with manager responses
    data = [
        ("1001", "Issue text 0", "ISSUE", "FIXED"),
        ("1002", "Issue text 1", "ISSUE", "FIXED"),
        ("1003", "Issue text 2", "ISSUE", "REPORTED"),
        ("1004", "Issue text 3", "ISSUE", "CHECKING"),
        ("1005", "Issue text 4", "ISSUE", "NON-ISSUE"),
        ("1006", "Issue text 5", "ISSUE", ""),           # empty = pending
        ("1007", "Issue text 6", "ISSUE", ""),           # empty = pending
    ]
    for row_idx, (sid, text, tstatus, mstatus) in enumerate(data, 2):
        ws.cell(row_idx, 1, sid)
        ws.cell(row_idx, 2, text)
        ws.cell(row_idx, 3, tstatus)
        ws.cell(row_idx, 4, mstatus)

    # Also add some NO ISSUE rows (these should NOT count as active issues)
    for i, sid in enumerate(["2001", "2002", "2003"], 9):
        ws.cell(sid_row := row_idx + 1 + i, 1, sid)
        ws.cell(sid_row, 2, f"OK text {i}")
        ws.cell(sid_row, 3, "NO ISSUE")
        ws.cell(sid_row, 4, "")

    path = folder / f"Master_{category}.xlsx"
    wb.save(path)
    wb.close()
    return path


class TestProofOverallVsActive:
    """PROOF: Overall QA data and Active masterfile data are correctly separated."""

    def test_full_scenario(self, tmp_path):
        """The complete scenario from the docstring."""
        en_folder = tmp_path / "Masterfolder_EN"
        cn_folder = tmp_path / "Masterfolder_CN"
        en_folder.mkdir()
        cn_folder.mkdir()

        # Create QA file (simulates what count_sheet_stats reads)
        qa_dir = tmp_path / "QAfolder"
        qa_dir.mkdir()
        _make_qa_file(qa_dir, "Alice", "Quest")

        # Create masterfile (only 7 of 10 issues survive)
        _make_masterfile(en_folder, "Quest")

        # ---- OVERALL: from QA file ----
        from core.tracker_update import count_sheet_stats
        from openpyxl import load_workbook

        qa_path = qa_dir / "Alice_Quest" / "Alice_Quest.xlsx"
        qa_wb = load_workbook(qa_path)
        qa_ws = qa_wb[qa_wb.sheetnames[0]]
        overall = count_sheet_stats(qa_ws, "Quest", is_english=True, sheet_name="Quest")
        qa_wb.close()

        # Verify OVERALL stats
        assert overall["total"] == 20, f"Total rows: {overall['total']}"
        assert overall["issue"] == 10, f"OVERALL issues: {overall['issue']}"
        assert overall["no_issue"] == 5, f"NO ISSUE: {overall['no_issue']}"
        assert overall["blocked"] == 3, f"BLOCKED: {overall['blocked']}"
        assert overall["korean"] == 2, f"KOREAN: {overall['korean']}"

        # ---- ACTIVE: from masterfile ----
        from tracker.masterfile_pending import build_pending_from_masterfiles

        active = build_pending_from_masterfiles(en_folder, cn_folder)

        # Verify ACTIVE stats
        assert "Alice" in active, f"Alice not in active: {active.keys()}"
        assert "Quest" in active["Alice"], f"Quest not in Alice's active: {active['Alice'].keys()}"

        a = active["Alice"]["Quest"]
        assert a["active_issues"] == 7, f"ACTIVE issues: {a['active_issues']} (expected 7)"
        assert a["pending"] == 3, f"PENDING: {a['pending']} (expected 3: 2 empty + 1 CHECKING)"
        assert a["fixed"] == 2, f"FIXED: {a['fixed']} (expected 2)"
        assert a["reported"] == 1, f"REPORTED: {a['reported']} (expected 1)"
        assert a["checking"] == 1, f"CHECKING: {a['checking']} (expected 1)"
        assert a["nonissue"] == 1, f"NON-ISSUE: {a['nonissue']} (expected 1)"

        # ---- STALE: derived ----
        stale = overall["issue"] - a["active_issues"]
        assert stale == 3, f"STALE issues: {stale} (expected 3 — removed from game)"

        # ---- INVARIANTS ----
        # Active issues = pending_unresolved + fixed + reported + nonissue
        # (CHECKING is subset of pending, not separate in this sum)
        resolved = a["fixed"] + a["reported"] + a["nonissue"]
        pending_unresolved = a["active_issues"] - resolved
        assert pending_unresolved == a["pending"], (
            f"Pending mismatch: {pending_unresolved} != {a['pending']}"
        )

        # CHECKING is a subset of pending
        assert a["checking"] <= a["pending"], (
            f"CHECKING ({a['checking']}) > PENDING ({a['pending']})"
        )

    def test_two_testers_two_categories(self, tmp_path):
        """Two testers, two categories, fully independent."""
        en_folder = tmp_path / "Masterfolder_EN"
        cn_folder = tmp_path / "Masterfolder_CN"
        en_folder.mkdir()
        cn_folder.mkdir()

        # Master_Quest: Alice has 3 issues (1 fixed, 2 pending)
        wb = Workbook()
        del wb["Sheet"]
        ws = wb.create_sheet("Quest")
        ws.cell(1, 1, "STRINGID"); ws.cell(1, 2, "TEXT")
        ws.cell(1, 3, "TESTER_STATUS_Alice"); ws.cell(1, 4, "STATUS_Alice")
        ws.cell(2, 1, "Q1"); ws.cell(2, 2, "T1"); ws.cell(2, 3, "ISSUE"); ws.cell(2, 4, "FIXED")
        ws.cell(3, 1, "Q2"); ws.cell(3, 2, "T2"); ws.cell(3, 3, "ISSUE"); ws.cell(3, 4, "")
        ws.cell(4, 1, "Q3"); ws.cell(4, 2, "T3"); ws.cell(4, 3, "ISSUE"); ws.cell(4, 4, "")
        wb.save(en_folder / "Master_Quest.xlsx")
        wb.close()

        # Master_Item: Bob has 4 issues (2 fixed, 1 reported, 1 pending)
        wb = Workbook()
        del wb["Sheet"]
        ws = wb.create_sheet("Item")
        ws.cell(1, 1, "STRINGID"); ws.cell(1, 2, "TEXT")
        ws.cell(1, 3, "TESTER_STATUS_Bob"); ws.cell(1, 4, "STATUS_Bob")
        ws.cell(2, 1, "I1"); ws.cell(2, 2, "T1"); ws.cell(2, 3, "ISSUE"); ws.cell(2, 4, "FIXED")
        ws.cell(3, 1, "I2"); ws.cell(3, 2, "T2"); ws.cell(3, 3, "ISSUE"); ws.cell(3, 4, "FIXED")
        ws.cell(4, 1, "I3"); ws.cell(4, 2, "T3"); ws.cell(4, 3, "ISSUE"); ws.cell(4, 4, "REPORTED")
        ws.cell(5, 1, "I4"); ws.cell(5, 2, "T4"); ws.cell(5, 3, "ISSUE"); ws.cell(5, 4, "")
        wb.save(en_folder / "Master_Item.xlsx")
        wb.close()

        active = build_pending_from_masterfiles(en_folder, cn_folder)

        # Alice: Quest only
        assert active["Alice"]["Quest"]["active_issues"] == 3
        assert active["Alice"]["Quest"]["fixed"] == 1
        assert active["Alice"]["Quest"]["pending"] == 2
        assert "Item" not in active["Alice"]

        # Bob: Item only
        assert active["Bob"]["Item"]["active_issues"] == 4
        assert active["Bob"]["Item"]["fixed"] == 2
        assert active["Bob"]["Item"]["reported"] == 1
        assert active["Bob"]["Item"]["pending"] == 1
        assert "Quest" not in active["Bob"]

    def test_all_resolved_zero_pending(self, tmp_path):
        """When manager resolved everything, pending = 0."""
        en_folder = tmp_path / "Masterfolder_EN"
        cn_folder = tmp_path / "Masterfolder_CN"
        en_folder.mkdir()
        cn_folder.mkdir()

        wb = Workbook()
        del wb["Sheet"]
        ws = wb.create_sheet("Character")
        ws.cell(1, 1, "STRINGID"); ws.cell(1, 2, "TEXT")
        ws.cell(1, 3, "TESTER_STATUS_Carol"); ws.cell(1, 4, "STATUS_Carol")
        ws.cell(2, 1, "C1"); ws.cell(2, 2, "T1"); ws.cell(2, 3, "ISSUE"); ws.cell(2, 4, "FIXED")
        ws.cell(3, 1, "C2"); ws.cell(3, 2, "T2"); ws.cell(3, 3, "ISSUE"); ws.cell(3, 4, "REPORTED")
        ws.cell(4, 1, "C3"); ws.cell(4, 2, "T3"); ws.cell(4, 3, "ISSUE"); ws.cell(4, 4, "NON-ISSUE")
        wb.save(en_folder / "Master_Character.xlsx")
        wb.close()

        active = build_pending_from_masterfiles(en_folder, cn_folder)

        c = active["Carol"]["Character"]
        assert c["active_issues"] == 3
        assert c["pending"] == 0, f"Should be 0 pending, got {c['pending']}"
        assert c["fixed"] == 1
        assert c["reported"] == 1
        assert c["nonissue"] == 1


# Import here so test collection works even if module has issues
from tracker.masterfile_pending import build_pending_from_masterfiles
