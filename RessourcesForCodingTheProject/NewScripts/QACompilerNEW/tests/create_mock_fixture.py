#!/usr/bin/env python3
"""
Create mock Excel fixture for testing Script-type preprocessing.

This script creates a mock_script_qa.xlsx file with:
- Sheet named "Sequencer" or "Dialog"
- Columns: EventName, Text, Translation, STATUS, MEMO
- Mix of rows with different STATUS values
"""

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_mock_fixture():
    """Create the mock fixture Excel file."""
    output_path = Path(__file__).parent / "fixtures" / "mock_script_qa.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create "Sequencer" sheet
    ws_seq = wb.create_sheet("Sequencer")

    # Header row with styling
    headers = ["EventName", "Text", "Translation", "STATUS", "MEMO"]
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws_seq.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Data rows with various STATUS values
    test_rows = [
        # Row 2: ISSUE
        ("SEQ_001", "Hello there!", "Translation 1", "ISSUE", "Needs fix"),
        # Row 3: NON-ISSUE
        ("SEQ_002", "Goodbye!", "Translation 2", "NON-ISSUE", "Looks OK"),
        # Row 4: Empty STATUS (should be skipped)
        ("SEQ_003", "Skip me", "Translation 3", "", ""),
        # Row 5: None STATUS (should be skipped)
        ("SEQ_004", "Also skip me", "Translation 4", None, ""),
        # Row 6: ISSUE
        ("SEQ_005", "Another issue", "Translation 5", "ISSUE", "Check this"),
        # Row 7: BLOCKED
        ("SEQ_006", "Blocked item", "Translation 6", "BLOCKED", "Waiting"),
        # Row 8-49: Empty STATUS (should be skipped)
    ]

    # Add more empty rows to simulate large file
    for i in range(8, 50):
        test_rows.append((f"SEQ_{i:03d}", f"Text line {i}", f"Trans {i}", "", ""))

    # Row 50: ISSUE (test sparse rows)
    test_rows.append(("SEQ_050", "Sparse issue at row 50", "Translation 50", "ISSUE", "Deep in file"))

    # More empty rows
    for i in range(51, 100):
        test_rows.append((f"SEQ_{i:03d}", f"Text line {i}", f"Trans {i}", None, ""))

    # Row 100: NON-ISSUE (test another sparse row)
    test_rows.append(("SEQ_100", "Sparse non-issue at row 100", "Translation 100", "NON-ISSUE", "Very deep"))

    # Write data rows
    for row_idx, row_data in enumerate(test_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if value is not None:  # Only write non-None values
                ws_seq.cell(row=row_idx, column=col_idx, value=value)

    # Create "Dialog" sheet with similar structure
    ws_dialog = wb.create_sheet("Dialog")

    # Copy headers
    for col, header in enumerate(headers, 1):
        cell = ws_dialog.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Add some dialog rows
    dialog_rows = [
        ("DLG_001", "Dialog line 1", "Dialog trans 1", "ISSUE", "Dialog issue"),
        ("DLG_002", "Dialog line 2", "Dialog trans 2", "", ""),
        ("DLG_003", "Dialog line 3", "Dialog trans 3", "NON-ISSUE", "OK"),
        ("DLG_004", "Dialog line 4", "Dialog trans 4", "KOREAN", "Korean text"),
    ]

    for row_idx, row_data in enumerate(dialog_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if value is not None:
                ws_dialog.cell(row=row_idx, column=col_idx, value=value)

    # Set column widths
    for ws in [ws_seq, ws_dialog]:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20

    # Save
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Sequencer rows: {ws_seq.max_row}")
    print(f"  Dialog rows: {ws_dialog.max_row}")

    # Count expected STATUS rows
    seq_with_status = sum(1 for r in test_rows if r[3] and str(r[3]).strip())
    print(f"  Sequencer rows with STATUS: {seq_with_status}")
    print(f"  Dialog rows with STATUS: 3")  # ISSUE, NON-ISSUE, KOREAN

    return output_path


if __name__ == "__main__":
    create_mock_fixture()
