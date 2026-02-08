"""
Tests for replicate_duplicate_row_data() from core/excel_ops.py
===============================================================

Tests the duplicate-row post-processing function that:
- Groups master rows by content key (category-dependent)
- For each group of 2+ identical rows, finds first non-empty value per user column
- Copies that value to all empty cells in the group
- Returns count of cells filled

All tests use in-memory openpyxl workbooks (no disk I/O).
"""

import sys
from pathlib import Path

# Ensure project root is on sys.path so core/ and config imports resolve
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
from openpyxl import Workbook
from core.excel_ops import replicate_duplicate_row_data


# =============================================================================
# HELPERS
# =============================================================================

def _make_standard_wb(rows, extra_headers=None):
    """Create a workbook with standard category headers + data rows.

    Default headers:
        STRINGID | TRANSLATION | COMMENT_alice | STATUS_alice
        | MANAGER_COMMENT_alice | SCREENSHOT_alice | TESTER_STATUS_alice

    Args:
        rows: List of dicts with keys matching header names. Missing keys
              leave the cell empty (None).
        extra_headers: Optional list of additional header names appended
                       after the defaults.

    Returns:
        openpyxl.Workbook with a single sheet named "Sheet1".
    """
    headers = [
        "STRINGID", "TRANSLATION",
        "COMMENT_alice", "STATUS_alice",
        "MANAGER_COMMENT_alice", "SCREENSHOT_alice",
        "TESTER_STATUS_alice",
    ]
    if extra_headers:
        headers.extend(extra_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers in row 1
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Write data rows starting at row 2
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            val = row_data.get(h)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    return wb


def _make_item_wb(rows, extra_headers=None):
    """Create a workbook with Item category headers.

    Headers:
        STRINGID | ITEMNAME(ENG) | ITEMDESC(ENG) | COMMENT_alice
        | STATUS_alice | MANAGER_COMMENT_alice
    """
    headers = [
        "STRINGID", "ITEMNAME(ENG)", "ITEMDESC(ENG)",
        "COMMENT_alice", "STATUS_alice", "MANAGER_COMMENT_alice",
    ]
    if extra_headers:
        headers.extend(extra_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            val = row_data.get(h)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    return wb


def _make_script_wb(rows, extra_headers=None):
    """Create a workbook with Script (Sequencer/Dialog) category headers.

    Headers:
        TEXT | EVENTNAME | COMMENT_alice | STATUS_alice | MANAGER_COMMENT_alice
    """
    headers = [
        "TEXT", "EVENTNAME",
        "COMMENT_alice", "STATUS_alice", "MANAGER_COMMENT_alice",
    ]
    if extra_headers:
        headers.extend(extra_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            val = row_data.get(h)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    return wb


def _cell_val(wb, sheet, row, header):
    """Read a cell value by header name from the workbook."""
    ws = wb[sheet]
    # Find column by header
    for col in range(1, (ws.max_column or 0) + 1):
        if ws.cell(1, col).value == header:
            return ws.cell(row, col).value
    return None


# =============================================================================
# TEST 1: Standard duplicates replicated
# =============================================================================

class TestStandardDuplicatesReplicated:
    """3 rows with same STRINGID+Translation. Row 1 has COMMENT_alice and
    STATUS_alice filled. Rows 2-3 empty. After post-process all 3 have
    the same data."""

    def test_comment_and_status_replicated(self):
        rows = [
            {"STRINGID": "1001", "TRANSLATION": "Hello World",
             "COMMENT_alice": "Looks good", "STATUS_alice": "ISSUE"},
            {"STRINGID": "1001", "TRANSLATION": "Hello World"},
            {"STRINGID": "1001", "TRANSLATION": "Hello World"},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # Rows 3 and 4 (Excel rows) should now have the replicated data
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Looks good"
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_alice") == "Looks good"
        assert _cell_val(wb, "Sheet1", 3, "STATUS_alice") == "ISSUE"
        assert _cell_val(wb, "Sheet1", 4, "STATUS_alice") == "ISSUE"

        # Original row 2 untouched
        assert _cell_val(wb, "Sheet1", 2, "COMMENT_alice") == "Looks good"
        assert _cell_val(wb, "Sheet1", 2, "STATUS_alice") == "ISSUE"

        # 2 empty cells per column * 2 columns = 4 cells filled
        assert filled == 4


# =============================================================================
# TEST 2: Manager data replicated
# =============================================================================

class TestManagerDataReplicated:
    """3 duplicate rows. Row 1 has COMMENT_alice, MANAGER_COMMENT_alice,
    and STATUS_alice (manager status like 'FIXED'). Rows 2-3 empty.
    ALL columns get replicated."""

    def test_all_user_columns_replicated(self):
        rows = [
            {"STRINGID": "2001", "TRANSLATION": "Good morning",
             "COMMENT_alice": "Fixed typo",
             "STATUS_alice": "FIXED",
             "MANAGER_COMMENT_alice": "Confirmed fix",
             "SCREENSHOT_alice": "screen01.png",
             "TESTER_STATUS_alice": "ISSUE"},
            {"STRINGID": "2001", "TRANSLATION": "Good morning"},
            {"STRINGID": "2001", "TRANSLATION": "Good morning"},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        for excel_row in [3, 4]:
            assert _cell_val(wb, "Sheet1", excel_row, "COMMENT_alice") == "Fixed typo"
            assert _cell_val(wb, "Sheet1", excel_row, "STATUS_alice") == "FIXED"
            assert _cell_val(wb, "Sheet1", excel_row, "MANAGER_COMMENT_alice") == "Confirmed fix"
            assert _cell_val(wb, "Sheet1", excel_row, "SCREENSHOT_alice") == "screen01.png"
            assert _cell_val(wb, "Sheet1", excel_row, "TESTER_STATUS_alice") == "ISSUE"

        # 5 columns * 2 empty rows = 10 cells filled
        assert filled == 10


# =============================================================================
# TEST 3: No overwrite of existing data
# =============================================================================

class TestNoOverwriteExisting:
    """3 duplicate rows. Row 1 has COMMENT_alice='Comment A'.
    Row 2 has COMMENT_alice='Comment B'. Row 3 empty.
    Verify row 1 keeps 'Comment A', row 2 keeps 'Comment B',
    row 3 gets 'Comment A' (first non-empty)."""

    def test_existing_values_preserved(self):
        rows = [
            {"STRINGID": "3001", "TRANSLATION": "Farewell",
             "COMMENT_alice": "Comment A"},
            {"STRINGID": "3001", "TRANSLATION": "Farewell",
             "COMMENT_alice": "Comment B"},
            {"STRINGID": "3001", "TRANSLATION": "Farewell"},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        assert _cell_val(wb, "Sheet1", 2, "COMMENT_alice") == "Comment A"
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Comment B"
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_alice") == "Comment A"

        # Only row 3 (Excel row 4) was empty -> 1 cell filled
        assert filled == 1


# =============================================================================
# TEST 4: Non-duplicates untouched
# =============================================================================

class TestNonDuplicatesUntouched:
    """3 rows with DIFFERENT StringID+Translation. Each has different data.
    Verify nothing changes."""

    def test_unique_rows_not_modified(self):
        rows = [
            {"STRINGID": "4001", "TRANSLATION": "Alpha",
             "COMMENT_alice": "Comment Alpha"},
            {"STRINGID": "4002", "TRANSLATION": "Beta",
             "COMMENT_alice": "Comment Beta"},
            {"STRINGID": "4003", "TRANSLATION": "Gamma"},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        assert _cell_val(wb, "Sheet1", 2, "COMMENT_alice") == "Comment Alpha"
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Comment Beta"
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_alice") is None

        assert filled == 0


# =============================================================================
# TEST 5: Multiple users replicated
# =============================================================================

class TestMultipleUsersReplicated:
    """3 duplicate rows. Row 1 has COMMENT_alice filled. Row 2 has
    COMMENT_bob filled. Row 3 empty. After: Row 3 gets both
    COMMENT_alice and COMMENT_bob."""

    def test_multi_user_columns_filled(self):
        extra = ["COMMENT_bob", "STATUS_bob"]
        rows = [
            {"STRINGID": "5001", "TRANSLATION": "Welcome",
             "COMMENT_alice": "Alice's note"},
            {"STRINGID": "5001", "TRANSLATION": "Welcome",
             "COMMENT_bob": "Bob's note"},
            {"STRINGID": "5001", "TRANSLATION": "Welcome"},
        ]
        wb = _make_standard_wb(rows, extra_headers=extra)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # Row 2 (Excel): alice filled, bob empty -> bob gets filled
        assert _cell_val(wb, "Sheet1", 2, "COMMENT_alice") == "Alice's note"
        assert _cell_val(wb, "Sheet1", 2, "COMMENT_bob") == "Bob's note"

        # Row 3 (Excel): alice empty, bob filled -> alice gets filled
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Alice's note"
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_bob") == "Bob's note"

        # Row 4 (Excel): both empty -> both get filled
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_alice") == "Alice's note"
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_bob") == "Bob's note"

        # COMMENT_alice: empty in rows 3,4 = 2 fills
        # COMMENT_bob: empty in rows 2,4 = 2 fills
        # Total = 4
        assert filled == 4


# =============================================================================
# TEST 6: Item category duplicates
# =============================================================================

class TestItemCategoryDuplicates:
    """Item category. 3 rows with same ItemName+ItemDesc+StringID.
    Verify duplication works using Item-specific content key."""

    def test_item_duplicates_replicated(self):
        rows = [
            {"STRINGID": "6001", "ITEMNAME(ENG)": "Iron Sword",
             "ITEMDESC(ENG)": "A sturdy blade",
             "COMMENT_alice": "Name OK", "STATUS_alice": "NO ISSUE"},
            {"STRINGID": "6001", "ITEMNAME(ENG)": "Iron Sword",
             "ITEMDESC(ENG)": "A sturdy blade"},
            {"STRINGID": "6001", "ITEMNAME(ENG)": "Iron Sword",
             "ITEMDESC(ENG)": "A sturdy blade"},
        ]
        wb = _make_item_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Item", is_english=True)

        for excel_row in [3, 4]:
            assert _cell_val(wb, "Sheet1", excel_row, "COMMENT_alice") == "Name OK"
            assert _cell_val(wb, "Sheet1", excel_row, "STATUS_alice") == "NO ISSUE"

        # 2 columns * 2 empty rows = 4
        assert filled == 4

    def test_different_items_not_grouped(self):
        """Items with different names should NOT be grouped together."""
        rows = [
            {"STRINGID": "6001", "ITEMNAME(ENG)": "Iron Sword",
             "ITEMDESC(ENG)": "A sturdy blade",
             "COMMENT_alice": "Sword OK"},
            {"STRINGID": "6002", "ITEMNAME(ENG)": "Steel Shield",
             "ITEMDESC(ENG)": "Heavy protection"},
            {"STRINGID": "6003", "ITEMNAME(ENG)": "Wooden Staff",
             "ITEMDESC(ENG)": "Magic focus"},
        ]
        wb = _make_item_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Item", is_english=True)

        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") is None
        assert _cell_val(wb, "Sheet1", 4, "COMMENT_alice") is None
        assert filled == 0


# =============================================================================
# TEST 7: Script category duplicates
# =============================================================================

class TestScriptCategoryDuplicates:
    """Script category (Sequencer). 3 rows with same Text+EventName.
    Verify duplication works using script-specific content key."""

    def test_script_duplicates_replicated(self):
        rows = [
            {"TEXT": "Run away!", "EVENTNAME": "EVT_ESCAPE_01",
             "COMMENT_alice": "Urgent tone", "STATUS_alice": "ISSUE"},
            {"TEXT": "Run away!", "EVENTNAME": "EVT_ESCAPE_01"},
            {"TEXT": "Run away!", "EVENTNAME": "EVT_ESCAPE_01"},
        ]
        wb = _make_script_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Sequencer", is_english=True)

        for excel_row in [3, 4]:
            assert _cell_val(wb, "Sheet1", excel_row, "COMMENT_alice") == "Urgent tone"
            assert _cell_val(wb, "Sheet1", excel_row, "STATUS_alice") == "ISSUE"

        # 2 columns * 2 empty rows = 4
        assert filled == 4

    def test_dialog_category_also_works(self):
        """Dialog is also a script-type category."""
        rows = [
            {"TEXT": "Hello there", "EVENTNAME": "DLG_GREET_01",
             "COMMENT_alice": "Friendly"},
            {"TEXT": "Hello there", "EVENTNAME": "DLG_GREET_01"},
        ]
        wb = _make_script_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Dialog", is_english=True)

        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Friendly"
        assert filled == 1


# =============================================================================
# TEST 8: STATUS sheet skipped
# =============================================================================

class TestStatusSheetSkipped:
    """Ensure the STATUS sheet is never processed."""

    def test_status_sheet_ignored(self):
        wb = Workbook()
        # Create a "STATUS" sheet with duplicate rows
        ws_status = wb.active
        ws_status.title = "STATUS"

        headers = ["STRINGID", "TRANSLATION", "COMMENT_alice", "STATUS_alice"]
        for col_idx, h in enumerate(headers, start=1):
            ws_status.cell(row=1, column=col_idx, value=h)

        # 3 duplicate rows with data only in row 1
        for row_idx in [2, 3, 4]:
            ws_status.cell(row=row_idx, column=1, value="9001")
            ws_status.cell(row=row_idx, column=2, value="Status text")

        ws_status.cell(row=2, column=3, value="Should not replicate")
        ws_status.cell(row=2, column=4, value="ISSUE")

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # STATUS sheet should be completely skipped
        assert filled == 0
        assert ws_status.cell(row=3, column=3).value is None
        assert ws_status.cell(row=4, column=3).value is None

    def test_status_sheet_skipped_but_other_sheets_processed(self):
        """STATUS is skipped but normal sheets still get processed."""
        wb = Workbook()

        # STATUS sheet (should be skipped)
        ws_status = wb.active
        ws_status.title = "STATUS"
        ws_status.cell(1, 1, "STRINGID")
        ws_status.cell(1, 2, "TRANSLATION")
        ws_status.cell(1, 3, "COMMENT_alice")
        for r in [2, 3]:
            ws_status.cell(r, 1, "S001")
            ws_status.cell(r, 2, "Status text")
        ws_status.cell(2, 3, "Status comment")

        # Normal sheet (should be processed)
        ws_normal = wb.create_sheet("QuestData")
        ws_normal.cell(1, 1, "STRINGID")
        ws_normal.cell(1, 2, "TRANSLATION")
        ws_normal.cell(1, 3, "COMMENT_alice")
        for r in [2, 3]:
            ws_normal.cell(r, 1, "N001")
            ws_normal.cell(r, 2, "Normal text")
        ws_normal.cell(2, 3, "Normal comment")

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # STATUS sheet untouched
        assert ws_status.cell(3, 3).value is None

        # Normal sheet replicated
        assert ws_normal.cell(3, 3).value == "Normal comment"
        assert filled == 1


# =============================================================================
# TEST 9: Returns correct count
# =============================================================================

class TestReturnsCorrectCount:
    """Verify the return value equals the exact number of cells filled."""

    def test_zero_when_no_duplicates(self):
        rows = [
            {"STRINGID": "A", "TRANSLATION": "One", "COMMENT_alice": "C1"},
            {"STRINGID": "B", "TRANSLATION": "Two", "COMMENT_alice": "C2"},
        ]
        wb = _make_standard_wb(rows)
        assert replicate_duplicate_row_data(wb, "Quest", is_english=True) == 0

    def test_zero_when_all_cells_already_filled(self):
        rows = [
            {"STRINGID": "X", "TRANSLATION": "Same",
             "COMMENT_alice": "C1", "STATUS_alice": "ISSUE"},
            {"STRINGID": "X", "TRANSLATION": "Same",
             "COMMENT_alice": "C2", "STATUS_alice": "NO ISSUE"},
        ]
        wb = _make_standard_wb(rows)
        assert replicate_duplicate_row_data(wb, "Quest", is_english=True) == 0

    def test_exact_count_mixed_scenario(self):
        """Multiple groups, multiple columns, verify precise count."""
        extra = ["COMMENT_bob"]
        rows = [
            # Group A: 3 rows, alice comment filled in row 1, bob in row 2
            {"STRINGID": "GA", "TRANSLATION": "GroupA",
             "COMMENT_alice": "A-alice"},
            {"STRINGID": "GA", "TRANSLATION": "GroupA",
             "COMMENT_bob": "A-bob"},
            {"STRINGID": "GA", "TRANSLATION": "GroupA"},
            # Group B: 2 rows, alice comment filled in row 1
            {"STRINGID": "GB", "TRANSLATION": "GroupB",
             "COMMENT_alice": "B-alice"},
            {"STRINGID": "GB", "TRANSLATION": "GroupB"},
        ]
        wb = _make_standard_wb(rows, extra_headers=extra)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # Group A:
        #   COMMENT_alice: empty in rows 3,4 -> 2 fills
        #   COMMENT_bob: empty in rows 2,4 -> 2 fills
        # Group B:
        #   COMMENT_alice: empty in row 6 -> 1 fill
        #   COMMENT_bob: no data in group -> 0 fills
        # Total: 2 + 2 + 1 = 5
        assert filled == 5

    def test_empty_workbook_returns_zero(self):
        """Workbook with only headers (no data rows) returns 0."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(1, 1, "STRINGID")
        ws.cell(1, 2, "TRANSLATION")
        ws.cell(1, 3, "COMMENT_alice")

        assert replicate_duplicate_row_data(wb, "Quest", is_english=True) == 0

    def test_single_row_returns_zero(self):
        """Single data row can never be a duplicate group."""
        rows = [
            {"STRINGID": "SOLO", "TRANSLATION": "Only one",
             "COMMENT_alice": "Solo comment"},
        ]
        wb = _make_standard_wb(rows)
        assert replicate_duplicate_row_data(wb, "Quest", is_english=True) == 0


# =============================================================================
# EDGE CASES
# =============================================================================

class TestEdgeCases:
    """Additional edge cases for robustness."""

    def test_whitespace_only_treated_as_empty(self):
        """Cells containing only whitespace should be treated as empty
        and overwritten by the replication."""
        rows = [
            {"STRINGID": "W001", "TRANSLATION": "Whitespace test",
             "COMMENT_alice": "Real comment"},
            {"STRINGID": "W001", "TRANSLATION": "Whitespace test",
             "COMMENT_alice": "   "},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        # Whitespace-only cell should be overwritten
        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "Real comment"
        assert filled == 1

    def test_multiple_sheets_processed_independently(self):
        """Each sheet's duplicates are independent."""
        wb = Workbook()

        # Sheet1
        ws1 = wb.active
        ws1.title = "Sheet1"
        for col, h in enumerate(["STRINGID", "TRANSLATION", "COMMENT_alice"], 1):
            ws1.cell(1, col, h)
        for r in [2, 3]:
            ws1.cell(r, 1, "S1")
            ws1.cell(r, 2, "Sheet1 text")
        ws1.cell(2, 3, "Sheet1 comment")

        # Sheet2
        ws2 = wb.create_sheet("Sheet2")
        for col, h in enumerate(["STRINGID", "TRANSLATION", "COMMENT_alice"], 1):
            ws2.cell(1, col, h)
        for r in [2, 3]:
            ws2.cell(r, 1, "S2")
            ws2.cell(r, 2, "Sheet2 text")
        ws2.cell(2, 3, "Sheet2 comment")

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=True)

        assert ws1.cell(3, 3).value == "Sheet1 comment"
        assert ws2.cell(3, 3).value == "Sheet2 comment"
        assert filled == 2

    def test_cn_language_flag(self):
        """is_english=False should still work (content key uses same
        STRINGID + TRANSLATION column detection by header name)."""
        rows = [
            {"STRINGID": "CN01", "TRANSLATION": "Chinese text",
             "COMMENT_alice": "CN comment"},
            {"STRINGID": "CN01", "TRANSLATION": "Chinese text"},
        ]
        wb = _make_standard_wb(rows)

        filled = replicate_duplicate_row_data(wb, "Quest", is_english=False)

        assert _cell_val(wb, "Sheet1", 3, "COMMENT_alice") == "CN comment"
        assert filled == 1
