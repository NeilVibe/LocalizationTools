"""
Integration tests for tracker.masterfile_pending — verifying stale issues
are excluded and only masterfile-resident rows count as active.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from tracker.masterfile_pending import build_pending_from_masterfiles


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_master(folder: Path, filename: str, sheets: dict):
    """
    Create a mock masterfile with given sheets.

    sheets = {
        "SheetName": {
            "headers": ["Col1", "Col2", ...],
            "rows": [["val1", "val2", ...], ...],
        },
    }
    """
    wb = Workbook()
    first = True
    for sheet_name, spec in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for j, hdr in enumerate(spec["headers"], start=1):
            ws.cell(1, j, hdr)
        for i, row in enumerate(spec["rows"], start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)

    path = folder / filename
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_dirs(tmp_path):
    en = tmp_path / "Masterfolder_EN"
    cn = tmp_path / "Masterfolder_CN"
    en.mkdir()
    cn.mkdir()
    return en, cn


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_stale_issues_excluded_from_active(tmp_dirs):
    """
    QA file might have 10 ISSUE rows, but masterfile only has 3 of those
    StringIDs. Active issues should be 3 (from masterfile), not 10.

    This test creates a masterfile with exactly 3 ISSUE rows and verifies
    that only those 3 are counted — the function reads masterfiles, not
    QA files, so stale issues that didn't make it into the master are
    naturally excluded.
    """
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_UserA", "STATUS_UserA",
                "COMMENT_UserA", "MEMO_UserA", "SCREENSHOT_UserA",
            ],
            "rows": [
                # ISSUE + empty status + comment -> pending
                ["KR1", "EN1", "S001", "ISSUE", None, "Typo here", None, None],
                # ISSUE + FIXED + comment -> fixed
                ["KR2", "EN2", "S002", "ISSUE", "FIXED", "Was wrong", None, None],
                # ISSUE + CHECKING + comment -> checking + pending
                ["KR3", "EN3", "S003", "ISSUE", "CHECKING", "Grammar", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    user_a = result["UserA"]["Quest"]
    assert user_a["active_issues"] == 3
    # Row 1 (empty) -> pending, Row 3 (CHECKING) -> also pending
    assert user_a["pending"] == 2
    assert user_a["fixed"] == 1
    assert user_a["checking"] == 1
    assert user_a["reported"] == 0
    assert user_a["nonissue"] == 0


def test_multiple_categories_independent(tmp_dirs):
    """
    Two masterfiles (Master_Quest.xlsx, Master_Item.xlsx) with different
    testers. Verify counts are independent per category.
    """
    en, cn = tmp_dirs

    # Quest masterfile: Alice has 2 issues
    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", None, "Wrong word", None, None],
                ["KR2", "EN2", "S002", "ISSUE", "FIXED", "Typo", None, None],
            ],
        },
    })

    # Item masterfile: Bob has 3 issues
    _make_master(en, "Master_Item.xlsx", {
        "Items": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Bob", "STATUS_Bob",
                "COMMENT_Bob", "MEMO_Bob", "SCREENSHOT_Bob",
            ],
            "rows": [
                ["KR1", "EN1", "S101", "ISSUE", "REPORTED", "Bad format", None, None],
                ["KR2", "EN2", "S102", "ISSUE", None, "Missing text", None, None],
                ["KR3", "EN3", "S103", "ISSUE", "CHECKING", "Under review", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    # Alice only in Quest
    assert "Quest" in result["Alice"]
    assert "Item" not in result["Alice"]
    alice_quest = result["Alice"]["Quest"]
    assert alice_quest["active_issues"] == 2
    assert alice_quest["pending"] == 1
    assert alice_quest["fixed"] == 1

    # Bob only in Item
    assert "Item" in result["Bob"]
    assert "Quest" not in result["Bob"]
    bob_item = result["Bob"]["Item"]
    assert bob_item["active_issues"] == 3
    assert bob_item["reported"] == 1
    # pending = 1 (empty) + 1 (CHECKING) = 2
    assert bob_item["pending"] == 2
    assert bob_item["checking"] == 1


def test_manager_resolved_all_no_pending(tmp_dirs):
    """
    Masterfile where all ISSUE rows have FIXED/REPORTED/NON-ISSUE status.
    Verify pending=0.
    """
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_UserA", "STATUS_UserA",
                "COMMENT_UserA", "MEMO_UserA", "SCREENSHOT_UserA",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", "FIXED", "Was wrong", None, None],
                ["KR2", "EN2", "S002", "ISSUE", "REPORTED", "Reported upstream", None, None],
                ["KR3", "EN3", "S003", "ISSUE", "NON-ISSUE", "Intended", None, None],
                ["KR4", "EN4", "S004", "ISSUE", "FIXED", "Corrected", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    user_a = result["UserA"]["Quest"]
    assert user_a["active_issues"] == 4
    assert user_a["pending"] == 0
    assert user_a["fixed"] == 2
    assert user_a["reported"] == 1
    assert user_a["nonissue"] == 1


def test_mixed_tester_statuses_only_issues_counted(tmp_dirs):
    """
    Masterfile with rows where TESTER_STATUS is ISSUE, NO ISSUE, BLOCKED,
    KOREAN, empty. Only ISSUE rows should appear in active counts.
    """
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_UserA", "STATUS_UserA",
                "COMMENT_UserA", "MEMO_UserA", "SCREENSHOT_UserA",
            ],
            "rows": [
                # ISSUE -> counted (has comment)
                ["KR1", "EN1", "S001", "ISSUE", None, "Wrong text", None, None],
                ["KR2", "EN2", "S002", "ISSUE", "FIXED", "Typo fixed", None, None],
                # NO ISSUE -> NOT counted
                ["KR3", "EN3", "S003", "NO ISSUE", None, None, None, None],
                # BLOCKED -> NOT counted
                ["KR4", "EN4", "S004", "BLOCKED", None, None, None, None],
                # KOREAN -> NOT counted
                ["KR5", "EN5", "S005", "KOREAN", None, None, None, None],
                # Empty -> NOT counted
                ["KR6", "EN6", "S006", None, None, None, None, None],
                # Another ISSUE -> counted (has comment)
                ["KR7", "EN7", "S007", "ISSUE", "REPORTED", "Bad format", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    user_a = result["UserA"]["Quest"]
    # Only 3 rows have TESTER_STATUS = ISSUE
    assert user_a["active_issues"] == 3
    assert user_a["pending"] == 1
    assert user_a["fixed"] == 1
    assert user_a["reported"] == 1
