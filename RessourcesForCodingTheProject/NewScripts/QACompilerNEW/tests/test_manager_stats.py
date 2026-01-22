"""
Test manager_stats collection and lookup logic with mock data.
"""
import sys
import tempfile
import shutil
from pathlib import Path
from openpyxl import Workbook

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


def create_mock_master_script(folder: Path):
    """Create a mock Master_Script.xlsx with STATUS_ columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sequencer_Sheet1"

    # Headers: EventName, Text, Translation, STATUS_Alice, STATUS_Bob
    ws.cell(1, 1, "EventName")
    ws.cell(1, 2, "Text")
    ws.cell(1, 3, "Translation")
    ws.cell(1, 4, "STATUS_Alice")
    ws.cell(1, 5, "STATUS_Bob")

    # Data rows with manager statuses
    data = [
        ("EVT001", "Hello", "Bonjour", "FIXED", "REPORTED"),
        ("EVT002", "World", "Monde", "FIXED", "CHECKING"),
        ("EVT003", "Test", "Test", "REPORTED", "NON-ISSUE"),
        ("EVT004", "Game", "Jeu", None, "FIXED"),
        ("EVT005", "Play", "Jouer", "CHECKING", None),
    ]
    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    # Add a Dialog sheet too
    ws2 = wb.create_sheet("Dialog_Sheet1")
    ws2.cell(1, 1, "EventName")
    ws2.cell(1, 2, "Text")
    ws2.cell(1, 3, "Translation")
    ws2.cell(1, 4, "STATUS_Alice")
    ws2.cell(1, 5, "STATUS_Charlie")

    data2 = [
        ("DLG001", "Hi", "Salut", "FIXED", "FIXED"),
        ("DLG002", "Bye", "Au revoir", "NON ISSUE", "REPORTED"),  # Note: space not hyphen
    ]
    for i, row in enumerate(data2, start=2):
        for j, val in enumerate(row, start=1):
            ws2.cell(i, j, val)

    master_path = folder / "Master_Script.xlsx"
    wb.save(master_path)
    wb.close()
    print(f"Created: {master_path}")
    return master_path


def create_mock_master_quest(folder: Path):
    """Create a mock Master_Quest.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MainQuest"

    ws.cell(1, 1, "Korean")
    ws.cell(1, 2, "English")
    ws.cell(1, 3, "STRINGID")
    ws.cell(1, 4, "STATUS_Alice")
    ws.cell(1, 5, "STATUS_Dave")

    data = [
        ("퀘스트1", "Quest 1", "Q001", "FIXED", "FIXED"),
        ("퀘스트2", "Quest 2", "Q002", "REPORTED", None),
    ]
    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    master_path = folder / "Master_Quest.xlsx"
    wb.save(master_path)
    wb.close()
    print(f"Created: {master_path}")
    return master_path


def test_collect_manager_stats():
    """Test that manager_stats correctly collects counts."""
    import config

    # Create temp folders
    temp_dir = Path(tempfile.mkdtemp())
    master_en = temp_dir / "Masterfolder_EN"
    master_cn = temp_dir / "Masterfolder_CN"
    master_en.mkdir()
    master_cn.mkdir()

    # Patch config paths
    original_en = config.MASTER_FOLDER_EN
    original_cn = config.MASTER_FOLDER_CN
    config.MASTER_FOLDER_EN = master_en
    config.MASTER_FOLDER_CN = master_cn

    try:
        # Create mock masters (EN only for simplicity)
        create_mock_master_script(master_en)
        create_mock_master_quest(master_en)

        # Import and run collection
        from core.compiler import collect_manager_stats_for_tracker
        result = collect_manager_stats_for_tracker()

        print("\n=== COLLECTION RESULTS ===")
        print(f"Categories collected: {list(result.keys())}")

        for cat, users in result.items():
            print(f"\n{cat}:")
            for user, stats in users.items():
                print(f"  {user}: {stats}")

        # Verify Script category
        assert "Script" in result, "Script category should exist"
        script = result["Script"]

        # Alice: 2 FIXED (Seq) + 1 FIXED (Dlg) + 1 REPORTED (Seq) + 1 CHECKING (Seq) + 1 NON ISSUE (Dlg)
        # = 3 fixed, 1 reported, 1 checking, 1 nonissue
        assert "Alice" in script, "Alice should be in Script"
        print(f"\nAlice Script stats: {script['Alice']}")
        assert script["Alice"]["fixed"] == 3, f"Alice fixed should be 3, got {script['Alice']['fixed']}"
        assert script["Alice"]["reported"] == 1, f"Alice reported should be 1, got {script['Alice']['reported']}"
        assert script["Alice"]["checking"] == 1, f"Alice checking should be 1, got {script['Alice']['checking']}"
        assert script["Alice"]["nonissue"] == 1, f"Alice nonissue should be 1, got {script['Alice']['nonissue']}"

        # Bob: 1 REPORTED + 1 CHECKING + 1 NON-ISSUE + 1 FIXED = 1 fixed, 1 reported, 1 checking, 1 nonissue
        assert "Bob" in script, "Bob should be in Script"
        print(f"Bob Script stats: {script['Bob']}")
        assert script["Bob"]["fixed"] == 1
        assert script["Bob"]["reported"] == 1
        assert script["Bob"]["checking"] == 1
        assert script["Bob"]["nonissue"] == 1

        # Charlie: 1 FIXED + 1 REPORTED
        assert "Charlie" in script, "Charlie should be in Script"
        print(f"Charlie Script stats: {script['Charlie']}")
        assert script["Charlie"]["fixed"] == 1
        assert script["Charlie"]["reported"] == 1

        # Verify Quest category
        assert "Quest" in result, "Quest category should exist"
        quest = result["Quest"]
        assert "Alice" in quest
        assert quest["Alice"]["fixed"] == 1
        assert quest["Alice"]["reported"] == 1

        print("\n=== ALL TESTS PASSED ===")

    finally:
        # Restore config
        config.MASTER_FOLDER_EN = original_en
        config.MASTER_FOLDER_CN = original_cn
        # Cleanup
        shutil.rmtree(temp_dir)


def test_lookup_phase():
    """Test that lookup correctly matches tester entries to manager stats."""
    print("\n=== LOOKUP PHASE TEST ===")

    # Simulate manager_stats (what would be collected)
    manager_stats = {
        "Script": {
            "Alice": {"fixed": 3, "reported": 1, "checking": 1, "nonissue": 1, "lang": "EN"},
            "Bob": {"fixed": 1, "reported": 1, "checking": 1, "nonissue": 1, "lang": "EN"},
        },
        "Quest": {
            "Alice": {"fixed": 1, "reported": 1, "checking": 0, "nonissue": 0, "lang": "EN"},
        }
    }

    # Simulate daily_entries (tester stats from QA folders)
    from config import get_target_master_category

    daily_entries = [
        {"user": "Alice", "category": "Sequencer"},  # Should map to Script
        {"user": "Bob", "category": "Dialog"},       # Should map to Script
        {"user": "Alice", "category": "Quest"},      # Should map to Quest
        {"user": "Charlie", "category": "Sequencer"},  # Charlie not in manager_stats
        {"user": "Alice", "category": "Knowledge"},    # Knowledge not in manager_stats
    ]

    for entry in daily_entries:
        user = entry["user"]
        category = entry["category"]
        lookup_category = get_target_master_category(category)

        category_stats = manager_stats.get(lookup_category, {})
        user_stats = category_stats.get(user, {"fixed": 0, "reported": 0, "checking": 0, "nonissue": 0})

        has_stats = any(user_stats[k] > 0 for k in ["fixed", "reported", "checking", "nonissue"])

        if has_stats:
            print(f"  HIT: {user}/{category} -> {lookup_category} -> {user_stats}")
        else:
            reason = "NO_CAT" if lookup_category not in manager_stats else ("NO_USER" if user not in category_stats else "ZERO")
            print(f"  MISS: {user}/{category} -> {lookup_category} -> {reason}")

    print("\n=== LOOKUP TEST COMPLETE ===")


if __name__ == "__main__":
    print("Testing manager_stats collection and lookup...\n")
    test_lookup_phase()
    print()
    test_collect_manager_stats()
