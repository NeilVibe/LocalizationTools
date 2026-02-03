#!/usr/bin/env python3
"""
Test Script-Type Preprocessing
==============================

Tests for preprocess_script_category() and build_master_from_universe()
to ensure they handle edge cases and produce correct output.

Run with: python3 tests/test_script_preprocessing.py
"""

import sys
import traceback
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


class TestResult:
    """Simple test result tracker."""
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.errors = []

    def pass_test(self, name: str):
        self.passed += 1
        print(f"  [PASS] {name}")

    def fail_test(self, name: str, reason: str):
        self.failed += 1
        self.errors.append((name, reason))
        print(f"  [FAIL] {name}: {reason}")

    def summary(self):
        total = self.passed + self.failed
        print(f"\n{'='*60}")
        print(f"Test Results: {self.passed}/{total} passed")
        if self.errors:
            print(f"\nFailed tests:")
            for name, reason in self.errors:
                print(f"  - {name}: {reason}")
        print(f"{'='*60}")
        return self.failed == 0


def test_preprocess_script_category():
    """Test preprocess_script_category with mock fixture."""
    print("\n--- Testing preprocess_script_category ---")
    result = TestResult()

    # Import the function
    from core.compiler import preprocess_script_category

    fixture_path = Path(__file__).parent / "fixtures" / "mock_script_qa.xlsx"

    if not fixture_path.exists():
        print(f"  [ERROR] Fixture not found: {fixture_path}")
        print(f"  Run create_mock_fixture.py first!")
        return False

    # Create mock qa_folders structure
    qa_folders = [{
        "xlsx_path": fixture_path,
        "username": "test_user",
        "category": "Sequencer",
        "folder_path": fixture_path.parent,
        "images": [],
    }]

    try:
        # Run preprocessing
        universe = preprocess_script_category(qa_folders, is_english=True)

        # Test 1: Returns dict with expected keys
        expected_keys = {"rows", "row_count", "source_files", "errors", "headers", "num_columns"}
        if isinstance(universe, dict) and expected_keys.issubset(universe.keys()):
            result.pass_test("Returns dict with expected keys (including Phase B keys)")
        else:
            result.fail_test("Returns dict with expected keys", f"Got: {type(universe)}, keys: {universe.keys() if isinstance(universe, dict) else 'N/A'}")

        # Test 2: Found expected number of rows with STATUS
        # Expected: 6 from Sequencer (ISSUE, NON-ISSUE, ISSUE, BLOCKED, ISSUE, NON-ISSUE)
        #         + 3 from Dialog (ISSUE, NON-ISSUE, KOREAN)
        expected_count = 9
        actual_count = universe.get("row_count", 0)
        if actual_count == expected_count:
            result.pass_test(f"Found {expected_count} rows with STATUS")
        else:
            result.fail_test(f"Found {expected_count} rows with STATUS", f"Got {actual_count}")

        # Test 3: Rows dict is properly keyed
        rows = universe.get("rows", {})
        sample_key = ("SEQ_001", "Hello there!")
        if sample_key in rows:
            result.pass_test("Rows keyed by (eventname, text)")
        else:
            result.fail_test("Rows keyed by (eventname, text)", f"Key {sample_key} not found")

        # Test 4: Source tracking works
        if sample_key in rows:
            sources = rows[sample_key].get("sources", [])
            if sources and len(sources) > 0:
                result.pass_test("Source tracking works")
            else:
                result.fail_test("Source tracking works", f"No sources found for {sample_key}")
        else:
            result.fail_test("Source tracking works", "Sample key not found")

        # Test 5: Empty STATUS rows are skipped
        empty_key = ("SEQ_003", "Skip me")
        if empty_key not in rows:
            result.pass_test("Empty STATUS rows are skipped")
        else:
            result.fail_test("Empty STATUS rows are skipped", f"Key {empty_key} was not skipped")

        # Test 6: Sparse rows are found (row 50 and row 100)
        sparse_key_50 = ("SEQ_050", "Sparse issue at row 50")
        sparse_key_100 = ("SEQ_100", "Sparse non-issue at row 100")
        if sparse_key_50 in rows and sparse_key_100 in rows:
            result.pass_test("Sparse rows at row 50 and 100 found")
        else:
            missing = []
            if sparse_key_50 not in rows:
                missing.append("row 50")
            if sparse_key_100 not in rows:
                missing.append("row 100")
            result.fail_test("Sparse rows at row 50 and 100 found", f"Missing: {missing}")

        # Test 7: Errors list is included
        if "errors" in universe:
            result.pass_test("Errors list included in result")
        else:
            result.fail_test("Errors list included in result", "No errors key")

        # Test 8 (Phase B): full_row data is stored
        if sample_key in rows:
            full_row = rows[sample_key].get("full_row")
            if full_row and isinstance(full_row, list) and len(full_row) > 0:
                result.pass_test("full_row data stored for each row")
            else:
                result.fail_test("full_row data stored for each row", f"Got: {type(full_row)}")
        else:
            result.fail_test("full_row data stored for each row", "Sample key not found")

        # Test 9 (Phase B): headers collected
        headers = universe.get("headers", {})
        if "Sequencer" in headers and len(headers["Sequencer"]) > 0:
            result.pass_test("Sheet headers collected")
        else:
            result.fail_test("Sheet headers collected", f"Got headers keys: {list(headers.keys())}")

        # Test 10 (Phase B): num_columns collected
        num_cols = universe.get("num_columns", {})
        if "Sequencer" in num_cols and num_cols["Sequencer"] > 0:
            result.pass_test("Sheet num_columns collected")
        else:
            result.fail_test("Sheet num_columns collected", f"Got: {num_cols}")

    except Exception as e:
        print(f"  [ERROR] Exception during test: {e}")
        traceback.print_exc()
        result.fail_test("No exceptions", str(e))

    return result.summary()


def test_build_master_from_universe():
    """Test build_master_from_universe with mock fixture."""
    print("\n--- Testing build_master_from_universe ---")
    result = TestResult()

    from core.compiler import preprocess_script_category, build_master_from_universe
    import tempfile
    import shutil

    fixture_path = Path(__file__).parent / "fixtures" / "mock_script_qa.xlsx"

    if not fixture_path.exists():
        print(f"  [ERROR] Fixture not found: {fixture_path}")
        return False

    # Create mock qa_folders structure
    qa_folders = [{
        "xlsx_path": fixture_path,
        "username": "test_user",
        "category": "Sequencer",
        "folder_path": fixture_path.parent,
        "images": [],
    }]

    # Create temp master folder
    temp_dir = Path(tempfile.mkdtemp(prefix="test_master_"))

    try:
        # First, run preprocessing
        universe = preprocess_script_category(qa_folders, is_english=True)

        # Test 1: build_master_from_universe doesn't crash
        try:
            master_wb, master_path = build_master_from_universe("Sequencer", universe, temp_dir)
            result.pass_test("build_master_from_universe completes without error")
        except Exception as e:
            result.fail_test("build_master_from_universe completes without error", str(e))
            traceback.print_exc()
            return result.summary()

        # Test 2: Returns a workbook
        if master_wb is not None:
            result.pass_test("Returns a Workbook")
        else:
            result.fail_test("Returns a Workbook", "Got None")
            return result.summary()

        # Test 3: Contains expected sheets
        if "Sequencer" in master_wb.sheetnames:
            result.pass_test("Contains Sequencer sheet")
        else:
            result.fail_test("Contains Sequencer sheet", f"Sheets: {master_wb.sheetnames}")

        if "Dialog" in master_wb.sheetnames:
            result.pass_test("Contains Dialog sheet")
        else:
            result.fail_test("Contains Dialog sheet", f"Sheets: {master_wb.sheetnames}")

        # Test 4: Sequencer sheet has correct row count (header + data rows)
        if "Sequencer" in master_wb.sheetnames:
            ws_seq = master_wb["Sequencer"]
            expected_seq_rows = 6 + 1  # 6 with STATUS + 1 header
            actual_seq_rows = ws_seq.max_row
            if actual_seq_rows == expected_seq_rows:
                result.pass_test(f"Sequencer has {expected_seq_rows} rows (6 data + header)")
            else:
                result.fail_test(f"Sequencer has {expected_seq_rows} rows", f"Got {actual_seq_rows}")

        # Test 5: Dialog sheet has correct row count
        if "Dialog" in master_wb.sheetnames:
            ws_dialog = master_wb["Dialog"]
            expected_dialog_rows = 3 + 1  # 3 with STATUS + 1 header
            actual_dialog_rows = ws_dialog.max_row
            if actual_dialog_rows == expected_dialog_rows:
                result.pass_test(f"Dialog has {expected_dialog_rows} rows (3 data + header)")
            else:
                result.fail_test(f"Dialog has {expected_dialog_rows} rows", f"Got {actual_dialog_rows}")

        # Test 6: STATUS/COMMENT/SCREENSHOT/STRINGID columns are removed from master
        if "Sequencer" in master_wb.sheetnames:
            ws_seq = master_wb["Sequencer"]
            removed_cols = set()
            for col in range(1, ws_seq.max_column + 1):
                header = ws_seq.cell(row=1, column=col).value
                if header and str(header).strip().upper() in ("STATUS", "COMMENT", "SCREENSHOT", "STRINGID"):
                    removed_cols.add(str(header).strip().upper())
            if not removed_cols:
                result.pass_test("STATUS/COMMENT/SCREENSHOT/STRINGID columns removed")
            else:
                result.fail_test("STATUS/COMMENT/SCREENSHOT/STRINGID columns removed", f"Still found: {removed_cols}")

        # Test 7: Data values are preserved
        if "Sequencer" in master_wb.sheetnames:
            ws_seq = master_wb["Sequencer"]
            # Find EventName column
            eventname_col = None
            for col in range(1, ws_seq.max_column + 1):
                if ws_seq.cell(row=1, column=col).value == "EventName":
                    eventname_col = col
                    break
            if eventname_col:
                first_val = ws_seq.cell(row=2, column=eventname_col).value
                if first_val and str(first_val).startswith("SEQ_"):
                    result.pass_test("Data values preserved in master")
                else:
                    result.fail_test("Data values preserved in master", f"First EventName: {first_val}")
            else:
                result.fail_test("Data values preserved in master", "EventName column not found")

        # Cleanup
        master_wb.close()

    except Exception as e:
        print(f"  [ERROR] Exception during test: {e}")
        traceback.print_exc()
        result.fail_test("No exceptions", str(e))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    return result.summary()


def test_edge_cases():
    """Test edge cases like empty files, missing columns, etc."""
    print("\n--- Testing Edge Cases ---")
    result = TestResult()

    from core.compiler import preprocess_script_category
    import tempfile
    import os

    # Test 1: Empty qa_folders list
    try:
        universe = preprocess_script_category([], is_english=True)
        if universe["row_count"] == 0 and "errors" in universe:
            result.pass_test("Empty qa_folders handled gracefully")
        else:
            result.fail_test("Empty qa_folders handled gracefully", f"Got: {universe}")
    except Exception as e:
        result.fail_test("Empty qa_folders handled gracefully", str(e))

    # Test 2: Missing xlsx_path
    try:
        qa_folders = [{"username": "test", "category": "Test"}]  # Missing xlsx_path
        universe = preprocess_script_category(qa_folders, is_english=True)
        if universe["row_count"] == 0:
            result.pass_test("Missing xlsx_path handled gracefully")
        else:
            result.fail_test("Missing xlsx_path handled gracefully", f"Got row_count: {universe['row_count']}")
    except Exception as e:
        result.fail_test("Missing xlsx_path handled gracefully", str(e))

    # Test 3: Non-existent file
    try:
        qa_folders = [{
            "xlsx_path": Path("/nonexistent/file.xlsx"),
            "username": "test",
        }]
        universe = preprocess_script_category(qa_folders, is_english=True)
        if universe["row_count"] == 0:
            result.pass_test("Non-existent file handled gracefully")
        else:
            result.fail_test("Non-existent file handled gracefully", f"Got row_count: {universe['row_count']}")
    except Exception as e:
        result.fail_test("Non-existent file handled gracefully", str(e))

    # Test 4: Empty workbook (no sheets)
    try:
        # Create temp empty workbook
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        temp_path = Path(temp_path)

        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        # Create empty sheet
        wb.create_sheet("EmptySheet")
        wb.save(temp_path)
        wb.close()

        qa_folders = [{"xlsx_path": temp_path, "username": "test"}]
        universe = preprocess_script_category(qa_folders, is_english=True)

        if universe["row_count"] == 0:
            result.pass_test("Empty workbook handled gracefully")
        else:
            result.fail_test("Empty workbook handled gracefully", f"Got row_count: {universe['row_count']}")

        temp_path.unlink()

    except Exception as e:
        result.fail_test("Empty workbook handled gracefully", str(e))
        traceback.print_exc()

    # Test 5: Workbook with missing required columns
    try:
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        temp_path = Path(temp_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sequencer"
        # Only add some columns, not all required ones
        ws.cell(1, 1, "SomeColumn")
        ws.cell(1, 2, "AnotherColumn")
        ws.cell(2, 1, "Value1")
        ws.cell(2, 2, "Value2")
        wb.save(temp_path)
        wb.close()

        qa_folders = [{"xlsx_path": temp_path, "username": "test"}]
        universe = preprocess_script_category(qa_folders, is_english=True)

        if universe["row_count"] == 0:
            result.pass_test("Missing columns handled gracefully")
        else:
            result.fail_test("Missing columns handled gracefully", f"Got row_count: {universe['row_count']}")

        temp_path.unlink()

    except Exception as e:
        result.fail_test("Missing columns handled gracefully", str(e))
        traceback.print_exc()

    return result.summary()


def test_build_prefiltered_rows():
    """Test build_prefiltered_rows helper."""
    print("\n--- Testing build_prefiltered_rows ---")
    result = TestResult()

    from core.compiler import preprocess_script_category, build_prefiltered_rows

    fixture_path = Path(__file__).parent / "fixtures" / "mock_script_qa.xlsx"

    if not fixture_path.exists():
        print(f"  [ERROR] Fixture not found: {fixture_path}")
        return False

    qa_folders = [{
        "xlsx_path": fixture_path,
        "username": "test_user",
        "category": "Sequencer",
        "folder_path": fixture_path.parent,
        "images": [],
    }]

    try:
        universe = preprocess_script_category(qa_folders, is_english=True)

        # Test 1: Returns rows for valid user/file/sheet
        rows = build_prefiltered_rows(universe, fixture_path, "Sequencer", "test_user")
        if rows and isinstance(rows, list) and len(rows) > 0:
            result.pass_test(f"Returns {len(rows)} prefiltered rows for Sequencer")
        else:
            result.fail_test("Returns prefiltered rows for Sequencer", f"Got: {rows}")

        # Test 2: Returns None for non-existent user
        rows = build_prefiltered_rows(universe, fixture_path, "Sequencer", "nonexistent_user")
        if rows is None:
            result.pass_test("Returns None for non-existent user")
        else:
            result.fail_test("Returns None for non-existent user", f"Got: {rows}")

        # Test 3: Returns None for non-existent sheet
        rows = build_prefiltered_rows(universe, fixture_path, "NonExistentSheet", "test_user")
        if rows is None:
            result.pass_test("Returns None for non-existent sheet")
        else:
            result.fail_test("Returns None for non-existent sheet", f"Got: {rows}")

        # Test 4: Rows are sorted
        rows = build_prefiltered_rows(universe, fixture_path, "Sequencer", "test_user")
        if rows and rows == sorted(rows):
            result.pass_test("Rows are sorted in ascending order")
        else:
            result.fail_test("Rows are sorted in ascending order", f"Got: {rows}")

    except Exception as e:
        print(f"  [ERROR] Exception during test: {e}")
        traceback.print_exc()
        result.fail_test("No exceptions", str(e))

    return result.summary()


def main():
    """Run all tests."""
    print("=" * 60)
    print("QACompiler Script Preprocessing Tests")
    print("=" * 60)

    all_passed = True

    # Run each test suite
    all_passed &= test_preprocess_script_category()
    all_passed &= test_build_master_from_universe()
    all_passed &= test_build_prefiltered_rows()
    all_passed &= test_edge_cases()

    print("\n" + "=" * 60)
    if all_passed:
        print("ALL TESTS PASSED!")
    else:
        print("SOME TESTS FAILED - see above for details")
    print("=" * 60)

    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
