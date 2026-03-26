"""
Tests for tracker.masterfile_pending — masterfile-based pending issue extraction.

TDD: Written BEFORE the implementation module exists.
"""
from __future__ import annotations

import os
import sys
import time
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from tracker.masterfile_pending import build_pending_from_masterfiles


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_master(folder: Path, filename: str, sheets: dict, mtime: float | None = None):
    """
    Create a mock masterfile.

    sheets = {
        "SheetName": {
            "headers": ["Col1", "Col2", ...],
            "rows": [
                ["val1", "val2", ...],
                ...
            ],
        },
        ...
    }
    """
    wb = Workbook()
    first = True
    for sheet_name, spec in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for j, hdr in enumerate(spec["headers"], start=1):
            ws.cell(1, j, hdr)
        for i, row in enumerate(spec["rows"], start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)

    path = folder / filename
    wb.save(path)
    wb.close()

    if mtime is not None:
        os.utime(path, (mtime, mtime))
    return path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_dirs(tmp_path):
    en = tmp_path / "Masterfolder_EN"
    cn = tmp_path / "Masterfolder_CN"
    en.mkdir()
    cn.mkdir()
    return en, cn


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_single_master_one_tester(tmp_dirs):
    """1 masterfile, 1 tester, 4 rows with different statuses.

    Real masterfiles have comment columns after STATUS (COMMENT, MEMO, SCREENSHOT).
    ISSUE rows without any comment → phantom issue → auto-nonissue.
    """
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                # ISSUE + empty status + has comment -> pending
                ["KR1", "EN1", "S001", "ISSUE", None, "Typo in line 3", None, None],
                # ISSUE + FIXED + has comment -> fixed
                ["KR2", "EN2", "S002", "ISSUE", "FIXED", "Wrong word", None, None],
                # ISSUE + CHECKING + has comment -> pending + checking
                ["KR3", "EN3", "S003", "ISSUE", "CHECKING", "Grammar issue", None, None],
                # No ISSUE -> skip entirely
                ["KR4", "EN4", "S004", None, "FIXED", None, None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    alice = result["Alice"]["Quest"]
    assert alice["active_issues"] == 3
    # Row 1 (empty status, has comment) -> pending, Row 3 (CHECKING) -> also pending
    assert alice["pending"] == 2
    assert alice["fixed"] == 1
    assert alice["checking"] == 1
    assert alice["reported"] == 0
    assert alice["nonissue"] == 0


def test_phantom_issue_no_comment(tmp_dirs):
    """ISSUE without any comment in comment columns → auto-nonissue (phantom)."""
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                # ISSUE + no comment → phantom → nonissue
                ["KR1", "EN1", "S001", "ISSUE", None, None, None, None],
                # ISSUE + has comment → real pending
                ["KR2", "EN2", "S002", "ISSUE", None, "Real issue here", None, None],
                # ISSUE + FIXED + no comment → phantom → nonissue (even though FIXED)
                ["KR3", "EN3", "S003", "ISSUE", "FIXED", None, None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    alice = result["Alice"]["Quest"]
    assert alice["active_issues"] == 3
    assert alice["pending"] == 1   # only row 2 (has comment, empty status)
    assert alice["nonissue"] == 2  # rows 1 and 3 (no comment → phantom)
    assert alice["fixed"] == 0     # row 3 is phantom, not counted as fixed


def test_two_testers_same_master(tmp_dirs):
    """2 testers with independent columns in the same masterfile."""
    en, cn = tmp_dirs

    _make_master(en, "Master_Item.xlsx", {
        "Items": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
                "TESTER_STATUS_Bob", "STATUS_Bob",
                "COMMENT_Bob", "MEMO_Bob", "SCREENSHOT_Bob",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", "FIXED", "Wrong word", None, None, "ISSUE", "REPORTED", "Bad format", None, None],
                ["KR2", "EN2", "S002", "ISSUE", None, "Typo found", None, None, None, "FIXED", None, None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    alice = result["Alice"]["Item"]
    assert alice["active_issues"] == 2
    assert alice["fixed"] == 1
    assert alice["pending"] == 1  # row 2 empty status, has comment

    bob = result["Bob"]["Item"]
    assert bob["active_issues"] == 1  # only row 1 has ISSUE
    assert bob["reported"] == 1


def test_latest_file_wins_dedup(tmp_dirs):
    """2 Master_Quest.xlsx at different paths, only latest mtime used."""
    en, cn = tmp_dirs

    sub1 = en / "sub1"
    sub2 = en / "sub2"
    sub1.mkdir()
    sub2.mkdir()

    old_time = time.time() - 3600  # 1 hour ago
    new_time = time.time()

    # Old file: Alice has 2 issues
    _make_master(sub1, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", None, "Bug here", None, None],
                ["KR2", "EN2", "S002", "ISSUE", None, "Another bug", None, None],
            ],
        },
    }, mtime=old_time)

    # New file: Alice has 1 issue, fixed
    _make_master(sub2, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", "FIXED", "Was a typo", None, None],
            ],
        },
    }, mtime=new_time)

    result = build_pending_from_masterfiles(en, cn)

    alice = result["Alice"]["Quest"]
    assert alice["active_issues"] == 1
    assert alice["fixed"] == 1
    assert alice["pending"] == 0


def test_multi_sheet_master(tmp_dirs):
    """Master_System.xlsx with System + Help sheets, aggregated under 'System'.

    Note: Script and Face categories are EXCLUDED from active pending
    (same zeroing logic as total.py), so we test with System instead.
    """
    en, cn = tmp_dirs

    _make_master(en, "Master_System.xlsx", {
        "System": {
            "headers": [
                "StringID", "Text", "Translation",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["SYS1", "Setting1", "설정1", "ISSUE", "FIXED", "Wrong label", None, None],
                ["SYS2", "Setting2", "설정2", "ISSUE", None, "Missing text", None, None],
            ],
        },
        "Help": {
            "headers": [
                "StringID", "Text", "Translation",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["HLP1", "Help1", "도움말1", "ISSUE", "REPORTED", "Unclear help text", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    # Both sheets aggregate under "System" category
    alice = result["Alice"]["System"]
    assert alice["active_issues"] == 3
    assert alice["fixed"] == 1
    assert alice["pending"] == 1
    assert alice["reported"] == 1


def test_script_face_excluded(tmp_dirs):
    """Script and Face categories are excluded from active pending."""
    en, cn = tmp_dirs

    _make_master(en, "Master_Script.xlsx", {
        "Sequencer": {
            "headers": ["EventName", "Text", "TESTER_STATUS_Alice", "STATUS_Alice"],
            "rows": [["EVT1", "Hello", "ISSUE", ""]],
        },
    })
    _make_master(en, "Master_Face.xlsx", {
        "Face": {
            "headers": ["StringID", "Text", "TESTER_STATUS_Bob", "STATUS_Bob"],
            "rows": [["F1", "Face1", "ISSUE", ""]],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    # Both should be excluded — empty result
    assert result == {}, f"Script/Face should be excluded but got: {result}"


def test_empty_folders(tmp_dirs):
    """No masterfiles -> empty result."""
    en, cn = tmp_dirs
    result = build_pending_from_masterfiles(en, cn)
    assert result == {}


def test_skip_status_sheet(tmp_dirs):
    """STATUS sheet in masterfile is skipped."""
    en, cn = tmp_dirs

    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", None, "Real issue", None, None],
            ],
        },
        "STATUS": {
            "headers": [
                "TESTER_STATUS_Ghost", "STATUS_Ghost",
                "COMMENT_Ghost", "MEMO_Ghost", "SCREENSHOT_Ghost",
            ],
            "rows": [
                ["ISSUE", None, "Ghost comment", None, None],
                ["ISSUE", "FIXED", "Ghost fixed", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    # Ghost should NOT appear (STATUS sheet skipped)
    assert "Ghost" not in result
    # Alice from MainQuest should be there
    assert result["Alice"]["Quest"]["active_issues"] == 1


def test_non_issue_variants(tmp_dirs):
    """Both 'NON-ISSUE' and 'NON ISSUE' count as nonissue."""
    en, cn = tmp_dirs

    _make_master(en, "Master_Knowledge.xlsx", {
        "Knowledge": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", "NON-ISSUE", "Not a bug", None, None],
                ["KR2", "EN2", "S002", "ISSUE", "NON ISSUE", "Intended", None, None],
                ["KR3", "EN3", "S003", "ISSUE", "non-issue", "By design", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    alice = result["Alice"]["Knowledge"]
    assert alice["nonissue"] == 3
    assert alice["pending"] == 0


def test_en_and_cn_folders(tmp_dirs):
    """Files in both Masterfolder_EN and CN, merged correctly."""
    en, cn = tmp_dirs

    # EN folder: Quest with Alice
    _make_master(en, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ENG)", "STRINGID",
                "TESTER_STATUS_Alice", "STATUS_Alice",
                "COMMENT_Alice", "MEMO_Alice", "SCREENSHOT_Alice",
            ],
            "rows": [
                ["KR1", "EN1", "S001", "ISSUE", "FIXED", "Typo fixed", None, None],
            ],
        },
    })

    # CN folder: Quest with Bob (different tester)
    _make_master(cn, "Master_Quest.xlsx", {
        "MainQuest": {
            "headers": [
                "Korean", "Translation (ZHO-CN)", "STRINGID",
                "TESTER_STATUS_Bob", "STATUS_Bob",
                "COMMENT_Bob", "MEMO_Bob", "SCREENSHOT_Bob",
            ],
            "rows": [
                ["KR1", "CN1", "S001", "ISSUE", None, "Wrong translation", None, None],
                ["KR2", "CN2", "S002", "ISSUE", "REPORTED", "Grammar error", None, None],
            ],
        },
    })

    result = build_pending_from_masterfiles(en, cn)

    # Alice from EN
    assert result["Alice"]["Quest"]["active_issues"] == 1
    assert result["Alice"]["Quest"]["fixed"] == 1

    # Bob from CN
    assert result["Bob"]["Quest"]["active_issues"] == 2
    assert result["Bob"]["Quest"]["pending"] == 1
    assert result["Bob"]["Quest"]["reported"] == 1
