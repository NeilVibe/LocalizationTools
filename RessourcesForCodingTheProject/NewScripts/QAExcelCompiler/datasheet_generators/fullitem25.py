
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Item-data extractor / LQA helper – COMPLETE REBUILD v3.14
(NEW: ItemDesc from KnowledgeKey with fallback to ItemInfo.ItemDesc)

v3.14 Changes:
- ItemDesc priority: KnowledgeKey -> KnowledgeInfo.Desc
- ItemDesc fallback: ItemInfo.ItemDesc (if no KnowledgeKey or Desc not found)

v3.13 Changes:
- ItemName: from ItemInfo.ItemName (unchanged)
- ItemDesc: from KnowledgeKey -> KnowledgeInfo.Desc (NEW!)
"""

from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable, Set
from collections import defaultdict
from dataclasses import dataclass

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION – FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo\iteminfo"
)
ITEMGROUPINFO_FILE = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo\iteminfo\itemgroupinfo.staticinfo.xml"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"
)
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\cd_lambda\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)

# NEW: Knowledge folder for ItemDesc lookup via KnowledgeKey
KNOWLEDGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo\knowledgeinfo"
)


if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "ItemData_Map_All"
LOG_FILE = base_path / "item_scan.log"


TEXTFILES_FOLDER = OUTPUT_FOLDER / "ExecuteFiles"
PRIMARY_LQA_FOLDER = OUTPUT_FOLDER / "Item_Full_LQA"
SECONDARY_LQA_FOLDER = OUTPUT_FOLDER / "Item_Sorted_LQA"


# Clustering settings
MERGE_UP_THRESHOLD = 50       # Groups with < this many items merge into parent
FOLDER_MIN_THRESHOLD = 100    # Folders with < this many items merge into Others
MAX_COMMANDS_PER_FILE = 300   # Split files if larger
MIN_FOLDER_DEPTH = 1          # BLUE = minimum folder depth (ceiling)

# Special folder keys
OTHERS_KEY = "OTHERS"
MONSTER_ITEM_KEY = "MONSTER_ITEM"

# Monster substring for extraction (v3.10: substring check, not prefix)
MONSTER_SUBSTRING = "mon_"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING – FILE + CONSOLE
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("ItemLQA")
log.setLevel(logging.DEBUG)

# _file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
# _file_handler.setFormatter(
    # logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
# )
# _file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

# log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ─────────────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class ItemData:
    """Complete data for a single item."""
    strkey: str
    item_name: str       # KOR original (from ItemInfo.ItemName)
    item_desc: str       # KOR original (from KnowledgeKey -> KnowledgeInfo.Desc)
    group_key: str
    group_name_kor: str  # KOR original

# For Primary rows we now carry:
# depth, GroupKey,
# GroupName(KOR), GroupName(ENG), GroupName(LOC),
# ItemKey, Item#, ItemName(KOR), ItemName(ENG), ItemName(LOC),
# ItemDesc(KOR), ItemDesc(ENG), ItemDesc(LOC),
# StringID, DebugCommand, is_group
PrimaryRow = Tuple[
    int, str,
    str, str, str,
    str, str,
    str, str, str,
    str, str, str,
    str, str, bool
]

# ─────────────────────────────────────────────────────────────────────────────
# PLACEHOLDER NORMALIZATION / KOREAN DETECTION / FILENAME SANITIZATION
# ─────────────────────────────────────────────────────────────────────────────

_korean_re = re.compile(r'[\uAC00-\uD7AF]')
_unsafe_chars_re = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)   # matches any whitespace

def normalize_placeholders(text: str) -> str:
    """
    1) Remove '#…' suffix inside {...} placeholders.
    2) Collapse all whitespace (space, tab, NBSP, newline, etc.) to ONE space.
    3) Trim leading/trailing spaces.
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)   # step 1
    text = _whitespace_re.sub(' ', text).strip()        # step 2+3
    return text

def contains_korean(text: str) -> bool:
    return bool(_korean_re.search(text))

def is_good_translation(text: str) -> bool:
    return bool(text) and not contains_korean(text)

def sanitize_filename(name: str) -> str:
    if not name:
        return "UNKNOWN"
    clean = _unsafe_chars_re.sub('', name)
    clean = clean.replace(' ', '_')
    clean = re.sub(r'_+', '_', clean)
    clean = clean.strip('_')
    if len(clean) > 50:
        clean = clean[:50]
    return clean if clean else "UNKNOWN"

# ─────────────────────────────────────────────────────────────────────────────
# XML UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def fix_bad_entities(xml_text: str) -> str:
    return _bad_entity_re.sub("&amp;", xml_text)

def _preprocess_newlines(raw_content: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1)
        inner = inner.replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw_content, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(
        r'="([^"]*<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*&[^ltgapoqu][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )
    tag_stack: List[str] = []
    tag_open = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    fixed_lines: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            tag_stack.append(mo.group(1))
            fixed_lines.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if tag_stack and tag_stack[-1] == mc.group(1):
                tag_stack.pop()
                fixed_lines.append(line)
            else:
                if tag_stack:
                    fixed_lines.append(f"</{tag_stack.pop()}>")
                else:
                    fixed_lines.append(line)
            continue
        if stripped.startswith("</>"):
            if tag_stack:
                fixed_lines.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            else:
                fixed_lines.append(line)
            continue
        fixed_lines.append(line)
    while tag_stack:
        fixed_lines.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed_lines)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    parser_strict = ET.XMLParser(huge_tree=True)
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=parser_strict)
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retrying with recovery", path.name)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Even recovery parse failed: %s", path)
            return None

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn

def autofit(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len * 1.15 + 2, max_width)

# ─────────────────────────────────────────────────────────────────────────────
# HIERARCHY FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def calc_depth(node: str, parent_of: Dict[str, str]) -> int:
    d = 0
    visited: Set[str] = set()
    while node in parent_of and parent_of[node]:
        if node in visited:
            break
        visited.add(node)
        node = parent_of[node]
        d += 1
    return d

def get_depth_color(depth: int) -> str:
    if depth == 0:
        return "YELLOW"
    elif depth == 1:
        return "BLUE"
    elif depth == 2:
        return "GREEN"
    else:
        return "RED"

def build_children_map(parent_of: Dict[str, str]) -> Dict[str, List[str]]:
    children: Dict[str, List[str]] = {}
    for child, parent in parent_of.items():
        children.setdefault(parent, []).append(child)
    return children

# ─────────────────────────────────────────────────────────────────────────────
# DISPLAY NAME FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def get_display_name(
    group_key: str,
    group_names: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> str:
    """Get display name: ENG preferred, else raw group_key."""
    if group_key == OTHERS_KEY:
        return "Others"
    if group_key == MONSTER_ITEM_KEY:
        return "Monster_Item"

    # try to pull an English translation
    kor_name = group_names.get(group_key, "")
    if kor_name:
        normalized = normalize_placeholders(kor_name)
        eng_name = eng_tbl.get(normalized, ("", ""))[0]
        if eng_name and is_good_translation(eng_name):
            return sanitize_filename(eng_name)

    # fallback to raw key (strip common prefix)
    if group_key.lower().startswith("itemgroup_"):
        fallback = group_key[10:]
    else:
        fallback = group_key

    return sanitize_filename(fallback)

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ─────────────────────────────────────────────────────────────────────────────
def load_single_language(folder: Path, lang_code: str) -> Dict[str, Tuple[str, str]]:
    tbl: Dict[str, Tuple[str, str]] = {}
    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith(f"languagedata_{lang_code}"):
            continue
        log.info("Loading language file [%s] – %s", lang_code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            continue
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            tr = el.get("Str") or ""
            sid = el.get("StringId") or ""
            if not origin:
                continue
            normalized_origin = normalize_placeholders(origin)
            if normalized_origin in tbl:
                existing_tr, _ = tbl[normalized_origin]
                if is_good_translation(tr) and not is_good_translation(existing_tr):
                    tbl[normalized_origin] = (tr, sid)
            else:
                tbl[normalized_origin] = (tr, sid)
    log.info("Language %s loaded – %d entries", lang_code.upper(), len(tbl))
    return tbl

def parse_language_folder(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    log.info("Scanning language folder: %s", folder)
    langs: Dict[str, Dict[str, Tuple[str, str]]] = {}
    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_") or stem.startswith("languagedata_kor"):
            continue
        code = stem.split("_", 1)[1].lower()
        if code in langs:
            continue
        log.info("Loading language file [%s] – %s", code.upper(), path.name)
        root = parse_xml_file(path)
        if root is None:
            continue
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            tr = el.get("Str") or ""
            sid = el.get("StringId") or ""
            if not origin:
                continue
            normalized_origin = normalize_placeholders(origin)
            if normalized_origin in tbl:
                existing_tr, _ = tbl[normalized_origin]
                if is_good_translation(tr) and not is_good_translation(existing_tr):
                    tbl[normalized_origin] = (tr, sid)
            else:
                tbl[normalized_origin] = (tr, sid)
        langs[code] = tbl
        log.info("Language %s loaded – %d entries", code.upper(), len(tbl))
    log.info("Language tables loaded: %d", len(langs))
    return langs

# ─────────────────────────────────────────────────────────────────────────────
# STRING-KEY TABLE
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Aborting – cannot parse StringKeyTable")
    tbl: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        num = el.get("Key") or ""
        sk = el.get("StrKey") or ""
        if num and sk:
            tbl[sk.lower()] = num
    log.info("StringKeyTable entries: %d", len(tbl))
    return tbl

# ─────────────────────────────────────────────────────────────────────────────
# KNOWLEDGE DESCRIPTIONS (NEW in v3.13)
# ─────────────────────────────────────────────────────────────────────────────
def load_knowledge_descriptions(folder: Path) -> Dict[str, str]:
    """
    Parse knowledge XML files and build KnowledgeKey -> Desc mapping.

    Looks for KnowledgeInfo elements with StrKey and Desc attributes.
    Example:
        <KnowledgeInfo StrKey="Knowledge_Collection_Prop_Viewingstone"
                       Name="수석"
                       Desc="바람과 물, 그리고 오랜 세월이 빚어낸..." />

    Returns:
        Dict mapping KnowledgeKey (StrKey) to Desc text
    """
    log.info("Loading knowledge descriptions from: %s", folder)
    knowledge_map: Dict[str, str] = {}

    if not folder.exists():
        log.warning("Knowledge folder does not exist: %s", folder)
        return knowledge_map

    file_count = 0
    for path in iter_xml_files(folder):
        root = parse_xml_file(path)
        if root is None:
            continue
        file_count += 1

        # Find all KnowledgeInfo elements
        for el in root.iter("KnowledgeInfo"):
            strkey = el.get("StrKey") or ""
            desc = el.get("Desc") or ""

            if strkey and strkey not in knowledge_map:
                # Store the description (can be empty if no Desc attribute)
                knowledge_map[strkey] = desc

    log.info("Knowledge descriptions loaded: %d entries from %d files",
             len(knowledge_map), file_count)
    return knowledge_map

# ─────────────────────────────────────────────────────────────────────────────
# MASTER ITEM-GROUP INFO – FIRST OCCURRENCE WINS
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_groups(path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    log.info("Parsing ItemGroupInfo master: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse ItemGroupInfo master file")
    group_name: Dict[str, str] = {}
    parent_of: Dict[str, str] = {}
    duplicate_count = 0
    for el in root.iter("ItemGroupInfo"):
        sk = el.get("StrKey") or ""
        name = el.get("GroupName") or ""
        if sk:
            if sk not in group_name:
                group_name[sk] = name
            p = el.getparent()
            if p is not None and p.tag == "ItemGroupInfo":
                parent_key = p.get("StrKey") or ""
                if parent_key and sk not in parent_of:
                    parent_of[sk] = parent_key
                elif parent_key:
                    duplicate_count += 1
    for el in root.iter("ChildGroupInfo"):
        child = el.get("StrKey") or ""
        name = el.get("GroupName") or ""
        p = el.getparent()
        if child:
            if name and child not in group_name:
                group_name[child] = name
            if p is not None and p.tag == "ItemGroupInfo":
                parent_key = p.get("StrKey") or ""
                if parent_key and child not in parent_of:
                    parent_of[child] = parent_key
                elif parent_key:
                    duplicate_count += 1
    log.info(
        "Groups parsed: %d  |  parent links: %d  |  duplicates ignored: %d",
        len(group_name), len(parent_of), duplicate_count
    )
    return group_name, parent_of

# ─────────────────────────────────────────────────────────────────────────────
# RESOURCE SCAN (UPDATED in v3.13 - uses KnowledgeKey for ItemDesc)
# ─────────────────────────────────────────────────────────────────────────────
def scan_resource_folder(
    folder: Path,
    knowledge_desc_map: Dict[str, str]
) -> Tuple[Dict[str, List[Tuple[str, str, str]]], Dict[str, str]]:
    """
    Scan resource folder for items.

    NEW in v3.13:
    - ItemName: from ItemInfo.ItemName (unchanged)
    - ItemDesc: from KnowledgeKey -> knowledge_desc_map lookup
    - If no KnowledgeKey or KnowledgeKey not found, ItemDesc is BLANK

    Args:
        folder: Path to item resource folder
        knowledge_desc_map: Dict mapping KnowledgeKey -> Desc

    Returns:
        Tuple of (group_items, scanned_group_names)
    """
    log.info("Scanning resource folder for items: %s", folder)
    group_items: Dict[str, List[Tuple[str, str, str]]] = {}
    scanned_group_names: Dict[str, str] = {}

    # Stats for logging
    items_with_knowledge = 0
    items_without_knowledge = 0

    for idx, path in enumerate(iter_xml_files(folder), 1):
        root = parse_xml_file(path)
        if root is None:
            continue
        for g_el in root.iter("ItemGroupInfo"):
            g_key = g_el.get("StrKey") or ""
            g_name = g_el.get("GroupName") or ""
            if not g_key:
                continue
            if g_name and g_key not in scanned_group_names:
                scanned_group_names[g_key] = g_name
            bucket = group_items.setdefault(g_key, [])
            for item in g_el.iter("ItemInfo"):
                ik = item.get("StrKey") or ""
                name = item.get("ItemName") or ""

                # Get description: KnowledgeKey priority, fallback to ItemDesc
                knowledge_key = item.get("KnowledgeKey") or ""
                item_desc_attr = item.get("ItemDesc") or ""

                if knowledge_key:
                    # Priority: Look up description from KnowledgeKey
                    desc = knowledge_desc_map.get(knowledge_key, "")
                    if desc:
                        items_with_knowledge += 1
                    else:
                        # KnowledgeKey exists but no Desc found, fallback to ItemDesc
                        desc = item_desc_attr
                        items_without_knowledge += 1
                else:
                    # No KnowledgeKey: fallback to ItemDesc attribute
                    desc = item_desc_attr
                    items_without_knowledge += 1

                if ik:
                    bucket.append((ik, name, desc))
        if idx % 200 == 0:
            log.debug("... scanned %d XML files", idx)

    total_items = sum(len(v) for v in group_items.values())
    log.info(
        "Group-item mapping built: Groups=%d  Total items=%d",
        len(group_items), total_items,
    )
    log.info(
        "  Items with KnowledgeKey desc: %d  |  Items using ItemDesc fallback: %d",
        items_with_knowledge, items_without_knowledge
    )
    return group_items, scanned_group_names

# ─────────────────────────────────────────────────────────────────────────────
# DEPTH-BASED CLUSTERING
# ─────────────────────────────────────────────────────────────────────────────
def apply_depth_based_clustering(
    group_items: Dict[str, List[Tuple[str, str, str]]],
    parent_of: Dict[str, str],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> Dict[str, Dict[str, List[str]]]:
    log.info("Applying depth-based clustering (threshold: %d items)", MERGE_UP_THRESHOLD)
    group_item_count: Dict[str, int] = {}
    group_depth: Dict[str, int] = {}
    for group_key, items_list in group_items.items():
        group_item_count[group_key] = len(items_list)
        group_depth[group_key] = calc_depth(group_key, parent_of)

    log.info("Initial group counts by depth:")
    depth_summary: Dict[int, Tuple[int, int]] = defaultdict(lambda: (0, 0))
    for gk, count in group_item_count.items():
        d = group_depth[gk]
        g, i = depth_summary[d]
        depth_summary[d] = (g + 1, i + count)
    for d in sorted(depth_summary):
        groups, items = depth_summary[d]
        log.info("  Depth %d (%s): %d groups, %d items", d, get_depth_color(d), groups, items)

    merge_target: Dict[str, str] = {gk: gk for gk in group_items}
    groups_by_depth = sorted(group_items.keys(), key=lambda g: -group_depth.get(g, 0))
    accumulated_counts = dict(group_item_count)
    merge_count = skip_at_ceiling = skip_above_threshold = skip_no_parent = 0

    for gk in groups_by_depth:
        depth = group_depth[gk]
        count = accumulated_counts.get(gk, 0)
        if depth <= MIN_FOLDER_DEPTH:
            skip_at_ceiling += 1
            continue
        if count < MERGE_UP_THRESHOLD:
            parent = parent_of.get(gk, "")
            if parent:
                pdepth = group_depth.get(parent, 0)
                if pdepth < depth:
                    merge_target[gk] = parent
                    merge_count += 1
                    accumulated_counts[parent] = accumulated_counts.get(parent, 0) + count
                    log.debug("  MERGE: %s→%s", get_display_name(gk, group_names, eng_tbl),
                              get_display_name(parent, group_names, eng_tbl))
            else:
                skip_no_parent += 1
        else:
            skip_above_threshold += 1

    log.info("Merge: %d, at ceiling: %d, ≥thr: %d, no parent: %d",
             merge_count, skip_at_ceiling, skip_above_threshold, skip_no_parent)

    def resolve(gk: str, seen: Set[str]) -> str:
        if gk in seen:
            return gk
        tgt = merge_target.get(gk, gk)
        if tgt == gk:
            return gk
        seen.add(gk)
        return resolve(tgt, seen)

    for gk in merge_target:
        merge_target[gk] = resolve(gk, set())

    def find_folder_at_blue(gk: str) -> str:
        cur = gk
        seen: Set[str] = set()
        while True:
            if cur in seen:
                return OTHERS_KEY
            seen.add(cur)
            d = calc_depth(cur, parent_of)
            if d == 1:
                return cur
            if d == 0:
                return OTHERS_KEY
            cur = parent_of.get(cur, "")
            if not cur:
                return OTHERS_KEY

    structure: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for group_key, items_list in group_items.items():
        sub = merge_target[group_key]
        folder = find_folder_at_blue(sub)
        for ik, _, _ in items_list:
            structure[folder][sub].append(ik)

    log.info("Final structure:")
    for fk, subs in structure.items():
        log.info("  %s: %d subgroups, %d items",
                 get_display_name(fk, group_names, eng_tbl), len(subs),
                 sum(len(it) for it in subs.values()))
    return dict(structure)

# ─────────────────────────────────────────────────────────────────────────────
# MONSTER ITEM EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
def extract_monster_items(
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> Dict[str, Dict[str, List[str]]]:
    log.info("Extracting mon_ items to Monster_Item")
    new_struct: Dict[str, Dict[str, List[str]]] = {}
    monster: Dict[str, List[str]] = {}
    for fk, subs in structure.items():
        keep: Dict[str, List[str]] = {}
        for sk, items in subs.items():
            if MONSTER_SUBSTRING in sk.lower():
                monster[sk] = items
                log.debug("  Extracted %s", sk)
            else:
                keep[sk] = items
        if keep:
            new_struct[fk] = keep
    if monster:
        new_struct[MONSTER_ITEM_KEY] = monster
        log.info("  %d subgroups moved to Monster_Item", len(monster))
    return new_struct

# ─────────────────────────────────────────────────────────────────────────────
# SMALL FOLDER CONSOLIDATION
# ─────────────────────────────────────────────────────────────────────────────
def consolidate_small_folders(
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> Dict[str, Dict[str, List[str]]]:
    log.info("Consolidating small folders (<%d)", FOLDER_MIN_THRESHOLD)
    totals = {fk: sum(len(it) for it in subs.values()) for fk, subs in structure.items()}
    small = []
    keep = []
    special = []
    for fk, tot in totals.items():
        if fk in (OTHERS_KEY, MONSTER_ITEM_KEY):
            special.append(fk)
        elif tot < FOLDER_MIN_THRESHOLD:
            small.append(fk)
            log.info("  Small: %s (%d)→Others",
                     get_display_name(fk, group_names, eng_tbl), tot)
        else:
            keep.append(fk)
    if not small:
        log.info("  None to consolidate")
        return structure

    new_struct: Dict[str, Dict[str, List[str]]] = {}
    for fk in keep + special:
        new_struct[fk] = structure[fk]

    others_sub: Dict[str, List[str]] = {}
    if OTHERS_KEY in new_struct:
        others_sub = dict(new_struct[OTHERS_KEY])
    for fk in small:
        for sk, items in structure[fk].items():
            others_sub.setdefault(sk, []).extend(items)
    if others_sub:
        new_struct[OTHERS_KEY] = others_sub
        log.info("  Others now %d subgroups, %d items",
                 len(others_sub), sum(len(it) for it in others_sub.values()))
    return new_struct

# ─────────────────────────────────────────────────────────────────────────────
# BUILD ITEM DATA
# ─────────────────────────────────────────────────────────────────────────────
def build_item_data(
    group_items: Dict[str, List[Tuple[str, str, str]]],
    group_names: Dict[str, str],
) -> Dict[str, ItemData]:
    items: Dict[str, ItemData] = {}
    for gk, lst in group_items.items():
        kor = group_names.get(gk, "")
        for ik, iname, idesc in lst:
            items[ik] = ItemData(
                strkey=ik,
                item_name=iname,
                item_desc=idesc,
                group_key=gk,
                group_name_kor=kor,
            )
    log.info("Built data for %d items", len(items))
    return items

# ─────────────────────────────────────────────────────────────────────────────
# PRIMARY EXCEL FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def propagate_group_names(rows: List[PrimaryRow]) -> List[PrimaryRow]:
    result: List[PrimaryRow] = []
    last = ["", "", ""]
    for row in rows:
        depth, gk, gkor, geng, gloc, ik, inum, nkor, neng, nloc, dkor, deng, dloc, sid, is_group = row
        if gkor:
            last = [gkor, geng, gloc]
        else:
            gkor, geng, gloc = last
        result.append((depth, gk, gkor, geng, gloc,
                       ik, inum, nkor, neng, nloc, dkor, deng, dloc, sid, is_group))
    return result

def build_rows_for_language(
    code: str,
    group_names: Dict[str, str],
    parent_of: Dict[str, str],
    group_items: Dict[str, List[Tuple[str, str, str]]],
    lang_tbl: Dict[str, Tuple[str, str]],
    id_table: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> List[PrimaryRow]:
    log.info("Building rows for %s", code.upper())
    def t(tbl: Dict[str, Tuple[str, str]], text: str) -> str:
        if not text:
            return ""
        return tbl.get(normalize_placeholders(text), ("", ""))[0]
    rows: List[PrimaryRow] = []
    children_of = build_children_map(parent_of)
    all_groups = set(group_items)
    master_roots = sorted(g for g in group_names if g not in parent_of)
    seen: Set[str] = set()

    def recurse(gk: str):
        if gk in seen:
            return
        seen.add(gk)
        depth = calc_depth(gk, parent_of)
        kor = group_names.get(gk, "")
        eng = t(eng_tbl, kor)
        loc = t(lang_tbl, kor)
        rows.append((depth, gk, kor, eng, loc,
                     "", "", "", "", "", "", "", "", "", True))
        for ik, iname, idesc in sorted(group_items.get(gk, []), key=lambda x: x[0]):
            ieng = t(eng_tbl, iname)
            iloc = t(lang_tbl, iname)
            deng = t(eng_tbl, idesc)
            dloc = t(lang_tbl, idesc)
            num = id_table.get(ik.lower(), "<MISSING>")
            sid = lang_tbl.get(normalize_placeholders(iname), ("", ""))[1]
            dbg = f"/create item {ik}"
            rows.append((depth+1, gk, kor, eng, loc,
                         ik, num, iname, ieng, iloc, idesc, deng, dloc, sid, False))
        for child in sorted(children_of.get(gk, [])):
            recurse(child)

    for root in master_roots:
        recurse(root)
    for orphan in sorted(all_groups - seen):
        recurse(orphan)

    log.info("Total rows for %s: %d", code.upper(), len(rows))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ─────────────────────────────────────────────────────────────────────────────
_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),  # YELLOW
    1: PatternFill("solid", fgColor="D9E1F2"),  # BLUE
    2: PatternFill("solid", fgColor="E2EFDA"),  # GREEN
}
_item_fill = PatternFill("solid", fgColor="FCE4D6")
_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def write_primary_sheet(
    wb: Workbook,
    lang_code: str,
    rows: List[PrimaryRow],
    hide_depth_col: bool = True,
) -> None:
    """
    Create the primary LQA sheet for one language.
    Adds extra columns: STATUS / COMMENT / STRINGID / SCREENSHOT
    STATUS column has drop-down list: ISSUE / NO ISSUE / BLOCKED (or blank)
    """
    def _add_status_validation(sh, status_col_idx: int, max_row: int) -> None:
        """Adds drop-down list to STATUS column."""
        col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED,KOREAN"',
            allow_blank=True,
            showErrorMessage=True,
        )
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)
        sh.add_data_validation(dv)

    title = lang_code.upper()
    ws = wb.create_sheet(title=title[:31])

    # -------------------- Build headers --------------------
    headers = [
        "Depth",
        "GroupKey",
        "GroupName(KOR)",
        "GroupName(ENG)",
        "ItemKey",
        "Item#",
        "ItemName(KOR)",
        "ItemName(ENG)",
    ]
    if lang_code != "eng":
        headers.append(f"ItemName({title})")

    headers += ["ItemDesc(KOR)", "ItemDesc(ENG)"]
    if lang_code != "eng":
        headers.append(f"ItemDesc({title})")

    headers += ["StringID", "DebugCommand"]

    # Extra columns
    headers += ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]

    # Write header row
    for col, txt in enumerate(headers, start=1):
        cell = ws.cell(1, col, txt)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border

    # -------------------- Write data rows (with deduplication) --------------------
    is_eng = lang_code.lower() == "eng"
    seen_keys = set()
    duplicates_removed = 0
    r = 2

    for row in rows:
        (
            depth, gk, gkor, geng, gloc,
            ik, num, nkor, neng, nloc,
            dkor, deng, dloc, sid, is_group
        ) = row

        # Deduplication: skip if (Korean, Translation, STRINGID) already seen
        # nkor = ItemName(KOR), neng = ItemName(ENG), nloc = ItemName(LOC)
        trans = neng if is_eng else nloc
        dedup_key = (nkor, trans, sid)
        if dedup_key in seen_keys:
            duplicates_removed += 1
            continue
        seen_keys.add(dedup_key)

        vals = [depth, gk, gkor, geng]
        vals += [ik, num, nkor, neng]
        if lang_code != "eng":
            vals.append(nloc)
        vals += [dkor, deng]
        if lang_code != "eng":
            vals.append(dloc)
        vals += [sid, f"/create item {ik}" if not is_group else ""]

        # Extra columns default values
        vals += ["", "", sid, ""]

        for cidx, val in enumerate(vals, start=1):
            c = ws.cell(r, cidx, val)
            if isinstance(val, int):
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(
                    indent=(depth if cidx == 1 else 0),
                    wrap_text=True,
                    vertical="top",
                )
            c.border = _border
            if is_group:
                c.fill = _depth_fill.get(depth, _depth_fill[2])
                if cidx in (3, 4):
                    c.font = _bold_font
            else:
                c.fill = _item_fill
        r += 1

    if duplicates_removed > 0:
        log.info("    Removed %d duplicate rows (Korean+Translation+STRINGID)", duplicates_removed)

    # -------------------- Sheet cosmetics --------------------
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{r-1}"
    ws.freeze_panes = "A2"
    autofit(ws)

    if hide_depth_col:
        ws.column_dimensions["A"].hidden = True

    # -------------------- Add STATUS drop-down --------------------
    status_col_idx = headers.index("STATUS") + 1
    _add_status_validation(ws, status_col_idx, ws.max_row)

    # Force STRINGID column to text format (prevents scientific notation)
    stringid_col_idx = headers.index("STRINGID") + 1
    for row in range(2, r):
        ws.cell(row, stringid_col_idx).number_format = '@'

# ─────────────────────────────────────────────────────────────────────────────
# TEXT FILE GENERATION
# ─────────────────────────────────────────────────────────────────────────────
def write_text_files(
    out_folder: Path,
    structure: Dict[str, Dict[str, List[str]]],
    group_names: Dict[str, str],
    eng_tbl: Dict[str, Tuple[str, str]],
) -> Dict[str, List[Tuple[str, str, List[str]]]]:
    out_folder.mkdir(parents=True, exist_ok=True)
    folder_files: Dict[str, List[Tuple[str, str, List[str]]]] = {}

    for fk, subs in sorted(structure.items()):
        folder_display = get_display_name(fk, group_names, eng_tbl) or fk
        folder_path = out_folder / folder_display
        folder_path.mkdir(parents=True, exist_ok=True)
        file_list: List[Tuple[str, str, List[str]]] = []

        if fk == OTHERS_KEY:
            all_items = sorted({ik for lst in subs.values() for ik in lst})
            chunks = [all_items[i:i+MAX_COMMANDS_PER_FILE]
                      for i in range(0, len(all_items), MAX_COMMANDS_PER_FILE)]
            for idx, chunk in enumerate(chunks, 1):
                fn = f"Others_{idx}.txt"
                with open(folder_path/fn, "w", encoding="utf-8") as f:
                    f.write("/reset inventory\n/expandinventory 2 300\n")
                    for ik in chunk:
                        f.write(f"/create item {ik}\n")
                    f.write("/reset alert\n")
                file_list.append((fn, OTHERS_KEY, chunk))
            log.info("Folder %s: %d files (%d items)",
                     folder_display, len(chunks), len(all_items))
        else:
            disp_map: Dict[str, List[str]] = {}
            key_map: Dict[str, str] = {}
            for sk, lst in subs.items():
                disp = get_display_name(sk, group_names, eng_tbl) or sk
                disp_map.setdefault(disp, []).extend(lst)
                key_map.setdefault(disp, sk)
            for disp, all_iks in sorted(disp_map.items()):
                chunked = [all_iks[i:i+MAX_COMMANDS_PER_FILE]
                           for i in range(0, len(all_iks), MAX_COMMANDS_PER_FILE)]
                for idx, chunk in enumerate(chunked, 1):
                    fn = disp + (f"_{idx}.txt" if len(chunked)>1 else ".txt")
                    with open(folder_path/fn, "w", encoding="utf-8") as f:
                        f.write("/reset inventory\n/expandinventory 2 300\n")
                        for ik in sorted(set(chunk)):
                            f.write(f"/create item {ik}\n")
                        f.write("/reset alert\n")
                    file_list.append((fn, key_map[disp], chunk))
            log.info("Folder %s: %d text files", folder_display, len(file_list))

        folder_files[folder_display] = file_list
    return folder_files

# ─────────────────────────────────────────────────────────────────────────────
# SECONDARY EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────
def write_secondary_excel(
    out_path: Path,
    folder_files: Dict[str, List[Tuple[str, str, List[str]]]],
    items: Dict[str, ItemData],
    group_names: Dict[str, str],
    lang_tbl: Dict[str, Tuple[str, str]],
    lang_code: str,
    eng_tbl: Dict[str, Tuple[str, str]],
) -> None:
    """
    Secondary Excel with STATUS column drop-down list: ISSUE / NO ISSUE / BLOCKED (or blank)
    """
    def _add_status_validation(sh, status_col_idx: int, max_row: int) -> None:
        """Adds drop-down list to STATUS column."""
        col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED,KOREAN"',
            allow_blank=True,
            showErrorMessage=True,
        )
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)
        sh.add_data_validation(dv)

    wb = Workbook()
    wb.remove(wb.active)
    code = lang_code.upper()

    def t(tbl: Dict[str, Tuple[str, str]], text: str) -> str:
        return tbl.get(normalize_placeholders(text or ""), ("", ""))[0]

    width_map = {
        "Filename": 33,
        "ItemName(KOR)": 30,
        "ItemDesc(KOR)": 40,
        "ItemName(ENG)": 30,
        "ItemDesc(ENG)": 40,
        f"ItemName({code})": 30,
        f"ItemDesc({code})": 40,
        "STATUS": 15,
        "COMMENT": 50,
        "STRINGID": 25,
        "SCREENSHOT": 25,
    }

    for folder_display, flist in sorted(folder_files.items()):
        if not flist:
            continue
        base = folder_display[:31]
        title = base
        cnt = 1
        while title in {ws.title for ws in wb.worksheets}:
            title = f"{base[:28]}_{cnt}"
            cnt += 1
        ws = wb.create_sheet(title)

        headers = [
            "Filename",
            "SubGroup",
            "ItemName(KOR)", "ItemDesc(KOR)",
            "ItemName(ENG)", "ItemDesc(ENG)",
        ]
        if lang_code != "eng":
            headers += [f"ItemName({code})", f"ItemDesc({code})"]

        headers += ["ItemKey", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]

        # Write header row
        for col_idx, txt in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, txt)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border

        hide_cols = {"ItemKey"} if lang_code == "eng" else {"ItemKey", "ItemName(ENG)", "ItemDesc(ENG)"}
        for idx, h in enumerate(headers, 1):
            if h in hide_cols:
                ws.column_dimensions[get_column_letter(idx)].hidden = True

        rows_accum: List[Tuple[str, str, List[str]]] = []
        seen_keys = set()
        duplicates_removed = 0
        is_eng = lang_code.lower() == "eng"

        for fn, subgroup_key, iks in flist:
            for ik in sorted(iks):
                itm = items.get(ik)
                if not itm:
                    continue
                sub_disp = get_display_name(itm.group_key, group_names, eng_tbl)
                data_map = {
                    "Filename": fn,
                    "SubGroup": sub_disp,
                    "ItemName(KOR)": itm.item_name,
                    "ItemDesc(KOR)": itm.item_desc,
                    "ItemName(ENG)": t(eng_tbl, itm.item_name),
                    "ItemDesc(ENG)": t(eng_tbl, itm.item_desc),
                    "ItemKey": ik,
                    "STATUS": "",
                    "COMMENT": "",
                    "STRINGID": lang_tbl.get(normalize_placeholders(itm.item_name), ("", ""))[1],
                    "SCREENSHOT": "",
                }
                if lang_code != "eng":
                    data_map[f"ItemName({code})"] = t(lang_tbl, itm.item_name)
                    data_map[f"ItemDesc({code})"] = t(lang_tbl, itm.item_desc)

                # Deduplication: skip if (Korean, Translation, STRINGID) already seen
                korean = data_map["ItemName(KOR)"]
                trans = data_map["ItemName(ENG)"] if is_eng else data_map.get(f"ItemName({code})", "")
                sid = data_map["STRINGID"]
                dedup_key = (korean, trans, sid)
                if dedup_key in seen_keys:
                    duplicates_removed += 1
                    continue
                seen_keys.add(dedup_key)

                row_vals = [data_map.get(h, "") for h in headers]
                rows_accum.append((sub_disp, ik, row_vals))

        if duplicates_removed > 0:
            log.info("    Removed %d duplicate rows (Korean+Translation+STRINGID)", duplicates_removed)

        rows_accum.sort(key=lambda x: (x[0], x[1]))
        fill_a = PatternFill("solid", fgColor="E2EFDA")
        fill_b = PatternFill("solid", fgColor="FCE4D6")
        current_fill = fill_a
        last_sub = None
        excel_row = 2

        for sub, _, row_vals in rows_accum:
            if last_sub is not None and sub != last_sub:
                current_fill = fill_b if current_fill == fill_a else fill_a
            last_sub = sub
            for col_idx, val in enumerate(row_vals, 1):
                header = headers[col_idx - 1]
                cell = ws.cell(excel_row, col_idx, val)
                cell.fill = current_fill
                cell.border = _border
                if header == "STATUS":
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = _bold_font
            excel_row += 1

        for idx, header in enumerate(headers, 1):
            width = width_map.get(header, 20)
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row-1}"
        ws.freeze_panes = "A2"

        # -------------------- Add STATUS drop-down --------------------
        status_col_idx = headers.index("STATUS") + 1
        _add_status_validation(ws, status_col_idx, ws.max_row)

        # Force STRINGID column to text format (prevents scientific notation)
        stringid_col_idx = headers.index("STRINGID") + 1
        for row in range(2, excel_row):
            ws.cell(row, stringid_col_idx).number_format = '@'

        log.info("  Sheet '%s': %d rows", title, excel_row-2)

    wb.save(out_path)
    log.info("Secondary Excel saved: %s", out_path.name)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("="*70)
    log.info("Item LQA Extractor – v3.13 (KnowledgeKey-based ItemDesc)")
    log.info("="*70)

    OUTPUT_FOLDER.mkdir(exist_ok=True)
    TEXTFILES_FOLDER.mkdir(exist_ok=True)
    PRIMARY_LQA_FOLDER.mkdir(exist_ok=True)
    SECONDARY_LQA_FOLDER.mkdir(exist_ok=True)
    log.info("Output: %s", OUTPUT_FOLDER)

    # Load English for display names
    eng_tbl = load_single_language(LANGUAGE_FOLDER, "eng")
    if not eng_tbl:
        log.warning("No English table → using raw KOR names")

    lang_tables = parse_language_folder(LANGUAGE_FOLDER)
    if not lang_tables:
        sys.exit("No language files found!")
    id_tbl = load_string_key_table(STRINGKEYTABLE_FILE)
    master_names, parent_of = parse_master_groups(ITEMGROUPINFO_FILE)

    # NEW: Load knowledge descriptions BEFORE scanning resources
    knowledge_desc_map = load_knowledge_descriptions(KNOWLEDGE_FOLDER)

    # Pass knowledge_desc_map to scan_resource_folder
    group_items, scanned_names = scan_resource_folder(RESOURCE_FOLDER, knowledge_desc_map)

    all_names = {**scanned_names, **master_names}
    all_names[OTHERS_KEY] = "Others"
    all_names[MONSTER_ITEM_KEY] = "Monster_Item"
    log.info("Total group names: %d", len(all_names))

    items = build_item_data(group_items, all_names)

    struct = apply_depth_based_clustering(group_items, parent_of, all_names, eng_tbl)
    struct = extract_monster_items(struct, all_names, eng_tbl)
    struct = consolidate_small_folders(struct, all_names, eng_tbl)

    log.info("Writing text files…")
    folder_files = write_text_files(TEXTFILES_FOLDER, struct, all_names, eng_tbl)

    log.info("Processing languages…")
    total = len(lang_tables)
    for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
        log.info("(%d/%d) Language %s", idx, total, code.upper())

        rows = build_rows_for_language(code, all_names, parent_of,
                                       group_items, tbl, id_tbl, eng_tbl)
        rows = propagate_group_names(rows)

        wb = Workbook()
        wb.remove(wb.active)
        write_primary_sheet(wb, code, rows)
        wb.save(PRIMARY_LQA_FOLDER / f"Item_LQA_{code.upper()}.xlsx")
        log.info("  Primary Excel saved: Item_LQA_%s.xlsx", code.upper())

        secondary_path = SECONDARY_LQA_FOLDER / f"ITEM_WORKING_LQA_{code.upper()}.xlsx"
        write_secondary_excel(secondary_path, folder_files,
                              items, all_names, tbl, code, eng_tbl)

    log.info("Done!")
    log.info("="*70)

if __name__ == "__main__":
    main()
