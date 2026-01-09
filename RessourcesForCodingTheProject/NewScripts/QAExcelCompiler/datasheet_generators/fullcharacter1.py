#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CharacterInfo Extractor v1
--------------------------
• Extracts CharacterInfo nodes from staticinfo files
• Groups by filename pattern (characterinfo_X_*.staticinfo → X tab)
• Multi-language support with Original (KR) + English (ENG) + OtherLanguage columns
• Includes COMMAND column: /create character {StrKey}

Tab Organization:
  characterinfo_npc.staticinfo        ]
  characterinfo_npc_shop.staticinfo   ] → NPC tab
  characterinfo_npc_unique.staticinfo ]
  
  characterinfo_monster.staticinfo        ] → MONSTER tab
  characterinfo_monster_unique.staticinfo ]
"""

from __future__ import annotations

import logging
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "Character_LQA_All"
LOG_FILE = base_path / "character_scan.log"

# ──────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────
log = logging.getLogger("CharacterLQA")
log.setLevel(logging.DEBUG)

# _file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
# _file_handler.setFormatter(
    # logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
# )
# _file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

# log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ──────────────────────────────────────────────────────────────────────
# NORMALIZATION (FOR TRANSLATION MATCHING)
# ──────────────────────────────────────────────────────────────────────
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)

def normalize_placeholders(text: str) -> str:
    """
    1) Remove '#…' suffix inside {...} placeholders.
    2) Collapse all whitespace to ONE space.
    3) Trim leading/trailing spaces.
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text

# ──────────────────────────────────────────────────────────────────────
# KOREAN DETECTION
# ──────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')

def contains_korean(text: str) -> bool:
    """Check if text contains any Korean (Hangul) characters."""
    return bool(_korean_re.search(text))

def is_good_translation(text: str) -> bool:
    """
    A good translation is non-empty and does NOT contain Korean.
    """
    return bool(text) and not contains_korean(text)

# ──────────────────────────────────────────────────────────────────────
# XML SANITIZATION
# ──────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r"&(?!lt;|gt;|amp;|apos;|quot;)")

def _fix_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)

def _escape_newlines_in_seg(txt: str) -> str:
    def repl(m: re.Match) -> str:
        seg = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{seg}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, txt, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = _fix_entities(raw)
    raw = _escape_newlines_in_seg(raw)

    # Escape stray < inside attribute values
    raw = re.sub(
        r'="([^"]*?<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*?&[^ltgapoqu"][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    # Fix orphan/broken close tags
    tag_stack: List[str] = []
    o_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    c_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed: List[str] = []
    for line in raw.splitlines():
        s = line.strip()
        m_open = o_re.match(s)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed.append(line)
            continue
        m_close = c_re.match(s)
        if m_close:
            while tag_stack and tag_stack[-1] != m_close.group(1):
                fixed.append(f"</{tag_stack.pop()}>")
            if tag_stack:
                tag_stack.pop()
            fixed.append(line)
            continue
        if s.startswith("</>") and tag_stack:
            fixed.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            continue
        fixed.append(line)
    while tag_stack:
        fixed.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None

    wrapped = f"<ROOT>\n{sanitize_xml(raw)}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retry with recover", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True, recover=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse failed: %s", path)
            return None

# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────
def iter_characterinfo_files(root: Path) -> Iterable[Path]:
    """Find all characterinfo_*.staticinfo.xml files"""
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().startswith("characterinfo_") and fn.lower().endswith(".staticinfo.xml"):
                yield Path(dp) / fn

def get_group_key(filename: str) -> str:
    """
    Extract group key from filename.
    
    characterinfo_npc.staticinfo.xml → npc
    characterinfo_npc_shop.staticinfo.xml → npc
    characterinfo_monster_unique.staticinfo.xml → monster
    """
    # Remove prefix and suffix
    stem = filename.lower()
    if stem.startswith("characterinfo_"):
        stem = stem[len("characterinfo_"):]
    if stem.endswith(".staticinfo.xml"):
        stem = stem[:-len(".staticinfo.xml")]
    
    # Get first segment (before first underscore, or whole thing if no underscore)
    parts = stem.split("_")
    return parts[0] if parts else "unknown"

# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ──────────────────────────────────────────────────────────────────────
@dataclass
class CharacterItem:
    strkey: str
    name: str  # CharacterName (Korean)
    source_file: str  # For debugging

# ──────────────────────────────────────────────────────────────────────
# CHARACTER EXTRACTION
# ──────────────────────────────────────────────────────────────────────
def extract_characters_from_file(path: Path) -> List[CharacterItem]:
    """Extract all CharacterInfo nodes from a file"""
    characters: List[CharacterItem] = []
    
    root_el = parse_xml_file(path)
    if root_el is None:
        return characters
    
    for node in root_el.iter("CharacterInfo"):
        strkey = node.get("StrKey") or ""
        name = node.get("CharacterName") or ""
        
        # Skip entries without both key and name
        if not strkey or not name:
            continue
        
        characters.append(CharacterItem(
            strkey=strkey,
            name=name,
            source_file=path.name
        ))
    
    return characters

def build_character_groups(folder: Path) -> Dict[str, List[CharacterItem]]:
    """
    Build character groups organized by filename pattern.
    
    Returns:
        Dict mapping group_key (e.g., 'npc', 'monster') to list of CharacterItems
    """
    groups: Dict[str, List[CharacterItem]] = defaultdict(list)
    seen_strkeys: set = set()  # Deduplication
    
    log.info("Scanning for characterinfo files...")
    
    file_count = 0
    for path in sorted(iter_characterinfo_files(folder)):
        file_count += 1
        group_key = get_group_key(path.name)
        
        characters = extract_characters_from_file(path)
        
        added = 0
        for char in characters:
            if char.strkey in seen_strkeys:
                continue
            seen_strkeys.add(char.strkey)
            groups[group_key].append(char)
            added += 1
        
        log.debug("  %s → %s: %d characters", path.name, group_key.upper(), added)
    
    log.info("Scanned %d files, found %d groups", file_count, len(groups))
    for key, items in sorted(groups.items()):
        log.info("  • %s: %d characters", key.upper(), len(items))
    
    return dict(groups)

# ──────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ──────────────────────────────────────────────────────────────────────
def load_language_tables(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    """
    Load all non-Korean language tables with normalized placeholder keys.
    
    Returns dict: lang_code → {normalized_origin → (translation, string_id)}
    """
    tables: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for path in sorted(folder.iterdir()):
        if not path.is_file():
            continue
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}
        duplicates_improved = 0

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            normalized_origin = normalize_placeholders(origin)

            # Smart duplicate handling
            if normalized_origin in tbl:
                existing_tr, existing_sid = tbl[normalized_origin]
                existing_is_good = is_good_translation(existing_tr)
                new_is_good = is_good_translation(tr)

                if new_is_good and not existing_is_good:
                    tbl[normalized_origin] = (tr, sid)
                    duplicates_improved += 1
            else:
                tbl[normalized_origin] = (tr, sid)

        tables[lang] = tbl
        log.info(
            "Language %s loaded – %d entries (%d duplicates improved)",
            lang.upper(), len(tbl), duplicates_improved
        )

    if not tables:
        log.warning("No localisation tables found!")

    return tables

# ──────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ──────────────────────────────────────────────────────────────────────
_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_normal_font = Font()
_bold_font = Font(bold=True)

_row_fill_even = PatternFill("solid", fgColor="F2F2F2")
_row_fill_odd = PatternFill("solid", fgColor="FFFFFF")

_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def _apply_cell_style(cell, fill: PatternFill, font: Font, center: bool = False) -> None:
    """Apply styling to a cell"""
    cell.fill = fill
    cell.font = font
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _border

# ──────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ──────────────────────────────────────────────────────────────────────
def write_workbook(
    groups: Dict[str, List[CharacterItem]],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with one sheet per group.

    Columns:
      - Original (KR)
      - English (ENG)
      - Translation (OTHER) [skipped for ENG workbook]
      - COMMAND
      - STATUS (drop-down list: ISSUE / NO ISSUE / BLOCKED, or blank — user cannot type anything else)
      - COMMENT
      - STRINGID
      - SCREENSHOT
    """
    wb = Workbook()
    wb.remove(wb.active)

    is_eng = lang_code.lower() == "eng"

    # Helper to add STATUS drop-down to a sheet ----------------------------
    def _add_status_validation(sh, status_col_idx: int, max_row: int) -> None:
        """
        Adds a strict data-validation list (ISSUE/NO ISSUE/BLOCKED) to the STATUS column.
        Users can only pick one of the three options or leave the cell blank.
        Manual typing of other values is blocked.
        """
        col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED"',
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid STATUS",
            error="Please select ISSUE, NO ISSUE, or BLOCKED from the list."
        )
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)
        sh.add_data_validation(dv)

    for group_key in sorted(groups.keys()):
        characters = groups[group_key]
        if not characters:
            continue

        title = group_key.upper()[:31]
        title = re.sub(r"[\\/*?:\[\]]", "_", title)
        sheet = wb.create_sheet(title=title)

        # Header row
        headers = []
        h1 = sheet.cell(1, 1, "Original (KR)")
        h2 = sheet.cell(1, 2, "English (ENG)")
        headers = [h1, h2]

        col_idx = 3
        if not is_eng:
            h3 = sheet.cell(1, col_idx, f"Translation ({lang_code.upper()})")
            headers.append(h3)
            col_idx += 1

        extra_headers = ["COMMAND", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
        for name in extra_headers:
            headers.append(sheet.cell(1, col_idx, name))
            col_idx += 1

        for c in headers:
            c.font = _header_font
            c.fill = _header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border
        sheet.row_dimensions[1].height = 25

        # Write data rows (with deduplication)
        seen_keys = set()
        duplicates_removed = 0
        r_idx = 2

        for char in characters:
            fill = _row_fill_even if r_idx % 2 == 0 else _row_fill_odd
            normalized_name = normalize_placeholders(char.name)

            trans_eng, sid_eng = eng_tbl.get(normalized_name, ("", ""))
            trans_other, sid_other = ("", "")
            if not is_eng and lang_tbl is not None:
                trans_other, sid_other = lang_tbl.get(normalized_name, ("", ""))

            # Deduplication: skip if (Korean, Translation, STRINGID) already seen
            trans = trans_eng if is_eng else trans_other
            sid = sid_eng if is_eng else sid_other
            dedup_key = (char.name, trans, sid)
            if dedup_key in seen_keys:
                duplicates_removed += 1
                continue
            seen_keys.add(dedup_key)

            command = f"/create character {char.strkey}"

            col = 1
            c_orig = sheet.cell(r_idx, col, char.name)
            _apply_cell_style(c_orig, fill, _normal_font)
            col += 1

            c_eng = sheet.cell(r_idx, col, trans_eng)
            _apply_cell_style(c_eng, fill, _normal_font)
            col += 1

            if not is_eng:
                c_other = sheet.cell(r_idx, col, trans_other)
                _apply_cell_style(c_other, fill, _normal_font)
                col += 1

            c_command = sheet.cell(r_idx, col, command)
            _apply_cell_style(c_command, fill, _bold_font)
            col += 1

            c_status = sheet.cell(r_idx, col, "")
            _apply_cell_style(c_status, fill, _normal_font, center=True)
            col += 1

            c_comment = sheet.cell(r_idx, col, "")
            _apply_cell_style(c_comment, fill, _normal_font)
            col += 1

            c_stringid = sheet.cell(r_idx, col, sid_other if not is_eng else sid_eng)
            _apply_cell_style(c_stringid, fill, _bold_font)
            col += 1

            c_screenshot = sheet.cell(r_idx, col, "")
            _apply_cell_style(c_screenshot, fill, _normal_font)

            r_idx += 1

        # Log duplicates removed
        if duplicates_removed > 0:
            log.info("    Removed %d duplicate rows (Korean+Translation+STRINGID)", duplicates_removed)

        last_row = r_idx - 1  # Actual last row written

        # Column widths
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].hidden = not is_eng
        sheet.column_dimensions["B"].width = 40
        if not is_eng:
            sheet.column_dimensions["C"].width = 40
            sheet.column_dimensions["D"].width = 50
            sheet.column_dimensions["E"].width = 11
            sheet.column_dimensions["F"].width = 50
            sheet.column_dimensions["G"].width = 25
            sheet.column_dimensions["H"].width = 20
        else:
            sheet.column_dimensions["C"].width = 50
            sheet.column_dimensions["D"].width = 11
            sheet.column_dimensions["E"].width = 50
            sheet.column_dimensions["F"].width = 25
            sheet.column_dimensions["G"].width = 20

        # Add strict STATUS validation
        status_col_idx = 5 if not is_eng else 4
        _add_status_validation(sheet, status_col_idx, last_row)

        # Force STRINGID column to text format (prevents scientific notation)
        stringid_col_idx = 7 if not is_eng else 6
        for row in range(2, last_row + 1):
            sheet.cell(row, stringid_col_idx).number_format = '@'

        actual_rows = last_row - 1  # Rows written (excluding header)
        log.info("  Sheet '%s': %d rows", title, actual_rows)

    if wb.worksheets:
        wb.save(out_path)
        log.info("→ Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))
    else:
        log.warning("→ Skipped: %s (no data)", out_path.name)

# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 60)
    log.info("CharacterInfo LQA Extractor v1")
    log.info("=" * 60)

    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # 1. Build character groups from StaticInfo
    log.info("")
    log.info("Scanning for characters from: %s", RESOURCE_FOLDER)
    groups = build_character_groups(RESOURCE_FOLDER)

    if not groups:
        sys.exit("ERROR: No character groups found!")

    # 2. Load language tables
    log.info("")
    log.info("Loading language tables from: %s", LANGUAGE_FOLDER)
    lang_tables = load_language_tables(LANGUAGE_FOLDER)

    if not lang_tables:
        sys.exit("ERROR: No language tables found!")

    # Extract English table for universal ENG column
    eng_tbl = lang_tables.get("eng", {})

    # 3. Generate workbooks
    log.info("")
    log.info("Generating Excel workbooks...")
    total = len(lang_tables)

    for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
        log.info("")
        log.info("(%d/%d) Processing language: %s", idx, total, code.upper())
        out_xlsx = OUTPUT_FOLDER / f"Character_LQA_{code.upper()}.xlsx"
        if code.lower() == "eng":
            write_workbook(groups, eng_tbl, None, code, out_xlsx)
        else:
            write_workbook(groups, eng_tbl, tbl, code, out_xlsx)

    log.info("")
    log.info("=" * 60)
    log.info("DONE – %d workbook(s) generated in %s", total, OUTPUT_FOLDER)
    log.info("=" * 60)

if __name__ == "__main__":
    main()
