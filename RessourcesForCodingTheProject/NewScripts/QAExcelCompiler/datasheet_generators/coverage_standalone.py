#!/usr/bin/env python3
"""
Standalone Coverage Calculator
==============================
Post-process coverage calculation for monolith datasheet outputs.

Run this AFTER generating all datasheets with the monolith generators.
Reads ENG Excel files from output folders in the same directory.

Usage:
    python coverage_standalone.py
"""

import re
import sys
from pathlib import Path
from typing import Dict, Set, Tuple, List
from dataclasses import dataclass, field

from openpyxl import load_workbook
from lxml import etree as ET

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).parent

# Output folders (relative to script location)
OUTPUT_FOLDERS = {
    "Character":  SCRIPT_DIR / "Character_LQA_All",
    "Quest":      SCRIPT_DIR / "QuestData_Map_All",
    "Item":       SCRIPT_DIR / "ItemData_Map_All" / "Item_Full_LQA",
    "Knowledge":  SCRIPT_DIR / "Knowledge_LQA_All",
    "Skill":      SCRIPT_DIR / "Skill_LQA_All",
    "Region":     SCRIPT_DIR / "Region_LQA_v3",
    "Gimmick":    SCRIPT_DIR / "Gimmick_LQA_Output",
    "Help":       SCRIPT_DIR / "GameAdvice_LQA_All",
}

# ENG file patterns per category
ENG_FILE_PATTERNS = {
    "Character":  "Character_LQA_ENG.xlsx",
    "Quest":      "Quest_LQA_ENG.xlsx",
    "Item":       "*_ENG*.xlsx",  # Multiple files
    "Knowledge":  "Knowledge_LQA_ENG.xlsx",
    "Skill":      "LQA_Skill_ENG.xlsx",
    "Region":     "Region_LQA_ENG.xlsx",
    "Gimmick":    "Gimmick_LQA_ENG.xlsx",
    "Help":       "LQA_GameAdvice_ENG.xlsx",
}

# Korean column index (1-based) per category - column A = 1
# Can be int for single column, or tuple for multiple columns
KOREAN_COLUMN = {
    "Character":  1,       # Original (KR)
    "Quest":      1,       # Original
    "Item":       (6, 8),  # ItemName(KOR)=col F, ItemDesc(KOR)=col H
    "Knowledge":  1,       # Original (KR)
    "Skill":      1,       # Original (KR)
    "Region":     1,       # Original (KR)
    "Gimmick":    1,       # Original (KR)
    "Help":       1,       # Original (KR)
}

# Language data folder
LANGUAGE_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc")

# Export folder (for analyzing unconsumed strings)
EXPORT_FOLDER = Path(r"F:\perforce\cd\mainline\resource\GameData\stringtable\export__")

# Voice Recording Sheet folder
VOICE_RECORDING_FOLDER = Path(r"F:\perforce\cd\mainline\resource\editordata\VoiceRecordingSheet__")

# Non-priority folders (under System/) - excluded from CLEAN report
NON_PRIORITY_FOLDERS = {"ItemGroup", "Gimmick", "MultiChange"}

# =============================================================================
# HELPERS
# =============================================================================

_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)


def normalize_placeholders(text: str) -> str:
    """Normalize text: remove placeholder suffixes, collapse whitespace."""
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text


def count_korean_words(text: str) -> int:
    """Count words in Korean text (whitespace-separated tokens)."""
    if not text:
        return 0
    tokens = text.split()
    return len(tokens) if tokens else (1 if text.strip() else 0)


def count_words_in_set(strings: Set[str]) -> int:
    """Count total words in a set of strings."""
    return sum(count_korean_words(s) for s in strings)


# =============================================================================
# XML PARSING (from monolith)
# =============================================================================

_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')


def _fix_bad_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)


def _preprocess_newlines(raw: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw, flags=re.DOTALL)


def sanitize_xml(raw: str) -> str:
    raw = _fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"',
                 lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                 lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', raw)
    tag_open = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    stack: List[str] = []
    out: List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            stack.append(mo.group(1))
            out.append(line)
            continue
        mc = tag_close.match(stripped)
        if mc:
            if stack and stack[-1] == mc.group(1):
                stack.pop()
                out.append(line)
            else:
                out.append(stack and f"</{stack.pop()}>" or line)
            continue
        if stripped.startswith("</>"):
            out.append(stack and line.replace("</>", f"</{stack.pop()}>") or line)
            continue
        out.append(line)
    while stack:
        out.append(f"</{stack.pop()}>")
    return "\n".join(out)


def parse_xml_file(path: Path):
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        try:
            return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(recover=True, huge_tree=True))
        except ET.XMLSyntaxError:
            return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_master_language_data(language_folder: Path) -> Tuple[Set[str], int, Dict[str, str]]:
    """Load all StrOrigin values from language data.

    Returns:
        (master_strings set, word_count, origin_to_stringid mapping)
    """
    print(f"Loading master language data from: {language_folder}")

    master_strings: Set[str] = set()
    origin_to_stringid: Dict[str, str] = {}

    if not language_folder.exists():
        print(f"  ERROR: Language folder not found")
        return master_strings, 0, origin_to_stringid

    lang_files = sorted(language_folder.glob("languagedata_*.xml"))
    if not lang_files:
        print(f"  ERROR: No language data files found")
        return master_strings, 0, origin_to_stringid

    # Prefer English
    target_file = None
    for f in lang_files:
        if "eng" in f.stem.lower():
            target_file = f
            break
    if target_file is None:
        target_file = lang_files[0]

    print(f"  Using: {target_file.name}")

    root = parse_xml_file(target_file)
    if root is None:
        print(f"  ERROR: Failed to parse language file")
        return master_strings, 0, origin_to_stringid

    for loc in root.iter("LocStr"):
        origin = loc.get("StrOrigin") or ""
        sid = loc.get("StringId") or ""
        if origin:
            normalized = normalize_placeholders(origin)
            if normalized:
                master_strings.add(normalized)
                if sid:
                    origin_to_stringid[normalized] = sid

    total_words = count_words_in_set(master_strings)
    print(f"  Loaded {len(master_strings):,} unique strings ({total_words:,} words)")
    print(f"  Mapped {len(origin_to_stringid):,} strings to StringIds")

    return master_strings, total_words, origin_to_stringid


def load_voice_recording_sheet(folder: Path) -> Set[str]:
    """Load StrOrigin values from VoiceRecordingSheet Excel file."""
    print(f"Loading VoiceRecordingSheet from: {folder}")

    voice_strings: Set[str] = set()

    if not folder.exists():
        print(f"  WARNING: VoiceRecordingSheet folder not found")
        return voice_strings

    excel_files = sorted(folder.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not excel_files:
        print(f"  WARNING: No Excel files found")
        return voice_strings

    target_file = excel_files[0]
    print(f"  Using: {target_file.name}")

    try:
        wb = load_workbook(target_file, read_only=True, data_only=True)

        for sheet in wb.worksheets:
            header_row = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = row
                break

            if header_row is None:
                continue

            str_origin_col = None
            for idx, header in enumerate(header_row):
                if header and "StrOrigin" in str(header):
                    str_origin_col = idx
                    break

            if str_origin_col is None:
                continue

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > str_origin_col:
                    value = row[str_origin_col]
                    if value:
                        normalized = normalize_placeholders(str(value))
                        if normalized:
                            voice_strings.add(normalized)

        wb.close()

    except Exception as e:
        print(f"  ERROR: {e}")

    print(f"  Loaded {len(voice_strings):,} strings from VoiceRecordingSheet")
    return voice_strings


def load_korean_from_excel(xlsx_path: Path, korean_cols = 1) -> Set[str]:
    """Load Korean strings from specified column(s) of Excel file.

    Args:
        xlsx_path: Path to Excel file
        korean_cols: int for single column, or tuple of ints for multiple columns (1-based)
    """
    strings: Set[str] = set()

    # Normalize to tuple
    if isinstance(korean_cols, int):
        cols = (korean_cols,)
    else:
        cols = korean_cols

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if not row:
                    continue
                for col in cols:
                    if len(row) >= col:
                        value = row[col - 1]  # Convert to 0-based
                        if value:
                            normalized = normalize_placeholders(str(value))
                            if normalized:
                                strings.add(normalized)

        wb.close()

    except Exception as e:
        print(f"    ERROR loading {xlsx_path.name}: {e}")

    return strings


def collect_category_strings() -> Dict[str, Set[str]]:
    """Collect Korean strings from all category output folders."""
    print()
    print("=" * 70)
    print("Collecting Korean strings from datasheet outputs")
    print("=" * 70)

    category_strings: Dict[str, Set[str]] = {}

    for category, folder in OUTPUT_FOLDERS.items():
        print(f"\n{category}:")

        if not folder.exists():
            print(f"  SKIP: Folder not found: {folder.name}")
            continue

        pattern = ENG_FILE_PATTERNS.get(category, "*.xlsx")
        korean_col = KOREAN_COLUMN.get(category, 1)

        # Find matching files
        if "*" in pattern:
            files = list(folder.glob(pattern))
        else:
            files = [folder / pattern] if (folder / pattern).exists() else []

        if not files:
            print(f"  SKIP: No ENG files found matching {pattern}")
            continue

        category_set: Set[str] = set()

        for xlsx_file in files:
            strings = load_korean_from_excel(xlsx_file, korean_col)
            category_set.update(strings)
            print(f"  {xlsx_file.name}: {len(strings):,} strings")

        category_strings[category] = category_set
        print(f"  TOTAL: {len(category_set):,} unique strings")

    return category_strings


# =============================================================================
# COVERAGE CALCULATION
# =============================================================================

@dataclass
class CategoryCoverage:
    name: str
    unique_strings: int = 0
    word_count: int = 0
    sub_categories: List["CategoryCoverage"] = field(default_factory=list)


@dataclass
class CoverageReport:
    categories: List[CategoryCoverage] = field(default_factory=list)
    total_master_strings: int = 0
    total_master_words: int = 0
    total_covered_strings: int = 0
    total_covered_words: int = 0


def calculate_coverage(
    master_strings: Set[str],
    master_word_count: int,
    category_strings: Dict[str, Set[str]],
    voice_sheet_strings: Set[str],
) -> Tuple[CoverageReport, Set[str]]:
    """Calculate coverage using consume technique.

    Returns:
        (CoverageReport, remaining unconsumed strings)
    """
    print()
    print("Calculating coverage with consume technique...")

    report = CoverageReport(
        total_master_strings=len(master_strings),
        total_master_words=master_word_count,
    )

    remaining = master_strings.copy()

    # Process order
    category_order = ["Character", "Quest", "Item", "Knowledge", "Skill", "Region", "Gimmick", "Help"]

    for category_name in category_order:
        if category_name not in category_strings:
            continue

        cat_strings = category_strings[category_name]

        # Consume: intersection with remaining
        consumed = cat_strings & remaining
        remaining -= consumed

        word_count = count_words_in_set(consumed)

        cat_coverage = CategoryCoverage(
            name=category_name,
            unique_strings=len(consumed),
            word_count=word_count,
        )

        # Quest: add voice sheet as sub-category
        if category_name == "Quest" and voice_sheet_strings:
            voice_consumed = voice_sheet_strings & remaining
            remaining -= voice_consumed
            voice_word_count = count_words_in_set(voice_consumed)

            voice_coverage = CategoryCoverage(
                name="Voice Sheet",
                unique_strings=len(voice_consumed),
                word_count=voice_word_count,
            )
            cat_coverage.sub_categories.append(voice_coverage)

            report.total_covered_strings += len(voice_consumed)
            report.total_covered_words += voice_word_count

        report.categories.append(cat_coverage)
        report.total_covered_strings += len(consumed)
        report.total_covered_words += word_count

    return report, remaining


def print_coverage_report(report: CoverageReport) -> None:
    """Print formatted coverage report."""
    width = 75

    print()
    print("=" * width)
    print("                    LANGUAGE DATA COVERAGE REPORT")
    print("=" * width)
    print()
    print("NOTE: All counts are UNIQUE strings (duplicates removed via normalization)")
    print()

    print(f"{'Category':<20} {'Unique Strings':>15} {'Words Covered':>15} {'% Coverage':>12}")
    print("-" * width)

    for cat in report.categories:
        pct = (cat.unique_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
        print(f"{cat.name:<20} {cat.unique_strings:>15,} {cat.word_count:>15,} {pct:>11.1f}%")

        for sub in cat.sub_categories:
            sub_pct = (sub.unique_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
            print(f"  {'└─ ' + sub.name:<17} {sub.unique_strings:>15,} {sub.word_count:>15,} {sub_pct:>11.1f}%")

    print()
    print("=" * width)

    string_pct = (report.total_covered_strings / report.total_master_strings * 100) if report.total_master_strings > 0 else 0
    word_pct = (report.total_covered_words / report.total_master_words * 100) if report.total_master_words > 0 else 0

    print(f"TOTAL COVERAGE:  {report.total_covered_strings:,} / {report.total_master_strings:,} strings ({string_pct:.1f}%)")
    print(f"                 {report.total_covered_words:,} / {report.total_master_words:,} words ({word_pct:.1f}%)")
    print("=" * width)
    print()


# =============================================================================
# UNCONSUMED STRINGS ANALYSIS
# =============================================================================

def build_export_index(export_folder: Path, depth: int = 2) -> Dict[str, str]:
    """Build index mapping StringId → category (folder path with subfolder).

    Args:
        export_folder: Path to export folder
        depth: How many folder levels to include (default 2 = TopFolder/SubFolder)

    Returns:
        Dict mapping StringId → category path (e.g., "World/NPC", "System/UI")
    """
    print(f"\nBuilding export index from: {export_folder}")
    print(f"  Category depth: {depth} levels")

    idx: Dict[str, str] = {}

    if not export_folder.exists():
        print(f"  WARNING: Export folder not found")
        return idx

    xml_count = 0
    for xml in export_folder.rglob("*.xml"):
        try:
            root = parse_xml_file(xml)
            if root is None:
                continue
        except Exception:
            continue

        xml_count += 1
        parts = xml.relative_to(export_folder).parts

        # Build category from folder path up to specified depth
        # e.g., depth=2 gives "World/NPC" or "Dialog/MainQuest"
        if parts:
            # Take up to 'depth' folder parts (excluding the filename)
            folder_parts = parts[:-1] if len(parts) > 1 else parts[:1]
            cat_parts = folder_parts[:depth]
            cat = "/".join(cat_parts) if cat_parts else parts[0]
        else:
            cat = "Unknown"

        # Remap System/ItemGroup, System/Gimmick, System/MultiChange → "Non-Priority"
        if cat.startswith("System/"):
            subfolder = cat.split("/", 1)[1] if "/" in cat else ""
            if subfolder in NON_PRIORITY_FOLDERS:
                cat = "Non-Priority"

        for loc in root.iter("LocStr"):
            sid = loc.get("StringId")
            if sid:
                idx[sid] = cat

    print(f"  Indexed {len(idx):,} StringIds from {xml_count:,} XML files")
    return idx


@dataclass
class UnconsumedAnalysis:
    """Analysis of unconsumed strings by category."""
    category_breakdown: Dict[str, int] = field(default_factory=dict)
    category_words: Dict[str, int] = field(default_factory=dict)
    total_strings: int = 0
    total_words: int = 0
    unmapped_strings: int = 0  # Strings with no StringId


def analyze_unconsumed(
    remaining: Set[str],
    origin_to_stringid: Dict[str, str],
    export_index: Dict[str, str],
) -> UnconsumedAnalysis:
    """Analyze unconsumed strings by looking up their StringIds in export.

    Args:
        remaining: Set of unconsumed StrOrigin strings
        origin_to_stringid: Mapping from StrOrigin → StringId
        export_index: Mapping from StringId → category (from export folder)

    Returns:
        UnconsumedAnalysis with breakdown by category
    """
    print("\nAnalyzing unconsumed strings...")

    analysis = UnconsumedAnalysis(
        total_strings=len(remaining),
        total_words=count_words_in_set(remaining),
    )

    category_strings: Dict[str, Set[str]] = {}

    for origin in remaining:
        sid = origin_to_stringid.get(origin)
        if not sid:
            analysis.unmapped_strings += 1
            cat = "No StringId"
        else:
            cat = export_index.get(sid, "Not in Export")

        if cat not in category_strings:
            category_strings[cat] = set()
        category_strings[cat].add(origin)

    # Calculate counts and words per category
    for cat, strings in category_strings.items():
        analysis.category_breakdown[cat] = len(strings)
        analysis.category_words[cat] = count_words_in_set(strings)

    return analysis


def print_unconsumed_report(analysis: UnconsumedAnalysis) -> None:
    """Print formatted report of unconsumed strings by subfolder.

    Shows TWO reports:
    - RAW REPORT: All unconsumed strings
    - CLEAN REPORT: Excludes "None", "No StringId", "Not in Export", "Non-Priority"
    """
    width = 90

    # Categories to exclude from CLEAN report
    excluded_cats = {"None", "No StringId", "Not in Export", "Non-Priority"}

    # Calculate CLEAN totals
    clean_strings = 0
    clean_words = 0
    for cat, count in analysis.category_breakdown.items():
        if cat not in excluded_cats:
            clean_strings += count
            clean_words += analysis.category_words.get(cat, 0)

    # Sort by count descending
    sorted_cats = sorted(
        analysis.category_breakdown.items(),
        key=lambda x: x[1],
        reverse=True
    )

    # =========================================================================
    # RAW REPORT
    # =========================================================================
    print()
    print("=" * width)
    print("              UNCONSUMED STRINGS - RAW REPORT (All Categories)")
    print("=" * width)
    print()
    print("NOTE: All counts are UNIQUE strings (no duplicates)")
    print()

    print(f"Total unconsumed (RAW): {analysis.total_strings:,} unique strings ({analysis.total_words:,} words)")
    if analysis.unmapped_strings > 0:
        print(f"  (includes {analysis.unmapped_strings:,} strings with no StringId in language data)")
    print()

    print(f"{'Folder/Subfolder':<45} {'Strings':>12} {'Words':>12} {'% of RAW':>15}")
    print("-" * width)

    for cat, count in sorted_cats:
        words = analysis.category_words.get(cat, 0)
        pct = (count / analysis.total_strings * 100) if analysis.total_strings > 0 else 0
        # Truncate long paths
        display_cat = cat if len(cat) <= 44 else cat[:41] + "..."
        # Mark excluded categories
        marker = " [EXCLUDED]" if cat in excluded_cats else ""
        print(f"{display_cat:<45} {count:>12,} {words:>12,} {pct:>14.1f}%{marker}")

    print("=" * width)

    # =========================================================================
    # CLEAN REPORT
    # =========================================================================
    print()
    print("=" * width)
    print("              UNCONSUMED STRINGS - CLEAN REPORT (Priority Only)")
    print("=" * width)
    print()
    print("EXCLUDED: None, No StringId, Not in Export, Non-Priority")
    print("          (Non-Priority = System/ItemGroup, System/Gimmick, System/MultiChange)")
    print()

    print(f"Total unconsumed (CLEAN): {clean_strings:,} unique strings ({clean_words:,} words)")
    print()

    print(f"{'Folder/Subfolder':<45} {'Strings':>12} {'Words':>12} {'% of CLEAN':>15}")
    print("-" * width)

    for cat, count in sorted_cats:
        if cat in excluded_cats:
            continue
        words = analysis.category_words.get(cat, 0)
        pct = (count / clean_strings * 100) if clean_strings > 0 else 0
        # Truncate long paths
        display_cat = cat if len(cat) <= 44 else cat[:41] + "..."
        print(f"{display_cat:<45} {count:>12,} {words:>12,} {pct:>14.1f}%")

    print("=" * width)
    print()


# =============================================================================
# MAIN
# =============================================================================

def main():
    print()
    print("=" * 70)
    print("         STANDALONE COVERAGE CALCULATOR FOR QA DATASHEETS")
    print("=" * 70)
    print()
    print(f"Script location: {SCRIPT_DIR}")
    print()

    # 1. Load master language data (with StrOrigin → StringId mapping)
    master_strings, master_word_count, origin_to_stringid = load_master_language_data(LANGUAGE_FOLDER)
    if not master_strings:
        print("ERROR: No master language data - cannot calculate coverage")
        sys.exit(1)

    # 2. Load voice recording sheet
    voice_strings = load_voice_recording_sheet(VOICE_RECORDING_FOLDER)

    # 3. Collect Korean strings from all category outputs
    category_strings = collect_category_strings()

    if not category_strings:
        print("\nERROR: No category data found - make sure datasheets are generated first")
        sys.exit(1)

    # 4. Calculate coverage (returns report + remaining unconsumed)
    report, remaining = calculate_coverage(
        master_strings,
        master_word_count,
        category_strings,
        voice_strings,
    )

    # 5. Print coverage report
    print_coverage_report(report)

    # 6. Analyze unconsumed strings (if any)
    if remaining:
        # Build export index to categorize unconsumed strings
        export_index = build_export_index(EXPORT_FOLDER)

        # Analyze where unconsumed strings come from
        unconsumed_analysis = analyze_unconsumed(
            remaining,
            origin_to_stringid,
            export_index,
        )

        # Print unconsumed report
        print_unconsumed_report(unconsumed_analysis)
    else:
        print("\n*** 100% COVERAGE - No unconsumed strings! ***\n")

    print("Done!")


if __name__ == "__main__":
    main()
