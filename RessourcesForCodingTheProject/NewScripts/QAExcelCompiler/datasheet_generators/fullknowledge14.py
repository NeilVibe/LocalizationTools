#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Knowledge-data extractor – COMPLETE REBUILD v4 with ENG + OtherLanguage columns
-----------------------------------------------
• ONE Excel sheet per TOP-LEVEL KnowledgeGroupInfo ("mega root" like 인물, 생태, etc.)
• Everything nested inside that mega root goes into that single sheet
• Proper indentation based on hierarchy depth
• Handles both:
  - KnowledgeInfo nested directly inside KnowledgeGroupInfo
  - Knowledge nodes in CharacterInfo that reference groups via KnowledgeGroupKey
• Multi-language support (one workbook per language)
• Always includes Original (KR) + English (ENG) + OtherLanguage columns
  - For English workbook, only Original (KR) + English (ENG)

FIXES in v2:
• Placeholder normalization for translation matching
  (removes #DisplayName suffixes like {Staticinfo:Knowledge:Knowledge_Kliff#클리프})
• Enhanced styling: Bigger cells/fonts for depth 0 (master) and depth 1 (sub-group)
• Bright purple color for depth 1 groups

FIXES in v3:
• Fixed extract_character_knowledge() to also extract LevelData descriptions
  (previously only extract_knowledge_info() did this, causing missing descriptions
  for Knowledge nodes inside CharacterInfo like Balder)

FIXES in v4:
• Smart duplicate handling in language table loading:
  - When multiple entries have the same origin text, we now pick the BEST translation
  - "Best" = non-empty AND does NOT contain Korean characters
  - This fixes cases where first match was empty but later matches had real translations

UPDATE v5:
• Added English column next to Original (KR) in every output
• For non-English languages, includes three columns: Original (KR) / English (ENG) / Translation (OTHER)
• For English, only two columns: Original (KR) / English (ENG)
"""

from __future__ import annotations

import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\cd_lambda\resource\GameData\stringtable\loc"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "Knowledge_LQA_All"
LOG_FILE = base_path / "knowledge_scan.log"

IGNORE_LIST = {
    "해제/면역 연금 포션",
}

# ──────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────
log = logging.getLogger("KnowledgeLQA")
log.setLevel(logging.DEBUG)

# _file_handler = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
# _file_handler.setFormatter(
    # logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
# )
# _file_handler.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

# log.addHandler(_file_handler)
log.addHandler(_console_handler)
log.propagate = False

# ──────────────────────────────────────────────────────────────────────
#  NORMALIZATION (FIX FOR TRANSLATION MATCHING)
# ──────────────────────────────────────────────────────────────────────
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)   # matches any whitespace

def normalize_placeholders(text: str) -> str:
    """
    1) Remove '#…' suffix inside {...} placeholders.
    2) Collapse all whitespace (space, tab, NBSP, newline, etc.) to ONE space.
    3) Trim leading/trailing spaces.
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)   # step 1
    text = _whitespace_re.sub(' ', text).strip()        # step 2+3
    return text


# ──────────────────────────────────────────────────────────────────────
# KOREAN DETECTION (v4 FIX)
# ──────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')

def contains_korean(text: str) -> bool:
    """
    Check if text contains any Korean (Hangul) characters.

    Used to filter out translations that are still in Korean
    when we expect them to be in another language.
    """
    return bool(_korean_re.search(text))

def is_good_translation(text: str) -> bool:
    """
    Check if a translation is "good" (usable).

    A good translation is:
    - Non-empty
    - Does NOT contain Korean characters

    This helps us pick the best translation when duplicates exist.
    """
    return bool(text) and not contains_korean(text)

# ──────────────────────────────────────────────────────────────────────
# XML SANITIZATION
# ──────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r"&(?!lt;|gt;|amp;|apos;|quot;)")

def _fix_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)

def _escape_newlines_in_seg(txt: str) -> str:
    def repl(m: re.Match) -> str:
        seg = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{seg}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, txt, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = _fix_entities(raw)
    raw = _escape_newlines_in_seg(raw)

    # Escape stray < inside attribute values
    raw = re.sub(
        r'="([^"]*?<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*?&[^ltgapoqu"][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    # Fix orphan/broken close tags
    tag_stack: List[str] = []
    o_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    c_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed: List[str] = []
    for line in raw.splitlines():
        s = line.strip()
        m_open = o_re.match(s)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed.append(line)
            continue
        m_close = c_re.match(s)
        if m_close:
            while tag_stack and tag_stack[-1] != m_close.group(1):
                fixed.append(f"</{tag_stack.pop()}>")
            if tag_stack:
                tag_stack.pop()
            fixed.append(line)
            continue
        if s.startswith("</>") and tag_stack:
            fixed.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            continue
        fixed.append(line)
    while tag_stack:
        fixed.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None

    wrapped = f"<ROOT>\n{sanitize_xml(raw)}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retry with recover", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True, recover=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse failed: %s", path)
            return None

# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn

def should_ignore(text: str) -> bool:
    return any(bad in text for bad in IGNORE_LIST)

# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ──────────────────────────────────────────────────────────────────────
@dataclass
class LevelLine:
    level: str
    desc: str

@dataclass
class KnowledgeItem:
    strkey: str
    name: str
    desc: str
    levels: List[LevelLine] = field(default_factory=list)

@dataclass
class GroupNode:
    strkey: str
    name: str
    desc: str
    icon: bool
    children: List["GroupNode"] = field(default_factory=list)
    knowledges: List[KnowledgeItem] = field(default_factory=list)

# ──────────────────────────────────────────────────────────────────────
# KNOWLEDGE EXTRACTION HELPERS
# ──────────────────────────────────────────────────────────────────────
def extract_knowledge_info(node: ET._Element) -> Optional[KnowledgeItem]:
    """Extract KnowledgeInfo element (nested inside KnowledgeGroupInfo)"""
    strkey = node.get("StrKey") or ""
    name = node.get("Name") or ""
    desc = node.get("Desc") or ""

    if not name and not desc:
        return None

    ki = KnowledgeItem(strkey=strkey, name=name, desc=desc)

    for lvl in node.findall("LevelData"):
        lvl_no = lvl.get("Level") or ""
        lvl_desc = lvl.get("Desc") or ""
        if lvl_desc and not should_ignore(lvl_desc):
            ki.levels.append(LevelLine(level=lvl_no, desc=lvl_desc))

    return ki

def extract_character_knowledge(node: ET._Element) -> Optional[KnowledgeItem]:
    """
    Extract Knowledge element (inside CharacterInfo, references group via KnowledgeGroupKey)

    v3 FIX: Now also extracts LevelData children, just like extract_knowledge_info()
    """
    strkey = node.get("StrKey") or ""
    name = node.get("Name") or ""
    desc = node.get("Desc") or ""

    if not name and not desc:
        return None

    ki = KnowledgeItem(strkey=strkey, name=name, desc=desc)

    # v3 FIX: Extract LevelData descriptions (previously missing!)
    for lvl in node.findall("LevelData"):
        lvl_no = lvl.get("Level") or ""
        lvl_desc = lvl.get("Desc") or ""
        if lvl_desc and not should_ignore(lvl_desc):
            ki.levels.append(LevelLine(level=lvl_no, desc=lvl_desc))

    return ki

# ──────────────────────────────────────────────────────────────────────
# BUILD HIERARCHY
# ──────────────────────────────────────────────────────────────────────
def build_hierarchy(folder: Path) -> Tuple[List[GroupNode], Dict[str, GroupNode]]:
    """
    Two-phase parsing:
    1. Build the nested group hierarchy from KnowledgeGroupInfo structure
    2. Find Knowledge nodes with KnowledgeGroupKey and attach them to groups

    Returns:
        - mega_roots: List of top-level GroupNode (these become Excel tabs)
        - groups_by_key: Dict for lookup by StrKey
    """
    groups_by_key: Dict[str, GroupNode] = {}
    mega_roots_by_key: Dict[str, GroupNode] = {}  # Use dict to avoid duplicates
    pending_knowledges: List[Tuple[str, KnowledgeItem]] = []  # (group_key, knowledge)
    seen_knowledge_keys: set = set()  # Track seen knowledge to avoid duplicates

    log.info("Phase 1: Building group hierarchy...")

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        def parse_group(node: ET._Element) -> Optional[GroupNode]:
            """Recursively parse KnowledgeGroupInfo and its nested children"""
            if node.tag != "KnowledgeGroupInfo":
                return None

            strkey = node.get("StrKey") or ""
            name = node.get("GroupName") or ""
            desc = node.get("Desc") or ""
            icon = bool(node.get("KnowledgeGroupIcon"))

            # Skip if already processed (deduplication)
            if strkey and strkey in groups_by_key:
                existing = groups_by_key[strkey]
                # Still process children to potentially add more
                for ch in node:
                    if ch.tag == "KnowledgeInfo":
                        ki = extract_knowledge_info(ch)
                        if ki and ki.strkey not in seen_knowledge_keys:
                            existing.knowledges.append(ki)
                            if ki.strkey:
                                seen_knowledge_keys.add(ki.strkey)
                return None  # Don't create duplicate

            group = GroupNode(
                strkey=strkey,
                name=name,
                desc=desc,
                icon=icon,
            )

            if strkey:
                groups_by_key[strkey] = group

            # Process children
            for ch in node:
                if ch.tag == "KnowledgeGroupInfo":
                    child_group = parse_group(ch)
                    if child_group:
                        group.children.append(child_group)
                elif ch.tag == "KnowledgeInfo":
                    ki = extract_knowledge_info(ch)
                    if ki:
                        # Avoid duplicate knowledge items
                        if ki.strkey and ki.strkey in seen_knowledge_keys:
                            continue
                        group.knowledges.append(ki)
                        if ki.strkey:
                            seen_knowledge_keys.add(ki.strkey)

            return group

        def find_top_level_groups(node: ET._Element):
            """Find KnowledgeGroupInfo that are direct children of root"""
            for ch in node:
                if ch.tag == "KnowledgeGroupInfo":
                    strkey = ch.get("StrKey") or ""
                    # Skip if already seen as mega root
                    if strkey and strkey in mega_roots_by_key:
                        # But still parse to catch any new children
                        parse_group(ch)
                        continue

                    group = parse_group(ch)
                    if group and group.strkey:
                        mega_roots_by_key[group.strkey] = group
                else:
                    # Recurse to find groups inside other structures
                    find_top_level_groups(ch)

        def find_character_knowledge(node: ET._Element):
            """Find Knowledge nodes inside CharacterInfo etc."""
            for ch in node:
                if ch.tag == "Knowledge":
                    group_key = ch.get("KnowledgeGroupKey") or ""
                    strkey = ch.get("StrKey") or ""
                    if group_key:
                        # Skip duplicates
                        if strkey and strkey in seen_knowledge_keys:
                            continue
                        ki = extract_character_knowledge(ch)
                        if ki:
                            pending_knowledges.append((group_key, ki))
                            if strkey:
                                seen_knowledge_keys.add(strkey)
                # Also check CharacterInfo for embedded Knowledge
                if ch.tag == "CharacterInfo":
                    for inner in ch:
                        if inner.tag == "Knowledge":
                            group_key = inner.get("KnowledgeGroupKey") or ""
                            strkey = inner.get("StrKey") or ""
                            if group_key:
                                if strkey and strkey in seen_knowledge_keys:
                                    continue
                                ki = extract_character_knowledge(inner)
                                if ki:
                                    pending_knowledges.append((group_key, ki))
                                    if strkey:
                                        seen_knowledge_keys.add(strkey)
                find_character_knowledge(ch)

        find_top_level_groups(root_el)
        find_character_knowledge(root_el)

    # Convert to ordered list (preserve discovery order)
    mega_roots = list(mega_roots_by_key.values())

    log.info("Phase 1 complete: %d mega roots, %d total groups", len(mega_roots), len(groups_by_key))

    # Phase 2: Attach pending Knowledge items to their groups
    log.info("Phase 2: Attaching %d character knowledge items...", len(pending_knowledges))
    attached = 0
    for group_key, ki in pending_knowledges:
        if group_key in groups_by_key:
            groups_by_key[group_key].knowledges.append(ki)
            attached += 1
        else:
            log.debug("Group not found for Knowledge: %s -> %s", ki.name, group_key)

    log.info("Phase 2 complete: %d knowledge items attached", attached)

    return mega_roots, groups_by_key

# ──────────────────────────────────────────────────────────────────────
# ROW GENERATION
# ──────────────────────────────────────────────────────────────────────
# (depth, text, needs_translation, is_icon_group, is_name_attribute)
RowItem = Tuple[int, str, bool, bool, bool]

def emit_group_rows(group: GroupNode, depth: int) -> List[RowItem]:
    """Generate rows for a group and all its nested content"""
    rows: List[RowItem] = []

    # Emit group name
    if group.name and not should_ignore(group.name):
        rows.append((depth, group.name, True, group.icon, True))

    # Emit group description
    if group.desc and not should_ignore(group.desc):
        rows.append((depth + 1, group.desc, True, group.icon, False))

    # Emit child groups (recursively)
    for child in group.children:
        rows.extend(emit_group_rows(child, depth + 1))

    # Emit knowledge items in this group
    for kn in group.knowledges:
        rows.extend(emit_knowledge_rows(kn, depth + 1))

    return rows

def emit_knowledge_rows(kn: KnowledgeItem, depth: int) -> List[RowItem]:
    """Generate rows for a single knowledge item"""
    rows: List[RowItem] = []

    # Knowledge name
    if kn.name and not should_ignore(kn.name):
        rows.append((depth, kn.name, True, False, True))

    # Knowledge description
    if kn.desc and not should_ignore(kn.desc):
        rows.append((depth + 1, kn.desc, True, False, False))

    # Level data
    for lvl in kn.levels:
        label = f"Level {lvl.level}"
        if not should_ignore(label):
            rows.append((depth + 1, label, False, False, False))
        if lvl.desc and not should_ignore(lvl.desc):
            rows.append((depth + 2, lvl.desc, True, False, False))

    return rows

# ──────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES (v4 FIX: Smart duplicate handling)
# ──────────────────────────────────────────────────────────────────────
def load_language_tables(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    """
    Load all non-Korean language tables with normalized placeholder keys.

    v4 FIX: Smart duplicate handling
    - When the same origin text appears multiple times, we pick the BEST translation
    - "Best" = non-empty AND does NOT contain Korean characters
    - This ensures we get actual translations instead of empty or Korean entries
    """
    tables: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}
        duplicates_improved = 0  # Counter for logging

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            # NORMALIZE: Remove #DisplayName suffixes from placeholders
            normalized_origin = normalize_placeholders(origin)

            # v4 FIX: Smart duplicate handling
            if normalized_origin in tbl:
                existing_tr, existing_sid = tbl[normalized_origin]
                existing_is_good = is_good_translation(existing_tr)
                new_is_good = is_good_translation(tr)

                if new_is_good and not existing_is_good:
                    # New translation is better! Replace.
                    tbl[normalized_origin] = (tr, sid)
                    duplicates_improved += 1
                    log.debug(
                        "Duplicate improved: '%s' | OLD: '%s' → NEW: '%s'",
                        normalized_origin[:50], existing_tr[:30], tr[:30]
                    )
                # else: keep existing
            else:
                tbl[normalized_origin] = (tr, sid)

        tables[lang] = tbl
        log.info(
            "Language %s loaded – %d entries (%d duplicates improved)",
            lang.upper(), len(tbl), duplicates_improved
        )

    if not tables:
        log.warning("No localisation tables found!")

    return tables

# ──────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ──────────────────────────────────────────────────────────────────────
# Depth 0: Master group (tab name) - BRIGHT YELLOW, BIG
_depth0_fill = PatternFill("solid", fgColor="FFD700")
_depth0_font = Font(bold=True, size=14)
_depth0_row_height = 40

# Depth 1: Sub-group just below master - BRIGHT PURPLE, BIG
_depth1_fill = PatternFill("solid", fgColor="9966FF")
_depth1_font = Font(bold=True, size=12, color="FFFFFF")
_depth1_row_height = 35

# Other depths - normal styling with subtle colors
_depth_fills: Dict[int, PatternFill] = {
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FCE4D6"),
    4: PatternFill("solid", fgColor="DDEBF7"),
    5: PatternFill("solid", fgColor="FFF2CC"),
}
_icon_fill = PatternFill("solid", fgColor="9BC2E6")
_no_colour_fill = PatternFill("solid", fgColor="FFFDEB")

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")
_bold_font = Font(bold=True)
_normal_font = Font()

_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def _get_style_for_depth(depth: int, is_icon: bool) -> Tuple[PatternFill, Font, Optional[float]]:
    """
    Returns (fill, font, row_height) for a given depth.
    row_height is None for default height.
    """
    if depth == 0:
        return _depth0_fill, _depth0_font, _depth0_row_height
    elif depth == 1:
        return _depth1_fill, _depth1_font, _depth1_row_height
    else:
        fill = _depth_fills.get(depth)
        if fill is None and is_icon:
            fill = _icon_fill
        if fill is None:
            fill = _no_colour_fill
        return fill, _bold_font, None

def _apply_cell_style(cell, fill: PatternFill, indent: int, font: Font) -> None:
    """Apply styling to a cell"""
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
    cell.border = _border

# ──────────────────────────────────────────────────────────────────────
# EXCEL WRITER WITH ENG + OtherLanguage
# ──────────────────────────────────────────────────────────────────────
def write_workbook(
    mega_roots: List[GroupNode],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with one sheet per mega root.

    Columns:
      - Original (KR)
      - English (ENG)
      - Translation (OTHER) [skipped for ENG workbook]
      - STATUS   (data-validated, only: ISSUE / NO ISSUE / BLOCKED)
      - COMMENT
      - STRINGID
      - SCREENSHOT
    """
    wb = Workbook()
    wb.remove(wb.active)

    is_eng = lang_code.lower() == "eng"

    for root_group in mega_roots:
        rows = emit_group_rows(root_group, 0)
        if not rows:
            log.debug("Skipping empty mega root: %s", root_group.name)
            continue

        # ─── Sheet title ──────────────────────────────────────────────
        normalized_name = normalize_placeholders(root_group.name)
        title_eng = eng_tbl.get(normalized_name, ("", ""))[0].strip()
        title_orig = root_group.name.strip() or "Sheet"
        title = (title_eng or title_orig)[:31]
        title = re.sub(r"[\\/*?:\[\]]", "_", title)

        base_title = title
        c = 1
        while title in wb.sheetnames:
            title = f"{base_title[:28]}_{c}"
            c += 1

        ws = wb.create_sheet(title)

        # ─── Header row ───────────────────────────────────────────────
        headers: List[ET._Element] = []
        h1 = ws.cell(1, 1, "Original (KR)")
        h2 = ws.cell(1, 2, "English (ENG)")
        headers.extend([h1, h2])

        if not is_eng:
            h3 = ws.cell(1, 3, f"Translation ({lang_code.upper()})")
            headers.append(h3)

        start_extra_col = len(headers) + 1
        extra_names = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
        for idx, name in enumerate(extra_names, start=start_extra_col):
            headers.append(ws.cell(1, idx, name))

        for hcell in headers:
            hcell.font = _header_font
            hcell.fill = _header_fill
            hcell.alignment = Alignment(horizontal="center", vertical="center")
            hcell.border = _border
        ws.row_dimensions[1].height = 25

        # ─── Data rows (with deduplication) ───────────────────────────
        seen_keys = set()
        duplicates_removed = 0
        r_idx = 2

        for (depth, text, needs_tr, is_icon, is_name_attr) in rows:
            fill, font, rh = _get_style_for_depth(depth, is_icon)
            if depth >= 2:
                if is_name_attr or fill != _no_colour_fill:
                    font = _bold_font
                else:
                    font = _normal_font

            norm_text = normalize_placeholders(text)

            eng_tr, sid_eng = ("", "")
            if needs_tr:
                eng_tr, sid_eng = eng_tbl.get(norm_text, ("", ""))

            other_tr, sid_other = ("", "")
            if needs_tr and not is_eng and lang_tbl is not None:
                other_tr, sid_other = lang_tbl.get(norm_text, ("", ""))

            # Deduplication: skip if (Korean, Translation, STRINGID) already seen
            trans = eng_tr if is_eng else other_tr
            sid = sid_eng if is_eng else sid_other
            dedup_key = (text, trans, sid)
            if dedup_key in seen_keys:
                duplicates_removed += 1
                continue
            seen_keys.add(dedup_key)

            # Write core columns
            c_orig = ws.cell(r_idx, 1, text)
            c_eng  = ws.cell(r_idx, 2, eng_tr)
            c_other = None
            if not is_eng:
                c_other = ws.cell(r_idx, 3, other_tr)

            # Extra columns
            col_off = 2 if is_eng else 3  # index of ENG or OTHER
            c_status     = ws.cell(r_idx, col_off + 1, "")
            c_comment    = ws.cell(r_idx, col_off + 2, "")
            c_stringid   = ws.cell(r_idx, col_off + 3, sid_other if not is_eng else sid_eng)
            c_screenshot = ws.cell(r_idx, col_off + 4, "")

            # Styling
            _apply_cell_style(c_orig, fill, depth, font)
            _apply_cell_style(c_eng,  fill, depth, font)
            if c_other:
                _apply_cell_style(c_other, fill, depth, font)
            _apply_cell_style(c_status,     fill, 0, _normal_font)
            _apply_cell_style(c_comment,    fill, 0, _normal_font)
            _apply_cell_style(c_stringid,   fill, 0, _bold_font)
            _apply_cell_style(c_screenshot, fill, 0, _normal_font)

            if rh is not None:
                ws.row_dimensions[r_idx].height = rh

            r_idx += 1

        # Log duplicates removed
        if duplicates_removed > 0:
            log.info("    Removed %d duplicate rows (Korean+Translation+STRINGID)", duplicates_removed)

        # ─── Column widths / visibility ───────────────────────────────
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["A"].hidden = False
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["B"].hidden = not is_eng

        if not is_eng:
            ws.column_dimensions["C"].width = 60   # OTHER
            ws.column_dimensions["D"].width = 11   # STATUS
            ws.column_dimensions["E"].width = 70   # COMMENT
            ws.column_dimensions["F"].width = 25   # STRINGID
            ws.column_dimensions["G"].width = 20   # SCREENSHOT
        else:
            ws.column_dimensions["C"].width = 11   # STATUS
            ws.column_dimensions["D"].width = 70   # COMMENT
            ws.column_dimensions["E"].width = 25   # STRINGID
            ws.column_dimensions["F"].width = 20   # SCREENSHOT

        # ─── Data-validation for STATUS column ────────────────────────
        status_col_idx = (3 if is_eng else 4)     # 1-based column number
        status_letter  = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED,KOREAN"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Entry",
            error="Select one of: ISSUE, NO ISSUE, BLOCKED",
            errorStyle="stop",
        )
        ws.add_data_validation(dv)
        last_row = r_idx - 1  # Actual last row written (after dedup)
        dv.add(f"${status_letter}$2:${status_letter}${last_row}")

        # Force STRINGID column to text format (prevents scientific notation)
        stringid_col_idx = 5 if is_eng else 6  # E for ENG, F for non-ENG
        for row in range(2, last_row + 1):
            ws.cell(row, stringid_col_idx).number_format = '@'

        actual_rows = last_row - 1  # Rows written (excluding header)
        log.info("  Sheet '%s': %d rows", title, actual_rows)

    if wb.worksheets:
        wb.save(out_path)
        log.info("→ Saved: %s (%d sheets)", out_path.name, len(wb.worksheets))
    else:
        log.warning("→ Skipped: %s (no data)", out_path.name)

# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 60)
    log.info("Knowledge LQA Extractor – REBUILD v5")
    log.info("=" * 60)

    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # 1. Build hierarchy from StaticInfo
    log.info("")
    log.info("Building knowledge hierarchy from: %s", RESOURCE_FOLDER)
    mega_roots, groups_by_key = build_hierarchy(RESOURCE_FOLDER)

    if not mega_roots:
        sys.exit("ERROR: No mega root groups found!")

    log.info("")
    log.info("Mega roots found (these become Excel tabs):")
    for root in mega_roots:
        log.info("  • %s (%s)", root.name, root.strkey)

    # 2. Load language tables
    log.info("")
    log.info("Loading language tables from: %s", LANGUAGE_FOLDER)
    lang_tables = load_language_tables(LANGUAGE_FOLDER)

    if not lang_tables:
        sys.exit("ERROR: No language tables found!")

    # Extract English table for universal ENG column
    eng_tbl = lang_tables.get("eng", {})

    # 3. Generate workbooks
    log.info("")
    log.info("Generating Excel workbooks...")
    total = len(lang_tables)

    for idx, (code, tbl) in enumerate(lang_tables.items(), 1):
        log.info("")
        log.info("(%d/%d) Processing language: %s", idx, total, code.upper())
        out_xlsx = OUTPUT_FOLDER / f"Knowledge_LQA_{code.upper()}.xlsx"
        if code.lower() == "eng":
            # English workbook: only Original + ENG
            write_workbook(mega_roots, eng_tbl, None, code, out_xlsx)
        else:
            # Other languages: Original + ENG + Other
            write_workbook(mega_roots, eng_tbl, tbl, code, out_xlsx)

    log.info("")
    log.info("=" * 60)
    log.info("DONE – %d workbook(s) generated in %s", total, OUTPUT_FOLDER)
    log.info("=" * 60)

if __name__ == "__main__":
    main()