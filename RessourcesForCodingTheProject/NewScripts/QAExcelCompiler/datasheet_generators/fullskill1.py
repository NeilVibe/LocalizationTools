#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Skill-data extractor – v1.0
---------------------------
Extracts SkillInfo entries and links them to KnowledgeInfo for detailed descriptions.

Structure:
- SkillInfo: Key, StrKey, SkillName, SkillDesc, LearnKnowledgeKey, MaxLevel
- KnowledgeInfo: Linked via LearnKnowledgeKey, contains Name, Desc, nested children

Priority Rule:
- When multiple SkillInfo share the same LearnKnowledgeKey, the one with highest MaxLevel wins
- Losers use their own SkillDesc instead of the linked KnowledgeInfo

Output:
- ONE Excel sheet with all skills
- Indentation: SkillName (depth 0) -> Knowledge Name/Title (depth 1) -> Desc (depth 2) -> Sub-skills (depth 3+)
- Columns: Original (KR) | English (ENG) | Translation (LOC) | STATUS | COMMENT | STRINGID | SCREENSHOT
"""

from __future__ import annotations

import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Set

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "Skill_LQA_All"
LOG_FILE = base_path / "skill_scan.log"

# ──────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────
log = logging.getLogger("SkillLQA")
log.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

log.addHandler(_console_handler)
log.propagate = False

# ──────────────────────────────────────────────────────────────────────
# NORMALIZATION
# ──────────────────────────────────────────────────────────────────────
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)

def normalize_placeholders(text: str) -> str:
    """
    1) Remove '#...' suffix inside {...} placeholders.
    2) Collapse all whitespace (space, tab, NBSP, newline, etc.) to ONE space.
    3) Trim leading/trailing spaces.
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text

# ──────────────────────────────────────────────────────────────────────
# KOREAN DETECTION
# ──────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')

def contains_korean(text: str) -> bool:
    return bool(_korean_re.search(text))

def is_good_translation(text: str) -> bool:
    return bool(text) and not contains_korean(text)

# ──────────────────────────────────────────────────────────────────────
# XML SANITIZATION
# ──────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r"&(?!lt;|gt;|amp;|apos;|quot;)")

def _fix_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)

def _escape_newlines_in_seg(txt: str) -> str:
    def repl(m: re.Match) -> str:
        seg = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{seg}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, txt, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = _fix_entities(raw)
    raw = _escape_newlines_in_seg(raw)

    # Escape stray < inside attribute values
    raw = re.sub(
        r'="([^"]*?<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*?&[^ltgapoqu"][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    # Fix orphan/broken close tags
    tag_stack: List[str] = []
    o_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    c_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed: List[str] = []
    for line in raw.splitlines():
        s = line.strip()
        m_open = o_re.match(s)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed.append(line)
            continue
        m_close = c_re.match(s)
        if m_close:
            while tag_stack and tag_stack[-1] != m_close.group(1):
                fixed.append(f"</{tag_stack.pop()}>")
            if tag_stack:
                tag_stack.pop()
            fixed.append(line)
            continue
        if s.startswith("</>") and tag_stack:
            fixed.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            continue
        fixed.append(line)
    while tag_stack:
        fixed.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None

    wrapped = f"<ROOT>\n{sanitize_xml(raw)}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s - retry with recover", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True, recover=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse failed: %s", path)
            return None

# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn

# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ──────────────────────────────────────────────────────────────────────
@dataclass
class KnowledgeNode:
    """KnowledgeInfo node - can be parent or child"""
    strkey: str
    name: str
    desc: str
    parent_key: str = ""  # Reference to parent KnowledgeInfo (from LevelData.KnowledgeList)
    unlock_level: int = 0  # Level at which this unlocks (from KnowledgeList like "Knowledge_X(2)")
    children: List["KnowledgeNode"] = field(default_factory=list)

@dataclass
class SkillItem:
    """Single SkillInfo entry"""
    key: str
    strkey: str
    skill_name: str
    skill_desc: str
    learn_knowledge_key: str
    max_level: int
    # Will be filled after knowledge linking
    knowledge: Optional[KnowledgeNode] = None
    use_own_desc: bool = False  # True if this skill lost priority for its knowledge key

# ──────────────────────────────────────────────────────────────────────
# EXTRACTION
# ──────────────────────────────────────────────────────────────────────
# Regex to parse KnowledgeList like "Knowledge_UnArmedMastery_I(2)"
_knowledge_list_re = re.compile(r'^([A-Za-z0-9_]+)\((\d+)\)$')

def extract_skill_data(folder: Path) -> Tuple[List[SkillItem], Dict[str, KnowledgeNode]]:
    """
    Scan StaticInfo folder for SkillInfo and KnowledgeInfo elements.

    Knowledge parent-child relationships are determined by LevelData.KnowledgeList:
    - KnowledgeList="Knowledge_UnArmedMastery_I(2)" means this knowledge
      is a child of Knowledge_UnArmedMastery_I, unlocked at level 2

    Returns:
    - List of SkillItem
    - Dict of KnowledgeNode by StrKey (with children attached)
    """
    skills: List[SkillItem] = []
    knowledge_map: Dict[str, KnowledgeNode] = {}
    seen_skill_keys: Set[str] = set()

    log.info("Scanning for Skill/Knowledge data in: %s", folder)

    # First pass: collect all KnowledgeInfo (flat, with parent references)
    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        for kn_el in root_el.iter("KnowledgeInfo"):
            strkey = kn_el.get("StrKey") or ""
            name = kn_el.get("Name") or ""
            desc = kn_el.get("Desc") or ""

            # Skip if already seen or no content
            if not strkey or strkey in knowledge_map:
                continue
            if not name and not desc:
                continue

            # Check for parent reference in LevelData
            parent_key = ""
            unlock_level = 0
            for level_data in kn_el.findall("LevelData"):
                knowledge_list = level_data.get("KnowledgeList") or ""
                if knowledge_list:
                    # Parse "Knowledge_X(N)" format
                    match = _knowledge_list_re.match(knowledge_list)
                    if match:
                        parent_key = match.group(1)
                        unlock_level = int(match.group(2))
                        break

            knowledge_map[strkey] = KnowledgeNode(
                strkey=strkey,
                name=name,
                desc=desc,
                parent_key=parent_key,
                unlock_level=unlock_level,
            )

    log.info("Found %d KnowledgeInfo entries", len(knowledge_map))

    # Build parent-child relationships
    for kn in knowledge_map.values():
        if kn.parent_key and kn.parent_key in knowledge_map:
            parent = knowledge_map[kn.parent_key]
            parent.children.append(kn)

    # Sort children by unlock_level
    for kn in knowledge_map.values():
        kn.children.sort(key=lambda c: c.unlock_level)

    # Count parents (nodes with children)
    parents_with_children = sum(1 for kn in knowledge_map.values() if kn.children)
    log.info("Built parent-child relationships: %d parents with children", parents_with_children)

    # Second pass: collect SkillInfo from ONLY skillinfo_pc.staticinfo.xml
    # This restricts output to player character skills only
    skill_file = folder / "skillinfo" / "skillinfo_pc.staticinfo.xml"
    if not skill_file.exists():
        log.error("Skill file not found: %s", skill_file)
        return skills, knowledge_map

    log.info("Reading skills from: %s", skill_file.name)
    root_el = parse_xml_file(skill_file)
    if root_el is not None:
        for skill_el in root_el.iter("SkillInfo"):
            key = skill_el.get("Key") or ""
            strkey = skill_el.get("StrKey") or ""
            skill_name = skill_el.get("SkillName") or ""
            skill_desc = skill_el.get("SkillDesc") or ""
            learn_knowledge_key = skill_el.get("LearnKnowledgeKey") or ""
            max_level_str = skill_el.get("MaxLevel") or "1"

            try:
                max_level = int(max_level_str)
            except ValueError:
                max_level = 1

            # Skip duplicates
            if strkey and strkey in seen_skill_keys:
                continue
            if strkey:
                seen_skill_keys.add(strkey)

            # Skip if no localizable content
            if not skill_name and not skill_desc:
                continue

            skills.append(SkillItem(
                key=key,
                strkey=strkey,
                skill_name=skill_name,
                skill_desc=skill_desc,
                learn_knowledge_key=learn_knowledge_key,
                max_level=max_level,
            ))

    log.info("Found %d SkillInfo entries", len(skills))

    # Third pass: resolve LearnKnowledgeKey priority
    # When multiple skills have the same LearnKnowledgeKey, highest MaxLevel wins
    knowledge_key_to_skills: Dict[str, List[SkillItem]] = {}
    for skill in skills:
        if skill.learn_knowledge_key:
            if skill.learn_knowledge_key not in knowledge_key_to_skills:
                knowledge_key_to_skills[skill.learn_knowledge_key] = []
            knowledge_key_to_skills[skill.learn_knowledge_key].append(skill)

    # Resolve priorities
    for kn_key, skill_list in knowledge_key_to_skills.items():
        if len(skill_list) > 1:
            # Sort by max_level descending, pick highest
            skill_list.sort(key=lambda s: s.max_level, reverse=True)
            winner = skill_list[0]
            log.debug("Priority for %s: %s (MaxLevel=%d) wins over %d others",
                      kn_key, winner.strkey, winner.max_level, len(skill_list) - 1)
            for loser in skill_list[1:]:
                loser.use_own_desc = True

    # Link knowledge to skills
    linked_count = 0
    for skill in skills:
        if skill.learn_knowledge_key and not skill.use_own_desc:
            if skill.learn_knowledge_key in knowledge_map:
                skill.knowledge = knowledge_map[skill.learn_knowledge_key]
                linked_count += 1

    log.info("Linked %d skills to KnowledgeInfo", linked_count)

    return skills, knowledge_map

# ──────────────────────────────────────────────────────────────────────
# ROW GENERATION
# ──────────────────────────────────────────────────────────────────────
# (depth, text, needs_translation)
RowItem = Tuple[int, str, bool]

def emit_knowledge_child_rows(child: KnowledgeNode, depth: int) -> List[RowItem]:
    """Emit rows for a knowledge child (sub-skill unlocked at certain level)"""
    rows: List[RowItem] = []

    # Child name (with unlock level indicator)
    if child.name:
        rows.append((depth, child.name, True))

    # Child description
    if child.desc:
        rows.append((depth + 1, child.desc, True))

    # Recursively emit nested children (children of children)
    for nested in child.children:
        rows.extend(emit_knowledge_child_rows(nested, depth + 1))

    return rows

def emit_skill_rows(skill: SkillItem) -> List[RowItem]:
    """
    Generate rows for a single skill with proper nesting.

    Structure:
    - Depth 0: SkillName (Gold - parent skill)
    - Depth 1: Parent description (Light blue)
    - Depth 1: Child knowledge Name (Light blue - sub-skills)
    - Depth 2: Child knowledge Desc (Light green)
    """
    rows: List[RowItem] = []

    # Depth 0: SkillName (parent)
    if skill.skill_name:
        rows.append((0, skill.skill_name, True))

    if skill.knowledge and not skill.use_own_desc:
        # Has linked KnowledgeInfo
        kn = skill.knowledge

        # Depth 1: Parent Knowledge Desc (or Skill Desc if different)
        if kn.desc:
            rows.append((1, kn.desc, True))
        elif skill.skill_desc:
            rows.append((1, skill.skill_desc, True))

        # Emit children (sub-skills) - these are other KnowledgeInfo that reference this one
        for child in kn.children:
            rows.extend(emit_knowledge_child_rows(child, 1))
    else:
        # No knowledge link or lost priority - use SkillDesc directly
        if skill.skill_desc:
            rows.append((1, skill.skill_desc, True))

    return rows

def emit_rows(skills: List[SkillItem]) -> List[RowItem]:
    """Generate all rows from skill list."""
    rows: List[RowItem] = []

    for skill in skills:
        rows.extend(emit_skill_rows(skill))

    # Postprocess: drop empty rows (whitespace-only text)
    rows = [(d, t, n) for (d, t, n) in rows if t and t.strip()]

    return rows

# ──────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ──────────────────────────────────────────────────────────────────────
def load_language_tables(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    """
    Load all non-Korean language tables with normalized placeholder keys.
    Returns: {lang_code: {normalized_korean: (translation, stringid)}}
    """
    tables: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}
        duplicates_improved = 0

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            normalized_origin = normalize_placeholders(origin)

            if normalized_origin in tbl:
                existing_tr, existing_sid = tbl[normalized_origin]
                existing_is_good = is_good_translation(existing_tr)
                new_is_good = is_good_translation(tr)

                if new_is_good and not existing_is_good:
                    tbl[normalized_origin] = (tr, sid)
                    duplicates_improved += 1
            else:
                tbl[normalized_origin] = (tr, sid)

        tables[lang] = tbl
        log.info("Language %s loaded - %d entries", lang.upper(), len(tbl))

    return tables

# ──────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ──────────────────────────────────────────────────────────────────────
_depth0_fill = PatternFill("solid", fgColor="FFD700")  # Gold for skill names
_depth0_font = Font(bold=True, size=12)
_depth0_row_height = 35

_depth1_fill = PatternFill("solid", fgColor="B4C6E7")  # Light blue for knowledge names
_depth1_font = Font(bold=True, size=11)
_depth1_row_height = 25

_depth2_fill = PatternFill("solid", fgColor="E2EFDA")  # Light green for descriptions
_depth2_font = Font(size=10)
_depth2_row_height = None  # Auto

_depth3_fill = PatternFill("solid", fgColor="FCE4D6")  # Light orange for sub-skills
_depth3_font = Font(bold=True, size=10)

_depth4_fill = PatternFill("solid", fgColor="DDEBF7")  # Light blue for sub-skill desc
_depth4_font = Font(size=10)

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")

_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def _get_style_for_depth(depth: int) -> Tuple[PatternFill, Font, Optional[float]]:
    if depth == 0:
        return _depth0_fill, _depth0_font, _depth0_row_height
    elif depth == 1:
        return _depth1_fill, _depth1_font, _depth1_row_height
    elif depth == 2:
        return _depth2_fill, _depth2_font, _depth2_row_height
    elif depth == 3:
        return _depth3_fill, _depth3_font, None
    else:
        return _depth4_fill, _depth4_font, None

# ──────────────────────────────────────────────────────────────────────
# AUTO-FIT HELPER
# ──────────────────────────────────────────────────────────────────────
def autofit_worksheet(ws, min_width: int = 10, max_width: int = 80, row_height_per_line: float = 15.0) -> None:
    """
    Auto-fit column widths and row heights based on cell content.
    """
    # Calculate optimal column widths
    col_widths: Dict[str, float] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                content_len = len(str(cell.value))
                width = min(max(content_len * 1.1 + 2, min_width), max_width)
                col_widths[col_letter] = max(col_widths.get(col_letter, min_width), width)

    # Apply column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Calculate optimal row heights (considering text wrap)
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        max_lines = 1
        for cell in row:
            if cell.value and cell.alignment and cell.alignment.wrap_text:
                col_letter = get_column_letter(cell.column)
                col_width = col_widths.get(col_letter, max_width)
                chars_per_line = max(col_width * 0.9, 10)
                content_len = len(str(cell.value))
                lines = max(1, int(content_len / chars_per_line) + 1)
                max_lines = max(max_lines, lines)
        # Set row height (minimum 20, scale by lines)
        ws.row_dimensions[row_idx].height = max(20, max_lines * row_height_per_line)

# ──────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ──────────────────────────────────────────────────────────────────────
def write_workbook(
    rows: List[RowItem],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with ONE sheet (Skills).

    Columns:
      - Original (KR)
      - English (ENG)
      - Translation (OTHER) [skipped for ENG workbook]
      - STATUS
      - COMMENT
      - STRINGID
      - SCREENSHOT
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"

    is_eng = lang_code.lower() == "eng"

    # ─── Header row ───────────────────────────────────────────────
    headers: List = []
    h1 = ws.cell(1, 1, "Original (KR)")
    h2 = ws.cell(1, 2, "English (ENG)")
    headers.extend([h1, h2])

    if not is_eng:
        h3 = ws.cell(1, 3, f"Translation ({lang_code.upper()})")
        headers.append(h3)

    start_extra_col = len(headers) + 1
    extra_names = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    for idx, name in enumerate(extra_names, start=start_extra_col):
        headers.append(ws.cell(1, idx, name))

    for hcell in headers:
        hcell.font = _header_font
        hcell.fill = _header_fill
        hcell.alignment = Alignment(horizontal="center", vertical="center")
        hcell.border = _border
    ws.row_dimensions[1].height = 25

    # ─── Column widths / visibility ─────────────────────────────────
    ws.column_dimensions["A"].width = 50  # Original (KR)
    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["B"].width = 50  # English (ENG)
    ws.column_dimensions["B"].hidden = not is_eng  # Hide ENG column for non-English workbooks
    if not is_eng:
        ws.column_dimensions["C"].width = 50  # Translation
        ws.column_dimensions["D"].width = 12  # STATUS
        ws.column_dimensions["E"].width = 40  # COMMENT
        ws.column_dimensions["F"].width = 20  # STRINGID
        ws.column_dimensions["G"].width = 20  # SCREENSHOT
    else:
        ws.column_dimensions["C"].width = 12  # STATUS
        ws.column_dimensions["D"].width = 40  # COMMENT
        ws.column_dimensions["E"].width = 20  # STRINGID
        ws.column_dimensions["F"].width = 20  # SCREENSHOT

    # ─── Data validation for STATUS ───────────────────────────────
    status_col = 4 if not is_eng else 3
    dv = DataValidation(
        type="list",
        formula1='"ISSUE,NO ISSUE,BLOCKED"',
        allow_blank=True,
    )
    dv.error = "Invalid status"
    dv.prompt = "Select status"
    ws.add_data_validation(dv)

    # ─── Write data rows ──────────────────────────────────────────
    # Deduplication DISABLED - keep all rows

    for row_idx, (depth, text, needs_trans) in enumerate(rows, start=2):
        normalized = normalize_placeholders(text)
        eng_tr, sid = eng_tbl.get(normalized, ("", ""))
        loc_tr = ""
        if lang_tbl:
            loc_tr, sid = lang_tbl.get(normalized, (loc_tr, sid))

        actual_row = row_idx  # Direct row index (no dedup skipping)

        fill, font, row_height = _get_style_for_depth(depth)
        indent = depth

        # Column A: Original (KR)
        c1 = ws.cell(actual_row, 1, text)
        c1.fill = fill
        c1.font = font
        c1.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c1.border = _border

        # Column B: English (ENG)
        c2 = ws.cell(actual_row, 2, eng_tr)
        c2.fill = fill
        c2.font = font
        c2.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c2.border = _border

        col_offset = 2

        # Column C: Translation (if not ENG)
        if not is_eng:
            c3 = ws.cell(actual_row, 3, loc_tr)
            c3.fill = fill
            c3.font = font
            c3.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
            c3.border = _border
            col_offset = 3

        # STATUS, COMMENT, STRINGID, SCREENSHOT
        for extra_idx, val in enumerate(["", "", sid, ""], start=col_offset + 1):
            cell = ws.cell(actual_row, extra_idx, val)
            cell.border = _border
            cell.alignment = Alignment(vertical="center")
            # STRINGID as text to prevent scientific notation
            if extra_idx == col_offset + 3:  # STRINGID column
                cell.number_format = '@'

        # Apply STATUS data validation
        dv.add(ws.cell(actual_row, status_col))

    # ─── Auto-fit columns and rows ────────────────────────────────
    autofit_worksheet(ws)

    # ─── Save ─────────────────────────────────────────────────────
    wb.save(out_path)
    log.info("Saved: %s (%d rows)", out_path.name, len(rows))

# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 70)
    log.info("Skill LQA Extractor - v1.0")
    log.info("=" * 70)

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    # 1. Extract Skill and Knowledge data
    skills, knowledge_map = extract_skill_data(RESOURCE_FOLDER)
    if not skills:
        log.warning("No Skill data found!")
        return

    # 2. Generate rows
    rows = emit_rows(skills)
    log.info("Generated %d rows", len(rows))

    # 3. Load language tables
    lang_tables = load_language_tables(LANGUAGE_FOLDER)
    eng_tbl = lang_tables.get("eng", {})

    if not eng_tbl:
        log.warning("English language table not found!")

    # 4. Write workbooks (one per language)
    # Always write English
    write_workbook(
        rows, eng_tbl, None, "eng",
        OUTPUT_FOLDER / "LQA_Skill_ENG.xlsx"
    )

    # Write other languages
    for lang_code, lang_tbl in lang_tables.items():
        if lang_code.lower() == "eng":
            continue
        write_workbook(
            rows, eng_tbl, lang_tbl, lang_code,
            OUTPUT_FOLDER / f"LQA_Skill_{lang_code.upper()}.xlsx"
        )

    log.info("=" * 70)
    log.info("Done! Output folder: %s", OUTPUT_FOLDER)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
