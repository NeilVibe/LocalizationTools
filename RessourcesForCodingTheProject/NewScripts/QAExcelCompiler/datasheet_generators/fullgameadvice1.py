#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GameAdvice-data extractor – v1.0
--------------------------------
Extracts GameAdviceGroupInfo / GameAdviceInfo entries (tutorial tips, help text).

Structure:
- GameAdviceGroupInfo (parent): StrKey, GroupName
  - GameAdviceInfo (child): StrKey, Title, Desc

Output:
- ONE Excel sheet with all data
- Parent-child indentation (depth 0 = group, depth 1 = item title, depth 2 = item desc)
- Columns: Original (KR) | English (ENG) | Translation (LOC) | STATUS | COMMENT | STRINGID | SCREENSHOT
"""

from __future__ import annotations

import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from lxml import etree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────
RESOURCE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo"
)
LANGUAGE_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
)

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "GameAdvice_LQA_All"
LOG_FILE = base_path / "gameadvice_scan.log"

# ──────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────
log = logging.getLogger("GameAdviceLQA")
log.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
_console_handler.setLevel(logging.INFO)

log.addHandler(_console_handler)
log.propagate = False

# ──────────────────────────────────────────────────────────────────────
# NORMALIZATION
# ──────────────────────────────────────────────────────────────────────
_placeholder_suffix_re = re.compile(r'\{([^#}]+)#[^}]+\}')
_whitespace_re = re.compile(r'\s+', flags=re.UNICODE)

def normalize_placeholders(text: str) -> str:
    """
    1) Remove '#…' suffix inside {...} placeholders.
    2) Collapse all whitespace (space, tab, NBSP, newline, etc.) to ONE space.
    3) Trim leading/trailing spaces.
    """
    if not text:
        return ""
    text = _placeholder_suffix_re.sub(r'{\1}', text)
    text = _whitespace_re.sub(' ', text).strip()
    return text

# ──────────────────────────────────────────────────────────────────────
# KOREAN DETECTION
# ──────────────────────────────────────────────────────────────────────
_korean_re = re.compile(r'[\uAC00-\uD7AF]')

def contains_korean(text: str) -> bool:
    return bool(_korean_re.search(text))

def is_good_translation(text: str) -> bool:
    return bool(text) and not contains_korean(text)

# ──────────────────────────────────────────────────────────────────────
# XML SANITIZATION
# ──────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r"&(?!lt;|gt;|amp;|apos;|quot;)")

def _fix_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)

def _escape_newlines_in_seg(txt: str) -> str:
    def repl(m: re.Match) -> str:
        seg = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r", "")
        return f"<seg>{seg}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, txt, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = _fix_entities(raw)
    raw = _escape_newlines_in_seg(raw)

    # Escape stray < inside attribute values
    raw = re.sub(
        r'="([^"]*?<[^"]*)"',
        lambda m: '="' + m.group(1).replace("<", "&lt;") + '"',
        raw,
    )
    raw = re.sub(
        r'="([^"]*?&[^ltgapoqu"][^"]*)"',
        lambda m: '="' + m.group(1).replace("&", "&amp;") + '"',
        raw,
    )

    # Fix orphan/broken close tags
    tag_stack: List[str] = []
    o_re = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    c_re = re.compile(r"</([A-Za-z0-9_]+)>")

    fixed: List[str] = []
    for line in raw.splitlines():
        s = line.strip()
        m_open = o_re.match(s)
        if m_open:
            tag_stack.append(m_open.group(1))
            fixed.append(line)
            continue
        m_close = c_re.match(s)
        if m_close:
            while tag_stack and tag_stack[-1] != m_close.group(1):
                fixed.append(f"</{tag_stack.pop()}>")
            if tag_stack:
                tag_stack.pop()
            fixed.append(line)
            continue
        if s.startswith("</>") and tag_stack:
            fixed.append(line.replace("</>", f"</{tag_stack.pop()}>"))
            continue
        fixed.append(line)
    while tag_stack:
        fixed.append(f"</{tag_stack.pop()}>")
    return "\n".join(fixed)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None

    wrapped = f"<ROOT>\n{sanitize_xml(raw)}\n</ROOT>"
    try:
        return ET.fromstring(wrapped.encode("utf-8"), parser=ET.XMLParser(huge_tree=True))
    except ET.XMLSyntaxError:
        log.debug("Strict parse failed: %s – retry with recover", path)
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(huge_tree=True, recover=True),
            )
        except ET.XMLSyntaxError:
            log.exception("Recovery parse failed: %s", path)
            return None

# ──────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────
def iter_xml_files(root: Path) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            if fn.lower().endswith(".xml"):
                yield Path(dp) / fn

# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ──────────────────────────────────────────────────────────────────────
@dataclass
class AdviceItem:
    strkey: str
    title: str
    desc: str

@dataclass
class AdviceGroup:
    strkey: str
    group_name: str
    items: List[AdviceItem] = field(default_factory=list)

# ──────────────────────────────────────────────────────────────────────
# EXTRACTION
# ──────────────────────────────────────────────────────────────────────
def extract_gameadvice_data(folder: Path) -> List[AdviceGroup]:
    """
    Scan StaticInfo folder for GameAdviceGroupInfo/GameAdviceInfo elements.
    Returns list of AdviceGroup with nested AdviceItems.
    """
    groups: List[AdviceGroup] = []
    seen_group_keys: set = set()
    seen_item_keys: set = set()

    log.info("Scanning for GameAdvice data in: %s", folder)

    for path in sorted(iter_xml_files(folder)):
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        # Find all GameAdviceGroupInfo elements
        for group_el in root_el.iter("GameAdviceGroupInfo"):
            strkey = group_el.get("StrKey") or ""
            group_name = group_el.get("GroupName") or ""

            # Skip duplicates
            if strkey and strkey in seen_group_keys:
                continue
            if strkey:
                seen_group_keys.add(strkey)

            group = AdviceGroup(strkey=strkey, group_name=group_name)

            # Find GameAdviceInfo children
            for item_el in group_el.iter("GameAdviceInfo"):
                item_strkey = item_el.get("StrKey") or ""
                title = item_el.get("Title") or ""
                desc = item_el.get("Desc") or ""

                # Skip duplicates
                if item_strkey and item_strkey in seen_item_keys:
                    continue
                if item_strkey:
                    seen_item_keys.add(item_strkey)

                # Skip if no content
                if not title and not desc:
                    continue

                group.items.append(AdviceItem(
                    strkey=item_strkey,
                    title=title,
                    desc=desc,
                ))

            # Only add group if it has items or a name
            if group.items or group.group_name:
                groups.append(group)

    log.info("Found %d groups with %d total items",
             len(groups), sum(len(g.items) for g in groups))

    return groups

# ──────────────────────────────────────────────────────────────────────
# ROW GENERATION
# ──────────────────────────────────────────────────────────────────────
# (depth, text, needs_translation)
RowItem = Tuple[int, str, bool]

def emit_rows(groups: List[AdviceGroup]) -> List[RowItem]:
    """Generate rows with proper indentation."""
    rows: List[RowItem] = []

    for group in groups:
        # Emit group name (depth 0)
        if group.group_name:
            rows.append((0, group.group_name, True))

        # Emit items
        for item in group.items:
            # Title (depth 1)
            if item.title:
                rows.append((1, item.title, True))

            # Description (depth 2)
            if item.desc:
                rows.append((2, item.desc, True))

    # Postprocess: drop empty rows (whitespace-only text)
    rows = [(d, t, n) for (d, t, n) in rows if t and t.strip()]

    return rows

# ──────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ──────────────────────────────────────────────────────────────────────
def load_language_tables(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    """
    Load all non-Korean language tables with normalized placeholder keys.
    Returns: {lang_code: {normalized_korean: (translation, stringid)}}
    """
    tables: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for path in iter_xml_files(folder):
        stem = path.stem.lower()
        if not stem.startswith("languagedata_"):
            continue
        if stem.endswith("kor"):
            continue

        lang = stem.split("_", 1)[1]
        root_el = parse_xml_file(path)
        if root_el is None:
            continue

        tbl: Dict[str, Tuple[str, str]] = {}
        duplicates_improved = 0

        for loc in root_el.iter("LocStr"):
            origin = loc.get("StrOrigin") or ""
            tr = loc.get("Str") or ""
            sid = loc.get("StringId") or ""

            if not origin:
                continue

            normalized_origin = normalize_placeholders(origin)

            if normalized_origin in tbl:
                existing_tr, existing_sid = tbl[normalized_origin]
                existing_is_good = is_good_translation(existing_tr)
                new_is_good = is_good_translation(tr)

                if new_is_good and not existing_is_good:
                    tbl[normalized_origin] = (tr, sid)
                    duplicates_improved += 1
            else:
                tbl[normalized_origin] = (tr, sid)

        tables[lang] = tbl
        log.info("Language %s loaded – %d entries", lang.upper(), len(tbl))

    return tables

# ──────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ──────────────────────────────────────────────────────────────────────
_depth0_fill = PatternFill("solid", fgColor="FFD700")  # Gold for groups
_depth0_font = Font(bold=True, size=12)
_depth0_row_height = 35

_depth1_fill = PatternFill("solid", fgColor="B4C6E7")  # Light blue for titles
_depth1_font = Font(bold=True, size=11)
_depth1_row_height = 25

_depth2_fill = PatternFill("solid", fgColor="E2EFDA")  # Light green for descriptions
_depth2_font = Font(size=10)
_depth2_row_height = None  # Auto

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill("solid", fgColor="4F81BD")

_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def _get_style_for_depth(depth: int) -> Tuple[PatternFill, Font, Optional[float]]:
    if depth == 0:
        return _depth0_fill, _depth0_font, _depth0_row_height
    elif depth == 1:
        return _depth1_fill, _depth1_font, _depth1_row_height
    else:
        return _depth2_fill, _depth2_font, _depth2_row_height

# ──────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ──────────────────────────────────────────────────────────────────────
def write_workbook(
    rows: List[RowItem],
    eng_tbl: Dict[str, Tuple[str, str]],
    lang_tbl: Optional[Dict[str, Tuple[str, str]]],
    lang_code: str,
    out_path: Path,
) -> None:
    """
    Write one workbook with ONE sheet (GameAdvice).

    Columns:
      - Original (KR)
      - English (ENG)
      - Translation (OTHER) [skipped for ENG workbook]
      - STATUS
      - COMMENT
      - STRINGID
      - SCREENSHOT
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "GameAdvice"

    is_eng = lang_code.lower() == "eng"

    # ─── Header row ───────────────────────────────────────────────
    headers: List = []
    h1 = ws.cell(1, 1, "Original (KR)")
    h2 = ws.cell(1, 2, "English (ENG)")
    headers.extend([h1, h2])

    if not is_eng:
        h3 = ws.cell(1, 3, f"Translation ({lang_code.upper()})")
        headers.append(h3)

    start_extra_col = len(headers) + 1
    extra_names = ["STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    for idx, name in enumerate(extra_names, start=start_extra_col):
        headers.append(ws.cell(1, idx, name))

    for hcell in headers:
        hcell.font = _header_font
        hcell.fill = _header_fill
        hcell.alignment = Alignment(horizontal="center", vertical="center")
        hcell.border = _border
    ws.row_dimensions[1].height = 25

    # ─── Column widths / visibility ─────────────────────────────────
    ws.column_dimensions["A"].width = 50  # Original (KR)
    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["B"].width = 50  # English (ENG)
    ws.column_dimensions["B"].hidden = not is_eng  # Hide ENG column for non-English workbooks
    if not is_eng:
        ws.column_dimensions["C"].width = 50  # Translation
        ws.column_dimensions["D"].width = 12  # STATUS
        ws.column_dimensions["E"].width = 40  # COMMENT
        ws.column_dimensions["F"].width = 20  # STRINGID
        ws.column_dimensions["G"].width = 20  # SCREENSHOT
    else:
        ws.column_dimensions["C"].width = 12  # STATUS
        ws.column_dimensions["D"].width = 40  # COMMENT
        ws.column_dimensions["E"].width = 20  # STRINGID
        ws.column_dimensions["F"].width = 20  # SCREENSHOT

    # ─── Data validation for STATUS ───────────────────────────────
    status_col = 4 if not is_eng else 3
    dv = DataValidation(
        type="list",
        formula1='"ISSUE,NO ISSUE,BLOCKED"',
        allow_blank=True,
    )
    dv.error = "Invalid status"
    dv.prompt = "Select status"
    ws.add_data_validation(dv)

    # ─── Write data rows ──────────────────────────────────────────
    seen_rows: set = set()  # Deduplication: (korean, translation, stringid)

    for row_idx, (depth, text, needs_trans) in enumerate(rows, start=2):
        normalized = normalize_placeholders(text)
        eng_tr, sid = eng_tbl.get(normalized, ("", ""))
        loc_tr = ""
        if lang_tbl:
            loc_tr, sid = lang_tbl.get(normalized, (loc_tr, sid))

        # Deduplication
        dedup_key = (text, eng_tr if is_eng else loc_tr, sid)
        if dedup_key in seen_rows:
            continue
        seen_rows.add(dedup_key)

        fill, font, row_height = _get_style_for_depth(depth)
        indent = depth

        # Column A: Original (KR)
        c1 = ws.cell(row_idx, 1, text)
        c1.fill = fill
        c1.font = font
        c1.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c1.border = _border

        # Column B: English (ENG)
        c2 = ws.cell(row_idx, 2, eng_tr)
        c2.fill = fill
        c2.font = font
        c2.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
        c2.border = _border

        col_offset = 2

        # Column C: Translation (if not ENG)
        if not is_eng:
            c3 = ws.cell(row_idx, 3, loc_tr)
            c3.fill = fill
            c3.font = font
            c3.alignment = Alignment(indent=indent, wrap_text=True, vertical="center")
            c3.border = _border
            col_offset = 3

        # STATUS, COMMENT, STRINGID, SCREENSHOT
        for extra_idx, val in enumerate(["", "", sid, ""], start=col_offset + 1):
            cell = ws.cell(row_idx, extra_idx, val)
            cell.border = _border
            cell.alignment = Alignment(vertical="center")
            # STRINGID as text to prevent scientific notation
            if extra_idx == col_offset + 3:  # STRINGID column
                cell.number_format = '@'

        # Apply STATUS data validation
        dv.add(ws.cell(row_idx, status_col))

        # Row height
        if row_height:
            ws.row_dimensions[row_idx].height = row_height

    # ─── Save ─────────────────────────────────────────────────────
    wb.save(out_path)
    log.info("Saved: %s (%d rows)", out_path.name, len(rows))

# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 70)
    log.info("GameAdvice LQA Extractor – v1.0")
    log.info("=" * 70)

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    # 1. Extract GameAdvice data
    groups = extract_gameadvice_data(RESOURCE_FOLDER)
    if not groups:
        log.warning("No GameAdvice data found!")
        return

    # 2. Generate rows
    rows = emit_rows(groups)
    log.info("Generated %d rows", len(rows))

    # 3. Load language tables
    lang_tables = load_language_tables(LANGUAGE_FOLDER)
    eng_tbl = lang_tables.get("eng", {})

    if not eng_tbl:
        log.warning("English language table not found!")

    # 4. Write workbooks (one per language)
    # Always write English
    write_workbook(
        rows, eng_tbl, None, "eng",
        OUTPUT_FOLDER / "LQA_GameAdvice_ENG.xlsx"
    )

    # Write other languages
    for lang_code, lang_tbl in lang_tables.items():
        if lang_code.lower() == "eng":
            continue
        write_workbook(
            rows, eng_tbl, lang_tbl, lang_code,
            OUTPUT_FOLDER / f"LQA_GameAdvice_{lang_code.upper()}.xlsx"
        )

    log.info("=" * 70)
    log.info("Done! Output folder: %s", OUTPUT_FOLDER)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
