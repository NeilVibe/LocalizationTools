#!/usr/bin/env python3
"""
QA Excel Compiler (Robust Version with Image Compilation)
==========================================================
Compiles QA tester Excel files into master sheets with image consolidation.

Works with ANY Excel structure - finds columns dynamically.

Usage:
    python3 compile_qa.py

Input:  QAfolder/{Username}_{Category}/
        - One .xlsx file per folder
        - Images referenced in SCREENSHOT column

Output: Masterfolder/Master_{Category}.xlsx
        Masterfolder/Images/{Username}_{Category}_{original}.png

Categories: Quest, Knowledge, Item, Node, System

Comment Handling:
- REPLACE mode: New comments replace old (no append/concatenation)
- Timestamp: Uses file's last modified time (not current time)
- Format: comment text\\n---\\nstringid:\\n<value>\\n(updated: YYMMDD HHMM)
- Duplicate check: Splits on '---' delimiter to compare original comment text
"""

import os
import sys
import shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# === CONFIGURATION ===
# Handle PyInstaller frozen state - exe unpacks to temp dir, need exe's actual location
if getattr(sys, 'frozen', False):
    # Running as compiled executable (.exe)
    SCRIPT_DIR = Path(sys.executable).parent
else:
    # Running as normal Python script
    SCRIPT_DIR = Path(__file__).parent
QA_FOLDER = SCRIPT_DIR / "QAfolder"

# Language-separated Master folders
MASTER_FOLDER_EN = SCRIPT_DIR / "Masterfolder_EN"
MASTER_FOLDER_CN = SCRIPT_DIR / "Masterfolder_CN"
IMAGES_FOLDER_EN = MASTER_FOLDER_EN / "Images"
IMAGES_FOLDER_CN = MASTER_FOLDER_CN / "Images"

# Tester→Language mapping file
TESTER_MAPPING_FILE = SCRIPT_DIR / "languageTOtester_list.txt"

# Progress Tracker at root level (combines all)
TRACKER_PATH = SCRIPT_DIR / "LQA_Tester_ProgressTracker.xlsx"

# Migration folders (for Transfer feature)
QA_FOLDER_OLD = SCRIPT_DIR / "QAfolderOLD"
QA_FOLDER_NEW = SCRIPT_DIR / "QAfolderNEW"

CATEGORIES = ["Quest", "Knowledge", "Item", "Region", "System", "Character"]

# Translation column positions by category
# For matching rows during transfer
TRANSLATION_COLS = {
    # Quest, Knowledge, Character, Region: Col 2 for ENG, Col 3 for other
    "Quest": {"eng": 2, "other": 3},
    "Knowledge": {"eng": 2, "other": 3},
    "Character": {"eng": 2, "other": 3},
    "Region": {"eng": 2, "other": 3},
    # Item: Col 5 for ENG (ItemName), Col 7 for other
    "Item": {"eng": 5, "other": 7},
    # System: Manually created sheets, Translation in Col 1
    # Structure: Col1=Translation, Col2=STATUS, Col3=COMMENT, Col4=STRINGID, Col5=SCREENSHOT
    "System": {"eng": 1, "other": 1},
}

# Item Description column positions (for Item-specific matching)
# Used alongside TRANSLATION_COLS for stricter Item matching
ITEM_DESC_COLS = {
    # Item: Col 6 for ENG (ItemDesc), Col 8 for other
    "eng": 6,
    "other": 8,
}

# Supported image extensions
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}

# Valid STATUS values (only these count as "filled")
VALID_STATUS = ["ISSUE", "NO ISSUE", "BLOCKED", "KOREAN"]

# Valid MANAGER STATUS values (for manager workflow)
VALID_MANAGER_STATUS = ["FIXED", "REPORTED", "CHECKING", "NON-ISSUE"]

# === TRACKER CONFIGURATION ===
TRACKER_FILENAME = "LQA_Tester_ProgressTracker.xlsx"

TRACKER_STYLES = {
    "title_color": "FFD700",       # Gold
    "header_color": "87CEEB",      # Light blue
    "subheader_color": "D3D3D3",   # Light gray
    "alt_row_color": "F5F5F5",     # Alternating gray
    "total_row_color": "E6E6E6",   # Total row gray
    "border_color": "000000",      # Black
}

CHART_COLORS = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5", "A5A5A5"]

# Characters that trigger Excel formula interpretation
EXCEL_FORMULA_CHARS = ('=', '+', '@')


def sanitize_for_excel(value):
    """
    Prevent Excel formula injection from user content.

    Excel interprets cells starting with =, +, @ as formulas.
    Prepending a single quote (') tells Excel to treat it as text.
    The quote is invisible in the cell but prevents formula parsing.

    Args:
        value: Cell value to sanitize

    Returns:
        Sanitized value safe for Excel
    """
    if not value:
        return value
    text = str(value)
    if text and text[0] in EXCEL_FORMULA_CHARS:
        return "'" + text  # Excel's native escape - invisible in cell
    return text


def contains_korean(text):
    """
    Check if text contains Korean characters (Hangul).

    Korean Unicode ranges:
    - Hangul Syllables: U+AC00 to U+D7AF (most common, 11,172 chars)
    - Hangul Jamo: U+1100 to U+11FF (archaic/combining)
    - Hangul Compatibility Jamo: U+3130 to U+318F

    Args:
        text: Text to check

    Returns:
        True if text contains Korean characters, False otherwise
    """
    if not text:
        return False
    for char in str(text):
        # Hangul Syllables (most common)
        if '\uAC00' <= char <= '\uD7AF':
            return True
        # Hangul Jamo
        if '\u1100' <= char <= '\u11FF':
            return True
        # Hangul Compatibility Jamo
        if '\u3130' <= char <= '\u318F':
            return True
    return False


def count_words_english(text):
    """
    Count words in English text.

    Splits by whitespace and counts tokens.
    Returns 0 if text contains Korean (untranslated).

    Args:
        text: Text to count words in

    Returns:
        int: Word count (0 if Korean detected or empty)
    """
    if not text or contains_korean(text):
        return 0
    return len(str(text).split())


def count_chars_chinese(text):
    """
    Count characters in Chinese text (excluding whitespace).

    For CJK languages, character count is more meaningful than word count.
    Returns 0 if text contains Korean (untranslated).

    Args:
        text: Text to count characters in

    Returns:
        int: Character count excluding whitespace (0 if Korean detected or empty)
    """
    if not text or contains_korean(text):
        return 0
    # Remove all whitespace characters
    cleaned = str(text).replace(" ", "").replace("\n", "").replace("\t", "").replace("\r", "")
    return len(cleaned)


def find_file_in_folder(filename, folder):
    """
    Search for file in folder (case-insensitive).

    Used by hyperlink auto-fixer to find actual file when
    only the filename text is present without a hyperlink.

    Args:
        filename: Filename to search for
        folder: Path to folder to search in

    Returns:
        Actual filename if found (preserving case), None otherwise
    """
    if not folder or not Path(folder).exists():
        return None
    filename_lower = str(filename).lower()
    try:
        for f in os.listdir(folder):
            if f.lower() == filename_lower:
                return f
    except OSError:
        pass
    return None


def repair_excel_filters(filepath):
    """
    Repair Excel file by stripping corrupted auto-filter XML.

    Excel files are ZIP archives. This extracts, removes autoFilter
    elements from worksheet XML, and repacks to a temp file.

    Args:
        filepath: Path to corrupted Excel file

    Returns:
        Path to repaired temp file, or None if repair failed
    """
    import zipfile
    import tempfile
    import re

    try:
        # Create temp file for repaired Excel
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)

        with zipfile.ZipFile(filepath, 'r') as zip_in:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for item in zip_in.namelist():
                    data = zip_in.read(item)

                    # Strip autoFilter from worksheet XML files
                    if item.startswith('xl/worksheets/') and item.endswith('.xml'):
                        content = data.decode('utf-8')
                        # Remove autoFilter elements (including corrupted ones)
                        content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                        content = re.sub(r'<autoFilter[^/]*/>', '', content)
                        data = content.encode('utf-8')

                    zip_out.writestr(item, data)

        return Path(temp_path)
    except Exception as e:
        print(f"    Repair failed: {e}")
        return None


def safe_load_workbook(filepath, **kwargs):
    """
    Safely load an Excel workbook, handling common corruption issues.

    ROBUST: Auto-repairs corrupted auto-filters by stripping them from XML.

    Handles:
    - Corrupted auto-filter values (ValueError: must be numerical or wildcard)
    - Invalid XML in worksheets

    Args:
        filepath: Path to the Excel file
        **kwargs: Additional arguments passed to openpyxl.load_workbook

    Returns:
        Workbook object

    Raises:
        Exception for unrecoverable errors
    """
    try:
        # First attempt: normal load
        return openpyxl.load_workbook(filepath, **kwargs)
    except ValueError as e:
        error_msg = str(e)

        # Handle corrupted filter values - AUTO REPAIR
        if "numerical or a string containing a wildcard" in error_msg or "could not read worksheets" in error_msg:
            print(f"    WARNING: Corrupted filters detected, auto-repairing...")

            # Repair by stripping autoFilter XML
            repaired_path = repair_excel_filters(filepath)
            if repaired_path:
                try:
                    wb = openpyxl.load_workbook(repaired_path, **kwargs)
                    # Clean up temp file
                    try:
                        os.unlink(repaired_path)
                    except:
                        pass
                    print(f"    SUCCESS: File repaired and loaded (filters stripped)")
                    return wb
                except Exception as repair_error:
                    print(f"    Repair load failed: {repair_error}")
                    try:
                        os.unlink(repaired_path)
                    except:
                        pass

            # If repair failed, give manual instructions
            print(f"    ERROR: Auto-repair failed for '{filepath.name}'")
            print(f"           Manual fix: Open in Excel → Data → Clear → Save")
            raise ValueError(f"Corrupted Excel file: {filepath.name}. Auto-repair failed.")

        # Re-raise other ValueErrors
        raise
    except Exception as e:
        # Log and re-raise other exceptions
        print(f"    ERROR loading '{filepath}': {e}")
        raise


def load_tester_mapping():
    """
    Load tester→language mapping from languageTOtester_list.txt.

    File format:
        ENG
        김동헌
        황하연
        ...

        ZHO-CN
        김춘애
        최문석
        ...

    Returns: Dict {tester_name: "EN" or "CN"}
    """
    mapping = {}
    current_lang = None

    if not TESTER_MAPPING_FILE.exists():
        print(f"WARNING: Mapping file not found: {TESTER_MAPPING_FILE}")
        return mapping

    with open(TESTER_MAPPING_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line == "ENG":
                current_lang = "EN"
            elif line == "ZHO-CN":
                current_lang = "CN"
            elif current_lang:
                mapping[line] = current_lang

    print(f"  Loaded {len(mapping)} tester→language mappings")
    return mapping


def get_master_folder(username, tester_mapping):
    """Get the correct master folder based on tester's language."""
    lang = tester_mapping.get(username, "EN")  # Default to EN if not found
    if lang == "CN":
        return MASTER_FOLDER_CN, IMAGES_FOLDER_CN
    return MASTER_FOLDER_EN, IMAGES_FOLDER_EN


def ensure_master_folders():
    """Create both EN and CN master folders if they don't exist."""
    MASTER_FOLDER_EN.mkdir(exist_ok=True)
    MASTER_FOLDER_CN.mkdir(exist_ok=True)
    IMAGES_FOLDER_EN.mkdir(exist_ok=True)
    IMAGES_FOLDER_CN.mkdir(exist_ok=True)


def discover_qa_folders():
    """
    Find all QA folders and parse their metadata.

    Folder format: {Username}_{Category}/
    Each folder contains: one .xlsx file + images

    Returns: List of dicts with {folder_path, xlsx_path, username, category, images}
    """
    results = []

    if not QA_FOLDER.exists():
        print(f"ERROR: QAfolder not found: {QA_FOLDER}")
        return results

    for folder in QA_FOLDER.iterdir():
        if not folder.is_dir():
            continue

        # Skip hidden/temp folders
        if folder.name.startswith('.') or folder.name.startswith('~'):
            continue

        # Parse folder name: {Username}_{Category}
        parts = folder.name.split("_")
        if len(parts) < 2:
            print(f"WARN: Invalid folder name format: {folder.name} (expected: Username_Category)")
            continue

        username = parts[0]
        category = "_".join(parts[1:])  # Handle categories with underscores (though unlikely)

        if category not in CATEGORIES:
            print(f"WARN: Unknown category '{category}' in folder {folder.name}")
            continue

        # Find xlsx file (must be exactly 1)
        xlsx_files = list(folder.glob("*.xlsx"))
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

        if len(xlsx_files) == 0:
            print(f"WARN: No xlsx in folder: {folder.name}")
            continue
        if len(xlsx_files) > 1:
            print(f"WARN: Multiple xlsx in folder: {folder.name}, using first: {xlsx_files[0].name}")

        xlsx_path = xlsx_files[0]

        # Find images
        images = [f for f in folder.iterdir()
                  if f.is_file() and f.suffix.lower() in IMAGE_EXTENSIONS]

        results.append({
            "folder_path": folder,
            "xlsx_path": xlsx_path,
            "username": username,
            "category": category,
            "images": images
        })

    return results


def copy_images_with_unique_names(qa_folder_info, images_folder):
    """
    Copy images from QA folder to Images/ with original names.

    Args:
        qa_folder_info: Dict with {folder_path, username, category, images}
        images_folder: Target Images folder (EN or CN)

    Returns: Dict mapping original_name -> original_name
    """
    images = qa_folder_info["images"]

    image_mapping = {}

    for img_path in images:
        original_name = img_path.name
        # Keep original name - hyperlink depends on it
        dest_path = images_folder / original_name

        # Copy image (overwrite if exists - same image from different users)
        shutil.copy2(img_path, dest_path)

        image_mapping[original_name] = original_name  # Same name

    if image_mapping:
        print(f"    Copied {len(image_mapping)} images to Images/")

    return image_mapping


def find_column_by_header(ws, header_name, case_insensitive=True):
    """
    Find column index by header name.

    Args:
        ws: Worksheet
        header_name: Header to search for
        case_insensitive: Match case-insensitively

    Returns: Column index (1-based) or None if not found
    """
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header_str = str(header).strip()
            if case_insensitive:
                if header_str.upper() == header_name.upper():
                    return col
            else:
                if header_str == header_name:
                    return col
    return None


def get_or_create_master(category, master_folder, template_file=None):
    """
    ALWAYS create fresh master from template file.

    FULL REBUILD: Delete old master and create fresh from first QA file.
    This ensures structure is always up-to-date with latest QA files.

    Manager status is preserved via dict (extracted before this function runs).

    Args:
        category: Category name (Quest, Knowledge, etc.)
        master_folder: Target Master folder (EN or CN)
        template_file: Path to first QA file to use as template

    Returns: openpyxl Workbook, master_path
    """
    master_path = master_folder / f"Master_{category}.xlsx"

    # ALWAYS delete old master and create fresh
    if master_path.exists():
        print(f"  Deleting old master: {master_path.name} (rebuilding fresh)")
        master_path.unlink()

    if template_file:
        print(f"  Creating new master from: {template_file.name}")
        wb = safe_load_workbook(template_file)

        # DELETE STATUS, COMMENT, SCREENSHOT, STRINGID columns entirely (CLEAN START)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # CRITICAL: Clear any AutoFilter from QA file (testers may have filters applied)
            if ws.auto_filter.ref:
                ws.auto_filter.ref = None

            # Find columns to delete (must delete from right to left to preserve indices)
            cols_to_delete = []

            status_col = find_column_by_header(ws, "STATUS")
            comment_col = find_column_by_header(ws, "COMMENT")
            screenshot_col = find_column_by_header(ws, "SCREENSHOT")
            stringid_col = find_column_by_header(ws, "STRINGID")

            if status_col:
                cols_to_delete.append(status_col)
            if comment_col:
                cols_to_delete.append(comment_col)
            if screenshot_col:
                cols_to_delete.append(screenshot_col)
            if stringid_col:
                cols_to_delete.append(stringid_col)

            # Sort descending (delete from right to left)
            cols_to_delete.sort(reverse=True)

            for col in cols_to_delete:
                ws.delete_cols(col)
                print(f"    Deleted column {col} from {sheet_name}")

        print(f"    Master cleaned: STATUS/COMMENT/SCREENSHOT/STRINGID removed")
        return wb, master_path
    else:
        print(f"  ERROR: No template file for {category}")
        return None, master_path


def get_or_create_user_comment_column(ws, username):
    """
    Find or create COMMENT_{username} column using MAX_COLUMN + 1.

    ROBUST: Always adds new columns at the far right (max_column + 1).
    Works with ANY Excel structure.

    BEAUTIFUL: Adds color, bold, and border formatting to header.

    Args:
        ws: Worksheet
        username: User identifier

    Returns: Column index (1-based)
    """
    col_name = f"COMMENT_{username}"

    # Check if column already exists
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip() == col_name:
            return col

    # MAX_COLUMN + 1: Add new column at the far right
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col)
    cell.value = col_name

    # Beautiful formatting for COMMENT_{User} header
    # Light blue background
    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    # Bold font
    cell.font = Font(bold=True, color="000000")
    # Center alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Nice border
    cell.border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )

    # Set column width for readability (only width, preserve hidden state of other columns)
    col_letter = get_column_letter(new_col)
    ws.column_dimensions[col_letter].width = 35
    # Explicitly ensure this new column is visible (don't inherit hidden state)
    ws.column_dimensions[col_letter].hidden = False

    print(f"    Created column: {col_name} at {get_column_letter(new_col)} (styled)")
    return new_col


def get_or_create_user_screenshot_column(ws, username, after_comment_col):
    """
    Find or create SCREENSHOT_{username} column immediately after COMMENT_{username}.

    Args:
        ws: Worksheet
        username: User identifier
        after_comment_col: Column index of COMMENT_{username} (screenshot goes right after)

    Returns: Column index (1-based)
    """
    col_name = f"SCREENSHOT_{username}"

    # Check if column already exists
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip() == col_name:
            return col

    # Insert new column right after COMMENT_{username}
    new_col = after_comment_col + 1

    # If there's already data after the comment column, we need to insert
    # Actually, since we're using MAX_COLUMN + 1 for COMMENT, SCREENSHOT should be MAX_COLUMN + 1 now
    # But to be safe, let's just use max_column + 1 for screenshot too
    new_col = ws.max_column + 1

    cell = ws.cell(row=1, column=new_col)
    cell.value = col_name

    # Light blue background for screenshot columns (matching COMMENT style)
    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    # Bold font
    cell.font = Font(bold=True, color="000000")
    # Center alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Nice border (matching COMMENT style)
    cell.border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )

    # Set column width
    col_letter = get_column_letter(new_col)
    ws.column_dimensions[col_letter].width = 40
    ws.column_dimensions[col_letter].hidden = False

    print(f"    Created column: {col_name} at {get_column_letter(new_col)} (styled)")
    return new_col


def get_or_create_user_status_column(ws, username, after_comment_col):
    """
    Find or create STATUS_{username} column immediately after COMMENT_{username}.

    This is for MANAGER workflow - managers mark FIXED/REPORTED/CHECKING.
    Adds a dropdown list (data validation) so managers can only select valid values.

    Args:
        ws: Worksheet
        username: User identifier
        after_comment_col: Column index of COMMENT_{username} (status goes right after)

    Returns: Column index (1-based)
    """
    col_name = f"STATUS_{username}"

    # Check if column already exists
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip() == col_name:
            # Ensure dropdown validation is up-to-date (might be old formula without NON-ISSUE)
            col_letter = get_column_letter(col)
            # Remove any existing validation for this column
            to_remove = []
            for existing_dv in ws.data_validations.dataValidation:
                if col_letter in str(existing_dv.sqref):
                    to_remove.append(existing_dv)
            for dv in to_remove:
                ws.data_validations.dataValidation.remove(dv)
            # Add fresh validation with current options (includes NON-ISSUE)
            dv = DataValidation(
                type="list",
                formula1='"FIXED,REPORTED,CHECKING,NON-ISSUE"',
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid Status",
                error="Please select: FIXED, REPORTED, CHECKING, or NON-ISSUE"
            )
            # Use actual row count + buffer instead of hardcoded 1000
            last_row = max(ws.max_row, 10) + 50  # Buffer for future rows
            dv.add(f"{col_letter}2:{col_letter}{last_row}")
            ws.add_data_validation(dv)
            return col

    # Add new column at max_column + 1
    new_col = ws.max_column + 1

    cell = ws.cell(row=1, column=new_col)
    cell.value = col_name

    # Light green background for manager STATUS columns (distinct from tester columns)
    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    # Bold font
    cell.font = Font(bold=True, color="000000")
    # Center alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Nice border (green tint)
    cell.border = Border(
        left=Side(style='medium', color='228B22'),
        right=Side(style='medium', color='228B22'),
        top=Side(style='medium', color='228B22'),
        bottom=Side(style='medium', color='228B22')
    )

    # Set column width
    col_letter = get_column_letter(new_col)
    ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions[col_letter].hidden = False

    # Add dropdown data validation for FIXED/REPORTED/CHECKING/NON-ISSUE
    # Use actual row count + buffer instead of hardcoded 1000 (prevents Excel repair warnings)
    last_row = max(ws.max_row, 10) + 50  # Buffer for future rows
    dv = DataValidation(
        type="list",
        formula1='"FIXED,REPORTED,CHECKING,NON-ISSUE"',
        allow_blank=True,
        showDropDown=False,  # False = show dropdown arrow, True = hide it (confusing API)
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select: FIXED, REPORTED, CHECKING, or NON-ISSUE",
        promptTitle="Manager Status",
        prompt="Select status: FIXED, REPORTED, CHECKING, or NON-ISSUE"
    )
    dv.add(f"{col_letter}2:{col_letter}{last_row}")
    ws.add_data_validation(dv)

    print(f"    Created column: {col_name} at {get_column_letter(new_col)} (manager status - dropdown)")
    return new_col


def get_or_create_tester_status_column(ws, username, after_comment_col):
    """
    Find or create TESTER_STATUS_{username} column immediately after COMMENT_{username}.

    This column stores the ORIGINAL status from the QA file (ISSUE, NO ISSUE, BLOCKED, KOREAN).
    It is HIDDEN by default - used for filtering logic, not for manager review.

    The column is placed between COMMENT_{username} and STATUS_{username} (manager status).

    Args:
        ws: Worksheet
        username: User identifier
        after_comment_col: Column index of COMMENT_{username} (tester status goes right after)

    Returns: Column index (1-based)
    """
    col_name = f"TESTER_STATUS_{username}"

    # Check if column already exists
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip() == col_name:
            return col

    # Add new column at max_column + 1
    new_col = ws.max_column + 1

    cell = ws.cell(row=1, column=new_col)
    cell.value = col_name

    # Gray background for TESTER_STATUS columns (hidden, for filtering only)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    # Bold font
    cell.font = Font(bold=True, color="000000")
    # Center alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Subtle border
    cell.border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )

    # Set column width and HIDE IT
    col_letter = get_column_letter(new_col)
    ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions[col_letter].hidden = True  # HIDDEN by default

    print(f"    Created column: {col_name} at {get_column_letter(new_col)} (tester status - HIDDEN)")
    return new_col


def format_comment(new_comment, string_id=None, existing_comment=None, file_mod_time=None):
    """
    Format comment with StringID and file modification time.

    Format (with stringid):
        <comment text>
        ---
        stringid:
        <stringid value>
        (updated: YYMMDD HHMM)

    Format (without stringid):
        <comment text>
        ---
        (updated: YYMMDD HHMM)

    REPLACE MODE: If comment text differs, REPLACE entirely (no append).
    DUPLICATE CHECK: Split on '---' delimiter to extract original comment for comparison.
    """
    if not new_comment or str(new_comment).strip() == "":
        return existing_comment  # Return existing if no new comment

    new_text = str(new_comment).strip()

    # Check if this exact comment text already exists (avoid re-update on re-run)
    if existing_comment:
        existing_str = str(existing_comment)
        # Split on --- delimiter to extract original comment
        if "\n---\n" in existing_str:
            existing_original = existing_str.split("\n---\n")[0].strip()
            if existing_original == new_text:
                return existing_comment  # Duplicate, skip

    # Format timestamp from file modification time (or fallback to now)
    if file_mod_time:
        timestamp = file_mod_time.strftime("%y%m%d %H%M")
    else:
        timestamp = datetime.now().strftime("%y%m%d %H%M")

    # Build formatted comment: text + delimiter + metadata
    if string_id and str(string_id).strip():
        # With stringid: comment -> --- -> stringid: -> value -> (updated)
        return f"{new_text}\n---\nstringid:\n{str(string_id).strip()}\n(updated: {timestamp})"
    else:
        # Without stringid: comment -> --- -> (updated)
        return f"{new_text}\n---\n(updated: {timestamp})"


def get_row_signature(ws, row, exclude_cols=None):
    """
    Get a signature for a row based on non-empty cells (excluding specified columns).
    Used for fallback row matching.

    Args:
        ws: Worksheet
        row: Row number
        exclude_cols: Set of column indices to exclude (STATUS, COMMENT, SCREENSHOT)

    Returns: Tuple of (col_index, value) for non-empty cells
    """
    if exclude_cols is None:
        exclude_cols = set()

    signature = []
    for col in range(1, ws.max_column + 1):
        if col in exclude_cols:
            continue
        val = ws.cell(row=row, column=col).value
        if val and str(val).strip():
            signature.append((col, str(val).strip()))

    return tuple(signature)


def find_matching_row_fallback(master_ws, qa_signature, exclude_cols, start_row=2):
    """
    Find a row in master that matches QA row by 2+ cell matching.

    Args:
        master_ws: Master worksheet
        qa_signature: Signature from QA row
        exclude_cols: Columns to exclude from matching
        start_row: Start searching from this row

    Returns: Row number or None
    """
    if len(qa_signature) < 2:
        return None  # Need at least 2 cells to match

    for row in range(start_row, master_ws.max_row + 1):
        master_sig = get_row_signature(master_ws, row, exclude_cols)

        # Count matching cells
        matches = 0
        for (col, val) in qa_signature:
            if (col, val) in master_sig:
                matches += 1

        # 2+ cell match = found
        if matches >= 2:
            return row

    return None


def sort_worksheet_az(ws, sort_column=1):
    """
    Sort worksheet rows A-Z by specified column.

    Used for EN Item category to match tester's A-Z sorted files.
    Preserves header row (row 1), sorts data rows (row 2+).

    Args:
        ws: Worksheet to sort
        sort_column: Column index to sort by (1-based, default=1 for column A)
    """
    from copy import copy

    # CRITICAL: Clear any AutoFilter before sorting (testers may have filters applied)
    if ws.auto_filter.ref:
        ws.auto_filter.ref = None

    # Get all data rows (skip header)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append({
                'value': cell.value,
                'font': copy(cell.font) if cell.font else None,
                'fill': copy(cell.fill) if cell.fill else None,
                'border': copy(cell.border) if cell.border else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
                'hyperlink': cell.hyperlink.target if cell.hyperlink else None
            })
        # Store sort key (first column value) and row data
        sort_key = ws.cell(row=row, column=sort_column).value
        sort_key = str(sort_key).lower() if sort_key else ""
        data_rows.append((sort_key, row_data))

    # Sort by first column (A-Z)
    data_rows.sort(key=lambda x: x[0])

    # Write back sorted data
    for new_row_idx, (sort_key, row_data) in enumerate(data_rows, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            cell = ws.cell(row=new_row_idx, column=col_idx)
            cell.value = cell_data['value']
            if cell_data['font']:
                cell.font = cell_data['font']
            if cell_data['fill']:
                cell.fill = cell_data['fill']
            if cell_data['border']:
                cell.border = cell_data['border']
            if cell_data['alignment']:
                cell.alignment = cell_data['alignment']
            if cell_data['hyperlink']:
                cell.hyperlink = cell_data['hyperlink']


def find_matching_row_item_fallback(master_ws, qa_ws, qa_row, start_row=2):
    """
    Item category fallback: Match by 4+ of first 6 columns.

    Used when STRINGID is not available for Item category.
    Compares first 6 columns between QA row and master rows.
    If 4 or more values match, it's considered a match.

    Args:
        master_ws: Master worksheet
        qa_ws: QA worksheet
        qa_row: Row number in QA worksheet
        start_row: Start searching from this row in master

    Returns: Row number or None
    """
    # Get first 6 column values from QA row
    qa_values = []
    for col in range(1, 7):  # Columns 1-6
        val = qa_ws.cell(row=qa_row, column=col).value
        if val is not None:
            qa_values.append((col, str(val).strip()))
        else:
            qa_values.append((col, None))

    # Search master for matching row
    for master_row in range(start_row, master_ws.max_row + 1):
        matches = 0
        for col, qa_val in qa_values:
            master_val = master_ws.cell(row=master_row, column=col).value
            if master_val is not None:
                master_val = str(master_val).strip()
            else:
                master_val = None

            # Both None or both equal = match
            if qa_val == master_val:
                matches += 1
            elif qa_val and master_val and qa_val == master_val:
                matches += 1

        # 4+ matches out of 6 = found
        if matches >= 4:
            return master_row

    return None


def process_sheet(master_ws, qa_ws, username, category, image_mapping=None, xlsx_path=None, manager_status=None):
    """
    Process a single sheet: copy COMMENT and SCREENSHOT from QA to master, collect STATUS stats.

    FULL REBUILD VERSION:
    - Master is ALWAYS fresh from first QA file (same structure)
    - Uses INDEX for row matching (no fallback needed)
    - Applies manager status by looking up COMMENT TEXT in dict
    - Creates columns: COMMENT_{User} → STATUS_{User} → SCREENSHOT_{User}

    Args:
        master_ws: Master worksheet
        qa_ws: QA worksheet
        username: User identifier
        category: Category name
        image_mapping: Dict mapping original_name -> new_name (for hyperlink transformation)
        xlsx_path: Path to QA xlsx file (for file modification time)
        manager_status: Dict of {comment_text: {user: status}} for this sheet (keyed by comment text)

    Returns: Dict with {comments: n, screenshots: n, stats: {...}, manager_restored: n}
    """
    if image_mapping is None:
        image_mapping = {}
    if manager_status is None:
        manager_status = {}

    # Get file modification time for timestamp
    file_mod_time = None
    if xlsx_path:
        file_mod_time = datetime.fromtimestamp(xlsx_path.stat().st_mtime)

    # Find columns dynamically in QA worksheet
    qa_status_col = find_column_by_header(qa_ws, "STATUS")
    qa_comment_col = find_column_by_header(qa_ws, "COMMENT")
    qa_screenshot_col = find_column_by_header(qa_ws, "SCREENSHOT")
    qa_stringid_col = find_column_by_header(qa_ws, "STRINGID")

    # Find or create user columns in master in correct order:
    # COMMENT_{username} → TESTER_STATUS_{username} (hidden) → STATUS_{username} → SCREENSHOT_{username}
    master_comment_col = get_or_create_user_comment_column(master_ws, username)
    master_tester_status_col = get_or_create_tester_status_column(master_ws, username, master_comment_col)
    master_user_status_col = get_or_create_user_status_column(master_ws, username, master_comment_col)
    master_screenshot_col = get_or_create_user_screenshot_column(master_ws, username, master_user_status_col)

    result = {
        "comments": 0,
        "screenshots": 0,
        "stats": {"issue": 0, "no_issue": 0, "blocked": 0, "korean": 0, "total": 0},
        "manager_restored": 0
    }

    # Process each row by INDEX (master is fresh from QA, same structure)
    for qa_row in range(2, qa_ws.max_row + 1):  # Skip header
        # Skip empty rows - check column 1 (first column always has data if row is valid)
        first_col_value = qa_ws.cell(row=qa_row, column=1).value
        if first_col_value is None or str(first_col_value).strip() == "":
            continue  # Skip empty rows

        result["stats"]["total"] += 1
        master_row = qa_row  # Direct index matching (always works - fresh rebuild)

        # Get QA STATUS (for stats AND to determine which rows to show/hide)
        # Compile comments for ALL statuses: ISSUE (visible), BLOCKED/KOREAN/NO ISSUE (hidden)
        should_compile_comment = False
        status_type = None
        if qa_status_col:
            qa_status = qa_ws.cell(row=qa_row, column=qa_status_col).value
            if qa_status:
                status_upper = str(qa_status).strip().upper()
                if status_upper == "ISSUE":
                    result["stats"]["issue"] += 1
                    should_compile_comment = True
                    status_type = "ISSUE"
                elif status_upper == "NO ISSUE":
                    result["stats"]["no_issue"] += 1
                    should_compile_comment = True
                    status_type = "NO ISSUE"
                elif status_upper == "BLOCKED":
                    result["stats"]["blocked"] += 1
                    should_compile_comment = True
                    status_type = "BLOCKED"
                elif status_upper == "KOREAN":
                    result["stats"]["korean"] += 1
                    should_compile_comment = True
                    status_type = "KOREAN"

        # Write TESTER STATUS to hidden column (for filtering/hiding logic)
        if status_type:
            tester_status_cell = master_ws.cell(row=master_row, column=master_tester_status_col)
            tester_status_cell.value = status_type
            tester_status_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get QA COMMENT and STRINGID, copy to master with styling
        # Compile comments for ALL statuses: ISSUE (visible), BLOCKED/KOREAN/NO ISSUE (hidden)
        comment_text_for_lookup = None  # For manager status lookup
        if qa_comment_col and should_compile_comment:
            qa_comment = qa_ws.cell(row=qa_row, column=qa_comment_col).value
            if qa_comment and str(qa_comment).strip():
                # Extract comment text for manager status lookup
                comment_text_for_lookup = extract_comment_text(qa_comment)

                # Get StringID if available
                string_id = None
                if qa_stringid_col:
                    string_id = qa_ws.cell(row=qa_row, column=qa_stringid_col).value

                # Get existing comment in master
                existing = master_ws.cell(row=master_row, column=master_comment_col).value

                # Format and update (REPLACE mode, includes stringid + file mod time)
                new_value = format_comment(qa_comment, string_id, existing, file_mod_time)

                # NO prefix needed - TESTER_STATUS column handles status info
                # Styling distinguishes status types visually

                # ONLY write if there's actually new content
                if new_value != existing:
                    cell = master_ws.cell(row=master_row, column=master_comment_col)
                    cell.value = sanitize_for_excel(new_value)  # Prevent Excel formula injection

                    # Apply styling based on status type
                    if status_type == "ISSUE":
                        # Light blue fill for ISSUE (visible)
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif status_type == "BLOCKED":
                        # Orange fill for BLOCKED (will be hidden)
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                        cell.font = Font(bold=True, color="FF8C00")
                    elif status_type == "KOREAN":
                        # Purple fill for KOREAN (will be hidden)
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        cell.font = Font(bold=True, color="800080")
                    elif status_type == "NO ISSUE":
                        # Light green fill for NO ISSUE (will be hidden)
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        cell.font = Font(bold=True, color="2E7D32")

                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin', color='87CEEB'),
                        right=Side(style='thin', color='87CEEB'),
                        top=Side(style='thin', color='87CEEB'),
                        bottom=Side(style='thin', color='87CEEB')
                    )

                    result["comments"] += 1

        # Transfer SCREENSHOT to SCREENSHOT_{username} with hyperlink transformation
        if qa_screenshot_col:
            qa_screenshot_cell = qa_ws.cell(row=qa_row, column=qa_screenshot_col)
            screenshot_value = qa_screenshot_cell.value
            screenshot_hyperlink = qa_screenshot_cell.hyperlink

            if screenshot_value and str(screenshot_value).strip():
                master_screenshot_cell = master_ws.cell(row=master_row, column=master_screenshot_col)
                existing_screenshot = master_screenshot_cell.value

                # Determine the new value and hyperlink target
                new_screenshot_value = None
                new_screenshot_target = None
                is_warning = False  # Orange for missing images

                if screenshot_hyperlink and screenshot_hyperlink.target:
                    # Extract original filename from hyperlink target
                    original_target = screenshot_hyperlink.target
                    original_name = os.path.basename(original_target)

                    # Transform to new path - ALWAYS use Images/ prefix
                    # AUTO-FIX: Try case-insensitive match if exact match fails
                    if original_name in image_mapping:
                        new_name = image_mapping[original_name]
                        new_screenshot_value = new_name
                        new_screenshot_target = f"Images/{new_name}"
                    else:
                        # Try case-insensitive match
                        matched_name = None
                        for img_name in image_mapping.keys():
                            if img_name.lower() == original_name.lower():
                                matched_name = img_name
                                break
                        if matched_name:
                            new_name = image_mapping[matched_name]
                            new_screenshot_value = new_name
                            new_screenshot_target = f"Images/{new_name}"
                        else:
                            # Image not found in mapping - still use Images/ prefix
                            new_screenshot_value = original_name
                            new_screenshot_target = f"Images/{original_name}"
                            is_warning = True
                else:
                    # No hyperlink, just copy value (might be just text)
                    # AUTO-FIX: Try case-insensitive match if exact match fails
                    original_name = str(screenshot_value).strip()
                    if original_name in image_mapping:
                        new_name = image_mapping[original_name]
                        new_screenshot_value = new_name
                        new_screenshot_target = f"Images/{new_name}"
                    else:
                        # Try case-insensitive match in image_mapping
                        matched_name = None
                        for img_name in image_mapping.keys():
                            if img_name.lower() == original_name.lower():
                                matched_name = img_name
                                break
                        if matched_name:
                            new_name = image_mapping[matched_name]
                            new_screenshot_value = new_name
                            new_screenshot_target = f"Images/{new_name}"
                        else:
                            # No mapping - still use Images/ prefix
                            new_screenshot_value = original_name
                            new_screenshot_target = f"Images/{original_name}"
                            is_warning = True

                # Check if hyperlink needs updating
                existing_hyperlink = master_screenshot_cell.hyperlink.target if master_screenshot_cell.hyperlink else None
                needs_update = (new_screenshot_value != existing_screenshot) or (new_screenshot_target != existing_hyperlink)

                if needs_update:
                    master_screenshot_cell.value = sanitize_for_excel(new_screenshot_value)
                    if new_screenshot_target:
                        master_screenshot_cell.hyperlink = new_screenshot_target

                    # Apply styling: blue fill + blue border
                    master_screenshot_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    master_screenshot_cell.alignment = Alignment(horizontal='left', vertical='center')
                    master_screenshot_cell.border = Border(
                        left=Side(style='thin', color='87CEEB'),
                        right=Side(style='thin', color='87CEEB'),
                        top=Side(style='thin', color='87CEEB'),
                        bottom=Side(style='thin', color='87CEEB')
                    )

                    # Hyperlink font: blue for valid, orange for warning
                    if new_screenshot_target:
                        if is_warning:
                            master_screenshot_cell.font = Font(color="FF6600", underline="single")
                        else:
                            master_screenshot_cell.font = Font(color="0000FF", underline="single")

                    result["screenshots"] += 1

        # Apply manager STATUS by looking up COMMENT TEXT in dict
        if comment_text_for_lookup and manager_status:
            if comment_text_for_lookup in manager_status:
                user_statuses = manager_status[comment_text_for_lookup]
                if username in user_statuses:
                    status_value = user_statuses[username]
                    status_cell = master_ws.cell(row=master_row, column=master_user_status_col)
                    status_cell.value = status_value
                    # Style based on status value
                    status_cell.alignment = Alignment(horizontal='center', vertical='center')
                    if status_value == "FIXED":
                        status_cell.font = Font(bold=True, color="228B22")  # Forest green
                    elif status_value == "REPORTED":
                        status_cell.font = Font(bold=True, color="FF8C00")  # Dark orange
                    elif status_value == "CHECKING":
                        status_cell.font = Font(bold=True, color="0000FF")  # Blue
                    elif status_value == "NON-ISSUE":
                        status_cell.font = Font(bold=True, color="808080")  # Gray
                    result["manager_restored"] += 1

    return result


def hide_empty_comment_rows(wb, context_rows=1, debug=False):
    """
    Post-process: Hide rows/sheets/columns based on tester and manager status.

    This allows focusing on ISSUE rows that need attention while preserving all data.
    Rows are hidden (not deleted) so they can be unhidden in Excel if needed.

    Features:
    - Sheet hiding: Sheets with NO comments at all are hidden entirely
    - Column hiding: COMMENT_{User} columns with no data in a sheet are hidden
      (also hides paired SCREENSHOT_{User}, STATUS_{User}, and TESTER_STATUS_{User} columns)
    - TESTER_STATUS columns: Always hidden (internal use for filtering)
    - Row hiding based on TESTER STATUS:
      - ISSUE rows: Visible (these are the active issues)
      - BLOCKED/KOREAN/NO ISSUE rows: Hidden
    - Row hiding based on MANAGER STATUS:
      - FIXED/NON-ISSUE: Hidden (issue resolved)
      - REPORTED/CHECKING: Visible (still being tracked)
      - Empty: Visible (pending manager review)

    Args:
        wb: Master workbook
        context_rows: Number of rows above/below visible rows to keep visible (default: 1)
        debug: If True, print debug info about what's being detected

    Returns: Tuple of (rows_hidden, sheets_hidden, hidden_columns_total)
    """
    hidden_rows = 0
    hidden_sheets = []
    hidden_columns_total = 0

    # === PHASE 0: RESET all sheets to visible first (fixes re-run UNHIDE bug) ===
    # Previously hidden sheets that now have content must be shown
    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue
        wb[sheet_name].sheet_state = 'visible'
        if debug:
            print(f"    [DEBUG] Reset sheet '{sheet_name}' to visible")

    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue

        ws = wb[sheet_name]

        # CRITICAL: Clear any AutoFilter from QA files before applying our logic
        # Testers may have filters/sorting applied - we want clean output
        if ws.auto_filter.ref:
            ws.auto_filter.ref = None
            if debug:
                print(f"    [DEBUG] Cleared AutoFilter from sheet: {sheet_name}")

        # Find all COMMENT_{User} columns
        comment_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).startswith("COMMENT_"):
                comment_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found comment column: {header} at col {col}")

        if not comment_cols:
            if debug:
                print(f"    [DEBUG] No COMMENT_ columns found in {sheet_name}")
            continue

        # === RESET all COMMENT_*, SCREENSHOT_*, STATUS_* columns to visible first ===
        # Fixes re-run UNHIDE bug: previously hidden columns that now have content must be shown
        # NOTE: TESTER_STATUS_* columns STAY HIDDEN (they're internal for filtering only)
        for col in comment_cols:
            header = ws.cell(row=1, column=col).value
            username = str(header).replace("COMMENT_", "") if header else ""
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].hidden = False

            # Also unhide paired SCREENSHOT_{User} and STATUS_{User} columns
            # But keep TESTER_STATUS_{User} hidden!
            for search_col in range(1, ws.max_column + 1):
                search_header = ws.cell(row=1, column=search_col).value
                if search_header:
                    if str(search_header) == f"SCREENSHOT_{username}" or str(search_header) == f"STATUS_{username}":
                        search_col_letter = get_column_letter(search_col)
                        ws.column_dimensions[search_col_letter].hidden = False
                    # TESTER_STATUS always stays hidden
                    elif str(search_header) == f"TESTER_STATUS_{username}":
                        search_col_letter = get_column_letter(search_col)
                        ws.column_dimensions[search_col_letter].hidden = True
            if debug:
                print(f"    [DEBUG] Reset column group for user '{username}' to visible (TESTER_STATUS stays hidden)")

        # First pass: Find rows that have comments AND track which columns have comments
        rows_with_comments = set()
        cols_with_comments = set()  # Track which COMMENT_ columns have data

        for row in range(2, ws.max_row + 1):
            for col in comment_cols:
                value = ws.cell(row=row, column=col).value
                # Check for any content (not just whitespace)
                if value is not None and str(value).strip():
                    rows_with_comments.add(row)
                    cols_with_comments.add(col)
                    if debug and row <= 10:  # Only debug first 10 rows
                        print(f"    [DEBUG] Row {row} has comment in col {col}: {repr(str(value)[:30])}")

        # If NO comments in entire sheet, hide the sheet tab
        if not rows_with_comments:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has {len(comment_cols)} COMMENT_ columns but ALL are empty - HIDING SHEET")
            ws.sheet_state = 'hidden'
            hidden_sheets.append(sheet_name)
            continue
        else:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has {len(rows_with_comments)} rows with comments - keeping visible")

        # Column hiding: Hide COMMENT_{User} columns that are entirely empty in this sheet
        # Also hide paired SCREENSHOT_{User} and STATUS_{User} columns
        hidden_cols_this_sheet = 0
        for col in comment_cols:
            if col not in cols_with_comments:
                # This COMMENT_ column is empty in this sheet - hide it
                header = ws.cell(row=1, column=col).value
                username = str(header).replace("COMMENT_", "") if header else ""

                # Hide the COMMENT_{User} column
                col_letter = ws.cell(row=1, column=col).column_letter
                ws.column_dimensions[col_letter].hidden = True
                hidden_cols_this_sheet += 1

                # Find and hide paired SCREENSHOT_{User}, STATUS_{User}, and TESTER_STATUS_{User} columns
                for search_col in range(1, ws.max_column + 1):
                    search_header = ws.cell(row=1, column=search_col).value
                    if search_header:
                        if str(search_header) == f"SCREENSHOT_{username}":
                            search_col_letter = ws.cell(row=1, column=search_col).column_letter
                            ws.column_dimensions[search_col_letter].hidden = True
                            hidden_cols_this_sheet += 1
                        elif str(search_header) == f"STATUS_{username}":
                            search_col_letter = ws.cell(row=1, column=search_col).column_letter
                            ws.column_dimensions[search_col_letter].hidden = True
                            hidden_cols_this_sheet += 1
                        elif str(search_header) == f"TESTER_STATUS_{username}":
                            search_col_letter = ws.cell(row=1, column=search_col).column_letter
                            ws.column_dimensions[search_col_letter].hidden = True
                            hidden_cols_this_sheet += 1

                if debug:
                    print(f"    [DEBUG] Hidden empty column group for user: {username}")

        hidden_columns_total += hidden_cols_this_sheet

        # === FIND TESTER_STATUS_{User} columns for tester status hiding ===
        # Hide rows where tester marked BLOCKED, KOREAN, or NO ISSUE
        # Only show: ISSUE rows
        tester_status_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).startswith("TESTER_STATUS_"):
                tester_status_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found tester status column: {header} at col {col}")

        # Find rows to hide due to tester status (BLOCKED, KOREAN, NO ISSUE)
        # A row is hidden only if ALL tester statuses are non-ISSUE (or empty)
        # If ANY user marked it as ISSUE, the row should be visible
        rows_non_issue_by_tester = set()
        rows_with_issue_status = set()
        TESTER_HIDE_STATUSES = {"BLOCKED", "KOREAN", "NO ISSUE"}
        for row in range(2, ws.max_row + 1):
            has_issue = False
            has_any_status = False
            for col in tester_status_cols:
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    has_any_status = True
                    status_upper = str(value).strip().upper()
                    if status_upper == "ISSUE":
                        has_issue = True
                        rows_with_issue_status.add(row)
                        break  # Found ISSUE, no need to check more columns
            # Only hide if has status but NO ISSUE status found
            if has_any_status and not has_issue:
                rows_non_issue_by_tester.add(row)
                if debug and row <= 20:
                    print(f"    [DEBUG] Row {row} has tester status but no ISSUE - will hide")

        if debug:
            if rows_with_issue_status:
                print(f"    [DEBUG] {len(rows_with_issue_status)} rows with ISSUE tester status (will show)")
            if rows_non_issue_by_tester:
                print(f"    [DEBUG] {len(rows_non_issue_by_tester)} rows with only non-ISSUE tester status (will hide)")

        # === FIND STATUS_{User} columns for manager status hiding ===
        # Hide rows where manager marked FIXED or NON-ISSUE
        # Keep visible: REPORTED, CHECKING, or empty (pending)
        manager_status_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            # STATUS_ but NOT TESTER_STATUS_
            if header and str(header).startswith("STATUS_") and not str(header).startswith("TESTER_STATUS_"):
                manager_status_cols.append(col)
                if debug:
                    print(f"    [DEBUG] Found manager status column: {header} at col {col}")

        # Find rows to hide due to manager status (FIXED, NON-ISSUE only)
        # REPORTED and CHECKING stay visible!
        rows_resolved_by_manager = set()
        MANAGER_HIDE_STATUSES = {"FIXED", "NON-ISSUE"}  # NOT REPORTED - that stays visible!
        for row in range(2, ws.max_row + 1):
            for col in manager_status_cols:
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip().upper() in MANAGER_HIDE_STATUSES:
                    rows_resolved_by_manager.add(row)
                    if debug and row <= 20:
                        print(f"    [DEBUG] Row {row} has manager status '{value}' - will hide")
                    break  # One resolved status is enough to hide

        if debug and rows_resolved_by_manager:
            print(f"    [DEBUG] {len(rows_resolved_by_manager)} rows resolved by manager (FIXED/NON-ISSUE)")

        # === COMBINE HIDING RULES ===
        # Hide if:
        #   1. Row has non-ISSUE tester status (BLOCKED/KOREAN/NO ISSUE)
        #   2. OR row has manager status FIXED or NON-ISSUE
        # Show if:
        #   - Row has ISSUE tester status AND manager status is empty/REPORTED/CHECKING
        rows_to_hide = rows_non_issue_by_tester | rows_resolved_by_manager

        # Build set of rows to keep visible
        # Start with rows that have comments and are ISSUE status, minus resolved rows
        rows_to_show = (rows_with_comments - rows_non_issue_by_tester) - rows_resolved_by_manager
        for row in list(rows_to_show):  # Use list() to avoid modifying set during iteration
            # Add context rows above (only if they're not in hide set)
            for offset in range(1, context_rows + 1):
                if row - offset >= 2 and row - offset not in rows_to_hide:
                    rows_to_show.add(row - offset)
            # Add context rows below (only if they're not in hide set)
            for offset in range(1, context_rows + 1):
                if row + offset <= ws.max_row and row + offset not in rows_to_hide:
                    rows_to_show.add(row + offset)

        # === SHEET HIDING: If NO rows to show after filtering, hide entire sheet ===
        # This catches sheets where ALL comments are non-ISSUE or all resolved by manager
        if not rows_to_show:
            if debug:
                print(f"    [DEBUG] Sheet '{sheet_name}' has comments but NONE are visible ISSUE rows - HIDING SHEET")
            ws.sheet_state = 'hidden'
            hidden_sheets.append(sheet_name)
            continue  # Skip row hiding for this sheet

        # Pass 1: UNHIDE all rows first (clear any previous hiding)
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].hidden = False

        # Pass 2: Hide rows not in the show set
        for row in range(2, ws.max_row + 1):
            if row not in rows_to_show:
                ws.row_dimensions[row].hidden = True
                hidden_rows += 1

    return hidden_rows, hidden_sheets, hidden_columns_total


def autofit_rows_with_wordwrap(wb, default_row_height=15, chars_per_line=50):
    """
    Apply word wrap to all cells and autofit row heights based on content.

    Args:
        wb: Workbook
        default_row_height: Default height for single-line rows
        chars_per_line: Estimated characters per line (for height calculation)
    """
    for sheet_name in wb.sheetnames:
        if sheet_name == "STATUS":
            continue

        ws = wb[sheet_name]

        for row in range(1, ws.max_row + 1):
            max_lines = 1  # Track max lines needed for this row

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                # Apply word wrap to all cells
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Calculate lines needed based on content
                if cell.value:
                    content = str(cell.value)
                    # Count explicit line breaks
                    explicit_lines = content.count('\n') + 1
                    # Estimate wrapped lines based on length
                    longest_line = max(len(line) for line in content.split('\n')) if content else 0
                    wrapped_lines = max(1, (longest_line // chars_per_line) + 1)
                    # Total lines needed
                    total_lines = explicit_lines + wrapped_lines - 1
                    max_lines = max(max_lines, total_lines)

            # Set row height based on content (15 points per line is standard)
            calculated_height = max_lines * default_row_height
            # Cap at reasonable max (300 points)
            ws.row_dimensions[row].height = min(calculated_height, 300)


def update_status_sheet(wb, users, user_stats):
    """
    Create/update STATUS sheet with completion tracking and detailed stats.

    Args:
        wb: Master workbook
        users: Set of usernames
        user_stats: {username: {total: n, issue: n, no_issue: n, blocked: n}}
    """
    # Create or clear STATUS sheet
    if "STATUS" in wb.sheetnames:
        del wb["STATUS"]

    ws = wb.create_sheet("STATUS", 0)  # Insert at position 0 (first tab)

    # Styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold/Yellow
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center')

    # Headers
    headers = ["User", "Completion %", "Total Rows", "ISSUE #", "NO ISSUE %", "BLOCKED %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    # Data rows
    for row_idx, user in enumerate(sorted(users), 2):
        stats = user_stats.get(user, {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0})

        total = stats["total"]
        issue = stats["issue"]
        no_issue = stats["no_issue"]
        blocked = stats["blocked"]
        completed = issue + no_issue + blocked

        # Calculate percentages
        completion_pct = round(completed / total * 100, 1) if total > 0 else 0
        no_issue_pct = round(no_issue / total * 100, 1) if total > 0 else 0
        blocked_pct = round(blocked / total * 100, 1) if total > 0 else 0

        # Write data
        row_data = [
            user,
            f"{completion_pct}%",
            total,
            issue,  # Raw number for ISSUE
            f"{no_issue_pct}%",
            f"{blocked_pct}%"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.alignment = center

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    print(f"  Updated STATUS sheet with {len(users)} users (first tab, yellow header)")


# =============================================================================
# MANAGER STATUS PREPROCESS FUNCTIONS
# =============================================================================

def extract_comment_text(full_comment):
    """
    Extract the actual comment text from formatted comment.

    Comment format: "comment text\n---\nstringid:\n10001\n(updated: ...)"
    Returns: Just the comment text before "---"
    """
    if not full_comment:
        return ""
    comment_str = str(full_comment).strip()
    if "---" in comment_str:
        return comment_str.split("---")[0].strip()
    return comment_str


def collect_manager_status(master_folder):
    """
    Read existing Master files and collect all STATUS_{User} values.

    KEYED BY COMMENT TEXT - so we can match after rebuild.

    This is the PREPROCESS step - runs before compilation to preserve
    manager-entered status values (FIXED/REPORTED/CHECKING).

    Args:
        master_folder: Which Master folder to scan (EN or CN)

    Returns: Dict structure
    {
        "Quest": {
            "Sheet1": {
                "comment text here": {
                    "John": "FIXED",
                    "Alice": "REPORTED"
                }
            }
        }
    }
    """
    manager_status = {}

    for category in CATEGORIES:
        master_path = master_folder / f"Master_{category}.xlsx"
        if not master_path.exists():
            continue

        try:
            wb = safe_load_workbook(master_path)
            manager_status[category] = {}

            for sheet_name in wb.sheetnames:
                if sheet_name == "STATUS":
                    continue

                ws = wb[sheet_name]
                manager_status[category][sheet_name] = {}

                # Find all COMMENT_{User} and STATUS_{User} columns
                comment_cols = {}  # username -> col
                status_cols = {}   # username -> col
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header:
                        header_str = str(header)
                        if header_str.startswith("COMMENT_"):
                            username = header_str.replace("COMMENT_", "")
                            comment_cols[username] = col
                        elif header_str.startswith("STATUS_"):
                            username = header_str.replace("STATUS_", "")
                            status_cols[username] = col

                if not status_cols:
                    continue

                # Collect values per row, keyed by comment text
                for row in range(2, ws.max_row + 1):
                    for username, status_col in status_cols.items():
                        status_value = ws.cell(row=row, column=status_col).value
                        if status_value and str(status_value).strip().upper() in VALID_MANAGER_STATUS:
                            # Get the paired comment
                            comment_col = comment_cols.get(username)
                            if comment_col:
                                full_comment = ws.cell(row=row, column=comment_col).value
                                comment_text = extract_comment_text(full_comment)
                                if comment_text:
                                    # Store: dict[comment_text][user] = status
                                    if comment_text not in manager_status[category][sheet_name]:
                                        manager_status[category][sheet_name][comment_text] = {}
                                    manager_status[category][sheet_name][comment_text][username] = str(status_value).strip().upper()

            wb.close()

        except Exception as e:
            print(f"  WARN: Error reading {master_path.name} for preprocess: {e}")

    return manager_status


def collect_manager_stats_for_tracker():
    """
    Read all Master files (EN + CN) and count FIXED/REPORTED/CHECKING/NON-ISSUE per user per category.

    Returns: Dict structure
    {
        "Quest": {
            "John": {"fixed": 5, "reported": 2, "checking": 1, "nonissue": 0, "lang": "EN"},
            "Alice": {"fixed": 3, "reported": 1, "checking": 0, "nonissue": 2, "lang": "CN"}
        }
    }
    """
    manager_stats = defaultdict(lambda: defaultdict(lambda: {"fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "lang": "EN"}))

    # Load tester mapping to get language
    tester_mapping = load_tester_mapping()

    # Scan both EN and CN folders
    for master_folder in [MASTER_FOLDER_EN, MASTER_FOLDER_CN]:
        for category in CATEGORIES:
            master_path = master_folder / f"Master_{category}.xlsx"
            if not master_path.exists():
                continue

            try:
                wb = safe_load_workbook(master_path)

                for sheet_name in wb.sheetnames:
                    if sheet_name == "STATUS":
                        continue

                    ws = wb[sheet_name]

                    # Find all STATUS_{User} columns
                    status_cols = {}
                    for col in range(1, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if header and str(header).startswith("STATUS_"):
                            username = str(header).replace("STATUS_", "")
                            status_cols[username] = col

                    if not status_cols:
                        continue

                    # Count status values per user
                    for row in range(2, ws.max_row + 1):
                        for username, col in status_cols.items():
                            value = ws.cell(row=row, column=col).value
                            if value:
                                status_upper = str(value).strip().upper()
                                if status_upper == "FIXED":
                                    manager_stats[category][username]["fixed"] += 1
                                elif status_upper == "REPORTED":
                                    manager_stats[category][username]["reported"] += 1
                                elif status_upper == "CHECKING":
                                    manager_stats[category][username]["checking"] += 1
                                elif status_upper == "NON-ISSUE":
                                    manager_stats[category][username]["nonissue"] += 1
                            # Set language from mapping
                            manager_stats[category][username]["lang"] = tester_mapping.get(username, "EN")

                wb.close()

            except Exception as e:
                print(f"  WARN: Error reading {master_path.name} for manager stats: {e}")

    return manager_stats


# =============================================================================
# TRACKER FUNCTIONS - LQA User Progress Tracker
# =============================================================================

def get_or_create_tracker():
    """
    Load existing tracker or create new one with sheets.

    Returns: (workbook, path)
    """
    tracker_path = TRACKER_PATH  # Root level

    if tracker_path.exists():
        wb = safe_load_workbook(tracker_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        # Create sheets in order (no GRAPHS - charts embedded in DAILY/TOTAL)
        wb.create_sheet("DAILY", 0)
        wb.create_sheet("TOTAL", 1)
        wb.create_sheet("_DAILY_DATA", 2)
        # Remove default sheet
        wb.remove(default_sheet)
        # Hide data sheet
        wb["_DAILY_DATA"].sheet_state = 'hidden'

    return wb, tracker_path


def update_daily_data_sheet(wb, daily_entries, manager_stats=None):
    """
    Update hidden _DAILY_DATA sheet with new entries including manager stats.

    Args:
        wb: Tracker workbook
        daily_entries: List of dicts with {date, user, category, total_rows, done, issues, no_issue, blocked}
        manager_stats: Dict of {category: {user: {fixed, reported, checking}}} from manager status

    Mode: REPLACE - same (date, user, category) overwrites existing row
    """
    if manager_stats is None:
        manager_stats = {}

    ws = wb["_DAILY_DATA"]

    # Ensure headers exist (now includes TotalRows, Fixed, Reported, Checking, NonIssue, WordCount, Korean)
    # Schema: Date, User, Category, TotalRows, Done, Issues, NoIssue, Blocked, Fixed, Reported, Checking, NonIssue, WordCount, Korean
    if ws.cell(1, 1).value != "Date" or ws.max_column < 14:
        headers = ["Date", "User", "Category", "TotalRows", "Done", "Issues", "NoIssue", "Blocked", "Fixed", "Reported", "Checking", "NonIssue", "WordCount", "Korean"]
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)

    # Build index of existing rows: (date, user, category) -> row_number
    existing = {}
    for row in range(2, ws.max_row + 1):
        key = (
            ws.cell(row, 1).value,  # Date
            ws.cell(row, 2).value,  # User
            ws.cell(row, 3).value   # Category
        )
        if key[0] is not None:  # Only index non-empty rows
            existing[key] = row

    # Update or insert entries
    for entry in daily_entries:
        key = (entry["date"], entry["user"], entry["category"])

        if key in existing:
            row = existing[key]
        else:
            row = ws.max_row + 1
            existing[key] = row

        # Get manager stats for this user/category
        category = entry["category"]
        user = entry["user"]
        user_manager_stats = manager_stats.get(category, {}).get(user, {"fixed": 0, "reported": 0, "checking": 0, "nonissue": 0})

        # Schema: Date, User, Category, TotalRows, Done, Issues, NoIssue, Blocked, Fixed, Reported, Checking, NonIssue, WordCount, Korean
        ws.cell(row, 1, entry["date"])
        ws.cell(row, 2, entry["user"])
        ws.cell(row, 3, entry["category"])
        ws.cell(row, 4, entry.get("total_rows", 0))  # TotalRows (universe)
        ws.cell(row, 5, entry["done"])               # Done (completed)
        ws.cell(row, 6, entry["issues"])
        ws.cell(row, 7, entry["no_issue"])
        ws.cell(row, 8, entry["blocked"])
        ws.cell(row, 9, user_manager_stats["fixed"])
        ws.cell(row, 10, user_manager_stats["reported"])
        ws.cell(row, 11, user_manager_stats["checking"])
        ws.cell(row, 12, user_manager_stats["nonissue"])
        ws.cell(row, 13, entry.get("word_count", 0))  # WordCount (words for EN, chars for CN)
        ws.cell(row, 14, entry.get("korean", 0))      # Korean status count


def build_daily_sheet(wb):
    """
    Build DAILY sheet from _DAILY_DATA.

    Separate EN and CN sections with full status breakdown:
    - Done, Issues, No Issue, Blocked, Korean, Words/Chars
    Plus Manager Stats (Fixed, Reported, Checking, Pending).
    """
    # Delete and recreate sheet to handle merged cells properly
    if "DAILY" in wb.sheetnames:
        del wb["DAILY"]
    ws = wb.create_sheet("DAILY", 0)

    data_ws = wb["_DAILY_DATA"]

    # Load tester mapping for EN/CN separation
    tester_mapping = load_tester_mapping()

    # Read raw data and aggregate by (date, user)
    # Full schema: Date(1), User(2), Category(3), TotalRows(4), Done(5), Issues(6), NoIssue(7), Blocked(8),
    #              Fixed(9), Reported(10), Checking(11), NonIssue(12), WordCount(13), Korean(14)
    daily_data = defaultdict(lambda: defaultdict(lambda: {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    }))
    users = set()

    for row in range(2, data_ws.max_row + 1):
        date = data_ws.cell(row, 1).value
        user = data_ws.cell(row, 2).value
        total_rows = data_ws.cell(row, 4).value or 0
        done = data_ws.cell(row, 5).value or 0
        issues = data_ws.cell(row, 6).value or 0
        no_issue = data_ws.cell(row, 7).value or 0
        blocked = data_ws.cell(row, 8).value or 0
        fixed = data_ws.cell(row, 9).value or 0
        reported = data_ws.cell(row, 10).value or 0
        checking = data_ws.cell(row, 11).value or 0
        nonissue = data_ws.cell(row, 12).value or 0
        word_count = data_ws.cell(row, 13).value or 0
        korean = data_ws.cell(row, 14).value or 0

        if date and user:
            daily_data[date][user]["total_rows"] += total_rows
            daily_data[date][user]["done"] += done
            daily_data[date][user]["issues"] += issues
            daily_data[date][user]["no_issue"] += no_issue
            daily_data[date][user]["blocked"] += blocked
            daily_data[date][user]["korean"] += korean
            daily_data[date][user]["fixed"] += fixed
            daily_data[date][user]["reported"] += reported
            daily_data[date][user]["checking"] += checking
            daily_data[date][user]["nonissue"] += nonissue
            daily_data[date][user]["word_count"] += word_count
            users.add(user)

    if not users:
        ws.cell(1, 1, "No data yet")
        return

    # Separate EN and CN users
    en_users = sorted([u for u in users if tester_mapping.get(u, "EN") == "EN"])
    cn_users = sorted([u for u in users if tester_mapping.get(u) == "CN"])
    dates = sorted(daily_data.keys())

    # === Calculate DAILY DELTAS from cumulative values ===
    # Each date's data is cumulative - to get daily work, subtract previous day's cumulative
    default_data = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "word_count": 0
    }
    daily_delta = defaultdict(lambda: defaultdict(lambda: default_data.copy()))

    for i, date in enumerate(dates):
        for user in users:
            current = daily_data[date].get(user, default_data.copy())

            if i == 0:
                prev = default_data.copy()
            else:
                prev_date = dates[i - 1]
                prev = daily_data[prev_date].get(user, default_data.copy())

            # Calculate delta (ensure non-negative)
            daily_delta[date][user]["total_rows"] = current["total_rows"]
            daily_delta[date][user]["done"] = max(0, current["done"] - prev["done"])
            daily_delta[date][user]["issues"] = max(0, current["issues"] - prev["issues"])
            daily_delta[date][user]["no_issue"] = max(0, current["no_issue"] - prev["no_issue"])
            daily_delta[date][user]["blocked"] = max(0, current["blocked"] - prev["blocked"])
            daily_delta[date][user]["korean"] = max(0, current["korean"] - prev["korean"])
            daily_delta[date][user]["fixed"] = max(0, current["fixed"] - prev["fixed"])
            daily_delta[date][user]["reported"] = max(0, current["reported"] - prev["reported"])
            daily_delta[date][user]["checking"] = max(0, current["checking"] - prev["checking"])
            daily_delta[date][user]["nonissue"] = max(0, current["nonissue"] - prev["nonissue"])
            daily_delta[date][user]["word_count"] = max(0, current["word_count"] - prev["word_count"])

    # Styles
    en_title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for EN
    cn_title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red for CN
    manager_title_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    header_fill = PatternFill(start_color=TRACKER_STYLES["header_color"], end_color=TRACKER_STYLES["header_color"], fill_type="solid")
    subheader_fill = PatternFill(start_color=TRACKER_STYLES["subheader_color"], end_color=TRACKER_STYLES["subheader_color"], fill_type="solid")
    alt_fill = PatternFill(start_color=TRACKER_STYLES["alt_row_color"], end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid")
    border = Border(
        left=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        right=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        top=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        bottom=Side(style='thin', color=TRACKER_STYLES["border_color"])
    )
    thick_side = Side(style='thick', color='000000')
    thin_side = Side(style='thin', color=TRACKER_STYLES["border_color"])
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Columns per user: Done, Issues, No Issue, Blocked, Korean, Words/Chars
    tester_headers = ["Done", "Issues", "NoIssue", "Blocked", "Korean"]
    cols_per_user = len(tester_headers) + 1  # +1 for Words/Chars

    def build_daily_section(ws, start_row, section_title, title_fill, users_list, is_english):
        """Build a daily stats section for EN or CN users."""
        if not users_list:
            return start_row

        current_row = start_row
        word_label = "Words" if is_english else "Chars"

        # Calculate total columns: Date(1) + users * cols_per_user
        total_cols = 1 + len(users_list) * cols_per_user

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, section_title)
        title_cell.fill = title_fill
        title_cell.font = white_bold
        title_cell.alignment = center
        current_row += 1

        # User names row (merged across their columns)
        ws.cell(current_row, 1, "")  # Empty cell above Date
        col = 2
        for user in users_list:
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + cols_per_user - 1)
            cell = ws.cell(current_row, col, user)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
            # Style merged cells
            for c in range(col + 1, col + cols_per_user):
                ws.cell(current_row, c).fill = header_fill
                ws.cell(current_row, c).border = border
            col += cols_per_user
        current_row += 1

        # Sub-headers row: Date, then per user columns
        date_cell = ws.cell(current_row, 1, "Date")
        date_cell.fill = subheader_fill
        date_cell.font = bold
        date_cell.alignment = center
        date_cell.border = border

        col = 2
        for user in users_list:
            user_headers = tester_headers + [word_label]
            for i, header in enumerate(user_headers):
                cell = ws.cell(current_row, col + i, header)
                cell.fill = subheader_fill
                cell.font = bold
                cell.alignment = center
                if i == 0:
                    cell.border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
                else:
                    cell.border = border
            col += cols_per_user
        current_row += 1

        # Data rows
        for idx, date in enumerate(dates):
            # Date column - format as MM/DD
            if isinstance(date, str) and len(date) >= 10:
                display_date = date[5:7] + "/" + date[8:10]
            else:
                display_date = str(date)

            date_cell = ws.cell(current_row, 1, display_date)
            date_cell.alignment = center
            date_cell.border = border
            if idx % 2 == 1:
                date_cell.fill = alt_fill

            col = 2
            for user in users_list:
                user_data = daily_delta[date].get(user, default_data.copy())
                done_val = user_data["done"]
                issues_val = user_data["issues"]
                no_issue_val = user_data["no_issue"]
                blocked_val = user_data["blocked"]
                korean_val = user_data["korean"]
                word_val = user_data["word_count"]

                values = [done_val, issues_val, no_issue_val, blocked_val, korean_val, word_val]
                for i, val in enumerate(values):
                    display_val = val if val > 0 else "--"
                    cell = ws.cell(current_row, col + i, display_val)
                    cell.alignment = center
                    if i == 0:
                        cell.border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
                    else:
                        cell.border = border
                    if idx % 2 == 1:
                        cell.fill = alt_fill
                col += cols_per_user

            current_row += 1

        return current_row + 1  # Extra row for spacing

    def build_manager_section(ws, start_row):
        """Build manager stats section (aggregated across all users)."""
        current_row = start_row
        manager_headers = ["Date", "Fixed", "Reported", "NonIssue", "Checking", "Pending"]
        total_cols = len(manager_headers)

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, "MANAGER STATS")
        title_cell.fill = manager_title_fill
        title_cell.font = bold
        title_cell.alignment = center
        current_row += 1

        # Headers row
        for i, header in enumerate(manager_headers, 1):
            cell = ws.cell(current_row, i, header)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        current_row += 1

        # Data rows
        for idx, date in enumerate(dates):
            if isinstance(date, str) and len(date) >= 10:
                display_date = date[5:7] + "/" + date[8:10]
            else:
                display_date = str(date)

            # Aggregate manager stats across ALL users for this date
            day_fixed = sum(daily_delta[date].get(u, default_data.copy())["fixed"] for u in users)
            day_reported = sum(daily_delta[date].get(u, default_data.copy())["reported"] for u in users)
            day_checking = sum(daily_delta[date].get(u, default_data.copy())["checking"] for u in users)
            day_nonissue = sum(daily_delta[date].get(u, default_data.copy())["nonissue"] for u in users)
            day_issues = sum(daily_delta[date].get(u, default_data.copy())["issues"] for u in users)
            day_pending = max(0, day_issues - day_fixed - day_reported - day_checking - day_nonissue)

            row_values = [display_date, day_fixed, day_reported, day_nonissue, day_checking, day_pending]
            for i, val in enumerate(row_values, 1):
                display_val = val if (i == 1 or val > 0) else "--"
                cell = ws.cell(current_row, i, display_val)
                cell.alignment = center
                cell.border = border
                if idx % 2 == 1:
                    cell.fill = alt_fill

            current_row += 1

        return current_row + 1

    # Build sections
    current_row = 1

    # EN DAILY STATS
    if en_users:
        current_row = build_daily_section(ws, current_row, "EN DAILY STATS", en_title_fill, en_users, is_english=True)

    # CN DAILY STATS
    if cn_users:
        current_row = build_daily_section(ws, current_row, "CN DAILY STATS", cn_title_fill, cn_users, is_english=False)

    # MANAGER STATS
    current_row = build_manager_section(ws, current_row)

    # Set column widths
    PADDING = 2
    ws.column_dimensions['A'].width = 8  # Date column

    # Auto-width for other columns based on max users
    max_users = max(len(en_users), len(cn_users)) if en_users or cn_users else 1
    for col_idx in range(2, 2 + max_users * cols_per_user):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 8  # Compact width for numbers


def build_ranking_table(ws, start_row, user_data, tester_mapping):
    """
    Build Ranking Tables with weighted scoring - separate EN and CN rankings.

    Score = 70% Completion + 30% Actual Issues
    - Completion = Done / TotalRows * 100
    - Actual Issues = (Issues - NonIssue) / Issues * 100 (clamped 0-100)

    Args:
        ws: Worksheet to write to
        start_row: Row to start building
        user_data: Dict of user -> {total_rows, done, issues, nonissue, ...}
        tester_mapping: Dict of username -> language

    Returns:
        next_row: Row after this section
    """
    if not user_data:
        return start_row

    # Styles
    en_title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for EN
    cn_title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red for CN
    header_fill = PatternFill(start_color=TRACKER_STYLES["header_color"],
                              end_color=TRACKER_STYLES["header_color"], fill_type="solid")
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for #1
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Silver for #2
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")  # Bronze for #3
    alt_fill = PatternFill(start_color=TRACKER_STYLES["alt_row_color"],
                           end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid")
    border = Border(
        left=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        right=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        top=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        bottom=Side(style='thin', color=TRACKER_STYLES["border_color"])
    )
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Calculate scores for each user (SCALAR based, not percentage)
    en_scores = []
    cn_scores = []
    for user, data in user_data.items():
        done = data.get("done", 0)
        issues = data.get("issues", 0)
        nonissue = data.get("nonissue", 0)

        # Actual Issues (scalar) = Issues - NonIssue
        actual_issues = max(0, issues - nonissue)

        # Weighted score: 70% Done + 30% Actual Issues (SCALAR values)
        score = round(0.7 * done + 0.3 * actual_issues, 1)

        lang = tester_mapping.get(user, "EN")
        user_score = {
            "user": user,
            "lang": lang,
            "done": done,
            "actual_issues": actual_issues,
            "score": score
        }

        # Split by language
        if lang == "CN":
            cn_scores.append(user_score)
        else:
            en_scores.append(user_score)

    # Sort each list by score descending
    en_scores.sort(key=lambda x: (-x["score"], -x["done"], x["user"]))
    cn_scores.sort(key=lambda x: (-x["score"], -x["done"], x["user"]))

    current_row = start_row
    total_cols = 5  # Rank, User, Done, Actual Issues, Score
    headers = ["Rank", "User", "Done", "Actual Issues", "Score"]

    def build_ranking_section(scores_list, title, title_fill):
        """Build a single ranking section (EN or CN)."""
        nonlocal current_row

        if not scores_list:
            return

        # Title row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, title)
        title_cell.fill = title_fill
        title_cell.font = white_bold
        title_cell.alignment = center
        current_row += 1

        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        current_row += 1

        # Data rows
        for idx, user_score in enumerate(scores_list):
            rank = idx + 1

            # Rank cell with medal colors
            rank_cell = ws.cell(current_row, 1, rank)
            rank_cell.alignment = center
            rank_cell.border = border
            if rank == 1:
                rank_cell.fill = gold_fill
                rank_cell.font = bold
            elif rank == 2:
                rank_cell.fill = silver_fill
                rank_cell.font = bold
            elif rank == 3:
                rank_cell.fill = bronze_fill
                rank_cell.font = bold
            elif idx % 2 == 1:
                rank_cell.fill = alt_fill

            # User
            user_cell = ws.cell(current_row, 2, user_score["user"])
            user_cell.alignment = center
            user_cell.border = border
            if rank <= 3:
                user_cell.font = bold
            if idx % 2 == 1 and rank > 3:
                user_cell.fill = alt_fill

            # Done (scalar)
            done_cell = ws.cell(current_row, 3, user_score["done"])
            done_cell.alignment = center
            done_cell.border = border
            if idx % 2 == 1 and rank > 3:
                done_cell.fill = alt_fill

            # Actual Issues (scalar)
            ai_cell = ws.cell(current_row, 4, user_score["actual_issues"])
            ai_cell.alignment = center
            ai_cell.border = border
            if idx % 2 == 1 and rank > 3:
                ai_cell.fill = alt_fill

            # Score
            score_cell = ws.cell(current_row, 5, user_score["score"])
            score_cell.number_format = '0.0'
            score_cell.alignment = center
            score_cell.border = border
            score_cell.font = bold
            if idx % 2 == 1 and rank > 3:
                score_cell.fill = alt_fill

            current_row += 1

    # Build EN Ranking
    build_ranking_section(en_scores, "EN RANKING (70% Done + 30% Actual Issues)", en_title_fill)

    if en_scores and cn_scores:
        current_row += 1  # Gap between tables

    # Build CN Ranking
    build_ranking_section(cn_scores, "CN RANKING (70% Done + 30% Actual Issues)", cn_title_fill)

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 6   # Rank
    ws.column_dimensions[get_column_letter(2)].width = 12  # User
    ws.column_dimensions[get_column_letter(3)].width = 12  # Completion
    ws.column_dimensions[get_column_letter(4)].width = 14  # Actual Issues
    ws.column_dimensions[get_column_letter(5)].width = 8   # Score

    return current_row


def build_category_breakdown_section(ws, start_row, latest_data, users_list, is_english, tester_mapping):
    """
    Build Category Breakdown pivot table for EN or CN testers.

    Shows per-category completion % and word/character count for each user.

    Args:
        ws: Worksheet to write to
        start_row: Row to start building
        latest_data: Dict of (user, category) -> {done, total_rows, word_count, ...}
        users_list: List of usernames for this section (EN or CN)
        is_english: True for EN (shows "Words"), False for CN (shows "Chars")
        tester_mapping: Dict of username -> language

    Returns:
        next_row: Row after this section
    """
    if not users_list:
        return start_row

    # Styles
    title_fill = PatternFill(start_color="4472C4" if is_english else "C00000",
                             end_color="4472C4" if is_english else "C00000", fill_type="solid")
    header_fill = PatternFill(start_color=TRACKER_STYLES["header_color"],
                              end_color=TRACKER_STYLES["header_color"], fill_type="solid")
    alt_fill = PatternFill(start_color=TRACKER_STYLES["alt_row_color"],
                           end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid")
    total_fill = PatternFill(start_color=TRACKER_STYLES["total_row_color"],
                             end_color=TRACKER_STYLES["total_row_color"], fill_type="solid")
    border = Border(
        left=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        right=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        top=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        bottom=Side(style='thin', color=TRACKER_STYLES["border_color"])
    )
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Build pivot: user -> {category -> {done%, word_count}}
    pivot = {}
    for (user, category), data in latest_data.items():
        if user not in users_list:
            continue
        if user not in pivot:
            pivot[user] = {}
        total_rows = data["total_rows"]
        done = data["done"]
        pct = round(done / total_rows * 100, 1) if total_rows > 0 else 0
        pivot[user][category] = {
            "pct": pct,
            "count": data.get("word_count", 0)
        }

    # Categories to show (use global CATEGORIES)
    categories = CATEGORIES

    # Calculate column count: User + (Done%/Count) * categories + Total
    # Each category has 2 sub-columns: Done% and Words/Chars
    count_label = "Words" if is_english else "Chars"
    total_cols = 1 + len(categories) * 2 + 1  # User + categories*2 + Total

    current_row = start_row

    # Title row
    title_text = f"EN CATEGORY BREAKDOWN ({count_label})" if is_english else f"CN CATEGORY BREAKDOWN ({count_label})"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
    title_cell = ws.cell(current_row, 1, title_text)
    title_cell.fill = title_fill
    title_cell.font = white_bold
    title_cell.alignment = center
    current_row += 1

    # Header row 1: Category names (merged across 2 columns each)
    ws.cell(current_row, 1, "User").fill = header_fill
    ws.cell(current_row, 1).font = bold
    ws.cell(current_row, 1).alignment = center
    ws.cell(current_row, 1).border = border

    col = 2
    for cat in categories:
        # Merge 2 cells for category name
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
        cell = ws.cell(current_row, col, cat)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border
        ws.cell(current_row, col + 1).border = border
        col += 2

    # Total column header
    ws.cell(current_row, col, f"Total {count_label}").fill = header_fill
    ws.cell(current_row, col).font = bold
    ws.cell(current_row, col).alignment = center
    ws.cell(current_row, col).border = border
    current_row += 1

    # Header row 2: Done% / Words|Chars sub-headers
    ws.cell(current_row, 1, "").border = border  # Empty under "User"

    col = 2
    for cat in categories:
        cell1 = ws.cell(current_row, col, "Done%")
        cell1.fill = header_fill
        cell1.alignment = center
        cell1.border = border
        cell2 = ws.cell(current_row, col + 1, count_label)
        cell2.fill = header_fill
        cell2.alignment = center
        cell2.border = border
        col += 2

    ws.cell(current_row, col, "").border = border  # Empty under Total
    current_row += 1

    # Data rows
    category_totals = {cat: {"done": 0, "total_rows": 0, "count": 0} for cat in categories}
    grand_total_count = 0

    for idx, user in enumerate(sorted(users_list)):
        user_total_count = 0

        # User name
        cell = ws.cell(current_row, 1, user)
        cell.alignment = center
        cell.border = border
        if idx % 2 == 1:
            cell.fill = alt_fill

        col = 2
        for cat in categories:
            user_cat_data = pivot.get(user, {}).get(cat, None)

            if user_cat_data:
                pct = user_cat_data["pct"]
                count = user_cat_data["count"]
                user_total_count += count

                # Aggregate for category totals
                # Need original done/total_rows for accurate category total %
                for (u, c), data in latest_data.items():
                    if u == user and c == cat:
                        category_totals[cat]["done"] += data["done"]
                        category_totals[cat]["total_rows"] += data["total_rows"]
                        category_totals[cat]["count"] += data.get("word_count", 0)
                        break

                # Done%
                cell1 = ws.cell(current_row, col, pct)
                cell1.number_format = '0.0"%"'
                cell1.alignment = center
                cell1.border = border
                if idx % 2 == 1:
                    cell1.fill = alt_fill

                # Word/Char count
                cell2 = ws.cell(current_row, col + 1, count)
                cell2.number_format = '#,##0'
                cell2.alignment = center
                cell2.border = border
                if idx % 2 == 1:
                    cell2.fill = alt_fill
            else:
                # No data for this category - show "-"
                cell1 = ws.cell(current_row, col, "-")
                cell1.alignment = center
                cell1.border = border
                if idx % 2 == 1:
                    cell1.fill = alt_fill

                cell2 = ws.cell(current_row, col + 1, "-")
                cell2.alignment = center
                cell2.border = border
                if idx % 2 == 1:
                    cell2.fill = alt_fill

            col += 2

        # Total count for this user
        grand_total_count += user_total_count
        cell = ws.cell(current_row, col, user_total_count)
        cell.number_format = '#,##0'
        cell.alignment = center
        cell.border = border
        if idx % 2 == 1:
            cell.fill = alt_fill

        current_row += 1

    # TOTAL row
    cell = ws.cell(current_row, 1, "TOTAL")
    cell.fill = total_fill
    cell.font = bold
    cell.alignment = center
    cell.border = border

    col = 2
    for cat in categories:
        cat_data = category_totals[cat]
        if cat_data["total_rows"] > 0:
            cat_pct = round(cat_data["done"] / cat_data["total_rows"] * 100, 1)
            cell1 = ws.cell(current_row, col, cat_pct)
            cell1.number_format = '0.0"%"'
        else:
            cell1 = ws.cell(current_row, col, "-")
        cell1.fill = total_fill
        cell1.font = bold
        cell1.alignment = center
        cell1.border = border

        if cat_data["count"] > 0:
            cell2 = ws.cell(current_row, col + 1, cat_data["count"])
            cell2.number_format = '#,##0'
        else:
            cell2 = ws.cell(current_row, col + 1, "-")
        cell2.fill = total_fill
        cell2.font = bold
        cell2.alignment = center
        cell2.border = border

        col += 2

    # Grand total count
    cell = ws.cell(current_row, col, grand_total_count)
    cell.number_format = '#,##0'
    cell.fill = total_fill
    cell.font = bold
    cell.alignment = center
    cell.border = border

    current_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12  # User column
    col_letter_idx = 2
    for cat in categories:
        ws.column_dimensions[get_column_letter(col_letter_idx)].width = 8  # Done%
        ws.column_dimensions[get_column_letter(col_letter_idx + 1)].width = 10  # Words/Chars
        col_letter_idx += 2
    ws.column_dimensions[get_column_letter(col_letter_idx)].width = 12  # Total

    return current_row


def build_total_sheet(wb):
    """
    Build TOTAL sheet from _DAILY_DATA.

    Aggregates by user across all dates and categories.
    Includes both tester stats and manager stats (Fixed, Reported, Checking, Pending).
    Separates EN and CN testers into distinct sections.
    """
    # Delete and recreate sheet to handle merged cells properly
    if "TOTAL" in wb.sheetnames:
        del wb["TOTAL"]
    ws = wb.create_sheet("TOTAL", 1)

    data_ws = wb["_DAILY_DATA"]

    # Load tester mapping to separate EN/CN
    tester_mapping = load_tester_mapping()

    # Read raw data and aggregate by user (now includes manager stats + total_rows + nonissue)
    # IMPORTANT: _DAILY_DATA stores CUMULATIVE values per (date, user, category)
    # We must only use the LATEST date's data for each (user, category) to avoid double-counting

    # First pass: find the latest date for each (user, category)
    latest_data = {}  # (user, category) -> {date, row_data}

    # Schema: Date(1), User(2), Category(3), TotalRows(4), Done(5), Issues(6), NoIssue(7), Blocked(8), Fixed(9), Reported(10), Checking(11), NonIssue(12), WordCount(13)
    for row in range(2, data_ws.max_row + 1):
        date = data_ws.cell(row, 1).value
        user = data_ws.cell(row, 2).value
        category = data_ws.cell(row, 3).value

        if not user or not date:
            continue

        key = (user, category)

        # Keep the row with the latest date for each (user, category)
        if key not in latest_data or str(date) > str(latest_data[key]["date"]):
            latest_data[key] = {
                "date": date,
                "category": category,  # Store category for breakdown tables
                "total_rows": data_ws.cell(row, 4).value or 0,
                "done": data_ws.cell(row, 5).value or 0,
                "issues": data_ws.cell(row, 6).value or 0,
                "no_issue": data_ws.cell(row, 7).value or 0,
                "blocked": data_ws.cell(row, 8).value or 0,
                "fixed": data_ws.cell(row, 9).value or 0,
                "reported": data_ws.cell(row, 10).value or 0,
                "checking": data_ws.cell(row, 11).value or 0,
                "nonissue": data_ws.cell(row, 12).value or 0,
                "word_count": data_ws.cell(row, 13).value or 0,  # Words (EN) or Chars (CN)
                "korean": data_ws.cell(row, 14).value or 0  # Korean status count
            }

    # Second pass: aggregate latest data by user (sum across categories)
    user_data = defaultdict(lambda: {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0,
        "fixed": 0, "reported": 0, "checking": 0, "nonissue": 0, "korean": 0
    })

    for (user, category), data in latest_data.items():
        user_data[user]["total_rows"] += data["total_rows"]
        user_data[user]["done"] += data["done"]
        user_data[user]["issues"] += data["issues"]
        user_data[user]["no_issue"] += data["no_issue"]
        user_data[user]["blocked"] += data["blocked"]
        user_data[user]["fixed"] += data["fixed"]
        user_data[user]["reported"] += data["reported"]
        user_data[user]["checking"] += data["checking"]
        user_data[user]["nonissue"] += data["nonissue"]
        user_data[user]["korean"] += data.get("korean", 0)

    if not user_data:
        ws.cell(1, 1, "No data yet")
        return

    # Separate users by language
    en_users = sorted([u for u in user_data.keys() if tester_mapping.get(u, "EN") == "EN"])
    cn_users = sorted([u for u in user_data.keys() if tester_mapping.get(u) == "CN"])
    all_users = sorted(user_data.keys())

    # Styles
    title_fill = PatternFill(start_color=TRACKER_STYLES["title_color"], end_color=TRACKER_STYLES["title_color"], fill_type="solid")
    en_title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for EN
    cn_title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red for CN
    header_fill = PatternFill(start_color=TRACKER_STYLES["header_color"], end_color=TRACKER_STYLES["header_color"], fill_type="solid")
    manager_header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    alt_fill = PatternFill(start_color=TRACKER_STYLES["alt_row_color"], end_color=TRACKER_STYLES["alt_row_color"], fill_type="solid")
    total_fill = PatternFill(start_color=TRACKER_STYLES["total_row_color"], end_color=TRACKER_STYLES["total_row_color"], fill_type="solid")
    border = Border(
        left=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        right=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        top=Side(style='thin', color=TRACKER_STYLES["border_color"]),
        bottom=Side(style='thin', color=TRACKER_STYLES["border_color"])
    )
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Total columns: 6 tester + 4 manager = 10 (removed Completion/Actual Issues - already in Category Breakdown)
    tester_headers = ["User", "Done", "Issues", "No Issue", "Blocked", "Korean"]
    manager_headers = ["Fixed", "Reported", "Checking", "Pending"]
    total_cols = len(tester_headers) + len(manager_headers)

    def build_section(ws, start_row, section_title, section_fill, users_list, user_data, user_row_tracker=None):
        """Build a section (EN or CN) with title, headers, data rows, and subtotal.

        user_row_tracker: Optional list to append (row_num, user) tuples for chart references.
        """
        if not users_list:
            return start_row  # No data for this section

        current_row = start_row

        # Section Title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, section_title)
        title_cell.fill = section_fill
        title_cell.font = white_bold
        title_cell.alignment = center
        current_row += 1

        # Headers
        for col, header in enumerate(tester_headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border

        manager_start_col = len(tester_headers) + 1
        for col, header in enumerate(manager_headers, manager_start_col):
            cell = ws.cell(current_row, col, header)
            cell.fill = manager_header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        current_row += 1

        # Data rows
        section_total = {
            "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
            "fixed": 0, "reported": 0, "checking": 0, "pending": 0, "nonissue": 0
        }

        for idx, user in enumerate(users_list):
            data = user_data[user]
            total_rows = data["total_rows"]
            done = data["done"]
            issues = data["issues"]
            no_issue = data["no_issue"]
            blocked = data["blocked"]
            korean = data.get("korean", 0)
            fixed = data["fixed"]
            reported = data["reported"]
            checking = data["checking"]
            nonissue = data["nonissue"]
            # Pending = Issues - Fixed - Reported - Checking - NonIssue
            pending = issues - fixed - reported - checking - nonissue
            if pending < 0:
                pending = 0

            # Accumulate section totals
            section_total["total_rows"] += total_rows
            section_total["done"] += done
            section_total["issues"] += issues
            section_total["no_issue"] += no_issue
            section_total["blocked"] += blocked
            section_total["korean"] += korean
            section_total["fixed"] += fixed
            section_total["reported"] += reported
            section_total["checking"] += checking
            section_total["pending"] += pending
            section_total["nonissue"] += nonissue

            # Row data: User, Done, Issues, No Issue, Blocked, Korean, Fixed, Reported, Checking, Pending
            row_data = [user, done, issues, no_issue, blocked, korean, fixed, reported, checking, pending]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(current_row, col, value)
                cell.alignment = center
                cell.border = border
                if idx % 2 == 1:
                    cell.fill = alt_fill

            # Track this row for chart references
            if user_row_tracker is not None:
                user_row_tracker.append((current_row, user))

            current_row += 1

        # Section subtotal row
        st = section_total
        subtotal_data = [
            "SUBTOTAL", st["done"], st["issues"], st["no_issue"], st["blocked"], st["korean"],
            st["fixed"], st["reported"], st["checking"], st["pending"]
        ]

        for col, value in enumerate(subtotal_data, 1):
            cell = ws.cell(current_row, col, value)
            cell.fill = total_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        current_row += 1

        return current_row, section_total

    # Build EN section
    current_row = 1
    en_total = None
    user_data_rows = []  # Track (row_num, user) for chart references to main table
    if en_users:
        current_row, en_total = build_section(ws, current_row, "EN TESTER STATS", en_title_fill, en_users, user_data, user_data_rows)
        current_row += 1  # Empty row between sections

    # Build CN section
    cn_total = None
    if cn_users:
        current_row, cn_total = build_section(ws, current_row, "CN TESTER STATS", cn_title_fill, cn_users, user_data, user_data_rows)
        current_row += 1  # Empty row before grand total

    # Grand total row (combines EN + CN)
    grand_total = {
        "total_rows": 0, "done": 0, "issues": 0, "no_issue": 0, "blocked": 0, "korean": 0,
        "fixed": 0, "reported": 0, "checking": 0, "pending": 0, "nonissue": 0
    }
    for t in [en_total, cn_total]:
        if t:
            for key in grand_total:
                grand_total[key] += t[key]

    if grand_total["total_rows"] > 0:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        gt_title = ws.cell(current_row, 1, "GRAND TOTAL (ALL LANGUAGES)")
        gt_title.fill = title_fill
        gt_title.font = Font(bold=True, size=12)
        gt_title.alignment = center
        current_row += 1

        gt = grand_total
        total_row_data = [
            "TOTAL", gt["done"], gt["issues"], gt["no_issue"], gt["blocked"], gt["korean"],
            gt["fixed"], gt["reported"], gt["checking"], gt["pending"]
        ]

        for col, value in enumerate(total_row_data, 1):
            cell = ws.cell(current_row, col, value)
            cell.fill = total_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        current_row += 1

    # Set column widths with auto-sizing + padding
    PADDING = 2
    all_headers = tester_headers + manager_headers
    for col, header in enumerate(all_headers, 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = len(header) + PADDING

    # NOTE: Charts removed from TOTAL tab - replaced with Ranking Table

    breakdown_start_row = current_row + 2

    # === Add Category Breakdown Tables (EN and CN) ===
    # These show per-category completion % and word/char counts

    # EN Category Breakdown (if there are EN users)
    if en_users:
        breakdown_start_row = build_category_breakdown_section(
            ws, breakdown_start_row, latest_data, en_users,
            is_english=True, tester_mapping=tester_mapping
        )
        breakdown_start_row += 2  # Gap between tables

    # CN Category Breakdown (if there are CN users)
    if cn_users:
        breakdown_start_row = build_category_breakdown_section(
            ws, breakdown_start_row, latest_data, cn_users,
            is_english=False, tester_mapping=tester_mapping
        )
        breakdown_start_row += 2  # Gap before ranking

    # === Add Ranking Table ===
    # Shows tester ranking based on 70% Completion + 30% Actual Issues
    breakdown_start_row = build_ranking_table(ws, breakdown_start_row, user_data, tester_mapping)


# build_graphs_sheet removed - charts now embedded in DAILY/TOTAL


# =============================================================================
# TRANSFER FEATURE - Migrate tester work from OLD to NEW QA file structures
# =============================================================================

def sanitize_stringid_for_match(value):
    """
    Normalize STRINGID for comparison during transfer.

    Handles:
    - None values
    - INT vs STRING type mismatch
    - Scientific notation (e.g., "1.23E+15")
    - Leading/trailing whitespace
    """
    if value is None:
        return ""
    s = str(value).strip()
    # Handle scientific notation
    if 'e' in s.lower():
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    return s


def is_english_file(xlsx_path):
    """Detect if file is English based on filename."""
    name = str(xlsx_path).upper()
    # Check for explicit language markers
    if "_ENG" in name or "_EN." in name or "ENG_" in name:
        return True
    # Check for non-English markers
    non_eng = ["_FRE", "_ZHO", "_JPN", "_DEU", "_SPA", "_ITA", "_POR", "_RUS", "_KOR"]
    for marker in non_eng:
        if marker in name:
            return False
    # Default to English if no markers found
    return True


def get_translation_column(category, is_english):
    """
    Get translation column index based on category and language.

    Returns:
        int: 1-based column index for the translation column
    """
    cols = TRANSLATION_COLS.get(category, {"eng": 2, "other": 3})
    return cols["eng"] if is_english else cols["other"]


def discover_qa_folders_in(base_folder):
    """
    Discover QA folders in a given base folder.

    Returns:
        list of dict: [{username, category, xlsx_path, folder_path, images}, ...]
    """
    folders = []
    if not base_folder.exists():
        return folders

    for folder in base_folder.iterdir():
        if not folder.is_dir():
            continue

        # Parse folder name: {Username}_{Category}
        folder_name = folder.name
        if "_" not in folder_name:
            continue

        parts = folder_name.rsplit("_", 1)
        if len(parts) != 2:
            continue

        username, category = parts
        if category not in CATEGORIES:
            continue

        # Find xlsx file (skip temp files starting with ~)
        xlsx_files = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~")]
        if not xlsx_files:
            continue

        # Use the largest xlsx file (likely the real one)
        xlsx_path = max(xlsx_files, key=lambda x: x.stat().st_size)

        # Collect images
        images = [f for f in folder.iterdir()
                  if f.suffix.lower() in IMAGE_EXTENSIONS]

        folders.append({
            "username": username,
            "category": category,
            "xlsx_path": xlsx_path,
            "folder_path": folder,
            "images": images,
        })

    return folders


def find_matching_row_for_transfer(old_row_data, new_ws, category, is_english):
    """
    Find matching row in NEW file for OLD row data.

    Uses 2-step cascade:
    1. STRINGID + Translation match (exact)
    2. Translation only match (fallback)

    Args:
        old_row_data: dict with {stringid, translation, row_num}
        new_ws: New worksheet to search in
        category: Category name for column detection
        is_english: Whether file is English

    Returns:
        tuple: (new_row_num, match_type) or (None, None)
        match_type: "stringid+trans" or "trans_only"
    """
    old_stringid = sanitize_stringid_for_match(old_row_data.get("stringid"))
    old_trans = str(old_row_data.get("translation", "")).strip()

    if not old_trans:
        return None, None

    trans_col = get_translation_column(category, is_english)
    stringid_col = find_column_by_header(new_ws, "STRINGID")

    # Step 1: Try STRINGID + Translation match
    if old_stringid and stringid_col:
        for row in range(2, new_ws.max_row + 1):
            new_stringid = sanitize_stringid_for_match(new_ws.cell(row, stringid_col).value)
            new_trans = str(new_ws.cell(row, trans_col).value or "").strip()

            if new_stringid == old_stringid and new_trans == old_trans:
                return row, "stringid+trans"

    # Step 2: Fall back to Translation only
    for row in range(2, new_ws.max_row + 1):
        new_trans = str(new_ws.cell(row, trans_col).value or "").strip()
        if new_trans == old_trans:
            return row, "trans_only"

    return None, None


def find_matching_row_for_item_transfer(old_row_data, new_ws, is_english):
    """
    Item-specific matching: requires BOTH ItemName AND ItemDesc to match.

    Uses 2-step cascade:
    1. ItemName + ItemDesc + STRINGID (all 3 must match)
    2. ItemName + ItemDesc (both translations must match, fallback)
    3. No match -> return None

    This is stricter than the generic transfer because Item descriptions
    now come from KnowledgeKey (not ItemDesc attribute), so we need both
    ItemName and ItemDesc to identify the correct row.

    Args:
        old_row_data: dict with {stringid, item_name, item_desc, row_num}
        new_ws: New worksheet to search in
        is_english: Whether file is English

    Returns:
        tuple: (new_row_num, match_type) or (None, None)
        match_type: "name+desc+stringid" or "name+desc"
    """
    old_stringid = sanitize_stringid_for_match(old_row_data.get("stringid"))
    old_item_name = str(old_row_data.get("item_name", "")).strip()
    old_item_desc = str(old_row_data.get("item_desc", "")).strip()

    # Need at least ItemName to match
    if not old_item_name:
        return None, None

    # Get column positions for Item
    name_col = TRANSLATION_COLS["Item"]["eng"] if is_english else TRANSLATION_COLS["Item"]["other"]
    desc_col = ITEM_DESC_COLS["eng"] if is_english else ITEM_DESC_COLS["other"]
    stringid_col = find_column_by_header(new_ws, "STRINGID")

    # Step 1: ItemName + ItemDesc + STRINGID (strictest match)
    if old_stringid and stringid_col:
        for row in range(2, new_ws.max_row + 1):
            new_stringid = sanitize_stringid_for_match(new_ws.cell(row, stringid_col).value)
            new_item_name = str(new_ws.cell(row, name_col).value or "").strip()
            new_item_desc = str(new_ws.cell(row, desc_col).value or "").strip()

            if (new_stringid == old_stringid and
                new_item_name == old_item_name and
                new_item_desc == old_item_desc):
                return row, "name+desc+stringid"

    # Step 2: ItemName + ItemDesc only (fallback)
    for row in range(2, new_ws.max_row + 1):
        new_item_name = str(new_ws.cell(row, name_col).value or "").strip()
        new_item_desc = str(new_ws.cell(row, desc_col).value or "").strip()

        if new_item_name == old_item_name and new_item_desc == old_item_desc:
            return row, "name+desc"

    # Step 3: No match
    return None, None


def transfer_sheet_data(old_ws, new_ws, category, is_english, old_folder_path=None):
    """
    Transfer COMMENT/STATUS/SCREENSHOT from old sheet to new sheet.

    For Item category: Uses stricter matching with ItemName + ItemDesc + STRINGID.
    For other categories: Uses STRINGID + Translation matching.

    Args:
        old_ws: Source worksheet (OLD)
        new_ws: Target worksheet (NEW)
        category: Category name
        is_english: Whether file is English
        old_folder_path: Path to old folder (for hyperlink auto-fix)

    Returns:
        dict: {total, stringid_match, trans_only, unmatched,
               name_desc_stringid_match, name_desc_match} (Item-specific stats)
    """
    stats = {
        "total": 0,
        "stringid_match": 0,
        "trans_only": 0,
        "unmatched": 0,
        # Item-specific stats
        "name_desc_stringid_match": 0,
        "name_desc_match": 0,
    }

    is_item_category = category.lower() == "item"

    # Find columns in OLD file
    old_trans_col = get_translation_column(category, is_english)
    old_stringid_col = find_column_by_header(old_ws, "STRINGID")
    old_comment_col = find_column_by_header(old_ws, "COMMENT")
    old_status_col = find_column_by_header(old_ws, "STATUS")
    old_screenshot_col = find_column_by_header(old_ws, "SCREENSHOT")

    # For Item category, also get ItemDesc column
    old_desc_col = None
    if is_item_category:
        old_desc_col = ITEM_DESC_COLS["eng"] if is_english else ITEM_DESC_COLS["other"]

    # Find columns in NEW file
    new_comment_col = find_column_by_header(new_ws, "COMMENT")
    new_status_col = find_column_by_header(new_ws, "STATUS")
    new_screenshot_col = find_column_by_header(new_ws, "SCREENSHOT")

    if not old_comment_col and not old_status_col:
        return stats  # Nothing to transfer

    # Process each row in OLD file
    for old_row in range(2, old_ws.max_row + 1):
        old_comment = old_ws.cell(old_row, old_comment_col).value if old_comment_col else None
        old_status = old_ws.cell(old_row, old_status_col).value if old_status_col else None
        old_screenshot_cell = old_ws.cell(old_row, old_screenshot_col) if old_screenshot_col else None
        old_screenshot = old_screenshot_cell.value if old_screenshot_cell else None
        old_screenshot_hyperlink = old_screenshot_cell.hyperlink if old_screenshot_cell else None

        # Skip rows with no work
        if not old_comment and not old_status:
            continue

        stats["total"] += 1

        # Build old row data for matching
        old_row_data = {
            "stringid": old_ws.cell(old_row, old_stringid_col).value if old_stringid_col else None,
            "translation": old_ws.cell(old_row, old_trans_col).value,
            "row_num": old_row,
        }

        # For Item category, add ItemName and ItemDesc for stricter matching
        if is_item_category:
            old_row_data["item_name"] = old_ws.cell(old_row, old_trans_col).value
            old_row_data["item_desc"] = old_ws.cell(old_row, old_desc_col).value if old_desc_col else ""

        # Find matching row in NEW file
        if is_item_category:
            # Use Item-specific matching (ItemName + ItemDesc + STRINGID)
            new_row, match_type = find_matching_row_for_item_transfer(old_row_data, new_ws, is_english)
        else:
            # Use generic matching (STRINGID + Translation)
            new_row, match_type = find_matching_row_for_transfer(old_row_data, new_ws, category, is_english)

        if new_row is None:
            stats["unmatched"] += 1
            continue

        # Track match type stats
        if match_type == "stringid+trans":
            stats["stringid_match"] += 1
        elif match_type == "trans_only":
            stats["trans_only"] += 1
        elif match_type == "name+desc+stringid":
            stats["name_desc_stringid_match"] += 1
        elif match_type == "name+desc":
            stats["name_desc_match"] += 1

        # Transfer data to NEW row
        if new_comment_col and old_comment:
            new_ws.cell(new_row, new_comment_col, old_comment)
        if new_status_col and old_status:
            new_ws.cell(new_row, new_status_col, old_status)
        if new_screenshot_col and old_screenshot:
            new_cell = new_ws.cell(new_row, new_screenshot_col, old_screenshot)
            # Also transfer the hyperlink (not just the display text)
            if old_screenshot_hyperlink:
                new_cell.hyperlink = old_screenshot_hyperlink.target
            elif old_folder_path:
                # AUTO-FIX: If no hyperlink, try to find file in old folder
                filename = str(old_screenshot).strip()
                actual_file = find_file_in_folder(filename, old_folder_path)
                if actual_file:
                    # Create hyperlink to the actual file
                    new_cell.hyperlink = actual_file

    return stats


def detect_duplicate_translations(old_wb, category, is_english):
    """
    Detect translations that appear multiple times with DIFFERENT comments.

    This helps identify potential data loss during Transfer Step 2 (translation-only match)
    where we take the first match but other comments might be lost.

    Args:
        old_wb: OLD workbook
        category: Category name
        is_english: Whether file is English

    Returns:
        dict: {sheet_name: [(translation, [comment1, comment2, ...]), ...]}
    """
    duplicates = {}

    for sheet_name in old_wb.sheetnames:
        old_ws = old_wb[sheet_name]
        trans_col = get_translation_column(category, is_english)
        comment_col = find_column_by_header(old_ws, "COMMENT")

        if not comment_col:
            continue

        # Group comments by translation
        trans_to_comments = {}
        for row in range(2, old_ws.max_row + 1):
            trans = str(old_ws.cell(row, trans_col).value or "").strip()
            comment = str(old_ws.cell(row, comment_col).value or "").strip()

            if not trans or not comment:
                continue

            if trans not in trans_to_comments:
                trans_to_comments[trans] = set()
            trans_to_comments[trans].add(comment)

        # Find translations with multiple different comments
        sheet_duplicates = []
        for trans, comments in trans_to_comments.items():
            if len(comments) > 1:
                sheet_duplicates.append((trans, sorted(comments)))

        if sheet_duplicates:
            duplicates[sheet_name] = sheet_duplicates

    return duplicates


def write_duplicate_translation_report(duplicates, output_folder, username, category):
    """
    Write a report file listing translations with different comments.

    Args:
        duplicates: dict from detect_duplicate_translations()
        output_folder: Path to user's output folder
        username: Tester username
        category: Category name
    """
    if not duplicates:
        return None

    report_path = output_folder / "DUPLICATE_TRANSLATION_REPORT.txt"

    total_count = sum(len(items) for items in duplicates.values())

    lines = [
        "=" * 70,
        "DUPLICATE TRANSLATION REPORT",
        "=" * 70,
        f"Tester: {username}",
        f"Category: {category}",
        f"Total duplicates: {total_count}",
        "",
        "These translations appear multiple times with DIFFERENT comments.",
        "During transfer, only the FIRST comment was kept.",
        "Review this list to ensure no important feedback was lost.",
        "",
        "-" * 70,
        "",
    ]

    for sheet_name, items in sorted(duplicates.items()):
        lines.append(f"Sheet: {sheet_name}")
        lines.append("-" * 40)

        for trans, comments in items:
            # Truncate long translations for readability
            trans_display = trans[:80] + "..." if len(trans) > 80 else trans
            lines.append(f"\nTranslation: {trans_display}")
            lines.append("Comments:")
            for i, comment in enumerate(comments, 1):
                comment_display = comment[:100] + "..." if len(comment) > 100 else comment
                lines.append(f"  {i}. {comment_display}")

        lines.append("")

    lines.append("=" * 70)
    lines.append("End of report")

    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return report_path


def transfer_folder_data(old_folder, new_folder, output_dir, tester_mapping):
    """
    Transfer all sheet data from OLD folder to NEW folder, save to output.

    Args:
        old_folder: dict with folder info from discover_qa_folders_in()
        new_folder: dict with folder info from discover_qa_folders_in()
        output_dir: Path to QAfolder (output)
        tester_mapping: dict {tester_name: "EN" or "CN"} from load_tester_mapping()

    Returns:
        dict: {(username, category): {total, stringid_match, trans_only, unmatched,
               name_desc_stringid_match, name_desc_match}}
    """
    username = old_folder["username"]
    category = old_folder["category"]
    # Use tester mapping to determine language (not filename detection)
    is_english = tester_mapping.get(username, "EN") == "EN"
    is_item_category = category.lower() == "item"

    # Load workbooks
    old_wb = safe_load_workbook(old_folder["xlsx_path"])
    new_wb = safe_load_workbook(new_folder["xlsx_path"])

    combined_stats = {
        "total": 0,
        "stringid_match": 0,
        "trans_only": 0,
        "unmatched": 0,
        # Item-specific stats
        "name_desc_stringid_match": 0,
        "name_desc_match": 0,
    }

    # Process each sheet that exists in both
    for sheet_name in old_wb.sheetnames:
        if sheet_name not in new_wb.sheetnames:
            print(f"    WARN: Sheet '{sheet_name}' not in NEW file, skipping")
            continue

        old_ws = old_wb[sheet_name]
        new_ws = new_wb[sheet_name]

        # Pass old folder path for hyperlink auto-fix
        sheet_stats = transfer_sheet_data(old_ws, new_ws, category, is_english, old_folder["folder_path"])

        # Accumulate stats
        for key in combined_stats:
            combined_stats[key] += sheet_stats.get(key, 0)

        if sheet_stats["total"] > 0:
            if is_item_category:
                # Item-specific output: name+desc+stringid and name+desc matches
                print(f"    Sheet '{sheet_name}': {sheet_stats['name_desc_stringid_match']}+{sheet_stats['name_desc_match']} transferred, {sheet_stats['unmatched']} unmatched")
            else:
                # Generic output: stringid+trans and trans_only matches
                print(f"    Sheet '{sheet_name}': {sheet_stats['stringid_match']}+{sheet_stats['trans_only']} transferred, {sheet_stats['unmatched']} unmatched")

    # Create output folder
    output_folder = output_dir / f"{username}_{category}"
    output_folder.mkdir(parents=True, exist_ok=True)

    # Detect duplicate translations with different comments
    duplicates = detect_duplicate_translations(old_wb, category, is_english)
    if duplicates:
        report_path = write_duplicate_translation_report(duplicates, output_folder, username, category)
        if report_path:
            total_dups = sum(len(items) for items in duplicates.values())
            print(f"    WARNING: {total_dups} translations have different comments!")
            print(f"    Report: {report_path.name}")

    # Save the new workbook with transferred data
    output_xlsx = output_folder / new_folder["xlsx_path"].name
    new_wb.save(output_xlsx)

    # Copy images from OLD folder
    for img in old_folder["images"]:
        dest = output_folder / img.name
        shutil.copy2(img, dest)

    old_wb.close()
    new_wb.close()

    return {(username, category): combined_stats}


def print_transfer_report(stats):
    """Print formatted transfer report to terminal."""
    print()
    print("═" * 79)
    print("                              TRANSFER REPORT")
    print("═" * 79)
    print()
    print(f"{'Tester':<20}{'Category':<12}{'Total':>7}{'Exact':>10}{'Fallback':>10}{'Success %':>12}")
    print("─" * 79)

    grand_total = 0
    grand_exact = 0
    grand_fallback = 0

    for (tester, category), data in sorted(stats.items()):
        total = data["total"]
        is_item = category.lower() == "item"

        if is_item:
            # Item uses name+desc+stringid and name+desc matching
            exact = data.get("name_desc_stringid_match", 0)
            fallback = data.get("name_desc_match", 0)
        else:
            # Others use stringid+trans and trans_only matching
            exact = data.get("stringid_match", 0)
            fallback = data.get("trans_only", 0)

        success = (exact + fallback) / total * 100 if total > 0 else 0

        print(f"{tester:<20}{category:<12}{total:>7}{exact:>10}{fallback:>10}{success:>11.1f}%")

        grand_total += total
        grand_exact += exact
        grand_fallback += fallback

    print("─" * 79)
    grand_success = (grand_exact + grand_fallback) / grand_total * 100 if grand_total > 0 else 0
    print(f"{'TOTAL':<20}{'':<12}{grand_total:>7}{grand_exact:>10}{grand_fallback:>10}{grand_success:>11.1f}%")
    print("═" * 79)
    print()
    print("Legend:")
    print("  Exact    = Strong match (STRINGID+Trans for most, ItemName+ItemDesc+STRINGID for Item)")
    print("  Fallback = Weaker match (Trans only for most, ItemName+ItemDesc for Item)")
    unmatched = grand_total - grand_exact - grand_fallback
    print(f"  Unmatched = {unmatched} rows (not transferred)")
    print()


def transfer_qa_files():
    """
    Main transfer function: QAfolderOLD -> QAfolderNEW -> QAfolder

    Returns:
        bool: True if transfer completed successfully
    """
    print()
    print("=" * 60)
    print("QA File Transfer (OLD -> NEW structure)")
    print("=" * 60)

    # Check folders exist
    if not QA_FOLDER_OLD.exists():
        print(f"ERROR: QAfolderOLD not found at {QA_FOLDER_OLD}")
        print("Please create QAfolderOLD/ with your OLD QA files.")
        return False

    if not QA_FOLDER_NEW.exists():
        print(f"ERROR: QAfolderNEW not found at {QA_FOLDER_NEW}")
        print("Please create QAfolderNEW/ with your NEW (empty) QA files.")
        return False

    # Discover folders
    old_folders = discover_qa_folders_in(QA_FOLDER_OLD)
    new_folders = discover_qa_folders_in(QA_FOLDER_NEW)

    if not old_folders:
        print("ERROR: No valid QA folders found in QAfolderOLD/")
        return False

    if not new_folders:
        print("ERROR: No valid QA folders found in QAfolderNEW/")
        return False

    print(f"Found {len(old_folders)} OLD folder(s), {len(new_folders)} NEW folder(s)")

    # Load tester→language mapping (same as Build uses)
    print("\nLoading tester→language mapping...")
    tester_mapping = load_tester_mapping()

    # Build lookup for NEW folders
    new_lookup = {f"{f['username']}_{f['category']}": f for f in new_folders}

    # Transfer each OLD folder
    all_stats = {}

    for old_folder in old_folders:
        key = f"{old_folder['username']}_{old_folder['category']}"
        username = old_folder['username']
        lang = tester_mapping.get(username, "EN")
        in_mapping = username in tester_mapping
        print(f"\nTransferring: {key} -> {lang}{'' if in_mapping else ' (not in mapping, default)'}")

        if key not in new_lookup:
            print(f"  WARN: No matching NEW folder for {key}, skipping")
            continue

        new_folder = new_lookup[key]
        folder_stats = transfer_folder_data(old_folder, new_folder, QA_FOLDER, tester_mapping)
        all_stats.update(folder_stats)

    # Print report
    if all_stats:
        print_transfer_report(all_stats)

    print("Transfer complete!")
    print(f"Output: {QA_FOLDER}")
    print("You can now run 'Build Masterfiles' to compile.")

    return True


# =============================================================================
# GUI - Main application window with Transfer and Build buttons
# =============================================================================

def run_gui():
    """Launch the GUI application."""
    import tkinter as tk
    from tkinter import ttk, messagebox

    class QACompilerGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("QA Excel Compiler")
            self.root.geometry("450x250")
            self.root.resizable(False, False)

            # Title
            title_label = tk.Label(
                root,
                text="QA Excel Compiler",
                font=("Arial", 16, "bold")
            )
            title_label.pack(pady=15)

            # Transfer button
            transfer_btn = ttk.Button(
                root,
                text="Transfer QA Files\n(QAfolderOLD → QAfolder)",
                command=self.do_transfer,
                width=35
            )
            transfer_btn.pack(pady=10, ipady=10)

            # Build button
            build_btn = ttk.Button(
                root,
                text="Build Masterfiles\n(QAfolder → Masterfolder)",
                command=self.do_build,
                width=35
            )
            build_btn.pack(pady=10, ipady=10)

            # Status label
            self.status_var = tk.StringVar(value="Ready")
            status_label = tk.Label(
                root,
                textvariable=self.status_var,
                font=("Arial", 10),
                fg="gray"
            )
            status_label.pack(pady=15)

        def do_transfer(self):
            self.status_var.set("Transferring...")
            self.root.update()
            try:
                success = transfer_qa_files()
                if success:
                    self.status_var.set("Transfer complete!")
                    messagebox.showinfo("Success", "Transfer completed!\nCheck console for details.")
                else:
                    self.status_var.set("Transfer failed - check console")
                    messagebox.showerror("Error", "Transfer failed.\nCheck console for details.")
            except Exception as e:
                self.status_var.set("Error occurred")
                messagebox.showerror("Error", f"Transfer failed:\n{str(e)}")

        def do_build(self):
            self.status_var.set("Building...")
            self.root.update()
            try:
                main()
                self.status_var.set("Build complete!")
                messagebox.showinfo("Success", "Build completed!\nCheck console for details.")
            except Exception as e:
                self.status_var.set("Error occurred")
                messagebox.showerror("Error", f"Build failed:\n{str(e)}")

    root = tk.Tk()
    app = QACompilerGUI(root)
    root.mainloop()


def process_category(category, qa_folders, master_folder, images_folder, lang_label, manager_status=None):
    """
    Process all QA folders for one category.

    Args:
        category: Category name (Quest, Knowledge, etc.)
        qa_folders: List of dicts with {folder_path, xlsx_path, username, category, images}
        master_folder: Target Master folder (EN or CN)
        images_folder: Target Images folder (EN or CN)
        lang_label: Language label for display ("EN" or "CN")
        manager_status: Dict of {sheet_name: {row: {user: status}}} from preprocess (optional)

    Returns:
        List of daily_entry dicts for tracker
    """
    if manager_status is None:
        manager_status = {}
    print(f"\n{'='*50}")
    print(f"Processing: {category} [{lang_label}] ({len(qa_folders)} folders)")
    print(f"{'='*50}")

    daily_entries = []  # NEW: Collect entries for tracker

    # Get or create master (in correct language folder)
    first_xlsx = qa_folders[0]["xlsx_path"]
    master_wb, master_path = get_or_create_master(category, master_folder, first_xlsx)

    if master_wb is None:
        return daily_entries

    # For EN Item category: Sort master A-Z by ItemName(ENG) column to match tester's sorted files
    # CN Item uses original indexing (no sorting)
    if category.lower() == "item" and lang_label == "EN":
        print("  [Item EN] Sorting master sheets A-Z by ItemName(ENG) column...")
        for sheet_name in master_wb.sheetnames:
            if sheet_name != "STATUS":
                ws = master_wb[sheet_name]
                # Find ItemName(ENG) column
                sort_col = find_column_by_header(ws, "ItemName(ENG)")
                if sort_col:
                    sort_worksheet_az(ws, sort_column=sort_col)
                    print(f"    Sorted {sheet_name} by column {sort_col} (ItemName(ENG))")
                else:
                    print(f"    WARNING: ItemName(ENG) column not found in {sheet_name}, skipping sort")

    # Track users and aggregated stats
    all_users = set()
    user_stats = defaultdict(lambda: {"total": 0, "issue": 0, "no_issue": 0, "blocked": 0, "korean": 0})
    total_images = 0
    total_screenshots = 0

    # NEW: Track word/char counts per user for category breakdown
    user_wordcount = defaultdict(int)  # username -> word count (EN) or char count (CN)
    is_english = (lang_label == "EN")
    trans_col_key = "eng" if is_english else "other"
    trans_col = TRANSLATION_COLS.get(category, {"eng": 2, "other": 3}).get(trans_col_key, 2)

    # Process each QA folder
    for qf in qa_folders:
        username = qf["username"]
        xlsx_path = qf["xlsx_path"]
        folder_path = qf["folder_path"]
        all_users.add(username)

        print(f"\n  Folder: {folder_path.name}")

        # Get file modification date for tracker
        file_mod_time = datetime.fromtimestamp(xlsx_path.stat().st_mtime)
        file_mod_date = file_mod_time.strftime("%Y-%m-%d")

        # Copy images to correct language folder
        image_mapping = copy_images_with_unique_names(qf, images_folder)
        total_images += len(image_mapping)

        # Load xlsx
        qa_wb = safe_load_workbook(xlsx_path)

        # For EN Item category: Sort input A-Z to match master ordering
        # This ensures consistent row alignment regardless of tester's file order
        if category.lower() == "item" and lang_label == "EN":
            for sheet_name in qa_wb.sheetnames:
                if sheet_name != "STATUS":
                    qa_ws = qa_wb[sheet_name]
                    sort_col = find_column_by_header(qa_ws, "ItemName(ENG)")
                    if sort_col:
                        sort_worksheet_az(qa_ws, sort_column=sort_col)

        for sheet_name in qa_wb.sheetnames:
            # Skip STATUS sheets
            if sheet_name == "STATUS":
                continue

            # Check if sheet exists in master
            if sheet_name not in master_wb.sheetnames:
                print(f"    WARN: Sheet '{sheet_name}' not in master, skipping")
                continue

            # Process sheet with image mapping for hyperlink transformation + xlsx_path for mod time
            # Also pass manager_status for this sheet to preserve manager entries
            sheet_manager_status = manager_status.get(sheet_name, {})
            result = process_sheet(master_wb[sheet_name], qa_wb[sheet_name], username, category, image_mapping, xlsx_path, sheet_manager_status)
            stats = result["stats"]

            manager_info = f", manager_restored:{result['manager_restored']}" if result.get('manager_restored', 0) > 0 else ""
            screenshot_info = f", screenshots:{result['screenshots']}" if result.get('screenshots', 0) > 0 else ""
            korean_info = f", korean:{stats['korean']}" if stats.get('korean', 0) > 0 else ""
            print(f"    {sheet_name}: {result['comments']} comments, {stats['issue']} issues, {stats['no_issue']} OK, {stats['blocked']} blocked{korean_info}{manager_info}{screenshot_info}")

            total_screenshots += result.get('screenshots', 0)

            # Aggregate stats for this user across all sheets
            user_stats[username]["total"] += stats["total"]
            user_stats[username]["issue"] += stats["issue"]
            user_stats[username]["no_issue"] += stats["no_issue"]
            user_stats[username]["blocked"] += stats["blocked"]
            user_stats[username]["korean"] += stats.get("korean", 0)

            # NEW: Count words (EN) or characters (CN) from translation column
            # ONLY count rows where STATUS is filled (DONE rows)
            qa_ws = qa_wb[sheet_name]
            qa_status_col = find_column_by_header(qa_ws, "STATUS")
            for row in range(2, qa_ws.max_row + 1):
                # Only count if STATUS is filled (row is DONE)
                if qa_status_col:
                    status_val = qa_ws.cell(row, qa_status_col).value
                    if not status_val or str(status_val).strip().upper() not in ["ISSUE", "NO ISSUE", "BLOCKED", "KOREAN"]:
                        continue  # Skip rows not marked as done
                cell_value = qa_ws.cell(row, trans_col).value
                if is_english:
                    user_wordcount[username] += count_words_english(cell_value)
                else:
                    user_wordcount[username] += count_chars_chinese(cell_value)

        qa_wb.close()

        # NEW: Collect entry for tracker (after processing all sheets for this user)
        daily_entries.append({
            "date": file_mod_date,
            "user": username,
            "category": category,
            "lang": lang_label,  # EN or CN for tracker separation
            "total_rows": user_stats[username]["total"],  # Total universe of rows
            "done": user_stats[username]["issue"] + user_stats[username]["no_issue"] + user_stats[username]["blocked"] + user_stats[username]["korean"],
            "issues": user_stats[username]["issue"],
            "no_issue": user_stats[username]["no_issue"],
            "blocked": user_stats[username]["blocked"],
            "korean": user_stats[username]["korean"],
            "word_count": user_wordcount[username]  # Words (EN) or Characters (CN)
        })

    # Update STATUS sheet (first tab, with stats)
    update_status_sheet(master_wb, all_users, user_stats)

    # Post-process: Hide rows/sheets/columns with no comments (focus on issues)
    hidden_rows, hidden_sheets, hidden_columns = hide_empty_comment_rows(master_wb)

    # Apply word wrap and autofit row heights
    autofit_rows_with_wordwrap(master_wb)

    # Save master
    master_wb.save(master_path)
    print(f"\n  Saved: {master_path}")
    if hidden_sheets:
        print(f"  Hidden sheets (no comments): {', '.join(hidden_sheets)}")
    if hidden_columns > 0:
        print(f"  Hidden: {hidden_columns} empty user columns (unhide in Excel if needed)")
    if hidden_rows > 0:
        print(f"  Hidden: {hidden_rows} rows with no comments (unhide in Excel if needed)")
    if total_images > 0:
        print(f"  Images: {total_images} copied to Images/, {total_screenshots} hyperlinks updated")

    return daily_entries  # NEW: Return entries for tracker


def main():
    """Main entry point."""
    print("="*60)
    print("QA Excel Compiler (EN/CN Separation + Manager Status)")
    print("="*60)
    print("Features:")
    print("  - Folder-based input: QAfolder/{Username}_{Category}/")
    print("  - Language separation: Masterfolder_EN/ and Masterfolder_CN/")
    print("  - Auto-routing testers by language mapping")
    print("  - Manager workflow: STATUS_{User} for FIXED/REPORTED/CHECKING")
    print("  - Manager status preservation on re-compile")
    print("  - Auto-hide rows with no comments (focus on issues)")
    print("  - Combined Progress Tracker at root level")
    print()

    # Load tester→language mapping
    print("Loading tester→language mapping...")
    tester_mapping = load_tester_mapping()

    # Ensure folders exist
    ensure_master_folders()

    # PREPROCESS: Collect manager status from existing Master files (both EN and CN)
    # Keyed by comment text for matching after rebuild
    print("\nCollecting manager status from existing Master files...")
    manager_status_en = collect_manager_status(MASTER_FOLDER_EN)
    manager_status_cn = collect_manager_status(MASTER_FOLDER_CN)
    # Count unique comment-status pairs
    total_en = sum(sum(len(statuses) for statuses in sheets.values()) for sheets in manager_status_en.values())
    total_cn = sum(sum(len(statuses) for statuses in sheets.values()) for sheets in manager_status_cn.values())
    if total_en > 0 or total_cn > 0:
        print(f"  Found {total_en} EN + {total_cn} CN manager status entries to preserve")
    else:
        print("  No existing manager status entries found")

    # Discover QA folders
    qa_folders = discover_qa_folders()

    if not qa_folders:
        print("\nNo valid QA folders found in QAfolder/")
        print("Expected format: QAfolder/{Username}_{Category}/")
        print("  - Each folder should contain one .xlsx file")
        print("  - Images should be in the same folder")
        print(f"Valid categories: {', '.join(CATEGORIES)}")
        return

    print(f"Found {len(qa_folders)} QA folder(s)")

    # Count total images
    total_images = sum(len(qf["images"]) for qf in qa_folders)
    if total_images > 0:
        print(f"Total images to process: {total_images}")

    # Group by category AND language
    by_category_en = defaultdict(list)
    by_category_cn = defaultdict(list)
    print("\nRouting testers by language:")
    for qf in qa_folders:
        username = qf["username"].strip()  # Strip whitespace for safety
        lang = tester_mapping.get(username, "EN")
        in_mapping = username in tester_mapping
        print(f"  {username} ({qf['category']}) -> {lang}{'' if in_mapping else ' (not in mapping, default)'}")
        if lang == "CN":
            by_category_cn[qf["category"]].append(qf)
        else:
            by_category_en[qf["category"]].append(qf)

    # Process each category for EN and CN
    all_daily_entries = []

    # Process EN
    for category in CATEGORIES:
        if category in by_category_en:
            category_manager_status = manager_status_en.get(category, {})
            entries = process_category(
                category, by_category_en[category],
                MASTER_FOLDER_EN, IMAGES_FOLDER_EN, "EN",
                category_manager_status
            )
            all_daily_entries.extend(entries)

    # Process CN
    for category in CATEGORIES:
        if category in by_category_cn:
            category_manager_status = manager_status_cn.get(category, {})
            entries = process_category(
                category, by_category_cn[category],
                MASTER_FOLDER_CN, IMAGES_FOLDER_CN, "CN",
                category_manager_status
            )
            all_daily_entries.extend(entries)

    # Show skipped categories
    for category in CATEGORIES:
        if category not in by_category_en and category not in by_category_cn:
            print(f"\nSKIP: No folders for category '{category}'")

    # Update LQA User Progress Tracker
    if all_daily_entries:
        print("\n" + "="*60)
        print("Updating LQA User Progress Tracker...")
        print("="*60)

        # Collect manager stats for tracker (FIXED/REPORTED/CHECKING counts)
        manager_stats = collect_manager_stats_for_tracker()

        tracker_wb, tracker_path = get_or_create_tracker()
        update_daily_data_sheet(tracker_wb, all_daily_entries, manager_stats)
        build_daily_sheet(tracker_wb)
        build_total_sheet(tracker_wb)

        # Remove GRAPHS sheet if it exists (deprecated)
        if "GRAPHS" in tracker_wb.sheetnames:
            del tracker_wb["GRAPHS"]

        tracker_wb.save(tracker_path)

        print(f"  Saved: {tracker_path}")
        print(f"  Sheets: DAILY (with chart), TOTAL (with chart)")

    print("\n" + "="*60)
    print("Compilation complete!")
    print(f"Output EN: {MASTER_FOLDER_EN}")
    print(f"Output CN: {MASTER_FOLDER_CN}")
    if all_daily_entries:
        print(f"Tracker: {TRACKER_PATH}")
    print("="*60)


if __name__ == "__main__":
    # Check for command-line arguments
    if len(sys.argv) > 1:
        arg = sys.argv[1].lower()
        if arg in ("--build", "-b", "build"):
            main()
        elif arg in ("--transfer", "-t", "transfer"):
            transfer_qa_files()
        elif arg in ("--help", "-h", "help"):
            print("QA Excel Compiler")
            print()
            print("Usage:")
            print("  python compile_qa.py          Launch GUI")
            print("  python compile_qa.py --build  Run build directly")
            print("  python compile_qa.py --transfer  Run transfer directly")
        else:
            print(f"Unknown argument: {arg}")
            print("Use --help for usage information")
    else:
        # Default: Launch GUI
        run_gui()
