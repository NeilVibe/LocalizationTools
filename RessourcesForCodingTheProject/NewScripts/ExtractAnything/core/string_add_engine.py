"""String Add engine – add missing LocStr nodes from source to target folder."""

from __future__ import annotations

import logging
import os
import shutil
from pathlib import Path

from lxml import etree

import config
from . import xml_parser
from .text_utils import normalize_text

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Key helpers
# ---------------------------------------------------------------------------

def _make_key(sid: str, so: str) -> tuple:
    """Return exact key for a StringID + StrOrigin pair (space-sensitive)."""
    return (sid.lower(), normalize_text(so))


def _collect_keys_from_root(root) -> set[tuple]:
    """Collect all (StringID, StrOrigin) keys from a parsed root element."""
    keys: set[tuple] = set()

    for elem in xml_parser.iter_locstr(root):
        _, sid = xml_parser.get_attr(elem, config.STRINGID_ATTRS)
        if not sid:
            continue
        _, so = xml_parser.get_attr(elem, config.STRORIGIN_ATTRS)
        so = so or ""
        keys.add(_make_key(sid, so))

    return keys


def _collect_source_entries(source_path: Path) -> list[dict]:
    """Parse source XML and return entries with raw_attribs preserved."""
    raw = xml_parser.read_xml_raw(source_path)
    if raw is None:
        return []

    root = xml_parser.parse_root_from_string(raw)
    entries: list[dict] = []

    for elem in xml_parser.iter_locstr(root):
        _, sid = xml_parser.get_attr(elem, config.STRINGID_ATTRS)
        if not sid:
            continue
        _, so = xml_parser.get_attr(elem, config.STRORIGIN_ATTRS)
        _, sv = xml_parser.get_attr(elem, config.STR_ATTRS)
        so = so or ""
        sv = sv or ""
        raw_attribs = dict(elem.attrib)
        entries.append({
            "string_id": sid,
            "str_origin": so,
            "str_value": sv,
            "key": _make_key(sid, so),
            "raw_attribs": raw_attribs,
        })

    return entries


def _collect_source_entries_from_file(source_path: Path) -> list[dict]:
    """Parse source XML or Excel and return entries with key field (auto-detect)."""
    suffix = source_path.suffix.lower()

    if suffix == ".xml":
        return _collect_source_entries(source_path)
    elif suffix in (".xlsx", ".xls"):
        from .input_parser import parse_input_file
        entries, _ = parse_input_file(source_path)
        # Add the key field expected by add logic
        for entry in entries:
            sid = entry.get("string_id", "")
            so = entry.get("str_origin", "")
            entry["key"] = _make_key(sid, so)
            # Ensure raw_attribs has standard attrs for XML output
            if not entry.get("raw_attribs"):
                entry["raw_attribs"] = {}
            ra = entry["raw_attribs"]
            if "StringId" not in ra and sid:
                ra["StringId"] = sid
            if "StrOrigin" not in ra and so:
                ra["StrOrigin"] = so
            sv = entry.get("str_value", "")
            if "Str" not in ra and sv:
                ra["Str"] = sv
        return entries
    else:
        logger.warning("Unsupported source type: %s", source_path.name)
        return []


def _dedup_source_entries(entries: list[dict]) -> list[dict]:
    """Remove source-internal duplicates (exact key only, space-sensitive)."""
    deduped: list[dict] = []
    seen: set[tuple] = set()
    for entry in entries:
        if entry["key"] in seen:
            continue
        seen.add(entry["key"])
        deduped.append(entry)
    return deduped


# ---------------------------------------------------------------------------
# Core: add missing entries to a single target file
# ---------------------------------------------------------------------------

def _add_to_target(
    source_entries: list[dict],
    target_path: Path,
    *,
    log_fn=None,
) -> tuple[int, list[dict]]:
    """Diff *source_entries* vs *target_path* and append missing entries.

    *source_entries* must already be deduped (call ``_dedup_source_entries``).
    Returns ``(count_added, report)``.
    """
    try:
        tree, root = xml_parser.parse_tree_from_file(target_path)
    except Exception as exc:
        if log_fn:
            log_fn(f"  Cannot parse {target_path.name}: {exc}", "warning")
        return 0, []

    target_keys = _collect_keys_from_root(root)

    # Skip files with no existing LocStr (likely not a languagedata file)
    if not target_keys:
        if log_fn:
            log_fn(f"  {target_path.name}: skipped (no LocStr entries)", "warning")
        return 0, []

    # Diff — exact match only (space-sensitive)
    missing: list[dict] = []
    for entry in source_entries:
        if entry["key"] in target_keys:
            continue
        missing.append(entry)

    if not missing:
        return 0, []

    # Detect indent from existing elements
    existing = list(root)
    if existing:
        sample_tail = existing[0].tail or ""
        indent = sample_tail.replace("\n", "") if "\n" in sample_tail else "    "
    else:
        indent = "    "
    child_tail = "\n" + indent
    closing_tail = "\n"

    if existing:
        existing[-1].tail = child_tail

    # Append missing entries
    report: list[dict] = []
    for i, entry in enumerate(missing):
        raw = entry["raw_attribs"]
        elem = etree.SubElement(root, "LocStr", **{k: str(v) for k, v in raw.items()})
        is_last = (i == len(missing) - 1)
        elem.tail = closing_tail if is_last else child_tail
        report.append({
            "string_id": entry["string_id"],
            "status": "ADDED",
            "target_file": target_path.name,
        })

    xml_parser.write_xml_tree(tree, target_path)
    return len(report), report


def _add_to_excel_target(
    source_entries: list[dict],
    target_path: Path,
    *,
    log_fn=None,
) -> tuple[int, list[dict]]:
    """Diff *source_entries* vs *target_path* (Excel) and append missing rows.

    Returns ``(count_added, report)``.
    """
    import openpyxl
    from .excel_reader import detect_headers
    from .text_utils import convert_linebreaks_for_xml, br_to_newline

    try:
        wb = openpyxl.load_workbook(str(target_path))
        ws = wb.active
    except Exception as exc:
        if log_fn:
            log_fn(f"  Cannot open {target_path.name}: {exc}", "warning")
        return 0, []

    hdrs = detect_headers(ws)
    sid_col = hdrs.get("stringid")
    so_col = hdrs.get("strorigin")
    str_col = hdrs.get("str")

    if sid_col is None:
        if log_fn:
            log_fn(f"  {target_path.name}: no StringID column — skipped", "warning")
        wb.close()
        return 0, []

    # Collect existing keys from target
    target_keys: set[tuple] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_col]).strip() if sid_col < len(row) and row[sid_col] else ""
        if not sid:
            continue
        so = str(row[so_col]).strip() if so_col is not None and so_col < len(row) and row[so_col] else ""
        so = convert_linebreaks_for_xml(so)
        target_keys.add(_make_key(sid, so))

    if not target_keys:
        if log_fn:
            log_fn(f"  {target_path.name}: skipped (no entries)", "warning")
        wb.close()
        return 0, []

    # Find missing
    missing = [e for e in source_entries if e["key"] not in target_keys]
    if not missing:
        wb.close()
        return 0, []

    # Append rows — use detected column positions
    report: list[dict] = []
    for entry in missing:
        next_row = ws.max_row + 1
        sid = entry.get("string_id", "")
        so = entry.get("str_origin", "")
        sv = entry.get("str_value", "")

        # Convert <br/> to newline for Excel display
        so_excel = br_to_newline(so)
        sv_excel = br_to_newline(sv)

        # Write into detected column positions (StringID as text format)
        cell_sid = ws.cell(row=next_row, column=sid_col + 1, value=sid)
        cell_sid.number_format = '@'

        if so_col is not None:
            ws.cell(row=next_row, column=so_col + 1, value=so_excel)
        if str_col is not None:
            ws.cell(row=next_row, column=str_col + 1, value=sv_excel)

        report.append({
            "string_id": sid,
            "status": "ADDED",
            "target_file": target_path.name,
        })

    # Backup before write
    bak_path = target_path.with_suffix(target_path.suffix + ".bak")
    try:
        shutil.copy2(target_path, bak_path)
    except Exception as exc:
        logger.error("CANNOT create backup of %s: %s — aborting write", target_path.name, exc)
        if log_fn:
            log_fn(f"  SKIPPED {target_path.name}: backup failed", "error")
        wb.close()
        return 0, []

    wb.save(str(target_path))
    wb.close()
    return len(report), report


# ---------------------------------------------------------------------------
# Public: file-to-folder add
# ---------------------------------------------------------------------------

def add_missing_folder(
    source_path: Path,
    target_folder: Path,
    *,
    log_fn=None,
    progress_fn=None,
) -> tuple[int, list[dict]]:
    """Add missing entries from *source_path* to all XMLs in *target_folder*.

    Source is parsed once, then each target XML is processed individually.
    Returns ``(total_added, full_report)``.
    """
    if log_fn:
        log_fn(f"Source: {source_path.name}")
        log_fn(f"Target folder: {target_folder}")

    source_entries = _collect_source_entries_from_file(source_path)
    if not source_entries:
        if log_fn:
            log_fn("No entries found in source.", "warning")
        return 0, []

    source_entries = _dedup_source_entries(source_entries)
    if log_fn:
        log_fn(f"Source: {len(source_entries):,} unique entries")

    # Scan for both XML and Excel targets
    target_files: list[Path] = []
    for pat in ("*.xml", "*.xlsx"):
        target_files.extend(target_folder.rglob(pat))
    target_files = sorted(set(f for f in target_files
                              if not f.name.startswith("~$")
                              and not _is_same_file(source_path, f)))

    total = len(target_files)
    if not target_files:
        if log_fn:
            log_fn("No XML/Excel files in target folder.", "warning")
        return 0, []

    if log_fn:
        log_fn(f"Found {total} target files (XML + Excel)")

    full_report: list[dict] = []
    total_added = 0

    for i, fpath in enumerate(target_files, 1):
        if progress_fn:
            progress_fn(i * 100 // total)

        suffix = fpath.suffix.lower()
        if suffix == ".xml":
            added, file_report = _add_to_target(source_entries, fpath, log_fn=log_fn)
        elif suffix in (".xlsx", ".xls"):
            added, file_report = _add_to_excel_target(source_entries, fpath, log_fn=log_fn)
        else:
            continue

        if file_report:
            total_added += added
            full_report.extend(file_report)
            if log_fn:
                log_fn(f"  {fpath.name}: {added} added", "success")
        else:
            if log_fn:
                log_fn(f"  {fpath.name}: nothing to add")

    return total_added, full_report


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def write_add_report(report: list[dict], output_dir: Path) -> Path:
    """Write a plain-text add report."""
    output_dir.mkdir(parents=True, exist_ok=True)
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"add_report_{ts}.txt"

    lines = [f"String Add Report – {ts}", "=" * 60, ""]
    for r in report:
        target = r.get("target_file", "")
        lines.append(f"  {r['string_id']:40s} {r['status']:10s} {target}")
    lines.append("")
    lines.append(f"Total: {len(report)} entries processed")

    report_path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("Report written → %s", report_path)
    return report_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_same_file(a: Path, b: Path) -> bool:
    """Check if two paths point to the same file (handles symlinks, case)."""
    try:
        return os.path.samefile(a, b)
    except OSError:
        return False
