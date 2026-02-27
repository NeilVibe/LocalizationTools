"""Read Excel files – entry extraction, blacklist loading, erase-key loading."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

import config
from .text_utils import convert_linebreaks_for_xml

logger = logging.getLogger(__name__)


def _is_temp_file(path: Path) -> bool:
    """Return True for Excel temp files (``~$`` prefix)."""
    return path.name.startswith("~$")


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _match_header(cell_value: str | None, variants: set[str]) -> bool:
    if cell_value is None:
        return False
    return cell_value.strip().lower() in variants


def detect_headers(ws, *, extra_variants: dict[str, set[str]] | None = None):
    """Detect column indices for known headers in the first row.

    Returns ``{logical_name: col_index}`` where *logical_name* is one of
    ``stringid``, ``strorigin``, ``str``, ``eventname``, or any key from
    *extra_variants*.
    """
    mapping: dict[str, int] = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if header_row is None:
        return mapping

    builtins = {
        "stringid": config.STRINGID_HEADERS,
        "strorigin": config.STRORIGIN_HEADERS,
        "str": config.STR_HEADERS,
        "eventname": config.EVENTNAME_HEADERS,
    }
    if extra_variants:
        builtins.update(extra_variants)

    for cell in header_row:
        val = str(cell.value).strip() if cell.value is not None else ""
        low = val.lower()
        for name, variants in builtins.items():
            if low in variants and name not in mapping:
                mapping[name] = cell.column - 1  # 0-based
                break
    return mapping


# ---------------------------------------------------------------------------
# Entry reading
# ---------------------------------------------------------------------------

def _report_headers(ws, hdrs: dict[str, int], excel_name: str, log_fn) -> None:
    """Log detected vs actual headers so user knows exactly what matched."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    actual = [str(v).strip() for v in header_row if v is not None and str(v).strip()] if header_row else []

    found = []
    if "stringid" in hdrs:
        found.append("StringID")
    if "strorigin" in hdrs:
        found.append("StrOrigin")
    if "str" in hdrs:
        found.append("Str")

    missing = []
    if "stringid" not in hdrs:
        missing.append("StringID")
    if "strorigin" not in hdrs:
        missing.append("StrOrigin")
    if "str" not in hdrs:
        missing.append("Str")

    if found:
        log_fn(f"  {excel_name}: columns found: {', '.join(found)}")
    if missing:
        log_fn(
            f"  {excel_name}: columns NOT found: {', '.join(missing)} "
            f"— actual headers: {actual}",
            "warning",
        )


def read_entries_from_excel(
    excel_path: Path,
    language: str = "EXCEL",
    log_fn=None,
) -> list[dict]:
    """Read LocStr-like entries from an .xlsx file.

    Returns list of dicts with keys: ``string_id``, ``str_origin``,
    ``str_value``, ``raw_attribs``, ``language``, ``source_file``.
    """
    if _is_temp_file(excel_path):
        logger.warning("Skipping Excel temp file: %s", excel_path.name)
        if log_fn:
            log_fn(f"  Skipping temp file: {excel_path.name}", "warning")
        return []

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    hdrs = detect_headers(ws)

    sid_col = hdrs.get("stringid")
    so_col = hdrs.get("strorigin")
    str_col = hdrs.get("str")

    if log_fn:
        _report_headers(ws, hdrs, excel_path.name, log_fn)

    if sid_col is None:
        logger.warning("No StringID column found in %s", excel_path.name)
        if log_fn:
            log_fn(
                f"  {excel_path.name}: SKIPPED — no StringID column. "
                f"Expected one of: {', '.join(sorted(config.STRINGID_HEADERS))}",
                "error",
            )
        wb.close()
        return []

    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_col]).strip() if sid_col is not None and sid_col < len(row) and row[sid_col] else ""
        if not sid:
            continue
        so = str(row[so_col]).strip() if so_col is not None and so_col < len(row) and row[so_col] else ""
        sv = str(row[str_col]).strip() if str_col is not None and str_col < len(row) and row[str_col] else ""

        # Convert Excel newlines (Alt+Enter) to <br/> for XML matching/writing
        so = convert_linebreaks_for_xml(so)
        sv = convert_linebreaks_for_xml(sv)

        entries.append({
            "string_id": sid,
            "str_origin": so,
            "str_value": sv,
            "raw_attribs": {},
            "language": language,
            "source_file": str(excel_path),
        })

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# Blacklist reading
# ---------------------------------------------------------------------------

def read_blacklist_from_excel(
    excel_path: Path,
    valid_codes: set[str],
) -> tuple[dict[str, list[str]], list[str]]:
    """Read a blacklist Excel where columns are language codes.

    Returns ``({LANG: [terms]}, [warnings])``.
    """
    if _is_temp_file(excel_path):
        logger.warning("Skipping Excel temp file: %s", excel_path.name)
        return {}, ["Skipped temp file"]

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    warnings: list[str] = []

    # Detect language columns from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        wb.close()
        return {}, ["Empty workbook"]

    lang_cols: dict[int, str] = {}  # col_idx → LANG
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        code = str(cell_val).strip().upper()
        if code in valid_codes:
            lang_cols[idx] = code
        elif code:
            warnings.append(f"Unknown language column: {cell_val}")

    if not lang_cols:
        wb.close()
        return {}, ["No valid language columns found"]

    result: dict[str, list[str]] = {lang: [] for lang in lang_cols.values()}
    seen: dict[str, set[str]] = {lang: set() for lang in lang_cols.values()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        for col_idx, lang in lang_cols.items():
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            term = str(row[col_idx]).strip()
            if not term:
                continue
            key = term.lower()
            if key not in seen[lang]:
                seen[lang].add(key)
                result[lang].append(term)

    wb.close()
    return result, warnings


# ---------------------------------------------------------------------------
# Erase-key reading
# ---------------------------------------------------------------------------

def read_erase_keys_from_excel(
    excel_path: Path,
) -> tuple[set[tuple[str, str]], set[tuple[str, str]]]:
    """Read StringID + StrOrigin pairs from Excel for string erasing.

    Returns ``(keys, nospace_keys)`` where each key is
    ``(sid_lower, normalized_strorigin)``.
    """
    if _is_temp_file(excel_path):
        logger.warning("Skipping Excel temp file: %s", excel_path.name)
        return set(), set()

    from .text_utils import normalize_text, normalize_nospace

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    hdrs = detect_headers(ws)

    sid_col = hdrs.get("stringid")
    so_col = hdrs.get("strorigin")

    if sid_col is None:
        wb.close()
        return set(), set()

    keys: set[tuple[str, str]] = set()
    nospace_keys: set[tuple[str, str]] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_col]).strip() if sid_col is not None and sid_col < len(row) and row[sid_col] else ""
        if not sid:
            continue
        so = str(row[so_col]).strip() if so_col is not None and so_col < len(row) and row[so_col] else ""
        nt = normalize_text(so)
        keys.add((sid.lower(), nt))
        nospace_keys.add((sid.lower(), normalize_nospace(nt)))

    wb.close()
    return keys, nospace_keys
