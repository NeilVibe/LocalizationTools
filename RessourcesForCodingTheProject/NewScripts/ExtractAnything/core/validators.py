"""Modular folder/file validation for ExtractAnything.

Each tab has different requirements.  Validators return a
:class:`ValidationResult` with status, stats, and messages that the
GUI logs verbatim.

Inspired by QuickTranslate's ``source_scanner.validate_*`` family but
adapted to ExtractAnything's multi-tab design where each operation
accepts different input shapes.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import config
from . import xml_parser
from .language_utils import (
    discover_valid_codes,
    extract_language_suffix,
    get_valid_codes,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Result container
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    """Outcome of a folder/file validation check."""

    ok: bool = True
    languages: list[str] = field(default_factory=list)
    file_count: int = 0
    entry_count: int = 0
    messages: list[tuple[str, str]] = field(default_factory=list)  # (msg, tag)

    # Per-file detail  (filename -> count-or-error)
    file_details: dict[str, int | str] = field(default_factory=dict)

    def add(self, msg: str, tag: str = "info") -> None:
        self.messages.append((msg, tag))

    def log_to(self, log_fn) -> None:
        """Replay all messages through *log_fn(msg, tag)*."""
        for msg, tag in self.messages:
            log_fn(msg, tag)


# ---------------------------------------------------------------------------
# LOC folder
# ---------------------------------------------------------------------------

def validate_loc_folder(loc_folder: Path | None, log_fn=None) -> ValidationResult:
    """Validate a LOC folder: expect ``languagedata_*.xml`` files.

    Reports detected language codes and file count.
    """
    r = ValidationResult()

    if not loc_folder or not loc_folder.is_dir():
        r.ok = False
        r.add("LOC folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    codes = discover_valid_codes(loc_folder)

    if not codes:
        r.ok = False
        r.add(f"LOC: No languagedata_*.xml files found in {loc_folder.name}/", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    r.languages = sorted(codes.keys())
    r.file_count = len(codes)

    r.add(f"LOC: {r.file_count} languages detected: {', '.join(r.languages)}", "success")

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# EXPORT folder
# ---------------------------------------------------------------------------

def validate_export_folder(export_folder: Path | None, log_fn=None) -> ValidationResult:
    """Validate an EXPORT folder: expect ``*.loc.xml`` files."""
    r = ValidationResult()

    if not export_folder or not export_folder.is_dir():
        r.ok = False
        r.add("EXPORT folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    loc_xml = list(export_folder.rglob("*.loc.xml"))
    r.file_count = len(loc_xml)

    if r.file_count == 0:
        r.ok = False
        r.add(f"EXPORT: No .loc.xml files found in {export_folder.name}/", "error")
    else:
        # Count top-level categories
        categories: set[str] = set()
        for f in loc_xml:
            try:
                rel = f.relative_to(export_folder)
                if len(rel.parts) > 1:
                    categories.add(rel.parts[0])
            except ValueError:
                pass
        r.add(
            f"EXPORT: {r.file_count:,} .loc.xml files, "
            f"{len(categories)} categories ({', '.join(sorted(categories)[:8])}"
            f"{'...' if len(categories) > 8 else ''})",
            "success",
        )

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# XML source folder  (languagedata XMLs — Long String, No-Voice, etc.)
# ---------------------------------------------------------------------------

def validate_xml_source_folder(
    folder: Path | None,
    loc_folder: Path | None = None,
    *,
    label: str = "Source",
    log_fn=None,
) -> ValidationResult:
    """Validate a folder of languagedata XMLs/Excel (lightweight — no content parsing).

    Detects language suffixes from filenames, counts files per language.
    Does NOT parse content (that happens during the actual operation).
    This keeps the browse callback fast even for large folders.
    """
    r = ValidationResult()

    if not folder or not folder.is_dir():
        r.ok = False
        r.add(f"{label}: folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    valid_codes = get_valid_codes(loc_folder)

    # Collect XML + Excel files (multi-glob with dedup)
    _is_valid = lambda f: f.is_file() and not f.name.startswith("~$")
    xml_files = [f for f in folder.rglob("*.xml") if _is_valid(f)]
    xlsx_files = [f for f in folder.rglob("*.xlsx") if _is_valid(f)]
    all_files = sorted(set(xml_files + xlsx_files))

    if not all_files:
        r.ok = False
        r.add(f"{label}: No XML or Excel files found in {folder.name}/", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    # Per-language grouping by filename suffix only (fast — no I/O)
    lang_counts: dict[str, int] = {}  # lang -> file count

    for fpath in all_files:
        lang = extract_language_suffix(fpath.name, valid_codes) or "UNKNOWN"
        lang_counts[lang] = lang_counts.get(lang, 0) + 1

    r.file_count = len(all_files)
    r.languages = sorted(k for k in lang_counts if k != "UNKNOWN")

    # Summary line with type breakdown
    type_parts: list[str] = []
    if xml_files:
        type_parts.append(f"{len(xml_files)} XML")
    if xlsx_files:
        type_parts.append(f"{len(xlsx_files)} Excel")
    type_desc = " + ".join(type_parts)

    r.add(
        f"{label}: {type_desc} file{'s' if r.file_count != 1 else ''}, "
        f"{len(r.languages)} language{'s' if len(r.languages) != 1 else ''} detected",
        "success",
    )

    # Per-language breakdown
    for lang in sorted(lang_counts):
        fc = lang_counts[lang]
        r.add(f"  {lang}: {fc} file{'s' if fc != 1 else ''}", "info")

    if "UNKNOWN" in lang_counts:
        r.add(
            f"  Tip: set LOC folder to improve language detection "
            f"({lang_counts['UNKNOWN']} unrecognized)",
            "warning",
        )

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Excel source  (entry-format: StringID, StrOrigin, Str columns)
# ---------------------------------------------------------------------------

def validate_excel_source(
    path: Path | None,
    *,
    label: str = "Source",
    require_strorigin: bool = True,
    log_fn=None,
) -> ValidationResult:
    """Validate an Excel file has the expected columns for entry reading."""
    r = ValidationResult()

    if not path or not path.is_file():
        r.ok = False
        r.add(f"{label}: file does not exist.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    if path.name.startswith("~$"):
        r.ok = False
        r.add(f"{label}: {path.name} is a temp file (close Excel first).", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active

        from .excel_reader import detect_headers
        hdrs = detect_headers(ws)
        wb.close()
    except Exception as exc:
        r.ok = False
        r.add(f"{label}: Cannot read {path.name} — {exc}", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    found = []
    missing = []

    if "stringid" in hdrs:
        found.append("StringID")
    else:
        missing.append("StringID")

    if require_strorigin:
        if "strorigin" in hdrs:
            found.append("StrOrigin")
        else:
            missing.append("StrOrigin")

    if "str" in hdrs:
        found.append("Str")

    if missing:
        r.ok = False
        r.add(f"{label}: Missing required columns: {', '.join(missing)}", "error")
        r.add(f"  Found: {', '.join(found) if found else 'none'}", "info")
    else:
        r.add(f"{label}: Columns OK ({', '.join(found)})", "success")

    r.file_count = 1
    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Diff: Excel column validation per compare mode (BLOCKING)
# ---------------------------------------------------------------------------

# Maps each compare mode to the set of logical column names required.
_DIFF_MODE_COLUMNS: dict[str, set[str]] = {
    "Full (all attributes)":       {"stringid", "strorigin", "str"},
    "StrOrigin + StringID":        {"stringid", "strorigin"},
    "StrOrigin + StringID + Str":  {"stringid", "strorigin", "str"},
    "StringID + Str":              {"stringid", "str"},
    "StrOrigin Diff":              {"stringid", "strorigin"},
    "Str Diff":                    {"stringid", "str"},
}

# Human-readable display names for logical column names.
_COLUMN_DISPLAY: dict[str, str] = {
    "stringid":  "StringID",
    "strorigin": "StrOrigin",
    "str":       "Str",
}


def validate_diff_excel_columns(
    path: Path,
    mode: str,
    *,
    label: str = "File",
    log_fn=None,
) -> ValidationResult:
    """Check that an Excel file has the columns required by *mode*.

    XML files always pass (they carry ``raw_attribs``).
    Returns ``ok=False`` with a clear error when columns are missing.
    """
    r = ValidationResult()

    suffix = path.suffix.lower()
    if suffix == ".xml":
        # XML always has full attributes — skip validation
        r.add(f"{label}: {path.name} — XML (column check skipped)", "info")
        if log_fn:
            r.log_to(log_fn)
        return r

    required = _DIFF_MODE_COLUMNS.get(mode)
    if required is None:
        # Unknown mode — let it through (shouldn't happen)
        if log_fn:
            r.log_to(log_fn)
        return r

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        from .excel_reader import detect_headers
        hdrs = detect_headers(ws)
        wb.close()
    except Exception as exc:
        r.ok = False
        r.add(f"{label}: Cannot read {path.name} — {exc}", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    found = sorted((_COLUMN_DISPLAY[c] for c in required if c in hdrs), key=str.lower)
    missing = sorted((_COLUMN_DISPLAY[c] for c in required if c not in hdrs), key=str.lower)

    if missing:
        r.ok = False
        req_display = ", ".join(sorted((_COLUMN_DISPLAY[c] for c in required), key=str.lower))
        r.add(
            f"{label}: {path.name} — BLOCKED — mode '{mode}' requires columns: {req_display}",
            "error",
        )
        r.add(
            f"  Found: {', '.join(found) if found else 'none'} | Missing: {', '.join(missing)}",
            "info",
        )
    else:
        r.add(
            f"{label}: {path.name} — columns OK for '{mode}' ({', '.join(found)})",
            "success",
        )

    if log_fn:
        r.log_to(log_fn)
    return r


def validate_diff_folder_excel_columns(
    folder: Path,
    mode: str,
    *,
    label: str = "Folder",
    log_fn=None,
) -> ValidationResult:
    """Check ALL Excel files in *folder* have columns required by *mode*.

    Returns ``ok=False`` if ANY Excel file fails, listing all failures.
    If no Excel files exist in the folder, returns ``ok=True``.
    """
    r = ValidationResult()

    xlsx_files = sorted(
        f for f in folder.rglob("*.xlsx")
        if f.is_file() and not f.name.startswith("~$")
    )

    if not xlsx_files:
        # No Excel files — nothing to validate
        r.add(f"{label}: no Excel files to validate", "info")
        if log_fn:
            r.log_to(log_fn)
        return r

    failures: list[str] = []
    for xlsx in xlsx_files:
        sub = validate_diff_excel_columns(xlsx, mode, label=label, log_fn=None)
        r.messages.extend(sub.messages)
        if not sub.ok:
            failures.append(xlsx.name)
            r.ok = False

    if failures:
        r.add(
            f"{label}: {len(failures)} Excel file(s) failed column check for mode '{mode}'",
            "error",
        )

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Excel source folder  (multiple XML/Excel for erase keys)
# ---------------------------------------------------------------------------

def validate_source_keys_folder(
    folder: Path | None,
    *,
    label: str = "Source Keys",
    log_fn=None,
) -> ValidationResult:
    """Validate a folder of XML/Excel files for erase keys (StringID+StrOrigin)."""
    r = ValidationResult()

    if not folder or not folder.is_dir():
        r.ok = False
        r.add(f"{label}: folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    xml_files = [f for f in folder.rglob("*.xml") if f.is_file() and not f.name.startswith("~$")]
    xlsx_files = [f for f in folder.rglob("*.xlsx") if f.is_file() and not f.name.startswith("~$")]

    total = len(xml_files) + len(xlsx_files)
    if total == 0:
        r.ok = False
        r.add(f"{label}: No XML or Excel files found in {folder.name}/", "error")
    else:
        r.file_count = total
        parts = []
        if xml_files:
            parts.append(f"{len(xml_files)} XML")
        if xlsx_files:
            parts.append(f"{len(xlsx_files)} Excel")
        r.add(f"{label}: {' + '.join(parts)} file{'s' if total != 1 else ''} found", "success")

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Blacklist Excel (special format: columns = language codes)
# ---------------------------------------------------------------------------

def validate_blacklist_excel(
    path: Path | None,
    loc_folder: Path | None = None,
    *,
    log_fn=None,
) -> ValidationResult:
    """Validate a blacklist Excel: header row must be language codes."""
    r = ValidationResult()

    if not path or not path.is_file():
        r.ok = False
        r.add("Blacklist: file does not exist.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    if path.name.startswith("~$"):
        r.ok = False
        r.add(f"Blacklist: {path.name} is a temp file (close Excel first).", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    valid_codes = get_valid_codes(loc_folder)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            r.ok = False
            r.add("Blacklist: Empty workbook.", "error")
            wb.close()
            if log_fn:
                r.log_to(log_fn)
            return r

        matched_langs: list[str] = []
        unknown_cols: list[str] = []

        for cell_val in header_row:
            if cell_val is None:
                continue
            code = str(cell_val).strip().upper()
            if not code:
                continue
            if code in valid_codes:
                matched_langs.append(code)
            else:
                unknown_cols.append(str(cell_val).strip())

        # Count data rows
        data_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        wb.close()

    except Exception as exc:
        r.ok = False
        r.add(f"Blacklist: Cannot read {path.name} — {exc}", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    r.languages = sorted(matched_langs)
    r.file_count = 1
    r.entry_count = data_rows

    if not matched_langs:
        r.ok = False
        r.add("Blacklist: No valid language columns found in header row.", "error")
        if unknown_cols:
            r.add(f"  Unknown columns: {', '.join(unknown_cols)}", "warning")
        r.add("  Expected: column headers matching language codes (ENG, FRE, KR, ...)", "info")
    else:
        r.add(
            f"Blacklist: {len(matched_langs)} language{'s' if len(matched_langs) != 1 else ''} "
            f"({', '.join(r.languages)}), {data_rows} data row{'s' if data_rows != 1 else ''}",
            "success",
        )
        if unknown_cols:
            r.add(f"  Ignored columns: {', '.join(unknown_cols)}", "warning")

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Blacklist folder (multiple Excel files)
# ---------------------------------------------------------------------------

def validate_blacklist_folder(
    folder: Path | None,
    loc_folder: Path | None = None,
    *,
    log_fn=None,
) -> ValidationResult:
    """Validate a folder of blacklist Excel files."""
    r = ValidationResult()

    if not folder or not folder.is_dir():
        r.ok = False
        r.add("Blacklist folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    xlsx_files = sorted(f for f in folder.rglob("*.xlsx") if f.is_file() and not f.name.startswith("~$"))
    if not xlsx_files:
        r.ok = False
        r.add(f"Blacklist: No .xlsx files found in {folder.name}/", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    all_langs: set[str] = set()
    total_entries = 0
    errors: list[str] = []
    for xlsx in xlsx_files:
        sub = validate_blacklist_excel(xlsx, loc_folder)
        all_langs.update(sub.languages)
        total_entries += sub.entry_count
        r.file_count += 1
        if not sub.ok:
            errors.append(xlsx.name)
            r.ok = False

    r.languages = sorted(all_langs)
    r.entry_count = total_entries
    r.add(
        f"Blacklist folder: {r.file_count} Excel file{'s' if r.file_count != 1 else ''}, "
        f"{len(r.languages)} language{'s' if len(r.languages) != 1 else ''} "
        f"({', '.join(r.languages)}), {total_entries:,} data rows",
        "success" if r.languages and not errors else "warning",
    )
    for fname in errors:
        r.add(f"  {fname}: validation failed", "error")

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Generic folder  (File Erase — any files)
# ---------------------------------------------------------------------------

def validate_generic_folder(
    folder: Path | None,
    *,
    label: str = "Folder",
    log_fn=None,
) -> ValidationResult:
    """Validate a folder exists and has files."""
    r = ValidationResult()

    if not folder or not folder.is_dir():
        r.ok = False
        r.add(f"{label}: folder does not exist or is not a directory.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    file_count = sum(1 for f in folder.rglob("*") if f.is_file())
    r.file_count = file_count

    if file_count == 0:
        r.ok = False
        r.add(f"{label}: folder is empty.", "warning")
    else:
        r.add(f"{label}: {file_count:,} file{'s' if file_count != 1 else ''}", "success")

    if log_fn:
        r.log_to(log_fn)
    return r


# ---------------------------------------------------------------------------
# Diff: file validation (XML or Excel)
# ---------------------------------------------------------------------------

def validate_diff_file(
    path: Path | None,
    *,
    label: str = "File",
    log_fn=None,
) -> ValidationResult:
    """Validate a single file for diff (XML or Excel with LocStr entries)."""
    r = ValidationResult()

    if not path or not path.is_file():
        r.ok = False
        r.add(f"{label}: file does not exist.", "error")
        if log_fn:
            r.log_to(log_fn)
        return r

    suffix = path.suffix.lower()
    if suffix == ".xml":
        try:
            raw = xml_parser.read_xml_raw(path)
            if raw is None:
                r.ok = False
                r.add(f"{label}: Cannot read {path.name}", "error")
            else:
                root = xml_parser.parse_root_from_string(raw)
                count = sum(1 for _ in xml_parser.iter_locstr(root))
                r.entry_count = count
                if count > 0:
                    r.add(f"{label}: {path.name} — {count:,} LocStr entries", "success")
                else:
                    r.ok = False
                    r.add(f"{label}: {path.name} — 0 LocStr entries", "warning")
        except Exception as exc:
            r.ok = False
            r.add(f"{label}: {path.name} parse error — {exc}", "error")

    elif suffix in (".xlsx", ".xls"):
        sub = validate_excel_source(path, label=label, require_strorigin=False, log_fn=None)
        r.ok = sub.ok
        r.messages = sub.messages

    else:
        r.ok = False
        r.add(f"{label}: Unsupported file type ({suffix})", "error")

    r.file_count = 1
    if log_fn:
        r.log_to(log_fn)
    return r
