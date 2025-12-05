"""
Script Name: cell_parser.py
Created: 2025-12-05
Purpose: Parse cells with <<header::value>> format and expand into columns

Input: Excel file with cells containing <<header::value>><<header2::value2>>...
Output: Excel file with headers as columns, values expanded, nice formatting

Usage:
    python cell_parser.py

    - File dialog opens automatically
    - Select Excel file
    - Output saved as [filename]_parsed.xlsx

Dependencies:
    pip install openpyxl
"""

import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Missing Dependency", "openpyxl not installed!\n\nRun: pip install openpyxl")
    sys.exit(1)


def parse_cell(cell_value, row_num, col_num):
    """Parse <<header::value>> patterns from cell text, return dict and warnings"""
    warnings = []

    if not cell_value:
        return {}, warnings

    cell_str = str(cell_value)

    # Find all <<...>> blocks
    blocks = re.findall(r'<<[^>]*>>', cell_str)

    result = {}

    for block in blocks:
        # Check if block has proper format <<header::value>>
        match = re.match(r'<<([^:]*)(::)?([^>]*)>>', block)

        if not match:
            warnings.append(f"[Row {row_num}, Col {col_num}] Malformed block: {block}")
            continue

        header = match.group(1).strip() if match.group(1) else ""
        separator = match.group(2)
        value = match.group(3).strip() if match.group(3) else ""

        if not header:
            warnings.append(f"[Row {row_num}, Col {col_num}] Empty header in block: {block}")
            continue

        if not separator:
            warnings.append(f"[Row {row_num}, Col {col_num}] Missing '::' separator in block: {block}")
            continue

        if not value:
            warnings.append(f"[Row {row_num}, Col {col_num}] Empty value for header '{header}'")

        result[header] = value

    # Check for potential malformed patterns (unclosed or weird stuff)
    unclosed = re.findall(r'<<[^>]*(?:$|(?!>>))', cell_str)
    for u in unclosed:
        if u not in [b.rstrip('>') for b in blocks]:
            warnings.append(f"[Row {row_num}, Col {col_num}] Possible unclosed block: {u[:30]}...")

    return result, warnings


def get_headers_from_first_row(ws):
    """Extract headers from first row to determine initial column order"""
    first_cell = ws.cell(row=1, column=1).value

    if not first_cell:
        return []

    pattern = r'<<([^:]+)::'
    matches = re.findall(pattern, str(first_cell))

    return [h.strip() for h in matches]


def apply_styles(ws, num_headers):
    """Apply nice formatting to the worksheet"""
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")

    header_font = Font(bold=True, size=11)

    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    for col in range(1, num_headers + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = green_fill
        cell.font = header_font
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, num_headers + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if row % 2 == 0:
                cell.fill = orange_fill

    # Column widths
    for col in range(1, num_headers + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)

        for row in range(1, min(ws.max_row + 1, 50)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 15)


def process_file(input_file):
    """Process the Excel file and return result info"""

    print(f"\n{'='*60}")
    print(f"CELL PARSER - Processing")
    print(f"{'='*60}")
    print(f"Input: {input_file}")
    print()

    # Load workbook
    print("[1/6] Loading workbook...")
    try:
        wb_in = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"  ✗ Failed to open file: {e}")
        return False, f"Cannot open file:\n{e}"

    ws_in = wb_in.active
    print(f"  ✓ Loaded sheet '{ws_in.title}' ({ws_in.max_row} rows, {ws_in.max_column} cols)")

    # Check if file has data
    if ws_in.max_row < 1:
        print("  ✗ File is empty!")
        return False, "File is empty!"

    # Get headers from first row
    print("\n[2/6] Extracting headers from first row...")
    headers = get_headers_from_first_row(ws_in)

    if not headers:
        print("  ⚠ No headers in first row, scanning for patterns...")
        for row in range(1, min(ws_in.max_row + 1, 10)):
            for col in range(1, ws_in.max_column + 1):
                cell_val = ws_in.cell(row=row, column=col).value
                if cell_val and '<<' in str(cell_val) and '::' in str(cell_val):
                    pattern = r'<<([^:]+)::'
                    matches = re.findall(pattern, str(cell_val))
                    if matches:
                        headers = [h.strip() for h in matches]
                        print(f"  ✓ Found headers in row {row}, col {col}")
                        break
            if headers:
                break

    if not headers:
        print("  ✗ No <<header::value>> pattern found!")
        return False, "No <<header::value>> pattern found!\n\nExpected format:\n<<Korean::안녕>><<English::Hello>>"

    print(f"  ✓ Initial headers ({len(headers)}): {headers}")

    # Create header to column mapping (mutable - can add new headers)
    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}
    next_col = len(headers) + 1
    new_headers_added = []

    # Create output workbook
    print("\n[3/6] Creating output workbook...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Parsed"

    # Write initial headers
    for header, col in header_to_col.items():
        ws_out.cell(row=1, column=col, value=header)
    print(f"  ✓ Created output sheet")

    # Process rows
    print("\n[4/6] Processing cells...")
    out_row = 2
    rows_processed = 0
    all_warnings = []

    total_cells = ws_in.max_row * ws_in.max_column
    cells_checked = 0
    last_progress = 0

    for row in range(1, ws_in.max_row + 1):
        for col in range(1, ws_in.max_column + 1):
            cells_checked += 1

            # Progress update every 10%
            progress = int((cells_checked / total_cells) * 100)
            if progress >= last_progress + 10:
                print(f"  ... {progress}% ({cells_checked}/{total_cells} cells)")
                last_progress = progress

            cell_value = ws_in.cell(row=row, column=col).value

            if not cell_value:
                continue

            parsed, warnings = parse_cell(cell_value, row, col)
            all_warnings.extend(warnings)

            if not parsed:
                continue

            # Write values, adding new headers if needed
            for header, value in parsed.items():
                if header not in header_to_col:
                    # New header found - add to far right
                    header_to_col[header] = next_col
                    ws_out.cell(row=1, column=next_col, value=header)
                    new_headers_added.append(header)
                    print(f"  + New header found: '{header}' -> Column {next_col}")
                    next_col += 1

                target_col = header_to_col[header]
                ws_out.cell(row=out_row, column=target_col, value=value)

            out_row += 1
            rows_processed += 1

    print(f"  ✓ Processed {rows_processed} data rows")

    if new_headers_added:
        print(f"  ✓ Added {len(new_headers_added)} new headers: {new_headers_added}")

    if rows_processed == 0:
        print("  ✗ No data rows found!")
        return False, "No data rows found with <<header::value>> pattern!"

    # Print warnings
    if all_warnings:
        print(f"\n[!] Warnings ({len(all_warnings)}):")
        for w in all_warnings[:20]:  # Show first 20
            print(f"  ⚠ {w}")
        if len(all_warnings) > 20:
            print(f"  ... and {len(all_warnings) - 20} more warnings")

    # Apply styles
    print("\n[5/6] Applying formatting...")
    total_headers = len(header_to_col)
    apply_styles(ws_out, total_headers)
    print(f"  ✓ Styled {total_headers} columns, {out_row - 1} rows")

    # Save output
    print("\n[6/6] Saving output...")
    output_file = Path(input_file).stem + "_parsed.xlsx"
    output_path = Path(input_file).parent / output_file

    try:
        wb_out.save(output_path)
    except Exception as e:
        print(f"  ✗ Failed to save: {e}")
        return False, f"Cannot save output:\n{e}"

    print(f"  ✓ Saved to: {output_path}")

    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"  Headers: {total_headers} ({len(headers)} initial + {len(new_headers_added)} new)")
    print(f"  Rows: {rows_processed}")
    print(f"  Warnings: {len(all_warnings)}")
    print(f"  Output: {output_path}")
    print(f"{'='*60}\n")

    return True, f"Success!\n\nHeaders: {total_headers}\nRows: {rows_processed}\nWarnings: {len(all_warnings)}\n\nSaved to:\n{output_path}"


def main():
    """Main - launches file dialog immediately"""

    print("\n" + "="*60)
    print("CELL PARSER")
    print("Parse <<header::value>> cells into columns")
    print("="*60)

    # Setup tkinter
    root = tk.Tk()
    root.withdraw()

    # Show file dialog
    print("\nOpening file dialog...")
    input_file = filedialog.askopenfilename(
        title="Select Excel file with <<header::value>> cells",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    if not input_file:
        print("No file selected. Exiting.")
        sys.exit(0)

    # Process
    success, message = process_file(input_file)

    # Show result
    if success:
        messagebox.showinfo("Done", message)
    else:
        messagebox.showerror("Error", message)


if __name__ == "__main__":
    main()
