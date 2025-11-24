"""
Script Name: glossary_sniffer_1124.py
Created: 2025-11-24
Updated: 2025-11-24 (Multi-language support added)
Purpose: Extract glossary terms from LANGUAGE DATA FOLDER and find them in Excel text lines
Input:
    1. LANGUAGE DATA FOLDER with languagedata_*.xml files (13 languages)
    2. Excel file with lines to analyze (Korean text)
Output: Excel file with original lines + glossary terms + translations in ALL languages
Reference: QuickSearch0818.py (Aho-Corasick), wordcount1.py (folder walking)

Usage:
    python glossary_sniffer_1124.py

    1. Run the script
    2. Select LANGUAGE DATA FOLDER (with languagedata_KOR.xml, languagedata_ENG.xml, etc.)
    3. Script extracts glossary from KOR file and maps to all language translations
    4. Select Excel file (lines to analyze)
    5. Script searches for glossary terms and maps to all languages
    6. Select output location
    7. Results saved with 2 + N language columns:
       - Column 1: Original Line (Korean)
       - Column 2: Glossary Terms Found (StrOrigin from KOR)
       - Columns 3+: One column per language (ENG, FRA, GER, SPA, ITA, POR, RUS, POL, TUR, THA, JPN, CHS, CHT)

Dependencies:
    pip install openpyxl lxml pyahocorasick

Author: Generated with Claude Code
"""

import os
import re
import string
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import openpyxl
from lxml import etree
import ahocorasick

# Configuration - HARDCODED RULES (based on QuickSearch0818)
DEFAULT_LENGTH_THRESHOLD = 15  # Max chars for glossary term (Korean names/terms)
MIN_OCCURRENCE = 2  # Term must appear at least this many times to be glossary
FILTER_SENTENCES = True  # Remove entries ending with punctuation (.?!)
FILTER_PUNCTUATION = True  # Remove entries containing ANY punctuation (including …)
WORD_BOUNDARIES = True  # Only match complete words (not inside other words)

def iter_language_files(folder: Path):
    """
    Walk folder and yield all languagedata_*.xml files.
    Pattern from wordcount1.py lines 43-47.

    Args:
        folder: Path to language data folder

    Yields:
        Path: Path to each languagedata_*.xml file found
    """
    for dirpath, _, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().startswith("languagedata_") and fn.lower().endswith(".xml"):
                yield Path(dirpath) / fn

def extract_language_code(xml_path: Path) -> str:
    """
    Extract language code from filename.
    Example: "languagedata_ENG.xml" → "ENG"
    Pattern from wordcount1.py lines 135-139.

    Args:
        xml_path: Path to XML file

    Returns:
        str: Language code (e.g., "ENG", "FRA", "KOR")
    """
    stem = xml_path.stem
    parts = stem.split("_", 1)
    if len(parts) == 2:
        return parts[1].upper()
    return "UNKNOWN"

def extract_multilanguage_glossary(folder_path, length_threshold=DEFAULT_LENGTH_THRESHOLD, min_occurrence=MIN_OCCURRENCE):
    """
    Extract glossary from LANGUAGE DATA FOLDER (multi-language support).

    HARDCODED FILTERING RULES (from QuickSearch0818):
    1. Length < 15 characters (Korean names/terms are usually short)
    2. No punctuation endings (.?!)
    3. No punctuation inside term (except spaces/hyphens for multi-word)
    4. Must appear >= 2 times (min_occurrence)
    5. Non-empty strings only

    Args:
        folder_path: Path to language data folder
        length_threshold: Maximum character length for glossary terms
        min_occurrence: Minimum times a term must appear to be considered glossary

    Returns:
        tuple: (glossary_terms: list, multi_lang_map: dict, language_codes: list)
            - glossary_terms: List of StrOrigin values (from KOR)
            - multi_lang_map: Dict mapping StrOrigin → {lang_code: translation}
            - language_codes: List of language codes found (sorted)

    Example multi_lang_map:
    {
        '클리프': {'ENG': 'Kliff', 'FRA': 'Kliff', 'GER': 'Kliff', ...},
        '칼파데': {'ENG': 'Calphade', 'FRA': 'Calphade', ...}
    }
    """
    print(f"\n📖 Extracting glossary from folder: {Path(folder_path).name}")

    try:
        # 1. Find all languagedata_*.xml files
        xml_files = list(iter_language_files(Path(folder_path)))
        print(f"   Found {len(xml_files)} language files")

        if not xml_files:
            raise ValueError("No languagedata_*.xml files found in folder!")

        # 2. Group by language code
        files_by_lang = {}
        for xml_path in xml_files:
            lang_code = extract_language_code(xml_path)
            files_by_lang[lang_code] = xml_path
            print(f"      - {lang_code}: {xml_path.name}")

        # 3. Extract StrOrigin from KOR file (reference)
        if 'KOR' not in files_by_lang:
            raise ValueError("No Korean (KOR) language file found! Need languagedata_KOR.xml")

        kor_path = files_by_lang['KOR']
        print(f"\n   Using KOR file as reference: {kor_path.name}")
        tree = etree.parse(kor_path)

        # Build StrOrigin list (for filtering)
        all_terms = []
        for locstr in tree.xpath('//LocStr'):
            str_origin = locstr.get('StrOrigin', '').strip()
            if str_origin:
                all_terms.append(str_origin)

        print(f"   Found {len(all_terms):,} total StrOrigin entries from KOR")

        # Filter glossary (same rules as before)
        glossary_terms = filter_glossary_terms(all_terms, length_threshold, min_occurrence)

        print(f"   ✅ Filtered to {len(glossary_terms):,} glossary terms (min {min_occurrence} occurrences)")

        # 4. Build multi-language mapping
        multi_lang_map = {term: {} for term in glossary_terms}

        print(f"\n   Building multi-language mapping...")
        for lang_code, xml_path in sorted(files_by_lang.items()):
            print(f"      Processing {lang_code}...", end=" ")
            tree = etree.parse(xml_path)

            mapped_count = 0
            for locstr in tree.xpath('//LocStr'):
                str_origin = locstr.get('StrOrigin', '').strip()
                str_value = locstr.get('Str', '').strip()

                # Only include terms that passed filtering
                if str_origin in multi_lang_map:
                    multi_lang_map[str_origin][lang_code] = str_value
                    mapped_count += 1

            print(f"{mapped_count:,} terms mapped")

        # Get sorted language codes (excluding KOR, put it first)
        language_codes = sorted([lang for lang in files_by_lang.keys() if lang != 'KOR'])
        language_codes = ['KOR'] + language_codes  # KOR first

        print(f"\n   ✅ Multi-language glossary built: {len(glossary_terms):,} terms × {len(language_codes)} languages")

        return glossary_terms, multi_lang_map, language_codes

    except Exception as e:
        print(f"   ❌ Error extracting glossary: {e}")
        raise

def extract_glossary_from_xml(xml_path, length_threshold=DEFAULT_LENGTH_THRESHOLD, min_occurrence=MIN_OCCURRENCE):
    """
    Extract glossary terms from XML StrOrigin attributes + build mapping to Str values.

    HARDCODED FILTERING RULES (from QuickSearch0818):
    1. Length < 15 characters (Korean names/terms are usually short)
    2. No punctuation endings (.?!)
    3. No punctuation inside term (except spaces/hyphens for multi-word)
    4. Must appear >= 2 times (min_occurrence)
    5. Non-empty strings only

    Args:
        xml_path: Path to XML file
        length_threshold: Maximum character length for glossary terms
        min_occurrence: Minimum times a term must appear to be considered glossary

    Returns:
        tuple: (glossary_terms: list, glossary_map: dict)
            - glossary_terms: List of StrOrigin values for Aho-Corasick
            - glossary_map: Dict mapping StrOrigin → Str values (Korean → English)
    """
    print(f"\n📖 Extracting glossary from: {Path(xml_path).name}")

    try:
        tree = etree.parse(xml_path)
        all_terms = []
        term_to_str_map = {}  # NEW: Store mapping StrOrigin → Str

        # Extract all StrOrigin attributes AND Str values from LocStr elements
        for locstr in tree.xpath('//LocStr'):
            str_origin = locstr.get('StrOrigin', '').strip()
            str_value = locstr.get('Str', '').strip()

            if str_origin:
                all_terms.append(str_origin)
                term_to_str_map[str_origin] = str_value  # Map Korean → English

        print(f"   Found {len(all_terms):,} total StrOrigin entries")

        # Filter to keep only glossary-worthy terms (with occurrence counting)
        glossary_terms = filter_glossary_terms(all_terms, length_threshold, min_occurrence)

        # Build final mapping (only for terms that passed filtering)
        glossary_map = {term: term_to_str_map[term] for term in glossary_terms}

        print(f"   ✅ Filtered to {len(glossary_terms):,} glossary terms (min {min_occurrence} occurrences)")
        return glossary_terms, glossary_map  # NEW: Return both

    except Exception as e:
        print(f"   ❌ Error parsing XML: {e}")
        raise

def filter_glossary_terms(terms, length_threshold, min_occurrence):
    """
    Filter terms to keep only glossary-worthy entries.

    HARDCODED FILTERING RULES (from QuickSearch0818 lines 2113-2149):
    1. Length < threshold (15 chars for Korean)
    2. Non-empty strings
    3. Does NOT end with punctuation (.?!)
    4. Does NOT contain ANY punctuation (including special ellipsis …)
    5. Must appear >= min_occurrence times (2+)

    Two-pass filtering (like QuickSearch0818):
    - Pass 1: Basic filtering + count occurrences
    - Pass 2: Apply min_occurrence filter

    Args:
        terms: List of terms to filter
        length_threshold: Maximum character length
        min_occurrence: Minimum times term must appear

    Returns:
        list: Filtered unique terms (sorted by frequency, most common first)
    """
    # Pass 1: Basic filtering + count occurrences
    count_map = {}

    for term in terms:
        # Skip empty
        if not term:
            continue

        # Skip if too long
        if len(term) >= length_threshold:
            continue

        # Skip if ends with punctuation (likely a sentence)
        if FILTER_SENTENCES and re.search(r'[.?!]\s*$', term.strip()):
            continue

        # Skip if contains ANY punctuation (except spaces and hyphens for multi-word terms)
        # Based on QuickSearch0818 pattern: line 2140
        if FILTER_PUNCTUATION:
            # Allow only alphanumeric, spaces, hyphens
            # Filter out other punctuation including special ellipsis …
            if any(ch in string.punctuation.replace('-', '').replace(' ', '') for ch in term) or '…' in term:
                continue

        # Count occurrences
        count_map[term] = count_map.get(term, 0) + 1

    # Pass 2: Apply min_occurrence filter
    filtered = [term for term, count in count_map.items() if count >= min_occurrence]

    # Sort by frequency (most common first)
    filtered.sort(key=lambda t: count_map[t], reverse=True)

    return filtered

def build_ahocorasick_automaton(glossary_terms):
    """
    Build Aho-Corasick automaton from glossary terms for fast multi-pattern matching.

    Pattern from QuickSearch0818.py lines 2217-2224

    Args:
        glossary_terms: List of terms to search for

    Returns:
        Aho-Corasick automaton object
    """
    print(f"\n🔧 Building Aho-Corasick automaton with {len(glossary_terms):,} terms...")

    automaton = ahocorasick.Automaton()

    for term in glossary_terms:
        # Add term to automaton (key=term, value=term)
        automaton.add_word(term, term)

    # Finalize the automaton
    automaton.make_automaton()

    print(f"   ✅ Automaton built successfully")
    return automaton

def search_line_for_glossary(line, automaton):
    """
    Search a single line for glossary terms using Aho-Corasick + word boundaries.

    Pattern from QuickSearch0818.py lines 2250-2260
    + Word boundary validation (only match complete words, not substrings)

    Args:
        line: Text line to search
        automaton: Aho-Corasick automaton

    Returns:
        list: Unique glossary terms found (sorted by position for consistent output)
    """
    if not line:
        return []

    matches = []
    seen = set()

    # Use Aho-Corasick to find all matches
    for end_index, term in automaton.iter(line):
        if term in seen:
            continue

        # Word boundary validation (if enabled)
        if WORD_BOUNDARIES:
            # Check if match is at word boundaries
            # Find start position
            start_index = end_index - len(term) + 1

            # Check character before match (if exists)
            if start_index > 0:
                char_before = line[start_index - 1]
                # If alphanumeric before, it's inside a word - skip
                if char_before.isalnum():
                    continue

            # Check character after match (if exists)
            if end_index + 1 < len(line):
                char_after = line[end_index + 1]
                # If alphanumeric after, it's inside a word - skip
                if char_after.isalnum():
                    continue

        matches.append(term)
        seen.add(term)

    return matches

def resolve_overlapping_matches(matches, line):
    """
    For overlapping matches, prefer longest match.

    Example: If both "Duke" and "Duke Elenor" match, keep only "Duke Elenor"

    Args:
        matches: List of matched terms
        line: Original line text

    Returns:
        list: Filtered matches with overlaps resolved
    """
    if len(matches) <= 1:
        return matches

    # Sort by length (longest first)
    sorted_matches = sorted(matches, key=len, reverse=True)

    final_matches = []
    for match in sorted_matches:
        # Check if this match is a substring of any already accepted match
        is_substring = False
        for accepted in final_matches:
            if match in accepted and match != accepted:
                is_substring = True
                break

        if not is_substring:
            final_matches.append(match)

    # Sort by appearance order in line for consistent output
    final_matches.sort(key=lambda m: line.find(m))

    return final_matches

def process_excel_lines(excel_path, automaton, multi_lang_map, language_codes):
    """
    Read Excel, search each line for glossary terms, map to all language translations.

    Args:
        excel_path: Path to input Excel file
        automaton: Aho-Corasick automaton
        multi_lang_map: Dict mapping StrOrigin → {lang_code: translation}
        language_codes: List of language codes (e.g., ['KOR', 'ENG', 'FRA', ...])

    Returns:
        list: Tuples of (original_line, glossary_terms_found, translations_by_lang)
            - translations_by_lang: Dict mapping lang_code → list of translations
    """
    print(f"\n📊 Processing Excel file: {Path(excel_path).name}")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active

        results = []
        total_matches = 0

        # Process each row (skip header if exists)
        for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Get first column value (line to analyze)
            line = str(row[0]) if row[0] else ""

            if not line or line == "None":
                # Empty line: empty matches and empty translations for all languages
                empty_translations = {lang: [] for lang in language_codes}
                results.append((line, [], empty_translations))
                continue

            # Search for glossary terms
            matches = search_line_for_glossary(line, automaton)

            # Resolve overlapping matches (prefer longest)
            matches = resolve_overlapping_matches(matches, line)

            # NEW: Build translations for ALL languages
            translations_by_lang = {}
            for lang_code in language_codes:
                lang_translations = [
                    multi_lang_map.get(term, {}).get(lang_code, '')
                    for term in matches
                ]
                translations_by_lang[lang_code] = lang_translations

            results.append((line, matches, translations_by_lang))

            if matches:
                total_matches += 1

            # Progress indicator
            if idx % 100 == 0:
                print(f"   Processed {idx:,} lines ({total_matches:,} with matches)...")

        print(f"   ✅ Processed {len(results):,} total lines")
        print(f"   ✅ Found glossary terms in {total_matches:,} lines ({total_matches/len(results)*100:.1f}%)")

        wb.close()
        return results

    except Exception as e:
        print(f"   ❌ Error reading Excel: {e}")
        raise

def write_results_to_excel(results, output_path, language_codes):
    """
    Write results to Excel with multiple columns (2 + N languages):
    - Column 1: Original Line
    - Column 2: Glossary Terms Found (StrOrigin, comma-separated)
    - Columns 3+: One column per language (comma-separated translations)

    Args:
        results: List of (original_line, glossary_terms, translations_by_lang) tuples
        output_path: Path to save output Excel
        language_codes: List of language codes (e.g., ['KOR', 'ENG', 'FRA', ...])
    """
    print(f"\n💾 Saving results to: {Path(output_path).name}")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Glossary Analysis"

        # Header row: Original Line + Glossary Terms + one column per language
        headers = ["Original Line", "Glossary Terms (StrOrigin)"] + language_codes
        ws.append(headers)

        # Format header (bold)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for line, matches, translations_by_lang in results:
            # Column 1: Original Line
            # Column 2: Glossary Terms Found (comma-separated)
            glossary_str = ", ".join(matches) if matches else ""

            row_data = [line, glossary_str]

            # Columns 3+: One column per language (comma-separated translations)
            for lang_code in language_codes:
                lang_translations = translations_by_lang.get(lang_code, [])
                translation_str = ", ".join(lang_translations) if lang_translations else ""
                row_data.append(translation_str)

            ws.append(row_data)

        # Auto-size columns
        ws.column_dimensions['A'].width = 80  # Original Line
        ws.column_dimensions['B'].width = 50  # Glossary Terms

        # Auto-size language columns
        for i, lang_code in enumerate(language_codes, start=3):
            col_letter = chr(64 + i)  # C, D, E, F, G, ...
            ws.column_dimensions[col_letter].width = 40

        wb.save(output_path)
        print(f"   ✅ Saved {len(results):,} rows with {len(language_codes)} language columns")

    except Exception as e:
        print(f"   ❌ Error saving Excel: {e}")
        raise

def main():
    """
    Main execution (MULTI-LANGUAGE):
    1. Pick LANGUAGE DATA FOLDER (with languagedata_*.xml files)
    2. Build multi-language glossary with smart filtering
    3. Build Aho-Corasick automaton
    4. Pick Excel input file (lines to analyze)
    5. Search for glossary terms in each line
    6. Map to ALL language translations
    7. Save results to Excel (2 + N language columns)
    """
    print("="*70)
    print("🔍 GlossarySniffer - Glossary Term Extraction & Search")
    print("="*70)

    # Tkinter setup
    root = tk.Tk()
    root.withdraw()

    try:
        # Step 1: Select LANGUAGE DATA FOLDER
        print("\n📂 Step 1: Select Language Data Folder (with languagedata_*.xml files)")
        folder_path = filedialog.askdirectory(
            title="Select Language Data Folder"
        )

        if not folder_path:
            print("❌ No folder selected. Exiting.")
            return

        # Step 2: Extract multi-language glossary + build mapping
        glossary_terms, multi_lang_map, language_codes = extract_multilanguage_glossary(
            folder_path,
            DEFAULT_LENGTH_THRESHOLD
        )

        if not glossary_terms:
            messagebox.showerror("Error", "No glossary terms found in language data folder!")
            return

        # Show some sample terms with translations
        print(f"\n   Sample glossary terms across languages:")
        for term in glossary_terms[:5]:
            translations = []
            for lang in language_codes[:5]:  # Show first 5 languages
                trans = multi_lang_map.get(term, {}).get(lang, '')
                translations.append(f"{lang}:{trans}")
            print(f"      - {term} → [{', '.join(translations)}]")
        if len(glossary_terms) > 5:
            print(f"      ... and {len(glossary_terms)-5} more")

        # Step 3: Build Aho-Corasick automaton
        automaton = build_ahocorasick_automaton(glossary_terms)

        # Step 4: Select Excel input file
        print("\n📂 Step 2: Select Excel input file (lines to analyze)")
        excel_path = filedialog.askopenfilename(
            title="Select Excel File (Lines to Analyze)",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )

        if not excel_path:
            print("❌ No Excel file selected. Exiting.")
            return

        # Step 5: Process Excel and search for glossary terms
        results = process_excel_lines(excel_path, automaton, multi_lang_map, language_codes)

        # Step 6: Select output location
        print("\n📂 Step 3: Select output location")
        output_path = filedialog.asksaveasfilename(
            title="Save Results As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{Path(excel_path).stem}_glossary_analysis.xlsx"
        )

        if not output_path:
            print("❌ No output location selected. Exiting.")
            return

        # Step 7: Write results
        write_results_to_excel(results, output_path, language_codes)

        # Success!
        print("\n" + "="*70)
        print("✅ SUCCESS!")
        print("="*70)
        print(f"📊 Glossary Terms: {len(glossary_terms):,}")
        print(f"🌍 Languages: {len(language_codes)} ({', '.join(language_codes)})")
        print(f"📄 Lines Processed: {len(results):,}")
        print(f"💾 Output: {output_path}")
        print("="*70)

        messagebox.showinfo(
            "Success",
            f"Multi-language glossary analysis complete!\n\n"
            f"Glossary Terms: {len(glossary_terms):,}\n"
            f"Languages: {len(language_codes)} ({', '.join(language_codes[:5])}...)\n"
            f"Lines Processed: {len(results):,}\n\n"
            f"Results saved to:\n{output_path}"
        )

    except Exception as e:
        error_msg = f"Error: {str(e)}"
        print(f"\n❌ {error_msg}")
        messagebox.showerror("Error", error_msg)
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
