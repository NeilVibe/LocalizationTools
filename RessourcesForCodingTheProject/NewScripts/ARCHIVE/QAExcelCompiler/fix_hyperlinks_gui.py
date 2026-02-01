#!/usr/bin/env python3
"""
GUI tool to fix SCREENSHOT hyperlinks in Excel files.

Select a folder → finds all .xlsx files → fixes hyperlinks → saves in place.
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path
from openpyxl import load_workbook


def find_screenshot_column(ws):
    """Find SCREENSHOT column by header name."""
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header and "SCREENSHOT" in str(header).upper():
            return col
    return None


def find_file_in_folder(filename, folder):
    """Search for file in folder (case-insensitive)."""
    filename_lower = filename.lower()
    for f in os.listdir(folder):
        if f.lower() == filename_lower:
            return f
    return None


def fix_excel_file(excel_path, log_func):
    """Fix hyperlinks in one Excel file. Returns (fixed, missing) counts."""
    folder = excel_path.parent
    total_fixed = 0
    total_missing = 0

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        log_func(f"  ERROR loading: {e}")
        return 0, 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        screenshot_col = find_screenshot_column(ws)

        if not screenshot_col:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, screenshot_col)
            value = cell.value

            if not value or cell.hyperlink:
                continue

            filename = str(value).strip()
            actual_filename = find_file_in_folder(filename, folder)

            if actual_filename:
                cell.hyperlink = actual_filename
                total_fixed += 1
            else:
                total_missing += 1
                log_func(f"    [MISSING] {sheet_name} row {row}: {filename}")

    if total_fixed > 0:
        wb.save(excel_path)

    return total_fixed, total_missing


class App:
    def __init__(self, root):
        self.root = root
        root.title("Fix Screenshot Hyperlinks")
        root.geometry("600x400")

        # Frame for button
        btn_frame = tk.Frame(root, pady=10)
        btn_frame.pack(fill=tk.X)

        self.select_btn = tk.Button(
            btn_frame,
            text="Select Folder & Fix Hyperlinks",
            command=self.select_and_fix,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            padx=20,
            pady=10
        )
        self.select_btn.pack()

        # Log area
        self.log = scrolledtext.ScrolledText(root, height=20, font=("Consolas", 10))
        self.log.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    def log_msg(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.root.update()

    def select_and_fix(self):
        folder = filedialog.askdirectory(title="Select folder with Excel files")
        if not folder:
            return

        self.log.delete(1.0, tk.END)
        self.log_msg(f"Folder: {folder}")
        self.log_msg("-" * 50)

        folder_path = Path(folder)
        xlsx_files = list(folder_path.glob("*.xlsx"))

        if not xlsx_files:
            self.log_msg("No .xlsx files found in this folder.")
            return

        self.log_msg(f"Found {len(xlsx_files)} Excel file(s)\n")

        grand_fixed = 0
        grand_missing = 0
        files_modified = 0

        for xlsx in xlsx_files:
            self.log_msg(f"Processing: {xlsx.name}")
            fixed, missing = fix_excel_file(xlsx, self.log_msg)

            if fixed > 0:
                self.log_msg(f"  -> Fixed {fixed} hyperlinks")
                files_modified += 1
            elif missing == 0:
                self.log_msg(f"  -> No changes needed")

            grand_fixed += fixed
            grand_missing += missing

        self.log_msg("")
        self.log_msg("=" * 50)
        self.log_msg(f"DONE!")
        self.log_msg(f"  Files modified: {files_modified}")
        self.log_msg(f"  Hyperlinks fixed: {grand_fixed}")
        self.log_msg(f"  Missing images: {grand_missing}")

        if grand_fixed > 0:
            messagebox.showinfo("Done", f"Fixed {grand_fixed} hyperlinks in {files_modified} file(s)!")
        else:
            messagebox.showinfo("Done", "No hyperlinks needed fixing.")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
