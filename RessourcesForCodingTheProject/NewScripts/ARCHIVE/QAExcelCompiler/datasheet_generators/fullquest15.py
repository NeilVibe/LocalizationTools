#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Quest-data extractor with legacy grouping (Main / Faction / Daily / Challenge / Minigame),
new "/complete" + "/teleport" logic, robust sheet merging, QuestStart support,
two-step faction lookup (FactionKeyList → filename → factioninfo),
and an "Others" fallback tab.

EXTRA – TELEPORT POST-PROCESSOR
───────────────────────────────
After every language workbook is built the script post-processes the sheets with
/teleport information that is stored in the fixed file

        Quest_LQA_ENG_1231_seon_final_final.xlsx      (== 'reference file')

Changes (vs previous revision):
1. The reference file may contain **duplicate StringKeys that live on different
   sheets/tabs**.  
   → A teleport string is now associated with the **pair (SheetName, StringKey)**.
2. When augmenting a freshly built workbook the script only looks into the
   sub-map that corresponds to the **current target sheet's title** (trimmed to
   Excel's 31-char limit).  
   Hence identical StringKeys on different tabs no longer collide.
3. No duplicate /teleport is ever appended.

Details of teleport mapping:
• In the reference workbook each sheet is scanned.
• Column C (index 2)  – StringKey (lookup key).  
• Column E (index 4)  – teleport text (value to append).
• Resulting structure returned by load_teleport_map():

        {
            "SheetTitle1" : { "1000659" : "/teleport -6718 534 -5319", ... },
            "Challenge Quest" : { "1000659" : "/create item DarkLeader_TwoHandSword", ... },
            ...
        }

When processing a target workbook sheet "SheetTitle1" (trimmed to 31 chars) we
fetch teleport_map.get("SheetTitle1", {}) and pass that to apply_teleport_map().
Only hits from this sub-map are considered; cross-sheet duplicates are ignored.

All header cells are light-blue & bold.
"""

from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable, Any

from lxml import etree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION – FOLDERS / FILES
# ─────────────────────────────────────────────────────────────────────────────
QUESTGROUPINFO_FILE = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\questgroupinfo.staticinfo.xml"
)
SCENARIO_FOLDER   = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\scenario"
)
FACTION_FOLDER    = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\quest\faction"
)
CHALLENGE_FOLDER  = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\Challenge"
)
LANGUAGE_FOLDER   = Path(
    r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"
)
STRINGKEYTABLE_FILE = Path(
    r"F:\perforce\cd\mainline\resource\editordata\StaticInfo__\StaticInfo_StringKeyTable.xml"
)
SEQUENCER_FOLDER  = Path(
    r"F:\perforce\cd\mainline\resource\sequencer\stageseq"
)
FACTIONINFO_FOLDER = Path(
    r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\factioninfo"
)

# ── minigame file ───────────────────────────────────────────────────────────
MINIGAME_FILE = Path(
    r"F:\perforce\cd\mainline\resource\GameData\staticinfo_quest\contents\contents_minigame.staticinfo.xml"
)

# ── teleport lookup ─────────────────────────────────────────────────────────
TELEPORT_SOURCE_FILE = Path.cwd() / "Quest_LQA_ENG_1231_seon_final_final.xlsx"
# ────────────────────────────────────────────────────────────────────────────

if getattr(sys, "frozen", False):
    base_path = Path(sys.executable).parent
else:
    base_path = Path(__file__).parent

OUTPUT_FOLDER = base_path / "QuestData_Map_All"
LOG_FILE      = base_path / "quest_scan.log"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("QuestLQA")
log.setLevel(logging.DEBUG)
# fh = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
# fh.setFormatter(logging.Formatter(
    # "%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S"))
# fh.setLevel(logging.DEBUG)
ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(logging.Formatter("%(levelname)-8s  %(message)s"))
ch.setLevel(logging.INFO)
# log.addHandler(fh)
log.addHandler(ch)
log.propagate = False

# ─────────────────────────────────────────────────────────────────────────────
# XML SANITISATION
# ─────────────────────────────────────────────────────────────────────────────
_bad_entity_re = re.compile(r'&(?!lt;|gt;|amp;|apos;|quot;)')

def _fix_bad_entities(txt: str) -> str:
    return _bad_entity_re.sub("&amp;", txt)

def _preprocess_newlines(raw: str) -> str:
    def repl(m: re.Match) -> str:
        inner = m.group(1).replace("\n", "&lt;br/&gt;").replace("\r\n", "&lt;br/&gt;")
        return f"<seg>{inner}</seg>"
    return re.sub(r"<seg>(.*?)</seg>", repl, raw, flags=re.DOTALL)

def sanitize_xml(raw: str) -> str:
    raw = _fix_bad_entities(raw)
    raw = _preprocess_newlines(raw)
    raw = re.sub(r'="([^"]*<[^"]*)"',
                 lambda m: '="' + m.group(1).replace("<", "&lt;") + '"', raw)
    raw = re.sub(r'="([^"]*&[^ltgapoqu][^"]*)"',
                 lambda m: '="' + m.group(1).replace("&", "&amp;") + '"', raw)
    tag_open  = re.compile(r"<([A-Za-z0-9_]+)(\s[^>]*)?>")
    tag_close = re.compile(r"</([A-Za-z0-9_]+)>")
    stack: List[str] = []
    out:   List[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        mo = tag_open.match(stripped)
        if mo:
            stack.append(mo.group(1)); out.append(line); continue
        mc = tag_close.match(stripped)
        if mc:
            if stack and stack[-1] == mc.group(1):
                stack.pop(); out.append(line)
            else:
                out.append(stack and f"</{stack.pop()}>" or line)
            continue
        if stripped.startswith("</>"):
            out.append(stack and line.replace("</>", f"</{stack.pop()}>") or line)
            continue
        out.append(line)
    while stack:
        out.append(f"</{stack.pop()}>")
    return "\n".join(out)

def parse_xml_file(path: Path) -> Optional[ET._Element]:
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception:
        log.exception("Cannot read %s", path)
        return None
    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"
    try:
        return ET.fromstring(
            wrapped.encode("utf-8"),
            parser=ET.XMLParser(huge_tree=True)
        )
    except ET.XMLSyntaxError:
        try:
            return ET.fromstring(
                wrapped.encode("utf-8"),
                parser=ET.XMLParser(recover=True, huge_tree=True)
            )
        except ET.XMLSyntaxError:
            log.exception("Parse failed for %s", path)
            return None

# ─────────────────────────────────────────────────────────────────────────────
# GENERIC HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def iter_xml_files(
    root: Path,
    suffixes: Tuple[str, ...] = (".xml", ".seqc")
) -> Iterable[Path]:
    for dp, _, files in os.walk(root):
        for fn in files:
            low = fn.lower()
            if any(low.endswith(suf) for suf in suffixes):
                yield Path(dp) / fn

def _parse_vec3(s: str) -> Optional[Tuple[float, float, float]]:
    parts = [p for p in s.replace(",", " ").split() if p]
    if len(parts) != 3:
        return None
    try:
        return tuple(float(x) for x in parts)  # type: ignore[return-value]
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE TABLES
# ─────────────────────────────────────────────────────────────────────────────
def parse_language_folder(folder: Path) -> Dict[str, Dict[str, Tuple[str, str]]]:
    """
    Returns:
      { lang_code : { origin_string : (translated_text , string_id) } }
    """
    log.info("Scanning language folder: %s", folder)
    langs: Dict[str, Dict[str, Tuple[str, str]]] = {}
    for p in iter_xml_files(folder, (".xml",)):
        stem = p.stem.lower()
        if not stem.startswith("languagedata_") or stem.startswith("languagedata_kor"):
            continue
        code = stem.split("_", 1)[1].lower()
        root = parse_xml_file(p)
        if root is None:
            log.error("Parse error: %s", p)
            continue
        tbl: Dict[str, Tuple[str, str]] = {}
        for el in root.iter("LocStr"):
            origin = el.get("StrOrigin") or ""
            sid    = el.get("StringId")  or ""
            txt    = el.get("Str")       or ""
            if origin:
                tbl[origin] = (txt, sid)
        langs[code] = tbl
    log.info("Language tables loaded: %d", len(langs))
    return langs

# ─────────────────────────────────────────────────────────────────────────────
# STRING-KEY TABLE
# ─────────────────────────────────────────────────────────────────────────────
def load_string_key_table(path: Path) -> Dict[str, str]:
    log.info("Loading StringKeyTable: %s", path)
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse StringKeyTable")
    tbl: Dict[str, str] = {}
    for el in root.iter("StringKeyMap"):
        num = el.get("Key") or ""
        sk  = el.get("StrKey") or ""
        if num and sk:
            tbl[sk.lower()] = num
    log.info("StringKeyTable entries: %d", len(tbl))
    return tbl

# ─────────────────────────────────────────────────────────────────────────────
# STAGE → SEQUENCER MAP & POSITION MAP
# ─────────────────────────────────────────────────────────────────────────────
def _norm_seq_name(raw: str) -> str:
    return Path(raw).stem.lower()

def build_stage_to_seq_map(
    folders: List[Path]
) -> Tuple[Dict[str, str], Dict[str, str]]:
    log.info("Building Stage→Sequencer map for: %s", folders)
    mapping: Dict[str, str] = {}
    name_map: Dict[str, str] = {}
    for root_folder in folders:
        for p in iter_xml_files(root_folder, (".xml",)):
            rt = parse_xml_file(p)
            if rt is None:
                continue
            for st in rt.iter("Stage"):
                key = (st.get("StrKey") or "").lower()
                seq_tag = st.find("SequencerStage")
                if not key or seq_tag is None:
                    continue
                raw = seq_tag.get("Sequencer") or ""
                if not raw:
                    continue
                seq = _norm_seq_name(raw)
                mapping.setdefault(key, seq)
                if not key.startswith("mission_"):
                    mapping.setdefault(f"mission_{key}", seq)
                else:
                    mapping.setdefault(key[len("mission_"):], seq)
                name = st.get("Name") or ""
                if name:
                    name_map.setdefault(name.lower(), seq)
    log.info("  → Stage2Seq: %d ; Name→Seq: %d",
             len(mapping), len(name_map))
    return mapping, name_map

def build_seq_position_map(
    root_folder: Path
) -> Dict[str, Tuple[float, float, float]]:
    log.info("Scanning Sequencer folder for positions: %s", root_folder)
    seq_pos: Dict[str, Tuple[float, float, float]] = {}
    for idx, p in enumerate(iter_xml_files(root_folder, (".seqc",)), 1):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        seq_el = rt if rt.tag == "Sequencer" else next(rt.iter("Sequencer"), None)
        if seq_el is None:
            continue
        vec = _parse_vec3(seq_el.get("Position") or "")
        if vec:
            seq_pos[p.stem.lower()] = vec
        if idx % 500 == 0:
            log.debug("… parsed %d sequencer files", idx)
    log.info("Sequencer positions parsed: %d", len(seq_pos))
    return seq_pos

# ─────────────────────────────────────────────────────────────────────────────
# FACTION INFO + QUEST→FACTION MAP + NODE→FACTION MAP
# ─────────────────────────────────────────────────────────────────────────────
def parse_faction_info(
    folder: Path
) -> Tuple[Dict[str, Tuple[str, str]], Dict[str, str], Dict[str, str]]:
    """
    Returns:
      faction_map:   key = faction_strkey.lower()       → (faction_name, orderByString)
      quest_to_fac:  key = questKey.lower()             → faction_strkey.lower()
      node_to_fac:   key = factionnode_strkey.lower()   → faction_strkey.lower()
    """
    log.info("Scanning faction info folder: %s", folder)
    faction_map: Dict[str, Tuple[str, str]] = {}
    quest_to_faction: Dict[str, str] = {}
    node_to_faction: Dict[str, str] = {}
    for p in iter_xml_files(folder, (".xml",)):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        for fac in rt.iter("Faction"):
            fac_sk = (fac.get("StrKey") or "").lower()
            fac_name = fac.get("Name") or ""
            order = None
            for dev in fac.iter("Dev"):
                order = dev.get("OrderByString")
                if order:
                    break
            if fac_sk and fac_name and order:
                faction_map[fac_sk] = (fac_name, order)
            if fac_sk:
                for qel in fac.iter("Quest"):
                    qk = (qel.get("QuestKey") or "").lower()
                    if qk:
                        quest_to_faction[qk] = fac_sk
                for fn in fac.iter("FactionNode"):
                    node_sk = (fn.get("StrKey") or "").lower()
                    if node_sk:
                        node_to_faction[node_sk] = fac_sk
    log.info("Faction infos loaded: %d ; Quests→Faction: %d ; Nodes→Faction: %d",
             len(faction_map), len(quest_to_faction), len(node_to_faction))
    return faction_map, quest_to_faction, node_to_faction

# ─────────────────────────────────────────────────────────────────────────────
# COMMAND BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_command(
    el: ET._Element,
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> str:
    mission_sk = (
        (el.get("StrKey") or "")
        or (el.get("StartMission") or "")
        or (el.get("QuestStart")   or "")
    ).strip()
    if not mission_sk:
        return ""
    if el.tag == "Mission":
        prefix = "prevmission"
    elif el.tag == "SubMission":
        prefix = "prevsubmission"
    else:
        prefix = "prevquest"
    skey = id_tbl.get(mission_sk.lower())
    if skey:
        complete = f"/complete {prefix} {skey}"
    else:
        complete = f"/complete prevquestgroup {mission_sk}"
    seq = stage2seq.get(mission_sk.lower())
    if not seq:
        for sub in el.iter("ExecuteStage"):
            sk = (sub.get("StageKey") or "").lower()
            seq = stage2seq.get(sk)
            if seq:
                break
    if not seq:
        nm = el.get("Name") or ""
        seq = name_map.get(nm.lower())
    if not seq:
        return complete
    pos = seq_pos.get(seq)
    if not pos:
        return complete
    teleport = "/teleport " + " ".join(f"{c:g}" for c in pos)
    return f"{complete}\n{teleport}"

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLES / HELPERS
# ─────────────────────────────────────────────────────────────────────────────
Row = Tuple[int, str, str, str, str, str, bool, bool, str, str, str, str]


def dedupe_rows(rows: List[Row], is_eng: bool) -> List[Row]:
    """
    DISABLED for Quest - keep all duplicates since quest content is important.

    Quest text may appear multiple times in different contexts, and testers
    need to review each occurrence separately.
    """
    # Return rows as-is without deduplication
    return rows

def _dedupe_rows_disabled(rows: List[Row], is_eng: bool) -> List[Row]:
    """
    [DISABLED] Remove duplicate rows based on (Korean, Translation, STRINGID).

    For ENG: key = (orig, eng, sid) = indices (1, 2, 5)
    For non-ENG: key = (orig, loc, sid) = indices (1, 3, 5)

    Keeps first occurrence of each duplicate.
    """
    seen = set()
    result = []
    duplicates = 0

    for row in rows:
        orig = row[1]  # Korean
        trans = row[2] if is_eng else row[3]  # ENG or LOC
        sid = row[5]   # STRINGID

        key = (orig, trans, sid)
        if key in seen:
            duplicates += 1
            continue

        seen.add(key)
        result.append(row)

    if duplicates > 0:
        log.info("    Removed %d duplicate rows (Korean+Translation+STRINGID)", duplicates)

    return result


_depth_fill = {
    0: PatternFill("solid", fgColor="FFD966"),
    1: PatternFill("solid", fgColor="D9E1F2"),
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FCE4D6"),
    4: PatternFill("solid", fgColor="9C27B0"),
}
_icon_fill    = PatternFill("solid", fgColor="9BC2E6")
_light_yellow = PatternFill("solid", fgColor="FFFDEB")
_border       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
_bold_font    = Font(bold=True)
_header_font  = Font(bold=True)
_header_fill  = PatternFill("solid", fgColor="9BC2E6")

def _tr(text: str, tbl: Dict[str, Tuple[str, str]]) -> str:
    return tbl.get(text, ("", ""))[0]

def _sid(text: str, tbl: Dict[str, Tuple[str, str]]) -> str:
    return tbl.get(text, ("", ""))[1]

def autofit(ws) -> None:
    fixed_widths = {
        "COMMENT": 50,
        "SCREENSHOT": 20,
        "STRINGID": 25,
        "StringKey": 13,
        "ENG": 40
    }
    headers = [c.value for c in ws[1]]
    for col_idx, col in enumerate(ws.columns, start=1):
        letter = col[0].column_letter
        header_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else None
        if header_name in fixed_widths:
            ws.column_dimensions[letter].width = fixed_widths[header_name]
            continue
        max_len = 0
        for cell in col[1:]:
            if cell.value not in (None, ""):
                for line in str(cell.value).splitlines():
                    max_len = max(max_len, len(line))
        ws.column_dimensions[letter].width = 10 if max_len == 0 else min(max_len + 2, 80)

# ─────────────────────────────────────────────────────────────────────────────
# TELEPORT LOOKUP (per-sheet mapping)
# ─────────────────────────────────────────────────────────────────────────────
def load_teleport_map(path: Path) -> Dict[str, Dict[str, str]]:
    """
    Returns:
        { sheet_title_trimmed_to_31 : { StringKey : teleport_string } }
    """
    log.info("Loading teleport lookup from: %s", path)
    if not path.is_file():
        log.warning("Teleport source file not found – skipping.")
        return {}

    wb = load_workbook(path, data_only=True)
    mapping: Dict[str, Dict[str, str]] = {}

    for ws in wb.worksheets:
        sheet_key = ws.title[:31]  # mimic target sheet title trimming
        sub_map: Dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 5:
                continue
            key_cell, tele_cell = row[2], row[4]
            if key_cell is None or tele_cell in (None, ""):
                continue
            sk = str(key_cell).strip()
            teleport = str(tele_cell).strip()
            if sk and teleport:
                sub_map.setdefault(sk, teleport)
        if sub_map:
            mapping[sheet_key] = sub_map
    log.info("Teleport entries loaded: %d sheets, %d total rows",
             len(mapping), sum(len(v) for v in mapping.values()))
    return mapping

def apply_teleport_map(rows: List[Row], tp_submap: Dict[str, str]) -> List[Row]:
    """
    Append teleport strings to Command field when StringKey matches entry in
    tp_submap (which is already sheet-specific). Avoid duplicates.
    """
    out: List[Row] = []
    for (
        depth, orig, eng, loc, sk, sid, icon, bold,
        cmd, status, comment, shot
    ) in rows:
        new_cmd = cmd
        if sk:
            tele = tp_submap.get(sk)
            if tele and tele not in (cmd or ""):
                new_cmd = f"{cmd}\n{tele}" if cmd else tele
        out.append((
            depth, orig, eng, loc, sk, sid, icon, bold,
            new_cmd, status, comment, shot
        ))
    return out

# ─────────────────────────────────────────────────────────────────────────────
# QUEST-GROUP META (for Challenge)
# ─────────────────────────────────────────────────────────────────────────────
def parse_group_meta(path: Path) -> Dict[str, str]:
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse group meta")
    meta: Dict[str, str] = {}
    for el in root.iter("QuestGroup"):
        sk = (el.get("StrKey") or "").lower()
        nm = el.get("Name") or ""
        if sk:
            meta[sk] = nm
    return meta

# ─────────────────────────────────────────────────────────────────────────────
# MASTER QUESTS PARSING
# ─────────────────────────────────────────────────────────────────────────────
def parse_master_quests(path: Path) -> Tuple[List[str], Dict[str, str]]:
    root = parse_xml_file(path)
    if root is None:
        sys.exit("Cannot parse QuestGroupInfo")
    order: List[str] = []
    q2g: Dict[str, str] = {}
    seen: set[str] = set()
    for el in root.iter():
        grp = el.get("Group")
        if grp:
            qk = el.tag
            if qk not in seen:
                order.append(qk)
                q2g[qk] = grp
                seen.add(qk)
    return order, q2g

# ─────────────────────────────────────────────────────────────────────────────
# ROW BUILDERS (unchanged logic, teleport appended later)
# ─────────────────────────────────────────────────────────────────────────────
def build_rows_main(
    scenario_folder: Path,
    quest_order: List[str],
    quest_groups: Dict[str, str],
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    group_meta: Dict[str, str],
) -> List[Row]:
    log.info("Building rows for MAIN quests …")
    quests: Dict[str, ET._Element] = {}
    for p in iter_xml_files(scenario_folder, (".xml",)):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        for q in rt.iter():
            if q.get("Name") and (q.get("StartMission") or q.get("QuestStart")):
                quests[q.tag] = q

    rows: List[Row] = []
    curg = None
    for qk in quest_order:
        q = quests.get(qk)
        if q is None:
            continue
        grp = quest_groups.get(qk, "")
        if grp != curg:
            curg = grp
            kr_group = group_meta.get(grp.lower(), grp)
            rows.append((
                0,
                kr_group,
                _tr(kr_group, eng_tbl),
                _tr(kr_group, lang_tbl),
                "",                                # StringKey
                _sid(kr_group, eng_tbl),           # StringID
                False, True,                       # icon, bold
                "", "", "", ""                     # cmd, status, comment, screenshot
            ))

        kor  = q.get("Name") or ""
        skn  = id_tbl.get(qk.lower(), "")
        rows.append((
            1,
            kor,
            _tr(kor, eng_tbl),
            _tr(kor, lang_tbl),
            skn,
            _sid(kor, eng_tbl),
            bool(q.get("StageIcon")), True,
            "", "", "", ""                       # cmd, status, comment, screenshot
        ))

        for m in q.findall("Mission"):
            mk  = m.get("Name")   or ""
            msk = m.get("StrKey") or ""
            cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
            rows.append((
                2,
                mk,
                _tr(mk, eng_tbl),
                _tr(mk, lang_tbl),
                id_tbl.get(msk.lower(), msk),
                _sid(mk, eng_tbl),
                False, True,
                cmd, "", "", ""
            ))
            for s in m.findall("SubMission"):
                sk  = s.get("Name")   or ""
                ssk = s.get("StrKey") or ""
                cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    3,
                    sk,
                    _tr(sk, eng_tbl),
                    _tr(sk, lang_tbl),
                    id_tbl.get(ssk.lower(), ssk),
                    _sid(sk, eng_tbl),
                    False, True,
                    cmd2, "", "", ""
                ))
    log.info("Main quest rows: %d", len(rows))
    return rows

def _bump_depth(rows: List[Row], inc: int = 1) -> List[Row]:
    return [(r[0] + inc,) + r[1:] for r in rows]

def build_rows_simple(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    faction_map: Dict[str, Tuple[str, str]],
    quest2fac: Dict[str, str],
    node2fac: Dict[str, str],
) -> Dict[str, List[Row]]:
    log.info("Building rows for FACTION quests …")

    def find_faction(el: ET._Element) -> str:
        for sub in el.iter():
            if "FactionNodeKeyList" in sub.attrib:
                parts = [v.strip() for v in (sub.get("FactionNodeKeyList") or "").split() if v.strip()]
                if parts:
                    fk = node2fac.get(parts[0].lower())
                    if fk:
                        return fk
        for sub in el.iter():
            if "FactionKeyList" in sub.attrib:
                parts = [v.strip() for v in (sub.get("FactionKeyList") or "").split(",") if v.strip()]
                if parts:
                    return parts[0].lower()
        return quest2fac.get(el.tag.lower(), "")

    grouping: Dict[str, List[Row]] = {}
    for p in iter_xml_files(folder, (".xml",)):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        for q in rt.iter():
            if not (q.get("Name") and (q.get("StartMission") or q.get("QuestStart"))):
                continue
            fk = find_faction(q)
            if not fk:
                continue
            kor  = q.get("Name") or ""
            rows: List[Row] = []
            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(q.tag.lower(), ""),
                _sid(kor, eng_tbl),
                bool(q.get("StageIcon")), True,
                "", "", "", ""
            ))
            for m in q.findall("Mission"):
                mk  = m.get("Name") or ""
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))
                for s in m.findall("SubMission"):
                    sk  = s.get("Name") or ""
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))
            grouping.setdefault(fk, []).extend(_bump_depth(rows, 1))
    log.info("Faction grouping keys: %d", len(grouping))
    return grouping

def build_rows_daily(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> List[Row]:
    log.info("Building rows for DAILY quests …")
    rows: List[Row] = []
    for p in iter_xml_files(folder, (".xml",)):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        for q in rt.iter():
            if not (q.get("Name") and q.get("StartMission")):
                continue
            grp = (q.get("Group") or "").lower()
            if "daily" not in grp:
                continue
            kor  = q.get("Name") or ""
            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(q.tag.lower(), ""),
                _sid(kor, eng_tbl),
                bool(q.get("StageIcon")), True,
                "", "", "", ""
            ))
            for m in q.findall("Mission"):
                mk  = m.get("Name") or ""
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))
                for s in m.findall("SubMission"):
                    sk  = s.get("Name") or ""
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))
    log.info("Daily rows: %d", len(rows))
    return rows

_CHT_TP     = {"어비스", "탐험"}
_itemkey_re = re.compile(r'ItemKey\(\s*([A-Za-z0-9_]+)\s*\)')

def _extract_item_command(el: ET._Element) -> str:
    last: Optional[str] = None
    for sub in el.iter():
        for n, v in sub.attrib.items():
            if "EventCondition" in n or n == "EventCondition":
                m = _itemkey_re.findall(v)
                if m:
                    last = m[-1]
    return f"/create item {last}" if last else ""

def build_rows_challenge(
    folder: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
    group_meta: Dict[str, str],
) -> List[Row]:
    log.info("Building rows for CHALLENGE quests …")
    rows: List[Row] = []
    for p in iter_xml_files(folder, (".xml",)):
        rt = parse_xml_file(p)
        if rt is None:
            continue
        for q in rt:
            if not (q.get("Name") and q.get("StartMission")):
                continue
            group_code = (q.get("Group") or "").lower()
            group_name = group_meta.get(group_code, group_code)

            rows.append((
                4,
                group_name, _tr(group_name, eng_tbl), _tr(group_name, lang_tbl),
                "", _sid(group_name, eng_tbl),
                False, True,
                "", "", "", ""
            ))

            kor = q.get("Name") or ""
            use_tp = group_name in _CHT_TP
            cmd = build_command(q, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(q)
            rows.append((
                0,
                kor, _tr(kor, eng_tbl), _tr(kor, lang_tbl),
                id_tbl.get(q.tag.lower(), ""),
                _sid(kor, eng_tbl),
                False, True, cmd, "", "", ""
            ))
            for m in q.findall("Mission"):
                mk  = m.get("Name") or ""
                msk = m.get("StrKey") or ""
                cmd2 = build_command(m, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(m)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd2, "", "", ""
                ))
                for s in m.findall("SubMission"):
                    sk  = s.get("Name") or ""
                    ssk = s.get("StrKey") or ""
                    cmd3 = build_command(s, id_tbl, stage2seq, name_map, seq_pos) if use_tp else _extract_item_command(s)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd3, "", "", ""
                    ))
    log.info("Challenge rows: %d", len(rows))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# MINIGAME QUEST BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_rows_minigame(
    minigame_file: Path,
    lang_tbl: Dict[str, Tuple[str, str]],
    eng_tbl: Dict[str, Tuple[str, str]],
    id_tbl: Dict[str, str],
    stage2seq: Dict[str, str],
    name_map: Dict[str, str],
    seq_pos: Dict[str, Tuple[float, float, float]],
) -> List[Row]:
    """
    Parse the minigame XML file.
    
    Two element types:
    1. <Contents_MiniGame_*_Mission> with Group="Challenge_Minigame"
       - Has Name, StartMission, contains <Mission> with <SubMission>
    2. <Contents_MiniGame_*> with Group="MiniGame"  
       - Has Name, contains <Stage> elements with StrKey, Name, Desc, StagePosition
    """
    log.info("Building rows for MINIGAME quests from: %s", minigame_file)
    
    if not minigame_file.is_file():
        log.warning("Minigame file not found: %s", minigame_file)
        return []
    
    rt = parse_xml_file(minigame_file)
    if rt is None:
        log.error("Failed to parse minigame file: %s", minigame_file)
        return []
    
    rows: List[Row] = []
    
    for el in rt:
        tag = el.tag
        name = el.get("Name") or ""
        group = (el.get("Group") or "").lower()
        
        if not name:
            continue
        
        # Type 1: Mission-based minigames (Challenge_Minigame)
        if "_mission" in tag.lower() and el.get("StartMission"):
            # Quest header row
            rows.append((
                0,
                name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                id_tbl.get(tag.lower(), ""),
                _sid(name, eng_tbl),
                bool(el.get("StageIcon")), True,
                "", "", "", ""
            ))
            
            # Process Mission children
            for m in el.findall("Mission"):
                mk = m.get("Name") or ""
                msk = m.get("StrKey") or ""
                cmd = build_command(m, id_tbl, stage2seq, name_map, seq_pos)
                rows.append((
                    1,
                    mk, _tr(mk, eng_tbl), _tr(mk, lang_tbl),
                    id_tbl.get(msk.lower(), msk),
                    _sid(mk, eng_tbl),
                    False, True, cmd, "", "", ""
                ))
                
                # Process SubMission children
                for s in m.findall("SubMission"):
                    sk = s.get("Name") or ""
                    ssk = s.get("StrKey") or ""
                    cmd2 = build_command(s, id_tbl, stage2seq, name_map, seq_pos)
                    rows.append((
                        2,
                        sk, _tr(sk, eng_tbl), _tr(sk, lang_tbl),
                        id_tbl.get(ssk.lower(), ssk),
                        _sid(sk, eng_tbl),
                        False, True, cmd2, "", "", ""
                    ))
        
        # Type 2: Stage-based minigames (MiniGame)
        elif group == "minigame" or el.find("Stage") is not None:
            # Quest header row
            rows.append((
                0,
                name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                id_tbl.get(tag.lower(), ""),
                _sid(name, eng_tbl),
                bool(el.get("StageIcon")), True,
                "", "", "", ""
            ))
            
            # Process Stage children
            for stage in el.findall("Stage"):
                stage_name = stage.get("Name") or ""
                stage_sk = stage.get("StrKey") or ""
                stage_desc = stage.get("Desc") or ""
                
                # Build teleport command from StagePosition if available
                cmd = ""
                stage_pos_str = stage.get("StagePosition") or ""
                if stage_pos_str:
                    pos = _parse_vec3(stage_pos_str)
                    if pos:
                        cmd = "/teleport " + " ".join(f"{c:g}" for c in pos)
                
                # If no StagePosition, try to get from SequencerStage
                if not cmd:
                    seq_stage = stage.find("SequencerStage")
                    if seq_stage is not None:
                        seq_name = seq_stage.get("Sequencer") or ""
                        if seq_name:
                            seq_key = _norm_seq_name(seq_name)
                            pos = seq_pos.get(seq_key)
                            if pos:
                                cmd = "/teleport " + " ".join(f"{c:g}" for c in pos)
                
                rows.append((
                    1,
                    stage_name, _tr(stage_name, eng_tbl), _tr(stage_name, lang_tbl),
                    id_tbl.get(stage_sk.lower(), stage_sk),
                    _sid(stage_name, eng_tbl),
                    bool(stage.get("StageIcon")), False,
                    cmd, "", "", ""
                ))
    
    log.info("Minigame rows: %d", len(rows))
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────
def write_sheet(
    wb: Workbook,
    title: str,
    rows: List[Row],
    tgt_code: str
) -> None:
    ws = wb.create_sheet(title=title[:31])
    is_eng = tgt_code.lower().startswith("eng")
    headers = (
        ["Original", "ENG"] +
        ([] if is_eng else [tgt_code.upper()]) +
        ["StringKey", "Command", "STATUS", "COMMENT", "STRINGID", "SCREENSHOT"]
    )

    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(1, col, hdr)
        cell.font      = _header_font
        cell.fill      = _header_fill
        cell.border    = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    indent_stop = headers.index("StringKey") + 1

    r = 2
    for (
        depth, orig, eng, loc, sk, sid,
        icon, bold, cmd, status, comment, shot
    ) in rows:
        if is_eng:
            vals = (orig, eng, sk, cmd, status, comment, sid, shot)
        else:
            vals = (orig, eng, loc, sk, cmd, status, comment, sid, shot)

        fg = _icon_fill if icon else _depth_fill.get(depth, _light_yellow)
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(r, col_idx, v)
            cell.alignment = Alignment(
                indent=(depth if col_idx < indent_stop else 0),
                wrap_text=True,
                vertical="top"
            )
            cell.border = _border
            cell.fill   = fg
            if bold:
                cell.font = _bold_font
        r += 1

    ws.freeze_panes = "A2"
    autofit(ws)
    if not is_eng:
        eng_idx = headers.index("ENG") + 1
        ws.column_dimensions[ws.cell(row=1, column=eng_idx).column_letter].hidden = True

    # ─────────────────────────────────────────────────────────────
    # Add STATUS drop-down validation
    # ─────────────────────────────────────────────────────────────
    def _add_status_validation(sh, status_col_idx: int, max_row: int) -> None:
        col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"ISSUE,NO ISSUE,BLOCKED,KOREAN"',
            allow_blank=True,
            showErrorMessage=True
        )
        rng = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(rng)
        sh.add_data_validation(dv)

    status_col_idx = headers.index("STATUS") + 1
    _add_status_validation(ws, status_col_idx, ws.max_row)

    # ─────────────────────────────────────────────────────────────
    # Force STRINGID column to text format (prevents scientific notation)
    # ─────────────────────────────────────────────────────────────
    stringid_col_idx = headers.index("STRINGID") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, stringid_col_idx).number_format = '@'

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("Quest extraction started")
    OUTPUT_FOLDER.mkdir(exist_ok=True)
    log.info("Output folder: %s", OUTPUT_FOLDER)

    # 1. Teleport lookup (per-sheet maps)
    teleport_map = load_teleport_map(TELEPORT_SOURCE_FILE)

    # 2. Language tables
    lang_tables = parse_language_folder(LANGUAGE_FOLDER)
    if not lang_tables:
        sys.exit("No language files parsed.")
    eng_code = next((c for c in lang_tables if c.startswith("eng")), None)
    if eng_code is None:
        sys.exit("English missing.")
    eng_tbl = lang_tables[eng_code]

    # 3. StringKey table
    id_tbl = load_string_key_table(STRINGKEYTABLE_FILE)

    # 4. Quest group info & meta
    quest_order, quest_groups = parse_master_quests(QUESTGROUPINFO_FILE)
    group_meta = parse_group_meta(QUESTGROUPINFO_FILE)

    # 5. Sequencer maps
    stage2seq_all, name_map = build_stage_to_seq_map([
        SCENARIO_FOLDER, FACTION_FOLDER, CHALLENGE_FOLDER
    ])
    seq_pos = build_seq_position_map(SEQUENCER_FOLDER)

    # 6. Faction info
    faction_info_map, quest2fac_map, node2fac_map = parse_faction_info(FACTIONINFO_FOLDER)

    # 7. Process each language
    for idx, (code, lang_tbl) in enumerate(lang_tables.items(), 1):
        log.info("(%d/%d) Language %s", idx, len(lang_tables), code.upper())

        rows_main = build_rows_main(
            SCENARIO_FOLDER, quest_order, quest_groups,
            lang_tbl, eng_tbl, id_tbl,
            stage2seq_all, name_map, seq_pos,
            group_meta
        )

        faction_groups = build_rows_simple(
            FACTION_FOLDER, lang_tbl, eng_tbl,
            id_tbl, stage2seq_all, name_map, seq_pos,
            faction_info_map, quest2fac_map, node2fac_map
        )

        rows_daily = build_rows_daily(
            FACTION_FOLDER, lang_tbl, eng_tbl,
            id_tbl, stage2seq_all, name_map, seq_pos
        )

        rows_chal = build_rows_challenge(
            CHALLENGE_FOLDER, lang_tbl, eng_tbl,
            id_tbl, stage2seq_all, name_map, seq_pos,
            group_meta
        )

        rows_minigame = build_rows_minigame(
            MINIGAME_FILE, lang_tbl, eng_tbl,
            id_tbl, stage2seq_all, name_map, seq_pos
        )

        sheet_data: Dict[str, List[Row]] = {"Main Quest": rows_main}

        def parse_order(ordstr: str) -> Tuple[int, object]:
            m = re.search(r"(\d+)", ordstr)
            if m:
                return (0, int(m.group(1)))
            return (1, ordstr)

        sorted_factions = sorted(
            [(fk, name, ordstr) for fk, (name, ordstr) in faction_info_map.items()],
            key=lambda tup: parse_order(tup[2])
        )

        faction_sheet_ordstr: Dict[str, str] = {}
        for fk, name, ordstr in sorted_factions:
            rows = faction_groups.get(fk)
            if not rows:
                continue
            sheet_name = ordstr[:31]
            faction_sheet_ordstr[sheet_name] = ordstr
            header_row: Row = (
                0, name, _tr(name, eng_tbl), _tr(name, lang_tbl),
                "", _sid(name, eng_tbl),
                False, True, "", "", "", ""
            )
            sheet_data.setdefault(sheet_name, []).append(header_row)
            sheet_data[sheet_name].extend(rows)

        leftover_keys = set(faction_groups) - set(faction_info_map)
        if leftover_keys:
            others_rows: List[Row] = []
            for fk in leftover_keys:
                others_rows.extend(faction_groups[fk])
            sheet_data["Others"] = [
                (0, "Others", "", "", "", "", False, True, "", "", "", "")
            ] + others_rows

        if rows_daily:
            sheet_data["Daily Quest"] = rows_daily
        sheet_data["Challenge Quest"] = rows_chal
        if rows_minigame:
            sheet_data["Minigame Quest"] = rows_minigame

        # 7.1  Apply teleport per-sheet
        for sheet_name in list(sheet_data.keys()):
            sub_map = teleport_map.get(sheet_name, {})
            sheet_data[sheet_name] = apply_teleport_map(sheet_data[sheet_name], sub_map)

        # 8. Write workbook
        wb = Workbook()
        wb.remove(wb.active)

        ordered = ["Main Quest"]
        ordered += sorted(faction_sheet_ordstr, key=lambda sn: parse_order(faction_sheet_ordstr[sn]))
        if "Others" in sheet_data:
            ordered.append("Others")
        if "Daily Quest" in sheet_data:
            ordered.append("Daily Quest")
        ordered.append("Challenge Quest")
        if "Minigame Quest" in sheet_data:
            ordered.append("Minigame Quest")

        is_eng = code.lower().startswith("eng")
        for title in ordered:
            if title in sheet_data and sheet_data[title]:
                deduped = dedupe_rows(sheet_data[title], is_eng)
                write_sheet(wb, title, deduped, code)

        out = OUTPUT_FOLDER / f"Quest_LQA_{code.upper()}.xlsx"
        wb.save(out)
        log.info("Excel saved: %s", out.name)

    log.info("All done – processed %d language(s).", len(lang_tables))

if __name__ == "__main__":
    main()
