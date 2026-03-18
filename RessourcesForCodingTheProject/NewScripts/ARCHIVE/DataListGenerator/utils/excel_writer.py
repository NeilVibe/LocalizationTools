"""
Excel Writing Utilities for DataListGenerator
==============================================
Provides styling constants and Excel generation functions.
"""

from pathlib import Path
from typing import List, Dict
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# STYLING CONSTANTS
# =============================================================================

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# Yellow fill for section headers in concatenated list
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BOLD_FONT = Font(bold=True)


def auto_fit_columns(ws, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, max_width)


@dataclass
class DataEntry:
    """A generic data entry (faction, skill, etc.)."""
    name: str
    entry_type: str
    source_file: str


def write_data_list_excel(
    entries: List[DataEntry],
    output_path: Path,
    sheet_name: str,
    headers: List[str],
    type_colors: Dict[str, PatternFill]
) -> None:
    """Generate a data list Excel file with styling.

    Args:
        entries: List of DataEntry objects
        output_path: Path to save the Excel file
        sheet_name: Name of the worksheet
        headers: Column headers (e.g., ["Name", "Type", "Source File"])
        type_colors: Mapping of entry_type to PatternFill for coloring
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data
    for row_idx, entry in enumerate(entries, start=2):
        cell_name = ws.cell(row=row_idx, column=1, value=entry.name)
        cell_name.border = THIN_BORDER

        cell_type = ws.cell(row=row_idx, column=2, value=entry.entry_type)
        cell_type.border = THIN_BORDER
        if entry.entry_type in type_colors:
            cell_type.fill = type_colors[entry.entry_type]

        cell_source = ws.cell(row=row_idx, column=3, value=entry.source_file)
        cell_source.border = THIN_BORDER

    # Auto-fit columns
    auto_fit_columns(ws)

    # Add auto-filter
    ws.auto_filter.ref = f"A1:C{len(entries) + 1}"

    # Ensure output directory exists and save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Saved: {output_path}")
