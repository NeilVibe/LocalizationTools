"""
Linkage Module - SIMPLIFIED (Following DatasheetGenerator Pattern)

Data Collection Strategy:
1. Build knowledge lookup table FIRST (StrKey -> Name, UITextureName)
2. Collect ALL entries, don't filter by DDS existence
3. Mark entries that have verified images for prioritization
4. Support 3 modes: MAP (Region), CHARACTER, ITEM

Key difference from before: We DON'T filter out entries without images.
We collect everything, prioritize those with images.
"""

import re
import logging
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

from .xml_parser import parse_xml, iter_xml_files

log = logging.getLogger(__name__)


# =============================================================================
# DATA MODE ENUM
# =============================================================================

class DataMode(Enum):
    """Application data modes."""
    MAP = "map"
    CHARACTER = "character"
    ITEM = "item"
    AUDIO = "audio"


# =============================================================================
# KNOWLEDGE LOOKUP (Built First - Like DatasheetGenerator)
# =============================================================================

@dataclass
class KnowledgeLookup:
    """Knowledge entry from KnowledgeInfo XML."""
    strkey: str
    name: str  # Korean display name
    desc: str  # Description
    ui_texture_name: str  # UITextureName attribute
    group_key: str = ""
    source_file: str = ""


# =============================================================================
# DATA STRUCTURES (Unified - All Modes)
# =============================================================================

@dataclass
class DataEntry:
    """
    Unified data entry for all modes.

    Used for MAP (Region), CHARACTER, and ITEM modes.
    Image is OPTIONAL - entries without images are still collected.
    """
    strkey: str
    name_kr: str  # Korean name (from Knowledge lookup)
    desc_kr: str  # Description
    ui_texture_name: str  # May be empty
    dds_path: Optional[Path]  # None if no DDS found
    has_image: bool  # True if DDS exists

    # Mode-specific fields
    position: Optional[Tuple[float, float, float]] = None  # MAP mode: WorldPosition
    group: str = ""  # Group/Category name
    knowledge_key: str = ""  # Key used for Knowledge lookup
    source_file: str = ""
    entry_type: str = ""  # Type of entry (for sorting)

    # CHARACTER-specific fields
    use_macro: str = ""  # Race/Gender (e.g., "Macro_NPC_Human_Male")
    age: str = ""  # Age (e.g., "Adult", "Child")
    job: str = ""  # Job (e.g., "Job_Scholar")

    # ITEM-specific fields
    string_id: str = ""  # StringID from language table

    # AUDIO-specific fields (export path category tree)
    export_path: str = ""  # Relative dir from export root (e.g., "Dialog/QuestDialog")
    xml_order: int = 0  # Element position within source XML (0, 1, 2...)

    @property
    def position_2d(self) -> Optional[Tuple[float, float]]:
        """Get 2D position (x, z) for map display."""
        if self.position:
            return (self.position[0], self.position[2])
        return None


# =============================================================================
# DDS INDEX (Texture Lookup)
# =============================================================================

class DDSIndex:
    """Index of available DDS files for texture lookup."""

    def __init__(self):
        self._texture_folder: Optional[Path] = None
        self._dds_files: Dict[str, Path] = {}  # lowercase name -> full path
        self._scanned = False

    def scan_folder(self, texture_folder: Path, progress_callback=None) -> int:
        """Scan texture folder for DDS files."""
        self._texture_folder = texture_folder
        self._dds_files.clear()

        if not texture_folder or not texture_folder.exists():
            log.warning("Texture folder not found: %s", texture_folder)
            return 0

        if progress_callback:
            progress_callback(f"Scanning {texture_folder}...")

        count = 0
        for dds_path in texture_folder.rglob("*.dds"):
            name_lower = dds_path.stem.lower()
            self._dds_files[name_lower] = dds_path
            self._dds_files[dds_path.name.lower()] = dds_path
            count += 1

        self._scanned = True
        log.info("Indexed %d DDS files from %s", count, texture_folder)

        # DEBUG: Check for specific files that user reported issues with
        test_names = [
            "cd_knowledgeimage_node_hexe_sanctuary",
            "cd_knowledgeimage_node_dem_demenisscastle",
        ]
        for test_name in test_names:
            if test_name in self._dds_files:
                log.info("DEBUG: Found '%s' -> %s", test_name, self._dds_files[test_name])
            else:
                log.warning("DEBUG: NOT FOUND in index: '%s'", test_name)
                # Try to find similar keys
                similar = [k for k in self._dds_files.keys() if "hexe" in k or "sanctuary" in k]
                if similar:
                    log.warning("DEBUG: Similar keys found: %s", similar[:5])

        return count

    def find(self, ui_texture_name: str) -> Optional[Path]:
        """Find DDS file for texture name."""
        if not ui_texture_name:
            return None

        original_name = ui_texture_name
        name = ui_texture_name.lower().strip()

        # DEBUG: Log specific lookups
        if "hexe" in name.lower() or "sanctuary" in name.lower():
            log.info("DEBUG find(): Looking for '%s' -> normalized: '%s'", original_name, name)

        # Remove path components
        if '/' in name or '\\' in name:
            name = name.replace('\\', '/').split('/')[-1]

        # Remove .dds extension if present for lookup
        lookup_name = name
        if lookup_name.endswith('.dds'):
            lookup_name = lookup_name[:-4]

        # Try without extension
        if lookup_name in self._dds_files:
            return self._dds_files[lookup_name]

        # Try with .dds extension
        if name in self._dds_files:
            return self._dds_files[name]

        # Try adding .dds
        name_with_ext = lookup_name + '.dds'
        if name_with_ext in self._dds_files:
            return self._dds_files[name_with_ext]

        # FALLBACK: Try partial/fuzzy matching for similar names
        for key in self._dds_files.keys():
            if lookup_name in key or key in lookup_name:
                log.info("DDS fuzzy match: '%s' -> '%s'", original_name, key)
                return self._dds_files[key]

        # Log ALL misses with detail
        log.warning("DDS NOT FOUND: '%s' (tried: '%s')", original_name, lookup_name)

        return None

    @property
    def is_scanned(self) -> bool:
        return self._scanned

    @property
    def file_count(self) -> int:
        return len(self._dds_files) // 2  # Each file indexed twice


# =============================================================================
# LINKAGE RESOLVER (SIMPLIFIED)
# =============================================================================

class LinkageResolver:
    """
    Resolves and collects data following DatasheetGenerator pattern.

    Pattern:
    1. Build knowledge lookup table FIRST
    2. Collect all entries (don't filter by image)
    3. Mark entries with verified images for prioritization
    """

    def __init__(self):
        self._dds_index = DDSIndex()

        # Knowledge lookup (built first!)
        self._knowledge_lookup: Dict[str, KnowledgeLookup] = {}

        # Unified data collection (all modes use DataEntry)
        self._entries: Dict[str, DataEntry] = {}

        # Audio index (for AUDIO mode only)
        self._audio_index: Optional['AudioIndex'] = None
        self._audio_category_tree: Optional[Dict] = None

        # Track which mode's data is currently loaded
        self._current_mode: Optional[DataMode] = None

        # Stats
        self._stats = {
            'knowledge_loaded': 0,
            'entries_total': 0,
            'entries_with_image': 0,
            'entries_without_image': 0,
        }

    @property
    def current_mode(self) -> Optional[DataMode]:
        """Return which mode's data is currently loaded."""
        return self._current_mode

    # =========================================================================
    # DDS INDEX
    # =========================================================================

    def scan_textures(self, texture_folder: Path, progress_callback=None) -> int:
        """Scan texture folder for DDS files."""
        return self._dds_index.scan_folder(texture_folder, progress_callback)

    @property
    def dds_index(self) -> DDSIndex:
        return self._dds_index

    # =========================================================================
    # STEP 1: BUILD KNOWLEDGE LOOKUP (Like DatasheetGenerator)
    # =========================================================================

    def build_knowledge_lookup(
        self,
        knowledge_folder: Path,
        progress_callback=None
    ) -> int:
        """
        Build knowledge lookup table from KnowledgeInfo XML files.

        This MUST be called first before loading other data!
        Maps: StrKey -> (Name, UITextureName, Desc)
        """
        self._knowledge_lookup.clear()

        if not knowledge_folder or not knowledge_folder.exists():
            log.error("Knowledge folder not found: %s", knowledge_folder)
            return 0

        log.info("Building knowledge lookup from: %s", knowledge_folder)
        count = 0

        for path in iter_xml_files(knowledge_folder):
            if progress_callback:
                progress_callback(f"Reading {path.name}...")

            root = parse_xml(path)
            if root is None:
                continue

            # Parse ALL KnowledgeInfo elements
            for ki in root.iter("KnowledgeInfo"):
                strkey = (ki.get("StrKey") or "").strip()
                if not strkey:
                    continue

                name = (ki.get("Name") or "").strip()
                desc = (ki.get("Desc") or "").replace("<br/>", "\n").strip()
                ui_texture = (ki.get("UITextureName") or "").strip()
                group_key = (ki.get("KnowledgeGroupKey") or "").strip()

                # DEBUG: Log specific entries that user reported issues with
                if "hexe" in strkey.lower() or "sanctuary" in strkey.lower():
                    log.info("DEBUG Knowledge: StrKey='%s', UITextureName='%s' (from %s)",
                             strkey, ui_texture, path.name)

                # CRITICAL DEBUG: Check if we're getting wrong UITextureName from parent
                if "null" in ui_texture.lower() or not ui_texture:
                    # Check if parent has a different UITextureName
                    parent = ki.getparent()
                    if parent is not None:
                        parent_tex = parent.get("UITextureName", "")
                        child_tex = ki.get("UITextureName", "")
                        log.warning("NESTING ISSUE? StrKey=%s, got UITexture='%s', parent has='%s', direct child attr='%s'",
                                   strkey, ui_texture, parent_tex, child_tex)

                lookup = KnowledgeLookup(
                    strkey=strkey,
                    name=name,
                    desc=desc,
                    ui_texture_name=ui_texture,
                    group_key=group_key,
                    source_file=path.name
                )

                # IMPORTANT: Don't overwrite if existing entry has UITextureName and new one doesn't
                existing = self._knowledge_lookup.get(strkey)
                if existing and existing.ui_texture_name and not ui_texture:
                    log.debug("Keeping existing UITexture for %s: '%s' (skipping empty)", strkey, existing.ui_texture_name)
                    continue  # Keep the existing entry with valid UITextureName

                self._knowledge_lookup[strkey] = lookup
                count += 1

                # DEBUG: Log specific entry AND check for parent attribute issues
                if "DemenissCastle" in strkey or "null" in ui_texture.lower():
                    log.info("DEBUG: Knowledge StrKey=%s, UITexture=%s (file=%s)", strkey, ui_texture, path.name)
                    # Check if parent element has UITextureName that might be interfering
                    parent = ki.getparent()
                    if parent is not None:
                        parent_texture = parent.get("UITextureName", "")
                        if parent_texture:
                            log.warning("DEBUG: Parent element '%s' also has UITextureName='%s'", parent.tag, parent_texture)

        self._stats['knowledge_loaded'] = count
        log.info("Knowledge lookup: %d entries", count)

        # Log sample entries with textures
        with_texture = sum(1 for k in self._knowledge_lookup.values() if k.ui_texture_name)
        log.info("  - With UITextureName: %d", with_texture)

        return count

    def get_knowledge(self, strkey: str) -> Optional[KnowledgeLookup]:
        """Get knowledge entry by StrKey."""
        return self._knowledge_lookup.get(strkey)

    # =========================================================================
    # STEP 2: LOAD ENTRIES (All Modes)
    # =========================================================================

    def load_map_data(
        self,
        faction_folder: Path,
        progress_callback=None
    ) -> int:
        """
        Load MAP (Region) data - KNOWLEDGE FIRST, then FactionNode for positions.

        Priority: KnowledgeInfo (has UITextureName) > FactionNode (has position)
        """
        self._entries.clear()
        self._current_mode = DataMode.MAP
        count = 0
        with_image = 0

        # =================================================================
        # STEP 1: Load ALL KnowledgeInfo entries FIRST (they have UITextureName!)
        # =================================================================
        log.info("Loading Knowledge entries first (they have images)...")
        for strkey, knowledge in self._knowledge_lookup.items():
            if not knowledge.name:
                continue

            ui_texture = knowledge.ui_texture_name
            dds_path = self._dds_index.find(ui_texture) if ui_texture else None
            has_image = dds_path is not None

            # DEBUG: Log specific entries
            if "hexe" in strkey.lower() or "sanctuary" in strkey.lower():
                log.info("DEBUG load_map: StrKey='%s', UITexture='%s', dds_path=%s, has_image=%s",
                         strkey, ui_texture, dds_path, has_image)

            if has_image:
                with_image += 1

            entry = DataEntry(
                strkey=strkey,
                name_kr=knowledge.name,
                desc_kr=knowledge.desc,
                ui_texture_name=ui_texture,
                dds_path=dds_path,
                has_image=has_image,
                position=None,  # Will be updated from FactionNode if available
                knowledge_key=strkey,
                source_file=knowledge.source_file,
                entry_type="Knowledge",
            )

            self._entries[strkey] = entry
            count += 1

        log.info("Loaded %d Knowledge entries (%d with images)", count, with_image)

        # =================================================================
        # STEP 2: Process FactionNodes to ADD position data
        # =================================================================
        if not faction_folder or not faction_folder.exists():
            log.warning("Faction folder not found: %s", faction_folder)
        else:
            log.info("Adding position data from FactionNodes: %s", faction_folder)
            positions_added = 0

            for path in iter_xml_files(faction_folder):
                if progress_callback:
                    progress_callback(f"Loading {path.name}...")

                root = parse_xml(path)
                if root is None:
                    continue

                for fn in root.iter("FactionNode"):
                    strkey = (fn.get("StrKey") or "").strip()
                    knowledge_key = (fn.get("KnowledgeKey") or "").strip()

                    # Parse position
                    position = None
                    pos_str = (fn.get("WorldPosition") or "").strip()
                    if pos_str:
                        parts = re.split(r"[,\s]+", pos_str)
                        if len(parts) >= 3:
                            try:
                                position = (float(parts[0]), float(parts[1]), float(parts[2]))
                            except ValueError:
                                pass

                    # Try to find matching entry by KnowledgeKey first, then by StrKey
                    existing = self._entries.get(knowledge_key) or self._entries.get(strkey)

                    if existing and position:
                        # UPDATE existing entry with position
                        existing.position = position
                        positions_added += 1
                    elif strkey and strkey not in self._entries:
                        # NEW entry from FactionNode (no matching Knowledge)
                        name_kr = (fn.get("Name") or "").strip()
                        desc_kr = (fn.get("Desc") or "").replace("<br/>", "\n").strip()
                        node_type = (fn.get("Type") or "").strip()

                        # Try to get UITextureName from Knowledge lookup
                        ui_texture = ""
                        if knowledge_key:
                            knowledge = self._knowledge_lookup.get(knowledge_key)
                            if knowledge:
                                if not name_kr:
                                    name_kr = knowledge.name
                                if not desc_kr:
                                    desc_kr = knowledge.desc
                                ui_texture = knowledge.ui_texture_name

                        dds_path = self._dds_index.find(ui_texture) if ui_texture else None
                        has_image = dds_path is not None

                        if has_image:
                            with_image += 1

                        entry = DataEntry(
                            strkey=strkey,
                            name_kr=name_kr,
                            desc_kr=desc_kr,
                            ui_texture_name=ui_texture,
                            dds_path=dds_path,
                            has_image=has_image,
                            position=position,
                            knowledge_key=knowledge_key,
                            source_file=path.name,
                            entry_type=node_type,
                        )

                        self._entries[strkey] = entry
                        count += 1

            log.info("Added position to %d entries from FactionNodes", positions_added)

        self._stats['entries_total'] = count
        self._stats['entries_with_image'] = with_image
        self._stats['entries_without_image'] = count - with_image
        # Mode-specific stats for app.py compatibility
        self._stats['faction_nodes_verified'] = with_image
        self._stats['faction_nodes_skipped'] = count - with_image

        log.info("Loaded %d MAP entries: %d with image, %d without",
                 count, with_image, count - with_image)

        # =================================================================
        # DDS DIAGNOSTIC REPORT - Shows exactly what's happening
        # =================================================================
        log.info("=" * 60)
        log.info("DDS DIAGNOSTIC REPORT")
        log.info("=" * 60)
        log.info("DDS Index: %d files indexed from %s",
                 self._dds_index.file_count, self._dds_index._texture_folder)

        # Sample entries: show 5 WITH image and 5 WITHOUT
        with_img_samples = []
        without_img_samples = []
        for strkey, entry in self._entries.items():
            if entry.has_image and len(with_img_samples) < 5:
                with_img_samples.append((strkey, entry.ui_texture_name, entry.dds_path))
            elif not entry.has_image and len(without_img_samples) < 5:
                without_img_samples.append((strkey, entry.ui_texture_name))
            if len(with_img_samples) >= 5 and len(without_img_samples) >= 5:
                break

        log.info("--- SAMPLE ENTRIES WITH IMAGE (HIT) ---")
        for strkey, ui_tex, dds_path in with_img_samples:
            log.info("  HIT: %s -> %s -> %s", strkey, ui_tex, dds_path)

        log.info("--- SAMPLE ENTRIES WITHOUT IMAGE (MISS) ---")
        for strkey, ui_tex in without_img_samples:
            if ui_tex:
                # Has UITextureName but DDS not found - this is a PROBLEM
                log.warning("  MISS: %s -> UITexture='%s' (DDS NOT FOUND!)", strkey, ui_tex)
            else:
                # No UITextureName in XML - expected
                log.info("  MISS: %s -> (no UITextureName in XML)", strkey)

        log.info("=" * 60)

        return count

    def load_character_data(
        self,
        character_folder: Path,
        progress_callback=None
    ) -> int:
        """Load CHARACTER data from CharacterInfo XML files."""
        self._entries.clear()
        self._current_mode = DataMode.CHARACTER
        count = 0
        with_image = 0

        if not character_folder or not character_folder.exists():
            log.error("Character folder not found: %s", character_folder)
            return 0

        log.info("Loading CHARACTER data from: %s", character_folder)

        for path in iter_xml_files(character_folder):
            if progress_callback:
                progress_callback(f"Loading {path.name}...")

            root = parse_xml(path)
            if root is None:
                continue

            for char in root.iter("CharacterInfo"):
                strkey = (char.get("StrKey") or "").strip()
                if not strkey:
                    continue

                if strkey in self._entries:
                    continue

                name_kr = (char.get("CharacterName") or "").strip()
                group = (char.get("CharacterGroup") or "").strip()

                # Extract CHARACTER-specific fields
                use_macro = (char.get("UseMacro") or "").strip()
                age = (char.get("Age") or "").strip()
                job = (char.get("Job") or "").strip()

                # Get UITextureName from nested Knowledge or UIIconPath
                ui_texture = ""
                desc_kr = ""

                for knowledge in char.iter("Knowledge"):
                    ui_texture = (knowledge.get("UITextureName") or "").strip()
                    desc_kr = (knowledge.get("Desc") or "").replace("<br/>", "\n").strip()
                    knowledge_name = (knowledge.get("Name") or "").strip()
                    if knowledge_name:
                        name_kr = knowledge_name
                    if ui_texture:
                        break

                if not ui_texture:
                    ui_texture = (char.get("UIIconPath") or "").strip()

                # Check DDS
                dds_path = self._dds_index.find(ui_texture) if ui_texture else None
                has_image = dds_path is not None

                if has_image:
                    with_image += 1

                entry = DataEntry(
                    strkey=strkey,
                    name_kr=name_kr,
                    desc_kr=desc_kr,
                    ui_texture_name=ui_texture,
                    dds_path=dds_path,
                    has_image=has_image,
                    group=group,
                    source_file=path.name,
                    entry_type="Character",
                    use_macro=use_macro,
                    age=age,
                    job=job,
                )

                self._entries[strkey] = entry
                count += 1

        self._stats['entries_total'] = count
        self._stats['entries_with_image'] = with_image
        self._stats['entries_without_image'] = count - with_image
        # Mode-specific stats for app.py compatibility
        self._stats['characters_verified'] = with_image
        self._stats['characters_skipped'] = count - with_image

        log.info("Loaded %d CHARACTER entries: %d with image, %d without",
                 count, with_image, count - with_image)

        return count

    def load_item_data(
        self,
        knowledge_folder: Path,
        progress_callback=None
    ) -> int:
        """Load ITEM data directly from KnowledgeInfo (already in lookup)."""
        self._entries.clear()
        self._current_mode = DataMode.ITEM
        count = 0
        with_image = 0

        log.info("Loading ITEM data from knowledge lookup...")

        # Track current group
        current_group = ""

        # Re-parse to get group info
        if knowledge_folder and knowledge_folder.exists():
            for path in iter_xml_files(knowledge_folder):
                if progress_callback:
                    progress_callback(f"Loading {path.name}...")

                root = parse_xml(path)
                if root is None:
                    continue

                # Reset group for each file to avoid leaking across files
                current_group = ""

                for elem in root.iter():
                    if elem.tag == "KnowledgeGroupInfo":
                        current_group = (elem.get("GroupName") or "").strip()

                    elif elem.tag == "KnowledgeInfo":
                        strkey = (elem.get("StrKey") or "").strip()
                        if not strkey or strkey in self._entries:
                            continue

                        name_kr = (elem.get("Name") or "").strip()
                        desc_kr = (elem.get("Desc") or "").replace("<br/>", "\n").strip()
                        ui_texture = (elem.get("UITextureName") or "").strip()

                        if not name_kr:
                            continue

                        dds_path = self._dds_index.find(ui_texture) if ui_texture else None
                        has_image = dds_path is not None

                        if has_image:
                            with_image += 1

                        entry = DataEntry(
                            strkey=strkey,
                            name_kr=name_kr,
                            desc_kr=desc_kr,
                            ui_texture_name=ui_texture,
                            dds_path=dds_path,
                            has_image=has_image,
                            group=current_group,
                            source_file=path.name,
                            entry_type="Item",
                        )

                        self._entries[strkey] = entry
                        count += 1

        self._stats['entries_total'] = count
        self._stats['entries_with_image'] = with_image
        self._stats['entries_without_image'] = count - with_image
        # Mode-specific stats for app.py compatibility
        self._stats['items_verified'] = with_image
        self._stats['items_skipped'] = count - with_image

        log.info("Loaded %d ITEM entries: %d with image, %d without",
                 count, with_image, count - with_image)

        return count

    # =========================================================================
    # ACCESSORS
    # =========================================================================

    @property
    def entries(self) -> Dict[str, DataEntry]:
        """Get all entries."""
        return self._entries

    @property
    def knowledge_lookup(self) -> Dict[str, KnowledgeLookup]:
        """Get knowledge lookup table."""
        return self._knowledge_lookup

    @property
    def stats(self) -> dict:
        return self._stats.copy()

    @property
    def audio_category_tree(self) -> Optional[Dict]:
        """Get cached audio category tree (built during load_audio_data)."""
        return self._audio_category_tree

    def get_entry(self, strkey: str) -> Optional[DataEntry]:
        """Get entry by StrKey."""
        return self._entries.get(strkey)

    # Legacy compatibility (mode-aware - returns empty if wrong mode)
    @property
    def faction_nodes(self) -> Dict[str, DataEntry]:
        """Legacy: Get MAP entries (only valid if current_mode == MAP)."""
        if self._current_mode != DataMode.MAP:
            return {}
        return self._entries

    @property
    def characters(self) -> Dict[str, DataEntry]:
        """Legacy: Get CHARACTER entries (only valid if current_mode == CHARACTER)."""
        if self._current_mode != DataMode.CHARACTER:
            return {}
        return self._entries

    @property
    def items(self) -> Dict[str, DataEntry]:
        """Legacy: Get ITEM entries (only valid if current_mode == ITEM)."""
        if self._current_mode != DataMode.ITEM:
            return {}
        return self._entries

    def get_node(self, strkey: str) -> Optional[DataEntry]:
        """Legacy: Get node by strkey."""
        return self._entries.get(strkey)

    def get_character(self, strkey: str) -> Optional[DataEntry]:
        """Legacy: Get character by strkey."""
        return self._entries.get(strkey)

    def get_item(self, strkey: str) -> Optional[DataEntry]:
        """Legacy: Get item by strkey."""
        return self._entries.get(strkey)

    # =========================================================================
    # AUDIO MODE DATA LOADING
    # =========================================================================

    def load_audio_data(
        self,
        audio_folder: Path,
        export_folder: Path,
        loc_folder: Path,
        vrs_folder: Optional[Path] = None,
        progress_callback=None
    ) -> int:
        """
        Load AUDIO data - WEM files with linked script lines.

        Args:
            audio_folder: Folder containing WEM files
            export_folder: Export folder with SoundEventName -> StrOrigin mappings
            loc_folder: Localization folder with language XML files
            vrs_folder: VoiceRecordingSheet folder for chronological ordering

        Returns:
            Number of audio entries loaded
        """
        self._entries.clear()
        self._current_mode = DataMode.AUDIO
        count = 0

        if not audio_folder or not audio_folder.exists():
            log.error("Audio folder not found: %s", audio_folder)
            return 0

        if progress_callback:
            progress_callback("Scanning WEM files...")

        # Build audio index
        audio_index = AudioIndex()
        wem_count = audio_index.scan_folder(audio_folder, progress_callback)
        log.info("Found %d WEM files", wem_count)

        if wem_count == 0:
            return 0

        # Load event name -> StrOrigin mappings from export folder
        if progress_callback:
            progress_callback("Loading event mappings...")
        event_count = audio_index.load_event_mappings(export_folder, progress_callback)
        log.info("Loaded %d event mappings", event_count)

        # Apply VRS ordering (reorders xml_order based on VoiceRecordingSheet)
        if vrs_folder and vrs_folder.exists():
            if progress_callback:
                progress_callback("Applying VRS chronological order...")
            vrs_count = audio_index.load_vrs_order(vrs_folder, progress_callback)
            log.info("VRS reordered %d events", vrs_count)

        # Load StrOrigin -> script line mappings from BOTH KOR and ENG loc files
        if progress_callback:
            progress_callback("Loading script lines (KOR + ENG)...")
        script_count = audio_index.load_script_lines(loc_folder, progress_callback)
        log.info("Loaded %d script lines", script_count)

        # Store audio_index for SearchEngine access
        self._audio_index = audio_index

        # Create entries for each audio file
        if progress_callback:
            progress_callback("Building audio entries...")

        for event_name, wem_path in audio_index.wem_files.items():
            # Get script lines for this event
            script_kr = audio_index.get_script_kor(event_name)
            script_eng = audio_index.get_script_eng(event_name)
            str_origin = audio_index.get_str_origin(event_name)

            entry = DataEntry(
                strkey=event_name,
                name_kr=event_name,
                desc_kr=script_kr,
                ui_texture_name="",  # No image for audio
                dds_path=wem_path,  # Store WEM path here
                has_image=False,
                group=str_origin,
                knowledge_key=script_eng,  # Store ENG script here for audio mode
                source_file=wem_path.name,
                entry_type="Audio",
                export_path=audio_index.get_export_path(event_name),
                xml_order=audio_index.get_xml_order(event_name),
            )

            self._entries[event_name] = entry
            count += 1

        # Cache category tree for GUI
        self._audio_category_tree = audio_index.build_category_tree()

        self._stats['entries_total'] = count
        self._stats['entries_with_image'] = 0
        self._stats['entries_without_image'] = count
        self._stats['audio_with_script'] = sum(1 for e in self._entries.values() if e.desc_kr or e.knowledge_key)
        self._stats['audio_without_script'] = count - self._stats['audio_with_script']

        log.info("Loaded %d AUDIO entries: %d with script, %d without",
                 count, self._stats['audio_with_script'], self._stats['audio_without_script'])

        return count


# =============================================================================
# AUDIO INDEX (For AUDIO mode)
# =============================================================================

@dataclass
class AudioEntry:
    """Audio entry with event name and linked script."""
    event_name: str
    wem_path: Path
    str_origin: str = ""
    script_kr: str = ""
    script_translated: str = ""
    duration: Optional[float] = None


class AudioIndex:
    """
    Index of WEM audio files with event name -> script line mappings.

    Data flow:
    1. WEM files: event_name.wem -> Path
    2. Export XMLs: SoundEventName -> StrOrigin
    3. ENG Loc XML: StrOrigin (KOR text) -> Str (ENG text)
    """

    def __init__(self):
        self._wem_files: Dict[str, Path] = {}  # event_name (lowercase) -> wem path
        self._event_to_origin: Dict[str, str] = {}  # event_name -> StrOrigin key
        self._origin_to_kor: Dict[str, str] = {}  # StrOrigin key -> KOR script
        self._origin_to_eng: Dict[str, str] = {}  # StrOrigin key -> ENG script

        # Export path category tree tracking
        self._event_to_export_path: Dict[str, str] = {}  # event_name -> relative dir
        self._event_to_xml_order: Dict[str, int] = {}  # event_name -> order in file
        self._category_counts: Dict[str, int] = {}  # rel_dir -> count

    @property
    def wem_files(self) -> Dict[str, Path]:
        """Get WEM file mapping."""
        return self._wem_files

    def scan_folder(self, audio_folder: Path, progress_callback=None) -> int:
        """
        Scan folder for WEM files.

        Args:
            audio_folder: Folder containing WEM files

        Returns:
            Number of WEM files found
        """
        self._wem_files.clear()

        if not audio_folder or not audio_folder.exists():
            log.warning("Audio folder not found: %s", audio_folder)
            return 0

        if progress_callback:
            progress_callback(f"Scanning {audio_folder}...")

        count = 0
        for wem_path in audio_folder.rglob("*.wem"):
            event_name = wem_path.stem.lower()
            self._wem_files[event_name] = wem_path
            count += 1

        log.info("Found %d WEM files", count)
        return count

    def load_event_mappings(self, export_folder: Path, progress_callback=None) -> int:
        """
        Load SoundEventName -> StrOrigin mappings from export XMLs.

        Args:
            export_folder: Folder containing export XML files

        Returns:
            Number of mappings loaded
        """
        self._event_to_origin.clear()
        self._event_to_export_path.clear()
        self._event_to_xml_order.clear()
        self._category_counts.clear()

        if not export_folder or not export_folder.exists():
            log.warning("Export folder not found: %s", export_folder)
            return 0

        if progress_callback:
            progress_callback(f"Loading event mappings from {export_folder}...")

        count = 0
        element_order = 0  # Global counter across all files — preserves ordering
        for xml_path in export_folder.rglob("*.xml"):
            if progress_callback:
                progress_callback(f"Parsing {xml_path.name}...")

            root = parse_xml(xml_path)
            if root is None:
                continue

            # Compute relative dir from export root once per file
            # Normalize backslashes to forward slashes for cross-platform consistency
            rel_dir = str(xml_path.relative_to(export_folder).parent).replace("\\", "/")
            if rel_dir == ".":
                rel_dir = ""

            for elem in root.iter():
                event_name = (elem.get("SoundEventName") or "").strip()
                str_origin = (elem.get("StrOrigin") or "").strip()

                if event_name and str_origin:
                    event_lower = event_name.lower()
                    self._event_to_origin[event_lower] = str_origin

                    # Track export path and XML element order
                    self._event_to_export_path[event_lower] = rel_dir
                    self._event_to_xml_order[event_lower] = element_order
                    element_order += 1

                    # Accumulate category counts
                    self._category_counts[rel_dir] = self._category_counts.get(rel_dir, 0) + 1
                    count += 1

        log.info("Loaded %d event -> StrOrigin mappings (%d categories)", count, len(self._category_counts))
        return count

    def load_vrs_order(self, vrs_folder: Path, progress_callback=None) -> int:
        """
        Load VoiceRecordingSheet EventName ordering and apply to xml_order.

        Finds the "EventName" column by header name (flexible, not hardcoded index).
        Events found in VRS get their row position as xml_order.
        Events NOT in VRS keep their original xml_order offset to sort after VRS entries.

        Args:
            vrs_folder: Folder containing VoiceRecordingSheet Excel files

        Returns:
            Number of events reordered by VRS
        """
        if not vrs_folder or not vrs_folder.exists():
            log.warning("VRS folder not found: %s", vrs_folder)
            return 0

        # Find most recent Excel file
        xlsx_files = sorted(vrs_folder.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
        if not xlsx_files:
            log.warning("No Excel files found in VRS folder: %s", vrs_folder)
            return 0

        vrs_file = xlsx_files[0]
        log.info("Loading VRS order from: %s", vrs_file.name)

        if progress_callback:
            progress_callback(f"Loading VRS order from {vrs_file.name}...")

        try:
            from openpyxl import load_workbook
        except ImportError:
            log.warning("openpyxl not installed — VRS ordering disabled")
            return 0

        try:
            wb = load_workbook(vrs_file, data_only=True, read_only=True)
            ws = wb.active

            # Find EventName column by header (row 1)
            event_col_idx = None
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                for idx, cell_value in enumerate(header_row):
                    if cell_value and str(cell_value).strip().lower() == "eventname":
                        event_col_idx = idx
                        break

            if event_col_idx is None:
                log.warning("'EventName' column not found in VRS header: %s", header_row)
                wb.close()
                return 0

            log.info("Found 'EventName' at column index %d", event_col_idx)

            # Build event_name -> VRS row position
            vrs_order: Dict[str, int] = {}
            position = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > event_col_idx:
                    event_name = row[event_col_idx]
                    if event_name:
                        event_lower = str(event_name).strip().lower()
                        if event_lower and event_lower not in vrs_order:
                            vrs_order[event_lower] = position
                            position += 1

            wb.close()
            log.info("Loaded %d EventNames from VRS", len(vrs_order))

            # Apply VRS order: matched events get VRS position,
            # unmatched events get offset to sort after all VRS entries
            vrs_max = len(vrs_order)
            reordered = 0
            for event_name, current_order in list(self._event_to_xml_order.items()):
                vrs_pos = vrs_order.get(event_name)
                if vrs_pos is not None:
                    self._event_to_xml_order[event_name] = vrs_pos
                    reordered += 1
                else:
                    # Offset unmatched events beyond all VRS entries
                    self._event_to_xml_order[event_name] = vrs_max + current_order

            log.info("VRS reordered %d / %d events (%d unmatched appended after)",
                     reordered, len(self._event_to_xml_order),
                     len(self._event_to_xml_order) - reordered)
            return reordered

        except Exception as e:
            log.error("Error loading VRS: %s", e)
            return 0

    def load_script_lines(
        self,
        loc_folder: Path,
        progress_callback=None
    ) -> int:
        """
        Load script lines from BOTH KOR and ENG localization XMLs.

        ENG LOC: StrOrigin (key) -> Str (English text)
        KOR LOC: StrOrigin (key) -> Str (Korean text)

        Args:
            loc_folder: Folder containing languagedata_*.xml files

        Returns:
            Number of script lines loaded
        """
        self._origin_to_kor.clear()
        self._origin_to_eng.clear()

        if not loc_folder or not loc_folder.exists():
            log.warning("Loc folder not found: %s", loc_folder)
            return 0

        total_count = 0

        # Load KOR LOC file
        kor_file = self._find_lang_file(loc_folder, 'kor')
        if kor_file:
            if progress_callback:
                progress_callback(f"Loading KOR script lines from {kor_file.name}...")

            root = parse_xml(kor_file)
            if root is not None:
                for elem in root.iter():
                    str_origin = (elem.get("StrOrigin") or "").strip()
                    text = (elem.get("Str") or "").strip()
                    if str_origin and text:
                        self._origin_to_kor[str_origin] = text
                        total_count += 1

            log.info("Loaded %d KOR script lines", len(self._origin_to_kor))

        # Load ENG LOC file
        eng_file = self._find_lang_file(loc_folder, 'eng')
        if eng_file:
            if progress_callback:
                progress_callback(f"Loading ENG script lines from {eng_file.name}...")

            root = parse_xml(eng_file)
            if root is not None:
                for elem in root.iter():
                    str_origin = (elem.get("StrOrigin") or "").strip()
                    text = (elem.get("Str") or "").strip()
                    if str_origin and text:
                        self._origin_to_eng[str_origin] = text

            log.info("Loaded %d ENG script lines", len(self._origin_to_eng))

        return total_count

    def _find_lang_file(self, loc_folder: Path, lang_code: str) -> Optional[Path]:
        """Find language file for given code."""
        # Try uppercase
        lang_file = loc_folder / f"languagedata_{lang_code.upper()}.xml"
        if lang_file.exists():
            return lang_file

        # Try lowercase
        lang_file = loc_folder / f"languagedata_{lang_code.lower()}.xml"
        if lang_file.exists():
            return lang_file

        # Try finding by pattern
        for f in loc_folder.glob("languagedata_*.xml"):
            if lang_code.lower() in f.name.lower():
                return f

        log.warning("Language file not found for %s in %s", lang_code, loc_folder)
        return None

    def find(self, event_name: str) -> Optional[Path]:
        """
        Find WEM file path for event name.

        Args:
            event_name: Event name to look up

        Returns:
            Path to WEM file or None
        """
        return self._wem_files.get(event_name.lower())

    def get_str_origin(self, event_name: str) -> str:
        """Get StrOrigin key for event name."""
        return self._event_to_origin.get(event_name.lower(), "")

    def get_script_kor(self, event_name: str) -> str:
        """
        Get Korean script line for event name.

        Args:
            event_name: Event name to look up

        Returns:
            Korean script text or empty string
        """
        str_origin = self._event_to_origin.get(event_name.lower())
        if not str_origin:
            return ""
        return self._origin_to_kor.get(str_origin, "")

    def get_script_eng(self, event_name: str) -> str:
        """
        Get English script line for event name.

        Args:
            event_name: Event name to look up

        Returns:
            English script text or empty string
        """
        str_origin = self._event_to_origin.get(event_name.lower())
        if not str_origin:
            return ""
        return self._origin_to_eng.get(str_origin, "")

    def get_script_line(self, event_name: str) -> str:
        """
        Get script line for event name (KOR preferred, ENG fallback).

        Args:
            event_name: Event name to look up

        Returns:
            Script line text or empty string
        """
        # Try KOR first, then ENG
        kor = self.get_script_kor(event_name)
        if kor:
            return kor
        return self.get_script_eng(event_name)

    def get_export_path(self, event_name: str) -> str:
        """Get export path category for event name."""
        return self._event_to_export_path.get(event_name.lower(), "")

    def get_xml_order(self, event_name: str) -> int:
        """Get XML element order for event name."""
        return self._event_to_xml_order.get(event_name.lower(), 0)

    @property
    def category_counts(self) -> Dict[str, int]:
        """Get category path -> entry count mapping."""
        return self._category_counts

    def build_category_tree(self) -> Dict:
        """Build hierarchical tree from flat category paths.

        Returns:
            Nested dict: {"_count": N, "children": {"name": {...}, ...}}
        """
        tree: Dict = {"_count": 0, "children": {}}

        for rel_dir, count in sorted(self._category_counts.items()):
            tree["_count"] += count

            if not rel_dir:
                # Root-level entries
                node = tree.setdefault("children", {}).setdefault("(root)", {"_count": 0, "children": {}})
                node["_count"] += count
                continue

            parts = rel_dir.replace("\\", "/").split("/")
            current = tree
            for part in parts:
                child = current["children"].setdefault(part, {"_count": 0, "children": {}})
                child["_count"] += count
                current = child

        return tree

    @property
    def file_count(self) -> int:
        """Get number of indexed WEM files."""
        return len(self._wem_files)

    @property
    def mapped_count(self) -> int:
        """Get number of events with script line mappings."""
        count = 0
        for event_name in self._wem_files:
            if self.get_script_line(event_name):
                count += 1
        return count
