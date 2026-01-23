"""
Faction & Glossary Tool Suite v2.0
==================================
A 2-button GUI application:

Button 1: Faction List Generator + Glossary Autofiltering
    - Parses faction XML files from factioninfo folder
    - Generates FactionList.xlsx with all faction names
    - (Optional) Filters a selected Excel file by removing rows containing faction names

Button 2: List Concatenator + Multi-language Translator
    - Reads FIRST COLUMN from each Excel file in a selected folder
    - Concatenates all lists with bold yellow file name headers
    - Translates ALL entries to ALL supported languages using StrOrigin matching
    - Outputs CombinedList.xlsx and Translated_*.xlsx files

Data Sources:
- Faction XML: F:\perforce\cd\mainline\resource\GameData\StaticInfo\factioninfo\*.xml
- LOC files: F:\perforce\cd\mainline\resource\GameData\stringtable\loc\languagedata_*.xml
- Configurable via settings.json (drive letter)
"""

import json
import sys
import os
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from dataclasses import dataclass
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from lxml import etree as ET
except ImportError:
    import xml.etree.ElementTree as ET
    print("Note: Using standard XML parser. Install lxml for better performance.")

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import translation utilities
from translation_utils import (
    discover_language_files,
    load_language_table,
    load_all_language_tables,
    translate_list,
    get_translation_stats
)

# =============================================================================
# CONFIGURATION
# =============================================================================

# Detect if running as PyInstaller executable
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent


def load_settings() -> dict:
    """Load runtime settings from settings.json.

    Settings file format:
    {
        "drive_letter": "F",
        "version": "2.0",
        "paths": {
            "factioninfo": "...",
            "loc_folder": "..."
        }
    }
    """
    settings_file = SCRIPT_DIR / "settings.json"

    if not settings_file.exists():
        return {}

    try:
        with open(settings_file, 'r', encoding='utf-8') as f:
            settings = json.load(f)

        # Validate drive_letter if present
        if 'drive_letter' in settings:
            drive = settings['drive_letter']
            if not isinstance(drive, str) or len(drive) != 1 or not drive.isalpha():
                print(f"  WARNING: Invalid drive_letter in settings.json: '{drive}'. Using default F:")
                del settings['drive_letter']

        return settings
    except json.JSONDecodeError as e:
        print(f"  WARNING: Invalid JSON in settings.json: {e}. Using defaults.")
        return {}
    except Exception as e:
        print(f"  WARNING: Error reading settings.json: {e}. Using defaults.")
        return {}


def apply_drive_letter(path_str: str, drive_letter: str) -> str:
    """Replace the default F: drive with the configured drive letter."""
    if path_str.startswith("F:") or path_str.startswith("f:"):
        return f"{drive_letter.upper()}:{path_str[2:]}"
    return path_str


# Load settings at module import time
_SETTINGS = load_settings()
_DRIVE_LETTER = _SETTINGS.get('drive_letter', 'F')

if _DRIVE_LETTER != 'F':
    print(f"  Using custom drive letter: {_DRIVE_LETTER}:")

# Path configurations
FACTIONINFO_FOLDER = Path(apply_drive_letter(
    _SETTINGS.get('paths', {}).get('factioninfo',
        r"F:\perforce\cd\mainline\resource\GameData\StaticInfo\factioninfo"),
    _DRIVE_LETTER
))

LOC_FOLDER = Path(apply_drive_letter(
    _SETTINGS.get('paths', {}).get('loc_folder',
        r"F:\perforce\cd\mainline\resource\GameData\stringtable\loc"),
    _DRIVE_LETTER
))

# Output folder
OUTPUT_FOLDER = SCRIPT_DIR / "Output"
TRANSLATIONS_FOLDER = OUTPUT_FOLDER / "Translations"


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class FactionEntry:
    """A single faction entry."""
    name: str
    faction_type: str  # "FactionGroup", "Faction", or "FactionNode"
    source_file: str


# =============================================================================
# STYLING CONSTANTS
# =============================================================================

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# Yellow fill for section headers in concatenated list
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BOLD_FONT = Font(bold=True)

TYPE_COLORS = {
    "FactionGroup": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "Faction": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    "FactionNode": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
}


# =============================================================================
# XML PARSING (Faction Files)
# =============================================================================

def sanitize_xml(raw: str) -> str:
    """Remove illegal XML characters that can cause parse errors."""
    import re
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', raw)


def parse_xml_file(path: Path) -> Optional[ET._Element]:
    """Parse XML file, return root element or None on error."""
    try:
        raw = path.read_text(encoding="utf-8")
    except Exception as e:
        print(f"  Error reading {path.name}: {e}")
        return None

    cleaned = sanitize_xml(raw)
    wrapped = f"<ROOT>\n{cleaned}\n</ROOT>"

    try:
        if hasattr(ET, 'XMLParser'):
            parser = ET.XMLParser(huge_tree=True, recover=True)
            return ET.fromstring(wrapped.encode("utf-8"), parser=parser)
        else:
            return ET.fromstring(wrapped)
    except Exception as e:
        print(f"  Error parsing {path.name}: {e}")
        return None


def iter_xml_files(folder: Path) -> List[Path]:
    """Get all XML files in folder (non-recursive)."""
    if not folder.exists():
        print(f"  WARNING: Folder not found: {folder}")
        return []
    return sorted([f for f in folder.glob("*.xml") if f.is_file()])


def extract_faction_nodes_recursive(elem, source_file: str, entries: List[FactionEntry]) -> None:
    """Recursively extract FactionNode names from an element and its children."""
    name = elem.get("Name") or ""

    if name:
        entries.append(FactionEntry(
            name=name,
            faction_type="FactionNode",
            source_file=source_file
        ))

    for child in elem:
        if child.tag == "FactionNode":
            extract_faction_nodes_recursive(child, source_file, entries)


def parse_faction_xml_files(folder: Path) -> List[FactionEntry]:
    """Parse all faction XML files and extract faction names."""
    entries: List[FactionEntry] = []
    seen_names: Set[str] = set()

    xml_files = iter_xml_files(folder)

    if not xml_files:
        print(f"  No XML files found in {folder}")
        return entries

    print(f"  Found {len(xml_files)} XML files to parse...")

    for path in xml_files:
        root = parse_xml_file(path)
        if root is None:
            continue

        source_file = path.name

        # Parse FactionGroups
        for fg_elem in root.iter("FactionGroup"):
            group_name = fg_elem.get("GroupName") or ""

            if group_name and group_name not in seen_names:
                entries.append(FactionEntry(
                    name=group_name,
                    faction_type="FactionGroup",
                    source_file=source_file
                ))
                seen_names.add(group_name)

            for f_elem in fg_elem.iter("Faction"):
                faction_name = f_elem.get("Name") or ""

                if faction_name and faction_name not in seen_names:
                    entries.append(FactionEntry(
                        name=faction_name,
                        faction_type="Faction",
                        source_file=source_file
                    ))
                    seen_names.add(faction_name)

                for node_elem in f_elem:
                    if node_elem.tag == "FactionNode":
                        extract_faction_nodes_recursive(node_elem, source_file, entries)

        # Parse standalone Factions
        for f_elem in root:
            if f_elem.tag == "Faction":
                faction_name = f_elem.get("Name") or ""

                if faction_name and faction_name not in seen_names:
                    entries.append(FactionEntry(
                        name=faction_name,
                        faction_type="Faction",
                        source_file=source_file
                    ))
                    seen_names.add(faction_name)

                for node_elem in f_elem:
                    if node_elem.tag == "FactionNode":
                        extract_faction_nodes_recursive(node_elem, source_file, entries)

    # Deduplicate
    unique_entries = []
    final_seen = set()
    for entry in entries:
        if entry.name not in final_seen:
            unique_entries.append(entry)
            final_seen.add(entry.name)

    return unique_entries


# =============================================================================
# EXCEL GENERATION (Faction List)
# =============================================================================

def generate_faction_list_excel(faction_data: List[FactionEntry], output_path: Path) -> None:
    """Generate FactionList.xlsx with all faction names."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faction List"

    headers = ["Faction Name", "Type", "Source File"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, entry in enumerate(faction_data, start=2):
        cell_name = ws.cell(row=row_idx, column=1, value=entry.name)
        cell_name.border = THIN_BORDER

        cell_type = ws.cell(row=row_idx, column=2, value=entry.faction_type)
        cell_type.border = THIN_BORDER
        if entry.faction_type in TYPE_COLORS:
            cell_type.fill = TYPE_COLORS[entry.faction_type]

        cell_source = ws.cell(row=row_idx, column=3, value=entry.source_file)
        cell_source.border = THIN_BORDER

    for col_idx in range(1, 4):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    ws.auto_filter.ref = f"A1:C{len(faction_data) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Saved: {output_path}")


# =============================================================================
# EXCEL FILTERING (Button 1 - Optional)
# =============================================================================

def filter_excel_by_factions(
    input_path: Path,
    faction_names: Set[str],
    output_path: Path
) -> Tuple[int, int, List[Tuple[int, str]]]:
    """Filter an Excel file by removing rows containing faction names."""
    wb = load_workbook(input_path)

    total_removed = 0
    removed_details: List[Tuple[int, str]] = []
    faction_names_lower = {name.lower() for name in faction_names if name}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_to_delete = []

        for row_idx in range(2, ws.max_row + 1):
            row_values = []
            matched_faction = None

            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_values.append(str(cell_value))

            row_text = " ".join(row_values).lower()

            for faction_name in faction_names_lower:
                if faction_name and faction_name in row_text:
                    matched_faction = faction_name
                    break

            if matched_faction:
                rows_to_delete.append((row_idx, matched_faction))

        for row_idx, matched in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
            removed_details.append((row_idx, matched))
            total_removed += 1

    total_rows = sum(ws.max_row - 1 for ws in wb.worksheets)

    wb.save(output_path)
    print(f"  Saved filtered file: {output_path}")

    return total_rows + total_removed, total_removed, removed_details


# =============================================================================
# LIST CONCATENATION (Button 2)
# =============================================================================

def read_first_column_from_excel(file_path: Path) -> List[str]:
    """
    Read first column from an Excel file, skipping header row.

    Args:
        file_path: Path to Excel file

    Returns:
        List of values from first column (row 2 onwards)
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        values = []

        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value is not None:
                values.append(str(cell_value).strip())

        wb.close()
        return values
    except Exception as e:
        print(f"  Error reading {file_path.name}: {e}")
        return []


def concatenate_excel_lists(
    folder_path: Path,
    output_path: Path
) -> Tuple[List[str], List[Tuple[str, List[str]]]]:
    """
    Read all Excel files in folder and concatenate their first columns.

    Args:
        folder_path: Folder containing Excel files
        output_path: Path for output CombinedList.xlsx

    Returns:
        Tuple of:
        - all_items: Flat list of all items (no headers)
        - sections: List of (filename, items) for each file
    """
    excel_files = sorted(folder_path.glob("*.xlsx"))

    if not excel_files:
        print(f"  No Excel files found in {folder_path}")
        return [], []

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined List"

    # Set column header
    header_cell = ws.cell(row=1, column=1, value="Item")
    header_cell.fill = HEADER_FILL
    header_cell.font = HEADER_FONT
    header_cell.alignment = HEADER_ALIGNMENT
    header_cell.border = THIN_BORDER

    all_items: List[str] = []
    sections: List[Tuple[str, List[str]]] = []
    current_row = 2

    for excel_file in excel_files:
        # Skip temp files
        if excel_file.name.startswith("~$"):
            continue

        file_name = excel_file.stem  # Filename without extension
        items = read_first_column_from_excel(excel_file)

        if not items:
            continue

        sections.append((file_name, items))

        # Add section header (file name with yellow background)
        header_cell = ws.cell(row=current_row, column=1, value=f"[{file_name}]")
        header_cell.fill = YELLOW_FILL
        header_cell.font = BOLD_FONT
        header_cell.border = THIN_BORDER
        current_row += 1

        # Add items
        for item in items:
            ws.cell(row=current_row, column=1, value=item).border = THIN_BORDER
            all_items.append(item)
            current_row += 1

    # Auto-fit column width
    max_length = 0
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions['A'].width = min(max_length + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Saved: {output_path}")

    return all_items, sections


# =============================================================================
# TRANSLATION OUTPUT (Button 2)
# =============================================================================

def generate_translation_excel(
    korean_items: List[str],
    lang_code: str,
    lang_table: Dict[str, str],
    output_path: Path
) -> Dict[str, int]:
    """
    Generate a translation Excel file for a single language.

    Args:
        korean_items: List of Korean strings to translate
        lang_code: Language code (e.g., "eng", "fre")
        lang_table: Translation lookup table
        output_path: Output file path

    Returns:
        Stats dict: {"translated": N, "missing": M}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Translation_{lang_code.upper()}"

    # Headers
    headers = ["SourceText (Korean)", f"Translation ({lang_code.upper()})"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Translate and write
    translated_count = 0
    missing_count = 0

    # Remove duplicates while preserving order
    seen = set()
    unique_items = []
    for item in korean_items:
        if item and item not in seen:
            seen.add(item)
            unique_items.append(item)

    for row_idx, korean in enumerate(unique_items, start=2):
        # Source (Korean)
        ws.cell(row=row_idx, column=1, value=korean).border = THIN_BORDER

        # Translation
        translation = lang_table.get(korean, "NO_TRANSLATION")
        target_cell = ws.cell(row=row_idx, column=2, value=translation)
        target_cell.border = THIN_BORDER

        if translation == "NO_TRANSLATION":
            target_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            missing_count += 1
        else:
            translated_count += 1

    # Auto-fit columns
    for col_idx in range(1, 3):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {"translated": translated_count, "missing": missing_count}


# =============================================================================
# GUI APPLICATION
# =============================================================================

class FactionToolGUI:
    """Main GUI for Faction & Glossary Tool Suite."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Faction & Glossary Tool Suite v2.0")
        self.root.geometry("700x500")
        self.root.resizable(True, True)

        self._create_widgets()

    def _create_widgets(self):
        """Build the GUI."""
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title = ttk.Label(
            main_frame,
            text="Faction & Glossary Tool Suite",
            font=("Helvetica", 16, "bold")
        )
        title.pack(pady=(0, 15))

        # Button 1 Frame
        btn1_frame = ttk.LabelFrame(
            main_frame,
            text="Button 1: Faction List Generator",
            padding="10"
        )
        btn1_frame.pack(fill=tk.X, pady=10)

        ttk.Label(
            btn1_frame,
            text="• Extracts faction names from XML files\n• Optional: Filter Excel by faction names",
            justify=tk.LEFT
        ).pack(anchor=tk.W)

        ttk.Button(
            btn1_frame,
            text="Generate Faction List",
            command=self._do_faction_list,
            width=30
        ).pack(pady=(10, 0))

        # Button 2 Frame
        btn2_frame = ttk.LabelFrame(
            main_frame,
            text="Button 2: List Concatenator + Translator",
            padding="10"
        )
        btn2_frame.pack(fill=tk.X, pady=10)

        ttk.Label(
            btn2_frame,
            text="• Combines Excel lists from folder\n• Translates to all LOC languages",
            justify=tk.LEFT
        ).pack(anchor=tk.W)

        ttk.Button(
            btn2_frame,
            text="Concatenate & Translate",
            command=self._do_concat_translate,
            width=30
        ).pack(pady=(10, 0))

        # Progress
        self.progress_var = tk.DoubleVar()
        self.progress_label = ttk.Label(main_frame, text="")
        self.progress_label.pack(pady=(15, 0))

        self.progress_bar = ttk.Progressbar(
            main_frame,
            variable=self.progress_var,
            maximum=100
        )
        self.progress_bar.pack(fill=tk.X, pady=5)

        # Status log
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="5")
        status_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.status_text = tk.Text(status_frame, height=10, width=80, state=tk.DISABLED)
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=scrollbar.set)

        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Paths info
        paths_frame = ttk.LabelFrame(main_frame, text="Configured Paths", padding="5")
        paths_frame.pack(fill=tk.X)

        ttk.Label(
            paths_frame,
            text=f"Faction XML: {FACTIONINFO_FOLDER}\nLOC Folder: {LOC_FOLDER}",
            font=("Courier", 8)
        ).pack(anchor=tk.W)

    def _log(self, message: str):
        """Add message to status log."""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _update_progress(self, current: int, total: int, message: str):
        """Update progress bar and label."""
        percent = (current / total * 100) if total > 0 else 0
        self.progress_var.set(percent)
        self.progress_label.config(text=message)
        self.root.update_idletasks()

    def _clear_status(self):
        """Clear status log."""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.progress_var.set(0)
        self.progress_label.config(text="")

    # -------------------------------------------------------------------------
    # Button 1: Faction List Generator
    # -------------------------------------------------------------------------

    def _do_faction_list(self):
        """Execute Button 1: Faction List Generator + optional filtering."""
        self._clear_status()
        self._log("=" * 50)
        self._log("FACTION LIST GENERATOR")
        self._log("=" * 50)

        # Ask for optional input file
        input_file = filedialog.askopenfilename(
            title="Select Excel file to filter (or Cancel to just generate faction list)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if input_file:
            self._log(f"Selected file: {Path(input_file).name}")
        else:
            self._log("No file selected - will only generate faction list")

        # Check faction folder
        if not FACTIONINFO_FOLDER.exists():
            self._log(f"\nERROR: Faction folder not found!")
            self._log(f"  Expected: {FACTIONINFO_FOLDER}")
            self._log(f"  Check settings.json drive_letter")
            messagebox.showerror(
                "Error",
                f"Faction folder not found!\n\n{FACTIONINFO_FOLDER}\n\nCheck settings.json"
            )
            return

        # Parse faction files
        self._log(f"\nParsing faction XML files...")
        self._update_progress(20, 100, "Parsing faction XML...")

        faction_data = parse_faction_xml_files(FACTIONINFO_FOLDER)

        if not faction_data:
            self._log("ERROR: No faction data extracted!")
            messagebox.showerror("Error", "No faction data found in XML files!")
            return

        # Count by type
        type_counts = {}
        for entry in faction_data:
            type_counts[entry.faction_type] = type_counts.get(entry.faction_type, 0) + 1

        self._log(f"Extracted {len(faction_data)} unique faction names:")
        for ftype, count in sorted(type_counts.items()):
            self._log(f"  - {ftype}: {count}")

        # Generate FactionList.xlsx
        self._log("\nGenerating FactionList.xlsx...")
        self._update_progress(50, 100, "Generating Excel...")

        faction_list_path = OUTPUT_FOLDER / "FactionList.xlsx"
        generate_faction_list_excel(faction_data, faction_list_path)

        # Filter input file if provided
        filtered_path = None
        total_rows = 0
        removed_rows = 0

        if input_file:
            self._log(f"\nFiltering {Path(input_file).name}...")
            self._update_progress(70, 100, "Filtering Excel...")

            faction_names = {entry.name for entry in faction_data}
            input_path = Path(input_file)
            output_name = f"Filtered_{input_path.name}"
            filtered_path = OUTPUT_FOLDER / output_name

            try:
                total_rows, removed_rows, removed_details = filter_excel_by_factions(
                    input_path, faction_names, filtered_path
                )
                self._log(f"  Rows processed: {total_rows}")
                self._log(f"  Rows removed: {removed_rows}")
                self._log(f"  Remaining: {total_rows - removed_rows}")

                # Generate report
                if removed_details:
                    report_path = OUTPUT_FOLDER / f"FilterReport_{input_path.stem}.txt"
                    with open(report_path, 'w', encoding='utf-8') as f:
                        f.write(f"Faction Filter Report\n")
                        f.write(f"Input: {input_path.name}\n")
                        f.write(f"Total rows removed: {removed_rows}\n\n")
                        f.write("Removed rows (original row number, matched faction):\n")
                        for row_num, faction in removed_details[:100]:
                            f.write(f"  Row {row_num}: {faction}\n")
                        if len(removed_details) > 100:
                            f.write(f"  ... and {len(removed_details) - 100} more\n")
                    self._log(f"  Report: {report_path.name}")

            except Exception as e:
                self._log(f"ERROR filtering file: {e}")

        # Complete
        self._update_progress(100, 100, "Complete!")
        self._log("\n" + "=" * 50)
        self._log("COMPLETE!")
        self._log("=" * 50)

        # Show completion message
        message = f"Faction List Generator Complete!\n\n"
        message += f"Generated: {faction_list_path.name}\n"
        message += f"Total factions found: {len(faction_data)}\n"

        if filtered_path:
            message += f"\nFiltered file: {filtered_path.name}\n"
            message += f"Original rows: {total_rows}\n"
            message += f"Rows removed: {removed_rows}"

        message += f"\n\nOutput folder: {OUTPUT_FOLDER}"
        messagebox.showinfo("Complete", message)

    # -------------------------------------------------------------------------
    # Button 2: List Concatenator + Translator
    # -------------------------------------------------------------------------

    def _do_concat_translate(self):
        """Execute Button 2: Concatenate lists and translate to all languages."""
        self._clear_status()
        self._log("=" * 50)
        self._log("LIST CONCATENATOR + TRANSLATOR")
        self._log("=" * 50)

        # Select folder with Excel files
        folder_path = filedialog.askdirectory(
            title="Select folder containing Excel files"
        )

        if not folder_path:
            self._log("Cancelled - no folder selected")
            return

        folder_path = Path(folder_path)
        self._log(f"Selected folder: {folder_path}")

        # Check for Excel files
        excel_files = list(folder_path.glob("*.xlsx"))
        excel_files = [f for f in excel_files if not f.name.startswith("~$")]

        if not excel_files:
            self._log("ERROR: No Excel files found in folder!")
            messagebox.showerror("Error", "No Excel files found in the selected folder!")
            return

        self._log(f"Found {len(excel_files)} Excel files")

        # Step 1: Concatenate lists
        self._log("\nStep 1: Concatenating Excel lists...")
        self._update_progress(10, 100, "Concatenating lists...")

        combined_path = OUTPUT_FOLDER / "CombinedList.xlsx"
        all_items, sections = concatenate_excel_lists(folder_path, combined_path)

        if not all_items:
            self._log("ERROR: No items found in Excel files!")
            messagebox.showerror("Error", "No items found in the Excel files!")
            return

        self._log(f"  Total items: {len(all_items)}")
        self._log(f"  Unique items: {len(set(all_items))}")
        for filename, items in sections:
            self._log(f"    [{filename}]: {len(items)} items")

        # Step 2: Load all language tables
        self._log("\nStep 2: Loading language data...")
        self._update_progress(20, 100, "Loading language data...")

        if not LOC_FOLDER.exists():
            self._log(f"ERROR: LOC folder not found: {LOC_FOLDER}")
            messagebox.showerror(
                "Error",
                f"LOC folder not found!\n\n{LOC_FOLDER}\n\nCheck settings.json"
            )
            return

        lang_files = discover_language_files(LOC_FOLDER)

        if not lang_files:
            self._log("ERROR: No language files found!")
            messagebox.showerror("Error", "No language files found in LOC folder!")
            return

        self._log(f"Found {len(lang_files)} language files")

        # Step 3: Generate translations for each language
        self._log("\nStep 3: Generating translations...")

        TRANSLATIONS_FOLDER.mkdir(parents=True, exist_ok=True)

        total_langs = len(lang_files)
        translated_langs = []
        total_translated = 0
        total_missing = 0

        for idx, (lang_code, xml_path) in enumerate(lang_files.items()):
            progress_pct = 20 + int(75 * (idx + 1) / total_langs)
            self._update_progress(progress_pct, 100, f"Translating to {lang_code.upper()}...")

            # Load language table
            lang_table = load_language_table(xml_path)

            if not lang_table:
                self._log(f"  Skipping {lang_code.upper()} - no data loaded")
                continue

            # Generate translation file
            output_path = TRANSLATIONS_FOLDER / f"Translated_{lang_code.upper()}.xlsx"
            stats = generate_translation_excel(all_items, lang_code, lang_table, output_path)

            translated_langs.append(lang_code.upper())
            total_translated += stats["translated"]
            total_missing += stats["missing"]

            self._log(f"  {lang_code.upper()}: {stats['translated']} translated, {stats['missing']} missing")

        # Complete
        self._update_progress(100, 100, "Complete!")
        self._log("\n" + "=" * 50)
        self._log("COMPLETE!")
        self._log("=" * 50)
        self._log(f"Generated {len(translated_langs)} translation files")
        self._log(f"Output folder: {TRANSLATIONS_FOLDER}")

        # Show completion message
        unique_count = len(set(all_items))
        avg_translated = total_translated / len(translated_langs) if translated_langs else 0
        avg_missing = total_missing / len(translated_langs) if translated_langs else 0

        message = f"List Concatenator + Translator Complete!\n\n"
        message += f"Source files: {len(sections)}\n"
        message += f"Total items: {len(all_items)}\n"
        message += f"Unique items: {unique_count}\n\n"
        message += f"Languages translated: {len(translated_langs)}\n"
        message += f"Avg. translated per lang: {int(avg_translated)}\n"
        message += f"Avg. missing per lang: {int(avg_missing)}\n\n"
        message += f"Output: {TRANSLATIONS_FOLDER}"

        messagebox.showinfo("Complete", message)

    def run(self):
        """Start the GUI."""
        self.root.mainloop()


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    print("=" * 60)
    print("Faction & Glossary Tool Suite v2.0")
    print("=" * 60)

    # Ensure output folders exist
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    TRANSLATIONS_FOLDER.mkdir(parents=True, exist_ok=True)

    # Launch GUI
    app = FactionToolGUI()
    app.run()


if __name__ == "__main__":
    main()
