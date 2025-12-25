"""
Windows Excel File Tests

Tests for Excel file handling on Windows.
Uses openpyxl for .xlsx files.
"""

import os
import platform
import tempfile
from pathlib import Path

import pytest


pytestmark = pytest.mark.skipif(
    platform.system() != "Windows",
    reason="Windows-only tests"
)


class TestExcelImport:
    """Test Excel import functionality."""

    @pytest.fixture
    def openpyxl(self):
        """Get openpyxl or skip if not installed."""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_openpyxl_importable(self, openpyxl):
        """openpyxl should be importable."""
        assert openpyxl is not None

    def test_create_workbook(self, openpyxl):
        """Should be able to create a workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        assert ws['A1'].value == "Test"

    def test_save_and_load_workbook(self, openpyxl):
        """Should be able to save and load workbook."""
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "StringID"
        ws['B1'] = "Source"
        ws['C1'] = "Target"
        ws['A2'] = "STR_001"
        ws['B2'] = "Hello"
        ws['C2'] = "안녕하세요"

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            wb.close()

            # Load and verify
            wb2 = openpyxl.load_workbook(temp_path)
            ws2 = wb2.active
            assert ws2['A1'].value == "StringID"
            assert ws2['C2'].value == "안녕하세요"
            wb2.close()
        finally:
            os.unlink(temp_path)

    def test_korean_in_cells(self, openpyxl):
        """Korean text in cells should work."""
        wb = openpyxl.Workbook()
        ws = wb.active

        korean_texts = [
            "안녕하세요",
            "게임 로컬라이제이션",
            "{0}님 환영합니다",
            "레벨: {level}",
        ]

        for i, text in enumerate(korean_texts, 1):
            ws.cell(row=i, column=1, value=text)

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            wb.close()

            # Reload and verify
            wb2 = openpyxl.load_workbook(temp_path)
            ws2 = wb2.active
            for i, text in enumerate(korean_texts, 1):
                assert ws2.cell(row=i, column=1).value == text
            wb2.close()
        finally:
            os.unlink(temp_path)

    def test_large_file_handling(self, openpyxl):
        """Should handle larger files (1000 rows)."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Create 1000 rows
        for i in range(1, 1001):
            ws.cell(row=i, column=1, value=f"STR_{i:04d}")
            ws.cell(row=i, column=2, value=f"Source text {i}")
            ws.cell(row=i, column=3, value=f"번역 텍스트 {i}")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            wb.close()

            # Verify file was created and has content
            assert Path(temp_path).exists()
            assert Path(temp_path).stat().st_size > 10000  # Should be reasonably sized

            # Reload and spot check
            wb2 = openpyxl.load_workbook(temp_path)
            ws2 = wb2.active
            assert ws2.cell(row=500, column=1).value == "STR_0500"
            assert ws2.cell(row=1000, column=3).value == "번역 텍스트 1000"
            wb2.close()
        finally:
            os.unlink(temp_path)


class TestExcelPaths:
    """Test Excel file paths on Windows."""

    @pytest.fixture
    def openpyxl(self):
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_save_to_appdata(self, openpyxl):
        """Should be able to save Excel to AppData."""
        appdata = os.environ.get("APPDATA")
        save_path = Path(appdata) / "LocaNext" / "exports"
        save_path.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Test Export"

        excel_file = save_path / "test_export.xlsx"
        wb.save(str(excel_file))
        wb.close()

        assert excel_file.exists()
        excel_file.unlink()

    def test_path_with_spaces(self, openpyxl):
        """Should handle paths with spaces."""
        appdata = os.environ.get("APPDATA")
        save_path = Path(appdata) / "LocaNext" / "test folder with spaces"
        save_path.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Spaces Test"

        excel_file = save_path / "file with spaces.xlsx"
        wb.save(str(excel_file))
        wb.close()

        assert excel_file.exists()

        # Reload
        wb2 = openpyxl.load_workbook(str(excel_file))
        assert wb2.active['A1'].value == "Spaces Test"
        wb2.close()

        excel_file.unlink()
        save_path.rmdir()

    def test_korean_filename(self, openpyxl):
        """Should handle Korean filenames."""
        appdata = os.environ.get("APPDATA")
        save_path = Path(appdata) / "LocaNext" / "exports"
        save_path.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "한글 파일명 테스트"

        excel_file = save_path / "한글파일명.xlsx"
        wb.save(str(excel_file))
        wb.close()

        assert excel_file.exists()

        # Reload
        wb2 = openpyxl.load_workbook(str(excel_file))
        assert wb2.active['A1'].value == "한글 파일명 테스트"
        wb2.close()

        excel_file.unlink()
